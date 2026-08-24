from nicegui import ui, app
import intranet_data
import intranet_logger
import datetime
import re
import unicodedata
from intranet_ui_utils import refreshable_na_klienta


def _bez_diakritiky(s: str) -> str:
    """Malá písmena bez diakritiky (pro porovnávání názvů listů)."""
    return ''.join(
        c for c in unicodedata.normalize('NFKD', str(s))
        if not unicodedata.combining(c)
    ).lower()

POBOCKY = [
    'Pardubice', 'Praha', 'Jilemnice', 'Most', 'Liberec', 'Hodonín',
    'Winelife', 'Zlín', 'Ostrava', 'Olomouc', 'České Budějovice',
    'Plzeň', 'Horšovský Týn', 'Nová Role', 'Přefakturace', 'Bourárna', 'Kamiony',
]

POBOCKY_KLICE = {
    'Pardubice':        'pardubice',
    'Praha':            'praha',
    'Jilemnice':        'jilemnice',
    'Most':             'most',
    'Liberec':          'liberec',
    'Hodonín':          'hodonin',
    'Winelife':         'winelife',
    'Zlín':             'zlin',
    'Ostrava':          'ostrava',
    'Olomouc':          'olomouc',
    'České Budějovice': 'ceske_budejovice',
    'Plzeň':            'plzen',
    'Horšovský Týn':    'horsovsky_tyn',
    'Nová Role':        'nova_role',
    'Přefakturace':     'prefakturace',
    'Bourárna':         'bourarna',
    'Kamiony':          'kamiony',
}

# Vnořené pobočky: rodičovská pobočka → [(zobrazený_název, datový_klíč)]
# Pozn.: datový klíč 'Brno' (středisko 030) je vnořený jen pod Hodonínem;
# 'Winelife' (středisko 028) je samostatná pobočka v POBOCKY, ne podzáložka.
_VNORENE_POBOCKY: dict[str, list[tuple[str, str]]] = {
    'Hodonín': [('030 - Brno', 'Brno')],
    'Ostrava': [('027-B.Olomouc', 'B.Olomouc'), ('031-B.Jeseník', 'B.Jeseník')],
}

OSTATNI_PROVOZY: list[tuple[str, str]] = [
    ('Praha-Becher', 'Praha-Becher'),
    ('326-B.Ostrava CC', '326-B.Ostrava CC'),
    ('077-Benstar rest.', '077-Benstar rest.'),
    ('082-B.Gastro OVA', '082-B.Gastro OVA'),
    ('097-Restaur.Pha', '097-Restaur.Pha'),
    ('081-Gastrostud.Pha', '081-Gastrostud.Pha'),
    ('Supervizor', 'Supervizor'),
    ('Údržba', 'Údržba'),
]

MESICE_DB = [
    'leden', 'unor', 'brezen', 'duben', 'kveten', 'cerven',
    'cervenec', 'srpen', 'zari', 'rijen', 'listopad', 'prosinec',
]
MESICE_NAZVY = [
    'Leden', 'Únor', 'Březen', 'Duben', 'Květen', 'Červen',
    'Červenec', 'Srpen', 'Září', 'Říjen', 'Listopad', 'Prosinec',
]

VYCHOZI_OSNOVA = [
    ('501.100', 'Spotřeba materiálu'),
    ('501.110', 'Spotřeba materiálu-tonery'),
    ('501.200', 'Spotřeba mat. kancelářské pot.'),
    ('501.300', 'Pohonné hmoty'),
    ('501.400', 'Spotřeba materiálu-auta'),
    ('501.500', 'DHM'),
    ('501.510', 'DHM do 3 tis.Kč'),
    ('501.520', 'DHM zápůjčka'),
    ('02x.xxx', 'Investice rozpočtem'),
    ('501.600', 'Spotřeba OOPP - prac.oděvy'),
    ('501.700', 'Spotřeba materiálu - reklama'),
    ('501.900', 'Spotřeba mater. nedaňová'),
    ('502.100', 'Elektrická energie'),
    ('502.100', 'Elektrická energie-odtížit CC Praha 74100 kWh'),
    ('502.200', 'Zemní plyn'),
    ('502.200', 'Zemní plyn - odtížit G'),
    ('503.100', 'Vodné'),
    ('503.100', 'Vodné - odtížit G'),
    ('503.200', 'Teplo'),
    ('511.100', 'Opravy vozidel'),
    ('511.2xx', 'Opravy nemovitostí'),
    ('511.300', 'Opravy ostatního majetku'),
    ('512.100', 'Cestovné'),
    ('513.900', 'Náklady na reprezentaci'),
    ('518.100', 'Poštovné'),
    ('518.200', 'Telefony pevné + internet'),
    ('518.210', 'Telefony mobilní'),
    ('518.210', 'Telefony mobilní car control'),
    ('518.230', 'Služby IT'),
    ('518.300', 'Nájemné - nebytové prostory'),
    ('518.300', 'Nájemné - oblouk.hala+TVIN'),
    ('518.310', 'Nájemné - ostatní'),
    ('518.320', 'Služby k nájemnému'),
    ('518.400', 'Reklama a propagace'),
    ('518.410', 'Reklama - letákové akce'),
    ('518.500', 'Služby ostatní'),
    ('518.510', 'Přepravné'),
    ('518.520', 'Školení'),
    ('518.530', 'Odpadové hospodářství'),
    ('518.530', 'Stočné'),
    ('518.540', 'Zprostředkování, poradenství'),
    ('518.550', 'Úklid a ostraha objektů'),
    ('518.560', 'Subdodavatelé'),
    ('518.570', 'Stravenky, provize'),
    ('518.600', 'Leasing -auta'),
    ('518.600', 'Leasing-ostatní'),
    ('518.700', 'Audit, daně, právník'),
    ('518.800', 'Software'),
    ('518.900', 'Ostatní služby nedaňové'),
    ('527.100', 'Náklady na stravování'),
    ('527.200', 'Ostatní sociální náklady'),
    ('531.100', 'Daň silniční'),
    ('532.100', 'Dań z nemovitostí'),
    ('538.100', 'Poplatky státu, obcím'),
    ('538.200', 'Poplatky - mýtné'),
    ('538.300', 'Poplatky - ostatní'),
    ('543.000', 'Dary'),
    ('544.100', 'Smluvní pokuty'),
    ('545.xxx', 'Pokuty'),
    ('548.xxx', 'Pojištění'),
    ('548.30',  'Náhrady škod'),
    ('548.900', 'Ost.provoz.náklady'),
    ('568.101', 'Poplatky GP QR platby'),
    ('',        'Kamionová doprava'),
    ('602.xxx', 'Vratka-stravenky vybr.navíc'),
    ('648.',    'Poj.plnění'),
    ('648.',    'Druhotné suroviny'),
    ('602.103', 'Sazka'),
    ('644.100', 'Úroky a sml.pokuty                     V'),
    ('644.100', 'Úroky a sml.pokuty neuhraz.  N'),
    ('501.',    'Centrála'),
]

# ag-Grid value formatters / cell style (JavaScript strings)
_CS_FMT = (
    "function(p){"
    "if(p.value===null||p.value===undefined)return '—';"
    "var n=parseFloat(p.value);"
    "if(isNaN(n)||n===0)return '—';"
    "return new Intl.NumberFormat('cs-CZ',{minimumFractionDigits:2,maximumFractionDigits:2}).format(n);"
    "}"
)
_DIFF_STYLE = (
    "function(p){"
    "if(!p.value||p.value===0)return{};"
    "return p.value>0?{color:'#dc2626',fontWeight:'bold'}:{color:'#16a34a',fontWeight:'bold'};"
    "}"
)
_CELKEM_GETTER = (
    "function(p){"
    "var fs=['leden','unor','brezen','duben','kveten','cerven','cervenec','srpen','zari','rijen','listopad','prosinec'];"
    "return fs.reduce(function(s,f){return s+(parseFloat(p.data[f])||0)},0);"
    "}"
)
# Styl pro připnutý součtový řádek dole (tučně, zvýrazněné pozadí).
# Pozn.: v NiceGUI 3.8 se JS funkce do ag-gridu předávají klíčem s prefixem ":".
_PINNED_BOTTOM_STYLE = (
    "function(p){"
    "if(p.node&&p.node.rowPinned==='bottom')"
    "return{fontWeight:'bold',backgroundColor:'#f1f5f9',borderTop:'2px solid #94a3b8'};"
    "return null;}"
)

# Pevná výška gridu s vnitřním scrollováním (responzivní dle výšky okna)
_GRID_STYLE = 'height: calc(100vh - 340px); min-height: 380px'

# Auto-šířka sloupců (ag-Grid v32). Cíl: nikdy neoříznout text ani čísla a
# zároveň nenechat vpravo prázdné místo – funguje i po zoomu prohlížeče a
# u více gridů zobrazených naráz (Obraty/Zisk).
#   1) autoSizeAllColumns() → každý sloupec na šířku svého obsahu (vč. záhlaví)
#   2) sizeColumnsToFit() s columnLimits, kde minWidth každého sloupce = jeho
#      šířka podle obsahu. Fit pak může sloupce jen ZVĚTŠIT (nikdy pod obsah →
#      nic se neuseče): je-li místo navíc, roztáhne je na celou šířku (a sám
#      započítá svislý posuvník → žádná mezera); je-li obsah širší, nechá je na
#      šířce obsahu a zobrazí vodorovný posuvník. Vše per-grid (p.api), bez
#      měření DOM. rAF zajistí výpočet až po dokončení layoutu.
# Šířky sloupců počítáme sami z obsahu (canvas) – včetně PINNED součtového řádku
# (autoSizeAllColumns ho ignoruje → součty se ořezávaly) a nejdelšího slova hlavičky
# (hlavička se zalamuje, takže ji nemusíme měřit celou). Nic se neořízne, sloupce
# jsou kompaktní. Header-based minWidth (dle délky textu) se cappuje na 80 px, aby
# nebránilo zúžení; menší explicitní minWidth (ikony) se zachová.
_AUTOSIZE_FIT = (
    "function(p){"
    "var api=p.api;"
    "if(!api||!api.getColumns||!api.sizeColumnsToFit)return;"
    "requestAnimationFrame(function(){"
    "try{"
    "var cols=api.getColumns().filter(function(c){return c.isVisible();});"
    "if(!cols.length)return;"
    "var pad=24,smp=document.querySelector('.ag-cell');"
    "if(smp){var cs=getComputedStyle(smp);pad=(parseFloat(cs.paddingLeft)||8)+(parseFloat(cs.paddingRight)||8);}"
    "pad+=10;"
    "var normal=[];api.forEachNode(function(n){if(n&&n.data)normal.push(n.data);});"
    "var pinned=[],pbc=api.getPinnedBottomRowCount?api.getPinnedBottomRowCount():0,i;"
    "for(i=0;i<pbc;i++){var rn=api.getPinnedBottomRow(i);if(rn&&rn.data)pinned.push(rn.data);}"
    "var ctx=document.createElement('canvas').getContext('2d'),FF='Arial,Helvetica,sans-serif';"
    "function W(s,b){ctx.font=(b?'bold ':'')+'13px '+FF;return ctx.measureText(s).width;}"
    "function F(cd,v,d){try{if(typeof cd.valueFormatter==='function'){var r=cd.valueFormatter({value:v,data:d,colDef:cd,api:api});return r==null?'':(''+r);}}catch(e){}return v==null?'':(''+v);}"
    "var lim=cols.map(function(c){"
    "var cd=c.getColDef(),id=c.getColId(),f=cd.field||id,mx=0,k;"
    "(''+(cd.headerName==null?'':cd.headerName)).split(/\\s+/).forEach(function(wd){var x=W(wd,true);if(x>mx)mx=x;});"
    "for(k=0;k<normal.length;k++){var a=W(F(cd,normal[k][f],normal[k]),false);if(a>mx)mx=a;}"
    "for(k=0;k<pinned.length;k++){var b=W(F(cd,pinned[k][f],pinned[k]),true);if(b>mx)mx=b;}"
    "var need=Math.ceil(mx)+pad,fl=44;"
    "if(cd.minWidth){fl=cd.minWidth<80?cd.minWidth:80;}"
    "if(need<fl)need=fl;"
    "return{key:id,minWidth:need};"
    "});"
    "api.sizeColumnsToFit({columnLimits:lim});"
    "}catch(e){try{api.sizeColumnsToFit();}catch(e2){}}"
    "});"
    "}"
)


# ─── Zámky měsíců (publikace podrobných nákladů) ─────────────────────────────

def _zamky_key(pobocka: str, rok: int) -> str:
    return f'naklady_zamky_{pobocka}_{rok}'

def _nacti_zamky(pobocka: str, rok: int) -> list[int]:
    """Vrátí list zamčených měsíců (1-12) pro pobočku a rok."""
    return app.storage.general.get(_zamky_key(pobocka, rok)) or []

def _uloz_zamky(pobocka: str, rok: int, zamcene: list[int]):
    app.storage.general[_zamky_key(pobocka, rok)] = sorted(zamcene)

def _zamknout_data(rows: list[dict], zamcene: list[int]) -> list[dict]:
    """Vynuluje zamčené měsíce v datech (pro ne-účetní uživatele).
    Skryje i komentáře u zamčených (nepublikovaných) měsíců."""
    if not zamcene:
        return rows
    zamcene_db = [MESICE_DB[m - 1] for m in zamcene if 1 <= m <= 12]
    out = []
    for r in rows:
        r2 = dict(r)
        kom = dict(r2.get('_komentare') or {})   # kopie, ať nemutujeme originál
        for db in zamcene_db:
            r2[db] = 0
            kom.pop(db, None)
        r2['_komentare'] = kom
        r2['celkem'] = sum(_s(r2[m]) for m in MESICE_DB)
        out.append(r2)
    return out


# ─── Tabulka nákladů – souhrn z datového souboru ─────────────────────────────

# Mapování názvů poboček z Excel souboru na interní klíče
POBOCKY_EXCEL_MAPPING = {
    '010 - Pardubice':        'Pardubice',
    '011 - Praha':            'Praha',
    '012 - Jilemnice':        'Jilemnice',
    '013 - Most':             'Most',
    '014 - Liberec':          'Liberec',
    '017 - Horšovský Týn':    'Horšovský Týn',
    '019 - Hodonín':          'Hodonín',
    '020 - Zlín':             'Zlín',
    '026 - Ostrava':          'Ostrava',
    '027-B.Olomouc':          'B.Olomouc',
    '028 - Brno - WINE LIFE': 'Winelife',
    '031-B.Jeseník':          'B.Jeseník',
    '032 - Olomouc':          'Olomouc',
    '033 - České Budějovice': 'České Budějovice',
    '034 - Plzeň':            'Plzeň',
    '037 - Nová Role':        'Nová Role',
    '089 - Přefakturace':     'Přefakturace',
    '089 - Přefakturace JIP': 'Přefakturace',
    '015 - Bourárna':         'Bourárna',
    '090 - Kamiony':          'Kamiony',
}

# Obrácené mapování: interní název pobočky → popis z Excelu (např. 'Praha' → '011 - Praha')
_POBOCKY_EXCEL_REVERSE = {v: k for k, v in POBOCKY_EXCEL_MAPPING.items()}
_POBOCKY_EXCEL_REVERSE['Olomouc'] = '032 - Olomouc VO'
_POBOCKY_EXCEL_REVERSE['Winelife'] = '028 - WINE LIFE'
_POBOCKY_EXCEL_REVERSE['Ostrava'] = '026 - Ostrava VO'

# Normalizované mapování „Pobočka popis" → pobočka (malá písmena, sloučené mezery),
# aby import snesl drobné rozdíly v zápisu (velikost písmen, vícenásobné mezery).
_POBOCKY_EXCEL_NORM = {
    ' '.join(str(k).split()).lower(): v for k, v in POBOCKY_EXCEL_MAPPING.items()
}

# Mapování čísla střediska (prefix názvu listu) → interní klíč pobočky
_STREDISKA_SHEET_MAPPING = {
    '010': 'Pardubice',
    '011': 'Praha',
    '012': 'Jilemnice',
    '013': 'Most',
    '014': 'Liberec',
    '017': 'Horšovský Týn',
    '019': 'Hodonín',
    '020': 'Zlín',
    '026': 'Ostrava',
    '027': 'B.Olomouc',
    '028': 'Winelife',
    '030': 'Brno',  # Brno – vnořené středisko pod Hodonínem (vlastní datový klíč)
    '031': 'B.Jeseník',
    '032': 'Olomouc',
    '033': 'České Budějovice',
    '034': 'Plzeň',
    '037': 'Nová Role',
    '089': 'Přefakturace',
    '326': '326-B.Ostrava CC',
    '077': '077-Benstar rest.',
    '082': '082-B.Gastro OVA',
    '097': '097-Restaur.Pha',
    '081': '081-Gastrostud.Pha',
}
_STREDISKA_NAME_FALLBACK = {
    'kamiony': 'Kamiony',
    'bourárna': 'Bourárna',
    'bourarna': 'Bourárna',
    'praha-becher': 'Praha-Becher',
    'supervizor': 'Supervizor',
    'údržba': 'Údržba',
    'udrzba': 'Údržba',
}
_PREDPIS_RE = re.compile(r'^(\d[\dx]*\.[\dx]*)\s+(.*\S)')

# (db_sloupec, zobrazený_název, index_v_excelu, šířka_px)
# Názvy sloupců přesně dle souboru „Tabulka nákladů_Vzor.xlsx".
_SOUHRN_COLS = [
    ('obrat_nc',            'Obrat v NC',                 5,  130),
    ('zisk_vydajove',       'Zisk výdajové doklady',      7,  180),
    ('zisk_prijmove',       'Zisk příjmové doklady',      11, 180),
    ('naklad_odvody',       'Náklad odvody sítě',         27, 160),
    ('kompenzace',          'Kompenzace',                 44, 120),
    ('zr',                  'Zr',                         19,  80),
    ('odberatelske_bonusy', 'Odběratelské bonusy',        25, 165),
    ('sankce',              'Sankce',                     23, 100),
    ('vysledek',            'Výsledek',                   54, 125),
    ('mzdy',                'Mzdy',                       53, 115),
    # po mzdy se vloží vypočítaný sloupec "Podíl mzdy v %"
    ('inventura',           'Inventura',                  14, 115),
    ('inventura_zaplaceno', 'Inventura z toho zaplaceno', 17, 200),
    ('likvidace',           'Likvidace',                  15, 110),
    ('nahrady',             'Náhrady',                    16, 105),
    ('repre_spotreba',      'Repre + spotřeba',           18, 145),
    ('naklady',             'Náklady',                    52, 110),
    ('total',               'Total',                      55, 115),
    # po total se vloží vypočítaný sloupec "%"
    ('zasoby',              'Zásoby',                     26, 130),
]

# Kg sloupce – ukládají se do DB pro grafy „Přehled pro vedení", ale NEzobrazují
# se v tabulce nákladů (proto nejsou v _SOUHRN_COLS). (db_sloupec, index_v_excelu)
_KG_COLS = [
    ('obrat_kg_celkem',      56),
    ('obrat_kg',             57),
    ('obrat_kg_dcery',       58),
    ('obrat_kg_site',        59),
    ('obrat_kg_staropramen', 60),
    ('obrat_kg_ostatni',     61),
    ('operace',              63),
    ('operace_dcery',        64),
    ('operace_site',         65),
]
# Kompletní sada importovaných sloupců (zobrazované + Kg): (db_sloupec, index_v_excelu)
_IMPORT_COLS = [(c, idx) for c, _, idx, _ in _SOUHRN_COLS] + _KG_COLS

_PCT_FMT = (
    "function(p){"
    "if(p.value===null||p.value===undefined||isNaN(parseFloat(p.value)))return '—';"
    "return (parseFloat(p.value)*100).toFixed(2)+' %';"
    "}"
)
# Formátování pro Tabulku nákladů – přesně dle Vzoru:
#   hodnoty „#,##0" (bez desetinných míst), procenta „0,00%" (cs-CZ).
_SOUHRN_FMT_NUM = (
    "function(p){"
    "if(p.value===null||p.value===undefined||p.value==='')return '—';"
    "var n=parseFloat(p.value);"
    "if(isNaN(n)||n===0)return '—';"
    "return new Intl.NumberFormat('cs-CZ',{maximumFractionDigits:0}).format(n);"
    "}"
)
_SOUHRN_FMT_PCT = (
    "function(p){"
    "if(p.value===null||p.value===undefined||p.value===''||isNaN(parseFloat(p.value)))return '—';"
    "var n=parseFloat(p.value);if(n===0)return '—';"
    "return new Intl.NumberFormat('cs-CZ',{minimumFractionDigits:2,maximumFractionDigits:2}).format(n*100)+'%';"
    "}"
)
_SOUHRN_CELKEM_STYLE = (
    "function(p){"
    "return (p.data&&p.data._je_celkem)?"
    "{fontWeight:'bold',backgroundColor:'#f1f5f9'}:{};"
    "}"
)
_SOUHRN_BOLD_CELKEM_STYLE = (
    "function(p){"
    "var s=(p.data&&p.data._je_celkem)?"
    "{fontWeight:'bold',backgroundColor:'#f1f5f9'}:{fontWeight:'bold'};"
    "return s;"
    "}"
)

# ── Barvy a styly přesně dle „Tabulka nákladů_Vzor.xlsx" ──────────────────────
# Motivové barvy s tintem 0.8: accent1→modrá, accent3→zelená, accent6→broskvová.
_TN_BLUE  = '#DCE6F2'   # hlavička + součtové řádky „Celkem"
_TN_GREEN = '#EBF1DE'   # sloupec „Výsledek"
_TN_PEACH = '#FDEADA'   # sloupce „Total" a „%"

# cellStyle pro datové buňky (součtový řádek „Celkem" = modrý tučný).
_TN_CS_NORMAL = (
    "function(p){"
    "if(p.data&&p.data._je_celkem)return{fontWeight:'bold',backgroundColor:'#DCE6F2'};"
    "return null;}"
)
# Sloupec „Rok": rok tučně (zobrazí se jen na 1. řádku), „Celkem" modrý tučný.
_TN_CS_ROK = (
    "function(p){"
    "if(p.data&&p.data._je_celkem)return{fontWeight:'bold',backgroundColor:'#DCE6F2'};"
    "return p.value?{fontWeight:'bold'}:null;}"
)
# Sloupec „Výsledek": zelený i v součtovém řádku.
_TN_CS_GREEN = (
    "function(p){"
    "var s={backgroundColor:'#EBF1DE'};"
    "if(p.data&&p.data._je_celkem)s.fontWeight='bold';"
    "return s;}"
)
# Sloupce „Total"/„%": broskvové; v součtovém řádku modré (dle Vzoru).
_TN_CS_PEACH = (
    "function(p){"
    "if(p.data&&p.data._je_celkem)return{fontWeight:'bold',backgroundColor:'#DCE6F2'};"
    "return{backgroundColor:'#FDEADA'};}"
)
# CSS pro hlavičku (barvy + tučně + font Arial) – aplikuje se na grid s třídou .tn-grid.
_TN_CSS = (
    ".tn-grid .ag-header-cell{background-color:#DCE6F2!important;}"
    ".tn-grid .ag-header-cell .ag-header-cell-text{font-weight:700;}"
    ".tn-grid .ag-header-cell.tn-h-green{background-color:#EBF1DE!important;}"
    ".tn-grid .ag-header-cell.tn-h-peach{background-color:#FDEADA!important;}"
    ".tn-grid .ag-cell,.tn-grid .ag-header-cell{font-family:Arial,Helvetica,sans-serif;}"
    ".tn-grid{height:auto!important;}"   # domLayout autoHeight – tabulka roste dle počtu měsíců
    # hlavička drží nahoře i při scrollu stránky (autoHeight nemá vnitřní scroll)
    ".tn-grid .ag-header{position:sticky;top:0;z-index:5;}"
    ".tn-grid .ag-root-wrapper{overflow:visible!important;}"
    ".tn-grid .ag-root{overflow:visible!important;}"
)


# ─── DB funkce ────────────────────────────────────────────────────────────────

def inicializace_vysledky_db():
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS vysledky_naklady (
                id              INT AUTO_INCREMENT PRIMARY KEY,
                pobocka         VARCHAR(100) NOT NULL,
                rok             SMALLINT     NOT NULL,
                poradi          INT          NOT NULL DEFAULT 0,
                ucetni_predpis  VARCHAR(100) NOT NULL DEFAULT '',
                nazev_predpisu  TEXT         NOT NULL,
                leden           DECIMAL(15,2) NOT NULL DEFAULT 0,
                unor            DECIMAL(15,2) NOT NULL DEFAULT 0,
                brezen          DECIMAL(15,2) NOT NULL DEFAULT 0,
                duben           DECIMAL(15,2) NOT NULL DEFAULT 0,
                kveten          DECIMAL(15,2) NOT NULL DEFAULT 0,
                cerven          DECIMAL(15,2) NOT NULL DEFAULT 0,
                cervenec        DECIMAL(15,2) NOT NULL DEFAULT 0,
                srpen           DECIMAL(15,2) NOT NULL DEFAULT 0,
                zari            DECIMAL(15,2) NOT NULL DEFAULT 0,
                rijen           DECIMAL(15,2) NOT NULL DEFAULT 0,
                listopad        DECIMAL(15,2) NOT NULL DEFAULT 0,
                prosinec        DECIMAL(15,2) NOT NULL DEFAULT 0,
                INDEX idx_pb_rok     (pobocka, rok),
                INDEX idx_pb_rok_por (pobocka, rok, poradi)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS tabulka_nakladu_mesice (
                pobocka              VARCHAR(100)  NOT NULL,
                rok                  SMALLINT      NOT NULL,
                mesic                TINYINT       NOT NULL,
                obrat_nc             DECIMAL(18,2) NOT NULL DEFAULT 0,
                zisk_vydajove        DECIMAL(18,2) NOT NULL DEFAULT 0,
                zisk_prijmove        DECIMAL(18,2) NOT NULL DEFAULT 0,
                naklad_odvody        DECIMAL(18,2) NOT NULL DEFAULT 0,
                kompenzace           DECIMAL(18,2) NOT NULL DEFAULT 0,
                zr                   DECIMAL(18,2) NOT NULL DEFAULT 0,
                odberatelske_bonusy  DECIMAL(18,2) NOT NULL DEFAULT 0,
                sankce               DECIMAL(18,2) NOT NULL DEFAULT 0,
                vysledek             DECIMAL(18,2) NOT NULL DEFAULT 0,
                mzdy                 DECIMAL(18,2) NOT NULL DEFAULT 0,
                inventura            DECIMAL(18,2) NOT NULL DEFAULT 0,
                inventura_zaplaceno  DECIMAL(18,2) NOT NULL DEFAULT 0,
                likvidace            DECIMAL(18,2) NOT NULL DEFAULT 0,
                nahrady              DECIMAL(18,2) NOT NULL DEFAULT 0,
                repre_spotreba       DECIMAL(18,2) NOT NULL DEFAULT 0,
                naklady              DECIMAL(18,2) NOT NULL DEFAULT 0,
                total                DECIMAL(18,2) NOT NULL DEFAULT 0,
                zasoby               DECIMAL(18,2) NOT NULL DEFAULT 0,
                obrat_kg_celkem      DECIMAL(18,2) NOT NULL DEFAULT 0,
                obrat_kg             DECIMAL(18,2) NOT NULL DEFAULT 0,
                obrat_kg_dcery       DECIMAL(18,2) NOT NULL DEFAULT 0,
                obrat_kg_site        DECIMAL(18,2) NOT NULL DEFAULT 0,
                obrat_kg_staropramen DECIMAL(18,2) NOT NULL DEFAULT 0,
                obrat_kg_ostatni     DECIMAL(18,2) NOT NULL DEFAULT 0,
                operace              DECIMAL(18,2) NOT NULL DEFAULT 0,
                operace_dcery        DECIMAL(18,2) NOT NULL DEFAULT 0,
                operace_site         DECIMAL(18,2) NOT NULL DEFAULT 0,
                PRIMARY KEY (pobocka, rok, mesic),
                INDEX idx_tnm_pb (pobocka)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS vysledky_obraty_zisk (
                pobocka        VARCHAR(100)  NOT NULL,
                list           VARCHAR(20)   NOT NULL,
                sekce_poradi   INT           NOT NULL DEFAULT 0,
                sekce_nazev    VARCHAR(160)  NOT NULL DEFAULT '',
                je_procento    TINYINT       NOT NULL DEFAULT 0,
                metrika_klic   VARCHAR(20)   NOT NULL,
                metrika_poradi INT           NOT NULL DEFAULT 0,
                metrika_nazev  VARCHAR(200)  NOT NULL DEFAULT '',
                rok            SMALLINT      NOT NULL,
                mesic          TINYINT       NOT NULL,
                hodnota        DECIMAL(22,4) NOT NULL DEFAULT 0,
                PRIMARY KEY (pobocka, list, sekce_poradi, metrika_klic, rok, mesic),
                INDEX idx_oz_pb_list (pobocka, list)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS vysledky_oz_data (
                pobocka  VARCHAR(100)  NOT NULL,
                list     VARCHAR(10)   NOT NULL,   -- 'obraty' | 'zisk'
                skupina  VARCHAR(20)   NOT NULL,   -- 'mesice' | 'dcery' | 'site' | 'zalozni_sklad'
                sekce    VARCHAR(10)   NOT NULL,   -- 'kc' | 'kg'
                rok      SMALLINT      NOT NULL,
                mesic    TINYINT       NOT NULL,   -- 1-12 = měsíce, 0 = Celkový součet
                hodnota  DECIMAL(22,4) NOT NULL DEFAULT 0,
                PRIMARY KEY (pobocka, list, skupina, sekce, rok, mesic),
                INDEX idx_ozd_pb (pobocka, list)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS vysledky_naklady_historie (
                id            BIGINT AUTO_INCREMENT PRIMARY KEY,
                radek_id      INT          NOT NULL,
                pobocka       VARCHAR(100) NOT NULL,
                rok           SMALLINT     NOT NULL,
                pole          VARCHAR(40)  NOT NULL,
                stara_hodnota TEXT,
                nova_hodnota  TEXT,
                uzivatel      VARCHAR(150) NOT NULL DEFAULT '',
                cas           DATETIME     NOT NULL DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_hist_radek  (radek_id),
                INDEX idx_hist_pb_rok (pobocka, rok)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS vysledky_mesic_status (
                rok      SMALLINT     NOT NULL,
                mesic    TINYINT      NOT NULL,
                oddeleni VARCHAR(20)  NOT NULL,   -- 'ao' | 'ucetni'
                hotovo   TINYINT      NOT NULL DEFAULT 0,
                kdo      VARCHAR(150) NOT NULL DEFAULT '',
                cas      DATETIME     NOT NULL DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (rok, mesic, oddeleni)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        # Komentáře k jednotlivým buňkám podrobných nákladů (jako v Excelu).
        # 1 komentář na buňku = (radek_id, pole); pole = měsíc / 'ucetni_predpis' / 'nazev_predpisu'.
        cur.execute("""
            CREATE TABLE IF NOT EXISTS vysledky_naklady_komentare (
                id        INT AUTO_INCREMENT PRIMARY KEY,
                radek_id  INT          NOT NULL,
                pole      VARCHAR(40)  NOT NULL,
                komentar  TEXT         NOT NULL,
                uzivatel  VARCHAR(150) NOT NULL DEFAULT '',
                cas       DATETIME     NOT NULL DEFAULT CURRENT_TIMESTAMP,
                UNIQUE KEY uniq_radek_pole (radek_id, pole),
                INDEX idx_kom_radek (radek_id)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        # Komentáře / diskuze k celé pobočce (chat). Více zpráv na pobočku,
        # každá nese kdo (uzivatel + user_id) a kdy (cas). Vidí ji každý,
        # kdo má přístup do detailu dané pobočky.
        cur.execute("""
            CREATE TABLE IF NOT EXISTS vysledky_pobocka_komentare (
                id        INT AUTO_INCREMENT PRIMARY KEY,
                pobocka   VARCHAR(100) NOT NULL,
                komentar  TEXT         NOT NULL,
                uzivatel  VARCHAR(150) NOT NULL DEFAULT '',
                user_id   INT          NOT NULL DEFAULT 0,
                cas       DATETIME     NOT NULL DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_pk_pobocka (pobocka, cas)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        # Migrace: doplnění Kg sloupců do existující tabulky (grafy pro vedení).
        cur.execute('SHOW COLUMNS FROM tabulka_nakladu_mesice')
        _existing = {r[0] for r in cur.fetchall()}
        for _kgc in ('obrat_kg_celkem', 'obrat_kg', 'obrat_kg_dcery',
                     'obrat_kg_site', 'obrat_kg_staropramen',
                     'obrat_kg_ostatni', 'operace', 'operace_dcery',
                     'operace_site'):
            if _kgc not in _existing:
                cur.execute(f'ALTER TABLE tabulka_nakladu_mesice '
                            f'ADD COLUMN {_kgc} DECIMAL(18,2) NOT NULL DEFAULT 0')
        # Migrace: skupina VARCHAR(10)→VARCHAR(20) kvůli 'zalozni_sklad'
        # (užší sloupec hodnotu tiše ořezával na 'zalozni_sk') + oprava
        # už oříznutých řádků.
        cur.execute("SHOW COLUMNS FROM vysledky_oz_data LIKE 'skupina'")
        _sk = cur.fetchone()
        if _sk and 'varchar(10)' in str(_sk[1]).lower():
            cur.execute('ALTER TABLE vysledky_oz_data MODIFY skupina VARCHAR(20) NOT NULL')
            cur.execute("UPDATE vysledky_oz_data SET skupina='zalozni_sklad' "
                        "WHERE skupina='zalozni_sk'")
        conn.commit()
    except Exception as exc:
        print(f'[vysledky] DB init error: {exc}')
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()


def _s(v) -> float:
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def nacti_naklady(pobocka: str, rok: int) -> list[dict]:
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(
            'SELECT * FROM vysledky_naklady WHERE pobocka=%s AND rok=%s ORDER BY poradi',
            (pobocka, rok),
        )
        rows = cur.fetchall()
        for r in rows:
            r['celkem'] = sum(_s(r[m]) for m in MESICE_DB)
            r['_komentare'] = {}
        # Komentáře k buňkám (jako v Excelu) – samostatně, aby chybějící tabulka
        # (starší DB) nezhatila načtení samotných nákladů.
        try:
            ids = [r['id'] for r in rows if r.get('id')]
            if ids:
                ph = ','.join(['%s'] * len(ids))
                cur.execute(
                    f'SELECT radek_id, pole, komentar FROM vysledky_naklady_komentare '
                    f'WHERE radek_id IN ({ph})', ids)
                kmap: dict = {}
                for kr in cur.fetchall():
                    kmap.setdefault(kr['radek_id'], {})[kr['pole']] = kr['komentar']
                for r in rows:
                    r['_komentare'] = kmap.get(r['id'], {})
        except Exception as exc:
            print(f'[vysledky] nacti_naklady komentare error: {exc}')
        return rows
    except Exception as exc:
        print(f'[vysledky] nacti_naklady error: {exc}')
        return []
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()


def inicializuj_osnovu(pobocka: str, rok: int):
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.execute(
            'SELECT COUNT(*) FROM vysledky_naklady WHERE pobocka=%s AND rok=%s',
            (pobocka, rok),
        )
        if cur.fetchone()[0] > 0:
            return
        for i, (predpis, nazev) in enumerate(VYCHOZI_OSNOVA):
            cur.execute(
                'INSERT INTO vysledky_naklady (pobocka, rok, poradi, ucetni_predpis, nazev_predpisu) VALUES (%s,%s,%s,%s,%s)',
                (pobocka, rok, i, predpis, nazev),
            )
        conn.commit()
    except Exception as exc:
        print(f'[vysledky] inicializuj_osnovu error: {exc}')
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()


def uloz_hodnotu(row_id: int, mesic: str, hodnota: float):
    if mesic not in MESICE_DB:
        return
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.execute(f'UPDATE vysledky_naklady SET {mesic}=%s WHERE id=%s', (hodnota, row_id))
        conn.commit()
    except Exception as exc:
        print(f'[vysledky] uloz_hodnotu error: {exc}')
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()


def uloz_predpis(row_id: int, ucetni_predpis: str, nazev_predpisu: str):
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.execute(
            'UPDATE vysledky_naklady SET ucetni_predpis=%s, nazev_predpisu=%s WHERE id=%s',
            (ucetni_predpis or '', nazev_predpisu or '', row_id),
        )
        conn.commit()
    except Exception as exc:
        print(f'[vysledky] uloz_predpis error: {exc}')
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()


def zaznam_historie(radek_id: int, pobocka: str, rok: int, pole: str,
                    stara, nova, uzivatel: str):
    """Zapíše jednu změnu řádku nákladů do historie (kdo, co, z čeho na co)."""
    if not radek_id:
        return
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.execute(
            'INSERT INTO vysledky_naklady_historie '
            '(radek_id, pobocka, rok, pole, stara_hodnota, nova_hodnota, uzivatel) '
            'VALUES (%s,%s,%s,%s,%s,%s,%s)',
            (radek_id, pobocka, int(rok), pole,
             None if stara is None else str(stara),
             None if nova is None else str(nova),
             uzivatel or ''),
        )
        conn.commit()
    except Exception as exc:
        print(f'[vysledky] zaznam_historie error: {exc}')
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()


def nacti_historii(radek_id: int) -> list[dict]:
    """Vrátí historii úprav řádku (nejnovější první)."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(
            'SELECT pole, stara_hodnota, nova_hodnota, uzivatel, cas '
            'FROM vysledky_naklady_historie WHERE radek_id=%s '
            'ORDER BY cas DESC, id DESC',
            (radek_id,),
        )
        return cur.fetchall()
    except Exception as exc:
        print(f'[vysledky] nacti_historii error: {exc}')
        return []
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()


def nacti_komentar(radek_id: int, pole: str) -> dict | None:
    """Vrátí jeden komentář k buňce (text, uzivatel, cas) nebo None."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return None
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(
            'SELECT komentar, uzivatel, cas FROM vysledky_naklady_komentare '
            'WHERE radek_id=%s AND pole=%s',
            (int(radek_id), pole),
        )
        return cur.fetchone()
    except Exception as exc:
        print(f'[vysledky] nacti_komentar error: {exc}')
        return None
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()


def uloz_komentar(radek_id: int, pole: str, text: str, uzivatel: str):
    """Vloží/aktualizuje komentář k buňce (1 komentář na buňku)."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.execute(
            'INSERT INTO vysledky_naklady_komentare (radek_id, pole, komentar, uzivatel, cas) '
            'VALUES (%s,%s,%s,%s,NOW()) '
            'ON DUPLICATE KEY UPDATE komentar=VALUES(komentar), uzivatel=VALUES(uzivatel), cas=NOW()',
            (int(radek_id), pole, text or '', uzivatel or ''),
        )
        conn.commit()
    except Exception as exc:
        print(f'[vysledky] uloz_komentar error: {exc}')
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()


def smaz_komentar(radek_id: int, pole: str):
    """Smaže komentář k buňce."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.execute(
            'DELETE FROM vysledky_naklady_komentare WHERE radek_id=%s AND pole=%s',
            (int(radek_id), pole),
        )
        conn.commit()
    except Exception as exc:
        print(f'[vysledky] smaz_komentar error: {exc}')
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()


def nacti_pobocka_komentare(pobocka: str) -> list[dict]:
    """Vrátí diskuzi (chat) k pobočce chronologicky – nejstarší první."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(
            'SELECT id, komentar, uzivatel, user_id, cas '
            'FROM vysledky_pobocka_komentare WHERE pobocka=%s '
            'ORDER BY cas ASC, id ASC',
            (pobocka,),
        )
        return cur.fetchall()
    except Exception as exc:
        print(f'[vysledky] nacti_pobocka_komentare error: {exc}')
        return []
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()


def pridej_pobocka_komentar(pobocka: str, text: str, uzivatel: str, user_id: int):
    """Přidá jeden komentář (zprávu) do diskuze pobočky."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.execute(
            'INSERT INTO vysledky_pobocka_komentare (pobocka, komentar, uzivatel, user_id, cas) '
            'VALUES (%s,%s,%s,%s,NOW())',
            (pobocka, text or '', uzivatel or '', int(user_id or 0)),
        )
        conn.commit()
    except Exception as exc:
        print(f'[vysledky] pridej_pobocka_komentar error: {exc}')
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()


def smaz_pobocka_komentar(komentar_id: int):
    """Smaže jeden komentář z diskuze pobočky."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.execute(
            'DELETE FROM vysledky_pobocka_komentare WHERE id=%s',
            (int(komentar_id),),
        )
        conn.commit()
    except Exception as exc:
        print(f'[vysledky] smaz_pobocka_komentar error: {exc}')
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()


def nacti_mesic_status(rok: int) -> dict:
    """Stav dokončení měsíců za rok pro obě oddělení.
    Vrací {(mesic, oddeleni): {'hotovo': bool, 'kdo': str, 'cas': datetime}}."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return {}
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(
            'SELECT mesic, oddeleni, hotovo, kdo, cas FROM vysledky_mesic_status WHERE rok=%s',
            (rok,),
        )
        out = {}
        for r in cur.fetchall():
            out[(int(r['mesic']), r['oddeleni'])] = {
                'hotovo': bool(r['hotovo']), 'kdo': r['kdo'], 'cas': r['cas'],
            }
        return out
    except Exception as exc:
        print(f'[vysledky] nacti_mesic_status error: {exc}')
        return {}
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()


def uloz_mesic_status(rok: int, mesic: int, oddeleni: str, hotovo: bool, kdo: str):
    """Nastaví stav dokončení měsíce pro oddělení ('ao' | 'ucetni')."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.execute(
            'INSERT INTO vysledky_mesic_status (rok, mesic, oddeleni, hotovo, kdo, cas) '
            'VALUES (%s,%s,%s,%s,%s,NOW()) '
            'ON DUPLICATE KEY UPDATE hotovo=VALUES(hotovo), kdo=VALUES(kdo), cas=NOW()',
            (int(rok), int(mesic), oddeleni, 1 if hotovo else 0, kdo or ''),
        )
        conn.commit()
    except Exception as exc:
        print(f'[vysledky] uloz_mesic_status error: {exc}')
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()


def pridej_radek_synced(pobocka: str, aktualni_rok: int, predpis: str = '', nazev: str = '') -> bool:
    """Přidá řádek do aktuálního roku a synchronně i do minulého roku (pokud chybí)."""
    minuly_rok = aktualni_rok - 1
    conn = intranet_data.get_db_connection()
    if not conn:
        return False
    try:
        cur = conn.cursor()
        # Nový pořadí pro aktuální rok
        cur.execute(
            'SELECT COALESCE(MAX(poradi),-1) FROM vysledky_naklady WHERE pobocka=%s AND rok=%s',
            (pobocka, aktualni_rok),
        )
        novy = cur.fetchone()[0] + 1
        cur.execute(
            'INSERT INTO vysledky_naklady (pobocka,rok,poradi,ucetni_predpis,nazev_predpisu) VALUES (%s,%s,%s,%s,%s)',
            (pobocka, aktualni_rok, novy, predpis, nazev),
        )
        # Přidej odpovídající nulový řádek do minulého roku (synchronizace porovnání)
        cur.execute(
            'SELECT COALESCE(MAX(poradi),-1) FROM vysledky_naklady WHERE pobocka=%s AND rok=%s',
            (pobocka, minuly_rok),
        )
        max_min = cur.fetchone()[0]
        if max_min < novy:
            cur.execute(
                'INSERT INTO vysledky_naklady (pobocka,rok,poradi,ucetni_predpis,nazev_predpisu) VALUES (%s,%s,%s,%s,%s)',
                (pobocka, minuly_rok, novy, predpis, nazev),
            )
        conn.commit()
        return True
    except Exception as exc:
        print(f'[vysledky] pridej_radek_synced error: {exc}')
        return False
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()


def smaz_radek_synced(row_id: int, pobocka: str, aktualni_rok: int):
    """Smaže řádek z aktuálního roku a odpovídající řádek z minulého roku (dle pořadí)."""
    minuly_rok = aktualni_rok - 1
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.execute(
            'SELECT poradi FROM vysledky_naklady WHERE id=%s AND pobocka=%s AND rok=%s',
            (row_id, pobocka, aktualni_rok),
        )
        result = cur.fetchone()
        if not result:
            return
        poradi = result[0]
        # ID odpovídajícího řádku v minulém roce (kvůli úklidu jeho komentářů)
        cur.execute(
            'SELECT id FROM vysledky_naklady WHERE pobocka=%s AND rok=%s AND poradi=%s',
            (pobocka, minuly_rok, poradi),
        )
        _mr = cur.fetchone()
        minuly_id = _mr[0] if _mr else None
        # Smaž komentáře k buňkám mazaných řádků (aktuální i odpovídající minulý)
        cur.execute('DELETE FROM vysledky_naklady_komentare WHERE radek_id=%s', (row_id,))
        if minuly_id:
            cur.execute('DELETE FROM vysledky_naklady_komentare WHERE radek_id=%s', (minuly_id,))
        # Smaž z aktuálního roku a přečísluj
        cur.execute('DELETE FROM vysledky_naklady WHERE id=%s', (row_id,))
        cur.execute(
            'UPDATE vysledky_naklady SET poradi=poradi-1 WHERE pobocka=%s AND rok=%s AND poradi>%s',
            (pobocka, aktualni_rok, poradi),
        )
        # Smaž odpovídající řádek z minulého roku a přečísluj
        cur.execute(
            'DELETE FROM vysledky_naklady WHERE pobocka=%s AND rok=%s AND poradi=%s',
            (pobocka, minuly_rok, poradi),
        )
        cur.execute(
            'UPDATE vysledky_naklady SET poradi=poradi-1 WHERE pobocka=%s AND rok=%s AND poradi>%s',
            (pobocka, minuly_rok, poradi),
        )
        conn.commit()
    except Exception as exc:
        print(f'[vysledky] smaz_radek_synced error: {exc}')
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()


def _slouc_naklady_vice_pobocek(klice: list[str], rok: int) -> dict:
    """Sečte měsíční náklady více poboček po (předpis, název) pro daný rok.

    Vrací mapu (předpis, název) → řádek se sečtenými měsíci a metadaty
    (pořadí dle prvního výskytu). Slouží k porovnání oblasti složené z více
    poboček (např. Ostrava + 027-B.Olomouc + 031-B.Jeseník).
    """
    out: dict = {}
    for klic in klice:
        for r in nacti_naklady(klic, rok):
            k = ((r.get('ucetni_predpis') or '').strip(),
                 (r.get('nazev_predpisu') or '').strip())
            cil = out.get(k)
            if cil is None:
                cil = {
                    'ucetni_predpis': r.get('ucetni_predpis', ''),
                    'nazev_predpisu': r.get('nazev_predpisu', ''),
                    '_poradi': r.get('poradi', 0),
                }
                for m in MESICE_DB:
                    cil[m] = 0.0
                out[k] = cil
            for m in MESICE_DB:
                cil[m] = _s(cil[m]) + _s(r.get(m, 0))
    return out


def nacti_porovnani(pobocka: str, aktualni_rok: int) -> list[dict]:
    """Vrátí kombinovaná data obou let seřazená dle pořadí pro záložku Porovnání.

    U poboček s vnořenými provozy (``_VNORENE_POBOCKY``) se náklady sčítají přes
    rodičovskou pobočku i všechny vnořené pobočky dohromady, aby porovnání
    odpovídalo celé oblasti, ne jen rodičovské pobočce
    (Ostrava → 027-B.Olomouc + 031-B.Jeseník, Hodonín → 030-Brno).
    """
    # Obě roky spojujeme VŽDY podle (účetní předpis, název) – NE podle pořadí
    # řádku. Historická data minulého roku mohou mít jinou osnovu (jiný počet
    # i pořadí řádků) než aktuální rok; spojení přes `poradi` pak hodnoty posune
    # (např. částka „Elektrická energie" 2025 se ukázala u „Zemní plyn"). Pro
    # vnořené pobočky se navíc sečtou náklady rodiče i všech vnořených provozů.
    klice = [pobocka] + [klic for _, klic in _VNORENE_POBOCKY.get(pobocka, [])]
    return _porovnani_z_klicu(klice, aktualni_rok)


def _porovnani_z_klicu(klice: list[str], aktualni_rok: int) -> list[dict]:
    """Sestaví meziroční porovnání podrobných nákladů sečtených přes ``klice``
    (datové klíče poboček). Spojuje aktuální a minulý rok podle (předpis, název)."""
    minuly_rok = aktualni_rok - 1
    akt  = _slouc_naklady_vice_pobocek(klice, aktualni_rok)
    prev = _slouc_naklady_vice_pobocek(klice, minuly_rok)
    all_keys = sorted(
        set(akt) | set(prev),
        key=lambda k: ((akt.get(k) or prev.get(k))['_poradi'], k[0], k[1]),
    )
    result = []
    for key in all_keys:
        a = akt.get(key, {})
        p = prev.get(key, {})
        row: dict = {
            'ucetni_predpis': a.get('ucetni_predpis') or p.get('ucetni_predpis', ''),
            'nazev_predpisu': a.get('nazev_predpisu') or p.get('nazev_predpisu', ''),
        }
        for m in MESICE_DB:
            av = _s(a.get(m, 0))
            pv = _s(p.get(m, 0))
            row[f'akt_{m}']  = av
            row[f'min_{m}']  = pv
            row[f'diff_{m}'] = av - pv
        row['akt_celkem']  = sum(row[f'akt_{m}']  for m in MESICE_DB)
        row['min_celkem']  = sum(row[f'min_{m}']  for m in MESICE_DB)
        row['diff_celkem'] = row['akt_celkem'] - row['min_celkem']
        result.append(row)
    return result


def _soucet_naklady(rows: list[dict]) -> dict:
    """Připravený součtový řádek (CELKEM) pro grid aktuálního/minulého roku."""
    total: dict = {'ucetni_predpis': '', 'nazev_predpisu': 'CELKEM', '_soucet': True}
    for m in MESICE_DB:
        total[m] = sum(_s(r.get(m)) for r in rows)
    total['celkem'] = sum(total[m] for m in MESICE_DB)
    return total


def _soucet_porovnani(rows: list[dict]) -> dict:
    """Připravený součtový řádek (CELKEM) pro grid porovnání let."""
    total: dict = {'ucetni_predpis': '', 'nazev_predpisu': 'CELKEM', '_soucet': True}
    for m in MESICE_DB:
        total[f'akt_{m}']  = sum(_s(r.get(f'akt_{m}'))  for r in rows)
        total[f'min_{m}']  = sum(_s(r.get(f'min_{m}'))  for r in rows)
        total[f'diff_{m}'] = sum(_s(r.get(f'diff_{m}')) for r in rows)
    total['akt_celkem']  = sum(_s(r.get('akt_celkem'))  for r in rows)
    total['min_celkem']  = sum(_s(r.get('min_celkem'))  for r in rows)
    total['diff_celkem'] = sum(_s(r.get('diff_celkem')) for r in rows)
    return total


def _xlsx_sloupce(coldefs: list[dict]) -> list[tuple]:
    """Grid columnDefs → sloupce pro XLSX: (nadpis, field, typ, šířka ve znacích).

    Skupiny (Leden → 2025/2024/Δ) se zploští na „Leden 2025"…, technické sloupce
    (`_historie`) vypadnou, zámkové emoji v hlavičce měsíce se ořízne.
    Sloupec „Celkem" v gridu nemá `field` (počítá ho valueGetter) → mapuje se
    na `celkem`, které do řádků doplní `_export`.
    """
    out: list[tuple] = []
    for cd in coldefs:
        deti = cd.get('children')
        for c in (deti or [cd]):
            field = c.get('field') or ('celkem' if c.get(':valueGetter') else None)
            if not field or field.startswith('_'):
                continue
            nadpis = str(c.get('headerName') or field).lstrip('🔒🔓 ')
            if deti:
                nadpis = f"{cd.get('headerName') or ''} {nadpis}".strip()
            typ = 'money' if c.get('type') == 'numericColumn' else 'text'
            sirka = max(10, round(int(c.get('width') or 110) / 7))   # px → znaky
            out.append((nadpis, field, typ, sirka))
    return out


def _pridej_filtr_nazvu(rows: list[dict], grid, soucet_fn, nazev_souboru: str,
                        s_exportem: bool = True):
    """Vytvoří u sloupce „Název" rozbalovací zaškrtávací filtr (jako v Excelu)
    a (volitelně) tlačítko pro export dat do XLSX.

    Po změně výběru přefiltruje řádky v gridu a přepočítá připnutý součtový
    řádek dole. Prázdný výběr = zobrazí všechny řádky.
    Export respektuje aktuálně vyfiltrovaná data zobrazená v gridu;
    ``s_exportem=False`` tlačítko exportu vynechá (např. souhrn pro majitele).

    Vrací funkci ``zobrazene()`` → aktuálně zobrazené (vyfiltrované) řádky,
    aby volající mohl po editaci přepočítat součet nad správnou množinou.
    """
    nazvy = sorted({
        (r.get('nazev_predpisu') or '').strip()
        for r in rows
        if (r.get('nazev_predpisu') or '').strip()
    })

    sel_ref: dict = {}

    def _zobrazene() -> list[dict]:
        vyb = sel_ref.get('sel').value if sel_ref.get('sel') else None
        if vyb:
            vyber = set(vyb)
            return [r for r in rows if (r.get('nazev_predpisu') or '').strip() in vyber]
        return rows

    def _aplikuj(vybrane):
        if vybrane:
            vyber = set(vybrane)
            filtr_rows = [r for r in rows if (r.get('nazev_predpisu') or '').strip() in vyber]
        else:
            filtr_rows = rows
        grid.options['rowData'] = filtr_rows
        grid.options['pinnedBottomRowData'] = [soucet_fn(filtr_rows)]
        grid.update()

    def _on_change(e):
        vyb = e.value or []
        # Kompaktní zobrazení místo chipsů, aby se pole nezvětšovalo
        if vyb:
            sel.props(f'display-value="{len(vyb)} vybráno"')
        else:
            sel.props(remove='display-value')
        _aplikuj(vyb)

    async def _export():
        # Proč ne CSV z ag-gridu: `exportDataAsCsv` prohnal hodnoty valueFormatterem,
        # takže do souboru šlo „1 234 567,50" s NBSP (U+00A0) jako oddělovačem tisíců.
        # Excel NBSP nerozpozná → buňka je TEXT, nejde sčítat (a čísla pod 1000 přitom
        # prošla, takže se to projevilo zákeřně). Stavíme rovnou .xlsx: čísla zůstávají
        # čísly, hlavička a šířky sloupců jsou vždy stejné → nic se nemusí upravovat.
        from intranet_sankce import _export_xlsx
        cols = _xlsx_sloupce(grid.options.get('columnDefs') or [])
        radky = _zobrazene()
        if any(f == 'celkem' for _n, f, _t, _w in cols):
            radky = [dict(r, celkem=sum(_s(r.get(m)) for m in MESICE_DB)) for r in radky]
        await _export_xlsx(cols, radky, soucet_fn(radky), nazev_souboru, 'Náklady')

    with ui.row().classes('items-center gap-2'):
        ui.label('Název:').classes('text-sm text-gray-600')
        sel = ui.select(
            nazvy, multiple=True, clearable=True,
            with_input=True,
        ).props('dense outlined options-dense behavior=menu') \
         .classes('w-64') \
         .tooltip('Filtr podle názvu předpisu (jako v Excelu)')
        sel_ref['sel'] = sel
        sel.on_value_change(_on_change)
        if s_exportem:
            ui.button(icon='download', text='Export XLSX', on_click=_export) \
              .props('color=secondary outline dense no-caps') \
              .tooltip('Export zobrazených dat do Excelu (respektuje filtr, '
                       'čísla jsou čísla – jdou rovnou sčítat)')

    return _zobrazene


# ─── Obraty / Zisk – parser per-pobočkového souboru (data GIST) ───────────────

def _oz_m_of(v):
    """Vrátí číslo měsíce 1–12, pokud buňka odpovídá měsíci, jinak None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        iv = int(v)
        return iv if (iv == v and 1 <= iv <= 12) else None
    s = str(v).strip()
    if s.isdigit():
        iv = int(s)
        return iv if 1 <= iv <= 12 else None
    return None


def _oz_num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace('\xa0', '').replace(' ', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def _oz_year(v):
    if isinstance(v, (int, float)):
        iv = int(v)
        if iv == v and 2000 <= iv <= 2100:
            return iv
    s = str(v).strip() if v is not None else ''
    if s.isdigit():
        iv = int(s)
        if 2000 <= iv <= 2100:
            return iv
    return None


def _importuj_souhrn_sync(zdroj, pobocka: str) -> tuple[int, str]:
    """Import měsíčního souhrnu nákladů z per-pobočkového souboru
    „Tabulka nákladů_DATA.xlsx" do tabulky tabulka_nakladu_mesice.

    `zdroj` = cesta nebo file-like (BytesIO). Volejte přes asyncio.to_thread –
    blokující I/O! Importují se pouze řádky patřící zadané pobočce; dříve nahraná
    data této pobočky se nahradí, aby zobrazení odpovídalo nahranému souboru.
    Vrátí (počet_uložených_řádků, chybová_zpráva_nebo_prázdný_řetězec)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(zdroj, data_only=True, read_only=True)
        ws = wb.active
    except Exception as exc:
        return 0, f'Nelze otevřít soubor: {exc}'

    # Pobočka může mít více dílčích jednotek (např. „011 - Praha" + „903 - Praha,
    # záložní sklad") sdílejících stejný popis pobočky. Hodnoty za měsíc proto
    # sčítáme přes všechny dílčí jednotky → jeden řádek na (rok, měsíc).
    nalezene_pobocky: set[str] = set()
    agg: dict[tuple[int, int], list[float]] = {}
    try:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            pb = POBOCKY_EXCEL_MAPPING.get(row[0])
            if pb is None and row[0] is not None:
                pb = _POBOCKY_EXCEL_NORM.get(' '.join(str(row[0]).split()).lower())
            if pb is None:
                continue
            nalezene_pobocky.add(pb)
            if pb != pobocka:
                continue
            rok, mesic = row[1], row[4]
            if not (rok and mesic):
                continue
            key = (int(rok), int(mesic))
            vals = [_s(row[idx]) for _, idx in _IMPORT_COLS]
            acc = agg.get(key)
            if acc is None:
                agg[key] = vals
            else:
                for i, v in enumerate(vals):
                    acc[i] += v
    except Exception as exc:
        return 0, f'Chyba čtení dat: {exc}'
    finally:
        try:
            wb.close()
        except Exception:
            pass

    radky: list[tuple] = [
        tuple([pobocka, rok, mesic, *vals])
        for (rok, mesic), vals in sorted(agg.items())
    ]

    if not radky:
        if nalezene_pobocky:
            return 0, (f'Soubor neobsahuje data pro pobočku „{pobocka}". '
                       f'Nalezena data poboček: {", ".join(sorted(nalezene_pobocky))}.')
        return 0, ('V souboru nebyly nalezeny žádné rozpoznatelné řádky '
                   '(zkontrolujte, že jde o soubor „Tabulka nákladů_DATA.xlsx").')

    conn = intranet_data.get_db_connection()
    if not conn:
        return 0, 'Nelze se připojit k databázi.'

    col_names    = ', '.join(c for c, _ in _IMPORT_COLS)
    placeholders = ', '.join(['%s'] * len(_IMPORT_COLS))
    updates      = ', '.join(f'{c}=VALUES({c})' for c, _ in _IMPORT_COLS)
    sql = (
        f"INSERT INTO tabulka_nakladu_mesice "
        f"(pobocka, rok, mesic, {col_names}) "
        f"VALUES (%s, %s, %s, {placeholders}) "
        f"ON DUPLICATE KEY UPDATE {updates}"
    )
    try:
        cur = conn.cursor()
        cur.execute('DELETE FROM tabulka_nakladu_mesice WHERE pobocka=%s', (pobocka,))
        cur.executemany(sql, radky)
        conn.commit()
        return len(radky), ''
    except Exception as exc:
        return 0, str(exc)
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()


def _importuj_souhrn_vse_sync(zdroj) -> tuple[int, list[str], str]:
    """Import dat pro VŠECHNY rozpoznané pobočky z jednoho souboru
    „Tabulka nákladů_DATA.xlsx". Řádky se rozdělí podle „Pobočka popis" a dílčí
    jednotky se sečtou. Data každé nalezené pobočky se nahradí.
    Vrací (počet_řádků, seznam_poboček, chybová_zpráva)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(zdroj, data_only=True, read_only=True)
        ws = wb.active
    except Exception as exc:
        return 0, [], f'Nelze otevřít soubor: {exc}'

    agg: dict[tuple[str, int, int], list[float]] = {}
    try:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            pb = POBOCKY_EXCEL_MAPPING.get(row[0])
            if pb is None and row[0] is not None:
                pb = _POBOCKY_EXCEL_NORM.get(' '.join(str(row[0]).split()).lower())
            if pb is None:
                continue
            rok, mesic = row[1], row[4]
            if not (rok and mesic):
                continue
            key = (pb, int(rok), int(mesic))
            vals = [_s(row[idx]) for _, idx in _IMPORT_COLS]
            acc = agg.get(key)
            if acc is None:
                agg[key] = vals
            else:
                for i, v in enumerate(vals):
                    acc[i] += v
    except Exception as exc:
        return 0, [], f'Chyba čtení dat: {exc}'
    finally:
        try:
            wb.close()
        except Exception:
            pass

    if not agg:
        return 0, [], ('V souboru nebyly nalezeny žádné rozpoznatelné pobočky '
                      '(zkontrolujte, že jde o „Tabulka nákladů_DATA.xlsx").')

    pobocky = sorted({k[0] for k in agg})
    radky = [tuple([pb, rok, mes, *vals]) for (pb, rok, mes), vals in sorted(agg.items())]

    conn = intranet_data.get_db_connection()
    if not conn:
        return 0, [], 'Nelze se připojit k databázi.'
    col_names    = ', '.join(c for c, _ in _IMPORT_COLS)
    placeholders = ', '.join(['%s'] * len(_IMPORT_COLS))
    updates      = ', '.join(f'{c}=VALUES({c})' for c, _ in _IMPORT_COLS)
    sql = (
        f"INSERT INTO tabulka_nakladu_mesice (pobocka, rok, mesic, {col_names}) "
        f"VALUES (%s, %s, %s, {placeholders}) ON DUPLICATE KEY UPDATE {updates}"
    )
    try:
        cur = conn.cursor()
        fmt = ', '.join(['%s'] * len(pobocky))
        cur.execute(f'DELETE FROM tabulka_nakladu_mesice WHERE pobocka IN ({fmt})', pobocky)
        cur.executemany(sql, radky)
        conn.commit()
        return len(radky), pobocky, ''
    except Exception as exc:
        return 0, [], str(exc)
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()


def _strediska_najdi_pobocku(sheet_name: str) -> str | None:
    m = re.match(r'^(\d{2,3})', sheet_name.strip())
    if m:
        return _STREDISKA_SHEET_MAPPING.get(m.group(1))
    # porovnání bez diakritiky – list „Kamióny" se shodne s „kamiony" apod.
    sn = _bez_diakritiky(sheet_name)
    for klic, val in _STREDISKA_NAME_FALLBACK.items():
        if _bez_diakritiky(klic) in sn:
            return val
    return None


def _strediska_parsuj_excel_sync(zdroj) -> tuple[list[dict], list[str], int, str]:
    """Naparsuje Excel soubor středisek.
    Vrací (matched_list, unmatched_names, rok, chyba)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(zdroj, data_only=True, read_only=True)
    except Exception as exc:
        return [], [], 0, f'Nelze otevřít soubor: {exc}'

    matched: dict[str, dict] = {}
    unmatched: list[str] = []
    detected_year = None

    try:
        for ws in wb.worksheets:
            sn = ws.title or ''
            pobocka = _strediska_najdi_pobocku(sn)
            if pobocka is None:
                unmatched.append(sn)
                continue
            ym = re.search(r'(20\d{2})', sn)
            if ym and detected_year is None:
                detected_year = int(ym.group(1))

            rows: list[dict] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue
                if not row or row[0] is None:
                    continue
                text = str(row[0]).strip()
                if not text:
                    continue
                pm = _PREDPIS_RE.match(text)
                predpis, nazev = (pm.group(1), pm.group(2).strip()) if pm else ('', text)
                mesice_vals = {}
                for mi, db in enumerate(MESICE_DB):
                    v = row[mi + 1] if len(row) > mi + 1 else None
                    mesice_vals[db] = _s(v) if v is not None else 0.0
                rows.append({'ucetni_predpis': predpis, 'nazev_predpisu': nazev, **mesice_vals})

            if rows:
                if pobocka in matched:
                    matched[pobocka]['rows'].extend(rows)
                    matched[pobocka]['sheet_names'].append(sn)
                else:
                    matched[pobocka] = {'pobocka': pobocka, 'sheet_names': [sn], 'rows': rows}
    except Exception as exc:
        return [], [], 0, f'Chyba čtení dat: {exc}'
    finally:
        try:
            wb.close()
        except Exception:
            pass

    if detected_year is None:
        detected_year = datetime.datetime.now().year - 1
    return list(matched.values()), unmatched, detected_year, ''


def _strediska_importuj_sync(matched: list[dict], rok: int) -> tuple[int, int, str]:
    """Importuje naparsovaná data středisek do vysledky_naklady.
    Vrací (počet_řádků, počet_poboček, chyba)."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return 0, 0, 'Nelze se připojit k databázi.'
    total = 0
    try:
        cur = conn.cursor()
        for entry in matched:
            pb = entry['pobocka']
            cur.execute(
                'DELETE FROM vysledky_naklady WHERE pobocka=%s AND rok=%s',
                (pb, rok),
            )
            for poradi, r in enumerate(entry['rows']):
                vals = [_s(r.get(m, 0)) for m in MESICE_DB]
                cur.execute(
                    'INSERT INTO vysledky_naklady '
                    '(pobocka, rok, poradi, ucetni_predpis, nazev_predpisu, '
                    'leden, unor, brezen, duben, kveten, cerven, '
                    'cervenec, srpen, zari, rijen, listopad, prosinec) '
                    'VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)',
                    (pb, rok, poradi, r['ucetni_predpis'], r['nazev_predpisu'], *vals),
                )
                total += 1
        conn.commit()
        return total, len(matched), ''
    except Exception as exc:
        return 0, 0, str(exc)
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()


def _nacti_prijemce_aktualizace() -> list[dict]:
    """Najde uživatele s právem ČTENÍ Výsledků poboček (přiřazeným přímo,
    přes pracovní pozici nebo oddělení) a vrátí seznam
    [{'email', 'pobocky': [názvy], 'vse': bool}]. „vse"=přístup ke všem pobočkám
    (AO / účetní / majitel / superправо). Příjemci se odvozují z práv čtení
    jednotlivých poboček – přesně dle toho, co každý vidí."""
    rel = ['vse', 'vysledky_ao', 'vysledky_ucetni', 'vysledky_majitel'] \
        + [f'vysledky_pobocka_{POBOCKY_KLICE[p]}' for p in POBOCKY]
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor()
        ph = ', '.join(['%s'] * len(rel))
        cur.execute(f"""
            SELECT u.iduser, u.email, x.pname FROM (
                SELECT utp.user_iduser AS uid, LOWER(p.name) AS pname
                FROM user_To_privileges utp
                JOIN privileges p ON utp.privileges_idprivileges = p.idprivileges
                WHERE LOWER(p.name) IN ({ph})
                UNION
                SELECT utj.user_iduser, LOWER(p.name)
                FROM user_To_jobPosition utj
                JOIN jobPosition_To_privileges jtp ON utj.jobPosition_idjobPosition = jtp.jobPosition_idjobPosition
                JOIN privileges p ON jtp.privileges_idprivileges = p.idprivileges
                WHERE LOWER(p.name) IN ({ph})
                UNION
                SELECT dtu.user_iduser, LOWER(p.name)
                FROM department_To_user dtu
                JOIN department_To_privileges dtp ON dtu.department_iddepartment = dtp.department_iddepartment
                JOIN privileges p ON dtp.privileges_idprivileges = p.idprivileges
                WHERE LOWER(p.name) IN ({ph})
            ) x
            JOIN user u ON u.iduser = x.uid
            WHERE u.email IS NOT NULL AND u.email <> ''
        """, rel * 3)
        raw = cur.fetchall()
    except Exception as exc:
        print(f'[vysledky] _nacti_prijemce_aktualizace error: {exc}')
        return []
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()

    # Agregace dle e-mailu (jeden uživatel = jeden e-mail, sloučíme jeho práva)
    by_email: dict[str, dict] = {}
    for _uid, email, pname in raw:
        e = (email or '').strip()
        if not e:
            continue
        d = by_email.setdefault(e.lower(), {'email': e, 'prava': set()})
        d['prava'].add(pname)

    klic_to_pob = {POBOCKY_KLICE[p]: p for p in POBOCKY}
    vysledek: list[dict] = []
    for d in by_email.values():
        pr = d['prava']
        if pr & {'vse', 'vysledky_ao', 'vysledky_ucetni', 'vysledky_majitel'}:
            vysledek.append({'email': d['email'], 'pobocky': list(POBOCKY), 'vse': True})
        else:
            pob = [klic_to_pob[k[len('vysledky_pobocka_'):]]
                   for k in pr
                   if k.startswith('vysledky_pobocka_')
                   and k[len('vysledky_pobocka_'):] in klic_to_pob]
            if pob:
                vysledek.append({'email': d['email'],
                                 'pobocky': [p for p in POBOCKY if p in pob],
                                 'vse': False})
    vysledek.sort(key=lambda r: r['email'].lower())
    return vysledek


# Lidské popisky práv „přístup ke všem pobočkám" pro přehled práv.
_PRAVO_ROLE_LABEL = {
    'vse': 'Superadmin (vše)',
    'vysledky_ao': 'AO',
    'vysledky_ucetni': 'Účetní – hlavní',
    'vysledky_ucetni_bezna': 'Účetní – čtení',
    'vysledky_majitel': 'Majitel',
}
# Všechny tyto role zpřístupňují všechny pobočky (viz pristupne_pobocky).
_PRAVO_ROLE_VSE = set(_PRAVO_ROLE_LABEL)


def _nacti_prehled_prav() -> list[dict]:
    """Pro hlavního administrátora: jak mají jednotliví uživatelé nastavena
    práva na pobočky Výsledků (přiřazená přímo / přes pracovní pozici / přes
    oddělení). Vrací seznam dictů seřazený dle jména:
      {'jmeno', 'email', 'aktivni', 'role': [labely], 'vse': bool,
       'pobocky': [názvy poboček], 'zdroje': [labely]}.
    Role Superadmin/AO/Účetní/Majitel = přístup ke všem pobočkám.
    """
    role_klice = list(_PRAVO_ROLE_LABEL)
    rel = role_klice + [f'vysledky_pobocka_{POBOCKY_KLICE[p]}' for p in POBOCKY]
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor()
        ph = ', '.join(['%s'] * len(rel))
        cur.execute(f"""
            SELECT x.uid, u.name, u.surname, u.email, u.is_active, x.pname, x.zdroj FROM (
                SELECT utp.user_iduser AS uid, LOWER(p.name) AS pname, 'přímo' AS zdroj
                FROM user_To_privileges utp
                JOIN privileges p ON utp.privileges_idprivileges = p.idprivileges
                WHERE LOWER(p.name) IN ({ph})
                UNION
                SELECT utj.user_iduser, LOWER(p.name), 'pozice'
                FROM user_To_jobPosition utj
                JOIN jobPosition_To_privileges jtp ON utj.jobPosition_idjobPosition = jtp.jobPosition_idjobPosition
                JOIN privileges p ON jtp.privileges_idprivileges = p.idprivileges
                WHERE LOWER(p.name) IN ({ph})
                UNION
                SELECT dtu.user_iduser, LOWER(p.name), 'oddělení'
                FROM department_To_user dtu
                JOIN department_To_privileges dtp ON dtu.department_iddepartment = dtp.department_iddepartment
                JOIN privileges p ON dtp.privileges_idprivileges = p.idprivileges
                WHERE LOWER(p.name) IN ({ph})
            ) x
            JOIN user u ON u.iduser = x.uid
        """, rel * 3)
        raw = cur.fetchall()
    except Exception as exc:
        print(f'[vysledky] _nacti_prehled_prav error: {exc}')
        return []
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()

    klic_to_pob = {POBOCKY_KLICE[p]: p for p in POBOCKY}
    by_user: dict = {}
    for uid, name, surname, email, aktivni, pname, zdroj in raw:
        d = by_user.setdefault(uid, {
            'jmeno': f'{name or ""} {surname or ""}'.strip() or (email or f'#{uid}'),
            'email': (email or '').strip(),
            'aktivni': bool(aktivni),
            'role': set(), 'pobocky': set(), 'zdroje': set(),
        })
        if pname in _PRAVO_ROLE_LABEL:
            d['role'].add(pname)
        elif pname.startswith('vysledky_pobocka_'):
            klic = pname[len('vysledky_pobocka_'):]
            if klic in klic_to_pob:
                d['pobocky'].add(klic_to_pob[klic])
        d['zdroje'].add(zdroj)

    vysledek: list[dict] = []
    for d in by_user.values():
        vysledek.append({
            'jmeno': d['jmeno'],
            'email': d['email'],
            'aktivni': d['aktivni'],
            'role': [_PRAVO_ROLE_LABEL[r] for r in role_klice if r in d['role']],
            'vse': bool(d['role'] & _PRAVO_ROLE_VSE),
            'pobocky': [p for p in POBOCKY if p in d['pobocky']],
            'zdroje': sorted(d['zdroje']),
        })
    vysledek.sort(key=lambda r: r['jmeno'].lower())
    return vysledek


def _rozesli_aktualizace_sync(prijemci: list[dict]) -> tuple[int, int]:
    """Rozešle příjemcům e-mail, že Výsledky poboček jsou vyplněné a k dispozici.
    Každému uvede pobočky, na které má přístup. Vrací (odesláno, chyb).
    Blokující (SMTP) – volat přes asyncio.to_thread."""
    import intranet_emaily
    sent = fail = 0
    for r in prijemci:
        if r.get('vse'):
            pob_txt = 'všechny pobočky'
        else:
            pob_txt = ', '.join(
                _POBOCKY_EXCEL_REVERSE.get(p, p) if p not in ('Bourárna', 'Kamiony') else p
                for p in r.get('pobocky', [])
            ) or '—'
        text = (
            "Dobrý den,\n\n"
            "Výsledky poboček byly aktualizovány – všechna data jsou vyplněná "
            "a k dispozici v aplikaci Moje JIPka v sekci „Výsledky poboček\".\n\n"
            f"Máte přístup k: {pob_txt}.\n\n"
            "S pozdravem\nMoje JIPka"
        )
        try:
            ok = intranet_emaily.odesli_upozorneni_email(
                r['email'], 'Výsledky poboček – data jsou k dispozici', text)
        except Exception as exc:
            print(f'[vysledky] rozeslani e-mailu {r["email"]} selhalo: {exc}')
            ok = False
        if ok:
            sent += 1
        else:
            fail += 1
    return sent, fail


def nacti_souhrn(pobocka: str) -> list[dict]:
    """Načte všechna data z tabulka_nakladu_mesice pro danou pobočku."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(
            'SELECT * FROM tabulka_nakladu_mesice WHERE pobocka=%s ORDER BY rok, mesic',
            (pobocka,),
        )
        return cur.fetchall()
    except Exception as exc:
        print(f'[vysledky] nacti_souhrn error: {exc}')
        return []
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()


def nacti_dostupne_roky() -> list[int]:
    """Roky dostupné v tabulka_nakladu_mesice (sestupně) – pro přehled majitele."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor()
        cur.execute('SELECT DISTINCT rok FROM tabulka_nakladu_mesice ORDER BY rok DESC')
        return [int(r[0]) for r in cur.fetchall()]
    except Exception as exc:
        print(f'[vysledky] nacti_dostupne_roky error: {exc}')
        return []
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()


def nacti_vysledky_roky() -> list[int]:
    """Roky dostupné v podrobných nákladech (vysledky_naklady), sestupně –
    pro záložku „Podrobné náklady" v přehledu pro vedení."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor()
        cur.execute('SELECT DISTINCT rok FROM vysledky_naklady ORDER BY rok DESC')
        return [int(r[0]) for r in cur.fetchall()]
    except Exception as exc:
        print(f'[vysledky] nacti_vysledky_roky error: {exc}')
        return []
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()


def _priprav_souhrn_pro_grid(raw: list[dict]) -> list[dict]:
    """Z raw DB řádků sestaví řádky pro ag-Grid:
    • přidá _rok_disp, _mesic_disp, _je_celkem
    • vypočítá _podil_mzdy a _procento
    • vloží roční součtové řádky."""
    if not raw:
        return []
    from collections import defaultdict
    roky_dict: dict[int, list] = defaultdict(list)
    for r in raw:
        roky_dict[int(r['rok'])].append(r)

    result = []
    for rok in sorted(roky_dict.keys()):
        mesice = sorted(roky_dict[rok], key=lambda x: int(x['mesic']))
        totals = {c: 0.0 for c, _, _, _ in _SOUHRN_COLS}
        first  = True
        for m in mesice:
            mzdy  = _s(m.get('mzdy'))
            vys   = _s(m.get('vysledek'))
            tot   = _s(m.get('total'))
            obn   = _s(m.get('obrat_nc'))
            row   = {k: (float(v) if isinstance(v, (int, float)) and k not in ('rok', 'mesic') else v)
                     for k, v in m.items()}
            row['_rok_disp']   = str(rok) if first else ''
            row['_mesic_disp'] = str(int(m['mesic']))
            row['_je_celkem']  = False
            row['_podil_mzdy'] = round(mzdy / vys, 6) if vys else None
            row['_procento']   = round(tot  / obn, 6) if obn else None
            result.append(row)
            first = False
            for c, _, _, _ in _SOUHRN_COLS:
                totals[c] += _s(m.get(c))

        # Roční součtový řádek
        t_mzdy = totals['mzdy']
        t_vys  = totals['vysledek']
        t_tot  = totals['total']
        t_obn  = totals['obrat_nc']
        total_row: dict = dict(totals)
        total_row['_rok_disp']   = f'{rok} Celkem'
        total_row['_mesic_disp'] = ''
        total_row['_je_celkem']  = True
        total_row['_podil_mzdy'] = round(t_mzdy / t_vys, 6) if t_vys else None
        total_row['_procento']   = round(t_tot  / t_obn, 6) if t_obn else None
        result.append(total_row)

    return result


# ─── ag-Grid column definitions ───────────────────────────────────────────────

_HIST_EYE_RENDERER = (
    "function(p){"
    "if(p.node&&p.node.rowPinned)return '';"
    "return '<span title=\"Historie úprav\" "
    "style=\"cursor:pointer;font-size:15px;opacity:.6\">👁️</span>';"
    "}"
)

# Řazení sloupce „Předpis" dle čísla účetní osnovy: přirozené (numeric) řazení
# kódů (501.20 < 501.100), prázdné předpisy vždy dole (v obou směrech).
_PREDPIS_COMPARATOR = (
    "function(a,b,na,nb,desc){"
    "var sa=(a==null?'':String(a)).trim(),sb=(b==null?'':String(b)).trim();"
    "if(sa===''&&sb==='')return 0;"
    "if(sa==='')return desc?-1:1;"
    "if(sb==='')return desc?1:-1;"
    "return sa.localeCompare(sb,'cs',{numeric:true,sensitivity:'base'});"
    "}"
)


def _col_defs_naklady(editovat: bool, zamcene: list[int] | None = None,
                      s_historii: bool = True) -> list[dict]:
    _edit = "function(p){return %s && !(p.node&&p.node.rowPinned);}" \
            % ('true' if editovat else 'false')
    _show_zamky = zamcene is not None
    zamcene = zamcene or []
    cols: list[dict] = [
        {
            'headerName': 'Předpis', 'field': 'ucetni_predpis',
            'pinned': 'left', 'width': 120, 'minWidth': _tn_minw('Předpis'), ':editable': _edit,
            'cellStyle': {'fontFamily': 'monospace', 'fontSize': '12px'},
            'sortable': True, ':comparator': _PREDPIS_COMPARATOR,
            'headerTooltip': 'Klikem seřadíte dle čísla účetní osnovy',
        },
        {
            'headerName': 'Název', 'field': 'nazev_predpisu',
            'pinned': 'left', 'width': 280, 'minWidth': _tn_minw('Název'), ':editable': _edit,
            'sortable': True,
        },
    ]
    if s_historii:
        cols.append({
            'headerName': '', 'field': '_historie',
            'pinned': 'left', 'width': 46, 'minWidth': 46, 'maxWidth': 46,
            'sortable': False, 'editable': False, 'resizable': False,
            'suppressSizeToFit': True, 'suppressAutoSize': True,
            ':cellRenderer': _HIST_EYE_RENDERER,
            'cellStyle': {'textAlign': 'center', 'cursor': 'pointer', 'padding': '0'},
            'headerTooltip': 'Historie úprav řádku',
        })
    for i, (nazev, db) in enumerate(zip(MESICE_NAZVY, MESICE_DB), 1):
        je_zamcen = i in zamcene
        header = f'🔒 {nazev}' if je_zamcen else f'🔓 {nazev}'
        h_cls = 'mesic-header zamceny-mesic' if je_zamcen else 'mesic-header'
        cols.append({
            'headerName': header if _show_zamky else nazev,
            'field': db, 'width': 115, 'minWidth': _tn_minw(nazev),
            ':editable': _edit, 'type': 'numericColumn',
            ':valueFormatter': _CS_FMT,
            'headerClass': h_cls if _show_zamky else '',
        })
    cols.append({
        'headerName': 'Celkem', ':valueGetter': _CELKEM_GETTER,
        'pinned': 'right', 'width': 145, 'minWidth': _tn_minw('Celkem'),
        'editable': False, 'type': 'numericColumn',
        ':valueFormatter': _CS_FMT,
        'cellStyle': {'fontWeight': 'bold'},
    })
    return cols

_ZAMEK_CSS = (
    ".mesic-header{cursor:pointer;}"
    ".mesic-header .ag-header-cell-label{justify-content:center;cursor:pointer;}"
    ".zamceny-mesic{background:#fff3e0!important;color:#e65100!important;}"
)

# ─── Komentáře k buňkám (jako v Excelu) ──────────────────────────────────────
# Pole, ke kterým lze přidat komentář (12 měsíců + Předpis + Název).
_KOMENTAR_POLE = set(MESICE_DB) | {'ucetni_predpis', 'nazev_predpisu'}

# Červený růžek v pravém horním rohu buňky s komentářem (jako v Excelu).
# Pozn.: .ag-cell má v AG Gridu position:absolute, slouží tedy jako kotva pro
# ::after – position:relative bychom nesměli nastavovat (rozbil by layout sloupců).
_KOMENTAR_CSS = (
    ".komentare-grid .ag-cell.ma-komentar::after{"
    "content:'';position:absolute;top:0;right:0;width:0;height:0;"
    "border-top:9px solid #dc2626;border-left:9px solid transparent;"
    "pointer-events:none;z-index:1;}"
)
# cellClassRule „ma-komentar": buňka má komentář, pokud existuje _komentare[field].
_KOMENTAR_CLASSRULE = (
    "function(p){"
    "if(p.node&&p.node.rowPinned)return false;"
    "var k=p.data&&p.data._komentare;"
    "return !!(k&&p.colDef&&p.colDef.field&&k[p.colDef.field]);"
    "}"
)
# Tooltip = text komentáře (zobrazí se při najetí myší na buňku).
_KOMENTAR_TOOLTIP = (
    "function(p){"
    "if(p.node&&p.node.rowPinned)return null;"
    "var k=p.data&&p.data._komentare;"
    "if(k&&p.colDef&&p.colDef.field&&k[p.colDef.field])return k[p.colDef.field];"
    "return null;}"
)


def _col_defs_porovnani(aktualni_rok: int) -> list[dict]:
    minuly_rok = aktualni_rok - 1
    cols: list[dict] = [
        {'headerName': 'Předpis', 'field': 'ucetni_predpis', 'pinned': 'left', 'width': 120, 'minWidth': _tn_minw('Předpis'),
         'cellStyle': {'fontFamily': 'monospace', 'fontSize': '12px'},
         'sortable': True, ':comparator': _PREDPIS_COMPARATOR,
         'headerTooltip': 'Klikem seřadíte dle čísla účetní osnovy'},
        {'headerName': 'Název',   'field': 'nazev_predpisu', 'pinned': 'left', 'width': 280, 'minWidth': _tn_minw('Název'),
         'sortable': True},
    ]
    for nazev, db in zip(MESICE_NAZVY, MESICE_DB):
        cols.append({
            'headerName': nazev,
            'children': [
                {'headerName': str(aktualni_rok), 'field': f'akt_{db}',  'width': 110, 'minWidth': 100, 'type': 'numericColumn', ':valueFormatter': _CS_FMT},
                {'headerName': str(minuly_rok),   'field': f'min_{db}',  'width': 110, 'minWidth': 100, 'type': 'numericColumn', ':valueFormatter': _CS_FMT},
                {'headerName': 'Δ',               'field': f'diff_{db}', 'width': 100, 'minWidth': 90, 'type': 'numericColumn', ':valueFormatter': _CS_FMT, ':cellStyle': _DIFF_STYLE},
            ],
        })
    cols.append({
        'headerName': 'Celkem', 'pinned': 'right',
        'children': [
            {'headerName': str(aktualni_rok), 'field': 'akt_celkem',  'width': 145, 'minWidth': 120, 'type': 'numericColumn', ':valueFormatter': _CS_FMT, 'cellStyle': {'fontWeight': 'bold'}},
            {'headerName': str(minuly_rok),   'field': 'min_celkem',  'width': 145, 'minWidth': 120, 'type': 'numericColumn', ':valueFormatter': _CS_FMT, 'cellStyle': {'fontWeight': 'bold'}},
            {'headerName': 'Δ',               'field': 'diff_celkem', 'width': 130, 'minWidth': 100, 'type': 'numericColumn', ':valueFormatter': _CS_FMT, ':cellStyle': _DIFF_STYLE},
        ],
    })
    return cols


def _tn_minw(text: str, floor: int = 100) -> int:
    """Minimální šířka sloupce (px), aby se text hlavičky neořízl.
    Počítá s tučným Arialem ~14px (~8,5 px/znak) + odsazení buňky.
    ag-grid `minWidth` se respektuje při ručním zmenšení i při auto-fitu –
    sloupec tedy nejde zmenšit tak, aby zakryl část textu."""
    return max(floor, round(len(text) * 8.5) + 34)


def _col_defs_souhrn() -> list[dict]:
    """Definice sloupců pro záložku Tabulka nákladů (souhrn z datového souboru).

    JS funkce (formátování, styl buněk) se v NiceGUI 3.8 do ag-gridu předávají
    klíčem s prefixem „:" (ne dictem {'function': …}, ten se nepřevede).
    Každý sloupec má `minWidth` dle délky hlavičky → nejde zmenšit přes text."""
    cols: list[dict] = [
        {
            'headerName': 'Rok',   'field': '_rok_disp',
            'pinned': 'left', 'width': 130, 'minWidth': 130,   # pojme „2025 Celkem"
            ':cellStyle': _TN_CS_ROK,
        },
        {
            'headerName': 'Měsíc', 'field': '_mesic_disp',
            'pinned': 'left', 'width': 105, 'minWidth': 105,   # pojme „Listopad"/„Prosinec"
            'cellDataType': 'text',   # zabrání „Invalid Number" v řádku Celkem (prázdná buňka)
            ':cellStyle': _TN_CS_NORMAL,
        },
    ]
    for db_col, header, _, width in _SOUHRN_COLS:
        # Barevné zvýraznění sloupců přesně dle Vzoru:
        #   Výsledek = zelená, Total = broskvová, ostatní = bílá / modrá „Celkem".
        if db_col == 'vysledek':
            cs, hcls = _TN_CS_GREEN, 'tn-h-green'
        elif db_col == 'total':
            cs, hcls = _TN_CS_PEACH, 'tn-h-peach'
        else:
            cs, hcls = _TN_CS_NORMAL, None
        mw = _tn_minw(header)
        col: dict = {
            'headerName': header, 'field': db_col,
            'width': max(width, mw), 'minWidth': mw, 'type': 'numericColumn',
            ':valueFormatter': _SOUHRN_FMT_NUM,
            ':cellStyle': cs,
        }
        if hcls:
            col['headerClass'] = hcls
        cols.append(col)
        # Vložit vypočítané sloupce na správné místo
        if db_col == 'mzdy':
            mw2 = _tn_minw('Podíl mzdy v %')
            cols.append({
                'headerName': 'Podíl mzdy v %',
                'field': '_podil_mzdy',
                'width': max(125, mw2), 'minWidth': mw2, 'type': 'numericColumn',
                ':valueFormatter': _SOUHRN_FMT_PCT,
                ':cellStyle': _TN_CS_NORMAL,
            })
        elif db_col == 'total':
            # Sloupec „%" – broskvový, stejně jako „Total".
            mw3 = _tn_minw('%', 70)
            cols.append({
                'headerName': '%',
                'field': '_procento',
                'width': max(90, mw3), 'minWidth': mw3, 'type': 'numericColumn',
                'headerClass': 'tn-h-peach',
                ':valueFormatter': _SOUHRN_FMT_PCT,
                ':cellStyle': _TN_CS_PEACH,
            })
    return cols


# ─── Přehled pro vedení – grafy za pobočky ───────────────────────

def nacti_graf_mesicni(roky: list[int], pobocky: list[str]) -> dict[tuple[int, int], dict]:
    """Měsíční součty metrik (vč. Kg) přes vybrané pobočky, pro dané roky.
    Vrací {(rok, mesic): {metrika: hodnota}}."""
    if not roky or not pobocky:
        return {}
    conn = intranet_data.get_db_connection()
    if not conn:
        return {}
    try:
        cur = conn.cursor(dictionary=True)
        ph_y = ', '.join(['%s'] * len(roky))
        ph_p = ', '.join(['%s'] * len(pobocky))
        cur.execute(
            f'SELECT rok, mesic, '
            f'SUM(mzdy) AS mzdy, SUM(naklady) AS naklady, SUM(total) AS total, '
            f'SUM(vysledek) AS vysledek, SUM(repre_spotreba) AS repre_spotreba, '
            f'SUM(obrat_kg_celkem) AS obrat_kg_celkem, SUM(obrat_kg) AS obrat_kg, '
            f'SUM(obrat_kg_dcery) AS obrat_kg_dcery, SUM(obrat_kg_site) AS obrat_kg_site, '
            f'SUM(obrat_kg_staropramen) AS obrat_kg_staropramen '
            f'FROM tabulka_nakladu_mesice '
            f'WHERE rok IN ({ph_y}) AND pobocka IN ({ph_p}) '
            f'GROUP BY rok, mesic',
            list(roky) + list(pobocky),
        )
        return {(int(r['rok']), int(r['mesic'])): r for r in cur.fetchall()}
    except Exception as exc:
        print(f'[vysledky] nacti_graf_mesicni error: {exc}')
        return {}
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()


def nacti_graf_naklad_kg_pobocky(rok: int, pobocky: list[str]) -> dict[str, dict]:
    """Náklad na 1 Kg = (Mzdy + Náklady) / Obrat v Kg celkem, per pobočka a měsíc
    pro daný rok. Vrací {pobocka: {mesic: hodnota|None}}."""
    if not rok or not pobocky:
        return {}
    conn = intranet_data.get_db_connection()
    if not conn:
        return {}
    try:
        cur = conn.cursor(dictionary=True)
        ph_p = ', '.join(['%s'] * len(pobocky))
        cur.execute(
            f'SELECT pobocka, mesic, SUM(mzdy) AS m, SUM(naklady) AS n, '
            f'SUM(obrat_kg_celkem) AS kg '
            f'FROM tabulka_nakladu_mesice WHERE rok=%s AND pobocka IN ({ph_p}) '
            f'GROUP BY pobocka, mesic',
            [rok] + list(pobocky),
        )
        res: dict[str, dict] = {}
        for r in cur.fetchall():
            kg = _s(r['kg'])
            res.setdefault(r['pobocka'], {})[int(r['mesic'])] = (
                (_s(r['m']) + _s(r['n'])) / kg) if kg else None
        return res
    except Exception as exc:
        print(f'[vysledky] nacti_graf_naklad_kg_pobocky error: {exc}')
        return {}
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()


# ── Přehled poboček – Souhrnné / Mezisoučty tabulky ──────────────────────────

_PREHLED_DB_COLS = [
    'total', 'vysledek', 'naklady', 'mzdy', 'nahrady', 'likvidace', 'inventura',
    'repre_spotreba', 'zasoby', 'obrat_kg_celkem', 'obrat_kg', 'obrat_kg_dcery',
    'obrat_kg_site', 'obrat_kg_staropramen', 'obrat_kg_ostatni',
    'operace', 'operace_dcery', 'operace_site',
]

_PREHLED_METRIKY = [
    ('total', 'Total'),
    ('vysledek', 'Výsledek'),
    ('naklady', 'Náklady'),
    ('mzdy', 'Mzdy'),
    ('nahrady', 'Náhrady'),
    ('repre_spotreba', 'Repre + spotřeba'),
    ('zasoby', 'Zásoby'),
    ('naklad_kg', 'Náklad na 1 Kg'),
    ('obrat_kg_celkem', 'Obrat v Kg celkem'),
    ('obrat_kg', 'Obrat v Kg'),
    ('obrat_kg_dcery', 'Obrat v Kg dcery'),
    ('obrat_kg_site', 'Obrat v Kg sítě'),
    ('obrat_kg_staropramen', 'Obrat v kg Staropramen'),
    ('obrat_kg_ostatni', 'Obrat v Kg Ostatní'),
    ('operace_celkem', 'Operace celkem'),
]

_PREHLED_GRID_CSS = ".prehled-grid{height:auto!important;}"

# Ukotvená hlavička pro mezisoučty (grid v autoHeight = scrolluje se přes
# stránku, takže hlavičku přilepujeme k viewportu přes position:sticky).
# Aby se sticky počítalo vůči stránce, nesmí vnější obaly gridu (předkové
# hlavičky) ustanovit vlastní scroll-container → overflow:visible.
# Vnitřní viewporty (.ag-body-viewport / .ag-center-cols-viewport) jsou
# sourozenci hlavičky a NESMÍ se měnit – drží vodorovný posun a tím i
# ukotvení pinned sloupců „Pobočka"/„Měsíc".
# Mírné zvětšení nejmenších textů v Přehledu poboček (drobné popisky, mřížky).
_PREHLED_ZOOM_CSS = (
    ".prehled-zoom .text-xs{font-size:0.85rem!important;line-height:1.35!important;}"
    ".prehled-zoom .text-sm{font-size:0.95rem!important;}"
    ".prehled-zoom .ag-cell,.prehled-zoom .ag-header-cell{font-size:0.85rem!important;}"
    ".prehled-tabs .q-tab__label{font-size:1rem!important;}"
)

_PREHLED_FMT_NUM2 = (
    "function(p){"
    "if(p.value===null||p.value===undefined||p.value==='')return '—';"
    "var n=parseFloat(p.value);"
    "if(isNaN(n)||n===0)return '—';"
    "return new Intl.NumberFormat('cs-CZ',{maximumFractionDigits:2}).format(n);"
    "}"
)


def _prehled_doplni_pocitane(row, rok):
    """Počítaná pole souhrnu: Náklad na 1 Kg a Operace celkem.
    Náklad na 1 Kg dle pivotu SOUHRN_TOTAL = (Mzdy + Náklady) / Obrat v Kg celkem."""
    nak = row.get(f'naklady_{rok}')
    mzdy = row.get(f'mzdy_{rok}')
    kg = row.get(f'obrat_kg_celkem_{rok}')
    row[f'naklad_kg_{rok}'] = (round((_s(mzdy) + _s(nak)) / kg, 6)
                               if (kg and nak is not None) else None)
    row[f'operace_celkem_{rok}'] = (_s(row.get(f'operace_{rok}'))
                                    + _s(row.get(f'operace_dcery_{rok}'))
                                    + _s(row.get(f'operace_site_{rok}')))


def nacti_prehled_souhrn(roky: list[int], pobocky: list[str],
                         mesice: list[int] | None = None) -> tuple[list[dict], dict]:
    """Roční součty metrik per pobočka pro Souhrnné zobrazení.
    `mesice` (1-12) zúží součty jen na vybrané měsíce; prázdné = celý rok."""
    if not roky or not pobocky:
        return [], {}
    conn = intranet_data.get_db_connection()
    if not conn:
        return [], {}
    try:
        cur = conn.cursor(dictionary=True)
        ph_y = ', '.join(['%s'] * len(roky))
        ph_p = ', '.join(['%s'] * len(pobocky))
        sums = ', '.join(f'SUM({c}) AS {c}' for c in _PREHLED_DB_COLS)
        where = f'rok IN ({ph_y}) AND pobocka IN ({ph_p})'
        params = list(roky) + list(pobocky)
        if mesice:
            where += f' AND mesic IN ({", ".join(["%s"] * len(mesice))})'
            params += list(mesice)
        cur.execute(
            f'SELECT pobocka, rok, {sums} '
            f'FROM tabulka_nakladu_mesice WHERE {where} '
            f'GROUP BY pobocka, rok ORDER BY pobocka, rok',
            params,
        )
        raw = cur.fetchall()
    except Exception as exc:
        print(f'[vysledky] nacti_prehled_souhrn error: {exc}')
        return [], {}
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()
    from collections import defaultdict
    by_pob: dict = defaultdict(dict)
    for r in raw:
        by_pob[r['pobocka']][int(r['rok'])] = r
    rows: list[dict] = []
    totals: dict = {f'{col}_{rok}': 0.0 for col in _PREHLED_DB_COLS for rok in roky}
    for pob in pobocky:
        row: dict = {'_pobocka': _POBOCKY_EXCEL_REVERSE.get(pob, pob), '_je_celkem': False}
        for rok in roky:
            d = by_pob[pob].get(rok, {})
            for col in _PREHLED_DB_COLS:
                row[f'{col}_{rok}'] = _s(d.get(col))
            _prehled_doplni_pocitane(row, rok)
        rows.append(row)
        for col in _PREHLED_DB_COLS:
            for rok in roky:
                totals[f'{col}_{rok}'] += row.get(f'{col}_{rok}') or 0
    total_row: dict = {'_pobocka': 'Celkový součet', '_je_celkem': True}
    total_row.update(totals)
    for rok in roky:
        _prehled_doplni_pocitane(total_row, rok)
    return rows, total_row


def nacti_prehled_mezisoucty(roky: list[int], pobocky: list[str],
                             mesice: list[int] | None = None) -> tuple[list[dict], dict]:
    """Měsíční rozpad metrik per pobočka (všech 12 měsíců, příp. jen vybrané)
    × vybrané roky. Vrací (rows, celkovy_soucet_row)."""
    if not roky or not pobocky:
        return [], {}
    conn = intranet_data.get_db_connection()
    if not conn:
        return [], {}
    try:
        cur = conn.cursor(dictionary=True)
        ph_y = ', '.join(['%s'] * len(roky))
        ph_p = ', '.join(['%s'] * len(pobocky))
        sums = ', '.join(f'SUM({c}) AS {c}' for c in _PREHLED_DB_COLS)
        where = f'rok IN ({ph_y}) AND pobocka IN ({ph_p})'
        params = list(roky) + list(pobocky)
        if mesice:
            where += f' AND mesic IN ({", ".join(["%s"] * len(mesice))})'
            params += list(mesice)
        cur.execute(
            f'SELECT pobocka, rok, mesic, {sums} '
            f'FROM tabulka_nakladu_mesice WHERE {where} '
            f'GROUP BY pobocka, rok, mesic ORDER BY pobocka, mesic, rok',
            params,
        )
        raw = cur.fetchall()
    except Exception as exc:
        print(f'[vysledky] nacti_prehled_mezisoucty error: {exc}')
        return [], {}
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()
    mesice_zobr = sorted(mesice) if mesice else list(range(1, 13))
    from collections import defaultdict
    by_pob_m: dict = defaultdict(lambda: defaultdict(dict))
    for r in raw:
        by_pob_m[r['pobocka']][int(r['mesic'])][int(r['rok'])] = r
    rows: list[dict] = []
    totals: dict = {f'{col}_{rok}': 0.0 for col in _PREHLED_DB_COLS for rok in roky}
    for pob in pobocky:
        first = True
        for m in mesice_zobr:
            row: dict = {
                '_pobocka': _POBOCKY_EXCEL_REVERSE.get(pob, pob) if first else '',
                '_mesic': str(m),
                '_je_celkem': False,
            }
            first = False
            d = by_pob_m[pob].get(m, {})
            for rok in roky:
                dr = d.get(rok, {})
                for col in _PREHLED_DB_COLS:
                    v = _s(dr.get(col))
                    row[f'{col}_{rok}'] = v
                    totals[f'{col}_{rok}'] += v or 0
                _prehled_doplni_pocitane(row, rok)
            rows.append(row)
    total_row: dict = {'_pobocka': 'Celkový součet', '_mesic': '', '_je_celkem': True}
    total_row.update(totals)
    for rok in roky:
        _prehled_doplni_pocitane(total_row, rok)
    return rows, total_row


def _col_defs_prehled_souhrn(roky: list[int]) -> list[dict]:
    """Sloupce souhrnu – pro každou metriku jeden podsloupec na vybraný rok."""
    cols: list[dict] = [
        {'headerName': 'Pobočka', 'field': '_pobocka', 'pinned': 'left',
         'width': 200, 'minWidth': 180, ':cellStyle': _SOUHRN_CELKEM_STYLE},
    ]
    for mk, nazev in _PREHLED_METRIKY:
        fmt = _PREHLED_FMT_NUM2 if mk == 'naklad_kg' else _SOUHRN_FMT_NUM
        cols.append({
            'headerName': nazev,
            'children': [
                {'headerName': str(rok), 'field': f'{mk}_{rok}',
                 'width': 140, 'minWidth': 105, 'type': 'numericColumn',
                 ':valueFormatter': fmt, ':cellStyle': _SOUHRN_CELKEM_STYLE}
                for rok in roky
            ],
        })
    return cols


def _col_defs_prehled_mezisoucty(roky: list[int]) -> list[dict]:
    """Sloupce mezisoučtů – Pobočka + Měsíc, pak metrika × vybraný rok."""
    cols: list[dict] = [
        {'headerName': 'Pobočka', 'field': '_pobocka', 'pinned': 'left',
         'width': 200, 'minWidth': 180, ':cellStyle': _SOUHRN_CELKEM_STYLE},
        {'headerName': 'Měsíc', 'field': '_mesic', 'pinned': 'left',
         'width': 80, 'minWidth': 60, 'cellDataType': 'text',
         ':cellStyle': _SOUHRN_CELKEM_STYLE},
    ]
    for mk, nazev in _PREHLED_METRIKY:
        fmt = _PREHLED_FMT_NUM2 if mk == 'naklad_kg' else _SOUHRN_FMT_NUM
        cols.append({
            'headerName': nazev,
            'children': [
                {'headerName': str(rok), 'field': f'{mk}_{rok}',
                 'width': 140, 'minWidth': 105, 'type': 'numericColumn',
                 ':valueFormatter': fmt, ':cellStyle': _SOUHRN_CELKEM_STYLE}
                for rok in roky
            ],
        })
    return cols


# ── Přehled poboček – KT „Obraty v Kg a Operace" ─────────────────────────────
# Leaf sloupce sčítané z DB; mezisoučty (celkem bez sítí, celkem) a rozdíly %
# se dopočítávají (počítaná pole kontingenční tabulky).
_OBOP_KG_LEAF = ['obrat_kg', 'obrat_kg_dcery', 'obrat_kg_staropramen',
               'obrat_kg_ostatni', 'obrat_kg_site']
_OBOP_OP_LEAF = ['operace', 'operace_dcery', 'operace_site']

# Pozadí dle Vzoru: mezisoučty světle modré, „celkem" tmavě modré (jako pivot).
_OBOP_BLUE_L = '#DBE5F1'   # theme4 tint 0.8 – „celkem bez sítí", „sítě", % hlavičky
_OBOP_BLUE_M = '#B8CCE4'   # theme4 tint 0.6 – „celkem" + součtový řádek %

_OBOP_CSS = (
    ".obop-grid .ag-header-cell.obop-h-light,.obop-grid .ag-header-group-cell.obop-h-light"
    "{background-color:#DBE5F1!important;}"
    ".obop-grid .ag-header-cell.obop-h-med,.obop-grid .ag-header-group-cell.obop-h-med"
    "{background-color:#B8CCE4!important;}"
    ".obop-grid .ag-header-group-cell-label,.obop-grid .ag-header-cell-label"
    "{font-weight:700;justify-content:center;text-align:center;}"
    ".obop-grid .ag-cell,.obop-grid .ag-header-cell,.obop-grid .ag-header-group-cell"
    "{font-family:Calibri,Arial,sans-serif;}"
)

# Číslo „#,##0" (0 = „0"), procento „0,00%" – přesně dle Vzoru.
_OBOP_FMT_NUM = (
    "function(p){"
    "if(p.value===null||p.value===undefined||p.value==='')return '—';"
    "var n=parseFloat(p.value);if(isNaN(n)||n===0)return '—';"
    "return new Intl.NumberFormat('cs-CZ',{maximumFractionDigits:0}).format(n);"
    "}"
)


def _obop_diff(old: float, new: float):
    """Rozdíl (new−old)/old; 0 když obě 0; None když old=0 a new≠0."""
    if old:
        return (new - old) / old
    return 0.0 if not new else None


def _obop_cell_style(barva: str) -> str:
    """cellStyle pro datovou buňku: pevné modré pozadí (L/M) nebo bílé;
    součtový řádek („_je_celkem") = tučně."""
    if barva:
        return ("function(p){var s={backgroundColor:'" + barva + "'};"
                "if(p.data&&p.data._je_celkem)s.fontWeight='bold';return s;}")
    return ("function(p){return (p.data&&p.data._je_celkem)?"
            "{fontWeight:'bold'}:null;}")


def _obop_colorscale_style(values: list) -> str:
    """3-barevná škála jako ve Vzoru: min=#F8696B (červená), 50. percentil=
    #FFEB84 (žlutá), max=#63BE7B (zelená). Total řádek bez škály (tmavě modrý).
    `values` = všechny % hodnoty obou/tří sloupců tabulky (bez total řádku)."""
    base = ("function(p){if(p.data&&p.data._je_celkem)return"
            "{fontWeight:'bold',backgroundColor:'" + _OBOP_BLUE_M + "'};return null;}")
    vals = sorted(v for v in values if isinstance(v, (int, float)))
    if len(vals) < 2 or vals[0] == vals[-1]:
        return base
    lo, hi = vals[0], vals[-1]
    m = len(vals)
    mid = vals[m // 2] if m % 2 else (vals[m // 2 - 1] + vals[m // 2]) / 2
    return (
        "function(p){"
        "if(p.data&&p.data._je_celkem)return{fontWeight:'bold',backgroundColor:'" + _OBOP_BLUE_M + "'};"
        "if(p.data&&p.data._bez_skaly)return null;"
        "var v=parseFloat(p.value);if(isNaN(v))return null;"
        f"var lo={lo!r},mid={mid!r},hi={hi!r};"
        "var R=[248,105,107],Y=[255,235,132],G=[99,190,123];"
        "function mix(a,b,t){t=t<0?0:(t>1?1:t);"
        "return 'rgb('+Math.round(a[0]+(b[0]-a[0])*t)+','+Math.round(a[1]+(b[1]-a[1])*t)+','+Math.round(a[2]+(b[2]-a[2])*t)+')';}"
        "var bg;"
        "if(v<=lo)bg='rgb(248,105,107)';"
        "else if(v>=hi)bg='rgb(99,190,123)';"
        "else if(v<=mid)bg=mix(R,Y,(mid>lo)?(v-lo)/(mid-lo):1);"
        "else bg=mix(Y,G,(hi>mid)?(v-mid)/(hi-mid):0);"
        "return{backgroundColor:bg};}"
    )


def nacti_prehled_obraty_operace(roky: list[int], pobocky: list[str],
                                 mesice: list[int] | None = None
                                 ) -> tuple[list[dict], dict, int, int]:
    """KT „Obraty v Kg a Operace" – součty per pobočka × 2 roky (starší → novější)
    přes vybrané měsíce, s počítanými poli. Zdroj: tabulka_nakladu_mesice
    (z „Tabulka nákladů_DATA"). Vrací (rows, total_row, rok_old, rok_new)."""
    if not roky or not pobocky:
        return [], {}, 0, 0
    eff = sorted(roky)[-2:]
    rok_old, rok_new = (eff[0], eff[-1]) if len(eff) == 2 else (eff[0], eff[0])
    cols = _OBOP_KG_LEAF + _OBOP_OP_LEAF
    conn = intranet_data.get_db_connection()
    if not conn:
        return [], {}, rok_old, rok_new
    eff_uniq = sorted({rok_old, rok_new})
    try:
        cur = conn.cursor(dictionary=True)
        ph_y = ', '.join(['%s'] * len(eff_uniq))
        ph_p = ', '.join(['%s'] * len(pobocky))
        sums = ', '.join(f'SUM({c}) AS {c}' for c in cols)
        where = f'rok IN ({ph_y}) AND pobocka IN ({ph_p})'
        params = list(eff_uniq) + list(pobocky)
        if mesice:
            where += f' AND mesic IN ({", ".join(["%s"] * len(mesice))})'
            params += list(mesice)
        cur.execute(
            f'SELECT pobocka, rok, {sums} '
            f'FROM tabulka_nakladu_mesice WHERE {where} '
            f'GROUP BY pobocka, rok',
            params,
        )
        raw = cur.fetchall()
    except Exception as exc:
        print(f'[vysledky] nacti_prehled_obraty_operace error: {exc}')
        return [], {}, rok_old, rok_new
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()

    from collections import defaultdict
    by_pob: dict = defaultdict(dict)
    for r in raw:
        by_pob[r['pobocka']][int(r['rok'])] = r

    def _vypln(row: dict, src_by_rok: dict):
        for r in (rok_old, rok_new):
            d = src_by_rok.get(r, {})
            zak = _s(d.get('obrat_kg')); dcery = _s(d.get('obrat_kg_dcery'))
            staro = _s(d.get('obrat_kg_staropramen')); ost = _s(d.get('obrat_kg_ostatni'))
            site = _s(d.get('obrat_kg_site'))
            bezsiti = zak + dcery + staro + ost
            row[f'okg_zak_{r}'] = zak
            row[f'okg_dcery_{r}'] = dcery
            row[f'okg_staro_{r}'] = staro
            row[f'okg_ost_{r}'] = ost
            row[f'okg_bezsiti_{r}'] = bezsiti
            row[f'okg_site_{r}'] = site
            row[f'okg_celkem_{r}'] = bezsiti + site
            ozak = _s(d.get('operace')); odcery = _s(d.get('operace_dcery'))
            osite = _s(d.get('operace_site'))
            obez = ozak + odcery
            row[f'op_zak_{r}'] = ozak
            row[f'op_dcery_{r}'] = odcery
            row[f'op_bezsiti_{r}'] = obez
            row[f'op_site_{r}'] = osite
            row[f'op_celkem_{r}'] = obez + osite
        for klic in ('okg_bezsiti', 'okg_site', 'okg_celkem',
                     'op_bezsiti', 'op_site', 'op_celkem'):
            row[f'{klic}_diff'] = _obop_diff(row[f'{klic}_{rok_old}'],
                                           row[f'{klic}_{rok_new}'])

    rows: list[dict] = []
    agg_tot: dict = defaultdict(dict)
    for pob in pobocky:
        src = by_pob.get(pob, {})
        nazev = _POBOCKY_EXCEL_REVERSE.get(pob, pob)
        row = {'_pobocka': nazev, '_je_celkem': False,
               '_bez_skaly': 'prefakturace' in _bez_diakritiky(nazev)}
        _vypln(row, src)
        rows.append(row)
        for r in (rok_old, rok_new):
            d = src.get(r, {})
            tt = agg_tot[r]
            for c in cols:
                tt[c] = tt.get(c, 0.0) + _s(d.get(c))
    rows.sort(key=lambda x: x['_pobocka'])
    total = {'_pobocka': 'Celkový součet', '_je_celkem': True}
    _vypln(total, {r: agg_tot.get(r, {}) for r in (rok_old, rok_new)})
    return rows, total, rok_old, rok_new


def _obop_grp(name: str, klic: str, rok_old: int, rok_new: int, barva: str = '') -> dict:
    """Skupina metriky se dvěma podsloupci (starší/novější rok)."""
    hcls = {'L': 'obop-h-light', 'M': 'obop-h-med'}.get(
        'L' if barva == _OBOP_BLUE_L else 'M' if barva == _OBOP_BLUE_M else '', '')
    return {
        'headerName': name,
        'headerClass': hcls,
        'children': [
            {'headerName': str(r), 'field': f'{klic}_{r}', 'type': 'numericColumn',
             'width': 102, 'minWidth': 70, ':valueFormatter': _OBOP_FMT_NUM,
             ':cellStyle': _obop_cell_style(barva)}
            for r in (rok_old, rok_new)
        ],
    }


def _obop_pct_col(name: str, field: str, pct_style: str) -> dict:
    return {
        'headerName': name, 'field': field, 'type': 'numericColumn',
        'headerClass': 'obop-h-light', 'width': 116, 'minWidth': 96,
        'wrapHeaderText': True, 'autoHeaderHeight': True,
        ':valueFormatter': _SOUHRN_FMT_PCT, ':cellStyle': pct_style,
    }


def _col_defs_obop_kg(rok_old: int, rok_new: int, pct_style: str) -> list[dict]:
    cols = [
        {'headerName': 'Pobočka popis', 'field': '_pobocka', 'pinned': 'left',
         'width': 168, 'minWidth': 138, ':cellStyle': _obop_cell_style('')},
        _obop_grp('Obrat v Kg zákazník', 'okg_zak', rok_old, rok_new),
        _obop_grp('Obrat v Kg dcery', 'okg_dcery', rok_old, rok_new),
        _obop_grp('Obrat v Kg Staropramen', 'okg_staro', rok_old, rok_new),
        _obop_grp('Obrat v Kg Ostatní', 'okg_ost', rok_old, rok_new),
        _obop_grp('Obrat v Kg celkem BEZ SÍTÍ', 'okg_bezsiti', rok_old, rok_new, _OBOP_BLUE_L),
        _obop_grp('Obrat v Kg sítě', 'okg_site', rok_old, rok_new, _OBOP_BLUE_L),
        _obop_grp('Obrat v Kg celkem', 'okg_celkem', rok_old, rok_new, _OBOP_BLUE_M),
        _obop_pct_col(f'Obrat v Kg celkem bez sítě rozdíl {rok_new}×{rok_old} %',
                    'okg_bezsiti_diff', pct_style),
        _obop_pct_col(f'Obrat v Kg sítě rozdíl {rok_new}×{rok_old} %',
                    'okg_site_diff', pct_style),
        _obop_pct_col(f'Obrat v Kg celkem rozdíl {rok_new}×{rok_old} %',
                    'okg_celkem_diff', pct_style),
    ]
    return cols


def _col_defs_obop_op(rok_old: int, rok_new: int, pct_style: str) -> list[dict]:
    cols = [
        {'headerName': 'Pobočka popis', 'field': '_pobocka', 'pinned': 'left',
         'width': 168, 'minWidth': 138, ':cellStyle': _obop_cell_style('')},
        _obop_grp('Operace zákazník', 'op_zak', rok_old, rok_new),
        _obop_grp('Operace dcery', 'op_dcery', rok_old, rok_new),
        _obop_grp('Operace celkem BEZ SÍTÍ', 'op_bezsiti', rok_old, rok_new, _OBOP_BLUE_L),
        _obop_grp('Operace sítě', 'op_site', rok_old, rok_new, _OBOP_BLUE_L),
        _obop_grp('Operace celkem', 'op_celkem', rok_old, rok_new, _OBOP_BLUE_M),
        _obop_pct_col(f'Operace celkem bez sítě rozdíl {rok_new}×{rok_old} %',
                    'op_bezsiti_diff', pct_style),
        _obop_pct_col(f'Operace sítě rozdíl {rok_new}×{rok_old} %',
                    'op_site_diff', pct_style),
        _obop_pct_col(f'Operace celkem rozdíl {rok_new}×{rok_old} %',
                    'op_celkem_diff', pct_style),
    ]
    return cols


# ── Přehled poboček – KT „Přehled AO" (porovnání 2 nejnovějších let) ─────────
# Pivot s rozdíly Kč i % pro Total/Výsledek/Náklady/Mzdy/Naklad na 1 Kg, plus
# 2 sloupce (starší/novější rok) pro Náhrady/Repre+spotřeba/Zásoby/Obraty/Operace.
# Vizuál odpovídá PivotStyleLight16 (accent2 = oranžová #E97132) – styl bez
# výplní řádků, jen oranžové okraje hlavičky/první sloupec/total a bold.
_AO_ACCENT = '#E97132'
_AO_ACCENT_LIGHT = '#FBE5D6'

_AO_CSS = (
    ".ao-grid .ag-header-cell,.ao-grid .ag-header-group-cell"
    f"{{background:#fff!important;border-bottom:2px solid {_AO_ACCENT}!important;}}"
    ".ao-grid .ag-header-group-cell-label,.ao-grid .ag-header-cell-label"
    "{font-weight:700;justify-content:center;text-align:center;}"
    ".ao-grid .ag-cell{background:#fff;}"
    ".ao-grid .ag-cell,.ao-grid .ag-header-cell,.ao-grid .ag-header-group-cell"
    "{font-family:Calibri,Arial,sans-serif;}"
)

# Číslo „#,##0" bez desetinných míst, „#,##0.00" pro Náklad na 1 Kg, procento „0,0%".
# Nuly (i prázdné/nedefinované hodnoty) se zobrazují jako „—".
_AO_FMT_NUM = (
    "function(p){"
    "if(p.value===null||p.value===undefined||p.value==='')return '—';"
    "var n=parseFloat(p.value);if(isNaN(n)||n===0)return '—';"
    "return new Intl.NumberFormat('cs-CZ',{maximumFractionDigits:0}).format(n);"
    "}"
)
_AO_FMT_NUM2 = (
    "function(p){"
    "if(p.value===null||p.value===undefined||p.value==='')return '—';"
    "var n=parseFloat(p.value);if(isNaN(n)||n===0)return '—';"
    "return new Intl.NumberFormat('cs-CZ',{minimumFractionDigits:2,maximumFractionDigits:2}).format(n);"
    "}"
)
_AO_FMT_PCT = (
    "function(p){"
    "if(p.value===null||p.value===undefined||p.value==='')return '—';"
    "var n=parseFloat(p.value);if(isNaN(n)||n===0)return '—';"
    "return new Intl.NumberFormat('cs-CZ',{minimumFractionDigits:1,maximumFractionDigits:1}).format(n*100)+'%';"
    "}"
)
# Datová buňka: bílá výplň; total řádek = bold + horní oranžový okraj + světlá výplň.
_AO_CELL_STYLE = (
    "function(p){"
    "if(p.data&&p.data._je_celkem)return{fontWeight:'bold',"
    f"borderTop:'2px solid {_AO_ACCENT}',backgroundColor:'{_AO_ACCENT_LIGHT}'}};"
    "return null;}"
)
# První sloupec (Pobočka): bold + pravý oranžový okraj; total řádek navíc bold + horní okraj.
_AO_FIRSTCOL_STYLE = (
    "function(p){"
    "var s={fontWeight:'bold',"
    f"borderRight:'2px solid {_AO_ACCENT}'}};"
    "if(p.data&&p.data._je_celkem){"
    f"s.borderTop='2px solid {_AO_ACCENT}';"
    f"s.backgroundColor='{_AO_ACCENT_LIGHT}';}}"
    "return s;}"
)

# Podmíněné formátování dle Excelu (Good/Bad styl): zelená/červená.
_AO_CF_GOOD_BG = '#C6EFCE'
_AO_CF_GOOD_FG = '#006100'
_AO_CF_BAD_BG  = '#FFC7CE'
_AO_CF_BAD_FG  = '#9C0006'


def _ao_cf_style(positive_is: str) -> str:
    """cellStyle pro rozdílový sloupec s podmíněným formátováním (1:1 dle Excelu).
    `positive_is='good'` → kladná hodnota = zelená, záporná = červená (např. Total %).
    `positive_is='bad'`  → kladná hodnota = červená, záporná = zelená (Náklady, Mzdy,
    Total Kč, Výsledek Kč, NK – v Excelu označeno červeně, když rozdíl > 0).
    `positive_is='neutral'` → bez podmíněného barvení (jen styl total řádku) –
    pro metriky bez jednoznačné „lepší/horší" interpretace (např. Zásoby).
    Total („Celkový součet") řádek si zachovává vlastní orange-themed styl."""
    if positive_is == 'neutral':
        return _AO_CELL_STYLE
    if positive_is == 'good':
        pbg, pfg = _AO_CF_GOOD_BG, _AO_CF_GOOD_FG
        nbg, nfg = _AO_CF_BAD_BG, _AO_CF_BAD_FG
    else:
        pbg, pfg = _AO_CF_BAD_BG, _AO_CF_BAD_FG
        nbg, nfg = _AO_CF_GOOD_BG, _AO_CF_GOOD_FG
    return (
        "function(p){"
        "if(p.data&&p.data._je_celkem)return{fontWeight:'bold',"
        f"borderTop:'2px solid {_AO_ACCENT}',backgroundColor:'{_AO_ACCENT_LIGHT}'}};"
        "var v=parseFloat(p.value);if(isNaN(v))return null;"
        f"if(v>0)return{{backgroundColor:'{pbg}',color:'{pfg}'}};"
        f"if(v<0)return{{backgroundColor:'{nbg}',color:'{nfg}'}};"
        "return null;}"
    )


def _ao_grp_diff(name: str, klic: str, rok_old: int, rok_new: int,
                 kc_pos: str = 'good', pct_pos: str = 'good',
                 fmt: str = _AO_FMT_NUM) -> dict:
    """Skupina 4 sloupců: starší rok / novější rok / Rozdíl Kč / Rozdíl %.
    `kc_pos`/`pct_pos` ovládá podmíněné formátování posledních 2 sloupců –
    'good' = > 0 zelená / < 0 červená; 'bad' = obráceně. Viz [[mojejipka_prehled_ao]]."""
    return {
        'headerName': name,
        'children': [
            {'headerName': str(rok_old), 'field': f'{klic}_{rok_old}',
             'type': 'numericColumn', 'width': 130, 'minWidth': 110,
             ':valueFormatter': fmt, ':cellStyle': _AO_CELL_STYLE},
            {'headerName': str(rok_new), 'field': f'{klic}_{rok_new}',
             'type': 'numericColumn', 'width': 130, 'minWidth': 110,
             ':valueFormatter': fmt, ':cellStyle': _AO_CELL_STYLE},
            {'headerName': 'Rozdíl Kč', 'field': f'{klic}_diff_kc',
             'type': 'numericColumn', 'width': 130, 'minWidth': 110,
             ':valueFormatter': fmt, ':cellStyle': _ao_cf_style(kc_pos)},
            {'headerName': 'Rozdíl %', 'field': f'{klic}_diff_pct',
             'type': 'numericColumn', 'width': 100, 'minWidth': 90,
             'wrapHeaderText': True, 'autoHeaderHeight': True,
             ':valueFormatter': _AO_FMT_PCT, ':cellStyle': _ao_cf_style(pct_pos)},
        ],
    }


def _ao_grp_nk(rok_old: int, rok_new: int) -> dict:
    """Skupina pro Průměr z Náklad na 1 Kg (2 desetinná místa + rozdíl + %).
    Rozdíl > 0 = červená (víc Kč na kg = horší), < 0 = zelená."""
    return {
        'headerName': 'Průměr z Náklad na 1 Kg',
        'children': [
            {'headerName': str(rok_old), 'field': f'naklad_kg_{rok_old}',
             'type': 'numericColumn', 'width': 120, 'minWidth': 100,
             ':valueFormatter': _AO_FMT_NUM2, ':cellStyle': _AO_CELL_STYLE},
            {'headerName': str(rok_new), 'field': f'naklad_kg_{rok_new}',
             'type': 'numericColumn', 'width': 120, 'minWidth': 100,
             ':valueFormatter': _AO_FMT_NUM2, ':cellStyle': _AO_CELL_STYLE},
            {'headerName': 'Rozdíl', 'field': 'naklad_kg_diff',
             'type': 'numericColumn', 'width': 110, 'minWidth': 95,
             ':valueFormatter': _AO_FMT_NUM2, ':cellStyle': _ao_cf_style('bad')},
            {'headerName': 'Rozdíl %', 'field': 'naklad_kg_diff_pct',
             'type': 'numericColumn', 'width': 100, 'minWidth': 90,
             'wrapHeaderText': True, 'autoHeaderHeight': True,
             ':valueFormatter': _AO_FMT_PCT, ':cellStyle': _ao_cf_style('bad')},
        ],
    }


def _col_defs_prehled_ao(rok_old: int, rok_new: int) -> list[dict]:
    """Sloupce KT „Přehled AO" – pořadí dle listu „porovnání 2026x2025".
    Barevné CF dle obchodní logiky (NE 1:1 dle Excelu, který měl u Total/Výsledku
    Kč a % nekonzistentní/inverzní pravidla):
      • Total, Výsledek = „čím víc, tím líp" → růst (>0) zelená, pokles (<0) červená.
      • Náklady, Mzdy, Náklad na 1 Kg = „čím míň, tím líp" → pokles (<0) zelená,
        růst (>0) červená."""
    return [
        {'headerName': 'Pobočka popis', 'field': '_pobocka', 'pinned': 'left',
         'width': 200, 'minWidth': 170, ':cellStyle': _AO_FIRSTCOL_STYLE},
        # Total/Výsledek = tržby/zisk: růst (>0) zelená, pokles (<0) červená (Kč i %)
        _ao_grp_diff('Total',    'total',    rok_old, rok_new,
                     kc_pos='good', pct_pos='good'),
        _ao_grp_diff('Výsledek', 'vysledek', rok_old, rok_new,
                     kc_pos='good', pct_pos='good'),
        # Náklady/Mzdy = náklady: růst (>0) červená, pokles (<0) zelená (Kč i %)
        _ao_grp_diff('Náklady',  'naklady',  rok_old, rok_new,
                     kc_pos='bad', pct_pos='bad'),
        _ao_grp_diff('Mzdy',     'mzdy',     rok_old, rok_new,
                     kc_pos='bad', pct_pos='bad'),
        # Náhrady / Likvidace / Inventura / Repre + spotřeba = náklady →
        # růst (>0) červená, pokles (<0) zelená. Likvidace = sl. „Likvidace",
        # Inventura = sl. „Inventura" z importu Tabulky nákladů (tabulka_nakladu_mesice).
        _ao_grp_diff('Náhrady',          'nahrady',        rok_old, rok_new,
                     kc_pos='bad', pct_pos='bad'),
        _ao_grp_diff('Likvidace',        'likvidace',      rok_old, rok_new,
                     kc_pos='bad', pct_pos='bad'),
        _ao_grp_diff('Inventura',        'inventura',      rok_old, rok_new,
                     kc_pos='bad', pct_pos='bad'),
        _ao_grp_diff('Repre + spotřeba', 'repre_spotreba', rok_old, rok_new,
                     kc_pos='bad', pct_pos='bad'),
        # Zásoby = bez jednoznačné interpretace → rozdíl bez barvení
        _ao_grp_diff('Zásoby',           'zasoby',         rok_old, rok_new,
                     kc_pos='neutral', pct_pos='neutral'),
        _ao_grp_nk(rok_old, rok_new),
        # Obraty v Kg / Operace = výkon → růst (>0) zelená, pokles červená
        _ao_grp_diff('Obrat v Kg celkem',     'obrat_kg_celkem',     rok_old, rok_new,
                     kc_pos='good', pct_pos='good'),
        _ao_grp_diff('Obrat v Kg',            'obrat_kg',            rok_old, rok_new,
                     kc_pos='good', pct_pos='good'),
        _ao_grp_diff('Obrat v Kg dcery',      'obrat_kg_dcery',      rok_old, rok_new,
                     kc_pos='good', pct_pos='good'),
        _ao_grp_diff('Obrat v Kg sítě',       'obrat_kg_site',       rok_old, rok_new,
                     kc_pos='good', pct_pos='good'),
        _ao_grp_diff('Obrat v Kg Staropramen','obrat_kg_staropramen',rok_old, rok_new,
                     kc_pos='good', pct_pos='good'),
        _ao_grp_diff('Obrat v Kg Ostatní',    'obrat_kg_ostatni',    rok_old, rok_new,
                     kc_pos='good', pct_pos='good'),
        _ao_grp_diff('Operace celkem',        'operace_celkem',      rok_old, rok_new,
                     kc_pos='good', pct_pos='good'),
    ]


def _nacti_naklad_kg_prumer(roky: list[int], pobocky: list[str],
                            mesice: list[int] | None = None
                            ) -> tuple[dict, dict]:
    """Per (pobocka, rok): průměr měsíčního (Mzdy+Náklady)/obrat_kg_celkem přes
    měsíce s validním obratem. Zároveň total: per rok průměr přes všechny
    (pobočka, měsíc). Vrací (per_pob_rok, per_rok)."""
    if not roky or not pobocky:
        return {}, {}
    conn = intranet_data.get_db_connection()
    if not conn:
        return {}, {}
    try:
        cur = conn.cursor(dictionary=True)
        ph_y = ', '.join(['%s'] * len(roky))
        ph_p = ', '.join(['%s'] * len(pobocky))
        where = f'rok IN ({ph_y}) AND pobocka IN ({ph_p})'
        params = list(roky) + list(pobocky)
        if mesice:
            where += f' AND mesic IN ({", ".join(["%s"] * len(mesice))})'
            params += list(mesice)
        cur.execute(
            f'SELECT pobocka, rok, mesic, SUM(mzdy) AS m, SUM(naklady) AS n, '
            f'SUM(obrat_kg_celkem) AS kg FROM tabulka_nakladu_mesice '
            f'WHERE {where} GROUP BY pobocka, rok, mesic',
            params,
        )
        raw = cur.fetchall()
    except Exception as exc:
        print(f'[vysledky] _nacti_naklad_kg_prumer error: {exc}')
        return {}, {}
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()
    from collections import defaultdict
    by_pob_rok = defaultdict(list)
    by_rok = defaultdict(list)
    for r in raw:
        kg = _s(r['kg'])
        if kg <= 0:
            continue
        v = (_s(r['m']) + _s(r['n'])) / kg
        by_pob_rok[(r['pobocka'], int(r['rok']))].append(v)
        by_rok[int(r['rok'])].append(v)
    per_pob = {k: sum(v) / len(v) for k, v in by_pob_rok.items() if v}
    per_rok = {r: sum(v) / len(v) for r, v in by_rok.items() if v}
    return per_pob, per_rok


def nacti_prehled_ao(roky: list[int], pobocky: list[str],
                     mesice: list[int] | None = None
                     ) -> tuple[list[dict], dict, int, int]:
    """KT „Přehled AO" – porovnání 2 nejnovějších let (z `roky`) per pobočka.
    Zdroj: tabulka_nakladu_mesice. Vrací (rows, total_row, rok_old, rok_new)."""
    if not roky or not pobocky:
        return [], {}, 0, 0
    eff = sorted(roky)[-2:]
    rok_old, rok_new = (eff[0], eff[-1]) if len(eff) == 2 else (eff[0], eff[0])
    eff_uniq = sorted({rok_old, rok_new})

    # Roční součty per pobočka (znovupoužití souhrn loaderu)
    s_rows, s_total = nacti_prehled_souhrn(eff_uniq, pobocky, mesice)
    if not s_rows:
        return [], {}, rok_old, rok_new

    # Naklad na 1 Kg = průměr měsíčního ratio (dle Excelu „Průměr z Náklad na 1 Kg")
    nk_per_pob, nk_per_rok = _nacti_naklad_kg_prumer(eff_uniq, pobocky, mesice)

    # Mapování pobočka_excel (UI) → pobočka_db (klíč v `pobocky` listu)
    excel_to_db = {_POBOCKY_EXCEL_REVERSE.get(p, p): p for p in pobocky}

    # Rozdíl Kč + % počítáme pro VŠECHNY metriky zobrazené v Přehledu AO
    # (operace_celkem se dopočítá v _op_celkem ještě před _diff).
    diff_metriky = (
        'total', 'vysledek', 'naklady', 'mzdy', 'nahrady', 'likvidace', 'inventura',
        'repre_spotreba', 'zasoby', 'obrat_kg_celkem', 'obrat_kg', 'obrat_kg_dcery',
        'obrat_kg_site', 'obrat_kg_staropramen', 'obrat_kg_ostatni', 'operace_celkem',
    )

    def _op_celkem(row):
        for r in (rok_old, rok_new):
            row[f'operace_celkem_{r}'] = (_s(row.get(f'operace_{r}'))
                                          + _s(row.get(f'operace_dcery_{r}'))
                                          + _s(row.get(f'operace_site_{r}')))

    def _diff(row):
        # Procento = (new - old) / |old| — abs() v jmenovateli zajistí, že znaménko
        # procentního rozdílu odpovídá znaménku Kč rozdílu i pro zápornou základnu
        # (např. Total ČB: 2025=-1.2M, 2026=-5.7M, rozdíl=-4.5M → bez abs() by se
        # vyšlo +365 % a CF by buňku obarvilo zeleně, ačkoli se výsledek zhoršil).
        for col in diff_metriky:
            old_v = _s(row.get(f'{col}_{rok_old}'))
            new_v = _s(row.get(f'{col}_{rok_new}'))
            row[f'{col}_diff_kc'] = new_v - old_v
            row[f'{col}_diff_pct'] = ((new_v - old_v) / abs(old_v)) if old_v else None
        old_nk = row.get(f'naklad_kg_{rok_old}')
        new_nk = row.get(f'naklad_kg_{rok_new}')
        if old_nk is not None and new_nk is not None:
            row['naklad_kg_diff'] = new_nk - old_nk
            row['naklad_kg_diff_pct'] = ((new_nk - old_nk) / abs(old_nk)) if old_nk else None
        else:
            row['naklad_kg_diff'] = None
            row['naklad_kg_diff_pct'] = None

    rows: list[dict] = []
    for src in s_rows:
        row = dict(src)
        pob_db = excel_to_db.get(row.get('_pobocka'), row.get('_pobocka'))
        for r in (rok_old, rok_new):
            row[f'naklad_kg_{r}'] = nk_per_pob.get((pob_db, r))
        _op_celkem(row)
        _diff(row)
        rows.append(row)

    total = dict(s_total or {})
    if total:
        for r in (rok_old, rok_new):
            total[f'naklad_kg_{r}'] = nk_per_rok.get(r)
        _op_celkem(total)
        _diff(total)
        total['_je_celkem'] = True
        total.setdefault('_pobocka', 'Celkový součet')

    return rows, total, rok_old, rok_new


# Barevná mapa: tmavší = starší rok, světlejší = novější rok
_GRAF_BARVY: dict[str, tuple[str, str]] = {
    'mzdy':               ('#2E7D32', '#66BB6A'),   # zelená
    'naklady':            ('#00695C', '#4DB6AC'),   # teal
    'total':              ('#E65100', '#FFB74D'),   # oranžová
    'obrat_kg_celkem':    ('#4527A0', '#9575CD'),   # fialová
    'vysledek':           ('#1565C0', '#64B5F6'),   # modrá
    'obrat_kg':           ('#283593', '#7986CB'),   # indigo
    'obrat_kg_dcery':     ('#00838F', '#4DD0E1'),   # cyan
    'obrat_kg_site':      ('#558B2F', '#AED581'),   # limetková
    'obrat_kg_staropramen': ('#6A1B9A', '#CE93D8'), # purpurová
    'repre_spotreba':     ('#AD1457', '#F06292'),   # růžová
}

def _graf_barva(db: str, rok: int, roky: list[int]) -> str:
    """Vrátí barvu pro metriku a rok (tmavá pro starší, světlá pro novější)."""
    dark, light = _GRAF_BARVY.get(db, ('#546E7A', '#90A4AE'))
    return dark if rok == roky[0] else light


def _graf_mesicni_option(nadpis: str, metriky: list[tuple[str, str]],
                         mes: dict, roky: list[int], typ: str = 'line',
                         mesic_od: int = 1, mesic_do: int = 12) -> dict:
    """ECharts: osa X = měsíce, série = metrika × rok (součet vybraných poboček)."""
    rozsah = list(range(mesic_od, mesic_do + 1))
    series = []
    for db, label in metriky:
        for rok in roky:
            series.append({
                'name': f'{label} {rok}',
                'type': typ,
                'smooth': typ == 'line',
                'emphasis': {'focus': 'series'},
                'itemStyle': {'color': _graf_barva(db, rok, roky)},
                'lineStyle': {'color': _graf_barva(db, rok, roky)} if typ == 'line' else {},
                'data': [round(_s((mes.get((rok, m)) or {}).get(db)), 0) for m in rozsah],
            })
    return {
        'title': {'text': nadpis, 'left': 'center', 'top': 8, 'textStyle': {'fontSize': 14}},
        'tooltip': {'trigger': 'axis'},
        'legend': {'type': 'scroll', 'orient': 'vertical', 'right': 0, 'top': 40, 'bottom': 20},
        'grid': {'top': 40, 'left': 72, 'right': 180, 'bottom': 36},
        'xAxis': {'type': 'category', 'data': [MESICE_NAZVY[m - 1] for m in rozsah]},
        'yAxis': {'type': 'value'},
        'series': series,
    }


# Paleta barev pro linky jednotlivých poboček (dost odlišných odstínů)
_GRAF_POBOCKY_PALETA = [
    '#1565C0', '#E65100', '#2E7D32', '#6A1B9A', '#00838F', '#C62828',
    '#558B2F', '#4527A0', '#AD1457', '#00695C', '#EF6C00', '#283593',
    '#9E9D24', '#5D4037', '#0277BD', '#7B1FA2', '#D84315', '#316B83',
]


def _graf_naklad_kg_pobocky_option(nadpis: str, perpob: dict, pobocky: list[str],
                                   mesic_od: int = 1, mesic_do: int = 12) -> dict:
    """ECharts spojnicový: osa X = měsíce, jedna čára na pobočku.
    Náklad na 1 Kg = (Mzdy + Náklady) / Obrat v Kg celkem za daný rok."""
    rozsah = list(range(mesic_od, mesic_do + 1))
    series = []
    for i, pob in enumerate(pobocky):
        mp = perpob.get(pob, {})
        data = [round(mp[m], 2) if mp.get(m) is not None else None for m in rozsah]
        nazev = _POBOCKY_EXCEL_REVERSE.get(pob, pob) if pob not in ('Bourárna', 'Kamiony') else pob
        clr = _GRAF_POBOCKY_PALETA[i % len(_GRAF_POBOCKY_PALETA)]
        series.append({
            'name': nazev,
            'type': 'line',
            'smooth': True,
            'connectNulls': True,
            'emphasis': {'focus': 'series'},
            'itemStyle': {'color': clr},
            'lineStyle': {'color': clr},
            'data': data,
        })
    return {
        'title': {'text': nadpis, 'left': 'center', 'top': 8, 'textStyle': {'fontSize': 14}},
        'tooltip': {'trigger': 'axis'},
        'legend': {'type': 'scroll', 'orient': 'vertical', 'right': 0, 'top': 40, 'bottom': 20},
        'grid': {'top': 40, 'left': 72, 'right': 200, 'bottom': 36},
        'xAxis': {'type': 'category', 'data': [MESICE_NAZVY[m - 1] for m in rozsah]},
        'yAxis': {'type': 'value'},
        'series': series,
    }


def _graf_naklad_kg_porovnani_option(nadpis: str, perpob_old: dict, perpob_new: dict,
                                     rok_old: int, rok_new: int, pobocky: list[str],
                                     mesic_od: int = 1, mesic_do: int = 12) -> dict:
    """ECharts spojnicový: osa X = pobočky, dvě čáry = PRŮMĚRNÝ Náklad na 1 Kg
    za rok_old a rok_new přes zvolený rozsah měsíců (porovnání let)."""
    rozsah = list(range(mesic_od, mesic_do + 1))

    def _prumer(mp: dict):
        vals = [mp[m] for m in rozsah if mp.get(m) is not None]
        return round(sum(vals) / len(vals), 2) if vals else None

    nazvy = [pob if pob in ('Bourárna', 'Kamiony') else _POBOCKY_EXCEL_REVERSE.get(pob, pob)
             for pob in pobocky]
    data_old = [_prumer(perpob_old.get(pob, {})) for pob in pobocky]
    data_new = [_prumer(perpob_new.get(pob, {})) for pob in pobocky]

    def _serie(nazev, data, barva, pozice):
        return {
            'name': nazev, 'type': 'line', 'smooth': False, 'connectNulls': True,
            'symbol': 'circle', 'symbolSize': 6, 'emphasis': {'focus': 'series'},
            'itemStyle': {'color': barva}, 'lineStyle': {'color': barva, 'width': 2},
            'label': {'show': True, 'fontSize': 9, 'position': pozice,
                      'color': barva, 'formatter': '{c}'},
            'data': data,
        }

    series = []
    if rok_old != rok_new:
        series.append(_serie(f'Průměr z Náklad na 1 Kg {rok_old}', data_old, '#2563eb', 'bottom'))
    series.append(_serie(f'Průměr z Náklad na 1 Kg {rok_new}', data_new, '#ea580c', 'top'))

    return {
        'title': {'text': nadpis, 'left': 'center', 'top': 8, 'textStyle': {'fontSize': 14}},
        'tooltip': {'trigger': 'axis'},
        'legend': {'top': 30, 'left': 'center'},
        'grid': {'top': 70, 'left': 60, 'right': 30, 'bottom': 96},
        'xAxis': {'type': 'category', 'data': nazvy,
                  'axisLabel': {'rotate': 35, 'fontSize': 10, 'interval': 0}},
        'yAxis': {'type': 'value', 'name': 'Kč / Kg'},
        'series': series,
    }


def _vykresli_prehled(user_id: int, user_name: str, pristupne_pobocky: list[str], on_open, je_ao: bool = False):
    """Přehled pro vedení – záložky: Grafické / Souhrnné / Mezisoučty / Obraty Kg+Operace / Přehled AO."""
    roky_all = nacti_dostupne_roky()
    if not roky_all:
        with ui.column().classes('items-center py-20 gap-3'):
            ui.icon('insights', size='5rem', color='grey-3')
            ui.label('Žádná data').classes('text-xl font-semibold text-gray-400')
            ui.label('Zatím nebyla nahrána žádná data Tabulky nákladů.') \
                .classes('text-sm text-gray-400 text-center')
        return
    roky = sorted(roky_all[:2])  # dva nejnovější roky (starší → novější)
    rok_old, rok_new = roky[0], roky[-1]

    ui.add_css(_PREHLED_ZOOM_CSS)

    # ── Záložky ───────────────────────────────────────────────────────────────
    # Aktivní záložka se ukládá → po obnovení stránky zůstane uživatel ve své sekci.
    _prehled_tab_key = f'vysledky_prehled_tab_{user_id}'
    _prehled_taby = ('grafy', 'souhrn', 'mezisoucty', 'obraty_operace', 'ao', 'podrobne')
    _prehled_tab_init = app.storage.user.get(_prehled_tab_key)
    if _prehled_tab_init not in _prehled_taby:
        _prehled_tab_init = 'grafy'

    with ui.tabs().props(
        'align=left active-color=primary indicator-color=primary'
    ).classes('w-full border-b border-gray-200 prehled-tabs') as tabs_prehled:
        ui.tab('grafy',       label='📊 Grafické zobrazení')
        ui.tab('souhrn',      label='📋 Souhrnné zobrazení')
        ui.tab('mezisoucty',  label='📊 Mezisoučty zobrazení')
        ui.tab('obraty_operace', label='⚖️ Obraty v Kg a Operace')
        ui.tab('ao',          label='🧮 Přehled AO')
        ui.tab('podrobne',    label='🔍 Podrobné náklady')
    tabs_prehled.on_value_change(
        lambda e: app.storage.user.update({_prehled_tab_key: e.value}))

    with ui.tab_panels(tabs_prehled, value=_prehled_tab_init).classes('w-full pt-4 prehled-zoom'):

        # ── 1) Grafické zobrazení (stávající grafy s filtrem) ─────────────────
        with ui.tab_panel('grafy'):
            # ── Sdílený default grafů (nastavuje AO) ─────────────────────────
            gdk_pob  = 'prehled_graf_default_pob'
            gdk_od   = 'prehled_graf_default_od'
            gdk_do   = 'prehled_graf_default_do'
            gdk_ver  = 'prehled_graf_default_ver'
            gd_pob   = app.storage.general.get(gdk_pob)
            gd_od    = app.storage.general.get(gdk_od)
            gd_do    = app.storage.general.get(gdk_do)
            gd_ver   = app.storage.general.get(gdk_ver, 0)

            fk = f'vysledky_prehled_pobocka_{user_id}'
            fk_od = f'vysledky_prehled_mesic_od_{user_id}'
            fk_do = f'vysledky_prehled_mesic_do_{user_id}'
            fk_gver = f'prehled_graf_dver_{user_id}'

            # Filtr poboček grafů je seznam (multi-select). Prázdný seznam =
            # všechny přístupné pobočky (součet). Podporuje i starou uloženou
            # hodnotu (jediná pobočka jako řetězec).
            def _na_seznam_pob(raw):
                if isinstance(raw, str):
                    raw = [raw]
                if not isinstance(raw, (list, tuple, set)):
                    raw = []
                return [p for p in pristupne_pobocky if p in raw]

            gd_pob_list = _na_seznam_pob(gd_pob)

            # Při nové verzi defaultu → přepsat user filtr
            user_gver = app.storage.user.get(fk_gver, 0)
            if user_gver < gd_ver:
                app.storage.user[fk_gver] = gd_ver
                if gd_pob is not None:
                    app.storage.user[fk] = list(gd_pob_list)
                if gd_od is not None:
                    app.storage.user[fk_od] = gd_od
                if gd_do is not None:
                    app.storage.user[fk_do] = gd_do

            # Normalizuj uložený filtr na seznam přístupných poboček (prázdné = vše)
            app.storage.user[fk] = _na_seznam_pob(app.storage.user.get(fk))
            if app.storage.user.get(fk_od) not in range(1, 13):
                app.storage.user[fk_od] = 1
            if app.storage.user.get(fk_do) not in range(1, 13):
                app.storage.user[fk_do] = 12
            mesice_options = {i: str(i) for i in range(1, 13)}

            @ui.refreshable
            def _grafy():
                vyb = _na_seznam_pob(app.storage.user.get(fk))
                sel = vyb or list(pristupne_pobocky)   # prázdné = všechny pobočky
                m_od = app.storage.user.get(fk_od, 1)
                m_do = app.storage.user.get(fk_do, 12)
                if m_od > m_do:
                    m_od, m_do = m_do, m_od
                mes = nacti_graf_mesicni(roky, sel)
                vsechny = (not vyb) or len(vyb) == len(pristupne_pobocky)
                pob_txt = ('všechny pobočky' if vsechny else ', '.join(
                    _POBOCKY_EXCEL_REVERSE.get(p, p) if p not in ('Bourárna', 'Kamiony') else p
                    for p in vyb))
                ui.label(
                    f'Pobočky: {pob_txt} · roky {" vs ".join(str(r) for r in roky)}'
                ).classes('text-xs text-gray-400 italic mb-2')
                def _chart(opt):
                    ui.echart(opt).classes('w-full').style('height: 380px')
                with ui.grid(columns=2).classes('w-full gap-4'):
                    _chart(_graf_mesicni_option(
                        'Mzdy / Náklady / Total / Obrat v Kg (měsíčně)',
                        [('mzdy', 'Mzdy'), ('naklady', 'Náklady'), ('total', 'Total'),
                         ('obrat_kg_celkem', 'Obrat v Kg celkem')], mes, roky, typ='bar',
                        mesic_od=m_od, mesic_do=m_do))
                    _chart(_graf_mesicni_option(
                        'Total / Výsledek / Náklady / Mzdy (měsíčně)',
                        [('total', 'Total'), ('vysledek', 'Výsledek'),
                         ('naklady', 'Náklady'), ('mzdy', 'Mzdy')], mes, roky, typ='line',
                        mesic_od=m_od, mesic_do=m_do))
                    _chart(_graf_mesicni_option(
                        'Obrat v Kg: celkem / dcery / sítě / Staropramen (měsíčně)',
                        [('obrat_kg', 'Obrat v Kg'), ('obrat_kg_dcery', 'Obrat v Kg dcery'),
                         ('obrat_kg_site', 'Obrat v Kg sítě'),
                         ('obrat_kg_staropramen', 'Obrat v Kg Staropramen')], mes, roky, typ='line',
                        mesic_od=m_od, mesic_do=m_do))
                    _chart(_graf_mesicni_option(
                        'Repre + spotřeba (měsíčně)',
                        [('repre_spotreba', 'Repre + spotřeba')], mes, roky, typ='bar',
                        mesic_od=m_od, mesic_do=m_do))

            with ui.row().classes('items-center gap-3 mb-3 flex-wrap'):
                # AO ovládání výchozího filtru – refreshable, aby se zámek/štítek
                # a přepnutí tlačítka projevily hned (bez obnovení stránky).
                if je_ao:
                    @ui.refreshable
                    def _graf_default_ovladani():
                        cur_pob = _na_seznam_pob(app.storage.user.get(fk))
                        cur_od  = app.storage.user.get(fk_od, 1)
                        cur_do  = app.storage.user.get(fk_do, 12)
                        d_pob   = app.storage.general.get(gdk_pob)
                        d_od    = app.storage.general.get(gdk_od)
                        d_do    = app.storage.general.get(gdk_do)
                        d_ver   = app.storage.general.get(gdk_ver, 0)
                        d_pob_list = _na_seznam_pob(d_pob)
                        d_set   = d_ver > 0 and d_pob is not None
                        je_def  = (d_set and d_pob_list == cur_pob
                                   and d_od == cur_od and d_do == cur_do)

                        def _set_graf_default(_):
                            nv = app.storage.general.get(gdk_ver, 0) + 1
                            app.storage.general[gdk_pob] = _na_seznam_pob(app.storage.user.get(fk))
                            app.storage.general[gdk_od]  = app.storage.user.get(fk_od, 1)
                            app.storage.general[gdk_do]  = app.storage.user.get(fk_do, 12)
                            app.storage.general[gdk_ver] = nv
                            app.storage.user[fk_gver]    = nv
                            ui.notify('Výchozí filtr grafů nastaven pro všechny.', type='positive')
                            _graf_default_ovladani.refresh()

                        def _clear_graf_default(_):
                            nv = app.storage.general.get(gdk_ver, 0) + 1
                            app.storage.general[gdk_pob] = None
                            app.storage.general[gdk_od]  = None
                            app.storage.general[gdk_do]  = None
                            app.storage.general[gdk_ver] = nv
                            app.storage.user[fk_gver]    = nv
                            ui.notify('Výchozí filtr grafů odstraněn.', type='info')
                            _graf_default_ovladani.refresh()

                        if not je_def:
                            ui.button('Nastavit jako výchozí', icon='lock',
                                      on_click=_set_graf_default) \
                                .props('flat dense color=primary').classes('text-xs') \
                                .tooltip('Uloží aktuální filtr jako výchozí pro všechny uživatele')
                        if d_set:
                            _gd_txt = ('všechny pobočky'
                                       if (not d_pob_list) or len(d_pob_list) == len(pristupne_pobocky)
                                       else ', '.join(
                                           _POBOCKY_EXCEL_REVERSE.get(p, p) if p not in ('Bourárna', 'Kamiony') else p
                                           for p in d_pob_list))
                            ui.label(f'🔒 Výchozí: {_gd_txt} · měs. {d_od}–{d_do}') \
                                .classes('text-xs text-blue-500 italic')
                            if je_def:
                                ui.button('Zrušit výchozí', icon='lock_open',
                                          on_click=_clear_graf_default) \
                                    .props('flat dense color=grey-6').classes('text-xs')

                def _refresh_graf_extras():
                    if je_ao:
                        _graf_default_ovladani.refresh()

                sel_pob = ui.select(
                    pristupne_pobocky,
                    value=_na_seznam_pob(app.storage.user.get(fk)),
                    label='Pobočky', multiple=True, with_input=True, clearable=True,
                ).props('dense outlined options-dense').style('min-width: 280px') \
                 .tooltip('Vyberte jednu či více poboček. Prázdné = všechny pobočky (součet).')
                if not _na_seznam_pob(app.storage.user.get(fk)):
                    sel_pob.props('display-value="Všechny pobočky"')
                sel_od = ui.select(
                    mesice_options, value=app.storage.user.get(fk_od, 1), label='Od měsíce',
                ).props('dense outlined options-dense').style('min-width: 100px')
                sel_do = ui.select(
                    mesice_options, value=app.storage.user.get(fk_do, 12), label='Do měsíce',
                ).props('dense outlined options-dense').style('min-width: 100px')

                def _zmena(e):
                    vals = _na_seznam_pob(e.value)
                    app.storage.user[fk] = vals
                    if vals:
                        sel_pob.props(remove='display-value')
                    else:
                        sel_pob.props('display-value="Všechny pobočky"')
                    _grafy.refresh()
                    _refresh_graf_extras()
                sel_pob.on_value_change(_zmena)
                def _zmena_od(e):
                    app.storage.user[fk_od] = e.value
                    _grafy.refresh()
                    _refresh_graf_extras()
                sel_od.on_value_change(_zmena_od)
                def _zmena_do(e):
                    app.storage.user[fk_do] = e.value
                    _grafy.refresh()
                    _refresh_graf_extras()
                sel_do.on_value_change(_zmena_do)

                if je_ao:
                    _graf_default_ovladani()

            _grafy()

            # ── Náklad na 1 Kg (všechny pobočky) – VLASTNÍ měsíční filtr + výchozí ──
            nkdk_od, nkdk_do, nkdk_ver = (
                'prehled_nk_default_od', 'prehled_nk_default_do', 'prehled_nk_default_ver')
            fk_nk_od   = f'prehled_nk_mesic_od_{user_id}'
            fk_nk_do   = f'prehled_nk_mesic_do_{user_id}'
            fk_nk_gver = f'prehled_nk_dver_{user_id}'

            nk_def_ver = app.storage.general.get(nkdk_ver, 0)
            if app.storage.user.get(fk_nk_gver, 0) < nk_def_ver:
                app.storage.user[fk_nk_gver] = nk_def_ver
                _nod = app.storage.general.get(nkdk_od)
                _ndo = app.storage.general.get(nkdk_do)
                if _nod is not None:
                    app.storage.user[fk_nk_od] = _nod
                if _ndo is not None:
                    app.storage.user[fk_nk_do] = _ndo
            if app.storage.user.get(fk_nk_od) not in range(1, 13):
                app.storage.user[fk_nk_od] = 1
            if app.storage.user.get(fk_nk_do) not in range(1, 13):
                app.storage.user[fk_nk_do] = 12

            # Náklad na 1 Kg se nepočítá pro provozy bez obratu v Kg
            _nk_pobocky = [p for p in pristupne_pobocky
                           if p not in ('Bourárna', 'Kamiony', 'Přefakturace')]

            @ui.refreshable
            def _graf_nk():
                m_od = app.storage.user.get(fk_nk_od, 1)
                m_do = app.storage.user.get(fk_nk_do, 12)
                if m_od > m_do:
                    m_od, m_do = m_do, m_od
                perpob_new = nacti_graf_naklad_kg_pobocky(rok_new, _nk_pobocky)
                perpob_old = nacti_graf_naklad_kg_pobocky(rok_old, _nk_pobocky)
                ui.echart(_graf_naklad_kg_porovnani_option(
                    f'Náklad na 1 Kg = (Mzdy + Náklady) / Obrat v Kg celkem · '
                    f'průměr {rok_old} vs {rok_new} · po pobočkách (měs. {m_od}–{m_do})',
                    perpob_old, perpob_new, rok_old, rok_new,
                    _nk_pobocky, mesic_od=m_od, mesic_do=m_do
                )).classes('w-full').style('height: 460px')

            ui.separator().classes('mt-4 mb-2')
            with ui.row().classes('items-center gap-3 mb-2 flex-wrap'):
                ui.label('Náklad na 1 Kg (všechny pobočky):') \
                    .classes('text-sm font-semibold text-gray-700')
                # AO ovládání výchozího zobrazení – vykreslí se na konci řádku
                if je_ao:
                    @ui.refreshable
                    def _nk_ovladani():
                        cur_od = app.storage.user.get(fk_nk_od, 1)
                        cur_do = app.storage.user.get(fk_nk_do, 12)
                        dod = app.storage.general.get(nkdk_od)
                        ddo = app.storage.general.get(nkdk_do)
                        dver = app.storage.general.get(nkdk_ver, 0)
                        je_def = (dver > 0 and dod is not None and cur_od == dod and cur_do == ddo)

                        def _set_nk_def(_):
                            nv = app.storage.general.get(nkdk_ver, 0) + 1
                            app.storage.general[nkdk_od]  = app.storage.user.get(fk_nk_od, 1)
                            app.storage.general[nkdk_do]  = app.storage.user.get(fk_nk_do, 12)
                            app.storage.general[nkdk_ver] = nv
                            app.storage.user[fk_nk_gver]  = nv
                            ui.notify('Výchozí zobrazení grafu Náklad na 1 Kg nastaveno pro všechny.',
                                      type='positive')
                            _nk_ovladani.refresh()

                        def _clear_nk_def(_):
                            nv = app.storage.general.get(nkdk_ver, 0) + 1
                            app.storage.general[nkdk_od]  = None
                            app.storage.general[nkdk_do]  = None
                            app.storage.general[nkdk_ver] = nv
                            app.storage.user[fk_nk_gver]  = nv
                            ui.notify('Výchozí zobrazení grafu Náklad na 1 Kg odstraněno.', type='info')
                            _nk_ovladani.refresh()

                        if not je_def:
                            ui.button('Nastavit jako výchozí', icon='lock', on_click=_set_nk_def) \
                                .props('flat dense color=primary').classes('text-xs') \
                                .tooltip('Uloží rozsah měsíců jako výchozí zobrazení tohoto grafu pro všechny')
                        if dver > 0 and dod is not None:
                            ui.label(f'🔒 Výchozí: měs. {dod}–{ddo}') \
                                .classes('text-xs text-blue-500 italic')
                            if je_def:
                                ui.button('Zrušit výchozí', icon='lock_open', on_click=_clear_nk_def) \
                                    .props('flat dense color=grey-6').classes('text-xs')

                def _refresh_nk_extras():
                    if je_ao:
                        _nk_ovladani.refresh()

                nk_od = ui.select(
                    mesice_options, value=app.storage.user.get(fk_nk_od, 1), label='Od měsíce',
                ).props('dense outlined options-dense').style('min-width: 100px')
                nk_do = ui.select(
                    mesice_options, value=app.storage.user.get(fk_nk_do, 12), label='Do měsíce',
                ).props('dense outlined options-dense').style('min-width: 100px')

                def _zm_nk_od(e):
                    app.storage.user[fk_nk_od] = e.value
                    _graf_nk.refresh()
                    _refresh_nk_extras()
                nk_od.on_value_change(_zm_nk_od)
                def _zm_nk_do(e):
                    app.storage.user[fk_nk_do] = e.value
                    _graf_nk.refresh()
                    _refresh_nk_extras()
                nk_do.on_value_change(_zm_nk_do)

                if je_ao:
                    _nk_ovladani()
            _graf_nk()

        # ── 2) Souhrnné zobrazení ────────────────────────────────────────────
        with ui.tab_panel('souhrn'):
            fk_sp   = f'prehled_souhrn_pobocky_{user_id}'
            fk_sr   = f'prehled_souhrn_roky_{user_id}'
            fk_sm   = f'prehled_souhrn_mesice_{user_id}'
            fk_sver = f'prehled_souhrn_dver_{user_id}'
            sdk_roky, sdk_mesice, sdk_ver = (
                'prehled_souhrn_default_roky', 'prehled_souhrn_default_mesice',
                'prehled_souhrn_default_ver')

            # Sdílený výchozí filtr (rok + měsíc) – při nové verzi přepíše uživatelův
            sd_ver = app.storage.general.get(sdk_ver, 0)
            if app.storage.user.get(fk_sver, 0) < sd_ver:
                app.storage.user[fk_sver] = sd_ver
                _sdr = app.storage.general.get(sdk_roky)
                _sdm = app.storage.general.get(sdk_mesice)
                if _sdr is not None:
                    app.storage.user[fk_sr] = list(_sdr)
                if _sdm is not None:
                    app.storage.user[fk_sm] = list(_sdm)

            sel_sp = app.storage.user.get(fk_sp) or []

            @ui.refreshable
            def _souhrn_grid():
                vyb = app.storage.user.get(fk_sp) or []
                pob = [p for p in pristupne_pobocky if p in vyb] if vyb else pristupne_pobocky
                sel_r = [r for r in (app.storage.user.get(fk_sr) or []) if r in roky_all]
                eff_roky = sorted(sel_r) if sel_r else roky  # výchozí = 2 nejnovější
                sel_m = sorted(m for m in (app.storage.user.get(fk_sm) or []) if 1 <= m <= 12)
                ui.label(
                    'Roky: ' + ', '.join(str(r) for r in eff_roky)
                    + ' · Měsíce: ' + ('celý rok' if not sel_m
                                        else ', '.join(str(m) for m in sel_m))
                ).classes('text-xs text-gray-400 italic mb-1')
                s_rows, s_total = nacti_prehled_souhrn(eff_roky, pob, sel_m)
                if not s_rows:
                    ui.label('Žádná data pro souhrnné zobrazení.').classes('text-gray-400 italic py-4')
                    return
                ui.add_css(_PREHLED_GRID_CSS)
                ui.aggrid({
                    'columnDefs': _col_defs_prehled_souhrn(eff_roky),
                    'rowData': s_rows,
                    'pinnedBottomRowData': [s_total] if s_total else [],
                    'defaultColDef': {'resizable': True, 'sortable': True, 'wrapHeaderText': True, 'autoHeaderHeight': True},
                    'rowHeight': 34,
                    'suppressMovableColumns': True,
                    'domLayout': 'autoHeight',
                    ':onFirstDataRendered': _AUTOSIZE_FIT,
                    ':onGridSizeChanged': _AUTOSIZE_FIT,
                    ':getRowStyle': (
                        "function(p){"
                        "if(p.data&&p.data._je_celkem)"
                        "return{fontWeight:'bold',backgroundColor:'#DCE6F2',"
                        "borderTop:'1px solid #b8cce4'};"
                        "return null;}"
                    ),
                }).classes('w-full prehled-grid')

            pob_opts = {p: _POBOCKY_EXCEL_REVERSE.get(p, p) for p in pristupne_pobocky}
            rok_opts = {r: str(r) for r in roky_all}
            mes_opts = {i: str(i) for i in range(1, 13)}
            with ui.row().classes('items-center gap-3 mb-3 flex-wrap'):
                # AO ovládání výchozího zobrazení (rok + měsíc) – vykreslí se na konci řádku
                if je_ao:
                    @ui.refreshable
                    def _souhrn_default_ovladani():
                        cur_r = list(app.storage.user.get(fk_sr) or [])
                        cur_m = list(app.storage.user.get(fk_sm) or [])
                        dr = app.storage.general.get(sdk_roky)
                        dm = app.storage.general.get(sdk_mesice)
                        dv = app.storage.general.get(sdk_ver, 0)
                        je_def = (dv > 0 and (dr or dm)
                                  and cur_r == list(dr or []) and cur_m == list(dm or []))

                        def _set_def(_):
                            nv = app.storage.general.get(sdk_ver, 0) + 1
                            app.storage.general[sdk_roky]   = list(app.storage.user.get(fk_sr) or [])
                            app.storage.general[sdk_mesice] = list(app.storage.user.get(fk_sm) or [])
                            app.storage.general[sdk_ver]    = nv
                            app.storage.user[fk_sver]       = nv
                            ui.notify('Výchozí zobrazení (rok + měsíc) nastaveno pro všechny.', type='positive')
                            _souhrn_default_ovladani.refresh()

                        def _clear_def(_):
                            nv = app.storage.general.get(sdk_ver, 0) + 1
                            app.storage.general[sdk_roky]   = None
                            app.storage.general[sdk_mesice] = None
                            app.storage.general[sdk_ver]    = nv
                            app.storage.user[fk_sver]       = nv
                            ui.notify('Výchozí zobrazení odstraněno.', type='info')
                            _souhrn_default_ovladani.refresh()

                        if not je_def:
                            ui.button('Nastavit jako výchozí', icon='lock', on_click=_set_def) \
                                .props('flat dense color=primary').classes('text-xs') \
                                .tooltip('Uloží rok + měsíc jako výchozí zobrazení pro všechny uživatele')
                        if dv > 0 and (dr or dm):
                            _parts = []
                            if dr:
                                _parts.append('rok ' + ', '.join(str(r) for r in dr))
                            if dm:
                                _parts.append('měs. ' + ', '.join(str(m) for m in dm))
                            ui.label('🔒 Výchozí: ' + ' · '.join(_parts)) \
                                .classes('text-xs text-blue-500 italic')
                            if je_def:
                                ui.button('Zrušit výchozí', icon='lock_open', on_click=_clear_def) \
                                    .props('flat dense color=grey-6').classes('text-xs')

                def _refresh_souhrn_extras():
                    if je_ao:
                        _souhrn_default_ovladani.refresh()

                w_sr = ui.select(
                    rok_opts, value=list(app.storage.user.get(fk_sr) or []), label='Rok',
                    multiple=True,
                ).props('dense outlined options-dense use-chips').style('min-width: 160px') \
                 .tooltip('Prázdné = 2 nejnovější roky')
                w_sm = ui.select(
                    mes_opts, value=list(app.storage.user.get(fk_sm) or []), label='Měsíce',
                    multiple=True,
                ).props('dense outlined options-dense use-chips').style('min-width: 160px') \
                 .tooltip('Prázdné = celý rok')
                w_sp = ui.select(
                    pob_opts, value=list(sel_sp), label='Pobočky',
                    multiple=True,
                ).props('dense outlined options-dense use-chips').style('min-width: 280px')

                if (app.storage.user.get(fk_sr) or app.storage.user.get(fk_sm)
                        or app.storage.user.get(fk_sp)):
                    def _reset_souhrn(_):
                        app.storage.user[fk_sr] = []
                        app.storage.user[fk_sm] = []
                        app.storage.user[fk_sp] = []
                        _souhrn_grid.refresh()
                        _refresh_souhrn_extras()
                    ui.button('Vše', icon='clear_all', on_click=_reset_souhrn) \
                        .props('flat dense color=grey-7').classes('text-xs')

                def _zm_sr(e):
                    app.storage.user[fk_sr] = list(e.value) if e.value else []
                    _souhrn_grid.refresh()
                    _refresh_souhrn_extras()
                w_sr.on_value_change(_zm_sr)
                def _zm_sm(e):
                    app.storage.user[fk_sm] = list(e.value) if e.value else []
                    _souhrn_grid.refresh()
                    _refresh_souhrn_extras()
                w_sm.on_value_change(_zm_sm)
                def _zm_sp(e):
                    app.storage.user[fk_sp] = list(e.value) if e.value else []
                    _souhrn_grid.refresh()
                w_sp.on_value_change(_zm_sp)

                if je_ao:
                    _souhrn_default_ovladani()
            _souhrn_grid()

        # ── 3) Mezisoučty zobrazení (rozpad po měsících) ─────────────────────
        with ui.tab_panel('mezisoucty'):
            fk_mp   = f'prehled_mezisoucty_pobocky_{user_id}'
            fk_mr   = f'prehled_mezisoucty_roky_{user_id}'
            fk_mm   = f'prehled_mezisoucty_mesice_{user_id}'
            fk_mver = f'prehled_mez_dver_{user_id}'
            mdk_roky, mdk_mesice, mdk_ver = (
                'prehled_mez_default_roky', 'prehled_mez_default_mesice', 'prehled_mez_default_ver')

            # Sdílený výchozí filtr (rok + měsíc) – při nové verzi přepíše uživatelův
            md_ver = app.storage.general.get(mdk_ver, 0)
            if app.storage.user.get(fk_mver, 0) < md_ver:
                app.storage.user[fk_mver] = md_ver
                _dr = app.storage.general.get(mdk_roky)
                _dm = app.storage.general.get(mdk_mesice)
                if _dr is not None:
                    app.storage.user[fk_mr] = list(_dr)
                if _dm is not None:
                    app.storage.user[fk_mm] = list(_dm)

            sel_mp = app.storage.user.get(fk_mp) or []

            @ui.refreshable
            def _mezisoucty_grid():
                vyb = app.storage.user.get(fk_mp) or []
                pob = [p for p in pristupne_pobocky if p in vyb] if vyb else pristupne_pobocky
                sel_r = [r for r in (app.storage.user.get(fk_mr) or []) if r in roky_all]
                eff_roky = sorted(sel_r) if sel_r else roky
                sel_m = sorted(m for m in (app.storage.user.get(fk_mm) or []) if 1 <= m <= 12)
                ui.label(
                    'Roky: ' + ', '.join(str(r) for r in eff_roky)
                    + ' · Měsíce: ' + ('všech 12' if not sel_m
                                        else ', '.join(str(m) for m in sel_m))
                ).classes('text-xs text-gray-400 italic mb-1')
                m_rows, m_total = nacti_prehled_mezisoucty(eff_roky, pob, sel_m)
                if not m_rows:
                    ui.label('Žádná data pro mezisoučty.').classes('text-gray-400 italic py-4')
                    return
                # Bez `domLayout:'autoHeight'` → normální layout s interním scrollem,
                # ag-Grid automaticky udrží hlavičky ukotvené při vertikálním scrollu
                # (Total řádek je pinned bottom, zůstává viditelný také). Pevnou výšku
                # dáme inline – třídu `prehled-grid` (height:auto!important) zde nesmíme
                # použít, jinak by se grid srazil na nulu a nic by se nezobrazilo.
                ui.aggrid({
                    'columnDefs': _col_defs_prehled_mezisoucty(eff_roky),
                    'rowData': m_rows,
                    'pinnedBottomRowData': [m_total] if m_total else [],
                    'defaultColDef': {'resizable': True, 'sortable': False, 'wrapHeaderText': True, 'autoHeaderHeight': True},
                    'rowHeight': 34,
                    'suppressMovableColumns': True,
                    ':onFirstDataRendered': _AUTOSIZE_FIT,
                    ':onGridSizeChanged': _AUTOSIZE_FIT,
                    ':getRowStyle': (
                        "function(p){"
                        "if(p.data&&p.data._je_celkem)"
                        "return{fontWeight:'bold',backgroundColor:'#DCE6F2',"
                        "borderTop:'1px solid #b8cce4'};"
                        "return null;}"
                    ),
                }).classes('w-full prehled-mez').style('height: calc(100vh - 360px); min-height: 440px')

            pob_opts_m = {p: _POBOCKY_EXCEL_REVERSE.get(p, p) for p in pristupne_pobocky}
            rok_opts_m = {r: str(r) for r in roky_all}
            mes_opts_m = {i: str(i) for i in range(1, 13)}
            with ui.row().classes('items-center gap-3 mb-3 flex-wrap'):
                # AO ovládání výchozího zobrazení (rok + měsíc) – vykreslí se na konci řádku
                if je_ao:
                    @ui.refreshable
                    def _mez_default_ovladani():
                        cur_r = list(app.storage.user.get(fk_mr) or [])
                        cur_m = list(app.storage.user.get(fk_mm) or [])
                        dr = app.storage.general.get(mdk_roky)
                        dm = app.storage.general.get(mdk_mesice)
                        dv = app.storage.general.get(mdk_ver, 0)
                        je_def = (dv > 0 and (dr or dm)
                                  and cur_r == list(dr or []) and cur_m == list(dm or []))

                        def _set_def(_):
                            nv = app.storage.general.get(mdk_ver, 0) + 1
                            app.storage.general[mdk_roky]   = list(app.storage.user.get(fk_mr) or [])
                            app.storage.general[mdk_mesice] = list(app.storage.user.get(fk_mm) or [])
                            app.storage.general[mdk_ver]    = nv
                            app.storage.user[fk_mver]       = nv
                            ui.notify('Výchozí zobrazení (rok + měsíc) nastaveno pro všechny.', type='positive')
                            _mez_default_ovladani.refresh()

                        def _clear_def(_):
                            nv = app.storage.general.get(mdk_ver, 0) + 1
                            app.storage.general[mdk_roky]   = None
                            app.storage.general[mdk_mesice] = None
                            app.storage.general[mdk_ver]    = nv
                            app.storage.user[fk_mver]       = nv
                            ui.notify('Výchozí zobrazení odstraněno.', type='info')
                            _mez_default_ovladani.refresh()

                        if not je_def:
                            ui.button('Nastavit jako výchozí', icon='lock', on_click=_set_def) \
                                .props('flat dense color=primary').classes('text-xs') \
                                .tooltip('Uloží rok + měsíc jako výchozí zobrazení pro všechny uživatele')
                        if dv > 0 and (dr or dm):
                            _parts = []
                            if dr:
                                _parts.append('rok ' + ', '.join(str(r) for r in dr))
                            if dm:
                                _parts.append('měs. ' + ', '.join(str(m) for m in dm))
                            ui.label('🔒 Výchozí: ' + ' · '.join(_parts)) \
                                .classes('text-xs text-blue-500 italic')
                            if je_def:
                                ui.button('Zrušit výchozí', icon='lock_open', on_click=_clear_def) \
                                    .props('flat dense color=grey-6').classes('text-xs')

                def _refresh_mez_extras():
                    if je_ao:
                        _mez_default_ovladani.refresh()

                w_mr = ui.select(
                    rok_opts_m, value=list(app.storage.user.get(fk_mr) or []), label='Rok',
                    multiple=True,
                ).props('dense outlined options-dense use-chips').style('min-width: 150px') \
                 .tooltip('Prázdné = 2 nejnovější roky')
                w_mm = ui.select(
                    mes_opts_m, value=list(app.storage.user.get(fk_mm) or []), label='Měsíce',
                    multiple=True,
                ).props('dense outlined options-dense use-chips').style('min-width: 150px') \
                 .tooltip('Prázdné = všech 12 měsíců')
                w_mp = ui.select(
                    pob_opts_m, value=list(sel_mp), label='Pobočky',
                    multiple=True,
                ).props('dense outlined options-dense use-chips').style('min-width: 260px')

                def _zm_mr(e):
                    app.storage.user[fk_mr] = list(e.value) if e.value else []
                    _mezisoucty_grid.refresh()
                    _refresh_mez_extras()
                w_mr.on_value_change(_zm_mr)
                def _zm_mm(e):
                    app.storage.user[fk_mm] = list(e.value) if e.value else []
                    _mezisoucty_grid.refresh()
                    _refresh_mez_extras()
                w_mm.on_value_change(_zm_mm)
                def _zm_mp(e):
                    app.storage.user[fk_mp] = list(e.value) if e.value else []
                    _mezisoucty_grid.refresh()
                w_mp.on_value_change(_zm_mp)

                if je_ao:
                    _mez_default_ovladani()

            # Šipky pro horizontální posun (mnoho sloupců se nevejde na šířku) –
            # scrollují vnitřní horizontální viewport ag-Gridu (třída `prehled-mez`).
            def _mez_scroll(akce: str):
                sel = (".prehled-mez .ag-body-horizontal-scroll-viewport,"
                       ".prehled-mez .ag-center-cols-viewport")
                if akce == 'start':
                    op = "v.scrollTo({left:0,behavior:'smooth'});"
                elif akce == 'end':
                    op = "v.scrollTo({left:v.scrollWidth,behavior:'smooth'});"
                elif akce == 'left':
                    op = "v.scrollBy({left:-320,behavior:'smooth'});"
                else:
                    op = "v.scrollBy({left:320,behavior:'smooth'});"
                ui.run_javascript(
                    f"(function(){{var v=document.querySelector(\"{sel}\");"
                    f"if(v){{{op}}}}})();"
                )

            with ui.row().classes('items-center gap-0.5 mb-1'):
                ui.label('posun:').classes('text-xs text-gray-500 mr-1')
                ui.button(icon='first_page', on_click=lambda: _mez_scroll('start')) \
                    .props('flat dense round size=sm color=dark').tooltip('Začátek')
                ui.button(icon='chevron_left', on_click=lambda: _mez_scroll('left')) \
                    .props('flat dense round size=sm color=dark').tooltip('Posun vlevo')
                ui.button(icon='chevron_right', on_click=lambda: _mez_scroll('right')) \
                    .props('flat dense round size=sm color=dark').tooltip('Posun vpravo')
                ui.button(icon='last_page', on_click=lambda: _mez_scroll('end')) \
                    .props('flat dense round size=sm color=dark').tooltip('Konec')

            _mezisoucty_grid()

        # ── 4) Obraty v Kg a Operace (KT s počítanými poli) ──────────────────
        with ui.tab_panel('obraty_operace'):
            # Jedna nezávislá sekce = tabulka + vlastní filtr + vlastní AO výchozí.
            # Voláno 2× (Obrat v Kg / Operace), aby měla každá tabulka svůj filtr.
            def _obop_sekce(klic: str, nadpis: str, col_defs_fn, diff_fields: tuple):
                fk_r   = f'prehled_obop_{klic}_roky_{user_id}'
                fk_m   = f'prehled_obop_{klic}_mesice_{user_id}'
                fk_p   = f'prehled_obop_{klic}_pobocky_{user_id}'
                fk_ver = f'prehled_obop_{klic}_dver_{user_id}'
                dk_roky   = f'prehled_obop_{klic}_default_roky'
                dk_mesice = f'prehled_obop_{klic}_default_mesice'
                dk_ver    = f'prehled_obop_{klic}_default_ver'

                # Sdílený výchozí filtr (rok + měsíc) – při nové verzi přepíše uživatelův
                sd_ver = app.storage.general.get(dk_ver, 0)
                if app.storage.user.get(fk_ver, 0) < sd_ver:
                    app.storage.user[fk_ver] = sd_ver
                    _dr = app.storage.general.get(dk_roky)
                    _dm = app.storage.general.get(dk_mesice)
                    if _dr is not None:
                        app.storage.user[fk_r] = list(_dr)
                    if _dm is not None:
                        app.storage.user[fk_m] = list(_dm)

                @ui.refreshable
                def _grid():
                    vyb = app.storage.user.get(fk_p) or []
                    pob = [p for p in pristupne_pobocky if p in vyb] if vyb else pristupne_pobocky
                    sel_r = [r for r in (app.storage.user.get(fk_r) or []) if r in roky_all]
                    eff_roky = sorted(sel_r)[-2:] if len(sel_r) >= 2 else roky
                    sel_m = sorted(m for m in (app.storage.user.get(fk_m) or []) if 1 <= m <= 12)
                    rows, total, r_old, r_new = nacti_prehled_obraty_operace(eff_roky, pob, sel_m)
                    ui.label(
                        f'Porovnání: {r_new} × {r_old} · Měsíce: '
                        + ('celý rok' if not sel_m else ', '.join(str(m) for m in sel_m))
                    ).classes('text-xs text-gray-400 italic mb-1')
                    if not rows:
                        ui.label('Žádná data pro zobrazení.').classes('text-gray-400 italic py-4')
                        return
                    ui.add_css(_PREHLED_GRID_CSS)
                    ui.add_css(_OBOP_CSS)
                    vals = [row[k] for row in rows if not row.get('_bez_skaly')
                            for k in diff_fields]
                    style = _obop_colorscale_style(vals)
                    ui.aggrid({
                        'columnDefs': col_defs_fn(r_old, r_new, style),
                        'rowData': rows,
                        'pinnedBottomRowData': [total] if total else [],
                        'defaultColDef': {'resizable': True, 'sortable': True, 'wrapHeaderText': True, 'autoHeaderHeight': True},
                        'rowHeight': 30,
                        'suppressMovableColumns': True,
                        'domLayout': 'autoHeight',
                        ':onFirstDataRendered': _AUTOSIZE_FIT,
                        ':onGridSizeChanged': _AUTOSIZE_FIT,
                    }).classes('w-full prehled-grid obop-grid')

                pob_opts = {p: _POBOCKY_EXCEL_REVERSE.get(p, p) for p in pristupne_pobocky}
                rok_opts = {r: str(r) for r in roky_all}
                mes_opts = {i: str(i) for i in range(1, 13)}
                ui.label(nadpis).classes('text-base font-semibold text-gray-700 mt-1 mb-1')
                with ui.row().classes('items-center gap-3 mb-2 flex-wrap'):
                    # AO ovládání výchozího zobrazení (rok + měsíc) této tabulky
                    if je_ao:
                        @ui.refreshable
                        def _def_ovladani():
                            cur_r = list(app.storage.user.get(fk_r) or [])
                            cur_m = list(app.storage.user.get(fk_m) or [])
                            dr = app.storage.general.get(dk_roky)
                            dm = app.storage.general.get(dk_mesice)
                            dv = app.storage.general.get(dk_ver, 0)
                            je_def = (dv > 0 and (dr or dm)
                                      and cur_r == list(dr or []) and cur_m == list(dm or []))

                            def _set_def(_):
                                nv = app.storage.general.get(dk_ver, 0) + 1
                                app.storage.general[dk_roky]   = list(app.storage.user.get(fk_r) or [])
                                app.storage.general[dk_mesice] = list(app.storage.user.get(fk_m) or [])
                                app.storage.general[dk_ver]    = nv
                                app.storage.user[fk_ver]       = nv
                                ui.notify(f'Výchozí zobrazení „{nadpis}" (rok + měsíc) nastaveno pro všechny.', type='positive')
                                _def_ovladani.refresh()

                            def _clear_def(_):
                                nv = app.storage.general.get(dk_ver, 0) + 1
                                app.storage.general[dk_roky]   = None
                                app.storage.general[dk_mesice] = None
                                app.storage.general[dk_ver]    = nv
                                app.storage.user[fk_ver]       = nv
                                ui.notify(f'Výchozí zobrazení „{nadpis}" odstraněno.', type='info')
                                _def_ovladani.refresh()

                            if not je_def:
                                ui.button('Nastavit jako výchozí', icon='lock', on_click=_set_def) \
                                    .props('flat dense color=primary').classes('text-xs') \
                                    .tooltip('Uloží rok + měsíc jako výchozí zobrazení této tabulky pro všechny uživatele')
                            if dv > 0 and (dr or dm):
                                _parts = []
                                if dr:
                                    _parts.append('rok ' + ', '.join(str(r) for r in dr))
                                if dm:
                                    _parts.append('měs. ' + ', '.join(str(m) for m in dm))
                                ui.label('🔒 Výchozí: ' + ' · '.join(_parts)) \
                                    .classes('text-xs text-blue-500 italic')
                                if je_def:
                                    ui.button('Zrušit výchozí', icon='lock_open', on_click=_clear_def) \
                                        .props('flat dense color=grey-6').classes('text-xs')

                    def _refresh_extras():
                        if je_ao:
                            _def_ovladani.refresh()

                    w_r = ui.select(
                        rok_opts, value=list(app.storage.user.get(fk_r) or []), label='Rok',
                        multiple=True,
                    ).props('dense outlined options-dense use-chips').style('min-width: 160px') \
                     .tooltip('Prázdné = 2 nejnovější roky; porovnávají se vždy 2 nejnovější vybrané')
                    w_m = ui.select(
                        mes_opts, value=list(app.storage.user.get(fk_m) or []), label='Měsíce',
                        multiple=True,
                    ).props('dense outlined options-dense use-chips').style('min-width: 160px') \
                     .tooltip('Prázdné = celý rok')
                    w_p = ui.select(
                        pob_opts, value=list(app.storage.user.get(fk_p) or []), label='Pobočky',
                        multiple=True,
                    ).props('dense outlined options-dense use-chips').style('min-width: 280px')

                    if (app.storage.user.get(fk_r) or app.storage.user.get(fk_m)
                            or app.storage.user.get(fk_p)):
                        def _reset(_):
                            app.storage.user[fk_r] = []
                            app.storage.user[fk_m] = []
                            app.storage.user[fk_p] = []
                            _grid.refresh()
                            _refresh_extras()
                        ui.button('Vše', icon='clear_all', on_click=_reset) \
                            .props('flat dense color=grey-7').classes('text-xs')

                    def _zm_r(e):
                        app.storage.user[fk_r] = list(e.value) if e.value else []
                        _grid.refresh()
                        _refresh_extras()
                    w_r.on_value_change(_zm_r)

                    def _zm_m(e):
                        app.storage.user[fk_m] = list(e.value) if e.value else []
                        _grid.refresh()
                        _refresh_extras()
                    w_m.on_value_change(_zm_m)

                    def _zm_p(e):
                        app.storage.user[fk_p] = list(e.value) if e.value else []
                        _grid.refresh()
                    w_p.on_value_change(_zm_p)

                    if je_ao:
                        _def_ovladani()
                _grid()

            _obop_sekce('kg', 'Obrat v Kg', _col_defs_obop_kg,
                      ('okg_bezsiti_diff', 'okg_site_diff', 'okg_celkem_diff'))
            ui.separator().classes('my-4')
            _obop_sekce('op', 'Operace', _col_defs_obop_op,
                      ('op_bezsiti_diff', 'op_site_diff', 'op_celkem_diff'))

        # ── 5) Přehled AO (porovnání 2 nejnovějších let) ─────────────────────
        with ui.tab_panel('ao'):
            fk_ao_p   = f'prehled_ao_pobocky_{user_id}'
            fk_ao_m   = f'prehled_ao_mesice_{user_id}'
            fk_ao_ver = f'prehled_ao_dver_{user_id}'
            dk_ao_m   = 'prehled_ao_default_mesice'
            dk_ao_ver = 'prehled_ao_default_ver'

            # Sdílený výchozí filtr (jen měsíce) – při nové verzi přepíše uživatelův
            sd_ver_ao = app.storage.general.get(dk_ao_ver, 0)
            if app.storage.user.get(fk_ao_ver, 0) < sd_ver_ao:
                app.storage.user[fk_ao_ver] = sd_ver_ao
                _ddm = app.storage.general.get(dk_ao_m)
                if _ddm is not None:
                    app.storage.user[fk_ao_m] = list(_ddm)

            @ui.refreshable
            def _ao_grid():
                vyb = app.storage.user.get(fk_ao_p) or []
                pob = [p for p in pristupne_pobocky if p in vyb] if vyb else pristupne_pobocky
                sel_m = sorted(m for m in (app.storage.user.get(fk_ao_m) or []) if 1 <= m <= 12)
                ao_rows, ao_total, r_old, r_new = nacti_prehled_ao(roky, pob, sel_m)
                ui.label(
                    f'Porovnání: {r_new} × {r_old} · Měsíce: '
                    + ('celý rok' if not sel_m else ', '.join(str(m) for m in sel_m))
                ).classes('text-xs text-gray-400 italic mb-1')
                if not ao_rows:
                    ui.label('Žádná data pro Přehled AO.').classes('text-gray-400 italic py-4')
                    return
                ui.add_css(_PREHLED_GRID_CSS)
                ui.add_css(_AO_CSS)
                ui.aggrid({
                    'columnDefs': _col_defs_prehled_ao(r_old, r_new),
                    'rowData': ao_rows,
                    'pinnedBottomRowData': [ao_total] if ao_total else [],
                    'defaultColDef': {'resizable': True, 'sortable': True,
                                      'wrapHeaderText': True, 'autoHeaderHeight': True},
                    'rowHeight': 32,
                    'suppressMovableColumns': True,
                    'domLayout': 'autoHeight',
                    ':onFirstDataRendered': _AUTOSIZE_FIT,
                    ':onGridSizeChanged': _AUTOSIZE_FIT,
                }).classes('w-full prehled-grid ao-grid')

            pob_opts = {p: _POBOCKY_EXCEL_REVERSE.get(p, p) for p in pristupne_pobocky}
            mes_opts = {i: str(i) for i in range(1, 13)}
            with ui.row().classes('items-center gap-3 mb-3 flex-wrap'):
                # AO ovládání výchozího filtru měsíců
                if je_ao:
                    @ui.refreshable
                    def _ao_default_ovladani():
                        cur_m = list(app.storage.user.get(fk_ao_m) or [])
                        dm = app.storage.general.get(dk_ao_m)
                        dv = app.storage.general.get(dk_ao_ver, 0)
                        je_def = (dv > 0 and dm and cur_m == list(dm or []))

                        def _set_def(_):
                            nv = app.storage.general.get(dk_ao_ver, 0) + 1
                            app.storage.general[dk_ao_m]   = list(app.storage.user.get(fk_ao_m) or [])
                            app.storage.general[dk_ao_ver] = nv
                            app.storage.user[fk_ao_ver]    = nv
                            ui.notify('Výchozí měsíce „Přehled AO" nastaveny pro všechny.',
                                      type='positive')
                            _ao_default_ovladani.refresh()

                        def _clear_def(_):
                            nv = app.storage.general.get(dk_ao_ver, 0) + 1
                            app.storage.general[dk_ao_m]   = None
                            app.storage.general[dk_ao_ver] = nv
                            app.storage.user[fk_ao_ver]    = nv
                            ui.notify('Výchozí měsíce „Přehled AO" odstraněny.', type='info')
                            _ao_default_ovladani.refresh()

                        if not je_def:
                            ui.button('Nastavit jako výchozí', icon='lock', on_click=_set_def) \
                                .props('flat dense color=primary').classes('text-xs') \
                                .tooltip('Uloží měsíce jako výchozí zobrazení této tabulky pro všechny')
                        if dv > 0 and dm:
                            ui.label('🔒 Výchozí: měs. ' + ', '.join(str(m) for m in dm)) \
                                .classes('text-xs text-blue-500 italic')
                            if je_def:
                                ui.button('Zrušit výchozí', icon='lock_open', on_click=_clear_def) \
                                    .props('flat dense color=grey-6').classes('text-xs')

                def _refresh_ao_extras():
                    if je_ao:
                        _ao_default_ovladani.refresh()

                w_ao_m = ui.select(
                    mes_opts, value=list(app.storage.user.get(fk_ao_m) or []), label='Měsíce',
                    multiple=True,
                ).props('dense outlined options-dense use-chips').style('min-width: 160px') \
                 .tooltip('Prázdné = celý rok')
                w_ao_p = ui.select(
                    pob_opts, value=list(app.storage.user.get(fk_ao_p) or []), label='Pobočky',
                    multiple=True,
                ).props('dense outlined options-dense use-chips').style('min-width: 280px')

                if app.storage.user.get(fk_ao_m) or app.storage.user.get(fk_ao_p):
                    def _reset_ao(_):
                        app.storage.user[fk_ao_m] = []
                        app.storage.user[fk_ao_p] = []
                        _ao_grid.refresh()
                        _refresh_ao_extras()
                    ui.button('Vše', icon='clear_all', on_click=_reset_ao) \
                        .props('flat dense color=grey-7').classes('text-xs')

                def _zm_ao_m(e):
                    app.storage.user[fk_ao_m] = list(e.value) if e.value else []
                    _ao_grid.refresh()
                    _refresh_ao_extras()
                w_ao_m.on_value_change(_zm_ao_m)

                def _zm_ao_p(e):
                    app.storage.user[fk_ao_p] = list(e.value) if e.value else []
                    _ao_grid.refresh()
                w_ao_p.on_value_change(_zm_ao_p)

                if je_ao:
                    _ao_default_ovladani()
            _ao_grid()

        # ── 6) Podrobné náklady (souhrn účetní osnovy přes pobočky) ──────────
        # Meziroční porovnání podrobných nákladů (vysledky_naklady) sečtených
        # přes všechny / vybrané pobočky – VO + Kamiony + Bourárna + Přefakturace
        # (tj. všechny dlaždice kromě skrytých Ostatních provozů).
        with ui.tab_panel('podrobne'):
            fk_pn_p = f'prehled_podrobne_pobocky_{user_id}'
            fk_pn_r = f'prehled_podrobne_rok_{user_id}'

            pn_roky = nacti_vysledky_roky()
            _akt_rok_pn = datetime.datetime.now().year
            if not pn_roky:
                pn_roky = [_akt_rok_pn]
            pn_rok_default = _akt_rok_pn if _akt_rok_pn in pn_roky else pn_roky[0]

            # Výběr poboček zahrnuje i vnořené provozy jako samostatné položky
            # (Ostrava → 027-B.Olomouc, 031-B.Jeseník; Hodonín → 030-Brno),
            # aby je šlo filtrovat zvlášť. Každá položka = vlastní datový klíč;
            # prázdný výběr = součet všech (vč. vnořených, každý jednou).
            vsechny_klice_pn: list[str] = []
            labely_pn: dict[str, str] = {}
            for p in pristupne_pobocky:
                vsechny_klice_pn.append(p)
                labely_pn[p] = (_POBOCKY_EXCEL_REVERSE.get(p, p)
                                if p not in ('Bourárna', 'Kamiony') else p)
                for disp, klic in _VNORENE_POBOCKY.get(p, []):
                    if klic not in labely_pn:
                        vsechny_klice_pn.append(klic)
                        labely_pn[klic] = f'↳ {disp}'

            @ui.refreshable
            def _podrobne_grid():
                vyb = app.storage.user.get(fk_pn_p) or []
                pob = [k for k in vsechny_klice_pn if k in vyb] if vyb else list(vsechny_klice_pn)
                rok = app.storage.user.get(fk_pn_r)
                if rok not in pn_roky:
                    rok = pn_rok_default
                minuly = rok - 1
                rows = _porovnani_z_klicu(pob, rok)
                vsechny = (not vyb) or len(pob) == len(vsechny_klice_pn)
                pob_txt = ('všechny pobočky' if vsechny else ', '.join(
                    labely_pn.get(k, k).lstrip('↳ ') for k in pob))
                with ui.row().classes('items-center gap-3 mb-1 flex-wrap'):
                    ui.icon('compare_arrows', color='teal', size='sm')
                    ui.label(f'Porovnání {rok} vs {minuly}') \
                        .classes('text-sm font-semibold text-gray-700')
                    ui.label(f'Pobočky: {pob_txt}').classes('text-xs text-gray-400 italic')
                    ui.label('Δ = Aktuální − Minulý · 🔴 nárůst  🟢 pokles') \
                        .classes('text-xs text-gray-500')
                    filtr_slot = ui.row().classes('items-center gap-1 ml-auto')
                if not rows:
                    ui.label('Žádná data podrobných nákladů pro vybrané pobočky.') \
                        .classes('text-gray-400 italic py-4')
                    return
                grid = ui.aggrid({
                    'columnDefs': _col_defs_porovnani(rok),
                    'rowData': rows,
                    'pinnedBottomRowData': [_soucet_porovnani(rows)],
                    'defaultColDef': {'resizable': True, 'sortable': False,
                                      'wrapHeaderText': True, 'autoHeaderHeight': True},
                    'rowHeight': 34,
                    'suppressMovableColumns': True,
                    ':getRowStyle': _PINNED_BOTTOM_STYLE,
                    ':onFirstDataRendered': _AUTOSIZE_FIT,
                    ':onGridSizeChanged': _AUTOSIZE_FIT,
                }).classes('w-full').style(_GRID_STYLE)
                with filtr_slot:
                    _pridej_filtr_nazvu(rows, grid, _soucet_porovnani,
                                        f'podrobne_naklady_vse_{rok}', s_exportem=False)

            pob_opts_pn = {k: labely_pn[k] for k in vsechny_klice_pn}
            rok_opts_pn = {r: str(r) for r in pn_roky}
            with ui.row().classes('items-center gap-3 mb-3 flex-wrap'):
                w_pn_r = ui.select(
                    rok_opts_pn,
                    value=(app.storage.user.get(fk_pn_r)
                           if app.storage.user.get(fk_pn_r) in pn_roky else pn_rok_default),
                    label='Rok',
                ).props('dense outlined options-dense').style('min-width: 120px') \
                 .tooltip('Porovnává se vybraný rok proti předchozímu')
                # Bez `use-chips` – při výběru všech poboček by se vypsal dlouhý
                # seznam chipsů. Místo toho kompaktní `display-value` (viz _pn_stav).
                w_pn_p = ui.select(
                    pob_opts_pn, value=list(app.storage.user.get(fk_pn_p) or []),
                    label='Pobočky', multiple=True,
                ).props('dense outlined options-dense behavior=menu').style('min-width: 240px') \
                 .tooltip('Prázdné = všechny pobočky (souhrn)')

                # Stav výběru: zelená fajfka = vybrány všechny pobočky;
                # jinak šedá. Kompaktní `display-value` místo dlouhého výpisu chipsů.
                def _pn_stav():
                    vyb = app.storage.user.get(fk_pn_p) or []
                    je_vse = bool(vyb) and len(vyb) >= len(vsechny_klice_pn)
                    if je_vse:
                        w_pn_p.props('display-value="Všechny pobočky"')
                    elif vyb:
                        w_pn_p.props(f'display-value="{len(vyb)} vybráno"')
                    else:
                        w_pn_p.props(remove='display-value')
                    btn_vse.props(f'color={"positive" if je_vse else "grey-7"}')

                # Přepínač „Všechny pobočky": kliknutím se vybere vše (zelená
                # fajfka), opětovným kliknutím se výběr zruší (zešediví).
                def _toggle_vse(_):
                    vyb = app.storage.user.get(fk_pn_p) or []
                    je_vse = bool(vyb) and len(vyb) >= len(vsechny_klice_pn)
                    nove = [] if je_vse else list(vsechny_klice_pn)
                    app.storage.user[fk_pn_p] = nove
                    w_pn_p.set_value(nove)
                    _pn_stav()
                    _podrobne_grid.refresh()
                btn_vse = ui.button('Všechny pobočky', icon='done_all', on_click=_toggle_vse) \
                    .props('flat dense no-caps').classes('text-xs') \
                    .tooltip('Přepnout výběr všech poboček (zapnout/vypnout)')

                def _zm_pn_r(e):
                    app.storage.user[fk_pn_r] = e.value
                    _podrobne_grid.refresh()
                w_pn_r.on_value_change(_zm_pn_r)

                def _zm_pn_p(e):
                    app.storage.user[fk_pn_p] = list(e.value) if e.value else []
                    _pn_stav()
                    _podrobne_grid.refresh()
                w_pn_p.on_value_change(_zm_pn_p)

                _pn_stav()   # výchozí stav (display-value + barva fajfky)
            _podrobne_grid()


# ─── UI funkce ────────────────────────────────────────────────────────────────

def _stav_bunka(rok: int, mesic: int, oddeleni: str, info, editable: bool,
                user_name: str, refresh):
    """Jedna měsíční dlaždice stavu (zelená=hotovo, červená=ne). Editovatelná
    pro příslušné oddělení – klik přepne stav."""
    hotovo = bool(info and info.get('hotovo'))
    bg = '#16a34a' if hotovo else '#dc2626'
    cls = 'flex items-center justify-center rounded text-white select-none'
    if editable:
        cls += ' cursor-pointer hover:opacity-80 transition-opacity'
    box = ui.element('div').classes(cls).style(
        f'width:46px;height:36px;background:{bg};font-size:13px;font-weight:700')
    with box:
        ui.label(('✓ ' if hotovo else '') + f'{mesic:02d}')
    if hotovo and info and info.get('kdo'):
        cas = info.get('cas')
        cas_txt = cas.strftime('%d.%m.%Y %H:%M') if hasattr(cas, 'strftime') else ''
        popis = f'Dokončil: {info["kdo"]}' + (f' · {cas_txt}' if cas_txt else '')
    else:
        popis = 'Klikněte pro označení jako dokončené' if editable else 'Zatím nedokončeno'
    # Nativní HTML title místo QTooltip: panel se při každém kliknutí celý
    # překresluje (_board.refresh()), takže by se QTooltip odpojoval od mizející
    # kotvy a házel do konzole „Anchor: target #cXXX not found". Title tento
    # problém nemá. (Uvozovky v textu nahradíme, aby nerozbily prop.)
    box.props(f'title="{popis.replace(chr(34), chr(39))}"')
    if editable:
        def _toggle(_=None, m=mesic, o=oddeleni, h=hotovo):
            uloz_mesic_status(rok, m, o, not h, user_name)
            intranet_logger.log_activity(
                user_name, 'Výsledky poboček',
                f'Stav {o} {rok}-{m:02d}: {"dokončeno" if not h else "zrušeno dokončení"}')
            refresh()
        box.on('click', _toggle)


def _vykresli_stav_mesicu(rok: int, je_ao: bool, je_ucetni_dept: bool, user_name: str):
    """Přehledový panel stavu vyplnění po měsících za obě oddělení."""
    oddeleni = [
        ('ao',     'Analytické oddělení', je_ao),
        ('ucetni', 'Účetní oddělení',     je_ucetni_dept),
    ]

    @ui.refreshable
    def _board():
        status = nacti_mesic_status(rok)
        with ui.card().classes('w-full p-4 mb-6 rounded-2xl border border-gray-100'):
            with ui.row().classes('items-center gap-2 mb-1'):
                ui.icon('event_available', color='primary')
                ui.label(f'Stav vyplnění výsledků za rok {rok}') \
                    .classes('text-lg font-bold text-gray-800')
            ui.label('🟢 měsíc dokončen · 🔴 ještě nedokončeno') \
                .classes('text-xs text-gray-500 mb-2')
            for odd, nazev, editable in oddeleni:
                hot = sum(1 for m in range(1, 13)
                          if status.get((m, odd), {}).get('hotovo'))
                with ui.row().classes('items-center gap-1.5 mb-2 flex-wrap'):
                    with ui.column().classes('gap-0').style('width: 200px'):
                        ui.label(nazev).classes('font-semibold text-sm text-gray-700')
                        ui.label(f'{hot}/12 dokončeno'
                                 + ('' if editable else ' · jen pro čtení')) \
                            .classes('text-xs text-gray-400')
                    for m in range(1, 13):
                        _stav_bunka(rok, m, odd, status.get((m, odd)),
                                    editable, user_name, _board.refresh)

    _board()


def _vykresli_ostatni_provozy(user_id: int, user_name: str, je_ao: bool, je_ucetni: bool):
    aktualni_rok = datetime.datetime.now().year
    minuly_rok = aktualni_rok - 1

    for _, data_key in OSTATNI_PROVOZY:
        inicializuj_osnovu(data_key, aktualni_rok)
        inicializuj_osnovu(data_key, minuly_rok)

    with ui.tabs().props('align=left active-color=primary indicator-color=primary') \
            .classes('w-full border-b border-gray-200') as tabs_op:
        for nazev, _ in OSTATNI_PROVOZY:
            ui.tab(nazev, label=f'🏪 {nazev.upper()}')

    with ui.tab_panels(tabs_op, value=OSTATNI_PROVOZY[0][0]).classes('w-full pt-4'):
        for nazev, data_key in OSTATNI_PROVOZY:
            with ui.tab_panel(nazev):
                _vykresli_podrobne_naklady(
                    data_key, aktualni_rok, minuly_rok,
                    user_id, user_name, je_ao, je_ucetni,
                )


@refreshable_na_klienta
def vykresli_vysledky(user_id: int, user_name: str, vsechna_prava: list):
    inicializace_vysledky_db()

    ma_vse         = 'vse' in vsechna_prava
    je_ao          = ma_vse or 'vysledky_ao'           in vsechna_prava
    # Účetní hlavní (vysledky_ucetni): vidí vše + edituje Podrobné náklady.
    # Účetní běžná (vysledky_ucetni_bezna): jen Podrobné náklady, pouze pro čtení.
    je_ucetni_h    = je_ao  or 'vysledky_ucetni'        in vsechna_prava   # může editovat
    je_ucetni_b    = ma_vse or 'vysledky_ucetni_bezna'  in vsechna_prava   # jen čtení
    je_ucetni_any  = je_ucetni_h or je_ucetni_b
    je_majitel     = ma_vse or 'vysledky_majitel'      in vsechna_prava
    # Pro stavový panel: účetní oddělení = jen hlavní účetní právo (nebo superadmin).
    je_ucetni_dept = ma_vse or 'vysledky_ucetni'       in vsechna_prava
    # Čtenář Ostatních provozů: VO vedoucí – vidí jen dlaždici a data Ostatních provozů (read-only).
    je_op_ctenar   = ma_vse or 'vysledky_pobocka_ostatni_provozy' in vsechna_prava
    # Běžná účetní bez širšího práva → v detailu pobočky jen Podrobné náklady (čtení)
    jen_podrobne   = je_ucetni_b and not (je_ao or je_ucetni_h or je_majitel)

    pristupne_pobocky = [
        p for p in POBOCKY
        if je_ao or je_ucetni_any or je_majitel or f'vysledky_pobocka_{POBOCKY_KLICE[p]}' in vsechna_prava
    ]

    if not pristupne_pobocky and not je_op_ctenar:
        with ui.column().classes('items-center py-24 gap-4'):
            ui.icon('lock', size='4rem', color='grey-4')
            ui.label('Nemáte přístup do žádné pobočky výsledků.').classes('text-gray-400 text-lg')
        return

    _state_key = f'vysledky_pobocka_{user_id}'

    @ui.refreshable
    def _panel():
        sel = app.storage.user.get(_state_key)

        if sel is None:
            ui.label('Výsledky poboček').classes('text-3xl font-bold text-gray-800 mb-6')

            # Stavový přehled měsíců – vidí AO i obě účetní role (čtení);
            # přepínat řádek může jen příslušné oddělení (hlavní účetní / AO).
            if je_ucetni_any:
                _vykresli_stav_mesicu(datetime.datetime.now().year, je_ao, je_ucetni_dept, user_name)

            if je_ao:
                import asyncio, io, inspect

                with ui.dialog() as _dlg_import, ui.card().classes('p-5 gap-3') \
                        .style('min-width: 440px'):
                    ui.label('Nahrát data').classes('text-lg font-bold text-gray-800')
                    ui.label(
                        'Nahrajte „Tabulka nákladů_DATA.xlsx" se všemi pobočkami – data se '
                        'automaticky rozdělí podle sloupce „Pobočka popis" na jednotlivé pobočky '
                        '(dílčí jednotky se sečtou). Data nahraných poboček se přepíší.'
                    ).classes('text-xs text-gray-500')

                    ui.label('Co se ze souboru naplní').classes('text-sm font-semibold text-gray-700')
                    w_imp_naklady = ui.checkbox('Přehled nákladů', value=True) \
                        .props('dense').classes('text-sm')
                    w_imp_oz = ui.checkbox(
                        'Obraty – Staropramen a Operace (po měsících, dcery, sítě)',
                        value=True).props('dense').classes('text-sm')

                    # ── Volitelné: výchozí zobrazení pro všechny nahrané pobočky ──
                    ui.separator()
                    ui.label('Výchozí zobrazení po importu (volitelné)') \
                        .classes('text-sm font-semibold text-gray-700')
                    ui.label(
                        'Vybraný rok a měsíce se po dokončení importu nastaví jako výchozí '
                        'filtr tabulky nákladů pro všechny nahrané pobočky najednou. '
                        'Ponecháte-li prázdné, výchozí zobrazení poboček se nezmění.'
                    ).classes('text-xs text-gray-500')
                    _def_roky_opts = sorted(
                        {*nacti_dostupne_roky(), datetime.datetime.now().year}, reverse=True)
                    with ui.row().classes('items-center gap-2'):
                        w_def_rok = ui.select(
                            {r: str(r) for r in _def_roky_opts},
                            value=[], label='Rok', multiple=True,
                        ).props('dense outlined options-dense use-chips').style('min-width: 150px')
                        w_def_mes = ui.select(
                            {i: str(i) for i in range(1, 13)},
                            value=[], label='Měsíce', multiple=True,
                        ).props('dense outlined options-dense use-chips').style('min-width: 150px')

                    async def _on_upload_vse(e):
                        if not (w_imp_naklady.value or w_imp_oz.value):
                            ui.notify('Vyberte, co se má ze souboru naplnit.',
                                      type='warning')
                            return
                        raw = await _precti_soubor(e)
                        if raw is None:
                            return
                        _dlg_import.close()
                        _upload_vse_el.reset()
                        count, pobocky_imp = 0, []
                        hlasky = []
                        if w_imp_naklady.value:
                            count, pobocky_imp, err = await asyncio.to_thread(
                                _importuj_souhrn_vse_sync, io.BytesIO(raw))
                            if err:
                                ui.notify(f'Chyba importu nákladů: {err}',
                                          type='negative', timeout=10000)
                                return
                            hlasky.append(f'náklady {count} záznamů / '
                                          f'{len(pobocky_imp)} poboček')
                        if w_imp_oz.value:
                            matched, preskocene, err = await _oz_spust_s_progressem(
                                _ozn_parsuj_data_sync, io.BytesIO(raw),
                                uvod='Načítám obraty…')
                            if err:
                                ui.notify(f'Chyba importu obratů: {err}',
                                          type='negative', timeout=10000)
                                return
                            oz_count, oz_npob, chyby = await _oz_spust_s_progressem(
                                _oz_centralni_importuj_sync, matched,
                                _ozn_zapis_pobocku, uvod='Ukládám obraty…')
                            hlasky.append(f'obraty {oz_count} hodnot / '
                                          f'{oz_npob} poboček'
                                          + (f' (chyby: {"; ".join(chyby)})'
                                             if chyby else ''))
                            for txt in (preskocene or []):
                                ui.notify(f'⚠️ {txt}', type='warning',
                                          position='top-right')
                        npob = len(pobocky_imp)
                        def_roky   = sorted(w_def_rok.value or [])
                        def_mesice = sorted(w_def_mes.value or [])
                        if def_roky or def_mesice:
                            for _pb in pobocky_imp:
                                _ver = app.storage.general.get(f'souhrn_default_ver_{_pb}', 0) + 1
                                app.storage.general[f'souhrn_default_rok_{_pb}']    = list(def_roky)
                                app.storage.general[f'souhrn_default_mesice_{_pb}'] = list(def_mesice)
                                app.storage.general[f'souhrn_default_ver_{_pb}']    = _ver
                        intranet_logger.log_activity(
                            user_name, 'Výsledky poboček',
                            f'Hromadný import Tabulka nákladů: {", ".join(hlasky)}'
                            + (f'; výchozí filtr – rok: {def_roky}, měsíce: {def_mesice}'
                               if (def_roky or def_mesice) else ''))
                        ui.notify(f'Import dokončen – {", ".join(hlasky)}.',
                                  type='positive', position='top-right')
                        if def_roky or def_mesice:
                            ui.notify(f'Výchozí zobrazení nastaveno pro {npob} poboček.',
                                      type='info', position='top-right')

                    _upload_vse_el = ui.upload(on_upload=_on_upload_vse, auto_upload=True,
                              max_file_size=50_000_000, label='Vybrat .xlsx soubor') \
                        .props('accept=.xlsx').classes('w-full')
                    with ui.row().classes('justify-end w-full'):
                        ui.button('Zavřít', on_click=_dlg_import.close).props('flat no-caps')

                # ── Dialog importu středisek (sdílený pro historická i aktuální data) ──
                _akt_rok = datetime.datetime.now().year

                def _vytvor_strediska_dialog(rok_override, nadpis, popis_text):
                    """Vytvoří dialog importu středisek. rok_override=None → rok se
                    rozpozná ze souboru (historická data); jinak se data uloží vždy
                    do zadaného roku (aktuální rok)."""
                    ref = {'data': None, 'unmatched': None, 'rok': None}
                    with ui.dialog() as dlg, ui.card().classes('p-5 gap-3') \
                            .style('min-width: 600px; max-width: 900px; max-height: 80vh; overflow-y: auto'):
                        ui.label(nadpis).classes('text-lg font-bold text-gray-800')
                        ui.label(popis_text).classes('text-xs text-gray-500')

                        async def _on_upload(e):
                            zdroj = None
                            for attr in ('content', 'file', 'stream', 'data', 'file_obj'):
                                val = getattr(e, attr, None)
                                if val is not None and hasattr(val, 'read'):
                                    zdroj = val
                                    break
                            if zdroj is None:
                                ui.notify('Nepodařilo se načíst obsah souboru.', type='negative')
                                return
                            try:
                                raw = zdroj.read()
                                if inspect.isawaitable(raw):
                                    raw = await raw
                            except Exception as exc:
                                ui.notify(f'Chyba čtení souboru: {exc}', type='negative')
                                return
                            matched, unmatched, det_rok, err = await asyncio.to_thread(
                                _strediska_parsuj_excel_sync, io.BytesIO(raw))
                            if err:
                                ui.notify(f'Chyba: {err}', type='negative', timeout=10000)
                                return
                            if not matched:
                                ui.notify('V souboru nebyly nalezeny žádné rozpoznatelné pobočky.', type='warning')
                                return
                            ref['data'] = matched
                            ref['unmatched'] = unmatched
                            ref['rok'] = rok_override if rok_override is not None else det_rok
                            _preview.refresh()

                        ui.upload(on_upload=_on_upload, auto_upload=True,
                                  max_file_size=50_000_000, label='Vybrat .xlsx soubor') \
                            .props('accept=.xlsx').classes('w-full')

                        @ui.refreshable
                        def _preview():
                            if ref['data'] is None:
                                return
                            matched = ref['data']
                            unmatched = ref['unmatched']
                            rok = ref['rok']

                            ui.separator()
                            ui.label(f'Rok importu: {rok}').classes('text-sm font-bold text-gray-700')

                            if matched:
                                ui.label(f'Rozpoznané pobočky ({len(matched)}):') \
                                    .classes('text-sm font-semibold text-green-700 mt-2')
                                for entry in matched:
                                    sheets = ', '.join(entry['sheet_names'])
                                    ui.label(
                                        f'✅ {entry["pobocka"]} ← {sheets} ({len(entry["rows"])} řádků)'
                                    ).classes('text-xs text-gray-700')

                            if unmatched:
                                ui.label(f'Nerozpoznané listy ({len(unmatched)}):') \
                                    .classes('text-sm font-semibold text-orange-700 mt-2')
                                for name in unmatched:
                                    ui.label(f'⚠️ {name}').classes('text-xs text-gray-500')

                            if matched:
                                total_rows = sum(len(e['rows']) for e in matched)
                                ui.label(
                                    f'Celkem bude importováno {total_rows} řádků '
                                    f'do {len(matched)} poboček pro rok {rok}. '
                                    f'Existující data podrobných nákladů za rok {rok} budou nahrazena.'
                                ).classes('text-sm text-red-600 font-semibold mt-3')

                                async def _potvrdit():
                                    count, npob, err = await asyncio.to_thread(
                                        _strediska_importuj_sync, ref['data'], ref['rok'])
                                    if err:
                                        ui.notify(f'Chyba importu: {err}', type='negative', timeout=10000)
                                        return
                                    intranet_logger.log_activity(
                                        user_name, 'Výsledky poboček',
                                        f'Import středisek {ref["rok"]}: {count} řádků, {npob} poboček')
                                    ui.notify(
                                        f'Import dokončen – {count} řádků pro {npob} poboček.',
                                        type='positive', position='top-right')
                                    dlg.close()
                                    _panel.refresh()

                                ui.button('Ano, importovat data', icon='check',
                                          on_click=_potvrdit) \
                                    .props('color=primary no-caps').classes('mt-2')

                        _preview()
                        with ui.row().classes('justify-end w-full'):
                            ui.button('Zavřít', on_click=dlg.close).props('flat no-caps')
                    return dlg

                _dlg_strediska = _vytvor_strediska_dialog(
                    None,
                    'Import historických dat středisek',
                    'Nahrajte soubor „Střediska RRRR.xlsx" – data se načtou do sekce '
                    '„Minulý rok" v podrobných nákladech. Rok se rozpozná z názvů listů, '
                    'pobočky podle čísla střediska.')
                _dlg_strediska_akt = _vytvor_strediska_dialog(
                    _akt_rok,
                    f'Import dat aktuálního roku ({_akt_rok})',
                    f'Nahrajte soubor „Střediska.xlsx" – data se uloží do sekce '
                    f'„Aktuální rok" ({_akt_rok}) v podrobných nákladech, vždy pro rok '
                    f'{_akt_rok} bez ohledu na rok v souboru. Pobočky se rozpoznají '
                    'podle čísla střediska v názvu listu.')

                # ── Dialog: centrální import Obratů a Zisků (Pobočkové GISTy) ──
                _oz_imp_ref = {'matched': None, 'unmatched': None}

                # Blokující overlay s progress barem – persistent dialog nad vším:
                # dokud běží parsování/import, nejde kliknout na nic jiného.
                _oz_prog = {'v': 0.0, 'txt': ''}
                with ui.dialog().props('persistent') as _dlg_oz_prog, \
                        ui.card().classes('p-5 gap-3').style('min-width: 420px'):
                    with ui.row().classes('items-center gap-2'):
                        ui.spinner(size='sm', color='primary')
                        _oz_prog_lbl = ui.label('') \
                            .classes('text-sm font-semibold text-gray-700')
                    _oz_prog_bar = ui.linear_progress(value=0.0, show_value=False) \
                        .props('instant-feedback color=primary size=10px rounded') \
                        .classes('w-full')

                async def _oz_spust_s_progressem(fn, *args, uvod: str):
                    """Spustí fn ve vlákně; overlay blokuje klikání a bar plynule
                    dotéká mezi checkpointy (načtení souboru je jeden dlouhý blok,
                    proto se mezi hlášeními posouvá po malých krocích)."""
                    _oz_prog['v'] = 0.0
                    _oz_prog['txt'] = uvod
                    _oz_prog_bar.set_value(0.0)
                    _oz_prog_lbl.set_text(uvod)
                    _dlg_oz_prog.open()

                    def _cb(v, txt):
                        _oz_prog['v'] = v
                        _oz_prog['txt'] = txt

                    task = asyncio.create_task(asyncio.to_thread(fn, *args, progress=_cb))
                    disp = 0.0
                    try:
                        while not task.done():
                            disp = min(max(disp + 0.008, _oz_prog['v']),
                                       _oz_prog['v'] + 0.25, 0.97)
                            _oz_prog_bar.set_value(round(disp, 3))
                            _oz_prog_lbl.set_text(_oz_prog['txt'])
                            await asyncio.sleep(0.1)
                        _oz_prog_bar.set_value(1.0)
                        return await task
                    finally:
                        _dlg_oz_prog.close()

                async def _precti_soubor(e):
                    """Obsah nahraného souboru jako bytes, nebo None + notifikace."""
                    zdroj = None
                    for attr in ('content', 'file', 'stream', 'data', 'file_obj'):
                        val = getattr(e, attr, None)
                        if val is not None and hasattr(val, 'read'):
                            zdroj = val
                            break
                    if zdroj is None:
                        ui.notify('Nepodařilo se načíst obsah souboru.', type='negative')
                        return None
                    try:
                        raw = zdroj.read()
                        if inspect.isawaitable(raw):
                            raw = await raw
                    except Exception as exc:
                        ui.notify(f'Chyba čtení souboru: {exc}', type='negative')
                        return None
                    return raw

                with ui.dialog() as _dlg_oz, ui.card().classes('p-5 gap-3') \
                        .style('min-width: 600px; max-width: 900px; max-height: 80vh; overflow-y: auto'):
                    ui.label('Import Obratů a Zisků všech poboček') \
                        .classes('text-lg font-bold text-gray-800')
                    ui.label(
                        'Nahrajte centrální soubor „Pobočkové GISTy.xlsx" – listy '
                        '„010 Obraty_prodeje", „010 Zisk_Marže", … se rozdělí na pobočky '
                        'podle čísla střediska v názvu listu a naplní záložky Obraty '
                        'a Zisk v detailu pobočky.'
                    ).classes('text-xs text-gray-500')

                    async def _on_upload_oz(e):
                        raw = await _precti_soubor(e)
                        if raw is None:
                            return
                        # Po dobu parsování se importní dialog schová – zůstane jen
                        # blokující overlay s progress barem; pak se otevře s náhledem.
                        _dlg_oz.close()
                        try:
                            matched, unmatched, err = await _oz_spust_s_progressem(
                                _oz_centralni_parsuj_sync, io.BytesIO(raw),
                                uvod='Načítám soubor…')
                        finally:
                            _oz_upload_el.reset()
                            _dlg_oz.open()
                        if err:
                            ui.notify(f'Chyba: {err}', type='negative', timeout=10000)
                            return
                        if not matched:
                            ui.notify('V souboru se nepodařilo spárovat žádnou pobočku.',
                                      type='warning')
                        _oz_imp_ref['matched'] = matched
                        _oz_imp_ref['unmatched'] = unmatched
                        # Po naparsování zůstane v dialogu jen přehled – pole pro
                        # výběr souboru se skryje (vrátí ho nové otevření dialogu).
                        if matched:
                            _oz_upload_el.set_visibility(False)
                        _oz_nahled.refresh()

                    _oz_upload_el = ui.upload(on_upload=_on_upload_oz, auto_upload=True,
                                              max_file_size=50_000_000,
                                              label='Vybrat .xlsx soubor') \
                        .props('accept=.xlsx').classes('w-full')

                    @ui.refreshable
                    def _oz_nahled():
                        matched = _oz_imp_ref['matched']
                        unmatched = _oz_imp_ref['unmatched']
                        if matched is None:
                            return
                        ui.separator()
                        if matched:
                            ui.label(f'Rozpoznané pobočky ({len(matched)}):') \
                                .classes('text-sm font-semibold text-green-700 mt-2')
                            for entry in matched:
                                roky = sorted({r[3] for r in entry['radky']})
                                pozn = f' – ⚠️ {", ".join(entry["pozn"])}' if entry['pozn'] else ''
                                ui.label(
                                    f'✅ {entry["cislo"]} {entry["pobocka"]} ← '
                                    f'{", ".join(entry["listy"])} ({len(entry["radky"])} hodnot, '
                                    f'roky {", ".join(map(str, roky))}){pozn}'
                                ).classes('text-xs text-gray-700')
                        if unmatched:
                            ui.label(f'Přeskočené listy ({len(unmatched)}):') \
                                .classes('text-sm font-semibold text-orange-700 mt-2')
                            for txt in unmatched:
                                ui.label(f'⚠️ {txt}').classes('text-xs text-gray-500')
                        if matched:
                            total = sum(len(entry['radky']) for entry in matched)
                            ui.label(
                                f'Celkem bude importováno {total} hodnot do {len(matched)} '
                                'poboček. Existující data Obratů a Zisků těchto poboček budou '
                                'nahrazena, pobočky mimo soubor zůstanou beze změny.'
                            ).classes('text-sm text-red-600 font-semibold mt-3')

                            async def _potvrdit_oz():
                                _dlg_oz.close()
                                count, npob, chyby = await _oz_spust_s_progressem(
                                    _oz_centralni_importuj_sync, _oz_imp_ref['matched'],
                                    uvod='Ukládám data…')
                                intranet_logger.log_activity(
                                    user_name, 'Výsledky poboček',
                                    f'Centrální import Obraty/Zisk: {count} hodnot, '
                                    f'{npob} poboček'
                                    + (f', chyby: {"; ".join(chyby)}' if chyby else ''))
                                if chyby:
                                    ui.notify(
                                        f'Importováno {npob} poboček ({count} hodnot). '
                                        f'Chyby: {"; ".join(chyby)}',
                                        type='warning', timeout=15000, position='top-right')
                                else:
                                    ui.notify(
                                        f'Import dokončen – {count} hodnot pro {npob} poboček.',
                                        type='positive', position='top-right')
                                _dlg_oz.close()

                            ui.button('Ano, importovat data', icon='check',
                                      on_click=_potvrdit_oz) \
                                .props('color=primary no-caps').classes('mt-2')

                    _oz_nahled()
                    with ui.row().classes('justify-end w-full'):
                        ui.button('Zavřít', on_click=_dlg_oz.close).props('flat no-caps')

                def _otevri_oz():
                    """Otevře importní dialog vždy v čistém stavu – bez náhledu
                    z minula a s viditelným polem pro výběr souboru."""
                    _oz_imp_ref['matched'] = None
                    _oz_imp_ref['unmatched'] = None
                    _oz_upload_el.reset()
                    _oz_upload_el.set_visibility(True)
                    _oz_nahled.refresh()
                    _dlg_oz.open()

                # ── Dialog: Rozeslat e-maily o aktualizaci ──────────────────
                _email_ref = {'prijemci': None}

                with ui.dialog() as _dlg_email, ui.card().classes('p-5 gap-3') \
                        .style('min-width: 560px; max-width: 820px; max-height: 80vh; overflow-y: auto'):
                    with ui.row().classes('items-center gap-2'):
                        ui.icon('mark_email_read', color='positive')
                        ui.label('Rozeslat e-maily o aktualizaci') \
                            .classes('text-lg font-bold text-gray-800')
                    ui.label(
                        'Odešle všem uživatelům s právem čtení Výsledků poboček informaci, '
                        'že jsou všechna data vyplněná a k dispozici. Každý dostane seznam '
                        'poboček, na které má přístup.'
                    ).classes('text-xs text-gray-500')

                    async def _potvrd_email():
                        prijemci = _email_ref.get('prijemci') or []
                        if not prijemci:
                            return
                        ui.notify('Odesílám e-maily…', type='ongoing', position='top-right')
                        sent, fail = await asyncio.to_thread(_rozesli_aktualizace_sync, prijemci)
                        intranet_logger.log_activity(
                            user_name, 'Výsledky poboček',
                            f'Rozeslány e-maily o aktualizaci: {sent} odesláno, {fail} chyb')
                        if fail:
                            ui.notify(f'Odesláno {sent} e-mailů, {fail} se nepodařilo odeslat.',
                                      type='warning', timeout=9000, position='top-right')
                        else:
                            ui.notify(f'Hotovo – odesláno {sent} e-mailů.',
                                      type='positive', position='top-right')
                        _dlg_email.close()

                    @ui.refreshable
                    def _email_preview():
                        prijemci = _email_ref.get('prijemci')
                        if prijemci is None:
                            with ui.row().classes('items-center gap-2 py-3'):
                                ui.spinner(size='sm')
                                ui.label('Načítám příjemce…').classes('text-sm text-gray-500')
                            return
                        if not prijemci:
                            ui.label('Nenašli se žádní příjemci s e-mailem a právem čtení.') \
                                .classes('text-sm text-orange-600 py-3')
                            return
                        ui.separator()
                        ui.label(f'Bude osloveno {len(prijemci)} příjemců:') \
                            .classes('text-sm font-semibold text-gray-700')
                        with ui.column().classes('gap-0 w-full'):
                            for r in prijemci[:60]:
                                pob = ('všechny pobočky' if r['vse']
                                       else ', '.join(
                                           _POBOCKY_EXCEL_REVERSE.get(p, p)
                                           if p not in ('Bourárna', 'Kamiony') else p
                                           for p in r['pobocky']) or '—')
                                ui.label(f'• {r["email"]} — {pob}') \
                                    .classes('text-xs text-gray-600')
                            if len(prijemci) > 60:
                                ui.label(f'… a další {len(prijemci) - 60}') \
                                    .classes('text-xs text-gray-400 italic')
                        ui.button('Ano, rozeslat e-maily', icon='send',
                                  on_click=_potvrd_email) \
                            .props('color=positive no-caps').classes('mt-2')

                    _email_preview()
                    with ui.row().classes('justify-end w-full'):
                        ui.button('Zavřít', on_click=_dlg_email.close).props('flat no-caps')

                    async def _otevri_email():
                        _email_ref['prijemci'] = None
                        _email_preview.refresh()
                        _dlg_email.open()
                        _email_ref['prijemci'] = await asyncio.to_thread(_nacti_prijemce_aktualizace)
                        _email_preview.refresh()

                # ── Dialog: Přehled práv na pobočky (jen hlavní administrátor) ──
                if ma_vse:
                    _prava_ref = {'data': None}

                    with ui.dialog() as _dlg_prava, ui.card().classes('p-5 gap-3') \
                            .style('min-width: 760px; max-width: 1120px; max-height: 85vh; overflow-y: auto'):
                        with ui.row().classes('items-center gap-2'):
                            ui.icon('admin_panel_settings', color='primary')
                            ui.label('Přehled práv na pobočky') \
                                .classes('text-lg font-bold text-gray-800')
                        ui.label(
                            'Jak mají jednotliví uživatelé nastavena práva ke čtení Výsledků '
                            'poboček – přiřazená přímo, přes pracovní pozici nebo přes oddělení. '
                            'Role Superadmin / AO / Účetní / Majitel = přístup ke všem pobočkám.'
                        ).classes('text-xs text-gray-500')

                        @ui.refreshable
                        def _prava_obsah():
                            data = _prava_ref.get('data')
                            if data is None:
                                with ui.row().classes('items-center gap-2 py-4'):
                                    ui.spinner(size='sm')
                                    ui.label('Načítám práva…').classes('text-sm text-gray-500')
                                return
                            if not data:
                                ui.label('Žádní uživatelé nemají nastavena práva na Výsledky poboček.') \
                                    .classes('text-sm text-orange-600 py-3')
                                return
                            rows = []
                            for d in data:
                                if d['vse']:
                                    pob_txt = 'Všechny pobočky'
                                else:
                                    pob_txt = ', '.join(
                                        _POBOCKY_EXCEL_REVERSE.get(p, p)
                                        if p not in ('Bourárna', 'Kamiony') else p
                                        for p in d['pobocky']) or '—'
                                rows.append({
                                    'uzivatel': d['jmeno'],
                                    'email': d['email'] or '—',
                                    'role': ', '.join(d['role']) or '—',
                                    'pobocky': pob_txt,
                                    'zdroj': ', '.join(d['zdroje']),
                                    'aktivni': 'Ano' if d['aktivni'] else 'Ne',
                                })
                            ui.label(f'Celkem {len(rows)} uživatelů s právem na Výsledky poboček.') \
                                .classes('text-xs text-gray-500')
                            grid_prava = ui.aggrid({
                                'columnDefs': [
                                    {'headerName': 'Uživatel', 'field': 'uzivatel',
                                     'minWidth': 170, 'filter': True, 'floatingFilter': True},
                                    {'headerName': 'E-mail', 'field': 'email',
                                     'minWidth': 210, 'filter': True, 'floatingFilter': True},
                                    {'headerName': 'Role', 'field': 'role',
                                     'minWidth': 150, 'filter': True},
                                    {'headerName': 'Pobočky', 'field': 'pobocky',
                                     'minWidth': 240, 'filter': True, 'floatingFilter': True,
                                     'wrapText': True, 'autoHeight': True},
                                    {'headerName': 'Zdroj práva', 'field': 'zdroj',
                                     'minWidth': 130, 'filter': True},
                                    {'headerName': 'Aktivní', 'field': 'aktivni',
                                     'width': 95, 'minWidth': 80,
                                     ':cellStyle': "function(p){return p.value==='Ne'"
                                                   "?{color:'#b91c1c'}:null;}"},
                                ],
                                'rowData': rows,
                                'defaultColDef': {'resizable': True, 'sortable': True},
                                'domLayout': 'autoHeight',
                            }).classes('w-full').style('max-height: 62vh; overflow-y: auto')

                            def _export_prava():
                                grid_prava.run_grid_method('exportDataAsCsv', {
                                    'fileName': 'prehled_prav_vysledky', 'columnSeparator': ';'})
                            ui.button('Export CSV', icon='download', on_click=_export_prava) \
                                .props('flat dense no-caps color=primary').classes('text-xs mt-1')

                        _prava_obsah()
                        with ui.row().classes('justify-end w-full'):
                            ui.button('Zavřít', on_click=_dlg_prava.close).props('flat no-caps')

                    async def _otevri_prava():
                        _prava_ref['data'] = None
                        _prava_obsah.refresh()
                        _dlg_prava.open()
                        _prava_ref['data'] = await asyncio.to_thread(_nacti_prehled_prav)
                        _prava_obsah.refresh()

                with ui.row().classes('w-full justify-end mb-4 pr-1 gap-2'):
                    if ma_vse:
                        ui.button('Přehled práv', icon='admin_panel_settings',
                                  on_click=_otevri_prava) \
                            .props('color=indigo outline no-caps')
                    ui.button('Nahrát data', icon='upload', on_click=_dlg_import.open) \
                        .props('color=primary outline no-caps')
                    ui.button('Import historických dat', icon='history',
                              on_click=_dlg_strediska.open) \
                        .props('color=teal outline no-caps')
                    ui.button(f'Import aktuálního roku ({_akt_rok})', icon='event_available',
                              on_click=_dlg_strediska_akt.open) \
                        .props('color=teal no-caps')
                    ui.button('Import Obratů a Zisků', icon='trending_up',
                              on_click=_otevri_oz) \
                        .props('color=deep-orange outline no-caps')
                    ui.button('Rozeslat e-maily o aktualizaci', icon='mark_email_read',
                              on_click=_otevri_email) \
                        .props('color=positive outline no-caps')

            with ui.grid(columns=4).classes('w-full gap-4'):
                if je_majitel:
                    with ui.card().classes(
                        'cursor-pointer hover:shadow-xl hover:-translate-y-0.5 transition-all duration-200 '
                        'p-6 items-center gap-1 rounded-2xl bg-indigo-600 hover:bg-indigo-700'
                    ) as card:
                        ui.label('📊').classes('text-5xl mb-1')
                        ui.label('Přehled poboček').classes('text-lg font-bold text-center text-white')
                        card.on('click', lambda: _otevri('__prehled__'))
                if je_ao or je_ucetni_any or je_op_ctenar or je_majitel:
                    with ui.card().classes(
                        'cursor-pointer hover:shadow-xl hover:-translate-y-0.5 transition-all duration-200 '
                        'p-6 items-center gap-1 rounded-2xl bg-teal-600 hover:bg-teal-700'
                    ) as card_op:
                        ui.label('🏭').classes('text-5xl mb-1')
                        ui.label('Ostatní provozy').classes('text-lg font-bold text-center text-white')
                        card_op.on('click', lambda: _otevri('__ostatni__'))
                for p in pristupne_pobocky:
                    with ui.card().classes(
                        'cursor-pointer hover:shadow-xl hover:-translate-y-0.5 transition-all duration-200 '
                        'p-6 items-center gap-2 rounded-2xl border border-gray-100'
                    ) as card:
                        ui.label('🏪').classes('text-5xl mb-1')
                        ui.label(_POBOCKY_EXCEL_REVERSE.get(p, p)
                                if p not in ('Bourárna', 'Kamiony') else p
                        ).classes('text-lg font-bold text-center text-gray-800')
                        card.on('click', lambda pb=p: _otevri(pb))
        elif sel == '__ostatni__':
            with ui.column().classes('w-full gap-0'):
                with ui.row().classes('items-center gap-3 mb-4'):
                    ui.button(icon='arrow_back', on_click=_zpet).props('flat round').tooltip('Zpět na výběr pobočky')
                    ui.label('Výsledky – Ostatní provozy').classes('text-2xl font-bold text-gray-800')
                _vykresli_ostatni_provozy(user_id, user_name, je_ao, je_ucetni_h)
        elif sel == '__prehled__':
            with ui.column().classes('w-full gap-0'):
                with ui.row().classes('items-center gap-3 mb-4'):
                    ui.button(icon='arrow_back', on_click=_zpet).props('flat round').tooltip('Zpět na výběr pobočky')
                    ui.label('Přehled poboček').classes('text-2xl font-bold text-gray-800')
                    _exp_tlacitko_zr(pristupne_pobocky, user_name)
                _vykresli_prehled(user_id, user_name, pristupne_pobocky, on_open=_otevri, je_ao=je_ao)
        else:
            with ui.column().classes('w-full gap-0'):
                with ui.row().classes('items-center gap-3 mb-4'):
                    ui.button(icon='arrow_back', on_click=_zpet).props('flat round').tooltip('Zpět na výběr pobočky')
                    _detail_nazev = _POBOCKY_EXCEL_REVERSE.get(sel, sel) if sel not in ('Bourárna', 'Kamiony') else sel
                    ui.label(f'Výsledky – {_detail_nazev}') \
                        .classes('text-2xl font-bold text-gray-800')
                _vykresli_detail_pobocky(sel, user_id, user_name, je_ao, je_ucetni_h, jen_podrobne)

    def _otevri(pb: str):
        app.storage.user[_state_key] = pb
        _panel.refresh()

    def _zpet():
        app.storage.user.pop(_state_key, None)
        _panel.refresh()

    _panel()


def _vykresli_detail_pobocky(pobocka: str, user_id: int, user_name: str, je_ao: bool,
                             je_ucetni: bool, jen_podrobne: bool = False):
    aktualni_rok = datetime.datetime.now().year
    minuly_rok   = aktualni_rok - 1

    # Které záložky pobočka zobrazuje:
    #   Přefakturace      → jen Tabulka nákladů
    #   Kamiony, Bourárna → Tabulka nákladů + Podrobné náklady (bez Obratů/Zisku)
    #   ostatní           → všechny čtyři
    if pobocka == 'Přefakturace':
        povolene = ['naklady']
    elif pobocka in ('Kamiony', 'Bourárna'):
        povolene = ['naklady', 'podrobne']
    else:
        povolene = ['obraty', 'zisk', 'naklady', 'podrobne']

    # Účetní běžná vidí jen Podrobné náklady (pouze pro čtení – je_ucetni=False)
    if jen_podrobne:
        povolene = [t for t in povolene if t == 'podrobne']
        if not povolene:
            ui.label('Tato pobočka nemá sekci Podrobné náklady.') \
                .classes('text-gray-400 italic py-6')
            return

    with ui.row().classes('w-full justify-end mb-2'):
        _exp_tlacitko_pobocka(pobocka, povolene, user_name)

    if 'podrobne' in povolene:
        inicializuj_osnovu(pobocka, aktualni_rok)
        inicializuj_osnovu(pobocka, minuly_rok)
        for _, klic in _VNORENE_POBOCKY.get(pobocka, []):
            inicializuj_osnovu(klic, aktualni_rok)
            inicializuj_osnovu(klic, minuly_rok)

    def _render_naklady():
        _vykresli_souhrn_nakladu(pobocka, je_ao, user_name, user_id)
        # Diskuze (chat) k pobočce – dole pod tabulkou nákladů; vidí ji a píše
        # do ní každý, kdo má k pobočce přístup.
        _vykresli_pobocka_komentare(pobocka, user_id, user_name, je_ao)

    _tab_defs = {
        'obraty':   ('📊 Obraty a prodeje',  lambda: _vykresli_obraty(pobocka, je_ao, user_name)),
        'zisk':     ('📈 Zisk / Marže',      lambda: _vykresli_zisk(pobocka, je_ao, user_name)),
        'naklady':  ('📋 Tabulka nákladů',   _render_naklady),
        'podrobne': ('🔍 Podrobné náklady',  lambda: _vykresli_podrobne_naklady(
            pobocka, aktualni_rok, minuly_rok, user_id, user_name, je_ao, je_ucetni)),
    }

    # Jediná sekce → bez záložkové lišty
    if len(povolene) == 1:
        _tab_defs[povolene[0]][1]()
    else:
        with ui.tabs().props('align=left active-color=primary indicator-color=primary').classes('w-full border-b border-gray-200') as tabs_pobocka:
            for key in povolene:
                ui.tab(key, label=_tab_defs[key][0])

        with ui.tab_panels(tabs_pobocka, value=povolene[0]).classes('w-full pt-4'):
            for key in povolene:
                with ui.tab_panel(key):
                    _tab_defs[key][1]()


@refreshable_na_klienta
def _vykresli_pobocka_komentare(pobocka: str, user_id: int, user_name: str, je_ao: bool):
    """Sekce „Komentáře" v detailu pobočky – jednoduchý chat pro všechny, kdo
    mají k pobočce přístup. U každé zprávy se eviduje, kdo ji napsal a kdy.
    Smazat zprávu může její autor nebo AO (moderace)."""
    zaznamy = nacti_pobocka_komentare(pobocka)

    def _smaz(zid):
        smaz_pobocka_komentar(zid)
        intranet_logger.log_activity(
            user_name, 'Výsledky poboček',
            f'Smazán komentář v diskuzi pobočky {pobocka} (ID {zid})')
        ui.notify('Komentář smazán.', type='info', position='top-right')
        _vykresli_pobocka_komentare.refresh()

    def _bublina(z):
        moje = int(z.get('user_id') or 0) == int(user_id) and int(user_id) > 0
        cas = z.get('cas')
        cas_txt = cas.strftime('%d.%m.%Y %H:%M') if hasattr(cas, 'strftime') else str(cas or '')
        with ui.row().classes('w-full ' + ('justify-end' if moje else 'justify-start')):
            with ui.column().classes('gap-0').style('max-width:78%'):
                if not moje:
                    ui.label(z.get('uzivatel') or '—') \
                        .classes('text-xs font-semibold text-gray-600 px-1')
                bg = 'background:#2563eb;color:#fff' if moje else 'background:#f1f5f9;color:#1e293b'
                with ui.element('div').classes('rounded-2xl px-3 py-2').style(bg):
                    ui.label(z.get('komentar') or '').classes('text-sm') \
                        .style('white-space:pre-wrap;word-break:break-word')
                with ui.row().classes('items-center gap-1 px-1 ' + ('self-end' if moje else '')):
                    ui.label(cas_txt).classes('text-gray-400').style('font-size:10px')
                    if moje or je_ao:
                        ui.button(icon='delete', on_click=lambda zid=z['id']: _smaz(zid)) \
                            .props('flat round dense size=xs color=grey-6') \
                            .tooltip('Smazat komentář')

    with ui.card().classes('w-full mt-6 p-0 rounded-2xl border border-gray-100 gap-0'):
        with ui.row().classes('items-center gap-2 px-5 pt-4 pb-1 w-full'):
            ui.icon('forum', color='primary').classes('text-2xl')
            with ui.column().classes('gap-0'):
                with ui.row().classes('items-center gap-2'):
                    ui.label('Komentáře').classes('text-lg font-bold text-gray-800')
                    ui.label(f'({len(zaznamy)})').classes('text-sm text-gray-400')

        ui.separator()

        if not zaznamy:
            with ui.column().classes('items-center w-full py-10 gap-2'):
                ui.icon('chat_bubble_outline', size='2.5rem', color='grey-4')
                ui.label('Zatím tu nikdo nic nenapsal. Napište první komentář!') \
                    .classes('text-sm text-gray-400')
        else:
            vyska = min(360, 70 + len(zaznamy) * 70)
            box = ui.scroll_area().classes('w-full').style(f'height:{vyska}px')
            with box:
                with ui.column().classes('w-full gap-2 px-3 py-2'):
                    for z in zaznamy:
                        _bublina(z)
            # po vykreslení sjet na nejnovější zprávu (dole)
            ui.timer(0.05, lambda: box.scroll_to(percent=1.0), once=True)

        ui.separator()
        with ui.row().classes('w-full items-end gap-2 px-4 py-3'):
            inp = ui.textarea(placeholder='Napište komentář… (Ctrl+Enter odešle)') \
                .props('outlined autogrow dense input-style=max-height:120px').classes('flex-1')

            def _odesli():
                txt = (inp.value or '').strip()
                if not txt:
                    return
                pridej_pobocka_komentar(pobocka, txt, user_name, user_id)
                intranet_logger.log_activity(
                    user_name, 'Výsledky poboček',
                    f'Komentář v diskuzi pobočky {pobocka}')
                inp.value = ''
                _vykresli_pobocka_komentare.refresh()

            inp.on('keydown.ctrl.enter', lambda _: _odesli())
            ui.button(icon='send', on_click=_odesli) \
                .props('round unelevated color=primary').tooltip('Odeslat (Ctrl+Enter)')


@refreshable_na_klienta
def _vykresli_souhrn_nakladu(pobocka: str, je_ao: bool, user_name: str, user_id: int = 0):
    # ── Data grid ────────────────────────────────────────────────────────────
    raw = nacti_souhrn(pobocka)

    if not raw:
        with ui.column().classes('items-center py-20 gap-3'):
            ui.icon('table_chart', size='5rem', color='grey-3')
            ui.label('Žádná data').classes('text-xl font-semibold text-gray-400')
            ui.label(
                'Data zatím nebyla nahrána. Požádejte AO o nahrání souboru na úvodní stránce.'
            ).classes('text-sm text-gray-400 text-center')
        return

    dostupne_roky = sorted({int(r['rok']) for r in raw}, reverse=True)

    # ── Sdílený default (nastavuje AO) ────────────────────────────────────────
    dk_rok    = f'souhrn_default_rok_{pobocka}'
    dk_mesice = f'souhrn_default_mesice_{pobocka}'
    dk_verze  = f'souhrn_default_ver_{pobocka}'
    default_roky   = app.storage.general.get(dk_rok) or []
    default_mesice = app.storage.general.get(dk_mesice) or []
    default_verze  = app.storage.general.get(dk_verze, 0)

    # ── Storage klíče pro filtr (per-user) ────────────────────────────────────
    fk_rok    = f'souhrn_rok_{pobocka}_{user_id}'
    fk_mesice = f'souhrn_mesice_{pobocka}_{user_id}'
    fk_dver   = f'souhrn_dver_{pobocka}_{user_id}'

    # Pokud admin změnil default (nová verze) → přepsat user filtr na nový default
    user_dver = app.storage.user.get(fk_dver, 0)
    if user_dver < default_verze:
        app.storage.user[fk_dver]   = default_verze
        app.storage.user[fk_rok]    = list(default_roky)
        app.storage.user[fk_mesice] = list(default_mesice)

    sel_roky   = app.storage.user.get(fk_rok) or []
    sel_mesice = app.storage.user.get(fk_mesice) or []
    # oprava staré single-value podoby
    if not isinstance(sel_roky, list):
        sel_roky = [sel_roky] if sel_roky else []
        app.storage.user[fk_rok] = sel_roky

    # ── Filtrování raw dat ────────────────────────────────────────────────────
    raw_f = raw
    if sel_roky:
        raw_f = [r for r in raw_f if int(r['rok']) in sel_roky]
    if sel_mesice:
        raw_f = [r for r in raw_f if int(r['mesic']) in sel_mesice]
    rows = _priprav_souhrn_pro_grid(raw_f)

    # ── Hlavička + filtr ──────────────────────────────────────────────────────
    with ui.row().classes('items-center gap-3 mb-3 flex-wrap'):
        _nazev_pob = _POBOCKY_EXCEL_REVERSE.get(pobocka, pobocka) if pobocka not in ('Bourárna', 'Kamiony') else pobocka
        ui.label(f'Pobočka: {_nazev_pob}') \
            .classes('text-base font-bold text-gray-800')

        rok_options = {r: str(r) for r in dostupne_roky}
        w_rok = ui.select(
            rok_options,
            value=list(sel_roky),
            label='Rok',
            multiple=True,
        ).props('dense outlined options-dense use-chips').style('min-width: 160px')

        mesice_opts = {i: str(i) for i in range(1, 13)}
        w_mes = ui.select(
            mesice_opts,
            value=list(sel_mesice),
            label='Měsíce',
            multiple=True,
        ).props('dense outlined options-dense use-chips').style('min-width: 160px')

        if sel_roky or sel_mesice:
            def _reset(_):
                app.storage.user[fk_rok]    = []
                app.storage.user[fk_mesice] = []
                _vykresli_souhrn_nakladu.refresh()
            ui.button('Vše', icon='clear_all', on_click=_reset) \
                .props('flat dense color=grey-7').classes('text-xs')

        # ── AO: nastavit/smazat výchozí filtr pro všechny ─────────────────────
        if je_ao:
            _je_default = (sel_roky == list(default_roky) and sel_mesice == list(default_mesice)
                           and (default_roky or default_mesice))

            def _set_default(_):
                nova_ver = app.storage.general.get(dk_verze, 0) + 1
                app.storage.general[dk_rok]    = list(sel_roky)
                app.storage.general[dk_mesice] = list(sel_mesice)
                app.storage.general[dk_verze]  = nova_ver
                app.storage.user[fk_dver]      = nova_ver
                ui.notify('Výchozí filtr nastaven pro všechny uživatele.', type='positive')
                _vykresli_souhrn_nakladu.refresh()

            def _clear_default(_):
                nova_ver = app.storage.general.get(dk_verze, 0) + 1
                app.storage.general[dk_rok]    = []
                app.storage.general[dk_mesice] = []
                app.storage.general[dk_verze]  = nova_ver
                app.storage.user[fk_dver]      = nova_ver
                ui.notify('Výchozí filtr odstraněn.', type='info')
                _vykresli_souhrn_nakladu.refresh()

            if not _je_default:
                ui.button('Nastavit jako výchozí', icon='lock',
                          on_click=_set_default) \
                    .props('flat dense color=primary').classes('text-xs') \
                    .tooltip('Uloží aktuální filtr (rok + měsíc) jako výchozí pro všechny uživatele')
            if default_roky or default_mesice:
                _def_txt_parts = []
                if default_roky:
                    _def_txt_parts.append(f'Rok: {", ".join(str(r) for r in default_roky)}')
                if default_mesice:
                    _def_txt_parts.append(f'Měs: {", ".join(str(m) for m in default_mesice)}')
                ui.label(f'🔒 Výchozí: {" · ".join(_def_txt_parts)}') \
                    .classes('text-xs text-blue-500 italic')
                if _je_default:
                    ui.button('Zrušit výchozí', icon='lock_open',
                              on_click=_clear_default) \
                        .props('flat dense color=grey-6').classes('text-xs')

        def _zmena_rok(e):
            app.storage.user[fk_rok] = list(e.value) if e.value else []
            _vykresli_souhrn_nakladu.refresh()
        w_rok.on_value_change(_zmena_rok)

        def _zmena_mes(e):
            app.storage.user[fk_mesice] = list(e.value) if e.value else []
            _vykresli_souhrn_nakladu.refresh()
        w_mes.on_value_change(_zmena_mes)

    if not rows:
        ui.label('Žádná data pro vybraný filtr.').classes('text-gray-400 italic py-4')
        return

    ui.add_css(_TN_CSS)
    ui.aggrid({
        'columnDefs': _col_defs_souhrn(),
        'rowData': rows,
        'defaultColDef': {'resizable': True, 'sortable': False, 'wrapHeaderText': True, 'autoHeaderHeight': True},
        'rowHeight': 34,
        'suppressMovableColumns': True,
        'domLayout': 'autoHeight',
        ':onFirstDataRendered': _AUTOSIZE_FIT,
        ':onGridSizeChanged': _AUTOSIZE_FIT,
        ':getRowStyle': (
            "function(p){"
            "if(p.data&&p.data._je_celkem)"
            "return{fontWeight:'bold',backgroundColor:'#DCE6F2',borderTop:'1px solid #b8cce4'};"
            "return null;}"
        ),
    }).classes('w-full tn-grid')


def _vykresli_podrobne_naklady(
    pobocka: str, aktualni_rok: int, minuly_rok: int,
    user_id: int, user_name: str, je_ao: bool, je_ucetni: bool,
):
    vnorene = _VNORENE_POBOCKY.get(pobocka, [])

    if vnorene:
        vsechny = [(pobocka, pobocka)] + vnorene
        ui.add_css('.naklady-tabs .q-tab__label{white-space:pre-line;text-align:center;line-height:1.3;}')

        with ui.tabs().props('align=left active-color=teal indicator-color=teal') \
                .classes('w-full naklady-tabs') as tabs_naklady:
            for i, (nazev, _) in enumerate(vsechny):
                t = 'akt' if i == 0 else f'akt_{i}'
                ui.tab(t, label=f'📅 {aktualni_rok} – Aktuální rok\n{nazev}')
            ui.tab('por', label='⚖️ Porovnání')
            for i, (nazev, _) in enumerate(vsechny):
                t = 'min' if i == 0 else f'min_{i}'
                ui.tab(t, label=f'📝 {minuly_rok} – Minulý rok\n{nazev}')

        def _make_akt_refreshable(klic):
            @ui.refreshable
            def _tbl():
                _tabulka_aktualni(klic, aktualni_rok, user_id, user_name, je_ucetni, _refresh_cb=_tbl.refresh)
            return _tbl

        akt_tbls = [_make_akt_refreshable(klic) for _, klic in vsechny]

        with ui.tab_panels(tabs_naklady, value='akt').classes('w-full pt-4'):
            for i, tbl in enumerate(akt_tbls):
                t = 'akt' if i == 0 else f'akt_{i}'
                with ui.tab_panel(t):
                    tbl()
            with ui.tab_panel('por'):
                _tabulka_porovnani(pobocka, aktualni_rok, je_ucetni)
            for i, (_, klic) in enumerate(vsechny):
                t = 'min' if i == 0 else f'min_{i}'
                with ui.tab_panel(t):
                    _tabulka_aktualni(klic, minuly_rok, user_id, user_name, je_ucetni,
                                      zamky_aktivni=False, povolit_pridat=False)
    else:
        @ui.refreshable
        def _tbl_akt():
            _tabulka_aktualni(pobocka, aktualni_rok, user_id, user_name, je_ucetni, _refresh_cb=_tbl_akt.refresh)

        with ui.tabs().props('align=left active-color=teal indicator-color=teal').classes('w-full') as tabs_naklady:
            ui.tab('akt', label=f'📅 {aktualni_rok} – Aktuální rok')
            ui.tab('min', label=f'📝 {minuly_rok} – Minulý rok')
            ui.tab('por', label='⚖️ Porovnání')

        with ui.tab_panels(tabs_naklady, value='akt').classes('w-full pt-4'):
            with ui.tab_panel('akt'):
                _tbl_akt()
            with ui.tab_panel('min'):
                _tabulka_aktualni(pobocka, minuly_rok, user_id, user_name, je_ucetni,
                                  zamky_aktivni=False, povolit_pridat=False)
            with ui.tab_panel('por'):
                _tabulka_porovnani(pobocka, aktualni_rok, je_ucetni)


def _hist_field_label(pole: str) -> str:
    """Lidský název editovaného pole pro historii."""
    if pole in MESICE_DB:
        return MESICE_NAZVY[MESICE_DB.index(pole)]
    return {'ucetni_predpis': 'Předpis', 'nazev_predpisu': 'Název'}.get(pole, pole)


def _hist_fmt_val(pole: str, v) -> str:
    """Naformátuje hodnotu z historie (čísla cs-CZ, text beze změny)."""
    if v is None or str(v) == '':
        return '—'
    if pole in MESICE_DB:
        # cs-CZ: NBSP oddělovač tisíců + čárka (shodně s Intl.NumberFormat jinde)
        try:
            return (f'{float(v):,.2f}'.replace(',',' ').replace('.', ','))
        except (TypeError, ValueError):
            return str(v)
    return str(v)


def _zobraz_historii_radku(row_id: int, nazev: str):
    """Dialog s historií úprav jednoho řádku nákladů (kdo, kdy, z čeho na co)."""
    zaznamy = nacti_historii(row_id)
    with ui.dialog() as dlg, ui.card().classes('p-5 gap-3') \
            .style('min-width: 560px; max-width: 820px; max-height: 80vh; overflow-y: auto'):
        with ui.row().classes('items-center gap-2'):
            ui.icon('history', color='teal')
            ui.label(f'Historie úprav – {nazev}'.strip(' –')).classes('text-lg font-bold text-gray-800')
        if not zaznamy:
            ui.label('Pro tento řádek zatím nebyly zaznamenány žádné úpravy.') \
                .classes('text-sm text-gray-500 italic py-4')
        else:
            columns = [
                {'name': 'cas', 'label': 'Čas', 'field': 'cas', 'align': 'left'},
                {'name': 'uzivatel', 'label': 'Uživatel', 'field': 'uzivatel', 'align': 'left'},
                {'name': 'pole', 'label': 'Pole', 'field': 'pole', 'align': 'left'},
                {'name': 'zmena', 'label': 'Změna', 'field': 'zmena', 'align': 'left'},
            ]
            t_rows = []
            for z in zaznamy:
                cas = z.get('cas')
                cas_txt = cas.strftime('%d.%m.%Y %H:%M') if hasattr(cas, 'strftime') else str(cas or '')
                pole = z.get('pole', '')
                t_rows.append({
                    'cas': cas_txt,
                    'uzivatel': z.get('uzivatel') or '—',
                    'pole': _hist_field_label(pole),
                    'zmena': f"{_hist_fmt_val(pole, z.get('stara_hodnota'))} "
                             f"→ {_hist_fmt_val(pole, z.get('nova_hodnota'))}",
                })
            ui.table(columns=columns, rows=t_rows, row_key='cas') \
                .classes('w-full').props('dense flat bordered')
            ui.label(f'Celkem záznamů: {len(zaznamy)}').classes('text-xs text-gray-400')
        with ui.row().classes('justify-end w-full'):
            ui.button('Zavřít', on_click=dlg.close).props('flat no-caps')
    dlg.open()


def _tabulka_aktualni(pobocka: str, rok: int, user_id: int, user_name: str, je_ucetni: bool,
                      zamky_aktivni: bool = True, povolit_pridat: bool = True,
                      _refresh_cb=None):
    """Editovatelná tabulka podrobných nákladů.
    • zamky_aktivni=True  → aktuální rok: zámky měsíců (klik na hlavičku), pro
      ne-účetní vynulují zamčené měsíce.
    • zamky_aktivni=False → historická data (minulý rok): bez zámků, celý rok
      viditelný; účetní může editovat hodnoty i názvy.
    • povolit_pridat → tlačítka Přidat/Smazat řádek (jen aktuální rok – osnova se
      odtud synchronizuje i do minulého roku)."""
    rows_raw = nacti_naklady(pobocka, rok)
    zamcene = _nacti_zamky(pobocka, rok) if zamky_aktivni else []

    # Účetní vidí skutečná data; ostatní mají zamčené měsíce vynulované.
    # Bez zámků (historická data) je celý rok viditelný pro všechny.
    rows = rows_raw if (je_ucetni or not zamky_aktivni) else _zamknout_data(rows_raw, zamcene)

    # naplní se po vytvoření filtru; vrací aktuálně zobrazené (vyfiltrované) řádky
    zobrazene_ref: dict = {}

    def _prepocet_souctu():
        zobr = zobrazene_ref.get('fn')
        disp = zobr() if zobr else rows
        grid.run_grid_method('setGridOption', 'pinnedBottomRowData',
                             [_soucet_naklady(disp)])

    def _on_cell_change(e):
        args   = e.args if hasattr(e, 'args') else {}
        field  = args.get('colId') or ''
        data   = args.get('data') or {}
        new_v  = args.get('newValue')
        old_v  = args.get('oldValue')
        row_id = data.get('id')
        if not row_id:
            return
        if field in MESICE_DB:
            try:
                if new_v in (None, ''):
                    val = 0.0
                else:
                    val = float(str(new_v).replace('\xa0', '').replace(' ', '').replace(',', '.'))
            except (TypeError, ValueError):
                val = 0.0
            if _s(old_v) != val:
                uloz_hodnotu(row_id, field, val)
                zaznam_historie(row_id, pobocka, rok, field, _s(old_v), val, user_name)
            for r in rows:
                if r.get('id') == row_id:
                    r[field]    = val
                    r['celkem'] = sum(_s(r[m]) for m in MESICE_DB)
                    break
            _prepocet_souctu()
        elif field == 'ucetni_predpis':
            if str(old_v or '') != str(new_v or ''):
                uloz_predpis(row_id, str(new_v or ''), data.get('nazev_predpisu', ''))
                zaznam_historie(row_id, pobocka, rok, 'ucetni_predpis',
                                old_v, new_v, user_name)
            for r in rows:
                if r.get('id') == row_id:
                    r['ucetni_predpis'] = str(new_v or '')
                    break
        elif field == 'nazev_predpisu':
            if str(old_v or '') != str(new_v or ''):
                uloz_predpis(row_id, data.get('ucetni_predpis', ''), str(new_v or ''))
                zaznam_historie(row_id, pobocka, rok, 'nazev_predpisu',
                                old_v, new_v, user_name)
            for r in rows:
                if r.get('id') == row_id:
                    r['nazev_predpisu'] = str(new_v or '')
                    break

    if je_ucetni and zamky_aktivni:
        ui.add_css(_ZAMEK_CSS)

    # Komentáře k buňkám (červený růžek + tooltip) – vidí všichni, přidává jen účetní.
    ui.add_css(_KOMENTAR_CSS)

    # Informační pruh pro historická data (minulý rok bez zámků)
    if not zamky_aktivni:
        with ui.row().classes('items-center gap-2 mb-1'):
            ui.icon('history', color='teal', size='sm')
            ui.label(
                f'Historická data za rok {rok} – '
                + ('hodnoty i názvy lze upravovat, změny se ukládají automaticky.'
                   if je_ucetni else 'celý rok, pouze pro čtení.')
            ).classes('text-sm text-gray-600 italic')

    # Toolbar (vykreslí se NAD gridem, aby nekolidoval s řádky)
    toolbar = ui.row().classes('w-full mb-3 gap-3 items-center')

    grid = ui.aggrid({
        # Očičko historie vidí jen Komentátor AO / Komentátor účetní (= je_ucetni).
        'columnDefs': _col_defs_naklady(je_ucetni, zamcene if (je_ucetni and zamky_aktivni) else None,
                                        s_historii=je_ucetni),
        'rowData': rows,
        'pinnedBottomRowData': [_soucet_naklady(rows)],
        'rowSelection': 'single' if je_ucetni else 'none',
        'stopEditingWhenCellsLoseFocus': True,
        'singleClickEdit': False,
        'defaultColDef': {
            'resizable': True, 'sortable': False, 'wrapHeaderText': True, 'autoHeaderHeight': True,
            # Komentáře k buňkám: růžek (cellClassRules) + tooltip s textem.
            'cellClassRules': {':ma-komentar': _KOMENTAR_CLASSRULE},
            ':tooltipValueGetter': _KOMENTAR_TOOLTIP,
        },
        'rowHeight': 34,
        'suppressMovableColumns': True,
        'tooltipShowDelay': 400,
        # Účetní: pravým klikem se otevře komentář → potlačíme nativní menu prohlížeče.
        'preventDefaultOnContextMenu': je_ucetni,
        ':getRowStyle': _PINNED_BOTTOM_STYLE,
        'undoRedoCellEditing': je_ucetni,
        'undoRedoCellEditingLimit': 50,
        'enterNavigatesVertically': True,
        'enterNavigatesVerticallyAfterEdit': True,
        ':onFirstDataRendered': _AUTOSIZE_FIT,
        ':onGridSizeChanged': _AUTOSIZE_FIT,
    }).classes('w-full komentare-grid').style(_GRID_STYLE)

    with toolbar:
        zobrazene_ref['fn'] = _pridej_filtr_nazvu(
            rows, grid, _soucet_naklady, f'naklady_{pobocka}_{rok}')
        if je_ucetni:
            with ui.row().classes('items-center gap-1 text-gray-400') \
                    .tooltip('Klikněte pravým tlačítkem na buňku (měsíc, Předpis nebo Název) '
                             'a přidejte k ní komentář – jako v Excelu.'):
                ui.icon('sticky_note_2', size='xs')
                ui.label('Pravý klik = komentář k buňce').classes('text-xs italic')

    # Očičko „Historie úprav" – jen pro Komentátora AO / účetního (je_ucetni).
    if je_ucetni:
        def _on_cell_clicked(e):
            a = e.args or {}
            if a.get('colId') != '_historie':
                return
            data = a.get('data') or {}
            rid = data.get('id')
            if rid:
                _zobraz_historii_radku(rid, data.get('nazev_predpisu', ''))
        grid.on('cellClicked', _on_cell_clicked)

        # ─── Komentáře k buňkám (pravý klik = jako v Excelu „Vložit komentář") ───
        def _aktualizuj_komentare_grid():
            """Po změně komentáře pošle do gridu aktuálně zobrazené řádky
            (respektuje filtr názvu) → překreslí růžky a tooltipy."""
            fn = zobrazene_ref.get('fn')
            disp = fn() if fn else rows
            grid.run_grid_method('setGridOption', 'rowData', disp)

        def _komentar_dialog(row_id: int, field: str, data_row: dict):
            existing = nacti_komentar(row_id, field)
            label = _hist_field_label(field)
            nazev_radku = (data_row or {}).get('nazev_predpisu', '') or '(bez názvu)'
            with ui.dialog() as dlg, ui.card().classes('p-5 gap-3') \
                    .style('min-width: 460px; max-width: 640px'):
                with ui.row().classes('items-center gap-2'):
                    ui.icon('sticky_note_2', color='amber-8')
                    ui.label('Komentář k buňce').classes('text-lg font-bold text-gray-800')
                ui.label(f'{nazev_radku}  •  {label}').classes('text-sm text-gray-600')
                ta = ui.textarea(
                    value=(existing or {}).get('komentar', '') or '',
                    placeholder='Napište komentář k této buňce…',
                ).props('outlined autogrow autofocus').classes('w-full')
                if existing:
                    kdo = existing.get('uzivatel') or '—'
                    cas = existing.get('cas')
                    cas_txt = cas.strftime('%d.%m.%Y %H:%M') if hasattr(cas, 'strftime') else str(cas or '')
                    ui.label(f'Naposledy upravil: {kdo} • {cas_txt}').classes('text-xs text-gray-400')

                def _uloz():
                    text = (ta.value or '').strip()
                    if not text:
                        if existing:
                            _smaz()
                        else:
                            dlg.close()
                        return
                    uloz_komentar(row_id, field, text, user_name)
                    for r in rows:
                        if r.get('id') == row_id:
                            r.setdefault('_komentare', {})[field] = text
                            break
                    intranet_logger.log_activity(
                        user_name, 'Výsledky poboček',
                        f'Komentář k buňce „{label}" (řádek ID {row_id}): {pobocka} {rok}',
                    )
                    _aktualizuj_komentare_grid()
                    ui.notify('Komentář uložen.', type='positive', position='top-right')
                    dlg.close()

                def _smaz():
                    smaz_komentar(row_id, field)
                    for r in rows:
                        if r.get('id') == row_id:
                            kom = r.get('_komentare')
                            if kom and field in kom:
                                del kom[field]
                            break
                    intranet_logger.log_activity(
                        user_name, 'Výsledky poboček',
                        f'Smazán komentář k buňce „{label}" (řádek ID {row_id}): {pobocka} {rok}',
                    )
                    _aktualizuj_komentare_grid()
                    ui.notify('Komentář smazán.', type='info', position='top-right')
                    dlg.close()

                with ui.row().classes('justify-end w-full gap-2 mt-1'):
                    if existing:
                        ui.button('Smazat', on_click=_smaz).props('flat color=negative no-caps')
                    ui.button('Zrušit', on_click=dlg.close).props('flat no-caps')
                    ui.button('Uložit', on_click=_uloz).props('color=primary unelevated no-caps')
            dlg.open()

        def _on_cell_context(e):
            a = e.args or {}
            if a.get('rowPinned'):
                return
            field = a.get('colId') or ''
            if field not in _KOMENTAR_POLE:
                return
            data = a.get('data') or {}
            rid = data.get('id')
            if rid:
                _komentar_dialog(rid, field, data)
        grid.on('cellContextMenu', _on_cell_context)

    if je_ucetni:
        grid.on('cellValueChanged', _on_cell_change)

        import json as _json
        _mesice_js = _json.dumps(MESICE_DB)
        _nazvy_js = _json.dumps(MESICE_NAZVY)
        _grid_id = grid.id
        _zamek_event = f'zamek_toggle_{_grid_id}'
        # Unikátní marker class na grid container — slouží pro scoped document-level
        # click delegation (`getElement(id).$el` v NiceGUI 3.8 nevracel spolehlivě
        # DOM root ag-gridu, proto cílíme přes CSS třídu, kterou NiceGUI dává na
        # vlastní container DIV — stejný mechanismus, jakým funguje `komentare-grid`).
        _grid_marker = f'zamek-g-{_grid_id}'
        if zamky_aktivni:
            grid.classes(_grid_marker)

        def _on_header_click(e):
            a = e.args
            col_id = a if isinstance(a, str) else (a[0] if isinstance(a, list) and a else '')
            if col_id not in MESICE_DB:
                return
            mi = MESICE_DB.index(col_id) + 1
            z = _nacti_zamky(pobocka, rok)
            if mi in z:
                z.remove(mi)
            else:
                z.append(mi)
            _uloz_zamky(pobocka, rok, z)
            # Přepsat jen text a CSS headeru přímo v DOM — žádný column reset
            zamky_js = _json.dumps(z)
            ui.run_javascript(
                f"(function(){{"
                f"var root=document.querySelector('.{_grid_marker}');if(!root)return;"
                f"var cols={_mesice_js};var names={_nazvy_js};var zamky={zamky_js};"
                "root.querySelectorAll('.ag-header-cell').forEach(function(hc){"
                "var cid=hc.getAttribute('col-id');"
                "var idx=cols.indexOf(cid);"
                "if(idx<0)return;"
                "var mi=idx+1;var locked=zamky.indexOf(mi)>=0;"
                "var txt=hc.querySelector('.ag-header-cell-text');"
                "if(txt)txt.textContent=(locked?'🔒 ':'🔓 ')+names[idx];"
                "if(locked){hc.classList.add('zamceny-mesic');}else{hc.classList.remove('zamceny-mesic');}"
                "});"
                f"}})();"
            )

        if zamky_aktivni:
            # Per-grid event name → každý grid má vlastní kanál, žádné kolize.
            ui.on(_zamek_event, _on_header_click)

        def _install_header_clicks():
            # Event delegation na grid containeru (CSS marker) — přežije re-render
            # headerů ag-Gridem (setGridOption rowData/pinnedBottomRowData), na
            # rozdíl od onclick připojeného k jednotlivým .ag-header-cell.
            # Retry polling do 5 s, kdyby grid ještě nebyl v DOM.
            ui.run_javascript(
                f"(function(){{"
                f"var marker='.{_grid_marker}';"
                f"var cols={_mesice_js};"
                f"var ev_name={_json.dumps(_zamek_event)};"
                "function tryInstall(){"
                "var root=document.querySelector(marker);"
                "if(!root)return false;"
                "if(root._zamekDelegate)return true;"
                "root._zamekDelegate=true;"
                "root.addEventListener('click',function(ev){"
                "var hc=ev.target&&ev.target.closest?ev.target.closest('.ag-header-cell'):null;"
                "if(!hc)return;"
                "if(!root.contains(hc))return;"
                "var cid=hc.getAttribute('col-id');"
                "if(cols.indexOf(cid)<0)return;"
                "ev.stopPropagation();"
                "emitEvent(ev_name,cid);"
                "},true);"
                "return true;"
                "}"
                "if(tryInstall())return;"
                "var iv=setInterval(function(){if(tryInstall())clearInterval(iv);},200);"
                "setTimeout(function(){clearInterval(iv);},5000);"
                f"}})();"
            )
        if zamky_aktivni:
            ui.timer(0.3, _install_header_clicks, once=True)

        def _pridat():
            ok = pridej_radek_synced(pobocka, rok)
            if ok:
                intranet_logger.log_activity(
                    user_name, 'Výsledky poboček',
                    f'Přidán nový řádek nákladů: {pobocka} {rok}',
                )
                ui.notify('Nový řádek byl přidán.', type='positive', position='top-right')
                if _refresh_cb: _refresh_cb()
            else:
                ui.notify('Chyba při přidávání řádku.', type='negative')

        async def _smazat():
            sel = await grid.get_selected_rows()
            if not sel:
                ui.notify('Nejprve klikněte na řádek, který chcete smazat.', type='warning')
                return
            row_id = sel[0].get('id')
            if not row_id:
                return

            def _potvrdit():
                smaz_radek_synced(row_id, pobocka, rok)
                intranet_logger.log_activity(
                    user_name, 'Výsledky poboček',
                    f'Smazán řádek ID {row_id}: {pobocka} {rok}',
                )
                ui.notify('Řádek byl smazán.', type='positive', position='top-right')
                if _refresh_cb: _refresh_cb()
                dlg.close()

            with ui.dialog() as dlg, ui.card().classes('p-6 gap-4'):
                ui.label('Smazat řádek?').classes('text-lg font-bold')
                predpis = sel[0].get('ucetni_predpis', '')
                nazev   = sel[0].get('nazev_predpisu', '')
                ui.label(f'{predpis} – {nazev}').classes('text-sm text-gray-600')
                ui.label('Řádek bude smazán i z minulého roku (synchronizace porovnání).').classes('text-xs text-orange-600')
                with ui.row().classes('justify-end gap-2 mt-2'):
                    ui.button('Zrušit', on_click=dlg.close).props('flat no-caps')
                    ui.button('Smazat', on_click=_potvrdit).props('color=negative no-caps')
            dlg.open()

        if povolit_pridat:
            with toolbar:
                ui.button(icon='add', text='Přidat řádek', on_click=_pridat).props('color=primary unelevated no-caps')
                ui.button(icon='delete', text='Smazat vybraný', on_click=_smazat).props('color=negative outline no-caps')
    else:
        with toolbar:
            ui.label('Pouze pro čtení.').classes('text-xs text-gray-400 italic')


@refreshable_na_klienta
def _tabulka_porovnani(pobocka: str, aktualni_rok: int, je_ucetni: bool = False):
    rows       = nacti_porovnani(pobocka, aktualni_rok)
    minuly_rok = aktualni_rok - 1
    if not je_ucetni:
        # Zámky se uplatňují jen na aktuální rok; minulý rok (historická data)
        # je bez zámků a zobrazuje se celý.
        zamky_akt = _nacti_zamky(pobocka, aktualni_rok)
        if zamky_akt:
            zdb_akt = [MESICE_DB[m - 1] for m in zamky_akt if 1 <= m <= 12]
            for r in rows:
                for db in zdb_akt:
                    r[f'akt_{db}'] = 0
                    r[f'diff_{db}'] = r.get(f'akt_{db}', 0) - r.get(f'min_{db}', 0)
                r['akt_celkem']  = sum(r.get(f'akt_{m}', 0) for m in MESICE_DB)
                r['min_celkem']  = sum(r.get(f'min_{m}', 0) for m in MESICE_DB)
                r['diff_celkem'] = r['akt_celkem'] - r['min_celkem']
    with ui.column().classes('w-full gap-2'):
        with ui.row().classes('items-center gap-3 mb-1 flex-wrap'):
            ui.icon('compare_arrows', color='teal', size='sm')
            ui.label(f'Porovnání {aktualni_rok} vs {minuly_rok}').classes('text-sm font-semibold text-gray-700')
            ui.label('Δ = Aktuální rok − Minulý rok').classes('text-xs text-gray-500')
            ui.label('🔴 nárůst nákladů  🟢 pokles nákladů').classes('text-xs text-gray-500')
            filtr_slot = ui.row().classes('items-center gap-1 ml-auto')
        grid = ui.aggrid({
            'columnDefs': _col_defs_porovnani(aktualni_rok),
            'rowData': rows,
            'pinnedBottomRowData': [_soucet_porovnani(rows)],
            'defaultColDef': {'resizable': True, 'sortable': False, 'wrapHeaderText': True, 'autoHeaderHeight': True},
            'rowHeight': 34,
            'suppressMovableColumns': True,
            ':getRowStyle': _PINNED_BOTTOM_STYLE,
            ':onFirstDataRendered': _AUTOSIZE_FIT,
            ':onGridSizeChanged': _AUTOSIZE_FIT,
        }).classes('w-full').style(_GRID_STYLE)
        with filtr_slot:
            _pridej_filtr_nazvu(rows, grid, _soucet_porovnani,
                                f'naklady_porovnani_{pobocka}_{aktualni_rok}')


# ─── Obraty / Zisk – UI ───────────────────────────────────────────────────────

# ═══ Obraty/Zisk – věrná reprodukce listů (6 tabulek/list) ═══════════════════
# Popisný text u jednotlivých skupin (zachován dle Excelu; filtry vynechány).
_OZ_SKUP_PORADI = ['mesice', 'dcery', 'site']

# Tabulky plněné z centrálního souboru „Tabulka nákladů_DATA.xlsx" (list DATA).
# (skupina, sekce, název sloupce v souboru, titulek karty)
_OZN_KARTY = [
    ('staropramen', 'kg',  'Obrat v Kg Staropramen', 'Meziroční porovnání prodejů Staropramen'),
    ('mesice',      'ops', 'Operace',                'Meziroční porovnání operací po měsících'),
    ('dcery',       'ops', 'Operace dcery',          'Meziroční porovnání prodejů na DCERY'),
    ('site',        'ops', 'Operace sítě',           'Meziroční porovnání prodejů na SÍTĚ'),
]
_OZ_TEXTY = {
    'obraty': {
        'titulek': 'Porovnání obratů pobočky',
        'sekce': [('kc', 'Obrat v Kč bez DPH'), ('kg', 'Obrat v Kg')],
        'skupiny': {
            'mesice': ('Meziroční porovnání obratů po měsících',
                       ['Výdajové daňové doklady',
                        'Včetně dokladů natížení centrála a drobného prodeje',
                        'Bez obalů', 'Nejsou započítané DCERY a SÍTĚ']),
            'dcery':  ('Meziroční porovnání prodejů na DCERY',
                       ['Prodeje na dcery = převodky + faktury na dceřiné společnosti']),
            'site':   ('Meziroční porovnání prodejů na SÍTĚ',
                       ['Aramark, Benzina, Compass Group, Delikomat, Eurobit, Eurooil, Jamaro,',
                        'JLV, Lagardere, MOL, OMV, Shell, Sodexo, UGO, Unixan, Very Goodies']),
        },
    },
    'zisk': {
        'titulek': 'Porovnání Nzisků pobočky',
        'sekce': [('kc', 'NZisk v Kč bez DPH')],
        'skupiny': {
            'mesice': ('Meziroční porovnání zisků po měsících',
                       ['Výdajové daňové doklady',
                        'Včetně dokladů natížení centrála a drobného prodeje',
                        'Bez obalů', 'Nejsou započítané DCERY a SÍTĚ']),
            'dcery':  ('Meziroční porovnání zisků na DCERY',
                       ['Zisk na dcery = převodky + faktury na dceřiné společnosti']),
            'site':   ('Meziroční porovnání zisků na SÍTĚ',
                       ['Aramark, Benzina, Compass Group, Delikomat, Eurobit, Eurooil, Jamaro,',
                        'JLV, Lagardere, MOL, OMV, Shell, Sodexo, UGO, Unixan, Very Goodies']),
        },
        'marze': ('Marže v % počítaná z celkových obratů a zisků',
                  {'mesice': 'bez sítí a dcer', 'dcery': 'na dcery', 'site': 'na sítě'}),
    },
}

# CSS / JS pro reprodukci (Arial, tmavá hlavička sloupce „Porovnání", zelené buňky).
_OZ_CSS = (
    ".oz-grid{height:auto!important;}"   # umožní domLayout autoHeight (bez posuvníku)
    ".oz-grid .ag-cell,.oz-grid .ag-header-cell{font-family:Arial,Helvetica,sans-serif;font-size:12px;}"
    ".oz-grid .ag-header-cell.oz-h-dark{background-color:#404040!important;}"
    ".oz-grid .ag-header-cell.oz-h-dark .ag-header-cell-text{color:#fff;font-weight:700;}"
)
_OZ_POROVNANI_GETTER = (
    "function(p){if(!p.data)return null;"
    "var a=parseFloat(p.data.y_old),b=parseFloat(p.data.y_new);"
    "if(isNaN(a)||a===0||isNaN(b))return null;return b/a;}"
)
# Obraty/zisky se zobrazují jako celá Kč/Kg (bez haléřů) – velká čísla (miliony až
# miliardy) by jinak roztáhla sloupce a ve 3-sloupcovém rozložení karet by se
# tabulka „srazila" a oříznul by se sloupec „Porovnání". Uložená hodnota i editace
# si plnou přesnost zachovávají (mění se jen formát zobrazení).
_OZ_FMT_NUM0 = (
    "function(p){"
    "if(p.value===null||p.value===undefined)return '—';"
    "var n=parseFloat(p.value);"
    "if(isNaN(n)||n===0)return '—';"
    "return new Intl.NumberFormat('cs-CZ',{maximumFractionDigits:0}).format(n);"
    "}"
)
_OZ_CELKEM_CS = (
    "function(p){return (p.data&&p.data._celkem)?{fontWeight:'bold'}:null;}"
)
_OZ_POROVNANI_CS = (
    "function(p){if(p.data&&p.data._celkem)"
    "return{backgroundColor:'#404040',color:'#fff',fontWeight:'bold'};"
    "return{backgroundColor:'#EBF1DE'};}"
)
# Parser zadané hodnoty: povolí mezery (oddělovač tisíců) i čárku jako desetinnou.
_OZ_VALUE_PARSER = (
    "function(p){var v=p.newValue;"
    "if(v===null||v===undefined||v==='')return null;"
    "var s=String(v).replace(/[\\s\\u00a0]/g,'').replace(',','.');"
    "var n=parseFloat(s);return isNaN(n)?null:n;}"
)
_OZ_COLMAP = {'mesice': (3, 4, 5), 'dcery': (9, 10, 11), 'site': (14, 15, 16)}  # (měsíc, rok1, rok2)
_OZ_SEKCE_TITUL = {
    'obraty': [('kc', 'Obrat v Kč bez DPH'), ('kg', 'Obrat v Kg')],
    'zisk':   [('kc', 'Zisk v Kč bez DPH')],
}


def nacti_oz_data(pobocka: str, list_klic: str) -> dict:
    """Vrací vnořený dict d[skupina][sekce][rok][mesic] = hodnota (mesic 0 = Celkem)."""
    d: dict = {}
    conn = intranet_data.get_db_connection()
    if not conn:
        return d
    try:
        cur = conn.cursor()
        cur.execute(
            'SELECT skupina, sekce, rok, mesic, hodnota FROM vysledky_oz_data '
            'WHERE pobocka=%s AND list=%s', (pobocka, list_klic),
        )
        for skup, sek, rok, mes, hod in cur.fetchall():
            d.setdefault(skup, {}).setdefault(sek, {}).setdefault(int(rok), {})[int(mes)] = _s(hod)
        return d
    except Exception as exc:
        print(f'[vysledky] nacti_oz_data error: {exc}')
        return d
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()


def uloz_oz_hodnota(pobocka, list_klic, skupina, sekce, rok, mesic, hodnota):
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.execute(
            'INSERT INTO vysledky_oz_data (pobocka,list,skupina,sekce,rok,mesic,hodnota) '
            'VALUES (%s,%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE hodnota=VALUES(hodnota)',
            (pobocka, list_klic, skupina, sekce, int(rok), int(mesic), float(hodnota)),
        )
        conn.commit()
    except Exception as exc:
        print(f'[vysledky] uloz_oz_hodnota error: {exc}')
    finally:
        try:
            cur.close()
        except Exception:
            pass
        conn.close()


def _oz_parsuj_sheets(sheets: dict, progress=None) -> tuple[list[tuple], str]:
    """Z dvojice listů {'obraty': ws, 'zisk': ws} vytáhne surové měsíční hodnoty
    (Obraty Kč+Kg, Zisk Kč). Sekce „Marže v %" se nečte – dopočítává se z Obratů
    a Zisku. progress = volitelný callback(podíl 0–1, název listu) před čtením
    každého listu. Vrací (radky, chyba); radky = (list, skupina, sekce, rok,
    mesic, hodnota) – bez pobočky, tu doplní volající."""
    radky: list[tuple] = []
    try:
        for si, (list_klic, ws) in enumerate(sheets.items()):
            if progress:
                try:
                    progress(si / len(sheets), ws.title)
                except Exception:
                    pass
            maxr = ws.max_row
            for sekce_key, titul in _OZ_SEKCE_TITUL[list_klic]:
                # Titulek sekce hledáme ve sloupcích VŠECH tří skupin
                # (C=měsíce, I=dcery, N=sítě), ne jen ve sloupci C. Některé
                # exporty (např. Praha) mají v Kg přehledu kvůli šablonovému
                # překlepu sloupec C chybně označený „Obrat v Kč bez DPH" a
                # správné „Obrat v Kg" zůstane jen ve sloupci N – při hledání
                # pouze ve sloupci C by se zarovnaný blok přeskočil a sekce „Kg"
                # by se navázala na pozdější detailní pivot bez dat sítí (sítě
                # by se neimportovaly). Bereme první řádek, kde titulek sedí v
                # libovolném sloupci skupiny A ZÁROVEŇ je pod ním řádek s roky –
                # tím se vyhneme falešným shodám bez dat.
                # tolerujeme volitelné „N" na začátku (soubory mají „Zisk…"
                # i „NZisk…") a malé odlišnosti v mezerách
                trow = None
                for r in range(1, maxr + 1):
                    hit = False
                    for col in (3, 9, 14):
                        cv = ws.cell(r, col).value
                        if cv is not None and str(cv).strip() in (titul, 'N' + titul):
                            hit = True
                            break
                    if hit and (_oz_year(ws.cell(r + 1, 4).value)
                                or _oz_year(ws.cell(r + 1, 5).value)):
                        trow = r
                        break
                if trow is None:
                    continue
                y1 = _oz_num(ws.cell(trow + 1, 4).value)
                y2 = _oz_num(ws.cell(trow + 1, 5).value)
                if not y1 or not y2:
                    continue
                y1, y2 = int(y1), int(y2)
                # Řádky určujeme podle POPISKU ve sloupci s číslem měsíce (cm),
                # ne podle pevného offsetu – jednotlivé bloky (měsíce/dcery/sítě)
                # mohou být u různých poboček svisle posunuté. Měsíc = číslo 1–12,
                # součtový řádek „Celkový součet" → mesic 0. Tím se zabrání tomu,
                # aby roční součet spadl do prosince (měsíc 12).
                for skup, (cm, c1, c2) in _OZ_COLMAP.items():
                    # Sken začínáme už od trow+1 – některé bloky (dcery/sítě) bývají
                    # o řádek výš než blok „měsíce" (rok/„Měsíc – číslo" se přeskočí,
                    # protože ve sloupci s číslem měsíce nemají 1–12).
                    for rr in range(trow + 1, min(maxr, trow + 22) + 1):
                        label = ws.cell(rr, cm).value
                        mnum = _oz_m_of(label)
                        if mnum is not None:
                            v1 = _oz_num(ws.cell(rr, c1).value)
                            v2 = _oz_num(ws.cell(rr, c2).value)
                            if v1 is not None:
                                radky.append((list_klic, skup, sekce_key, y1, mnum, v1))
                            if v2 is not None:
                                radky.append((list_klic, skup, sekce_key, y2, mnum, v2))
                        elif label is not None and 'součet' in str(label).lower():
                            cv1 = _oz_num(ws.cell(rr, c1).value)
                            cv2 = _oz_num(ws.cell(rr, c2).value)
                            if cv1 is not None:
                                radky.append((list_klic, skup, sekce_key, y1, 0, cv1))
                            if cv2 is not None:
                                radky.append((list_klic, skup, sekce_key, y2, 0, cv2))
                            break  # součtový řádek = konec bloku

            # ── Záložní sklad (zatím jen Praha) – samostatné pivoty s titulkem
            # „Meziroční porovnání prodejů na Záložní sklad". Blok nemá pevný
            # sloupec, hledá se pozice titulku; pod ním je blok filtrů, řádek
            # sekce („Obrat v Kg" / „Obrat v Kč bez DPH") a řádek „Měsíc – číslo"
            # s roky ve dvou sloupcích vpravo.
            maxc = min(ws.max_column, 30)
            for r in range(1, maxr + 1):
                for c in range(1, maxc + 1):
                    cv = ws.cell(r, c).value
                    # Jen titulek bloku („Meziroční porovnání … Záložní sklad") –
                    # popisky typu „Bez převodek na záložní sklad" se přeskočí.
                    s = str(cv).lower() if cv is not None else ''
                    if 'záložní sklad' not in s or 'porovnání' not in s:
                        continue
                    for rr in range(r + 1, min(maxr, r + 15) + 1):
                        mv = ws.cell(rr, c).value
                        if mv is None or not str(mv).strip().lower().startswith('měsíc'):
                            continue
                        y1 = _oz_year(ws.cell(rr, c + 1).value)
                        y2 = _oz_year(ws.cell(rr, c + 2).value)
                        if not y1 or not y2:
                            continue
                        sekce_lbl = str(ws.cell(rr - 1, c).value or '').lower()
                        sekce_key = 'kg' if 'kg' in sekce_lbl else 'kc'
                        for r2 in range(rr + 1, min(maxr, rr + 22) + 1):
                            label = ws.cell(r2, c).value
                            mnum = _oz_m_of(label)
                            if mnum is not None:
                                v1 = _oz_num(ws.cell(r2, c + 1).value)
                                v2 = _oz_num(ws.cell(r2, c + 2).value)
                                if v1 is not None:
                                    radky.append((list_klic, 'zalozni_sklad', sekce_key, y1, mnum, v1))
                                if v2 is not None:
                                    radky.append((list_klic, 'zalozni_sklad', sekce_key, y2, mnum, v2))
                            elif label is not None and 'součet' in str(label).lower():
                                cv1 = _oz_num(ws.cell(r2, c + 1).value)
                                cv2 = _oz_num(ws.cell(r2, c + 2).value)
                                if cv1 is not None:
                                    radky.append((list_klic, 'zalozni_sklad', sekce_key, y1, 0, cv1))
                                if cv2 is not None:
                                    radky.append((list_klic, 'zalozni_sklad', sekce_key, y2, 0, cv2))
                                break
                        break  # jeden datový blok na jeden titulek
    except Exception as exc:
        return [], f'Chyba čtení dat: {exc}'
    return radky, ''


def _oz_zapis_pobocku(conn, pobocka: str, radky: list[tuple]):
    """Nahradí data pobočky v vysledky_oz_data. Maže jen listy, které radky
    skutečně obsahují – pobočka s neúplným párem listů nepřijde o data druhého
    listu. Data z Tabulky nákladů (Staropramen/Operace) se nemažou, ta mají
    vlastní import. Jedna transakce na pobočku; při chybě rollback a výjimka nahoru."""
    listy = sorted({r[0] for r in radky})
    cur = conn.cursor()
    try:
        ph = ','.join(['%s'] * len(listy))
        cur.execute(f'DELETE FROM vysledky_oz_data WHERE pobocka=%s AND list IN ({ph}) '
                    "AND sekce<>'ops' AND skupina<>'staropramen'",
                    (pobocka, *listy))
        cur.executemany(
            'INSERT INTO vysledky_oz_data (pobocka,list,skupina,sekce,rok,mesic,hodnota) '
            'VALUES (%s,%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE hodnota=VALUES(hodnota)',
            [(pobocka,) + r for r in radky],
        )
        conn.commit()
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        raise
    finally:
        try:
            cur.close()
        except Exception:
            pass


def _oz_centralni_parsuj_sync(zdroj, progress=None) -> tuple[list[dict], list[str], str]:
    """Naparsuje centrální soubor („Pobočkové GISTy.xlsx") – listy
    „010 Obraty_prodeje" / „010 Zisk_Marže" seskupí podle čísla střediska
    v názvu listu a spáruje na pobočky přes _STREDISKA_SHEET_MAPPING.
    progress = volitelný callback(hotovo 0–1, popisek) volaný z vlákna.
    Vrací (matched, preskocene, chyba):
      matched    = [{'cislo', 'pobocka', 'listy', 'radky', 'pozn'}, …]
      preskocene = popisy listů, které se nepodařilo spárovat (nezpůsobí pád)."""
    def _p(v, txt):
        if progress:
            try:
                progress(v, txt)
            except Exception:
                pass
    _p(0.02, 'Otevírám soubor…')
    try:
        import openpyxl
        wb = openpyxl.load_workbook(zdroj, data_only=True)
    except Exception as exc:
        return [], [], f'Nelze otevřít soubor: {exc}'
    _p(0.4, 'Hledám listy poboček…')
    try:
        skupiny: dict[str, dict] = {}
        preskocene: list[str] = []
        for ws in wb.worksheets:
            nazev = (ws.title or '').strip()
            m = re.match(r'^(\d{3})(?!\d)', nazev)
            if not m:
                preskocene.append(f'{nazev} – název listu nezačíná číslem střediska')
                continue
            cislo = m.group(1)
            pobocka = _STREDISKA_SHEET_MAPPING.get(cislo)
            if pobocka is None:
                preskocene.append(f'{nazev} – neznámé číslo střediska {cislo}')
                continue
            if pobocka not in POBOCKY:
                preskocene.append(f'{nazev} – středisko {cislo} ({pobocka}) není pobočka Výsledků')
                continue
            zbytek = nazev[m.end():].lower()
            if 'obrat' in zbytek:
                lk = 'obraty'
            elif 'zisk' in zbytek or 'marže' in zbytek or 'marze' in zbytek:
                lk = 'zisk'
            else:
                preskocene.append(f'{nazev} – list není Obraty_prodeje ani Zisk_Marže')
                continue
            sk = skupiny.setdefault(cislo, {'pobocka': pobocka, 'sheets': {}, 'nazvy': []})
            if lk in sk['sheets']:
                preskocene.append(f'{nazev} – duplicitní list pro středisko {cislo}')
                continue
            sk['sheets'][lk] = ws
            sk['nazvy'].append(nazev)

        matched: list[dict] = []
        n = max(len(skupiny), 1)
        for i, cislo in enumerate(sorted(skupiny)):
            sk = skupiny[cislo]
            zakl = 0.4 + 0.55 * i / n
            _p(zakl, f'Čtu data – {sk["pobocka"]}…')
            radky, err = _oz_parsuj_sheets(
                sk['sheets'],
                progress=lambda podil, nazev, z=zakl: _p(
                    z + podil * 0.55 / n, f'Načítám list {nazev}…'))
            if err:
                preskocene.append(f'{cislo} {sk["pobocka"]} – {err}')
                continue
            if not radky:
                preskocene.append(f'{cislo} {sk["pobocka"]} – v listech nenalezena žádná data')
                continue
            pozn = ['chybí list ' + ('Zisk_Marže' if k == 'zisk' else 'Obraty_prodeje')
                    for k in ('obraty', 'zisk') if k not in sk['sheets']]
            matched.append({'cislo': cislo, 'pobocka': sk['pobocka'],
                            'listy': sorted(sk['nazvy']), 'radky': radky, 'pozn': pozn})
    finally:
        try:
            wb.close()
        except Exception:
            pass
    if not matched and not preskocene:
        return [], [], 'Soubor neobsahuje žádné listy.'
    _p(1.0, 'Hotovo')
    return matched, preskocene, ''


def _oz_centralni_importuj_sync(matched: list[dict], zapis=None,
                                progress=None) -> tuple[int, int, list[str]]:
    """Uloží naparsovaná data centrálního souboru – každá pobočka ve vlastní
    transakci, aby chyba jedné nezrušila ostatní. Pobočky mimo soubor se
    nemažou. zapis = funkce(conn, pobocka, radky) pro zápis (výchozí nahrazuje
    celé listy). progress = volitelný callback(hotovo 0–1, popisek) z vlákna.
    Vrací (počet hodnot, počet poboček, chyby po pobočkách)."""
    zapis = zapis or _oz_zapis_pobocku
    def _p(v, txt):
        if progress:
            try:
                progress(v, txt)
            except Exception:
                pass
    conn = intranet_data.get_db_connection()
    if not conn:
        return 0, 0, ['Nelze se připojit k databázi.']
    celkem = 0
    npob = 0
    chyby: list[str] = []
    try:
        for i, e in enumerate(matched):
            _p(i / max(len(matched), 1), f'Ukládám – {e["pobocka"]}…')
            try:
                zapis(conn, e['pobocka'], e['radky'])
                celkem += len(e['radky'])
                npob += 1
            except Exception as exc:
                chyby.append(f'{e["pobocka"]}: {exc}')
    finally:
        conn.close()
    _p(1.0, 'Hotovo')
    return celkem, npob, chyby


_OZN_ALIAS = {
    'brno - wine life': 'Winelife',
    'wine life': 'Winelife',
    'přefakturace jip': 'Přefakturace',
}


def _ozn_pobocka(popis) -> str | None:
    """„010 - Pardubice" → 'Pardubice'. Vrací None u neznámé pobočky."""
    s = re.sub(r'^\s*\d{3}\s*-\s*', '', str(popis or '')).strip()
    k = s.lower()
    if k in _OZN_ALIAS:
        return _OZN_ALIAS[k]
    for p in POBOCKY:
        if p.lower() == k:
            return p
    return None


def _ozn_parsuj_data_sync(zdroj, progress=None) -> tuple[list[dict], list[str], str]:
    """Naparsuje plochý centrální soubor („Tabulka nákladů_DATA.xlsx", list
    DATA): jeden řádek = pobočka × rok × měsíc. Dělí se podle sloupce „Pobočka
    popis" (víc středisek pod stejným popisem se sčítá), bere jen poslední dva
    roky. Vrací (matched, přeskočené, chyba)."""
    def _p(v, txt):
        if progress:
            try:
                progress(v, txt)
            except Exception:
                pass

    _p(0.02, 'Otevírám soubor…')
    try:
        import openpyxl
        wb = openpyxl.load_workbook(zdroj, data_only=True, read_only=True)
    except Exception as exc:
        return [], [], f'Nelze otevřít soubor: {exc}'
    try:
        ws = wb['DATA'] if 'DATA' in wb.sheetnames else wb.worksheets[0]
        hlavicka = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not hlavicka:
            return [], [], 'List je prázdný.'
        idx = {str(v).strip().lower(): i for i, v in enumerate(hlavicka) if v is not None}

        def _col(nazev):
            return idx.get(nazev.lower())

        i_pob, i_rok, i_mes = _col('Pobočka popis'), _col('Rok'), _col('Měsíc')
        chybi = [n for n, i in (('Pobočka popis', i_pob), ('Rok', i_rok), ('Měsíc', i_mes))
                 if i is None]
        metriky = []
        for skup, sekce, sloupec, _titul in _OZN_KARTY:
            i_val = _col(sloupec)
            if i_val is None:
                chybi.append(sloupec)
            else:
                metriky.append((skup, sekce, i_val))
        if chybi:
            return [], [], 'V souboru chybí sloupce: ' + ', '.join(chybi)

        rok_min = datetime.datetime.now().year - 1
        _p(0.15, 'Načítám řádky…')
        soucty: dict = {}
        nezname: dict[str, int] = {}
        pobocky_roky: dict[str, set] = {}
        for r in ws.iter_rows(min_row=2, values_only=True):
            rok = _oz_year(r[i_rok])
            if not rok or rok < rok_min:
                continue
            pobocka = _ozn_pobocka(r[i_pob])
            if not pobocka:
                popis = str(r[i_pob] or '').strip()
                if popis:
                    nezname[popis] = nezname.get(popis, 0) + 1
                continue
            mnum = _oz_m_of(r[i_mes])
            if not mnum:
                continue
            pobocky_roky.setdefault(pobocka, set()).add(rok)
            for skup, sekce, i_val in metriky:
                v = _oz_num(r[i_val])
                if not v:
                    continue
                for m in (mnum, 0):   # 0 = Celkový součet
                    k = (pobocka, skup, sekce, rok, m)
                    soucty[k] = soucty.get(k, 0.0) + v
    except Exception as exc:
        return [], [], f'Chyba čtení dat: {exc}'
    finally:
        try:
            wb.close()
        except Exception:
            pass

    _p(0.85, 'Skládám přehled…')
    radky_pob: dict[str, list] = {}
    for (pobocka, skup, sekce, rok, m), v in soucty.items():
        radky_pob.setdefault(pobocka, []).append(('obraty', skup, sekce, rok, m, v))
    matched = [{'pobocka': p, 'radky': sorted(radky_pob[p]),
                'roky': sorted(pobocky_roky.get(p, set()))}
               for p in sorted(radky_pob)]
    prazdne = [p for p in sorted(pobocky_roky) if p not in radky_pob]
    preskocene = [f'{p} – v souboru samé nuly' for p in prazdne]
    preskocene += [f'{popis} – neznámá pobočka ({n} řádků)'
                   for popis, n in sorted(nezname.items())]
    _p(1.0, 'Hotovo')
    return matched, preskocene, ''


def _ozn_zapis_pobocku(conn, pobocka: str, radky: list[tuple]):
    """Nahradí u pobočky jen data Staropramen/Operace – ručně zadané Kč/Kg
    a Záložní sklad zůstávají nedotčené. Jedna transakce na pobočku."""
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM vysledky_oz_data WHERE pobocka=%s AND list='obraty' "
                    "AND (sekce='ops' OR skupina='staropramen')", (pobocka,))
        if radky:
            cur.executemany(
                'INSERT INTO vysledky_oz_data (pobocka,list,skupina,sekce,rok,mesic,hodnota) '
                'VALUES (%s,%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE hodnota=VALUES(hodnota)',
                [(pobocka,) + r for r in radky])
        conn.commit()
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        raise
    finally:
        try:
            cur.close()
        except Exception:
            pass


def _oz_zjisti_roky(*datasets) -> tuple[int, int]:
    """Dva nejnovější roky ke srovnání (starší, novější). Vždy zahrne aktuální
    a minulý rok – aby šlo zadávat/porovnávat data za aktuální rok i u poboček,
    které mají v databázi jen starší roky (jinak by chyběl sloupec pro letošek)."""
    roky = set()
    for data in datasets:
        for skup in data.values():
            for sek in skup.values():
                roky.update(sek.keys())
    y = datetime.datetime.now().year
    roky.update({y - 1, y})
    roky = sorted(roky)
    return roky[-2], roky[-1]


def _oz_cols2(rok_old: int, rok_new: int, editable: bool, je_marze: bool) -> list[dict]:
    fmt = _SOUHRN_FMT_PCT if je_marze else _OZ_FMT_NUM0
    edit_js = "function(p){return %s;}" % ('true' if (editable and not je_marze) else 'false')
    cols = [
        {'headerName': 'Měsíc', 'field': 'mesic', 'width': 90, 'minWidth': 84, 'maxWidth': 110,
         'cellDataType': 'text', ':cellStyle': _OZ_CELKEM_CS},
        {'headerName': str(rok_old), 'field': 'y_old', 'width': 128, 'minWidth': 96,
         'type': 'numericColumn', 'cellDataType': False, ':editable': edit_js,
         ':valueParser': _OZ_VALUE_PARSER,
         ':valueFormatter': fmt, ':cellStyle': _OZ_CELKEM_CS},
        {'headerName': str(rok_new), 'field': 'y_new', 'width': 128, 'minWidth': 96,
         'type': 'numericColumn', 'cellDataType': False, ':editable': edit_js,
         ':valueParser': _OZ_VALUE_PARSER,
         ':valueFormatter': fmt, ':cellStyle': _OZ_CELKEM_CS},
    ]
    if not je_marze:
        cols.append({
            'headerName': 'Porovnání', 'field': 'porovnani', 'width': 96, 'minWidth': 80, 'maxWidth': 150,
            'type': 'numericColumn', 'headerClass': 'oz-h-dark',
            ':valueGetter': _OZ_POROVNANI_GETTER, ':valueFormatter': _SOUHRN_FMT_PCT,
            ':cellStyle': _OZ_POROVNANI_CS,
        })
    return cols


def _oz_rows2(vals_old: dict, vals_new: dict) -> list[dict]:
    rows = []
    for m in range(1, 13):
        rows.append({'mesic': str(m), 'y_old': vals_old.get(m), 'y_new': vals_new.get(m), '_celkem': False})
    rows.append({'mesic': 'Celkem', 'y_old': vals_old.get(0), 'y_new': vals_new.get(0), '_celkem': True})
    return rows


def _oz_karta(pobocka, list_klic, skup, sekce_key, rok_old, rok_new,
              vals_old, vals_new, titul, popis, editable, je_marze, popis_h=96):
    with ui.column().classes('gap-1 w-full'):
        # Popisový blok na pevné výšce → tabulky ve 3 sloupcích se zarovnají.
        with ui.column().classes('gap-0 w-full').style(f'height: {popis_h}px; overflow: hidden'):
            ui.label(titul).classes('text-sm font-bold leading-tight').style('color:#748C43')
            for line in popis:
                ui.label(line).classes('text-xs text-gray-500 leading-tight')
        grid = ui.aggrid({
            'columnDefs': _oz_cols2(rok_old, rok_new, editable, je_marze),
            'rowData': _oz_rows2(vals_old, vals_new),
            'defaultColDef': {'resizable': True, 'sortable': False, 'wrapHeaderText': True, 'autoHeaderHeight': True},
            'rowHeight': 30, 'suppressMovableColumns': True,
            'stopEditingWhenCellsLoseFocus': True,
            'domLayout': 'autoHeight',
            ':onFirstDataRendered': _AUTOSIZE_FIT,
            ':onGridSizeChanged': _AUTOSIZE_FIT,
        }).classes('w-full oz-grid')

        if editable and not je_marze:
            def _on_change(e):
                a = e.args or {}
                col = a.get('colId')
                if col not in ('y_old', 'y_new'):
                    return
                rok = rok_old if col == 'y_old' else rok_new
                ml = (a.get('data') or {}).get('mesic')
                mesic = 0 if ml == 'Celkem' else int(ml)
                nv = a.get('newValue')
                try:
                    val = (float(str(nv).replace('\xa0', '').replace(' ', '').replace(',', '.'))
                           if nv not in (None, '') else 0.0)
                except (TypeError, ValueError):
                    val = 0.0
                uloz_oz_hodnota(pobocka, list_klic, skup, sekce_key, rok, mesic, val)
            grid.on('cellValueChanged', _on_change)


def _oz_render_novy(pobocka: str, je_ao: bool, user_name: str, list_klic: str):
    ui.add_css(_OZ_CSS)
    txt = _OZ_TEXTY[list_klic]
    data = nacti_oz_data(pobocka, list_klic)
    obraty = nacti_oz_data(pobocka, 'obraty') if list_klic == 'zisk' else data
    rok_old, rok_new = _oz_zjisti_roky(data, obraty)

    ui.label(f'{txt["titulek"]} {pobocka} – {rok_new} × {rok_old}') \
        .classes('text-lg font-bold text-gray-800 mb-1 oz-titul-banner') \
        .style('background-color:#FFFFCC;padding:4px 10px;border-radius:4px')
    if je_ao:
        ui.label('Editovatelné jsou hodnoty 2025/2026 i Celkový součet; „Porovnání" a „Marže v %" se počítají.') \
            .classes('text-xs text-gray-400 italic mb-3')

    # ── Editovatelné sekce (Obraty: Kč+Kg; Zisk: Kč) ──────────────────────────
    for sekce_key, sekce_nazev in txt['sekce']:
        ui.label(sekce_nazev).classes('text-base font-bold text-gray-700 mt-3 mb-1')
        with ui.grid(columns=3).classes('w-full gap-4'):
            for skup in _OZ_SKUP_PORADI:
                titul, popis = txt['skupiny'][skup]
                sek = data.get(skup, {}).get(sekce_key, {})
                _oz_karta(pobocka, list_klic, skup, sekce_key, rok_old, rok_new,
                          sek.get(rok_old, {}), sek.get(rok_new, {}),
                          titul, popis, je_ao, je_marze=False)

    # ── Staropramen + Operace (import „Tabulka nákladů_DATA.xlsx") ────────────
    if list_klic == 'obraty':
        def _ozn_grid(karty):
            with ui.grid(columns=3).classes('w-full gap-4'):
                for skup, sekce_key, _sloupec, titul in karty:
                    sek = data.get(skup, {}).get(sekce_key, {})
                    # Operace: stejný popis jako u obratů (jen titulek mluví o operacích)
                    popis = txt['skupiny'][skup][1] if sekce_key == 'ops' else []
                    _oz_karta(pobocka, list_klic, skup, sekce_key, rok_old, rok_new,
                              sek.get(rok_old, {}), sek.get(rok_new, {}),
                              titul, popis, je_ao, je_marze=False,
                              popis_h=96 if popis else 28)

        _ozn_grid([k for k in _OZN_KARTY if k[1] != 'ops'])
        ui.label('Operace').classes('text-base font-bold text-gray-700 mt-4 mb-1')
        _ozn_grid([k for k in _OZN_KARTY if k[1] == 'ops'])

    # ── Záložní sklad – jen pobočky, které mají data (zatím Praha) ────────────
    zs = data.get('zalozni_sklad', {})
    if zs:
        ui.label('Meziroční porovnání prodejů na Záložní sklad') \
            .classes('text-base font-bold text-gray-700 mt-4 mb-1')
        with ui.grid(columns=3).classes('w-full gap-4'):
            for sekce_key, sekce_nazev in txt['sekce']:
                sek = zs.get(sekce_key, {})
                if not sek:
                    continue
                _oz_karta(pobocka, list_klic, 'zalozni_sklad', sekce_key, rok_old, rok_new,
                          sek.get(rok_old, {}), sek.get(rok_new, {}),
                          f'Záložní sklad – {sekce_nazev}',
                          ['Prodeje na záložní sklad (dceřiná společnost)'],
                          je_ao, je_marze=False, popis_h=48)

    # ── Marže v % (jen Zisk) – dopočítané ze Zisk ÷ Obraty ────────────────────
    if list_klic == 'zisk':
        marze_nazev, marze_pod = txt['marze']
        ui.label(marze_nazev).classes('text-base font-bold text-gray-700 mt-4 mb-1')

        def _marze_vals(skup, rok):
            zk = data.get(skup, {}).get('kc', {}).get(rok, {})
            ob = obraty.get(skup, {}).get('kc', {}).get(rok, {})
            out = {}
            for m in list(range(1, 13)) + [0]:
                o = ob.get(m)
                out[m] = (zk.get(m) / o) if (o and zk.get(m) is not None) else None
            return out

        with ui.grid(columns=3).classes('w-full gap-4'):
            for skup in _OZ_SKUP_PORADI:
                _oz_karta(pobocka, list_klic, skup, 'marze', rok_old, rok_new,
                          _marze_vals(skup, rok_old), _marze_vals(skup, rok_new),
                          f'Marže v % – {marze_pod[skup]}', [], False, je_marze=True, popis_h=28)


@refreshable_na_klienta
def _vykresli_obraty(pobocka: str, je_ao: bool, user_name: str):
    _oz_render_novy(pobocka, je_ao, user_name, 'obraty')


@refreshable_na_klienta
def _vykresli_zisk(pobocka: str, je_ao: bool, user_name: str):
    _oz_render_novy(pobocka, je_ao, user_name, 'zisk')


# ═══════════════════════════════════════════════════════════════════════════════
#  EXPORT DO XLSX – kompletní data pobočky + přehledový soubor ZR
# ═══════════════════════════════════════════════════════════════════════════════
# Formáty čísel bereme z `intranet_sankce._XLSX_FMT`, ať se exporty napříč
# aplikací chovají stejně. Vlastní zapisovač je tu proto, že `_xlsx_bytes`
# umí jen jeden list s jednou tabulkou – tady je potřeba sešit o mnoha listech
# a u obratů/zisků bloková osnova vedle sebe (měsíce / dcery / sítě).
#
# Listy „Obraty_prodeje" a „Zisk_Marže" se zapisují v rozložení, které umí
# přečíst `_oz_parsuj_sheets` (titulek sekce ve sloupci C/I/N, roky pod ním ve
# sloupcích D/E, měsíce a „Celkový součet" pod tím) → export jde znovu
# naimportovat. Bloky proto nesmí mít jinou vnitřní osnovu.

_EXP_OZ_ROZTEC = 18          # svislá rozteč bloků obratů/zisků (blok = 16 řádků)
# Barvy/písma dle vzorového sešitu „…_porovnání_obratů_….xlsx" od uživatelů:
# modrá hlavička, žlutý titulek, data Arial 10, součty tučně červeně.
_EXP_HLAVICKA_BG = '3399FF'  # modrý pruh hlavičky
_EXP_TITUL_BG = 'FFCC00'     # žlutý titulek listu
_EXP_TITUL_BG_SVETLY = 'FFFFCC'  # světle žlutý titulek na listech obratů/zisků
_EXP_SOUCET_BARVA = 'FF0000'  # červené součtové řádky

# Listy obratů/zisků kopírují barvy kontingenčních tabulek ze vzorového sešitu:
# hlavičky i „Celkový součet" tmavě šedé s bílým písmem, data v barvě skupiny.
_EXP_OZ_HLAVICKA_BG = '404040'
_EXP_OZ_BARVY = {                       # (titulek skupiny, výplň dat, světlá výplň marží)
    'mesice': ('76933C', 'EBF1DE', 'C4D79B'),
    'dcery': ('31859B', 'DAEEF3', 'B7DEE8'),
    'site': ('E26B0A', 'FDE9D9', 'FCD5B4'),
}
_EXP_OZ_BANNER = {'obraty': 'FFFFCC', 'zisk': 'DCE6F1'}


def _exp_slug(s: str) -> str:
    """Název souboru bez diakritiky a bez znaků, které rozbíjí Content-Disposition."""
    return re.sub(r'[^A-Za-z0-9]+', '_', _bez_diakritiky(s)).strip('_') or 'export'


def _exp_wb():
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)          # výchozí prázdný list nechceme
    return wb


def _exp_list(wb, nazev: str, pouzite: set):
    """Nový list s názvem ošetřeným dle pravidel Excelu (max 31 znaků, bez
    []:*?/\\, unikátní v sešitu)."""
    zaklad = re.sub(r'[\[\]:*?/\\]', '-', str(nazev)).strip()[:31] or 'List'
    jmeno, i = zaklad, 2
    while jmeno.lower() in pouzite:
        pripona = f'_{i}'
        jmeno = zaklad[:31 - len(pripona)] + pripona
        i += 1
    pouzite.add(jmeno.lower())
    return wb.create_sheet(title=jmeno)


def _exp_tabulka(ws, cols: list[tuple], rows: list[dict], soucet: dict | None = None,
                 r0: int = 1):
    """Zapíše tabulku: cols = [(nadpis, pole, typ, šířka)], typ dle `_XLSX_FMT`.
    Součtový řádek (a řádky s `_je_celkem`/`_soucet`) jsou tučně."""
    from intranet_sankce import _XLSX_FMT
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    tucne = Font(name='Calibri', size=11, bold=True, color=_EXP_SOUCET_BARVA)
    bezne = Font(name='Arial', size=10)
    hlavicka = Font(name='Calibri', size=11, bold=True)
    for i, (nadpis, _f, _t, sirka) in enumerate(cols, start=1):
        c = ws.cell(r0, i, nadpis)
        c.font = hlavicka
        c.fill = PatternFill('solid', fgColor=_EXP_HLAVICKA_BG)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = sirka
    ws.row_dimensions[r0].height = 30
    r = r0
    for row in list(rows) + ([soucet] if soucet else []):
        r += 1
        je_soucet = bool(row.get('_je_celkem') or row.get('_soucet')) or row is soucet
        for i, (_n, pole, typ, _w) in enumerate(cols, start=1):
            v = row.get(pole)
            if typ == 'text':
                v = '' if v is None else str(v)
            elif v in (None, ''):
                v = None
            else:
                try:
                    v = float(v)
                except (TypeError, ValueError):
                    v = None
            c = ws.cell(r, i, v)
            c.number_format = _XLSX_FMT[typ]
            c.font = tucne if je_soucet else bezne
    ws.freeze_panes = ws.cell(r0 + 1, 1).coordinate
    if r > r0:
        ws.auto_filter.ref = f'A{r0}:{get_column_letter(len(cols))}{r}'
    return r


# ── Sloupce jednotlivých tabulek ─────────────────────────────────────────────

def _exp_cols_souhrn(prefix: list[tuple] | None = None,
                     pocitane: bool = True) -> list[tuple]:
    """Tabulka nákladů = stejné sloupce (i pořadí a počítané) jako grid.
    `pocitane=False` pro surové DB řádky, kde „Podíl mzdy v %"/„%" nejsou."""
    cols = list(prefix) if prefix else []
    for db, popis, _idx, sirka in _SOUHRN_COLS:
        cols.append((popis, db, 'money', max(12, round(sirka / 7))))
        if not pocitane:
            continue
        if db == 'mzdy':
            cols.append(('Podíl mzdy v %', '_podil_mzdy', 'pct', 15))
        elif db == 'total':
            cols.append(('%', '_procento', 'pct', 10))
    return cols


def _exp_cols_naklady() -> list[tuple]:
    return ([('Účetní předpis', 'ucetni_predpis', 'text', 14),
             ('Název předpisu', 'nazev_predpisu', 'text', 42)]
            + [(nazev, db, 'money', 15) for db, nazev in zip(MESICE_DB, MESICE_NAZVY)]
            + [('Celkem', 'celkem', 'money', 17)])


def _exp_cols_porovnani(rok: int, minuly: int) -> list[tuple]:
    cols = [('Účetní předpis', 'ucetni_predpis', 'text', 14),
            ('Název předpisu', 'nazev_predpisu', 'text', 42)]
    for db, nazev in zip(MESICE_DB, MESICE_NAZVY):
        cols += [(f'{nazev} {rok}', f'akt_{db}', 'money', 15),
                 (f'{nazev} {minuly}', f'min_{db}', 'money', 15),
                 (f'{nazev} rozdíl', f'diff_{db}', 'money', 15)]
    cols += [(f'Celkem {rok}', 'akt_celkem', 'money', 17),
             (f'Celkem {minuly}', 'min_celkem', 'money', 17),
             ('Celkem rozdíl', 'diff_celkem', 'money', 17)]
    return cols


# ── Obraty / Zisk (bloková osnova kompatibilní s importem) ───────────────────

def _exp_oz_blok(ws, r0: int, sloupce: tuple, titul_skupiny: str, titul_sekce: str,
                 rok_old: int, rok_new: int, v_old: dict, v_new: dict, pct: bool = False,
                 skup: str = 'mesice'):
    """Jeden blok (karta v UI): titulek, „Rok", měsíce 1–12, Celkový součet
    a dopočítané „Porovnání" (nový rok ÷ starý rok).

    Barvy 1:1 dle vzorových kontingenčních tabulek: nadpis v barvě skupiny,
    dvouřádková hlavička a součet tmavě šedé s bílým písmem, data ve světlé
    barvě skupiny. Bloky marží (`pct`) mají místo šedé střední odstín skupiny
    a data bez výplně."""
    from openpyxl.styles import Font, PatternFill
    cm, c1, c2 = sloupce
    c3 = c2 + 1
    fmt = '0.00%' if pct else '#,##0.00'
    barva_titulku, vypln_dat, vypln_svetla = _EXP_OZ_BARVY[skup]
    nadpis = Font(name='Arial', size=11, bold=True, color=barva_titulku)
    hlavicka_f = Font(name='Arial', size=10, bold=True,
                      color='000000' if pct else 'FFFFFF')
    bezne = Font(name='Arial', size=10)
    hlavicka = PatternFill('solid', fgColor=vypln_svetla if pct else _EXP_OZ_HLAVICKA_BG)
    data_fill = None if pct else PatternFill('solid', fgColor=vypln_dat)
    for c in (cm, c1, c2, c3):
        ws.cell(r0 + 1, c).fill = hlavicka
        ws.cell(r0 + 2, c).fill = hlavicka
    ws.cell(r0, cm, titul_skupiny).font = nadpis
    ws.cell(r0 + 1, cm, titul_sekce).font = hlavicka_f
    ws.cell(r0 + 1, c1, 'Rok').font = hlavicka_f
    for c, v in ((cm, 'Měsíc - číslo'), (c1, rok_old), (c2, rok_new), (c3, 'Porovnání')):
        ws.cell(r0 + 2, c, v).font = hlavicka_f
    for i, mesic in enumerate(list(range(1, 13)) + [0]):
        r = r0 + 3 + i
        je_celkem = mesic == 0
        pismo = hlavicka_f if je_celkem else bezne
        vypln = hlavicka if je_celkem else data_fill
        ws.cell(r, cm, 'Celkový součet' if je_celkem else mesic).font = pismo
        for c, vals in ((c1, v_old), (c2, v_new)):
            v = vals.get(mesic)
            cell = ws.cell(r, c, None if v is None else float(v))
            cell.number_format = fmt
            cell.font = pismo
        a, b = v_old.get(mesic), v_new.get(mesic)
        pom = ws.cell(r, c3, (b / a) if (a and b is not None) else None)
        pom.number_format = '0.00%'
        pom.font = pismo
        if vypln is not None:
            for c in (cm, c1, c2, c3):
                ws.cell(r, c).fill = vypln


def _exp_list_oz(wb, pouzite: set, pobocka: str, list_klic: str):
    """List „Obraty_prodeje" / „Zisk_Marže" – stejné karty jako v UI."""
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    txt = _OZ_TEXTY[list_klic]
    data = nacti_oz_data(pobocka, list_klic)
    obraty = nacti_oz_data(pobocka, 'obraty') if list_klic == 'zisk' else data
    rok_old, rok_new = _oz_zjisti_roky(data, obraty)
    ws = _exp_list(wb, 'Obraty_prodeje' if list_klic == 'obraty' else 'Zisk_Marže', pouzite)
    nazev = _POBOCKY_EXCEL_REVERSE.get(pobocka, pobocka)
    titul = ws.cell(1, 1, f'{txt["titulek"]} {nazev} – {rok_new} × {rok_old}')
    titul.font = Font(name='Arial', size=16, bold=True)
    titul.fill = PatternFill('solid', fgColor=_EXP_OZ_BANNER[list_klic])
    for c in range(3, 18):
        ws.column_dimensions[get_column_letter(c)].width = 17

    r = 3
    for sekce_key, sekce_nazev in txt['sekce']:
        for skup in _OZ_SKUP_PORADI:
            sek = data.get(skup, {}).get(sekce_key, {})
            _exp_oz_blok(ws, r, _OZ_COLMAP[skup], txt['skupiny'][skup][0], sekce_nazev,
                         rok_old, rok_new, sek.get(rok_old, {}), sek.get(rok_new, {}),
                         skup=skup)
        r += _EXP_OZ_ROZTEC

    if list_klic == 'obraty':
        # Staropramen + Operace (data z „Tabulka nákladů_DATA.xlsx")
        for skup, sekce_key, sloupec, titul in _OZN_KARTY:
            sek = data.get(skup, {}).get(sekce_key, {})
            _exp_oz_blok(ws, r, _OZ_COLMAP[skup if sekce_key == 'ops' else 'mesice'],
                         titul, sloupec, rok_old, rok_new,
                         sek.get(rok_old, {}), sek.get(rok_new, {}),
                         skup=skup if sekce_key == 'ops' else 'mesice')
            if sekce_key != 'ops':
                r += _EXP_OZ_ROZTEC
        r += _EXP_OZ_ROZTEC

    zs = data.get('zalozni_sklad', {})
    for sekce_key, sekce_nazev in (txt['sekce'] if zs else []):
        sek = zs.get(sekce_key, {})
        if not sek:
            continue
        _exp_oz_blok(ws, r, _OZ_COLMAP['mesice'],
                     'Meziroční porovnání prodejů na Záložní sklad', sekce_nazev,
                     rok_old, rok_new, sek.get(rok_old, {}), sek.get(rok_new, {}))
        r += _EXP_OZ_ROZTEC

    if list_klic == 'zisk':
        # Marže se v UI nedrží v DB, dopočítává se Zisk ÷ Obrat – stejně i tady.
        marze_nazev, marze_pod = txt['marze']

        def _marze(skup, rok):
            zk = data.get(skup, {}).get('kc', {}).get(rok, {})
            ob = obraty.get(skup, {}).get('kc', {}).get(rok, {})
            return {m: ((zk.get(m) / ob[m]) if (ob.get(m) and zk.get(m) is not None) else None)
                    for m in list(range(1, 13)) + [0]}

        for skup in _OZ_SKUP_PORADI:
            _exp_oz_blok(ws, r, _OZ_COLMAP[skup], marze_nazev,
                         f'Marže v % – {marze_pod[skup]}', rok_old, rok_new,
                         _marze(skup, rok_old), _marze(skup, rok_new), pct=True, skup=skup)
        r += _EXP_OZ_ROZTEC
    return ws


# ── Sešit pobočky ────────────────────────────────────────────────────────────

def _export_pobocka_xlsx(pobocka: str, sekce: list[str]) -> bytes:
    """Kompletní data pobočky do XLSX. `sekce` = záložky, které uživatel v detailu
    reálně vidí (`povolene`) – export nikdy neobsahuje víc, než co je na obrazovce."""
    import io
    rok = datetime.datetime.now().year
    minuly = rok - 1
    nazev = _POBOCKY_EXCEL_REVERSE.get(pobocka, pobocka)
    wb = _exp_wb()
    pouzite: set = set()

    if 'obraty' in sekce:
        _exp_list_oz(wb, pouzite, pobocka, 'obraty')
    if 'zisk' in sekce:
        _exp_list_oz(wb, pouzite, pobocka, 'zisk')
    if 'naklady' in sekce:
        ws = _exp_list(wb, 'Tabulka nákladů', pouzite)
        _exp_tabulka(ws, _exp_cols_souhrn([('Rok', '_rok_disp', 'text', 14),
                                           ('Měsíc', '_mesic_disp', 'text', 9)]),
                     _priprav_souhrn_pro_grid(nacti_souhrn(pobocka)))
    if 'podrobne' in sekce:
        # Pořadí listů jako ve vzorovém souboru: aktuální rok, porovnání, minulý rok.
        # V názvu listu je krátký (interní) název pobočky – „010 - Pardubice 2026"
        # by přeteklo přes 31 znaků, které Excel na název listu dovolí.
        def _podrobne(r: int):
            for _popis, klic in [(nazev, pobocka)] + _VNORENE_POBOCKY.get(pobocka, []):
                rows = nacti_naklady(klic, r)
                ws2 = _exp_list(wb, f'Náklady podrobně {klic} {r}', pouzite)
                _exp_tabulka(ws2, _exp_cols_naklady(), rows, _soucet_naklady(rows))

        _podrobne(rok)
        rows = nacti_porovnani(pobocka, rok)
        ws = _exp_list(wb, f'Porovnání nákladů {rok}x{minuly}', pouzite)
        _exp_tabulka(ws, _exp_cols_porovnani(rok, minuly), rows, _soucet_porovnani(rows))
        _podrobne(minuly)

    if not wb.sheetnames:
        return b''
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Sešit ZR (přehled poboček) ───────────────────────────────────────────────

_EXP_KG_POPIS = dict(_PREHLED_METRIKY, **{
    'operace': 'Operace', 'operace_dcery': 'Operace dcery', 'operace_site': 'Operace sítě',
})


def _exp_cols_zr_data() -> list[tuple]:
    cols = _exp_cols_souhrn([('Pobočka', '_pobocka_nazev', 'text', 24),
                             ('Rok', 'rok', 'int', 8),
                             ('Číslo pobočky', '_cislo', 'text', 13),
                             ('Měsíc', 'mesic', 'int', 8)], pocitane=False)
    return cols + [(_EXP_KG_POPIS.get(db, db), db, 'num', 17) for db, _idx in _KG_COLS]


def _exp_cols_zr_souhrn(roky: list[int], mesic: bool) -> list[tuple]:
    cols = [('Pobočka', '_pobocka', 'text', 26)]
    if mesic:
        cols.append(('Měsíc', '_mesic', 'text', 9))
    for db, popis in _PREHLED_METRIKY:
        for rok in roky:
            cols.append((f'{popis} {rok}', f'{db}_{rok}', 'num', 18))
    return cols


def _export_zr_xlsx(pobocky: list[str], roky: list[int]) -> bytes:
    """Soubor ZR: plochý list DATA + počítané souhrny + list „Výsledek …" na
    každou pobočku. Jen pobočky, které uživatel vidí v Přehledu poboček."""
    import io
    wb = _exp_wb()
    pouzite: set = set()

    # DATA – všechny pobočky × roky × měsíce, všechny sloupce tabulky nákladů
    data_rows: list[dict] = []
    souhrny: list[tuple] = []
    for pob in pobocky:
        nazev = _POBOCKY_EXCEL_REVERSE.get(pob, pob)
        cislo = (re.match(r'\s*(\d+)', nazev) or [None, ''])[1]
        raw = [r for r in nacti_souhrn(pob) if int(r['rok']) in roky]
        for r in raw:
            r['_pobocka_nazev'] = nazev
            r['_cislo'] = cislo
        data_rows += raw
        souhrny.append((pob, _priprav_souhrn_pro_grid(raw)))
    _exp_tabulka(_exp_list(wb, 'DATA', pouzite), _exp_cols_zr_data(), data_rows)

    # SOUHRN listy – statické hodnoty (ne kontingenční tabulky) ze stejných
    # funkcí, jaké plní záložky Souhrn / Mezisoučty v Přehledu poboček.
    rows, total = nacti_prehled_souhrn(roky, pobocky)
    _exp_tabulka(_exp_list(wb, 'SOUHRN_TOTAL', pouzite),
                 _exp_cols_zr_souhrn(roky, mesic=False), rows, total or None)
    rows, total = nacti_prehled_mezisoucty(roky, pobocky)
    _exp_tabulka(_exp_list(wb, 'SOUHRN_mezisoučty', pouzite),
                 _exp_cols_zr_souhrn(roky, mesic=True), rows, total or None)

    cols_vysledek = _exp_cols_souhrn([('Rok', '_rok_disp', 'text', 14),
                                      ('Měsíc', '_mesic_disp', 'text', 9)])
    for nazev, rows in souhrny:
        _exp_tabulka(_exp_list(wb, f'Výsledek {nazev}', pouzite), cols_vysledek, rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Tlačítka ─────────────────────────────────────────────────────────────────

async def _exp_stahni(builder, args: tuple, soubor: str, user_name: str, zprava: str):
    """Sestaví sešit mimo event loop (openpyxl je blokující) a pošle ho přes HTTP
    – ui.download.content by u velkých sešitů narazil na limit WebSocketu."""
    import asyncio
    from intranet_sankce import _stahni_pres_http
    ui.notify('Připravuji export…', type='ongoing')
    try:
        data = await asyncio.to_thread(builder, *args)
    except Exception as exc:
        print(f'[vysledky] export error: {exc}')
        ui.notify(f'Export se nezdařil: {exc}', type='negative')
        return
    if not data:
        ui.notify('Není co exportovat.', type='warning')
        return
    _stahni_pres_http(data, soubor)
    intranet_logger.log_activity(user_name, 'Výsledky poboček', zprava)


def _exp_tlacitko_pobocka(pobocka: str, sekce: list[str], user_name: str):
    nazev = _POBOCKY_EXCEL_REVERSE.get(pobocka, pobocka)
    dnes = datetime.date.today().strftime('%Y-%m-%d')
    ui.button('Export do XLSX', icon='download',
              on_click=lambda: _exp_stahni(
                  _export_pobocka_xlsx, (pobocka, sekce),
                  f'{_exp_slug(nazev)}_vysledky_{dnes}.xlsx', user_name,
                  f'Export dat pobočky {nazev} do XLSX')) \
        .props('color=green outline no-caps') \
        .tooltip('Stáhne všechna zobrazená data pobočky do sešitu XLSX')


def _exp_tlacitko_zr(pobocky: list[str], user_name: str):
    rok = datetime.datetime.now().year
    roky = [rok - 1, rok]
    dnes = datetime.date.today().strftime('%d_%m_%Y')
    ui.button('Stáhnout ZR', icon='download',
              on_click=lambda: _exp_stahni(
                  _export_zr_xlsx, (list(pobocky), roky),
                  f'Vysledky_pobocek_{dnes}.xlsx', user_name,
                  f'Export souboru ZR ({rok - 1}+{rok}, {len(pobocky)} poboček)')) \
        .props('color=green outline no-caps') \
        .tooltip(f'Sešit ZR za roky {rok - 1} a {rok} – DATA, souhrny a listy poboček')
