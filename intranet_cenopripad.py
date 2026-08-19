# -*- coding: utf-8 -*-
"""
Modul Cenopřípad — kontrola nastavení letákových cen.

Hotovo:
  • DB schéma + denní import (OP + DATA_POROVNANI → master), ruční (správce).
  • Rozcestník s 5 dlaždicemi (role-gated nákup/obchod) + přepínač modulu.
  • Vstup žadatele: nahrání .xlsx šablony I paste z Excelu (TSV).
  • Engine vyhodnotí → verdikt zeleně/červeně/výjimka. OP a marže v % vidí JEN správce.
  • Workflow: žadatel oprav/„žádám 2. kontrolu" → správce schválí/zamítne → office zpracuje;
    e-maily (office na nový-ok + schváleno, správce na 2. kontrolu) best-effort.
  • Historie s filtry (název/žadatel/stav/datum) + CSV export (souhrn pro všechny,
    detailní vč. OP jen pro správce).

Zbývá (volitelné): automatický import přes intranet_jobs (potřebuje cestu k souborům).
Výpočet je v `cenopripad_vypocet.py` (ověřený self-testem).
"""
import asyncio
import base64
import datetime
import inspect
import io
import json
import os
import re
from decimal import Decimal, ROUND_DOWN

from nicegui import ui, app, context, run

import intranet_data
import intranet_emaily
import intranet_logger
import intranet_notifikace
import cenopripad_vypocet as cp


# Vzorový formulář ke stažení v dlaždici Porovnání (individuální ceny).
VZOR_IND_SOUBOR = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                               "cenopripad_vzor_ind_ceny.xlsx")
VZOR_IND_NAZEV = "Vzorovy_formular_individualni_ceny.xlsx"


# ============================================================================
# Konfigurace typů případů (dlaždic)
#   mapa = {normalizovaná hlavička šablony: pole pro engine}
#   povinne = pole, jejichž sloupec MUSÍ být v šabloně
# ============================================================================
TYPY = {
    "webportal": {
        "nazev": "Leták – webportál", "oddeleni": "nakup", "emoji": "🌐",
        "barva": "border-indigo-200",
        "mapa": {"kód": "kod", "název": "nazev", "akční pc": "akcni_pc",
                 "akční nc": "akcni_nc", "aktuální nc2": "aktualni_nc2"},
        "povinne": {"kod", "akcni_pc", "akcni_nc", "aktualni_nc2"},
    },
    "sklad6": {
        "nazev": "Leták – sklad 6", "oddeleni": "nakup", "emoji": "🏬",
        "barva": "border-teal-200",
        "mapa": {"kód": "kod", "název": "nazev", "pc-akce": "pc_akce",
                 "anc": "anc", "nc2": "nc2"},
        "povinne": {"kod", "pc_akce", "anc", "nc2"},
    },
    "mimoletak": {
        "nazev": "Mimoleták/Výprodej", "oddeleni": "nakup", "emoji": "🏷️",
        "barva": "border-amber-200",
        "mapa": {"kód": "kod", "název": "nazev", "akční fc": "akcni_fc",
                 "cc s dph": "cc_s_dph", "mo s dph": "mo_s_dph",
                 "vo bez dph": "vo_bez_dph"},
        "povinne": {"kod", "akcni_fc"},
    },
    "paima": {
        "nazev": "Paima – OP", "oddeleni": "nakup", "emoji": "📦",
        "barva": "border-cyan-200",
        "mapa": {"kód9": "kod", "název": "nazev", "op": "op"},
        "povinne": {"kod", "op"},
        # Sekce viditelná POUZE správcům (správce nákup, správce, hlavní správce);
        # žadatel ani office ji nevidí.
        "jen_spravce": True,
    },
    "porovnani": {
        "nazev": "Porovnání (individuální ceny)", "oddeleni": "obchod", "emoji": "⚖️",
        "barva": "border-rose-200",
        # Nové sloupce (Zákazník IČO, termín od/do, typy dokladů, pokračovat v hledání…)
        # jsou JEN informativní (ne počítaná pole) — engine je nepoužívá, jen se zobrazí
        # a exportují. Pořadí mapy = pořadí sloupců v náhledu (dle šablony).
        "mapa": {"kód produktu": "kod_produktu", "pc bez dph": "pc_bez_dph",
                 "bonus %": "bonus", "kompenzace": "kompenzace",
                 "zákazník ičo": "zakaznik_ico",
                 "zákazník jméno": "zakaznik",
                 "definice nastavení": "definice",
                 "výchozí cena (např. cp hodnota)": "vychozi_cena",
                 "přirážka/srážka": "prirazka",
                 "termín od": "termin_od", "termín do": "termin_do",
                 "typy dokladů": "typy_dokladu",
                 "pokračovat v hledání nižší ceny": "pokracovat_nizsi"},
        "povinne": {"kod_produktu", "pc_bez_dph"},
    },
    "ncfaktura": {
        "nazev": "NC faktura", "oddeleni": "nakup", "emoji": "🧾",
        "barva": "border-lime-200",
        # Dva podtypy formuláře (Formulář NC / Formulář NC a FC-IMPORT) se detekují
        # z hlavičky v dedikované funkci _zpracuj_ncfaktura (kvůli opakovaným
        # „datum od/do" se sloupce mapují pozičně). Zobrazení/export mají vlastní větve.
        "mapa": {"kód výrobku": "kod", "název výrobku": "nazev"},
        "povinne": {"kod"},
    },
}

# Manuál uživatele (verze 1.0, červen 2026) — zdroj: Cenopripad_Prirucka_Uzivatele.docx.
# Renderuje se v dialogu „Manuál" jako markdown.
MANUAL_MD = """\
# CENOPŘÍPAD

**Příručka uživatele**

*Kontrola nastavení letákových a individuálních cen*

> Verze 1.0  |  Červen 2026  |  Interní dokument

## Obsah
1. Co je Cenopřípad a proč vznikl
2. Přihlášení a orientace v aplikaci
3. Role a přístupová práva
4. Přehled dlaždic a jejich účel
5. Jak podat žádost (postup krok za krokem)
6. Stavy žádostí a barevné označení
7. Postup správce při druhé kontrole
8. Zpracování žádostí (Office)
9. Export a historie případů
10. Důležitá upozornění a pravidla
11. Nejčastější dotazy
12. Technické specifikace

## 1. Co je Cenopřípad a proč vznikl
Cenopřípad je interní aplikace v Moje JIPka, která nahrazuje ruční kontrolu letákových a individuálních cen v Excelu. Dříve probíhala veškerá kontrola manuálně – dnes ji provádí systém automaticky podle nastavených podmínek.

> **Co aplikace řeší:**<br>
> Kontrolu nastavení letákových cen (webportál, sklad 6, mimoleták)<br>
> Kontrolu individuálních cen (porovnání, Paima-OP)<br>
> Automatické vyhodnocení, zda jsou ceny v pořádku<br>
> Řízení toku schválení mezi žadatelem, správcem a office<br>
> Uchování kompletní historie všech případů

Díky aplikaci nemá žadatel přístup k citlivým datům (OP podmínkám a procentuálním porovnáním), které zůstávají skryté a viditelné pouze správci.

## 2. Přihlášení a orientace v aplikaci
Přihlaste se do aplikace Moje JIPka. Na úvodní obrazovce najdete dlaždici označenou:

> **🏷  Cenopřípad**<br>
> Kontrola nastavení cen

Po kliknutí na dlaždici Cenopřípad se zobrazí výběr jednotlivých modulů (dlaždic). Každý modul slouží pro jiný typ cenové kontroly.

## 3. Role a přístupová práva
Aplikace rozlišuje šest typů uživatelů. Každá role vidí pouze to, co potřebuje ke své práci.

| Role | Přístup k dlaždicím | Vidí OP a % podmínky? | Vidí tikety ostatních? |
|---|---|---|---|
| Žadatel Nákup | Letákové dlaždice (Nákup) | NE | NE – jen své vlastní |
| Žadatel Obchod | Porovnání individuálních cen | NE | NE – jen své vlastní |
| Office Nákup | Letákové dlaždice (Nákup) | NE – vidí jen stav | ANO – zpracovává |
| Office Obchod | Porovnání individuálních cen | NE – vidí jen stav | ANO – zpracovává |
| Správce Nákup | Letákové dlaždice (Nákup) | ANO – vše viditelné | ANO – vše |
| Správce | Všechny dlaždice | ANO – vše viditelné | ANO – vše |

> **KLÍČOVÉ PRAVIDLO – PŘÍSNÉ ZABEZPEČENÍ**<br>
> Nikdo kromě Správce nesmí mít přístup k souboru OP podmínek ani k procentuálnímu porovnání.<br>
> Toto omezení je technicky vynuceno aplikací a nelze ho obejít.

## 4. Přehled dlaždic a jejich účel
Po vstupu do Cenoprípadu vidíte tyto dlaždice (podle vaší role):

- **🌐  Leták – webportál** — Kontrola letákových cen - webovém portálu.
- **🏢  Leták – sklad 6** — Kontrola letákových cen - sklad 6.
- **🏷  Mimoleták** — Kontrola mimo letákových cen (CC, MO, VO nebo kombinace).
- **📦  Paima – OP** — Kontrola nastavení OP v systému Paima.
- **⚖️  Porovnání (ind. ceny)** — Kontrola individuálních cen zákazníkům VO.
- **⬆️  Import dat** — Denní import zdrojového souboru.

Poznámka: Karty, jejichž název začíná hvězdičkou (\\*), mohou vykazovat chybové hodnoty – aplikace je označí jako v pořádku. Jedná se o technicky očekávaný stav.

## 5. Jak podat žádost – postup krok za krokem

### 5.1  Příprava dat
Před vstupem do aplikace připravte tabulku, kterou budete vkládat ke kontrole. Každý typ dlaždice má svůj specifický formát:

- Leták – webportál: soubor ve formátu "Webportál – leták"
- Leták – sklad 6: soubor ve formátu "Sklad6 – leták"
- Mimoleták: soubor ve formátu "Mimoleták"
- Paima – OP: soubor ve formátu "Paima"
- Porovnání: soubor ve formátu "Porovnání"

> **Import dat (každé ráno)**<br>
> Před jakoukoli kontrolou je nutné provést denní import přes dlaždici "Import dat".<br>
> Tento import stahuje aktuální data ze skladu 6 (list data\\_porovnání).<br>
> Bez aktuálního importu mohou být výsledky kontroly nepřesné.<br>
> Import provádí pověřená osoba – obvykle ráno před začátkem práce nejpozději do 8:00. Z toho důvodu doporučuji provádět kontroly po 8:00 ráno každý den.

> **Testování nastavených cen**<br>
> V každé dlaždici se nachází při vytváření žádosti tlačítko „testovací“. Pokud zvolíte potvrzení touto cestou, tak testujete pro sebe nastavení cen, ale neodesíláte žádost oficiálně ke schválení.

### 5.2  Postup zadání žádosti
- Přihlaste se do Moje JIPka a klikněte na dlaždici Cenopřípad.
- Vyberte dlaždici odpovídající vašemu typu kontroly (např. Leták – webportál).
- Pojmenujte případ – každý případ musí mít jednoznačný název.
- Vložte připravenou tabulku ke kontrole.
- Klikněte na tlačítko "Porovnat a schválit".
- Počkejte na výsledek automatické kontroly.

### 5.3  Co se stane po odeslání

> **Výsledek: V POŘÁDKU (zelená)**<br>
> Stav v řádku konkrétní žádostí se označí zeleně a změní stav.<br>
> Systém automaticky odešle e-mail pracovníkům office, že mají nový případ ke zpracování.<br>
> Vy jako žadatel nemusíte dělat nic dalšího.

> **Výsledek: CHYBA (červená)**<br>
> Aplikace vám zobrazí hlášku, že žádost není v pořádku.<br>
> Máte dvě možnosti:<br>
> &nbsp;&nbsp;&nbsp;A) Opravit chybu a znovu odeslat ke kontrole<br>
> &nbsp;&nbsp;&nbsp;B) Kliknout na "Mám vše v pořádku a žádám o druhou kontrolu"<br>
> <br>
> Pokud zvolíte možnost B: případ se označí červeně a odesílá se e-mail správci.<br>
> Správce celý případ zkontroluje a rozhodne.

## 6. Stavy žádostí a barevné označení
Každá žádost prochází stavovým systémem. Barva a text stavu vám na první pohled říká, co se s případem děje.

| Stav | Barva | Význam | Kdo jedná |
|---|---|---|---|
| V pořádku | Zelená | Vše je nastaveno správně | Office zpracuje a dotáhne |
| Chyba – čeká na opravu | Červená | Některé ceny jsou mimo limity | Žadatel opraví nebo žádá druhou kontrolu |
| Čeká na schválení správce | Červená | Žadatel požádal o druhou kontrolu | Správce zkontroluje a rozhodne |
| Schváleno správcem | Zelená | Správce potvrdil OK | Office zpracuje a dotáhne |

## 7. Postup správce při druhé kontrole
Správce vstupuje do hry ve dvou situacích: buď sám provádí kontrolu, nebo mu žadatel předal případ k druhé kontrole.

### 7.1  Případ předaný žadatelem k druhé kontrole
- Správce obdrží e-mail s informací, že existuje případ čekající na schválení.
- Vstoupí do příslušné dlaždice a nalezne případ označený červeně.
- Jako jediný vidí veškeré detaily: OP podmínky i procentuální porovnání v tabulce.
- Má dvě možnosti:
- Opravit chybu sám a znovu odeslat ke kontrole – systém znovu vyhodnotí a při OK stavu pošle e-mail office.
- Potvrdit, že je vše v pořádku – případ se označí zeleně a odesílá se e-mail office ke zpracování.

### 7.2  Přímá kontrola správcem
Správce může kdykoli vstoupit do libovolné dlaždice, zadat žádost nebo zkontrolovat existující případy. Správce vidí tikety všech žadatelů a jejich aktuální stavy.

## 8. Zpracování žádostí (Office)
Pracovníci office dostávají e-mailové notifikace a jejich úkolem je dotáhnout schválené případy do nastavení cen.

### 8.1  Co office vidí
- Případy označené zeleně (ať schválené automaticky, nebo správcem)
- Stav vyřízení – pouze zda je v pořádku nebo ne
- Office NEVIDÍ OP podmínky ani procentuální porovnání

### 8.2  Postup office
- Office obdrží e-mail s informací o novém případu ke zpracování.
- Vstoupí do aplikace a ověří zelené označení se stavem ke kontrole příslušného případu.
- Na základě zeleného stavu provede nastavení cen.
- Případ označí jako ukončený (vyřízený).

> **Důležité pro office:**<br>
> Office se NEROZHODUJE na základě dat v tabulce – rozhoduje výhradně na základě stavu (zelená = zpracovat).<br>
> O správnosti cen rozhodla aplikace nebo správce. Office pouze provádí technické nastavení.

## 9. Export a historie případů
Aplikace uchovává kompletní historii všech zadaných případů. Nad seznamem případů jsou k dispozici filtry:

- Zadavatel – filtrování podle osoby, která případ zadala
- Datum zadání – filtrování podle data nebo rozsahu dat
- Stav vyřízení – zobrazení pouze otevřených, vyřešených, čekajících atd.
- Název případu – vyhledávání konkrétního případu podle jeho názvu

Veškerá data lze exportovat do souboru Excel pro další zpracování nebo archivaci.

> **Tip pro pojmenování případů:**<br>
> Vždy zadávejte smysluplný a jednoznačný název případu.<br>
> Doporučený formát: Typ\\_DatumPlatnosti\\_Popis (např. "Letak\\_WP\\_01072026\\_Letni akce")<br>
> Správné pojmenování výrazně usnadní pozdější vyhledávání v historii.

## 10. Důležitá upozornění a pravidla

> **TECHNICKÉ PODMÍNKY KONTROLY:**<br>
> Karty začínající \\* (hvězdičkou) mohou hlásit chybové hodnoty – je to v pořádku, systém je ignoruje.<br>
> Chyby typu "hodnota na text" nebo "dělení nulou" jsou automaticky vyhodnoceny jako chyba (pokud se netýkají karet s \\*).<br>
> Mimoleták: lze vyplnit CC, MO nebo VO nebo jejich kombinaci. Systém kontroluje pouze vyplněné sloupce.

> **PRŮBĚH SCHVALOVÁNÍ:**<br>
> Každý případ musí mít název.<br>
> Systém vždy upřednostní automatickou kontrolu – ručního schválení správce je potřeba pouze při neshodě.<br>
> E-mailové notifikace jsou automatické – nelze je nahradit ústní komunikací.<br>
> Celý průběh schválení je zaznamenán v historii a nelze ho zpětně upravit.

## 11. Nejčastější dotazy

**Proč nevidím OP data a procentuální porovnání?**
Jde o záměrné omezení. Tyto informace jsou obchodně citlivé a vidí je výhradně Správce. Vaše role nevyžaduje přístup k těmto datům.

**Co mám dělat, když si myslím, že jsou ceny správně, ale systém hlásí chybu?**
Klikněte na tlačítko "Mám vše v pořádku a žádám o druhou kontrolu". Případ přejde na Správce, který má přístup ke všem datům a posoudí situaci.

**Vidím pouze své vlastní tikety – je to správně?**
Ano. Žadatelé vidí pouze své vlastní případy. Celkový přehled mají pouze Office a Správce.

**Co mám udělat po obdržení e-mailu o novém případu ke zpracování?**
Vstupte do aplikace, najděte případ označený zeleně, stavem ke zpracování a proveďte nastavení cen. Poté případ označte jako vyřízený.

**Karta začíná hvězdičkou (\\*) a systém hlásí chybu – mám se bát?**
Ne. Karty začínající \\* smějí technicky vykazovat chybové hodnoty. Systém je automaticky bere jako v pořádku. Jde o záměrné chování aplikace.

## 12. Technické specifikace

- Kontrola cen

- V každé aplikaci se nachází stejný styl zadávání:
- Vkládání souboru přes plus nebo přetažení souboru do modrého pole (levá část).
- Lze v excelovém souboru označit veškerá data a pomoci ctrl+C vložit do prázdného pole ctrl+V (pravá část).

- V případě Mimoletáku, Paima = je potřeba nahrávat soubor. Jinak si nestáhne Office doplňující podmínky potřebné k nastavení.
- V případě Leták – sklad 6, Leták – webportál, Individuálních cen = lze použít obě možnosti
- Každý případ lze stornovat s důvodem

Soubory pro práci, nahrávání/kopírování do aplikace pro kontrolu.

- Leták – sklad 6 – formát stažený jako doposud – lze kompletně kopírovat do prázdného pole nebo nahrávat celý soubor ke kontrole
- Leták – webportál – formát stažený jako doposud – lze kompletně kopírovat do prázdného pole nebo nahrávat celý soubor ke kontrole
- Paima – formát jako doposud – výhradně nahrávat soubor ke kontrole
- Mimoleták – zachovaný formát, ale došlo k upravení názvů sloupců, výhradně nahrávat soubor
- Stáhněte si upravené soubory pro práci. Primárně Mimoleták a Ind Ceny, ostátní beze změny
- Pozor: Nesmí se soubory přejmenovat, musí se jmenovat stejně v případě nahrávání.

Soubor IND Cen:

- Kód produktu (bez komentáře)
- PC bez DPH – prodejní cena, za kterou požadujete prodávat zákazníkovi. Vždy částka v Kč (číslo), nikdy procento – procento patří jen do sloupce Přirážka/srážka. Soubor s procentem v PC systém nepřijme.
- Bonus % - Pokud má zákazník bonusové ohodnocení
- Kompenzace – V případě, že domluvená určitá kompenzace u tákazníka
- Zákazník jméno – Např. Hospoda u Lvice
- Definice nastavení – vyplňuje se pokud existuje – např. „PC 39 + 5 %“ (pokud vycházíte pouze z pevné ceny, tak nevyplňujete)
- Výchozí cena – navazuje na definici nastavení (PC 39) – napíšete cenu PC 39
- Přirážka/srážka – navazuje na definici nastavení (+ 5 %) – napíšete `5%`. Může být i přirážka/srážka v Kč (pak BEZ „%“, např. `10` = +10 Kč, `-2` = −2 Kč).
- Termín od – nastavení ceny (start)
- Termín do nastavení ceny (konec)
- Typ dokladů – bez komentáře
- Pokračování v hledání nižší ceny - ano/ne

Pozor:

- Procenta zadávejte SE znakem „%“ — např. 5 % → napište `5%`, srážka −3 % → `-3%`. Přirážku/srážku v Kč zadávejte BEZ „%“ (např. `10` = +10 Kč, `-2` = −2 Kč). Tím systém rozpozná, zda jde o procenta, nebo koruny.
- Pokud budete chtít nastavovat ceny na skupiny, a ne na kód vyplňte tabulku, zažádejte – vypadne Vám chyba a automaticky zažádejte o druhou kontrolu. Poté prověří správce.

---

*Cenopřípad – Příručka uživatele  |  Verze 1.0  |  Červen 2026*

*Interní dokument – nepublikovat mimo společnost*
"""


def _stahni_vzor_ind():
    """Stáhne vzorový formulář pro individuální ceny (.xlsx) z kořene projektu.
    Před stažením ukáže na 3 s dialog s upozorněním na manuál."""
    if not os.path.exists(VZOR_IND_SOUBOR):
        ui.notify("Vzorový formulář není k dispozici.", type="negative")
        return
    def _rozumim():
        dlg.close()
        ui.download.file(VZOR_IND_SOUBOR, VZOR_IND_NAZEV)

    with ui.dialog() as dlg, ui.card().classes("gap-4 p-6").style("max-width: 420px"):
        with ui.row().classes("items-center gap-3 no-wrap"):
            ui.icon("info", size="2rem").classes("text-emerald-600")
            ui.label("Před stažením").classes("text-lg font-bold")
        ui.label("Pro správné vyplnění formuláře je vysoce doporučeno použít manuál, "
                 "který je umístěn na vedlejším tlačítku.") \
            .classes("text-sm text-gray-600 leading-relaxed")
        with ui.row().classes("w-full justify-end"):
            ui.button("Rozumím", on_click=_rozumim).props("unelevated color=green")
    dlg.open()


def _dialog_manual():
    """Zobrazí manuál uživatele (markdown) v rolovatelném dialogu."""
    with ui.dialog() as dlg, ui.card().style(
            "max-width: 96vw; width: 900px; max-height: 90vh; overflow-y: auto"):
        with ui.row().classes("w-full items-center gap-2 sticky top-0 bg-white z-10 pb-2 "
                              "border-b border-gray-200"):
            ui.icon("menu_book", size="1.6rem").classes("text-emerald-600")
            ui.label("Manuál uživatele").classes("text-xl font-bold")
            ui.space()
            ui.button(icon="close", on_click=dlg.close) \
                .props("flat round dense").classes("text-gray-500")
        # .nicegui-markdown už styluje nadpisy/seznamy/citace/tabulky (rámečky, padding).
        # w-full + overflow-x-auto = na úzkém displeji se případná široká tabulka odroluje
        # vodorovně místo rozbití layoutu (jinak se buňky lámou a vejdou se).
        ui.markdown(MANUAL_MD).classes("w-full overflow-x-auto text-sm leading-relaxed")
    dlg.open()


_STAV_BADGE = {
    "vyhodnoceno_ok":    ("V pořádku", "green"),
    "vyhodnoceno_chyba": ("Není v pořádku", "red"),
    "ceka_na_spravce":   ("Čeká na správce", "orange"),
    "schvaleno":         ("Schváleno", "green"),
    "castecne_schvaleno": ("Částečně schváleno", "teal"),
    "zpracovano":        ("Zpracováno", "blue"),
    "uzavreno":          ("Uzavřeno", "dark"),
    "zamitnuto":         ("Zamítnuto", "red"),
    "stornovano":        ("Stornováno", "grey"),
    "delisting":         ("Delisting", "deep-orange"),
}


def _norm(h):
    """Normalizace hlavičky: malá písmena, bez vícenásobných mezer, bez koncové ':'."""
    s = " ".join(str(h or "").strip().lower().split())
    return s.rstrip(":").strip()


# ============================================================================
# Práva / viditelnost
# ============================================================================
def _je_spravce(p):
    # „Správce (vše)" i jeho varianta bez e-mailů mají STEJNÁ práva; liší se jen tím,
    # že varianta `_bez_emailu` se nezahrnuje do příjemců notifikací (viz _emaily_*).
    return ("vse" in p or "cenopripad_spravce" in p
            or "cenopripad_spravce_bez_emailu" in p)


def _vidi_import(p):
    """Smí do dlaždice „Import dat". Plný správce nebo vkladatel (jen nahrávání dat)."""
    return _je_spravce(p) or "cenopripad_vkladatel" in p


def _vidi_op(p):
    """Smí vidět OP a marže v % (tabulka). JEN plný správce nebo správce nákup."""
    return _je_spravce(p) or "cenopripad_spravce_nakup" in p


def _vidi_costprice(p):
    """Smí vidět sloupce CostPrice (Výběrová řízení) u ind. cen.
    Plný správce modulu (vč. varianty bez e-mailů) nebo Office obchod."""
    return _je_spravce(p) or "cenopripad_office_obchod" in p


def _je_office(p):
    return any(x in p for x in ("cenopripad_office_nakup", "cenopripad_office_obchod"))


def _vidi_vsechny_pripady(p):
    """Office/správce vidí všechny případy typu; žadatel jen své."""
    return _je_spravce(p) or _vidi_op(p) or _je_office(p)


def _vidi_oddeleni(p):
    """Smí vidět případy ostatních členů svého oddělení (právo přiřazené ODDĚLENÍ).
    Nedává plnou viditelnost — rozsah dopočítá `_kolegove_oddeleni` dle členství."""
    return "cenopripad_zobrazeni_oddeleni" in p


def _kolegove_oddeleni(user_id):
    """ID uživatelů ze stejných oddělení, kterým je přiřazeno právo
    'cenopripad_zobrazeni_oddeleni'. Tzn. spolučlenové oddělení, kde toto právo
    platí (vč. zadaného uživatele). Prázdné, pokud takové oddělení nemá."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT DISTINCT dtu2.user_iduser
            FROM department_To_user dtu1
            JOIN department_To_privileges dtp
              ON dtu1.department_iddepartment = dtp.department_iddepartment
            JOIN privileges p ON dtp.privileges_idprivileges = p.idprivileges
            JOIN department_To_user dtu2
              ON dtu2.department_iddepartment = dtu1.department_iddepartment
            WHERE dtu1.user_iduser = %s AND p.name = 'cenopripad_zobrazeni_oddeleni'
        """, (user_id,))
        return [r[0] for r in cur.fetchall()]
    except Exception as e:
        print(f"Chyba _kolegove_oddeleni: {e}")
        return []
    finally:
        conn.close()


# Prefix práva „Schvalovatel – oddělení" (záměrně krátký kvůli VARCHAR(45) v
# privileges.name). Stejný řetězec generuje intranet_prava.ziskej_kompletni_seznam_prav.
_SCHVAL_ODD_PREFIX = "cp_schval_odd_"

# Prefix práva „Hlavní vedoucí oddělení" (kategorie Manažer oddělení, generuje
# ho intranet_prava.ziskej_kompletni_seznam_prav stejně jako _SCHVAL_ODD_PREFIX).
_VEDOUCI_ODD_PREFIX = "hlavni_vedouci_"


def _schval_odd_slugs(p):
    """Slugy oddělení (lowercase názvy), pro která má uživatel právo schvalovat
    žádosti (právo 'cp_schval_odd_<oddělení>')."""
    return {x[len(_SCHVAL_ODD_PREFIX):] for x in p if x.startswith(_SCHVAL_ODD_PREFIX)}


def _vedouci_odd_slugs(p):
    """Slugy oddělení, kde je uživatel hlavní vedoucí (právo
    'hlavni_vedouci_<oddělení>'). Dává jen ČTENÍ případů daného oddělení —
    schvalování zůstává na 'cp_schval_odd_<oddělení>'."""
    return {x[len(_VEDOUCI_ODD_PREFIX):] for x in p if x.startswith(_VEDOUCI_ODD_PREFIX)}


def _oddeleni_uzivatele_lower(user_id):
    """Názvy oddělení uživatele (lowercase) — pro porovnání se slugy práv."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor()
        cur.execute("SELECT LOWER(d.name) FROM department_To_user dtu "
                    "JOIN department d ON dtu.department_iddepartment = d.iddepartment "
                    "WHERE dtu.user_iduser = %s", (user_id,))
        return [r[0] for r in cur.fetchall() if r[0]]
    except Exception as e:
        print(f"Chyba _oddeleni_uzivatele_lower: {e}")
        return []
    finally:
        conn.close()


def _je_schvalovatel_oddeleni(pripad, p):
    """True, pokud uživatel smí schvalovat tento případ jako schvalovatel oddělení
    žadatele (právo 'cp_schval_odd_<oddělení žadatele>')."""
    slugs = _schval_odd_slugs(p)
    if not slugs:
        return False
    return any(o in slugs for o in _oddeleni_uzivatele_lower(pripad.get("zadavatel_id")))


def _uzivatele_dle_odd_slugu(slugs):
    """ID všech uživatelů v oddělěních daných slugů (lowercase názvy)."""
    if not slugs:
        return []
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor()
        ph = ",".join(["%s"] * len(slugs))
        cur.execute(f"SELECT DISTINCT dtu.user_iduser FROM department d "
                    f"JOIN department_To_user dtu ON dtu.department_iddepartment = d.iddepartment "
                    f"WHERE LOWER(d.name) IN ({ph})", tuple(slugs))
        return [r[0] for r in cur.fetchall()]
    except Exception as e:
        print(f"Chyba _uzivatele_dle_odd_slugu: {e}")
        return []
    finally:
        conn.close()


def _zadavatele_ke_schvaleni(p):
    """ID uživatelů ze všech oddělení, která daný uživatel smí schvalovat
    (právo 'cp_schval_odd_<oddělení>'). Prázdné, pokud žádné takové právo nemá."""
    return _uzivatele_dle_odd_slugu(_schval_odd_slugs(p))


def _zadavatele_vedeni(p):
    """ID uživatelů z oddělení, kde je uživatel hlavní vedoucí — jeho podřízení.
    Prázdné, pokud žádné oddělení nevede."""
    return _uzivatele_dle_odd_slugu(_vedouci_odd_slugs(p))


def _podrizeni(user_id):
    """ID aktivních uživatelů, kteří mají `user_id` mezi přímými nadřízenými
    (tabulka user_manager — pole "Přímí nadřízení" v editaci uživatele).
    Nadřízený tak vidí případy své asistentky. Pouze čtení — právo schvalovat
    zůstává na 'cp_schval_odd_<oddělení>'."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor()
        cur.execute("SELECT um.user_id FROM user_manager um "
                    "JOIN user u ON u.iduser = um.user_id "
                    "WHERE um.manager_id = %s AND u.is_active = 1", (user_id,))
        return [r[0] for r in cur.fetchall()]
    except Exception as e:
        print(f"Chyba _podrizeni: {e}")
        return []
    finally:
        conn.close()


def _viditelne_typy(p):
    vidi_nakup = _je_spravce(p) or any(x in p for x in (
        "cenopripad_spravce_nakup",
        "cenopripad_zadatel_nakup", "cenopripad_office_nakup"))
    vidi_obchod = _je_spravce(p) or any(x in p for x in (
        "cenopripad_zadatel_obchod", "cenopripad_office_obchod"))
    # Schvalovatel oddělení musí na dlaždice, aby se k případům ke schválení dostal.
    schvalovatel = bool(_schval_odd_slugs(p))
    spravce = _vidi_op(p)   # správce nákup / správce / hlavní správce
    out = []
    for klic, cfg in TYPY.items():
        if cfg.get("jen_spravce") and not spravce:   # např. Paima — jen pro správce
            continue
        if cfg["oddeleni"] == "nakup" and (vidi_nakup or schvalovatel):
            out.append(klic)
        elif cfg["oddeleni"] == "obchod" and (vidi_obchod or schvalovatel):
            out.append(klic)
    return out


def _muze_zadat(typ, p):
    if _je_spravce(p):
        return True
    if TYPY[typ].get("jen_spravce"):   # Paima — zadávat smí jen správce
        return _vidi_op(p)
    if TYPY[typ]["oddeleni"] == "nakup":
        return "cenopripad_zadatel_nakup" in p or "cenopripad_spravce_nakup" in p
    return "cenopripad_zadatel_obchod" in p


# ============================================================================
# DB
# ============================================================================
def inicializace_cenopripad_db():
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS cenopripad_master (
                kod VARCHAR(40) PRIMARY KEY,
                sortiment VARCHAR(40),
                sortiment_popis VARCHAR(255),
                nazev VARCHAR(255),
                nc DOUBLE, nc2 DOUBLE, nc3 DOUBLE, nc4 DOUBLE, nc5 DOUBLE, dnc DOUBLE,
                dph DOUBLE, id_kod VARCHAR(40), op DOUBLE,
                INDEX idx_sortiment (sortiment)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS cenopripad_op (
                sortiment VARCHAR(40) PRIMARY KEY, op DOUBLE
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        # CostPrice (Výběrová řízení) — nákupní ceny v čase, párování dle unikátního kódu.
        # cp_a = aktuální, cp_b = +7 dní, cp_c = +14 dní, cp_d = +29 dní.
        cur.execute("""
            CREATE TABLE IF NOT EXISTS cenopripad_costprice (
                kod VARCHAR(40) COLLATE utf8mb4_bin PRIMARY KEY,
                cp_a DOUBLE, cp_b DOUBLE, cp_c DOUBLE, cp_d DOUBLE
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS cenopripad_import (
                id INT AUTO_INCREMENT PRIMARY KEY,
                kdy DATETIME DEFAULT CURRENT_TIMESTAMP,
                uzivatel VARCHAR(255),
                pocet_master INT DEFAULT 0, pocet_op INT DEFAULT 0,
                pocet_costprice INT DEFAULT 0,
                soubor_op VARCHAR(255), soubor_data VARCHAR(255),
                soubor_costprice VARCHAR(255),
                INDEX idx_kdy (kdy)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        # ── Kontrolní data letáků (samostatná dlaždice) ──────────────────────
        cur.execute("""
            CREATE TABLE IF NOT EXISTS letaky_pripady (
                id INT AUTO_INCREMENT PRIMARY KEY,
                cislo VARCHAR(20),
                nazev VARCHAR(255),                     -- název případu (zadává žadatel)
                kanal VARCHAR(4),                       -- VO / MO / SP
                druh VARCHAR(10) DEFAULT 'kontrolni',   -- kontrolni / finalni (vlastní práva)
                zadavatel_id INT, zadavatel_jmeno VARCHAR(255),
                datum_zadani DATETIME DEFAULT CURRENT_TIMESTAMP,
                stav VARCHAR(20) DEFAULT 'odeslano',     -- odeslano / zpracovano
                pocet_radku INT DEFAULT 0,
                pocet_chyb_k2 INT DEFAULT 0,
                soubor_nazev VARCHAR(255),
                sloupce_json TEXT,                      -- {hlavička: písmeno sloupce v Excelu}
                aktualizovano DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                INDEX idx_stav (stav), INDEX idx_zadavatel (zadavatel_id)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS letaky_radky (
                id INT AUTO_INCREMENT PRIMARY KEY,
                pripad_id INT, poradi INT,
                dod VARCHAR(40),
                kod VARCHAR(40),
                k2 VARCHAR(60),
                k2_chyba TINYINT DEFAULT 0,
                nc3_anc DOUBLE, nc_anc DOUBLE, rozdil_korekce DOUBLE, akcni_prirazka DOUBLE,
                data_json MEDIUMTEXT,
                INDEX idx_pripad (pripad_id)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        # Původní nahraný soubor (pro pozdější export „Označená sestava").
        cur.execute("""
            CREATE TABLE IF NOT EXISTS letaky_soubory (
                pripad_id INT PRIMARY KEY,
                nazev VARCHAR(255),
                data LONGBLOB
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        # Chat k sestavě (fáze 3) — připraveno.
        cur.execute("""
            CREATE TABLE IF NOT EXISTS letaky_chat (
                id INT AUTO_INCREMENT PRIMARY KEY,
                pripad_id INT, user_id INT, jmeno VARCHAR(255),
                zprava TEXT,
                kdy DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_pripad (pripad_id)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)

        # Případy (tikety)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS cenopripad_pripady (
                id INT AUTO_INCREMENT PRIMARY KEY,
                cislo VARCHAR(20),
                nazev VARCHAR(255) NOT NULL,
                typ VARCHAR(20), oddeleni VARCHAR(10),
                zadavatel_id INT, zadavatel_jmeno VARCHAR(255),
                datum_zadani DATETIME DEFAULT CURRENT_TIMESTAMP,
                stav VARCHAR(30) DEFAULT 'vyhodnoceno_ok',
                pocet_radku INT DEFAULT 0, pocet_chyb INT DEFAULT 0,
                vysledek_ok TINYINT DEFAULT 0,
                storno_duvod VARCHAR(500),
                poznamka VARCHAR(1000),
                zamitnuti_duvod VARCHAR(1000),
                aktualizovano DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                INDEX idx_typ (typ), INDEX idx_zadavatel (zadavatel_id), INDEX idx_stav (stav)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        # Řádky případu (OP a marže jen pro správce — gateuje UI, ne DB)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS cenopripad_radky (
                id INT AUTO_INCREMENT PRIMARY KEY,
                pripad_id INT, poradi INT,
                vstup_json TEXT,
                nazev_karty VARCHAR(255), sortiment_popis VARCHAR(255), kod VARCHAR(40),
                verdikt VARCHAR(10), duvod VARCHAR(255),
                op DOUBLE, marze_json TEXT,
                INDEX idx_pripad (pripad_id)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        # Původní nahraný soubor případu (ke stažení na konci cesty) — blob zvlášť,
        # ať seznam/detail dotazy netahají binárku.
        cur.execute("""
            CREATE TABLE IF NOT EXISTS cenopripad_soubory (
                pripad_id INT PRIMARY KEY,
                nazev VARCHAR(255),
                data LONGBLOB
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        # Přílohy/fotky k případu (žadatel je přikládá k žádosti o druhou kontrolu).
        # Víc příloh na případ → samostatná tabulka s vlastním id; data jako blob.
        cur.execute("""
            CREATE TABLE IF NOT EXISTS cenopripad_prilohy (
                id INT AUTO_INCREMENT PRIMARY KEY,
                pripad_id INT,
                nazev VARCHAR(255),
                typ VARCHAR(120),
                data LONGBLOB,
                kdo VARCHAR(255),
                kdy DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_pripad (pripad_id)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        # Historie/průběh případu („očičko") — každý krok workflow = 1 záznam.
        cur.execute("""
            CREATE TABLE IF NOT EXISTS cenopripad_historie (
                id INT AUTO_INCREMENT PRIMARY KEY,
                pripad_id INT,
                akce VARCHAR(80),
                detail VARCHAR(1000),
                kdo VARCHAR(255),
                kdy DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_pripad (pripad_id)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        # Migrace: sortiment_popis do starších tabulek
        for _t in ("cenopripad_master", "cenopripad_radky"):
            cur.execute("SELECT COUNT(*) FROM information_schema.COLUMNS WHERE "
                        "TABLE_SCHEMA=DATABASE() AND TABLE_NAME=%s AND "
                        "COLUMN_NAME='sortiment_popis'", (_t,))
            if cur.fetchone()[0] == 0:
                cur.execute(f"ALTER TABLE {_t} ADD COLUMN sortiment_popis VARCHAR(255)")
        # Migrace: nové sloupce případů
        for _col, _typ in (("storno_duvod", "VARCHAR(500)"), ("poznamka", "VARCHAR(1000)"),
                           ("testovaci", "TINYINT DEFAULT 0"),
                           ("poznamka_zadani", "VARCHAR(1000)"),
                           ("zamitnuti_duvod", "VARCHAR(1000)"),
                           ("pobocka", "VARCHAR(3) DEFAULT NULL")):
            cur.execute("SELECT COUNT(*) FROM information_schema.COLUMNS WHERE "
                        "TABLE_SCHEMA=DATABASE() AND TABLE_NAME='cenopripad_pripady' AND "
                        "COLUMN_NAME=%s", (_col,))
            if cur.fetchone()[0] == 0:
                cur.execute(f"ALTER TABLE cenopripad_pripady ADD COLUMN {_col} {_typ}")
        # Jednorázový backfill pobočky u starých případů — idempotentní přes WHERE p.pobocka IS NULL
        cur.execute("UPDATE cenopripad_pripady p JOIN user u ON p.zadavatel_id=u.iduser "
                    "SET p.pobocka=u.pobocka WHERE p.pobocka IS NULL")
        # Migrace: příznak neschválené položky (částečné schválení IND)
        cur.execute("SELECT COUNT(*) FROM information_schema.COLUMNS WHERE "
                    "TABLE_SCHEMA=DATABASE() AND TABLE_NAME='cenopripad_radky' AND "
                    "COLUMN_NAME='neschvaleno'")
        if cur.fetchone()[0] == 0:
            cur.execute("ALTER TABLE cenopripad_radky ADD COLUMN neschvaleno TINYINT DEFAULT 0")
        # Migrace: název případu do letákových sestav
        cur.execute("SELECT COUNT(*) FROM information_schema.COLUMNS WHERE "
                    "TABLE_SCHEMA=DATABASE() AND TABLE_NAME='letaky_pripady' AND "
                    "COLUMN_NAME='nazev'")
        if cur.fetchone()[0] == 0:
            cur.execute("ALTER TABLE letaky_pripady ADD COLUMN nazev VARCHAR(255) AFTER cislo")
        # Migrace: mapa „hlavička → písmeno sloupce v Excelu" (starým sestavám se dopočítá
        # z uloženého původního souboru při prvním otevření detailu).
        cur.execute("SELECT COUNT(*) FROM information_schema.COLUMNS WHERE "
                    "TABLE_SCHEMA=DATABASE() AND TABLE_NAME='letaky_pripady' AND "
                    "COLUMN_NAME='sloupce_json'")
        if cur.fetchone()[0] == 0:
            cur.execute("ALTER TABLE letaky_pripady ADD COLUMN sloupce_json TEXT")
        # Migrace: druh sestavy (kontrolní / finální) — staré sestavy jsou kontrolní.
        cur.execute("SELECT COUNT(*) FROM information_schema.COLUMNS WHERE "
                    "TABLE_SCHEMA=DATABASE() AND TABLE_NAME='letaky_pripady' AND "
                    "COLUMN_NAME='druh'")
        if cur.fetchone()[0] == 0:
            cur.execute("ALTER TABLE letaky_pripady ADD COLUMN druh VARCHAR(10) "
                        "DEFAULT 'kontrolni' AFTER kanal")
            cur.execute("UPDATE letaky_pripady SET druh='kontrolni' WHERE druh IS NULL")
        # Migrace: log importu — sloupce pro CostPrice (Výběrová řízení)
        for _col, _typ in (("pocet_costprice", "INT DEFAULT 0"),
                           ("soubor_costprice", "VARCHAR(255)")):
            cur.execute("SELECT COUNT(*) FROM information_schema.COLUMNS WHERE "
                        "TABLE_SCHEMA=DATABASE() AND TABLE_NAME='cenopripad_import' AND "
                        "COLUMN_NAME=%s", (_col,))
            if cur.fetchone()[0] == 0:
                cur.execute(f"ALTER TABLE cenopripad_import ADD COLUMN {_col} {_typ}")
        conn.commit()
        cur.close()
    except Exception as e:
        print(f"Chyba při inicializaci DB Cenopřípad: {e}")
    finally:
        conn.close()


def _str(v, n):
    return None if v is None else str(v)[:n]


# ---------------------------------------------------------------------------
# Import masteru (Fáze 1) — beze změny, běží v threadu
# ---------------------------------------------------------------------------
def _nacti_op_mapu_z_db(cur):
    """Existující OP mapa {sortiment: op} z DB — pro DATA-only import (zachová OP)."""
    cur.execute("SELECT sortiment, op FROM cenopripad_op")
    return {s: o for s, o in cur.fetchall()}


def _importuj_sync(op_raw, data_raw, op_name, data_name, user_name):
    """Import OP a/nebo DATA. Lze importovat i JEN JEDEN soubor:
      • jen DATA → master se přepočítá s EXISTUJÍCÍM OP z DB (OP zůstane beze změny;
        DATA porovnání se mění denně);
      • jen OP → nahradí se OP tabulka a u STÁVAJÍCÍHO masteru se přepočítá sloupec
        `op` přes sortiment (DATA řádky zůstanou; OP se mění zřídka);
      • oba → kompletní náhrada.
    Záznam importu uloží NULL u souboru, který se neimportoval.
    Vrací (pocet_master_v_db, pocet_op_v_db, chyba|None)."""
    if not op_raw and not data_raw:
        return 0, 0, "Nevybrán žádný soubor k importu."

    # Parsování mimo DB transakci (dlouhý parse ať nedrží zámky).
    op_mapa = None
    if op_raw:
        try:
            op_mapa = cp.nacti_op_mapu(op_raw)
        except Exception as e:
            return 0, 0, f"OP soubor nelze načíst: {e}"
        if not op_mapa:
            return 0, 0, "OP soubor je prázdný nebo má neznámý formát."

    # Master se staví z DATA + OP. OP bereme z nově nahraného souboru, jinak z DB.
    # OP mapu čteme krátkodobým spojením a hned ho vracíme — parse velkého xlsx
    # trvá minuty a spojení z poolu se přes něj držet nesmí (server ho mezitím
    # zabije přes wait_timeout a padne to až v conn.close()).
    master = None
    if data_raw:
        mapa = op_mapa
        if mapa is None:
            conn = intranet_data.get_db_connection()
            if not conn:
                return 0, 0, "Chyba připojení k databázi."
            try:
                cur = conn.cursor()
                mapa = _nacti_op_mapu_z_db(cur)
                cur.close()
            except Exception as e:
                return 0, 0, f"Chyba čtení OP mapy z databáze: {e}"
            finally:
                conn.close()
        try:
            master = cp.nacti_master(data_raw, mapa)
        except Exception as e:
            return 0, 0, f"DATA soubor nelze načíst (list DATA_POROVNANI?): {e}"
        if not master:
            return 0, 0, "V DATA souboru nejsou žádné řádky."

    conn = intranet_data.get_db_connection()
    if not conn:
        return 0, 0, "Chyba připojení k databázi."
    try:
        cur = conn.cursor()

        # 1) OP tabulka (jen když se OP importuje)
        if op_mapa is not None:
            cur.execute("DELETE FROM cenopripad_op")
            op_rows = [(_str(s, 40), v) for s, v in op_mapa.items()]
            for i in range(0, len(op_rows), 2000):
                cur.executemany("INSERT INTO cenopripad_op (sortiment, op) VALUES (%s, %s)",
                                op_rows[i:i + 2000])

        # 2) MASTER
        if master is not None:
            # DATA importována → kompletní náhrada masteru (op už dle `mapa`)
            cur.execute("DELETE FROM cenopripad_master")
            cols = ("kod", "sortiment", "sortiment_popis", "nazev", "nc", "nc2", "nc3", "nc4",
                    "nc5", "dnc", "dph", "id_kod", "op")
            sql = (f"INSERT INTO cenopripad_master ({','.join(cols)}) "
                   f"VALUES ({','.join(['%s'] * len(cols))})")
            davka = [(kod, _str(m["sortiment"], 40), _str(m["sortiment_popis"], 255),
                      _str(m["nazev"], 255),
                      m["nc"], m["nc2"], m["nc3"], m["nc4"], m["nc5"], m["dnc"],
                      m["dph"], _str(m["id"], 40), m["op"]) for kod, m in master.items()]
            for i in range(0, len(davka), 2000):
                cur.executemany(sql, davka[i:i + 2000])
        elif op_mapa is not None:
            # JEN OP (bez DATA) → přepočti sloupec op ve stávajícím masteru přes sortiment
            # (sortiment bez záznamu v novém OP → op = NULL; engine bere None jako 0).
            cur.execute("UPDATE cenopripad_master m "
                        "LEFT JOIN cenopripad_op o ON m.sortiment = o.sortiment "
                        "SET m.op = o.op")

        # 3) Záznam importu — NULL u souboru, který se neimportoval
        cur.execute("SELECT COUNT(*) FROM cenopripad_master")
        pm = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM cenopripad_op")
        po = cur.fetchone()[0]
        cur.execute("INSERT INTO cenopripad_import "
                    "(uzivatel, pocet_master, pocet_op, soubor_op, soubor_data) "
                    "VALUES (%s, %s, %s, %s, %s)",
                    (user_name, pm, po,
                     _str(op_name, 255) if op_raw else None,
                     _str(data_name, 255) if data_raw else None))
        conn.commit()
        cur.close()
        return pm, po, None
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return 0, 0, f"Chyba zápisu do databáze: {e}"
    finally:
        conn.close()


def pocty_v_db():
    conn = intranet_data.get_db_connection()
    if not conn:
        return (0, 0)
    try:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM cenopripad_master")
        m = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM cenopripad_op")
        o = cur.fetchone()[0]
        cur.close()
        return (m, o)
    except Exception:
        return (0, 0)
    finally:
        conn.close()


# ----------------------------------------------------------------------------
# CostPrice (Výběrová řízení) — import + lookup
# ----------------------------------------------------------------------------
def _parse_costprice(raw_bytes, filename=""):
    """Výběrová řízení (.xlsx) → {kod: (cp_a, cp_b, cp_c, cp_d)}. Sloupce dle hlavičky:
    první „Kód" + sloupce obsahující „CostPrice" v pořadí (A=aktuální, B=+7, C=+14, D=+29)."""
    rows = _nacti_rows(raw_bytes, filename)
    if not rows:
        return None, "Soubor je prázdný."
    hdr_idx = None
    for i, row in enumerate(rows[:40]):
        norm = [_norm(c) for c in row]
        if any(n in ("kód", "kod") for n in norm) and any("costprice" in n for n in norm):
            hdr_idx = i
            break
    if hdr_idx is None:
        return None, "Nenalezena hlavička se sloupci „Kód“ a „CostPrice“."
    header = [_norm(c) for c in rows[hdr_idx]]
    kod_i, cp_idx = None, []
    for j, h in enumerate(header):
        if kod_i is None and h in ("kód", "kod"):
            kod_i = j
        elif "costprice" in h:
            cp_idx.append(j)
    if kod_i is None or not cp_idx:
        return None, "Chybí sloupec „Kód“ nebo „CostPrice“."
    cp_idx = cp_idx[:4]
    out = {}
    for r in rows[hdr_idx + 1:]:
        if r is None:
            continue
        kod = _str(r[kod_i] if kod_i < len(r) else None, 40)
        if not kod:
            continue
        vals = [cp.parse_cislo(r[j]) if j < len(r) else None for j in cp_idx]
        while len(vals) < 4:
            vals.append(None)
        out[kod] = tuple(vals[:4])
    return out, None


def _importuj_costprice_sync(raw, name, user_name):
    """Plná náhrada tabulky CostPrice. Vrací (pocet, chyba|None)."""
    if not raw:
        return 0, "Nevybrán soubor."
    try:
        mapa, err = _parse_costprice(raw, name)
    except Exception as e:
        return 0, f"Soubor nelze načíst: {e}"
    if err:
        return 0, err
    conn = intranet_data.get_db_connection()
    if not conn:
        return 0, "Není připojení k databázi."
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM cenopripad_costprice")
        davka = [(k, a, b, c, d) for k, (a, b, c, d) in mapa.items()]
        for i in range(0, len(davka), 2000):
            cur.executemany("INSERT INTO cenopripad_costprice "
                            "(kod, cp_a, cp_b, cp_c, cp_d) VALUES (%s,%s,%s,%s,%s)",
                            davka[i:i + 2000])
        cur.execute("SELECT COUNT(*) FROM cenopripad_costprice")
        n = cur.fetchone()[0]
        cur.execute("INSERT INTO cenopripad_import "
                    "(uzivatel, pocet_costprice, soubor_costprice) VALUES (%s,%s,%s)",
                    (user_name, n, _str(name, 255)))
        conn.commit()
        cur.close()
        return n, None
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return 0, f"Chyba zápisu do databáze: {e}"
    finally:
        conn.close()


def pocet_costprice_v_db():
    conn = intranet_data.get_db_connection()
    if not conn:
        return 0
    try:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM cenopripad_costprice")
        n = cur.fetchone()[0]
        cur.close()
        return n
    except Exception:
        return 0
    finally:
        conn.close()


def _costprice_dle_kodu(kods):
    """{kod (kanonicky, 8 míst): (cp_a, cp_b, cp_c, cp_d)} pro zadané kódy.
    Kódy v případech i v CostPrice tabulce mohou být s vodicími nulami i bez
    → hledá se v obou tvarech; klíč výsledku je vždy zfill(8) (viz _export_kod),
    caller proto hledá přes _export_kod(kod)."""
    hledane = set()
    for k in dict.fromkeys(kods):
        nk = cp.normalizuj_kod(k)
        if not nk:
            continue
        hledane |= {nk, nk.zfill(8), nk.lstrip("0") or nk}
    if not hledane:
        return {}
    conn = intranet_data.get_db_connection()
    if not conn:
        return {}
    try:
        cur = conn.cursor()
        ph = ",".join(["%s"] * len(hledane))
        cur.execute(f"SELECT kod, cp_a, cp_b, cp_c, cp_d FROM cenopripad_costprice "
                    f"WHERE kod IN ({ph})", tuple(hledane))
        out = {}
        for r in cur.fetchall():
            out.setdefault(_export_kod(r[0]), (r[1], r[2], r[3], r[4]))
        cur.close()
        return out
    except Exception as e:
        print(f"[cenopripad] _costprice_dle_kodu: {e}")
        return {}
    finally:
        conn.close()


def posledni_import():
    conn = intranet_data.get_db_connection()
    if not conn:
        return None
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT * FROM cenopripad_import ORDER BY id DESC LIMIT 1")
        r = cur.fetchone()
        cur.close()
        return r
    except Exception:
        return None
    finally:
        conn.close()


def stav_masteru():
    """Co je AKTUÁLNĚ v masteru — poslední import ZVLÁŠŤ pro OP a pro DATA (každý
    soubor se mohl nahrát jindy). Vrací dict {'op': {nazev,kdy,uzivatel}|None,
    'data': {...}|None}. Bere poslední `cenopripad_import`, kde byl daný soubor
    vyplněn (od jednosouborového importu se nevyplněný ukládá jako NULL)."""
    out = {"op": None, "data": None}
    conn = intranet_data.get_db_connection()
    if not conn:
        return out
    try:
        cur = conn.cursor(dictionary=True)
        for klic, sl in (("op", "soubor_op"), ("data", "soubor_data")):
            cur.execute(f"SELECT {sl} AS nazev, kdy, uzivatel FROM cenopripad_import "
                        f"WHERE {sl} IS NOT NULL ORDER BY id DESC LIMIT 1")
            out[klic] = cur.fetchone()
        cur.close()
    except Exception:
        pass
    finally:
        conn.close()
    return out


# ---------------------------------------------------------------------------
# Vyhodnocení případu (Fáze 3a)
# ---------------------------------------------------------------------------
def _nacti_master_pro_kody(kody):
    """Načte z DB jen produkty pro dané kódy (efektivní místo 139 tis. řádků)."""
    kody = [k for k in kody if k]
    if not kody:
        return {}
    conn = intranet_data.get_db_connection()
    if not conn:
        return {}
    master = {}
    try:
        cur = conn.cursor(dictionary=True)
        for i in range(0, len(kody), 1000):
            chunk = kody[i:i + 1000]
            ph = ",".join(["%s"] * len(chunk))
            cur.execute(f"SELECT kod, sortiment, sortiment_popis, nazev, nc, nc2, nc3, nc4, "
                        f"nc5, dnc, dph, id_kod, op FROM cenopripad_master WHERE kod IN ({ph})", chunk)
            for r in cur.fetchall():
                master[r["kod"]] = {
                    "nazev": r["nazev"], "sortiment_popis": r["sortiment_popis"],
                    "nc": r["nc"], "nc2": r["nc2"], "nc3": r["nc3"],
                    "nc4": r["nc4"], "nc5": r["nc5"], "dnc": r["dnc"], "dph": r["dph"],
                    "id": r["id_kod"], "sortiment": r["sortiment"], "op": r["op"]}
        cur.close()
        return master
    except Exception as e:
        print(f"Chyba načtení masteru pro kódy: {e}")
        return master
    finally:
        conn.close()


def vyhodnot_pripad_z_db(typ, radky):
    """Posbírá kódy řádků, načte k nim master z DB a vyhodnotí přes engine."""
    kody = set()
    for r in radky:
        nk = cp.normalizuj_kod(r.get("kod") or r.get("kod_produktu"))
        if nk:
            kody.add(nk)
            kody.add(nk.zfill(8))
    master = _nacti_master_pro_kody(kody)
    vys = cp.vyhodnot_pripad(typ, radky, master)
    # doplň název karty + sortiment popis z masteru (pro zobrazení a export)
    for radek, rr in zip(radky, vys["radky"]):
        nk = cp.normalizuj_kod(radek.get("kod") or radek.get("kod_produktu"))
        m = (master.get(nk) or master.get(nk.zfill(8))) if nk else None
        if m:
            rr["nazev_master"] = m.get("nazev")
            rr["sortiment_popis"] = m.get("sortiment_popis")
    return vys


def _nove_karty(radky, vys):
    """Kódy, které engine nenašel v masteru = nové karty bez cen (typicky nahrané
    dřív, než se v noci natáhne ceník). Vrací [(kód 8 míst, název)] v pořadí
    souboru, bez duplicit. Bez další DB dotazu — bere se z už spočítaného výsledku."""
    out = {}
    for radek, rr in zip(radky, vys.get("radky") or []):
        if (rr.get("duvod") or "") != cp.DUVOD_NENALEZEN:
            continue
        k = _export_kod(radek.get("kod") or radek.get("kod_produktu"))
        if k:
            out.setdefault(k, (radek.get("nazev") or "").strip())
    return list(out.items())


async def _potvrd_nove_karty(nove):
    """Dialog před založením případu: seznam kódů, které nejsou v databázi.
    Vrací True = přesto odeslat, False = zrušit."""
    # persistent: klik mimo/ESC dialog nezavře — jinak by `await d` nikdy neskončil
    # a odesílací tlačítka by zůstala zablokovaná.
    with ui.dialog().props("persistent") as d, ui.card().classes("w-[560px] max-w-full"):
        ui.label("⚠️ Některé kódy nejsou v databázi").classes("text-lg font-bold text-amber-700")
        ui.label(f"Těchto {len(nove)} položek se nepodařilo najít v ceníku — jde nejspíš "
                 "o nové karty, které se do systému natáhnou až v noci. U takových "
                 "řádků nelze spočítat marži a případ bude označen jako chybný.") \
            .classes("text-sm text-gray-700")
        with ui.scroll_area().classes("w-full max-h-64 border rounded bg-red-50"):
            for k, nz in nove:
                with ui.row().classes("items-center gap-2 no-wrap"):
                    ui.label(k).classes("font-mono font-bold text-red-600")
                    if nz:
                        ui.label(nz).classes("text-sm text-gray-600 truncate")
        ui.label("Doporučení: nahrajte případ zítra, až budou karty i s cenami v systému.") \
            .classes("text-sm font-medium text-gray-800")
        with ui.row().classes("w-full justify-end gap-2 mt-2"):
            ui.button("Zrušit a nahrát zítra", on_click=lambda: d.submit(False)) \
                .props("flat color=grey-8")
            ui.button("Přesto odeslat", on_click=lambda: d.submit(True)) \
                .props("color=amber-8")
    try:
        return bool(await d)
    finally:
        d.delete()


def _nacti_rows(raw_bytes, filename):
    """Načte 1. list libovolného formátu (.xls / .xlsx / .xlsm / .xlsb) jako seznam
    řádků (každý = seznam buněk). Formát se pozná především z OBSAHU (magic bytes),
    aby fungoval i když se z uploadu nepředá název/přípona."""
    name = (filename or "").lower()
    head = raw_bytes[:8] if raw_bytes else b""
    je_ole2 = head[:4] == b"\xd0\xcf\x11\xe0"   # starý binární .xls (OLE2)
    je_zip = head[:4] == b"PK\x03\x04"           # .xlsx/.xlsm/.xlsb (ZIP kontejner)

    if je_ole2 or (name.endswith(".xls") and not je_zip):
        try:
            import xlrd
        except ImportError:
            raise RuntimeError("Pro soubory .xls je na serveru potřeba knihovna xlrd "
                               "(pip install xlrd).")
        wb = xlrd.open_workbook(file_contents=raw_bytes)
        sh = wb.sheet_by_index(0)
        return [sh.row_values(r) for r in range(sh.nrows)]

    if name.endswith(".xlsb"):
        from pyxlsb import open_workbook
        rows = []
        with open_workbook(io.BytesIO(raw_bytes)) as wb:
            with wb.get_sheet(wb.sheets[0]) as sh:
                for row in sh.rows():
                    d = {c.c: c.v for c in row}
                    rows.append([d.get(i) for i in range((max(d) + 1) if d else 0)])
        return rows

    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    # Procentně formátované buňky (např. „-5 %") Excel ukládá jako číslo (-0,05).
    # Sjednotíme je s vkládáním přes schránku — vrátíme zobrazený text „-5%“,
    # aby parser i kontrola poznaly procenta (číslo bez „%" = koruny).
    out = []
    for row in ws.iter_rows():
        radek = []
        for c in row:
            v = c.value
            fmt = c.number_format or ""
            if isinstance(v, (int, float)) and not isinstance(v, bool) and "%" in fmt:
                v = f"{v * 100:g}%"
            radek.append(v)
        out.append(radek)
    return out


def _zpracuj_rows(rows, typ):
    """Najde řádek s hlavičkou (i pod „preamblem" s názvem akce/termíny), namapuje
    sloupce dle mapy typu a vrátí (radky, chyba|None).
    U mimolétáku bývají CC/MO/VO pod skupinami B99/K99/J99 (řádek nad hlavičkou)
    jako opakované „S DPH"/„bez DPH" → mapují se podle pořadí skupin."""
    cfg = TYPY[typ]
    if not rows:
        return None, "Soubor je prázdný."
    if typ == "ncfaktura":
        return _zpracuj_ncfaktura(rows)
    # 1) řádek hlavičky = první (max. do 40. řádku), jehož sloupce pokryjí povinná pole
    hdr_idx = None
    for i, row in enumerate(rows[:40]):
        mapped = {cfg["mapa"][_norm(c)] for c in row if _norm(c) in cfg["mapa"]}
        if cfg["povinne"] <= mapped:
            hdr_idx = i
            break
    if hdr_idx is None:
        return None, ("Nenašel jsem řádek s hlavičkou (chybí sloupce pro: "
                      + ", ".join(sorted(cfg["povinne"])) + ").")
    header = rows[hdr_idx]
    idx_pole = {i: cfg["mapa"][_norm(c)] for i, c in enumerate(header) if _norm(c) in cfg["mapa"]}
    # 2) mimoleták: CC/MO/VO ze skupinového řádku nad hlavičkou (B99/K99/J99),
    #    pokud je hlavička nemá pojmenované přímo
    if typ == "mimoletak" and hdr_idx > 0 and not (
            {"cc_s_dph", "mo_s_dph", "vo_bez_dph"} & set(idx_pole.values())):
        prah = max(idx_pole) if idx_pole else 0
        skupiny = rows[hdr_idx - 1]
        kanaly = [i for i, c in enumerate(skupiny) if str(c).strip() and i > prah]
        for pole, ci in zip(["cc_s_dph", "mo_s_dph", "vo_bez_dph"], kanaly):
            idx_pole[ci] = pole
    # 3) datové řádky pod hlavičkou
    radky = []
    spatne_pc = []          # (číslo řádku v Excelu, původní text) – jen typ porovnani
    for _j, r in enumerate(rows[hdr_idx + 1:]):
        if r is None or all(c is None or str(c).strip() == "" for c in r):
            continue
        radek = {pole: (r[i] if i < len(r) else None) for i, pole in idx_pole.items()}
        kodval = radek.get("kod") or radek.get("kod_produktu")
        if kodval is None or str(kodval).strip() == "":
            continue
        if typ == "porovnani":
            # PC bez DPH musí být cena v Kč. Procento ('5%', buňka s procentním
            # formátem) by engine přečetl jako 0,05 Kč a do skladu by nešlo nic.
            # Prázdné PC neřešíme – to hlásí až verdikt („Chybí PC bez DPH“).
            _pc = radek.get("pc_bez_dph")
            if str(_pc or "").strip() and cp.parse_cena_kc(_pc) is None:
                spatne_pc.append((hdr_idx + _j + 2, str(_pc).strip()))
        if typ in ("porovnani", "mimoletak"):
            # Vodicí nuly hned při nahrání (Excel je z čísel utrhne): kód vždy
            # 8 míst, IČ dle pravidla „8 číslic před případným písmenem".
            # V tomto tvaru se řádek uloží do vstup_json i do sloupce kod.
            for _k in ("kod", "kod_produktu"):
                if str(radek.get(_k) or "").strip():
                    radek[_k] = _export_kod(radek[_k])
            if str(radek.get("zakaznik_ico") or "").strip():
                radek["zakaznik_ico"] = _export_ico(radek["zakaznik_ico"])
        radky.append(radek)
    if spatne_pc:
        ukazka = "; ".join(f"řádek {n}: „{t}“" for n, t in spatne_pc[:10])
        zbytek = f" … a dalších {len(spatne_pc) - 10}" if len(spatne_pc) > 10 else ""
        return None, ("Sloupec „PC bez DPH“ musí obsahovat cenu v Kč (číslo), "
                      "ne procento ani text. Opravte a nahrajte znovu – "
                      f"{ukazka}{zbytek}.")
    if not radky:
        return None, "Žádné datové řádky (s vyplněným kódem)."
    return radky, None


def _ncf_subtyp(radky):
    """Podtyp NC faktury z řádků ('nc' / 'fc'). Bere z prvního řádku se značkou."""
    for r in radky or []:
        try:
            vst = json.loads(r["vstup_json"]) if isinstance(r, dict) and r.get("vstup_json") else (
                r if isinstance(r, dict) else {})
        except Exception:
            vst = {}
        s = vst.get("_subtyp")
        if s:
            return s
    return "nc"


def _zpracuj_ncfaktura(rows):
    """NC faktura — detekce podtypu formuláře z hlavičky + poziční mapování sloupců.
      • „Formulář NC": Kód výrobku, Název výrobku, Cena pro nastavení, datum od/do,
        název IND.ceny.
      • „Formulář NC a FC-IMPORT": Kód, Název, NC-IMPORT, datum od/do, FC-IMPORT,
        datum od/do.
    Vrací (radky, chyba|None). Každý řádek nese '_subtyp' ('nc'/'fc')."""
    hdr_idx = None
    for i, row in enumerate(rows[:40]):
        norm = {_norm(c) for c in row if c is not None and str(c).strip()}
        if "kód výrobku" in norm and (
                "cena pro nastavení" in norm or
                ("nc-import" in norm and "fc-import" in norm)):
            hdr_idx = i
            break
    if hdr_idx is None:
        return None, ("Nerozpoznán formulář NC faktury — hlavička musí obsahovat "
                      "„Kód výrobku“ a buď „Cena pro nastavení“, nebo „NC-IMPORT“ a „FC-IMPORT“.")
    header = [_norm(c) for c in rows[hdr_idx]]

    def col(name):
        return header.index(name) if name in header else None

    def col_after(name, offset):
        i = col(name)
        return (i + offset) if i is not None else None

    je_fc = "nc-import" in header and "fc-import" in header
    subtyp = "fc" if je_fc else "nc"
    if subtyp == "nc":
        idx = {"kod": col("kód výrobku"), "nazev": col("název výrobku"),
               "cena_nastaveni": col("cena pro nastavení"),
               "datum_od": col("datum od"), "datum_do": col("datum do"),
               "nazev_ind": col("název ind.ceny")}
    else:
        # pořadí sloupců: Kód, Název, NC-IMPORT, datum od, datum do, FC-IMPORT, datum od, datum do
        idx = {"kod": col("kód výrobku"), "nazev": col("název výrobku"),
               "nc_import": col("nc-import"),
               "nc_datum_od": col_after("nc-import", 1),
               "nc_datum_do": col_after("nc-import", 2),
               "fc_import": col("fc-import"),
               "fc_datum_od": col_after("fc-import", 1),
               "fc_datum_do": col_after("fc-import", 2)}
    ostatni = [pole for pole in idx if pole != "kod"]
    radky = []
    for r in rows[hdr_idx + 1:]:
        if r is None or all(c is None or str(c).strip() == "" for c in r):
            continue
        radek = {"_subtyp": subtyp}
        for pole, ci in idx.items():
            radek[pole] = (r[ci] if ci is not None and ci < len(r) else None)
        if radek.get("kod") is None or str(radek["kod"]).strip() == "":
            continue
        # Vynech legendu/poznámky pod tabulkou (mají text jen ve sloupci kódu, ostatní
        # buňky prázdné) — platný řádek musí mít vyplněnou aspoň jednu další hodnotu.
        if not any(radek.get(pole) is not None and str(radek.get(pole)).strip()
                   for pole in ostatni):
            continue
        radky.append(radek)
    if not radky:
        return None, "Žádné datové řádky (s vyplněným kódem výrobku)."
    return radky, None


def _parsuj_sablonu(raw_bytes, typ, filename=""):
    """Naparsuje nahraný soubor (.xls/.xlsx/.xlsm/.xlsb) dle hlaviček."""
    try:
        rows = _nacti_rows(raw_bytes, filename)
    except Exception as e:
        return None, ("Soubor nejde načíst — je to platný Excel (.xlsx/.xls)? "
                      f"(detail: {e})")
    try:
        return _zpracuj_rows(rows, typ)
    except Exception as e:
        return None, f"Soubor nelze zpracovat: {e}"


def _parsuj_paste(text, typ):
    """Naparsuje data vložená přes schránku (TSV z Excelu, vč. případného preamblu).
    Parsuje přes `csv` (delimiter=TAB) — tím správně zpracuje i buňky s vnořeným
    zalomením řádku (Alt+Enter), které Excel do schránky obaluje uvozovkami.
    Naivní split('\\n') by takovou buňku rozlomil a posunul následující sloupce."""
    import csv
    reader = csv.reader(io.StringIO(text), delimiter="\t")
    rows = [r for r in reader if any(str(c).strip() for c in r)]
    if len(rows) < 2:
        return None, "Vložte hlavičku a alespoň jeden datový řádek (kopie z Excelu)."
    return _zpracuj_rows(rows, typ)


def _dalsi_cislo(cur):
    rok = str(datetime.date.today().year)
    cur.execute("SELECT MAX(CAST(SUBSTRING(cislo,5) AS UNSIGNED)) FROM cenopripad_pripady "
                "WHERE cislo LIKE %s", (rok + "%",))
    mx = cur.fetchone()[0] or 0
    return f"{rok}{int(mx) + 1:05d}"


def _uloz_pripad(typ, nazev, zadavatel_id, zadavatel_jmeno, vstup_radky, vysledek,
                 soubor_raw=None, soubor_nazev=None, testovaci=False, poznamka_zadani=None,
                 stav_override=None):
    """Uloží případ + jeho řádky (+ volitelně původní nahraný soubor).
    `testovaci` = zkušební případ (neodesílají se e-maily, vizuálně fialový).
    `poznamka_zadani` = důvod nahrání (u mimolétáku povinné).
    `stav_override` = vynutí stav případu místo verdiktu (např. 'delisting' přeskočí
    fázi kontroly a jde rovnou na office).
    Vrací (cislo, stav, chyba|None)."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return None, None, "Chyba připojení k databázi."
    try:
        cur = conn.cursor()
        cislo = _dalsi_cislo(cur)
        stav = stav_override or ("vyhodnoceno_ok" if vysledek["ok"] else "vyhodnoceno_chyba")
        # Denormalizace pobočky: snapshot z aktuální pobočky autora (zadavatele) při vzniku případu
        cur.execute("SELECT pobocka FROM user WHERE iduser=%s", (zadavatel_id,))
        _pob_row = cur.fetchone()
        _pobocka = _pob_row[0] if _pob_row else None
        cur.execute(
            "INSERT INTO cenopripad_pripady "
            "(cislo, nazev, typ, oddeleni, zadavatel_id, zadavatel_jmeno, stav, "
            " pocet_radku, pocet_chyb, vysledek_ok, testovaci, poznamka_zadani, pobocka) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (cislo, nazev, typ, TYPY[typ]["oddeleni"], zadavatel_id, zadavatel_jmeno,
             stav, vysledek["pocet"], vysledek["chyby"], 1 if vysledek["ok"] else 0,
             1 if testovaci else 0, _str(poznamka_zadani, 1000), _pobocka))
        pid = cur.lastrowid
        davka = []
        for i, (inp, rr) in enumerate(zip(vstup_radky, vysledek["radky"]), 1):
            karta = rr.get("nazev_master") or inp.get("nazev") or inp.get("zakaznik") or ""
            kod = inp.get("kod") or inp.get("kod_produktu") or ""
            davka.append((pid, i, json.dumps(inp, ensure_ascii=False, default=str),
                          _str(karta, 255), _str(rr.get("sortiment_popis"), 255),
                          _str(str(kod), 40), rr["verdikt"],
                          _str(rr["duvod"], 255), rr.get("op"),
                          json.dumps(rr, ensure_ascii=False, default=str)))
        cur.executemany(
            "INSERT INTO cenopripad_radky "
            "(pripad_id, poradi, vstup_json, nazev_karty, sortiment_popis, kod, verdikt, "
            "duvod, op, marze_json) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", davka)
        if soubor_raw:
            cur.execute("INSERT INTO cenopripad_soubory (pripad_id, nazev, data) "
                        "VALUES (%s, %s, %s)", (pid, _str(soubor_nazev, 255), soubor_raw))
        conn.commit()
        cur.close()
        zaznam_historie(
            pid, "Založeno", zadavatel_jmeno,
            ("DELISTING — přeskočena fáze kontroly, předáno rovnou office"
             if stav_override == "delisting"
             else ("Vyhodnoceno: vše v pořádku" if vysledek["ok"]
                   else f"Vyhodnoceno: NENÍ v pořádku — {vysledek['chyby']} z {vysledek['pocet']} řádků"))
            + (" · TESTOVACÍ případ" if testovaci else "")
            + (f" · Důvod nahrání: {poznamka_zadani}" if poznamka_zadani else ""))
        return cislo, stav, None
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return None, None, f"Chyba uložení případu: {e}"
    finally:
        conn.close()


def nacti_pripady(typ, prava, user_id, filtr_pobocka=None):
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        # Volitelný filtr pobočky — pouze zužuje (protíná) již viditelný seznam, nikdy nerozšiřuje.
        _pob = f" AND pobocka=%s" if filtr_pobocka else ""
        if _vidi_vsechny_pripady(prava):
            if filtr_pobocka:
                cur.execute(f"SELECT * FROM cenopripad_pripady WHERE typ=%s{_pob} ORDER BY id DESC",
                            (typ, filtr_pobocka))
            else:
                cur.execute("SELECT * FROM cenopripad_pripady WHERE typ=%s ORDER BY id DESC", (typ,))
        else:
            # Bez plné viditelnosti: vždy vlastní případy, navíc případy spolučlenů
            # oddělení (právo „Zobrazení oddělení"), žadatelů z oddělení, která smí
            # uživatel schvalovat (právo „Schvalovatel – oddělení"), podřízených
            # z oddělení, kde je hlavní vedoucí (právo „Hlavní vedoucí" — jen čtení),
            # a přímých podřízených dle „Přímí nadřízení" (asistentka → nadřízený,
            # plošně pro všechny typy případů — jen čtení).
            ids = {user_id}
            if _vidi_oddeleni(prava):
                ids |= set(_kolegove_oddeleni(user_id))
            ids |= set(_zadavatele_ke_schvaleni(prava))
            ids |= set(_zadavatele_vedeni(prava))
            ids |= set(_podrizeni(user_id))
            ph = ",".join(["%s"] * len(ids))
            _params = [typ, *ids]
            if filtr_pobocka:
                _params.append(filtr_pobocka)
            cur.execute(f"SELECT * FROM cenopripad_pripady WHERE typ=%s "
                        f"AND zadavatel_id IN ({ph}){_pob} ORDER BY id DESC", tuple(_params))
        r = cur.fetchall()
        cur.close()
        return r
    except Exception as e:
        print(f"Chyba načtení případů: {e}")
        return []
    finally:
        conn.close()


def nacti_radky(pripad_id):
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT * FROM cenopripad_radky WHERE pripad_id=%s ORDER BY poradi", (pripad_id,))
        r = cur.fetchall()
        cur.close()
        return r
    except Exception:
        return []
    finally:
        conn.close()


def _ma_soubor(pripad_id):
    conn = intranet_data.get_db_connection()
    if not conn:
        return False
    try:
        cur = conn.cursor()
        cur.execute("SELECT 1 FROM cenopripad_soubory WHERE pripad_id=%s LIMIT 1", (pripad_id,))
        je = cur.fetchone() is not None
        cur.close()
        return je
    except Exception:
        return False
    finally:
        conn.close()


def _stahovaci_nazev(nazev, data, pripad_id):
    """Název ke stažení se zaručenou tabulkovou příponou. Řeší prázdný/uříznutý
    název (starší případy uložené bez názvu kvůli NiceGUI 3.x e.file.name) i klamnou
    tečku uvnitř názvu — příponu doplní z obsahu (magic bytes: OLE2 → .xls, jinak
    ZIP kontejner → .xlsx). Rozumnou tabulkovou příponu respektuje (.xlsm/.xlsb/.xls)."""
    base = (nazev or "").strip() or f"cenopripad_{pripad_id}"
    if base.lower().endswith((".xlsx", ".xlsm", ".xlsb", ".xls")):
        return base
    head = bytes(data[:4]) if data else b""
    return base + (".xls" if head == b"\xd0\xcf\x11\xe0" else ".xlsx")


def nacti_soubor(pripad_id):
    """Vrátí (nazev, bytes) původního nahraného souboru případu, nebo None."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return None
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT nazev, data FROM cenopripad_soubory WHERE pripad_id=%s", (pripad_id,))
        r = cur.fetchone()
        cur.close()
        if not r or r["data"] is None:
            return None
        data = r["data"]
        if isinstance(data, (bytearray, memoryview)):
            data = bytes(data)
        return (_stahovaci_nazev(r["nazev"], data, pripad_id), data)
    except Exception as e:
        print(f"Chyba načtení souboru případu: {e}")
        return None
    finally:
        conn.close()


def uloz_prilohu(pripad_id, nazev, typ, data, kdo):
    """Uloží jednu přílohu/foto k případu. Vrací (ok, chyba|None)."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return False, "Chyba připojení k databázi."
    try:
        cur = conn.cursor()
        cur.execute("INSERT INTO cenopripad_prilohy (pripad_id, nazev, typ, data, kdo) "
                    "VALUES (%s, %s, %s, %s, %s)",
                    (pripad_id, _str(nazev, 255), _str(typ, 120), data, _str(kdo, 255)))
        conn.commit()
        cur.close()
        return True, None
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return False, f"Chyba uložení přílohy: {e}"
    finally:
        conn.close()


def nacti_prilohy(pripad_id):
    """Seznam příloh případu BEZ binárky (id, nazev, typ, kdo, kdy, velikost)."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT id, nazev, typ, kdo, kdy, LENGTH(data) AS velikost "
                    "FROM cenopripad_prilohy WHERE pripad_id=%s ORDER BY id", (pripad_id,))
        r = cur.fetchall()
        cur.close()
        return r
    except Exception:
        return []
    finally:
        conn.close()


def nacti_prilohu(priloha_id):
    """Vrátí (nazev, typ, bytes) jedné přílohy ke stažení/zobrazení, nebo None."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return None
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT nazev, typ, data FROM cenopripad_prilohy WHERE id=%s", (priloha_id,))
        r = cur.fetchone()
        cur.close()
        if not r or r["data"] is None:
            return None
        data = r["data"]
        if isinstance(data, (bytearray, memoryview)):
            data = bytes(data)
        return (r["nazev"] or f"priloha_{priloha_id}", r["typ"] or "", data)
    except Exception as e:
        print(f"Chyba načtení přílohy: {e}")
        return None
    finally:
        conn.close()


def smaz_prilohu(priloha_id):
    """Smaže jednu přílohu (správce nebo vlastník případu). Vrací (ok, chyba|None)."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return False, "Chyba připojení k databázi."
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM cenopripad_prilohy WHERE id=%s", (priloha_id,))
        conn.commit()
        cur.close()
        return True, None
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return False, f"Chyba mazání přílohy: {e}"
    finally:
        conn.close()


def zaznam_historie(pripad_id, akce, kdo, detail=None):
    """Zapíše JEDEN krok průběhu případu (best-effort, „očičko")."""
    if not pripad_id:
        return
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.execute("INSERT INTO cenopripad_historie (pripad_id, akce, detail, kdo) "
                    "VALUES (%s,%s,%s,%s)",
                    (pripad_id, _str(akce, 80), _str(detail, 1000), _str(kdo, 255) or ""))
        conn.commit()
        cur.close()
    except Exception as e:
        print(f"[cenopripad] zaznam_historie: {e}")
    finally:
        conn.close()


def nacti_historie(pripad_id):
    """Průběh případu chronologicky (od zadání po finální krok)."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT akce, detail, kdo, kdy FROM cenopripad_historie "
                    "WHERE pripad_id=%s ORDER BY id", (pripad_id,))
        r = cur.fetchall()
        cur.close()
        return r
    except Exception:
        return []
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Workflow (stavy) + e-maily (Fáze 4)
# ---------------------------------------------------------------------------
def nacti_pripad(pripad_id):
    conn = intranet_data.get_db_connection()
    if not conn:
        return None
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT * FROM cenopripad_pripady WHERE id=%s", (pripad_id,))
        r = cur.fetchone()
        cur.close()
        return r
    except Exception:
        return None
    finally:
        conn.close()


def zmen_stav(pripad_id, novy_stav):
    conn = intranet_data.get_db_connection()
    if not conn:
        return False
    try:
        cur = conn.cursor()
        cur.execute("UPDATE cenopripad_pripady SET stav=%s WHERE id=%s", (novy_stav, pripad_id))
        conn.commit()
        cur.close()
        return True
    except Exception as e:
        print(f"Chyba změny stavu případu: {e}")
        return False
    finally:
        conn.close()


def schval_pripad(pripad_id, vyrazene_poradi=None):
    """Schválí případ; `vyrazene_poradi` = poradí řádků, které správce NESCHVÁLIL.
    Vyřazené → `neschvaleno=1` (zašedlé, mimo export); ostatní → 0. Stav =
    'castecne_schvaleno' když je něco vyřazeno, jinak 'schvaleno'.
    Vrací (novy_stav, ok, chyba|None)."""
    vyr = [int(x) for x in (vyrazene_poradi or [])]
    conn = intranet_data.get_db_connection()
    if not conn:
        return None, False, "Chyba připojení k databázi."
    try:
        cur = conn.cursor()
        cur.execute("UPDATE cenopripad_radky SET neschvaleno=0 WHERE pripad_id=%s", (pripad_id,))
        if vyr:
            ph = ",".join(["%s"] * len(vyr))
            cur.execute(f"UPDATE cenopripad_radky SET neschvaleno=1 "
                        f"WHERE pripad_id=%s AND poradi IN ({ph})", [pripad_id] + vyr)
        novy_stav = "castecne_schvaleno" if vyr else "schvaleno"
        cur.execute("UPDATE cenopripad_pripady SET stav=%s WHERE id=%s", (novy_stav, pripad_id))
        conn.commit()
        cur.close()
        return novy_stav, True, None
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return None, False, f"Chyba schválení případu: {e}"
    finally:
        conn.close()


def oprav_pripad(pripad_id, typ, vstup_radky, vysledek, soubor_raw=None, soubor_nazev=None):
    """Přepíše řádky případu novým vyhodnocením (+ volitelně nahraný soubor).
    Vrací (stav, chyba|None)."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return None, "Chyba připojení k databázi."
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM cenopripad_radky WHERE pripad_id=%s", (pripad_id,))
        davka = []
        for i, (inp, rr) in enumerate(zip(vstup_radky, vysledek["radky"]), 1):
            karta = rr.get("nazev_master") or inp.get("nazev") or inp.get("zakaznik") or ""
            kod = inp.get("kod") or inp.get("kod_produktu") or ""
            davka.append((pripad_id, i, json.dumps(inp, ensure_ascii=False, default=str),
                          _str(karta, 255), _str(rr.get("sortiment_popis"), 255),
                          _str(str(kod), 40), rr["verdikt"],
                          _str(rr["duvod"], 255), rr.get("op"),
                          json.dumps(rr, ensure_ascii=False, default=str)))
        cur.executemany(
            "INSERT INTO cenopripad_radky "
            "(pripad_id, poradi, vstup_json, nazev_karty, sortiment_popis, kod, verdikt, "
            "duvod, op, marze_json) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", davka)
        stav = "vyhodnoceno_ok" if vysledek["ok"] else "vyhodnoceno_chyba"
        cur.execute("UPDATE cenopripad_pripady SET stav=%s, pocet_radku=%s, pocet_chyb=%s, "
                    "vysledek_ok=%s WHERE id=%s",
                    (stav, vysledek["pocet"], vysledek["chyby"],
                     1 if vysledek["ok"] else 0, pripad_id))
        if soubor_raw:
            cur.execute("INSERT INTO cenopripad_soubory (pripad_id, nazev, data) "
                        "VALUES (%s, %s, %s) ON DUPLICATE KEY UPDATE "
                        "nazev=VALUES(nazev), data=VALUES(data)",
                        (pripad_id, _str(soubor_nazev, 255), soubor_raw))
        conn.commit()
        cur.close()
        return stav, None
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return None, f"Chyba opravy případu: {e}"
    finally:
        conn.close()


def prepocti_pripad(pripad_id, typ):
    """Znovu vyhodnotí ULOŽENÉ řádky případu proti AKTUÁLNÍMU masteru (po re-importu dat),
    aniž by se musela data nahrávat znovu. Bere vstup_json každého řádku → vyhodnotí
    engine s aktuálním masterem (OP ze sloupce AJ, sortiment popis…) → přepíše řádky.
    Nahraný soubor zůstává. Vrací (stav, chyba|None)."""
    radky_db = nacti_radky(pripad_id)
    if not radky_db:
        return None, "Případ nemá žádné řádky k přepočtu."
    vstup_radky = []
    for r in radky_db:
        try:
            vstup_radky.append(json.loads(r["vstup_json"]) if r["vstup_json"] else {})
        except Exception:
            vstup_radky.append({})
    vys = vyhodnot_pripad_z_db(typ, vstup_radky)
    return oprav_pripad(pripad_id, typ, vstup_radky, vys)


def zadej_druhou_kontrolu(pripad_id, poznamka):
    """Žadatel žádá o 2. kontrolu: stav → 'ceka_na_spravce' + uloží poznámku. Vrací (ok, chyba)."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return False, "Chyba připojení k databázi."
    try:
        cur = conn.cursor()
        cur.execute("UPDATE cenopripad_pripady SET stav='ceka_na_spravce', poznamka=%s "
                    "WHERE id=%s", (_str(poznamka, 1000) or None, pripad_id))
        conn.commit()
        cur.close()
        return True, None
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return False, f"Chyba: {e}"
    finally:
        conn.close()


def zamitni_pripad(pripad_id, duvod):
    """Zamítne případ správcem: stav → 'zamitnuto' + uloží důvod zamítnutí
    (uvidí ho žadatel). Vrací (ok, chyba|None)."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return False, "Chyba připojení k databázi."
    try:
        cur = conn.cursor()
        cur.execute("UPDATE cenopripad_pripady SET stav='zamitnuto', zamitnuti_duvod=%s "
                    "WHERE id=%s", (_str(duvod, 1000) or None, pripad_id))
        conn.commit()
        cur.close()
        return True, None
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return False, f"Chyba zamítnutí: {e}"
    finally:
        conn.close()


def stornuj_pripad(pripad_id, duvod):
    """Stornuje případ: stav → 'stornovano' + uloží důvod. Vrací (ok, chyba|None)."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return False, "Chyba připojení k databázi."
    try:
        cur = conn.cursor()
        cur.execute("UPDATE cenopripad_pripady SET stav='stornovano', storno_duvod=%s "
                    "WHERE id=%s", (_str(duvod, 500), pripad_id))
        conn.commit()
        cur.close()
        return True, None
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return False, f"Chyba storna: {e}"
    finally:
        conn.close()


def prevod_na_ostry(pripad_id):
    """Převede testovací případ na ostrý (testovaci → 0). Stav i řádky zůstávají;
    případ pak pokračuje běžným procesem (notifikace řeší volající). Vrací (ok, chyba|None)."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return False, "Chyba připojení k databázi."
    try:
        cur = conn.cursor()
        cur.execute("UPDATE cenopripad_pripady SET testovaci=0 WHERE id=%s", (pripad_id,))
        conn.commit()
        cur.close()
        return True, None
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return False, f"Chyba převodu na ostrý případ: {e}"
    finally:
        conn.close()


def smaz_pripad(pripad_id):
    """Nevratně smaže případ + jeho řádky + nahraný soubor. Vrací (ok, chyba|None)."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return False, "Chyba připojení k databázi."
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM cenopripad_radky WHERE pripad_id=%s", (pripad_id,))
        cur.execute("DELETE FROM cenopripad_soubory WHERE pripad_id=%s", (pripad_id,))
        cur.execute("DELETE FROM cenopripad_prilohy WHERE pripad_id=%s", (pripad_id,))
        cur.execute("DELETE FROM cenopripad_historie WHERE pripad_id=%s", (pripad_id,))
        cur.execute("DELETE FROM cenopripad_pripady WHERE id=%s", (pripad_id,))
        conn.commit()
        cur.close()
        return True, None
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return False, f"Chyba mazání případu: {e}"
    finally:
        conn.close()


def _emaily_s_pravy(*role_keys):
    """E-maily aktivních uživatelů s některým z `role_keys` (NEBO 'vse'), z EFEKTIVNÍCH
    práv (osobní + pozice + ZDĚDĚNÁ z oddělení) — viz `intranet_data.ziskej_emaily_s_pravem`.
    NE `ziskej_vsechny_uzivatele()['prava']` (jen osobní → zděděné role by maily neměly)."""
    klice = [k for k in role_keys if k] + ["vse"]
    return list(dict.fromkeys(intranet_data.ziskej_emaily_s_pravem(*klice)))


def _emaily_office(oddeleni):
    return _emaily_s_pravy("cenopripad_office_nakup" if oddeleni == "nakup"
                           else "cenopripad_office_obchod")


def _emaily_spravce(oddeleni):
    if oddeleni == "nakup":
        return _emaily_s_pravy("cenopripad_spravce", "cenopripad_spravce_nakup")
    return _emaily_s_pravy("cenopripad_spravce")


def _app_url():
    try:
        u = (intranet_data.nacti_nastaveni_intranetu().get("app_url", "") or "").strip().rstrip("/")
        return f"{u}/cenopripad" if u else ""
    except Exception:
        return ""


def _posli_emaily_sync(prijemci, predmet, text):
    odkaz = _app_url()
    if odkaz:
        text = f"{text}\n\nOtevřít v portálu: {odkaz}"
    for p in prijemci:
        try:
            intranet_emaily.odesli_upozorneni_email(p, predmet, text)
        except Exception as e:
            print(f"[cenopripad] e-mail {p}: {e}")


def _odesli_emaily(prijemci, predmet, text):
    prijemci = [p for p in dict.fromkeys(prijemci) if p and "@" in p]
    if not prijemci:
        return
    try:
        asyncio.create_task(asyncio.to_thread(_posli_emaily_sync, prijemci, predmet, text))
    except RuntimeError:
        _posli_emaily_sync(prijemci, predmet, text)


def _email_uzivatele(uid):
    """E-mail uživatele dle user.iduser (pro notifikaci žadatele), nebo None."""
    if not uid:
        return None
    conn = intranet_data.get_db_connection()
    if not conn:
        return None
    try:
        cur = conn.cursor()
        cur.execute("SELECT email FROM user WHERE iduser=%s", (uid,))
        r = cur.fetchone()
        cur.close()
        return r[0] if r and r[0] and "@" in r[0] else None
    except Exception as e:
        print(f"[cenopripad] _email_uzivatele: {e}")
        return None
    finally:
        conn.close()


def _odesli_emaily_zadateli(pripad, predmet, text):
    """Notifikace ŽADATELI (na e-mail z user.iduser = zadavatel_id)."""
    em = _email_uzivatele(pripad.get("zadavatel_id"))
    if em:
        _odesli_emaily([em], predmet, text)


def _je_spravce_typu(typ, p):
    if _je_spravce(p):
        return True
    return TYPY[typ]["oddeleni"] == "nakup" and "cenopripad_spravce_nakup" in p


def _je_office_typu(typ, p):
    if _je_spravce(p):
        return True
    odd = TYPY[typ]["oddeleni"]
    return ("cenopripad_office_nakup" if odd == "nakup" else "cenopripad_office_obchod") in p


# ============================================================================
# UI
# ============================================================================
async def _precti_upload(e, up):
    zdroj = None
    for attr in ("content", "file", "stream", "data", "file_obj"):
        val = getattr(e, attr, None)
        if val is not None and hasattr(val, "read"):
            zdroj = val
            break
    if zdroj is None:
        ui.notify("Nepodařilo se načíst obsah souboru.", type="negative")
        up.reset()
        return None, None
    try:
        raw = zdroj.read()
        if inspect.isawaitable(raw):
            raw = await raw
    except Exception as exc:
        ui.notify(f"Chyba čtení souboru: {exc}", type="negative")
        up.reset()
        return None, None
    # NiceGUI 3.x má jméno na e.file.name (UploadEventArguments.file: FileUpload),
    # starší verze na e.name. Bez tohoto zůstane název prázdný → uložený soubor se
    # pak stahuje BEZ PŘÍPONY (fallback „cenopripad_<id>").
    name = (getattr(getattr(e, "file", None), "name", None)
            or getattr(e, "name", None) or "")
    up.reset()
    return raw, name


def _refresh():
    """Přerenderuje modul JEN pro aktuálního klienta.

    `vykresli_cenopripad` je per-klient refreshable uložený v `app.storage.client`
    (viz tam). Nelze použít modulový `@ui.refreshable`+`.refresh()`: ten by
    přerenderoval VŠECHNY připojené klienty, a to navíc v kontextu uživatele, který
    akci spustil (request_contextvar se v refresh smyčce nepřepíná), takže by se všem
    modul přepnul na sekci jednoho uživatele (jeho `app.storage.user`)."""
    fn = app.storage.client.get("_cenopripad_refresh")
    if fn:
        fn()


def _nav(kam):
    app.storage.user["cenopripad_pohled"] = kam
    _refresh()


def _trunc2(x):
    """Osekne na 2 desetinná místa směrem k nule (NEzaokrouhluje)."""
    return Decimal(str(x)).quantize(Decimal("0.01"), rounding=ROUND_DOWN)


def _pct(v):
    """Procenta na 2 desetinná místa, oseknutá (bez zaokrouhlení)."""
    return "—" if v is None else f"{_trunc2(Decimal(str(v)) * 100):.2f}".replace(".", ",") + " %"


def _castka(v):
    """Částka na 2 desetinná místa, oseknutá (bez zaokrouhlení)."""
    return "—" if v is None else f"{_trunc2(v):.2f}".replace(".", ",")


def _marze_text(rr):
    casti = []
    if rr.get("marze_netto") is not None:
        casti.append(f"netto {_pct(rr['marze_netto'])}")
    if rr.get("marze_anc") is not None:
        casti.append(f"ANC {_pct(rr['marze_anc'])}")
    if rr.get("marze_nc2") is not None:
        casti.append(f"NC2 {_pct(rr['marze_nc2'])}")
    if rr.get("soucet") is not None:
        casti.append(f"Σ {_pct(rr['soucet'])}")
    for k, lbl in (("marze_cc", "CC"), ("marze_mo", "MO"), ("marze_vo", "VO")):
        if rr.get(k) is not None:
            casti.append(f"{lbl} {_pct(rr[k])}")
    if rr.get("nck") is not None:
        casti.append(f"NCK {_castka(rr['nck'])}")
    if rr.get("vstup_op") is not None:
        casti.append(f"vstup OP {_pct(rr['vstup_op'])}")
    return ", ".join(casti) or "—"


_VERDIKT_TXT = {"OK": "✅ OK", "CHYBA": "❌ Chyba", "VYJIMKA": "⚠️ Výjimka"}

# Popisky vstupních polí (zadané ceny) — vidí je žadatel i office.
POLE_LABEL = {
    "kod": "Kód", "kod_produktu": "Kód produktu", "nazev": "Název",
    "akcni_pc": "Akční PC", "akcni_nc": "Akční NC", "aktualni_nc2": "Aktuální NC2",
    "pc_akce": "PC-akce", "anc": "ANC", "nc2": "NC2",
    "koef": "koef", "dph_in": "DPH",
    "akcni_fc": "Akční FC", "cc_s_dph": "CC s DPH", "mo_s_dph": "MO s DPH",
    "vo_bez_dph": "VO bez DPH", "pc_bez_dph": "PC bez DPH", "bonus": "Bonus %",
    "kompenzace": "Kompenzace", "zakaznik": "Zákazník", "op": "OP (nabídka)",
    "definice": "Definice nastavení", "vychozi_cena": "Výchozí cena",
    "prirazka": "Přirážka/Srážka",
    "zakaznik_ico": "Zákazník IČO", "termin_od": "Termín od", "termin_do": "Termín do",
    "typy_dokladu": "Typy dokladů", "pokracovat_nizsi": "Pokračovat v hledání nižší ceny",
    # NC faktura
    "cena_nastaveni": "Cena pro nastavení", "nazev_ind": "Název IND. ceny",
    "nc_import": "NC-IMPORT", "fc_import": "FC-IMPORT",
    "datum_od": "Datum od", "datum_do": "Datum do",
    "nc_datum_od": "NC datum od", "nc_datum_do": "NC datum do",
    "fc_datum_od": "FC datum od", "fc_datum_do": "FC datum do",
}


def _varovani_paste():
    """Jednorázové upozornění při vložení dat přes schránku (Ctrl+V) místo importu
    souboru — u velkých objemů může TSV z Excelu vést k nepřesnému výpočtu."""
    with ui.dialog() as d, ui.card().classes("p-5 rounded-2xl shadow-xl") \
            .style("max-width: 460px"):
        with ui.row().classes("items-center gap-3"):
            ui.icon("warning", color="amber-7").classes("text-3xl")
            ui.label("Upozornění k vložení dat").classes("text-lg font-bold text-gray-800")
        ui.label("Pokud jsou data velká, tento způsob vložení nemusí provést správný "
                 "výpočet, doporučujeme pro tyto případy import souboru.") \
            .classes("text-sm text-gray-700 mt-1")
        with ui.row().classes("w-full justify-end mt-3"):
            ui.button("Beru na vědomí", icon="check", on_click=d.close) \
                .props("unelevated no-caps") \
                .classes("bg-amber-600 hover:bg-amber-700 text-white rounded-lg px-4")
    d.open()


def _hlidej_paste(textarea):
    """Na vložení do textarey (Ctrl+V) ukáže JEDNOU za daný formulář varování."""
    stav = {"varovano": False}

    def _on_paste(_=None):
        if stav["varovano"]:
            return
        stav["varovano"] = True
        _varovani_paste()

    textarea.on("paste", _on_paste)


def _oprava_widget(pripad, typ, user_name, po_akci):
    """Re-upload/paste + přepočítání existujícího případu (žadatel i správce)."""
    drzeny = {"raw": None, "name": ""}

    async def _on_up(e):
        raw, name = await _precti_upload(e, up)
        if raw is not None:
            drzeny["raw"], drzeny["name"] = raw, name
            ui.notify("Soubor připraven k přepočtu.", type="info")

    up = ui.upload(on_upload=_on_up, auto_upload=True, max_file_size=20_000_000,
                   label="Vybrat soubor").props("accept=.xlsx,.xls,.xlsm").classes("w-full")
    paste_in = ui.textarea("…nebo vložte řádky z Excelu (vč. hlavičky)") \
        .props("outlined").classes("w-full")
    _hlidej_paste(paste_in)

    async def _prepocti():
        try:
            text = (paste_in.value or "").strip()
            if text:
                radky, err = await asyncio.to_thread(_parsuj_paste, text, typ)
            elif drzeny["raw"]:
                radky, err = await asyncio.to_thread(
                    _parsuj_sablonu, drzeny["raw"], typ, drzeny["name"])
            else:
                ui.notify("Nahrajte soubor nebo vložte data.", type="warning")
                return
            if err:
                ui.notify(err, type="negative", timeout=8000)
                return
            vys = await asyncio.to_thread(vyhodnot_pripad_z_db, typ, radky)
            soubor_raw = None if text else drzeny["raw"]
            _, err2 = await asyncio.to_thread(oprav_pripad, pripad["id"], typ, radky, vys,
                                              soubor_raw, drzeny["name"])
            if err2:
                ui.notify(err2, type="negative", timeout=8000)
                return
            intranet_logger.log_activity(user_name, "Cenopřípad",
                                         f"Oprava {pripad['cislo']} → {'ok' if vys['ok'] else 'chyba'}")
            zaznam_historie(
                pripad["id"], "Opraveno a přepočítáno", user_name,
                ("Vše v pořádku" if vys["ok"]
                 else f"Stále NENÍ v pořádku — {vys['chyby']} z {vys['pocet']} řádků"))
            if vys["ok"] and not pripad.get("testovaci"):   # testovací NEodesílá e-maily
                _kontr = _kontrola_z_vysledku(typ, radky, vys)
                _odesli_emaily(
                    _emaily_office(TYPY[typ]["oddeleni"]),
                    f"Cenopřípad {pripad['cislo']}: případ je v pořádku ke zpracování",
                    f"Případ „{pripad['nazev']}“ ({TYPY[typ]['nazev']}) byl opraven a je v pořádku. "
                    f"Můžete ho zpracovat a dotáhnout do nastavení."
                    + (_KONTROLA_VAROVANI_MAIL if _kontr == "chyba" else ""))
            zprava = ("✅ Přepočítáno — vše v pořádku." if vys["ok"]
                      else f"❌ Přepočítáno — {vys['chyby']} z {vys['pocet']} řádků špatně.")
            await po_akci(zprava, "positive" if vys["ok"] else "negative")
        except Exception as e:
            ui.notify(f"Neočekávaná chyba při přepočtu: {e}", type="negative", timeout=9000)

    ui.button("Přepočítat", icon="calculate", on_click=_prepocti) \
        .props("unelevated no-caps").classes("bg-emerald-600 text-white rounded-lg px-4 mt-1")


async def _stahni_prilohu(priloha_id, klient):
    res = await asyncio.to_thread(nacti_prilohu, priloha_id)
    try:
        with klient:   # po awaitu mohl být slot dialogu smazán
            if res:
                ui.download.content(res[2], res[0], res[1] or "application/octet-stream")
            else:
                ui.notify("Příloha nenalezena.", type="warning")
    except Exception:
        pass


async def _zobraz_prilohu(priloha_id, klient):
    res = await asyncio.to_thread(nacti_prilohu, priloha_id)
    if not res:
        return
    nazev, typ, data = res
    uri = f"data:{typ or 'image/jpeg'};base64,{base64.b64encode(data).decode('ascii')}"
    try:
        with klient:
            with ui.dialog() as d, ui.card().classes("p-2").style("max-width: 92vw"):
                ui.label(nazev).classes("text-sm font-medium px-1")
                ui.image(uri).classes("rounded").style("max-height: 80vh; width: auto")
                with ui.row().classes("w-full justify-end"):
                    ui.button("Zavřít", on_click=d.close).props("flat no-caps")
            d.open()
    except Exception:
        pass


def _sekce_prilohy(pripad_id, klient, muze_mazat=False):
    """Seznam příloh/fotek případu (vidí každý, kdo vidí případ) — zobrazit/stáhnout;
    mazat smí jen vlastník případu nebo správce daného typu (`muze_mazat`)."""
    @ui.refreshable
    def _obsah():
        pril = nacti_prilohy(pripad_id)
        if not pril:
            return
        with ui.column().classes("w-full bg-amber-50 border border-amber-200 rounded-lg "
                                 "p-3 mb-2 gap-1"):
            ui.label(f"Přílohy / fotky ({len(pril)})") \
                .classes("text-sm font-semibold text-amber-900")
            for p in pril:
                je_obr = (p.get("typ") or "").startswith("image/")
                with ui.row().classes("w-full items-center gap-2 no-wrap text-sm"):
                    ui.icon("image" if je_obr else "attach_file", size="1.1rem") \
                        .classes("text-amber-700")
                    ui.label(p["nazev"]).classes("truncate text-gray-800")
                    ui.label(f"· {p.get('kdo') or ''} · {_dt_cz(p.get('kdy'))}") \
                        .classes("text-gray-400 whitespace-nowrap")
                    ui.space()
                    if je_obr:
                        ui.button("Zobrazit", icon="visibility",
                                  on_click=lambda pid=p["id"]: _zobraz_prilohu(pid, klient)) \
                            .props("flat dense no-caps size=sm").classes("text-amber-800")
                    ui.button("Stáhnout", icon="download",
                              on_click=lambda pid=p["id"]: _stahni_prilohu(pid, klient)) \
                        .props("flat dense no-caps size=sm").classes("text-amber-800")
                    if muze_mazat:
                        ui.button(icon="delete",
                                  on_click=lambda pid=p["id"], nm=p["nazev"]: _potvrd_smaz(pid, nm)) \
                            .props("flat dense round size=sm").classes("text-red-500") \
                            .tooltip("Smazat přílohu")

    def _potvrd_smaz(priloha_id, nazev):
        with ui.dialog() as cd, ui.card().classes("p-4"):
            ui.label(f"Smazat přílohu „{nazev}“?").classes("font-medium")
            ui.label("Tuto akci nelze vrátit.").classes("text-sm text-gray-500")

            async def _ano():
                ok, err = await asyncio.to_thread(smaz_prilohu, priloha_id)
                cd.close()
                try:
                    with klient:
                        if ok:
                            _obsah.refresh()
                            ui.notify("Příloha smazána.", type="positive")
                        else:
                            ui.notify(err or "Smazání selhalo.", type="negative")
                except Exception:
                    pass

            with ui.row().classes("w-full justify-end gap-2 mt-2"):
                ui.button("Zrušit", on_click=cd.close).props("flat no-caps")
                ui.button("Smazat", icon="delete_forever", on_click=_ano) \
                    .props("unelevated no-caps").classes("bg-red-600 text-white")
        cd.open()

    _obsah()


_HISTORIE_IKONA = {
    "Založeno": ("upload_file", "blue"),
    "Žádost o druhou kontrolu": ("how_to_reg", "amber"),
    "Opraveno a přepočítáno": ("autorenew", "teal"),
    "Přepočítáno dle masteru": ("sync", "teal"),
    "Schváleno správcem": ("verified", "green"),
    "Částečně schváleno": ("rule", "teal"),
    "Zamítnuto správcem": ("block", "red"),
    "Zpracováno (office)": ("task_alt", "blue"),
    "Převedeno na ostrý případ": ("published_with_changes", "purple"),
    "Stornováno": ("cancel", "grey"),
}


def _dialog_historie(pripad):
    """„Očičko" — chronologický průběh případu od zadání po finální krok."""
    zaznamy = nacti_historie(pripad["id"])
    with ui.dialog() as dlg, ui.card().classes("p-5 gap-2") \
            .style("min-width: 540px; max-width: 760px; max-height: 82vh; overflow-y: auto"):
        with ui.row().classes("items-center gap-2"):
            ui.icon("history", color="indigo")
            ui.label(f"Průběh případu {pripad['cislo']}").classes("text-lg font-bold text-gray-800")
            ui.label(pripad.get("nazev") or "").classes("text-gray-500")
        if not zaznamy:
            ui.label("U tohoto případu zatím není žádný záznam průběhu "
                     "(starší případy z doby před touto funkcí historii nemají).") \
                .classes("text-sm text-gray-500 italic py-4")
        else:
            for r in zaznamy:
                ikona, bv = _HISTORIE_IKONA.get(r.get("akce"), ("fiber_manual_record", "grey"))
                with ui.row().classes("w-full items-start gap-3 border-b border-gray-100 py-2 no-wrap"):
                    ui.icon(ikona, color=f"{bv}-6").classes("text-xl mt-1 shrink-0")
                    with ui.column().classes("gap-0 flex-1 min-w-0"):
                        with ui.row().classes("items-center gap-2 flex-wrap"):
                            ui.label(r.get("akce") or "").classes("text-sm font-bold text-gray-800")
                            ui.label(_dt_cz(r.get("kdy"))) \
                                .classes("text-xs text-gray-400 font-mono")
                            ui.label(f"· {r.get('kdo') or '—'}").classes("text-xs text-gray-500")
                        if r.get("detail"):
                            ui.label(r["detail"]) \
                                .classes("text-sm text-gray-600 whitespace-pre-wrap")
        with ui.row().classes("justify-end w-full"):
            ui.button("Zavřít", on_click=dlg.close).props("flat no-caps")
    dlg.open()


def _dialog_detail(pripad, user_id, user_name, prava):
    klient = context.client   # stabilní reference pro notifikace po awaitech (viz _bezpecne_notify)
    radky = nacti_radky(pripad["id"])
    vidi_op = _vidi_op(prava)
    typ = pripad["typ"]
    oddeleni = pripad["oddeleni"]
    stav = pripad["stav"]
    je_muj = (pripad["zadavatel_id"] == user_id)
    je_test = bool(pripad.get("testovaci"))
    # Smí tento případ spravovat jako schvalovatel? Plný správce typu NEBO schvalovatel
    # oddělení žadatele (právo „Schvalovatel – oddělení").
    muze_schvalit = _je_spravce_typu(typ, prava) or _je_schvalovatel_oddeleni(pripad, prava)
    with ui.dialog() as dlg, ui.card().style(
            "max-width: 96vw; width: 1000px; max-height: 90vh; overflow-y: auto"
            + ("; border: 2px solid #a855f7; background: #faf5ff" if je_test else "")):
        with ui.row().classes("w-full items-center gap-2"):
            ui.label(f"Případ {pripad['cislo']}").classes("text-xl font-bold")
            ui.label(pripad["nazev"]).classes("text-gray-600")
            lbl, barva = _STAV_BADGE.get(stav, (stav, "grey"))
            ui.badge(lbl, color=barva).props("outline")
            if je_test:
                ui.badge("🧪 TESTOVACÍ", color="purple")
            ui.space()
            ui.button(icon="history", on_click=lambda: _dialog_historie(pripad)) \
                .props("flat round dense").classes("text-indigo-600") \
                .tooltip("Průběh případu (očičko)")
        ui.label(f"{TYPY.get(typ, {}).get('nazev', typ)} · "
                 f"{pripad['pocet_radku']} řádků, {pripad['pocet_chyb']} chyb · "
                 f"žadatel {pripad['zadavatel_jmeno']}").classes("text-sm text-gray-500 mb-2")
        if je_test:
            with ui.row().classes("w-full items-center gap-2 bg-purple-50 rounded p-2 mb-2 "
                                  "border border-purple-300"):
                ui.icon("science", color="purple-7")
                ui.label("TESTOVACÍ PŘÍPAD — neodesílají se žádné e-maily.") \
                    .classes("text-sm font-semibold text-purple-800")
        if stav == "stornovano":
            with ui.row().classes("w-full items-center gap-2 bg-gray-100 rounded p-2 mb-2"):
                ui.icon("cancel", color="grey-7")
                ui.label(f"STORNOVÁNO — důvod: {pripad.get('storno_duvod') or '(neuveden)'}") \
                    .classes("text-sm font-medium text-gray-700")
        if stav == "zamitnuto" and pripad.get("zamitnuti_duvod"):
            with ui.row().classes("w-full items-center gap-2 bg-red-50 rounded p-2 mb-2 "
                                  "border border-red-200"):
                ui.icon("block", color="red-7")
                ui.label(f"Důvod zamítnutí: {pripad['zamitnuti_duvod']}") \
                    .classes("text-sm font-medium text-red-900")
        if pripad.get("poznamka_zadani"):
            with ui.row().classes("w-full items-center gap-2 bg-blue-50 rounded p-2 mb-2 "
                                  "border border-blue-200"):
                ui.icon("info", color="blue-7")
                ui.label(f"Důvod nahrání: {pripad['poznamka_zadani']}") \
                    .classes("text-sm font-medium text-blue-900")
        if pripad.get("poznamka"):
            with ui.row().classes("w-full items-center gap-2 bg-amber-50 rounded p-2 mb-2 "
                                  "border border-amber-200"):
                ui.icon("sticky_note_2", color="amber-8")
                ui.label(f"Poznámka žadatele: {pripad['poznamka']}") \
                    .classes("text-sm font-medium text-amber-900")
        _sekce_prilohy(pripad["id"], klient, je_muj or muze_schvalit)

        # Vstupní (zadané) sloupce dle typu — vidí je všichni; OP/marže/Detail jen správce.
        vstup_pole = list(dict.fromkeys(TYPY.get(typ, {}).get("mapa", {}).values()))
        ukaz_nazev = "nazev" not in vstup_pole   # porovnání nemá název ve vstupu → z masteru
        # NC faktura: vstupní sloupce dle podtypu formuláře (nc / nc+fc).
        ncf_sub = _ncf_subtyp(radky) if typ == "ncfaktura" else None
        # DNC3 (master NC3) a detail porovnání = pohled pro office nákup/správce
        # (vidí celou frontu); žadatel vidí jen svá vstupní data a výsledný Stav.
        ncf_srovnani = typ == "ncfaktura" and _vidi_vsechny_pripady(prava)
        ncf_pole = []
        if typ == "ncfaktura":
            ncf_pole = (["cena_nastaveni", "datum_od", "datum_do", "nazev_ind"]
                        if ncf_sub == "nc" else
                        ["nc_import", "nc_datum_od", "nc_datum_do",
                         "fc_import", "fc_datum_od", "fc_datum_do"])
        cols = [{"name": "poradi", "label": "#", "field": "poradi", "align": "left"}]
        if ukaz_nazev:
            cols.append({"name": "karta", "label": "Název", "field": "karta", "align": "left"})
        for f in vstup_pole:
            cols.append({"name": f, "label": POLE_LABEL.get(f, f), "field": f, "align": "left"})
        for f in ncf_pole:
            cols.append({"name": f, "label": POLE_LABEL.get(f, f), "field": f, "align": "left"})
        if typ != "ncfaktura":
            cols.append({"name": "sortiment_popis", "label": "Sortiment popis",
                         "field": "sortiment_popis", "align": "left"})
        cols.append({"name": "verdikt", "label": "Stav", "field": "verdikt", "align": "left"})
        if ncf_srovnani:
            if ncf_sub == "nc":
                cols.append({"name": "ncf_dnc3", "label": "DNC3 (NC3)",
                             "field": "ncf_dnc3", "align": "right"})
            cols.append({"name": "ncf_detail", "label": "Detail",
                         "field": "ncf_detail", "align": "left"})
        if typ == "porovnani":
            # Samostatná kontrola nastavení ceny (mimo verdikt) — vidí všichni (jen zadané hodnoty).
            cols.append({"name": "kontrola", "label": "Kontrola",
                         "field": "kontrola", "align": "left"})
        if vidi_op:
            cols.append({"name": "op_master", "label": "OP", "field": "op_master", "align": "left"})
            if typ == "porovnani":
                # Sloupce z workbooku: P nejvýhodnější NC, Q netto, R % marže netto, S sleva z nej NC.
                cols += [
                    {"name": "nej_nc", "label": "Nejvýh. NC", "field": "nej_nc", "align": "right"},
                    {"name": "netto", "label": "Netto", "field": "netto", "align": "right"},
                    {"name": "marze_netto", "label": "% marže netto",
                     "field": "marze_netto", "align": "right"},
                    {"name": "sleva_nej_nc", "label": "Sleva z nej NC",
                     "field": "sleva_nej_nc", "align": "right"},
                ]
            else:
                cols.append({"name": "marze", "label": "Marže", "field": "marze", "align": "left"})
            cols.append({"name": "duvod", "label": "Detail", "field": "duvod", "align": "left"})
        # CostPrice (Výběrová řízení) — zcela vpravo, JEN správce modulu / Office obchod,
        # jen ind. ceny.
        vidi_costprice = (typ == "porovnani" and _vidi_costprice(prava))
        cp_map = {}
        if vidi_costprice:
            cp_map = _costprice_dle_kodu([r.get("kod") for r in radky])
            cols += [
                {"name": "cp_a", "label": "CostPrice", "field": "cp_a", "align": "right"},
                {"name": "cp_b", "label": "CostPrice +7", "field": "cp_b", "align": "right"},
                {"name": "cp_c", "label": "CostPrice +14", "field": "cp_c", "align": "right"},
                {"name": "cp_d", "label": "CostPrice +29", "field": "cp_d", "align": "right"},
            ]
        rows = []
        for r in radky:
            try:
                vst = json.loads(r["vstup_json"]) if r["vstup_json"] else {}
            except Exception:
                vst = {}
            neschv = bool(r.get("neschvaleno"))
            d = {"poradi": r["poradi"],
                 "verdikt": ("🚫 Neschváleno" if neschv
                             else _VERDIKT_TXT.get(r["verdikt"], r["verdikt"])),
                 "_verdikt": r["verdikt"], "_neschvaleno": neschv,
                 "sortiment_popis": r.get("sortiment_popis")}
            if ukaz_nazev:
                d["karta"] = r["nazev_karty"]
            for f in vstup_pole:
                if f == "prirazka":
                    d[f] = _prirazka_nahled(vst.get(f))
                elif f in ("kod", "kod_produktu"):
                    d[f] = _export_kod(vst.get(f)) or None   # vodicí nuly (8 míst)
                elif f == "zakaznik_ico":
                    d[f] = _export_ico(vst.get(f)) or None   # vodicí nuly IČ
                else:
                    d[f] = _nahled_datum(vst.get(f))
            for f in ncf_pole:
                val = vst.get(f)
                if "datum" in f:
                    d[f] = _nahled_datum(val)
                elif f in ("cena_nastaveni", "nc_import", "fc_import"):
                    d[f] = _castka(cp.parse_cislo(val))
                else:
                    d[f] = "" if val is None else val
            rr = {}
            if r["marze_json"] and (typ == "porovnani" or vidi_op or ncf_srovnani):
                try:
                    rr = json.loads(r["marze_json"])
                except Exception:
                    rr = {}
            if ncf_srovnani:
                if ncf_sub == "nc":
                    d["ncf_dnc3"] = _castka(rr.get("dnc3"))
                d["ncf_detail"] = r["duvod"]
            if typ == "porovnani":
                kv = rr.get("kontrola")
                d["kontrola"] = {"OK": "✅ OK", "Chyba": "❌ Chyba"}.get(kv, "—")
            if vidi_op:
                op_val = r["op"] if r["op"] is not None else rr.get("master_op")
                d["op_master"] = "" if op_val is None else _pct(op_val)
                if typ == "porovnani":
                    d["nej_nc"] = _castka(rr.get("nej_nc"))
                    d["netto"] = _castka(rr.get("netto"))
                    d["marze_netto"] = _pct(rr.get("marze_netto"))
                    d["sleva_nej_nc"] = _pct(rr.get("sleva_nej_nc"))
                else:
                    d["marze"] = _marze_text(rr)
                d["duvod"] = r["duvod"]
            # CostPrice + semafor: 🔴 když PC bez DPH < daná CostPrice, 🟢 když ≥.
            if vidi_costprice:
                pc = cp.parse_cislo(vst.get("pc_bez_dph"))
                cpv = cp_map.get(_export_kod(r.get("kod")))   # kanonický klíč (8 míst)
                cpcols = {}
                for name, val in zip(("cp_a", "cp_b", "cp_c", "cp_d"),
                                     (cpv if cpv else (None, None, None, None))):
                    d[name] = _castka(val) if val is not None else "—"
                    if pc is not None and val is not None:
                        cpcols[name] = ("color:#dc2626;font-weight:700" if pc < val
                                        else "color:#16a34a;font-weight:700")
                d["_cpcols"] = cpcols
            # Nová karta (kód není v ceníku) — kód červeně, ať to vidí i žadatel,
            # ne jen správce v (skrytém) sloupci Detail.
            if typ == "mimoletak" and (r["duvod"] or "") == cp.DUVOD_NENALEZEN:
                d["_novy_kod"] = True
                d.setdefault("_cpcols", {})["kod"] = "color:#dc2626;font-weight:700"
            rows.append(d)
        pocet_novych = sum(1 for d in rows if d.get("_novy_kod"))
        if pocet_novych:
            ui.label(f"⚠️ {pocet_novych} kódů (červeně) není v databázi — jde o nové "
                     "karty bez cen, proto u nich nejde spočítat marže. Ceny nahrajte "
                     "znovu zítra, až budou karty v systému.") \
                .classes("text-sm font-medium text-red-600 mb-1")
        pocet_chyb = sum(1 for d in rows if d.get("_verdikt") == "CHYBA")
        pocet_neschv = sum(1 for d in rows if d.get("_neschvaleno"))
        stav_filtru = {"jen_chyby": False, "jen_neschv": False}

        # Filtry „jen chyby" a „jen neschválené" se kombinují (AND); přepínač
        # neschválených se zobrazí jen u částečně schválených případů (jsou-li nějaké
        # neschválené řádky). Filtruje JEN náhled tabulky, ne export.
        def _aplikuj_filtr():
            vyb = rows
            if stav_filtru["jen_chyby"]:
                vyb = [d for d in vyb if d.get("_verdikt") == "CHYBA"]
            if stav_filtru["jen_neschv"]:
                vyb = [d for d in vyb if d.get("_neschvaleno")]
            tbl.rows = vyb
            tbl.update()

        if pocet_chyb or pocet_neschv:
            with ui.row().classes("items-center gap-4 mb-1"):
                if pocet_chyb:
                    def _filtr_chyby(e):
                        stav_filtru["jen_chyby"] = bool(e.value)
                        _aplikuj_filtr()
                    ui.switch(f"Zobrazit jen chyby ({pocet_chyb} z {len(rows)})",
                              on_change=_filtr_chyby).classes("text-sm")
                if pocet_neschv:
                    def _filtr_neschv(e):
                        stav_filtru["jen_neschv"] = bool(e.value)
                        _aplikuj_filtr()
                    ui.switch(f"Zobrazit jen neschválené ({pocet_neschv} z {len(rows)})",
                              on_change=_filtr_neschv).classes("text-sm")
        # Režim schvalování IND: správce odškrtává položky PŘÍMO v této tabulce
        # (vestavěný výběr řádků; vše předzaškrtnuté = schválit). Mimo schvalování je
        # tabulka jen pro čtení a neschválené řádky se zašedí/přeškrtnou.
        muze_schvalovat = (typ == "porovnani" and stav == "ceka_na_spravce"
                           and muze_schvalit)
        if muze_schvalovat:
            ui.label("Schvalování: zaškrtnuté řádky se schválí. ODŠKRTNĚTE ty, které "
                     "NEchcete schválit — zašednou, nebudou v exportu a případ dostane "
                     "stav „Částečně schváleno“.").classes("text-sm text-teal-700 font-medium mb-1")
            tbl = ui.table(columns=cols, rows=rows, row_key="poradi",
                           selection="multiple").classes("w-full")
            tbl.selected = list(rows)   # vše zaškrtnuté = schválit
        else:
            tbl = ui.table(columns=cols, rows=rows, row_key="poradi").classes("w-full")
            # Body slot: zachová zašednutí neschválených + obarví buňky CostPrice (semafor).
            tbl.add_slot("body", r'''
                <q-tr :props="props" :class="props.row._neschvaleno ? 'text-grey-5' : ''"
                      :style="props.row._neschvaleno ? 'text-decoration: line-through; opacity:.6' : ''">
                  <q-td v-for="col in props.cols" :key="col.name" :props="props"
                        :style="(props.row._cpcols && props.row._cpcols[col.name]) || ''">{{ col.value }}</q-td>
                </q-tr>
            ''')
        if not vidi_op:
            ui.label("OP a marže v % jsou skryté — vidí je pouze správce.") \
                .classes("text-xs text-gray-400 italic mt-1")

        def _bezpecne_notify(zprava, barva="positive", timeout=7000):
            # Po awaitu (to_thread) mohl být slot dialogu mezitím smazán refreshem
            # seznamu -> ui.notify přes context.client by spadl ('parent ... deleted').
            # Notifikujeme proto přes zachyceného klienta; když je pryč, mlčky vynecháme.
            try:
                with klient:
                    ui.notify(zprava, type=barva, timeout=timeout)
            except Exception:
                pass

        async def _po_akci(zprava=None, barva="positive"):
            try:
                dlg.close()
            except Exception:
                pass
            if zprava:
                _bezpecne_notify(zprava, barva)
            _refresh()

        ui.separator().classes("my-2")
        with ui.column().classes("w-full gap-2"):
            if je_test and stav != "stornovano" and (je_muj or _je_spravce_typu(typ, prava)):
                def _prevod_dialog():
                    with ui.dialog() as pdlg, ui.card().classes("p-4").style("width: 470px"):
                        ui.label(f"Převést případ {pripad['cislo']} na ostrý?") \
                            .classes("font-medium")
                        ui.label("Případ přestane být testovací a projde standardním procesem "
                                 "(schválení a dále). Podle aktuálního stavu se odešlou příslušné "
                                 "e-maily.").classes("text-sm text-gray-500")

                        async def _potvrd():
                            ok, err = await asyncio.to_thread(prevod_na_ostry, pripad["id"])
                            pdlg.close()
                            if not ok:
                                _bezpecne_notify(err or "Převod selhal.", "negative")
                                return
                            zaznam_historie(
                                pripad["id"], "Převedeno na ostrý případ", user_name,
                                "Testovací případ převeden na ostrý — pokračuje běžným procesem.")
                            # Úvodní notifikace jako u nového případu (dle aktuálního stavu).
                            if stav == "vyhodnoceno_ok":
                                _kontr = kontrola_stav_pro_pripady([pripad["id"]]).get(pripad["id"])
                                _odesli_emaily(
                                    _emaily_office(oddeleni),
                                    f"Cenopřípad {pripad['cislo']}: nový případ ke zpracování",
                                    f"Žadatel {pripad['zadavatel_jmeno']} převedl případ "
                                    f"„{pripad['nazev']}“ ({TYPY[typ]['nazev']}) z testovacího na "
                                    f"ostrý — je v pořádku. Můžete ho zpracovat a dotáhnout do "
                                    f"nastavení."
                                    + (_KONTROLA_VAROVANI_MAIL if _kontr == "chyba" else ""))
                            elif stav == "ceka_na_spravce":
                                _odesli_emaily(
                                    _emaily_spravce(oddeleni),
                                    f"Cenopřípad {pripad['cislo']}: žádost o druhou kontrolu",
                                    f"Případ „{pripad['nazev']}“ ({TYPY[typ]['nazev']}) byl převeden "
                                    f"z testovacího na ostrý a čeká na vaši druhou kontrolu.")
                            intranet_logger.log_activity(
                                user_name, "Cenopřípad", f"Převod na ostrý {pripad['cislo']}")
                            await _po_akci("Případ převeden na ostrý — pokračuje běžným procesem.",
                                           "positive")

                        with ui.row().classes("w-full justify-end gap-2 mt-2"):
                            ui.button("Zrušit", on_click=pdlg.close).props("flat no-caps")
                            ui.button("Převést na ostrý", icon="published_with_changes",
                                      on_click=_potvrd) \
                                .props("unelevated no-caps").classes("bg-purple-600 text-white")
                    pdlg.open()

                ui.button("Převést na ostrý případ", icon="published_with_changes",
                          on_click=_prevod_dialog) \
                    .props("unelevated no-caps") \
                    .classes("bg-purple-600 hover:bg-purple-700 text-white rounded-lg px-4") \
                    .tooltip("Z testovacího případu udělá ostrý — projde standardním "
                             "schvalovacím procesem.")

            if je_muj and stav in ("vyhodnoceno_chyba", "zamitnuto"):
                ui.label("Vaše žádost není v pořádku. Opravte ji, nebo požádejte o druhou kontrolu.") \
                    .classes("text-sm text-red-600 font-medium")

                def _zadej():
                    prilohy = []   # [{nazev, typ, data}] — drženo do odeslání
                    with ui.dialog() as zdlg, ui.card().classes("p-4").style("width: 520px"):
                        ui.label("Žádost o druhou kontrolu").classes("font-medium")
                        ui.label("Můžete připsat poznámku pro správce a přiložit přílohy / fotky "
                                 "(např. foto regálu, doklad).").classes("text-sm text-gray-500")
                        pozn_in = ui.textarea("Poznámka pro správce") \
                            .props("outlined autogrow").classes("w-full")

                        @ui.refreshable
                        def _seznam_pril():
                            if not prilohy:
                                ui.label("Zatím žádné přílohy.").classes("text-xs text-gray-400 italic")
                                return
                            for i, p in enumerate(prilohy):
                                with ui.row().classes("w-full items-center gap-2 no-wrap text-sm"):
                                    ui.icon("image" if (p["typ"] or "").startswith("image/")
                                            else "attach_file", size="1.1rem") \
                                        .classes("text-gray-500")
                                    ui.label(p["nazev"]).classes("truncate")
                                    ui.label(f"{len(p['data']) // 1024} kB").classes("text-gray-400")
                                    ui.space()
                                    ui.button(icon="close", on_click=lambda i=i: _smaz(i)) \
                                        .props("flat round dense size=sm").classes("text-red-500")

                        def _smaz(i):
                            if 0 <= i < len(prilohy):
                                prilohy.pop(i)
                                _seznam_pril.refresh()

                        async def _on_priloha(e):
                            f = getattr(e, "file", None)
                            if f is None or not hasattr(f, "read"):
                                return
                            try:
                                raw = await f.read()
                            except Exception as exc:
                                ui.notify(f"Chyba čtení přílohy: {exc}", type="negative")
                                return
                            if not raw:
                                return
                            if len(raw) > 15_000_000:
                                ui.notify(f"Příloha „{getattr(f, 'name', '')}“ je větší než 15 MB.",
                                          type="warning")
                                return
                            prilohy.append({"nazev": getattr(f, "name", "") or "priloha",
                                            "typ": getattr(f, "content_type", "") or "",
                                            "data": raw})
                            _seznam_pril.refresh()

                        ui.upload(on_upload=_on_priloha, auto_upload=True, multiple=True,
                                  max_file_size=15_000_000, label="Přidat přílohu / fotku") \
                            .props('accept="image/*,.pdf,.jpg,.jpeg,.png,.heic,.webp,'
                                   '.doc,.docx,.xls,.xlsx"').classes("w-full")
                        _seznam_pril()

                        async def _odesli():
                            pozn = (pozn_in.value or "").strip()
                            ok, err = await asyncio.to_thread(
                                zadej_druhou_kontrolu, pripad["id"], pozn)
                            if not ok:
                                _bezpecne_notify(err or "Akce selhala.", "negative")
                                return
                            for p in prilohy:   # ulož přílohy (best-effort)
                                await asyncio.to_thread(uloz_prilohu, pripad["id"], p["nazev"],
                                                        p["typ"], p["data"], user_name)
                            zaznam_historie(
                                pripad["id"], "Žádost o druhou kontrolu", user_name,
                                ((f"Poznámka: {pozn}" if pozn else "Bez poznámky")
                                 + (f" · {len(prilohy)} příloh" if prilohy else "")))
                            zdlg.close()
                            text = (f"Žadatel {pripad['zadavatel_jmeno']} žádá o druhou kontrolu "
                                    f"případu „{pripad['nazev']}“ ({TYPY[typ]['nazev']}). "
                                    f"Případ má {pripad['pocet_chyb']} chyb a čeká na posouzení.")
                            if pozn:
                                text += f"\n\nPoznámka žadatele: {pozn}"
                            if prilohy:
                                text += (f"\n\nPřiloženo {len(prilohy)} příloh(a/y) — "
                                         f"otevřete případ v portálu.")
                            if not je_test:   # testovací případ NEodesílá e-maily
                                _odesli_emaily(
                                    _emaily_spravce(oddeleni),
                                    f"Cenopřípad {pripad['cislo']}: žádost o druhou kontrolu", text)
                            intranet_logger.log_activity(
                                user_name, "Cenopřípad", f"Žádost o 2. kontrolu {pripad['cislo']}"
                                + (f", {len(prilohy)} příloh" if prilohy else "")
                                + (f": {pozn}" if pozn else ""))
                            await _po_akci("Odesláno správci ke druhé kontrole.", "warning")

                        with ui.row().classes("w-full justify-end gap-2 mt-2"):
                            ui.button("Zrušit", on_click=zdlg.close).props("flat no-caps")
                            ui.button("Odeslat ke kontrole", icon="how_to_reg", on_click=_odesli) \
                                .props("unelevated no-caps").classes("bg-amber-600 text-white")
                    zdlg.open()

                ui.button("Mám vše v pořádku — žádám o druhou kontrolu",
                          icon="how_to_reg", on_click=_zadej) \
                    .props("unelevated no-caps").classes("bg-amber-600 text-white rounded-lg px-4")
                with ui.expansion("Opravit a přepočítat", icon="autorenew").classes("w-full"):
                    _oprava_widget(pripad, typ, user_name, _po_akci)

            if muze_schvalit and stav == "ceka_na_spravce":
                async def _proved_schvaleni(vyrazene):
                    novy_stav, ok, err = await asyncio.to_thread(
                        schval_pripad, pripad["id"], vyrazene)
                    if not ok:
                        _bezpecne_notify(err or "Schválení selhalo.", "negative")
                        return
                    castecne = bool(vyrazene)
                    zaznam_historie(
                        pripad["id"],
                        "Částečně schváleno" if castecne else "Schváleno správcem", user_name,
                        f"{len(vyrazene)} položek neschváleno (zašedlé, mimo export)"
                        if castecne else None)
                    if not je_test:   # testovací případ NEodesílá e-maily
                        _kontr = kontrola_stav_pro_pripady([pripad["id"]]).get(pripad["id"])
                        _odesli_emaily(
                            _emaily_office(oddeleni),
                            f"Cenopřípad {pripad['cislo']}: "
                            f"{'částečně ' if castecne else ''}schváleno ke zpracování",
                            f"Správce {'ČÁSTEČNĚ ' if castecne else ''}schválil případ "
                            f"„{pripad['nazev']}“ ({TYPY[typ]['nazev']})."
                            + (f" Vyřazeno {len(vyrazene)} položek (nejsou v exportu)."
                               if castecne else "")
                            + " Můžete ho zpracovat a dotáhnout do nastavení."
                            + (_KONTROLA_VAROVANI_MAIL if _kontr == "chyba" else ""))
                        _odesli_emaily_zadateli(
                            pripad,
                            f"Cenopřípad {pripad['cislo']}: "
                            f"{'částečně ' if castecne else ''}schváleno",
                            f"Váš případ „{pripad['nazev']}“ ({TYPY[typ]['nazev']}) byl správcem "
                            + (f"ČÁSTEČNĚ schválen — {len(vyrazene)} položek nebylo schváleno "
                               "(jsou zašedlé a nejsou v exportu)." if castecne
                               else "schválen a předán office ke zpracování."))
                    intranet_logger.log_activity(
                        user_name, "Cenopřípad",
                        f"{'Částečně schváleno' if castecne else 'Schváleno'} {pripad['cislo']}")
                    await _po_akci(("Částečně schváleno — odesláno na office." if castecne
                                    else "Schváleno — odesláno na office."), "positive")

                async def _schval():
                    if typ != "porovnani":          # výběr položek jen u IND
                        await _proved_schvaleni([])
                        return
                    # Výběr se čte PŘÍMO z tabulky výše (zaškrtnuté = schválit).
                    vybrane = {row.get("poradi") for row in (tbl.selected or [])}
                    vyrazene = sorted(r["poradi"] for r in radky if r["poradi"] not in vybrane)
                    if len(vyrazene) == len(radky):
                        _bezpecne_notify("Není zaškrtnutá žádná položka ke schválení. "
                                         "Zaškrtněte aspoň jednu, nebo případ zamítněte.", "warning")
                        return
                    await _proved_schvaleni(vyrazene)

                def _zamitni():
                    with ui.dialog() as zdlg, ui.card().classes("p-4").style("width: 460px"):
                        ui.label(f"Zamítnout případ {pripad['cislo']} — {pripad['nazev']}?") \
                            .classes("font-medium")
                        ui.label("Můžete připsat poznámku jako důvod zamítnutí — uvidí ji žadatel.") \
                            .classes("text-sm text-gray-500")
                        duvod_in = ui.textarea("Důvod zamítnutí (nepovinné)") \
                            .props("outlined autogrow").classes("w-full")

                        async def _potvrd():
                            d = (duvod_in.value or "").strip()
                            ok, err = await asyncio.to_thread(zamitni_pripad, pripad["id"], d)
                            zdlg.close()
                            if not ok:
                                _bezpecne_notify(err or "Zamítnutí selhalo.", "negative")
                                return
                            zaznam_historie(pripad["id"], "Zamítnuto správcem", user_name,
                                            (f"Důvod: {d}" if d else None))
                            if not je_test:
                                _odesli_emaily_zadateli(
                                    pripad, f"Cenopřípad {pripad['cislo']}: zamítnuto",
                                    f"Váš případ „{pripad['nazev']}“ ({TYPY[typ]['nazev']}) byl správcem "
                                    f"zamítnut."
                                    + (f"\n\nDůvod zamítnutí: {d}" if d else "")
                                    + " Otevřete ho v aplikaci, opravte a přepočítejte, nebo "
                                    "znovu požádejte o druhou kontrolu.")
                            intranet_logger.log_activity(
                                user_name, "Cenopřípad",
                                f"Zamítnuto {pripad['cislo']}" + (f": {d}" if d else ""))
                            await _po_akci("Zamítnuto — vráceno žadateli.", "negative")

                        with ui.row().classes("w-full justify-end gap-2 mt-2"):
                            ui.button("Zrušit", on_click=zdlg.close).props("flat no-caps")
                            ui.button("Zamítnout", icon="block", on_click=_potvrd) \
                                .props("unelevated no-caps").classes("bg-red-600 text-white")
                    zdlg.open()

                with ui.row().classes("gap-2"):
                    ui.button("Schválit", icon="verified", on_click=_schval) \
                        .props("unelevated no-caps").classes("bg-green-600 text-white rounded-lg px-4")
                    ui.button("Zamítnout", icon="block", on_click=_zamitni) \
                        .props("unelevated no-caps").classes("bg-red-600 text-white rounded-lg px-4")
                with ui.expansion("Opravit a přepočítat (správce)", icon="autorenew").classes("w-full"):
                    _oprava_widget(pripad, typ, user_name, _po_akci)

            if _je_office_typu(typ, prava) and stav in ("vyhodnoceno_ok", "schvaleno",
                                                        "castecne_schvaleno", "delisting"):
                async def _zprac():
                    zmen_stav(pripad["id"], "zpracovano")
                    zaznam_historie(pripad["id"], "Zpracováno (office)", user_name,
                                    "Případ dotažen do nastavení — finální stav.")
                    if not je_test:
                        _odesli_emaily_zadateli(
                            pripad, f"Cenopřípad {pripad['cislo']}: zpracováno",
                            f"Váš případ „{pripad['nazev']}“ ({TYPY[typ]['nazev']}) byl zpracován "
                            f"a dotažen do nastavení. Tím je vyřízen.")
                    intranet_logger.log_activity(user_name, "Cenopřípad", f"Zpracováno {pripad['cislo']}")
                    await _po_akci("Označeno jako zpracováno.", "positive")

                ui.button("Označit jako zpracováno", icon="task_alt", on_click=_zprac) \
                    .props("unelevated no-caps").classes("bg-blue-600 text-white rounded-lg px-4")

            if (_je_office_typu(typ, prava) and stav == "zpracovano"
                    and TYPY[typ]["oddeleni"] == "obchod"):
                async def _uzavri():
                    zmen_stav(pripad["id"], "uzavreno")
                    zaznam_historie(pripad["id"], "Uzavřeno (office)", user_name,
                                    "Případ uzavřen — finální stav (bez notifikace).")
                    intranet_logger.log_activity(user_name, "Cenopřípad",
                                                 f"Uzavřeno {pripad['cislo']}")
                    await _po_akci("Případ uzavřen.", "positive")

                ui.button("Uzavřít", icon="lock", on_click=_uzavri) \
                    .props("unelevated no-caps").classes("bg-gray-700 text-white rounded-lg px-4") \
                    .tooltip("Převede případ ze stavu „Zpracováno“ na „Uzavřeno“ "
                             "(bez e-mailové notifikace).")

        # Office (nákup i obchod) smí stornovat případy svého oddělení (s povinným důvodem).
        smi_stornovat = (je_muj or _je_spravce_typu(typ, prava)
                         or _je_office_typu(typ, prava))
        with ui.row().classes("w-full justify-end mt-2 gap-2"):
            if stav != "stornovano" and smi_stornovat:
                def _storno_dialog():
                    with ui.dialog() as sdlg, ui.card().classes("p-4").style("width: 420px"):
                        ui.label(f"Stornovat případ {pripad['cislo']} — {pripad['nazev']}?") \
                            .classes("font-medium")
                        duvod_in = ui.textarea("Důvod storna *") \
                            .props("outlined autogrow").classes("w-full")

                        async def _potvrd_storno():
                            d = (duvod_in.value or "").strip()
                            if not d:
                                ui.notify("Uveďte důvod storna.", type="warning")
                                return
                            ok, err = await asyncio.to_thread(stornuj_pripad, pripad["id"], d)
                            sdlg.close()
                            if ok:
                                intranet_logger.log_activity(
                                    user_name, "Cenopřípad", f"Storno {pripad['cislo']}: {d}")
                                zaznam_historie(pripad["id"], "Stornováno", user_name,
                                                f"Důvod: {d}")
                                if not je_test:   # testovací případ NEodesílá e-maily
                                    _odesli_emaily_zadateli(
                                        pripad, f"Cenopřípad {pripad['cislo']}: stornováno",
                                        f"Případ „{pripad['nazev']}“ ({TYPY[typ]['nazev']}) byl "
                                        f"stornován.\n\nDůvod: {d}")
                                await _po_akci(f"Případ {pripad['cislo']} stornován.", "warning")
                            else:
                                _bezpecne_notify(err or "Storno selhalo.", "negative")

                        with ui.row().classes("w-full justify-end gap-2 mt-2"):
                            ui.button("Zrušit", on_click=sdlg.close).props("flat no-caps")
                            ui.button("Stornovat", icon="cancel", on_click=_potvrd_storno) \
                                .props("unelevated no-caps").classes("bg-orange-700 text-white")
                    sdlg.open()

                ui.button("Stornovat", icon="cancel", on_click=_storno_dialog) \
                    .props("flat no-caps").classes("text-orange-700")
            if _je_spravce_typu(typ, prava):
                def _smaz_dialog():
                    with ui.dialog() as cdlg, ui.card().classes("p-4"):
                        ui.label(f"Nevratně smazat případ {pripad['cislo']} — "
                                 f"{pripad['nazev']}?").classes("font-medium")
                        ui.label("Smaže se případ, jeho řádky i nahraný soubor.") \
                            .classes("text-sm text-gray-500")

                        async def _potvrd():
                            ok, err = await asyncio.to_thread(smaz_pripad, pripad["id"])
                            cdlg.close()
                            if ok:
                                intranet_logger.log_activity(
                                    user_name, "Cenopřípad", f"Smazán případ {pripad['cislo']}")
                                await _po_akci(f"Případ {pripad['cislo']} smazán.", "warning")
                            else:
                                _bezpecne_notify(err or "Smazání selhalo.", "negative")

                        with ui.row().classes("w-full justify-end gap-2 mt-2"):
                            ui.button("Zrušit", on_click=cdlg.close).props("flat no-caps")
                            ui.button("Smazat", icon="delete_forever", on_click=_potvrd) \
                                .props("unelevated no-caps").classes("bg-red-600 text-white")
                    cdlg.open()

                ui.button("Smazat případ", icon="delete", on_click=_smaz_dialog) \
                    .props("flat no-caps").classes("text-red-600")
            # Paima nahraný soubor OBSAHUJE sloupce OP a marže (%) → smí ho stáhnout
            # JEN správce s právem na OP (_vidi_op). Webportál a mimoleták soubor žádné
            # OP/% nemají (kód/název/ceny + info jako splatnosti) → stahuje ho i
            # office/žadatel ve stejné podobě, jak byl nahrán (vč. údajů, které se
            # do řádků neparsují). Master OP/DATA soubory se do DB neukládají vůbec,
            # takže je stáhnout nelze.
            if typ in ("webportal", "mimoletak", "ncfaktura") or (typ == "paima" and vidi_op):
                je_soubor = _ma_soubor(pripad["id"])

                async def _stahni_soubor():
                    res = await asyncio.to_thread(nacti_soubor, pripad["id"])
                    try:
                        with klient:   # po awaitu mohl být slot dialogu smazán
                            if res:
                                ui.download.content(res[1], res[0], "application/octet-stream")
                            else:
                                ui.notify("U tohoto případu není uložený žádný soubor.",
                                          type="warning")
                    except Exception:
                        pass

                _btn_soubor = ui.button("Stáhnout nahraný soubor", icon="file_download",
                                        on_click=_stahni_soubor) \
                    .props("flat no-caps").classes("text-indigo-600")
                if not je_soubor:
                    _btn_soubor.set_enabled(False)
                    _btn_soubor.tooltip("Soubor je uložený jen u případů nahraných/přepočítaných "
                                        "po zavedení této funkce. U starších (nebo vložených přes "
                                        "schránku) soubor není.")
            if typ in _SLEPE_SLOUPCE or typ == "ncfaktura":
                def _export_slepy():
                    # NC faktura: office/správce dostanou i DNC3+Detail (master NC3),
                    # žadatel jen zadaná data + Stav.
                    ui.download.content(
                        _radky_slepe_xlsx(pripad, radky, _vidi_vsechny_pripady(prava)),
                        f"cenopripad_{pripad['cislo']}_data.xlsx", XLSX_MIME)

                ui.button("Stáhnout data (bez OP)", icon="download", on_click=_export_slepy) \
                    .props("flat no-caps").classes("text-indigo-600") \
                    .tooltip("Zadaná data + stav ke sdílení (všechny řádky případu) — "
                             "BEZ OP, marží a interních detailů.")
            if vidi_op:
                def _export_case():
                    vybr = ([r for r in radky if r["verdikt"] == "CHYBA"]
                            if stav_filtru["jen_chyby"] else radky)
                    ui.download.content(_radky_xlsx(pripad, vybr),
                                        f"cenopripad_{pripad['cislo']}.xlsx", XLSX_MIME)

                ui.button("Export případu (vč. OP)", icon="download", on_click=_export_case) \
                    .props("flat no-caps").classes("text-rose-600") \
                    .tooltip("Exportuje aktuálně zobrazené řádky (respektuje filtr jen chyby).")
            if typ == "porovnani" and _vidi_vsechny_pripady(prava):
                def _export_sklad():
                    dlg_sk = ui.dialog()
                    with dlg_sk, ui.card().classes("w-96"):
                        ui.label("Export do skladu").classes("text-lg font-semibold")
                        ui.label("Zadejte skupinu (sloupec SKUPINA), např. CENY-PEKOS. "
                                 "Vyexportují se schválené řádky případu.") \
                            .classes("text-sm text-gray-600")
                        inp_sk = ui.input("Skupina", value="CENY-") \
                            .props("outlined dense autofocus").classes("w-full")
                        sel_fmt = ui.select(
                            {k: v[0] for k, v in SKLAD_FORMATY.items()},
                            value="xlsx", label="Formát souboru") \
                            .props("outlined dense").classes("w-full") \
                            .tooltip("Starší importy (skladový systém na Excelu 2003) "
                                     "neumí .xlsx — pak zvolte některou .xls variantu.")

                        def _stahni_sklad():
                            sk = (inp_sk.value or "").strip()
                            if not sk:
                                ui.notify("Vyplňte skupinu.", type="warning")
                                return
                            try:
                                data, pripona, mime = _sklad_soubor(
                                    radky, sk, sel_fmt.value or "xlsx")
                            except Exception as e:
                                # nejčastěji chybějící xlwt na serveru nebo limit řádků
                                _bezpecne_notify(f"Export selhal: {e}", "negative", 10000)
                                return
                            dlg_sk.close()
                            ui.download.content(
                                data,
                                f"cenopripad_{pripad['cislo']}_do_skladu.{pripona}", mime)

                        with ui.row().classes("justify-end w-full"):
                            ui.button("Zrušit", on_click=dlg_sk.close).props("flat no-caps")
                            ui.button("Stáhnout", on_click=_stahni_sklad).props("no-caps")
                    dlg_sk.open()

                ui.button("Export do skladu", icon="warehouse", on_click=_export_sklad) \
                    .props("flat no-caps").classes("text-teal-700") \
                    .tooltip("Schválené řádky ve formátu pro import do skladu (list „IC“).")
            if stav != "stornovano" and (je_muj or _je_spravce_typu(typ, prava)):
                async def _prepocti_master():
                    novy, err = await asyncio.to_thread(prepocti_pripad, pripad["id"], typ)
                    if err:
                        _bezpecne_notify(err, "negative", 8000)
                        return
                    intranet_logger.log_activity(
                        user_name, "Cenopřípad", f"Přepočet dle masteru {pripad['cislo']}")
                    zaznam_historie(pripad["id"], "Přepočítáno dle masteru", user_name,
                                    "Znovu vyhodnoceno proti aktuálním datům.")
                    await _po_akci("Případ přepočítán dle aktuálního masteru.", "positive")

                ui.button("Přepočítat dle aktuálního masteru", icon="sync",
                          on_click=_prepocti_master) \
                    .props("flat no-caps").classes("text-emerald-700") \
                    .tooltip("Znovu vyhodnotí uložené řádky proti AKTUÁLNÍM datům "
                             "(OP, marže, sortiment popis) — použij po re-importu dat. "
                             "Data se nenahrávají znovu, stav se vrátí na vyhodnoceno.")
            ui.button("Zavřít", on_click=dlg.close).props("flat no-caps")
    dlg.open()


def _karta_pripadu(p, user_id, user_name, prava):
    lbl, barva = _STAV_BADGE.get(p["stav"], (p["stav"], "grey"))
    je_test = bool(p.get("testovaci"))
    if je_test:   # testovací = celý fialový (přebíjí zeleno/červeno/storno)
        okraj, pozadi = "border-l-4 border-purple-500", "bg-purple-50 hover:bg-purple-100"
    elif p["stav"] == "delisting":
        okraj, pozadi = "border-l-4 border-orange-500", "bg-orange-50 hover:bg-orange-100"
    elif p["stav"] == "stornovano":
        okraj, pozadi = "border-l-4 border-gray-400", "hover:bg-gray-50"
    else:
        okraj = "border-l-4 border-green-400" if p["vysledek_ok"] else "border-l-4 border-red-400"
        pozadi = "hover:bg-gray-50"
    with ui.card().classes(
            f"w-full {okraj} {pozadi} cursor-pointer transition-colors p-3 rounded-lg") \
            .on("click", lambda p=p: _dialog_detail(p, user_id, user_name, prava)):
        with ui.row().classes("w-full items-center gap-3 no-wrap"):
            ui.label(p["cislo"]).classes("font-mono text-sm text-gray-500")
            ui.label(p["nazev"]).classes("font-semibold text-gray-800 flex-1")
            if je_test:
                ui.badge("🧪 TEST", color="purple")
            ui.badge(lbl, color=barva).props("outline")
            # Druhý stav = agregovaný sloupec „Kontrola" (jen IND a jen když je
            # vyplněný sloupec „definice nastavení"; jinak se nezobrazuje).
            _kontr = p.get("_kontrola")
            if _kontr == "chyba":
                ui.badge("Kontrola: Není v pořádku", color="red").props("outline")
            elif _kontr == "ok":
                ui.badge("Kontrola: V pořádku", color="green").props("outline")
            if p.get("_archiv"):   # IND: po platnosti
                ui.badge("🗄️ ARCHIV", color="grey-7")
            if p.get("poznamka"):
                ui.icon("sticky_note_2", color="amber-8").classes("text-lg") \
                    .tooltip(p["poznamka"])
            if p.get("_platnost_do"):   # IND: platnost „termín do"
                ui.label(f"do {p['_platnost_do'].strftime('%d.%m.%Y')}") \
                    .classes("text-xs " + ("text-red-500 font-semibold"
                                           if p.get("_archiv") else "text-gray-500")) \
                    .tooltip("Platnost (nejzazší termín do)")
            ui.label(f"{p['pocet_chyb']}/{p['pocet_radku']} chyb").classes("text-xs text-gray-500")
            ui.label(str(p["datum_zadani"])[:16]).classes("text-xs text-gray-400")
            # „Očičko" — průběh případu; stopPropagation, ať se neotevře i detail.
            ui.button(icon="history") \
                .props("flat round dense size=sm").classes("text-indigo-500") \
                .on("click", lambda p=p: _dialog_historie(p),
                    js_handler="(e) => { if (e && e.stopPropagation) e.stopPropagation();"
                               " emit(); }") \
                .tooltip("Průběh případu (očičko)")
        if p["stav"] == "stornovano" and p.get("storno_duvod"):
            ui.label(f"⊘ Storno: {p['storno_duvod']}").classes("text-xs text-gray-500 italic")
        if p.get("poznamka"):
            ui.label(f"📝 Poznámka: {p['poznamka']}").classes("text-xs text-amber-800 italic")


def _dt_cz(v):
    """Datetime z DB → 'dd.mm.rrrr HH:MM' (jinak str)."""
    try:
        return v.strftime("%d.%m.%Y %H:%M")
    except Exception:
        return str(v or "")


def _bez_pripony(nazev):
    """Zobrazovaný název bez tabulkové přípony (.xlsb/.xlsx/.xlsm/.xls/.csv)."""
    s = (nazev or "").strip()
    low = s.lower()
    for ext in (".xlsb", ".xlsx", ".xlsm", ".xls", ".csv"):
        if low.endswith(ext):
            return s[:-len(ext)]
    return s


def _panel_import(user_name, refresh_fn):
    """Jedna tabulka zdrojů importu. U každého řádku zelené „+" otevře výběr
    souboru; vedle se ukáže, co je nahráno / vybráno. Jedno tlačítko „Importovat"."""
    pm, po = pocty_v_db()
    pcp = pocet_costprice_v_db()
    stav_m = stav_masteru()
    # (klíč, popis, accept, max_size, DB-stav záznam, text počtu v DB)
    zdroje = [
        ("data", "\U0001F4E6 DATA (porovnání)", ".xlsb", 80_000_000, stav_m.get("data"), f"{pm} produktů"),
        ("op", "\U0001F3F7\uFE0F OP (obchodní podmínky)", ".xlsb", 20_000_000, stav_m.get("op"), f"{po} sortimentů"),
        ("costprice", "\U0001F4C8 CostPrice (Výběrová řízení)", ".xlsx", 80_000_000, None, f"{pcp} kódů"),
    ]
    drzeny = {k: {"raw": None, "name": ""} for k, *_ in zdroje}
    ups, stav_lbl = {}, {}

    def _vychozi_text(rec, dbtext):
        if rec and rec.get("nazev"):
            return f"Naimportováno: {dbtext} · {_bez_pripony(rec['nazev'])} · {_dt_cz(rec.get('kdy'))}"
        if dbtext and not dbtext.startswith("0 "):
            return f"Naimportováno: {dbtext}"
        return "— zatím nenaimportováno —"

    with ui.card().classes("w-full max-w-3xl p-6 rounded-2xl shadow-lg"):
        ui.label("Import dat").classes("text-2xl font-bold text-gray-800")
        ui.label("U každého zdroje klikněte na zelené „+“ a vyberte soubor. Vpravo je "
                 "vidět, co je aktuálně nahráno, případně právě vybráno. Pak klikněte "
                 "„Importovat“ — naimportuje se vše vybrané.") \
            .classes("text-sm text-gray-500 mb-4")

        spin = ui.spinner(size="lg").classes("text-blue-600")
        spin.set_visibility(False)

        def _refresh_btn():
            btn.set_enabled(any(d["raw"] for d in drzeny.values()))

        def _mk_click(up):
            return lambda: ui.run_javascript(
                f"const c = getElement({up.id});"
                f" const i = c && c.$el && c.$el.querySelector('input[type=file]');"
                f" if (i) i.click();")

        def _mk_handler(key):
            async def _h(e):
                raw, name = await _precti_upload(e, ups[key])
                if raw is not None:
                    drzeny[key]["raw"], drzeny[key]["name"] = raw, name
                    stav_lbl[key].text = f"✅ Vybráno k importu: {name}"
                    stav_lbl[key].classes(replace="text-sm text-emerald-700 font-medium break-words")
                    _refresh_btn()
            return _h

        with ui.element("div").classes("w-full grid items-center gap-x-4 gap-y-3") \
                .style("grid-template-columns: max-content max-content minmax(0,1fr)"):
            for key, popis, acc, maxsz, rec, dbtext in zdroje:
                ui.label(popis).classes("font-medium text-gray-700 whitespace-nowrap")
                up = ui.upload(on_upload=_mk_handler(key), auto_upload=True, max_file_size=maxsz) \
                    .props(f"accept={acc}").style("display:none")
                ups[key] = up
                ui.button(icon="add", on_click=_mk_click(up)) \
                    .props("round unelevated color=green dense").classes("shadow-sm") \
                    .style("transform:scale(0.75);transform-origin:center") \
                    .tooltip("Vybrat soubor")
                stav_lbl[key] = ui.label(_vychozi_text(rec, dbtext)) \
                    .classes("text-sm text-gray-500 break-words")

        async def _import_click():
            if not any(d["raw"] for d in drzeny.values()):
                ui.notify("Vyberte alespoň jeden soubor (zelené „+“).", type="warning")
                return
            btn.set_enabled(False)
            spin.set_visibility(True)
            vysl = []
            try:
                if drzeny["op"]["raw"] or drzeny["data"]["raw"]:
                    # CPU-bound parse (pyxlsb) v samostatném PROCESU — neblokuje event loop.
                    pm2, po2, err = await run.cpu_bound(
                        _importuj_sync, drzeny["op"]["raw"], drzeny["data"]["raw"],
                        drzeny["op"]["name"], drzeny["data"]["name"], user_name)
                    if err:
                        raise RuntimeError(err)
                    vysl.append(f"{pm2} produktů, {po2} OP")
                if drzeny["costprice"]["raw"]:
                    n, err = await run.cpu_bound(
                        _importuj_costprice_sync, drzeny["costprice"]["raw"],
                        drzeny["costprice"]["name"], user_name)
                    if err:
                        raise RuntimeError(err)
                    vysl.append(f"{n} kódů CostPrice")
            except Exception as ex:
                spin.set_visibility(False)
                btn.set_enabled(True)
                ui.notify(f"Import selhal: {ex}", type="negative", timeout=9000)
                return
            spin.set_visibility(False)
            ui.notify("Import OK — " + "; ".join(vysl), type="positive",
                      position="top-right", timeout=6000)
            intranet_logger.log_activity(user_name, "Cenopřípad", "Import dat: " + "; ".join(vysl))
            refresh_fn()

        with ui.row().classes("w-full justify-end mt-4"):
            btn = ui.button("Importovat", icon="cloud_upload", on_click=_import_click) \
                .props("unelevated no-caps") \
                .classes("bg-blue-600 hover:bg-blue-700 text-white font-semibold rounded-lg shadow-md px-5")
            btn.set_enabled(False)



def _formular_novy_pripad(typ, user_id, user_name, oddeleni):
    cfg = TYPY[typ]
    drzeny = {"raw": None, "name": ""}
    with ui.card().classes("w-full max-w-3xl p-5 rounded-2xl shadow-lg mb-6"):
        ui.label("Nový případ ke kontrole").classes("text-xl font-bold text-gray-800")
        ui.label(f"Nahrajte vyplněný formulář „{cfg['nazev']}“ (.xlsx/.xls) NEBO vložte řádky "
                 "z Excelu (včetně hlavičky). Aplikace ceny zkontroluje a uloží.") \
            .classes("text-sm text-gray-500 mb-2")
        nazev_in = ui.input("Název případu *").props("outlined dense").classes("w-full max-w-md")
        # U mimolétáku/výprodeje je POVINNÝ důvod nahrání (bez něj to nepustí dál).
        duvod_in = None
        if typ == "mimoletak":
            duvod_in = ui.textarea("Důvod nahrání * (proč se mimoleták/výprodej nahrává)") \
                .props("outlined autogrow").classes("w-full max-w-md")

        async def _on_up(e):
            raw, name = await _precti_upload(e, up)
            if raw is not None:
                drzeny["raw"], drzeny["name"] = raw, name
                soubor_lbl.text = f"Přiložený soubor: {name}"
                soubor_radek.set_visibility(True)
                ui.notify(f"Soubor připraven: {name}", type="info")

        def _odeber_soubor():
            drzeny["raw"], drzeny["name"] = None, ""
            soubor_radek.set_visibility(False)
            up.reset()

        with ui.row().classes("w-full gap-4 flex-wrap mt-2"):
            with ui.column().classes("flex-1 min-w-72"):
                ui.label("Nahrát šablonu (.xlsx)").classes("text-xs font-semibold text-gray-500")
                up = ui.upload(on_upload=_on_up, auto_upload=True, max_file_size=20_000_000,
                               label="Vybrat soubor (.xlsx/.xls)") \
                    .props("accept=.xlsx,.xls,.xlsm").classes("w-full")
            with ui.column().classes("flex-1 min-w-72"):
                ui.label("…nebo vložit z Excelu (vč. hlavičky)") \
                    .classes("text-xs font-semibold text-gray-500")
                paste_in = ui.textarea(
                    placeholder="Sem vložte (Ctrl+V) zkopírované řádky z Excelu…") \
                    .props("outlined").classes("w-full").style("min-height: 90px")
                _hlidej_paste(paste_in)
        # Viditelné potvrzení přiloženého souboru — ať je PŘED odesláním ke
        # schválení jasné, že soubor je nahraný (notifikace zmizí, tohle zůstane).
        with ui.row().classes("items-center gap-1 mt-1 bg-emerald-50 border "
                              "border-emerald-200 rounded px-2 py-1") as soubor_radek:
            ui.icon("attach_file", size="1.1rem").classes("text-emerald-700")
            soubor_lbl = ui.label("").classes("text-sm font-semibold text-emerald-800")
            ui.button(icon="close", on_click=_odeber_soubor) \
                .props("flat round dense size=sm").classes("text-emerald-700") \
                .tooltip("Odebrat soubor")
        soubor_radek.set_visibility(False)
        spin = ui.spinner(size="lg").classes("text-emerald-600")
        spin.set_visibility(False)

        tlacitka = []

        async def _submit(testovaci=False, delist=False):
            nazev = (nazev_in.value or "").strip()
            if not nazev:
                ui.notify("Zadejte název případu.", type="warning")
                return
            duvod = (duvod_in.value or "").strip() if duvod_in else ""
            if typ == "mimoletak" and not duvod:
                ui.notify("U mimolétáku/výprodeje je povinný důvod nahrání.", type="warning")
                return
            text = (paste_in.value or "").strip()
            for b in tlacitka:
                b.set_enabled(False)
            spin.set_visibility(True)
            try:
                if text:
                    radky, err = await asyncio.to_thread(_parsuj_paste, text, typ)
                elif drzeny["raw"]:
                    radky, err = await asyncio.to_thread(
                        _parsuj_sablonu, drzeny["raw"], typ, drzeny["name"])
                else:
                    ui.notify("Nahrajte soubor nebo vložte data.", type="warning")
                    return
                if err:
                    ui.notify(err, type="negative", timeout=8000)
                    return
                vys = await asyncio.to_thread(vyhodnot_pripad_z_db, typ, radky)
                # Mimoleták: nové karty (kód není v ceníku) nejde spočítat — žadatel
                # to musí vidět HNED, ne až od správce. Dá potvrdit, nebo zrušit.
                if typ == "mimoletak":
                    nove = _nove_karty(radky, vys)
                    if nove:
                        spin.set_visibility(False)
                        potvrzeno = await _potvrd_nove_karty(nove)
                        spin.set_visibility(True)
                        if not potvrzeno:
                            ui.notify("Odeslání zrušeno — nahrajte prosím zítra, "
                                      "až budou karty s cenami v systému.",
                                      type="warning", timeout=8000)
                            return
                soubor_raw = None if text else drzeny["raw"]   # ulož jen nahraný soubor, ne paste
                cislo, stav_p, err2 = await asyncio.to_thread(
                    _uloz_pripad, typ, nazev, user_id, user_name, radky, vys,
                    soubor_raw, drzeny["name"], testovaci, duvod or None,
                    "delisting" if delist else None)
                if err2:
                    ui.notify(err2, type="negative", timeout=8000)
                    return
            except Exception as e:
                ui.notify(f"Neočekávaná chyba při zpracování: {e}", type="negative", timeout=9000)
                return
            finally:
                spin.set_visibility(False)
                for b in tlacitka:
                    b.set_enabled(True)
            if delist:
                # Delisting přeskakuje kontrolu/schvalování → rovnou na office.
                if not testovaci:
                    _odesli_emaily(
                        _emaily_office(oddeleni),
                        f"Cenopřípad {cislo}: DELISTING ke zpracování",
                        f"Žadatel {user_name} vložil DELISTING „{nazev}“ ({cfg['nazev']}). "
                        f"Případ přeskočil fázi kontroly — můžete ho rovnou zpracovat a "
                        f"dotáhnout do nastavení.")
                ui.notify(f"Případ {cislo}: 🗑️ DELISTING ({vys['pocet']} řádků) — "
                          + ("🧪 TESTOVACÍ, e-maily se neodeslaly." if testovaci
                             else "předáno rovnou office ke zpracování."),
                          type="positive", position="top", timeout=8000)
                intranet_logger.log_activity(
                    user_name, "Cenopřípad",
                    f"Nový {'TEST ' if testovaci else ''}DELISTING případ {cislo} ({typ})")
                _odeber_soubor()
                nazev_in.value = ""
                paste_in.value = ""
                if duvod_in is not None:
                    duvod_in.value = ""
                _refresh()
                return
            if vys["ok"]:
                _kontr = _kontrola_z_vysledku(typ, radky, vys)
                if not testovaci:   # testovací případ NEodesílá žádné e-maily
                    _odesli_emaily(
                        _emaily_office(oddeleni),
                        f"Cenopřípad {cislo}: nový případ ke zpracování",
                        f"Žadatel {user_name} vložil případ „{nazev}“ ({cfg['nazev']}), který je "
                        f"v pořádku. Můžete ho zpracovat a dotáhnout do nastavení."
                        + (_KONTROLA_VAROVANI_MAIL if _kontr == "chyba" else ""))
                ui.notify(f"Případ {cislo}: ✅ VŠE V POŘÁDKU ({vys['pocet']} řádků). "
                          + ("🧪 TESTOVACÍ — e-maily se neodeslaly." if testovaci
                             else "Office byl informován."),
                          type="positive", position="top", timeout=8000)
            else:
                ui.notify(f"Případ {cislo}: ❌ NENÍ v pořádku — {vys['chyby']} z {vys['pocet']} "
                          "řádků. Otevřete případ v seznamu: opravte a přepočítejte, nebo "
                          "požádejte o druhou kontrolu."
                          + (" 🧪 TESTOVACÍ." if testovaci else ""),
                          type="negative", position="top", timeout=11000)
            intranet_logger.log_activity(
                user_name, "Cenopřípad",
                f"Nový {'TEST ' if testovaci else ''}případ {cislo} ({typ}) – "
                f"{'ok' if vys['ok'] else 'chyba'}")
            _odeber_soubor()
            nazev_in.value = ""
            paste_in.value = ""
            if duvod_in is not None:
                duvod_in.value = ""
            _refresh()

        with ui.row().classes("w-full justify-end mt-2 gap-2"):
            btn_test = ui.button("Testovací", icon="science", on_click=lambda: _submit(True)) \
                .props("outline no-caps") \
                .classes("text-purple-700 font-semibold rounded-lg px-4") \
                .tooltip("Založí zkušební případ — NEODESÍLAJÍ se žádné e-maily a "
                         "případ je v seznamu fialově.")
            tlacitka.append(btn_test)
            if typ == "porovnani":
                btn_delist = ui.button("Delist", icon="remove_shopping_cart",
                                       on_click=lambda: _submit(delist=True)) \
                    .props("outline no-caps") \
                    .classes("text-orange-700 font-semibold rounded-lg px-4") \
                    .tooltip("Delisting — přeskočí fázi kontroly a předá případ rovnou "
                             "office obchod (stav „Delisting“).")
                tlacitka.append(btn_delist)
            btn = ui.button("Porovnat a schválit", icon="fact_check",
                            on_click=lambda: _submit(False)) \
                .props("unelevated no-caps") \
                .classes("bg-emerald-600 hover:bg-emerald-700 text-white font-semibold rounded-lg shadow-md px-5")
            tlacitka.append(btn)


def _parse_datum(v):
    """Datum z volného textu/Excelu → datetime.date, jinak None. Zvládne
    '31.07.2026', '31. 7. 2026', '2026-07-31', '2026-07-31 00:00:00', '31/07/2026'."""
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    m = re.search(r'(\d{1,2})\.\s*(\d{1,2})\.\s*(\d{4})', s)        # dd.mm.yyyy
    if not m:
        m2 = re.search(r'(\d{4})-(\d{1,2})-(\d{1,2})', s)            # yyyy-mm-dd
        if m2:
            try:
                return datetime.date(int(m2.group(1)), int(m2.group(2)), int(m2.group(3)))
            except ValueError:
                return None
        m3 = re.search(r'(\d{1,2})/(\d{1,2})/(\d{4})', s)            # dd/mm/yyyy
        if m3:
            try:
                return datetime.date(int(m3.group(3)), int(m3.group(2)), int(m3.group(1)))
            except ValueError:
                return None
        return None
    try:
        return datetime.date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    except ValueError:
        return None


def platnosti_do_pro_pripady(case_ids):
    """{pripad_id: nejzazší 'termín do' (date)} z řádků (vstup_json). Platnost case =
    poslední datum, do kdy platí aspoň jeden řádek. Případy bez data → bez záznamu."""
    ids = [i for i in case_ids if i]
    if not ids:
        return {}
    conn = intranet_data.get_db_connection()
    if not conn:
        return {}
    out = {}
    try:
        cur = conn.cursor(dictionary=True)
        ph = ",".join(["%s"] * len(ids))
        cur.execute(f"SELECT pripad_id, vstup_json FROM cenopripad_radky "
                    f"WHERE pripad_id IN ({ph})", ids)
        for r in cur.fetchall():
            try:
                vst = json.loads(r["vstup_json"]) if r["vstup_json"] else {}
            except Exception:
                continue
            d = _parse_datum(vst.get("termin_do"))
            if d is None:
                continue
            pid = r["pripad_id"]
            if pid not in out or d > out[pid]:
                out[pid] = d
        cur.close()
        return out
    except Exception as e:
        print(f"[cenopripad] platnosti_do_pro_pripady: {e}")
        return out
    finally:
        conn.close()


def kontrola_stav_pro_pripady(case_ids):
    """{pripad_id: 'ok'|'chyba'} — agregovaný sloupec „Kontrola" z řádků případu.
    Počítá se POUZE pro případy, které mají aspoň v jednom řádku vyplněný sloupec
    „definice nastavení" (vstup_json['definice']); ostatní případy nemají záznam
    (a badge se u nich nezobrazí). 'chyba' = aspoň jeden řádek má Kontrola == 'Chyba'
    (marze_json['kontrola']); jinak 'ok'."""
    ids = [i for i in case_ids if i]
    if not ids:
        return {}
    conn = intranet_data.get_db_connection()
    if not conn:
        return {}
    ma_definici, ma_chybu = set(), set()
    try:
        cur = conn.cursor(dictionary=True)
        ph = ",".join(["%s"] * len(ids))
        cur.execute(f"SELECT pripad_id, vstup_json, marze_json FROM cenopripad_radky "
                    f"WHERE pripad_id IN ({ph})", ids)
        for r in cur.fetchall():
            pid = r["pripad_id"]
            try:
                vst = json.loads(r["vstup_json"]) if r["vstup_json"] else {}
            except Exception:
                vst = {}
            if str(vst.get("definice") or "").strip():
                ma_definici.add(pid)
            try:
                mj = json.loads(r["marze_json"]) if r["marze_json"] else {}
            except Exception:
                mj = {}
            if mj.get("kontrola") == "Chyba":
                ma_chybu.add(pid)
        cur.close()
        return {pid: ("chyba" if pid in ma_chybu else "ok") for pid in ma_definici}
    except Exception as e:
        print(f"[cenopripad] kontrola_stav_pro_pripady: {e}")
        return {}
    finally:
        conn.close()


# Varovná věta do „OK" e-mailů, když je agregovaný sloupec „Kontrola" špatně.
_KONTROLA_VAROVANI_MAIL = (
    "\n\nPOZOR: Samostatná kontrola nastavení ceny (sloupec „Kontrola“ — výchozí cena "
    "× přirážka vs. PC) NENÍ u některých položek v pořádku. Případ je sice technicky "
    "v pořádku (verdikt OK), ale před zpracováním prosím nastavení cen prověřte.")


def _kontrola_z_vysledku(typ, vstup_radky, vysledek):
    """Agregovaný sloupec „Kontrola" z čerstvě spočítaného výsledku (bez DB).
    Vrací 'ok' / 'chyba' / None — None když se nevyhodnocuje (jiný typ než porovnani,
    nebo v žádném řádku není vyplněn sloupec „definice nastavení")."""
    if typ != "porovnani":
        return None
    ma_def = any(str((r or {}).get("definice") or "").strip() for r in (vstup_radky or []))
    if not ma_def:
        return None
    chyba = any((rr or {}).get("kontrola") == "Chyba" for rr in vysledek.get("radky", []))
    return "chyba" if chyba else "ok"


def _filtruj(items, flt):
    nazev = flt["nazev"].strip().lower()
    out = []
    for p in items:
        if nazev and nazev not in (p["nazev"] or "").lower():
            continue
        if flt["zadavatel"] and p["zadavatel_jmeno"] != flt["zadavatel"]:
            continue
        if flt["stav"] and p["stav"] != flt["stav"]:
            continue
        d = str(p["datum_zadani"])[:10]
        if flt["od"] and d < flt["od"]:
            continue
        if flt["do"] and d > flt["do"]:
            continue
        plat = flt.get("platnost")          # IND: platné (neprošlé) / archiv (po platnosti)
        if plat == "platne" and p.get("_archiv"):
            continue
        if plat == "archiv" and not p.get("_archiv"):
            continue
        out.append(p)
    return out


def _pripady_csv(items):
    """CSV souhrnu případů (BEZ OP/%) — bezpečné pro všechny role. ; oddělovač + BOM."""
    import csv
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";")
    w.writerow(["Číslo", "Název", "Typ", "Oddělení", "Žadatel", "Datum zadání",
                "Stav", "Řádků", "Chyb", "Výsledek", "Storno důvod"])
    for p in items:
        w.writerow([p["cislo"], p["nazev"], TYPY.get(p["typ"], {}).get("nazev", p["typ"]),
                    p["oddeleni"], p["zadavatel_jmeno"], str(p["datum_zadani"]),
                    _STAV_BADGE.get(p["stav"], (p["stav"],))[0], p["pocet_radku"],
                    p["pocet_chyb"], "OK" if p["vysledek_ok"] else "Chyba",
                    p.get("storno_duvod") or ""])
    return ("﻿" + buf.getvalue()).encode("utf-8")


_DETAIL_HLAVICKA = ["Číslo", "Název případu", "Typ", "Žadatel", "Stav", "#", "Kód",
                    "Název karty", "Sortiment popis", "Verdikt", "OP", "Marže", "Detail"]

# Info (NEpočítané) sloupce navíc v detailním exportu „(vč. OP)" — type-aware, z vstup_json.
_DETAIL_EXTRA = {
    "porovnani": [("Zákazník IČO", "zakaznik_ico"), ("Termín od", "termin_od"),
                  ("Termín do", "termin_do"), ("Typy dokladů", "typy_dokladu"),
                  ("Pokračovat v hledání nižší ceny", "pokracovat_nizsi")],
}


def _detail_hlavicka(typ):
    return _DETAIL_HLAVICKA + [lbl for lbl, _ in _DETAIL_EXTRA.get(typ, [])]


_DETAIL_KOD_COL = 6   # index sloupce "Kód" v _DETAIL_HLAVICKA → textový formát v .xlsx


def _detail_text_sloupce(typ):
    """Indexy sloupců detailního exportu psaných jako TEXT (formát '@'):
    Kód + případné Zákazník IČO — kvůli vodicím nulám."""
    sl = [_DETAIL_KOD_COL]
    for i, (_, k) in enumerate(_DETAIL_EXTRA.get(typ, [])):
        if k == "zakaznik_ico":
            sl.append(len(_DETAIL_HLAVICKA) + i)
    return tuple(sl)


def _detail_radek(p, r):
    """Jeden řádek detailního exportu (list), nebo None pro neschválené položky."""
    if r.get("neschvaleno"):   # neschválené položky (částečné schválení) NEjdou do exportu
        return None
    try:
        rr = json.loads(r["marze_json"]) if r["marze_json"] else {}
    except Exception:
        rr = {}
    op_val = r["op"] if r["op"] is not None else rr.get("master_op")
    radek = [p["cislo"], p["nazev"], TYPY.get(p["typ"], {}).get("nazev", p["typ"]),
             p["zadavatel_jmeno"], _STAV_BADGE.get(p["stav"], (p["stav"],))[0],
             r["poradi"], _export_kod(r["kod"]), r["nazev_karty"], r.get("sortiment_popis"),
             r["verdikt"], "" if op_val is None else _pct(op_val), _marze_text(rr), r["duvod"]]
    extra = _DETAIL_EXTRA.get(p["typ"])
    if extra:   # info sloupce IND z vstup_json
        try:
            vst = json.loads(r["vstup_json"]) if r.get("vstup_json") else {}
        except Exception:
            vst = {}
        radek += [(_export_ico(vst.get(k)) if k == "zakaznik_ico"
                   else _xlsx_bunka(vst.get(k))) for _, k in extra]
    return radek


def _detailni_xlsx(items):
    """Detailní .xlsx vč. OP a marží přes víc případů — JEN pro správce."""
    typ = items[0]["typ"] if items else None
    radky = []
    for p in items:
        for r in nacti_radky(p["id"]):
            radek = _detail_radek(p, r)
            if radek is not None:
                radky.append(radek)
    return _xlsx_bytes(_detail_hlavicka(typ), radky, _detail_text_sloupce(typ))


def _radky_xlsx(p, radky):
    """Detailní .xlsx jednoho případu z KONKRÉTNÍCH řádků (respektuje filtr v detailu)."""
    out = [r2 for r2 in (_detail_radek(p, r) for r in radky) if r2 is not None]
    return _xlsx_bytes(_detail_hlavicka(p["typ"]), out, _detail_text_sloupce(p["typ"]))


# „Slepý" per-case export — JEN zadaná data + Stav (verdikt), BEZ OP/marží/Detailu.
# Bezpečný pro VŠECHNY role (office i žadatel ke sdílení). Sloupce přesně dle zadání;
# pořadí = Kód, Název, <vstupní ceny>, Stav. Klíče = pole ve `vstup_json`,
# "_nazev" = název z vstupu/masteru, "_stav" = verdikt řádku. Paima sem nepatří —
# u ní se stahuje rovnou celý nahraný soubor (viz „Stáhnout nahraný soubor").
_SLEPE_SLOUPCE = {
    "porovnani": [("Kód produktu", "kod_produktu"), ("Název", "_nazev"),
                  ("Prodejní cena (PC bez DPH)", "pc_bez_dph"), ("Bonus %", "bonus"),
                  ("Kompenzace", "kompenzace"), ("Zákazník IČO", "zakaznik_ico"),
                  ("Zákazník", "zakaznik"),
                  ("Definice nastavení", "definice"), ("Výchozí cena", "vychozi_cena"),
                  ("Přirážka/Srážka", "prirazka"), ("Termín od", "termin_od"),
                  ("Termín do", "termin_do"), ("Typy dokladů", "typy_dokladu"),
                  ("Pokračovat v hledání nižší ceny", "pokracovat_nizsi")],
    "mimoletak": [("Kód", "kod"), ("Název", "_nazev"), ("CC s DPH", "cc_s_dph"),
                  ("MO s DPH", "mo_s_dph"), ("VO bez DPH", "vo_bez_dph")],
    "webportal": [("Kód", "kod"), ("Název", "_nazev"), ("Akční PC", "akcni_pc"),
                  ("Akční NC", "akcni_nc"), ("Aktuální NC2", "aktualni_nc2")],
    "sklad6":    [("Kód", "kod"), ("Název", "_nazev"), ("PC-akce", "pc_akce"),
                  ("ANC", "anc"), ("NC2", "nc2")],
}
_VERDIKT_PLAIN = {"OK": "OK", "CHYBA": "Chyba", "VYJIMKA": "Výjimka"}


def _export_kod(v):
    """Kód do exportu: VŽDY 8 míst doplněných nulami zleva, jako řetězec.
    Např. 5656655 → '05656655'. Buňka s kódem se v .xlsx zapisuje s textovým
    formátem (viz `_xlsx_bytes`), takže Excel zachová úvodní nuly a neudělá
    z kódu číslo. Prázdný kód → prázdná buňka."""
    nk = cp.normalizuj_kod(v)
    if not nk:
        return ""
    return nk.zfill(8)


def _export_ico(v):
    """IČ s vodicími nulami (do náhledu i exportu): číselná část PŘED případným
    písmenem se doplní zleva nulami na 8 míst — '859657' → '00859657',
    '859657A1' → '00859657A1'. Čistě číselné IČ s 8+ místy a hodnoty v jiném
    tvaru zůstávají beze změny. Prázdné → prázdný řetězec. V .xlsx se buňka
    zapisuje s textovým formátem, aby Excel nuly zachoval."""
    nk = cp.normalizuj_kod(v)
    if not nk:
        return ""
    m = re.match(r"^(\d+)([A-Za-z].*)?$", nk)
    if not m:
        return nk
    return m.group(1).zfill(8) + (m.group(2) or "")


def _xlsx_bytes(hlavicka, radky, text_sloupce=(), nazev_listu=None):
    """Sestaví .xlsx (bytes): hlavička + řádky. Sloupce v `text_sloupce` (0-based
    indexy) se zapíšou jako TEXT (formát '@'), aby Excel zachoval úvodní nuly a
    neudělal z nich číslo (typicky sloupec s kódem). Datumové buňky dostanou CZ
    formát 'd.m.rrrr h:mm' (jinak by openpyxl použil ISO 'rrrr-mm-dd hh:mm:ss').
    `nazev_listu` pojmenuje list (některé importy jméno listu vyžadují)."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    if nazev_listu:
        ws.title = nazev_listu
    text_sloupce = set(text_sloupce)
    ws.append(list(hlavicka))
    for r in radky:
        ws.append(list(r))
    for ci in text_sloupce:   # textový formát na celém sloupci (vč. hlavičky)
        for bunka in ws.iter_cols(min_col=ci + 1, max_col=ci + 1):
            for c in bunka:
                c.number_format = "@"
    for row in ws.iter_rows(min_row=2):   # CZ formát data bez času (např. 31.12.2026)
        for c in row:
            if isinstance(c.value, (datetime.datetime, datetime.date)):
                c.number_format = "dd.mm.yyyy"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
XLS_MIME = "application/vnd.ms-excel"


def _xls_bytes(hlavicka, radky, text_sloupce=(), nazev_listu=None):
    """Starý binární .xls (Excel 97-2003, formát BIFF8) — pro importy, které
    novější .xlsx nepřečtou. Sloupce se chovají stejně jako u `_xlsx_bytes`:
    `text_sloupce` dostanou formát '@' (zachová vodicí nuly), data CZ formát.
    Vyžaduje knihovnu xlwt (pip install xlwt)."""
    try:
        import xlwt
    except ImportError:
        raise RuntimeError("Pro export do formátu Excel 97-2003 (.xls) je na serveru "
                           "potřeba knihovna xlwt (pip install xlwt). Zatím použijte "
                           "formát .xlsx nebo Excel 2003 XML.")
    # Tvrdé limity BIFF8 — radši srozumitelná hláška než rozbitý soubor.
    if len(radky) + 1 > 65536:
        raise RuntimeError(f"Formát .xls pojme 65 536 řádků, export jich má "
                           f"{len(radky) + 1}. Použijte formát .xlsx.")
    if len(hlavicka) > 256:
        raise RuntimeError("Formát .xls pojme 256 sloupců, export jich má víc.")
    wb = xlwt.Workbook(encoding="utf-8")
    ws = wb.add_sheet((nazev_listu or "List1")[:31])
    st_text = xlwt.easyxf(num_format_str="@")
    st_datum = xlwt.easyxf(num_format_str="DD.MM.YYYY")
    st_norm = xlwt.Style.default_style
    text_sloupce = set(text_sloupce)
    for ci, lbl in enumerate(hlavicka):
        ws.write(0, ci, lbl, st_text if ci in text_sloupce else st_norm)
    for ri, r in enumerate(radky, start=1):
        for ci, v in enumerate(r):
            if ci in text_sloupce:                # kód apod. vždy jako text
                ws.write(ri, ci, "" if v is None else str(v), st_text)
            elif isinstance(v, (datetime.datetime, datetime.date)):
                ws.write(ri, ci, v, st_datum)
            else:
                ws.write(ri, ci, "" if v is None else v, st_norm)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xml_ocisti(s):
    """Odstraní řídicí znaky, které XML nepovoluje (v datech se objeví jen omylem,
    ale Excel by na nich celý soubor odmítl otevřít)."""
    return "".join(c for c in s if c in "\t\n\r" or ord(c) >= 32)


def _xls_xml_bytes(hlavicka, radky, text_sloupce=(), nazev_listu=None):
    """Excel 2003 XML Spreadsheet (SpreadsheetML) — ukládá se s příponou .xls.
    Nepotřebuje žádnou knihovnu (píše se přes stdlib). Otevře ho Excel i
    LibreOffice; importy, které čekají binárku, ho ale nepřečtou — proto je to
    v dialogu jen jedna z voleb, ne výchozí."""
    from xml.sax.saxutils import escape, quoteattr
    text_sloupce = set(text_sloupce)

    def _bunka(v, ci):
        if ci in text_sloupce:   # text i pro čísla — kvůli vodicím nulám
            return ('<Cell ss:StyleID="sT"><Data ss:Type="String">'
                    f'{escape(_xml_ocisti("" if v is None else str(v)))}</Data></Cell>')
        if isinstance(v, (datetime.datetime, datetime.date)):
            d = v if isinstance(v, datetime.datetime) else \
                datetime.datetime(v.year, v.month, v.day)
            return ('<Cell ss:StyleID="sD"><Data ss:Type="DateTime">'
                    f'{d.strftime("%Y-%m-%dT%H:%M:%S.000")}</Data></Cell>')
        if isinstance(v, bool):   # dřív než int — bool je podtyp int
            return f'<Cell><Data ss:Type="Number">{int(v)}</Data></Cell>'
        if isinstance(v, (int, float)):
            return f'<Cell><Data ss:Type="Number">{v}</Data></Cell>'
        s = _xml_ocisti("" if v is None else str(v))
        if not s:
            return "<Cell/>"
        return f'<Cell><Data ss:Type="String">{escape(s)}</Data></Cell>'

    out = ['<?xml version="1.0" encoding="UTF-8"?>',
           '<?mso-application progid="Excel.Sheet"?>',
           '<Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet"'
           ' xmlns:o="urn:schemas-microsoft-com:office:office"'
           ' xmlns:x="urn:schemas-microsoft-com:office:excel"'
           ' xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet">',
           '<Styles>'
           '<Style ss:ID="Default" ss:Name="Normal">'
           '<Alignment ss:Vertical="Bottom"/></Style>'
           '<Style ss:ID="sT"><NumberFormat ss:Format="@"/></Style>'
           '<Style ss:ID="sD"><NumberFormat ss:Format="dd\\.mm\\.yyyy"/></Style>'
           '</Styles>',
           f'<Worksheet ss:Name={quoteattr((nazev_listu or "List1")[:31])}><Table>']
    out.append("<Row>" + "".join(_bunka(h, ci) for ci, h in enumerate(hlavicka)) + "</Row>")
    for r in radky:
        out.append("<Row>" + "".join(_bunka(v, ci) for ci, v in enumerate(r)) + "</Row>")
    out.append("</Table></Worksheet></Workbook>")
    return "\n".join(out).encode("utf-8")


def _xlsx_datum(s):
    """Řetězec uložený z data (json default=str, např. '2026-12-31 00:00:00' nebo
    '2026-12-31') → datetime, jinak None. Pro .xlsx, aby se datum zapsalo jako
    skutečné datum (a dostalo CZ formát), ne jako ISO text. Konzervativní —
    bere jen ISO tvary, ostatní řetězce (kódy, IČO, '06518605') nechá být."""
    if not isinstance(s, str):
        return None
    t = s.strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.datetime.strptime(t, fmt)
        except ValueError:
            continue
    return None


def _nahled_datum(v):
    """Hodnota do náhledové tabulky: datum (i uložené jako ISO text) → čisté CZ
    datum bez času 'dd.mm.rrrr' (sjednoceno s exportem, např. 31.12.2026).
    Ostatní hodnoty vrací beze změny."""
    d = v if isinstance(v, (datetime.datetime, datetime.date)) else _xlsx_datum(v)
    if d is None:
        return v
    return f"{d.day:02d}.{d.month:02d}.{d.year}"


def _xlsx_bunka(v):
    """Hodnota do .xlsx: čísla zůstávají čísly (Excel je zobrazí dle locale),
    celá čísla bez koncového .0, data (i uložená jako ISO text) → datetime
    (CZ formát řeší `_xlsx_bytes`), ostatní text beze změny, None → prázdná buňka."""
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v
    if isinstance(v, str):
        d = _xlsx_datum(v)
        return d if d is not None else v
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def _prirazka_procento(raw):
    """IND ceny, „Přirážka/Srážka": je-li zadaná v procentech, vrať desetinný zlomek
    (číslo/100): '5%' → 0.05, '-2,5 %' → -0.025, '+10%' → 0.1.
    Kč hodnoty (bez „%"), prázdné a neparsovatelné → None (nemění se)."""
    s = ("" if raw is None else str(raw)).strip()
    if not s or "%" not in s:
        return None
    try:
        return float(s.replace("%", "").replace(" ", "").replace(",", ".")) / 100.0
    except ValueError:
        return None


def _cz_cislo(v):
    """Číslo v českém formátu (desetinná čárka), bez zbytečných koncových nul.
    0.05 → '0,05', 1.0 → '1', -0.025 → '-0,025'."""
    txt = f"{v:.6f}".rstrip("0").rstrip(".")
    if txt in ("", "-", "-0"):
        txt = "0"
    return txt.replace(".", ",")


def _prirazka_nahled(raw):
    """Hodnota „Přirážka/Srážka" do NÁHLEDU: procenta ÷100 BEZ znaku % (5% → '0,05');
    Kč/text beze změny."""
    p = _prirazka_procento(raw)
    return _cz_cislo(p) if p is not None else _nahled_datum(raw)


def _prirazka_export(raw):
    """Hodnota „Přirážka/Srážka" do EXPORTU: procenta jako číslo ÷100 (5% → 0.05,
    Excel zobrazí dle locale jako „0,05"); Kč/text beze změny."""
    p = _prirazka_procento(raw)
    return p if p is not None else _xlsx_bunka(raw)


def _ncf_slepe_xlsx(p, radky, srovnani=False):
    """Export NC faktury. Vždy zadaná data + Stav (bezpečné pro všechny role).
    `srovnani=True` (office nákup/správce) přidá sloupce DNC3 (NC3) a Detail —
    master NC3 nevidí žadatel. Podtyp ('nc'/'fc') určuje sadu sloupců."""
    sub = _ncf_subtyp(radky)
    if sub == "fc":
        spec = [("Kód výrobku", "kod"), ("Název výrobku", "nazev"),
                ("NC-IMPORT", "nc_import"), ("NC datum od", "nc_datum_od"),
                ("NC datum do", "nc_datum_do"), ("FC-IMPORT", "fc_import"),
                ("FC datum od", "fc_datum_od"), ("FC datum do", "fc_datum_do")]
        ma_dnc = False
    else:
        spec = [("Kód výrobku", "kod"), ("Název výrobku", "nazev"),
                ("Cena pro nastavení", "cena_nastaveni"), ("Datum od", "datum_od"),
                ("Datum do", "datum_do"), ("Název IND. ceny", "nazev_ind")]
        ma_dnc = True
    hlavicka = ([lbl for lbl, _ in spec]
                + (["DNC3 (NC3)"] if (ma_dnc and srovnani) else [])
                + (["Detail"] if srovnani else []) + ["Stav"])
    out_radky = []
    for r in radky:
        if r.get("neschvaleno"):
            continue
        try:
            vst = json.loads(r["vstup_json"]) if r["vstup_json"] else {}
        except Exception:
            vst = {}
        try:
            rr = json.loads(r["marze_json"]) if r["marze_json"] else {}
        except Exception:
            rr = {}
        out = [(_export_kod(vst.get(src)) if src == "kod" else _xlsx_bunka(vst.get(src)))
               for _, src in spec]
        if ma_dnc and srovnani:
            out.append(_xlsx_bunka(rr.get("dnc3")))
        if srovnani:
            out.append(r["duvod"])
        out.append(_VERDIKT_PLAIN.get(r["verdikt"], r["verdikt"]))
        out_radky.append(out)
    return _xlsx_bytes(hlavicka, out_radky, (0,))   # sloupec Kód jako text


def _radky_slepe_xlsx(p, radky, srovnani=False):
    """„Slepý" .xlsx jednoho případu: jen zadaná data + Stav, BEZ OP/marží/Detailu.
    Sloupec s kódem je TEXT (8 míst, úvodní nuly). Vrací bytes, nebo None pro typ
    bez specifikace (paima)."""
    if p["typ"] == "ncfaktura":
        return _ncf_slepe_xlsx(p, radky, srovnani)
    spec = _SLEPE_SLOUPCE.get(p["typ"])
    if not spec:
        return None
    spec = spec + [("Stav", "_stav")]
    # Kód i IČO jako TEXT (formát '@'), aby Excel zachoval vodicí nuly.
    kod_cols = [i for i, (_, src) in enumerate(spec)
                if src in ("kod", "kod_produktu", "zakaznik_ico")]
    out_radky = []
    for r in radky:
        if r.get("neschvaleno"):   # neschválené položky nejsou v exportu
            continue
        try:
            vst = json.loads(r["vstup_json"]) if r["vstup_json"] else {}
        except Exception:
            vst = {}
        out = []
        for _, src in spec:
            if src == "_nazev":
                v = vst.get("nazev") or r.get("nazev_karty")
                out.append(_xlsx_bunka(v))
            elif src == "_stav":
                out.append(_xlsx_bunka(_VERDIKT_PLAIN.get(r["verdikt"], r["verdikt"])))
            elif src == "prirazka":
                out.append(_prirazka_export(vst.get(src)))   # procenta ÷100 (5% → 0,05)
            elif src in ("kod", "kod_produktu"):
                out.append(_export_kod(vst.get(src)))
            elif src == "zakaznik_ico":
                out.append(_export_ico(vst.get(src)))   # vodicí nuly IČ
            else:
                out.append(_xlsx_bunka(vst.get(src)))
        out_radky.append(out)
    return _xlsx_bytes([lbl for lbl, _ in spec], out_radky, kod_cols)


# ── Export „do skladu" (IND ceny) ────────────────────────────────────────────
# Formát 1:1 dle vzoru z Pekosu (list „IC"). Z případu se mění jen KLIC (kód),
# SLEVAKC (cena), POZNAMKA (název) a termíny; ostatní sloupce jsou konstanty
# dohodnuté se skladem. SKUPINA se zadává v dialogu při exportu.
_SKLAD_HLAVICKA = ["SKUPINA", "TYPKLICE", "KLIC", "DATUMOD", "DATUMDO", "SLEVAPROC",
                   "SLEVAKC", "TYPSLEVY", "POZNAMKA", "TYPZAOKR", "PORADI",
                   "ODCASTKY", "PREPINACE", "TYPOD", "TYPDO"]


def _sklad_cena(v):
    """PC bez DPH do sloupce SLEVAKC: číslo (čárka i tečka jako oddělovač).
    Procenta a neparsovatelné hodnoty do skladu nepatří → prázdná buňka.
    Stejná funkce hlídá PC už při nahrání (_zpracuj_rows), takže se sem
    procento nedostane."""
    return cp.parse_cena_kc(v)


# „Definice nastavení" se píše ručně ve tvaru „PC39 + 5%" — do SLEVAPROC patří
# jen to číslo za plusem. Když v zápisu plus není (třeba „PC39" nebo „PC39 - 5%"),
# zůstane buňka prázdná.
_SLEVAPROC_RE = re.compile(r"\+\s*(\d+(?:[.,]\d+)?)")


def _sklad_slevaproc(v):
    """Procento za „+" z definice nastavení. Vrací float, nebo None (= prázdno)."""
    s = ("" if v is None else str(v)).replace("\xa0", " ")
    m = _SLEVAPROC_RE.search(s)
    if not m:
        return None
    try:
        return float(m.group(1).replace(",", "."))
    except ValueError:
        return None


# Volby v dialogu Export do skladu: klíč → (popisek, přípona, MIME, funkce).
# Výchozí je .xlsx; .xls varianty jsou pro starší importy, které novější formát
# neumí přečíst. „XML" je Excel 2003 XML Spreadsheet — taky s příponou .xls,
# ale uvnitř text, ne binárka.
SKLAD_FORMATY = {
    "xlsx": ("Excel .xlsx (klasický)", "xlsx", XLSX_MIME, lambda *a, **k: _xlsx_bytes(*a, **k)),
    "xls": ("Excel 97-2003 .xls (binární)", "xls", XLS_MIME, lambda *a, **k: _xls_bytes(*a, **k)),
    "xml": ("Excel 2003 XML (.xls)", "xls", XLS_MIME, lambda *a, **k: _xls_xml_bytes(*a, **k)),
}


def _sklad_radky(radky, skupina):
    """Řádky případu → řádky ve formátu skladu (bez ohledu na cílový formát souboru).
    Bere jen schválené položky."""
    out_radky = []
    for r in radky:
        if r.get("neschvaleno"):   # neschválené položky nejsou v exportu
            continue
        try:
            vst = json.loads(r["vstup_json"]) if r["vstup_json"] else {}
        except Exception:
            vst = {}
        cena = _sklad_cena(vst.get("pc_bez_dph"))
        slevaproc = _sklad_slevaproc(vst.get("definice"))
        out_radky.append([
            skupina,
            "kód",
            _export_kod(vst.get("kod_produktu")),           # KLIC — text, vodicí nuly
            _xlsx_bunka(vst.get("termin_od")),
            _xlsx_bunka(vst.get("termin_do")),
            "" if slevaproc is None else slevaproc,         # SLEVAPROC
            "" if cena is None else cena,                   # SLEVAKC
            "cena Kč",                                      # TYPSLEVY
            _xlsx_bunka(vst.get("nazev") or r.get("nazev_karty")),
            9, 0, 0, 17152, 0, 0,   # TYPZAOKR/PORADI/ODCASTKY/PREPINACE/TYPOD/TYPDO
        ])
    return out_radky


def _sklad_soubor(radky, skupina, format="xlsx"):
    """Soubor pro import do skladu (typ „porovnani" = IND ceny), list „IC".
    Vrací (bytes, přípona, MIME). `format` = klíč ze `SKLAD_FORMATY`."""
    _popis, pripona, mime, zapis = SKLAD_FORMATY.get(format) or SKLAD_FORMATY["xlsx"]
    data = zapis(_SKLAD_HLAVICKA, _sklad_radky(radky, skupina), [2], nazev_listu="IC")
    return data, pripona, mime


PRIPADU_NA_STRANKU = 40   # max karet případů na jednu stránku seznamu


def _sub_view_typ(typ, user_id, user_name, prava):
    cfg = TYPY[typ]
    if _muze_zadat(typ, prava):
        _formular_novy_pripad(typ, user_id, user_name, cfg["oddeleni"])

    # Aktuální pobočka přihlášeného uživatele (pro volitelný filtr „Zadání z pobočky")
    _moje_pobocka = None
    try:
        _pc = intranet_data.get_db_connection()
        if _pc:
            _pcu = _pc.cursor()
            _pcu.execute("SELECT pobocka FROM user WHERE iduser=%s", (user_id,))
            _pr = _pcu.fetchone()
            _moje_pobocka = _pr[0] if _pr else None
            _pcu.close()
            _pc.close()
    except Exception:
        _moje_pobocka = None
    _stav_pobocka = {"on": bool(app.storage.user.get("ind_filtr_pobocka", False))}

    def _nacti_vsechny():
        # Filtr pobočky pouze pro IND (porovnani), jen když je zapnutý a uživatel pobočku má
        _fp = _moje_pobocka if (typ == "porovnani" and _stav_pobocka["on"] and _moje_pobocka) else None
        data = nacti_pripady(typ, prava, user_id, filtr_pobocka=_fp)
        if typ == "porovnani":   # IND: dopočítej platnost „do" (nejzazší termín do) + archiv
            _dnes = datetime.date.today()
            _ids = [p["id"] for p in data]
            _mapa_plat = platnosti_do_pro_pripady(_ids)
            _mapa_kontrola = kontrola_stav_pro_pripady(_ids)   # agregovaný sloupec „Kontrola"
            for p in data:
                _pd = _mapa_plat.get(p["id"])
                p["_platnost_do"] = _pd
                p["_archiv"] = bool(_pd and _pd < _dnes)
                p["_kontrola"] = _mapa_kontrola.get(p["id"])   # 'ok'|'chyba'|None (None = bez definice)
        return data

    vsechny = _nacti_vsechny()
    nadpis = ("Historie případů" if _vidi_vsechny_pripady(prava)
              else "Případy oddělení" if (_vidi_oddeleni(prava) or _schval_odd_slugs(prava)
                                          or _vedouci_odd_slugs(prava))
              else "Moje případy")
    ui.label(nadpis).classes("text-lg font-bold text-gray-700 mb-1")
    if not vsechny:
        with ui.column().classes("items-center py-10 gap-2 w-full"):
            ui.icon("inbox", size="3rem", color="grey-4")
            ui.label("Zatím žádné případy.").classes("text-gray-400")
        return

    flt = {"nazev": "", "zadavatel": "", "stav": "", "od": "", "do": "", "platnost": ""}
    strana = {"p": 1}   # stránkování seznamu (max PRIPADU_NA_STRANKU karet na stránku)
    zadavatele = sorted({p["zadavatel_jmeno"] for p in vsechny if p["zadavatel_jmeno"]})
    stav_opts = {"": "— stav —", **{s: _STAV_BADGE.get(s, (s,))[0]
                                    for s in sorted({p["stav"] for p in vsechny})}}

    @ui.refreshable
    def _seznam():
        items = _filtruj(vsechny, flt)
        with ui.row().classes("w-full items-center justify-between max-w-5xl mb-1"):
            ui.label(f"{len(items)} z {len(vsechny)} případů").classes("text-xs text-gray-500")

            with ui.row().classes("gap-1"):
                def _export():
                    ui.download.content(_pripady_csv(items), f"cenopripady_{typ}.csv", "text/csv")

                ui.button("Export souhrnu (CSV)", icon="download", on_click=_export) \
                    .props("flat no-caps").classes("text-blue-600")
                if _vidi_op(prava):
                    async def _export_det():
                        data = await asyncio.to_thread(_detailni_xlsx, items)
                        ui.download.content(data, f"cenopripady_{typ}_detail.xlsx", XLSX_MIME)

                    ui.button("Detailní export (vč. OP)", icon="download", on_click=_export_det) \
                        .props("flat no-caps").classes("text-rose-600")
        if not items:
            ui.label("Žádné případy neodpovídají filtru.").classes("text-gray-400 italic")
            return
        # Stránkování: max PRIPADU_NA_STRANKU karet na stránku
        pocet_stran = max(1, -(-len(items) // PRIPADU_NA_STRANKU))
        strana["p"] = min(strana["p"], pocet_stran)
        _zac = (strana["p"] - 1) * PRIPADU_NA_STRANKU
        strana_items = items[_zac:_zac + PRIPADU_NA_STRANKU]
        with ui.column().classes("w-full gap-2 max-w-5xl"):
            for p in strana_items:
                _karta_pripadu(p, user_id, user_name, prava)
        if pocet_stran > 1:
            def _zmen_stranku(p):
                strana["p"] = p
                _seznam.refresh()
            with ui.row().classes("w-full items-center justify-center gap-2 mt-2 max-w-5xl"):
                ui.label(f"{_zac + 1}–{_zac + len(strana_items)} z {len(items)}").classes("text-xs text-gray-400 mr-2")
                ui.pagination(1, pocet_stran, direction_links=True, value=strana["p"],
                              on_change=lambda e: _zmen_stranku(int(e.value))).props("max-pages=7 boundary-numbers")

    def _zmen(klic, e):
        flt[klic] = e.value or ""
        strana["p"] = 1   # změna filtru → zpět na první stránku
        _seznam.refresh()

    def _prepni_pobocku(e):
        # Zapnutí/vypnutí filtru pobočky → znovunačtení (zúží/rozšíří v rámci již viditelných)
        _stav_pobocka["on"] = bool(e.value)
        app.storage.user["ind_filtr_pobocka"] = _stav_pobocka["on"]
        nonlocal vsechny
        vsechny = _nacti_vsechny()
        strana["p"] = 1
        _seznam.refresh()

    # flex-nowrap = filtry zůstanou v jedné rovině (jinak se „platnost" zalomí na další
    # řádek); šířky zmenšené tak, aby se vše vešlo do max-w-5xl.
    with ui.row().classes("w-full gap-2 flex-nowrap items-end max-w-5xl mb-2"):
        ui.input("Název", on_change=lambda e: _zmen("nazev", e)) \
            .props("outlined dense clearable").classes("w-40")
        ui.select({"": "— žadatel —", **{z: z for z in zadavatele}}, value="",
                  on_change=lambda e: _zmen("zadavatel", e)).props("outlined dense").classes("w-40")
        ui.select(stav_opts, value="", on_change=lambda e: _zmen("stav", e)) \
            .props("outlined dense").classes("w-40")
        ui.input("Od", on_change=lambda e: _zmen("od", e)) \
            .props("outlined dense type=date").classes("w-40")
        ui.input("Do", on_change=lambda e: _zmen("do", e)) \
            .props("outlined dense type=date").classes("w-40")
        if typ == "porovnani":   # IND: filtr dle platnosti „do" (archiv = po platnosti)
            # ml-auto = odsazení doprava (zarovná „platnost" k pravému okraji řádku)
            ui.select({"": "— platnost —", "platne": "Platné (neprošlé)",
                       "archiv": "Archiv (po platnosti)"}, value="",
                      on_change=lambda e: _zmen("platnost", e)) \
                .props("outlined dense").classes("w-44 ml-auto")
            # Volitelný filtr „Zadání z pobočky" — jen když uživatel pobočku má; zužuje seznam
            if _moje_pobocka:
                ui.switch("Zadání z pobočky", value=_stav_pobocka["on"],
                          on_change=_prepni_pobocku).props("dense").classes("ml-2")

    _seznam()


# ============================================================================
# KONTROLNÍ DATA LETÁKŮ (samostatná dlaždice) — fáze 1
# ============================================================================
LETAKY_KANALY = {"VO": "Velkoobchod", "MO": "Maloobchod", "SP": "Svět potravin"}

# Druhy sestav. Zpracování i sloupce jsou stejné, liší se jen práva a seznam:
# finální sestavy vidí POUZE uživatel s právy „…_letaky_final" (nebo správce).
LETAKY_DRUHY = {"kontrolni": "Kontrolní sestava", "finalni": "Finální sestava"}

# Pravidla zvýraznění počítaných sloupců (dle podmíněného formátování „Označené sestavy").
# Seznam (operátor, práh, barva); pořadí = priorita (první platné vyhraje, jako „zastavit
# pokud platí" v Excelu). Hodnoty mimo všechna pravidla zůstávají bez barvy.
_LETAKY_PRAVIDLA = {
    "nc3_anc":        [(">", 0, "blue"), ("<", 0, "red")],
    "nc_anc":         [(">", 0.01, "red"), ("<", -0.25, "orange")],
    "rozdil_korekce": [(">", 0.03, "yellow"), ("<", -0.03, "orange")],
    "akcni_prirazka": [],   # bez zvýraznění (dle zadání „nic")
}
# Finální sestava má vlastní podmíněné formátování (přebírá se z Excelu 1:1).
# Sloupce, které tu nejsou, se řídí pravidly Kontrolní sestavy.
_LETAKY_PRAVIDLA_FINAL = {
    # NC3-ANC: přesně 0 = zeleně, pokles pod −1 % = červeně. Mezi tím (a nad 0) bez barvy.
    "nc3_anc": [("=", 0, "green"), ("<", -0.01, "excel_red")],
    # NC/ANC („Sleva NC"): sleva hlubší než 25 % = žlutě. Jinak nic — pravidlo
    # kontrolní sestavy „> 1 % červeně" se do finální nepřenáší.
    "nc_anc": [("<", -0.25, "amber")],
    # Akční přirážka („Přír.%"): pod −3 % růžově, jinak každá záporná žlutě.
    # Excel má prahy v procentních bodech (−3, 0), modul drží zlomek → −0,03.
    # Pořadí je i priorita: intervaly se překrývají, první platné pravidlo vyhraje.
    "akcni_prirazka": [("<", -0.03, "magenta"), ("<", 0, "amber")],
    # rozdíl korekce v % tu schválně není — nevyjmenované sloupce si drží
    # formátování kontrolní sestavy.
}
_LETAKY_BARVA = {"red": "#dc2626", "blue": "#2563eb", "orange": "#ea580c",
                 "yellow": "#facc15", "purple": "#7c3aed",
                 # odstíny z Excelu (podmíněné formátování „Označené sestavy")
                 "green": "#c6efce", "excel_red": "#ffc7ce", "amber": "#ffc000",
                 "magenta": "#ff00ff"}
# barva písma na barevné výplni buňky (žlutá = tmavé písmo, ostatní = bílé)
_LETAKY_TEXT = {"red": "#ffffff", "blue": "#ffffff", "orange": "#ffffff",
                "yellow": "#1e293b", "purple": "#ffffff",
                "green": "#006100", "excel_red": "#9c0006", "amber": "#000000",
                "magenta": "#000000"}
_LETAKY_BARVA_LABEL = {"red": "červené", "blue": "modré", "orange": "oranžové",
                       "yellow": "žluté", "purple": "fialové",
                       "green": "zelené", "excel_red": "červené", "amber": "žluté",
                       "magenta": "růžové"}


def _letaky_pravidla(col, druh="kontrolni"):
    """Pravidla zvýraznění sloupce pro daný druh sestavy."""
    if _letaky_druh(druh) == "finalni" and col in _LETAKY_PRAVIDLA_FINAL:
        return _LETAKY_PRAVIDLA_FINAL[col]
    return _LETAKY_PRAVIDLA.get(col, [])


def _letaky_klasifikuj(col, val, druh="kontrolni"):
    """Vrátí barvu (klíč) dle pravidel sloupce, nebo None."""
    if val is None:
        return None
    for op, prah, barva in _letaky_pravidla(col, druh):
        if (op == ">" and val > prah) or (op == "<" and val < prah) or \
           (op == ">=" and val >= prah) or (op == "<=" and val <= prah) or \
           (op == "=" and abs(val - prah) < 1e-9):
            return barva
    return None
_LETAKY_K2_ZNAK = {"VO": "4", "MO": "C", "SP": "3"}     # požadovaný znak kanálu v K2
_LETAKY_K2_ZAKAZ = ("-", "9", "0")                       # zakázané znaky v K2

# Sloupce ze sjetiny, kde NULA = chyba (červená buňka), po kanálech.
# VO: „Akční NC" (ANC, sloupec Q) + „Akční PC" (PC bez daně, sloupec T).
# MO/SP: „Akční NC" (ANC, sloupec Q) + „Akční PC s DPH" (PC vč. DPH, sloupec U).
_LETAKY_NULA_SLOUPCE = {
    "VO": ("Akční NC", "Akční PC"),
    "MO": ("Akční NC", "Akční PC s DPH"),
    "SP": ("Akční NC", "Akční PC s DPH"),
}
# Výplň nulových buněk = stejná světlá červená jako u chybného K2 (syté barvy z palety
# _LETAKY_BARVA zůstávají vyhrazené dopočítaným sloupcům).
_LETAKY_NULA_FILL = "#fee2e2"
_LETAKY_NULA_TEXT = "#b91c1c"

# Finální sestava navíc hlídá zápornou cenu („Akce s DPH" < −0,01 Kč) — sytě červeně.
_LETAKY_ZAPOR_SLOUPCE = ("Akční PC s DPH",)
_LETAKY_ZAPOR_PRAH = -0.01
_LETAKY_ZAPOR_FILL = "#ff0000"
_LETAKY_ZAPOR_TEXT = "#000000"


def _letaky_nula_chyba(kanal, hlavicka, hodnota):
    """True = buňka má být červená: sloupec je pro daný kanál hlídaný na nulu a hodnota je 0.
    Prázdná/nečíselná buňka se nebarví (nula musí být opravdu zapsaná)."""
    if hlavicka not in _LETAKY_NULA_SLOUPCE.get(kanal, ()):
        return False
    cislo = cp.parse_cislo(hodnota)
    return cislo is not None and abs(cislo) < 1e-9


def _letaky_hlidane_sloupce(kanal, druh="kontrolni"):
    """Hlavičky ze sjetiny, které můžou dostat barvu (pro cellStyle a filtr)."""
    sl = list(_LETAKY_NULA_SLOUPCE.get(kanal, ()))
    if _letaky_druh(druh) == "finalni":
        sl += [h for h in _LETAKY_ZAPOR_SLOUPCE if h not in sl]
    return tuple(sl)


def _letaky_bunka_styl(kanal, hlavicka, hodnota, druh="kontrolni"):
    """Barva buňky ze sjetiny jako (výplň, písmo), nebo None. Nula v hlídaném sloupci
    platí pro oba druhy; záporná cena jen pro finální sestavu."""
    if _letaky_nula_chyba(kanal, hlavicka, hodnota):
        return (_LETAKY_NULA_FILL, _LETAKY_NULA_TEXT)
    if _letaky_druh(druh) == "finalni" and hlavicka in _LETAKY_ZAPOR_SLOUPCE:
        cislo = cp.parse_cislo(hodnota)
        if cislo is not None and cislo < _LETAKY_ZAPOR_PRAH:
            return (_LETAKY_ZAPOR_FILL, _LETAKY_ZAPOR_TEXT)
    return None


def _letaky_druh(v):
    """Normalizace druhu sestavy (staré sestavy bez hodnoty = kontrolní)."""
    return "finalni" if v == "finalni" else "kontrolni"


def _letaky_sufix(druh):
    """Druh → přípona práv. Finální sestavy mají vlastní kopii práv (…_letaky_final)."""
    return "_final" if _letaky_druh(druh) == "finalni" else ""


def _letaky_pristup(p, druh=None):
    """Smí do dlaždice „Kontrolní data letáků" (vč. čtenáře — jen prohlížení + stažení).
    Bez `druh` = smí aspoň na jeden druh sestav."""
    if druh is None:
        return any(_letaky_pristup(p, d) for d in LETAKY_DRUHY)
    return (_letaky_muze_nahrat(p, druh)
            or f"cenopripad_ctenar_letaky{_letaky_sufix(druh)}" in p)


def _letaky_muze_nahrat(p, druh="kontrolni"):
    """Smí nahrávat sestavy daného druhu a psát komentáře (žadatel/office/správce). Čtenář NE."""
    s = _letaky_sufix(druh)
    return _je_spravce(p) or any(f"cenopripad_{r}_letaky{s}" in p
                                 for r in ("office", "zadatel"))


def _letaky_office(p, druh="kontrolni"):
    """Smí měnit stav na „Zpracováno" a mazat (Office – Letáková kontrola / správce)."""
    return _je_spravce(p) or f"cenopripad_office_letaky{_letaky_sufix(druh)}" in p


def _letaky_vidi_vse(p, druh="kontrolni"):
    """Vidí všechny sestavy daného druhu (ne jen své) — office, čtenář, správce."""
    s = _letaky_sufix(druh)
    return _je_spravce(p) or any(f"cenopripad_{r}_letaky{s}" in p
                                 for r in ("office", "ctenar"))


def _letaky_dostupne_druhy(p):
    """Druhy sestav, na které uživatel vidí (pořadí dle LETAKY_DRUHY)."""
    return [d for d in LETAKY_DRUHY if _letaky_pristup(p, d)]


def _excel_sloupec(idx):
    """Index sloupce (0-based) → písmeno v Excelu. 0 → A, 25 → Z, 26 → AA."""
    pismeno = ""
    idx += 1
    while idx > 0:
        idx, zbytek = divmod(idx - 1, 26)
        pismeno = chr(65 + zbytek) + pismeno
    return pismeno


def _letaky_pismena(header):
    """Z řádku hlavičky udělá mapu {název sloupce: písmeno v Excelu}. Duplicitní název
    si drží první výskyt (stejně jako parser, kde pozdější přepis nedává smysl)."""
    mapa = {}
    for j, h in enumerate(header):
        nazev = str(h).strip() if h is not None else ""
        if nazev and nazev not in mapa:
            mapa[nazev] = _excel_sloupec(j)
    return mapa


def _k2_text(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _k2_chyba(k2, kanal):
    """True = K2 je špatně (červeně): chybí znak kanálu, NEBO obsahuje zakázané, NEBO prázdné."""
    s = (k2 or "").strip()
    if not s:
        return True
    znak = _LETAKY_K2_ZNAK.get(kanal, "")
    if znak and znak.lower() not in s.lower():
        return True
    if any(z in s for z in _LETAKY_K2_ZAKAZ):
        return True
    return False


def _letaky_parse(raw_bytes, filename, kanal):
    """dohody-karty_v_akci.(xls/xlsx) → (radky, pismena_sloupcu, chyba|None). Páruje sloupce
    dle názvu v hlavičce (řádek s „Dodavatel" + „Kód"). Dopočítá pole a kontrolu K2 pro daný
    kanál. `pismena_sloupcu` = {hlavička: písmeno sloupce v Excelu} pro zobrazení v tabulce."""
    try:
        rows = _nacti_rows(raw_bytes, filename)
    except Exception as e:
        return None, None, f"Soubor nelze načíst: {e}"
    if not rows:
        return None, None, "Soubor je prázdný."
    hdr_idx = None
    for i, row in enumerate(rows[:40]):
        norm = {_norm(c) for c in row if c is not None and str(c).strip()}
        if "dodavatel" in norm and ("kód" in norm or "kod" in norm):
            hdr_idx = i
            break
    if hdr_idx is None:
        return None, None, "Nenalezena hlavička (chybí sloupce „Dodavatel“ a „Kód“)."
    header = [str(c).strip() if c is not None else "" for c in rows[hdr_idx]]
    pismena = _letaky_pismena(header)
    _num = cp.parse_cislo
    out = []
    for r in rows[hdr_idx + 1:]:
        if r is None or all(c is None or str(c).strip() == "" for c in r):
            continue
        d = {}
        for j, h in enumerate(header):
            if h:
                d[h] = r[j] if j < len(r) else None
        kod = _k2_text(d.get("Kód") or d.get("Kod"))
        if not kod:
            continue
        nc = _num(d.get("Aktuální NC"))
        nc3 = _num(d.get("Aktuální NC3"))
        anc = _num(d.get("Akční NC"))
        anck = _num(d.get("Akční NCK"))
        apc = _num(d.get("Akční PC"))
        kal = _num(d.get("Kalkulace"))
        k2 = _k2_text(d.get("Kód2") or d.get("K2"))
        # Vzorce ověřené proti „Označené sestavě":
        #   NC3-ANC = (NC3 − Akční NC) / NC3          (relativní, v %)
        #   NC/ANC  = Akční NC / NC − 1
        #   Akční přirážka = Akční PC / Akční NCK − 1
        #   rozdíl korekce v % = Akční přirážka − Kalkulace/100
        # Dělení nulou (vyloučený řádek s nulovými akčními cenami) → 0, ne prázdné.
        # None zůstává jen když chybí vstup úplně (nečíselná buňka).
        nc3_anc = ((nc3 - anc) / nc3) if nc3 else \
            (0.0 if (nc3 is not None and anc is not None) else None)
        nc_anc = (anc / nc - 1.0) if nc else \
            (0.0 if (nc is not None and anc is not None) else None)
        akcni_prirazka = (apc / anck - 1.0) if anck else \
            (0.0 if (apc is not None and anck is not None) else None)
        rozdil_korekce = (akcni_prirazka - kal / 100.0) \
            if (akcni_prirazka is not None and kal is not None) else None
        rec = {
            "data": d,
            "dod": _k2_text(d.get("Dodavatel") or d.get("Dod")),
            "kod": kod,
            "k2": k2,
            "k2_chyba": _k2_chyba(k2, kanal),
            "nc3_anc": nc3_anc,
            "nc_anc": nc_anc,
            "rozdil_korekce": rozdil_korekce,
            "akcni_prirazka": akcni_prirazka,
        }
        out.append(rec)
    if not out:
        return None, None, "Žádné datové řádky (s vyplněným kódem)."
    return out, pismena, None


def _letaky_uloz(kanal, radky, pismena, raw, nazev, zadavatel_id, zadavatel_jmeno,
                 nazev_pripadu="", druh="kontrolni"):
    conn = intranet_data.get_db_connection()
    if not conn:
        return None, "Není připojení k databázi."
    druh = _letaky_druh(druh)
    try:
        cur = conn.cursor()
        chyb = sum(1 for r in radky if r["k2_chyba"])
        cur.execute("INSERT INTO letaky_pripady "
                    "(nazev, kanal, druh, zadavatel_id, zadavatel_jmeno, stav, pocet_radku, "
                    " pocet_chyb_k2, soubor_nazev, sloupce_json) "
                    "VALUES (%s,%s,%s,%s,%s,'odeslano',%s,%s,%s,%s)",
                    (_str(nazev_pripadu, 255), kanal, druh, zadavatel_id, zadavatel_jmeno[:255],
                     len(radky), chyb, _str(nazev, 255),
                     json.dumps(pismena or {}, ensure_ascii=False)))
        pid = cur.lastrowid
        # Finální sestavy mají v čísle „F" (L00042 vs F00043) — sdílená číselná řada.
        cislo = f"{'F' if druh == 'finalni' else 'L'}{pid:05d}"
        cur.execute("UPDATE letaky_pripady SET cislo=%s WHERE id=%s", (cislo, pid))
        davka = [(pid, i, r["dod"][:40], r["kod"][:40], r["k2"][:60],
                  1 if r["k2_chyba"] else 0, r["nc3_anc"], r["nc_anc"],
                  r["rozdil_korekce"], r["akcni_prirazka"],
                  json.dumps(r["data"], ensure_ascii=False, default=str))
                 for i, r in enumerate(radky, 1)]
        for i in range(0, len(davka), 1000):
            cur.executemany(
                "INSERT INTO letaky_radky (pripad_id, poradi, dod, kod, k2, k2_chyba, "
                "nc3_anc, nc_anc, rozdil_korekce, akcni_prirazka, data_json) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", davka[i:i + 1000])
        if raw:
            cur.execute("INSERT INTO letaky_soubory (pripad_id, nazev, data) VALUES (%s,%s,%s)",
                        (pid, _str(nazev, 255), raw))
        conn.commit()
        cur.close()
        return pid, cislo
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return None, f"Chyba zápisu do databáze: {e}"
    finally:
        conn.close()


def _letaky_seznam(user_id, prava, druh="kontrolni"):
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    druh = _letaky_druh(druh)
    try:
        cur = conn.cursor(dictionary=True)
        # IFNULL kvůli sestavám nahraným před zavedením druhu (= kontrolní).
        if _letaky_vidi_vse(prava, druh):
            cur.execute("SELECT * FROM letaky_pripady WHERE IFNULL(druh,'kontrolni')=%s "
                        "ORDER BY id DESC", (druh,))
        else:
            cur.execute("SELECT * FROM letaky_pripady WHERE IFNULL(druh,'kontrolni')=%s "
                        "AND zadavatel_id=%s ORDER BY id DESC", (druh, user_id))
        r = cur.fetchall()
        cur.close()
        return r
    except Exception as e:
        print(f"[letaky] _letaky_seznam: {e}")
        return []
    finally:
        conn.close()


def _letaky_pripad(pid):
    conn = intranet_data.get_db_connection()
    if not conn:
        return None
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT * FROM letaky_pripady WHERE id=%s", (pid,))
        p = cur.fetchone()
        cur.close()
        return p
    finally:
        conn.close()


def _letaky_sloupce(p):
    """Mapa {hlavička: písmeno sloupce v Excelu} pro sestavu. Sestavám nahraným před
    zavedením `sloupce_json` se mapa dopočítá z uloženého původního souboru a rovnou uloží
    (dopočet proběhne jednou; když soubor chybí, vrátí prázdno a tabulka jede bez písmen)."""
    try:
        mapa = json.loads(p.get("sloupce_json") or "{}") or {}
    except Exception:
        mapa = {}
    if mapa:
        return mapa

    conn = intranet_data.get_db_connection()
    if not conn:
        return {}
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT nazev, data FROM letaky_soubory WHERE pripad_id=%s", (p["id"],))
        s = cur.fetchone()
        if not s or not s.get("data"):
            cur.close()
            return {}
        try:
            rows = _nacti_rows(s["data"], s.get("nazev") or "")
        except Exception:
            cur.close()
            return {}
        for row in rows[:40]:
            norm = {_norm(c) for c in row if c is not None and str(c).strip()}
            if "dodavatel" in norm and ("kód" in norm or "kod" in norm):
                mapa = _letaky_pismena(row)
                break
        if mapa:
            cur.execute("UPDATE letaky_pripady SET sloupce_json=%s WHERE id=%s",
                        (json.dumps(mapa, ensure_ascii=False), p["id"]))
            conn.commit()
        cur.close()
        return mapa
    except Exception:
        return {}
    finally:
        conn.close()


def _letaky_radky(pid):
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT * FROM letaky_radky WHERE pripad_id=%s ORDER BY poradi", (pid,))
        r = cur.fetchall()
        cur.close()
        return r
    finally:
        conn.close()


def _letaky_zpracuj(pid):
    conn = intranet_data.get_db_connection()
    if not conn:
        return False
    try:
        cur = conn.cursor()
        cur.execute("UPDATE letaky_pripady SET stav='zpracovano' WHERE id=%s", (pid,))
        conn.commit()
        cur.close()
        return True
    finally:
        conn.close()


def _letaky_smaz(pid):
    """Nevratně smaže sestavu vč. řádků, souboru a chatu. Vrací (ok, err)."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return False, "Bez DB spojení."
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM letaky_radky WHERE pripad_id=%s", (pid,))
        cur.execute("DELETE FROM letaky_soubory WHERE pripad_id=%s", (pid,))
        cur.execute("DELETE FROM letaky_chat WHERE pripad_id=%s", (pid,))
        cur.execute("DELETE FROM letaky_pripady WHERE id=%s", (pid,))
        conn.commit()
        cur.close()
        return True, None
    except Exception as e:
        return False, str(e)
    finally:
        conn.close()


def _letaky_chat_nacti(pid):
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT id,user_id,jmeno,zprava,kdy FROM letaky_chat "
                    "WHERE pripad_id=%s ORDER BY id ASC", (pid,))
        r = cur.fetchall()
        cur.close()
        return r
    finally:
        conn.close()


def _letaky_chat_pridej(pid, user_id, jmeno, zprava):
    conn = intranet_data.get_db_connection()
    if not conn:
        return False
    try:
        cur = conn.cursor()
        cur.execute("INSERT INTO letaky_chat (pripad_id,user_id,jmeno,zprava) "
                    "VALUES (%s,%s,%s,%s)", (pid, user_id, jmeno[:255], zprava))
        conn.commit()
        cur.close()
        return True
    except Exception as e:
        print(f"[letaky] _letaky_chat_pridej: {e}")
        return False
    finally:
        conn.close()


def _letaky_notifikuj_komentar(pripad, autor_id, autor_jmeno, zprava):
    """Po novém komentáři: e-mail + zvoneček všem s právy a žadateli (kromě autora)."""
    try:
        cislo = pripad.get("cislo") or ""
        popis = (zprava or "").strip()
        if len(popis) > 200:
            popis = popis[:200] + "…"
        druh = _letaky_druh(pripad.get("druh"))
        s = _letaky_sufix(druh)
        text = f"💬 {autor_jmeno} přidal(a) komentář k sestavě letáků {cislo}: {popis}"
        # Příjemci jen z práv daného druhu — finální sestavy nesmí „prosáknout"
        # notifikací k lidem bez práv Finální.
        prava = (f"cenopripad_office_letaky{s}", f"cenopripad_zadatel_letaky{s}",
                 "cenopripad_spravce", "vse")
        prijemci_id = intranet_data.ziskej_uzivatele_s_pravem(*prava) or {}
        ids = set(int(u) for u in prijemci_id if u is not None)
        if pripad.get("zadavatel_id"):
            ids.add(int(pripad["zadavatel_id"]))
        for uid in ids:
            try:
                if int(uid) == int(autor_id):
                    continue
            except (TypeError, ValueError):
                continue
            intranet_notifikace.pridej(uid, text, "info")
        # e-maily: práva + žadatel
        emaily = list(_emaily_s_pravy(*[p for p in prava if p != "vse"]))
        ez = _email_uzivatele(pripad.get("zadavatel_id"))
        if ez:
            emaily.append(ez)
        _odesli_emaily(emaily, f"Kontrolní data letáků {cislo} ({LETAKY_DRUHY[druh]}) "
                               "— nový komentář", text)
    except Exception as e:
        print(f"[letaky] _letaky_notifikuj_komentar: {e}")


def _letaky_notifikuj(cislo, kanal, chyb, pocet, zadavatel_jmeno, zadavatel_id,
                      druh="kontrolni"):
    """Zvoneček + e-mail všem s právy daného druhu (Office/žadatel Letáková kontrola,
    správce) kromě zadavatele."""
    try:
        druh = _letaky_druh(druh)
        s = _letaky_sufix(druh)
        prava = (f"cenopripad_office_letaky{s}", f"cenopripad_zadatel_letaky{s}",
                 "cenopripad_spravce", "vse")
        text = (f"📰 Nová {LETAKY_DRUHY[druh].lower()} letáků {cislo} "
                f"({LETAKY_KANALY.get(kanal, kanal)}) "
                f"od {zadavatel_jmeno}: {pocet} řádků, {chyb} chyb K2. "
                f"Prosím zkontrolujte a případně proveďte nápravu.")
        prijemci = intranet_data.ziskej_uzivatele_s_pravem(*prava)  # {id: jmeno}
        for uid in (prijemci or {}):
            try:
                if uid is None or int(uid) == int(zadavatel_id):
                    continue
            except (TypeError, ValueError):
                continue
            intranet_notifikace.pridej(uid, text, "info")
        _odesli_emaily(_emaily_s_pravy(*[p for p in prava if p != "vse"]),
                       f"Kontrolní data letáků — nová sestava {cislo} "
                       f"({LETAKY_DRUHY[druh]})", text)
    except Exception as e:
        print(f"[letaky] _letaky_notifikuj: {e}")


def _letaky_export_xlsx(pid):
    """Export sestavy do .xlsx jako „Označená sestava": všechny sloupce + dopočítaná
    pole, barevné zvýraznění (K2 + 4 sloupce dle znaménka), autofiltr, řazení dle Dod."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    p = _letaky_pripad(pid)
    radky = _letaky_radky(pid)
    if not radky:
        return None, "Sestava nemá data."

    # surová data + řazení dle Dod (sloupec A) vzestupně
    def _dodkey(r):
        try:
            return (0, float(r.get("dod")))
        except (TypeError, ValueError):
            return (1, str(r.get("dod") or ""))
    radky = sorted(radky, key=_dodkey)

    try:
        data_keys = list(json.loads(radky[0]["data_json"]).keys()) if radky[0].get("data_json") else []
    except Exception:
        data_keys = []
    comp = [("nc3_anc", "NC3-ANC", "pct"), ("nc_anc", "NC/ANC", "pct"),
            ("rozdil_korekce", "rozdíl korekce v %", "pct"), ("akcni_prirazka", "Akční přirážka", "pct")]
    headers = data_keys + [c[1] for c in comp]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "sjetina"
    bold = Font(bold=True)
    fill_k2 = PatternFill("solid", fgColor="FEE2E2")
    font_k2 = Font(color="B91C1C", bold=True)
    # barvy z engine (bez '#') pro openpyxl výplň + písmo
    barva_hex = {k: v.lstrip("#").upper() for k, v in _LETAKY_BARVA.items()}
    text_hex = {k: v.lstrip("#").upper() for k, v in _LETAKY_TEXT.items()}

    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = bold
        c.alignment = Alignment(wrap_text=False)
    k2_col = next((j for j, h in enumerate(data_keys, 1) if h in ("Kód2", "K2")), None)
    kanal = p.get("kanal")
    druh = _letaky_druh(p.get("druh"))      # finální sestava barví jinak

    for i, r in enumerate(radky, start=2):
        try:
            dd = json.loads(r["data_json"]) if r.get("data_json") else {}
        except Exception:
            dd = {}
        for j, h in enumerate(data_keys, 1):
            bunka = ws.cell(row=i, column=j, value=dd.get(h))
            # Nula v hlídaném sloupci (+ u finální záporná cena) → barevná buňka.
            styl = _letaky_bunka_styl(kanal, h, dd.get(h), druh)
            if styl:
                bunka.fill = PatternFill("solid", fgColor=styl[0].lstrip("#").upper())
                bunka.font = Font(color=styl[1].lstrip("#").upper(), bold=True)
        if k2_col and r.get("k2_chyba"):
            cc = ws.cell(row=i, column=k2_col)
            cc.fill = fill_k2
            cc.font = font_k2
        for off, (key, _lbl, typ) in enumerate(comp):
            j = len(data_keys) + 1 + off
            val = r.get(key)
            cell = ws.cell(row=i, column=j, value=val)
            cell.number_format = "0.00%" if typ == "pct" else "0.00"
            barva = _letaky_klasifikuj(key, val, druh)
            if barva:
                cell.fill = PatternFill("solid", fgColor=barva_hex[barva])
                cell.font = Font(color=text_hex[barva], bold=True)

    ws.freeze_panes = "A2"
    last_col = openpyxl.utils.get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

    buf = io.BytesIO()
    wb.save(buf)
    nazev = f"Oznacena_sestava_{p.get('cislo') or pid}.xlsx"
    return buf.getvalue(), nazev


def _letaky_prepni_druh(druh):
    """Přepnutí Kontrolní ↔ Finální sestavy (zavře i otevřený detail)."""
    app.storage.user["letaky_druh"] = _letaky_druh(druh)
    app.storage.user["letaky_detail"] = None
    _refresh()


def _letaky_aktivni_druh(prava):
    """Zvolený druh sestav, omezený na to, na co uživatel má práva."""
    druhy = _letaky_dostupne_druhy(prava)
    druh = app.storage.user.get("letaky_druh")
    if druh not in druhy:
        druh = druhy[0] if druhy else "kontrolni"
    return druh, druhy


def _letaky_view(user_id, user_name, prava):
    """Seznam sestav + detail (dle app.storage.user['letaky_detail'])."""
    det = app.storage.user.get("letaky_detail")
    if det:
        _letaky_detail(det, user_id, user_name, prava)
        return

    druh, druhy = _letaky_aktivni_druh(prava)
    # Přepínač se ukáže jen tomu, kdo má práva „Finální" (jinak vidí rovnou kontrolní).
    if len(druhy) > 1:
        with ui.row().classes("w-full mb-4"):
            ui.toggle({d: LETAKY_DRUHY[d] for d in druhy}, value=druh,
                      on_change=lambda e: _letaky_prepni_druh(e.value)) \
                .props("unelevated no-caps toggle-color=green-6")

    # ── Nahrání nové sestavy (čtenář nevidí — jen prohlíží/stahuje) ──
    if _letaky_muze_nahrat(prava, druh):
        drzeny = {"raw": None, "name": "", "kanal": "VO"}
        with ui.card().classes("w-full max-w-3xl p-5 rounded-2xl shadow-lg mb-4"):
            ui.label(f"Nahrát novou sestavu — {LETAKY_DRUHY[druh]}") \
                .classes("text-lg font-bold text-gray-800")
            ui.label("Vyberte kanál a nahrajte „dohody-karty_v_akci“ (.xls/.xlsx). "
                     "Sestava se vyhodnotí (K2 + počítaná pole) a uloží se stavem „Odesláno“.") \
                .classes("text-sm text-gray-500 mb-2")
            with ui.row().classes("w-full gap-3 items-center flex-wrap"):
                nazev_in = ui.input("Název případu") \
                    .props("outlined dense").classes("flex-1 min-w-64")
                kanal_sel = ui.select(LETAKY_KANALY, value="VO", label="Kanál") \
                    .props("outlined dense").classes("min-w-56")

            async def _on_up(e):
                up_spin.set_visibility(False)  # přenos na server dokončen
                raw, name = await _precti_upload(e, up)
                if raw is not None:
                    drzeny["raw"], drzeny["name"] = raw, name
                    stav.text = f"✅ Nahráno, připraveno ke kontrole: {name}"
                    stav.classes(replace="text-sm text-emerald-700 font-medium break-words")
                    btn.set_enabled(True)
                else:
                    stav.text = "❌ Nahrání souboru selhalo, zkuste to znovu."
                    stav.classes(replace="text-sm text-red-600 font-medium break-words")
                    btn.set_enabled(False)

            def _pick():
                return ui.run_javascript(
                    f"const c = getElement({up.id});"
                    f" const i = c && c.$el && c.$el.querySelector('input[type=file]');"
                    f" if (i) i.click();")

            with ui.row().classes("w-full items-center gap-3 mt-2"):
                up = ui.upload(on_upload=_on_up, auto_upload=True, max_file_size=40_000_000) \
                    .props("accept=.xls,.xlsx,.xlsm").style("display:none")
                ui.button(icon="add", on_click=_pick) \
                    .props("round unelevated color=green dense").classes("shadow-sm") \
                    .style("transform:scale(0.75);transform-origin:center") \
                    .tooltip("Vybrat soubor")
                with ui.column().classes("gap-0"):
                    ui.label("Nahrát soubor").classes("text-sm font-medium text-gray-700")
                    with ui.row().classes("items-center gap-2"):
                        up_spin = ui.spinner(size="sm").classes("text-emerald-600")
                        up_spin.set_visibility(False)
                        stav = ui.label("").classes("text-sm text-gray-500 break-words")

            # ── Ukazatel stavu vlastního přenosu souboru na server ──────────
            # QUploader pošle při výběru událost 'added' a auto_upload hned
            # spustí přenos; po dokončení dorazí 'uploaded' → _on_up. Mezitím
            # (u velkých .xls klidně desítky MB) ukážeme spinner + „Nahrávám…".
            def _zacatek_uploadu():
                drzeny["raw"] = None
                btn.set_enabled(False)
                up_spin.set_visibility(True)
                stav.text = "⏳ Nahrávám soubor na server…"
                stav.classes(replace="text-sm text-amber-700 font-medium break-words")

            def _upload_odmitnut():
                up_spin.set_visibility(False)
                stav.text = "❌ Soubor odmítnut (nepovolený typ nebo větší než 40 MB)."
                stav.classes(replace="text-sm text-red-600 font-medium break-words")

            up.on("added", lambda e: _zacatek_uploadu(), [])
            up.on("rejected", lambda e: _upload_odmitnut(), [])

            spin = ui.spinner(size="lg").classes("text-emerald-600")
            spin.set_visibility(False)

            async def _nahraj():
                nazev_pripadu = (nazev_in.value or "").strip()
                if not nazev_pripadu:
                    ui.notify("Zadejte název případu.", type="warning")
                    return
                if not drzeny["raw"]:
                    ui.notify("Vyberte soubor.", type="warning")
                    return
                kanal = kanal_sel.value or "VO"
                btn.set_enabled(False)
                spin.set_visibility(True)
                try:
                    radky, pismena, err = await run.cpu_bound(
                        _letaky_parse, drzeny["raw"], drzeny["name"], kanal)
                finally:
                    spin.set_visibility(False)
                if err:
                    ui.notify(f"Nahrání selhalo: {err}", type="negative", timeout=9000)
                    btn.set_enabled(True)
                    return
                pid, cislo = _letaky_uloz(kanal, radky, pismena, drzeny["raw"], drzeny["name"],
                                          user_id, user_name, nazev_pripadu, druh)
                if not pid:
                    ui.notify(f"Uložení selhalo: {cislo}", type="negative")
                    btn.set_enabled(True)
                    return
                chyb = sum(1 for r in radky if r["k2_chyba"])
                intranet_logger.log_activity(
                    user_name, "Cenopřípad",
                    f"Letáky: nová sestava {cislo} ({LETAKY_DRUHY[druh]}, {kanal}), "
                    f"{chyb} chyb K2")
                _letaky_notifikuj(cislo, kanal, chyb, len(radky), user_name, user_id, druh)
                ui.notify(f"Sestava {cislo} nahrána — {len(radky)} řádků, {chyb} chyb K2.",
                          type="positive", position="top")
                drzeny["raw"] = None
                _refresh()

            with ui.row().classes("w-full justify-end mt-2"):
                btn = ui.button("Nahrát a vyhodnotit", icon="cloud_upload", on_click=_nahraj) \
                    .props("unelevated no-caps") \
                    .classes("bg-emerald-600 text-white font-semibold rounded-lg px-5")
                btn.set_enabled(False)

    # ── Seznam sestav ───────────────────────────────────────────────
    sestavy = _letaky_seznam(user_id, prava, druh)
    if not sestavy:
        ui.label(f"Zatím žádné nahrané sestavy — {LETAKY_DRUHY[druh]}.") \
            .classes("text-gray-400 italic p-4")
        return
    with ui.column().classes("w-full gap-2"):
        for p in sestavy:
            stav_txt = "✅ Zpracováno" if p.get("stav") == "zpracovano" else "📤 Odesláno"
            barva = "border-green-300" if p.get("stav") == "zpracovano" else "border-blue-200"
            with ui.card().classes(f"w-full p-3 rounded-xl shadow-sm hover:shadow-md "
                                   f"transition-shadow cursor-pointer border-l-4 {barva}") \
                    .on("click", lambda pid=p["id"]: _letaky_otevri(pid)):
                with ui.row().classes("w-full items-center gap-3 flex-wrap"):
                    ui.label(p.get("cislo") or "—").classes("font-bold text-gray-800")
                    if p.get("nazev"):
                        ui.label(p.get("nazev")).classes("font-medium text-gray-700")
                    ui.label(LETAKY_KANALY.get(p.get("kanal"), p.get("kanal") or "")) \
                        .classes("text-sm px-2 py-0.5 rounded bg-gray-100 text-gray-700")
                    ui.label(stav_txt).classes("text-sm font-medium")
                    ui.space()
                    ui.label(f"{p.get('pocet_radku', 0)} řádků").classes("text-sm text-gray-500")
                    chyb = p.get("pocet_chyb_k2") or 0
                    ui.label(f"{chyb} chyb K2").classes(
                        "text-sm font-medium " + ("text-red-600" if chyb else "text-green-600"))
                    ui.label(f"{p.get('zadavatel_jmeno') or ''} · {_dt_cz(p.get('datum_zadani'))}") \
                        .classes("text-xs text-gray-400")


def _letaky_otevri(pid):
    app.storage.user["letaky_detail"] = pid
    _refresh()


def _letaky_detail(pid, user_id, user_name, prava):
    p = _letaky_pripad(pid)
    if not p:
        app.storage.user["letaky_detail"] = None
        _refresh()
        return
    # Druh případu, ne zvolený druh v seznamu — na detail se dá dostat i odkazem.
    druh = _letaky_druh(p.get("druh"))
    if not _letaky_pristup(prava, druh):
        app.storage.user["letaky_detail"] = None
        ui.notify(f"K sestavám „{LETAKY_DRUHY[druh]}“ nemáte oprávnění.", type="warning")
        _refresh()
        return
    radky = _letaky_radky(pid)

    with ui.row().classes("w-full items-center gap-3 mb-2"):
        ui.button(icon="arrow_back", on_click=lambda: _letaky_otevri(None)) \
            .props("flat round color=grey-7").tooltip("Zpět na seznam")
        ui.label(f"{p.get('cislo')} · {LETAKY_KANALY.get(p.get('kanal'), p.get('kanal'))}"
                 + (f" · {p.get('nazev')}" if p.get("nazev") else "")) \
            .classes("text-xl font-bold text-gray-800")
        ui.label(LETAKY_DRUHY[druh]).classes(
            "text-sm px-2 py-0.5 rounded font-medium "
            + ("bg-violet-100 text-violet-700" if druh == "finalni"
               else "bg-gray-100 text-gray-700"))
        ui.label("✅ Zpracováno" if p.get("stav") == "zpracovano" else "📤 Odesláno") \
            .classes("text-sm font-medium")
        ui.space()
        chyb = p.get("pocet_chyb_k2") or 0
        ui.label(f"{p.get('pocet_radku', 0)} řádků · {chyb} chyb K2").classes("text-sm text-gray-500")

        async def _export():
            data, nazev = await run.cpu_bound(_letaky_export_xlsx, pid)
            if not data:
                ui.notify(f"Export selhal: {nazev}", type="negative")
                return
            ui.download(data, nazev)
        ui.button("Export (Označená sestava)", icon="file_download", on_click=_export) \
            .props("outline no-caps").classes("text-emerald-700 font-semibold rounded-lg")

        if p.get("stav") != "zpracovano" and _letaky_office(prava, druh):
            def _zprac():
                _letaky_zpracuj(pid)
                ui.notify("Označeno jako zpracováno.", type="positive")
                _refresh()
            ui.button("Označit zpracováno", icon="task_alt", on_click=_zprac) \
                .props("unelevated no-caps").classes("bg-green-600 text-white rounded-lg")

        if _letaky_office(prava, druh):
            def _smaz_dialog():
                with ui.dialog() as cdlg, ui.card().classes("p-4"):
                    ui.label(f"Nevratně smazat sestavu {p.get('cislo')} — "
                             f"{LETAKY_KANALY.get(p.get('kanal'), p.get('kanal'))}?") \
                        .classes("font-medium")
                    ui.label("Smaže se sestava, její řádky, nahraný soubor i komentáře.") \
                        .classes("text-sm text-gray-500")

                    def _potvrd():
                        ok, err = _letaky_smaz(pid)
                        cdlg.close()
                        if ok:
                            intranet_logger.log_activity(
                                user_name, "Cenopřípad",
                                f"Letáky: smazána sestava {p.get('cislo')}")
                            ui.notify(f"Sestava {p.get('cislo')} smazána.", type="warning")
                            _letaky_otevri(None)
                        else:
                            ui.notify(err or "Smazání selhalo.", type="negative")

                    with ui.row().classes("w-full justify-end gap-2 mt-2"):
                        ui.button("Zrušit", on_click=cdlg.close).props("flat no-caps")
                        ui.button("Smazat", icon="delete_forever", on_click=_potvrd) \
                            .props("unelevated no-caps").classes("bg-red-600 text-white")
                cdlg.open()

            ui.button("Smazat sestavu", icon="delete", on_click=_smaz_dialog) \
                .props("flat no-caps").classes("text-red-600")

    if not radky:
        ui.label("Sestava nemá žádné řádky.").classes("text-gray-400 italic p-4")
        return

    # Tabulka: všechny sloupce ze souboru + dopočítaná pole; K2 červeně při chybě.
    try:
        data_keys = list(json.loads(radky[0]["data_json"]).keys()) if radky[0].get("data_json") else []
    except Exception:
        data_keys = []
    # pořadí: nejdřív klíčové, pak zbytek
    prvni = [k for k in ("Dodavatel", "Dodavatel popis", "Kód", "Balení", "Název",
                         "Kód2", "Aktuální NC", "Aktuální NC3", "Akční NC", "Akční NCK",
                         "Akční PC") if k in data_keys]
    zbytek = [k for k in data_keys if k not in prvni]
    poradi_klicu = prvni + zbytek
    _SLOUPCE = ["nc3_anc", "nc_anc", "rozdil_korekce", "akcni_prirazka"]
    _LABEL = {"nc3_anc": "NC3-ANC", "nc_anc": "NC/ANC",
              "rozdil_korekce": "rozdíl korekce", "akcni_prirazka": "Akční přirážka"}

    # cellStyle pro K2 (červené pozadí při chybě) a pro počítané sloupce (dle _bg/_fg).
    _CS_K2 = ("(p) => p.data && p.data._k2bad ? "
              "{backgroundColor:'#fee2e2',color:'#b91c1c',fontWeight:'700'} : null")
    _CS_COMP = ("(p) => { const f=p.colDef.field; const b=p.data && p.data._bg && p.data._bg[f];"
                " return b ? {backgroundColor:b, color:(p.data._fg[f]||'#000'), fontWeight:'700',"
                " whiteSpace:'nowrap'} : {whiteSpace:'nowrap'}; }")
    # Hlídaná buňka ze sjetiny (nula, u finální i záporná cena) — barvy nese řádek v _nula.
    _CS_NULA = ("(p) => { const s = p.data && p.data._nula && p.data._nula[p.colDef.field];"
                " return s ? {backgroundColor:s[0], color:s[1], fontWeight:'700'} : null; }")
    k2_field = next((f"d_{i}" for i, k in enumerate(poradi_klicu) if k in ("Kód2", "K2")), None)
    kanal = p.get("kanal")
    nula_fieldy = {f"d_{i}": k for i, k in enumerate(poradi_klicu)
                   if k in _letaky_hlidane_sloupce(kanal, druh)}

    # Číselné sloupce (ceny/množství) → zarovnat vpravo pro čitelnost.
    _NUM_KEYS  = {"Aktuální NC", "Aktuální NC3", "Akční NC", "Akční NCK", "Akční PC", "Balení"}
    # Dlouhé textové sloupce → omezit šířku, plný text v tooltipu.
    _WIDE_KEYS = {"Název", "Dodavatel popis", "Dodavatel"}

    # Písmena sloupců z původního Excelu (grid řadí sloupce jinak než soubor, takže
    # bez písmene se „sloupec Q" v tabulce nedá dohledat).
    pismena = _letaky_sloupce(p)

    # {field: (název, písmeno)} — pro přepínání zobrazení písmen v hlavičce.
    hlavicky = {}

    coldefs = [{"headerName": "#", "field": "poradi", "width": 64, "pinned": "left"}]
    for i, k in enumerate(poradi_klicu):
        pism = pismena.get(k)
        cd = {"headerName": f"{k} · {pism}" if pism else k, "field": f"d_{i}", "minWidth": 90}
        hlavicky[f"d_{i}"] = (k, pism)
        if pism:
            cd["headerTooltip"] = f"{k} — sloupec {pism} v původním souboru"
        if k in _NUM_KEYS:
            cd["type"] = "rightAligned"
        if k in _WIDE_KEYS:
            cd["maxWidth"] = 320
            cd["tooltipField"] = f"d_{i}"
        if f"d_{i}" == k2_field:
            cd[":cellStyle"] = _CS_K2
        elif f"d_{i}" in nula_fieldy:
            cd[":cellStyle"] = _CS_NULA
        coldefs.append(cd)
    for c in _SLOUPCE:
        coldefs.append({"headerName": _LABEL[c], "field": c, "minWidth": 120,
                        "type": "rightAligned", ":cellStyle": _CS_COMP})

    vsechny = []
    for r in radky:
        try:
            dd = json.loads(r["data_json"]) if r.get("data_json") else {}
        except Exception:
            dd = {}
        def _cell(v):
            # None/NaN → prázdné; celá čísla bez „.0" (jinak AG zobrazí „Invalid Number").
            if v is None:
                return ""
            if isinstance(v, float):
                if v != v:            # NaN
                    return ""
                if v.is_integer():
                    return str(int(v))
            return v
        row = {"poradi": r["poradi"]}
        for i, k in enumerate(poradi_klicu):
            row[f"d_{i}"] = _cell(dd.get(k))
        row["nc3_anc"] = _pct(r.get("nc3_anc"))
        row["nc_anc"] = _pct(r.get("nc_anc"))
        row["rozdil_korekce"] = _pct(r.get("rozdil_korekce")) if r.get("rozdil_korekce") is not None else "—"
        row["akcni_prirazka"] = _pct(r.get("akcni_prirazka"))
        barvy = {c: _letaky_klasifikuj(c, r.get(c), druh) for c in _SLOUPCE}
        row["_barva"] = {c: b for c, b in barvy.items() if b}
        row["_bg"] = {c: _LETAKY_BARVA[b] for c, b in barvy.items() if b}
        row["_fg"] = {c: _LETAKY_TEXT[b] for c, b in barvy.items() if b}
        row["_k2bad"] = bool(r.get("k2_chyba"))
        row["_nula"] = {}
        for _f, _k in nula_fieldy.items():
            _st = _letaky_bunka_styl(kanal, _k, dd.get(_k), druh)
            if _st:
                row["_nula"][_f] = list(_st)
        vsechny.append(row)

    # ── Filtry (přepínače) — K2 + nulové buňky + dopočítané sloupce ──
    filtr = {"k2": "vse", "nula": "vse"}
    filtr.update({c: "vse" for c in _SLOUPCE})

    def _projde(d):
        if filtr["k2"] == "ok" and d["_k2bad"]:
            return False
        if filtr["k2"] == "bad" and not d["_k2bad"]:
            return False
        if filtr["nula"] == "ok" and d["_nula"]:
            return False
        if filtr["nula"] == "bad" and not d["_nula"]:
            return False
        for c in _SLOUPCE:
            f = filtr[c]
            if f != "vse" and d["_barva"].get(c) != f:
                return False
        return True

    def _aplikuj():
        data = [d for d in vsechny if _projde(d)]
        grid.options["rowData"] = data
        grid.update()
        info_lbl.text = f"Zobrazeno {len(data)} z {len(vsechny)} řádků."

    def _set(klic, val):
        filtr[klic] = val
        _aplikuj()

    # Přepínání písmen sloupců v hlavičce (zapnuto = „Akční NC · Q", vypnuto = „Akční NC").
    zobraz = {"pismena": True}

    def _prepni_pismena():
        zap = not zobraz["pismena"]
        zobraz["pismena"] = zap
        for cd in grid.options["columnDefs"]:
            nazev, pism = hlavicky.get(cd.get("field"), (None, None))
            if nazev:
                cd["headerName"] = f"{nazev} · {pism}" if (zap and pism) else nazev
        grid.update()
        btn_pism.props(f"icon={'label' if zap else 'label_off'} "
                       f"color={'primary' if zap else 'grey-6'}")

    # Filtry NAD tabulkou.
    with ui.row().classes("items-center gap-3 mb-2 flex-wrap"):
        ui.select({"vse": "K2: vše", "ok": "K2: ok", "bad": "K2: špatně"},
                  value="vse", on_change=lambda e: _set("k2", e.value)) \
            .props("outlined dense").classes("min-w-40")
        # Filtr nulových buněk — jen u kanálů, které mají hlídané sloupce (VO).
        if nula_fieldy:
            _nula_popis = " / ".join(
                f"{k} ({pismena[k]})" if pismena.get(k) else k for k in nula_fieldy.values())
            # Finální sestava hlídá i zápornou cenu, proto obecnější popisky.
            _opt_nula = ({"vse": "Nuly: vše", "ok": "Nuly: bez nuly", "bad": "Nuly: jen nulové"}
                         if druh == "kontrolni" else
                         {"vse": "Vadné ceny: vše", "ok": "Vadné ceny: bez vady",
                          "bad": "Vadné ceny: jen vadné"})
            _tip = ("Nulová hodnota ve sloupci: " if druh == "kontrolni"
                    else "Nulová nebo záporná hodnota ve sloupci: ")
            ui.select(_opt_nula, value="vse", on_change=lambda e: _set("nula", e.value)) \
                .props("outlined dense").classes("min-w-44") \
                .tooltip(_tip + _nula_popis)
        for c in _SLOUPCE:
            pravidla = _letaky_pravidla(c, druh)
            if not pravidla:
                continue
            opts = {"vse": f"{_LABEL[c]}: vše"}
            for _op, _prah, _b in pravidla:
                opts[_b] = f"{_LABEL[c]}: {_LETAKY_BARVA_LABEL.get(_b, _b)}"
            ui.select(opts, value="vse", on_change=lambda e, c=c: _set(c, e.value)) \
                .props("outlined dense").classes("min-w-44")
        # Malé tlačítko: zapnout/vypnout písmena sloupců v hlavičce (jen když je co skrývat).
        if pismena:
            btn_pism = ui.button(icon="label", on_click=_prepni_pismena) \
                .props("flat dense round size=sm color=primary") \
                .tooltip("Zobrazit / skrýt písmena sloupců z Excelu (Q, T, U…)")
    info_lbl = ui.label(f"Zobrazeno {len(vsechny)} z {len(vsechny)} řádků.") \
        .classes("text-xs text-gray-500 mb-1")

    # AG Grid (virtualizace) — zvládá tisíce řádků a desítky sloupců rychle.
    grid = ui.aggrid({
        "columnDefs": coldefs,
        "rowData": vsechny,
        "defaultColDef": {"resizable": True, "sortable": True, "filter": True,
                          "suppressMovable": False,
                          # Zalomit dlouhé názvy hlaviček, ať jsou celé čitelné.
                          "wrapHeaderText": True, "autoHeaderHeight": True,
                          # Vypnout auto-odvození typu (AG v31+) — jinak číselný sloupec
                          # zobrazí textové/NaN buňky jako „Invalid Number".
                          "cellDataType": False},
        # Napevno srovnat šířku každého sloupce podle jeho obsahu (ne slepené 90 px).
        "autoSizeStrategy": {"type": "fitCellContents"},
        "rowHeight": 28,
        ":getRowId": "(p) => String(p.data.poradi)",
    }).classes("w-full").style("height: 68vh")

    # ── Diskuze k sestavě (chat) + e-mail při novém komentáři ───────────
    ui.separator().classes("my-4")
    with ui.row().classes("items-center gap-2"):
        ui.icon("forum", color="primary")
        ui.label("Diskuze k sestavě").classes("text-lg font-bold text-gray-800")

    box = ui.column().classes("w-full gap-2")

    @ui.refreshable
    def _chat():
        box.clear()
        zpravy = _letaky_chat_nacti(pid)
        with box:
            if not zpravy:
                ui.label("Zatím žádné komentáře.").classes("text-sm text-gray-400 italic")
            for z in zpravy:
                moje = (z.get("user_id") == user_id)
                cas = z.get("kdy")
                cas_txt = cas.strftime("%d.%m.%Y %H:%M") if hasattr(cas, "strftime") else str(cas or "")
                with ui.row().classes("w-full " + ("justify-end" if moje else "justify-start")):
                    with ui.column().classes("gap-0").style("max-width:78%"):
                        if not moje:
                            ui.label(z.get("jmeno") or "—") \
                                .classes("text-xs font-semibold text-gray-600 px-1")
                        bg = ("background:#2563eb;color:#fff" if moje
                              else "background:#f1f5f9;color:#1e293b")
                        with ui.element("div").classes("rounded-2xl px-3 py-2").style(bg):
                            ui.label(z.get("zprava") or "").classes("text-sm") \
                                .style("white-space:pre-wrap;word-break:break-word")
                        ui.label(cas_txt).classes("text-gray-400 px-1 " +
                                                  ("self-end" if moje else "")).style("font-size:10px")

    _chat()

    if _letaky_muze_nahrat(prava, druh):
        with ui.row().classes("w-full items-end gap-2 mt-2"):
            inp = ui.textarea(placeholder="Napiš komentář… (Ctrl+Enter odešle)") \
                .props("outlined autogrow dense").classes("flex-1")

            def _odeslat_koment():
                txt = (inp.value or "").strip()
                if not txt:
                    return
                if _letaky_chat_pridej(pid, user_id, user_name, txt):
                    inp.value = ""
                    _chat.refresh()
                    intranet_logger.log_activity(user_name, "Cenopřípad",
                                                 f"Letáky {p.get('cislo')}: komentář")
                    # e-mail + zvoneček mimo event loop
                    import asyncio as _a
                    _a.create_task(_a.to_thread(_letaky_notifikuj_komentar, p, user_id, user_name, txt))
                else:
                    ui.notify("Komentář se nepodařilo uložit.", type="negative")

            inp.on("keydown.ctrl.enter", _odeslat_koment)
            ui.button(icon="send", on_click=_odeslat_koment) \
                .props("round unelevated color=primary").tooltip("Odeslat (Ctrl+Enter)")
    else:
        ui.label("Máte právo pouze pro čtení — komentáře můžete číst, ne psát.") \
            .classes("text-sm text-gray-400 italic mt-2")


def vykresli_cenopripad(user_id, user_name, vsechna_prava):
    """Vstupní bod modulu — vytvoří per-klient refreshable a vykreslí ho.

    Refreshable je záměrně lokální (vzniká nová instance při každém vykreslení = pro
    každého klienta zvlášť) a jeho `.refresh` se uloží do `app.storage.client`, odkud
    ho přečte `_refresh()`. Tím se přerenderuje jen aktuální klient. Modulový
    `@ui.refreshable` zde dělal chybu: `refresh()` re-renderoval všechny připojené
    klienty v kontextu jednoho uživatele → všem přepínal sekci (sdílený stav přes
    `app.storage.user` + request_contextvar)."""
    @ui.refreshable
    def _obsah():
        _vykresli_cenopripad(user_id, user_name, vsechna_prava)

    app.storage.client["_cenopripad_refresh"] = _obsah.refresh
    _obsah()


def _vykresli_cenopripad(user_id, user_name, vsechna_prava):
    inicializace_cenopripad_db()
    typy = _viditelne_typy(vsechna_prava)
    pohled = app.storage.user.get("cenopripad_pohled")
    if pohled not in typy and pohled not in ("import", "letaky"):
        pohled = None
    if pohled == "import" and not _vidi_import(vsechna_prava):
        pohled = None
    if pohled == "letaky" and not _letaky_pristup(vsechna_prava):
        pohled = None

    with ui.row().classes("w-full items-center gap-3 mb-6"):
        if pohled:
            ui.button(icon="arrow_back", on_click=lambda: _nav(None)) \
                .props("flat round color=grey-7").tooltip("Zpět na přehled")
        ui.icon("sell", size="2.2rem").classes("text-emerald-600")
        with ui.column().classes("gap-0"):
            ui.label("Cenopřípad").classes("text-3xl font-extrabold text-gray-800")
            podtitul = "Kontrola nastavení letákových cen"
            if pohled in typy:
                podtitul = TYPY[pohled]["nazev"]
            elif pohled == "import":
                podtitul = "Import dat"
            elif pohled == "letaky":
                podtitul = "Kontrolní data letáků"
            ui.label(podtitul).classes("text-sm text-gray-500")
        ui.space()
        if pohled == "porovnani":   # IND ceny — vzorový formulář ke stažení
            ui.button("Vzorový formulář", icon="file_download", on_click=_stahni_vzor_ind) \
                .props("outline no-caps") \
                .classes("text-emerald-700 font-semibold rounded-lg") \
                .tooltip("Stáhne vzorový formulář pro individuální ceny (.xlsx).")
        # Generický manuál Cenopřípadu (ind. ceny aj.) — NE v „Kontrolní data letáků"
        # (ta má vlastní obsah; ind. ceny a letáky tak nesdílí stejný manuál).
        if pohled != "letaky":
            ui.button("Manuál", icon="menu_book", on_click=_dialog_manual) \
                .props("outline no-caps") \
                .classes("text-emerald-700 font-semibold rounded-lg") \
                .tooltip("Příručka uživatele — náhled manuálu modulu Cenopřípad.")

    if pohled == "import":
        _panel_import(user_name, _refresh)
        return
    if pohled == "letaky":
        _letaky_view(user_id, user_name, vsechna_prava)
        return
    if pohled in typy:
        _sub_view_typ(pohled, user_id, user_name, vsechna_prava)
        return

    if not typy and not _vidi_import(vsechna_prava):
        with ui.column().classes("items-center py-20 gap-3 w-full"):
            ui.icon("lock", size="4rem", color="grey-4")
            ui.label("Nemáte přístup k žádné dlaždici modulu Cenopřípad.") \
                .classes("text-lg text-gray-400")
        return

    with ui.row().classes("w-full gap-6 flex-wrap pt-2"):
        if _vidi_import(vsechna_prava):
            _tile("⬆️", "Import dat", "OP + DATA_POROVNANI", "border-blue-200",
                  lambda: _nav("import"))
        if _letaky_pristup(vsechna_prava):
            def _otevri_letaky():
                app.storage.user["letaky_detail"] = None
                _nav("letaky")
            _tile("📰", "Kontrolní data letáků", "letáková kontrola", "border-rose-200", _otevri_letaky)
        for klic in typy:
            c = TYPY[klic]
            _tile(c["emoji"], c["nazev"], "nákup" if c["oddeleni"] == "nakup" else "obchod",
                  c["barva"], lambda k=klic: _nav(k))


def _tile(emoji, nadpis, popis, barva, on_click):
    with ui.card().classes(
            "w-72 h-60 items-center justify-center shadow-xl hover:scale-105 "
            "transition-transform duration-300 cursor-pointer bg-white rounded-2xl border "
            + barva).on("click", on_click):
        ui.label(emoji).classes("text-6xl mb-3")
        ui.label(nadpis).classes("text-xl font-bold text-gray-800 text-center")
        if popis:
            ui.label(popis).classes("text-xs text-gray-500 text-center uppercase tracking-wide")
