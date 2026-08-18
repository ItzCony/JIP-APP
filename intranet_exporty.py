from nicegui import ui
import pandas as pd
import os
import time
import datetime
import random
import calendar
import openpyxl
import asyncio
import io as _io
import csv as _csv_mod
import re as _re
import intranet_data
import intranet_logger
import intranet_jobs


# ════════════════════════════════════════════════════════════════════════════
# OŘEZ ABSENCE NA OBDOBÍ FILTRU
# Vícedenní absence přesahující hranici zvoleného období (např. 29.06–05.07 při
# filtru za červen) se musí do součtů i detailu započítat jen svou poměrnou částí
# spadající do období — jinak se do součtů zanesou hodiny mimo filtr.
# ════════════════════════════════════════════════════════════════════════════
def _prac_dny_v_rozsahu(od, do):
    """Počet pracovních dnů (Po–Pá mimo státní svátky) v rozsahu [od, do] včetně."""
    if od > do:
        return 0
    import intranet_obsah  # lazy — obsah tahá NiceGUI/DB, nechceme v child procesu stavěčů
    return sum(
        1 for i in range((do - od).days + 1)
        if (od + datetime.timedelta(days=i)).weekday() < 5
        and not intranet_obsah.je_statni_svatek(od + datetime.timedelta(days=i))
    )


def _orez_volno_na_obdobi(d_from, d_to, suma_hodin, obd_od, obd_do):
    """Ořízne záznam volna na průnik s obdobím [obd_od, obd_do].

    Vrací (datum_od, datum_do, hodiny). Hodiny se přepočítají poměrem pracovních
    dnů (Po–Pá mimo státní svátky) spadajících do období, aby vícedenní absence
    přesahující hranici filtru nezanesla do součtů hodiny mimo zvolené období.
    Záznam ležící celý uvnitř období zůstává beze změny."""
    if d_from >= obd_od and d_to <= obd_do:
        return d_from, d_to, suma_hodin
    clip_od = max(d_from, obd_od)
    clip_do = min(d_to, obd_do)
    plne = _prac_dny_v_rozsahu(d_from, d_to)
    if plne <= 0:
        return clip_od, clip_do, suma_hodin
    cast = _prac_dny_v_rozsahu(clip_od, clip_do)
    return clip_od, clip_do, round(suma_hodin * cast / plne, 2)


# ════════════════════════════════════════════════════════════════════════════
# JEDNOTNÉ UI STAVEBNÍ PRVKY EXPORTNÍCH DIALOGŮ (moderní vzhled)
# ════════════════════════════════════════════════════════════════════════════
_TBL_PROPS = 'flat bordered hide-bottom :pagination="{rowsPerPage: 0}"'
_TBL_CLASSES = 'w-full mb-6 rounded-xl overflow-hidden text-base bg-white'


def _ui_hlavicka_dialogu(dlg, ikona, titulek, podtitulek, badge_classes):
    """Záhlaví dialogu: ikona v barevném odznaku, titul + podtitul, zavírací křížek."""
    with ui.row().classes('w-full items-center justify-between bg-white border border-gray-200 rounded-2xl px-5 py-3 shadow-sm mb-2'):
        with ui.row().classes('items-center gap-3'):
            with ui.element('div').classes(f'w-11 h-11 rounded-xl flex items-center justify-center {badge_classes}'):
                ui.icon(ikona, size='sm')
            with ui.column().classes('gap-0'):
                ui.label(titulek).classes('text-xl font-extrabold text-gray-800 leading-tight')
                ui.label(podtitulek).classes('text-xs text-gray-400')
        ui.button(icon='close', on_click=dlg.close).props('flat round dense').classes('text-gray-400 hover:text-gray-600')


def _ui_nadpis_sekce(cislo, titulek, poznamka=None):
    """Nadpis sekce náhledu s číselným odznakem."""
    with ui.row().classes('w-full items-center gap-3 mb-3 mt-6'):
        if cislo:
            ui.label(cislo).classes('w-7 h-7 rounded-full bg-gray-800 text-white text-sm font-bold flex items-center justify-center shrink-0')
        ui.label(titulek).classes('text-lg font-bold text-gray-800')
        if poznamka:
            ui.label(poznamka).classes('text-xs text-gray-400')


def _ui_prazdny_stav(text, ikona='search_off'):
    """Prázdný stav náhledu (nic nenalezeno / bez oprávnění)."""
    with ui.column().classes('w-full items-center py-14 gap-2'):
        ui.icon(ikona, size='3rem').classes('text-gray-300')
        ui.label(text).classes('text-gray-400 text-base')


def _ui_stat_chip(ikona, hodnota, popisek):
    """Malý statistický štítek nad souhrnem."""
    with ui.row().classes('items-center gap-2 px-3 py-1.5 rounded-lg border border-gray-200 bg-white text-gray-700'):
        ui.icon(ikona, size='xs').classes('text-gray-400')
        ui.label(hodnota).classes('font-bold text-sm')
        ui.label(popisek).classes('text-xs opacity-70')


# ════════════════════════════════════════════════════════════════════════════
# STAVĚČE SOUBORŮ (CPU-náročné) — funkce na úrovni modulu, aby šly spustit přes
# intranet_jobs.cpu() v odděleném PROCESU (jiné jádro, obejde GIL).
# Vstup i výstup MUSÍ být picklovatelné: list dictů / str / bytes / cesty.
# Nesmí sahat do DB ani do NiceGUI — jen čistá data → soubor.
# ════════════════════════════════════════════════════════════════════════════
def build_porovnani_xlsx(export_radky: list, nazev_souboru: str) -> str:
    """[BĚŽÍ V PROCESU] Sestaví Excel „Porovnání dovolené" z hotových řádků."""
    import io as _io2
    from openpyxl.styles import PatternFill, Font, Alignment
    df_exp = pd.DataFrame(export_radky)
    buf = _io2.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df_exp.to_excel(writer, index=False, sheet_name='Porovnání')
        ws = writer.sheets['Porovnání']
        # Formátování záhlaví
        hdr_fill = PatternFill('solid', fgColor='4F46E5')
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center')
        # Barevné řádky podle stavu
        fill_green = PatternFill('solid', fgColor='D1FAE5')
        fill_red   = PatternFill('solid', fgColor='FEE2E2')
        fill_gray  = PatternFill('solid', fgColor='F3F4F6')
        stav_col_idx = [c.column for c in ws[1] if c.value == 'Stav']
        stav_col = stav_col_idx[0] if stav_col_idx else None
        for row in ws.iter_rows(min_row=2):
            stav_cell = ws.cell(row=row[0].row, column=stav_col) if stav_col else None
            stav_val  = stav_cell.value if stav_cell else ''
            if stav_val == 'Shoduje se':   fill = fill_green
            elif stav_val == 'Neshoduje se': fill = fill_red
            else:                            fill = fill_gray
            for cell in row:
                cell.fill = fill
        # Šířky sloupců
        for col in ws.columns:
            max_w = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_w + 4, 40)
    buf.seek(0)
    os.makedirs('Exporty_Logy', exist_ok=True)
    cesta = os.path.join('Exporty_Logy', nazev_souboru)
    with open(cesta, 'wb') as f:
        f.write(buf.read())
    return cesta


def read_excel_vsechny_listy(raw_bytes: bytes) -> dict:
    """[BĚŽÍ V PROCESU] Načte všechny listy nahraného .xlsx z bytes.
    Vrací {nazev_listu: DataFrame} (DataFrame je picklovatelný)."""
    return pd.read_excel(_io.BytesIO(raw_bytes), sheet_name=None)


def build_dochazka_xlsx(vyfiltrovana, presczasy_detail_rows, zustatky_vypis,
                        val_report, val_od, val_do) -> str:
    """[BĚŽÍ V PROCESU] Sestaví docházkový Excel z hotových dat (bez DB).
    Data se připraví v hlavním procesu (gather přes jobs.io), sem přijdou už
    jen picklovatelné seznamy dictů → výstupem je cesta k .xlsx."""
    df_souhrn = pd.DataFrame()
    df_detail = pd.DataFrame()
    if vyfiltrovana:
        df = pd.DataFrame(vyfiltrovana)
        df_souhrn = df.pivot_table(index=['Oddělení', 'Jméno'], columns='Typ volna', values='Hodiny do součtu', aggfunc='sum', fill_value=0).reset_index()
        typy_sloupce = [c for c in df_souhrn.columns if c not in ['Oddělení', 'Jméno']]
        df_souhrn['Celkem hodin'] = df_souhrn[typy_sloupce].sum(axis=1)

        df_detail = df.sort_values(by=['Oddělení', 'Jméno', 'Od'])
        df_detail['Od'] = df_detail['Od'].apply(lambda x: x.strftime('%d.%m.%Y') if hasattr(x, 'strftime') else str(x))
        df_detail['Do'] = df_detail['Do'].apply(lambda x: x.strftime('%d.%m.%Y') if hasattr(x, 'strftime') else str(x))
        if 'Hodiny do součtu' in df_detail.columns:
            df_detail = df_detail.drop(columns=['Hodiny do součtu'])

    df_zustatky = pd.DataFrame(zustatky_vypis)
    if not df_zustatky.empty:
        df_zustatky = df_zustatky.sort_values(by=['Oddělení', 'Jméno'])

    slozka = "Exporty_Dochazka"
    os.makedirs(slozka, exist_ok=True)
    cesta = os.path.join(slozka, f"Export_Dochazky_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")

    try:
        val_od_fmt = datetime.datetime.strptime(val_od, '%Y-%m-%d').strftime('%d.%m.%Y') if val_od else '...'
        val_do_fmt = datetime.datetime.strptime(val_do, '%Y-%m-%d').strftime('%d.%m.%Y') if val_do else '...'
    except Exception:
        val_od_fmt, val_do_fmt = val_od, val_do

    with pd.ExcelWriter(cesta, engine='openpyxl') as writer:
        if val_report == 'komplet' and not df_souhrn.empty:
            df_souhrn.to_excel(writer, sheet_name='Souhrn absencí', index=False, startrow=2)
            ws = writer.sheets['Souhrn absencí']
            ws['A1'] = f"Souhrn absencí + přesčasů za období: {val_od_fmt} - {val_do_fmt}"
            ws['A1'].font = openpyxl.styles.Font(bold=True, size=12)
            df_detail.to_excel(writer, sheet_name='Detailní výpis', index=False)
        if not df_zustatky.empty:
            df_zustatky.to_excel(writer, sheet_name='Zůstatky dovolené', index=False)
        if presczasy_detail_rows:
            df_ot = pd.DataFrame(presczasy_detail_rows).sort_values(by=['Oddělení', 'Jméno', 'Datum od'])
            df_ot.to_excel(writer, sheet_name='Přesčasy detail', index=False)

    return cesta


def build_dochazka_ucetni_xlsx(vyfiltrovana, zustatky_vypis, val_report, val_od, val_do) -> str:
    """[BĚŽÍ V PROCESU] Účetní export docházky (Souhrn/Detail/Zůstatky/Zadání)
    z hotových picklovatelných dat → cesta k .xlsx. Bez DB a NiceGUI."""
    df_souhrn = pd.DataFrame()
    df_detail = pd.DataFrame()
    if vyfiltrovana:
        df = pd.DataFrame(vyfiltrovana)
        df_souhrn = df.pivot_table(index=['Společnost', 'Oddělení', 'Osobní číslo', 'Jméno a příjmení'], columns='Typ volna', values='Hodiny do součtu', aggfunc='sum', fill_value=0).reset_index()
        typy_sloupce = [c for c in df_souhrn.columns if c not in ['Společnost', 'Oddělení', 'Osobní číslo', 'Jméno a příjmení']]
        df_souhrn['Celkem hodin'] = df_souhrn[typy_sloupce].sum(axis=1)

        df_detail = df.sort_values(by=['Společnost', 'Oddělení', 'Jméno a příjmení', 'Od'])
        df_detail['Od'] = df_detail['Od'].apply(lambda x: x.strftime('%d.%m.%Y') if hasattr(x, 'strftime') else str(x))
        df_detail['Do'] = df_detail['Do'].apply(lambda x: x.strftime('%d.%m.%Y') if hasattr(x, 'strftime') else str(x))
        if 'Hodiny do součtu' in df_detail.columns:
            df_detail = df_detail.drop(columns=['Hodiny do součtu'])
        if 'Zadáno' in df_detail.columns:
            df_detail = df_detail.drop(columns=['Zadáno'])

    df_zustatky = pd.DataFrame(zustatky_vypis)
    if not df_zustatky.empty:
        df_zustatky = df_zustatky.sort_values(by=['Společnost', 'Oddělení', 'Jméno a příjmení'])

    slozka = "Exporty_Dochazka"
    os.makedirs(slozka, exist_ok=True)
    cesta = os.path.join(slozka, f"Export_Dochazky_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")

    try:
        val_od_fmt = datetime.datetime.strptime(val_od, '%Y-%m-%d').strftime('%d.%m.%Y') if val_od else '...'
        val_do_fmt = datetime.datetime.strptime(val_do, '%Y-%m-%d').strftime('%d.%m.%Y') if val_do else '...'
    except Exception:
        val_od_fmt, val_do_fmt = val_od, val_do

    df_zadani = pd.DataFrame([z for z in vyfiltrovana if z.get('Typ volna') in ('dov.', 'Náhradní volno')])
    if not df_zadani.empty:
        df_zadani = df_zadani.sort_values(by=['Zadáno', 'Jméno a příjmení'])
        df_zadani['Zadáno'] = df_zadani['Zadáno'].apply(lambda x: x.strftime('%d.%m.%Y %H:%M') if hasattr(x, 'strftime') else (str(x) if x else ''))
        df_zadani = df_zadani[['Společnost', 'Oddělení', 'Osobní číslo', 'Jméno a příjmení', 'Typ volna', 'Od', 'Do', 'Hodiny', 'Zadáno']]
        df_zadani['Od'] = df_zadani['Od'].apply(lambda x: x.strftime('%d.%m.%Y') if hasattr(x, 'strftime') else str(x))
        df_zadani['Do'] = df_zadani['Do'].apply(lambda x: x.strftime('%d.%m.%Y') if hasattr(x, 'strftime') else str(x))

    with pd.ExcelWriter(cesta, engine='openpyxl') as writer:
        if val_report == 'komplet' and not df_souhrn.empty:
            df_souhrn.to_excel(writer, sheet_name='Souhrn', index=False, startrow=2)
            ws = writer.sheets['Souhrn']
            ws['A1'] = f"Souhrn absencí a přesčasů za období: {val_od_fmt} - {val_do_fmt}"
            ws['A1'].font = openpyxl.styles.Font(bold=True, size=12)
            df_detail.to_excel(writer, sheet_name='Detailní výpis', index=False)
        if not df_zustatky.empty:
            df_zustatky.to_excel(writer, sheet_name='Zůstatky dovolené', index=False)
        if not df_zadani.empty:
            df_zadani.to_excel(writer, sheet_name='Zadání Dov+NV', index=False)

    return cesta


def build_ikos_csv(vsechna_volna, vsichni_uzivatele_komplet, val_role, val_jmeno,
                   val_od, val_do, val_sloucit, oddelovac, slozka) -> str:
    """[BĚŽÍ V PROCESU] Sestaví IKOS CSV ze schválených voln (filtr + zápis).
    Vstup picklovatelný (záznamy + slovník uživatelů + parametry) → cesta k CSV."""
    try:
        d_od = datetime.datetime.strptime(val_od, '%Y-%m-%d').date() if val_od else datetime.date.min
        d_do = datetime.datetime.strptime(val_do, '%Y-%m-%d').date() if val_do else datetime.date.max
    except Exception:
        raise ValueError('Neplatný formát data!')

    radky = []
    for v in vsechna_volna:
        if v['stav_id'] != 2: continue
        if not (v['to'] >= d_od and v['from'] <= d_do): continue

        cele_jmeno = f"{v['u_jmeno']} {v['u_prijmeni']}"
        odd_zaznamu = v.get('oddeleni') or 'Bez oddělení'
        odd_list = [o.strip() for o in odd_zaznamu.split(',')]

        if val_role != 'vse' and val_role not in odd_list: continue
        if val_jmeno != 'vse' and cele_jmeno != val_jmeno: continue

        udata = next((u for u in vsichni_uzivatele_komplet.values() if u['jmeno_cele'] == cele_jmeno), {})
        os_cislo = str(udata.get('id', '')).zfill(6)
        if val_sloucit:
            priznak = udata.get('priznak_nazev', '') or ''
            if priznak:
                os_cislo = f'{priznak}{os_cislo}'
        datum_od_s = v['from'].strftime('%d.%m.%Y') if hasattr(v['from'], 'strftime') else str(v['from'])
        datum_do_s = v['to'].strftime('%d.%m.%Y') if hasattr(v['to'], 'strftime') else str(v['to'])
        cas_od_s = _ikos_fmt_cas(v.get('cas_od'))
        cas_do_s = _ikos_fmt_cas(v.get('cas_do'))
        radky.append([os_cislo, datum_od_s, datum_do_s, cas_od_s, cas_do_s, v['typ']])

    if not radky:
        raise ValueError('Pro zadaný filtr nebyla nalezena žádná schválená volna k exportu.')

    os.makedirs(slozka, exist_ok=True)
    cesta = os.path.join(slozka, f"IKOS_Export_{time.strftime('%Y%m%d_%H%M%S')}.csv")
    with open(cesta, 'w', newline='', encoding='utf-8-sig') as f:
        w = _csv_mod.writer(f, delimiter=oddelovac)
        w.writerow(['Os. číslo', 'Datum od', 'Datum do', 'Čas od', 'Čas do', 'Druh volna'])
        w.writerows(radky)
    return cesta


def _ikos_fmt_cas(val):
    """Naformátuje čas (z DB typu TIME → timedelta) na 'HH:MM:SS' se zero-paddingem.
    Zvládne i datetime.time a textový vstup; prázdná hodnota → ''."""
    if not val:
        return ''
    # MySQL TIME sloupce přicházejí jako datetime.timedelta (např. 8:00:00 bez úvodní nuly)
    if isinstance(val, datetime.timedelta):
        celkem = int(val.total_seconds())
        h, zbytek = divmod(celkem, 3600)
        m, s = divmod(zbytek, 60)
        return f'{h:02d}:{m:02d}:{s:02d}'
    if isinstance(val, datetime.time):
        return val.strftime('%H:%M:%S')
    # Textový vstup typu '8:00', '8:00:00', '08:00:00' → normalizovat na HH:MM:SS
    casti = str(val).split(':')
    try:
        h = int(casti[0]); m = int(casti[1]) if len(casti) > 1 else 0; s = int(casti[2]) if len(casti) > 2 else 0
        return f'{h:02d}:{m:02d}:{s:02d}'
    except (ValueError, IndexError):
        return str(val)


def _fmt_cas_hhmm(val):
    """Čas na 'HH:MM'. MySQL TIME chodí jako timedelta a str() je bez úvodní nuly
    ('7:00:00'), takže prosté [:5] nechává visící dvojtečku ('7:00:')."""
    return _ikos_fmt_cas(val)[:5]

def vykresli_exportni_tlacitka(user_name, vsechna_prava, vsichni_uzivatele_komplet, typy_volna, aktualni_oddeleni_dict, ma_pristup_vsechny_slozky, ma_tisk_odd_vse_b, ma_tisk_typy_vse_b, ma_pristup_ucetnictvi, ukazat_tlacitko_export, povoleny_porovnani_oddeleni=None, ma_ikos_export=False):

    # Nově si načteme ID žadatele a seznam jeho "sledovaných uživatelů" (pro Majitele)
    muj_ucet = next((u for u in vsichni_uzivatele_komplet.values() if u['jmeno_cele'] == user_name), {})
    user_id = muj_ucet.get('id')
    moji_sledovani = muj_ucet.get('sledovani_uzivatele', [])

    if ukazat_tlacitko_export:
        povoleny_report_oddeleni_b = set()
        if ma_tisk_odd_vse_b or ma_pristup_vsechny_slozky:
            povoleny_report_oddeleni_b = set(aktualni_oddeleni_dict.keys())
        else:
            for r_nazev in aktualni_oddeleni_dict.keys():
                # Nově přidáno právo hlavní_vedouci
                if f'tisk_odd_{r_nazev.lower()}' in vsechna_prava or f'slozka_{r_nazev.lower()}' in vsechna_prava or f'hlavni_vedouci_{r_nazev.lower()}' in vsechna_prava:
                    povoleny_report_oddeleni_b.add(r_nazev)

        povoleny_report_typy_b = set()
        if ma_tisk_typy_vse_b:
            povoleny_report_typy_b = set(typy_volna.values())
        else:
            for t_nazev in typy_volna.values():
                if f'tisk_typ_{t_nazev.lower()}' in vsechna_prava:
                    povoleny_report_typy_b.add(t_nazev)

        with ui.dialog().props('maximized transition-show="slide-up" transition-hide="slide-down"') as tisk_dlg_bezny:
            with ui.card().classes('w-full h-full p-4 bg-gray-100 text-black overflow-y-auto'):
                _ui_hlavicka_dialogu(tisk_dlg_bezny, 'summarize', 'Exporty absencí',
                                     'Souhrny hodin, detailní výpisy a zůstatky dovolené',
                                     'bg-purple-100 text-purple-700')

                vsechny_role_pro_tisk_b = {}
                if ma_tisk_odd_vse_b or ma_pristup_vsechny_slozky or len(povoleny_report_oddeleni_b) > 1:
                    vsechny_role_pro_tisk_b['vse'] = 'Všechna povolená oddělení'
                for r in sorted(povoleny_report_oddeleni_b):
                    vsechny_role_pro_tisk_b[r] = r

                with ui.row().classes('w-full gap-3 mb-4 items-end bg-white p-4 rounded-2xl border border-gray-200 shadow-sm flex-wrap'):
                    with ui.input('Od data').classes('w-32 bg-white').props('outlined dense') as tisk_od_b:
                        with tisk_od_b.add_slot('append'):
                            ui.icon('edit_calendar').on('click', lambda: tisk_menu_od_b.open()).classes('cursor-pointer text-gray-500 hover:text-blue-500')
                        with ui.menu() as tisk_menu_od_b: ui.date().bind_value(tisk_od_b)

                    with ui.input('Do data').classes('w-32 bg-white').props('outlined dense') as tisk_do_b:
                        with tisk_do_b.add_slot('append'):
                            ui.icon('edit_calendar').on('click', lambda: tisk_menu_do_b.open()).classes('cursor-pointer text-gray-500 hover:text-blue-500')
                        with ui.menu() as tisk_menu_do_b: ui.date().bind_value(tisk_do_b)

                    dnes = datetime.date.today()
                    posledni_den = calendar.monthrange(dnes.year, dnes.month)[1]
                    tisk_od_b.value = f"{dnes.year}-{dnes.month:02d}-01"
                    tisk_do_b.value = f"{dnes.year}-{dnes.month:02d}-{posledni_den:02d}"

                    def nastav_minuly_mesic_b():
                        d = datetime.date.today()
                        rok_m, mesic_m = (d.year - 1, 12) if d.month == 1 else (d.year, d.month - 1)
                        pd_m = calendar.monthrange(rok_m, mesic_m)[1]
                        tisk_od_b.value = f"{rok_m}-{mesic_m:02d}-01"
                        tisk_do_b.value = f"{rok_m}-{mesic_m:02d}-{pd_m:02d}"

                    ui.button('Předchozí měsíc', icon='history', on_click=nastav_minuly_mesic_b).classes('bg-gray-100 hover:bg-gray-200 text-gray-600 h-10 px-3 text-sm rounded-lg border border-gray-200').props('flat')

                    tisk_role_b = ui.select(vsechny_role_pro_tisk_b, value=list(vsechny_role_pro_tisk_b.keys())[0] if vsechny_role_pro_tisk_b else None, label='Oddělení').classes('w-48 bg-white').props('outlined dense')
                    tisk_jmeno_b = ui.select({'vse': 'Všichni zaměstnanci'}, value='vse', label='Zaměstnanec', with_input=True).classes('w-48 bg-white').props('outlined dense')

                    def update_jmena_b(e=None):
                        vybrane_odd = tisk_role_b.value
                        nova_jmena = {'vse': 'Všichni zaměstnanci'}
                        seznam = []
                        for data_u in vsichni_uzivatele_komplet.values():
                            odd_uzivatele = [o.strip() for o in (data_u.get('oddeleni') or 'Bez oddělení').split(',')]
                            ma_pristup_k_uzivateli = False

                            if ma_tisk_odd_vse_b or ma_pristup_vsechny_slozky:
                                ma_pristup_k_uzivateli = True
                            else:
                                if any(o in povoleny_report_oddeleni_b for o in odd_uzivatele): ma_pristup_k_uzivateli = True
                                if user_id in data_u.get('manager_id', []): ma_pristup_k_uzivateli = True
                                if data_u['id'] in moji_sledovani: ma_pristup_k_uzivateli = True

                            if ma_pristup_k_uzivateli:
                                if vybrane_odd == 'vse' or vybrane_odd in odd_uzivatele:
                                    seznam.append(data_u['jmeno_cele'])

                        for jm in sorted(seznam):
                            nova_jmena[jm] = jm
                        tisk_jmeno_b.options = nova_jmena
                        if tisk_jmeno_b.value not in nova_jmena:
                            tisk_jmeno_b.value = 'vse'
                        tisk_jmeno_b.update()

                    def update_jmena_b_a_refresh(e=None):
                        update_jmena_b(e)
                        obnovit_report_data_b.refresh()

                    tisk_role_b.on_value_change(update_jmena_b_a_refresh)
                    update_jmena_b()

                    tisk_od_b.on_value_change(lambda e: obnovit_report_data_b.refresh())
                    tisk_do_b.on_value_change(lambda e: obnovit_report_data_b.refresh())
                    tisk_jmeno_b.on_value_change(lambda e: obnovit_report_data_b.refresh())

                    typy_pro_tisk_b = {}
                    if len(povoleny_report_typy_b) > 1:
                        typy_pro_tisk_b['vse'] = 'Všechny povolené typy'

                    for t_id, t_nazev in typy_volna.items():
                        if t_nazev in povoleny_report_typy_b:
                            typy_pro_tisk_b[t_id] = t_nazev

                    if intranet_data.nacti_nastaveni_intranetu().get('presczasy_zapnuty', True):
                        typy_pro_tisk_b['presczas'] = 'Přesčas'

                    tisk_typ_b = ui.select(typy_pro_tisk_b, value=list(typy_pro_tisk_b.keys())[0] if typy_pro_tisk_b else None, label='Typ volna').classes('w-48 bg-white').props('outlined dense')
                    tisk_typ_reportu_b = ui.select({'komplet': 'Kompletní výpis (Vše)', 'zustatky': 'Pouze zůstatky dovolené'}, value='komplet', label='Druh reportu').classes('w-56 bg-white').props('outlined dense')

                    tisk_typ_b.on_value_change(lambda e: obnovit_report_data_b.refresh())
                    tisk_typ_reportu_b.on_value_change(lambda e: obnovit_report_data_b.refresh())

                    # ── Progress UI (klasický export) ────────────────────────
                    with ui.row().classes('items-center gap-3') as _export_row_b:
                        _btn_export_b_ref = {}
                        with ui.element('div').classes('flex items-center gap-2 px-3 py-2 '
                                                       'bg-blue-50 border border-blue-200 '
                                                       'rounded-lg').style('display:none') as _prog_box_b:
                            ui.icon('download').classes('text-blue-500 text-lg animate-pulse flex-shrink-0')
                            with ui.column().classes('gap-1'):
                                with ui.element('div').classes(
                                        'w-40 h-2 bg-blue-200 rounded-full overflow-hidden'):
                                    _bar_b = ui.element('div').classes(
                                        'h-full bg-blue-500 rounded-full').style('width:0%;transition:width 0.1s')
                                _pct_b = ui.label('0 %').classes('text-xs font-bold text-blue-600 text-center')

                    async def proved_export_b():
                        btn = _btn_export_b_ref.get('btn')
                        if btn: btn.disable()
                        _bar_b.style('width:0%;transition:width 0.1s')
                        _pct_b.set_text('0 %')
                        _prog_box_b.style('display:flex')

                        val_od = tisk_od_b.value
                        val_do = tisk_do_b.value
                        val_role = tisk_role_b.value
                        val_jmeno = tisk_jmeno_b.value
                        val_typ = tisk_typ_b.value
                        val_report = tisk_typ_reportu_b.value

                        async def _sim_b():
                            v = 0.0
                            while v < 0.90:
                                await asyncio.sleep(0.12)
                                v = min(v + random.uniform(0.03, 0.08), 0.90)
                                _bar_b.style(f'width:{int(v*100)}%;transition:width 0.1s')
                                _pct_b.set_text(f'{int(v * 100)} %')

                        sim = asyncio.create_task(_sim_b())

                        def _zpracuj_excel():
                            vsechna_volna = intranet_data.ziskej_vsechna_volna_kalendar(jen_budouci=False)
                            dostupna = []
                            for v in vsechna_volna:
                                if v['stav_id'] not in (2, 4): continue
                                odd_list = [o.strip() for o in (v.get('oddeleni') or 'Bez oddělení').split(',')]

                                z_jmeno_cele = f"{v['u_jmeno']} {v['u_prijmeni']}"
                                z_data = next((u for u in vsichni_uzivatele_komplet.values() if u['jmeno_cele'] == z_jmeno_cele), {})
                                z_id = z_data.get('id')

                                ma_pristup_k_zaznamu = False
                                if ma_tisk_odd_vse_b or ma_pristup_vsechny_slozky:
                                    ma_pristup_k_zaznamu = True
                                else:
                                    if any(o in povoleny_report_oddeleni_b for o in odd_list): ma_pristup_k_zaznamu = True
                                    if user_id in z_data.get('manager_id', []): ma_pristup_k_zaznamu = True
                                    if z_id in moji_sledovani: ma_pristup_k_zaznamu = True

                                if ma_pristup_k_zaznamu: dostupna.append(v)

                            vyfiltrovana = []
                            try:
                                d_od_val = datetime.datetime.strptime(val_od, '%Y-%m-%d').date() if val_od else datetime.date.min
                                d_do_val = datetime.datetime.strptime(val_do, '%Y-%m-%d').date() if val_do else datetime.date.max
                            except Exception:
                                raise ValueError('Neplatný formát data!')

                            for d in dostupna:
                                if not (d['to'] >= d_od_val and d['from'] <= d_do_val): continue

                                odd_zaznamu = d.get('oddeleni') or 'Bez oddělení'
                                if val_role != 'vse' and val_role not in [o.strip() for o in odd_zaznamu.split(',')]: continue

                                cele_jmeno = f"{d['u_jmeno']} {d['u_prijmeni']}"
                                if val_jmeno != 'vse' and cele_jmeno != val_jmeno: continue

                                typ_original = d['typ']
                                if val_typ != 'vse':
                                    hledany_typ = typy_pro_tisk_b.get(val_typ)
                                    if typ_original != hledany_typ: continue
                                else:
                                    if typ_original not in povoleny_report_typy_b: continue

                                _o_od, _o_do, _o_hod = _orez_volno_na_obdobi(d['from'], d['to'], float(d['sumaHours']), d_od_val, d_do_val)
                                vyfiltrovana.append({
                                    'Oddělení': odd_zaznamu, 'Jméno': cele_jmeno, 'Typ volna': typ_original,
                                    'Od': _o_od, 'Do': _o_do,
                                    'Čas od': _fmt_cas_hhmm(d.get('cas_od')),
                                    'Čas do': _fmt_cas_hhmm(d.get('cas_do')),
                                    'Stav': d['stav_nazev'],
                                    'Hodiny do součtu': _o_hod if d['stav_id'] == 2 else 0.0,
                                    'Hodiny': _o_hod
                                })

                            # Přidej přesčasy (při výběru všech typů nebo explicitně "Přesčas")
                            presczasy_detail_rows = []
                            if intranet_data.nacti_nastaveni_intranetu().get('presczasy_zapnuty', True) and val_typ in ('vse', 'presczas'):
                                vsechny_ot_ex = intranet_data.ziskej_presczasy(None)
                                for ot in vsechny_ot_ex:
                                    if ot['stav_id'] == 4: continue  # stornováno
                                    if not (ot['datum_do'] >= d_od_val and ot['datum_od'] <= d_do_val): continue
                                    cele_jmeno_ot = f"{ot['u_jmeno']} {ot['u_prijmeni']}"
                                    odd_ot = ot.get('oddeleni') or 'Bez oddělení'
                                    odd_ot_list = [o.strip() for o in odd_ot.split(',')]
                                    z_data_ot = next((u for u in vsichni_uzivatele_komplet.values() if u['jmeno_cele'] == cele_jmeno_ot), {})
                                    z_id_ot = z_data_ot.get('id')
                                    ma_p = False
                                    if ma_tisk_odd_vse_b or ma_pristup_vsechny_slozky: ma_p = True
                                    else:
                                        if any(o in povoleny_report_oddeleni_b for o in odd_ot_list): ma_p = True
                                        if user_id in z_data_ot.get('manager_id', []): ma_p = True
                                        if z_id_ot in moji_sledovani: ma_p = True
                                    if not ma_p: continue
                                    if val_role != 'vse' and val_role not in odd_ot_list: continue
                                    if val_jmeno != 'vse' and val_jmeno != cele_jmeno_ot: continue
                                    cas_od_s = _fmt_cas_hhmm(ot.get('cas_od'))
                                    cas_do_s = _fmt_cas_hhmm(ot.get('cas_do'))
                                    vyfiltrovana.append({
                                        'Oddělení': odd_ot, 'Jméno': cele_jmeno_ot, 'Typ volna': 'Přesčas',
                                        'Od': ot['datum_od'], 'Do': ot['datum_do'],
                                        'Čas od': cas_od_s, 'Čas do': cas_do_s,
                                        'Stav': 'Schváleno',
                                        'Hodiny do součtu': float(ot['sumaHours']),
                                        'Hodiny': float(ot['sumaHours'])
                                    })
                                    presczasy_detail_rows.append({
                                        'Oddělení': odd_ot, 'Jméno': cele_jmeno_ot,
                                        'Datum od': ot['datum_od'].strftime('%d.%m.%Y') if hasattr(ot['datum_od'], 'strftime') else str(ot['datum_od']),
                                        'Datum do': ot['datum_do'].strftime('%d.%m.%Y') if hasattr(ot['datum_do'], 'strftime') else str(ot['datum_do']),
                                        'Čas od': cas_od_s, 'Čas do': cas_do_s,
                                        'Hodiny': float(ot['sumaHours']),
                                        'Důvod': ot.get('duvod') or ''
                                    })

                            if not vyfiltrovana and val_report == 'komplet':
                                raise ValueError('Pro zadaný filtr nebyla nalezena žádná data k exportu.')

                            zustatky_vypis = []
                            for mail, udata in vsichni_uzivatele_komplet.items():
                                jm = udata['jmeno_cele']
                                odd_uziv = udata.get('oddeleni') or 'Bez oddělení'
                                odd_list = [o.strip() for o in odd_uziv.split(',')]

                                ma_pristup_k_uzivateli = False
                                if ma_tisk_odd_vse_b or ma_pristup_vsechny_slozky: ma_pristup_k_uzivateli = True
                                else:
                                    if any(o in povoleny_report_oddeleni_b for o in odd_list): ma_pristup_k_uzivateli = True
                                    if user_id in udata.get('manager_id', []): ma_pristup_k_uzivateli = True
                                    if udata['id'] in moji_sledovani: ma_pristup_k_uzivateli = True

                                if not ma_pristup_k_uzivateli: continue
                                if val_role != 'vse' and val_role not in odd_list: continue
                                if val_jmeno != 'vse' and val_jmeno != jm: continue

                                vybrano = sum(z['Hodiny do součtu'] for z in vyfiltrovana if z['Jméno'] == jm and z['Typ volna'] == 'Dovolená')
                                zaklad = udata.get('base_vacation', 160.0)
                                prevod = udata.get('carried_over_vacation', 0.0)
                                narok = zaklad + prevod
                                zustatky_vypis.append({'Oddělení': odd_uziv, 'Jméno': jm, 'Základní nárok (h)': zaklad, 'Z loňska (h)': prevod, 'Celkem nárok (h)': narok, 'Vybráno (h)': vybrano, 'Zbývá (h)': narok - vybrano})

                            # Stavba Excelu probíhá zvlášť v PROCESU (build_dochazka_xlsx);
                            # tady vracíme jen připravená picklovatelná data.
                            return vyfiltrovana, presczasy_detail_rows, zustatky_vypis

                        try:
                            # 1) gather: DB + filtrování ve VLÁKNĚ (event-loop volná)
                            vyf, presc, zust = await intranet_jobs.io(_zpracuj_excel)
                            # 2) build: stavba .xlsx v PROCESU (jiné jádro, obejde GIL)
                            vytvorena_cesta = await intranet_jobs.cpu(
                                build_dochazka_xlsx, vyf, presc, zust, val_report, val_od, val_do,
                            )
                            sim.cancel()
                            _bar_b.style('width:100%;transition:width 0.2s')
                            _pct_b.set_text('100 %')
                            await asyncio.sleep(0.6)
                            ui.download(vytvorena_cesta)
                            intranet_logger.log_activity(user_name, "Export docházky", f"Stažen klasický docházkový export")
                            ui.notify('Export byl úspěšně stažen.', type='positive', position='top')
                        except ValueError as ve:
                            sim.cancel()
                            ui.notify(str(ve), type='warning', position='top')
                        except Exception as e:
                            sim.cancel()
                            ui.notify(f'Chyba exportu: {e}', type='negative', position='top')
                        finally:
                            _prog_box_b.style('display:none')
                            btn = _btn_export_b_ref.get('btn')
                            if btn: btn.enable()

                    _btn_export_b_ref['btn'] = ui.button(
                        'Exportovat do Excelu', icon='download', on_click=proved_export_b
                    ).props('unelevated').classes('bg-blue-600 hover:bg-blue-700 text-white font-bold h-10 px-5 rounded-lg text-sm')

                @ui.refreshable
                def obnovit_report_data_b():
                    vsechna_volna = intranet_data.ziskej_vsechna_volna_kalendar(jen_budouci=False)

                    dostupna = []
                    for v in vsechna_volna:
                        if v['stav_id'] not in (2, 4): continue
                        odd_list = [o.strip() for o in (v.get('oddeleni') or 'Bez oddělení').split(',')]

                        z_jmeno_cele = f"{v['u_jmeno']} {v['u_prijmeni']}"
                        z_data = next((u for u in vsichni_uzivatele_komplet.values() if u['jmeno_cele'] == z_jmeno_cele), {})
                        z_id = z_data.get('id')

                        ma_pristup_k_zaznamu = False
                        if ma_tisk_odd_vse_b or ma_pristup_vsechny_slozky:
                            ma_pristup_k_zaznamu = True
                        else:
                            if any(o in povoleny_report_oddeleni_b for o in odd_list): ma_pristup_k_zaznamu = True
                            if user_id in z_data.get('manager_id', []): ma_pristup_k_zaznamu = True
                            if z_id in moji_sledovani: ma_pristup_k_zaznamu = True

                        if ma_pristup_k_zaznamu: dostupna.append(v)

                    if not dostupna:
                        _ui_prazdny_stav('Pro vaše oprávnění nejsou k dispozici žádná data.', 'lock')
                        return

                    vyfiltrovana = []
                    try:
                        d_od_val = datetime.datetime.strptime(tisk_od_b.value, '%Y-%m-%d').date() if tisk_od_b.value else datetime.date.min
                        d_do_val = datetime.datetime.strptime(tisk_do_b.value, '%Y-%m-%d').date() if tisk_do_b.value else datetime.date.max
                    except Exception:
                        ui.label('Chybný formát data.').classes('text-red-500 font-bold mt-4')
                        return

                    for d in dostupna:
                        if not (d['to'] >= d_od_val and d['from'] <= d_do_val): continue

                        odd_zaznamu = d.get('oddeleni') or 'Bez oddělení'
                        if tisk_role_b.value != 'vse' and tisk_role_b.value not in [o.strip() for o in odd_zaznamu.split(',')]: continue

                        cele_jmeno = f"{d['u_jmeno']} {d['u_prijmeni']}"
                        if tisk_jmeno_b.value != 'vse' and cele_jmeno != tisk_jmeno_b.value: continue

                        typ_original = d['typ']
                        if tisk_typ_b.value != 'vse':
                            hledany_typ = typy_pro_tisk_b.get(tisk_typ_b.value)
                            if typ_original != hledany_typ: continue
                        else:
                            if typ_original not in povoleny_report_typy_b: continue

                        _o_od, _o_do, _o_hod = _orez_volno_na_obdobi(d['from'], d['to'], float(d['sumaHours']), d_od_val, d_do_val)
                        vyfiltrovana.append({
                            'Oddělení': odd_zaznamu, 'Jméno': cele_jmeno, 'Typ volna': typ_original,
                            'Od': _o_od, 'Do': _o_do,
                            'Čas od': _fmt_cas_hhmm(d.get('cas_od')),
                            'Čas do': _fmt_cas_hhmm(d.get('cas_do')),
                            'Stav': d['stav_nazev'],
                            'Hodiny do součtu': _o_hod if d['stav_id'] == 2 else 0.0,
                            'Hodiny': _o_hod
                        })

                    # Přidej přesčasy do přehledu (při výběru všech typů nebo explicitně "Přesčas")
                    presczasy_pro_nahled = []
                    if intranet_data.nacti_nastaveni_intranetu().get('presczasy_zapnuty', True) and tisk_typ_b.value in ('vse', 'presczas'):
                        vsechny_ot_nb = intranet_data.ziskej_presczasy(None)
                        for ot in vsechny_ot_nb:
                            if ot['stav_id'] == 4: continue
                            if not (ot['datum_do'] >= d_od_val and ot['datum_od'] <= d_do_val): continue
                            cele_jmeno_ot = f"{ot['u_jmeno']} {ot['u_prijmeni']}"
                            odd_ot = ot.get('oddeleni') or 'Bez oddělení'
                            odd_ot_list = [o.strip() for o in odd_ot.split(',')]
                            z_data_ot = next((u for u in vsichni_uzivatele_komplet.values() if u['jmeno_cele'] == cele_jmeno_ot), {})
                            z_id_ot = z_data_ot.get('id')
                            ma_p = False
                            if ma_tisk_odd_vse_b or ma_pristup_vsechny_slozky: ma_p = True
                            else:
                                if any(o in povoleny_report_oddeleni_b for o in odd_ot_list): ma_p = True
                                if user_id in z_data_ot.get('manager_id', []): ma_p = True
                                if z_id_ot in moji_sledovani: ma_p = True
                            if not ma_p: continue
                            if tisk_role_b.value != 'vse' and tisk_role_b.value not in odd_ot_list: continue
                            if tisk_jmeno_b.value != 'vse' and tisk_jmeno_b.value != cele_jmeno_ot: continue
                            cas_od_s = _fmt_cas_hhmm(ot.get('cas_od'))
                            cas_do_s = _fmt_cas_hhmm(ot.get('cas_do'))
                            vyfiltrovana.append({
                                'Oddělení': odd_ot, 'Jméno': cele_jmeno_ot, 'Typ volna': 'Přesčas',
                                'Od': ot['datum_od'], 'Do': ot['datum_do'],
                                'Čas od': cas_od_s, 'Čas do': cas_do_s,
                                'Stav': 'Schváleno',
                                'Hodiny do součtu': float(ot['sumaHours']),
                                'Hodiny': float(ot['sumaHours'])
                            })
                            presczasy_pro_nahled.append({
                                'Oddělení': odd_ot, 'Jméno': cele_jmeno_ot,
                                'Datum od': ot['datum_od'].strftime('%d.%m.%Y') if hasattr(ot['datum_od'], 'strftime') else str(ot['datum_od']),
                                'Datum do': ot['datum_do'].strftime('%d.%m.%Y') if hasattr(ot['datum_do'], 'strftime') else str(ot['datum_do']),
                                'Čas od': cas_od_s, 'Čas do': cas_do_s,
                                'Hodiny': float(ot['sumaHours']),
                                'Důvod': ot.get('duvod') or ''
                            })

                    if not vyfiltrovana and tisk_typ_reportu_b.value == 'komplet':
                        _ui_prazdny_stav('Pro zadaný filtr nebyla nalezena žádná data.')
                        return

                    if tisk_typ_reportu_b.value == 'komplet' and vyfiltrovana:
                        df = pd.DataFrame(vyfiltrovana)
                        try:
                            od_fmt = datetime.datetime.strptime(tisk_od_b.value, '%Y-%m-%d').strftime('%d.%m.%Y') if tisk_od_b.value else '...'
                            do_fmt = datetime.datetime.strptime(tisk_do_b.value, '%Y-%m-%d').strftime('%d.%m.%Y') if tisk_do_b.value else '...'
                        except:
                            od_fmt, do_fmt = tisk_od_b.value, tisk_do_b.value

                        with ui.row().classes('w-full gap-2 mt-1 flex-wrap'):
                            _ui_stat_chip('receipt_long', str(len(df)), 'záznamů')
                            _ui_stat_chip('group', str(df['Jméno'].nunique()), 'zaměstnanců')
                            _ui_stat_chip('schedule', f"{df['Hodiny do součtu'].sum():g} h", 'do součtu')
                        _ui_nadpis_sekce('1', 'Souhrn hodin podle zaměstnanců', f'{od_fmt} – {do_fmt} · storna nejsou do součtů započítána')

                        df_souhrn = df.pivot_table(index=['Oddělení', 'Jméno'], columns='Typ volna', values='Hodiny do součtu', aggfunc='sum', fill_value=0).reset_index()
                        typy_sloupce = [c for c in df_souhrn.columns if c not in ['Oddělení', 'Jméno']]
                        df_souhrn['Celkem hodin'] = df_souhrn[typy_sloupce].sum(axis=1)

                        cols_souhrn = [{'name': 'Oddělení', 'label': 'Oddělení', 'field': 'Oddělení', 'align': 'left'}, {'name': 'Jméno', 'label': 'Jméno', 'field': 'Jméno', 'align': 'left'}]
                        for col in typy_sloupce: cols_souhrn.append({'name': col, 'label': col, 'field': col, 'align': 'right'})
                        cols_souhrn.append({'name': 'Celkem hodin', 'label': 'Celkem hodin', 'field': 'Celkem hodin', 'align': 'right', 'classes': 'font-bold bg-gray-100'})

                        ui.table(columns=cols_souhrn, rows=df_souhrn.to_dict('records')).props(_TBL_PROPS).classes(_TBL_CLASSES)

                        _ui_nadpis_sekce('2', 'Detailní výpis')
                        df_detail = df.sort_values(by=['Oddělení', 'Jméno', 'Od'])
                        df_detail['Od'] = df_detail['Od'].apply(lambda x: x.strftime('%d.%m.%Y') if hasattr(x, 'strftime') else str(x))
                        df_detail['Do'] = df_detail['Do'].apply(lambda x: x.strftime('%d.%m.%Y') if hasattr(x, 'strftime') else str(x))

                        cols_detail = [{'name': 'Oddělení', 'label': 'Oddělení', 'field': 'Oddělení', 'align': 'left'}, {'name': 'Jméno', 'label': 'Jméno', 'field': 'Jméno', 'align': 'left'}, {'name': 'Od', 'label': 'Od data', 'field': 'Od', 'align': 'left'}, {'name': 'Do', 'label': 'Do data', 'field': 'Do', 'align': 'left'}, {'name': 'Čas od', 'label': 'Čas od', 'field': 'Čas od', 'align': 'center'}, {'name': 'Čas do', 'label': 'Čas do', 'field': 'Čas do', 'align': 'center'}, {'name': 'Typ volna', 'label': 'Typ', 'field': 'Typ volna', 'align': 'left'}, {'name': 'Stav', 'label': 'Stav', 'field': 'Stav', 'align': 'left'}, {'name': 'Hodiny', 'label': 'Zadané hodiny', 'field': 'Hodiny', 'align': 'right'}]
                        ui.table(columns=cols_detail, rows=df_detail.to_dict('records')).props(_TBL_PROPS).classes(_TBL_CLASSES)

                        if presczasy_pro_nahled:
                            _ui_nadpis_sekce('3', 'Přesčasy — detail')
                            df_ot_nb = pd.DataFrame(presczasy_pro_nahled).sort_values(by=['Oddělení', 'Jméno', 'Datum od'])
                            cols_ot_nb = [
                                {'name': 'Oddělení', 'label': 'Oddělení', 'field': 'Oddělení', 'align': 'left'},
                                {'name': 'Jméno', 'label': 'Jméno', 'field': 'Jméno', 'align': 'left'},
                                {'name': 'Datum od', 'label': 'Datum od', 'field': 'Datum od', 'align': 'left'},
                                {'name': 'Datum do', 'label': 'Datum do', 'field': 'Datum do', 'align': 'left'},
                                {'name': 'Čas od', 'label': 'Čas od', 'field': 'Čas od', 'align': 'center'},
                                {'name': 'Čas do', 'label': 'Čas do', 'field': 'Čas do', 'align': 'center'},
                                {'name': 'Hodiny', 'label': 'Hodiny', 'field': 'Hodiny', 'align': 'right'},
                                {'name': 'Důvod', 'label': 'Důvod', 'field': 'Důvod', 'align': 'left'},
                            ]
                            ui.table(columns=cols_ot_nb, rows=df_ot_nb.to_dict('records')).props(_TBL_PROPS).classes(_TBL_CLASSES)

                    _ui_nadpis_sekce('4' if tisk_typ_reportu_b.value == 'komplet' else None, 'Zůstatky dovolené', 'pro zadané období')

                    zustatky_vypis = []
                    for mail, udata in vsichni_uzivatele_komplet.items():
                        jm = udata['jmeno_cele']
                        odd_uziv = udata.get('oddeleni') or 'Bez oddělení'
                        odd_list = [o.strip() for o in odd_uziv.split(',')]

                        ma_pristup_k_uzivateli = False
                        if ma_tisk_odd_vse_b or ma_pristup_vsechny_slozky: ma_pristup_k_uzivateli = True
                        else:
                            if any(o in povoleny_report_oddeleni_b for o in odd_list): ma_pristup_k_uzivateli = True
                            if user_id in udata.get('manager_id', []): ma_pristup_k_uzivateli = True
                            if udata['id'] in moji_sledovani: ma_pristup_k_uzivateli = True

                        if not ma_pristup_k_uzivateli: continue

                        if tisk_role_b.value != 'vse' and tisk_role_b.value not in odd_list: continue
                        if tisk_jmeno_b.value != 'vse' and tisk_jmeno_b.value != jm: continue

                        vybrano = sum(z['Hodiny do součtu'] for z in vyfiltrovana if z['Jméno'] == jm and z['Typ volna'] == 'Dovolená')
                        zaklad = udata.get('base_vacation', 160.0)
                        prevod = udata.get('carried_over_vacation', 0.0)
                        narok = zaklad + prevod
                        zustatky_vypis.append({'Oddělení': odd_uziv, 'Jméno': jm, 'Základní nárok (h)': zaklad, 'Z loňska (h)': prevod, 'Celkem nárok (h)': narok, 'Vybráno (h)': vybrano, 'Zbývá (h)': narok - vybrano})

                    df_zustatky = pd.DataFrame(zustatky_vypis)
                    if not df_zustatky.empty:
                        df_zustatky = df_zustatky.sort_values(by=['Oddělení', 'Jméno'])
                        cols_zustatky = [{'name': 'Oddělení', 'label': 'Oddělení', 'field': 'Oddělení', 'align': 'left'}, {'name': 'Jméno', 'label': 'Zaměstnanec', 'field': 'Jméno', 'align': 'left', 'classes': 'font-bold'}, {'name': 'Základní nárok (h)', 'label': 'Základ', 'field': 'Základní nárok (h)', 'align': 'right'}, {'name': 'Z loňska (h)', 'label': 'Převod', 'field': 'Z loňska (h)', 'align': 'right'}, {'name': 'Celkem nárok (h)', 'label': 'Celkový nárok', 'field': 'Celkem nárok (h)', 'align': 'right'}, {'name': 'Vybráno (h)', 'label': 'Schváleno / Vybráno', 'field': 'Vybráno (h)', 'align': 'right'}, {'name': 'Zbývá (h)', 'label': 'Zbývá', 'field': 'Zbývá (h)', 'align': 'right', 'classes': 'font-bold text-blue-600'}]
                        ui.table(columns=cols_zustatky, rows=df_zustatky.to_dict('records')).props(_TBL_PROPS).classes(_TBL_CLASSES)
                    else: _ui_prazdny_stav('Pro zadaný filtr nebyly nalezeny žádné zůstatky k zobrazení.', 'inbox')

                obnovit_report_data_b()

    if ma_pristup_ucetnictvi:
        povoleny_report_oddeleni_u = set(aktualni_oddeleni_dict.keys())

        povoleny_report_typy_u = set()
        if 'vse' in vsechna_prava or 'ucetni_vse' in vsechna_prava:
            povoleny_report_typy_u = set(typy_volna.values())
        else:
            for t_nazev in typy_volna.values():
                if f'ucetni_typ_{t_nazev.lower()}' in vsechna_prava:
                    povoleny_report_typy_u.add(t_nazev)

        with ui.dialog().props('maximized transition-show="slide-up" transition-hide="slide-down"') as tisk_dlg_ucetni:
            with ui.card().classes('w-full h-full p-4 bg-gray-100 text-black overflow-y-auto'):
                _ui_hlavicka_dialogu(tisk_dlg_ucetni, 'calculate', 'Účetní modul',
                                     'Exporty a reporty pro mzdové podklady',
                                     'bg-blue-100 text-blue-700')

                vsechny_role_pro_tisk_u = {'vse': 'Všechna povolená oddělení'}
                for r in sorted(povoleny_report_oddeleni_u):
                    vsechny_role_pro_tisk_u[r] = r

                with ui.row().classes('w-full gap-3 mb-4 items-end bg-white p-4 rounded-2xl border border-gray-200 shadow-sm flex-wrap'):
                    with ui.input('Od data').classes('w-32 bg-white').props('outlined dense') as tisk_od_u:
                        with tisk_od_u.add_slot('append'):
                            ui.icon('edit_calendar').on('click', lambda: tisk_menu_od_u.open()).classes('cursor-pointer text-gray-500 hover:text-blue-500')
                        with ui.menu() as tisk_menu_od_u: ui.date().bind_value(tisk_od_u)

                    with ui.input('Do data').classes('w-32 bg-white').props('outlined dense') as tisk_do_u:
                        with tisk_do_u.add_slot('append'):
                            ui.icon('edit_calendar').on('click', lambda: tisk_menu_do_u.open()).classes('cursor-pointer text-gray-500 hover:text-blue-500')
                        with ui.menu() as tisk_menu_do_u: ui.date().bind_value(tisk_do_u)

                    dnes = datetime.date.today()
                    posledni_den = calendar.monthrange(dnes.year, dnes.month)[1]
                    tisk_od_u.value = f"{dnes.year}-{dnes.month:02d}-01"
                    tisk_do_u.value = f"{dnes.year}-{dnes.month:02d}-{posledni_den:02d}"

                    def nastav_minuly_mesic_u():
                        d = datetime.date.today()
                        rok_m, mesic_m = (d.year - 1, 12) if d.month == 1 else (d.year, d.month - 1)
                        pd_m = calendar.monthrange(rok_m, mesic_m)[1]
                        tisk_od_u.value = f"{rok_m}-{mesic_m:02d}-01"
                        tisk_do_u.value = f"{rok_m}-{mesic_m:02d}-{pd_m:02d}"

                    ui.button('Předchozí měsíc', icon='history', on_click=nastav_minuly_mesic_u).classes('bg-gray-100 hover:bg-gray-200 text-gray-600 h-10 px-3 text-sm rounded-lg border border-gray-200').props('flat')

                    tisk_role_u = ui.select(vsechny_role_pro_tisk_u, value=list(vsechny_role_pro_tisk_u.keys())[0] if vsechny_role_pro_tisk_u else None, label='Oddělení').classes('w-48 bg-white').props('outlined dense')
                    tisk_jmeno_u = ui.select({'vse': 'Všichni zaměstnanci'}, value='vse', label='Zaměstnanec', with_input=True).classes('w-48 bg-white').props('outlined dense')

                    def update_jmena_u(e=None):
                        vybrane_odd = tisk_role_u.value
                        nova_jmena = {'vse': 'Všichni zaměstnanci'}
                        seznam = []
                        for data_u in vsichni_uzivatele_komplet.values():
                            odd_uzivatele = [o.strip() for o in (data_u.get('oddeleni') or 'Bez oddělení').split(',')]
                            if vybrane_odd == 'vse' or vybrane_odd in odd_uzivatele:
                                seznam.append(data_u['jmeno_cele'])
                        for jm in sorted(seznam):
                            nova_jmena[jm] = jm
                        tisk_jmeno_u.options = nova_jmena
                        if tisk_jmeno_u.value not in nova_jmena:
                            tisk_jmeno_u.value = 'vse'
                        tisk_jmeno_u.update()

                    def update_jmena_u_a_refresh(e=None):
                        update_jmena_u(e)
                        obnovit_report_data_u.refresh()

                    tisk_role_u.on_value_change(update_jmena_u_a_refresh)
                    update_jmena_u()

                    tisk_od_u.on_value_change(lambda e: obnovit_report_data_u.refresh())
                    tisk_do_u.on_value_change(lambda e: obnovit_report_data_u.refresh())
                    tisk_jmeno_u.on_value_change(lambda e: obnovit_report_data_u.refresh())

                    typy_pro_tisk_u = {}
                    if len(povoleny_report_typy_u) > 1:
                        typy_pro_tisk_u['vse'] = 'Všechny povolené typy'

                    for t_id, t_nazev in typy_volna.items():
                        if t_nazev in povoleny_report_typy_u:
                            typy_pro_tisk_u[t_id] = t_nazev

                    tisk_typ_u = ui.select(typy_pro_tisk_u, value=list(typy_pro_tisk_u.keys())[0] if typy_pro_tisk_u else None, label='Typ volna').classes('w-48 bg-white').props('outlined dense')
                    tisk_typ_reportu_u = ui.select({'komplet': 'Kompletní výpis (Vše)', 'zustatky': 'Pouze zůstatky dovolené'}, value='komplet', label='Druh reportu').classes('w-56 bg-white').props('outlined dense')

                    tisk_typ_u.on_value_change(lambda e: obnovit_report_data_u.refresh())
                    tisk_typ_reportu_u.on_value_change(lambda e: obnovit_report_data_u.refresh())

                    # ── Progress UI (účetní export) ───────────────────────────
                    with ui.row().classes('items-center gap-3') as _export_row_u:
                        _btn_export_u_ref = {}
                        with ui.element('div').classes('flex items-center gap-2 px-3 py-2 '
                                                       'bg-blue-50 border border-blue-200 '
                                                       'rounded-lg').style('display:none') as _prog_box_u:
                            ui.icon('download').classes('text-blue-500 text-lg animate-pulse flex-shrink-0')
                            with ui.column().classes('gap-1'):
                                with ui.element('div').classes(
                                        'w-40 h-2 bg-blue-200 rounded-full overflow-hidden'):
                                    _bar_u = ui.element('div').classes(
                                        'h-full bg-blue-500 rounded-full').style('width:0%;transition:width 0.1s')
                                _pct_u = ui.label('0 %').classes('text-xs font-bold text-blue-600 text-center')

                    async def proved_export_u():
                        btn = _btn_export_u_ref.get('btn')
                        if btn: btn.disable()
                        _bar_u.style('width:0%;transition:width 0.1s')
                        _pct_u.set_text('0 %')
                        _prog_box_u.style('display:flex')

                        val_od = tisk_od_u.value
                        val_do = tisk_do_u.value
                        val_role = tisk_role_u.value
                        val_jmeno = tisk_jmeno_u.value
                        val_typ = tisk_typ_u.value
                        val_report = tisk_typ_reportu_u.value

                        async def _sim_u():
                            v = 0.0
                            while v < 0.90:
                                await asyncio.sleep(0.12)
                                v = min(v + random.uniform(0.03, 0.08), 0.90)
                                _bar_u.style(f'width:{int(v*100)}%;transition:width 0.1s')
                                _pct_u.set_text(f'{int(v * 100)} %')

                        sim = asyncio.create_task(_sim_u())

                        def _zpracuj_excel_u():
                            vsechna_volna = intranet_data.ziskej_vsechna_volna_kalendar(jen_budouci=False)
                            dostupna = []
                            for v in vsechna_volna:
                                if v['stav_id'] not in (2, 4): continue
                                dostupna.append(v)

                            vyfiltrovana = []
                            try:
                                d_od_val = datetime.datetime.strptime(val_od, '%Y-%m-%d').date() if val_od else datetime.date.min
                                d_do_val = datetime.datetime.strptime(val_do, '%Y-%m-%d').date() if val_do else datetime.date.max
                            except Exception:
                                raise ValueError('Neplatný formát data!')

                            for d in dostupna:
                                if not (d['to'] >= d_od_val and d['from'] <= d_do_val): continue

                                odd_zaznamu = d.get('oddeleni') or 'Bez oddělení'
                                if val_role != 'vse' and val_role not in [o.strip() for o in odd_zaznamu.split(',')]: continue

                                cele_jmeno = f"{d['u_jmeno']} {d['u_prijmeni']}"
                                if val_jmeno != 'vse' and cele_jmeno != val_jmeno: continue

                                typ_original = d['typ']
                                if val_typ != 'vse':
                                    hledany_typ = typy_pro_tisk_u.get(val_typ)
                                    if typ_original != hledany_typ: continue
                                else:
                                    if typ_original not in povoleny_report_typy_u: continue

                                typ_volna_export = "dov." if typ_original == 'Dovolená' else typ_original

                                udata_export = next((u for u in vsichni_uzivatele_komplet.values() if u['jmeno_cele'] == cele_jmeno), {})
                                osobni_cislo = udata_export.get('id', "")

                                # Správné zpracování společnosti (list -> text)
                                spolecnosti_list = [s['nazev'] for s in udata_export.get('spolecnosti', [])]
                                spolecnost = ", ".join(spolecnosti_list) if spolecnosti_list else "Nepřiřazeno"

                                _o_od, _o_do, _o_hod = _orez_volno_na_obdobi(d['from'], d['to'], float(d['sumaHours']), d_od_val, d_do_val)
                                vyfiltrovana.append({
                                    'Společnost': spolecnost, 'Oddělení': odd_zaznamu, 'Osobní číslo': osobni_cislo, 'Jméno a příjmení': cele_jmeno, 'Typ volna': typ_volna_export,
                                    'Od': _o_od, 'Do': _o_do, 'Stav': d['stav_nazev'],
                                    'Hodiny do součtu': _o_hod if d['stav_id'] == 2 else 0.0,
                                    'Hodiny': _o_hod,
                                    'Zadáno': d.get('created_at')
                                })

                            # Přesčasy
                            if intranet_data.nacti_nastaveni_intranetu().get('presczasy_zapnuty', True) and val_typ == 'vse':
                                vsechny_ot_u = intranet_data.ziskej_presczasy(None)
                                for ot in vsechny_ot_u:
                                    if ot['stav_id'] == 4: continue
                                    if not (ot['datum_do'] >= d_od_val and ot['datum_od'] <= d_do_val): continue
                                    cele_jmeno_ot = f"{ot['u_jmeno']} {ot['u_prijmeni']}"
                                    odd_ot = ot.get('oddeleni') or 'Bez oddělení'
                                    if val_role != 'vse' and val_role not in [o.strip() for o in odd_ot.split(',')]: continue
                                    if val_jmeno != 'vse' and val_jmeno != cele_jmeno_ot: continue
                                    udata_ot = next((u for u in vsichni_uzivatele_komplet.values() if u['jmeno_cele'] == cele_jmeno_ot), {})
                                    osobni_cislo_ot = udata_ot.get('id', "")
                                    spolecnosti_ot = ", ".join([s['nazev'] for s in udata_ot.get('spolecnosti', [])]) or "Nepřiřazeno"
                                    vyfiltrovana.append({
                                        'Společnost': spolecnosti_ot, 'Oddělení': odd_ot, 'Osobní číslo': osobni_cislo_ot, 'Jméno a příjmení': cele_jmeno_ot, 'Typ volna': 'Přesčas',
                                        'Od': ot['datum_od'], 'Do': ot['datum_do'], 'Stav': 'Schváleno',
                                        'Hodiny do součtu': float(ot['sumaHours']),
                                        'Hodiny': float(ot['sumaHours'])
                                    })

                            if not vyfiltrovana and val_report == 'komplet':
                                raise ValueError('Pro zadaný filtr nebyla nalezena žádná data k exportu.')

                            zustatky_vypis = []
                            for mail, udata in vsichni_uzivatele_komplet.items():
                                jm = udata['jmeno_cele']
                                osobni_cislo = udata['id']

                                spolecnosti_list = [s['nazev'] for s in udata.get('spolecnosti', [])]
                                spolecnost = ", ".join(spolecnosti_list) if spolecnosti_list else "Nepřiřazeno"

                                odd_uziv = udata.get('oddeleni') or 'Bez oddělení'
                                odd_list = [o.strip() for o in odd_uziv.split(',')]

                                if val_role != 'vse' and val_role not in odd_list: continue
                                if val_jmeno != 'vse' and val_jmeno != jm: continue

                                vybrano = sum(z['Hodiny do součtu'] for z in vyfiltrovana if z['Jméno a příjmení'] == jm and z['Typ volna'] == 'dov.')
                                zaklad = udata.get('base_vacation', 160.0)
                                prevod = udata.get('carried_over_vacation', 0.0)
                                narok = zaklad + prevod
                                zustatky_vypis.append({'Společnost': spolecnost, 'Oddělení': odd_uziv, 'Osobní číslo': osobni_cislo, 'Jméno a příjmení': jm, 'Základní nárok (h)': f"{zaklad:g}", 'Z loňska (h)': f"{prevod:g}", 'Celkem nárok (h)': f"{narok:g}", 'Vybráno (h)': f"{vybrano:g}", 'Zbývá (h)': f"{narok - vybrano:g}"})

                            # Stavba Excelu probíhá zvlášť v PROCESU (build_dochazka_ucetni_xlsx).
                            return vyfiltrovana, zustatky_vypis

                        try:
                            # 1) gather: DB + filtrování ve VLÁKNĚ (event-loop volná)
                            vyf, zust = await intranet_jobs.io(_zpracuj_excel_u)
                            # 2) build: stavba .xlsx v PROCESU (jiné jádro, obejde GIL)
                            vytvoreni_cesta = await intranet_jobs.cpu(
                                build_dochazka_ucetni_xlsx, vyf, zust, val_report, val_od, val_do,
                            )
                            sim.cancel()
                            _bar_u.style('width:100%;transition:width 0.2s')
                            _pct_u.set_text('100 %')
                            await asyncio.sleep(0.6)
                            ui.download(vytvoreni_cesta)
                            intranet_logger.log_activity(user_name, "Export účetní", f"Stažen účetní export docházky")
                            ui.notify('Export byl úspěšně stažen.', type='positive', position='top')
                        except ValueError as ve:
                            sim.cancel()
                            ui.notify(str(ve), type='warning', position='top')
                        except Exception as e:
                            sim.cancel()
                            ui.notify(f'Chyba exportu: {e}', type='negative', position='top')
                        finally:
                            _prog_box_u.style('display:none')
                            btn = _btn_export_u_ref.get('btn')
                            if btn: btn.enable()

                    _btn_export_u_ref['btn'] = ui.button(
                        'Exportovat do Excelu', icon='download', on_click=proved_export_u
                    ).props('unelevated').classes('bg-blue-600 hover:bg-blue-700 text-white font-bold h-10 px-5 rounded-lg text-sm')

                @ui.refreshable
                def obnovit_report_data_u():
                    vsechna_volna = intranet_data.ziskej_vsechna_volna_kalendar(jen_budouci=False)

                    dostupna = []
                    for v in vsechna_volna:
                        if v['stav_id'] not in (2, 4): continue
                        dostupna.append(v)

                    if not dostupna:
                        _ui_prazdny_stav('Pro vaše oprávnění nejsou k dispozici žádná data.', 'lock')
                        return

                    vyfiltrovana = []
                    try:
                        d_od_val = datetime.datetime.strptime(tisk_od_u.value, '%Y-%m-%d').date() if tisk_od_u.value else datetime.date.min
                        d_do_val = datetime.datetime.strptime(tisk_do_u.value, '%Y-%m-%d').date() if tisk_do_u.value else datetime.date.max
                    except Exception:
                        ui.label('Chybný formát data.').classes('text-red-500 font-bold mt-4')
                        return

                    for d in dostupna:
                        if not (d['to'] >= d_od_val and d['from'] <= d_do_val): continue

                        odd_zaznamu = d.get('oddeleni') or 'Bez oddělení'
                        if tisk_role_u.value != 'vse' and tisk_role_u.value not in [o.strip() for o in odd_zaznamu.split(',')]: continue

                        cele_jmeno = f"{d['u_jmeno']} {d['u_prijmeni']}"
                        if tisk_jmeno_u.value != 'vse' and cele_jmeno != tisk_jmeno_u.value: continue

                        typ_original = d['typ']
                        if tisk_typ_u.value != 'vse':
                            hledany_typ = typy_pro_tisk_u.get(tisk_typ_u.value)
                            if typ_original != hledany_typ: continue
                        else:
                            if typ_original not in povoleny_report_typy_u: continue

                        typ_volna_export = "dov." if typ_original == 'Dovolená' else typ_original

                        udata_export = next((u for u in vsichni_uzivatele_komplet.values() if u['jmeno_cele'] == cele_jmeno), {})
                        osobni_cislo = udata_export.get('id', "")

                        spolecnosti_list = [s['nazev'] for s in udata_export.get('spolecnosti', [])]
                        spolecnost = ", ".join(spolecnosti_list) if spolecnosti_list else "Nepřiřazeno"

                        _o_od, _o_do, _o_hod = _orez_volno_na_obdobi(d['from'], d['to'], float(d['sumaHours']), d_od_val, d_do_val)
                        vyfiltrovana.append({
                            'Společnost': spolecnost, 'Oddělení': odd_zaznamu, 'Osobní číslo': osobni_cislo, 'Jméno a příjmení': cele_jmeno, 'Typ volna': typ_volna_export,
                            'Od': _o_od, 'Do': _o_do, 'Stav': d['stav_nazev'],
                            'Hodiny do součtu': _o_hod if d['stav_id'] == 2 else 0.0,
                            'Hodiny': _o_hod,
                            'Zadáno': d.get('created_at')
                        })

                    # Přesčasy
                    if intranet_data.nacti_nastaveni_intranetu().get('presczasy_zapnuty', True) and tisk_typ_u.value == 'vse':
                        vsechny_ot_u_prev = intranet_data.ziskej_presczasy(None)
                        for ot in vsechny_ot_u_prev:
                            if ot['stav_id'] == 4: continue
                            if not (ot['datum_do'] >= d_od_val and ot['datum_od'] <= d_do_val): continue
                            cele_jmeno_ot = f"{ot['u_jmeno']} {ot['u_prijmeni']}"
                            odd_ot = ot.get('oddeleni') or 'Bez oddělení'
                            if tisk_role_u.value != 'vse' and tisk_role_u.value not in [o.strip() for o in odd_ot.split(',')]: continue
                            if tisk_jmeno_u.value != 'vse' and tisk_jmeno_u.value != cele_jmeno_ot: continue
                            udata_ot = next((u for u in vsichni_uzivatele_komplet.values() if u['jmeno_cele'] == cele_jmeno_ot), {})
                            osobni_cislo_ot = udata_ot.get('id', "")
                            spolecnosti_ot = ", ".join([s['nazev'] for s in udata_ot.get('spolecnosti', [])]) or "Nepřiřazeno"
                            vyfiltrovana.append({
                                'Společnost': spolecnosti_ot, 'Oddělení': odd_ot, 'Osobní číslo': osobni_cislo_ot, 'Jméno a příjmení': cele_jmeno_ot, 'Typ volna': 'Přesčas',
                                'Od': ot['datum_od'], 'Do': ot['datum_do'], 'Stav': 'Schváleno',
                                'Hodiny do součtu': float(ot['sumaHours']),
                                'Hodiny': float(ot['sumaHours'])
                            })

                    if not vyfiltrovana and tisk_typ_reportu_u.value == 'komplet':
                        _ui_prazdny_stav('Pro zadaný filtr nebyla nalezena žádná data.')
                        return

                    if tisk_typ_reportu_u.value == 'komplet' and vyfiltrovana:
                        df = pd.DataFrame(vyfiltrovana)
                        try:
                            od_fmt = datetime.datetime.strptime(tisk_od_u.value, '%Y-%m-%d').strftime('%d.%m.%Y') if tisk_od_u.value else '...'
                            do_fmt = datetime.datetime.strptime(tisk_do_u.value, '%Y-%m-%d').strftime('%d.%m.%Y') if tisk_do_u.value else '...'
                        except:
                            od_fmt, do_fmt = tisk_od_u.value, tisk_do_u.value

                        with ui.row().classes('w-full gap-2 mt-1 flex-wrap'):
                            _ui_stat_chip('receipt_long', str(len(df)), 'záznamů')
                            _ui_stat_chip('group', str(df['Jméno a příjmení'].nunique()), 'zaměstnanců')
                            _ui_stat_chip('schedule', f"{df['Hodiny do součtu'].sum():g} h", 'do součtu')
                        _ui_nadpis_sekce('1', 'Souhrn hodin podle zaměstnanců', f'{od_fmt} – {do_fmt}')

                        df_souhrn = df.pivot_table(index=['Oddělení', 'Osobní číslo', 'Jméno a příjmení'], columns='Typ volna', values='Hodiny do součtu', aggfunc='sum', fill_value=0).reset_index()
                        typy_sloupce = [c for c in df_souhrn.columns if c not in ['Oddělení', 'Osobní číslo', 'Jméno a příjmení']]
                        df_souhrn['Celkem hodin'] = df_souhrn[typy_sloupce].sum(axis=1)

                        cols_souhrn = [{'name': 'Společnost', 'label': 'Společnost', 'field': 'Společnost', 'align': 'left'}, {'name': 'Oddělení', 'label': 'Oddělení', 'field': 'Oddělení', 'align': 'left'}, {'name': 'Osobní číslo', 'label': 'Osobní číslo', 'field': 'Osobní číslo', 'align': 'left'}, {'name': 'Jméno a příjmení', 'label': 'Jméno a příjmení', 'field': 'Jméno a příjmení', 'align': 'left'}]
                        for col in typy_sloupce: cols_souhrn.append({'name': col, 'label': col, 'field': col, 'align': 'right'})
                        cols_souhrn.append({'name': 'Celkem hodin', 'label': 'Celkem hodin', 'field': 'Celkem hodin', 'align': 'right', 'classes': 'font-bold bg-gray-100'})

                        ui.table(columns=cols_souhrn, rows=df_souhrn.to_dict('records')).props(_TBL_PROPS).classes(_TBL_CLASSES)

                        _ui_nadpis_sekce('2', 'Detailní výpis absencí a přesčasů')
                        df_detail = df.sort_values(by=['Oddělení', 'Jméno a příjmení', 'Od'])
                        df_detail['Od'] = df_detail['Od'].apply(lambda x: x.strftime('%d.%m.%Y') if hasattr(x, 'strftime') else str(x))
                        df_detail['Do'] = df_detail['Do'].apply(lambda x: x.strftime('%d.%m.%Y') if hasattr(x, 'strftime') else str(x))
                        if 'Zadáno' in df_detail.columns:
                            df_detail = df_detail.drop(columns=['Zadáno'])

                        cols_detail = [
                            {'name': 'Společnost', 'label': 'Společnost', 'field': 'Společnost', 'align': 'left'},
                            {'name': 'Oddělení', 'label': 'Oddělení', 'field': 'Oddělení', 'align': 'left'},
                            {'name': 'Osobní číslo', 'label': 'Osobní číslo', 'field': 'Osobní číslo', 'align': 'left'},
                            {'name': 'Jméno a příjmení', 'label': 'Jméno a příjmení', 'field': 'Jméno a příjmení', 'align': 'left'},
                            {'name': 'Od', 'label': 'Od data', 'field': 'Od', 'align': 'left'},
                            {'name': 'Do', 'label': 'Do data', 'field': 'Do', 'align': 'left'},
                            {'name': 'Typ volna', 'label': 'Typ', 'field': 'Typ volna', 'align': 'left'},
                            {'name': 'Stav', 'label': 'Stav', 'field': 'Stav', 'align': 'left'},
                            {'name': 'Hodiny', 'label': 'Zadané hodiny', 'field': 'Hodiny', 'align': 'right'},
                        ]
                        ui.table(columns=cols_detail, rows=df_detail.to_dict('records')).props(_TBL_PROPS).classes(_TBL_CLASSES)

                    _ui_nadpis_sekce('3' if tisk_typ_reportu_u.value == 'komplet' else None, 'Zůstatky dovolené', 'pro zadané období')

                    zustatky_vypis = []
                    for mail, udata in vsichni_uzivatele_komplet.items():
                        jm = udata['jmeno_cele']
                        osobni_cislo = udata['id']

                        spolecnosti_list = [s['nazev'] for s in udata.get('spolecnosti', [])]
                        spolecnost = ", ".join(spolecnosti_list) if spolecnosti_list else "Nepřiřazeno"

                        odd_uziv = udata.get('oddeleni') or 'Bez oddělení'
                        odd_list = [o.strip() for o in odd_uziv.split(',')]

                        if tisk_role_u.value != 'vse' and tisk_role_u.value not in odd_list: continue
                        if tisk_jmeno_u.value != 'vse' and tisk_jmeno_u.value != jm: continue

                        vybrano = sum(z['Hodiny do součtu'] for z in vyfiltrovana if z['Jméno a příjmení'] == jm and z['Typ volna'] == 'dov.')
                        zaklad = udata.get('base_vacation', 160.0)
                        prevod = udata.get('carried_over_vacation', 0.0)
                        narok = zaklad + prevod
                        zustatky_vypis.append({'Společnost': spolecnost, 'Oddělení': odd_uziv, 'Osobní číslo': osobni_cislo, 'Jméno a příjmení': jm, 'Základní nárok (h)': f"{zaklad:g}", 'Z loňska (h)': f"{prevod:g}", 'Celkem nárok (h)': f"{narok:g}", 'Vybráno (h)': f"{vybrano:g}", 'Zbývá (h)': f"{narok - vybrano:g}"})

                    df_zustatky = pd.DataFrame(zustatky_vypis)
                    if not df_zustatky.empty:
                        df_zustatky = df_zustatky.sort_values(by=['Společnost', 'Oddělení', 'Jméno a příjmení'])
                        cols_zustatky = [{'name': 'Společnost', 'label': 'Společnost', 'field': 'Společnost', 'align': 'left'}, {'name': 'Oddělení', 'label': 'Oddělení', 'field': 'Oddělení', 'align': 'left'}, {'name': 'Osobní číslo', 'label': 'Osobní číslo', 'field': 'Osobní číslo', 'align': 'left'}, {'name': 'Jméno a příjmení', 'label': 'Zaměstnanec', 'field': 'Jméno a příjmení', 'align': 'left', 'classes': 'font-bold'}, {'name': 'Základní nárok (h)', 'label': 'Základ', 'field': 'Základní nárok (h)', 'align': 'right'}, {'name': 'Z loňska (h)', 'label': 'Převod', 'field': 'Z loňska (h)', 'align': 'right'}, {'name': 'Celkem nárok (h)', 'label': 'Celkový nárok', 'field': 'Celkem nárok (h)', 'align': 'right'}, {'name': 'Vybráno (h)', 'label': 'Schváleno / Vybráno', 'field': 'Vybráno (h)', 'align': 'right'}, {'name': 'Zbývá (h)', 'label': 'Zbývá', 'field': 'Zbývá (h)', 'align': 'right', 'classes': 'font-bold text-blue-600'}]
                        ui.table(columns=cols_zustatky, rows=df_zustatky.to_dict('records')).props(_TBL_PROPS).classes(_TBL_CLASSES)
                    else: _ui_prazdny_stav('Pro zadaný filtr nebyly nalezeny žádné zůstatky k zobrazení.', 'inbox')

                    # Sekce: datum zadání dovolené a náhradního volna
                    _ui_nadpis_sekce('4' if tisk_typ_reportu_u.value == 'komplet' else None, 'Dovolená a náhradní volno', 'datum zadání')

                    zadani_radky = [z for z in vyfiltrovana if z.get('Typ volna') in ('dov.', 'Náhradní volno')]
                    if zadani_radky:
                        df_zadani_prev = pd.DataFrame(zadani_radky).sort_values(by=['Zadáno', 'Jméno a příjmení'])
                        df_zadani_prev['Zadáno'] = df_zadani_prev['Zadáno'].apply(lambda x: x.strftime('%d.%m.%Y %H:%M') if hasattr(x, 'strftime') else (str(x) if x else '—'))
                        df_zadani_prev['Od'] = df_zadani_prev['Od'].apply(lambda x: x.strftime('%d.%m.%Y') if hasattr(x, 'strftime') else str(x))
                        df_zadani_prev['Do'] = df_zadani_prev['Do'].apply(lambda x: x.strftime('%d.%m.%Y') if hasattr(x, 'strftime') else str(x))
                        df_zadani_prev = df_zadani_prev[['Společnost', 'Oddělení', 'Osobní číslo', 'Jméno a příjmení', 'Typ volna', 'Od', 'Do', 'Hodiny', 'Zadáno']]
                        cols_zadani = [
                            {'name': 'Společnost', 'label': 'Společnost', 'field': 'Společnost', 'align': 'left'},
                            {'name': 'Oddělení', 'label': 'Oddělení', 'field': 'Oddělení', 'align': 'left'},
                            {'name': 'Osobní číslo', 'label': 'Osobní číslo', 'field': 'Osobní číslo', 'align': 'left'},
                            {'name': 'Jméno a příjmení', 'label': 'Jméno a příjmení', 'field': 'Jméno a příjmení', 'align': 'left'},
                            {'name': 'Typ volna', 'label': 'Typ', 'field': 'Typ volna', 'align': 'left'},
                            {'name': 'Od', 'label': 'Od', 'field': 'Od', 'align': 'left'},
                            {'name': 'Do', 'label': 'Do', 'field': 'Do', 'align': 'left'},
                            {'name': 'Hodiny', 'label': 'Hodiny', 'field': 'Hodiny', 'align': 'right'},
                            {'name': 'Zadáno', 'label': 'Zadáno', 'field': 'Zadáno', 'align': 'left', 'classes': 'font-bold text-indigo-700'},
                        ]
                        ui.table(columns=cols_zadani, rows=df_zadani_prev.to_dict('records')).props(_TBL_PROPS).classes(_TBL_CLASSES)
                    else:
                        _ui_prazdny_stav('Pro zadaný filtr nebyly nalezeny žádné záznamy dovolené ani náhradního volna.', 'inbox')

                obnovit_report_data_u()

    # =====================================================================
    # DIALOG POROVNÁNÍ — nahrání CSV a porovnání s aktuálním stavem v systému
    # =====================================================================
    _por_oddeleni = povoleny_porovnani_oddeleni or set()
    if _por_oddeleni:
        _dnes = datetime.date.today()
        _rok_zacatek = _dnes.replace(month=1, day=1).strftime('%Y-%m-%d')
        _dnes_str = _dnes.strftime('%Y-%m-%d')
        # Možnosti oddělení pro select: 'vse' jen pokud má přístup k více než jednomu
        _por_odd_serazena = sorted(_por_oddeleni)
        _por_odd_options = ({'vse': 'Všechna povolená oddělení'} if len(_por_oddeleni) > 1 else {}) | {o: o for o in _por_odd_serazena}
        _por_odd_vychozi = 'vse' if len(_por_oddeleni) > 1 else _por_odd_serazena[0]
        stav_por = {
            'data': None,
            'chyba': None,
            'datum_od': _rok_zacatek,
            'datum_do': _dnes_str,
            'oddeleni': _por_odd_vychozi,
        }

        with ui.dialog().props('maximized') as tisk_dlg_porovnani, ui.card().classes('w-full h-full rounded-none p-0 bg-gray-50'):

            # --- Záhlaví dialogu ---
            with ui.row().classes('w-full items-center justify-between bg-white border-b border-gray-200 px-6 py-4 shadow-sm'):
                with ui.row().classes('items-center gap-3'):
                    ui.icon('compare', size='md', color='indigo-600')
                    ui.label('Porovnání zůstatků dovolené').classes('text-2xl font-extrabold text-gray-800')
                ui.button(icon='close', on_click=tisk_dlg_porovnani.close).props('flat round').classes('text-gray-500')

            # --- Tělo dialogu (refreshable) ---
            @ui.refreshable
            def obsah_porovnani():
                if stav_por['data'] is None:
                    # ---- FÁZE 1: nahrání souboru + výběr období ----
                    with ui.column().classes('w-full h-full items-center justify-center gap-6 p-10'):
                        ui.icon('upload_file', size='5rem', color='indigo-400')
                        ui.label('Nahrajte CSV soubor s daty k porovnání').classes('text-xl font-bold text-gray-700')
                        ui.label('Soubor musí obsahovat sloupce ID (= osobní číslo) a DOV (= zůstatek dovolené v hodinách).').classes('text-gray-500 text-center max-w-lg')

                        # --- Výběr období ---
                        with ui.card().classes('p-5 rounded-xl border border-indigo-100 bg-white shadow-sm w-full max-w-lg'):
                            ui.label('Období pro výpočet vybráno').classes('text-sm font-bold text-indigo-700 uppercase tracking-wide mb-3')
                            ui.label('Systém sečte schválené dovolené s termínem zahájení v tomto rozsahu a porovná je se zůstatkem v CSV.').classes('text-xs text-gray-400 mb-4')
                            with ui.row().classes('w-full gap-4 items-end'):
                                with ui.column().classes('flex-1'):
                                    ui.label('Od').classes('text-xs text-gray-500 mb-1')
                                    inp_od = ui.input(value=stav_por['datum_od']).props('outlined dense type=date').classes('w-full')
                                    inp_od.on_value_change(lambda e: stav_por.update({'datum_od': e.value}))
                                with ui.column().classes('flex-1'):
                                    ui.label('Do').classes('text-xs text-gray-500 mb-1')
                                    inp_do = ui.input(value=stav_por['datum_do']).props('outlined dense type=date').classes('w-full')
                                    inp_do.on_value_change(lambda e: stav_por.update({'datum_do': e.value}))

                            # Rychlé zkratky pro výběr roku
                            with ui.row().classes('gap-2 mt-3 flex-wrap'):
                                ui.label('Rychlý výběr:').classes('text-xs text-gray-400 self-center')
                                for _r in range(_dnes.year, _dnes.year - 3, -1):
                                    def _nastav_rok(rok=_r):
                                        stav_por['datum_od'] = f'{rok}-01-01'
                                        stav_por['datum_do'] = f'{rok}-12-31'
                                        obsah_porovnani.refresh()
                                    ui.button(str(_r), on_click=_nastav_rok).props('flat dense size=sm').classes('text-indigo-600 font-bold border border-indigo-200 rounded')

                        # --- Výběr oddělení (jen pokud má přístup k více oddělením) ---
                        if len(_por_oddeleni) > 1:
                            with ui.card().classes('p-5 rounded-xl border border-indigo-100 bg-white shadow-sm w-full max-w-lg'):
                                ui.label('Oddělení').classes('text-sm font-bold text-indigo-700 uppercase tracking-wide mb-3')
                                ui.label('Zobrazit porovnání pouze pro vybrané oddělení, nebo pro všechna povolená najednou.').classes('text-xs text-gray-400 mb-3')
                                sel_odd = ui.select(
                                    options=_por_odd_options,
                                    value=stav_por['oddeleni'],
                                    label='Oddělení'
                                ).props('outlined dense').classes('w-full')
                                sel_odd.on_value_change(lambda e: stav_por.update({'oddeleni': e.value}))

                        if stav_por['chyba']:
                            ui.label(f'⚠ {stav_por["chyba"]}').classes('text-red-600 font-bold')

                        async def zpracovat_csv(e):
                            try:
                                raw = await e.file.read()
                                try:
                                    text = raw.decode('utf-8-sig')
                                except UnicodeDecodeError:
                                    try:
                                        text = raw.decode('cp1250')
                                    except Exception:
                                        stav_por['chyba'] = 'Nepodařilo se přečíst soubor. Zkontrolujte kódování (UTF-8 nebo Windows-1250).'
                                        obsah_porovnani.refresh()
                                        return
                            except Exception as ex:
                                stav_por['chyba'] = f'Chyba při čtení souboru: {ex}'
                                obsah_porovnani.refresh()
                                return

                            try:
                                # Detekce oddělovače
                                vzorek = text[:2048]
                                oddelovac = ';' if vzorek.count(';') >= vzorek.count(',') else ','
                                reader = _csv_mod.DictReader(_io.StringIO(text), delimiter=oddelovac)

                                # Najdi sloupce ID a DOV (case-insensitive, whitespace-tolerant)
                                col_id = next((f for f in (reader.fieldnames or []) if f.strip().upper() == 'ID'), None)
                                col_dov = next((f for f in (reader.fieldnames or []) if f.strip().upper() == 'DOV'), None)

                                if col_id is None or col_dov is None:
                                    stav_por['chyba'] = f'CSV neobsahuje požadované sloupce ID a DOV. Nalezené sloupce: {", ".join(reader.fieldnames or [])}'
                                    stav_por['data'] = None
                                    obsah_porovnani.refresh()
                                    return

                                csv_data = {}
                                for radek in reader:
                                    try:
                                        uid = int(str(radek[col_id]).strip())
                                        dov_str = str(radek[col_dov]).strip().replace(',', '.')
                                        dov = float(dov_str)
                                        csv_data[uid] = dov
                                    except (ValueError, KeyError):
                                        continue

                                if not csv_data:
                                    stav_por['chyba'] = 'Soubor neobsahuje žádná platná data.'
                                    stav_por['data'] = None
                                    obsah_porovnani.refresh()
                                    return

                                stav_por['chyba'] = None
                                stav_por['data'] = csv_data
                                obsah_porovnani.refresh()
                            except Exception as ex:
                                stav_por['chyba'] = f'Chyba při zpracování: {ex}'
                                stav_por['data'] = None
                                obsah_porovnani.refresh()

                        ui.upload(label='Vybrat soubor (.csv)', on_upload=zpracovat_csv, auto_upload=True).props('accept=.csv flat color=indigo').classes('w-80')

                else:
                    # ---- FÁZE 2: porovnání ----
                    csv_data = stav_por['data']

                    # Parsování zvoleného rozsahu datumů
                    try:
                        por_od = datetime.date.fromisoformat(stav_por['datum_od'])
                    except Exception:
                        por_od = datetime.date.today().replace(month=1, day=1)
                    try:
                        por_do = datetime.date.fromisoformat(stav_por['datum_do'])
                    except Exception:
                        por_do = datetime.date.today()

                    popis_obdobi = f"{por_od.strftime('%d.%m.%Y')} – {por_do.strftime('%d.%m.%Y')}"

                    # Filtrovat schválené dovolené v zadaném období (podle data zahájení)
                    vsechny_zadosti = intranet_data.ziskej_zadosti()
                    vybrano_obdobi = {}  # user_id → hodiny
                    for z in vsechny_zadosti:
                        if z.get('stav_id') == 2 and z.get('typ') == 'Dovolená':
                            try:
                                datum_z = z['from'] if isinstance(z['from'], datetime.date) else datetime.date.fromisoformat(str(z['from'])[:10])
                            except Exception:
                                continue
                            if por_od <= datum_z <= por_do:
                                uid_z = z.get('user_iduser')
                                if uid_z:
                                    vybrano_obdobi[uid_z] = vybrano_obdobi.get(uid_z, 0.0) + float(z.get('sumaHours', 0) or 0)

                    # Sestavit řádky pro porovnání — pouze povolená oddělení
                    zvolene_oddeleni = stav_por.get('oddeleni', 'vse')
                    radky = []
                    for mail, udata in vsichni_uzivatele_komplet.items():
                        uid = udata.get('id')
                        if uid is None:
                            continue
                        # Filtr podle práv: uživatel musí být v aspoň jednom povoleném oddělení
                        odd_uziv = [o.strip() for o in (udata.get('oddeleni') or '').split(',') if o.strip()]
                        if not any(o in _por_oddeleni for o in odd_uziv):
                            continue
                        # Filtr podle zvoleného oddělení v UI
                        if zvolene_oddeleni != 'vse' and zvolene_oddeleni not in odd_uziv:
                            continue
                        jmeno = udata.get('jmeno_cele', '')
                        zaklad = float(udata.get('base_vacation', 160.0) or 160.0)
                        prevod = float(udata.get('carried_over_vacation', 0.0) or 0.0)
                        vybrano = vybrano_obdobi.get(uid, 0.0)
                        zbyvajici_sys = zaklad + prevod - vybrano
                        z_csv = csv_data.get(uid)
                        if z_csv is None:
                            shoda = None  # není v CSV
                        else:
                            shoda = abs(zbyvajici_sys - z_csv) <= 0.05
                        radky.append({
                            'uid': uid, 'jmeno': jmeno,
                            'oddeleni': ', '.join(odd_uziv),
                            'zaklad': zaklad, 'prevod': prevod, 'vybrano': vybrano,
                            'zbyvajici_sys': zbyvajici_sys,
                            'z_csv': z_csv, 'shoda': shoda
                        })

                    radky.sort(key=lambda r: (r['shoda'] is not False, r['jmeno']))

                    # Statistiky
                    pocet_shoda = sum(1 for r in radky if r['shoda'] is True)
                    pocet_rozdil = sum(1 for r in radky if r['shoda'] is False)
                    pocet_chybi = sum(1 for r in radky if r['shoda'] is None)

                    # --- Horní lišta ---
                    with ui.row().classes('w-full items-center justify-between px-6 py-3 bg-white border-b border-gray-200 gap-4 flex-wrap'):
                        with ui.row().classes('gap-4 items-center flex-wrap'):
                            with ui.row().classes('items-center gap-2 bg-indigo-50 px-3 py-1 rounded-lg'):
                                ui.icon('date_range', size='sm', color='indigo-500')
                                ui.label(f'Období: {popis_obdobi}').classes('text-indigo-700 font-bold text-sm')
                            # Badge / select oddělení
                            if len(_por_oddeleni) > 1:
                                def _zmen_oddeleni_v2(e):
                                    stav_por['oddeleni'] = e.value
                                    obsah_porovnani.refresh()
                                ui.select(
                                    options=_por_odd_options,
                                    value=stav_por.get('oddeleni', 'vse'),
                                    label='Oddělení'
                                ).props('outlined dense').classes('min-w-[180px] bg-white').on_value_change(_zmen_oddeleni_v2)
                            else:
                                odd_nazev = _por_odd_serazena[0] if _por_odd_serazena else '—'
                                with ui.row().classes('items-center gap-2 bg-gray-50 px-3 py-1 rounded-lg border border-gray-200'):
                                    ui.icon('business', size='sm', color='gray-500')
                                    ui.label(odd_nazev).classes('text-gray-700 font-bold text-sm')
                            # Filtr klikacími čítači
                            filtr_por = {'hodnota': 'vse'}
                            telo_por = ui.element('div')  # reference na scrollovatelné tělo, definujeme níže

                            def _prekreslit_telo():
                                telo_por.clear()
                                with telo_por:
                                    _vykreslit_radky(filtr_por['hodnota'])

                            def _btn_filtr(label, hodnota, pocet, barva_btn, barva_text):
                                aktivni = filtr_por['hodnota'] == hodnota
                                styl = f'border-2 {"border-" + barva_btn + " bg-" + barva_btn.replace("-600","").replace("-400","") + "-50" if aktivni else "border-gray-200 bg-white"}'
                                def _klik(h=hodnota):
                                    filtr_por['hodnota'] = h
                                    _prekreslit_telo()
                                btn = ui.button(f'{label}: {pocet}', on_click=_klik).props('flat dense').classes(
                                    f'px-3 py-1 rounded-lg text-sm font-bold {barva_text} border-2 {"border-" + barva_btn if aktivni else "border-gray-200"}')
                                return btn

                            btn_vse    = _btn_filtr('Vše', 'vse', len(radky), 'gray-300', 'text-gray-600')
                            btn_shoda  = _btn_filtr('✅ Shoduje se', 'shoda', pocet_shoda, 'green-400', 'text-green-700')
                            btn_rozdil = _btn_filtr('❌ Neshoduje se', 'rozdil', pocet_rozdil, 'red-400', 'text-red-600')
                            btn_chybi  = _btn_filtr('⚪ Chybí v CSV', 'chybi', pocet_chybi, 'gray-400', 'text-gray-500')

                        def _reset_csv():
                            stav_por['data'] = None
                            stav_por['chyba'] = None
                            obsah_porovnani.refresh()

                        async def _exportovat():
                            # Sestavit řádky podle aktivního filtru
                            filtr = filtr_por['hodnota']
                            export_radky = []
                            for r in radky:
                                if filtr == 'shoda'  and r['shoda'] is not True:  continue
                                if filtr == 'rozdil' and r['shoda'] is not False: continue
                                if filtr == 'chybi'  and r['shoda'] is not None:  continue
                                stav_str = 'Shoduje se' if r['shoda'] is True else ('Neshoduje se' if r['shoda'] is False else 'Chybí v CSV')
                                rozdil_val = round(r['zbyvajici_sys'] - r['z_csv'], 2) if r['z_csv'] is not None else None
                                export_radky.append({
                                    'Osobní číslo':    r['uid'],
                                    'Jméno a příjmení': r['jmeno'],
                                    'Základ (h)':       r['zaklad'],
                                    'Převod (h)':       r['prevod'],
                                    f'Vybráno {por_od.year} (h)': r['vybrano'],
                                    'Zbývá – systém (h)': round(r['zbyvajici_sys'], 2),
                                    'DOV – CSV (h)':    r['z_csv'] if r['z_csv'] is not None else '',
                                    'Rozdíl (h)':       rozdil_val if rozdil_val is not None else '',
                                    'Stav':             stav_str,
                                })
                            if not export_radky:
                                ui.notify('Žádná data k exportu.', type='warning')
                                return

                            nazvy_filtru = {'vse': 'Vse', 'shoda': 'Shoduje_se', 'rozdil': 'Neshoduje_se', 'chybi': 'Chybi_v_CSV'}
                            nazev_souboru = f"Porovnani_dovolene_{nazvy_filtru.get(filtr, filtr)}_{por_od.strftime('%Y%m%d')}_{por_do.strftime('%Y%m%d')}.xlsx"

                            # Stavbu Excelu pustíme do odděleného PROCESU (jiné
                            # jádro) – UI ostatních uživatelů zůstane svižné i
                            # při velkém exportu. (Fallback na vlákno je v jobs.cpu.)
                            try:
                                cesta = await intranet_jobs.cpu(
                                    build_porovnani_xlsx, export_radky, nazev_souboru,
                                )
                            except Exception as ex:
                                ui.notify(f'Chyba při tvorbě exportu: {ex}', type='negative')
                                return
                            ui.download(cesta)
                            ui.notify(f'Export stažen: {nazev_souboru}', type='positive')

                        with ui.row().classes('gap-2'):
                            ui.button('Exportovat', icon='download', on_click=_exportovat).props('flat').classes('text-green-700 font-bold')
                            ui.button('Nahrát jiný soubor', icon='upload_file', on_click=_reset_csv).props('flat').classes('text-indigo-600')

                    # --- Split pohled (jeden scroll, obě strany synchronizované) ---
                    # Lepkavý záhlaví se sloupci
                    with ui.row().classes('w-full gap-0 bg-gray-100 border-b border-gray-300'):
                        with ui.row().classes('flex-1 px-3 py-2 gap-0 border-r border-gray-300'):
                            ui.label('Aktuální stav v systému').classes('flex-1 text-xs font-extrabold text-gray-600 uppercase tracking-wide')
                            ui.label('Základ').classes('w-16 text-right text-xs font-bold text-gray-500')
                            ui.label('Převod').classes('w-16 text-right text-xs font-bold text-gray-500')
                            ui.label(f'Vybráno ({por_od.year})').classes('w-20 text-right text-xs font-bold text-gray-500')
                            ui.label('Zbývá').classes('w-20 text-right text-xs font-extrabold text-blue-600')
                        with ui.row().classes('flex-1 px-3 py-2 gap-0'):
                            ui.label('Nahraný soubor (CSV)').classes('flex-1 text-xs font-extrabold text-gray-600 uppercase tracking-wide')
                            ui.label('DOV (h) z CSV').classes('w-28 text-right text-xs font-bold text-gray-500')
                            ui.label('Rozdíl').classes('w-24 text-right text-xs font-bold text-gray-500')

                    # Scrollovatelné tělo — jeden posuvník, oba sloupce se posouvají společně
                    telo_por = ui.element('div').classes('w-full overflow-y-auto').style('height: calc(100vh - 210px)')

                    def _vykreslit_radky(filtr):
                        zobrazit = []
                        for r in radky:
                            if filtr == 'shoda'  and r['shoda'] is not True:  continue
                            if filtr == 'rozdil' and r['shoda'] is not False: continue
                            if filtr == 'chybi'  and r['shoda'] is not None:  continue
                            zobrazit.append(r)

                        if not zobrazit:
                            with ui.column().classes('w-full items-center py-16 gap-2'):
                                ui.icon('search_off', size='3rem', color='gray-300')
                                ui.label('Žádné záznamy neodpovídají zvolenému filtru.').classes('text-gray-400 text-base')
                            return

                        for r in zobrazit:
                            bg    = 'bg-green-50'  if r['shoda'] is True  else ('bg-red-50'  if r['shoda'] is False else 'bg-white')
                            ind_l = 'border-l-4 border-green-400' if r['shoda'] is True else ('border-l-4 border-red-400' if r['shoda'] is False else 'border-l-4 border-gray-200')
                            ind_r = 'border-l-4 border-green-300' if r['shoda'] is True else ('border-l-4 border-red-300' if r['shoda'] is False else 'border-l-4 border-gray-100')

                            with ui.row().classes(f'w-full gap-0 {bg} border-b border-gray-100 hover:brightness-95'):
                                # Levá polovina – systémová data
                                with ui.row().classes(f'flex-1 px-3 py-2 items-center gap-0 {ind_l} border-r border-gray-200'):
                                    ui.label(str(r['uid'])).classes('w-12 text-xs text-gray-400 shrink-0')
                                    ui.label(r['jmeno']).classes('flex-1 font-semibold text-sm text-gray-800 truncate')
                                    ui.label(f"{r['zaklad']:g}").classes('w-16 text-right text-sm text-gray-500 shrink-0')
                                    ui.label(f"{r['prevod']:g}").classes('w-16 text-right text-sm text-gray-500 shrink-0')
                                    ui.label(f"{r['vybrano']:g}").classes('w-20 text-right text-sm text-gray-500 shrink-0')
                                    ui.label(f"{r['zbyvajici_sys']:g} h").classes('w-20 text-right text-sm font-bold text-blue-700 shrink-0')

                                # Pravá polovina – CSV data
                                with ui.row().classes(f'flex-1 px-3 py-2 items-center gap-0 {ind_r}'):
                                    ui.label(str(r['uid'])).classes('w-12 text-xs text-gray-400 shrink-0')
                                    ui.label(r['jmeno']).classes('flex-1 font-semibold text-sm text-gray-800 truncate')
                                    if r['z_csv'] is not None:
                                        barva_dov  = 'text-green-700' if r['shoda'] else 'text-red-600'
                                        barva_diff = 'text-green-600' if r['shoda'] else 'text-red-500 font-bold'
                                        ui.label(f"{r['z_csv']:g} h").classes(f'w-28 text-right text-sm font-bold {barva_dov} shrink-0')
                                        rozdil = r['zbyvajici_sys'] - r['z_csv']
                                        znak = '+' if rozdil > 0 else ''
                                        ui.label(f"{znak}{rozdil:g} h").classes(f'w-24 text-right text-xs {barva_diff} shrink-0')
                                    else:
                                        ui.label('—').classes('w-28 text-right text-gray-300 text-sm shrink-0')
                                        ui.label('').classes('w-24 shrink-0')

                    with telo_por:
                        _vykreslit_radky('vse')

            obsah_porovnani()

    # =====================================================================
    # IKOS EXPORT VOLNA
    # =====================================================================
    if ma_ikos_export:
        ikos_vsechna_odd = {}
        if len(aktualni_oddeleni_dict) > 1:
            ikos_vsechna_odd['vse'] = 'Všechna oddělení'
        for _r in sorted(aktualni_oddeleni_dict.keys()):
            ikos_vsechna_odd[_r] = _r

        with ui.dialog().props('maximized transition-show="slide-up" transition-hide="slide-down"') as ikos_dlg:
            with ui.card().classes('w-full h-full p-4 bg-gray-100 text-black overflow-y-auto'):
                _ui_hlavicka_dialogu(ikos_dlg, 'output', 'IKOS Export volna',
                                     'CSV export schválených voln pro mzdový systém IKOS',
                                     'bg-teal-100 text-teal-700')

                with ui.row().classes('w-full gap-3 mb-6 items-end bg-white p-4 rounded-2xl border border-gray-200 shadow-sm flex-wrap'):
                    with ui.input('Od data').classes('w-32 bg-white').props('outlined dense') as ikos_od:
                        with ikos_od.add_slot('append'):
                            ui.icon('edit_calendar').on('click', lambda: ikos_menu_od.open()).classes('cursor-pointer text-gray-500 hover:text-blue-500')
                        with ui.menu() as ikos_menu_od:
                            ui.date().bind_value(ikos_od)

                    with ui.input('Do data').classes('w-32 bg-white').props('outlined dense') as ikos_do:
                        with ikos_do.add_slot('append'):
                            ui.icon('edit_calendar').on('click', lambda: ikos_menu_do.open()).classes('cursor-pointer text-gray-500 hover:text-blue-500')
                        with ui.menu() as ikos_menu_do:
                            ui.date().bind_value(ikos_do)

                    _dnes_i = datetime.date.today()
                    ikos_od.value = f"{_dnes_i.year}-{_dnes_i.month:02d}-01"
                    ikos_do.value = f"{_dnes_i.year}-{_dnes_i.month:02d}-{calendar.monthrange(_dnes_i.year, _dnes_i.month)[1]:02d}"

                    def nastav_minuly_mesic_ikos():
                        d = datetime.date.today()
                        y, m = (d.year - 1, 12) if d.month == 1 else (d.year, d.month - 1)
                        ikos_od.value = f"{y}-{m:02d}-01"
                        ikos_do.value = f"{y}-{m:02d}-{calendar.monthrange(y, m)[1]:02d}"

                    ui.button('Předchozí měsíc', icon='history', on_click=nastav_minuly_mesic_ikos).classes('bg-gray-100 hover:bg-gray-200 text-gray-600 h-10 px-3 text-sm rounded-lg border border-gray-200').props('flat')

                    ikos_role_sel = ui.select(ikos_vsechna_odd, value=list(ikos_vsechna_odd.keys())[0] if ikos_vsechna_odd else None, label='Oddělení').classes('w-48 bg-white').props('outlined dense')
                    ikos_jmeno_sel = ui.select({'vse': 'Všichni zaměstnanci'}, value='vse', label='Zaměstnanec', with_input=True).classes('w-48 bg-white').props('outlined dense')

                    ikos_sloucit_cb = ui.checkbox(
                        'Sloučit příznak s os. číslem',
                        value=intranet_data.nacti_nastaveni_intranetu().get('ikos_sloucit_priznak', False),
                    ).classes('font-bold text-teal-700')

                    def update_ikos_jmena(e=None):
                        vybr = ikos_role_sel.value
                        opts = {'vse': 'Všichni zaměstnanci'}
                        for _ud in vsichni_uzivatele_komplet.values():
                            _odd_l = [o.strip() for o in (_ud.get('oddeleni') or 'Bez oddělení').split(',')]
                            if vybr == 'vse' or vybr in _odd_l:
                                opts[_ud['jmeno_cele']] = _ud['jmeno_cele']
                        ikos_jmeno_sel.options = dict(sorted(opts.items(), key=lambda x: (x[0] != 'vse', x[0])))
                        if ikos_jmeno_sel.value not in ikos_jmeno_sel.options:
                            ikos_jmeno_sel.value = 'vse'
                        ikos_jmeno_sel.update()

                    ikos_role_sel.on_value_change(lambda e: update_ikos_jmena(e))
                    update_ikos_jmena()

                    _btn_ikos_ref = {}
                    with ui.element('div').classes('flex items-center gap-2 px-3 py-2 bg-teal-50 border border-teal-200 rounded-lg').style('display:none') as _prog_ikos:
                        ui.icon('download').classes('text-teal-500 text-lg animate-pulse flex-shrink-0')
                        with ui.column().classes('gap-1'):
                            with ui.element('div').classes('w-40 h-2 bg-teal-200 rounded-full overflow-hidden'):
                                _bar_ikos = ui.element('div').classes('h-full bg-teal-500 rounded-full').style('width:0%;transition:width 0.1s')
                            _pct_ikos = ui.label('0 %').classes('text-xs font-bold text-teal-600 text-center')

                    async def proved_ikos_export():
                        btn = _btn_ikos_ref.get('btn')
                        if btn: btn.disable()
                        _bar_ikos.style('width:0%;transition:width 0.1s')
                        _pct_ikos.set_text('0 %')
                        _prog_ikos.style('display:flex')

                        val_od = ikos_od.value
                        val_do = ikos_do.value
                        val_role = ikos_role_sel.value
                        val_jmeno = ikos_jmeno_sel.value
                        val_sloucit = ikos_sloucit_cb.value

                        async def _sim_ikos():
                            v = 0.0
                            while v < 0.90:
                                await asyncio.sleep(0.12)
                                v = min(v + random.uniform(0.04, 0.10), 0.90)
                                _bar_ikos.style(f'width:{int(v*100)}%;transition:width 0.1s')
                                _pct_ikos.set_text(f'{int(v*100)} %')

                        sim = asyncio.create_task(_sim_ikos())

                        def _zpracuj_ikos():
                            # gather: jen DB čtení (běží ve vláknu)
                            nast_i = intranet_data.nacti_nastaveni_intranetu()
                            oddelovac = nast_i.get('ikos_oddelovac', ';')
                            slozka = nast_i.get('ikos_slozka', 'Exporty_Dochazka')
                            vsechna_volna = intranet_data.ziskej_vsechna_volna_kalendar(jen_budouci=False)
                            return vsechna_volna, oddelovac, slozka

                        try:
                            # 1) gather (DB) ve VLÁKNĚ; 2) filtr + zápis CSV v PROCESU
                            vsechna_volna, oddelovac, slozka = await intranet_jobs.io(_zpracuj_ikos)
                            cesta = await intranet_jobs.cpu(
                                build_ikos_csv, vsechna_volna, vsichni_uzivatele_komplet,
                                val_role, val_jmeno, val_od, val_do, val_sloucit, oddelovac, slozka,
                            )
                            sim.cancel()
                            _bar_ikos.style('width:100%;transition:width 0.2s')
                            _pct_ikos.set_text('100 %')
                            await asyncio.sleep(0.6)
                            ui.download(cesta)
                            intranet_logger.log_activity(user_name, "IKOS Export", "Stažen IKOS export docházky")
                            ui.notify('IKOS Export úspěšně stažen.', type='positive', position='top')
                        except ValueError as ve:
                            sim.cancel()
                            ui.notify(str(ve), type='warning', position='top')
                        except Exception as e:
                            sim.cancel()
                            ui.notify(f'Chyba IKOS exportu: {e}', type='negative', position='top')
                        finally:
                            _prog_ikos.style('display:none')
                            btn = _btn_ikos_ref.get('btn')
                            if btn: btn.enable()

                    _btn_ikos_ref['btn'] = ui.button('Exportovat IKOS CSV', icon='download', on_click=proved_ikos_export).props('unelevated').classes('bg-teal-600 hover:bg-teal-700 text-white font-bold h-10 px-5 rounded-lg text-sm')

                with ui.card().classes('p-4 bg-white border border-gray-200 rounded-2xl mt-4 shadow-sm'):
                    with ui.row().classes('items-center gap-2 mb-1'):
                        ui.icon('info', size='xs').classes('text-teal-600')
                        ui.label('Formát souboru').classes('text-sm font-bold text-teal-800')
                    ui.label('Os. číslo  ·  Datum od  ·  Datum do  ·  Čas od  ·  Čas do  ·  Druh volna').classes('text-sm text-teal-700 font-mono bg-teal-50 px-3 py-1.5 rounded-lg')
                    ui.label('Export zahrnuje pouze schválená volna všech typů. Oddělovač a cílová složka jsou v Nastavení portálu → Exporty → IKOS Export.').classes('text-xs text-gray-500 mt-2 italic')

    # =====================================================================
    # IMPORT REÁLNÝCH ZŮSTATKŮ DOVOLENÉ
    # =====================================================================
    ma_import_dovolene = 'vse' in vsechna_prava or 'import_dovolene' in vsechna_prava
    if ma_import_dovolene:
        with ui.dialog() as import_dov_dlg, ui.card().classes('w-[560px] p-6 rounded-2xl'):
            with ui.row().classes('items-center gap-3 mb-1'):
                with ui.element('div').classes('w-10 h-10 rounded-xl bg-emerald-100 text-emerald-700 flex items-center justify-center'):
                    ui.icon('upload_file', size='sm')
                ui.label('Import reálných zůstatků dovolené').classes('text-xl font-bold text-gray-800')
            ui.label('Nahrajte soubor .xlsx se sloupci „Osobní číslo" (příznak + číslo, např. JV319935) a „Zůstatek". Spárování probíhá podle osobního čísla a ověřuje se shoda příznaku se systémem.').classes('text-sm text-gray-500 mb-4')

            import_stav = {'vysledek': None, 'datum': datetime.date.today().strftime('%Y-%m-%d')}

            with ui.column().classes('w-full mb-4'):
                ui.label('Zůstatek platí ke dni *').classes('text-xs font-bold text-gray-600 uppercase tracking-wide mb-1')
                ui.label('Datum, ke kterému jsou zůstatky v souboru platné. Zobrazí se zaměstnancům u reálného zůstatku.').classes('text-xs text-gray-400 mb-2')
                inp_datum_dov = ui.input(value=import_stav['datum']).props('outlined dense type=date').classes('w-full')
                inp_datum_dov.on_value_change(lambda e: import_stav.update({'datum': e.value}))

            @ui.refreshable
            def _import_vysledek():
                if import_stav['vysledek']:
                    barva, text = import_stav['vysledek']
                    ui.label(text).classes(f'text-sm font-bold {barva} whitespace-pre-line mt-2')

            async def _zpracuj_import(e):
                try:
                    datum_ke_dni = datetime.date.fromisoformat(import_stav['datum'])
                except (ValueError, TypeError):
                    import_stav['vysledek'] = ('text-red-600', 'CHYBA: Zadejte platné datum, ke kterému zůstatky platí.')
                    _import_vysledek.refresh()
                    return
                nazev = getattr(e, 'name', '') or ''
                if nazev and not nazev.lower().endswith(('.xlsx', '.xls')):
                    import_stav['vysledek'] = ('text-red-600', 'CHYBA: Nahrajte soubor ve formátu .xlsx.')
                    _import_vysledek.refresh()
                    return
                try:
                    zdroj = getattr(e, 'content', None) or getattr(e, 'file', None)
                    cteni = zdroj.read()
                    raw = await cteni if asyncio.iscoroutine(cteni) else cteni
                    # Parsování .xlsx (CPU) do procesu – nezdrží ostatní uživatele.
                    listy = await intranet_jobs.cpu(read_excel_vsechny_listy, raw)
                except Exception as ex:
                    import_stav['vysledek'] = ('text-red-600', f'Chyba při čtení souboru: {ex}')
                    _import_vysledek.refresh()
                    return

                def _norm(s):
                    return str(s).strip().lower().replace('\xa0', ' ')

                mapa = {}
                preskoceno = 0
                for df in listy.values():
                    # Sloupec D – „Osobní číslo" (příznak + číslo, např. JV319935)
                    col_id = next((c for c in df.columns if _norm(c).startswith('os') and ('íslo' in _norm(c) or 'islo' in _norm(c))), None)
                    # Sloupec se zůstatkem
                    col_z = next((c for c in df.columns if 'statek' in _norm(c) or _norm(c).startswith('zůstat') or _norm(c).startswith('zustat')), None)
                    if col_id is None or col_z is None:
                        continue
                    for _, radek in df.iterrows():
                        try:
                            if pd.isna(radek[col_id]) or pd.isna(radek[col_z]):
                                preskoceno += 1
                                continue
                            # Z hodnoty „JV319935" vytáhneme číselnou část (= iduser)
                            # a písmenný příznak („JV").
                            surovy = str(radek[col_id]).strip()
                            cislice = _re.sub(r'\D', '', surovy)
                            if not cislice:
                                preskoceno += 1
                                continue
                            os_cislo = int(cislice)
                            priznak_f = _re.sub(r'\d', '', surovy).strip()
                            zustatek = float(str(radek[col_z]).strip().replace('\xa0', '').replace(' ', '').replace(',', '.'))
                            mapa[os_cislo] = {'zustatek': zustatek, 'priznak': priznak_f}
                        except (ValueError, TypeError):
                            preskoceno += 1

                if not mapa:
                    import_stav['vysledek'] = ('text-red-600', 'Soubor neobsahuje platná data (potřebné sloupce: „Osobní číslo" a „Zůstatek").')
                    _import_vysledek.refresh()
                    return

                # Náhled – spočítáme, kolik záznamů se spáruje se systémem (bez zápisu).
                datum_cz = datum_ke_dni.strftime('%d.%m.%Y')
                shodne, nenalezeno_n, _ = intranet_data.import_realnych_zustatku_dovolene(mapa, datum_ke_dni, nahled=True)
                if nenalezeno_n is None:
                    import_stav['vysledek'] = ('text-red-600', 'Chyba při čtení databáze.')
                    _import_vysledek.refresh()
                    return

                # Data jsou načtena – před zápisem se zeptáme na potvrzení přehrání DB.
                import_stav['cekajici'] = {'mapa': mapa, 'datum': datum_ke_dni, 'datum_cz': datum_cz, 'preskoceno': preskoceno, 'shodne': shodne}
                import_stav['vysledek'] = None
                _import_vysledek.refresh()
                _potvrdit_telo.refresh()
                potvrdit_dlg.open()

            async def _proved_zapis():
                ceka = import_stav.get('cekajici')
                if not ceka:
                    return
                potvrdit_dlg.close()
                import_stav['cekajici'] = None
                mapa = ceka['mapa']
                datum_ke_dni = ceka['datum']
                datum_cz = ceka['datum_cz']
                preskoceno = ceka['preskoceno']

                aktualizovano, nenalezeno, nesedi_priznak = intranet_data.import_realnych_zustatku_dovolene(mapa, datum_ke_dni)
                if nenalezeno is None:
                    import_stav['vysledek'] = ('text-red-600', 'Chyba při ukládání do databáze.')
                    _import_vysledek.refresh()
                    return
                intranet_logger.log_activity(user_name, "Import dovolené", f"Aktualizováno {aktualizovano} záznamů reálného zůstatku dovolené ke dni {datum_cz} (nespárováno {len(nenalezeno)}, neshoda příznaku {len(nesedi_priznak)})")
                hlaska = f'✅ Úspěšně importováno: {aktualizovano} zaměstnanců (ke dni {datum_cz}).'
                if nenalezeno:
                    hlaska += f'\n⚠ Nespárováno (osobní číslo není v systému): {len(nenalezeno)} – {", ".join(str(x) for x in nenalezeno[:20])}{" …" if len(nenalezeno) > 20 else ""}'
                if nesedi_priznak:
                    hlaska += f'\n⚠ Neimportováno (příznak v souboru nesouhlasí se systémem): {len(nesedi_priznak)} – {", ".join(nesedi_priznak[:20])}{" …" if len(nesedi_priznak) > 20 else ""}'
                if preskoceno:
                    hlaska += f'\nPřeskočeno prázdných/neplatných řádků: {preskoceno}.'
                import_stav['vysledek'] = ('text-green-700', hlaska)
                _import_vysledek.refresh()
                ui.notify(f'Importováno {aktualizovano} zůstatků dovolené.', type='positive', position='top')

            # Potvrzovací dialog – přehrání databáze aktuálními daty ze souboru.
            with ui.dialog() as potvrdit_dlg, ui.card().classes('w-[440px] p-6 rounded-2xl'):
                ui.label('Přehrát databázi aktuálními daty?').classes('text-lg font-bold text-gray-800 mb-1')

                @ui.refreshable
                def _potvrdit_telo():
                    ceka = import_stav.get('cekajici')
                    pocet = ceka['shodne'] if ceka else 0
                    datum_txt = ceka['datum_cz'] if ceka else ''
                    ui.label(f'Se systémem se spárovalo {pocet} záznamů (zůstatky ke dni {datum_txt}).').classes('text-sm text-gray-600')
                    ui.label('Tato akce přehraje stávající reálné zůstatky dovolené v databázi aktuálními daty ze souboru. Chcete pokračovat?').classes('text-sm text-gray-500 mt-2')
                _potvrdit_telo()

                with ui.row().classes('w-full justify-end mt-4 gap-2'):
                    ui.button('Zrušit', on_click=potvrdit_dlg.close).props('flat').classes('text-gray-600 font-bold')
                    ui.button('Ano, přehrát', on_click=_proved_zapis).classes('bg-emerald-500 hover:bg-emerald-600 text-white font-bold px-6')

            ui.upload(label='Vybrat soubor (.xlsx)', on_upload=_zpracuj_import, auto_upload=True).props('accept=".xlsx,.xls" flat color=green').classes('w-full')
            _import_vysledek()

            with ui.row().classes('w-full justify-end mt-4'):
                ui.button('Zavřít', on_click=import_dov_dlg.close).props('flat').classes('text-gray-600 font-bold px-6')

    # =====================================================================
    # TLAČÍTKA EXPORTŮ
    # =====================================================================
    if ma_pristup_ucetnictvi:
        ui.button('Účetní oddělení', icon='calculate', on_click=tisk_dlg_ucetni.open).props('unelevated').classes('bg-blue-50 hover:bg-blue-100 text-blue-700 border border-blue-200 rounded-lg px-4 h-10 font-bold shadow-sm')

    if ma_import_dovolene:
        ui.button('Import dovolené', icon='upload_file', on_click=import_dov_dlg.open).props('unelevated').classes('bg-emerald-50 hover:bg-emerald-100 text-emerald-700 border border-emerald-200 rounded-lg px-4 h-10 font-bold shadow-sm')

    if ukazat_tlacitko_export:
        ui.button('Export dat', icon='download', on_click=tisk_dlg_bezny.open).props('unelevated').classes('bg-purple-50 hover:bg-purple-100 text-purple-700 border border-purple-200 rounded-lg px-4 h-10 font-bold shadow-sm')

    if _por_oddeleni:
        ui.button('Porovnání', icon='compare', on_click=tisk_dlg_porovnani.open).props('unelevated').classes('bg-indigo-50 hover:bg-indigo-100 text-indigo-700 border border-indigo-200 rounded-lg px-4 h-10 font-bold shadow-sm')

    if ma_ikos_export:
        ui.button('IKOS Export', icon='output', on_click=ikos_dlg.open).props('unelevated').classes('bg-teal-50 hover:bg-teal-100 text-teal-700 border border-teal-200 rounded-lg px-4 h-10 font-bold shadow-sm')


def proved_automaticky_export_ikos():
    """Automatický IKOS export – voláno z background tasku v web_main.py."""
    try:
        nast = intranet_data.nacti_nastaveni_intranetu()
        oddelovac = nast.get('ikos_oddelovac', ';')
        slozka = nast.get('ikos_slozka', 'Exporty_Dochazka')
        prijemci = nast.get('ikos_prijemci', [])
        sloucit_priznak = nast.get('ikos_sloucit_priznak', False)

        vsichni = intranet_data.ziskej_vsechny_uzivatele()
        vsechna_volna = intranet_data.ziskej_vsechna_volna_kalendar(jen_budouci=False)

        dnes = datetime.date.today()
        pondeli_tohoto = dnes - datetime.timedelta(days=dnes.weekday())
        d_od = pondeli_tohoto - datetime.timedelta(days=7)   # pondělí předchozího týdne
        d_do = pondeli_tohoto - datetime.timedelta(days=1)   # neděle předchozího týdne

        radky = []
        for v in vsechna_volna:
            if v['stav_id'] != 2: continue
            if not (v['to'] >= d_od and v['from'] <= d_do): continue
            cele_jmeno = f"{v['u_jmeno']} {v['u_prijmeni']}"
            udata = next((u for u in vsichni.values() if u['jmeno_cele'] == cele_jmeno), {})
            os_cislo = str(udata.get('id', '')).zfill(6)
            if sloucit_priznak:
                priznak = udata.get('priznak_nazev', '') or ''
                if priznak:
                    os_cislo = f'{priznak}{os_cislo}'
            datum_od_s = v['from'].strftime('%d.%m.%Y') if hasattr(v['from'], 'strftime') else str(v['from'])
            datum_do_s = v['to'].strftime('%d.%m.%Y') if hasattr(v['to'], 'strftime') else str(v['to'])
            cas_od_s = _ikos_fmt_cas(v.get('cas_od'))
            cas_do_s = _ikos_fmt_cas(v.get('cas_do'))
            radky.append([os_cislo, datum_od_s, datum_do_s, cas_od_s, cas_do_s, v['typ']])

        if not radky:
            return False, 'Žádná schválená volna pro předchozí týden.'

        os.makedirs(slozka, exist_ok=True)
        cesta = os.path.join(slozka, f"IKOS_Export_{time.strftime('%Y%m%d_%H%M%S')}.csv")
        with open(cesta, 'w', newline='', encoding='utf-8-sig') as f:
            w = _csv_mod.writer(f, delimiter=oddelovac)
            w.writerow(['Os. číslo', 'Datum od', 'Datum do', 'Čas od', 'Čas do', 'Druh volna'])
            w.writerows(radky)

        # Rozeslat e-mailem příjemcům
        if prijemci:
            import intranet_emaily as _em
            obdobi_str = f"{d_od.strftime('%d.%m.%Y')} – {d_do.strftime('%d.%m.%Y')}"
            predmet = f"IKOS Export volna – {obdobi_str}"
            text = (f"Dobrý den,\n\nv příloze zasíláme automatický IKOS export volna za období {obdobi_str} "
                    f"({len(radky)} záznamů).\n\nS pozdravem,\nMoje JIPka")
            for prijemce in prijemci:
                try:
                    _em.odesli_email_s_prilohou(prijemce, predmet, text, cesta)
                except Exception as e_mail:
                    print(f'[ikos_exporty] Chyba e-mailu ({prijemce}): {e_mail}')

        return True, cesta
    except Exception as e:
        return False, str(e)