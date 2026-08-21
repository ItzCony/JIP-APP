from nicegui import ui, app
import intranet_data
import intranet_emaily
import intranet_notifikace
import intranet_prava
import intranet_logger
import intranet_narozeniny
import intranet_okruhy
import intranet_2fa
import intranet_lupa
import os
import zipfile
import pandas as pd
import datetime
import time
import calendar
import openpyxl
import mysql.connector
import decimal
import re
import asyncio
import csv
import io
import pyzipper
import tempfile

# České řazení: č/ř/š/ž jsou samostatná písmena, 'ch' po 'h'; á/é/í… řadí k základnímu písmenu.
# ponytail: vlastní klíč místo locale.strxfrm — locale 'cs' nemusí být na serveru dostupná
_CZ_PRIM = 'a b c č d e f g h ch i j k l m n o p q r ř s š t u v w x y z ž'.split()
_CZ_SEK = {'á': 'a', 'ď': 'd', 'é': 'e', 'ě': 'e', 'í': 'i', 'ň': 'n', 'ó': 'o', 'ť': 't', 'ú': 'u', 'ů': 'u', 'ý': 'y'}

def cz_razeni(text):
    t = (text or '').lower()
    klic = []
    i = 0
    while i < len(t):
        if t[i:i + 2] == 'ch':
            klic.append((_CZ_PRIM.index('ch'), 0))
            i += 2
            continue
        c = t[i]
        base = _CZ_SEK.get(c, c)
        if base in _CZ_PRIM:
            klic.append((_CZ_PRIM.index(base), 1 if c in _CZ_SEK else 0))
        else:
            klic.append((100 + ord(c), 0))  # čísla/symboly za písmena
        i += 1
    return klic

def formatuj_datum(d):
    if isinstance(d, datetime.date) or isinstance(d, datetime.datetime):
        return d.strftime('%d.%m.%Y')
    return str(d)

def formatuj_cas(d):
    if isinstance(d, datetime.datetime):
        return d.strftime('%d.%m.%Y v %H:%M')
    return ""

def ziskej_statni_svatky(rok, mesic):
    svatky = {
        (1, 1): "Den obnovy samost. státu",
        (1, 5): "Svátek práce",
        (8, 5): "Den vítězství",
        (5, 7): "Cyril a Metoděj",
        (6, 7): "Upálení mistra Jana Husa",
        (28, 9): "Den české státnosti",
        (28, 10): "Vznik samostatného státu",
        (17, 11): "Den boje za svobodu",
        (24, 12): "Štědrý den",
        (25, 12): "1. svátek vánoční",
        (26, 12): "2. svátek vánoční"
    }

    a = rok % 19
    b = rok // 100
    c = rok % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    velikonocni_mesic = (h + l - 7 * m + 114) // 31
    velikonocni_den = ((h + l - 7 * m + 114) % 31) + 1

    velikonocni_nedele = datetime.date(rok, velikonocni_mesic, velikonocni_den)
    velky_patek = velikonocni_nedele - datetime.timedelta(days=2)
    velikonocni_pondeli = velikonocni_nedele + datetime.timedelta(days=1)

    svatky[(velky_patek.day, velky_patek.month)] = "Velký pátek"
    svatky[(velikonocni_pondeli.day, velikonocni_pondeli.month)] = "Velikonoční pondělí"

    vysledek = {}
    for (d, m), nazev in svatky.items():
        if m == mesic:
            vysledek[d] = nazev
    return vysledek

def je_statni_svatek(datum):
    svatky_mesice = ziskej_statni_svatky(datum.year, datum.month)
    return datum.day in svatky_mesice

# ── Ikony kategorií práv (heuristika podle názvu kategorie) ───────────────────
_KAT_IKONY = (
    ('administrace', 'admin_panel_settings'),
    ('docházka',     'event_available'),
    ('složky',       'folder_shared'),
    ('manažer',      'manage_accounts'),
    ('export',       'import_export'),
    ('porovnání',    'compare'),
    ('kvíz',         'school'),
    ('účetní',       'account_balance'),
    ('veletrh',      'storefront'),
    ('aprovia',      'shopping_cart'),
    ('značky',       'how_to_vote'),
    ('prodejní',     'point_of_sale'),
    ('mzdy',         'payments'),
    ('narozeniny',   'cake'),
    ('směn',         'calendar_month'),
    ('komunika',     'forum'),
    ('plánogram',    'grid_on'),
    ('porady',       'task_alt'),
    ('úkol',         'task_alt'),
    ('výsledky',     'leaderboard'),
    ('sankce',       'gavel'),
    ('společenský',  'celebration'),
    ('vizitky',      'badge'),
    ('cenopřípad',   'sell'),
    ('helpdesk',     'support_agent'),
)

def _ikona_kategorie(kat: str) -> str:
    nizky = (kat or '').lower()
    for frag, ico in _KAT_IKONY:
        if frag in nizky:
            return ico
    return 'widgets'


def render_prava_kategorie(zakladni_prava, vybrana_prava_list, lazy=False):
    """Profesionální výběr práv ve stylu master–detail.

    Vlevo seznam kategorií s živými počítadly „vybráno/celkem", vpravo se
    vykreslují pouze práva právě zvolené kategorie (případně výsledky hledání).
    Díky tomu se nestaví stovky řádků najednou jako dřív → rychlejší a přehlednější.

    Vrací živou množinu vybraných klíčů (set) — volající ji čte přes
    ",".join(...) při ukládání, takže rozhraní zůstává zpětně kompatibilní.
    Klíče, které nejsou v katalogu (např. práva skrytých modulů), zůstávají
    v množině zachována, i když pro ně není řádek.
    """
    vybrane_set = set(vybrana_prava_list)

    # Seskupení katalogu podle kategorií.
    kategorie: dict = {}
    for k, v in zakladni_prava.items():
        kat = v.get('kategorie', 'Ostatní')
        kategorie.setdefault(kat, {})[k] = v
    kat_nazvy = sorted(kategorie.keys())

    if not kat_nazvy:
        ui.label('Žádná dostupná práva k přiřazení.').classes('text-gray-400 italic text-sm')
        return vybrane_set

    # Podskupiny uvnitř kategorie (např. Značky: Produkt / Provoz) – přepínač
    # nad seznamem práv. Práva bez 'podskupina' se zobrazují vždy.
    podskupiny: dict = {}
    for kat, prava in kategorie.items():
        vp = []
        for v in prava.values():
            p = v.get('podskupina')
            if p and p not in vp:
                vp.append(p)
        if len(vp) > 1:
            podskupiny[kat] = vp

    stav = {'kat': kat_nazvy[0], 'hledat': '',
            'pod': {k: v[0] for k, v in podskupiny.items()}}

    def _prava_kat(kat):
        """Práva kategorie po filtru aktivní podskupiny."""
        pod = stav['pod'].get(kat)
        if not pod:
            return kategorie[kat]
        return {k: v for k, v in kategorie[kat].items()
                if v.get('podskupina') in (None, pod)}

    # Pre-deklarace pro nonlocal v _postav_ui (lazy build) — closury níže
    # (_aktualizuj_total, vyber_kat) tyto názvy čtou.
    banner = total_label = search = None

    cat_row: dict = {}
    cat_icon: dict = {}
    cat_badge: dict = {}

    # ── Sdílené třídy ──────────────────────────────────────────────────────────
    _kat_base = ('w-full items-center gap-2 px-3 py-2.5 rounded-lg cursor-pointer '
                 'select-none transition-colors flex-nowrap')
    _kat_on = 'bg-blue-600 text-white shadow-sm'
    _kat_off = 'text-gray-700 hover:bg-gray-100'
    _card_base = ('w-full items-center gap-3 p-3 rounded-xl border cursor-pointer '
                  'select-none transition-all duration-150 flex-nowrap')
    _card_on = 'bg-emerald-50 border-emerald-300 shadow-sm'
    _card_off = 'bg-white border-gray-200 hover:bg-gray-50 hover:border-gray-300'

    def _pocet(kat):
        return sum(1 for kk in kategorie[kat] if kk in vybrane_set)

    def _aktualizuj_badge(kat):
        sel, tot = _pocet(kat), len(kategorie[kat])
        je_akt = (kat == stav['kat'] and not stav['hledat'])
        if je_akt:
            cls = 'bg-emerald-400 text-white' if sel else 'bg-white text-blue-700'
        else:
            cls = 'bg-emerald-100 text-emerald-700' if sel else 'bg-gray-100 text-gray-400'
        cat_badge[kat].set_text(f'{sel}/{tot}')
        cat_badge[kat].classes(
            replace=f'text-[11px] font-extrabold rounded-full px-2 py-0.5 shrink-0 {cls}')

    def _aktualizuj_total():
        celkem = sum(1 for kk in zakladni_prava if kk in vybrane_set)
        total_label.set_text(f'{celkem} vybráno')
        banner.set_visibility('vse' in vybrane_set)

    def _zvyrazni_kat():
        # Pozn.: NIKDY nepoužívat classes(replace=...) na ui.row – smaže to i
        # interní třídu „nicegui-row" (display:flex) a obsah by se poskládal pod sebe.
        for k in kat_nazvy:
            je = (k == stav['kat'] and not stav['hledat'])
            if je:
                cat_row[k].classes(remove=_kat_off, add=_kat_on)
                cat_icon[k].classes(remove='text-gray-400', add='text-white')
            else:
                cat_row[k].classes(remove=_kat_on, add=_kat_off)
                cat_icon[k].classes(remove='text-white', add='text-gray-400')
            _aktualizuj_badge(k)

    def _radek_prava(kat, key, meta):
        je = key in vybrane_set
        row = ui.row().classes(_card_base + ' ' + (_card_on if je else _card_off))
        with row:
            ico = ui.icon(meta.get('ikona', 'label_important'), size='sm') \
                    .classes('shrink-0 ' + ('text-emerald-600' if je else 'text-gray-400'))
            with ui.column().classes('flex-1 min-w-0 gap-0'):
                ui.label(meta.get('nazev', key)).classes('font-bold text-gray-800 text-sm leading-tight')
                if meta.get('popis'):
                    ui.label(meta['popis']).classes('text-xs text-gray-500 leading-tight')
        def toggle(e=None, key=key, kat=kat, row=row, ico=ico):
            # Jen prohození stavových tříd (remove/add) – žádné replace, ať
            # zůstane „nicegui-row" a text neskáče pod ikonu.
            if key in vybrane_set:
                vybrane_set.discard(key)
                row.classes(remove=_card_on, add=_card_off)
                ico.classes(remove='text-emerald-600', add='text-gray-400')
            else:
                vybrane_set.add(key)
                row.classes(remove=_card_off, add=_card_on)
                ico.classes(remove='text-gray-400', add='text-emerald-600')
            _aktualizuj_badge(kat)
            _aktualizuj_total()

        row.on('click', toggle)

    def vybrat_vse_kat(kat):
        for k in _prava_kat(kat):
            vybrane_set.add(k)
        _aktualizuj_badge(kat)
        _aktualizuj_total()
        render_detail.refresh()

    def zrusit_kat(kat):
        for k in _prava_kat(kat):
            vybrane_set.discard(k)
        _aktualizuj_badge(kat)
        _aktualizuj_total()
        render_detail.refresh()

    def vycistit_vse():
        for k in zakladni_prava:          # jen klíče katalogu – skrytá práva zůstanou
            vybrane_set.discard(k)
        for k in kat_nazvy:
            _aktualizuj_badge(k)
        _aktualizuj_total()
        render_detail.refresh()

    @ui.refreshable
    def render_detail():
        hledat = stav['hledat'].lower()

        # ── Režim hledání: ploché výsledky napříč kategoriemi ─────────────────
        if hledat:
            vysledky = [(kat, k, v)
                        for kat in kat_nazvy for k, v in kategorie[kat].items()
                        if hledat in (v.get('nazev') or '').lower()
                        or hledat in (v.get('popis') or '').lower()]
            if not vysledky:
                with ui.column().classes('w-full items-center justify-center py-12 gap-2 text-gray-400'):
                    ui.icon('search_off', size='lg')
                    ui.label('Nic nenalezeno.').classes('text-sm font-medium')
                return
            with ui.column().classes('w-full gap-2 max-h-[48vh] md:max-h-[58vh] overflow-y-auto pr-1'):
                ui.label(f'{len(vysledky)} výsledků').classes('text-xs text-gray-400 px-1')
                aktualni = None
                for kat, k, v in vysledky:
                    if kat != aktualni:
                        aktualni = kat
                        ui.label(kat).classes('text-[11px] font-extrabold uppercase tracking-wider '
                                              'text-blue-400 mt-2 px-1')
                    _radek_prava(kat, k, v)
            return

        # ── Režim kategorie ───────────────────────────────────────────────────
        kat = stav['kat']
        with ui.row().classes('w-full items-center justify-between gap-2 mb-2 px-1 flex-nowrap'):
            with ui.row().classes('items-center gap-2 min-w-0 flex-nowrap'):
                ui.icon(_ikona_kategorie(kat), size='sm').classes('text-blue-700 shrink-0')
                ui.label(kat).classes('font-extrabold text-blue-900 text-base truncate')
            with ui.row().classes('items-center gap-1 shrink-0 flex-nowrap'):
                ui.button('Vše', icon='done_all', on_click=lambda k=kat: vybrat_vse_kat(k)) \
                  .props('flat dense size=sm color=positive')
                ui.button('Zrušit', icon='remove_done', on_click=lambda k=kat: zrusit_kat(k)) \
                  .props('flat dense size=sm color=grey')
        if kat in podskupiny:
            def _pod(p, k=kat):
                stav['pod'][k] = p
                render_detail.refresh()
            with ui.row().classes('items-center gap-1 mb-2 px-1 flex-nowrap'):
                for p in podskupiny[kat]:
                    akt = stav['pod'].get(kat) == p
                    ui.button(p, on_click=lambda _=None, p=p: _pod(p)) \
                      .props('dense no-caps size=sm ' + ('unelevated' if akt else 'flat outline')) \
                      .classes('rounded-lg px-3 ' + ('bg-blue-600 text-white' if akt
                                                     else 'text-gray-600'))
        with ui.column().classes('w-full gap-2 max-h-[44vh] md:max-h-[54vh] overflow-y-auto pr-1'):
            for k, v in _prava_kat(kat).items():
                _radek_prava(kat, k, v)

    def on_search(e):
        stav['hledat'] = (e.value or '').strip()
        _zvyrazni_kat()
        render_detail.refresh()

    def vyber_kat(k):
        stav['kat'] = k
        if stav['hledat']:
            search.set_value('')          # vyvolá on_search → zruší hledání a překreslí
        else:
            _zvyrazni_kat()
            render_detail.refresh()

    # ── Sestavení UI ────────────────────────────────────────────────────────────
    def _postav_ui():
        nonlocal banner, total_label, search
        with ui.column().classes('w-full gap-3'):

            # Banner SuperAdmina (zobrazí se jen když je vybráno právo „vse").
            banner = ui.row().classes('w-full items-center gap-2 bg-amber-50 border border-amber-300 '
                                      'rounded-xl px-4 py-2 flex-nowrap')
            with banner:
                ui.icon('local_police', size='sm').classes('text-amber-600 shrink-0')
                ui.label('SuperAdmin — uživatel má neomezený přístup ke všem modulům. '
                         'Ostatní práva se ignorují.').classes('text-sm font-bold text-amber-800')

            # Nástrojová lišta: hledání + počítadlo + vyčistit.
            with ui.row().classes('w-full items-center gap-3 flex-wrap'):
                search = ui.input(placeholder='Hledat právo podle názvu nebo popisu…', on_change=on_search) \
                           .props('outlined dense clearable debounce=200') \
                           .classes('flex-1 min-w-[220px] bg-white')
                with ui.element('div').classes('flex items-center gap-1.5 bg-blue-50 border border-blue-200 '
                                               'rounded-lg px-3 py-1.5 shrink-0'):
                    ui.icon('verified_user', size='18px').classes('text-blue-600')
                    total_label = ui.label('0 vybráno').classes('text-sm font-extrabold text-blue-700 whitespace-nowrap')
                ui.button('Vyčistit vše', icon='backspace', on_click=vycistit_vse) \
                  .props('flat dense').classes('text-red-500 shrink-0')

            # Dvoupanel: vlevo kategorie, vpravo práva (na mobilu pod sebou).
            with ui.element('div').classes('w-full flex flex-col md:flex-row gap-4 items-start'):
                with ui.column().classes('w-full md:w-72 md:shrink-0 gap-1 max-h-[40vh] md:max-h-[60vh] '
                                         'overflow-y-auto p-2 bg-gray-50 border border-gray-200 rounded-xl'):
                    for kat in kat_nazvy:
                        r = ui.row().classes(_kat_base + ' ' + _kat_off)
                        with r:
                            ic = ui.icon(_ikona_kategorie(kat), size='sm').classes('shrink-0 text-gray-400')
                            ui.label(kat).classes('flex-1 text-sm font-bold leading-tight min-w-0')
                            bd = ui.label('0/0').classes('text-[11px] font-extrabold rounded-full px-2 py-0.5 '
                                                         'shrink-0 bg-gray-100 text-gray-400')
                        r.on('click', lambda e=None, k=kat: vyber_kat(k))
                        cat_row[kat], cat_icon[kat], cat_badge[kat] = r, ic, bd

                with ui.element('div').classes('flex-1 min-w-0 w-full'):
                    render_detail()

        # Počáteční stav (po vytvoření všech prvků).
        _zvyrazni_kat()
        _aktualizuj_total()

    # Lazy režim: těžké DOM (kategorie + detail) se postaví až při prvním
    # rozbalení. Set vybraných práv ale existuje hned → ukládání funguje
    # i bez rozbalení (vrátí výchozí výběr).
    if lazy:
        _pocet0 = sum(1 for kk in zakladni_prava if kk in vybrane_set)
        with ui.expansion(
            f'Osobní práva navíc — rozbalte pro výběr'
            + (f' ({_pocet0} vybráno)' if _pocet0 else ''),
            icon='admin_panel_settings',
        ).classes('w-full border border-slate-200 rounded-xl bg-white shadow-sm') as _exp:
            _telo = ui.column().classes('w-full pt-2')
        _st = {'done': False}
        def _lazy_open(ev):
            if ev.value and not _st['done']:
                _st['done'] = True
                with _telo:
                    _postav_ui()
        _exp.on_value_change(_lazy_open)
    else:
        _postav_ui()

    return vybrane_set

def vykresli_prirazena_prava(prava_keys, zakladni_prava, varianta='osobni'):
    """Přehledně vykreslí přiřazená práva uživatele — seskupená podle kategorií,
    s ikonou a celým popisem v tooltipu.

    prava_keys:     iterovatelná množina klíčů práv (např. {'kviz', 'smeny_admin'})
    zakladni_prava: katalog z intranet_prava.ziskej_kompletni_seznam_prav
    varianta:       'osobni' (zelená) | 'zdedena' (fialová) | 'oddeleni' (modrá)
    """
    if varianta == 'zdedena':
        chip_cls, ico_cls, kat_cls = (
            'bg-violet-50 text-violet-800 border-violet-200', 'text-violet-500', 'text-violet-400')
    elif varianta == 'oddeleni':
        chip_cls, ico_cls, kat_cls = (
            'bg-blue-50 text-blue-800 border-blue-100', 'text-blue-500', 'text-blue-400')
    else:
        chip_cls, ico_cls, kat_cls = (
            'bg-emerald-50 text-emerald-800 border-emerald-200', 'text-emerald-500', 'text-emerald-500')

    keys = [k for k in prava_keys if k]

    # SuperAdmin přebíjí všechna ostatní práva — zvýrazníme ho samostatně.
    if 'vse' in keys:
        meta = zakladni_prava.get('vse') if isinstance(zakladni_prava.get('vse'), dict) else {}
        with ui.element('div').classes('inline-flex items-center gap-2 bg-amber-100 text-amber-900 '
                                       'border border-amber-300 rounded-lg px-3 py-1.5 text-sm font-bold mt-1'):
            ui.icon(meta.get('ikona', 'local_police'), size='18px').classes('text-amber-600')
            ui.label(meta.get('nazev', 'Plný přístup (SuperAdmin)'))
        return

    # Seskupení podle kategorií katalogu.
    skupiny = {}
    for k in keys:
        meta = zakladni_prava.get(k) if isinstance(zakladni_prava.get(k), dict) else None
        kat = (meta or {}).get('kategorie', 'Ostatní')
        skupiny.setdefault(kat, []).append((k, meta))

    with ui.column().classes('w-full gap-2 mt-1'):
        for kat in sorted(skupiny):
            with ui.column().classes('w-full gap-1'):
                ui.label(kat).classes(f'text-[10px] font-bold uppercase tracking-wider {kat_cls}')
                with ui.element('div').classes('flex flex-wrap gap-1.5'):
                    for k, meta in sorted(skupiny[kat], key=lambda x: ((x[1] or {}).get('nazev') or x[0]).lower()):
                        nazev = (meta or {}).get('nazev') or k
                        popis = (meta or {}).get('popis') or ''
                        ikona = (meta or {}).get('ikona') or 'label_important'
                        chip = ui.element('div').classes(
                            f'inline-flex items-center gap-1 {chip_cls} border '
                            'text-xs font-medium px-2 py-1 rounded-md')
                        with chip:
                            ui.icon(ikona, size='16px').classes(ico_cls)
                            ui.label(nazev)
                        if popis:
                            chip.tooltip(popis)


@ui.refreshable
def vykresli_prehled(user_id, user_name, vsechna_prava):
    nastaveni = intranet_data.nacti_nastaveni_intranetu()

    ma_pristup_kviz = ("kviz" in vsechna_prava or "vse" in vsechna_prava) and nastaveni.get('kviz_zapnuty', True)
    ma_pristup_dochazka_zaklad_dlazdice = "dlazdice_dochazka_zaklad" in vsechna_prava or "vse" in vsechna_prava
    ma_pristup_veletrh = ("veletrh_pristup" in vsechna_prava or "veletrh_admin" in vsechna_prava or "vse" in vsechna_prava) and nastaveni.get('veletrh_zapnuty', True)
    ma_pristup_znacky = ("znacky_uzivatel" in vsechna_prava or "znacky_spravce" in vsechna_prava or "vse" in vsechna_prava) and nastaveni.get('znacky_zapnuty', True)
    ma_pristup_znacky_provoz = ("znacky_provoz_uzivatel" in vsechna_prava or "znacky_provoz_spravce" in vsechna_prava or "vse" in vsechna_prava) and nastaveni.get('znacky_provoz_zapnuty', True)
    ma_pristup_prod_akt = bool(
        set(vsechna_prava) & {
            'prodej_akt_ctenar', 'prodej_akt_zadavatel', 'prodej_akt_ucetni',
            'prodej_akt_ao', 'prodej_akt_schvalovatel', 'vse',
        }
    ) and nastaveni.get('prod_akt_zapnuty', True)

    ma_pristup_finance = ("vse" in vsechna_prava or any(p in vsechna_prava for p in ['nakup_uzivatel', 'nakup_schvalit', 'faktury_seznam_schvalit'])) and nastaveni.get('finance_zapnuty', True)
    ma_pristup_narozeniny = (
        "vse" in vsechna_prava or
        "narozeniny_sprava" in vsechna_prava
    ) and nastaveni.get('narozeniny_zapnuty', True)
    _ma_smeny = (
        "vse" in vsechna_prava or
        "smeny_admin" in vsechna_prava or
        "smeny_zobrazit" in vsechna_prava or
        any(p.startswith('hlavni_vedouci_') for p in vsechna_prava)
    ) and nastaveni.get('smeny_zapnuty', True)
    ma_pristup_komunikace = nastaveni.get('komunikace_zapnuty', True)
    ma_pristup_planogram  = (
        'vse' in vsechna_prava or
        'planogram_admin' in vsechna_prava or
        'planogram_pristup' in vsechna_prava
    ) and nastaveni.get('planogram_zapnuty', True)
    ma_pristup_ochutnavky = (
        'vse' in vsechna_prava or
        'ochutnavky_admin' in vsechna_prava or
        'ochutnavky_pristup' in vsechna_prava
    ) and nastaveni.get('ochutnavky_zapnuty', True)
    ma_pristup_ukolovnik = (
        'vse' in vsechna_prava or
        'ukolovnik_admin' in vsechna_prava or
        any(p.startswith('ukolovnik_') for p in vsechna_prava)
    ) and nastaveni.get('ukolovnik_zapnuty', True)
    ma_pristup_vysledky = (
        'vse' in vsechna_prava or
        'vysledky_ao' in vsechna_prava or
        'vysledky_ucetni' in vsechna_prava or
        'vysledky_ucetni_bezna' in vsechna_prava or
        'vysledky_majitel' in vsechna_prava or
        any(p.startswith('vysledky_pobocka_') for p in vsechna_prava)
    ) and nastaveni.get('vysledky_zapnuty', True)
    ma_pristup_sankce = (
        'vse' in vsechna_prava or
        'sankce_analytik' in vsechna_prava or
        'sankce_ucetni' in vsechna_prava or
        'sankce_nakup' in vsechna_prava or
        'sankce_ctenar' in vsechna_prava
    ) and nastaveni.get('sankce_zapnuty', True)
    ma_pristup_spolvecer = (
        'vse' in vsechna_prava or
        'spolvecer_ctenar' in vsechna_prava or
        'spolvecer_schvalovatel' in vsechna_prava or
        any(p.startswith('spolvecer_organizator_') for p in vsechna_prava)
    ) and nastaveni.get('spolvecer_zapnuty', True)
    # Žadatelem je automaticky každý uživatel → dlaždice pro všechny (když je modul zapnutý).
    ma_pristup_vizitky = nastaveni.get('vizitky_zapnuty', True)
    ma_pristup_cenopripad = (
        'vse' in vsechna_prava or
        any(p in vsechna_prava for p in (
            'cenopripad_zadatel_nakup', 'cenopripad_zadatel_obchod',
            'cenopripad_zadatel_letaky', 'cenopripad_office_letaky',
            'cenopripad_ctenar_letaky',
            'cenopripad_office_nakup', 'cenopripad_office_obchod',
            'cenopripad_spravce_nakup', 'cenopripad_spravce',
            'cenopripad_spravce_bez_emailu', 'cenopripad_vkladatel'))
    ) and nastaveni.get('cenopripad_zapnuty', True)
    ma_pristup_asm = (
        'vse' in vsechna_prava or
        any(p in vsechna_prava for p in (
            'asm_zadatel', 'asm_office', 'asm_spravce',
            'asm_spravce_bez_emailu', 'asm_vkladatel'))
    ) and nastaveni.get('asm_zapnuty', True)
    ma_pristup_lupa = (
        nastaveni.get('lupa_zapnuty', True)
        and intranet_lupa.ma_pristup(user_id, vsechna_prava)
    )
    ma_pristup_schuzky = (
        'vse' in vsechna_prava or
        any(p in vsechna_prava for p in (
            'schuzky_zadatel', 'schuzky_vedouci', 'schuzky_spravce'))
    ) and nastaveni.get('schuzky_zapnuty', True)
    is_admin = "vse" in vsechna_prava

    nazev_kviz       = "Zkouškový Kvíz"
    nazev_dok        = "Dokumenty"
    nazev_dochazka   = "Docházka a Volno"
    nazev_veletrh    = "Plán Veletrh 2027"
    nazev_finance    = "Aprovia"
    nazev_narozeniny = "Narozeniny"
    nazev_komunikace = "Komunikační portál"

    def naviguj_s_nacitanim(url, nazev_modulu):
        with ui.dialog().props('maximized persistent transition-show="fade" transition-hide="fade"') as dlg:
            with ui.column().classes('w-full h-full items-center justify-center bg-gray-50 m-0 p-0'):
                ui.spinner('dots', size='5rem', color='blue').classes('mb-4')
                ui.label('Otevírám modul').classes('text-sm font-bold text-gray-500 uppercase tracking-widest')
                ui.label(nazev_modulu).classes('text-3xl font-black text-blue-900')
        dlg.open()
        ui.timer(0.2, lambda: ui.navigate.to(url), once=True)

    ui.label('Vítejte na hlavní nástěnce').classes('text-4xl font-extrabold text-gray-800 mb-6')

    with ui.row().classes('gap-8 flex-wrap items-stretch'):
        if ma_pristup_kviz:
            with ui.card().classes('w-80 h-72 items-center justify-center shadow-xl hover:scale-105 transition-transform duration-300 cursor-pointer bg-white rounded-2xl border border-green-100').on('click', lambda: naviguj_s_nacitanim('/kviz', nazev_kviz)):
                ui.label('👨‍🎓').classes('text-7xl mb-6')
                ui.label(nazev_kviz).classes('text-2xl font-bold text-gray-800 mb-4 text-center')
                ui.button('Otevřít aplikaci', on_click=lambda: naviguj_s_nacitanim('/kviz', nazev_kviz)).classes('bg-green-500 hover:bg-green-600 text-white font-bold py-3 px-8 rounded-lg shadow-md')

        if ma_pristup_dochazka_zaklad_dlazdice:
            def jit_do_dochazky():
                app.storage.user['aktivni_slozka_dochazka'] = None
                app.storage.user['intranet_tab'] = 'dochazka'
            with ui.card().classes('w-80 h-72 items-center justify-center shadow-md hover:scale-105 transition-transform duration-300 cursor-pointer bg-white rounded-2xl border border-orange-100').on('click', jit_do_dochazky):
                ui.label('📅').classes('text-7xl mb-6')
                ui.label(nazev_dochazka).classes('text-2xl font-bold text-gray-800 mb-4 text-center')
                ui.button('Otevřít aplikaci', on_click=jit_do_dochazky).classes('bg-orange-500 hover:bg-orange-600 text-white font-bold py-3 px-8 rounded-lg shadow-md')

        if ma_pristup_veletrh:
            def jit_do_veletrhu():
                app.storage.user['intranet_tab'] = 'veletrh'
            with ui.card().classes('w-80 h-72 items-center justify-center shadow-xl hover:scale-105 transition-transform duration-300 cursor-pointer bg-white rounded-2xl border border-purple-200').on('click', jit_do_veletrhu):
                ui.label('🎪').classes('text-7xl mb-6')
                ui.label(nazev_veletrh).classes('text-2xl font-bold text-gray-800 mb-4 text-center')
                ui.button('Otevřít mapu a stánky', on_click=jit_do_veletrhu).classes('bg-purple-600 hover:bg-purple-700 text-white font-bold py-3 px-8 rounded-lg shadow-md')

        if ma_pristup_finance:
            def jit_do_financi():
                app.storage.user['intranet_tab'] = 'finance'
            with ui.card().classes('w-80 h-72 items-center justify-center shadow-xl hover:scale-105 transition-transform duration-300 cursor-pointer bg-white rounded-2xl border border-blue-200').on('click', jit_do_financi):
                ui.label('💼').classes('text-7xl mb-6')
                ui.label(nazev_finance).classes('text-2xl font-bold text-gray-800 mb-4 text-center')
                ui.button('Přejít na modul', on_click=jit_do_financi).classes('bg-blue-600 hover:bg-blue-700 text-white font-bold py-3 px-8 rounded-lg shadow-md')

        if ma_pristup_znacky:
            def jit_do_znacek():
                app.storage.user['intranet_tab'] = 'znacky'
            with ui.card().classes('w-80 h-72 items-center justify-center shadow-xl hover:scale-105 transition-transform duration-300 cursor-pointer bg-white rounded-2xl border border-purple-300').on('click', jit_do_znacek):
                ui.label('🏷️').classes('text-7xl mb-6')
                ui.label('Privátní značky JIP').classes('text-2xl font-bold text-gray-800 mb-4 text-center')
                ui.button('Přejít na hlasování', on_click=jit_do_znacek).classes('bg-purple-600 hover:bg-purple-700 text-white font-bold py-3 px-8 rounded-lg shadow-md')

        if ma_pristup_znacky_provoz:
            def jit_do_znacek_provoz():
                app.storage.user['intranet_tab'] = 'znacky_provoz'
            with ui.card().classes('w-80 h-72 items-center justify-center shadow-xl hover:scale-105 transition-transform duration-300 cursor-pointer bg-white rounded-2xl border border-blue-300').on('click', jit_do_znacek_provoz):
                ui.label('🏭').classes('text-7xl mb-6')
                ui.label('Hlas Provozu').classes('text-2xl font-bold text-gray-800 mb-4 text-center')
                ui.button('Přejít na hlasování', on_click=jit_do_znacek_provoz).classes('bg-blue-600 hover:bg-blue-700 text-white font-bold py-3 px-8 rounded-lg shadow-md')

        if ma_pristup_prod_akt:
            def jit_do_prod_akt():
                app.storage.user['intranet_tab'] = 'prod_akt'
            with ui.card().classes('w-80 h-72 items-center justify-center shadow-xl hover:scale-105 transition-transform duration-300 cursor-pointer bg-white rounded-2xl border border-emerald-300').on('click', jit_do_prod_akt):
                ui.label('📋').classes('text-7xl mb-6')
                ui.label('Prodejní aktivity').classes('text-2xl font-bold text-gray-800 mb-4 text-center')
                ui.button('Otevřít přehled', on_click=jit_do_prod_akt).classes('bg-emerald-600 hover:bg-emerald-700 text-white font-bold py-3 px-8 rounded-lg shadow-md')

        if ma_pristup_narozeniny:
            narozeniny_dnes = intranet_narozeniny.ziskej_pocet_narozenin_dnes(vsechna_prava)
            def jit_do_narozenin():
                app.storage.user['intranet_tab'] = 'narozeniny'
            with ui.card().classes('w-80 h-72 items-center justify-center shadow-xl hover:scale-105 transition-transform duration-300 cursor-pointer bg-white rounded-2xl border border-pink-200 relative').on('click', jit_do_narozenin):
                if narozeniny_dnes > 0:
                    ui.badge(str(narozeniny_dnes), color='red').classes('absolute top-3 right-3 text-lg font-black px-2')
                ui.label('🎂').classes('text-7xl mb-6')
                ui.label(nazev_narozeniny).classes('text-2xl font-bold text-gray-800 mb-4 text-center')
                ui.button('Přejít na přehled', on_click=jit_do_narozenin).classes('bg-pink-500 hover:bg-pink-600 text-white font-bold py-3 px-8 rounded-lg shadow-md')

        if _ma_smeny:
            def jit_do_smen():
                app.storage.user['intranet_tab'] = 'smeny'
            with ui.card().classes('w-80 h-72 items-center justify-center shadow-xl hover:scale-105 transition-transform duration-300 cursor-pointer bg-white rounded-2xl border border-teal-200').on('click', jit_do_smen):
                ui.label('⌨️').classes('text-7xl mb-6')
                ui.label('Plánování směn').classes('text-2xl font-bold text-gray-800 mb-4 text-center')
                ui.button('Otevřít plán', on_click=jit_do_smen).classes('bg-teal-600 hover:bg-teal-700 text-white font-bold py-3 px-8 rounded-lg shadow-md')

        if ma_pristup_komunikace:
            def jit_do_komunikace():
                app.storage.user['intranet_tab'] = 'komunikace'
            with ui.card().classes('w-80 h-72 items-center justify-center shadow-xl hover:scale-105 transition-transform duration-300 cursor-pointer bg-white rounded-2xl border border-cyan-200').on('click', jit_do_komunikace):
                ui.label('💬').classes('text-7xl mb-6')
                ui.label(nazev_komunikace).classes('text-2xl font-bold text-gray-800 mb-4 text-center')
                ui.button('Otevřít portál', on_click=jit_do_komunikace).classes('bg-cyan-600 hover:bg-cyan-700 text-white font-bold py-3 px-8 rounded-lg shadow-md')

        if ma_pristup_planogram:
            def jit_do_planogramu():
                app.storage.user['intranet_tab'] = 'planogram'
            with ui.card().classes('w-80 h-72 items-center justify-center shadow-xl hover:scale-105 transition-transform duration-300 cursor-pointer bg-white rounded-2xl border border-amber-200').on('click', jit_do_planogramu):
                ui.label('🚬').classes('text-7xl mb-6')
                ui.label('Plánogram tabákových výrobků').classes('text-xl font-bold text-gray-800 mb-4 text-center')
                ui.button('Otevřít plánogram', on_click=jit_do_planogramu).classes('bg-amber-600 hover:bg-amber-700 text-white font-bold py-3 px-8 rounded-lg shadow-md')

        if ma_pristup_ochutnavky:
            _pocet_ochutnavek = 0
            try:
                import intranet_ochutnavky as _och
                _pocet_ochutnavek = _och.pocet_aktualnich_akci()
            except Exception as e:
                print(f"[Nástěnka] Počet ochutnávek: {e}")

            def jit_do_ochutnavek():
                app.storage.user['intranet_tab'] = 'ochutnavky'
            with ui.card().classes('w-80 h-72 items-center justify-center shadow-xl hover:scale-105 transition-transform duration-300 cursor-pointer bg-white rounded-2xl border border-green-200').on('click', jit_do_ochutnavek):
                ui.label('🍽️').classes('text-7xl mb-4')
                ui.label('Ochutnávky MO a CC').classes('text-xl font-bold text-gray-800 mb-2 text-center')
                if _pocet_ochutnavek:
                    ui.label(f'{_pocet_ochutnavek} probíhá nebo se blíží').classes('text-sm font-bold text-green-700 mb-2')
                else:
                    ui.label('Aktuálně nic neprobíhá').classes('text-sm text-gray-400 mb-2')
                ui.button('Otevřít ochutnávky', on_click=jit_do_ochutnavek).classes('bg-green-600 hover:bg-green-700 text-white font-bold py-3 px-8 rounded-lg shadow-md')

        if ma_pristup_ukolovnik:
            try:
                import intranet_data as _id
                _conn = _id.get_db_connection()
                _pocet_ukolu = 0
                if _conn:
                    try:
                        _cur = _conn.cursor()
                        _cur.execute(
                            "SELECT COUNT(*) FROM ukolovnik_ukoly "
                            "WHERE prirazen_id=%s AND stav NOT IN ('Hotovo','Zrušen') AND termin <= CURDATE()",
                            (user_id,)
                        )
                        _row = _cur.fetchone()
                        _pocet_ukolu = int(_row[0]) if _row else 0
                    finally:
                        _cur.close(); _conn.close()
            except Exception:
                _pocet_ukolu = 0

            def jit_do_ukolovniku():
                app.storage.user['intranet_tab'] = 'ukolovnik'
            with ui.card().classes('w-80 h-72 items-center justify-center shadow-xl hover:scale-105 transition-transform duration-300 cursor-pointer bg-white rounded-2xl border border-indigo-200 relative').on('click', jit_do_ukolovniku):
                if _pocet_ukolu > 0:
                    ui.badge(str(_pocet_ukolu), color='red').classes('absolute top-3 right-3 text-lg font-black px-2')
                ui.label('📋').classes('text-7xl mb-6')
                ui.label('Porady a úkoly').classes('text-2xl font-bold text-gray-800 mb-4 text-center')
                ui.button('Otevřít přehled', on_click=jit_do_ukolovniku).classes('bg-indigo-600 hover:bg-indigo-700 text-white font-bold py-3 px-8 rounded-lg shadow-md')

        if ma_pristup_vysledky:
            def jit_do_vysledku():
                app.storage.user['intranet_tab'] = 'vysledky'
            with ui.card().classes('w-80 h-72 items-center justify-center shadow-xl hover:scale-105 transition-transform duration-300 cursor-pointer bg-white rounded-2xl border border-rose-200').on('click', jit_do_vysledku):
                ui.label('📊').classes('text-7xl mb-6')
                ui.label('Výsledky poboček').classes('text-2xl font-bold text-gray-800 mb-4 text-center')
                ui.button('Otevřít přehled', on_click=jit_do_vysledku).classes('bg-rose-600 hover:bg-rose-700 text-white font-bold py-3 px-8 rounded-lg shadow-md')

        if ma_pristup_sankce:
            def jit_do_sankci():
                app.storage.user['sankce_pohled'] = None  # po vstupu vždy přehled dvou dlaždic
                app.storage.user['intranet_tab'] = 'sankce'
            with ui.card().classes('w-80 h-72 items-center justify-center shadow-xl hover:scale-105 transition-transform duration-300 cursor-pointer bg-white rounded-2xl border border-red-200').on('click', jit_do_sankci):
                ui.label('⚖️').classes('text-7xl mb-6')
                ui.label('Sankce').classes('text-2xl font-bold text-gray-800 mb-4 text-center')
                ui.button('Otevřít přehled', on_click=jit_do_sankci).classes('bg-red-600 hover:bg-red-700 text-white font-bold py-3 px-8 rounded-lg shadow-md')

        if ma_pristup_spolvecer:
            def jit_do_spolvecer():
                app.storage.user[f'spolvecer_sel_{user_id}'] = None  # po vstupu vždy přehled dlaždic poboček
                app.storage.user['intranet_tab'] = 'spolvecer'
            with ui.card().classes('w-80 h-72 items-center justify-center shadow-xl hover:scale-105 transition-transform duration-300 cursor-pointer bg-white rounded-2xl border border-fuchsia-200').on('click', jit_do_spolvecer):
                ui.label('🎉').classes('text-7xl mb-6')
                ui.label('Spol. večer 2026').classes('text-2xl font-bold text-gray-800 mb-4 text-center')
                ui.button('Otevřít přehled', on_click=jit_do_spolvecer).classes('bg-fuchsia-600 hover:bg-fuchsia-700 text-white font-bold py-3 px-8 rounded-lg shadow-md')

        if ma_pristup_vizitky:
            def jit_do_vizitek():
                app.storage.user['intranet_tab'] = 'vizitky'
            with ui.card().classes('w-80 h-72 items-center justify-center shadow-xl hover:scale-105 transition-transform duration-300 cursor-pointer bg-white rounded-2xl border border-sky-200').on('click', jit_do_vizitek):
                ui.label('🪪').classes('text-7xl mb-6')
                ui.label('Vizitky a podpisy').classes('text-2xl font-bold text-gray-800 mb-4 text-center')
                ui.button('Otevřít', on_click=jit_do_vizitek).classes('bg-sky-600 hover:bg-sky-700 text-white font-bold py-3 px-8 rounded-lg shadow-md')

        if ma_pristup_schuzky:
            def jit_do_schuzek():
                app.storage.user['intranet_tab'] = 'schuzky'
            with ui.card().classes('w-80 h-72 items-center justify-center shadow-xl hover:scale-105 transition-transform duration-300 cursor-pointer bg-white rounded-2xl border border-indigo-200').on('click', jit_do_schuzek):
                ui.label('🗓️').classes('text-7xl mb-6')
                ui.label('Schůzky s vedoucím').classes('text-2xl font-bold text-gray-800 mb-4 text-center')
                ui.button('Otevřít', on_click=jit_do_schuzek).classes('bg-indigo-600 hover:bg-indigo-700 text-white font-bold py-3 px-8 rounded-lg shadow-md')

        if ma_pristup_cenopripad:
            def jit_do_cenopripad():
                app.storage.user['cenopripad_pohled'] = None  # po vstupu vždy rozcestník dlaždic
                app.storage.user['intranet_tab'] = 'cenopripad'
            with ui.card().classes('w-80 h-72 items-center justify-center shadow-xl hover:scale-105 transition-transform duration-300 cursor-pointer bg-white rounded-2xl border border-emerald-200').on('click', jit_do_cenopripad):
                ui.label('🏷️').classes('text-7xl mb-6')
                ui.label('Cenopřípad').classes('text-2xl font-bold text-gray-800 mb-4 text-center')
                ui.button('Otevřít přehled', on_click=jit_do_cenopripad).classes('bg-emerald-600 hover:bg-emerald-700 text-white font-bold py-3 px-8 rounded-lg shadow-md')

        if ma_pristup_asm:
            def jit_do_asm():
                app.storage.user['asm_pohled'] = None  # po vstupu vždy rozcestník dlaždic
                app.storage.user['intranet_tab'] = 'asm'
            with ui.card().classes('w-80 h-72 items-center justify-center shadow-xl hover:scale-105 transition-transform duration-300 cursor-pointer bg-white rounded-2xl border border-emerald-200').on('click', jit_do_asm):
                ui.label('📝').classes('text-7xl mb-6')
                ui.label('Formuláře ASM').classes('text-2xl font-bold text-gray-800 mb-4 text-center')
                ui.button('Otevřít přehled', on_click=jit_do_asm).classes('bg-emerald-600 hover:bg-emerald-700 text-white font-bold py-3 px-8 rounded-lg shadow-md')

        if ma_pristup_lupa:
            def jit_do_lupa():
                app.storage.user[f'lupa_pohled_{user_id}'] = None  # vždy rozcestník ASM
                app.storage.user['intranet_tab'] = 'lupa'
            with ui.card().classes('w-80 h-72 items-center justify-center shadow-xl hover:scale-105 transition-transform duration-300 cursor-pointer bg-white rounded-2xl border border-indigo-200').on('click', jit_do_lupa):
                ui.label('🔍').classes('text-7xl mb-6')
                ui.label('Lupou na obchod').classes('text-2xl font-bold text-gray-800 mb-4 text-center')
                ui.button('Otevřít přehled', on_click=jit_do_lupa).classes('bg-indigo-600 hover:bg-indigo-700 text-white font-bold py-3 px-8 rounded-lg shadow-md')


    # ==========================================
    # PRAVÝ SPODNÍ ROH — helpdesk
    # ==========================================
    with ui.page_sticky(position='bottom-right', x_offset=20, y_offset=20):
        def jit_na_helpdesk():
            app.storage.user['intranet_tab'] = 'helpdesk'

        ui.button('Podpora', icon='support_agent', on_click=jit_na_helpdesk) \
            .classes('bg-blue-600 hover:bg-blue-700 text-white font-bold rounded-full shadow-[0_10px_25px_rgba(37,99,235,0.4)] px-5 h-12 text-sm transition-transform hover:scale-105')

@ui.refreshable
def vykresli_osobni_nastaveni(user_id, user_email, user_name):
    ui.label('Osobní nastavení účtu').classes('text-4xl font-extrabold text-gray-800 mb-8')

    with ui.card().classes('w-full max-w-2xl p-8 shadow-lg bg-white rounded-xl'):
        ui.label('Změna hesla').classes('text-2xl font-bold mb-4 text-blue-800')
        input_stare = ui.input('Současné heslo', password=True).classes('w-full mb-4')
        input_nove = ui.input('Nové heslo', password=True).classes('w-full mb-4')

        async def zmenit_vlastni_heslo():
            if user_id == 999999: return ui.notify('Nelze měnit heslo nouzovému administrátorovi!', type='negative', position='top')

            id_u, jm, msg = await asyncio.to_thread(intranet_data.overit_prihlaseni, user_email, input_stare.value)

            if id_u:
                if intranet_data.heslo_je_silne(input_nove.value):
                    uspech = await asyncio.to_thread(intranet_data.nastav_heslo_a_zrus_priznak, user_id, input_nove.value)
                    if uspech:
                        intranet_logger.log_activity(user_name, "Osobní profil", "Uživatel si změnil heslo")
                        ui.notify('Vaše heslo bylo úspěšně změněno.', type='positive', position='top')
                        input_stare.value = ''
                        input_nove.value = ''
                    else:
                        ui.notify('Spojení s databází selhalo.', type='negative', position='top')
                else:
                    ui.notify('Heslo musí mít min. 8 znaků, obsahovat velké i malé písmeno a číslo.', type='warning', position='top')
            else:
                ui.notify('Současné heslo je chybné!', type='negative', position='top')

        ui.button('Uložit nové heslo', on_click=zmenit_vlastni_heslo).classes('bg-blue-600 hover:bg-blue-700 text-white font-bold mb-4')

        # ==========================================
        # DVOUFAKTOROVÉ OVĚŘENÍ (TOTP)
        # ==========================================
        ui.label('Dvoufaktorové ověření (2FA)').classes('text-2xl font-bold mt-8 mb-4 text-blue-800 border-t pt-8')

        if user_id == 999999:
            ui.label('Pro nouzového administrátora není 2FA dostupné.').classes('text-sm text-gray-500')
        else:
            ma_2fa = intranet_2fa.ma_aktivni_2fa(user_id)

            def _zobraz_zalozni_kody(kody, kontejner):
                """Vypíše záložní kódy do kontejneru + tlačítko pro zkopírování."""
                kontejner.clear()
                with kontejner:
                    ui.label('Záložní kódy (každý platí jen jednou). Uložte si je — zobrazí se pouze teď:') \
                        .classes('text-sm font-bold text-amber-700 mb-2')
                    with ui.grid(columns=2).classes('gap-x-8 gap-y-1 mb-3 bg-slate-50 rounded-lg p-4'):
                        for k in kody:
                            ui.label(k).classes('font-mono text-base text-slate-800')
                    ui.button('Zkopírovat kódy', icon='content_copy',
                              on_click=lambda: (ui.clipboard.write('\n'.join(kody)),
                                                ui.notify('Zkopírováno do schránky.', type='positive'))) \
                        .props('outline size=sm no-caps').classes('rounded-lg')

            if ma_2fa:
                with ui.row().classes('items-center gap-2 mb-2'):
                    ui.icon('verified_user', color='green-600', size='sm')
                    ui.label('Aktivní — přihlášení vyžaduje kód z autentifikační aplikace.') \
                        .classes('text-sm font-semibold text-green-700')
                zbyva_kodu = intranet_2fa.pocet_zaloznich_kodu(user_id)
                ui.label(f'Zbývající záložní kódy: {zbyva_kodu}').classes('text-sm text-gray-600 mb-3')
                pocet_duveryhodnych = intranet_2fa.pocet_duveryhodnych_zarizeni(user_id)
                if pocet_duveryhodnych:
                    ui.label(f'Zapamatovaná zařízení (bez kódu): {pocet_duveryhodnych}').classes('text-sm text-gray-600 mb-3')

                def dialog_s_heslem(titulek, popis, akce_po_overeni):
                    """Citlivé akce (vypnutí 2FA, nové kódy) vyžadují znovu heslo."""
                    with ui.dialog() as dlg, ui.card().classes('p-6 rounded-2xl w-full max-w-sm'):
                        ui.label(titulek).classes('text-xl font-bold text-slate-800 mb-1')
                        ui.label(popis).classes('text-sm text-slate-500 mb-4')
                        heslo_inp = ui.input('Současné heslo', password=True).classes('w-full mb-4').props('outlined dense')
                        vysledek_box = ui.column().classes('w-full')

                        async def potvrdit():
                            id_o, _jm, _msg = await asyncio.to_thread(
                                intranet_data.overit_prihlaseni, user_email, heslo_inp.value)
                            if not id_o:
                                ui.notify('Heslo je chybné!', type='negative', position='top')
                                return
                            await akce_po_overeni(dlg, vysledek_box)

                        with ui.row().classes('w-full justify-end gap-3') as tlacitka:
                            ui.button('Zrušit', on_click=dlg.close).props('flat no-caps').classes('text-slate-600 font-semibold rounded-lg')
                            ui.button('Potvrdit', on_click=potvrdit).props('unelevated no-caps').classes('bg-blue-600 hover:bg-blue-700 text-white font-semibold rounded-lg')
                    dlg.open()
                    return dlg

                def vypnout_2fa():
                    async def _akce(dlg, _box):
                        uspech = await asyncio.to_thread(intranet_2fa.deaktivuj_2fa, user_id)
                        if uspech:
                            intranet_logger.log_activity(user_name, "Osobní profil", "Vypnuto dvoufaktorové ověření (2FA)")
                            ui.notify('Dvoufaktorové ověření bylo vypnuto.', type='warning', position='top')
                            dlg.close()
                            ui.timer(0, vykresli_osobni_nastaveni.refresh, once=True)
                        else:
                            ui.notify('Spojení s databází selhalo.', type='negative', position='top')
                    dialog_s_heslem('Vypnout 2FA?', 'Účet bude chráněn už jen heslem.', _akce)

                def nove_zalozni_kody():
                    async def _akce(dlg, box):
                        kody = await asyncio.to_thread(intranet_2fa.generuj_nove_zalozni_kody, user_id)
                        if kody:
                            intranet_logger.log_activity(user_name, "Osobní profil", "Vygenerovány nové záložní kódy 2FA")
                            _zobraz_zalozni_kody(kody, box)
                        else:
                            ui.notify('Generování kódů selhalo.', type='negative', position='top')
                    dialog_s_heslem('Nové záložní kódy', 'Staré záložní kódy přestanou platit.', _akce)

                def odhlasit_zarizeni():
                    with ui.dialog() as dlg_z, ui.card().classes('p-6 rounded-2xl w-full max-w-sm'):
                        ui.label('Odhlásit ze všech zařízení?').classes('text-xl font-bold text-slate-800 mb-1')
                        ui.label('Na všech zapamatovaných zařízeních bude při příštím přihlášení znovu potřeba 2FA kód.') \
                            .classes('text-sm text-slate-500 mb-4')

                        async def _potvrd():
                            pocet = await asyncio.to_thread(intranet_2fa.zrus_duveryhodna_zarizeni, user_id)
                            app.storage.user.pop('totp_duvera_token', None)
                            dlg_z.close()
                            if pocet >= 0:
                                intranet_logger.log_activity(user_name, "Osobní profil", f"Odhlášení ze všech zapamatovaných zařízení 2FA ({pocet})")
                                ui.notify('Hotovo — všechna zapamatovaná zařízení byla odhlášena.', type='positive', position='top')
                                ui.timer(0, vykresli_osobni_nastaveni.refresh, once=True)
                            else:
                                ui.notify('Spojení s databází selhalo.', type='negative', position='top')

                        with ui.row().classes('w-full justify-end gap-3'):
                            ui.button('Zrušit', on_click=dlg_z.close).props('flat no-caps').classes('text-slate-600 font-semibold rounded-lg')
                            ui.button('Odhlásit', on_click=_potvrd).props('unelevated no-caps').classes('bg-red-600 hover:bg-red-700 text-white font-semibold rounded-lg')
                    dlg_z.open()

                with ui.row().classes('gap-3'):
                    ui.button('Nové záložní kódy', icon='refresh', on_click=nove_zalozni_kody) \
                        .props('outline no-caps').classes('rounded-lg text-blue-600')
                    if pocet_duveryhodnych:
                        ui.button('Odhlásit ze všech zařízení', icon='devices', on_click=odhlasit_zarizeni) \
                            .props('outline no-caps').classes('rounded-lg text-amber-700')
                    ui.button('Vypnout 2FA', icon='block', color='red', on_click=vypnout_2fa) \
                        .props('outline no-caps').classes('rounded-lg text-red-600')
            else:
                ui.label('Druhý faktor přihlášení — kód z autentifikační aplikace v telefonu '
                         '(Google Authenticator, Microsoft Authenticator, Aegis…). '
                         'I při úniku hesla se bez telefonu nikdo nepřihlásí.') \
                    .classes('text-sm text-gray-500 mb-3')

                def aktivovat_2fa():
                    with ui.dialog() as dlg, ui.card().classes('p-6 sm:p-8 rounded-2xl w-full max-w-md'):
                        stav = {'hotovo': False}
                        obsah = ui.column().classes('w-full items-center')

                        async def nacti_setup():
                            secret, uri, qr = await asyncio.to_thread(
                                intranet_2fa.zahaj_aktivaci, user_id, user_email)
                            obsah.clear()
                            if not secret:
                                with obsah:
                                    ui.label('Spojení s databází selhalo.').classes('text-red-600 font-bold')
                                return
                            with obsah:
                                ui.label('Aktivace dvoufaktorového ověření').classes('text-xl font-black text-slate-800 mb-1')
                                ui.label('1) Naskenujte QR kód v autentifikační aplikaci, nebo zadejte klíč ručně.') \
                                    .classes('text-sm text-slate-500 self-start')
                                if qr:
                                    ui.image(qr).classes('w-52 h-52 my-2')
                                else:
                                    ui.label('(QR kód není k dispozici — zadejte klíč ručně)') \
                                        .classes('text-xs text-amber-600 my-2')
                                ui.label('Klíč pro ruční zadání:').classes('text-xs text-slate-400 self-start')
                                ui.label(' '.join(secret[i:i + 4] for i in range(0, len(secret), 4))) \
                                    .classes('font-mono text-sm bg-slate-50 rounded-lg px-3 py-2 mb-3 break-all')
                                ui.label('2) Zadejte 6místný kód, který aplikace zobrazuje:') \
                                    .classes('text-sm text-slate-500 self-start mb-1')
                                kod_inp = ui.input('Kód z aplikace').classes('w-full mb-3') \
                                    .props('outlined dense inputmode=numeric autocomplete=one-time-code')
                                zapamatovat_inp = ui.checkbox('Pamatovat toto zařízení (na 30 dní zde nebude kód potřeba)') \
                                    .classes('w-full text-sm text-slate-600 mb-4 self-start')

                                async def potvrdit_aktivaci():
                                    kody = await asyncio.to_thread(
                                        intranet_2fa.potvrd_aktivaci, user_id, kod_inp.value)
                                    if not kody:
                                        ui.notify('Kód nesouhlasí — zkontrolujte aplikaci a zkuste to znovu.',
                                                  type='negative', position='top')
                                        return
                                    stav['hotovo'] = True
                                    intranet_logger.log_activity(user_name, "Osobní profil", "Aktivováno dvoufaktorové ověření (2FA)")
                                    # "Pamatovat toto zařízení": token uložíme do prohlížeče (relace),
                                    # jeho hash do DB — na tomto zařízení se pak 30 dní kód nevyžaduje.
                                    if zapamatovat_inp.value:
                                        _tok = await asyncio.to_thread(
                                            intranet_2fa.zaregistruj_duveryhodne_zarizeni, user_id, None)
                                        if _tok:
                                            app.storage.user['totp_duvera_token'] = _tok
                                    obsah.clear()
                                    with obsah:
                                        with ui.row().classes('items-center gap-2 mb-2 self-start'):
                                            ui.icon('verified_user', color='green-600')
                                            ui.label('2FA je aktivní!').classes('text-xl font-black text-green-700')
                                        kody_box = ui.column().classes('w-full')
                                        _zobraz_zalozni_kody(kody, kody_box)
                                        def hotovo():
                                            dlg.close()
                                            ui.timer(0, vykresli_osobni_nastaveni.refresh, once=True)
                                        ui.button('Hotovo, kódy mám uložené', icon='check', on_click=hotovo) \
                                            .props('unelevated no-caps') \
                                            .classes('bg-green-600 hover:bg-green-700 text-white font-semibold rounded-lg mt-4 self-end')

                                kod_inp.on('keydown.enter', potvrdit_aktivaci)
                                with ui.row().classes('w-full justify-end gap-3'):
                                    ui.button('Zrušit', on_click=dlg.close).props('flat no-caps').classes('text-slate-600 font-semibold rounded-lg')
                                    ui.button('Ověřit a zapnout', icon='verified_user', on_click=potvrdit_aktivaci) \
                                        .props('unelevated no-caps').classes('bg-blue-600 hover:bg-blue-700 text-white font-semibold rounded-lg')

                        with obsah:
                            ui.spinner('dots', size='3rem', color='blue').classes('my-8')
                        ui.timer(0, nacti_setup, once=True)
                    dlg.open()

                ui.button('Aktivovat 2FA', icon='phonelink_lock', on_click=aktivovat_2fa) \
                    .classes('bg-blue-600 hover:bg-blue-700 text-white font-bold')

        ui.label('Automatické odhlášení').classes('text-2xl font-bold mt-8 mb-4 text-blue-800 border-t pt-8')

        osobni_minuty = intranet_data.ziskej_osobni_auto_odhlaseni(user_id)

        ui.label('Odhlášení po zadaném počtu minut nečinnosti. Výchozí hodnota je 10 minut. '
                 'Čas lze upravit, ale automatické odhlašování nelze úplně vypnout (minimum 1 minuta).').classes('text-sm text-gray-500 mb-2')

        osobni_input = ui.number(
            'Minut nečinnosti (prázdné = výchozích 10 minut)',
            value=osobni_minuty if osobni_minuty and int(osobni_minuty) > 0 else 10,
            min=1, max=480, step=1,
        ).classes('w-full md:w-96 bg-white').props('outlined')

        async def ulozit_osobni_odhlaseni():
            raw = osobni_input.value
            val = int(raw) if raw is not None and raw != '' else 10
            # Vypnutí není povoleno – jakákoli neplatná/nulová hodnota spadne na výchozích 10 minut.
            if val < 1:
                val = 10
            uspech = await asyncio.to_thread(intranet_data.uloz_osobni_auto_odhlaseni, user_id, val)
            if uspech:
                osobni_input.value = val
                popis = f"{val} min"
                intranet_logger.log_activity(user_name, "Osobní profil", f"Osobní auto-odhlášení nastaveno na: {popis}")
                ui.notify(f'Uloženo. Odhlášení po nečinnosti: {popis}. Projeví se po příštím přihlášení.', type='positive')
            else:
                ui.notify('Chyba při ukládání.', type='negative')

        ui.button('Uložit', icon='save', on_click=ulozit_osobni_odhlaseni).classes('bg-amber-500 hover:bg-amber-600 text-white font-bold h-10 px-6 shadow-sm rounded-xl mt-2')

        ui.label('Narozeninové přání').classes('text-2xl font-bold mt-8 mb-4 text-blue-800 border-t pt-8')
        ui.label('Když je vypnuto, nepřijde vám automatický e-mail s přáním v den vašich narozenin.') \
            .classes('text-sm text-gray-500 mb-2')

        async def prepnout_narozeninove_prani(e):
            uspech = await asyncio.to_thread(intranet_data.uloz_email_narozeniny, user_id, e.value)
            if uspech:
                stav = 'zapnuto' if e.value else 'vypnuto'
                intranet_logger.log_activity(user_name, "Osobní profil", f"Narozeninové přání e-mailem: {stav}")
                ui.notify(f'Narozeninové přání e-mailem: {stav}.', type='positive')
            else:
                ui.notify('Chyba při ukládání.', type='negative')

        ui.switch('Zasílat mi přání k narozeninám e-mailem',
                  value=intranet_data.ziskej_email_narozeniny(user_id),
                  on_change=prepnout_narozeninove_prani)

# ==========================================
# --- DOCHÁZKA S KALENDÁŘEM A SLOŽKAMI ---
# ==========================================
# --- Evidence absencí: vzhled „timeline" (varianta 1) -------------------------
_TL_DNY = ('Po', 'Út', 'St', 'Čt', 'Pá', 'So', 'Ne')
_TL_MESICE = ('', 'Leden', 'Únor', 'Březen', 'Duben', 'Květen', 'Červen', 'Červenec',
              'Srpen', 'Září', 'Říjen', 'Listopad', 'Prosinec')
_TL_BARVY = {'Schváleno': '#16a34a', 'Čeká na schválení': '#d97706',
             'Zamítnuto': '#dc2626', 'Stornováno': '#9ca3af'}


def _tl_datum(z):
    """Datum řádku – absence má 'from', přesčas 'datum_od'. Vždy date kvůli řazení."""
    d = z.get('from') or z.get('datum_od')
    return d.date() if isinstance(d, datetime.datetime) else d


def _tl_mesic(d):
    ui.label(f'{_TL_MESICE[d.month]} {d.year}'.upper()).classes(
        'text-xs font-black tracking-widest text-gray-400 mt-4 mb-1 pl-1')


def _tl_karta(d, stav, storno_req=False, extra=''):
    """Datumová lišta + karta s barevným pruhem stavu. Vrací kartu k naplnění (`with`)."""
    with ui.row().classes('w-full items-stretch gap-3 mb-2 flex-nowrap'):
        with ui.column().classes('w-12 shrink-0 items-center justify-center gap-0'):
            ui.label(str(d.day) if d else '–').classes('text-2xl font-black text-gray-700 leading-none')
            ui.label(_TL_DNY[d.weekday()] if d else '').classes('text-[10px] font-bold text-gray-400')
        return ui.row().classes('flex-1 items-center justify-between gap-2 bg-white rounded-xl '
                                'border border-gray-200 p-3 hover:shadow-md ' + extra).style(
            'border-left:5px solid ' + ('#ea580c' if storno_req else _TL_BARVY.get(stav, '#9ca3af')))


@ui.refreshable
def vykresli_dochazku(user_id, user_name, vsechna_prava):
    _nast_doc = intranet_data.nacti_nastaveni_intranetu()
    presczasy_zapnuty = _nast_doc.get('presczasy_zapnuty', True)

    vsichni_uzivatele_komplet = intranet_data.ziskej_vsechny_uzivatele()
    muj_ucet = next((u for u in vsichni_uzivatele_komplet.values() if u['id'] == user_id), {})
    moji_sledovani = muj_ucet.get('sledovani_uzivatele', [])
    jsem_majitel = bool(moji_sledovani)

    ma_pristup_vsechny_slozky = 'vse' in vsechna_prava or 'slozky_vse' in vsechna_prava
    ma_pristup_vsechny_kalendare = 'vse' in vsechna_prava or 'kalendar_vse' in vsechna_prava

    ma_pristup_zadosti = 'vse' in vsechna_prava or 'dochazka_zadosti' in vsechna_prava
    is_global_admin = 'vse' in vsechna_prava or 'dochazka_admin' in vsechna_prava
    ma_pristup_mazani = 'vse' in vsechna_prava or 'dochazka_mazani' in vsechna_prava

    ma_pristup_ucetnictvi = 'vse' in vsechna_prava or 'ucetni_vse' in vsechna_prava or 'ucetni_pristup' in vsechna_prava

    ukazat_tlacitko_export = 'vse' in vsechna_prava or 'dochazka_export' in vsechna_prava or any(p.startswith('tisk_') for p in vsechna_prava)
    ma_ikos_export = 'vse' in vsechna_prava or 'dochazka_admin' in vsechna_prava or 'ikos_export' in vsechna_prava
    ma_jakekoliv_pravo_kalendar = 'vse' in vsechna_prava or ma_pristup_vsechny_kalendare or any(p.startswith('kalendar_') for p in vsechna_prava) or jsem_majitel

    typy_volna = intranet_data.ziskej_typy_volna()
    aktualni_oddeleni_dict = intranet_data.ziskej_vsechna_oddeleni()

    ma_tisk_typy_vse_b = 'vse' in vsechna_prava or 'tisk_vse' in vsechna_prava
    ma_tisk_odd_vse_b = 'vse' in vsechna_prava or 'tisk_odd_vse' in vsechna_prava

    povoleny_porovnani_oddeleni = set()
    if 'vse' in vsechna_prava or 'porovnani_vse' in vsechna_prava:
        povoleny_porovnani_oddeleni = set(aktualni_oddeleni_dict.keys())
    else:
        for _r_nazev in aktualni_oddeleni_dict.keys():
            if f'porovnani_odd_{_r_nazev.lower()}' in vsechna_prava:
                povoleny_porovnani_oddeleni.add(_r_nazev)

    # Holder pro cílený refresh žádostí — plněn po definici sub-refreshables
    _rf = {}

    def _prijemci_zadosti(cilovy_id, oddeleni_cil):
        """Komu chodí mail o žádosti daného uživatele: jeho manažeři + kdo ho sleduje,
        jinak fallback na hlavní vedoucí jeho oddělení / dochazka_admin."""
        cil = next((u for u in vsichni_uzivatele_komplet.values() if u['id'] == cilovy_id), {})
        prijemci = set()

        man_ids = cil.get('manager_id', [])
        if man_ids:
            for mail, udata in vsichni_uzivatele_komplet.items():
                if udata['id'] in man_ids and udata.get('email_nova_zadost', True):
                    prijemci.add(mail)

        for mail, udata in vsichni_uzivatele_komplet.items():
            if cilovy_id in udata.get('sledovani_uzivatele', []) and udata.get('email_nova_zadost', True):
                prijemci.add(mail)

        if not prijemci:
            # ponytail: 1 dotaz na všechna relevantní práva místo N× ziskej_prava_uzivatele v cyklu
            hledana_prava = ['vse', 'dochazka_admin']
            hledana_prava += [f'hlavni_vedouci_{o.lower()}' for o in oddeleni_cil if o]
            opravneni_ids = set(intranet_data.ziskej_uzivatele_s_pravem(*hledana_prava).keys())

            for mail, udata in vsichni_uzivatele_komplet.items():
                if udata['id'] == cilovy_id: continue
                if not udata.get('email_nova_zadost', True): continue
                if udata['id'] in opravneni_ids:
                    prijemci.add(mail)
        return prijemci

    _sub_refreshes = []

    def smazat_zadost_btn(zid):
        with ui.dialog() as dlg, ui.card().classes('p-6 rounded-xl w-full max-w-sm'):
            ui.label('Smazání žádosti').classes('text-xl font-bold mb-4 text-red-600')
            ui.label('Opravdu chcete tuto žádost nenávratně smazat?').classes('mb-6 text-gray-700')
            async def potvrdit():
                uspech = await asyncio.to_thread(intranet_data.smaz_zadost_volno, zid)
                if uspech:
                    intranet_logger.log_activity(user_name, "Docházka", f"Trvale smazána žádost ID: {zid}")
                    ui.notify('Žádost byla úspěšně smazána.', type='positive', position='top')
                    dlg.close()
                    intranet_data.invaliduj_cache_dochazky()
                    await asyncio.gather(
                        asyncio.to_thread(intranet_data.ziskej_vsechny_uzivatele),
                        asyncio.to_thread(intranet_data.ziskej_zadosti, None),
                    )
                    ui.timer(0, _rf.get('fn', vykresli_dochazku.refresh), once=True)
                else:
                    ui.notify('Chyba při mazání.', type='negative', position='top')
            with ui.row().classes('w-full justify-between'):
                ui.button('Zrušit', on_click=dlg.close).classes('bg-gray-400 text-white font-bold')
                ui.button('Smazat', on_click=potvrdit).classes('bg-red-600 text-white font-bold shadow-md')
        dlg.open()

    def stornovat_zadost_btn(zid, z_data):
        with ui.dialog() as dlg, ui.card().classes('p-6 rounded-xl w-full max-w-md'):
            ui.label('Stornování záznamu').classes('text-xl font-bold mb-4 text-orange-600')
            ui.label('Tato žádost již prošla procesem vyřízení. Nelze ji tedy beze stopy smazat, ale můžete ji stornovat.').classes('mb-4 text-gray-700 text-sm')
            duvod_input = ui.textarea('Zadejte důvod storna (povinné) *').classes('w-full mb-4').props('outlined')

            async def potvrdit():
                duvod = duvod_input.value.strip() if duvod_input.value else ""
                if not duvod:
                    ui.notify('Musíte zadat důvod storna!', type='warning')
                    return

                uspech = await asyncio.to_thread(intranet_data.stornuj_zadost_volno, zid, user_id, duvod)
                if uspech:
                    intranet_logger.log_activity(user_name, "Stornování žádosti", f"ID: {zid} | Důvod: {duvod}")
                    ui.notify('Žádost byla úspěšně stornována.', type='positive', position='top')
                    dlg.close()
                    intranet_data.invaliduj_cache_dochazky()
                    await asyncio.gather(
                        asyncio.to_thread(intranet_data.ziskej_vsechny_uzivatele),
                        asyncio.to_thread(intranet_data.ziskej_zadosti, None),
                    )
                    ui.timer(0, _rf.get('fn', vykresli_dochazku.refresh), once=True)

                    if z_data['user_iduser'] != user_id:
                        try:
                            zadatel_email, zadatel_data = None, None
                            for mail, u in vsichni_uzivatele_komplet.items():
                                if u['id'] == z_data['user_iduser']:
                                    zadatel_email, zadatel_data = mail, u
                                    break
                            if zadatel_email and zadatel_data.get('email_vyrizeni_zadosti', True):
                                predmet = f"STORNO: Vaše schválená žádost o {z_data['typ']} byla stornována"
                                _app_url = intranet_data.nacti_nastaveni_intranetu().get('app_url', '').strip()
                                _proklik = f"\n➡ Přejít do aplikace: {_app_url}\n\n" if _app_url else "\n\n"
                                text = f"Dobrý den {zadatel_data['jmeno']},\n\nVaše dříve schválená žádost o {z_data['typ']} (v termínu {formatuj_datum(z_data['from'])} až {formatuj_datum(z_data['to'])}) byla STORNOVÁNA vaším nadřízeným.\n\n"
                                text += f"Důvod storna: {duvod}{_proklik}S pozdravem,\nMoje JIPka"
                                asyncio.create_task(asyncio.to_thread(intranet_emaily.odesli_upozorneni_email, zadatel_email, predmet, text))
                            intranet_notifikace.pridej(
                                z_data['user_iduser'],
                                f'Vaše žádost o {z_data["typ"]} ({formatuj_datum(z_data["from"])} – {formatuj_datum(z_data["to"])}) byla stornována. Důvod: {duvod}',
                                'warning'
                            )
                        except Exception as e: print("Chyba e-mail storno:", e)
                else:
                    ui.notify('Chyba při stornování.', type='negative', position='top')

            with ui.row().classes('w-full justify-between'):
                ui.button('Zpět', on_click=dlg.close).classes('bg-gray-400 text-white font-bold')
                ui.button('Potvrdit storno', on_click=potvrdit).classes('bg-orange-600 hover:bg-orange-700 text-white font-bold shadow-md')
        dlg.open()

    def pozadat_storno_btn(zid, z_data):
        """Zaměstnanec žádá vedoucího o storno své už schválené absence."""
        with ui.dialog() as dlg, ui.card().classes('p-6 rounded-xl w-full max-w-md'):
            ui.label('Žádost o storno absence').classes('text-xl font-bold mb-4 text-orange-600')
            ui.label('Absence je schválená — storno musí potvrdit vedoucí. Napište prosím důvod.').classes('mb-4 text-gray-700')
            duvod_input = ui.textarea('Zadejte důvod storna (povinné) *').classes('w-full mb-4').props('outlined')

            async def potvrdit():
                duvod = duvod_input.value.strip() if duvod_input.value else ""
                if not duvod:
                    ui.notify('Musíte zadat důvod storna!', type='warning')
                    return

                uspech = await asyncio.to_thread(intranet_data.pozadej_o_storno, zid, user_id, duvod)
                if not uspech:
                    ui.notify('Chyba při odesílání žádosti o storno.', type='negative', position='top')
                    return

                intranet_logger.log_activity(user_name, "Stornování žádosti", f"Žádost o storno ID: {zid} | Důvod: {duvod}")
                ui.notify('Žádost o storno odeslána ke schválení.', type='positive', position='top')
                dlg.close()
                intranet_data.invaliduj_cache_dochazky()
                await asyncio.to_thread(intranet_data.ziskej_zadosti, None)
                ui.timer(0, _rf.get('fn', vykresli_dochazku.refresh), once=True)

                try:
                    oddeleni_cil = [o.strip() for o in (z_data.get('oddeleni') or '').split(',')]
                    predmet = f"ŽÁDOST O STORNO: {user_name}"
                    _app_url = intranet_data.nacti_nastaveni_intranetu().get('app_url', '').strip()
                    _proklik = f"\n\n➡ Přejít do aplikace: {_app_url}" if _app_url else ""
                    text = (f"Dobrý den,\n\n{user_name} žádá o stornování schválené absence "
                            f"\"{z_data['typ']}\" (v termínu {formatuj_datum(z_data['from'])} – {formatuj_datum(z_data['to'])}).\n\n"
                            f"Důvod storna: {duvod}\n\nProsím, vyřiďte žádost v Moje JIPka.{_proklik}\n\n"
                            f"Toto je automatická zpráva z portálu Moje JIPka.")
                    for p_email in _prijemci_zadosti(user_id, oddeleni_cil):
                        asyncio.create_task(asyncio.to_thread(intranet_emaily.odesli_upozorneni_email, p_email, predmet, text))
                        p_data = vsichni_uzivatele_komplet.get(p_email)
                        if p_data:
                            intranet_notifikace.pridej(
                                p_data['id'],
                                f'{user_name} žádá o storno absence {z_data["typ"]} ({formatuj_datum(z_data["from"])} – {formatuj_datum(z_data["to"])}). Důvod: {duvod}',
                                'warning'
                            )
                except Exception as e: print("Chyba e-mail žádost o storno:", e)

            with ui.row().classes('w-full justify-between'):
                ui.button('Zpět', on_click=dlg.close).classes('bg-gray-400 text-white font-bold')
                ui.button('Odeslat žádost', on_click=potvrdit).classes('bg-orange-600 hover:bg-orange-700 text-white font-bold shadow-md')
        dlg.open()

    def vyrid_storno_btn(zid, z_data, schvalit):
        """Vedoucí schválí (= absence se stornuje) nebo zamítne žádost o storno."""
        with ui.dialog() as dlg, ui.card().classes('p-6 rounded-xl w-full max-w-md'):
            ui.label('Schválení storna' if schvalit else 'Zamítnutí storna').classes(
                f'text-xl font-bold mb-2 {"text-orange-600" if schvalit else "text-red-600"}')
            ui.label(f'{z_data.get("u_jmeno", "")} {z_data.get("u_prijmeni", "")} — {z_data["typ"]} '
                     f'({formatuj_datum(z_data["from"])} – {formatuj_datum(z_data["to"])})').classes('text-sm text-gray-700')
            ui.label(f'Důvod žadatele: {z_data.get("storno_req_reason") or "—"}').classes('text-sm text-gray-500 mb-4')

            async def potvrdit():
                duvod = (z_data.get('storno_req_reason') or 'bez důvodu')
                if schvalit:
                    uspech = await asyncio.to_thread(intranet_data.stornuj_zadost_volno, zid, user_id, duvod)
                else:
                    uspech = await asyncio.to_thread(intranet_data.zrus_pozadavek_storna, zid)
                if not uspech:
                    ui.notify('Chyba při vyřízení.', type='negative', position='top')
                    return

                intranet_logger.log_activity(user_name, "Stornování žádosti",
                                             f"{'Schváleno' if schvalit else 'Zamítnuto'} storno ID: {zid} | Důvod: {duvod}")
                ui.notify('Storno provedeno.' if schvalit else 'Storno zamítnuto.', type='positive', position='top')
                dlg.close()
                intranet_data.invaliduj_cache_dochazky()
                await asyncio.gather(
                    asyncio.to_thread(intranet_data.ziskej_vsechny_uzivatele),
                    asyncio.to_thread(intranet_data.ziskej_zadosti, None),
                )
                ui.timer(0, _rf.get('fn', vykresli_dochazku.refresh), once=True)

                try:
                    zadatel_email, zadatel_data = None, None
                    for mail, u in vsichni_uzivatele_komplet.items():
                        if u['id'] == z_data['user_iduser']:
                            zadatel_email, zadatel_data = mail, u
                            break
                    _zprava = (f'Vaše žádost o storno absence {z_data["typ"]} '
                               f'({formatuj_datum(z_data["from"])} – {formatuj_datum(z_data["to"])}) byla '
                               f'{"schválena — absence je stornována" if schvalit else "zamítnuta — absence zůstává schválená"}.')
                    if zadatel_email and zadatel_data.get('email_vyrizeni_zadosti', True):
                        predmet = f"STORNO {'SCHVÁLENO' if schvalit else 'ZAMÍTNUTO'}: {z_data['typ']}"
                        _app_url = intranet_data.nacti_nastaveni_intranetu().get('app_url', '').strip()
                        _proklik = f"\n\n➡ Přejít do aplikace: {_app_url}" if _app_url else ""
                        text = f"Dobrý den {zadatel_data['jmeno']},\n\n{_zprava}{_proklik}\n\nS pozdravem,\nMoje JIPka"
                        asyncio.create_task(asyncio.to_thread(intranet_emaily.odesli_upozorneni_email, zadatel_email, predmet, text))
                    intranet_notifikace.pridej(z_data['user_iduser'], _zprava, 'success' if schvalit else 'warning')
                except Exception as e: print("Chyba e-mail vyřízení storna:", e)

            with ui.row().classes('w-full justify-between'):
                ui.button('Zpět', on_click=dlg.close).classes('bg-gray-400 text-white font-bold')
                ui.button('Potvrdit', on_click=potvrdit).classes(
                    f'{"bg-orange-600 hover:bg-orange-700" if schvalit else "bg-red-600 hover:bg-red-700"} text-white font-bold shadow-md')
        dlg.open()


    def upravit_zadost_btn(zid, z_data):
        with ui.dialog() as dlg_edit, ui.card().classes('p-6 rounded-xl w-full max-w-lg'):
            ui.label('Úprava schválené absence').classes('text-xl font-bold mb-1 text-blue-700')
            ui.label('Změny se okamžitě projeví — původní hodnoty nelze obnovit.').classes('text-xs text-gray-400 italic mb-4')

            typy = intranet_data.ziskej_typy_volna()

            with ui.row().classes('w-full gap-4 mb-3'):
                with ui.input('Datum od (RRRR-MM-DD)', value=str(z_data['from'])) as e_od:
                    e_od.classes('flex-1')
                    with e_od.add_slot('append'):
                        ui.icon('edit_calendar').on('click', lambda: menu_e_od.open()).classes('cursor-pointer hover:text-blue-500')
                    with ui.menu() as menu_e_od:
                        ui.date().bind_value(e_od)
                with ui.input('Datum do (RRRR-MM-DD)', value=str(z_data['to'])) as e_do:
                    e_do.classes('flex-1')
                    with e_do.add_slot('append'):
                        ui.icon('edit_calendar').on('click', lambda: menu_e_do.open()).classes('cursor-pointer hover:text-blue-500')
                    with ui.menu() as menu_e_do:
                        ui.date().bind_value(e_do)

            _e_cas_od_val = str(z_data['cas_od'])[:5] if z_data.get('cas_od') else ''
            _e_cas_do_val = str(z_data['cas_do'])[:5] if z_data.get('cas_do') else ''
            with ui.row().classes('w-full gap-4 mb-3 items-end'):
                e_cas_od = ui.input('Čas od (HH:MM)', value=_e_cas_od_val, placeholder='07:00').classes('flex-1').props('type=time')
                e_cas_do = ui.input('Čas do (HH:MM)', value=_e_cas_do_val, placeholder='15:30').classes('flex-1').props('type=time')
                e_typ = ui.select(typy, value=z_data.get('typ_id'), label='Typ volna').classes('flex-[2]')

            e_duvod = ui.textarea('Důvod úpravy *', placeholder='Povinné — zapište proč byla absence upravena').classes('w-full mb-2').props('outlined rows=2')

            def _vypocitej_delku_edit(d_od_str, d_do_str, cas_od_str, cas_do_str):
                try:
                    d1 = datetime.datetime.strptime(d_od_str, '%Y-%m-%d').date()
                    d2 = datetime.datetime.strptime(d_do_str, '%Y-%m-%d').date()
                    if d1 == d2 and cas_od_str and cas_do_str:
                        t1 = datetime.datetime.strptime(cas_od_str, '%H:%M').time()
                        t2 = datetime.datetime.strptime(cas_do_str, '%H:%M').time()
                        diff_min = (datetime.datetime.combine(d1, t2) - datetime.datetime.combine(d1, t1)).seconds // 60
                        if diff_min > 0:
                            return diff_min // 60, diff_min % 60
                    prac_dny = sum(
                        1 for i in range((d2 - d1).days + 1)
                        if (d1 + datetime.timedelta(days=i)).weekday() < 5
                        and not je_statni_svatek(d1 + datetime.timedelta(days=i))
                    )
                    return prac_dny * 8, 0
                except Exception:
                    return int(z_data.get('suma_hodin') or 0) or int(z_data['sumaHours']), int(z_data.get('suma_minut') or 0)

            async def potvrdit_upravu():
                if not e_od.value or not e_do.value:
                    return ui.notify('Vyplňte datum od/do!', type='warning', position='top')
                if not e_cas_od.value or not e_cas_do.value:
                    return ui.notify('Vyplňte čas od i čas do!', type='warning', position='top')
                if not e_duvod.value or not e_duvod.value.strip():
                    return ui.notify('Důvod úpravy je povinný!', type='warning', position='top')
                try:
                    d_od = datetime.datetime.strptime(e_od.value, '%Y-%m-%d').date()
                    d_do = datetime.datetime.strptime(e_do.value, '%Y-%m-%d').date()
                    if d_do < d_od:
                        return ui.notify('Datum "Do" nesmí být před datem "Od"!', type='negative', position='top')
                except Exception:
                    return ui.notify('Neplatný formát data!', type='negative', position='top')
                _e_cas_od = e_cas_od.value
                _e_cas_do = e_cas_do.value
                _eh_val, _em_val = _vypocitej_delku_edit(e_od.value, e_do.value, _e_cas_od or '', _e_cas_do or '')
                if _eh_val == 0 and _em_val == 0:
                    return ui.notify('Celková délka musí být alespoň 1 minuta!', type='warning', position='top')
                uspech = await asyncio.to_thread(
                    intranet_data.uprav_zadost_volno,
                    zid, user_id, e_od.value, e_do.value,
                    _eh_val, _em_val, e_typ.value, e_duvod.value.strip(),
                    _e_cas_od, _e_cas_do
                )
                if uspech:
                    _dur_str = f"{_eh_val}h {_em_val}min" if _eh_val and _em_val else (f"{_eh_val}h" if _eh_val else f"{_em_val}min")
                    _cas_str_e = f" ({_e_cas_od}–{_e_cas_do})" if _e_cas_od and _e_cas_do else ""
                    intranet_logger.log_activity(user_name, "Úprava absence",
                        f"ID: {zid} | {z_data['u_jmeno']} {z_data['u_prijmeni']} | "
                        f"Od: {e_od.value} | Do: {e_do.value}{_cas_str_e} | "
                        f"Typ: {typy.get(e_typ.value, e_typ.value)} | Délka: {_dur_str} | "
                        f"Důvod: {e_duvod.value.strip()[:120]}")
                    ui.notify('Absence byla upravena.', type='positive', position='top', icon='check_circle')
                    dlg_edit.close()
                    intranet_data.invaliduj_cache_dochazky()
                    await asyncio.gather(
                        asyncio.to_thread(intranet_data.ziskej_vsechny_uzivatele),
                        asyncio.to_thread(intranet_data.ziskej_zadosti, None),
                    )
                    ui.timer(0, _rf.get('fn', vykresli_dochazku.refresh), once=True)
                    if z_data['user_iduser'] != user_id:
                        try:
                            zadatel_email, zadatel_data = None, None
                            for mail, u in vsichni_uzivatele_komplet.items():
                                if u['id'] == z_data['user_iduser']:
                                    zadatel_email, zadatel_data = mail, u
                                    break
                            if zadatel_email and zadatel_data.get('email_vyrizeni_zadosti', True):
                                predmet = f"ÚPRAVA: Vaše schválená absence ({z_data['typ']}) byla upravena"
                                _app_url = intranet_data.nacti_nastaveni_intranetu().get('app_url', '').strip()
                                _proklik = f"\n➡ Přejít do aplikace: {_app_url}\n\n" if _app_url else "\n\n"
                                text = (f"Dobrý den {zadatel_data['jmeno']},\n\n"
                                        f"Vaše schválená absence ({z_data['typ']}) byla upravena vaším nadřízeným.\n\n"
                                        f"Nový termín: {formatuj_datum(e_od.value)} – {formatuj_datum(e_do.value)}{_cas_str_e}\n"
                                        f"Typ: {typy.get(e_typ.value, e_typ.value)}\n"
                                        f"Délka: {_dur_str}\n\n"
                                        f"Důvod úpravy: {e_duvod.value.strip()}{_proklik}S pozdravem,\nMoje JIPka")
                                asyncio.create_task(asyncio.to_thread(intranet_emaily.odesli_upozorneni_email, zadatel_email, predmet, text))
                        except Exception as ex:
                            print("Chyba e-mail úprava absence:", ex)
                else:
                    ui.notify('Chyba při úpravě absence.', type='negative', position='top')

            with ui.row().classes('w-full justify-between mt-2'):
                ui.button('Zpět', on_click=dlg_edit.close).classes('bg-gray-400 text-white font-bold')
                ui.button('Uložit úpravu', icon='save', on_click=potvrdit_upravu).classes('bg-blue-600 hover:bg-blue-700 text-white font-bold shadow-md')
        dlg_edit.open()

    def _stornovat_presczas_btn(pid):
        with ui.dialog() as dlg_ps, ui.card().classes('p-6 rounded-xl w-full max-w-md'):
            ui.label('Stornování přesčasu').classes('text-xl font-bold mb-4 text-orange-600')
            ui.label('Přesčas byl automaticky schválen. Stornováním jej zneplatníte.').classes('text-gray-600 text-sm mb-4')
            duvod_ps = ui.textarea('Důvod storna (povinné) *').classes('w-full mb-4').props('outlined')
            async def _potvrdit_storno_ps():
                d = duvod_ps.value.strip() if duvod_ps.value else ''
                if not d:
                    ui.notify('Vyplňte důvod storna!', type='warning')
                    return
                ok = await asyncio.to_thread(intranet_data.stornuj_presczas, pid, user_id, d)
                if ok:
                    intranet_logger.log_activity(user_name, "Přesčas", f"Stornován přesčas ID: {pid} | Důvod: {d}")
                    ui.notify('Přesčas stornován.', type='positive', position='top')
                    dlg_ps.close()
                    intranet_data.invaliduj_cache_dochazky()
                    await asyncio.to_thread(intranet_data.ziskej_presczasy, None)
                    ui.timer(0, _rf.get('fn', vykresli_dochazku.refresh), once=True)
                else:
                    ui.notify('Chyba při stornování.', type='negative', position='top')
            with ui.row().classes('w-full justify-between'):
                ui.button('Zpět', on_click=dlg_ps.close).classes('bg-gray-400 text-white font-bold')
                ui.button('Stornovat', on_click=_potvrdit_storno_ps).classes('bg-orange-600 hover:bg-orange-700 text-white font-bold shadow-md')
        dlg_ps.open()

    povoleny_schvalovat_oddeleni = []
    if is_global_admin:
        povoleny_schvalovat_oddeleni = list(aktualni_oddeleni_dict.keys())
    else:
        for r_nazev in aktualni_oddeleni_dict.keys():
            if f'hlavni_vedouci_{r_nazev.lower()}' in vsechna_prava:
                povoleny_schvalovat_oddeleni.append(r_nazev)

    jsem_neci_manazer = False
    for u in vsichni_uzivatele_komplet.values():
        if user_id in u.get('manager_id', []):
            jsem_neci_manazer = True
            break

    povoleny_sprava_oddeleni = []
    if ma_pristup_vsechny_slozky:
        povoleny_sprava_oddeleni = list(aktualni_oddeleni_dict.keys())
    else:
        for r_nazev in aktualni_oddeleni_dict.keys():
            if f'slozka_{r_nazev.lower()}' in vsechna_prava:
                povoleny_sprava_oddeleni.append(r_nazev)

    def filtruj_kalendar(data_podle_oddeleni):
        vyfiltrovane = {}
        for odd_nazev, polozky in data_podle_oddeleni.items():
            klic_kalendar = f"kalendar_{odd_nazev.lower()}"
            if ma_pristup_vsechny_kalendare or klic_kalendar in vsechna_prava or jsem_majitel:
                vyfiltrovane[odd_nazev] = polozky
        return vyfiltrovane

    with ui.dialog().props('full-width') as kalendar_dlg, ui.card().classes('w-full max-w-7xl mx-auto min-h-[70vh] p-4 sm:p-6 lg:p-8 bg-white rounded-xl'):
        if 'cal_year' not in app.storage.user:
            app.storage.user['cal_year'] = datetime.date.today().year
            app.storage.user['cal_month'] = datetime.date.today().month
            app.storage.user['cal_odd_filter'] = 'Všechna'
            app.storage.user['cal_typ_filter'] = 'Všechny'

        dostupna_oddeleni_kalendar = []
        for odd_nazev in aktualni_oddeleni_dict.keys():
            if odd_nazev.lower() == 'admin':
                continue
            if ma_pristup_vsechny_kalendare or f"kalendar_{odd_nazev.lower()}" in vsechna_prava:
                dostupna_oddeleni_kalendar.append(odd_nazev)

        filter_options = {}
        if jsem_majitel:
            filter_options['Sledovani'] = '👑 Moji sledovaní uživatelé (Majitel)'
        else:
            filter_options['Všechna'] = 'Všechna oddělení'
            for o in sorted(dostupna_oddeleni_kalendar):
                filter_options[o] = o

        aktualni_filtr = app.storage.user.get('cal_odd_filter')
        if aktualni_filtr not in filter_options:
            app.storage.user['cal_odd_filter'] = list(filter_options.keys())[0]

        dostupne_typy = sorted({(v.get('typ') or '').strip()
                                for v in intranet_data.ziskej_vsechna_volna_kalendar(False)
                                if (v.get('typ') or '').strip()})
        typ_options = {'Všechny': 'Všechny typy'}
        for t in dostupne_typy:
            typ_options[t] = t
        if app.storage.user.get('cal_typ_filter') not in typ_options:
            app.storage.user['cal_typ_filter'] = 'Všechny'

        # Ztlumená paleta barev pro typy absencí (nízký vzájemný kontrast → bez cirkusu)
        TYP_PALETA = ['sky', 'emerald', 'amber', 'violet', 'rose', 'teal', 'orange', 'indigo', 'lime', 'fuchsia', 'cyan', 'stone']
        typ_barvy = {}
        for _i, _t in enumerate(dostupne_typy):
            _fam = TYP_PALETA[_i % len(TYP_PALETA)]
            typ_barvy[_t] = (f'bg-{_fam}-200', f'bg-{_fam}-400', f'text-{_fam}-900')

        # Filtr oddělení má smysl jen když je z čeho vybírat (vedoucí více oddělení).
        zobraz_odd_filtr = (not jsem_majitel) and (len(dostupna_oddeleni_kalendar) > 1)
        # Filtr typu absence jen pro vedoucí oddělení (ne řadové členy).
        jsem_vedouci_oddeleni = jsem_majitel or ma_pristup_vsechny_kalendare or any(p.startswith('hlavni_vedouci_') for p in vsechna_prava)
        # Řadový člen vidí u kolegů jen veřejné typy; ostatní se maskují jako "Absence".
        VEREJNE_TYPY = {'dovolená', 'homeoffice'}
        @ui.refreshable
        def vykresli_mesic():
            try:
                y = app.storage.user['cal_year']
            except RuntimeError:
                return  # klient odpojen, refresh nemá komu kreslit
            m = app.storage.user['cal_month']
            vybrany_filter = app.storage.user.get('cal_odd_filter', 'Všechna')
            vybrany_typ = app.storage.user.get('cal_typ_filter', 'Všechny') if jsem_vedouci_oddeleni else 'Všechny'

            with ui.row().classes('w-full justify-center items-center gap-3 sm:gap-6 mb-5 bg-blue-50 p-2 rounded-xl'):
                def posun(d):
                    nm = m + d
                    ny = y
                    if nm < 1: nm = 12; ny -= 1
                    if nm > 12: nm = 1; ny += 1
                    app.storage.user['cal_year'] = ny
                    app.storage.user['cal_month'] = nm
                    vykresli_mesic.refresh()

                ui.button('<<<', on_click=lambda: posun(-1)).props('flat dense no-caps').classes('text-2xl font-bold text-blue-700 min-w-0 px-3')
                mesice_cz = ['', 'Leden', 'Únor', 'Březen', 'Duben', 'Květen', 'Červen', 'Červenec', 'Srpen', 'Září', 'Říjen', 'Listopad', 'Prosinec']
                ui.label(f'{mesice_cz[m]} {y}').classes('text-2xl font-bold text-gray-800 text-center w-52')
                ui.button('>>>', on_click=lambda: posun(1)).props('flat dense no-caps').classes('text-2xl font-bold text-blue-700 min-w-0 px-3')

            vsechna_volna = intranet_data.ziskej_vsechna_volna_kalendar(False)
            volna_podle_oddeleni = {}
            for v in vsechna_volna:
                odd_list = (v.get('oddeleni') or 'Bez oddělení').split(',')
                for odd in odd_list:
                    odd = odd.strip()
                    if not odd: odd = 'Bez oddělení'
                    volna_podle_oddeleni.setdefault(odd, []).append(v)

            volna_podle_oddeleni = filtruj_kalendar(volna_podle_oddeleni)

            stavy_barvy = {
                1: ('Čeká na schválení', 'bg-yellow-400', 'text-black'),
                2: ('Schváleno', 'bg-green-500', 'text-white'),
                3: ('Zamítnuto', 'bg-red-500', 'text-white'),
                4: ('Stornováno', 'bg-gray-400', 'text-white')
            }

            zobraz_typ_chips = bool(volna_podle_oddeleni and dostupne_typy and jsem_vedouci_oddeleni)
            if zobraz_typ_chips or zobraz_odd_filtr:
                def nastav_typ(t):
                    akt = app.storage.user.get('cal_typ_filter', 'Všechny')
                    app.storage.user['cal_typ_filter'] = 'Všechny' if akt == t else t
                    vykresli_mesic.refresh()
                with ui.row().classes('w-full gap-2 mb-4 flex-wrap items-center p-3 bg-gray-50 rounded-xl'):
                    if zobraz_typ_chips:
                        ui.label('Typ absence:').classes('font-bold text-gray-500 text-sm mr-1')
                        vse_akt = (vybrany_typ == 'Všechny')
                        vse_cls = 'items-center gap-1.5 px-3 py-1 rounded-full cursor-pointer transition-all select-none ' + ('bg-blue-600 text-white shadow' if vse_akt else 'bg-white text-gray-500 shadow-sm hover:bg-gray-100')
                        with ui.row().classes(vse_cls).on('click', lambda: nastav_typ('Všechny')):
                            ui.label('Vše').classes('text-xs font-bold')
                        for t in dostupne_typy:
                            c_bg, c_dot, c_txt = typ_barvy.get(t, ('bg-gray-200', 'bg-gray-400', 'text-gray-900'))
                            akt = (vybrany_typ == t)
                            chip_cls = 'items-center gap-1.5 px-3 py-1 rounded-full cursor-pointer transition-all select-none shadow-sm ' + (f'{c_bg} {c_txt} ring-2 ring-inset ring-gray-400' if akt else 'bg-white text-gray-500 hover:bg-gray-100')
                            with ui.row().classes(chip_cls).on('click', lambda t=t: nastav_typ(t)):
                                ui.element('div').classes(f'w-3 h-3 rounded-full {c_dot}')
                                ui.label(t).classes('text-xs font-bold')
                    if zobraz_odd_filtr:
                        with ui.row().classes('items-center ml-auto'):
                            cal_filter = ui.select(filter_options, value=app.storage.user.get('cal_odd_filter'), label='Filtrovat kalendář').bind_value(app.storage.user, 'cal_odd_filter').classes('w-64 bg-white').props('dense outlined rounded')
                            cal_filter.on_value_change(vykresli_mesic.refresh)

            pocet_dni = calendar.monthrange(y, m)[1]
            statni_svatky = ziskej_statni_svatky(y, m)
            dnes = datetime.date.today()
            dny_kratky = ['Po', 'Út', 'St', 'Čt', 'Pá', 'So', 'Ne']
            grid_cols = f'grid-template-columns: minmax(150px, 1.4fr) repeat({pocet_dni}, minmax(0, 1fr));'

            def den_kontext(d):
                wd = datetime.date(y, m, d).weekday()
                return (wd,
                        d == dnes.day and m == dnes.month and y == dnes.year,
                        d in statni_svatky,
                        wd >= 5)

            # ── Sesbírej absence po osobách (řádky matice) ──
            skupiny = {}  # odd_nazev -> { person_key: {'jmeno', 'dny': {den: {'stav_id','typ'}}} }

            # Roster: řádek pro každého aktivního zaměstnance viditelných oddělení,
            # i když v měsíci žádnou absenci nemá.
            for ud in intranet_data.ziskej_vsechny_uzivatele().values():
                if not ud['aktivni']:
                    continue
                if vybrany_filter == 'Sledovani' and ud['id'] not in moji_sledovani:
                    continue
                for odd in (ud['oddeleni'] or 'Bez oddělení').split(','):
                    odd = odd.strip() or 'Bez oddělení'
                    if odd.lower() == 'admin':
                        continue
                    if vybrany_filter not in ('Sledovani', 'Všechna') and odd != vybrany_filter:
                        continue
                    if not (jsem_majitel or ma_pristup_vsechny_kalendare or f"kalendar_{odd.lower()}" in vsechna_prava):
                        continue
                    skupiny.setdefault(odd, {})[ud['id']] = {'jmeno': ud['jmeno_cele'], 'dny': {}}

            for odd_nazev, list_volna in volna_podle_oddeleni.items():
                for v in list_volna:
                    if v['stav_id'] in (3, 4): continue

                    if vybrany_typ != 'Všechny' and (v.get('typ') or '').strip() != vybrany_typ:
                        continue

                    if vybrany_filter == 'Sledovani':
                        if v['user_iduser'] not in moji_sledovani:
                            continue
                    elif vybrany_filter != 'Všechna' and odd_nazev != vybrany_filter:
                        continue

                    typ_zobrazeny = v['typ']
                    if not jsem_vedouci_oddeleni and v.get('user_iduser') != user_id \
                            and (typ_zobrazeny or '').strip().lower() not in VEREJNE_TYPY:
                        typ_zobrazeny = 'Absence'

                    dny_v_mesici = {}
                    for i in range((v['to'] - v['from']).days + 1):
                        cur_d = v['from'] + datetime.timedelta(days=i)
                        if cur_d.year == y and cur_d.month == m:
                            dny_v_mesici[cur_d.day] = {'stav_id': v['stav_id'], 'typ': typ_zobrazeny}
                    if not dny_v_mesici:
                        continue

                    p_key = v.get('user_iduser') or f"{v['u_jmeno']} {v['u_prijmeni']}"
                    osoba = skupiny.setdefault(odd_nazev, {}).setdefault(p_key, {
                        'jmeno': f"{v['u_jmeno']} {v['u_prijmeni']}", 'dny': {}})
                    osoba['dny'].update(dny_v_mesici)

            if not skupiny:
                with ui.column().classes('w-full items-center py-16 gap-2'):
                    ui.icon('event_busy').classes('text-5xl text-gray-300')
                    ui.label('V tomto měsíci nejsou žádné absence.').classes('text-gray-400 font-bold')
            else:
                with ui.element('div').classes('w-full overflow-x-auto border border-gray-200 rounded-xl'):
                    with ui.element('div').classes('min-w-[900px]'):
                        # Hlavička: dny měsíce
                        with ui.element('div').classes('grid').style(grid_cols):
                            ui.label('Zaměstnanec').classes('sticky left-0 z-10 bg-gray-50 text-xs font-bold text-gray-500 uppercase px-3 py-2 flex items-center border-r border-b border-gray-200')
                            for d in range(1, pocet_dni + 1):
                                wd, is_today, is_holiday, is_weekend = den_kontext(d)
                                hc = 'flex flex-col items-center justify-center py-2.5 border-r border-b border-gray-200 '
                                if is_today: hc += 'bg-blue-600 text-white'
                                elif is_holiday: hc += 'bg-purple-100 text-purple-700'
                                elif is_weekend: hc += 'bg-red-50 text-red-500'
                                else: hc += 'bg-gray-50 text-gray-500'
                                with ui.element('div').classes(hc) as hcell:
                                    ui.label(dny_kratky[wd]).classes('text-[9px] font-bold uppercase leading-none opacity-80')
                                    ui.label(str(d)).classes('text-xs font-bold leading-tight')
                                if is_holiday: hcell.tooltip(statni_svatky[d])

                        vice_skupin = len(skupiny) > 1
                        for odd_nazev in sorted(skupiny):
                            osoby = skupiny[odd_nazev]
                            if vice_skupin:
                                with ui.element('div').classes('grid border-b border-gray-200').style(grid_cols):
                                    ui.label(odd_nazev).classes('sticky left-0 z-10 col-span-full bg-blue-50 text-xs font-bold text-blue-700 uppercase px-3 py-1.5')
                            for p_key in sorted(osoby, key=lambda k: osoby[k]['jmeno']):
                                osoba = osoby[p_key]
                                dny = osoba['dny']
                                with ui.element('div').classes('grid hover:bg-blue-50 min-h-[52px]').style(grid_cols):
                                    ui.label(osoba['jmeno']).classes('sticky left-0 z-10 bg-white text-sm font-medium text-gray-700 px-3 py-2 flex items-center truncate border-r border-b border-gray-100').tooltip(osoba['jmeno'])
                                    for d in range(1, pocet_dni + 1):
                                        wd, is_today, is_holiday, is_weekend = den_kontext(d)
                                        info = dny.get(d)
                                        cc = 'border-r border-b border-gray-100 '
                                        if info:
                                            cc += typ_barvy.get(info['typ'], ('bg-gray-300', '', ''))[0]
                                            if info['stav_id'] != 2:
                                                cc += ' opacity-40'
                                            stav_nazev = stavy_barvy.get(info['stav_id'], ('', '', ''))[0]
                                            tip = f"{osoba['jmeno']} · {info['typ']}"
                                            if stav_nazev:
                                                tip += f" · {stav_nazev}"
                                            ui.element('div').classes(cc).tooltip(tip)
                                        else:
                                            if is_today: cc += 'bg-blue-50'
                                            elif is_holiday: cc += 'bg-purple-50'
                                            elif is_weekend: cc += 'bg-gray-50'
                                            else: cc += 'bg-white'
                                            ui.element('div').classes(cc)

            with ui.row().classes('w-full justify-end mt-5'):
                ui.button('Zavřít kalendář', on_click=kalendar_dlg.close).classes('bg-gray-500 hover:bg-gray-600 text-white font-bold px-6 rounded-lg')

        vykresli_mesic()


    # --- HORNÍ HLAVIČKA DOCHÁZKY ---
    with ui.row().classes('w-full items-center justify-between mb-4'):
        with ui.column().classes('gap-0'):
            ui.label('Docházka a absence').classes('text-2xl font-bold text-gray-900 leading-tight')
            ui.label('Žádosti, přesčasy a schvalování na jednom místě.').classes('text-xs text-gray-500')

        with ui.row().classes('gap-4'):
            import intranet_exporty
            intranet_exporty.vykresli_exportni_tlacitka(
                user_name, vsechna_prava, vsichni_uzivatele_komplet, typy_volna,
                aktualni_oddeleni_dict, ma_pristup_vsechny_slozky, ma_tisk_odd_vse_b,
                ma_tisk_typy_vse_b, ma_pristup_ucetnictvi, ukazat_tlacitko_export,
                povoleny_porovnani_oddeleni, ma_ikos_export
            )

            if ma_jakekoliv_pravo_kalendar:
                ui.button('Kalendář absencí', icon='calendar_month', on_click=kalendar_dlg.open).classes('bg-blue-100 hover:bg-blue-200 text-blue-800 shadow-sm rounded-lg px-4 py-2 font-bold').props('flat')

    vsechny_zadosti = intranet_data.ziskej_zadosti(None)
    ceka_zadosti_vse = [z for z in vsechny_zadosti if z['stav_id'] == 1]
    vyrizene_zadosti_vse = [z for z in vsechny_zadosti if z['stav_id'] != 1]

    dostupne_vnitrni = []
    if ma_pristup_zadosti or povoleny_schvalovat_oddeleni or jsem_neci_manazer or jsem_majitel: dostupne_vnitrni.append('moje')
    if povoleny_sprava_oddeleni: dostupne_vnitrni.append('sprava')

    if not dostupne_vnitrni:
        ui.label('Nemáte oprávnění k žádnému modulu docházky.').classes('text-red-500 text-xl')
        return

    if app.storage.user.get('dochazka_vnitrni_tab') not in dostupne_vnitrni:
        app.storage.user['dochazka_vnitrni_tab'] = dostupne_vnitrni[0] if dostupne_vnitrni else None

    with ui.tabs().props('no-caps align=left dense active-color=red-6 indicator-color=red-6').bind_value(app.storage.user, 'dochazka_vnitrni_tab').classes('w-full justify-start border-b-2 border-gray-200 mb-4 text-gray-600') as main_tabs:
        if ma_pristup_zadosti or povoleny_schvalovat_oddeleni or jsem_neci_manazer or jsem_majitel: ui.tab('moje', label='Moje docházka').classes('text-sm font-semibold')
        if povoleny_sprava_oddeleni: ui.tab('sprava', label='Záznamy oddělení').classes('text-sm font-semibold')

    with ui.tab_panels(main_tabs, value=app.storage.user['dochazka_vnitrni_tab']).bind_value(app.storage.user, 'dochazka_vnitrni_tab').classes('w-full bg-transparent p-0'):

        # ==========================================================
        # TAB 1: MOJE DOCHÁZKA (Žádosti + Schvalování)
        # ==========================================================
        if ma_pristup_zadosti or povoleny_schvalovat_oddeleni or jsem_neci_manazer or jsem_majitel:
            with ui.tab_panel('moje'):

                muze_zadat_za_sebe = ma_pristup_zadosti
                muze_zadat_za_jine = bool(povoleny_schvalovat_oddeleni) or is_global_admin or jsem_neci_manazer or jsem_majitel

                if muze_zadat_za_sebe:
                    @ui.refreshable
                    def vykresli_dovolenky():
                        # ponytail: filtr nad cachovaným "vše" místo per-user DB dotazu v event-loopu
                        moje_z_d = [z for z in intranet_data.ziskej_zadosti(None) if z['user_iduser'] == user_id]
                        _aktualni_db = intranet_data.ziskej_vsechny_uzivatele()
                        _muj_ucet = next((u for u in _aktualni_db.values() if u['id'] == user_id), {})
                        current_year = datetime.date.today().year
                        moje_dov_vybrano = sum(float(z['sumaHours']) for z in moje_z_d if z['stav_id'] == 2 and z['typ'] == 'Dovolená' and z['from'].year == current_year)
                        moje_dov_ceka = sum(float(z['sumaHours']) for z in moje_z_d if z['stav_id'] == 1 and z['typ'] == 'Dovolená' and z['from'].year == current_year)
                        zaklad = _muj_ucet.get('base_vacation', 160.0)
                        prevod = _muj_ucet.get('carried_over_vacation', 0.0)
                        celkovy_narok = zaklad + prevod
                        moje_dov_zbyva = celkovy_narok - moje_dov_vybrano
                        _rz_datum = _muj_ucet.get('realny_zustatek_dovolene_datum')
                        if _rz_datum:
                            try:
                                _rz_datum_cz = datetime.date.fromisoformat(_rz_datum).strftime('%d.%m.%Y')
                            except (ValueError, TypeError):
                                _rz_datum_cz = _rz_datum
                            _rz_popis = f'Reálný zůstatek k {_rz_datum_cz}'
                        else:
                            _rz_popis = 'Reálný zůstatek'
                        realny_zustatek = _muj_ucet.get('realny_zustatek_dovolene')
                        _pct = int(round(100 * moje_dov_vybrano / celkovy_narok)) if celkovy_narok else 0
                        # --- HERO PRUH ZŮSTATKU ---
                        with ui.card().classes('w-full p-5 shadow-sm bg-white rounded-2xl border border-gray-200 mb-6'):
                            with ui.row().classes('w-full items-center gap-8 flex-nowrap max-md:flex-wrap'):
                                with ui.column().classes('gap-1 shrink-0'):
                                    ui.label('Zbývá dovolené').classes('text-xs text-gray-400 uppercase tracking-wider')
                                    with ui.row().classes('items-end gap-1'):
                                        ui.label(f'{moje_dov_zbyva:g}').classes('text-4xl font-extrabold text-gray-900 leading-none tracking-tight')
                                        ui.label('h').classes('text-sm font-semibold text-gray-500 mb-0.5')
                                with ui.column().classes('flex-1 gap-1 min-w-[240px]'):
                                    with ui.row().classes('w-full justify-between items-baseline'):
                                        ui.label(f'Vyčerpáno {moje_dov_vybrano:g} z {celkovy_narok:g} h').classes('text-xs text-gray-500')
                                        ui.label(f'{_pct} %').classes('text-xs font-bold text-gray-500')
                                    ui.linear_progress(min(_pct / 100, 1.0), show_value=False, size='8px').props('rounded color=blue-grey-9 track-color=grey-3').classes('w-full')
                                with ui.row().classes('gap-6 shrink-0'):
                                    for _lbl, _val, _cls in (
                                        ('Nárok', f'{zaklad:g} h', 'text-gray-900'),
                                        ('Převod', f'{prevod:g} h', 'text-gray-900'),
                                        ('Čerpáno', f'{moje_dov_vybrano:g} h', 'text-green-700'),
                                        ('Čeká', f'{moje_dov_ceka:g} h', 'text-yellow-700'),
                                        (_rz_popis, f'{realny_zustatek:g} h' if realny_zustatek is not None else '—', 'text-emerald-700'),
                                    ):
                                        with ui.column().classes('gap-0'):
                                            ui.label(_lbl).classes('text-xs text-gray-500')
                                            ui.label(_val).classes(f'text-lg font-bold {_cls} leading-tight')
                    vykresli_dovolenky()
                    _sub_refreshes.append(vykresli_dovolenky.refresh)

                podrizeni_pro_zadani = {}
                if muze_zadat_za_jine:
                    for mail, udata in vsichni_uzivatele_komplet.items():
                        if udata['id'] == user_id: continue
                        if 'admin' in udata['jmeno_cele'].lower() or 'admin' in udata.get('oddeleni', '').lower(): continue
                        odd_list = [o.strip() for o in (udata.get('oddeleni') or 'Bez oddělení').split(',')]
                        is_manager = user_id in udata.get('manager_id', [])
                        is_watched = udata['id'] in moji_sledovani
                        if is_global_admin or any(o in povoleny_schvalovat_oddeleni for o in odd_list) or is_manager or is_watched:
                            podrizeni_pro_zadani[udata['id']] = f"{udata['jmeno_cele']} ({udata.get('oddeleni', 'Bez oddělení')})"

                moznosti_zadani = {}
                if muze_zadat_za_sebe:
                    moznosti_zadani[user_id] = '🙋‍♂️ Za sebe'
                if podrizeni_pro_zadani:
                    for k, v in sorted(podrizeni_pro_zadani.items(), key=lambda item: item[1]):
                        moznosti_zadani[k] = f"👤 Za: {v}"

                moznosti_statistika = {}
                if len(podrizeni_pro_zadani) > 0:
                    moznosti_statistika['vse'] = '👥 Souhrn za všechny podřízené'
                moznosti_statistika[user_id] = '🙋‍♂️ Moje statistiky'
                for k, v in sorted(podrizeni_pro_zadani.items(), key=lambda item: item[1]):
                    moznosti_statistika[k] = f"👤 {v}"

                with ui.card().classes('w-full p-5 shadow-sm bg-white rounded-2xl mb-6 border border-gray-200'):
                    with ui.row().classes('w-full items-center justify-between mb-2'):
                        ui.label('Statistiky čerpání').classes('text-xs font-bold text-gray-400 uppercase tracking-widest')

                        with ui.row().classes('gap-4 items-center'):
                            with ui.input('Od data').classes('w-32 bg-white') as stat_od:
                                with stat_od.add_slot('append'):
                                    ui.icon('edit_calendar').on('click', lambda: menu_stat_od.open()).classes('cursor-pointer text-gray-500 hover:text-blue-500')
                                with ui.menu() as menu_stat_od: ui.date().bind_value(stat_od)
                            with ui.input('Do data').classes('w-32 bg-white') as stat_do:
                                with stat_do.add_slot('append'):
                                    ui.icon('edit_calendar').on('click', lambda: menu_stat_do.open()).classes('cursor-pointer text-gray-500 hover:text-blue-500')
                                with ui.menu() as menu_stat_do: ui.date().bind_value(stat_do)

                            dnes = datetime.date.today()
                            stat_od.value = f"{dnes.year}-01-01"
                            stat_do.value = f"{dnes.year}-12-31"

                            stat_kdo = ui.select(moznosti_statistika, value=user_id).classes('w-64 bg-white')

                            ui.button('Filtrovat', on_click=lambda: obnovit_statistiky.refresh()).classes('bg-gray-100 hover:bg-gray-200 text-gray-800 font-bold px-4 shadow-sm')

                    @ui.refreshable
                    def obnovit_statistiky():
                        try:
                            od_val = datetime.datetime.strptime(stat_od.value, '%Y-%m-%d').date() if stat_od.value else datetime.date.min
                            do_val = datetime.datetime.strptime(stat_do.value, '%Y-%m-%d').date() if stat_do.value else datetime.date.max
                        except Exception:
                            ui.label('Chybný formát data.').classes('text-red-500 font-bold mt-2')
                            return

                        vybrane_id = stat_kdo.value
                        platna_id = [user_id] + list(podrizeni_pro_zadani.keys())

                        _vsechny_z = intranet_data.ziskej_zadosti(None)
                        stat_data = {}
                        for z in _vsechny_z:
                            if 'admin' in z.get('u_jmeno', '').lower() or 'admin' in z.get('u_prijmeni', '').lower(): continue
                            if z['stav_id'] != 2: continue
                            if not (z['to'] >= od_val and z['from'] <= do_val): continue

                            zid = z['user_iduser']
                            if vybrane_id == 'vse':
                                if zid == user_id: continue
                                if zid not in platna_id: continue
                            else:
                                if zid != vybrane_id: continue

                            typ = z['typ']
                            if typ == 'Dovolená': continue

                            stat_data[typ] = stat_data.get(typ, 0.0) + float(z['sumaHours'])

                        # Souhrn zadaných přesčasů (mimo storno)
                        prescas_hodiny = 0.0
                        if presczasy_zapnuty:
                            for ot in intranet_data.ziskej_presczasy(None):
                                if ot['stav_id'] == 4: continue
                                if not (ot['datum_do'] >= od_val and ot['datum_od'] <= do_val): continue

                                ot_id = ot['user_iduser']
                                if vybrane_id == 'vse':
                                    if ot_id == user_id: continue
                                    if ot_id not in platna_id: continue
                                else:
                                    if ot_id != vybrane_id: continue

                                prescas_hodiny += float(ot['sumaHours'])

                        if not stat_data and prescas_hodiny == 0.0:
                            ui.label('Žádné čerpání jiného volna ve vybraném období.').classes('text-gray-400 italic mt-2 text-sm')
                        else:
                            with ui.row().classes('w-full gap-4 mt-2 flex-wrap'):
                                for typ, hodiny in sorted(stat_data.items()):
                                    with ui.card().classes('p-3 shadow-sm bg-blue-50 border border-blue-100 rounded-lg min-w-[120px] items-center justify-center'):
                                        ui.label(typ).classes('text-[10px] text-blue-500 font-bold uppercase tracking-wider mb-1 text-center')
                                        ui.label(f'{hodiny:g} h').classes('text-xl font-black text-blue-700')
                                if presczasy_zapnuty and prescas_hodiny > 0.0:
                                    with ui.card().classes('p-3 shadow-sm bg-blue-50 border border-blue-100 rounded-lg min-w-[120px] items-center justify-center'):
                                        ui.label('Přesčas').classes('text-[10px] text-blue-500 font-bold uppercase tracking-wider mb-1 text-center')
                                        ui.label(f'{prescas_hodiny:g} h').classes('text-xl font-black text-blue-700')

                    obnovit_statistiky()
                    _sub_refreshes.append(obnovit_statistiky.refresh)

                with ui.row().classes('w-full gap-8 items-start mb-8'):
                    if muze_zadat_za_sebe or muze_zadat_za_jine:
                        with ui.card().classes('flex-1 p-5 shadow-sm bg-white rounded-2xl border border-gray-200'):
                            # ── ABSENCE ──────────────────────────────────────
                            with ui.column().classes('w-full gap-0'):
                                ui.label('Nový záznam').classes('text-xs font-bold text-gray-400 uppercase tracking-widest mb-1')
                                ui.label('Vyplňte termín a typ volna. Žádost putuje vedoucímu ke schválení.').classes('text-xs text-gray-400 mb-4')

                                zadat_za_select = None
                                if len(moznosti_zadani) > 1 or (not muze_zadat_za_sebe and moznosti_zadani):
                                    zadat_za_select = ui.select(moznosti_zadani, value=list(moznosti_zadani.keys())[0], label='Kdo bude čerpat volno?') \
                                        .props('outlined dense options-dense').classes('w-full mb-1')
                                    ui.label('Záznam za podřízeného se ihned zařadí do vaší fronty, kde jej jedním klikem schválíte.').classes('text-xs text-gray-400 mb-4')

                                with ui.row().classes('w-full gap-3 mb-3'):
                                    with ui.input('Od data', placeholder='RRRR-MM-DD').props('outlined dense') as date_od:
                                        date_od.classes('flex-1')
                                        with date_od.add_slot('append'):
                                            ui.icon('event').on('click', lambda: menu_od.open()).classes('cursor-pointer text-gray-400 hover:text-blue-600')
                                    with ui.input('Do data', placeholder='RRRR-MM-DD').props('outlined dense') as date_do:
                                        date_do.classes('flex-1')
                                        with date_do.add_slot('append'):
                                            ui.icon('event').on('click', lambda: menu_do.open()).classes('cursor-pointer text-gray-400 hover:text-blue-600')

                                with ui.row().classes('w-full gap-3 mb-3 items-start'):
                                    cas_od_input = ui.input('Čas od', placeholder='07:00').classes('flex-1').props('type=time outlined dense')
                                    cas_do_input = ui.input('Čas do', placeholder='15:30').classes('flex-1').props('type=time outlined dense')
                                    typ_select = ui.select(typy_volna, value=list(typy_volna.keys())[0] if typy_volna else None, label='Typ volna') \
                                        .props('outlined dense options-dense').classes('flex-[2]')

                                def _vypocitej_delku_absence(d_od_str, d_do_str, cas_od_str, cas_do_str):
                                    try:
                                        d1 = datetime.datetime.strptime(d_od_str, '%Y-%m-%d').date()
                                        d2 = datetime.datetime.strptime(d_do_str, '%Y-%m-%d').date()
                                        if d1 == d2 and cas_od_str and cas_do_str:
                                            t1 = datetime.datetime.strptime(cas_od_str, '%H:%M').time()
                                            t2 = datetime.datetime.strptime(cas_do_str, '%H:%M').time()
                                            diff_min = (datetime.datetime.combine(d1, t2) - datetime.datetime.combine(d1, t1)).seconds // 60
                                            if diff_min > 0:
                                                return diff_min // 60, diff_min % 60
                                        prac_dny = sum(
                                            1 for i in range((d2 - d1).days + 1)
                                            if (d1 + datetime.timedelta(days=i)).weekday() < 5
                                            and not je_statni_svatek(d1 + datetime.timedelta(days=i))
                                        )
                                        return prac_dny * 8, 0
                                    except Exception:
                                        return 8, 0

                                with date_od:
                                    with ui.menu() as menu_od:
                                        ui.date().bind_value(date_od)
                                with date_do:
                                    with ui.menu() as menu_do:
                                        ui.date().bind_value(date_do)

                                _delka_lbl = ui.label().classes('w-full text-xs font-bold text-blue-700 bg-blue-50 border border-blue-100 rounded-lg px-3 py-2 mb-3')
                                _delka_lbl.set_visibility(False)

                                def _obnov_delku(_=None):
                                    if not (date_od.value and date_do.value):
                                        return _delka_lbl.set_visibility(False)
                                    _h, _m = _vypocitej_delku_absence(date_od.value, date_do.value, cas_od_input.value or '', cas_do_input.value or '')
                                    _delka_lbl.text = f"Celková délka: {_h}h {_m}min" if _h and _m else (f"Celková délka: {_h}h" if _h else f"Celková délka: {_m}min")
                                    _delka_lbl.set_visibility(True)
                                for _w in (date_od, date_do, cas_od_input, cas_do_input):
                                    _w.on_value_change(_obnov_delku)

                                async def odeslat_zadost():
                                    if not date_od.value or not date_do.value: return ui.notify('Vyplňte obě data!', type='warning', position='top')
                                    if not cas_od_input.value or not cas_do_input.value: return ui.notify('Vyplňte čas od i čas do!', type='warning', position='top')
                                    try:
                                        d_od = datetime.datetime.strptime(date_od.value, '%Y-%m-%d').date()
                                        d_do = datetime.datetime.strptime(date_do.value, '%Y-%m-%d').date()
                                        if d_do < d_od: return ui.notify('CHYBA: Datum "Do" nesmí být starší než datum "Od"!', type='negative', position='top')
                                    except Exception: return ui.notify('Neplatný formát data!', type='negative', position='top')
                                    _cas_od = cas_od_input.value
                                    _cas_do = cas_do_input.value
                                    _h_val, _m_val = _vypocitej_delku_absence(date_od.value, date_do.value, _cas_od or '', _cas_do or '')
                                    if _h_val == 0 and _m_val == 0:
                                        return ui.notify('Celková délka musí být alespoň 1 minuta!', type='warning', position='top')

                                    cilovy_id = zadat_za_select.value if zadat_za_select else user_id
                                    _dur_str_z = f"{_h_val}h {_m_val}min" if _h_val and _m_val else (f"{_h_val}h" if _h_val else f"{_m_val}min")
                                    _cas_str_z = f" ({_cas_od}–{_cas_do})" if _cas_od and _cas_do else ""

                                    # Kontrola překryvu s již nahlášenými absencemi (čekající + schválené)
                                    existujici_z = await asyncio.to_thread(intranet_data.ziskej_zadosti, cilovy_id)
                                    konflikty = [z for z in existujici_z if z['stav_id'] in (1, 2) and z['to'] >= d_od and z['from'] <= d_do]
                                    if konflikty:
                                        with ui.dialog() as potvrzeni_dialog, ui.card().classes('p-6 min-w-[350px]'):
                                            _kdo_text = 'máte' if cilovy_id == user_id else 'má vybraný uživatel'
                                            ui.label(f'V tyto dny již {_kdo_text} nahlášenou absenci, opravdu chcete zadat?').classes('text-base font-bold text-gray-800 mb-2')
                                            for k in konflikty:
                                                ui.label(f"• {k['typ']}: {k['from'].strftime('%d.%m.%Y')} – {k['to'].strftime('%d.%m.%Y')} ({k['stav']})").classes('text-sm text-gray-600')
                                            with ui.row().classes('w-full justify-end gap-2 mt-4'):
                                                ui.button('Zrušit', on_click=lambda: potvrzeni_dialog.submit(False)).props('flat no-caps').classes('text-gray-600')
                                                ui.button('Přesto zadat', on_click=lambda: potvrzeni_dialog.submit(True)).props('no-caps').classes('bg-blue-600 text-white')
                                        if not await potvrzeni_dialog:
                                            return

                                    uspech = await asyncio.to_thread(intranet_data.pridej_zadost_volno, cilovy_id, date_od.value, date_do.value, _h_val, _m_val, typ_select.value, _cas_od, _cas_do)
                                    if uspech:
                                        if cilovy_id == user_id:
                                            ui.notify('Žádost odeslána ke schválení.', type='positive', position='top')
                                            intranet_logger.log_activity(user_name, "Podání žádosti", f"Typ: {typy_volna.get(typ_select.value, typ_select.value)} | Od: {date_od.value} | Do: {date_do.value}{_cas_str_z} | Délka: {_dur_str_z}")
                                        else:
                                            ui.notify('Žádost za podřízeného vytvořena. Najdete ji níže ve schvalování.', type='positive', position='top')
                                            cil_jmeno = next((u['jmeno_cele'] for u in vsichni_uzivatele_komplet.values() if u['id'] == cilovy_id), str(cilovy_id))
                                            intranet_logger.log_activity(user_name, "Podání žádosti", f"Za: {cil_jmeno} | Typ: {typy_volna.get(typ_select.value, typ_select.value)} | Od: {date_od.value} | Do: {date_do.value}{_cas_str_z} | Délka: {_dur_str_z}")
                                        intranet_data.invaliduj_cache_dochazky()
                                        await asyncio.gather(
                                            asyncio.to_thread(intranet_data.ziskej_vsechny_uzivatele),
                                            asyncio.to_thread(intranet_data.ziskej_zadosti, None),
                                        )
                                        ui.timer(0, _rf.get('fn', vykresli_dochazku.refresh), once=True)

                                        try:
                                            cil_data = next((u for u in vsichni_uzivatele_komplet.values() if u['id'] == cilovy_id), None)
                                            if cil_data:
                                                oddeleni_cil = [o.strip() for o in (cil_data.get('oddeleni') or '').split(',')]
                                                prijemci_k_odeslani = _prijemci_zadosti(cilovy_id, oddeleni_cil)

                                                if prijemci_k_odeslani:
                                                    typ_nazev = typy_volna.get(typ_select.value, 'volno')
                                                    predmet = f"Nový záznam: {cil_data['jmeno_cele']}"
                                                    _app_url = intranet_data.nacti_nastaveni_intranetu().get('app_url', '').strip()
                                                    _proklik = f"\n\n➡ Přejít do aplikace: {_app_url}" if _app_url else ""
                                                    text = f"Dobrý den,\n\n{cil_data['jmeno_cele']} přidal záznam: {typ_nazev} v termínu od {date_od.value} do {date_do.value}{_cas_str_z} ({_dur_str_z}).\n\nProsím, přihlaste se do Moje JIPka k náhledu na záznam.{_proklik}\n\nToto je automatická zpráva z portálu Moje JIPka"
                                                    for p_email in prijemci_k_odeslani:
                                                        asyncio.create_task(asyncio.to_thread(intranet_emaily.odesli_upozorneni_email, p_email, predmet, text))
                                                        p_data = vsichni_uzivatele_komplet.get(p_email)
                                                        if p_data:
                                                            intranet_notifikace.pridej(
                                                                p_data['id'],
                                                                f'{cil_data["jmeno_cele"]} podal/a žádost o {typ_nazev} ({date_od.value} – {date_do.value}).',
                                                                'info'
                                                            )
                                        except Exception as e:
                                            print("Chyba při přípravě e-mailu (odeslat_záznam):", e)
                                    else: ui.notify('Chyba při ukládání záznamu.', type='negative', position='top')

                                ui.button('Odeslat záznam', icon='send', on_click=odeslat_zadost) \
                                    .props('unelevated').classes('bg-blue-600 hover:bg-blue-700 text-white font-bold w-full h-11 rounded-lg tracking-wider')

                    if muze_zadat_za_sebe:
                        with ui.card().classes('flex-[2] p-5 shadow-sm bg-white rounded-2xl border border-gray-200'):
                            ui.label('Moje záznamy').classes('text-xs font-bold text-gray-400 uppercase tracking-widest mb-3')
                            _flt = {'stav': None, 'rok': datetime.date.today().year}
                            @ui.refreshable
                            def vykresli_moje_seznam():
                                moje_z_s = intranet_data.ziskej_zadosti(user_id)
                                moje_ot = intranet_data.ziskej_presczasy(user_id) if presczasy_zapnuty else []
                                polozky = []
                                for z in moje_z_s:
                                    z['_druh'] = 'absence'
                                    polozky.append(z)
                                for ot in moje_ot:
                                    ot['_druh'] = 'presczas'
                                    ot['from'] = ot['datum_od']
                                    ot['to'] = ot['datum_do']
                                    polozky.append(ot)
                                polozky.sort(key=lambda x: _tl_datum(x) or datetime.date.min, reverse=True)
                                _tl_last = None

                                def _prepni_filtr(stav):
                                    _flt['stav'] = stav
                                    vykresli_moje_seznam.refresh()
                                def _prepni_rok(rok):
                                    _flt['rok'] = rok
                                    vykresli_moje_seznam.refresh()
                                _roky = sorted({_d.year for _d in (_tl_datum(z) for z in polozky) if _d}, reverse=True)
                                if _flt['rok'] not in _roky:
                                    _flt['rok'] = _roky[0] if _roky else None
                                if _flt['rok'] is not None:
                                    polozky = [z for z in polozky if (_tl_datum(z) or datetime.date.min).year == _flt['rok']]
                                with ui.row().classes('w-full items-center justify-between gap-2 mb-4'):
                                    with ui.row().classes('gap-2 flex-wrap'):
                                        for _ch_txt, _ch_stav in (('Vše', None), ('Čeká', 'Čeká na schválení'), ('Schváleno', 'Schváleno'),
                                                                  ('Zamítnuto', 'Zamítnuto'), ('Stornováno', 'Stornováno')):
                                            _ch_on = _flt['stav'] == _ch_stav
                                            ui.button(_ch_txt, on_click=lambda _, s=_ch_stav: _prepni_filtr(s)) \
                                                .props('unelevated dense no-caps rounded '
                                                       + ('color=green-6 text-color=white' if _ch_on else 'color=white text-color=grey-8')) \
                                                .classes(f'px-3 py-0 text-xs rounded-full border {"border-green-600 font-bold" if _ch_on else "border-gray-300"}')
                                    if _roky:
                                        ui.select({**{r: str(r) for r in _roky}, None: 'Vše'}, value=_flt['rok'],
                                                  on_change=lambda e: _prepni_rok(e.value)) \
                                            .props('dense borderless options-dense') \
                                            .classes('text-xs font-bold text-gray-500 min-w-[70px]')
                                if _flt['stav'] is not None:
                                    polozky = [z for z in polozky if z['stav'] == _flt['stav']]
                                    _tl_last = None

                                if not polozky:
                                    ui.label('Žádné záznamy v tomto filtru.' if _roky else 'Zatím nemáte žádné záznamy.').classes('text-gray-500 italic')
                                else:
                                    for z in polozky:
                                        _d = _tl_datum(z)
                                        if _d and (_d.year, _d.month) != _tl_last:
                                            _tl_last = (_d.year, _d.month)
                                            _tl_mesic(_d)
                                        if z['_druh'] == 'presczas':
                                            barva_ot = 'bg-gray-100 text-gray-600 border-gray-300' if z['stav_id'] == 4 else 'bg-green-100 text-green-700 border-green-300'
                                            ikona_ot = '🛑' if z['stav_id'] == 4 else '✅'
                                            cas_od_str = str(z['cas_od'])[:5] if z['cas_od'] else ''
                                            cas_do_str = str(z['cas_do'])[:5] if z['cas_do'] else ''
                                            with _tl_karta(_d, 'Stornováno' if z['stav_id'] == 4 else 'Schváleno'):
                                                with ui.column().classes('flex-1 gap-0'):
                                                    with ui.row().classes('items-center gap-2'):
                                                        ui.label('⏰ Přesčas').classes('font-bold text-orange-700 text-sm')
                                                        _zh, _zm = int(z.get('suma_hodin') or 0), int(z.get('suma_minut') or 0)
                                                        _zdur = f"{_zh}h {_zm}min" if _zh and _zm else (f"{_zh}h" if _zh else f"{_zm}min")
                                                        ui.label(_zdur).classes('text-xs bg-orange-100 text-orange-700 px-2 py-0.5 rounded-full font-bold')
                                                    ui.label(f"{formatuj_datum(z['datum_od'])} {cas_od_str} – {formatuj_datum(z['datum_do'])} {cas_do_str}").classes('text-xs text-gray-500')
                                                with ui.row().classes('items-center gap-1'):
                                                    ui.label(f"{ikona_ot} {z['stav']}").classes(f'px-2 py-0.5 rounded-full text-xs font-bold border {barva_ot}')
                                                    if z.get('duvod'):
                                                        with ui.button(icon='info', on_click=lambda: None).props('flat round dense size=xs color=orange'):
                                                            with ui.tooltip().classes('bg-white text-gray-800 shadow-xl border border-gray-200 p-3 rounded-xl max-w-xs'):
                                                                ui.label('Důvod přesčasu:').classes('font-bold text-xs text-orange-700 mb-1')
                                                                ui.label(z['duvod']).classes('text-xs')
                                                                if z.get('storno_reason'):
                                                                    ui.label('Důvod storna:').classes('font-bold text-xs text-red-600 mt-2 mb-1')
                                                                    ui.label(z['storno_reason']).classes('text-xs text-red-600')
                                                    if z['stav_id'] != 4 and (povoleny_schvalovat_oddeleni or jsem_neci_manazer or is_global_admin):
                                                        ui.button(icon='block', on_click=lambda pid=z['idovertimeRequest']: _stornovat_presczas_btn(pid)).props('flat color=orange dense size=xs').tooltip('Stornovat přesčas')
                                        else:
                                            barva_ab = 'bg-yellow-100 text-yellow-800 border-yellow-300' if z['stav'] == 'Čeká na schválení' else ('bg-green-100 text-green-800 border-green-300' if z['stav'] == 'Schváleno' else ('bg-gray-100 text-gray-800 border-gray-300' if z['stav'] == 'Stornováno' else 'bg-red-100 text-red-800 border-red-300'))
                                            ikona_ab = '⏳' if z['stav'] == 'Čeká na schválení' else ('✅' if z['stav'] == 'Schváleno' else ('🛑' if z['stav'] == 'Stornováno' else '❌'))
                                            with _tl_karta(_d, z['stav'], bool(z.get('storno_req_at'))):
                                                with ui.column().classes('gap-1'):
                                                    _ab_h, _ab_m = int(z.get('suma_hodin') or 0), int(z.get('suma_minut') or 0)
                                                    _ab_dur = f"{_ab_h}h {_ab_m}min" if _ab_h and _ab_m else (f"{_ab_h}h" if _ab_h else f"{_ab_m}min")
                                                    ui.label(f"{z['typ']} ({_ab_dur})").classes('font-bold text-gray-800')
                                                    _ab_cas_od = str(z['cas_od'])[:5] if z.get('cas_od') else ''
                                                    _ab_cas_do = str(z['cas_do'])[:5] if z.get('cas_do') else ''
                                                    _ab_cas_str = f"  {_ab_cas_od}–{_ab_cas_do}" if _ab_cas_od and _ab_cas_do else ""
                                                    ui.label(f"{formatuj_datum(z['from'])} do {formatuj_datum(z['to'])}{_ab_cas_str}").classes('text-sm text-gray-600')
                                                with ui.column().classes('items-end gap-1'):
                                                    with ui.row().classes('items-center gap-2'):
                                                        ui.label(f"{ikona_ab} {z['stav']}").classes(f'px-3 py-1 rounded-full text-sm font-bold border {barva_ab}')

                                                        zadatel_data_s = next((u for u in vsichni_uzivatele_komplet.values() if u['id'] == z['user_iduser']), {})
                                                        is_manager_zadatele = user_id in zadatel_data_s.get('manager_id', [])
                                                        is_watched_by_me = z['user_iduser'] in moji_sledovani
                                                        ma_pravo_na_oddeleni = (z.get('oddeleni') and any(r.strip() in povoleny_schvalovat_oddeleni for r in z.get('oddeleni').split(',')))
                                                        muze_mazat = (z['stav'] == 'Čeká na schválení' and z['user_iduser'] == user_id) or ma_pristup_mazani or is_global_admin or ma_pravo_na_oddeleni or is_manager_zadatele or is_watched_by_me

                                                        if muze_mazat:
                                                            if z['stav'] == 'Čeká na schválení':
                                                                ui.button(icon='delete', on_click=lambda zid=z['idleaveRequest']: smazat_zadost_btn(zid)).props('flat color=red padding=none size=sm').tooltip('Smazat žádost')
                                                            elif z['stav'] in ('Schváleno', 'Zamítnuto'):
                                                                if z['stav'] == 'Schváleno':
                                                                    ui.button(icon='edit', on_click=lambda zid=z['idleaveRequest'], zd=z: upravit_zadost_btn(zid, zd)).props('flat color=blue padding=none size=sm').tooltip('Upravit absenci')
                                                                ui.button(icon='block', on_click=lambda zid=z['idleaveRequest'], zd=z: stornovat_zadost_btn(zid, zd)).props('flat color=orange padding=none size=sm').tooltip('Stornovat žádost')

                                                        if (z['stav'] == 'Schváleno' and z['user_iduser'] == user_id
                                                                and not z.get('storno_req_at') and z['from'] > datetime.date.today()):
                                                            ui.button(icon='undo', on_click=lambda zid=z['idleaveRequest'], zd=z: pozadat_storno_btn(zid, zd)).props('flat color=orange padding=none size=sm').tooltip('Požádat o storno')

                                                    if z['stav_id'] != 1 and z['a_jmeno']:
                                                        cas = formatuj_cas(z['approved_at'])
                                                        ui.label(f"Vyřídil: {z['a_jmeno']} {z['a_prijmeni']} ({cas})").classes('text-xs text-gray-500 font-medium')
                                                    if z['stav_id'] in (3, 4) and z['rejection_reason']:
                                                        ui.label(f"Důvod: {z['rejection_reason']}").classes('text-xs text-red-600 italic')
                                                    if z.get('storno_req_at'):
                                                        ui.label(f"⏳ Žádost o storno odeslána: {z.get('storno_req_reason') or ''}").classes('text-xs text-orange-600 font-medium')
                            vykresli_moje_seznam()
                            _sub_refreshes.append(vykresli_moje_seznam.refresh)

                    def _smazat_presczas_btn(pid):
                        with ui.dialog() as dlg_p, ui.card().classes('p-6 rounded-xl w-full max-w-sm'):
                            ui.label('Smazat přesčas?').classes('text-xl font-bold text-red-600 mb-4')
                            ui.label('Záznam přesčasu bude nenávratně odstraněn.').classes('text-gray-600 mb-6')
                            async def _potvrdit_smazat():
                                ok = await asyncio.to_thread(intranet_data.smaz_presczas, pid)
                                if ok:
                                    intranet_logger.log_activity(user_name, "Přesčas", f"Smazán přesčas ID: {pid}")
                                    ui.notify('Záznam smazán.', type='info', position='top')
                                    dlg_p.close()
                                    intranet_data.invaliduj_cache_dochazky()
                                    await asyncio.to_thread(intranet_data.ziskej_presczasy, None)
                                    ui.timer(0, _rf.get('fn', vykresli_dochazku.refresh), once=True)
                                else:
                                    ui.notify('Chyba.', type='negative', position='top')
                            with ui.row().classes('w-full justify-between'):
                                ui.button('Zrušit', on_click=dlg_p.close).classes('bg-gray-400 text-white font-bold')
                                ui.button('Smazat', on_click=_potvrdit_smazat).classes('bg-red-600 text-white font-bold shadow-md')
                        dlg_p.open()

                if povoleny_schvalovat_oddeleni or jsem_neci_manazer or jsem_majitel:
                    ui.label('Záznamy ke schválení').classes('text-xs font-bold text-orange-500 uppercase tracking-widest mt-6 mb-3')

                    def _make_schvalit(zid, z_data):
                        async def handler():
                            uspech = await asyncio.to_thread(intranet_data.zmen_stav_zadosti, zid, 2, user_id)
                            if not uspech:
                                ui.notify('Chyba při schvalování. Zkontrolujte DB připojení.', type='negative', position='top')
                                return
                            intranet_logger.log_activity(user_name, "Schválení žádosti", f"ID: {zid} | {z_data.get('typ', '')} | {formatuj_datum(z_data['from'])} – {formatuj_datum(z_data['to'])}")
                            ui.notify('Reakce provedena.', type='positive', position='top')
                            intranet_data.invaliduj_cache_dochazky()
                            await asyncio.gather(
                                asyncio.to_thread(intranet_data.ziskej_vsechny_uzivatele),
                                asyncio.to_thread(intranet_data.ziskej_zadosti, None),
                            )
                            ui.timer(0, _rf.get('fn', vykresli_dochazku.refresh), once=True)
                            try:
                                zadatel_email, zadatel_data = None, None
                                for mail, u in vsichni_uzivatele_komplet.items():
                                    if u['id'] == z_data['user_iduser']:
                                        zadatel_email, zadatel_data = mail, u
                                        break
                                if zadatel_email and zadatel_data.get('email_vyrizeni_zadosti', True):
                                    predmet = "Na Váš záznam byla učiněna reakce"
                                    _app_url = intranet_data.nacti_nastaveni_intranetu().get('app_url', '').strip()
                                    _proklik = f"\n\n➡ Přejít do aplikace: {_app_url}" if _app_url else ""
                                    text = f"Dobrý den {zadatel_data['jmeno']},\n\nNa Váš záznam \"{z_data['typ']}\" (v termínu {formatuj_datum(z_data['from'])} až {formatuj_datum(z_data['to'])}) byla učiněna reakce.{_proklik}\n\nToto je automatická zpráva z portálu Moje JIPka"
                                    asyncio.create_task(asyncio.to_thread(intranet_emaily.odesli_upozorneni_email, zadatel_email, predmet, text))
                                intranet_notifikace.pridej(
                                    z_data['user_iduser'],
                                    f'Vaše žádost o {z_data["typ"]} ({formatuj_datum(z_data["from"])} – {formatuj_datum(z_data["to"])}) byla schválena.',
                                    'success'
                                )
                            except Exception as e:
                                print("Chyba při přípravě e-mailu (schvalit):", e)
                        return handler

                    def _make_zamitnout(zid, z_data):
                        def handler():
                            with ui.dialog() as dlg, ui.card().classes('p-6 rounded-xl w-full max-w-md'):
                                ui.label('Zamítnutí žádosti').classes('text-xl font-bold mb-4 text-red-600')
                                duvod_input = ui.textarea('Důvod zamítnutí (volitelný)').classes('w-full mb-4')

                                async def potvrdit():
                                    duvod_val = duvod_input.value or ""
                                    uspech = await asyncio.to_thread(intranet_data.zmen_stav_zadosti, zid, 3, user_id, duvod_val)
                                    if not uspech:
                                        ui.notify('Chyba při zamítání. Zkontrolujte DB připojení.', type='negative', position='top')
                                        return
                                    duvod_txt = f" | Důvod: {duvod_val.strip()}" if duvod_val.strip() else ""
                                    intranet_logger.log_activity(user_name, "Zamítnutí žádosti", f"ID: {zid} | {z_data.get('typ', '')} | {formatuj_datum(z_data['from'])} – {formatuj_datum(z_data['to'])}{duvod_txt}")
                                    ui.notify('Žádost zamítnuta.', type='negative', position='top')
                                    dlg.close()
                                    intranet_data.invaliduj_cache_dochazky()
                                    await asyncio.gather(
                                        asyncio.to_thread(intranet_data.ziskej_vsechny_uzivatele),
                                        asyncio.to_thread(intranet_data.ziskej_zadosti, None),
                                    )
                                    ui.timer(0, _rf.get('fn', vykresli_dochazku.refresh), once=True)
                                    try:
                                        zadatel_email, zadatel_data = None, None
                                        for mail, u in vsichni_uzivatele_komplet.items():
                                            if u['id'] == z_data['user_iduser']:
                                                zadatel_email, zadatel_data = mail, u
                                                break
                                        if zadatel_email and zadatel_data.get('email_vyrizeni_zadosti', True):
                                            predmet = "Na Váš záznam byla učiněna reakce"
                                            _app_url = intranet_data.nacti_nastaveni_intranetu().get('app_url', '').strip()
                                            _proklik = f"\n\n➡ Přejít do aplikace: {_app_url}" if _app_url else ""
                                            text = f"Dobrý den {zadatel_data['jmeno']},\n\nNa Váš záznam \"{z_data['typ']}\" (v termínu {formatuj_datum(z_data['from'])} až {formatuj_datum(z_data['to'])}) byla učiněna reakce."
                                            if duvod_val.strip():
                                                text += f"\n\nDůvod zamítnutí: {duvod_val.strip()}"
                                            text += f"{_proklik}\n\nToto je automatická zpráva z portálu Moje JIPka"
                                            asyncio.create_task(asyncio.to_thread(intranet_emaily.odesli_upozorneni_email, zadatel_email, predmet, text))
                                        intranet_notifikace.pridej(
                                            z_data['user_iduser'],
                                            f'Vaše žádost o {z_data["typ"]} ({formatuj_datum(z_data["from"])} – {formatuj_datum(z_data["to"])}) byla zamítnuta.'
                                            + (f' Důvod: {duvod_val.strip()}' if duvod_val.strip() else ''),
                                            'error'
                                        )
                                    except Exception as e:
                                        print("Chyba při přípravě e-mailu (zamitnout):", e)

                                with ui.row().classes('w-full justify-between'):
                                    ui.button('Zpět', on_click=dlg.close).classes('bg-gray-400')
                                    ui.button('Zamítnout', on_click=potvrdit).classes('bg-red-600 text-white font-bold')
                            dlg.open()
                        return handler

                    @ui.refreshable
                    def vykresli_ceka_admin():
                        vsechny_zc = intranet_data.ziskej_zadosti(None)
                        # Čerstvá uživatelská data (reálný zůstatek dovolené) pro tooltip u jména žadatele.
                        vsichni_uzivatele_komplet = intranet_data.ziskej_vsechny_uzivatele()
                        ceka_zc = [z for z in vsechny_zc if z['stav_id'] == 1]

                        def _zustatek_dovolene_tip(uid):
                            zd = next((u for u in vsichni_uzivatele_komplet.values() if u['id'] == uid), {})
                            rz = zd.get('realny_zustatek_dovolene')
                            if rz is None:
                                return 'Reálný zůstatek dovolené: nezadáno'
                            rz_dat = zd.get('realny_zustatek_dovolene_datum')
                            if rz_dat:
                                try:
                                    rz_dat = datetime.date.fromisoformat(rz_dat).strftime('%d.%m.%Y')
                                except (ValueError, TypeError):
                                    pass
                                return f'Zůstatek dovolené k {rz_dat}: {rz:g} h'
                            return f'Reálný zůstatek dovolené: {rz:g} h'
                        ma_vyjimku_sebe = 'vse' in vsechna_prava or 'dochazka_schvalovat_sebe' in vsechna_prava
                        ceka_admin = []
                        for z in ceka_zc:
                            if 'admin' in z.get('u_jmeno', '').lower() or 'admin' in z.get('u_prijmeni', '').lower(): continue
                            zd = next((u for u in vsichni_uzivatele_komplet.values() if u['id'] == z.get('user_iduser')), {})
                            is_mgr = user_id in zd.get('manager_id', [])
                            is_wtch = z.get('user_iduser') in moji_sledovani
                            ma_odd = (z.get('oddeleni') and any(r.strip() in povoleny_schvalovat_oddeleni for r in z.get('oddeleni').split(',')))
                            if (is_global_admin or ma_odd or is_mgr or is_wtch):
                                cele_jmeno = f"{z['u_jmeno']} {z['u_prijmeni']}"
                                if cele_jmeno != user_name or ma_vyjimku_sebe:
                                    ceka_admin.append(z)

                        if not ceka_admin:
                            ui.label('Žádné nové žádosti ke schválení.').classes('text-gray-500 italic mb-8')
                        else:
                            with ui.card().classes('w-full p-3 shadow-sm bg-white rounded-xl border border-orange-200 mb-8 overflow-y-auto').style('max-height: 760px'):
                                for z in ceka_admin:
                                    with _tl_karta(_tl_datum(z), 'Čeká na schválení'):
                                        with ui.column().classes('gap-1'):
                                            ui.label(f"{z['u_jmeno']} {z['u_prijmeni']}").classes('font-black text-gray-800 text-lg cursor-help').tooltip(_zustatek_dovolene_tip(z.get('user_iduser')))
                                            _ceka_h, _ceka_m = int(z.get('suma_hodin') or 0), int(z.get('suma_minut') or 0)
                                            _ceka_dur = f"{_ceka_h}h {_ceka_m}min" if _ceka_h and _ceka_m else (f"{_ceka_h}h" if _ceka_h else f"{_ceka_m}min")
                                            ui.label(f"{z['typ']} | Celkem: {_ceka_dur}").classes('font-bold text-gray-700')
                                            _ceka_cas_od = str(z['cas_od'])[:5] if z.get('cas_od') else ''
                                            _ceka_cas_do = str(z['cas_do'])[:5] if z.get('cas_do') else ''
                                            _ceka_cas = f"  {_ceka_cas_od}–{_ceka_cas_do}" if _ceka_cas_od and _ceka_cas_do else ""
                                            ui.label(f"Termín: {formatuj_datum(z['from'])} do {formatuj_datum(z['to'])}{_ceka_cas} (Podáno: {z['created_at'].strftime('%d.%m.%Y %H:%M')})").classes('text-sm text-gray-600')
                                        with ui.row().classes('gap-2 items-center'):
                                            ui.button('✅ SCHVÁLIT', color='green', on_click=_make_schvalit(z['idleaveRequest'], z)).classes('font-bold px-6 py-2 shadow-md')
                                            ui.button('❌ ZAMÍTNOUT', color='red', on_click=_make_zamitnout(z['idleaveRequest'], z)).classes('font-bold px-6 py-2 shadow-md')
                                            ui.button(icon='delete', on_click=lambda zid=z['idleaveRequest']: smazat_zadost_btn(zid)).props('flat color=red padding=none size=md').tooltip('Smazat žádost').classes('ml-4')
                    vykresli_ceka_admin()
                    _sub_refreshes.append(vykresli_ceka_admin.refresh)

                    ui.label('Historie schvalování').classes('text-2xl font-bold mb-4 text-blue-600')

                    @ui.refreshable
                    def vykresli_historie_sekce():
                        vsechny_zh = intranet_data.ziskej_zadosti(None)
                        vyrizene_zh = [z for z in vsechny_zh if z['stav_id'] != 1]
                        vsechny_ot = intranet_data.ziskej_presczasy(None) if presczasy_zapnuty else []

                        historie_admin = []
                        for z in vyrizene_zh:
                            if 'admin' in z.get('u_jmeno', '').lower() or 'admin' in z.get('u_prijmeni', '').lower(): continue
                            zd_h = next((u for u in vsichni_uzivatele_komplet.values() if u['id'] == z.get('user_iduser')), {})
                            is_mgr_h = user_id in zd_h.get('manager_id', [])
                            is_wtch_h = z.get('user_iduser') in moji_sledovani
                            ma_odd_h = (z.get('oddeleni') and any(r.strip() in povoleny_schvalovat_oddeleni for r in z.get('oddeleni').split(',')))
                            if is_global_admin or ma_odd_h or is_mgr_h or is_wtch_h:
                                z['_typ_zaznamu'] = 'volno'
                                historie_admin.append(z)

                        # Přidej přesčasy do historie (vždy schválené nebo stornované)
                        if presczasy_zapnuty:
                            for ot in vsechny_ot:
                                if 'admin' in ot.get('u_jmeno', '').lower() or 'admin' in ot.get('u_prijmeni', '').lower(): continue
                                zd_ot = next((u for u in vsichni_uzivatele_komplet.values() if u['id'] == ot.get('user_iduser')), {})
                                is_mgr_ot = user_id in zd_ot.get('manager_id', [])
                                is_wtch_ot = ot.get('user_iduser') in moji_sledovani
                                ma_odd_ot = (ot.get('oddeleni') and any(r.strip() in povoleny_schvalovat_oddeleni for r in ot.get('oddeleni').split(',')))
                                if is_global_admin or ma_odd_ot or is_mgr_ot or is_wtch_ot:
                                    ot['_typ_zaznamu'] = 'presczas'
                                    ot['from'] = ot['datum_od']; ot['to'] = ot['datum_do']
                                    ot['typ'] = 'Přesčas'; ot['a_jmeno'] = None
                                    historie_admin.append(ot)

                        historie_admin.sort(key=lambda x: x.get('created_at') or datetime.datetime.min, reverse=True)

                        if not historie_admin:
                            ui.label('Zatím žádná historie.').classes('text-gray-500 italic text-lg mb-8')
                        else:
                            osoby_options = sorted(set(f"{z['u_jmeno']} {z['u_prijmeni']}" for z in historie_admin))
                            with ui.row().classes('w-full gap-4 mb-4 items-end flex-wrap bg-gray-50 p-4 rounded-xl border border-gray-200'):
                                with ui.input('Od data').classes('w-36 bg-white') as hist_od:
                                    with hist_od.add_slot('append'):
                                        ui.icon('edit_calendar').on('click', lambda: menu_hist_od.open()).classes('cursor-pointer text-gray-500 hover:text-blue-500')
                                    with ui.menu() as menu_hist_od:
                                        ui.date().bind_value(hist_od)
                                with ui.input('Do data').classes('w-36 bg-white') as hist_do:
                                    with hist_do.add_slot('append'):
                                        ui.icon('edit_calendar').on('click', lambda: menu_hist_do.open()).classes('cursor-pointer text-gray-500 hover:text-blue-500')
                                    with ui.menu() as menu_hist_do:
                                        ui.date().bind_value(hist_do)
                                hist_osoba = ui.select(
                                    options=['Všichni'] + osoby_options,
                                    value='Všichni',
                                    label='Osoba',
                                    with_input=True,
                                ).classes('w-56 bg-white')
                                def _filtruj_historie():
                                    hist_page['p'] = 1
                                    render_historie.refresh()
                                ui.button('Filtrovat', icon='filter_alt', on_click=_filtruj_historie).classes(
                                    'bg-blue-600 hover:bg-blue-700 text-white font-bold h-[3.5rem] px-5 shadow-sm')
                                def _zrusit_filtr_historie():
                                    hist_od.set_value('')
                                    hist_do.set_value('')
                                    hist_osoba.set_value('Všichni')
                                    hist_page['p'] = 1
                                    render_historie.refresh()
                                ui.button(icon='clear', on_click=_zrusit_filtr_historie).props('flat round').tooltip('Zrušit filtr').classes('self-center text-gray-500')

                            HIST_NA_STRANU = 10
                            hist_page = {'p': 1}

                            @ui.refreshable
                            def render_historie():
                                try:
                                    od_val = datetime.datetime.strptime(hist_od.value, '%Y-%m-%d').date() if hist_od.value else None
                                except Exception:
                                    od_val = None
                                try:
                                    do_val = datetime.datetime.strptime(hist_do.value, '%Y-%m-%d').date() if hist_do.value else None
                                except Exception:
                                    do_val = None
                                osoba_val = hist_osoba.value if hist_osoba.value != 'Všichni' else None
                                filtered = []
                                for z in historie_admin:
                                    if osoba_val and f"{z['u_jmeno']} {z['u_prijmeni']}" != osoba_val: continue
                                    if od_val and z['to'] < od_val: continue
                                    if do_val and z['from'] > do_val: continue
                                    filtered.append(z)
                                if not filtered:
                                    ui.label('Žádné záznamy neodpovídají filtru.').classes('text-gray-500 italic p-4')
                                    return
                                pocet_stran = max(1, -(-len(filtered) // HIST_NA_STRANU))
                                if hist_page['p'] > pocet_stran: hist_page['p'] = pocet_stran
                                _od = (hist_page['p'] - 1) * HIST_NA_STRANU
                                strana = filtered[_od:_od + HIST_NA_STRANU]
                                ui.label(f'Zobrazeno {_od + 1}–{_od + len(strana)} z {len(filtered)} záznamů').classes('text-xs text-gray-400 mb-2 px-1')
                                with ui.card().classes('w-full p-3 shadow-sm bg-white rounded-xl border border-blue-200 mb-2 justify-start').style('overflow: visible'):
                                    for z in strana:
                                        je_presczas = z.get('_typ_zaznamu') == 'presczas'
                                        barva = 'text-green-600' if z['stav'] == 'Schváleno' else ('text-gray-500' if z['stav'] == 'Stornováno' else 'text-red-600')
                                        ikona = '⏰' if je_presczas and z['stav'] == 'Schváleno' else ('✅' if z['stav'] == 'Schváleno' else ('🛑' if z['stav'] == 'Stornováno' else '❌'))
                                        with _tl_karta(_tl_datum(z), z['stav'], bool(z.get('storno_req_at')), 'bg-orange-50' if je_presczas else ''):
                                            with ui.column().classes('w-1/4 gap-0'):
                                                ui.label(f"{z['u_jmeno']} {z['u_prijmeni']}").classes('font-bold text-gray-800')
                                                if je_presczas:
                                                    cas_od_h = str(z['cas_od'])[:5] if z.get('cas_od') else ''
                                                    cas_do_h = str(z['cas_do'])[:5] if z.get('cas_do') else ''
                                                    _zh2, _zm2 = int(z.get('suma_hodin') or 0), int(z.get('suma_minut') or 0)
                                                    _zdur2 = f"{_zh2}h {_zm2}min" if _zh2 and _zm2 else (f"{_zh2}h" if _zh2 else f"{_zm2}min")
                                                    ui.label(f"⏰ Přesčas  {cas_od_h}–{cas_do_h}  ({_zdur2})").classes('text-xs text-orange-600 font-semibold')
                                                else:
                                                    _ab_h2, _ab_m2 = int(z.get('suma_hodin') or 0), int(z.get('suma_minut') or 0)
                                                    _ab_dur2 = f"{_ab_h2}h {_ab_m2}min" if _ab_h2 and _ab_m2 else (f"{_ab_h2}h" if _ab_h2 else f"{_ab_m2}min")
                                                    _ab_cas_od2 = str(z['cas_od'])[:5] if z.get('cas_od') else ''
                                                    _ab_cas_do2 = str(z['cas_do'])[:5] if z.get('cas_do') else ''
                                                    _ab_cas_str2 = f"  {_ab_cas_od2}–{_ab_cas_do2}" if _ab_cas_od2 and _ab_cas_do2 else ""
                                                    ui.label(f"{z['typ']}{_ab_cas_str2}  ({_ab_dur2})").classes('text-xs text-gray-500')
                                            ui.label(f"{formatuj_datum(z['from'])} do {formatuj_datum(z['to'])}").classes('w-1/4 text-gray-600 self-center')
                                            with ui.column().classes('w-1/3 items-end gap-0'):
                                                with ui.row().classes('items-center gap-2'):
                                                    ui.label(f"{ikona} {z['stav']}").classes(f'{barva} font-bold text-sm')
                                                    if je_presczas:
                                                        if z.get('duvod'):
                                                            with ui.button(icon='info').props('flat round dense size=xs color=orange'):
                                                                with ui.tooltip().classes('bg-white text-gray-800 shadow-xl border border-gray-200 p-3 rounded-xl max-w-xs'):
                                                                    ui.label('Důvod přesčasu:').classes('font-bold text-xs text-orange-700 mb-1')
                                                                    ui.label(z['duvod']).classes('text-xs')
                                                                    if z.get('storno_reason'):
                                                                        ui.label('Důvod storna:').classes('font-bold text-xs text-red-600 mt-2 mb-1')
                                                                        ui.label(z['storno_reason']).classes('text-xs text-red-600')
                                                        if z['stav_id'] != 4:
                                                            ui.button(icon='block', on_click=lambda pid=z['idovertimeRequest']: _stornovat_presczas_btn(pid)).props('flat color=orange padding=none size=sm').tooltip('Stornovat přesčas')
                                                    else:
                                                        if z['stav'] == 'Čeká na schválení':
                                                            ui.button(icon='delete', on_click=lambda zid=z['idleaveRequest']: smazat_zadost_btn(zid)).props('flat color=red padding=none size=sm').tooltip('Smazat žádost')
                                                        elif z['stav'] in ('Schváleno', 'Zamítnuto'):
                                                            if z['stav'] == 'Schváleno':
                                                                ui.button(icon='edit', on_click=lambda zid=z['idleaveRequest'], zd=z: upravit_zadost_btn(zid, zd)).props('flat color=blue padding=none size=sm').tooltip('Upravit absenci')
                                                            ui.button(icon='block', on_click=lambda zid=z['idleaveRequest'], zd=z: stornovat_zadost_btn(zid, zd)).props('flat color=orange padding=none size=sm').tooltip('Stornovat žádost')
                                                        if z.get('storno_req_at') and z['stav'] == 'Schváleno':
                                                            ui.button(icon='check', on_click=lambda zid=z['idleaveRequest'], zd=z: vyrid_storno_btn(zid, zd, True)).props('flat color=orange padding=none size=sm').tooltip('Schválit storno')
                                                            ui.button(icon='close', on_click=lambda zid=z['idleaveRequest'], zd=z: vyrid_storno_btn(zid, zd, False)).props('flat color=grey padding=none size=sm').tooltip('Zamítnout storno')
                                                if not je_presczas:
                                                    cas = formatuj_cas(z['approved_at'])
                                                    if z.get('a_jmeno'):
                                                        ui.label(f"Vyřídil: {z['a_jmeno']} {z['a_prijmeni']} ({cas})").classes('text-xs text-gray-500 font-medium')
                                                    if z['stav_id'] in (3, 4) and z['rejection_reason']:
                                                        ui.label(f"Důvod: {z['rejection_reason']}").classes('text-xs text-red-500 italic mt-1')
                                                    if z.get('storno_req_at'):
                                                        ui.label(f"⏳ Žádost o storno: {z.get('storno_req_reason') or ''}").classes('text-xs text-orange-600 font-medium')
                                                elif z.get('storno_at'):
                                                    ui.label(f"Stornováno: {formatuj_cas(z['storno_at'])}").classes('text-xs text-gray-500 font-medium')
                                    # doplň prázdné sloty, ať má každá strana výšku 10 záznamů
                                    for _ in range(HIST_NA_STRANU - len(strana)):
                                        ui.element('div').classes('w-full mb-2 shrink-0').style('height: 74px')
                                    if pocet_stran > 1:
                                        def _zmen_stranu(e):
                                            hist_page['p'] = int(e.value or 1)
                                            render_historie.refresh()
                                        with ui.row().classes('w-full justify-center items-center mt-auto pt-2 border-t border-gray-200'):
                                            ui.pagination(1, pocet_stran, value=hist_page['p'], direction_links=True,
                                                          on_change=_zmen_stranu).props('color=blue-6 max-pages=7 boundary-numbers')
                                ui.element('div').classes('mb-8')
                            render_historie()
                    vykresli_historie_sekce()
                    _sub_refreshes.append(vykresli_historie_sekce.refresh)

                if _sub_refreshes:
                    def _combined_refresh():
                        for fn in _sub_refreshes: fn()
                    _rf['fn'] = _combined_refresh

        # ==========================================================
        # TAB 2: ZÁZNAMY ODDĚLENÍ (Pouze čtení záznamů a filtrace)
        # ==========================================================
        if povoleny_sprava_oddeleni:
            with ui.tab_panel('sprava'):

                @ui.refreshable
                def sprava_oddeleni_ui():
                    # Čerstvá data při každém refreshi (stejně jako „Moje docházka"):
                    # po importu reálných zůstatků i po vyřízení/stornu absencí se hned projeví nové hodnoty.
                    vsichni_uzivatele_komplet = intranet_data.ziskej_vsechny_uzivatele()
                    vyrizene_zadosti_vse = [z for z in intranet_data.ziskej_zadosti(None) if z['stav_id'] != 1]
                    selected = app.storage.user.get('aktivni_slozka_dochazka')

                    def set_oddeleni(oddeleni):
                        app.storage.user['aktivni_slozka_dochazka'] = oddeleni
                        sprava_oddeleni_ui.refresh()

                    if not selected:
                        ui.label('Vyberte oddělení pro zobrazení záznamů').classes('text-2xl font-bold mb-6 text-gray-800')
                        with ui.row().classes('gap-8 flex-wrap items-stretch'):
                            for r_nazev in povoleny_sprava_oddeleni:
                                if r_nazev.lower() == 'admin': continue
                                with ui.card().classes('w-64 h-48 items-center justify-center shadow-md hover:scale-105 transition-transform duration-300 cursor-pointer bg-blue-50 rounded-2xl border border-blue-200').on('click', lambda r=r_nazev: set_oddeleni(r)):
                                    ui.label('🗂️').classes('text-5xl mb-4')
                                    ui.label(r_nazev).classes('text-xl font-bold text-blue-800 text-center')

                    else:
                        with ui.row().classes('w-full justify-between items-center mb-8'):
                            ui.label(f'Záznamy a čerpání oddělení: {selected}').classes('text-3xl font-extrabold text-blue-800')
                            ui.button('⬅️ Zpět na přehled oddělení', on_click=lambda: set_oddeleni(None)).classes('bg-gray-500 hover:bg-gray-600 text-white font-bold px-6 py-2 shadow-md')

                        current_year = datetime.date.today().year
                        ui.label(f'📊 Přehled čerpání dovolené').classes('text-2xl font-bold mb-4 text-blue-600')
                        with ui.card().classes('max-w-3xl p-0 shadow-sm bg-white rounded-xl border border-gray-200 mb-8'):
                            uzivatele_v_odd = []
                            for mail, udata in vsichni_uzivatele_komplet.items():
                                if 'admin' in udata.get('oddeleni', '').lower() or 'admin' in mail.lower(): continue
                                if selected in [o.strip() for o in (udata.get('oddeleni') or '').split(',')]:
                                    uzivatele_v_odd.append(udata['jmeno_cele'])

                            vybrano_dovolena = {jm: 0.0 for jm in uzivatele_v_odd}
                            for z in vyrizene_zadosti_vse:
                                if 'admin' in z.get('u_jmeno', '').lower() or 'admin' in z.get('u_prijmeni', '').lower(): continue
                                if z['stav_id'] == 2 and z['typ'] == 'Dovolená' and z['from'].year == current_year:
                                    odd_zaznamu = [o.strip() for o in (z.get('oddeleni') or '').split(',')]
                                    if selected in odd_zaznamu:
                                        cele_jmeno = f"{z['u_jmeno']} {z['u_prijmeni']}"
                                        if cele_jmeno in vybrano_dovolena:
                                            vybrano_dovolena[cele_jmeno] += float(z['sumaHours'])

                            rows_dovolena = []
                            _rz_datumy = set()
                            for jm in sorted(uzivatele_v_odd):
                                vybrano = vybrano_dovolena[jm]
                                udata = next(u for u in vsichni_uzivatele_komplet.values() if u['jmeno_cele'] == jm)
                                zaklad = udata.get('base_vacation', 160.0)
                                prevod = udata.get('carried_over_vacation', 0.0)
                                narok = zaklad + prevod
                                zbyva = narok - vybrano
                                realny_zustatek = udata.get('realny_zustatek_dovolene')
                                rz_datum = udata.get('realny_zustatek_dovolene_datum')
                                if rz_datum:
                                    _rz_datumy.add(rz_datum)
                                rows_dovolena.append({
                                    'Jméno': jm,
                                    'Základní nárok': f"{zaklad:g} h",
                                    'Z loňska': f"{prevod:g} h",
                                    'Celkový nárok': f"{narok:g} h",
                                    'Vybráno': f"{vybrano:g} h",
                                    'Zbývá': f"{zbyva:g} h",
                                    'Reálný zůstatek': f"{realny_zustatek:g} h" if realny_zustatek is not None else '—'
                                })

                            # Datum reálného zůstatku do hlavičky sloupce (import nastavuje všem stejné datum)
                            _rz_label = 'Reálný zůstatek dovolené'
                            if _rz_datumy:
                                _rz_datum_h = max(_rz_datumy)
                                try:
                                    _rz_label = f"Zůstatek dovolené k {datetime.date.fromisoformat(_rz_datum_h).strftime('%d.%m.%Y')}"
                                except (ValueError, TypeError):
                                    _rz_label = f"Zůstatek dovolené k {_rz_datum_h}"

                            if rows_dovolena:
                                cols_dov = [
                                    {'name': 'Jméno', 'label': 'Zaměstnanec', 'field': 'Jméno', 'align': 'left', 'classes': 'font-bold'},
                                    {'name': 'Vybráno', 'label': 'Schváleno / Vybráno', 'field': 'Vybráno', 'align': 'right'},
                                    {'name': 'Zbývá', 'label': f'Předpokládaný zůstatek do konce {current_year}', 'field': 'Zbývá', 'align': 'right', 'classes': 'font-bold text-blue-600'},
                                    {'name': 'Reálný zůstatek', 'label': _rz_label, 'field': 'Reálný zůstatek', 'align': 'right', 'classes': 'font-bold text-emerald-700'}
                                ]
                                ui.table(columns=cols_dov, rows=rows_dovolena).props('hide-bottom :pagination="{rowsPerPage: 0}"').classes('max-w-3xl shadow-none text-base')
                            else:
                                ui.label('V tomto oddělení zatím nejsou zařazeni žádní uživatelé.').classes('p-4 text-gray-500 italic text-base')

                        ui.label('Detailní výpis všech schválených absencí').classes('text-2xl font-bold mb-4 text-gray-800 border-t pt-8')
                        with ui.row().classes('w-full gap-4 mb-4 items-end bg-gray-50 p-4 rounded-xl border border-gray-200'):
                            with ui.input('Od data (RRRR-MM-DD)').classes('w-40 bg-white') as table_od:
                                with table_od.add_slot('append'):
                                    ui.icon('edit_calendar').on('click', lambda: t_menu_od.open()).classes('cursor-pointer text-gray-500 hover:text-blue-500')
                                with ui.menu() as t_menu_od: ui.date().bind_value(table_od)

                            with ui.input('Do data (RRRR-MM-DD)').classes('w-40 bg-white') as table_do:
                                with table_do.add_slot('append'):
                                    ui.icon('edit_calendar').on('click', lambda: t_menu_do.open()).classes('cursor-pointer text-gray-500 hover:text-blue-500')
                                with ui.menu() as t_menu_do: ui.date().bind_value(table_do)

                            dnes = datetime.date.today()
                            table_od.value = f"{dnes.year}-01-01"
                            table_do.value = f"{dnes.year}-12-31"

                            ui.button('Filtrovat data', on_click=lambda: render_dept_table.refresh()).classes('bg-blue-600 hover:bg-blue-700 text-white font-bold h-[3.5rem] px-6 ml-4 shadow-sm')

                        @ui.refreshable
                        def render_dept_table():
                            try:
                                od_val = datetime.datetime.strptime(table_od.value, '%Y-%m-%d').date() if table_od.value else datetime.date.min
                                do_val = datetime.datetime.strptime(table_do.value, '%Y-%m-%d').date() if table_do.value else datetime.date.max
                            except Exception:
                                ui.label('Chybný formát data!').classes('text-red-500 font-bold mt-4')
                                return

                            vsechna_volna = intranet_data.ziskej_vsechna_volna_kalendar(False)
                            vyfiltrovana = []

                            for d in vsechna_volna:
                                if 'admin' in d.get('u_jmeno', '').lower() or 'admin' in d.get('u_prijmeni', '').lower(): continue
                                if d['stav_id'] not in (2, 4): continue

                                odd_zaznamu = d.get('oddeleni') or 'Bez oddělení'
                                if selected not in [r.strip() for r in odd_zaznamu.split(',')]: continue

                                if not (d['to'] >= od_val and d['from'] <= do_val): continue

                                cele_jmeno = f"{d['u_jmeno']} {d['u_prijmeni']}"
                                osobni_cislo = next((u['id'] for u in vsichni_uzivatele_komplet.values() if u['jmeno_cele'] == cele_jmeno), "")

                                vyfiltrovana.append({
                                    'Osobní číslo': osobni_cislo,
                                    'Jméno a příjmení': cele_jmeno,
                                    'Od': d['from'].strftime('%d.%m.%Y'),
                                    'Do': d['to'].strftime('%d.%m.%Y'),
                                    'Typ volna': "dov." if d['typ'] == 'Dovolená' else d['typ'],
                                    'Stav': d['stav_nazev'],
                                    'Hodiny': float(d['sumaHours'])
                                })

                            if not vyfiltrovana:
                                ui.label('Žádné schválené ani stornované záznamy pro zadané období.').classes('text-gray-500 italic text-xl mt-4')
                                return

                            df = pd.DataFrame(vyfiltrovana)
                            df = df.sort_values(by=['Jméno a příjmení', 'Od'])

                            cols = [
                                {'name': 'Osobní číslo', 'label': 'Osobní číslo', 'field': 'Osobní číslo', 'align': 'left'},
                                {'name': 'Jméno a příjmení', 'label': 'Zaměstnanec', 'field': 'Jméno a příjmení', 'align': 'left'},
                                {'name': 'Od', 'label': 'Od data', 'field': 'Od', 'align': 'left'},
                                {'name': 'Do', 'label': 'Do data', 'field': 'Do', 'align': 'left'},
                                {'name': 'Typ volna', 'label': 'Typ', 'field': 'Typ volna', 'align': 'left'},
                                {'name': 'Stav', 'label': 'Stav', 'field': 'Stav', 'align': 'left'},
                                {'name': 'Hodiny', 'label': 'Zadané hodiny', 'field': 'Hodiny', 'align': 'right'}
                            ]
                            ui.table(columns=cols, rows=df.to_dict('records')).props('hide-bottom :pagination="{rowsPerPage: 0}"').classes('w-full mt-4 shadow-sm text-base')

                        render_dept_table()

                sprava_oddeleni_ui()

# ==========================================
# --- SPRÁVA UŽIVATELŮ A ODDĚLENÍ ---
# ==========================================
# ── Export matice práv (jen superadmin, šifrovaný ZIP) ─────────────
# Matice = kdo má které právo a kterým kanálem. Písmena v buňce:
# P = přímo, R = přes pracovní pozici (roli), O = přes oddělení.
_MATICE_TMP_DIR = os.path.join(tempfile.gettempdir(), 'jip_prava')
_MATICE_MAX_BUNEK = 400_000

def _matice_tmp_cesta(jmeno: str) -> str:
    """Cesta v temp adresáři; při každém exportu smaže staré soubory."""
    os.makedirs(_MATICE_TMP_DIR, exist_ok=True)
    ted = time.time()
    for f in os.listdir(_MATICE_TMP_DIR):
        stara = os.path.join(_MATICE_TMP_DIR, f)
        try:
            if os.path.isfile(stara) and ted - os.path.getmtime(stara) > 1800:
                os.remove(stara)
        except OSError:
            pass
    return os.path.join(_MATICE_TMP_DIR, f'{int(ted * 1000)}_{jmeno}')

def _matice_zip_sync(matice: dict, katalog: dict, kategorie: list,
                     odd_filtr: list, jen_aktivni: bool, zdedene: bool,
                     plochy: bool, kdo: str) -> tuple:
    """Sestaví šifrovaný ZIP s maticí práv. Běží ve vlákně mimo event loop."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    tucne = Font(bold=True)
    svisle = Alignment(textRotation=90, vertical='bottom', horizontal='center')

    klice = [
        k for k, v in sorted(
            katalog.items(),
            key=lambda x: (x[1].get('kategorie', ''), x[1].get('podskupina', '') or '',
                           x[1].get('nazev', ''))
        )
        if not kategorie or v.get('kategorie') in kategorie
    ]
    uzivatele = [
        u for u in matice['uzivatele']
        if (u['aktivni'] or not jen_aktivni)
        and (not odd_filtr or set(u['oddeleni']) & set(odd_filtr))
    ]
    if not klice or not uzivatele:
        raise ValueError('Zvolenému filtru neodpovídá žádné právo ani uživatel.')
    if len(uzivatele) * len(klice) > _MATICE_MAX_BUNEK:
        raise ValueError('Matice je příliš velká — zužte kategorie práv nebo oddělení.')

    primo = matice['primo']
    role_prava = matice['role_prava']
    odd_prava = matice['odd_prava']

    wb = Workbook()
    wb.remove(wb.active)

    def _list(nazev: str, header: list, radky: list, sirky: dict = None,
              svisla_od: int = None):
        ws = wb.create_sheet(title=nazev[:31])
        ws.append(header)
        for i, bunka in enumerate(ws[1], 1):
            bunka.font = tucne
            if svisla_od and i >= svisla_od:
                bunka.alignment = svisle
        for r in radky:
            ws.append(r)
        if svisla_od:
            ws.row_dimensions[1].height = 150
            ws.freeze_panes = ws.cell(2, svisla_od).coordinate
        for i in range(1, len(header) + 1):
            pismeno = ws.cell(1, i).column_letter
            if svisla_od and i >= svisla_od:
                ws.column_dimensions[pismeno].width = 4
            else:
                ws.column_dimensions[pismeno].width = (sirky or {}).get(i, 22)
        return ws

    # --- List 1: Parametry (co přesně je v souboru) ---------------------
    _list('Parametry', ['Položka', 'Hodnota'], [
        ('Export', 'Matice práv'),
        ('Vytvořeno', datetime.datetime.now().strftime('%d.%m.%Y %H:%M')),
        ('Exportoval', kdo),
        ('Kategorie práv', ', '.join(kategorie) if kategorie else 'Všechny'),
        ('Oddělení', ', '.join(odd_filtr) if odd_filtr else 'Všechna'),
        ('Uživatelé', 'Jen aktivní' if jen_aktivni else 'Aktivní i neaktivní'),
        ('Zděděná práva (R/O)', 'Zahrnuta' if zdedene else 'Nezahrnuta (jen přímá)'),
        ('Počet uživatelů', len(uzivatele)),
        ('Počet práv', len(klice)),
        ('Legenda', 'P = přímo · R = přes pracovní pozici · O = přes oddělení'),
        ('Poznámka', 'Hlavní (skrytý) administrátor a servisní účet nejsou '
                     'v exportu. Práva kategorie „Administrace portálu“ se '
                     'nepřidělují přes UI, proto v matici nejsou.'),
    ], sirky={1: 26, 2: 90})

    # --- List 2: Uživatelé × práva -------------------------------------
    hlavicka = ['Jméno', 'E-mail', 'Aktivní', 'Oddělení', 'Role', 'Poznámka']
    radky = []
    pouzita_prava = set()
    plochy_radky = []
    for u in uzivatele:
        p = primo.get(u['id'], set())
        r = set().union(*[role_prava.get(x, set()) for x in u['role']]) if u['role'] else set()
        o = set().union(*[odd_prava.get(x, set()) for x in u['oddeleni']]) if u['oddeleni'] else set()
        pouzita_prava |= p | r | o
        vse = 'vse' in (p | r | o)
        radek = [u['jmeno'], u['email'], 'Ano' if u['aktivni'] else 'Ne',
                 ', '.join(u['oddeleni']), ', '.join(u['role']),
                 'VŠE (superadmin)' if vse else '']
        for k in klice:
            kanaly = 'P' if k in p else ''
            if zdedene:
                kanaly += 'R' if k in r else ''
                kanaly += 'O' if k in o else ''
            radek.append(kanaly)
            if plochy and kanaly:
                plochy_radky.append([
                    u['jmeno'], u['email'], katalog[k].get('nazev', k), k,
                    katalog[k].get('kategorie', ''), kanaly,
                ])
        radky.append(radek)
    _list('Uživatelé × práva',
          hlavicka + [katalog[k].get('nazev', k) for k in klice], radky,
          sirky={1: 26, 2: 30, 3: 9, 4: 24, 5: 24, 6: 18},
          svisla_od=len(hlavicka) + 1)

    # --- List 3+4: zdroj dědičnosti ------------------------------------
    _list('Role × práva', ['Role'] + [katalog[k].get('nazev', k) for k in klice],
          [[role] + ['X' if k in prava else '' for k in klice]
           for role, prava in sorted(role_prava.items())],
          sirky={1: 30}, svisla_od=2)
    _list('Oddělení × práva', ['Oddělení'] + [katalog[k].get('nazev', k) for k in klice],
          [[odd] + ['X' if k in prava else '' for k in klice]
           for odd, prava in sorted(odd_prava.items())],
          sirky={1: 30}, svisla_od=2)

    # --- List 5: katalog (vysvětlivka ke zkráceným hlavičkám) ----------
    _list('Katalog práv', ['Klíč', 'Kategorie', 'Podskupina', 'Název', 'Popis'],
          [[k, katalog[k].get('kategorie', ''), katalog[k].get('podskupina', '') or '',
            katalog[k].get('nazev', ''), katalog[k].get('popis', '')] for k in klice],
          sirky={1: 30, 2: 26, 3: 20, 4: 34, 5: 80})

    # --- List 6: plochý seznam (volitelně, pro filtrování/kontingenci) --
    if plochy:
        _list('Plochý seznam',
              ['Jméno', 'E-mail', 'Právo', 'Klíč', 'Kategorie', 'Kanál'],
              plochy_radky, sirky={1: 26, 2: 30, 3: 34, 4: 30, 5: 26, 6: 10})

    # --- List 7: práva v DB, která katalog nezná (jen když existují) ----
    neznama = sorted(pouzita_prava - set(katalog) - {'vse'} - set(intranet_prava.ADMIN_ONLY_PRAVA))
    if neznama:
        _list('Neznámá práva', ['Klíč práva v DB'], [[k] for k in neznama], sirky={1: 40})

    ted = datetime.datetime.now().strftime('%d_%m_%H_%M')
    xlsx_nazev = f'matice_prav_{ted}.xlsx'
    zip_nazev = f'matice_prav_{ted}.zip'
    buf = io.BytesIO()
    wb.save(buf)
    cesta = _matice_tmp_cesta(zip_nazev)
    with pyzipper.AESZipFile(
        cesta, 'w',
        compression=pyzipper.ZIP_DEFLATED,
        encryption=pyzipper.WZ_AES,
    ) as zf:
        # Heslo jen z env (intranet_data.export_zip_heslo) — stejný mechanismus
        # jako u exportu uživatelů, lze rotovat bez zásahu do kódu.
        zf.setpassword(intranet_data.export_zip_heslo())
        zf.writestr(xlsx_nazev, buf.getvalue())
    return cesta, zip_nazev, len(uzivatele), len(klice)

@ui.refreshable
def vykresli_spravu_uzivatelu(user_email, user_name, vsechna_prava=None):
    # Matici práv smí stáhnout jen superadmin — je to mapa, kudy do aplikace.
    _je_superadmin = "vse" in (vsechna_prava or [])
    with ui.row().classes('w-full items-center gap-4 mb-8'):
        ui.icon('manage_accounts', size='2.4rem').classes('text-indigo-600 bg-indigo-50 p-3 rounded-2xl shadow-sm')
        with ui.column().classes('gap-0'):
            ui.label('Správa uživatelů a oprávnění').classes('text-3xl md:text-4xl font-extrabold text-slate-800 leading-tight')
            ui.label('Uživatelé, oddělení, práva a příznaky na jednom místě.').classes('text-sm text-slate-500 font-medium')

    db = intranet_data.ziskej_vsechny_uzivatele()
    role = intranet_data.ziskej_vsechny_role()
    oddeleni = intranet_data.ziskej_vsechna_oddeleni()
    typy_v = intranet_data.ziskej_typy_volna()
    zakladni_prava = intranet_prava.ziskej_kompletni_seznam_prav(oddeleni, typy_v)

    _nast = intranet_data.nacti_nastaveni_intranetu()
    _skryte_kategorie = set()
    if not _nast.get('kviz_zapnuty', True):       _skryte_kategorie.add('Modul Kvíz')
    if not _nast.get('finance_zapnuty', True):    _skryte_kategorie.add('Modul Aprovia')
    if not _nast.get('veletrh_zapnuty', True):    _skryte_kategorie.add('Modul Veletrh')
    if not _nast.get('planogram_zapnuty', True):  _skryte_kategorie.add('Modul Plánogram tabáku')
    if not _nast.get('ochutnavky_zapnuty', True): _skryte_kategorie.add('Modul Ochutnávky MO a CC')
    # Značky mají jednu kategorii se dvěma podskupinami → skrývá se podskupina.
    _skryte_podskupiny = set()
    if not _nast.get('znacky_zapnuty', True):        _skryte_podskupiny.add(('Modul Značky', 'Produkt'))
    if not _nast.get('znacky_provoz_zapnuty', True): _skryte_podskupiny.add(('Modul Značky', 'Provoz'))
    if _skryte_kategorie or _skryte_podskupiny:
        zakladni_prava = {
            k: v for k, v in zakladni_prava.items()
            if v.get('kategorie') not in _skryte_kategorie
            and (v.get('kategorie'), v.get('podskupina')) not in _skryte_podskupiny
        }

    spolecnosti_db = intranet_data.ziskej_vsechny_spolecnosti()
    spolecnosti_options = {s['id']: s['nazev'] for s in spolecnosti_db}

    priznaky_db = intranet_data.ziskej_vsechny_priznaky()

    # ── Dialog správy příznaků ────────────────────────────────────────────────
    def _dialog_sprava_priznaku():
        with ui.dialog() as dlg_pr, ui.card().classes('p-6 rounded-2xl shadow-2xl w-full max-w-lg'):
            ui.label('Správa příznaků').classes('text-xl font-extrabold text-gray-800 mb-4')

            @ui.refreshable
            def _seznam_priznaku():
                aktualni = intranet_data.ziskej_vsechny_priznaky()
                if not aktualni:
                    ui.label('Žádné příznaky.').classes('text-gray-400 italic text-sm mb-3')
                else:
                    with ui.column().classes('w-full gap-2 mb-4'):
                        for pz in aktualni:
                            with ui.row().classes('w-full items-center gap-3'):
                                ui.element('div').style(
                                    f'width:18px;height:18px;border-radius:50%;background:{pz["barva"]};flex-shrink:0'
                                )
                                ui.label(pz['nazev']).classes('flex-1 text-sm font-medium text-gray-700')

                                async def _smaz(pid=pz['id'], pnazev=pz['nazev']):
                                    ok, msg = await asyncio.to_thread(intranet_data.smaz_priznak, pid)
                                    if ok:
                                        ui.notify(f'Příznak "{pnazev}" smazán.', type='positive')
                                        intranet_data.invaliduj_cache_sprava()
                                        _seznam_priznaku.refresh()
                                        ui.timer(0, vykresli_spravu_uzivatelu.refresh, once=True)
                                    else:
                                        ui.notify(f'Chyba: {msg}', type='negative')

                                ui.button(icon='delete', on_click=_smaz).props('flat round dense').classes('text-red-400 hover:text-red-600')

            _seznam_priznaku()

            ui.separator()
            ui.label('Přidat nový příznak').classes('text-sm font-bold text-gray-700 mt-3 mb-2')
            with ui.row().classes('w-full items-center gap-3'):
                input_nazev_pr = ui.input('Název příznaku').classes('flex-1').props('outlined dense')
                input_barva_pr = ui.color_input(label='Barva', value='#6366f1').classes('w-32')

            async def _pridat_priznak():
                nazev = input_nazev_pr.value.strip()
                if not nazev:
                    ui.notify('Zadejte název příznaku.', type='warning')
                    return
                ok, msg = await asyncio.to_thread(intranet_data.pridej_priznak, nazev, input_barva_pr.value)
                if ok:
                    ui.notify(f'Příznak "{nazev}" přidán.', type='positive')
                    input_nazev_pr.set_value('')
                    intranet_data.invaliduj_cache_sprava()
                    _seznam_priznaku.refresh()
                    ui.timer(0, vykresli_spravu_uzivatelu.refresh, once=True)
                else:
                    ui.notify(f'Chyba: {msg}', type='negative')

            with ui.row().classes('w-full justify-end gap-3 mt-4'):
                ui.button('Přidat', icon='add', on_click=_pridat_priznak).classes('bg-indigo-600 hover:bg-indigo-700 text-white font-bold')

            ui.separator().classes('my-4')
            ui.label('Hromadné přiřazení celému oddělení').classes('text-sm font-bold text-gray-700 mb-2')

            _odd_list = list(oddeleni.keys())
            _pr_hrom_opts = {pz['id']: pz['nazev'] for pz in intranet_data.ziskej_vsechny_priznaky()}

            with ui.row().classes('w-full items-center gap-3'):
                vyber_odd_hrom = ui.select(
                    _odd_list, label='Oddělení', value=_odd_list[0] if _odd_list else None
                ).classes('flex-1')
                vyber_pr_hrom = ui.select(
                    _pr_hrom_opts, label='Příznak'
                ).classes('flex-1')

            async def _prirad_oddeleni():
                odd = vyber_odd_hrom.value
                pid = vyber_pr_hrom.value
                if not odd or not pid:
                    ui.notify('Vyberte oddělení i příznak.', type='warning')
                    return
                ok, pocet = await asyncio.to_thread(intranet_data.nastav_priznak_oddeleni, odd, pid)
                if ok:
                    ui.notify(f'Příznak přiřazen {pocet} uživatelům oddělení „{odd}".', type='positive')
                    intranet_data.invaliduj_cache_sprava()
                    ui.timer(0, vykresli_seznam_uzivatelu.refresh, once=True)
                else:
                    ui.notify('Chyba při hromadném přiřazení.', type='negative')

            ui.button('Přiřadit oddělení', icon='group', on_click=_prirad_oddeleni) \
              .classes('bg-teal-600 hover:bg-teal-700 text-white font-bold mt-2')

            ui.separator().classes('my-4')
            with ui.row().classes('w-full justify-end'):
                ui.button('Zavřít', on_click=dlg_pr.close).props('flat').classes('text-gray-600')

        dlg_pr.open()

    with ui.card().classes('w-full max-w-4xl p-8 shadow-sm bg-white rounded-2xl mb-8 border border-slate-200'):
        with ui.row().classes('w-full items-center justify-between mb-6 pb-4 border-b border-slate-100'):
            with ui.row().classes('items-center gap-3'):
                ui.icon('person_add', size='1.6rem').classes('text-indigo-600 bg-indigo-50 p-2 rounded-xl')
                ui.label('Přidat nového uživatele').classes('text-2xl font-bold text-slate-800')
            with ui.row().classes('items-center gap-2'):
                ui.button('Spravovat příznaky', icon='label', on_click=_dialog_sprava_priznaku) \
                  .props('flat dense no-caps').classes('text-indigo-600 font-semibold rounded-lg')

        ui.label('Základní údaje').classes('text-xs font-bold text-slate-400 uppercase tracking-wider mb-2')
        with ui.row().classes('w-full gap-4 mb-4'):
            input_email = ui.input('E-mail (Přihlašovací jméno)').classes('flex-[2]').props('outlined dense')
            input_jmeno = ui.input('Jméno').classes('flex-1').props('outlined dense')
            input_prijmeni = ui.input('Příjmení').classes('flex-1').props('outlined dense')

        with ui.row().classes('w-full gap-4 mb-6'):
            input_heslo = ui.input('Heslo', password=True).classes('flex-1').props('outlined dense')
            input_osobni_cislo = ui.input('Osobní číslo (Pro exporty)').classes('flex-1').props('outlined dense type=number min=1 step=1')
            input_datum_narozeni = ui.input('Datum narození (RRRR-MM-DD)').classes('flex-1').props('outlined dense mask="####-##-##" fill-mask')

        ui.label('Zařazení').classes('text-xs font-bold text-slate-400 uppercase tracking-wider mb-2')
        with ui.row().classes('w-full gap-4 mb-4'):
            input_spolecnost = ui.select(spolecnosti_options, label='Společnosti', multiple=True).classes('flex-1').props('outlined dense use-chips clearable')
            vyber_role = ui.select(role + ['Bez role'], value='Bez role', label='Pracovní pozice').classes('flex-1').props('outlined dense')
            vyber_oddeleni = ui.select(list(oddeleni.keys()) + ['Bez oddělení'], value='Bez oddělení', label='Oddělení').classes('flex-1').props('outlined dense')
            _pr_options_novy = {pz['id']: pz['nazev'] for pz in uznaky_db} if (uznaky_db := intranet_data.ziskej_vsechny_priznaky()) else {}
            vyber_priznak_novy = ui.select({None: '— bez příznaku —', **_pr_options_novy}, value=None, label='Příznak').classes('flex-1').props('outlined dense')
            vyber_pobocka_novy = ui.select({None: '— bez pobočky —', **{_pb: _pb for _pb in intranet_data.POBOCKY}}, value=None, label='Pobočka').classes('flex-1').props('outlined dense')

        with ui.row().classes('w-full gap-4 mb-6'):
            manazeri = {u['id']: f"{u['jmeno_cele']} ({u.get('oddeleni', '')})" for u in sorted((u for u in db.values() if u['aktivni']), key=lambda u: (cz_razeni(u.get('prijmeni', '')), cz_razeni(u.get('jmeno_cele', ''))))}
            vyber_manazera = ui.select(manazeri, label='Přímí nadřízení (Schvalovatelé)', multiple=True, with_input=True).classes('flex-[2]').props('outlined dense use-chips clearable')
            input_zaklad = ui.number('Základ dovolené (h)', value=160.0).classes('flex-1').props('outlined dense')
            input_prevod = ui.number('Převod z loňska (h)', value=0.0).classes('flex-1').props('outlined dense')

        ui.label('Osobní práva navíc (Přepíší výchozí práva oddělení)').classes('text-xs font-bold text-slate-400 uppercase tracking-wider mb-2')
        vybrane_prava_novy = render_prava_kategorie(zakladni_prava, [], lazy=True)

        async def pridat_uzivatele():
            em, jm, pr, heslo = input_email.value.strip(), input_jmeno.value.strip(), input_prijmeni.value.strip(), input_heslo.value
            oc = int(input_osobni_cislo.value) if input_osobni_cislo.value else None

            manazer_ids = vyber_manazera.value if vyber_manazera.value else []
            spol_ids = input_spolecnost.value if input_spolecnost.value else []

            if not em or not heslo: return ui.notify('E-mail a heslo jsou povinné!', type='warning', position='top')

            if not intranet_data.heslo_je_silne(heslo):
                return ui.notify('Heslo musí mít min. 8 znaků, obsahovat velké i malé písmeno a číslo.', type='warning', position='top')

            db_lokalni = await asyncio.to_thread(intranet_data.ziskej_vsechny_uzivatele)
            if em in db_lokalni: return ui.notify('Uživatel s tímto e-mailem už existuje!', type='negative', position='top')

            status, msg = await asyncio.to_thread(
                intranet_data.pridej_uprav_uzivatele,
                em, jm, pr, heslo, vyber_role.value, vyber_oddeleni.value,
                ",".join(vybrane_prava_novy), True, input_zaklad.value, input_prevod.value, oc,
                manazer_ids, spol_ids, input_datum_narozeni.value or None,
                vyber_priznak_novy.value, vyber_pobocka_novy.value
            )

            if status:
                intranet_logger.log_activity(user_name, "Správa uživatelů", f"Vytvořen uživatel: {em}")
                ui.notify(f'Uživatel {jm} {pr} přidán.', type='positive', position='top')
                intranet_data.invaliduj_cache_sprava()
                await asyncio.to_thread(intranet_data.ziskej_vsechny_uzivatele)
                ui.timer(0, vykresli_seznam_uzivatelu.refresh, once=True)
                # Vyčisti formulář
                input_email.set_value('')
                input_jmeno.set_value('')
                input_prijmeni.set_value('')
                input_heslo.set_value('')
                input_osobni_cislo.set_value('')
                input_spolecnost.set_value([])
                vyber_role.set_value('Bez role')
                vyber_oddeleni.set_value('Bez oddělení')
                input_zaklad.set_value(160.0)
                input_prevod.set_value(0.0)
                vyber_manazera.set_value([])
                input_datum_narozeni.set_value('')
                vyber_pobocka_novy.set_value(None)
            else:
                ui.notify(msg, type='negative', position='top')

        with ui.row().classes('w-full justify-end mt-6 pt-4 border-t border-slate-100'):
            ui.button('Vytvořit uživatele', on_click=pridat_uzivatele, icon='person_add') \
              .props('unelevated no-caps').classes('bg-indigo-600 hover:bg-indigo-700 text-white font-semibold h-12 px-8 rounded-lg shadow-sm')

    with ui.card().classes('w-full p-8 shadow-sm bg-white rounded-2xl mb-8 border border-slate-200'):
        with ui.row().classes('w-full items-center gap-3 mb-6'):
            ui.icon('groups', size='1.6rem').classes('text-indigo-600 bg-indigo-50 p-2 rounded-xl')
            ui.label('Existující uživatelé').classes('text-2xl font-bold text-slate-800')

        # ── Export CSV ────────────────────────────────────────────────────
        _EXPORT_SLOUPCE = [
            ('id',                    'Osobní číslo'),
            ('jmeno',                 'Jméno'),
            ('prijmeni',              'Příjmení'),
            ('email',                 'E-mail'),
            ('spolecnosti',           'Společnost'),
            ('oddeleni',              'Oddělení'),
            ('role',                  'Pracovní pozice'),
            ('nadrizeni',             'Přímí nadřízení'),
            ('aktivni',               'Je v zam. poměru'),
            ('datum_narozeni',        'Datum narození'),
            ('base_vacation',         'Základ dovolené (h)'),
            ('carried_over_vacation', 'Převod z loňska (h)'),
            ('prava',                 'Osobní práva navíc'),
            ('priznak',              'Příznak'),
        ]
        _DEFAULT_EXPORT = {'id', 'jmeno', 'prijmeni', 'spolecnosti', 'oddeleni', 'aktivni'}

        def _dialog_export_csv():
            with ui.dialog() as dlg, \
                 ui.card().classes('p-6 rounded-2xl shadow-2xl w-full max-w-lg'):
                ui.label('Export uživatelů do CSV') \
                  .classes('text-xl font-extrabold text-gray-800 mb-1')
                ui.label('Vyberte sloupce, které chcete exportovat:') \
                  .classes('text-sm text-gray-500 mb-4')

                checkboxy: dict = {}

                def _vyber_vse(hodnota: bool):
                    for cb in checkboxy.values():
                        cb.set_value(hodnota)

                with ui.row().classes('items-center gap-6 mb-3'):
                    ui.checkbox('Vybrat vše', value=False,
                                on_change=lambda e: _vyber_vse(e.value)) \
                      .classes('font-bold text-gray-700')
                    cb_rozdelit = ui.checkbox(
                        'Rozdělit na listy podle společnosti',
                        value=False,
                    ).classes('font-bold text-blue-700')
                    def _on_sloucit_change(e):
                        if e.value:
                            checkboxy['priznak'].set_value(True)
                            checkboxy['id'].set_value(True)
                    cb_sloucit_priznak = ui.checkbox(
                        'Sloučit příznak s os. číslem',
                        value=False,
                        on_change=_on_sloucit_change,
                    ).classes('font-bold text-indigo-700')

                with ui.element('div').classes(
                    'grid gap-x-6 gap-y-1 mb-5'
                ).style('grid-template-columns: 1fr 1fr'):
                    for klic, label in _EXPORT_SLOUPCE:
                        checkboxy[klic] = ui.checkbox(
                            label, value=(klic in _DEFAULT_EXPORT)
                        )

                def _radek_uzivatele(d: dict, em: str, vybrane: list,
                                    snap_prava: dict, aktualni_db: dict,
                                    sloucit: bool = False) -> list:
                    """Sestaví jeden CSV řádek pro uživatele d."""
                    row = []
                    _sloucit_aktivni = sloucit and 'id' in vybrane and 'priznak' in vybrane
                    for klic in vybrane:
                        if _sloucit_aktivni and klic == 'priznak':
                            continue  # sloučeno do sloupce id
                        if klic == 'id':
                            if _sloucit_aktivni:
                                priznak = d.get('priznak_nazev', '') or ''
                                osobni = str(d.get('id', '')).zfill(6)
                                row.append(f'{priznak}{osobni}' if priznak else osobni)
                            else:
                                row.append(d.get('id', ''))
                        elif klic == 'jmeno':
                            row.append(d.get('jmeno', ''))
                        elif klic == 'prijmeni':
                            row.append(d.get('prijmeni', ''))
                        elif klic == 'email':
                            row.append(em)
                        elif klic == 'spolecnosti':
                            nazvy = [s['nazev'] for s in d.get('spolecnosti', [])]
                            row.append(', '.join(nazvy) if nazvy else '')
                        elif klic == 'oddeleni':
                            row.append(d.get('oddeleni', '') or '')
                        elif klic == 'role':
                            row.append(d.get('role', '') or '')
                        elif klic == 'nadrizeni':
                            jmena = []
                            for mid in d.get('manager_id', []):
                                mj = next(
                                    (u['jmeno_cele'] for u in aktualni_db.values()
                                     if u['id'] == mid), None
                                )
                                if mj:
                                    jmena.append(mj)
                            row.append(', '.join(jmena) if jmena else '')
                        elif klic == 'aktivni':
                            row.append(1 if d.get('aktivni') else 0)
                        elif klic == 'datum_narozeni':
                            row.append(d.get('datum_narozeni', '') or '')
                        elif klic == 'base_vacation':
                            row.append(d.get('base_vacation', '') or '')
                        elif klic == 'carried_over_vacation':
                            row.append(d.get('carried_over_vacation', '') or '')
                        elif klic == 'prava':
                            prava_raw = d.get('prava', '') or ''
                            prava_list = [
                                p.strip() for p in prava_raw.split(',') if p.strip()
                            ]
                            prava_names = []
                            for p in prava_list:
                                meta = snap_prava.get(p)
                                prava_names.append(
                                    meta['nazev'] if isinstance(meta, dict) else p
                                )
                            row.append(', '.join(prava_names) if prava_names else '')
                        elif klic == 'priznak':
                            row.append(d.get('priznak_nazev', '') or '')
                    return row

                def _sestav_zip_sync(vybrane: list, snap_prava: dict,
                                     rozdelit: bool, sloucit: bool) -> tuple:
                    """Těžká práce — běží ve vlákně mimo event loop."""
                    aktualni_db = intranet_data.ziskej_vsechny_uzivatele()
                    ted = datetime.datetime.now().strftime('%d_%m_%H_%M')
                    zip_nazev = f'export_uzivatele_{ted}.zip'
                    _sloucit_aktivni = sloucit and 'id' in vybrane and 'priznak' in vybrane
                    header = []
                    for k, lbl in _EXPORT_SLOUPCE:
                        if k not in vybrane:
                            continue
                        if _sloucit_aktivni and k == 'priznak':
                            continue  # přeskočit — sloučeno do sloupce id
                        if _sloucit_aktivni and k == 'id':
                            header.append('Příznak + Osobní číslo')
                        else:
                            header.append(lbl)

                    # Přeskočit adminské účty, seřadit abecedně
                    uzivatele = [
                        (em, d)
                        for em, d in sorted(
                            aktualni_db.items(), key=lambda x: x[1]['jmeno_cele']
                        )
                        if 'admin' not in d['jmeno_cele'].lower()
                        and 'admin' not in em.lower()
                    ]

                    zip_buf = io.BytesIO()
                    with pyzipper.AESZipFile(
                        zip_buf, 'w',
                        compression=pyzipper.ZIP_DEFLATED,
                        encryption=pyzipper.WZ_AES,
                    ) as zf:
                        # Heslo k šifrovanému ZIP exportu — jen z env (viz
                        # intranet_data.export_zip_heslo), lze rotovat bez zásahu do kódu.
                        zf.setpassword(intranet_data.export_zip_heslo())

                        if rozdelit:
                            # Seskup uživatele podle společností
                            skupiny: dict[str, list] = {}
                            for em, d in uzivatele:
                                spol_list = [s['nazev'] for s in d.get('spolecnosti', [])]
                                if not spol_list:
                                    spol_list = ['Nepřiřazeno']
                                for spol in spol_list:
                                    skupiny.setdefault(spol, []).append((em, d))

                            # Jeden Excel soubor, každá společnost = jeden list
                            wb = openpyxl.Workbook()
                            wb.remove(wb.active)  # odeber prázdný výchozí list

                            for spol_nazev in sorted(skupiny):
                                # Název listu max 31 znaků, bez zakázaných znaků
                                nazev_listu = re.sub(r'[\\/:*?\[\]]', '_', spol_nazev)[:31]
                                ws = wb.create_sheet(title=nazev_listu)
                                ws.append(header)
                                for em, d in skupiny[spol_nazev]:
                                    ws.append(
                                        _radek_uzivatele(d, em, vybrane,
                                                         snap_prava, aktualni_db, sloucit)
                                    )

                            xlsx_buf = io.BytesIO()
                            wb.save(xlsx_buf)
                            xlsx_nazev = f'export_uzivatele_{ted}.xlsx'
                            zf.writestr(xlsx_nazev, xlsx_buf.getvalue())
                        else:
                            buf = io.StringIO()
                            writer = csv.writer(buf, delimiter=';')
                            writer.writerow(header)
                            for em, d in uzivatele:
                                writer.writerow(
                                    _radek_uzivatele(d, em, vybrane,
                                                     snap_prava, aktualni_db, sloucit)
                                )
                            csv_nazev = f'export_uzivatele_{ted}.csv'
                            zf.writestr(csv_nazev,
                                        b'\xef\xbb\xbf' + buf.getvalue().encode('utf-8'))

                    return zip_buf.getvalue(), zip_nazev

                async def proved_export():
                    vybrane = [k for k, cb in checkboxy.items() if cb.value]
                    if not vybrane:
                        ui.notify('Vyberte alespoň jeden sloupec.', type='warning')
                        return

                    snap_prava = dict(zakladni_prava)
                    rozdelit = cb_rozdelit.value
                    sloucit = cb_sloucit_priznak.value
                    try:
                        zip_data, zip_nazev = await asyncio.to_thread(
                            _sestav_zip_sync, vybrane, snap_prava, rozdelit, sloucit
                        )
                    except Exception as ex:
                        ui.notify(f'Chyba při exportu: {ex}', type='negative')
                        return

                    vlastni_nazev = (input_nazev.value or '').strip()
                    if vlastni_nazev:
                        if not vlastni_nazev.lower().endswith('.zip'):
                            vlastni_nazev += '.zip'
                        zip_nazev = vlastni_nazev

                    ui.download(zip_data, zip_nazev)
                    dlg.close()

                input_nazev = ui.input(
                    label='Název souboru (volitelné)',
                    placeholder=f'export_uzivatele_{datetime.datetime.now().strftime("%d_%m_%H_%M")}',
                ).classes('w-full mb-2').props('outlined dense clearable')

                with ui.row().classes('w-full justify-end gap-3 mt-2'):
                    ui.button('Zrušit', on_click=dlg.close) \
                      .classes('bg-gray-200 hover:bg-gray-300 text-gray-800 font-bold')
                    ui.button('Exportovat', icon='download', on_click=proved_export) \
                      .classes('bg-gray-700 hover:bg-gray-800 text-white font-bold')

            dlg.open()

        # ── Dialog: porovnání se sestavou ────────────────────────────────
        def _dialog_porovnani():
            stav = {'radky': [], 'nacteno': False}

            with ui.dialog() as dlg_pv, \
                 ui.card().classes('p-6 rounded-2xl shadow-2xl flex flex-col gap-4') \
                           .style('width:520px;max-height:90vh;'
                                  'overflow-y:auto;overflow-x:hidden') as karta_pv:

                with ui.row().classes('w-full items-center justify-between'):
                    ui.label('Porovnání se sestavou zaměstnanců') \
                      .classes('text-xl font-extrabold text-gray-800')
                    ui.button(icon='close', on_click=dlg_pv.close).props('flat round')

                popis_el = ui.label(
                    'Nahrajte sestavu ve formátu CSV (CisloJIP;Jmeno;Prijmeni;…;je_v_zam_pomeru). '
                    'Porovnání probíhá přes klíč Příznak + Osobní číslo (6 číslic). '
                    'Aktivita účtu v systému se porovnává se sloupcem je_v_zam_pomeru '
                    '(0 = mimo poměr → k deaktivaci, 1 = v poměru). Akci lze provést přímo z přehledu.'
                ).classes('text-sm text-gray-500')

                @ui.refreshable
                def _vykresli():
                    # Ochranná obálka: chyba při kreslení = viditelný banner,
                    # ne tiše prázdný dialog.
                    try:
                        _vykresli_telo()
                    except Exception:
                        import traceback
                        import html as _html
                        tb = traceback.format_exc()
                        print(tb, flush=True)
                        with ui.column().classes(
                                'w-full gap-2 p-4 bg-red-50 border '
                                'border-red-300 rounded-lg'):
                            ui.label('⚠ Chyba při vykreslení porovnání') \
                              .classes('font-bold text-red-700')
                            ui.html(
                                f'<pre style="white-space:pre-wrap;font-size:11px;'
                                f'margin:0">{_html.escape(tb)}</pre>')

                def _vykresli_telo():
                    if not stav['nacteno']:
                        ui.label('Nahrajte soubor sestavy…') \
                          .classes('text-gray-400 italic text-center mt-8')
                        return

                    aktualni_db = intranet_data.ziskej_vsechny_uzivatele()

                    # ── Sestavení map ─────────────────────────────────────────
                    system_map: dict = {}
                    email_map:  dict = {}
                    for em, d in aktualni_db.items():
                        if 'admin' in d['jmeno_cele'].lower() or 'admin' in em.lower():
                            continue
                        priznak = d.get('priznak_nazev', '') or ''
                        if priznak:
                            klic = f"{priznak}{str(d['id']).zfill(6)}"
                            system_map[klic] = d
                            email_map[klic]  = em

                    csv_map: dict = {}
                    for r in stav['radky']:
                        klic = r.get('CisloJIP', '').strip()
                        if klic:
                            csv_map[klic] = r

                    def _je_zam(r):
                        return str(r.get('je_v_zam_pomeru', '')).strip() == '1'

                    sys_keys    = set(system_map)
                    csv_keys    = set(csv_map)
                    shodne_s    = sys_keys & csv_keys
                    pouze_sys_s = sys_keys - csv_keys
                    pouze_csv_s = csv_keys - sys_keys

                    # ── Klasifikace stavů ─────────────────────────────────────
                    # OK            : klíč je v obou + stav se shoduje
                    #                 (sys aktivní ↔ CSV pomer=1, sys neaktivní ↔ CSV pomer=0)
                    #                 nebo sys neaktivní + chybí v CSV (už správně deaktivován)
                    # K deaktivaci  : sys aktivní + (CSV pomer=0  NEBO  chybí v CSV)
                    # K reaktivaci  : sys neaktivní + CSV pomer=1
                    # Pouze v sestavě: klíč není v sys → potřeba založit nový účet
                    ok_set      = set()
                    k_deakt_set = set()
                    k_reakt_set = set()

                    for k in shodne_s:
                        sys_active   = bool(system_map[k].get('aktivni'))
                        csv_employed = _je_zam(csv_map[k])
                        if sys_active and not csv_employed:
                            k_deakt_set.add(k)
                        elif (not sys_active) and csv_employed:
                            k_reakt_set.add(k)
                        else:
                            ok_set.add(k)

                    for k in pouze_sys_s:
                        if system_map[k].get('aktivni'):
                            k_deakt_set.add(k)
                        else:
                            ok_set.add(k)   # už deaktivován a v sestavě chybí → v pořádku

                    # ── Návrhy k aktivaci / deaktivaci ────────────────────────
                    # Okno = aktuální + minulý týden (po–ne), oboustranně ohraničené:
                    #   od pondělí minulého týdne do neděle tohoto týdne.
                    # Aktivace : ne v systému + poměr=1 + nástup v okně.
                    # Deaktivace: v systému + aktivní + odchod v okně
                    #   (jen lidé, kteří odešli nebo odejdou v aktuálním/minulém týdnu).
                    def _parse_dat(s):
                        s = (s or '').strip()
                        for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%d.%m.%y'):
                            try:
                                return datetime.datetime.strptime(s, fmt).date()
                            except ValueError:
                                continue
                        return None

                    _dnes = datetime.date.today()
                    _pondeli_tohoto = _dnes - datetime.timedelta(days=_dnes.weekday())
                    _pocatek_okna   = _pondeli_tohoto - datetime.timedelta(days=7)
                    _konec_okna     = _pondeli_tohoto + datetime.timedelta(days=6)

                    def _v_okne(d):
                        return d is not None and _pocatek_okna <= d <= _konec_okna

                    navrhy_set = set()        # k aktivaci (dle nástup)
                    for k in pouze_csv_s:
                        rr = csv_map[k]
                        if _je_zam(rr) and _v_okne(_parse_dat(rr.get('nastup'))):
                            navrhy_set.add(k)

                    navrhy_deakt_set = set()  # k deaktivaci (dle odchod)
                    for k in sys_keys:
                        if not system_map[k].get('aktivni'):
                            continue
                        rr = csv_map.get(k)
                        if rr and _v_okne(_parse_dat(rr.get('odchod'))):
                            navrhy_deakt_set.add(k)

                    def _stav_for(k):
                        if k in k_deakt_set: return 'K deaktivaci'
                        if k in k_reakt_set: return 'K reaktivaci'
                        if k in pouze_csv_s: return 'Pouze v sestavě'
                        return 'OK'

                    _ROW_CLS = {
                        'OK':              '',
                        'K deaktivaci':    'bg-red-50',
                        'K reaktivaci':    'bg-emerald-50',
                        'Pouze v sestavě': 'bg-amber-50',
                    }

                    # ── Předpočítaná data: systémová strana ───────────────────
                    def _sys_row(k):
                        d = system_map[k]
                        stav_lbl = _stav_for(k)
                        return {
                            'cislo':    k,
                            'jmeno':    d.get('jmeno_cele', ''),
                            'email':    email_map.get(k, ''),
                            'odd':      d.get('oddeleni', '') or '',
                            'aktivni':  '✓' if d.get('aktivni') else '✗',
                            'stav':     stav_lbl,
                            '_row_cls': _ROW_CLS.get(stav_lbl, ''),
                        }

                    # ── Předpočítaná data: CSV strana ─────────────────────────
                    def _csv_row(k):
                        r = csv_map[k]
                        stav_lbl = _stav_for(k)
                        return {
                            'cislo':    k,
                            'jmeno':    r.get('Jmeno', ''),
                            'prijmeni': r.get('Prijmeni', ''),
                            'charakteristika': r.get('nakladovy_okruh', '') or '',
                            'charakteristika_nazev': intranet_okruhy.nazev_okruhu(
                                r.get('nakladovy_okruh', '')),
                            'nastup':   r.get('nastup', ''),
                            'odchod':   r.get('odchod', ''),
                            'pomer':    '✓' if _je_zam(r) else '✗',
                            'vypoved':  '✓' if str(r.get('je_ve_vypovedi', '')) == '1' else '',
                            'stav':     stav_lbl,
                            '_row_cls': _ROW_CLS.get(stav_lbl, ''),
                        }

                    sys_rows_vse = [_sys_row(k) for k in sorted(sys_keys)]
                    csv_rows_vse = [_csv_row(k) for k in sorted(csv_keys)]

                    # ── Unikátní prefixy pro filtr ────────────────────────────
                    import re as _re
                    all_prefixes = sorted({
                        m.group(1) for k in (sys_keys | csv_keys)
                        if (m := _re.match(r'^([A-Za-z]+)', k))
                    })

                    # ── Sdílený stav filtrů ───────────────────────────────────
                    f             = {'text': '', 'prefix': ''}
                    tbl_refs      = []   # (tbl, all_rows, strana) – split q-tabulky
                    filter_hooks  = []   # refresh fce vlastních seznamů (Návrhy)

                    def _match_filter(r):
                        """Vrátí True, pokud řádek (dict) projde aktuálním filtrem."""
                        prefix = f['prefix']
                        if prefix and not str(r.get('cislo', '')).startswith(prefix):
                            return False
                        txt = f['text'].lower()
                        if txt and not any(
                            txt in str(r.get(c, '')).lower()
                            for c in ('cislo', 'jmeno', 'prijmeni', 'charakteristika', 'email', 'odd', 'nastup')
                        ):
                            return False
                        return True

                    def _filtruj(rows, strana):
                        return [r for r in rows if _match_filter(r)]

                    def _update_all():
                        for tbl, rows, strana in tbl_refs:
                            tbl.rows = _filtruj(rows, strana)
                            tbl.update()
                        for hook in filter_hooks:
                            try:
                                hook()
                            except Exception:
                                pass

                    # ── Stat karty (klikatelné, přepínají záložky) ────────────
                    # tabs je definován níže, ale lambda ho zachytí až při volání ✓
                    with ui.row().classes('gap-4 mb-4 flex-wrap'):
                        for label, count, border, tname in [
                            ('OK',                 len(ok_set),           'border-green-500',   'ok'),
                            ('Návrhy k aktivaci',  len(navrhy_set),       'border-blue-500',    'navrhy'),
                            ('Návrhy k deaktivaci', len(navrhy_deakt_set), 'border-rose-500',   'navrhy_deakt'),
                        ]:
                            txt_cls = border.replace('border-', 'text-').replace('-500', '-700').replace('-400', '-700')
                            with ui.card() \
                                    .classes(f'cursor-pointer p-5 rounded-xl border-t-4 {border} '
                                             f'hover:shadow-lg transition-all select-none') \
                                    .style('min-width:150px') \
                                    .on('click', lambda tn=tname: tabs.set_value(tn)):
                                ui.label(str(count)).classes(f'text-4xl font-bold {txt_cls}')
                                ui.label(label).classes('text-sm text-gray-500 uppercase tracking-wide mt-1')

                    # ── Filtrovací lišta ──────────────────────────────────────
                    with ui.row().classes('items-center gap-3 mb-3 flex-wrap'):
                        def _on_search(e):
                            f['text'] = e.value or ''
                            _update_all()
                        inp_search = ui.input(
                            placeholder='Hledat… (kliknutí na řádek vyplní Číslo JIP)',
                            on_change=_on_search,
                        ).props('outlined dense clearable').classes('flex-1 max-w-sm')

                        # Prefix filtr
                        prefix_opts = {'': '— Vše —'} | {p: p for p in all_prefixes}
                        def _on_prefix(e):
                            f['prefix'] = e.value or ''
                            _update_all()
                        ui.select(prefix_opts, value='', label='Příznak',
                                  on_change=_on_prefix) \
                          .props('outlined dense').classes('w-36')

                    # ── Klik na řádek → vyplní hledání ───────────────────────
                    def _on_row_click(args):
                        try:
                            cislo = args[1].get('cislo', '')
                            if cislo:
                                inp_search.set_value(cislo)
                                f['text'] = cislo
                                _update_all()
                        except Exception:
                            pass

                    # ── Slot pro barevný badge ve sloupci Stav ────────────────
                    STAV_SLOT = r'''<q-td :props="props">
                        <span v-if="props.row.stav" :class="[
                            'text-xs font-semibold px-2 py-0.5 rounded-full',
                            props.row.stav === 'OK' ? 'bg-green-100 text-green-800' :
                            props.row.stav === 'K deaktivaci' ? 'bg-red-100 text-red-800' :
                            props.row.stav === 'K reaktivaci' ? 'bg-emerald-100 text-emerald-800' :
                            props.row.stav === 'Pouze v sestavě' ? 'bg-amber-100 text-amber-800' :
                            'bg-gray-100 text-gray-700'
                        ]">{{ props.row.stav }}</span>
                    </q-td>'''

                    # ── Slot pro Charakteristiku s nápovědou (název střediska) ─
                    CHARAKT_SLOT = r'''<q-td :props="props">
                        {{ props.value }}
                        <q-tooltip v-if="props.row.charakteristika_nazev">
                            {{ props.row.charakteristika_nazev }}
                        </q-tooltip>
                    </q-td>'''

                    # ── Sloupce ───────────────────────────────────────────────
                    COLS_SYS = [
                        {'name': 'cislo',   'label': 'Číslo JIP', 'field': 'cislo',   'align': 'left', 'sortable': True},
                        {'name': 'jmeno',   'label': 'Jméno celé','field': 'jmeno',   'align': 'left', 'sortable': True},
                        {'name': 'email',   'label': 'E-mail',    'field': 'email',   'align': 'left'},
                        {'name': 'odd',     'label': 'Oddělení',  'field': 'odd',     'align': 'left', 'sortable': True},
                        {'name': 'aktivni', 'label': 'Aktivní',   'field': 'aktivni', 'align': 'center'},
                        {'name': 'stav',    'label': 'Stav',      'field': 'stav',    'align': 'left', 'sortable': True},
                    ]
                    COLS_CSV = [
                        {'name': 'cislo',    'label': 'Číslo JIP',   'field': 'cislo',    'align': 'left', 'sortable': True},
                        {'name': 'jmeno',    'label': 'Jméno',       'field': 'jmeno',    'align': 'left', 'sortable': True},
                        {'name': 'prijmeni', 'label': 'Příjmení',    'field': 'prijmeni', 'align': 'left', 'sortable': True},
                        {'name': 'charakteristika', 'label': 'Charakteristika', 'field': 'charakteristika', 'align': 'left', 'sortable': True},
                        {'name': 'nastup',   'label': 'Nástup',      'field': 'nastup',   'align': 'left'},
                        {'name': 'odchod',   'label': 'Odchod',      'field': 'odchod',   'align': 'left'},
                        {'name': 'pomer',    'label': 'Zam. poměr',  'field': 'pomer',    'align': 'center'},
                        {'name': 'vypoved',  'label': 'Ve výpovědi', 'field': 'vypoved',  'align': 'center'},
                        {'name': 'stav',     'label': 'Stav',        'field': 'stav',     'align': 'left', 'sortable': True},
                    ]
                    # Bez virtual-scrollu: měření výšek v Quasaru selhává při
                    # zoomu prohlížeče ≠ 100 % (prázdná tabulka). Stránkování
                    # je odolné a 200+ řádků zvládne bez problémů.
                    TBL_P = ('flat bordered dense cursor-pointer '
                             ':row-class="row => row._row_cls"')

                    # ── Pomocník: split panel ─────────────────────────────────
                    def _split(l_rows, r_rows,
                               l_label='🖥 Systém', r_label='📄 Sestava CSV'):
                        with ui.row().classes('w-full gap-0 flex-1'):
                            # Levá polovina
                            with ui.column().classes('flex-1 min-w-0 border-r-2 border-gray-200 pr-3'):
                                ui.label(l_label).classes(
                                    'font-bold text-gray-500 text-xs uppercase tracking-wide mb-1')
                                tbl_l = ui.table(
                                    columns=COLS_SYS,
                                    rows=_filtruj(l_rows, 'sys'),
                                    row_key='cislo',
                                    pagination=25,
                                ).classes('w-full shadow-sm rounded-xl').props(TBL_P)
                                tbl_l.add_slot('body-cell-stav', STAV_SLOT)
                                tbl_l.on('rowClick', lambda e: _on_row_click(e.args))
                                tbl_refs.append((tbl_l, l_rows, 'sys'))

                            # Pravá polovina
                            with ui.column().classes('flex-1 min-w-0 pl-3'):
                                ui.label(r_label).classes(
                                    'font-bold text-gray-500 text-xs uppercase tracking-wide mb-1')
                                tbl_r = ui.table(
                                    columns=COLS_CSV,
                                    rows=_filtruj(r_rows, 'csv'),
                                    row_key='cislo',
                                    pagination=25,
                                ).classes('w-full shadow-sm rounded-xl').props(TBL_P)
                                tbl_r.add_slot('body-cell-stav', STAV_SLOT)
                                tbl_r.add_slot('body-cell-charakteristika', CHARAKT_SLOT)
                                tbl_r.on('rowClick', lambda e: _on_row_click(e.args))
                                tbl_refs.append((tbl_r, r_rows, 'csv'))

                    # ── Akce: změna aktivity v DB ─────────────────────────────
                    async def _proved_zmenu_aktivity(email, nova_aktivita):
                        _u = intranet_data.ziskej_vsechny_uzivatele().get(email)
                        if not _u:
                            return False, 'Uživatel neexistuje.'
                        out = await asyncio.to_thread(
                            intranet_data.pridej_uprav_uzivatele,
                            email, _u['jmeno'], _u['prijmeni'], None, _u['role'],
                            _u.get('oddeleni'), _u['prava'], nova_aktivita,
                            _u.get('base_vacation', 160.0),
                            _u.get('carried_over_vacation', 0.0),
                            None, _u.get('manager_id', []),
                            [s['id'] for s in _u.get('spolecnosti', [])],
                            _u.get('datum_narozeni') or None,
                            _u.get('priznak_id') or None,
                        )
                        if isinstance(out, tuple) and len(out) == 2:
                            return out
                        return bool(out), ''

                    async def _refresh_po_akci():
                        intranet_data.invaliduj_cache_sprava()
                        await asyncio.to_thread(intranet_data.ziskej_vsechny_uzivatele)
                        try:
                            vykresli_seznam_uzivatelu.refresh()
                        except Exception:
                            pass
                        _vykresli.refresh()

                    # ── Akční tabulka pro K deaktivaci / K reaktivaci ─────────
                    def _akcni_tabulka(klice, akce_label, akce_color, btn_class,
                                       nova_aktivita_bool, log_verb, hromadne_label,
                                       btn_icon, duvod_fn=None):
                        if not klice:
                            ui.label('Žádné záznamy v této kategorii.') \
                              .classes('text-gray-400 italic text-center mt-8')
                            return

                        rows_akce = []
                        for k in sorted(klice):
                            d = system_map[k]
                            csv_r = csv_map.get(k)
                            if duvod_fn is not None:
                                duvod = duvod_fn(k)
                            elif csv_r is None:
                                duvod = '⚠ Chybí v sestavě'
                            else:
                                pomer_val = str(csv_r.get('je_v_zam_pomeru', '?')).strip()
                                duvod = f'V sestavě je_v_zam_pomeru = {pomer_val}'
                            rows_akce.append({
                                'cislo':   k,
                                'jmeno':   d.get('jmeno_cele', ''),
                                'email':   email_map.get(k, ''),
                                'odd':     d.get('oddeleni', '') or '',
                                'aktivni': '✓' if d.get('aktivni') else '✗',
                                'nastup':  (csv_r.get('nastup', '') if csv_r else ''),
                                'duvod':   duvod,
                            })

                        def _proved_hromadne():
                            with ui.dialog() as cdlg, \
                                 ui.card().classes('p-6 rounded-xl w-full max-w-md'):
                                ui.label(hromadne_label) \
                                  .classes('text-lg font-bold mb-3 text-gray-800')
                                ui.label(f'Bude se týkat {len(rows_akce)} uživatelů. '
                                         'Hromadnou akci nelze jedním klikem vrátit.') \
                                  .classes('text-sm text-gray-600 mb-4')

                                async def _proved():
                                    cdlg.close()
                                    uspech = 0; chyba = 0
                                    for r in list(rows_akce):
                                        em = r['email']
                                        if not em:
                                            chyba += 1; continue
                                        ok, _msg = await _proved_zmenu_aktivity(em, nova_aktivita_bool)
                                        if ok:
                                            uspech += 1
                                            intranet_logger.log_activity(
                                                user_name, "Správa uživatelů",
                                                f"{log_verb} uživatel (porovnání): {em}")
                                        else:
                                            chyba += 1
                                    ui.notify(
                                        f'Hotovo: {uspech} úspěch, {chyba} chyba.',
                                        type='positive' if chyba == 0 else 'warning',
                                        position='top',
                                    )
                                    await _refresh_po_akci()

                                with ui.row().classes('w-full justify-end gap-3'):
                                    ui.button('Zrušit', on_click=cdlg.close) \
                                      .classes('bg-gray-200 hover:bg-gray-300 text-gray-800 font-bold')
                                    ui.button('Provést', on_click=_proved) \
                                      .classes(f'{btn_class} text-white font-bold')
                            cdlg.open()

                        with ui.row().classes('w-full items-center justify-between mb-3'):
                            ui.label(f'Počet záznamů: {len(rows_akce)}') \
                              .classes('text-sm font-semibold text-gray-600')
                            ui.button(f'{akce_label} všechny', icon='done_all',
                                      on_click=_proved_hromadne) \
                              .classes(f'{btn_class} text-white font-bold')

                        # ── Stav řazení dle data nástupu ──────────────────────
                        sort_stav  = {'dir': None}   # None (původní) | 'asc' | 'desc'
                        hotovo_set = set()           # čísla JIP s již provedenou akcí

                        def _nastup_key(s):
                            s = (s or '').strip()
                            if not s:
                                return datetime.date.min
                            for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%d.%m.%y'):
                                try:
                                    return datetime.datetime.strptime(s, fmt).date()
                                except ValueError:
                                    continue
                            return datetime.date.min

                        def _serazene_radky():
                            zaklad = [r for r in rows_akce if _match_filter(r)]
                            d = sort_stav['dir']
                            if d == 'asc':
                                return sorted(zaklad, key=lambda r: _nastup_key(r['nastup']))
                            if d == 'desc':
                                return sorted(zaklad, key=lambda r: _nastup_key(r['nastup']),
                                              reverse=True)
                            return zaklad

                        def _toggle_sort():
                            sort_stav['dir'] = {None: 'asc', 'asc': 'desc',
                                                'desc': None}[sort_stav['dir']]
                            sort_ind.set_text({None: '⇅', 'asc': '▲',
                                               'desc': '▼'}[sort_stav['dir']])
                            _vykresli_radky.refresh()

                        # ── Hlavička sloupců ──────────────────────────────────
                        with ui.row().classes(
                            'w-full items-center gap-2 px-3 py-2 bg-gray-100 '
                            'rounded-t-lg font-semibold text-xs text-gray-600 '
                            'uppercase tracking-wide border border-gray-200'
                        ).style('flex-wrap:nowrap'):
                            ui.label('Číslo JIP').classes('w-28 shrink-0')
                            ui.label('Jméno / E-mail').classes('flex-1 min-w-0')
                            ui.label('Oddělení').classes('w-32 shrink-0')
                            with ui.row().classes(
                                'w-24 shrink-0 items-center gap-1 cursor-pointer '
                                'select-none hover:text-indigo-600'
                            ).on('click', _toggle_sort):
                                ui.label('Nástup')
                                sort_ind = ui.label('⇅').classes('text-gray-400')
                            ui.label('Aktivní').classes('w-20 text-center shrink-0')
                            ui.label('Důvod').classes('flex-1 min-w-0')
                            ui.label('Akce').classes('w-44 text-right shrink-0')

                        # ── Hotový badge (po úspěšné akci) ───────────────────
                        if nova_aktivita_bool:   # K reaktivaci → zelený výsledek
                            done_label   = 'Aktivován'
                            done_icon    = 'check_circle'
                            done_cls     = 'bg-emerald-100 text-emerald-700 border border-emerald-300'
                            done_row_cls = 'bg-emerald-50'
                        else:                    # K deaktivaci → červený výsledek
                            done_label   = 'Deaktivován'
                            done_icon    = 'block'
                            done_cls     = 'bg-red-100 text-red-700 border border-red-300'
                            done_row_cls = 'bg-red-50'

                        def _vykresli_hotovo(holder, row_el):
                            holder.clear()
                            with holder:
                                with ui.row().classes(
                                    f'w-full justify-center items-center gap-2 px-3 py-1 '
                                    f'rounded-lg font-bold text-sm {done_cls}'
                                ):
                                    ui.icon(done_icon, size='sm')
                                    ui.label(done_label)
                            row_el.classes(add=done_row_cls,
                                           remove='hover:bg-gray-50')

                        # ── Seznam řádků (řaditelný dle data nástupu) ─────────
                        @ui.refreshable
                        def _vykresli_radky():
                            for r in _serazene_radky():
                                je_hotovo = r['cislo'] in hotovo_set
                                row_el = ui.row().classes(
                                    'w-full items-center gap-2 px-3 py-2 '
                                    'border-t border-gray-100 text-sm'
                                    + ('' if je_hotovo else ' hover:bg-gray-50')
                                ).style('flex-wrap:nowrap')
                                if je_hotovo:
                                    row_el.classes(add=done_row_cls)
                                with row_el:
                                    ui.label(r['cislo']).classes(
                                        'w-28 shrink-0 font-mono text-gray-800')
                                    with ui.column().classes('flex-1 min-w-0 gap-0'):
                                        ui.label(r['jmeno']) \
                                          .classes('font-semibold text-gray-800 truncate')
                                        ui.label(r['email']) \
                                          .classes('text-xs text-gray-500 truncate')
                                    ui.label(r['odd']).classes(
                                        'w-32 shrink-0 text-gray-700 truncate')
                                    ui.label(r['nastup'] or '—').classes(
                                        'w-24 shrink-0 text-gray-700 truncate')
                                    _akt_cls = ('text-green-600' if r['aktivni'] == '✓'
                                                else 'text-red-600')
                                    ui.label(r['aktivni']).classes(
                                        f'w-20 shrink-0 text-center font-bold {_akt_cls}')
                                    ui.label(r['duvod']).classes(
                                        'flex-1 min-w-0 text-xs text-gray-600 truncate')

                                    btn_holder = ui.row().classes(
                                        'w-44 shrink-0 justify-end items-center gap-0')

                                    if je_hotovo:
                                        _vykresli_hotovo(btn_holder, row_el)
                                        continue

                                    async def _do_akce(em=r['email'],
                                                       jm=r['jmeno'],
                                                       cis=r['cislo'],
                                                       holder=btn_holder,
                                                       rowel=row_el):
                                        if not em:
                                            ui.notify('U záznamu chybí e-mail.',
                                                      type='negative', position='top')
                                            return
                                        ok, msg = await _proved_zmenu_aktivity(
                                            em, nova_aktivita_bool)
                                        if ok:
                                            hotovo_set.add(cis)
                                            intranet_logger.log_activity(
                                                user_name, "Správa uživatelů",
                                                f"{log_verb} uživatel (porovnání): {em}")
                                            ui.notify(f"{log_verb}: {jm}",
                                                      type='positive', position='top')
                                            _vykresli_hotovo(holder, rowel)
                                            # Tichá invalidace cache + refresh hl. seznamu uživatelů
                                            intranet_data.invaliduj_cache_sprava()
                                            await asyncio.to_thread(
                                                intranet_data.ziskej_vsechny_uzivatele)
                                            try:
                                                vykresli_seznam_uzivatelu.refresh()
                                            except Exception:
                                                pass
                                        else:
                                            ui.notify(f'Chyba: {msg}',
                                                      type='negative', position='top')

                                    with btn_holder:
                                        ui.button(akce_label, icon=btn_icon,
                                                  on_click=_do_akce) \
                                          .props('dense unelevated no-caps') \
                                          .classes(f'w-full {btn_class} '
                                                   f'text-white font-bold')

                        with ui.column().classes(
                            'w-full gap-0 border-l border-r border-b border-gray-200 '
                            'rounded-b-lg bg-white shrink-0'
                        ).style('max-height:60vh;overflow-y:auto'):
                            _vykresli_radky()
                        filter_hooks.append(_vykresli_radky.refresh)

                    # ── Tabulka pro založení nových účtů (V sestavě, ne v systému) ─
                    def _zalozeni_tabulka(klice):
                        if not klice:
                            ui.label('Žádní noví zaměstnanci k založení.') \
                              .classes('text-gray-400 italic text-center mt-8')
                            return

                        hotovo_set = set()   # čísla JIP s již založeným účtem

                        # ── Pomocné parsery ───────────────────────────────────
                        def _parse_datum(s):
                            s = (s or '').strip()
                            if not s or s in ('0', '00.00.0000', '0000-00-00'):
                                return None
                            for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%d.%m.%y'):
                                try:
                                    return datetime.datetime.strptime(s, fmt) \
                                                   .strftime('%Y-%m-%d')
                                except ValueError:
                                    continue
                            return None

                        def _gen_heslo():
                            import secrets, string
                            znaky = (secrets.choice(string.ascii_uppercase),
                                     secrets.choice(string.ascii_lowercase),
                                     secrets.choice(string.digits))
                            zbytek = [secrets.choice(string.ascii_letters + string.digits)
                                      for _ in range(7)]
                            sez = list(znaky) + zbytek
                            secrets.SystemRandom().shuffle(sez)
                            return ''.join(sez)

                        # ── Náhledový dialog před založením účtu ───────────────
                        def _otevri_nahled(k, holder, row_el):
                            r        = csv_map[k]
                            priznak, cislo = intranet_narozeniny._parsuj_cislo_jip(k)
                            priznak  = (priznak or '').upper()
                            jmeno    = (r.get('Jmeno') or '').strip()
                            prijmeni = (r.get('Prijmeni') or '').strip()
                            nastup   = (r.get('nastup') or '').strip()
                            dn       = _parse_datum(r.get('DatumNarozeni'))
                            email_def = intranet_narozeniny._vygeneruj_email(jmeno, prijmeni)

                            pr_db        = intranet_data.ziskej_vsechny_priznaky()
                            pr_map       = {(p['nazev'] or '').upper(): p for p in pr_db}
                            pr_names     = sorted(pr_map.keys())
                            if priznak and priznak not in pr_names:
                                pr_names = [priznak] + pr_names
                            role_opts = list(role) + ['Bez role']
                            odd_opts  = list(oddeleni.keys()) + ['Bez oddělení']

                            with ui.dialog() as ndlg, \
                                 ui.card().classes('p-6 rounded-2xl shadow-2xl w-full max-w-2xl') \
                                          .style('max-height:90vh;overflow-y:auto'):
                                ui.label('Náhled – založení účtu') \
                                  .classes('text-xl font-extrabold text-gray-800 mb-1')
                                ui.label(f'Číslo JIP „{k}" se rozloží na příznak + osobní číslo. '
                                         'Všechna pole lze před založením upravit; u číselníků '
                                         'vyber z nabídky z databáze.') \
                                  .classes('text-sm text-gray-500 mb-4')

                                def _pole(popis):
                                    row = ui.row().classes('w-full items-center gap-3 py-1')
                                    with row:
                                        ui.label(popis).classes(
                                            'w-52 shrink-0 text-xs uppercase tracking-wide '
                                            'text-gray-500 font-semibold')
                                    return row

                                with ui.column().classes(
                                    'w-full gap-0 bg-gray-50 rounded-xl p-4 mb-4 '
                                    'border border-gray-200'):
                                    with _pole('Příznak (2 písmena)'):
                                        priznak_in = ui.select(
                                            pr_names, value=priznak if priznak else None,
                                            with_input=True,
                                            new_value_mode='add-unique') \
                                          .props('outlined dense').classes('flex-1')
                                    with _pole('Osobní číslo (iduser)'):
                                        os_in = ui.number(value=cislo, step=1, min=1) \
                                          .props('outlined dense').classes('flex-1')
                                    with _pole('Jméno'):
                                        jmeno_in = ui.input(value=jmeno) \
                                          .props('outlined dense').classes('flex-1')
                                    with _pole('Příjmení'):
                                        prijmeni_in = ui.input(value=prijmeni) \
                                          .props('outlined dense').classes('flex-1')
                                    with _pole('Datum narození'):
                                        datum_in = ui.input(value=dn or '',
                                                            placeholder='RRRR-MM-DD') \
                                          .props('outlined dense mask="####-##-##" fill-mask') \
                                          .classes('flex-1')
                                    with _pole('Nástup (jen pro přehled)'):
                                        ui.label(nastup or '—') \
                                          .classes('flex-1 text-sm font-mono text-gray-700')
                                    with _pole('Role'):
                                        role_in = ui.select(role_opts, value='Bez role',
                                                            with_input=True) \
                                          .props('outlined dense').classes('flex-1')
                                    with _pole('Oddělení'):
                                        odd_in = ui.select(odd_opts, value='Bez oddělení',
                                                          with_input=True) \
                                          .props('outlined dense').classes('flex-1')
                                    with _pole('Společnosti'):
                                        spol_in = ui.select(spolecnosti_options, multiple=True,
                                                           value=[]) \
                                          .props('outlined dense use-chips clearable') \
                                          .classes('flex-1')
                                    with _pole('Stav účtu'):
                                        akt_in = ui.switch('Aktivní', value=True) \
                                          .classes('flex-1')
                                    with _pole('Základ dovolené (h)'):
                                        zaklad_in = ui.number(value=160.0, step=0.5) \
                                          .props('outlined dense').classes('flex-1')

                                email_in = ui.input('E-mail (přihlašovací jméno)',
                                                    value=email_def) \
                                             .props('outlined dense').classes('w-full mb-2')
                                heslo_in = ui.input('Počáteční heslo', value=_gen_heslo()) \
                                             .props('outlined dense').classes('w-full')
                                ui.label('Heslo si poznamenejte – po založení ho lze změnit '
                                         'přes „Změnit heslo". Min. 8 znaků, velké i malé '
                                         'písmeno a číslo.') \
                                  .classes('text-xs text-gray-400 mt-1 mb-3')

                                async def _zaloz():
                                    # Načtení a validace upravených hodnot
                                    cislo_val = None
                                    try:
                                        if os_in.value not in (None, ''):
                                            cislo_val = int(os_in.value)
                                    except (TypeError, ValueError):
                                        cislo_val = None
                                    if cislo_val is None:
                                        ui.notify('Vyplňte platné osobní číslo (iduser).',
                                                  type='negative', position='top')
                                        return
                                    priznak_val = str(priznak_in.value or '').strip().upper()
                                    if len(priznak_val) != 2 or not priznak_val.isalpha():
                                        ui.notify('Příznak musí být přesně 2 písmena.',
                                                  type='warning', position='top')
                                        return
                                    jmeno_val    = (jmeno_in.value or '').strip()
                                    prijmeni_val = (prijmeni_in.value or '').strip()
                                    dn_val       = (datum_in.value or '').strip() or None
                                    role_val     = role_in.value or 'Bez role'
                                    odd_val      = odd_in.value or 'Bez oddělení'
                                    spol_ids     = spol_in.value or []
                                    aktivni_val  = bool(akt_in.value)
                                    try:
                                        zaklad_val = float(zaklad_in.value)
                                    except (TypeError, ValueError):
                                        zaklad_val = 160.0
                                    em = (email_in.value or '').strip().lower()
                                    if not em:
                                        ui.notify('Vyplňte e-mail.', type='warning',
                                                  position='top')
                                        return
                                    if not intranet_data.heslo_je_silne(heslo_in.value or ''):
                                        ui.notify('Heslo musí mít min. 8 znaků, velké i malé '
                                                  'písmeno a číslo.', type='warning',
                                                  position='top')
                                        return

                                    db_now = await asyncio.to_thread(
                                        intranet_data.ziskej_vsechny_uzivatele)
                                    if em in db_now:
                                        ui.notify(f'E-mail {em} už v systému existuje – '
                                                  'účet nebude založen.',
                                                  type='negative', position='top')
                                        return

                                    # Příznak – případně založit nový
                                    pr_exist = pr_map.get(priznak_val)
                                    priznak_id = pr_exist['id'] if pr_exist else None
                                    if priznak_id is None:
                                        okp, resp = await asyncio.to_thread(
                                            intranet_data.pridej_priznak, priznak_val, '#6366f1')
                                        if not okp:
                                            ui.notify(f'Nepodařilo se založit příznak '
                                                      f'„{priznak_val}": {resp}',
                                                      type='negative', position='top')
                                            return
                                        try:
                                            priznak_id = int(resp)
                                        except (TypeError, ValueError):
                                            priznak_id = None

                                    status, msg = await asyncio.to_thread(
                                        intranet_data.pridej_uprav_uzivatele,
                                        em, jmeno_val, prijmeni_val, heslo_in.value, role_val,
                                        odd_val, '', aktivni_val, zaklad_val, 0.0,
                                        cislo_val, [], spol_ids, dn_val, priznak_id)

                                    if status:
                                        hotovo_set.add(k)
                                        intranet_logger.log_activity(
                                            user_name, "Správa uživatelů",
                                            f"Založen účet (porovnání): {em} "
                                            f"[JIP {priznak_val}{str(cislo_val).zfill(6)}, "
                                            f"os.č. {cislo_val}]")
                                        ui.notify(f'Účet založen: {jmeno_val} {prijmeni_val} '
                                                  f'({em})',
                                                  type='positive', position='top')
                                        ndlg.close()
                                        _oznac_zalozeno(holder, row_el)
                                        intranet_data.invaliduj_cache_sprava()
                                        await asyncio.to_thread(
                                            intranet_data.ziskej_vsechny_uzivatele)
                                        try:
                                            vykresli_seznam_uzivatelu.refresh()
                                        except Exception:
                                            pass
                                    else:
                                        ui.notify(f'Chyba při zakládání: {msg}',
                                                  type='negative', position='top')

                                with ui.row().classes('w-full justify-end gap-3 mt-2'):
                                    ui.button('Zrušit', on_click=ndlg.close) \
                                      .classes('bg-gray-200 hover:bg-gray-300 '
                                               'text-gray-800 font-bold')
                                    ui.button('Vytvořit účet', icon='person_add',
                                              on_click=_zaloz) \
                                      .classes('bg-blue-600 hover:bg-blue-700 '
                                               'text-white font-bold')
                            ndlg.open()

                        def _oznac_zalozeno(holder, row_el):
                            holder.clear()
                            with holder:
                                with ui.row().classes(
                                    'w-full justify-center items-center gap-2 px-3 py-1 '
                                    'rounded-lg font-bold text-sm '
                                    'bg-blue-100 text-blue-700 border border-blue-300'):
                                    ui.icon('how_to_reg', size='sm')
                                    ui.label('Účet založen')
                            row_el.classes(add='bg-blue-50', remove='hover:bg-gray-50')

                        rows_new = []
                        for k in sorted(klice):
                            r = csv_map[k]
                            priznak, cislo = intranet_narozeniny._parsuj_cislo_jip(k)
                            rows_new.append({
                                'cislo':    k,
                                'priznak':  (priznak or '').upper(),
                                'os_cislo': cislo,
                                'jmeno':    (r.get('Jmeno') or '').strip(),
                                'prijmeni': (r.get('Prijmeni') or '').strip(),
                                'charakteristika': (r.get('nakladovy_okruh') or '').strip(),
                                'nastup':   (r.get('nastup') or '').strip(),
                                'email':    intranet_narozeniny._vygeneruj_email(
                                                (r.get('Jmeno') or '').strip(),
                                                (r.get('Prijmeni') or '').strip()),
                            })

                        with ui.row().classes('w-full items-center mb-3'):
                            ui.label(f'Počet nových zaměstnanců: {len(rows_new)}') \
                              .classes('text-sm font-semibold text-gray-600')

                        # Hlavička
                        with ui.row().classes(
                            'w-full items-center gap-2 px-3 py-2 bg-gray-100 '
                            'rounded-t-lg font-semibold text-xs text-gray-600 '
                            'uppercase tracking-wide border border-gray-200'
                        ).style('flex-wrap:nowrap'):
                            ui.label('Číslo JIP').classes('w-24 shrink-0')
                            ui.label('Příznak → os. č.').classes('w-28 shrink-0')
                            ui.label('Jméno / Příjmení').classes('w-56 shrink-0')
                            with ui.label('Charakteristika').classes('w-28 shrink-0'):
                                ui.tooltip('Nákladový okruh – kód střediska. '
                                           'Najetím na hodnotu se zobrazí název střediska.')
                            ui.label('Nástup').classes('w-24 shrink-0')
                            ui.label('E-mail (návrh)').classes('flex-1 min-w-0')
                            ui.label('Akce').classes('w-40 text-right shrink-0')

                        @ui.refreshable
                        def _vykresli_radky_new():
                            for r in rows_new:
                                if not _match_filter(r):
                                    continue
                                row_el = ui.row().classes(
                                    'w-full items-center gap-2 px-3 py-2 '
                                    'border-t border-gray-100 hover:bg-gray-50 text-sm'
                                ).style('flex-wrap:nowrap')
                                with row_el:
                                    ui.label(r['cislo']).classes(
                                        'w-24 shrink-0 font-mono text-gray-800')
                                    _os = (str(r['os_cislo']) if r['os_cislo'] is not None
                                           else '⚠')
                                    ui.label(f"{r['priznak']} → {_os}").classes(
                                        'w-28 shrink-0 font-mono text-gray-700 truncate')
                                    with ui.column().classes('w-56 min-w-0 shrink-0 gap-0'):
                                        ui.label(f"{r['jmeno']} {r['prijmeni']}".strip()
                                                 or '—') \
                                          .classes('font-semibold text-gray-800 truncate')
                                    with ui.label(r['charakteristika'] or '—').classes(
                                            'w-28 shrink-0 text-gray-700 truncate'):
                                        _nazev_okr = intranet_okruhy.nazev_okruhu(
                                            r['charakteristika'])
                                        if _nazev_okr:
                                            ui.tooltip(_nazev_okr)
                                    ui.label(r['nastup'] or '—').classes(
                                        'w-24 shrink-0 text-gray-700 truncate')
                                    ui.label(r['email'] or '⚠ chybí jméno/příjmení').classes(
                                        'flex-1 min-w-0 text-xs text-gray-500 truncate')

                                    btn_holder = ui.row().classes(
                                        'w-40 shrink-0 justify-end items-center gap-0')
                                    with btn_holder:
                                        ui.button('Založit účet', icon='person_add',
                                                  on_click=lambda _=None, kk=r['cislo'],
                                                           h=btn_holder, rel=row_el:
                                                           _otevri_nahled(kk, h, rel)) \
                                          .props('dense unelevated no-caps') \
                                          .classes('w-full bg-blue-600 hover:bg-blue-700 '
                                                   'text-white font-bold')

                        with ui.column().classes(
                            'w-full gap-0 border-l border-r border-b border-gray-200 '
                            'rounded-b-lg bg-white shrink-0'
                        ).style('max-height:60vh;overflow-y:auto'):
                            _vykresli_radky_new()
                        filter_hooks.append(_vykresli_radky_new.refresh)

                    # ── Záložky ───────────────────────────────────────────────
                    _default_tab = ('navrhy' if navrhy_set else
                                    'navrhy_deakt' if navrhy_deakt_set else 'ok')
                    _platne_taby = {'ok', 'navrhy', 'navrhy_deakt'}
                    if stav.get('tab') not in _platne_taby:
                        stav['tab'] = _default_tab
                    with ui.tabs().classes('w-full border-b border-gray-200') as tabs:
                        ui.tab(name='ok',           label=f'✅ OK ({len(ok_set)})')
                        ui.tab(name='navrhy',       label=f'💡 Návrhy k aktivaci ({len(navrhy_set)})')
                        ui.tab(name='navrhy_deakt', label=f'⛔ Návrhy k deaktivaci ({len(navrhy_deakt_set)})')
                    tabs.set_value(stav['tab'])


                    # Obsah aktivní záložky vykreslujeme INLINE jako přímého sourozence
                    # lišty záložek (stejný kontext jako karty/taby, které se zobrazují
                    # správně). Přepínání = stav['tab'] + _vykresli.refresh().
                    # Vyhneme se tak Quasar „panel tracku", který posouval obsah doprava.
                    _t = stav['tab']
                    with ui.column().classes('w-full mt-2 shrink-0'):
                        if _t == 'ok':
                            _split(
                                [r for r in sys_rows_vse if r['stav'] == 'OK'],
                                [r for r in csv_rows_vse if r['stav'] == 'OK'],
                            )
                        elif _t == 'navrhy':
                            ui.label(
                                f'Noví zaměstnanci s nástupem v aktuálním nebo minulém '
                                f'týdnu ({_pocatek_okna.strftime("%d.%m.%Y")} – '
                                f'{_konec_okna.strftime("%d.%m.%Y")}), '
                                'kteří jsou v pracovním poměru (je_v_zam_pomeru = 1) '
                                'a dosud nejsou v systému. Tlačítkem „Založit účet" se '
                                'po náhledu vytvoří účet (Číslo JIP se rozloží na '
                                'příznak + osobní číslo).'
                            ).classes('text-sm text-blue-700 italic mb-2')
                            _zalozeni_tabulka(navrhy_set)
                        elif _t == 'navrhy_deakt':
                            ui.label(
                                f'Zaměstnanci, kteří odešli nebo odejdou v aktuálním '
                                f'nebo minulém týdnu ({_pocatek_okna.strftime("%d.%m.%Y")} – '
                                f'{_konec_okna.strftime("%d.%m.%Y")}) a v systému jsou '
                                'stále aktivní. Tlačítkem je deaktivujete.'
                            ).classes('text-sm text-rose-700 italic mb-2')
                            _akcni_tabulka(
                                navrhy_deakt_set, akce_label='Deaktivovat',
                                akce_color='red',
                                btn_class='bg-rose-600 hover:bg-rose-700',
                                nova_aktivita_bool=False, log_verb='Deaktivován',
                                hromadne_label='Hromadná deaktivace (dle odchodu)',
                                btn_icon='block',
                                duvod_fn=lambda k: (
                                    f"Odchod: {(csv_map[k].get('odchod') or '—').strip()}"),
                            )

                    def _on_tab(e):
                        stav['tab'] = (getattr(e, 'value', None) or tabs.value
                                       or _default_tab)
                        _vykresli.refresh()
                    tabs.on_value_change(_on_tab)

                async def _nahrat(e):
                    try:
                        raw = await e.file.read()
                        for enc in ('utf-8-sig', 'utf-8', 'cp1250', 'latin-1'):
                            try:
                                text = raw.decode(enc)
                                break
                            except UnicodeDecodeError:
                                continue
                        reader = csv.DictReader(io.StringIO(text), delimiter=';')
                        stav['radky'] = list(reader)
                        stav['nacteno'] = True
                        popis_el.set_visibility(False)
                        upload_el.set_visibility(False)
                        karta_pv.style('width:95vw;max-width:2100px;max-height:90vh;'
                                       'overflow-y:auto')
                    except Exception as ex:
                        ui.notify(f'Chyba při čtení souboru: {ex}', type='negative')
                        return
                    # Vykreslení mimo try – chyba renderu se ukáže v banneru,
                    # ne jako falešná „chyba čtení souboru“.
                    _vykresli.refresh()

                upload_el = ui.upload(label='Nahrát sestavu CSV', on_upload=_nahrat,
                                      auto_upload=True, max_files=1) \
                              .props('accept=.csv flat').classes('w-full')

                _vykresli()

            dlg_pv.open()

        # ── Dialog: Kdo má které právo? ───────────────────────────────────
        def _dialog_kdo_ma_pravo():
            # Katalog práv seřazený podle kategorie + názvu → přehledné volby.
            _opts = {
                k: f"{v.get('nazev', k)}  ·  {v.get('kategorie', '')}"
                for k, v in sorted(
                    zakladni_prava.items(),
                    key=lambda kv: (kv[1].get('kategorie', ''), kv[1].get('nazev', '')),
                )
            }
            stav = {'pravo': None, 'data': [], 'nacita': False, 'filtr': ''}

            def _chip(text, cls, ikona):
                with ui.row().classes('items-center gap-1 rounded-full px-2 py-0.5 border ' + cls):
                    ui.icon(ikona, size='0.9rem')
                    ui.label(text).classes('text-[11px] font-semibold whitespace-nowrap')

            def _radek_uziv(u):
                neaktivni = not u['aktivni']
                base = 'w-full items-center gap-3 p-3 rounded-xl border flex-nowrap '
                base += ('bg-gray-50 border-gray-200 opacity-60' if neaktivni
                         else 'bg-white border-slate-200')
                with ui.row().classes(base):
                    ui.icon('account_circle', size='1.8rem') \
                      .classes('shrink-0 ' + ('text-gray-400' if neaktivni else 'text-indigo-400'))
                    with ui.column().classes('gap-0 min-w-0 flex-1'):
                        with ui.row().classes('items-center gap-2 flex-nowrap min-w-0'):
                            ui.label(u['jmeno'] or '(bez jména)') \
                              .classes('font-bold text-gray-800 text-sm truncate')
                            if neaktivni:
                                ui.label('neaktivní') \
                                  .classes('shrink-0 text-[10px] font-bold text-red-500 bg-red-50 rounded px-1.5 py-0.5')
                        if u['email']:
                            ui.label(u['email']).classes('text-xs text-gray-400 truncate')
                    with ui.row().classes('items-center gap-1.5 shrink-0 flex-wrap justify-end max-w-[55%]'):
                        if u['primo']:
                            _chip('Přímo', 'bg-emerald-50 text-emerald-700 border-emerald-200', 'person')
                        for poz in u['pozice']:
                            _chip(f'Pozice: {poz}', 'bg-blue-50 text-blue-700 border-blue-200', 'badge')
                        for odd in u['oddeleni']:
                            _chip(f'Oddělení: {odd}', 'bg-amber-50 text-amber-700 border-amber-200', 'groups')

            with ui.dialog() as dlg_km, \
                 ui.card().classes('p-6 rounded-2xl shadow-2xl flex flex-col gap-4') \
                          .style('width:760px;max-width:95vw;max-height:90vh;'
                                 'overflow-y:auto;overflow-x:hidden'):

                with ui.row().classes('w-full items-center justify-between shrink-0'):
                    with ui.row().classes('items-center gap-3'):
                        ui.icon('policy', size='1.8rem') \
                          .classes('text-indigo-600 bg-indigo-50 p-2 rounded-xl')
                        with ui.column().classes('gap-0'):
                            ui.label('Kdo má které právo?') \
                              .classes('text-xl font-extrabold text-gray-800 leading-tight')
                            ui.label('Vyberte právo a uvidíte všechny, komu je přiřazeno — přímo, '
                                     'přes pracovní pozici nebo zděděno z oddělení.') \
                              .classes('text-xs text-gray-500')
                    ui.button(icon='close', on_click=dlg_km.close) \
                      .props('flat round dense').classes('text-gray-500')

                vyber = ui.select(_opts, label='Vyberte právo…', with_input=True) \
                          .classes('w-full shrink-0').props('outlined dense clearable')

                filtr_uziv = ui.input(placeholder='Filtrovat uživatele podle jména…') \
                               .classes('w-full shrink-0').props('outlined dense clearable prepend-icon=search')
                filtr_uziv.set_visibility(False)

                @ui.refreshable
                def _vysledky():
                    pr = stav['pravo']
                    if not pr:
                        with ui.column().classes('w-full items-center justify-center py-12 gap-2 text-gray-400'):
                            ui.icon('touch_app', size='2.5rem')
                            ui.label('Vyberte právo z nabídky výše.').classes('text-sm italic')
                        return
                    if stav['nacita']:
                        with ui.row().classes('w-full items-center justify-center py-12 gap-3 text-gray-500'):
                            ui.spinner(size='lg')
                            ui.label('Načítám…')
                        return

                    meta = zakladni_prava.get(pr, {})
                    with ui.row().classes('w-full items-start gap-3 p-4 rounded-xl bg-slate-50 border border-slate-200'):
                        ui.icon(meta.get('ikona', 'label'), size='1.6rem') \
                          .classes('text-indigo-600 shrink-0 mt-0.5')
                        with ui.column().classes('gap-1 min-w-0 flex-1'):
                            with ui.row().classes('items-center gap-2 flex-wrap'):
                                ui.label(meta.get('nazev', pr)).classes('font-extrabold text-slate-800')
                                if meta.get('kategorie'):
                                    ui.label(meta['kategorie']) \
                                      .classes('text-[11px] font-bold text-indigo-700 bg-indigo-100 rounded-full px-2 py-0.5')
                                ui.label(pr).classes('text-[11px] font-mono text-gray-400')
                            if meta.get('popis'):
                                ui.label(meta['popis']).classes('text-xs text-gray-500 leading-snug')

                    data = stav['data']
                    if not data:
                        with ui.column().classes('w-full items-center justify-center py-10 gap-2 text-gray-400'):
                            ui.icon('person_off', size='2.2rem')
                            ui.label('Toto právo nemá přiřazeno žádný uživatel.').classes('text-sm italic')
                        return

                    f = (stav['filtr'] or '').lower()
                    vybrani = [u for u in data if f in u['jmeno'].lower()] if f else data
                    akt = sum(1 for u in data if u['aktivni'])
                    ui.label(f"Celkem {len(data)} uživatelů · {akt} aktivních"
                             + (f" · zobrazeno {len(vybrani)}" if f else "")) \
                      .classes('text-sm font-bold text-slate-600')

                    with ui.column().classes('w-full gap-2'):
                        for u in vybrani:
                            _radek_uziv(u)

                async def _on_select(e):
                    pr = e.value
                    stav['pravo'] = pr
                    stav['filtr'] = ''
                    filtr_uziv.set_value('')
                    if not pr:
                        stav['data'] = []
                        filtr_uziv.set_visibility(False)
                        _vysledky.refresh()
                        return
                    stav['nacita'] = True
                    _vysledky.refresh()
                    data = await asyncio.to_thread(intranet_data.ziskej_uzivatele_pravo_detail, pr)
                    stav['data'] = data
                    stav['nacita'] = False
                    filtr_uziv.set_visibility(bool(data))
                    _vysledky.refresh()

                def _on_filtr(e):
                    stav['filtr'] = e.value or ''
                    _vysledky.refresh()

                vyber.on_value_change(_on_select)
                filtr_uziv.on_value_change(_on_filtr)
                # Vlastní scroll oblast výsledků s PŘÍMOU max výškou (ne flex-1).
                # flex-1 = flex-basis:0 → v kartě bez pevné výšky box zkolaboval
                # na 0 a seznam nebyl vidět (šel až po extrémním oddálení).
                # `max-h-[65vh]` drží flex-basis auto → box se roztáhne dle obsahu
                # a nad limitem scrolluje.
                with ui.column().classes('w-full max-h-[65vh] overflow-y-auto gap-3 pr-1'):
                    _vysledky()

            dlg_km.open()

        # ── Dialog: Export matice práv (viz _matice_zip_sync nahoře) ──────
        async def _dialog_export_matice():
            if not _je_superadmin:
                ui.notify('Matici práv smí stáhnout jen administrátor.', type='negative')
                return
            katalog = intranet_prava.ziskej_kompletni_seznam_prav(oddeleni, typy_v)
            kategorie_opts = sorted({v.get('kategorie', '') for v in katalog.values() if v.get('kategorie')})
            odd_opts = sorted(oddeleni.keys())
            with ui.dialog() as dlg, ui.card().classes('w-[560px] max-w-full p-5 gap-3'):
                ui.label('Export matice práv').classes('text-lg font-bold text-gray-800')
                ui.label('Kdo má které právo a odkud plyne. Výstup je šifrovaný ZIP '
                         '(heslo drží IT).').classes('text-xs text-gray-500')
                kat_in = ui.select(kategorie_opts, multiple=True,
                                   label='Kategorie práv (prázdné = všechny)') \
                           .classes('w-full').props('dense outlined use-chips')
                odd_in = ui.select(odd_opts, multiple=True,
                                   label='Oddělení (prázdné = všechna)') \
                           .classes('w-full').props('dense outlined use-chips')
                akt_in = ui.checkbox('Jen aktivní uživatelé', value=True)
                ded_in = ui.checkbox('Zahrnout zděděná práva (role, oddělení)', value=True)
                plochy_in = ui.checkbox('Přidat plochý seznam (pro kontingenční tabulku)', value=False)
                stav_lbl = ui.label('').classes('text-xs text-gray-500')

                async def _spust():
                    stav_lbl.text = 'Připravuji export…'
                    try:
                        matice = await asyncio.to_thread(intranet_data.ziskej_matici_prav)
                        cesta, jmeno, poc_u, poc_p = await asyncio.to_thread(
                            _matice_zip_sync, matice, katalog, list(kat_in.value or []),
                            list(odd_in.value or []), akt_in.value, ded_in.value,
                            plochy_in.value, user_name,
                        )
                    except ValueError as e:
                        stav_lbl.text = ''
                        ui.notify(str(e), type='warning')
                        return
                    except RuntimeError as e:
                        stav_lbl.text = ''
                        ui.notify(str(e), type='negative')
                        return
                    # Přes HTTP, ne WebSocket — ui.download.content padá nad ~1 MB.
                    ui.download.file(cesta, jmeno)
                    intranet_logger.log_activity(
                        user_name, 'Export práv',
                        f'Export matice práv ({poc_u} uživatelů × {poc_p} práv)')
                    stav_lbl.text = ''
                    dlg.close()

                with ui.row().classes('w-full justify-end gap-2 mt-1'):
                    ui.button('Zrušit', on_click=dlg.close).props('flat color=grey')
                    ui.button('Stáhnout ZIP', icon='lock', on_click=_spust) \
                      .props('color=indigo-7 unelevated')
            dlg.open()

        # ── Řádek: filtr + tlačítka ───────────────────────────────────────
        with ui.row().classes('w-full items-center gap-3 mb-6'):
            filtr_input = ui.input('Hledat uživatele...').classes('flex-1 max-w-md').props('outlined dense clearable debounce=400').props('prepend-icon=search')
            ui.space()
            ui.button('Kdo má právo?', icon='policy', on_click=_dialog_kdo_ma_pravo) \
              .props('unelevated no-caps').classes('bg-emerald-600 hover:bg-emerald-700 text-white font-semibold rounded-lg shadow-sm')
            ui.button('Porovnání sestavy', icon='compare_arrows', on_click=_dialog_porovnani) \
              .props('unelevated no-caps').classes('bg-indigo-600 hover:bg-indigo-700 text-white font-semibold rounded-lg shadow-sm')
            ui.button('Export CSV', icon='download', on_click=_dialog_export_csv) \
              .props('outline no-caps').classes('text-slate-700 font-semibold rounded-lg')
            if _je_superadmin:
                ui.button('Matice práv', icon='grid_on', on_click=_dialog_export_matice) \
                  .props('outline no-caps').classes('text-slate-700 font-semibold rounded-lg')

        # Bez hledání se vykreslí jen prvních _LIMIT_UZIV uživatelů — jinak by se
        # při vstupu do sekce stavěly stovky expansion komponent najednou (pomalé
        # první vykreslení). Hledání seznam zúží, takže ukáže všechny shody.
        _LIMIT_UZIV = 40
        _zobraz_vse = {'v': False}

        @ui.refreshable
        def vykresli_seznam_uzivatelu():
            hledany_text = filtr_input.value.lower() if filtr_input.value else ""
            aktualni_db = intranet_data.ziskej_vsechny_uzivatele()
            # Mapa id→jméno (zruší O(n²) dohledávání nadřízených v cyklu níže).
            id_to_jmeno = {u['id']: u['jmeno_cele'] for u in aktualni_db.values()}

            # Obsah karty uživatele se staví LÍNĚ až při prvním rozbalení – jinak
            # by se pro stovky uživatelů vykreslovaly tisíce prvků (chipy práv,
            # tlačítka, dialogy) hned při otevření sekce → dlouhé vykreslování.
            def _build_user_body(e_m, d, container):
                _pz_nazev = d.get('priznak_nazev', '')
                _pz_barva = d.get('priznak_barva', '')
                with container:
                    with ui.row().classes('w-full justify-between items-start p-4'):
                        with ui.column().classes('flex-1 gap-2'):
                            with ui.row().classes('items-center gap-3'):
                                ui.label(f"Osobní číslo: {d['id']}").classes('font-bold text-gray-700')
                                if _pz_nazev:
                                    ui.badge(_pz_nazev).props(f'style="background:{_pz_barva};color:#fff"')

                            nazvy_spolecnosti = [s['nazev'] for s in d.get('spolecnosti', [])]
                            txt_spolecnosti = ", ".join(nazvy_spolecnosti) if nazvy_spolecnosti else "Nepřiřazeno"
                            ui.label(f"Společnosti: {txt_spolecnosti}").classes('font-bold text-gray-700')

                            man_jmena_list = [id_to_jmeno.get(mid, "Neznámý") for mid in d.get('manager_id', [])]
                            man_jmeno_text = ", ".join(man_jmena_list) if man_jmena_list else "Nenastaven"
                            ui.label(f"Přímí nadřízení: {man_jmeno_text}").classes('font-bold text-gray-700')

                            ui.label(f"Oddělení: {d.get('oddeleni', 'Bez oddělení')}").classes('font-bold text-gray-700')

                            inherited_odd_prava = set()
                            if d.get('oddeleni', 'Bez oddělení') != 'Bez oddělení':
                                for o in d.get('oddeleni').split(','):
                                    o = o.strip()
                                    if o in oddeleni and oddeleni[o]:
                                        inherited_odd_prava.update(p.strip() for p in oddeleni[o]['prava'].split(',') if p.strip())
                            if inherited_odd_prava:
                                with ui.row().classes('items-center gap-2 mt-3'):
                                    ui.icon('groups', size='18px').classes('text-violet-500')
                                    ui.label('Práva zděděná z oddělení').classes('text-sm font-bold text-gray-600')
                                    ui.label(str(len(inherited_odd_prava))).classes('text-xs font-bold text-violet-700 bg-violet-100 rounded-full px-2 py-0.5')
                                vykresli_prirazena_prava(inherited_odd_prava, zakladni_prava, varianta='zdedena')

                            osobni_prava_list = [p.strip() for p in d['prava'].split(',') if p.strip()] if d['prava'] else []
                            with ui.row().classes('items-center gap-2 mt-4'):
                                ui.icon('person', size='18px').classes('text-emerald-500')
                                ui.label('Osobní práva navíc').classes('text-sm font-bold text-gray-600')
                                if osobni_prava_list:
                                    ui.label(str(len(osobni_prava_list))).classes('text-xs font-bold text-emerald-700 bg-emerald-100 rounded-full px-2 py-0.5')
                            if not osobni_prava_list:
                                ui.label('Žádná').classes('text-sm font-bold text-gray-400 italic')
                            else:
                                vykresli_prirazena_prava(osobni_prava_list, zakladni_prava, varianta='osobni')

                            ui.label('Dovolená:').classes('text-sm text-gray-500 mt-4')
                            ui.label(f"Základ: {d.get('base_vacation', 0)} h | Převod: {d.get('carried_over_vacation', 0)} h").classes('font-bold text-gray-700')

                        with ui.column().classes('gap-2 items-end'):
                            def edit_u(cil_mail=e_m):
                                uzivatel_db = intranet_data.ziskej_vsechny_uzivatele()
                                with ui.dialog() as dlg, ui.card().classes('p-6 rounded-2xl w-full max-w-4xl max-h-[85vh] overflow-y-auto'):
                                    with ui.row().classes('items-center gap-3 mb-5'):
                                        ui.icon('manage_accounts', size='1.5rem').classes('text-indigo-600 bg-indigo-50 p-2 rounded-xl')
                                        with ui.column().classes('gap-0'):
                                            ui.label('Úprava uživatele').classes('text-xl font-bold text-slate-800 leading-tight')
                                            ui.label(cil_mail).classes('text-sm text-slate-500')

                                    with ui.row().classes('w-full gap-4 mb-4'):
                                        n_j = ui.input('Jméno', value=uzivatel_db[cil_mail]['jmeno']).classes('flex-1').props('outlined dense')
                                        n_p = ui.input('Příjmení', value=uzivatel_db[cil_mail]['prijmeni']).classes('flex-1').props('outlined dense')

                                        aktualni_spol_ids = [s['id'] for s in uzivatel_db[cil_mail].get('spolecnosti', [])]
                                        n_spolecnost = ui.select(
                                            spolecnosti_options,
                                            value=aktualni_spol_ids,
                                            label='Společnosti',
                                            multiple=True
                                        ).classes('flex-1').props('outlined dense use-chips clearable')

                                    with ui.row().classes('w-full gap-4 mb-4'):
                                        n_od = ui.select(list(oddeleni.keys()) + ['Bez oddělení'], value=uzivatel_db[cil_mail].get('oddeleni', 'Bez oddělení'), label='Oddělení', multiple=True).classes('flex-1').props('outlined dense use-chips')

                                        manazeri_edit = {u['id']: f"{u['jmeno_cele']} ({u.get('oddeleni', '')})" for u in sorted((u for u in uzivatel_db.values() if u['aktivni'] and u['id'] != uzivatel_db[cil_mail]['id']), key=lambda u: (cz_razeni(u.get('prijmeni', '')), cz_razeni(u.get('jmeno_cele', ''))))}
                                        aktualni_manazeri = uzivatel_db[cil_mail].get('manager_id', [])
                                        n_man = ui.select(manazeri_edit, value=aktualni_manazeri, label='Přímí nadřízení', multiple=True, with_input=True).classes('flex-1').props('outlined dense use-chips clearable')

                                    with ui.row().classes('w-full gap-4 mb-4'):
                                        n_zaklad = ui.number('Základ dovolené (h)', value=uzivatel_db[cil_mail].get('base_vacation', 160.0)).classes('flex-1').props('outlined dense')
                                        n_prevod = ui.number('Převod z loňska (h)', value=uzivatel_db[cil_mail].get('carried_over_vacation', 0.0)).classes('flex-1').props('outlined dense')

                                    with ui.row().classes('w-full gap-4 mb-4'):
                                        n_datum_narozeni = ui.input(
                                            'Datum narození (RRRR-MM-DD)',
                                            value=uzivatel_db[cil_mail].get('datum_narozeni', '')
                                        ).classes('flex-1').props('outlined dense mask="####-##-##" fill-mask')
                                        _pr_opts_edit = {None: '— bez příznaku —', **{pz['id']: pz['nazev'] for pz in intranet_data.ziskej_vsechny_priznaky()}}
                                        n_priznak = ui.select(
                                            _pr_opts_edit,
                                            value=uzivatel_db[cil_mail].get('priznak_id'),
                                            label='Příznak'
                                        ).classes('flex-1').props('outlined dense')
                                        n_pobocka = ui.select(
                                            {None: '— bez pobočky —', **{_pb: _pb for _pb in intranet_data.POBOCKY}},
                                            value=uzivatel_db[cil_mail].get('pobocka'),
                                            label='Pobočka'
                                        ).classes('flex-1').props('outlined dense')

                                    ui.label('Osobní práva navíc').classes('text-xs font-bold text-slate-400 uppercase tracking-wider mt-2 mb-2')
                                    a_prv = [p.strip() for p in uzivatel_db[cil_mail]['prava'].split(',')] if uzivatel_db[cil_mail]['prava'] else []
                                    vybrane_prava_edit = render_prava_kategorie(zakladni_prava, a_prv)

                                    async def potvrdit_u():
                                        od_str = ",".join(n_od.value) if isinstance(n_od.value, list) else n_od.value
                                        pr_str = ",".join(vybrane_prava_edit)
                                        m_ids = n_man.value if n_man.value else []
                                        s_ids = n_spolecnost.value if n_spolecnost.value else []

                                        await asyncio.to_thread(
                                            intranet_data.pridej_uprav_uzivatele,
                                            cil_mail, n_j.value, n_p.value, None, uzivatel_db[cil_mail]['role'],
                                            od_str, pr_str, uzivatel_db[cil_mail]['aktivni'], n_zaklad.value, n_prevod.value,
                                            None, m_ids, s_ids, n_datum_narozeni.value or None,
                                            n_priznak.value, n_pobocka.value
                                        )
                                        intranet_logger.log_activity(user_name, "Správa uživatelů", f"Upraveny údaje uživatele: {cil_mail}")

                                        ui.notify(f'Uživatel upraven.', type='positive', position='top')
                                        dlg.close()
                                        intranet_data.invaliduj_cache_sprava()
                                        await asyncio.to_thread(intranet_data.ziskej_vsechny_uzivatele)
                                        ui.timer(0, vykresli_seznam_uzivatelu.refresh, once=True)

                                    with ui.row().classes('w-full justify-end gap-3 mt-6 pt-4 border-t border-slate-100'):
                                        ui.button('Zrušit', on_click=dlg.close).props('flat no-caps').classes('text-slate-600 font-semibold rounded-lg')
                                        ui.button('Uložit změny', icon='save', on_click=potvrdit_u).props('unelevated no-caps').classes('bg-indigo-600 hover:bg-indigo-700 text-white font-semibold h-12 px-6 rounded-lg shadow-sm')
                                dlg.open()

                            def prepni_aktivitu(cil_mail=e_m):
                                _u = intranet_data.ziskej_vsechny_uzivatele()[cil_mail]
                                a = not _u['aktivni']
                                async def _proved_prepnuti():
                                    await asyncio.to_thread(
                                        intranet_data.pridej_uprav_uzivatele,
                                        cil_mail, _u['jmeno'], _u['prijmeni'], None, _u['role'],
                                        _u.get('oddeleni'), _u['prava'], a, _u.get('base_vacation', 160.0),
                                        _u.get('carried_over_vacation', 0.0), None, _u.get('manager_id', []),
                                        [s['id'] for s in _u.get('spolecnosti', [])]
                                    )
                                    intranet_logger.log_activity(user_name, "Správa uživatelů", f"{'Aktivován' if a else 'Deaktivován'} uživatel: {cil_mail}")
                                    ui.notify(f"Účet {'aktivován' if a else 'deaktivován'}.", type='info', position='top')
                                    intranet_data.invaliduj_cache_sprava()
                                    await asyncio.to_thread(intranet_data.ziskej_vsechny_uzivatele)
                                    ui.timer(0, vykresli_seznam_uzivatelu.refresh, once=True)
                                return _proved_prepnuti()

                            def smaz_u(cil_mail=e_m):
                                with ui.dialog() as dlg, ui.card().classes('p-6 rounded-2xl w-full max-w-sm'):
                                    with ui.row().classes('items-center gap-3 mb-2'):
                                        ui.icon('warning', size='1.5rem').classes('text-red-600 bg-red-50 p-2 rounded-xl')
                                        ui.label('Opravdu smazat?').classes('text-xl font-bold text-slate-800')
                                    ui.label('Trvalé smazání uživatele i jeho přiřazení. Nelze vrátit zpět.').classes('text-sm text-slate-500 mb-4')
                                    async def potvrdit():
                                        await asyncio.to_thread(intranet_data.smaz_uzivatele, cil_mail)
                                        intranet_logger.log_activity(user_name, "Správa uživatelů", f"Trvale smazán uživatel: {cil_mail}")
                                        ui.notify(f'Smazáno.', type='positive', position='top')
                                        dlg.close()
                                        intranet_data.invaliduj_cache_sprava()
                                        await asyncio.to_thread(intranet_data.ziskej_vsechny_uzivatele)
                                        ui.timer(0, vykresli_seznam_uzivatelu.refresh, once=True)
                                    with ui.row().classes('w-full justify-end gap-3'):
                                        ui.button('Zrušit', on_click=dlg.close).props('flat no-caps').classes('text-slate-600 font-semibold rounded-lg')
                                        ui.button('Smazat', icon='delete', on_click=potvrdit).props('unelevated no-caps').classes('bg-red-600 hover:bg-red-700 text-white font-semibold rounded-lg shadow-sm')
                                dlg.open()

                            def otevri_dialog_zmena_hesla(cil_mail=e_m):
                                _u = intranet_data.ziskej_vsechny_uzivatele()[cil_mail]
                                with ui.dialog() as dlg, ui.card().classes('p-6 rounded-2xl w-full max-w-md'):
                                    with ui.row().classes('items-center gap-3 mb-4'):
                                        ui.icon('key', size='1.4rem').classes('text-orange-600 bg-orange-50 p-2 rounded-xl')
                                        ui.label(f'Nové heslo pro: {_u["jmeno_cele"]}').classes('text-lg font-bold text-slate-800')
                                    nove_h = ui.input('Zadejte nové heslo', password=True).classes('w-full mb-6').props('outlined dense')
                                    async def potvrdit():
                                        if not intranet_data.heslo_je_silne(nove_h.value):
                                            return ui.notify('Heslo musí mít min. 8 znaků, obsahovat velké i malé písmeno a číslo.', type='warning', position='top')

                                        await asyncio.to_thread(
                                            intranet_data.pridej_uprav_uzivatele,
                                            cil_mail, _u['jmeno'], _u['prijmeni'], nove_h.value, _u['role'],
                                            _u.get('oddeleni'), _u['prava'], _u['aktivni'],
                                            _u.get('base_vacation', 160.0), _u.get('carried_over_vacation', 0.0),
                                            None, _u.get('manager_id', []), [s['id'] for s in _u.get('spolecnosti', [])]
                                        )
                                        intranet_logger.log_activity(user_name, "Správa uživatelů", f"Administrátor změnil heslo uživateli: {cil_mail}")
                                        ui.notify(f'Heslo změněno.', type='positive', position='top')
                                        dlg.close()
                                    with ui.row().classes('w-full justify-end gap-3'):
                                        ui.button('Zrušit', on_click=dlg.close).props('flat no-caps').classes('text-slate-600 font-semibold rounded-lg')
                                        ui.button('Uložit heslo', icon='save', on_click=potvrdit).props('unelevated no-caps').classes('bg-emerald-600 hover:bg-emerald-700 text-white font-semibold rounded-lg shadow-sm')
                                dlg.open()

                            def zrus_2fa(cil_mail=e_m, cil_id=d['id']):
                                """Zruší uživateli 2FA — pro případ ztraceného telefonu / záložních kódů."""
                                with ui.dialog() as dlg, ui.card().classes('p-6 rounded-2xl w-full max-w-sm'):
                                    with ui.row().classes('items-center gap-3 mb-2'):
                                        ui.icon('phonelink_erase', size='1.5rem').classes('text-amber-600 bg-amber-50 p-2 rounded-xl')
                                        ui.label('Zrušit dvoufaktorové ověření?').classes('text-xl font-bold text-slate-800')
                                    ui.label('Účet bude chráněn už jen heslem. Uživatel si může 2FA znovu aktivovat v osobním nastavení.').classes('text-sm text-slate-500 mb-4')
                                    async def potvrdit():
                                        uspech = await asyncio.to_thread(intranet_2fa.deaktivuj_2fa, cil_id)
                                        if uspech:
                                            intranet_logger.log_activity(user_name, "Správa uživatelů", f"Administrátor zrušil 2FA uživateli: {cil_mail}")
                                            ui.notify('2FA zrušeno.', type='positive', position='top')
                                            dlg.close()
                                            ui.timer(0, vykresli_seznam_uzivatelu.refresh, once=True)
                                        else:
                                            ui.notify('Spojení s databází selhalo.', type='negative', position='top')
                                    with ui.row().classes('w-full justify-end gap-3'):
                                        ui.button('Zrušit', on_click=dlg.close).props('flat no-caps').classes('text-slate-600 font-semibold rounded-lg')
                                        ui.button('Zrušit 2FA', icon='phonelink_erase', on_click=potvrdit).props('unelevated no-caps').classes('bg-amber-600 hover:bg-amber-700 text-white font-semibold rounded-lg shadow-sm')
                                dlg.open()

                            if user_email != e_m:
                                with ui.row().classes('gap-2'):
                                    ui.button('Upravit', icon='edit', on_click=edit_u).props('outline size=sm no-caps').classes('w-32 rounded-lg text-indigo-600')
                                    ui.button('Změnit heslo', icon='key', on_click=otevri_dialog_zmena_hesla).props('outline size=sm no-caps color=orange').classes('w-32 rounded-lg text-orange-600')
                                with ui.row().classes('gap-2'):
                                    if d['aktivni']: ui.button('Deaktivovat', icon='block', color='orange', on_click=prepni_aktivitu).props('outline size=sm no-caps').classes('w-32 rounded-lg')
                                    else: ui.button('Aktivovat', icon='check_circle', color='green', on_click=prepni_aktivitu).props('outline size=sm no-caps').classes('w-32 rounded-lg')
                                    ui.button('Smazat', icon='delete', color='red', on_click=smaz_u).props('outline size=sm no-caps').classes('w-32 rounded-lg text-red-600')
                                if intranet_2fa.ma_aktivni_2fa(d['id']):
                                    with ui.row().classes('gap-2'):
                                        ui.button('Zrušit 2FA', icon='phonelink_erase', color='amber', on_click=zrus_2fa).props('outline size=sm no-caps').classes('w-32 rounded-lg text-amber-700')

            def _lazy_handler(e_m, d, telo):
                _st = {'done': False}
                def _h(ev):
                    if ev.value and not _st['done']:
                        _st['done'] = True
                        _build_user_body(e_m, d, telo)
                return _h

            # 1) Filtr (bez adminů, dle hledání) — levné, jen Python.
            filtrovani = []
            for e_m, d in sorted(aktualni_db.items(), key=lambda x: (cz_razeni(x[1].get('prijmeni', '')), cz_razeni(x[1]['jmeno_cele']))):
                if 'admin' in d['jmeno_cele'].lower() or 'admin' in e_m.lower(): continue
                if hledany_text:
                    osobni_cislo = str(d.get('id', ''))
                    oddeleni_text = (d.get('oddeleni') or '').lower()
                    if (hledany_text not in e_m.lower()
                            and hledany_text not in d['jmeno_cele'].lower()
                            and hledany_text not in osobni_cislo
                            and hledany_text not in oddeleni_text):
                        continue
                filtrovani.append((e_m, d))

            # 2) Slice — bez hledání cap na _LIMIT_UZIV (pokud uživatel nerozbalil vše).
            celkem = len(filtrovani)
            orezano = (not hledany_text) and (not _zobraz_vse['v']) and celkem > _LIMIT_UZIV
            k_zobrazeni = filtrovani[:_LIMIT_UZIV] if orezano else filtrovani

            if not k_zobrazeni:
                ui.label('Žádní uživatelé neodpovídají hledání.').classes('text-slate-400 italic text-sm py-4')

            # 3) Vykreslení (expansion + lazy tělo).
            for e_m, d in k_zobrazeni:
                barva = "border-slate-200 hover:border-indigo-300" if d['aktivni'] else "border-red-300 bg-red-50"
                exp = ui.expansion(
                    f"{d['jmeno_cele']} ({e_m}) {'[DEAKTIVOVÁN]' if not d['aktivni'] else ''}",
                    icon='account_circle',
                ).classes(f'w-full border rounded-xl mb-2.5 shadow-sm transition-colors {barva}')
                with exp:
                    telo = ui.column().classes('w-full')
                exp.on_value_change(_lazy_handler(e_m, d, telo))

            # 4) Patička: info o oříznutí + „Zobrazit vše".
            if orezano:
                def _zobraz_vsechny():
                    _zobraz_vse['v'] = True
                    vykresli_seznam_uzivatelu.refresh()
                with ui.row().classes('w-full items-center justify-center gap-3 mt-4 pt-4 border-t border-slate-100'):
                    ui.label(f'Zobrazeno {_LIMIT_UZIV} z {celkem} uživatelů — zužte hledáním.').classes('text-sm text-slate-500')
                    ui.button(f'Zobrazit vše ({celkem})', icon='expand_more', on_click=_zobraz_vsechny) \
                      .props('outline no-caps').classes('text-indigo-600 font-semibold rounded-lg')

        filtr_input.on_value_change(vykresli_seznam_uzivatelu.refresh)
        vykresli_seznam_uzivatelu()

    with ui.card().classes('w-full p-8 shadow-sm bg-white rounded-2xl border border-slate-200'):
        with ui.row().classes('w-full items-center gap-3 mb-6'):
            ui.icon('domain', size='1.6rem').classes('text-indigo-600 bg-indigo-50 p-2 rounded-xl')
            with ui.column().classes('gap-0'):
                ui.label('Správa oddělení').classes('text-2xl font-bold text-slate-800 leading-tight')
                ui.label('Nastavení hromadných práv pro celá oddělení.').classes('text-sm text-slate-500')

        @ui.refreshable
        def vykresli_oddeleni_vnitrni():
            with ui.card().classes('w-full bg-indigo-50/40 border border-indigo-100 p-6 mb-8 rounded-2xl'):
                ui.label('Vytvořit nové oddělení').classes('font-bold text-xl mb-4 text-slate-800')
                n_o = ui.input('Název nového oddělení').classes('w-full max-w-md mb-4 bg-white').props('outlined dense')

                is_maj = ui.switch('Oddělení Majitelů (oranžová barva s výběrem sledovaných lidí v kalendáři)').classes('mb-4 font-semibold text-orange-600')

                ui.label('Základní práva pro toto oddělení').classes('text-xs font-bold text-slate-400 uppercase tracking-wider mb-2')
                vybrane_prava_odd = render_prava_kategorie(zakladni_prava, [])

                async def potvrdit_o():
                    if not n_o.value: return ui.notify('Zadejte název!', type='warning')
                    nazev_o = n_o.value.strip()
                    await asyncio.to_thread(intranet_data.pridej_uprav_oddeleni, nazev_o, ",".join(vybrane_prava_odd), is_maj.value)
                    intranet_logger.log_activity(user_name, "Správa uživatelů", f"Vytvořeno nové oddělení: {nazev_o}")
                    ui.notify('Oddělení přidáno.', type='positive')
                    n_o.value = ''
                    intranet_data.invaliduj_cache_sprava()
                    await asyncio.to_thread(intranet_data.ziskej_vsechna_oddeleni)
                    ui.timer(0, vykresli_oddeleni_vnitrni.refresh, once=True)
                ui.button('Přidat oddělení', on_click=potvrdit_o, icon='add_business').props('unelevated no-caps').classes('bg-indigo-600 hover:bg-indigo-700 text-white font-semibold h-12 px-6 mt-4 rounded-lg shadow-sm')

            o_db = intranet_data.ziskej_vsechna_oddeleni()
            if not o_db:
                ui.label('Zatím nejsou vytvořena žádná oddělení.').classes('italic text-gray-500')
            else:
                uzivatele_v_odd_seznam = {n: [] for n in o_db.keys()}
                for u in db.values():
                    if 'admin' in u.get('oddeleni', '').lower() or 'admin' in u.get('email', '').lower(): continue
                    if u['aktivni'] and u.get('oddeleni'):
                        for o in u['oddeleni'].split(','):
                            o = o.strip()
                            if o in uzivatele_v_odd_seznam:
                                uzivatele_v_odd_seznam[o].append(u['jmeno_cele'])

                for k in uzivatele_v_odd_seznam:
                    uzivatele_v_odd_seznam[k].sort()

                with ui.grid(columns=1).classes('w-full gap-6 md:grid-cols-2 xl:grid-cols-3 mt-4'):
                    for nazev, o_data in sorted(o_db.items()):
                        if nazev.lower() == 'admin': continue

                        pr = o_data['prava']
                        is_majitele = o_data['is_majitele']

                        hlavni_vedouci_list = []
                        for u in db.values():
                            if not u['aktivni']: continue
                            osobni_prava = [p.strip().lower() for p in (u.get('prava') or '').split(',')]
                            if f'hlavni_vedouci_{nazev.lower()}' in osobni_prava:
                                hlavni_vedouci_list.append(u['jmeno_cele'])

                        hlavni_txt = ", ".join(hlavni_vedouci_list) if hlavni_vedouci_list else "Nenastaven"

                        card_bg = 'bg-orange-600' if is_majitele else 'bg-indigo-600'

                        with ui.card().classes('w-full p-0 bg-white border border-slate-200 shadow-sm hover:shadow-lg transition-shadow rounded-2xl flex flex-col overflow-hidden'):
                            with ui.row().classes(f'w-full {card_bg} text-white p-4 justify-between items-center'):
                                with ui.row().classes('items-center gap-2'):
                                    ui.icon('star' if is_majitele else 'business', size='sm').classes('opacity-90')
                                    ui.label(nazev).classes('font-extrabold text-lg tracking-wide')

                                with ui.row().classes('gap-1'):
                                    def edit_o(n=nazev, p=pr, i_m=is_majitele):
                                        with ui.dialog() as dlg_odd, ui.card().classes('p-6 rounded-2xl w-full max-w-3xl max-h-[85vh] overflow-y-auto'):
                                            with ui.row().classes('items-center gap-3 mb-4'):
                                                ui.icon('domain', size='1.4rem').classes('text-indigo-600 bg-indigo-50 p-2 rounded-xl')
                                                ui.label(f'Úprava oddělení: {n}').classes('text-xl font-bold text-slate-800')

                                            i_m_switch = ui.switch('Oddělení Majitelů (oranžová barva)', value=i_m).classes('mb-4 font-semibold text-orange-600')

                                            ui.label('Základní práva oddělení').classes('text-xs font-bold text-slate-400 uppercase tracking-wider mb-2')

                                            p_list = [x.strip() for x in p.split(',')] if p else []
                                            vybrane_prava_odd_edit = render_prava_kategorie(zakladni_prava, p_list)

                                            async def potvrdit_edit_odd():
                                                pr_str = ",".join(vybrane_prava_odd_edit)
                                                await asyncio.to_thread(intranet_data.pridej_uprav_oddeleni, n, pr_str, i_m_switch.value)
                                                intranet_logger.log_activity(user_name, "Správa uživatelů", f"Upravena práva pro oddělení: {n}")
                                                ui.notify('Práva oddělení upravena.', type='positive', position='top')
                                                dlg_odd.close()
                                                intranet_data.invaliduj_cache_sprava()
                                                await asyncio.to_thread(intranet_data.ziskej_vsechna_oddeleni)
                                                ui.timer(0, vykresli_oddeleni_vnitrni.refresh, once=True)

                                            with ui.row().classes('w-full justify-end gap-3 mt-6 pt-4 border-t border-slate-100'):
                                                ui.button('Zrušit', on_click=dlg_odd.close).props('flat no-caps').classes('text-slate-600 font-semibold rounded-lg')
                                                ui.button('Uložit změny', icon='save', on_click=potvrdit_edit_odd).props('unelevated no-caps').classes('bg-indigo-600 hover:bg-indigo-700 text-white font-semibold h-12 px-6 rounded-lg shadow-sm')
                                        dlg_odd.open()

                                    def smaz_o(n=nazev):
                                        with ui.dialog() as d_del, ui.card().classes('p-6 rounded-2xl'):
                                            with ui.row().classes('items-center gap-3 mb-4'):
                                                ui.icon('warning', size='1.4rem').classes('text-red-600 bg-red-50 p-2 rounded-xl')
                                                ui.label(f'Smazat oddělení {n}?').classes('font-bold text-xl text-slate-800')
                                            async def on_yes():
                                                await asyncio.to_thread(intranet_data.smaz_oddeleni, n)
                                                intranet_logger.log_activity(user_name, "Správa uživatelů", f"Smazáno oddělení: {n}")
                                                ui.notify('Smazáno.', type='info')
                                                d_del.close()
                                                intranet_data.invaliduj_cache_sprava()
                                                await asyncio.to_thread(intranet_data.ziskej_vsechna_oddeleni)
                                                ui.timer(0, vykresli_oddeleni_vnitrni.refresh, once=True)
                                            with ui.row().classes('w-full justify-end gap-3'):
                                                ui.button('Zrušit', on_click=d_del.close).props('flat no-caps').classes('text-slate-600 font-semibold rounded-lg')
                                                ui.button('Smazat', icon='delete', on_click=on_yes).props('unelevated no-caps').classes('bg-red-600 hover:bg-red-700 text-white font-semibold rounded-lg shadow-sm')
                                        d_del.open()

                                    ui.button(icon='edit', on_click=lambda n=nazev, p=pr, i_m=is_majitele: edit_o(n, p, i_m)).props('flat dense text-color=white').tooltip('Upravit oddělení')
                                    ui.button(icon='delete', on_click=lambda n=nazev: smaz_o(n)).props('flat dense text-color=white').tooltip('Smazat oddělení')

                            with ui.column().classes('p-5 w-full flex-1 gap-4 bg-white'):

                                # Informace o HLAVNÍM VEDOUCÍM
                                with ui.row().classes('items-center gap-3 w-full'):
                                    ui.icon('star', size='sm', color='orange-500').classes('bg-orange-50 p-2 rounded-full')
                                    with ui.column().classes('gap-0'):
                                        ui.label('Hlavní vedoucí oddělení').classes('text-[10px] text-gray-500 font-bold uppercase tracking-wider')
                                        ui.label(hlavni_txt).classes('text-sm font-bold text-gray-800')

                                seznam_lidi = uzivatele_v_odd_seznam[nazev]
                                pocet_lidi = len(seznam_lidi)

                                def ukaz_seznam_lidi(odd_n, seznam, odd_je_majitel):
                                    with ui.dialog() as dlg, ui.card().classes('p-6 rounded-2xl w-full max-w-3xl'):
                                        with ui.row().classes('items-center gap-3 mb-4'):
                                            ui.icon('groups', size='1.4rem').classes('text-indigo-600 bg-indigo-50 p-2 rounded-xl')
                                            ui.label(f'Zaměstnanci: {odd_n}').classes('text-xl font-bold text-slate-800')
                                        if not seznam:
                                            ui.label('Oddělení je zatím prázdné.').classes('italic text-gray-500')
                                        else:
                                            with ui.column().classes('w-full max-h-[60vh] overflow-y-auto gap-3 pr-2'):
                                                for jmeno in seznam:
                                                    u_data = next((u for u in db.values() if u['jmeno_cele'] == jmeno), None)
                                                    if u_data:
                                                        with ui.card().classes('w-full bg-gray-50 border border-gray-200 p-4 shadow-sm'):
                                                            with ui.row().classes('items-center gap-3 w-full mb-2'):
                                                                ui.icon('person', size='sm', color='gray-500')
                                                                ui.label(jmeno).classes('font-bold text-gray-800 text-lg')

                                                            if odd_je_majitel:
                                                                opts_id_name = {u_val['id']: u_val['jmeno_cele'] for u_val in db.values() if u_val['aktivni'] and u_val['id'] != u_data['id']}

                                                                def save_watched(e, uid=u_data['id']):
                                                                    intranet_data.uloz_sledovane_uzivatele(uid, e.value)
                                                                    ui.notify(f'Uloženo.', type='positive')

                                                                ui.select(opts_id_name, value=u_data.get('sledovani_uzivatele', []), multiple=True, label='Vyberte uživatele pro globální kalendář a schvalování', on_change=save_watched).classes('w-full').props('use-chips clearable')

                                        ui.button('Zavřít', on_click=dlg.close).classes('w-full mt-6 bg-gray-200 hover:bg-gray-300 text-gray-800 font-bold rounded-xl h-12')
                                    dlg.open()

                                with ui.row().classes('items-center gap-3 w-full p-2 -mx-2 rounded-xl hover:bg-blue-50 cursor-pointer transition-colors').on('click', lambda n=nazev, s=seznam_lidi, im=is_majitele: ukaz_seznam_lidi(n, s, im)):
                                    ui.icon('groups', size='sm', color='blue-500').classes('bg-blue-100 p-2 rounded-full')
                                    with ui.column().classes('gap-0'):
                                        ui.label('Aktivní zaměstnanci (Klikněte)').classes('text-[10px] text-blue-500 font-bold uppercase tracking-wider')
                                        ui.label(str(pocet_lidi)).classes('text-sm font-bold text-gray-800')

                                pr_list = [p.strip() for p in pr.split(',') if p.strip()] if pr else []
                                with ui.row().classes('items-center gap-2 mt-2 border-t border-gray-100 pt-4 w-full'):
                                    ui.label('Výchozí práva oddělení').classes('text-[10px] font-bold text-gray-400 uppercase tracking-wider')
                                    if pr_list:
                                        ui.label(str(len(pr_list))).classes('text-xs font-bold text-blue-700 bg-blue-100 rounded-full px-2 py-0.5')
                                if not pr_list:
                                    ui.label('Žádná zvláštní práva.').classes('text-sm text-gray-400 italic')
                                else:
                                    vykresli_prirazena_prava(pr_list, zakladni_prava, varianta='oddeleni')

        vykresli_oddeleni_vnitrni()