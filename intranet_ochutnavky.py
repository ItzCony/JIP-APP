# intranet_ochutnavky.py
# ═══════════════════════════════════════════════════════════════════════════════
#   Ochutnávky – Maloobchod a Svět potravin
# ═══════════════════════════════════════════════════════════════════════════════
import os
import re
import time
import uuid
import asyncio
import threading
import unicodedata
from datetime import date, datetime, timedelta

import intranet_data
import intranet_notifikace
from nicegui import ui
from intranet_ui_utils import refreshable_na_klienta

# ─────────────────────────────────────────────────────────────────────────────
# KONSTANTY
# ─────────────────────────────────────────────────────────────────────────────
PERM_ADMIN   = 'ochutnavky_admin'
PERM_PRISTUP = 'ochutnavky_pristup'
PRILOHY_DIR  = 'ochutnavky_prilohy'

MAX_PRILOHA = 25 * 1024 * 1024   # 25 MB
POVOLENE_PRIPONY = {'.pdf', '.jpg', '.jpeg', '.png', '.gif', '.webp', '.heic'}

# Divize – v UI všude celým textem, nikdy zkratkou.
DIVIZE = {
    'MO':  'Maloobchod',
    'SP':  'Svět potravin',
    'OBE': 'Maloobchod + Svět potravin',
}

# Pobočky dle seznamu prodejen na intranetu (Prodejny MO / Prodejny C&C)
POBOCKY_MO = [
    '103 - Jičín Korrekt "H"', '107 - Slaný', '115 - Horní Slavkov',
    '116 - Karlovy Vary Vyhlídka', '117 - Sokolov', '123 - Chomutov Jiráskova',
    '133 - Sezemice', '139 - Svoboda nad Úpou', '140 - Hrádek nad Nisou',
    '142 - Harrachov "H"', '145 - M. Lázně Husova', '147 - Žlutice',
    '148 - M. Lázně Hroznatova', '154 - Františkovy Lázně',
    '163 - Jáchymov', '165 - Prachatice', '172 - Sadská', '173 - Blatná',
    '174 - Špindlerův Mlýn', '175 - Pardubice Smilova',
]
POBOCKY_SP = [
    '300 - Pardubice', '302 - Praha Bořanovice', '306 - Svoboda nad Úpou',
    '308 - Karlovy Vary', '310 - Sušice NEW', '311 - Polička',
    '313 - Náchod', '316 - Jilemnice', '317 - České Budějovice',
    '318 - Zlín', '319 - Olomouc',
    '320 - Hodonín', '321 - Břeclav', '325 - Brno [HALBR-Váhy+4MAX]',
    '326 - Ostrava [4Max]', '340 - ELKO Berounka', '341 - ELKO Domažlice',
    '342 - ELKO Kladno', '343 - ELKO Kralovice', '344 - ELKO Nepomuk',
    '345 - ELKO Klatovy', '346 - ELKO Cheb',
]

# Přejmenování poboček oproti první verzi modulu – migrace uložených dat.
# Pořadí je důležité: '175 - Pardubice' musí odejít dřív, než vznikne
# '300 - Pardubice' (jinak by se hodnoty na chvíli překrývaly).
_MIGRACE_POBOCEK = [
    ('175 - Pardubice', '175 - Pardubice Smilova'),
    ('103 - Jičín', '103 - Jičín Korrekt "H"'),
    ('116 - Karlovy Vary - Vyhlídka', '116 - Karlovy Vary Vyhlídka'),
    ('123 - Chom.Jiráskova', '123 - Chomutov Jiráskova'),
    ('142 - Harrachov', '142 - Harrachov "H"'),
    ('145 - Mariánské Lázně Husova', '145 - M. Lázně Husova'),
    ('148 - Mariánské Lázně Hroznatova', '148 - M. Lázně Hroznatova'),
    ('174 - Špindlerův mlýn', '174 - Špindlerův Mlýn'),
    ('300 - Pardubice SP', '300 - Pardubice'),
    ('302 - Praha Bořanovice SP', '302 - Praha Bořanovice'),
    ('308 - Karlovy Vary SP', '308 - Karlovy Vary'),
    ('310 - Sušice SP', '310 - Sušice NEW'),
    ('311 - Polička SP', '311 - Polička'),
    ('316 - Jilemnice SP', '316 - Jilemnice'),
    ('318 - Zlín SP', '318 - Zlín'),
    ('319 - Olomouc SP', '319 - Olomouc'),
    ('325 - Brno SP', '325 - Brno [HALBR-Váhy+4MAX]'),
    ('326 - Ostrava SP', '326 - Ostrava [4Max]'),
]

# Barvy stavů akce
_STAVY = {
    'probiha':  ('Probíhá',      'bg-green-100 text-green-800 border-green-300'),
    'brzy':     ('Následujících 7 dní', 'bg-amber-100 text-amber-800 border-amber-300'),
    'budouci':  ('Plánováno',    'bg-blue-50 text-blue-700 border-blue-200'),
    'probehlo': ('Proběhlo',     'bg-gray-100 text-gray-500 border-gray-200'),
}

_DB_INIT = False


# ═══════════════════════════════════════════════════════════════════════════════
# POMOCNÉ FUNKCE (bez DB – testovatelné)
# ═══════════════════════════════════════════════════════════════════════════════
def pobocky_pro_divizi(divize):
    """Seznam poboček pro zvolenou divizi."""
    if divize == 'MO':
        return list(POBOCKY_MO)
    if divize == 'SP':
        return list(POBOCKY_SP)
    return list(POBOCKY_MO) + list(POBOCKY_SP)


def _na_datum(d):
    """Cokoliv (str/date/datetime) → date, jinak None."""
    if not d:
        return None
    if isinstance(d, datetime):
        return d.date()
    if isinstance(d, date):
        return d
    try:
        return datetime.strptime(str(d)[:10], '%Y-%m-%d').date()
    except Exception:
        return None


def _fmt_d(d):
    d = _na_datum(d)
    return d.strftime('%d.%m.%Y') if d else ''


def _fmt_dt(d):
    if isinstance(d, datetime):
        return d.strftime('%d.%m.%Y %H:%M')
    return str(d or '')


def obdobi_text(od, do):
    """Jeden den → jedno datum, jinak 'od – do'."""
    od, do = _na_datum(od), _na_datum(do)
    if not od:
        return ''
    if not do or do == od:
        return _fmt_d(od)
    return f'{_fmt_d(od)} – {_fmt_d(do)}'


def stav_akce(od, do, dnes=None):
    """probehlo | probiha | brzy (do 7 dnů) | budouci"""
    dnes = dnes or date.today()
    od, do = _na_datum(od), _na_datum(do)
    if not od:
        return 'budouci'
    do = do or od
    if do < dnes:
        return 'probehlo'
    if od <= dnes:
        return 'probiha'
    if od <= dnes + timedelta(days=7):
        return 'brzy'
    return 'budouci'


def _bezpecny_nazev(nazev):
    """Očistí název souboru – zabrání průchodu adresářem."""
    nazev = os.path.basename(str(nazev or 'soubor'))
    nazev = re.sub(r'[^A-Za-z0-9._\- ]+', '_', nazev)
    return nazev[:80] or 'soubor'


# ═══════════════════════════════════════════════════════════════════════════════
# DB INICIALIZACE
# ═══════════════════════════════════════════════════════════════════════════════
def _init_db():
    global _DB_INIT
    if _DB_INIT:
        return
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    cur = None
    try:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS ochutnavky (
                id          VARCHAR(36)  PRIMARY KEY,
                datum_od    DATE         NOT NULL,
                datum_do    DATE         NOT NULL,
                nazev       VARCHAR(200) NOT NULL,
                popis       TEXT,
                poznamka    TEXT,
                divize      VARCHAR(3)   NOT NULL,
                dodavatel   VARCHAR(200),
                sortiment   VARCHAR(200),
                autor_id    INT,
                autor_jmeno VARCHAR(200),
                vytvoreno   DATETIME     DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_obdobi (datum_od, datum_do)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci ENGINE=InnoDB
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS ochutnavky_pobocky (
                id            VARCHAR(36) PRIMARY KEY,
                ochutnavka_id VARCHAR(36) NOT NULL,
                pobocka       VARCHAR(80) NOT NULL,
                UNIQUE KEY uk_op (ochutnavka_id, pobocka),
                INDEX idx_pob (pobocka)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci ENGINE=InnoDB
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS ochutnavky_prilohy (
                id            VARCHAR(36)  PRIMARY KEY,
                ochutnavka_id VARCHAR(36)  NOT NULL,
                nazev         VARCHAR(255) NOT NULL,
                soubor        VARCHAR(255) NOT NULL,
                autor         VARCHAR(200),
                datum         DATETIME     DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_akce (ochutnavka_id)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci ENGINE=InnoDB
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS ochutnavky_chat (
                id         VARCHAR(36)  PRIMARY KEY,
                user_id    INT          NOT NULL,
                user_jmeno VARCHAR(200),
                text       TEXT         NOT NULL,
                datum      DATETIME     DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_datum (datum)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci ENGINE=InnoDB
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS ochutnavky_poznamky (
                id            VARCHAR(36) PRIMARY KEY,
                ochutnavka_id VARCHAR(36) NOT NULL,
                user_id       INT,
                user_jmeno    VARCHAR(200),
                text          TEXT        NOT NULL,
                datum         DATETIME    DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_akce_pozn (ochutnavka_id)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci ENGINE=InnoDB
        """)
        # Migrace názvů poboček na aktuální seznam prodejen z intranetu
        for stary, novy in _MIGRACE_POBOCEK:
            cur.execute('UPDATE IGNORE ochutnavky_pobocky SET pobocka = %s '
                        'WHERE pobocka = %s', (novy, stary))
        conn.commit()
        _DB_INIT = True
    except Exception as e:
        print(f'[Ochutnavky] DB init: {e}')
    finally:
        if cur:  cur.close()
        if conn: conn.close()


# ═══════════════════════════════════════════════════════════════════════════════
# DATA – akce
# ═══════════════════════════════════════════════════════════════════════════════
RAZENI_VOLBY = {
    'datum_asc':   'Datum – nejstarší první',
    'datum_desc':  'Datum – nejnovější první',
    'dodavatel':   'Dodavatel A → Z',
    'dodavatel_z': 'Dodavatel Z → A',
}


def _bez_diakritiky(text):
    return ''.join(z for z in unicodedata.normalize('NFKD', str(text or ''))
                   if not unicodedata.combining(z))


def _abc(text):
    """Řadicí klíč – bez diakritiky a bez ohledu na velikost písmen."""
    return _bez_diakritiky(text).casefold().strip()


def _serad(akce, razeni='datum_asc'):
    """Seřadí akce dle volby uživatele (v Pythonu kvůli české abecedě)."""
    seznam = sorted(akce, key=lambda a: (a['datum_od'] or date.min,
                                         _abc(a.get('nazev'))))
    if razeni == 'datum_desc':
        seznam.sort(key=lambda a: a['datum_od'] or date.min, reverse=True)
    elif razeni in ('dodavatel', 'dodavatel_z'):
        # akce bez vyplněného dodavatele vždy až na konec seznamu
        s_dodavatelem = [a for a in seznam if _abc(a.get('dodavatel'))]
        bez = [a for a in seznam if not _abc(a.get('dodavatel'))]
        s_dodavatelem.sort(key=lambda a: _abc(a['dodavatel']),
                           reverse=(razeni == 'dodavatel_z'))
        seznam = s_dodavatelem + bez
    return seznam


def pocet_aktualnich_akci(dnu_dopredu=7):
    """Počet akcí, které právě probíhají nebo začnou do X dnů (pro dlaždici)."""
    _init_db()
    conn = intranet_data.get_db_connection()
    if not conn:
        return 0
    cur = None
    try:
        cur = conn.cursor()
        dnes = date.today()
        cur.execute(
            "SELECT COUNT(*) FROM ochutnavky "
            "WHERE datum_do >= %s AND datum_od <= %s",
            (dnes, dnes + timedelta(days=dnu_dopredu))
        )
        row = cur.fetchone()
        return int(row[0]) if row else 0
    except Exception as e:
        print(f"[Ochutnávky] pocet_aktualnich_akci: {e}")
        return 0
    finally:
        if cur:
            try: cur.close()
            except Exception: pass
        if conn:
            try: conn.close()
            except Exception: pass


def _vyhovuje_hledani(akce, dotaz):
    """Fulltext bez ohledu na diakritiku a velikost písmen.
    Prohledává název, dodavatele, sortiment, popis, poznámku i provozovny."""
    if not dotaz:
        return True
    seno = ' '.join(str(akce.get(k) or '') for k in
                    ('nazev', 'dodavatel', 'sortiment', 'popis', 'poznamka'))
    seno += ' ' + ' '.join(akce.get('pobocky') or [])
    seno = _abc(seno)
    # každé slovo dotazu musí být obsaženo (pořadí nerozhoduje)
    return all(s in seno for s in _abc(dotaz).split())


def _nacti_akce(divize=None, pobocka=None, od=None, do=None, jen_aktualni=False,
                razeni='datum_asc', hledat=None):
    """Seznam akcí dle filtru; každá má klíč 'pobocky' (list) a 'prilohy_pocet'."""
    _init_db()
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    cur = None
    try:
        cur = conn.cursor(dictionary=True)
        kde, params = [], []
        if divize:
            # 'Maloobchod + Svět potravin' patří do obou přehledů
            kde.append('(o.divize = %s OR o.divize = %s)')
            params += [divize, 'OBE']
        # pobocka může být jedna provozovna (str) i seznam provozoven
        pob_seznam = ([pobocka] if isinstance(pobocka, str)
                      else [p for p in (pobocka or []) if p])
        if pob_seznam:
            ph_pob = ','.join(['%s'] * len(pob_seznam))
            kde.append('EXISTS (SELECT 1 FROM ochutnavky_pobocky p '
                       f'WHERE p.ochutnavka_id = o.id AND p.pobocka IN ({ph_pob}))')
            params += pob_seznam
        if od:
            kde.append('o.datum_do >= %s')
            params.append(od)
        if do:
            kde.append('o.datum_od <= %s')
            params.append(do)
        # při vyhledávání se prohledává celá historie (jinak by dotaz
        # „všechny akce dodavatele X“ vrátil jen ty budoucí)
        if jen_aktualni and not hledat:
            kde.append('o.datum_do >= CURDATE()')
        sql = 'SELECT o.* FROM ochutnavky o'
        if kde:
            sql += ' WHERE ' + ' AND '.join(kde)
        sql += ' ORDER BY o.datum_od, o.nazev'
        cur.execute(sql, params)
        akce = cur.fetchall()
        if not akce:
            return []

        ids = [a['id'] for a in akce]
        ph = ','.join(['%s'] * len(ids))
        cur.execute(f'SELECT ochutnavka_id, pobocka FROM ochutnavky_pobocky '
                    f'WHERE ochutnavka_id IN ({ph}) ORDER BY pobocka', ids)
        mapa = {}
        for r in cur.fetchall():
            mapa.setdefault(r['ochutnavka_id'], []).append(r['pobocka'])
        cur.execute(f'SELECT ochutnavka_id, COUNT(*) AS c FROM ochutnavky_prilohy '
                    f'WHERE ochutnavka_id IN ({ph}) GROUP BY ochutnavka_id', ids)
        pocty = {r['ochutnavka_id']: r['c'] for r in cur.fetchall()}
        cur.execute(f'SELECT ochutnavka_id, COUNT(*) AS c FROM ochutnavky_poznamky '
                    f'WHERE ochutnavka_id IN ({ph}) GROUP BY ochutnavka_id', ids)
        pocty_p = {r['ochutnavka_id']: r['c'] for r in cur.fetchall()}

        for a in akce:
            a['pobocky'] = mapa.get(a['id'], [])
            a['prilohy_pocet'] = pocty.get(a['id'], 0)
            a['poznamky_pocet'] = pocty_p.get(a['id'], 0)
        if hledat:
            akce = [a for a in akce if _vyhovuje_hledani(a, hledat)]
        return _serad(akce, razeni)
    except Exception as e:
        print(f'[Ochutnavky] _nacti_akce: {e}')
        return []
    finally:
        if cur:  cur.close()
        if conn: conn.close()


def _uloz_akce(data, pobocky, akce_id=None):
    """Vloží novou nebo přepíše existující akci. Vrátí id, nebo None při chybě."""
    _init_db()
    conn = intranet_data.get_db_connection()
    if not conn:
        return None
    cur = None
    try:
        cur = conn.cursor()
        if akce_id:
            cur.execute("""
                UPDATE ochutnavky SET datum_od=%s, datum_do=%s, nazev=%s, popis=%s,
                       poznamka=%s, divize=%s, dodavatel=%s, sortiment=%s
                WHERE id=%s
            """, (data['datum_od'], data['datum_do'], data['nazev'], data['popis'],
                  data['poznamka'], data['divize'], data['dodavatel'],
                  data['sortiment'], akce_id))
            cur.execute('DELETE FROM ochutnavky_pobocky WHERE ochutnavka_id=%s',
                        (akce_id,))
        else:
            akce_id = str(uuid.uuid4())
            cur.execute("""
                INSERT INTO ochutnavky (id, datum_od, datum_do, nazev, popis, poznamka,
                                        divize, dodavatel, sortiment,
                                        autor_id, autor_jmeno)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (akce_id, data['datum_od'], data['datum_do'], data['nazev'],
                  data['popis'], data['poznamka'], data['divize'], data['dodavatel'],
                  data['sortiment'], data.get('autor_id'), data.get('autor_jmeno')))
        for p in pobocky:
            cur.execute('INSERT INTO ochutnavky_pobocky (id, ochutnavka_id, pobocka) '
                        'VALUES (%s,%s,%s)', (str(uuid.uuid4()), akce_id, p))
        conn.commit()
        return akce_id
    except Exception as e:
        print(f'[Ochutnavky] _uloz_akce: {e}')
        return None
    finally:
        if cur:  cur.close()
        if conn: conn.close()


def _smaz_akce(akce_id):
    _init_db()
    for pr in _nacti_prilohy(akce_id):
        _smaz_soubor(pr['soubor'])
    conn = intranet_data.get_db_connection()
    if not conn:
        return False
    cur = None
    try:
        cur = conn.cursor()
        cur.execute('DELETE FROM ochutnavky_prilohy WHERE ochutnavka_id=%s', (akce_id,))
        cur.execute('DELETE FROM ochutnavky_poznamky WHERE ochutnavka_id=%s', (akce_id,))
        cur.execute('DELETE FROM ochutnavky_pobocky WHERE ochutnavka_id=%s', (akce_id,))
        cur.execute('DELETE FROM ochutnavky WHERE id=%s', (akce_id,))
        conn.commit()
        return True
    except Exception as e:
        print(f'[Ochutnavky] _smaz_akce: {e}')
        return False
    finally:
        if cur:  cur.close()
        if conn: conn.close()


# ═══════════════════════════════════════════════════════════════════════════════
# DATA – přílohy
# ═══════════════════════════════════════════════════════════════════════════════
def _nacti_prilohy(akce_id):
    _init_db()
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    cur = None
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute('SELECT * FROM ochutnavky_prilohy WHERE ochutnavka_id=%s '
                    'ORDER BY datum', (akce_id,))
        return cur.fetchall()
    except Exception:
        return []
    finally:
        if cur:  cur.close()
        if conn: conn.close()


def _uloz_prilohu(akce_id, nazev, obsah, autor):
    """Uloží soubor na disk + záznam do DB. Vrátí (ok, hlaska)."""
    nazev = _bezpecny_nazev(nazev)
    pripona = os.path.splitext(nazev)[1].lower()
    if pripona not in POVOLENE_PRIPONY:
        return False, 'Povolené jsou pouze obrázky a PDF.'
    if len(obsah) > MAX_PRILOHA:
        return False, 'Soubor je větší než 25 MB.'

    os.makedirs(PRILOHY_DIR, exist_ok=True)
    soubor = f'{uuid.uuid4().hex}{pripona}'
    try:
        with open(os.path.join(PRILOHY_DIR, soubor), 'wb') as f:
            f.write(obsah)
    except Exception as e:
        print(f'[Ochutnavky] zápis přílohy: {e}')
        return False, 'Soubor se nepodařilo uložit.'

    conn = intranet_data.get_db_connection()
    if not conn:
        _smaz_soubor(soubor)
        return False, 'Databáze není dostupná.'
    cur = None
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO ochutnavky_prilohy (id, ochutnavka_id, nazev, soubor, autor)
            VALUES (%s,%s,%s,%s,%s)
        """, (str(uuid.uuid4()), akce_id, nazev, soubor, autor))
        conn.commit()
        return True, 'Příloha nahrána.'
    except Exception as e:
        print(f'[Ochutnavky] _uloz_prilohu: {e}')
        _smaz_soubor(soubor)
        return False, 'Přílohu se nepodařilo uložit.'
    finally:
        if cur:  cur.close()
        if conn: conn.close()


def _smaz_soubor(soubor):
    try:
        cesta = os.path.join(PRILOHY_DIR, _bezpecny_nazev(soubor))
        if os.path.exists(cesta):
            os.remove(cesta)
    except Exception as e:
        print(f'[Ochutnavky] _smaz_soubor: {e}')


def _smaz_prilohu(priloha_id):
    _init_db()
    conn = intranet_data.get_db_connection()
    if not conn:
        return False
    cur = None
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute('SELECT soubor FROM ochutnavky_prilohy WHERE id=%s', (priloha_id,))
        row = cur.fetchone()
        cur.execute('DELETE FROM ochutnavky_prilohy WHERE id=%s', (priloha_id,))
        conn.commit()
        if row:
            _smaz_soubor(row['soubor'])
        return True
    except Exception as e:
        print(f'[Ochutnavky] _smaz_prilohu: {e}')
        return False
    finally:
        if cur:  cur.close()
        if conn: conn.close()


# ═══════════════════════════════════════════════════════════════════════════════
# DATA – chat + notifikace
# ═══════════════════════════════════════════════════════════════════════════════
def _nacti_chat():
    _init_db()
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    cur = None
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute('SELECT * FROM ochutnavky_chat ORDER BY datum DESC LIMIT 100')
        return cur.fetchall()
    except Exception:
        return []
    finally:
        if cur:  cur.close()
        if conn: conn.close()


def _pridej_chat(user_id, user_jmeno, text):
    _init_db()
    conn = intranet_data.get_db_connection()
    if not conn:
        return False
    cur = None
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO ochutnavky_chat (id, user_id, user_jmeno, text)
            VALUES (%s,%s,%s,%s)
        """, (str(uuid.uuid4()), user_id, user_jmeno, text))
        conn.commit()
        return True
    except Exception as e:
        print(f'[Ochutnavky] _pridej_chat: {e}')
        return False
    finally:
        if cur:  cur.close()
        if conn: conn.close()


# ═══════════════════════════════════════════════════════════════════════════════
# DATA – poznámky k akci (mohou přidávat zapisovatelé i čtenáři)
# ═══════════════════════════════════════════════════════════════════════════════
def _nacti_poznamky(akce_id):
    _init_db()
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    cur = None
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute('SELECT * FROM ochutnavky_poznamky WHERE ochutnavka_id=%s '
                    'ORDER BY datum ASC', (akce_id,))
        return cur.fetchall()
    except Exception:
        return []
    finally:
        if cur:  cur.close()
        if conn: conn.close()


def _pridej_poznamku(akce_id, user_id, user_jmeno, text):
    _init_db()
    conn = intranet_data.get_db_connection()
    if not conn:
        return False
    cur = None
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO ochutnavky_poznamky
                   (id, ochutnavka_id, user_id, user_jmeno, text)
            VALUES (%s,%s,%s,%s,%s)
        """, (str(uuid.uuid4()), akce_id, user_id, user_jmeno, text))
        conn.commit()
        return True
    except Exception as e:
        print(f'[Ochutnavky] _pridej_poznamku: {e}')
        return False
    finally:
        if cur:  cur.close()
        if conn: conn.close()


def _smaz_poznamku(pozn_id):
    _init_db()
    conn = intranet_data.get_db_connection()
    if not conn:
        return False
    cur = None
    try:
        cur = conn.cursor()
        cur.execute('DELETE FROM ochutnavky_poznamky WHERE id=%s', (pozn_id,))
        conn.commit()
        return True
    except Exception as e:
        print(f'[Ochutnavky] _smaz_poznamku: {e}')
        return False
    finally:
        if cur:  cur.close()
        if conn: conn.close()


# ═══════════════════════════════════════════════════════════════════════════════
# NOTIFIKACE (zvoneček)
# ═══════════════════════════════════════════════════════════════════════════════
def _rozesli(text, krome_id=None, typ='info'):
    """Zapíše notifikaci do zvonečku všem uživatelům s právem k modulu.

    Autor akce (krome_id) notifikaci nedostane. Volá se na pozadí (vlákno).
    """
    try:
        # ziskej_uzivatele_s_pravem vrací dict {id: jméno}
        prijemci = intranet_data.ziskej_uzivatele_s_pravem(
            PERM_ADMIN, PERM_PRISTUP, 'vse')
        for uid in prijemci:
            if krome_id is not None and int(uid) == int(krome_id):
                continue
            intranet_notifikace.pridej(uid, text, typ)
    except Exception as e:
        print(f'[Ochutnavky] _rozesli: {e}')


def _notifikuj(text, krome_id=None, typ='info'):
    """Rozešle notifikaci na pozadí, aby neblokovala UI."""
    threading.Thread(target=_rozesli, args=(text, krome_id, typ),
                     daemon=True).start()


def _zkrat(text, limit=90):
    text = ' '.join((text or '').split())
    return text if len(text) <= limit else text[:limit].rstrip() + '…'


def _notifikuj_pristupove(user_name, krome_id=None):
    """Ruční rozeslání informace o aktualizaci (tlačítko pro Nákup Office)."""
    _rozesli(
        f'🍽️ {user_name} aktualizoval/a Ochutnávky – Maloobchod a Svět '
        f'potravin. Zkontrolujte plán ochutnávek na svých pobočkách.',
        krome_id=krome_id)


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORT DO PDF  (HTML → Chromium/Playwright, stejný postup jako sankce/vizitky)
# ═══════════════════════════════════════════════════════════════════════════════
EXPORT_DIR = 'Exporty_Ochutnavky'
_PDF_LOCK = None

_MESICE = ['leden', 'únor', 'březen', 'duben', 'květen', 'červen', 'červenec',
           'srpen', 'září', 'říjen', 'listopad', 'prosinec']

# Barvy stavů pro tisk (Tailwind třídy v PDF nefungují – čistý inline CSS)
_STAVY_PDF = {
    'probiha':  ('Probíhá',      '#dcfce7', '#15803d', '#86efac'),
    'brzy':     ('Následujících 7 dní', '#fef3c7', '#b45309', '#fcd34d'),
    'budouci':  ('Plánováno',    '#eff6ff', '#1d4ed8', '#bfdbfe'),
    'probehlo': ('Proběhlo',     '#f3f4f6', '#6b7280', '#e5e7eb'),
}


def _get_pdf_lock():
    """Semafor – v jeden okamžik běží jen jedno Chromium (paměť serveru)."""
    global _PDF_LOCK
    if _PDF_LOCK is None:
        _PDF_LOCK = asyncio.Semaphore(1)
    return _PDF_LOCK


def _esc(t):
    return (str(t if t is not None else '')
            .replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;'))


def _pobocky_filtru(filtr):
    """Filtr provozoven jako seznam (podporuje str i list)."""
    p = filtr.get('pobocka')
    if isinstance(p, str):
        return [p]
    return [x for x in (p or []) if x]


def popis_filtru(filtr):
    """Textový popis aktuálně nastaveného filtru do hlavičky PDF."""
    pob = _pobocky_filtru(filtr)
    c = [DIVIZE.get(filtr.get('divize')) if filtr.get('divize')
         else 'Všechny divize',
         ', '.join(pob) if pob else 'Všechny provozovny']
    od, do = filtr.get('od'), filtr.get('do')
    if od and do:
        c.append(f'{_fmt_d(od)} – {_fmt_d(do)}')
    elif od:
        c.append(f'od {_fmt_d(od)}')
    elif do:
        c.append(f'do {_fmt_d(do)}')
    if filtr.get('hledat'):
        c.append(f'hledáno: „{filtr["hledat"]}“')
    c.append('včetně proběhlých'
             if filtr.get('hledat') or not filtr.get('jen_aktualni')
             else 'jen aktuální a budoucí')
    c.append('řazeno: ' + RAZENI_VOLBY.get(filtr.get('razeni'),
                                           RAZENI_VOLBY['datum_asc']))
    return '   •   '.join(c)


def _nazev_souboru(filtr, pripona='pdf'):
    """Název souboru – nese v sobě filtr, ať se exporty nepřepisují."""
    c = ['Ochutnavky']
    if filtr.get('divize'):
        c.append({'MO': 'Maloobchod', 'SP': 'SvetPotravin'}
                 .get(filtr['divize'], filtr['divize']))
    pob = _pobocky_filtru(filtr)
    if len(pob) == 1:
        # bez diakritiky – ať název souboru projde všemi systémy
        c.append(re.sub(r'[^A-Za-z0-9]+', '_',
                        _bez_diakritiky(pob[0])).strip('_'))
    elif pob:
        c.append(f'{len(pob)}_provozoven')
    c.append(datetime.now().strftime('%Y-%m-%d'))
    return '_'.join(c) + '.' + pripona


_PDF_CSS = """
<style>
  @page { size: A4 landscape; }
  * { box-sizing: border-box; }
  body { margin: 0; font-family: 'Segoe UI', 'DejaVu Sans', Arial, sans-serif;
         color: #1f2937; font-size: 9.5pt; }
  .hlav { background: linear-gradient(135deg, #1e3a8a 0%, #2563eb 100%);
          color: #fff; padding: 14px 18px; border-radius: 10px; }
  .hlav h1 { margin: 0; font-size: 17pt; font-weight: 800; letter-spacing: -.3px; }
  .hlav .pod { margin-top: 2px; font-size: 9pt; opacity: .85; }
  .hlav .filtr { margin-top: 9px; font-size: 8.5pt; background: rgba(255,255,255,.16);
                 border: 1px solid rgba(255,255,255,.28); border-radius: 6px;
                 padding: 5px 9px; display: inline-block; }
  .souhrn { display: flex; gap: 8px; margin: 10px 0 4px; flex-wrap: wrap; }
  .kpi { border: 1px solid #e5e7eb; border-radius: 8px; padding: 6px 12px;
         background: #f9fafb; min-width: 92px; }
  .kpi .c { font-size: 15pt; font-weight: 800; color: #1e3a8a; line-height: 1.1; }
  .kpi .p { font-size: 7.5pt; color: #6b7280; text-transform: uppercase;
            letter-spacing: .5px; font-weight: 700; }
  table { width: 100%; border-collapse: collapse; margin-top: 8px; }
  thead { display: table-header-group; }
  th { background: #1e3a8a; color: #fff; font-size: 7.5pt; text-transform: uppercase;
       letter-spacing: .6px; text-align: left; padding: 6px 7px; font-weight: 700; }
  td { padding: 6px 7px; border-bottom: 1px solid #e5e7eb; vertical-align: top; }
  tbody tr:nth-child(even) td { background: #f9fafb; }
  tr { page-break-inside: avoid; }
  .badge { display: inline-block; padding: 2px 7px; border-radius: 999px;
           font-size: 7.5pt; font-weight: 700; border: 1px solid; white-space: nowrap; }
  .nazev { font-weight: 700; color: #111827; }
  .popis { color: #6b7280; font-size: 8pt; margin-top: 1px; }
  .obdobi { font-weight: 700; color: #374151; white-space: nowrap; }
  .chip { display: inline-block; background: #eef2ff; color: #3730a3;
          border: 1px solid #c7d2fe; border-radius: 999px; padding: 1px 6px;
          font-size: 7.5pt; margin: 0 3px 3px 0; white-space: nowrap; }
  .mala { color: #6b7280; font-size: 8pt; }
  .mesic td { background: #eef2ff !important; color: #1e3a8a; font-weight: 800;
              font-size: 8.5pt; text-transform: uppercase; letter-spacing: .8px;
              padding: 5px 7px; border-bottom: 2px solid #c7d2fe; }
  .sekce { margin-top: 14px; page-break-inside: avoid; }
  .sekce h2 { margin: 0; font-size: 11pt; color: #1e3a8a; font-weight: 800;
              border-left: 4px solid #2563eb; padding-left: 8px; }
  .sekce h2 span { font-size: 8pt; color: #6b7280; font-weight: 600; }
  .pozn { color: #92400e; font-size: 8pt; margin-top: 2px; }
  .prazdno { color: #9ca3af; padding: 20px; text-align: center; }
</style>
"""


def _badge(stav):
    popisek, bg, fg, br = _STAVY_PDF[stav]
    return (f'<span class="badge" style="background:{bg};color:{fg};'
            f'border-color:{br}">{popisek}</span>')


def _bunky_akce(a, s_pobockami=True):
    """Buňky jednoho řádku tabulky (bez <tr>)."""
    stav = stav_akce(a['datum_od'], a['datum_do'])
    popis = _zkrat(a.get('popis'), 150)
    pozn = _zkrat(a.get('poznamka'), 110)
    extra = []
    if a.get('prilohy_pocet'):
        extra.append(f'{a["prilohy_pocet"]} příl.')
    if a.get('poznamky_pocet'):
        extra.append(f'{a["poznamky_pocet"]} pozn.')
    h = [f'<td>{_badge(stav)}</td>',
         f'<td class="obdobi">{_esc(obdobi_text(a["datum_od"], a["datum_do"]))}</td>',
         '<td><div class="nazev">' + _esc(a['nazev']) + '</div>'
         + (f'<div class="popis">{_esc(popis)}</div>' if popis else '')
         + (f'<div class="pozn">Poznámka: {_esc(pozn)}</div>' if pozn else '')
         + '</td>',
         f'<td class="mala">{_esc(DIVIZE.get(a["divize"], a["divize"]))}</td>',
         f'<td class="mala">{_esc(a.get("dodavatel"))}</td>',
         f'<td class="mala">{_esc(a.get("sortiment"))}</td>']
    if s_pobockami:
        h.append('<td>' + ''.join(f'<span class="chip">{_esc(p)}</span>'
                                  for p in a['pobocky']) + '</td>')
    h.append(f'<td class="mala">{_esc(", ".join(extra))}</td>')
    return ''.join(h)


def _hlavicka_tabulky(s_pobockami=True):
    sl = ['<th style="width:78px">Stav</th>',
          '<th style="width:112px">Období</th>',
          '<th>Ochutnávka</th>',
          '<th style="width:96px">Divize</th>',
          '<th style="width:104px">Dodavatel</th>',
          '<th style="width:104px">Sortiment</th>']
    if s_pobockami:
        sl.append('<th style="width:250px">Provozovny</th>')
    sl.append('<th style="width:62px">Přílohy</th>')
    return '<thead><tr>' + ''.join(sl) + '</tr></thead>'


def pdf_html(akce, filtr, user_name):
    """Kompletní HTML exportu – hlavička, souhrn a tabulka dle zvoleného pohledu."""
    dnes = date.today()
    pocty_stavu = {}
    pobocky_vsech = set()
    for a in akce:
        s = stav_akce(a['datum_od'], a['datum_do'], dnes)
        pocty_stavu[s] = pocty_stavu.get(s, 0) + 1
        pobocky_vsech.update(a['pobocky'])
    data_od = min((_na_datum(a['datum_od']) for a in akce), default=None)
    data_do = max((_na_datum(a['datum_do']) or _na_datum(a['datum_od'])
                   for a in akce), default=None)

    kpi = [('Ochutnávek', len(akce)), ('Provozoven', len(pobocky_vsech))]
    for klic in ('probiha', 'brzy', 'budouci', 'probehlo'):
        if pocty_stavu.get(klic):
            kpi.append((_STAVY_PDF[klic][0], pocty_stavu[klic]))
    souhrn = ''.join(f'<div class="kpi"><div class="c">{v}</div>'
                     f'<div class="p">{_esc(p)}</div></div>' for p, v in kpi)

    obdobi = (f'Období plánu: {_fmt_d(data_od)} – {_fmt_d(data_do)}'
              if data_od else '')

    telo = []
    if filtr.get('pohled') == 'pobocka':
        # Seskupeno po provozovnách
        pobocky = (_pobocky_filtru(filtr)
                   or pobocky_pro_divizi(filtr.get('divize')))
        for p in pobocky:
            akce_p = [a for a in akce if p in a['pobocky']]
            if not akce_p:
                continue
            radky = ''.join(f'<tr>{_bunky_akce(a, s_pobockami=False)}</tr>'
                            for a in akce_p)
            telo.append(
                f'<div class="sekce"><h2>{_esc(p)} '
                f'<span>({len(akce_p)} ochutnávek)</span></h2>'
                f'<table>{_hlavicka_tabulky(False)}<tbody>{radky}</tbody></table>'
                f'</div>')
    else:
        # Chronologicky, se zvýrazněním měsíců
        radky, mesic = [], None
        for a in akce:
            d = _na_datum(a['datum_od'])
            klic = (d.year, d.month) if d else None
            if klic and klic != mesic:
                mesic = klic
                radky.append(f'<tr class="mesic"><td colspan="8">'
                             f'{_MESICE[klic[1] - 1]} {klic[0]}</td></tr>')
            radky.append(f'<tr>{_bunky_akce(a)}</tr>')
        telo.append(f'<table>{_hlavicka_tabulky()}'
                    f'<tbody>{"".join(radky)}</tbody></table>')

    if not telo:
        telo = ['<div class="prazdno">Žádné ochutnávky neodpovídají filtru.</div>']

    return (
        '<!DOCTYPE html><html lang="cs"><head><meta charset="utf-8">'
        '<title>Plán ochutnávek</title>' + _PDF_CSS +
        '</head><body>'
        '<div class="hlav">'
        '<h1>Plán ochutnávek – Maloobchod a Svět potravin</h1>'
        f'<div class="pod">Vygeneroval {_esc(user_name)} '
        f'{datetime.now().strftime("%d.%m.%Y %H:%M")}'
        + (f'   •   {_esc(obdobi)}' if obdobi else '') + '</div>'
        f'<div class="filtr">Filtr: {_esc(popis_filtru(filtr))}</div>'
        '</div>'
        f'<div class="souhrn">{souhrn}</div>'
        + ''.join(telo) +
        '</body></html>'
    )


async def _render_pdf(html: str) -> bytes:
    """Vyrenderuje HTML do PDF přes Playwright/Chromium (max 1 render naráz)."""
    try:
        from playwright.async_api import async_playwright  # noqa
    except ImportError as e:
        raise RuntimeError(
            'Knihovna Playwright není nainstalovaná na serveru.\n'
            'Doinstalujte:\n  pip install playwright\n'
            '  playwright install chromium\n'
            '  (Linux: sudo playwright install-deps chromium)\n'
            f'Detail: {e}')

    async with _get_pdf_lock():
        async with async_playwright() as p:
            try:
                browser = await p.chromium.launch()
            except Exception as e:
                raise RuntimeError(
                    f'Nepodařilo se spustit Chromium: {e}\n'
                    'Spusťte na serveru: playwright install chromium')
            try:
                page = await browser.new_page()
                try:
                    await page.set_content(html, wait_until='load')
                    return await page.pdf(
                        format='A4', landscape=True, print_background=True,
                        margin={'top': '10mm', 'right': '9mm',
                                'bottom': '13mm', 'left': '9mm'},
                        display_header_footer=True,
                        header_template='<div></div>',
                        footer_template=(
                            '<div style="width:100%;font-size:7pt;color:#9ca3af;'
                            'font-family:Arial,sans-serif;padding:0 9mm;'
                            'display:flex;justify-content:space-between;">'
                            '<span>Ochutnávky – Maloobchod a Svět potravin</span>'
                            '<span>Strana <span class="pageNumber"></span> / '
                            '<span class="totalPages"></span></span></div>'))
                finally:
                    await page.close()
            finally:
                await browser.close()


def _uloz_export(data: bytes, nazev: str) -> str:
    """Uloží PDF na server (stahuje se po HTTP, ne přes WS) a uklidí staré exporty."""
    os.makedirs(EXPORT_DIR, exist_ok=True)
    ted = time.time()
    for f in os.listdir(EXPORT_DIR):
        cesta_f = os.path.join(EXPORT_DIR, f)
        try:
            if os.path.isfile(cesta_f) and ted - os.path.getmtime(cesta_f) > 3600:
                os.remove(cesta_f)
        except Exception:
            pass
    cesta = os.path.join(EXPORT_DIR, f'{int(ted * 1000)}_{nazev}')
    with open(cesta, 'wb') as fh:
        fh.write(data)
    return cesta


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORT DO EXCELU  (openpyxl – 3 listy: Plán / Po provozovnách / Souhrn)
# ═══════════════════════════════════════════════════════════════════════════════
# stav → (výplň, barva písma) v Excelu
_XLS_STAVY = {
    'probiha':  ('Probíhá',              'DCFCE7', '15803D'),
    'brzy':     ('Následujících 7 dní',  'FEF3C7', 'B45309'),
    'budouci':  ('Plánováno',            'EFF6FF', '1D4ED8'),
    'probehlo': ('Proběhlo',             'F3F4F6', '6B7280'),
}
_XLS_MODRA = '1E3A8A'


def xlsx_bytes(akce, filtr, user_name):
    """Sestaví .xlsx dle aktuálního filtru (běží v threadu). Vrací bytes."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    dnes = date.today()
    bily_bold = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill = PatternFill('solid', fgColor=_XLS_MODRA)
    tenka = Side(style='thin', color='D1D5DB')
    ramecek = Border(left=tenka, right=tenka, top=tenka, bottom=tenka)
    zebra = PatternFill('solid', fgColor='F8FAFC')
    stred = Alignment(horizontal='center', vertical='center')
    zalom = Alignment(vertical='top', wrap_text=True)
    vlevo = Alignment(vertical='center')

    def _hlavicka(ws, sloupce, podnadpis):
        """Titulek + popis filtru + hlavička tabulky. Data začínají řádkem 7."""
        posl = get_column_letter(len(sloupce))
        ws.merge_cells(f'A1:{posl}1')
        c = ws['A1']
        c.value = 'Ochutnávky – Maloobchod a Svět potravin'
        c.font = Font(bold=True, size=16, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor=_XLS_MODRA)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[1].height = 30

        ws.merge_cells(f'A2:{posl}2')
        ws['A2'].value = podnadpis
        ws['A2'].font = Font(bold=True, size=10, color='374151')

        ws.merge_cells(f'A3:{posl}3')
        ws['A3'].value = 'Filtr:   ' + popis_filtru(filtr)
        ws['A3'].font = Font(size=9, color='6B7280')

        ws.merge_cells(f'A4:{posl}4')
        ws['A4'].value = (f'Vygeneroval: {user_name}   •   '
                          f'{datetime.now().strftime("%d.%m.%Y %H:%M")}   •   '
                          f'záznamů: {len(akce)}')
        ws['A4'].font = Font(size=9, color='6B7280')

        for i, (nazev, sirka) in enumerate(sloupce, start=1):
            b = ws.cell(row=6, column=i, value=nazev)
            b.font, b.fill, b.border = bily_bold, hdr_fill, ramecek
            b.alignment = Alignment(horizontal='center', vertical='center',
                                    wrap_text=True)
            ws.column_dimensions[get_column_letter(i)].width = sirka
        ws.row_dimensions[6].height = 26
        ws.freeze_panes = 'A7'
        ws.auto_filter.ref = f'A6:{posl}{6 + max(len(akce), 1)}'
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_title_rows = '6:6'

    def _zapis(ws, r, hodnoty, stav=None, vyska=None):
        """Jeden datový řádek – rámečky, zebra, obarvený sloupec Stav."""
        for i, v in enumerate(hodnoty, start=1):
            b = ws.cell(row=r, column=i, value=v)
            b.border = ramecek
            b.alignment = zalom if isinstance(v, str) and len(v) > 40 else vlevo
            if isinstance(v, date):
                b.number_format = 'DD.MM.YYYY'
                b.alignment = stred
            elif isinstance(v, int):
                b.alignment = stred
            if r % 2 == 0:
                b.fill = zebra
        if stav:
            popis, vypln, pismo = _XLS_STAVY[stav]
            b = ws.cell(row=r, column=1)
            b.value = popis
            b.fill = PatternFill('solid', fgColor=vypln)
            b.font = Font(bold=True, color=pismo, size=9)
            b.alignment = stred
        if vyska:
            ws.row_dimensions[r].height = vyska

    wb = openpyxl.Workbook()

    # ── List 1: Plán ochutnávek ────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Plán ochutnávek'
    _hlavicka(ws, [('Stav', 20), ('Datum od', 12), ('Datum do', 12),
                   ('Název akce', 32), ('Divize', 24), ('Dodavatel', 22),
                   ('Sortiment', 22), ('Provozovny', 46), ('Počet\nprovozoven', 11),
                   ('Popis', 40), ('Poznámka', 30), ('Příloh', 8),
                   ('Poznámek', 10), ('Zadal', 22)],
              'Plán ochutnávek na provozovnách')
    r = 7
    for a in akce:
        od, do = _na_datum(a['datum_od']), _na_datum(a['datum_do'])
        _zapis(ws, r, [
            '', od, do or od, a['nazev'] or '',
            DIVIZE.get(a['divize'], a['divize'] or ''),
            a.get('dodavatel') or '', a.get('sortiment') or '',
            ', '.join(a['pobocky']), len(a['pobocky']),
            a.get('popis') or '', a.get('poznamka') or '',
            int(a.get('prilohy_pocet') or 0), int(a.get('poznamky_pocet') or 0),
            a.get('autor_jmeno') or '',
        ], stav=stav_akce(a['datum_od'], a['datum_do'], dnes))
        r += 1

    # ── List 2: Po provozovnách (1 řádek = akce × provozovna, pro kontingenčky)
    ws2 = wb.create_sheet('Po provozovnách')
    _hlavicka(ws2, [('Stav', 20), ('Provozovna', 30), ('Datum od', 12),
                    ('Datum do', 12), ('Název akce', 32), ('Divize', 24),
                    ('Dodavatel', 22), ('Sortiment', 22), ('Poznámka', 30)],
              'Rozpad plánu po jednotlivých provozovnách')
    radky = []
    for a in akce:
        for p in a['pobocky']:
            radky.append((p, _na_datum(a['datum_od']), a))
    radky.sort(key=lambda x: (x[0], x[1] or date.min))
    r = 7
    for p, od, a in radky:
        do = _na_datum(a['datum_do'])
        _zapis(ws2, r, [
            '', p, od, do or od, a['nazev'] or '',
            DIVIZE.get(a['divize'], a['divize'] or ''),
            a.get('dodavatel') or '', a.get('sortiment') or '',
            a.get('poznamka') or '',
        ], stav=stav_akce(a['datum_od'], a['datum_do'], dnes))
        r += 1
    ws2.auto_filter.ref = f'A6:I{6 + max(len(radky), 1)}'

    # ── List 3: Souhrn ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet('Souhrn')
    ws3.merge_cells('A1:C1')
    ws3['A1'].value = 'Souhrn exportu'
    ws3['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    ws3['A1'].fill = PatternFill('solid', fgColor=_XLS_MODRA)
    ws3['A1'].alignment = Alignment(vertical='center', indent=1)
    ws3.row_dimensions[1].height = 30
    ws3.merge_cells('A2:C2')
    ws3['A2'].value = 'Filtr:   ' + popis_filtru(filtr)
    ws3['A2'].font = Font(size=9, color='6B7280')
    for sl, sirka in (('A', 42), ('B', 14), ('C', 14)):
        ws3.column_dimensions[sl].width = sirka

    pocty_stavu, per_pob, per_dod = {}, {}, {}
    for a in akce:
        s = stav_akce(a['datum_od'], a['datum_do'], dnes)
        pocty_stavu[s] = pocty_stavu.get(s, 0) + 1
        for p in a['pobocky']:
            per_pob[p] = per_pob.get(p, 0) + 1
        d = (a.get('dodavatel') or '—').strip()
        per_dod[d] = per_dod.get(d, 0) + 1

    r = 4

    def _sekce(nadpis, dvojice):
        nonlocal r
        b = ws3.cell(row=r, column=1, value=nadpis)
        b.font, b.fill, b.border = bily_bold, hdr_fill, ramecek
        b2 = ws3.cell(row=r, column=2, value='Počet')
        b2.font, b2.fill, b2.border = bily_bold, hdr_fill, ramecek
        b2.alignment = stred
        r += 1
        for k, v in dvojice:
            c1 = ws3.cell(row=r, column=1, value=k)
            c2 = ws3.cell(row=r, column=2, value=v)
            c1.border = c2.border = ramecek
            c2.alignment = stred
            r += 1
        r += 1

    _sekce('Celkem', [('Ochutnávek', len(akce)),
                      ('Dotčených provozoven', len(per_pob)),
                      ('Dodavatelů', len(per_dod))])
    _sekce('Podle stavu', [(_XLS_STAVY[k][0], pocty_stavu[k])
                           for k in ('probiha', 'brzy', 'budouci', 'probehlo')
                           if pocty_stavu.get(k)])
    _sekce('Podle provozovny', sorted(per_pob.items()))
    _sekce('Podle dodavatele',
           sorted(per_dod.items(), key=lambda x: (-x[1], x[0])))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# UI
# ═══════════════════════════════════════════════════════════════════════════════
def _omez_datum_od(kalendar, min_datum):
    """V kalendáři zůstanou klikatelná jen data od min_datum výš.
    min_datum = None → omezení zrušeno (filtr musí umět i minulost)."""
    if min_datum is None:
        kalendar.props(remove=':options')
    else:
        kalendar.props(f''':options="d => d >= '{min_datum.strftime('%Y/%m/%d')}'"''')


def _pole_datum(label, sirka='w-44', min_datum=None):
    """Datum výhradně přes kalendář – ruční psaní je zakázáno (readonly)."""
    with ui.input(label).classes(f'{sirka} bg-white cursor-pointer') \
            .props('outlined dense readonly input-class=cursor-pointer') as inp:
        inp.on('click', lambda: menu.open())
        with inp.add_slot('append'):
            with ui.row().classes('items-center gap-1 flex-nowrap'):
                ui.icon('close') \
                    .classes('cursor-pointer text-gray-400 hover:text-red-500') \
                    .bind_visibility_from(inp, 'value') \
                    .on('click.stop', lambda: inp.set_value(None))
                ui.icon('edit_calendar') \
                    .classes('cursor-pointer text-gray-500 hover:text-blue-500')
        with ui.menu() as menu:
            kal = ui.date().bind_value(inp).props('today-btn') \
                .on('update:model-value', lambda: menu.close())
    inp.kalendar = kal
    if min_datum is not None:
        _omez_datum_od(kal, min_datum)
    return inp


@refreshable_na_klienta
def vykresli_ochutnavky(user_id, user_name, vsechna_prava):
    _init_db()
    os.makedirs(PRILOHY_DIR, exist_ok=True)

    je_admin   = 'vse' in vsechna_prava or PERM_ADMIN in vsechna_prava
    ma_pristup = je_admin or PERM_PRISTUP in vsechna_prava

    if not ma_pristup:
        with ui.column().classes('w-full items-center py-24 gap-4'):
            ui.icon('lock').classes('text-5xl text-gray-300')
            ui.label('Nemáte přístup k ochutnávkám.').classes('text-xl text-gray-500')
        return

    filtr = {'divize': None, 'pobocka': [], 'od': None, 'do': None,
             'jen_aktualni': True, 'pohled': 'datum', 'razeni': 'datum_asc',
             'hledat': None}

    # ── Dialog: formulář akce (nová i úprava) ────────────────────────────────
    ctx = {'akce_id': None, 'vybrane': set(), 'min_datum': date.today()}

    with ui.dialog().props('persistent') as dlg_form, \
            ui.card().classes('w-full !max-w-3xl p-6 gap-3'):
        form_titulek = ui.label('Nová ochutnávka') \
            .classes('text-xl font-black text-blue-900')

        ui.label('Období akce').classes('text-xs font-bold text-gray-600 uppercase')
        with ui.row().classes('items-center gap-3'):
            f_od = _pole_datum('Datum od', min_datum=date.today())
            f_do = _pole_datum('Datum do (nepovinné)', min_datum=date.today())
            ui.label('Jednodenní akce → vyplňte pouze „Datum od“.') \
                .classes('text-xs text-gray-400')

        f_nazev = ui.input('Název akce').classes('w-full').props('outlined dense')
        f_popis = ui.textarea('Popis ochutnávky') \
            .classes('w-full').props('outlined dense autogrow')
        with ui.row().classes('w-full gap-3'):
            f_dodavatel = ui.input('Dodavatel').classes('flex-1').props('outlined dense')
            f_sortiment = ui.input('Sortiment').classes('flex-1').props('outlined dense')
        f_poznamka = ui.textarea('Poznámka') \
            .classes('w-full').props('outlined dense autogrow')

        ui.separator()
        ui.label('Divize').classes('text-xs font-bold text-gray-600 uppercase')

        def _zmena_divize(_=None):
            povolene = set(pobocky_pro_divizi(f_divize.value))
            ctx['vybrane'] &= povolene
            _vyber_pobocek.refresh()

        f_divize = ui.toggle(DIVIZE, value='MO', on_change=_zmena_divize) \
            .props('no-caps unelevated').classes('w-full')

        with ui.row().classes('w-full items-center gap-2 mt-1'):
            ui.label('Provozovny').classes('text-xs font-bold text-gray-600 uppercase')
            ui.button('Vybrat vše', on_click=lambda: (
                ctx['vybrane'].update(pobocky_pro_divizi(f_divize.value)),
                _vyber_pobocek.refresh())).props('flat dense no-caps') \
                .classes('text-xs text-blue-600')
            ui.button('Zrušit výběr', on_click=lambda: (
                ctx['vybrane'].clear(),
                _vyber_pobocek.refresh())).props('flat dense no-caps') \
                .classes('text-xs text-gray-500')

        @ui.refreshable
        def _vyber_pobocek():
            def _prepni(p):
                ctx['vybrane'].symmetric_difference_update({p})
                _vyber_pobocek.refresh()

            for nadpis, seznam in (('Maloobchod', POBOCKY_MO),
                                   ('Svět potravin – C&C', POBOCKY_SP)):
                if not set(seznam) & set(pobocky_pro_divizi(f_divize.value)):
                    continue
                ui.label(nadpis).classes('text-xs font-bold text-gray-500 mt-2')
                with ui.row().classes('w-full gap-1 flex-wrap'):
                    for p in seznam:
                        vybrano = p in ctx['vybrane']
                        # ui.label místo ui.button — Quasar q-btn vnucuje bílý text
                        ui.label(p).classes(
                            'text-xs rounded-full px-3 py-1 border cursor-pointer '
                            'select-none whitespace-nowrap transition ' + (
                                'bg-green-600 text-white border-green-700 font-bold'
                                if vybrano else
                                'bg-white text-gray-700 border-gray-300 '
                                'hover:bg-green-50 hover:border-green-400')) \
                            .on('click', lambda _, _p=p: _prepni(_p))
            ui.label(f'Vybráno provozoven: {len(ctx["vybrane"])}') \
                .classes('text-xs text-gray-500 mt-1')

        with ui.column().classes('w-full gap-0 max-h-64 overflow-auto'):
            _vyber_pobocek()

        def _uloz():
            od = _na_datum(f_od.value)
            do = _na_datum(f_do.value) or od
            if not od:
                ui.notify('Vyplňte datum od.', type='warning'); return
            if do < od:
                ui.notify('Datum do nesmí být dříve než datum od.', type='warning'); return
            if od < ctx['min_datum']:
                ui.notify('Datum akce nelze zadat do minulosti.', type='warning'); return
            if not (f_nazev.value or '').strip():
                ui.notify('Vyplňte název akce.', type='warning'); return
            if not ctx['vybrane']:
                ui.notify('Vyberte alespoň jednu provozovnu.', type='warning'); return

            ok = _uloz_akce({
                'datum_od': od, 'datum_do': do,
                'nazev': f_nazev.value.strip(),
                'popis': (f_popis.value or '').strip(),
                'poznamka': (f_poznamka.value or '').strip(),
                'divize': f_divize.value,
                'dodavatel': (f_dodavatel.value or '').strip(),
                'sortiment': (f_sortiment.value or '').strip(),
                'autor_id': user_id, 'autor_jmeno': user_name,
            }, sorted(ctx['vybrane']), ctx['akce_id'])
            if not ok:
                ui.notify('Uložení se nezdařilo.', type='negative'); return
            je_uprava = bool(ctx['akce_id'])
            nazev_akce = f_nazev.value.strip()
            dlg_form.close()
            ui.notify('✅ Ochutnávka uložena.', type='positive')
            _notifikuj(
                (f'✏️ {user_name} upravil/a ochutnávku „{nazev_akce}“ '
                 f'({obdobi_text(od, do)}).') if je_uprava else
                (f'🍽️ {user_name} zadal/a novou ochutnávku „{nazev_akce}“ '
                 f'({obdobi_text(od, do)}).'),
                krome_id=user_id)
            _prehled.refresh()
            _rychly_vyber.refresh()
            _upozorneni.refresh()

        with ui.row().classes('w-full justify-end gap-2 mt-2'):
            ui.button('Zrušit', on_click=dlg_form.close).props('flat') \
                .classes('text-gray-500')
            ui.button('Potvrdit a vložit', icon='check', on_click=_uloz) \
                .classes('bg-blue-600 text-white font-bold px-5')

    def _sladit_datum_do():
        """„Datum do“ nesmí být dřív než „Datum od“ (a nikdy v minulosti)."""
        od = _na_datum(f_od.value)
        _omez_datum_od(f_do.kalendar, max(od, ctx['min_datum']) if od
                       else ctx['min_datum'])

    f_od.on_value_change(lambda _: _sladit_datum_do())

    def _otevri_form(akce=None):
        ctx['akce_id'] = akce['id'] if akce else None
        ctx['vybrane'] = set(akce['pobocky']) if akce else set()
        form_titulek.set_text('Úprava ochutnávky' if akce else 'Nová ochutnávka')
        # nová akce nesmí začínat v minulosti; u úpravy staré akce
        # se původní (už proběhlé) datum musí dát znovu uložit
        dnes = date.today()
        puvodni_od = _na_datum(akce['datum_od']) if akce else None
        ctx['min_datum'] = min(dnes, puvodni_od) if puvodni_od else dnes
        _omez_datum_od(f_od.kalendar, ctx['min_datum'])
        f_od.set_value(str(_na_datum(akce['datum_od'])) if akce else None)
        f_do.set_value(str(_na_datum(akce['datum_do'])) if akce else None)
        _sladit_datum_do()
        f_nazev.set_value(akce['nazev'] if akce else '')
        f_popis.set_value((akce.get('popis') or '') if akce else '')
        f_poznamka.set_value((akce.get('poznamka') or '') if akce else '')
        f_dodavatel.set_value((akce.get('dodavatel') or '') if akce else '')
        f_sortiment.set_value((akce.get('sortiment') or '') if akce else '')
        f_divize.set_value(akce['divize'] if akce else 'MO')
        _vyber_pobocek.refresh()
        dlg_form.open()

    # ── Dialog: detail akce + přílohy ────────────────────────────────────────
    det = {'akce': None}
    with ui.dialog() as dlg_detail, ui.card().classes('w-full !max-w-2xl p-6 gap-2'):
        @ui.refreshable
        def _detail_obsah():
            a = det['akce']
            if not a:
                return
            ui.label(a['nazev']).classes('text-xl font-black text-blue-900')
            ui.label(f'{obdobi_text(a["datum_od"], a["datum_do"])}  ·  '
                     f'{DIVIZE.get(a["divize"], a["divize"])}') \
                .classes('text-sm text-gray-500')
            if a.get('popis'):
                ui.label(a['popis']).classes('text-sm text-gray-700 whitespace-pre-wrap')
            with ui.row().classes('gap-4 flex-wrap'):
                if a.get('dodavatel'):
                    ui.label(f'Dodavatel: {a["dodavatel"]}').classes('text-sm text-gray-600')
                if a.get('sortiment'):
                    ui.label(f'Sortiment: {a["sortiment"]}').classes('text-sm text-gray-600')
            if a.get('poznamka'):
                with ui.card().classes('w-full bg-amber-50 border border-amber-200 p-3'):
                    ui.label('Poznámka').classes('text-xs font-bold text-amber-700')
                    ui.label(a['poznamka']) \
                        .classes('text-sm text-gray-700 whitespace-pre-wrap')

            ui.label('Provozovny').classes('text-xs font-bold text-gray-500 uppercase mt-2')
            with ui.row().classes('w-full gap-1 flex-wrap'):
                for p in a['pobocky']:
                    ui.label(p).classes('text-xs bg-gray-100 text-gray-700 '
                                        'rounded-full px-3 py-1')

            ui.separator().classes('my-2')
            ui.label('Přílohy (foto, PDF)') \
                .classes('text-xs font-bold text-gray-500 uppercase')
            prilohy = _nacti_prilohy(a['id'])
            if not prilohy:
                ui.label('Zatím žádné přílohy.').classes('text-sm text-gray-400 italic')
            for pr in prilohy:
                with ui.row().classes('w-full items-center gap-2'):
                    ui.icon('attach_file').classes('text-gray-400')
                    ui.link(pr['nazev'], f'/{PRILOHY_DIR}/{pr["soubor"]}',
                            new_tab=True).classes('text-sm text-blue-600 flex-1')
                    ui.label(f'{pr["autor"] or ""} · {_fmt_dt(pr["datum"])}') \
                        .classes('text-xs text-gray-400')
                    if je_admin or (pr['autor'] or '') == user_name:
                        ui.button(icon='delete', on_click=lambda _, _p=pr: (
                            _smaz_prilohu(_p['id']), _detail_obsah.refresh(),
                            _prehled.refresh())) \
                            .props('flat dense round').classes('text-red-400')

            async def _nahraj(e):
                obsah = await e.file.read()
                a_akt = det['akce']
                ok, hlaska = _uloz_prilohu(a_akt['id'], e.file.name,
                                           obsah, user_name)
                ui.notify(hlaska, type='positive' if ok else 'negative')
                if ok:
                    _notifikuj(
                        f'📎 {user_name} přidal/a přílohu „{e.file.name}“ '
                        f'k ochutnávce „{a_akt["nazev"]}“.', krome_id=user_id)
                    _detail_obsah.refresh()
                    _prehled.refresh()

            ui.upload(auto_upload=True, on_upload=_nahraj,
                      max_file_size=MAX_PRILOHA) \
                .props('accept="image/*,.pdf" label="Nahrát přílohu"') \
                .classes('w-full')

            # ── Poznámky k akci (zapisovatelé i čtenáři) ──────────────────
            ui.separator().classes('my-2')
            ui.label('Poznámky k akci') \
                .classes('text-xs font-bold text-gray-500 uppercase')
            poznamky = _nacti_poznamky(a['id'])
            if not poznamky:
                ui.label('Zatím žádné poznámky.') \
                    .classes('text-sm text-gray-400 italic')
            for pz in poznamky:
                with ui.row().classes('w-full items-start gap-2 bg-gray-50 '
                                      'rounded-lg px-3 py-2'):
                    with ui.column().classes('gap-0 flex-1 min-w-0'):
                        ui.label(pz['user_jmeno'] or '?') \
                            .classes('text-xs font-bold text-blue-700')
                        ui.label(pz['text']) \
                            .classes('text-sm text-gray-700 whitespace-pre-wrap')
                    ui.label(_fmt_dt(pz['datum'])) \
                        .classes('text-xs text-gray-400 flex-shrink-0')
                    if je_admin or pz.get('user_id') == user_id:
                        ui.button(icon='delete', on_click=lambda _, _z=pz: (
                            _smaz_poznamku(_z['id']), _detail_obsah.refresh())) \
                            .props('flat dense round').classes('text-red-400')

            with ui.row().classes('w-full gap-2 items-center mt-1'):
                pozn_input = ui.input(placeholder='Přidat poznámku k akci…') \
                    .classes('flex-1').props('outlined dense')

                def _pridej_pozn():
                    txt = (pozn_input.value or '').strip()
                    if not txt:
                        return
                    a_akt = det['akce']
                    if not _pridej_poznamku(a_akt['id'], user_id, user_name, txt):
                        ui.notify('Poznámku se nepodařilo uložit.', type='negative')
                        return
                    pozn_input.set_value('')
                    _notifikuj(
                        f'📝 {user_name} přidal/a poznámku k ochutnávce '
                        f'„{a_akt["nazev"]}“: {_zkrat(txt)}', krome_id=user_id)
                    _detail_obsah.refresh()

                pozn_input.on('keydown.enter', lambda: _pridej_pozn())
                ui.button(icon='send', on_click=_pridej_pozn) \
                    .props('flat round').classes('text-blue-500')

            with ui.row().classes('w-full justify-end gap-2 mt-2'):
                if je_admin:
                    ui.button('Upravit akci', icon='edit', on_click=lambda: (
                        dlg_detail.close(), _otevri_form(det['akce']))) \
                        .props('flat').classes('text-blue-600')
                ui.button('Zavřít', on_click=dlg_detail.close).props('flat') \
                    .classes('text-gray-500')

        _detail_obsah()

    def _otevri_detail(a):
        det['akce'] = a
        _detail_obsah.refresh()
        dlg_detail.open()

    smaz_ctx = {'akce': None}
    with ui.dialog() as dlg_smaz, ui.card().classes('p-6 gap-3'):
        smaz_popis = ui.label('').classes('text-sm text-gray-700')

        def _potvrd_smazani():
            a = smaz_ctx['akce']
            dlg_smaz.close()
            if a and _smaz_akce(a['id']):
                ui.notify('Ochutnávka smazána.', type='warning')
                _notifikuj(
                    f'🗑️ {user_name} smazal/a ochutnávku „{a["nazev"]}“ '
                    f'({obdobi_text(a["datum_od"], a["datum_do"])}).',
                    krome_id=user_id, typ='warning')
                _prehled.refresh()
                _rychly_vyber.refresh()
                _upozorneni.refresh()
            else:
                ui.notify('Smazání se nezdařilo.', type='negative')

        with ui.row().classes('w-full justify-end gap-2'):
            ui.button('Zrušit', on_click=dlg_smaz.close).props('flat') \
                .classes('text-gray-500')
            ui.button('Smazat', icon='delete', on_click=_potvrd_smazani) \
                .classes('bg-red-600 text-white font-bold')

    def _smaz(a):
        smaz_ctx['akce'] = a
        smaz_popis.set_text(
            f'Opravdu smazat ochutnávku „{a["nazev"]}“ '
            f'({obdobi_text(a["datum_od"], a["datum_do"])})? '
            f'Smažou se i její přílohy.')
        dlg_smaz.open()

    # ── Řádek akce ───────────────────────────────────────────────────────────
    def _radek(a, zobraz_pobocky=True):
        stav = stav_akce(a['datum_od'], a['datum_do'])
        popisek, barva = _STAVY[stav]
        with ui.row().classes(
                'w-full items-center gap-3 px-3 py-2 rounded-xl border '
                f'{barva} hover:shadow-sm cursor-pointer') \
                .on('click', lambda _, _a=a: _otevri_detail(_a)):
            ui.label(popisek).classes('text-xs font-bold w-28 flex-shrink-0')
            ui.label(obdobi_text(a['datum_od'], a['datum_do'])) \
                .classes('text-sm font-bold w-48 flex-shrink-0 text-gray-700')
            with ui.column().classes('gap-0 flex-1 min-w-0'):
                ui.label(a['nazev']).classes('text-sm font-bold text-gray-800 truncate')
                detaily = ' · '.join(x for x in (
                    DIVIZE.get(a['divize'], a['divize']),
                    a.get('dodavatel'), a.get('sortiment')) if x)
                ui.label(detaily).classes('text-xs text-gray-500 truncate')
            if zobraz_pobocky:
                with ui.row().classes('gap-1 flex-wrap max-w-md justify-end'):
                    for p in a['pobocky'][:6]:
                        ui.label(p).classes('text-xs bg-white/70 text-gray-600 '
                                            'rounded-full px-2 py-0.5 border border-gray-200')
                    if len(a['pobocky']) > 6:
                        ui.label(f'+{len(a["pobocky"]) - 6}') \
                            .classes('text-xs text-gray-500 px-1')
            if a['prilohy_pocet']:
                ui.label(f'📎 {a["prilohy_pocet"]}').classes('text-xs text-gray-500')
            if a.get('poznamky_pocet'):
                ui.label(f'💬 {a["poznamky_pocet"]}').classes('text-xs text-gray-500')
            if a.get('poznamka'):
                ui.icon('sticky_note_2').classes('text-amber-500 text-sm')
            if je_admin:
                # click.stop → neotevře se zároveň detail akce pod tímto řádkem
                ui.icon('delete').classes(
                    'text-red-400 hover:text-red-600 cursor-pointer text-lg') \
                    .on('click.stop', lambda _, _a=a: _smaz(_a))

    # ── Export do PDF (respektuje aktuální filtry) ───────────────────────────
    async def _export_pdf():
        akce = _nacti_akce(divize=filtr['divize'], pobocka=filtr['pobocka'],
                           od=filtr['od'], do=filtr['do'],
                           jen_aktualni=bool(filtr['jen_aktualni']),
                           razeni=filtr['razeni'], hledat=filtr['hledat'])
        if not akce:
            ui.notify('Není co exportovat – filtru neodpovídá žádná ochutnávka.',
                      type='warning')
            return
        notif = ui.notification('Připravuji PDF…', type='ongoing',
                                position='top-right', spinner=True, timeout=None)
        try:
            data = await _render_pdf(pdf_html(akce, filtr, user_name))
            nazev = _nazev_souboru(filtr)
            ui.download.file(_uloz_export(data, nazev), nazev)
            notif.dismiss()
            ui.notify(f'✅ PDF vytvořeno ({len(akce)} ochutnávek).',
                      type='positive')
        except Exception as e:
            notif.dismiss()
            ui.notify(f'Nepodařilo se vytvořit PDF: {e}', type='negative',
                      position='top', timeout=15000, multi_line=True)

    # ── Export do Excelu (respektuje aktuální filtry) ────────────────────────
    async def _export_xlsx():
        akce = _nacti_akce(divize=filtr['divize'], pobocka=filtr['pobocka'],
                           od=filtr['od'], do=filtr['do'],
                           jen_aktualni=bool(filtr['jen_aktualni']),
                           razeni=filtr['razeni'], hledat=filtr['hledat'])
        if not akce:
            ui.notify('Není co exportovat – filtru neodpovídá žádná ochutnávka.',
                      type='warning')
            return
        notif = ui.notification('Připravuji Excel…', type='ongoing',
                                position='top-right', spinner=True, timeout=None)
        try:
            data = await asyncio.to_thread(xlsx_bytes, akce, filtr, user_name)
            nazev = _nazev_souboru(filtr, 'xlsx')
            ui.download.file(_uloz_export(data, nazev), nazev)
            notif.dismiss()
            ui.notify(f'✅ Excel vytvořen ({len(akce)} ochutnávek).',
                      type='positive')
        except Exception as e:
            notif.dismiss()
            ui.notify(f'Nepodařilo se vytvořit Excel: {e}', type='negative',
                      position='top', timeout=15000, multi_line=True)

    # ── Hlavička ─────────────────────────────────────────────────────────────
    with ui.column().classes('w-full gap-0'):
        with ui.row().classes('w-full items-center justify-between '
                              'px-6 py-4 border-b border-gray-200'):
            with ui.column().classes('gap-0'):
                ui.label('Ochutnávky – Maloobchod a Svět potravin') \
                    .classes('text-2xl font-black text-blue-900')
                ui.label('Plán ochutnávek na provozovnách') \
                    .classes('text-sm text-gray-500')
            with ui.row().classes('items-center gap-2'):
                ui.button('Export do PDF', icon='picture_as_pdf',
                          on_click=_export_pdf) \
                    .props('outline').classes('text-red-700 font-bold px-4') \
                    .tooltip('Vyexportuje seznam přesně podle nastavených filtrů')
                ui.button('Export do Excelu', icon='table_view',
                          on_click=_export_xlsx) \
                    .props('outline').classes('text-green-700 font-bold px-4') \
                    .tooltip('Sešit .xlsx (Plán / Po provozovnách / Souhrn) '
                             'podle nastavených filtrů')
                if je_admin:
                    ui.button(
                        icon='send', text='Rozeslat informaci o aktualizaci',
                        on_click=lambda: (
                            threading.Thread(target=_notifikuj_pristupove,
                                             args=(user_name, user_id),
                                             daemon=True).start(),
                            ui.notify('✅ Notifikace rozeslána všem uživatelům '
                                      's přístupem.', type='positive')
                        )).classes('bg-blue-600 text-white font-bold shadow-md px-5')
                    ui.button('Nová ochutnávka', icon='add',
                              on_click=lambda: _otevri_form()) \
                        .classes('bg-green-600 text-white font-bold px-5')

        with ui.column().classes('w-full gap-4 p-6'):

            # ── Upozornění: následujících 7 dní ──────────────────────────────
            @ui.refreshable
            def _upozorneni():
                dnes = date.today()
                nadchazejici = [
                    a for a in _nacti_akce(jen_aktualni=True)
                    if _na_datum(a['datum_od']) <= dnes + timedelta(days=7)
                ]
                if not nadchazejici:
                    return
                with ui.card().classes('w-full bg-amber-50 border-2 border-amber-300 '
                                       'p-4 gap-2 shadow-sm'):
                    with ui.row().classes('items-center gap-2'):
                        ui.icon('notifications_active').classes('text-amber-600')
                        ui.label('Ochutnávky probíhající a v následujících 7 dnech') \
                            .classes('text-sm font-black text-amber-800 uppercase')
                    with ui.row().classes('w-full gap-2 flex-wrap'):
                        for a in nadchazejici:
                            with ui.card().classes(
                                    'p-3 gap-0 bg-white border border-amber-200 '
                                    'cursor-pointer hover:shadow-md min-w-64') \
                                    .on('click', lambda _, _a=a: _otevri_detail(_a)):
                                ui.label(obdobi_text(a['datum_od'], a['datum_do'])) \
                                    .classes('text-xs font-bold text-amber-700')
                                ui.label(a['nazev']) \
                                    .classes('text-sm font-bold text-gray-800')
                                ui.label(', '.join(a['pobocky'])) \
                                    .classes('text-xs text-gray-500 max-w-xs truncate')

            _upozorneni()

            # ── Filtr ────────────────────────────────────────────────────────
            with ui.row().classes('w-full gap-3 items-end bg-white p-4 rounded-2xl '
                                  'border border-gray-200 flex-wrap'):
                def _nastav(klic, hodnota):
                    filtr[klic] = (list(hodnota or []) if klic == 'pobocka'
                                   else (hodnota or None))
                    if klic == 'divize':
                        dostupne = pobocky_pro_divizi(filtr['divize'])
                        vyber_pob.set_options({p: p for p in dostupne})
                        # ponech jen provozovny, které do zvolené divize patří
                        zbyle = [p for p in filtr['pobocka'] if p in dostupne]
                        if zbyle != filtr['pobocka']:
                            filtr['pobocka'] = zbyle
                            vyber_pob.set_value(zbyle)
                    _rychly_vyber.refresh()
                    _prehled.refresh()

                with ui.column().classes('gap-0'):
                    hledat_in = ui.input(
                        label='Hledat', placeholder='dodavatel, název akce…') \
                        .classes('w-64').props('outlined dense clearable '
                                               'debounce=300')
                    hledat_in.on_value_change(
                        lambda e: _nastav('hledat', (e.value or '').strip()))
                    with hledat_in.add_slot('prepend'):
                        ui.icon('search').classes('text-gray-400')
                    ui.label('hledá se i v proběhlých akcích') \
                        .classes('text-xs text-blue-500 pl-1') \
                        .bind_visibility_from(hledat_in, 'value')

                ui.select({None: 'Všechny divize',
                           'MO': 'Maloobchod', 'SP': 'Svět potravin'},
                          value=None, label='Divize',
                          on_change=lambda e: _nastav('divize', e.value)) \
                    .classes('w-52').props('outlined dense')
                vyber_pob = ui.select(
                    {p: p for p in pobocky_pro_divizi(None)},
                    value=[], label='Provozovny',
                    multiple=True, with_input=True, clearable=True,
                    on_change=lambda e: _nastav('pobocka', e.value)) \
                    .classes('w-72').props('outlined dense use-chips')
                fil_od = _pole_datum('Od data', 'w-36')
                fil_od.on_value_change(lambda e: _nastav('od', _na_datum(e.value)))
                fil_do = _pole_datum('Do data', 'w-36')
                fil_do.on_value_change(lambda e: _nastav('do', _na_datum(e.value)))
                ui.select(RAZENI_VOLBY, value='datum_asc', label='Řazení',
                          on_change=lambda e: _nastav('razeni', e.value)) \
                    .classes('w-56').props('outlined dense')
                ui.switch('Jen aktuální a budoucí', value=True,
                          on_change=lambda e: _nastav('jen_aktualni', e.value)) \
                    .classes('text-sm')
                ui.toggle({'datum': 'Podle data', 'pobocka': 'Podle provozoven'},
                          value='datum',
                          on_change=lambda e: _nastav('pohled', e.value)) \
                    .props('no-caps dense unelevated').classes('ml-auto')

            # ── Rychlý výběr provozovny (bez filtrování) ─────────────────────
            def _klik_pobocka(p):
                """Klik na dlaždici provozovny = přidá/odebere ji z filtru."""
                if p is None:
                    filtr['pobocka'] = []
                elif p in filtr['pobocka']:
                    filtr['pobocka'] = [x for x in filtr['pobocka'] if x != p]
                else:
                    filtr['pobocka'] = filtr['pobocka'] + [p]
                vyber_pob.set_value(list(filtr['pobocka']))
                _rychly_vyber.refresh()
                _prehled.refresh()

            @ui.refreshable
            def _rychly_vyber():
                # akce dle ostatních filtrů (bez filtru pobočky) + počty na provozovnu
                akce_mapa = _nacti_akce(divize=filtr['divize'], od=filtr['od'],
                                        do=filtr['do'],
                                        jen_aktualni=bool(filtr['jen_aktualni']),
                                        razeni=filtr['razeni'],
                                        hledat=filtr['hledat'])
                pocty = {}
                for a in akce_mapa:
                    for p in a['pobocky']:
                        pocty[p] = pocty.get(p, 0) + 1

                with ui.column().classes('w-full gap-2 bg-white p-4 rounded-2xl '
                                         'border border-gray-200'):
                    with ui.row().classes('w-full items-center gap-2'):
                        ui.icon('bolt').classes('text-amber-500')
                        ui.label('Rychlý výběr provozovny') \
                            .classes('text-xs font-bold text-gray-500 '
                                     'uppercase tracking-widest')
                        ui.label('— klikej a přidávej i více provozoven zároveň') \
                            .classes('text-xs text-gray-400')
                        ui.space()
                        if filtr['pobocka']:
                            ui.label(f'vybráno: {len(filtr["pobocka"])}').classes(
                                'text-xs font-bold text-green-700')
                            ui.label('✕ Zrušit výběr').classes(
                                'text-xs text-gray-500 cursor-pointer '
                                'hover:text-red-600 select-none') \
                                .on('click', lambda _: _klik_pobocka(None))

                    for nadpis, seznam in (('Maloobchod', POBOCKY_MO),
                                           ('Svět potravin – C&C', POBOCKY_SP)):
                        if not set(seznam) & set(pobocky_pro_divizi(filtr['divize'])):
                            continue
                        with ui.row().classes('w-full gap-1 items-start flex-wrap'):
                            ui.label(nadpis).classes('text-xs font-bold text-gray-400 '
                                                     'w-28 pt-1 shrink-0')
                            for p in seznam:
                                vybrano = p in filtr['pobocka']
                                pocet = pocty.get(p, 0)
                                # ui.label místo ui.button — q-btn vnucuje bílý text
                                chip = ui.label(
                                    f'{p}  ({pocet})' if pocet else p).classes(
                                    'text-xs rounded-full px-3 py-1 border '
                                    'cursor-pointer select-none whitespace-nowrap '
                                    'transition ' + (
                                        'bg-green-600 text-white border-green-700 '
                                        'font-bold'
                                        if vybrano else
                                        ('bg-white text-gray-700 border-gray-300 '
                                         'hover:bg-green-50 hover:border-green-400'
                                         if pocet else
                                         'bg-gray-50 text-gray-400 border-gray-200 '
                                         'hover:bg-green-50 hover:text-gray-600')))
                                chip.on('click', lambda _, _p=p: _klik_pobocka(_p))

            _rychly_vyber()

            # ── Přehled ──────────────────────────────────────────────────────
            @ui.refreshable
            def _prehled():
                akce = _nacti_akce(divize=filtr['divize'], pobocka=filtr['pobocka'],
                                   od=filtr['od'], do=filtr['do'],
                                   jen_aktualni=bool(filtr['jen_aktualni']),
                                   razeni=filtr['razeni'], hledat=filtr['hledat'])
                if not akce:
                    with ui.column().classes('w-full items-center py-12 gap-2'):
                        ui.icon('event_busy').classes('text-4xl text-gray-300')
                        ui.label('Žádné ochutnávky neodpovídají filtru.') \
                            .classes('text-gray-400')
                    return

                if filtr['pohled'] == 'datum':
                    with ui.column().classes('w-full gap-2'):
                        for a in akce:
                            _radek(a)
                    return

                # Pohled podle provozoven
                pobocky = (list(filtr['pobocka'])
                           or pobocky_pro_divizi(filtr['divize']))
                with ui.column().classes('w-full gap-3'):
                    for p in pobocky:
                        akce_p = [a for a in akce if p in a['pobocky']]
                        if not akce_p:
                            continue
                        with ui.card().classes('w-full p-3 gap-2 border border-gray-200'):
                            ui.label(p).classes('text-sm font-black text-blue-900')
                            for a in akce_p:
                                _radek(a, zobraz_pobocky=False)

            _prehled()

            # ── Chat ─────────────────────────────────────────────────────────
            ui.separator().classes('my-2')
            with ui.column().classes('w-full gap-2'):
                ui.label('Dotazy a poznámky') \
                    .classes('font-bold text-xs text-gray-500 uppercase tracking-widest')

                @ui.refreshable
                def _chat():
                    zpravy = _nacti_chat()
                    if not zpravy:
                        ui.label('Zatím žádné zprávy.') \
                            .classes('text-sm text-gray-400 italic')
                        return
                    with ui.column().classes('w-full gap-2'):
                        for z in zpravy:
                            with ui.row().classes('gap-2 items-start'):
                                ui.label(z['user_jmeno'] or '?') \
                                    .classes('text-xs font-bold text-blue-700 '
                                             'w-40 flex-shrink-0')
                                ui.label(z['text']) \
                                    .classes('text-sm text-gray-700 flex-1 '
                                             'whitespace-pre-wrap')
                                ui.label(_fmt_dt(z['datum'])) \
                                    .classes('text-xs text-gray-400 flex-shrink-0')

                _chat()

                with ui.row().classes('w-full gap-2 items-center mt-1'):
                    chat_input = ui.input(placeholder='Napište dotaz nebo poznámku…') \
                        .classes('flex-1').props('outlined dense')

                    def _odeslat():
                        text = (chat_input.value or '').strip()
                        if not text:
                            return
                        _pridej_chat(user_id, user_name, text)
                        chat_input.set_value('')
                        _notifikuj(
                            f'💬 {user_name} napsal/a v Ochutnávkách MO a CC: '
                            f'{_zkrat(text)}', krome_id=user_id)
                        _chat.refresh()

                    chat_input.on('keydown.enter', lambda: _odeslat())
                    ui.button(icon='send', on_click=_odeslat) \
                        .props('flat round').classes('text-blue-500')


# ═══════════════════════════════════════════════════════════════════════════════
# Self-check: python intranet_ochutnavky.py
# ═══════════════════════════════════════════════════════════════════════════════
if __name__ == '__main__':
    d = date(2026, 8, 4)
    assert stav_akce('2026-08-01', '2026-08-03', d) == 'probehlo'
    assert stav_akce('2026-08-01', '2026-08-04', d) == 'probiha'
    assert stav_akce(d, None, d) == 'probiha'
    assert stav_akce('2026-08-10', '2026-08-10', d) == 'brzy'
    assert stav_akce('2026-08-11', None, d) == 'brzy'      # přesně 7. den
    assert stav_akce('2026-08-12', None, d) == 'budouci'

    assert obdobi_text('2026-08-04', '2026-08-04') == '04.08.2026'
    assert obdobi_text('2026-08-04', None) == '04.08.2026'
    assert obdobi_text('2026-08-04', '2026-08-06') == '04.08.2026 – 06.08.2026'
    assert obdobi_text(None, None) == ''

    # Počty dle seznamu prodejen na intranetu (Prodejny MO / Prodejny C&C)
    assert len(pobocky_pro_divizi('MO')) == 20
    assert len(pobocky_pro_divizi('SP')) == 22
    assert len(pobocky_pro_divizi('OBE')) == 42
    assert pobocky_pro_divizi(None) == pobocky_pro_divizi('OBE')
    # žádné duplicity a žádný překryv mezi divizemi
    vse = pobocky_pro_divizi('OBE')
    assert len(set(vse)) == len(vse)
    assert not (set(POBOCKY_MO) & set(POBOCKY_SP))
    # každý kód právě jednou, MO = 1xx, C&C = 3xx
    kody = [p.split(' - ')[0] for p in vse]
    assert len(set(kody)) == len(kody)
    assert all(p[0] == '1' for p in POBOCKY_MO)
    assert all(p[0] == '3' for p in POBOCKY_SP)
    # migrace musí končit na existující pobočce a nesmí mířit sama na sebe
    for stary, novy in _MIGRACE_POBOCEK:
        assert novy in vse, novy
        assert stary != novy
        assert stary not in vse, stary

    assert _bezpecny_nazev('../../etc/passwd') == 'passwd'
    assert _bezpecny_nazev('foto akce.JPG') == 'foto akce.JPG'
    assert _bezpecny_nazev('') == 'soubor'
    print('OK')
