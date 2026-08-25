"""Modul Sankce — dvě sestavy nad importovanými daty z listu DATA:

  • Zamítnuté dodávky dodavatelem  (pro nákupní oddělení + analytika importuje)
  • Sankce k vystavení             (pro účtárnu + analytika importuje)

Data přibývají po týdnech. Každý import nese „období", podle kterého se pozná,
co je nové. Re-import stejného období dávku nahradí, ale ručně zadané stavy a
poznámky se zachovají (párují se přes row_hash).

Zámek filtru:
  • Zamítnuté dodávky – filtr na OBDOBÍ se zámkem. Analytik po importu nastaví
    „čerstvé" období jako výchozí pro všechny; kdokoliv si může rozkliknout vše.
  • Sankce k vystavení – období je sloupcem v řádku; výchozí (zamčený) filtr je
    na STAV: defaultně Nová data + Nevyfakturovat + Rozpracováno + Nákup + Provoz. Kdokoliv může
    zobrazit vše.

Role (intranet_prava.py):
  • sankce_analytik – import dat do obou sestav + nastavení zámku období.
  • sankce_ucetni   – Sankce k vystavení: mění stav + píše poznámky.
  • sankce_nakup    – Zamítnuté dodávky: píše poznámky.
  • sankce_ctenar   – obě sestavy jen pro čtení (filtry/řazení/export + diskuze).
  • sankce_tiket_XX – nákupčí s kódem XX: řeší tikety svých dodavatelů.
  • sankce_tiket_provoz   – tikety předané na provoz.
  • sankce_tiket_kontrola – druhotná kontrola: schvaluje storno sankce.
  • 'vse'           – vše.
"""

from nicegui import ui, app
import intranet_data
import intranet_logger
import intranet_notifikace
import intranet_emaily
from intranet_ui_utils import refreshable_na_klienta
import datetime
import hashlib
import io
import os
import re
import json
import time
import inspect
import tempfile
import unicodedata
import asyncio
import zipfile
from collections import OrderedDict
from html import escape as _esc

# =========================================================
# KONSTANTY
# =========================================================
# Pozn.: kódy v DB zůstávají historicky stejné ('nevyfakturovano' / 'vyfakturovano'),
# přejmenoval se jen LABEL — z popisu STAVU (Nevyfakturováno / Vyfakturováno) na
# AKCI, kterou má účtárna provést (Nevyfakturovat / Fakturovat). Pořadí klíčů
# odpovídá workflow a zároveň pořadí ve filtru i v editoru buňky.
STAV_LABEL = {
    'nova_data':         'Nová data',
    'nevyfakturovano':   'Nevyfakturovat',
    'rozpracovano':      'Rozpracováno',
    'nakup':             'Nákup',
    'provoz':            'Provoz',
    'vyfakturovano':     'Fakturovat',
    'storno':            'Stornovat',
    'odevzdano_uctarne': 'Odevzdáno účtárně',
}
STAV_LABEL_REV = {v: k for k, v in STAV_LABEL.items()}
# Výchozí (zamčený) filtr stavu — „věci, které čekají na vyřízení".
# Nová data se nově zahrnují, aby čerstvě importované řádky byly hned vidět.
# Nákup a Provoz = předáno oddělení, stále čeká na vyřízení → patří do výchozích.
STAV_DEFAULT = ['nova_data', 'nevyfakturovano', 'rozpracovano', 'nakup', 'provoz']

# Druhý, nezávislý stav „Aktivita 2" (poslední sloupec) — roletka se dvěma hodnotami.
STAV2_LABEL = {
    'v_procesu': 'V procesu',
    'uzavreno':  'Uzavřeno',
}
STAV2_LABEL_REV = {v: k for k, v in STAV2_LABEL.items()}
STAV2_DEFAULT = 'v_procesu'

_NAST_KLIC_ZAMEK = 'sankce_zamitnute_obdobi'         # globálně zamčené období (Zamítnuté)

# ── Zálohy / body obnovení sestavy „Sankce k vystavení" ──
# Celá tabulka se ukládá jako JSON snímek; zálohy NEJDOU stahovat, jen obnovit.
_ZALOHA_TABLE = 'sankce_vystaveni_zalohy'
_ZALOHA_INTERVAL_H = 4                               # jak často se dělá AUTO záloha (hodiny)
_ZALOHA_AUTO_MAX = 10                                # kolik AUTO záloh držet (přebytek se prořezává – padá nejstarší)
_ZALOHA_TYP_LABEL = {'auto': 'Automatická záloha', 'rucni': 'Ruční bod obnovy'}

# Mapování hlaviček listu DATA → DB sloupce (klíč = normalizovaný název hlavičky).
_MAPA_ZAMITNUTE = {
    'ico-dodavatel':      'ico',
    'jmenododavatele':    'jmeno_dodavatele',
    'kodzbozi':           'kod_zbozi',
    'nazevzbozi':         'nazev_zbozi',
    'datumpozadovano':    'datum_pozadovano',
    'typadresy':          'typ_adresy',
    'k2':                 'k2',
    'c.objednavky':       'cislo_objednavky',
    'id.pobocky':         'id_pobocky',
    'dodavatel':          'dodavatel',
    'nakupci(cen.)':      'nakupci',
    'objednanomj':        'objednano_mj',
    'dodanomj':           'dodano_mj',
    'odmitnutomj':        'odmitnuto_mj',
    'obj.-cena':          'obj_cena',
    'typsankce':          'typ_sankce',
    'hodn.sankce':        'hodn_sankce',
    'odmitnutokccelkem':  'odmitnuto_kc_celkem',
}
_MAPA_VYSTAVENI = {
    'ico-dodavatel':   'ico',
    'jmenododavatele': 'jmeno_dodavatele',
    'kodzbozi':        'kod_zbozi',
    'nazevzbozi':      'nazev_zbozi',
    'c.objednavky':    'cislo_objednavky',
    'id.pobocky':      'id_pobocky',
    'objednanomj':     'objednano_mj',
    'dodanomj':        'dodano_mj',
    'dod.pozdemj':     'dod_pozde_mj',
    'obj.-cena':       'obj_cena',
    'hodn.sankce':     'hodn_sankce',
    'nakupci(pob.)':   'nakupci_pob',
}

# Textové vs. číselné sloupce (pro správné uložení a formátování)
_CISLA_ZAMITNUTE = {'objednano_mj', 'dodano_mj', 'odmitnuto_mj', 'obj_cena',
                    'hodn_sankce', 'odmitnuto_kc_celkem'}
_CISLA_VYSTAVENI = {'objednano_mj', 'dodano_mj', 'dod_pozde_mj', 'obj_cena', 'hodn_sankce'}

# Sloupce, které v importovaném souboru chybět SMĚJÍ (import je nepovažuje za chybu).
_VOLITELNE = {'nakupci_pob'}

# Sloupce do připnutého součtového řádku „Celkem" — vše číselné KROMĚ obj_cena.
# (Obj.-cena je cena za 1 MJ; její sečtení napříč řádky nic nevypovídá.)
_SOUCET_ZAMITNUTE = _CISLA_ZAMITNUTE - {'obj_cena'}
_SOUCET_VYSTAVENI = _CISLA_VYSTAVENI - {'obj_cena'}


# =========================================================
# POMOCNÉ FUNKCE
# =========================================================
def _norm(s) -> str:
    """Normalizuje název hlavičky: bez diakritiky, malá písmena, bez mezer."""
    if s is None:
        return ''
    txt = unicodedata.normalize('NFKD', str(s))
    txt = ''.join(c for c in txt if not unicodedata.combining(c))
    return txt.lower().replace(' ', '').strip()


def _f(v):
    """Bezpečný převod na float (None → None)."""
    if v is None or v == '':
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace('\xa0', '').replace(' ', '').replace(',', '.'))
    except (TypeError, ValueError):
        return None


def _s(v) -> str:
    """Bezpečný převod na ořezaný řetězec."""
    if v is None:
        return ''
    return str(v).strip()


def _row_hash(tabulka: str, obdobi: str, casti: list) -> str:
    """Stabilní otisk řádku v rámci období — slouží k párování ručních úprav
    (stav, poznámka, historie) i po opětovném importu téhož období."""
    klic = tabulka + '|' + obdobi + '|' + '|'.join(_s(c) for c in casti)
    return hashlib.md5(klic.encode('utf-8')).hexdigest()


def _parse_obdobi_z_nazvu(nazev: str):
    """Zkusí z názvu souboru vytáhnout období typu „01.05 -15.05.2026".
    Vrací (od_iso, do_iso) nebo (None, None)."""
    if not nazev:
        return None, None
    m = re.search(r'(\d{1,2})\.(\d{1,2})\.?(\d{4})?\s*-\s*(\d{1,2})\.(\d{1,2})\.(\d{4})', nazev)
    if not m:
        return None, None
    d1, m1, y1, d2, m2, y2 = m.groups()
    y1 = y1 or y2
    try:
        od = datetime.date(int(y1), int(m1), int(d1)).isoformat()
        do = datetime.date(int(y2), int(m2), int(d2)).isoformat()
        return od, do
    except ValueError:
        return None, None


def _iso_na_cz(iso: str) -> str:
    try:
        d = datetime.date.fromisoformat(iso)
        return d.strftime('%d.%m.%Y')
    except (ValueError, TypeError):
        return iso or ''


def _obdobi_label(od_iso: str, do_iso: str) -> str:
    return f'{_iso_na_cz(od_iso)} – {_iso_na_cz(do_iso)}'


# ─── Tisk oznámení a podkladu do PDF (server-side: Playwright/Chromium) ────
def _cz_money(n) -> str:
    try:
        return f'{float(n):,.2f}'.replace(',', ' ').replace('.', ',')
    except (TypeError, ValueError):
        return ''


def _cz_num(n) -> str:
    try:
        f = float(n)
    except (TypeError, ValueError):
        return ''
    if f == int(f):
        return f'{int(f):,}'.replace(',', ' ')
    return f'{f:,.2f}'.replace(',', ' ').replace('.', ',')


# Číselník VO poboček (kód střediska → název) pro popisek v podkladu.
# Zdroj: číselník středisek v intranet_vysledky.py (POBOCKY_EXCEL_MAPPING).
# Držíme lokální kopii záměrně — intranet_vysledky je stránkový modul (@ui.page)
# a jeho import sem by tahal vedlejší efekty.
_POBOCKY_NAZVY = {
    '010': 'Pardubice',
    '011': 'Praha',
    '012': 'Jilemnice',
    '013': 'Most',
    '014': 'Liberec',
    '017': 'Horšovský Týn',
    '019': 'Hodonín',
    '020': 'Zlín',
    '026': 'Ostrava',
    '028': 'Brno – Wine Life',
    '032': 'Olomouc',
    '033': 'České Budějovice',
    '034': 'Plzeň',
    '037': 'Nová Role',
}


def _pobocka_nazev(v) -> str:
    """Kód střediska → název VO pobočky pro tisk.

    Snese '10', '010', '010 - Pardubice' i '10.0' (číslo z Excelu). Neznámý kód
    vrací '' — v podkladu pak zůstane jen holá hodnota z importu.
    POZOR: uložené `id_pobocky` se tímhle NIKDY nepřepisuje, vstupuje do row_hash.
    """
    m = re.search(r'\d+', _s(v))
    if not m:
        return ''
    return _POBOCKY_NAZVY.get(m.group(0).zfill(3), '')


def _safe_filename(s: str) -> str:
    s = unicodedata.normalize('NFKD', str(s or ''))
    s = ''.join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r'[^A-Za-z0-9_\- ]', '', s).strip().replace(' ', '_')
    return s or 'podklad'


_PDF_CSS = """
<style>
  #sankce-print-area * { box-sizing: border-box; }
  .podklad { font-family: Arial, sans-serif; color:#111; width: 190mm; }
  .podklad h1 { font-size: 18px; margin:0 0 2px; }
  .hdr { display:flex; justify-content:space-between; align-items:flex-start;
         border-bottom:2px solid #b91c1c; padding-bottom:8px; margin-bottom:10px; }
  .hdr .meta { font-size:12px; color:#374151; line-height:1.55; text-align:right; }
  .hdr .meta b { color:#111; }
  table.t { width:100%; border-collapse:collapse; font-size:11px; page-break-inside:auto; }
  table.t thead { display:table-header-group; }
  table.t tr { page-break-inside:avoid; break-inside:avoid; }
  table.t th { background:#fee2e2; color:#7f1d1d; border:1px solid #d1d5db; padding:5px 6px; text-align:left; vertical-align:top; }
  table.t td { border:1px solid #e5e7eb; padding:4px 6px; vertical-align:top; }
  table.t td.c { text-align:center; }
  table.t td.r { text-align:right; white-space:nowrap; }
  table.t td.mono { font-family:'Courier New',monospace; }
  table.t td.b { font-weight:bold; }
  table.t td.pob { text-align:center; line-height:1.25; }
  table.t td.pob .pob-kod { font-weight:bold; }
  table.t td.pob .pob-nazev { display:block; font-size:9px; color:#6b7280; }
  table.t tr:nth-child(even) td { background:#fafafa; }
  .total { margin-top:10px; text-align:right; font-size:13px; }
  .total b { font-size:16px; color:#7f1d1d; }
  .pozn { margin-top:18px; font-size:10px; color:#6b7280; }
</style>
"""


_OZNAMENI_CSS = """
<style>
  .oznameni { font-family: Arial, sans-serif; color:#111; width:190mm; font-size:12px; line-height:1.5; }
  .oznameni .ozn-hdr { text-align:center; border-bottom:2px solid #b91c1c; padding-bottom:8px; margin-bottom:16px; }
  .oznameni .ozn-firma { font-size:18px; font-weight:bold; color:#7f1d1d; }
  .oznameni .ozn-sub { font-size:10px; color:#374151; }
  .oznameni .ozn-meta { display:flex; justify-content:space-between; align-items:flex-start; margin-bottom:16px; }
  .oznameni .ozn-lbl { font-size:11px; color:#6b7280; }
  .oznameni .ozn-dod { font-weight:bold; font-size:14px; }
  .oznameni .ozn-ref { text-align:right; line-height:1.7; }
  .oznameni .ozn-vec { font-weight:bold; margin:6px 0 12px; }
  .oznameni p { margin:0 0 9px; text-align:justify; }
  .oznameni .ozn-podpis { margin-top:22px; line-height:1.5; }
  .oznameni .ozn-jmeno { font-weight:bold; }
  .oznameni .ozn-priloha { margin-top:16px; font-size:11px; color:#374151; }
</style>
"""


def _sankce_souhrn(rows: list):
    """Z řádků spočítá (hrubá_sankce, sleva_částka, po_slevě, jednotný_podíl|None).
    `jednotný_podíl` = společný podíl slevy, mají-li ho všechny řádky stejný (jinak None).
    Sleva je podíl (0,05 = 5 %); řádkové částky se nemění, koriguje se jen součet."""
    raw = sum(_f(r.get('hodn_sankce')) or 0 for r in rows)
    sleva_castka = sum((_f(r.get('hodn_sankce')) or 0) * (_f(r.get('sleva')) or 0) for r in rows)
    po_sleve = raw - sleva_castka
    podily = {round(_f(r.get('sleva')) or 0.0, 6) for r in rows}
    jednotny = podily.pop() if len(podily) == 1 else None
    return raw, sleva_castka, po_sleve, jednotny


def _oznameni_pdf_html(rows: list) -> str:
    """První list tisku — „Předběžné oznámení o uplatnění smluvní pokuty" dle šablony.
    Žlutá pole se vyplní z označených řádků: dodavatel, IČO, variabilní symbol
    (rozsah našich čísel), datum (dnes), období dat a celková výše pokuty (po slevě)."""
    dod = (rows[0].get('jmeno_dodavatele') or '').strip()
    ico = (rows[0].get('ico') or '').strip()
    obd = sorted({(r.get('obdobi') or '').strip() for r in rows if r.get('obdobi')})
    obd_txt = ', '.join(obd)
    dnes = datetime.date.today().strftime('%d.%m.%Y')
    raw, sleva_castka, po_sleve, jednotny = _sankce_souhrn(rows)
    # věta o výši pokuty — se slevou ukáže rozpad (smluvní pokuta − sleva = po slevě)
    _uvod = ('Za uvedené prodlení, resp. Nesplnění povinnosti s dodáním zboží, Vám tímto v souladu '
             'se smluvními podmínkami upravujícími naši spolupráci oznamujeme uplatnění smluvní pokuty, ')
    if sleva_castka > 0:
        _sl = f' (sleva {_cz_num(jednotny * 100)} %)' if jednotny else ''
        _vyse = (f'jejíž celková výše za uvedené období po zohlednění slevy{_sl} činí '
                 f'<b>{_cz_money(po_sleve)}</b> Kč (smluvní pokuta {_cz_money(raw)} Kč − sleva '
                 f'{_cz_money(sleva_castka)} Kč). ')
    else:
        _vyse = f'jejíž celková výše za uvedené období činí <b>{_cz_money(raw)}</b> Kč. '
    veta_sankce = _uvod + _vyse + 'Vyčíslení smluvní pokuty u jednotlivých objednávek je uvedeno v přiloženém podkladu.'
    # variabilní symbol = rozsah „našich čísel" označených řádků (min–max)
    cisla = sorted(_s(r.get('nase_cislo')) for r in rows if _s(r.get('nase_cislo')))
    vs = '' if not cisla else (cisla[0] if cisla[0] == cisla[-1] else f'{cisla[0]}–{cisla[-1]}')
    return (
        '<div class="oznameni">'
        '<div class="ozn-hdr">'
        '<div class="ozn-firma">JIP východočeská, a.s.</div>'
        '<div class="ozn-sub">Hradišťská 407, Polabiny, 533 52 Pardubice  ·  IČO: 27464822  ·  DIČ: CZ27464822</div>'
        '<div class="ozn-sub">zapsaná v obchodním rejstříku vedeném Krajským soudem v Hradci Králové, oddíl B, vložka 2413</div>'
        '</div>'
        '<div class="ozn-meta">'
        '<div class="ozn-adr">'
        '<div class="ozn-lbl">Adresát (dodavatel):</div>'
        f'<div class="ozn-dod">{_esc(dod)}</div>'
        f'<div>IČO: <b>{_esc(ico)}</b></div>'
        '</div>'
        '<div class="ozn-ref">'
        f'Naše značka: <b>{_esc(vs)}</b><br>'
        f'V Pardubicích dne {dnes}'
        '</div>'
        '</div>'
        '<div class="ozn-vec">Věc: Předběžné oznámení o uplatnění smluvní pokuty '
        'vyplývající z nedodávek zboží a výzva k vyjádření</div>'
        '<p>Vážení obchodní partneři,</p>'
        f'<p>v rámci kontroly plnění dodavatelských povinností za období <b>{_esc(obd_txt)}</b> '
        'jsme zjistili, že u níže specifikovaných objednávek nebyly Vaše dodavatelské povinnosti '
        'splněny řádně a včas, a to konkrétně došlo k [nedodání objednaného zboží / dodání zboží '
        'po sjednaném termínu dodání / dodání zboží v nižším než objednaném množství ….].</p>'
        '<p>Podrobný přehled dotčených objednávek a vyčíslení smluvní pokuty je uveden v příloze tohoto '
        'oznámení („Podklad ke smluvní pokutě"), který je jeho nedílnou součástí.</p>'
        f'<p>{veta_sankce}</p>'
        '<p>Toto oznámení Vám zasíláme s předstihem, než přistoupíme k fakturaci smluvní pokuty. Máte možnost '
        'se k uvedeným skutečnostem vyjádřit a doložit případné rozhodné okolnosti (zejména že zboží '
        'bylo dodáno v požadovaném termínu a množství, případně z důvodu vyšší moci apod.) ve lhůtě '
        '10 dnů ode dne doručení předběžného oznámení na e-mailovou adresu lenka.novakova@jip-napoje.cz.</p>'
        '<p>Věříme, že se jedná o ojedinělé pochybení a že se nám podaří situaci vyřešit ke '
        'spokojenosti obou stran.</p>'
        '<div class="ozn-podpis">S pozdravem<br>'
        '<span class="ozn-jmeno">Lenka Nováková</span><br>'
        'Vedoucí – správa produktových karet<br>'
        'JIP východočeská, a.s.<br>'
        'e-mail: lenka.novakova@jip-napoje.cz</div>'
        '<div class="ozn-priloha">Příloha: Podklad ke smluvní pokutě – přehled dotčených objednávek</div>'
        '</div>'
    )


def _vystaveni_pdf_html(rows: list) -> str:
    """Sestaví HTML podkladu ke smluvní pokutě za JEDNOHO dodavatele (z označených řádků)."""
    dod = (rows[0].get('jmeno_dodavatele') or '').strip()
    ico = (rows[0].get('ico') or '').strip()
    obd = sorted({(r.get('obdobi') or '').strip() for r in rows if r.get('obdobi')})
    obd_txt = ', '.join(obd)
    dnes = datetime.date.today().strftime('%d.%m.%Y')
    raw, sleva_castka, po_sleve, jednotny = _sankce_souhrn(rows)

    radky = []
    for i, r in enumerate(rows, 1):
        _pob_kod = _s(r.get('id_pobocky'))
        _pob_nazev = _pobocka_nazev(_pob_kod)
        _pob_html = f'<span class="pob-kod">{_esc(_pob_kod)}</span>'
        if _pob_nazev:
            _pob_html += f'<span class="pob-nazev">{_esc(_pob_nazev)}</span>'
        radky.append(
            '<tr>'
            f'<td class="c">{i}</td>'
            f'<td class="mono">{_esc(_s(r.get("kod_zbozi")))}</td>'
            f'<td>{_esc(_s(r.get("nazev_zbozi")))}</td>'
            f'<td class="c">{_esc(_s(r.get("cislo_objednavky")))}</td>'
            f'<td class="pob">{_pob_html}</td>'
            f'<td class="r">{_cz_num(r.get("objednano_mj"))}</td>'
            f'<td class="r">{_cz_num(r.get("dodano_mj"))}</td>'
            f'<td class="r">{_cz_num(r.get("dod_pozde_mj"))}</td>'
            f'<td class="r">{_cz_money(r.get("obj_cena"))}</td>'
            f'<td class="r b">{_cz_money(r.get("hodn_sankce"))}</td>'
            '</tr>'
        )
    return (
        '<div class="podklad">'
        '<div class="hdr">'
        '<div><h1>Podklad ke smluvní pokutě</h1>'
        '<div style="font-size:12px;color:#6b7280">Smluvní pokuta k vystavení dodavateli</div></div>'
        f'<div class="meta">Dodavatel: <b>{_esc(dod)}</b><br>'
        f'IČO: <b>{_esc(ico)}</b><br>'
        f'Období: <b>{_esc(obd_txt)}</b><br>'
        f'Vystaveno: <b>{dnes}</b></div>'
        '</div>'
        '<table class="t"><thead><tr>'
        '<th>#</th><th>Kód zboží</th><th>Název zboží</th><th>Č.obj.</th><th>Pobočka</th>'
        '<th>Obj. MJ</th><th>Dod. MJ</th><th>Dod. pozdě MJ</th><th>Obj.-cena</th><th>Hodn. sml. pokuty</th>'
        '</tr></thead><tbody>' + ''.join(radky) + '</tbody></table>'
        + _vystaveni_total_html(raw, sleva_castka, po_sleve, jednotny)
        + f'<div class="pozn">Počet položek: {len(rows)}. Vygenerováno z firemního portálu JIP.</div>'
        '</div>'
    )


def _vystaveni_total_html(raw, sleva_castka, po_sleve, jednotny) -> str:
    """Součtová část podkladu: bez slevy jen „Celková výše smluvní pokuty", se slevou rozpad
    Smluvní pokuta celkem − Sleva = Celkem po slevě."""
    if sleva_castka and sleva_castka > 0:
        sl_lbl = f'Sleva ({_cz_num(jednotny * 100)} %)' if jednotny else 'Sleva'
        return (
            '<div class="total">'
            f'Smluvní pokuta celkem: {_cz_money(raw)} Kč<br>'
            f'{sl_lbl}: − {_cz_money(sleva_castka)} Kč<br>'
            f'Celkem po slevě: <b>{_cz_money(po_sleve)} Kč</b>'
            '</div>'
        )
    return f'<div class="total">Celková výše smluvní pokuty: <b>{_cz_money(raw)} Kč</b></div>'


_render_lock = None  # globální semafor — serializuje server-side PDF render


def _get_render_lock() -> asyncio.Semaphore:
    """Vrátí globální semafor (max 1 souběžný render). Lazy init, aby modul šel
    importovat i před tím, než je živý event loop."""
    global _render_lock
    if _render_lock is None:
        _render_lock = asyncio.Semaphore(1)
    return _render_lock


_PDF_TMP_DIR = os.path.join(tempfile.gettempdir(), 'jip_sankce_pdf')


def _uklid_stare_pdf(max_age_s: int = 1800):
    """Smaže dočasné PDF/ZIP starší než ~30 min (mohou obsahovat citlivá data)."""
    try:
        now = time.time()
        for f in os.listdir(_PDF_TMP_DIR):
            fp = os.path.join(_PDF_TMP_DIR, f)
            try:
                if os.path.isfile(fp) and now - os.path.getmtime(fp) > max_age_s:
                    os.remove(fp)
            except OSError:
                pass
    except FileNotFoundError:
        pass


async def _smaz_soubor_pozdeji(path: str, delay: int = 300):
    """Po stažení (s rezervou) smaže konkrétní dočasný soubor."""
    try:
        await asyncio.sleep(delay)
        os.remove(path)
    except Exception:
        pass


def _stahni_pres_http(data: bytes, filename: str):
    """Stáhne data do prohlížeče přes HTTP, ne přes WebSocket — obejde WS payload
    limit (engine.io ~1 MB), na který padá ui.download.content u velkých souborů
    („message too large for WebSocket transmission"). Zapíše bajty do dočasného
    souboru, klientovi pošle přes WS jen URL, a soubor po chvíli uklidí."""
    os.makedirs(_PDF_TMP_DIR, exist_ok=True)
    _uklid_stare_pdf()
    tmp_path = os.path.join(_PDF_TMP_DIR, f'{int(time.time() * 1000)}_{filename}')
    with open(tmp_path, 'wb') as fh:
        fh.write(data)
    ui.download.file(tmp_path, filename)   # přes WS jde jen URL, samotná data po HTTP
    try:
        asyncio.create_task(_smaz_soubor_pozdeji(tmp_path))
    except RuntimeError:
        pass  # mimo běžící event loop – uklidí to _uklid_stare_pdf při příštím tisku


# =========================================================
# EXPORT DO XLSX (server-side, openpyxl)
# =========================================================
# Proč ne CSV z AG Gridu: `exportDataAsCsv` prohnal hodnoty valueFormatterem,
# takže do souboru šlo "1 234 567,50" — s NBSP (U+00A0) jako oddělovačem tisíců.
# Excel NBSP nerozpozná → buňka je TEXT, nejde sčítat ani filtrovat jako číslo
# (a projevilo se to zákeřně: čísla pod 1000 prošla, od 1000 výš se rozbila).
# Tady stavíme skutečný .xlsx: čísla zůstávají čísly (zobrazení řeší
# number_format podle locale uživatele), IČO / kódy / čísla objednávek jsou
# explicitně text (nepřijdou o vedoucí nuly), datum je datum. Součtový řádek
# je až pod daty a mimo rozsah autofiltru, takže se neplete mezi data.

_XLSX_FMT = {
    'money': '#,##0.00',
    'num':   '#,##0.###',
    'int':   '#,##0',
    'pct':   '0.00%',
    'text':  '@',
    'date':  'dd.mm.yyyy',
}
_XLSX_CISLA = ('money', 'num', 'int', 'pct')


def _xlsx_format(typ, val) -> str:
    """Formát buňky. U „num" (#,##0.###) Excel vykreslí desetinný oddělovač i u
    celého čísla („2,"), proto celá čísla dostanou formát bez desetin.
    """
    if typ == 'num' and isinstance(val, (int, float)) and float(val).is_integer():
        return _XLSX_FMT['int']
    return _XLSX_FMT.get(typ, 'General')
_DATUM_VZORY = ('%d.%m.%Y', '%Y-%m-%d', '%d.%m.%y', '%d/%m/%Y',
                '%d.%m.%Y %H:%M', '%Y-%m-%d %H:%M:%S')


def _xlsx_hodnota(v, typ):
    """Hodnota z DB → to, co má reálně být v buňce (podle typu sloupce)."""
    if v is None or v == '':
        return None
    if typ in _XLSX_CISLA:
        return _f(v)
    if typ == 'date':
        if isinstance(v, (datetime.datetime, datetime.date)):
            return v
        s = str(v).strip()
        for vzor in _DATUM_VZORY:
            try:
                return datetime.datetime.strptime(s, vzor).date()
            except ValueError:
                continue
        return s      # nerozpoznaný tvar → radši text než rozsypané datum
    return _s(v)


def _xlsx_bytes(cols: list, rows: list, total: dict = None, sheet: str = 'Data') -> bytes:
    """Postaví sešit: zmrazená hlavička s autofiltrem, data, součtový řádek.
    `cols` = [(nadpis, field, typ, šířka)], `total` = dict se součty (nebo None)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = (sheet or 'Data')[:31]

    hlav_font = Font(bold=True, color='FFFFFF', size=11)
    hlav_fill = PatternFill('solid', fgColor='334155')
    hlav_zar = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for ci, (nadpis, _fld, _typ, sirka) in enumerate(cols, start=1):
        c = ws.cell(row=1, column=ci, value=nadpis)
        c.font, c.fill, c.alignment = hlav_font, hlav_fill, hlav_zar
        ws.column_dimensions[get_column_letter(ci)].width = sirka
    ws.row_dimensions[1].height = 28

    for ri, radek in enumerate(rows, start=2):
        for ci, (_nad, field, typ, _w) in enumerate(cols, start=1):
            hod = _xlsx_hodnota(radek.get(field), typ)
            c = ws.cell(row=ri, column=ci, value=hod)
            c.number_format = _xlsx_format(typ, hod)

    posledni = 1 + len(rows)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(cols))}{posledni}'

    if total:
        # O řádek níž, tedy ZA rozsahem autofiltru — filtrování v Excelu
        # součet neschová a nepočítá ho mezi data.
        tr = posledni + 2
        tucne = Font(bold=True)
        linka = Border(top=Side(style='thin', color='334155'))
        for ci, (_nad, field, typ, _w) in enumerate(cols, start=1):
            hod = _xlsx_hodnota(total.get(field), typ)
            c = ws.cell(row=tr, column=ci, value=hod)
            c.number_format = _xlsx_format(typ, hod)
            c.font, c.border = tucne, linka

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _export_xlsx(cols: list, rows: list, total, zaklad: str, sheet: str = 'Data'):
    """Sestaví XLSX mimo event loop a pošle ho do prohlížeče přes HTTP."""
    data = await asyncio.to_thread(_xlsx_bytes, cols, rows, total, sheet)
    jmeno = f'{zaklad}_{datetime.datetime.now():%Y-%m-%d_%H%M}.xlsx'
    _stahni_pres_http(data, jmeno)


# ID řádků tak, jak je uživatel právě vidí — po filtrech v hlavičkách sloupců
# a v pořadí podle řazení. Z prohlížeče tahám JEN ID (ne celá data), ať
# nenarazím na limit WS payloadu; kompletní řádky si pak najdu v paměti.
# Pozn.: skládám konkatenací, ne `.format()`/f-stringem — JS tělo obsahuje
# složené závorky a formátovač by na nich spadl.
def _viditelne_ids_js(gid: int) -> str:
    return (
        "const c=getElement(" + str(gid) + ");"
        "const api=c&&(c.api||c.gridApi||(c.grid&&c.grid.api));"
        "if(!api||!api.forEachNodeAfterFilterAndSort)return null;"
        "const out=[];"
        "api.forEachNodeAfterFilterAndSort(function(n){"
        "if(n&&!n.group&&n.data&&n.data.id!=null)out.push(String(n.data.id));});"
        "return out;"
    )


async def _viditelne_ids(g):
    """Vrátí seznam ID v zobrazeném pořadí, nebo None když se to nepovedlo
    (jiná verze AG Gridu, timeout) — volající pak exportuje bez zohlednění
    filtrů v hlavičkách."""
    try:
        ids = await ui.run_javascript(_viditelne_ids_js(g.id), timeout=30)
    except Exception as e:
        print(f'[sankce] Export: nepodařilo se přečíst viditelné řádky: {e!r}')
        return None
    if ids is None:
        print('[sankce] Export: grid API nevrátilo seznam viditelných řádků.')
        return None
    return [str(x) for x in ids if x is not None]


def _serad_dle_ids(rows: list, ids):
    """Seřadí/ořízne řádky podle ID z prohlížeče. `ids=None` → beze změny."""
    if ids is None:
        return rows, False
    podle_id = {str(r.get('id')): r for r in rows}
    return [podle_id[i] for i in ids if i in podle_id], True


# ── Sloupce exportu ──────────────────────────────────────────────────────
# (nadpis, field, typ, šířka) — pořadí i názvy kopírují sestavu na obrazovce;
# typ určuje, jak se hodnota do buňky zapíše a jak se zobrazí.
_EXP_ZAMITNUTE = [
    ('Období',            'obdobi',              'text',  16),
    ('IČO',               'ico',                 'text',  12),
    ('Dodavatel',         'jmeno_dodavatele',    'text',  32),
    ('Kód zboží',         'kod_zbozi',           'text',  14),
    ('Název zboží',       'nazev_zbozi',         'text',  38),
    ('Datum pož.',        'datum_pozadovano',    'date',  13),
    ('Typ adresy',        'typ_adresy',          'text',  13),
    ('K2',                'k2',                  'text',  10),
    ('Č.obj.',            'cislo_objednavky',    'text',  14),
    ('Pobočka',           'id_pobocky',          'text',  11),
    ('Dodavatel (kód)',   'dodavatel',           'text',  14),
    ('Nákupčí',           'nakupci',             'text',  14),
    ('Objedn. MJ',        'objednano_mj',        'num',   12),
    ('Dodáno MJ',         'dodano_mj',           'num',   12),
    ('Odmít. MJ',         'odmitnuto_mj',        'num',   12),
    ('Obj.-cena',         'obj_cena',            'money', 13),
    ('Typ sankce',        'typ_sankce',          'text',  18),
    ('Hodn. sankce',      'hodn_sankce',         'money', 14),
    ('Odmít. Kč celkem',  'odmitnuto_kc_celkem', 'money', 16),
    ('Poznámka',          'poznamka',            'text',  36),
]

_EXP_VYSTAVENI = [
    ('Stav',           'stav_label',       'text',  20),
    ('Aktivita 2',     'stav2_label',      'text',  15),
    ('Období',         'obdobi',           'text',  16),
    ('IČO',            'ico',              'text',  12),
    ('Dodavatel',      'jmeno_dodavatele', 'text',  34),
    ('Kód zboží',      'kod_zbozi',        'text',  14),
    ('Název zboží',    'nazev_zbozi',      'text',  40),
    ('Č.obj.',         'cislo_objednavky', 'text',  14),
    ('Pobočka',        'id_pobocky',       'text',  11),
    ('Objedn. MJ',     'objednano_mj',     'num',   12),
    ('Dodáno MJ',      'dodano_mj',        'num',   12),
    ('Dod. pozdě MJ',  'dod_pozde_mj',     'num',   14),
    ('Obj.-cena',      'obj_cena',         'money', 13),
    ('Hodn. sankce',   'hodn_sankce',      'money', 14),
    ('Nákupčí (pob.)', 'nakupci_pob',      'text',  14),
    ('Sleva na sankci', 'sleva',           'pct',   14),
    ('Poznámka',       'poznamka',         'text',  36),
]

_EXP_SOUHRN = [
    ('IČ',             'ico',              'text',  13),
    ('Dodavatel',      'jmeno_dodavatele', 'text',  40),
    ('Období',         'obdobi',           'text',  16),
    ('Nákupčí (pob.)', 'nakupci_pob',      'text',  16),
    ('Položek',        'pocet',            'int',   10),
    ('Sankce celkem',  'sankce_celkem',    'money', 15),
    ('Sleva',          'sleva_castka',     'money', 14),
    ('Sankce uznaná',  'sankce_uznana',    'money', 15),
    ('Stav',           'stav_label',       'text',  20),
    ('Aktivita 2',     'stav2_label',      'text',  15),
]


def _slep_dokument(rows: list, jen_podklad: bool = False) -> str:
    """Sestaví kompletní HTML jednoho dokumentu (oznámení + podklad) k renderu.
    `jen_podklad=True` vynechá 1. list (oznámení).
    Vrací string připravený pro Playwright (DOCTYPE, head, body)."""
    oznameni = '' if jen_podklad else (
        _oznameni_pdf_html(rows) + '<div style="page-break-before:always"></div>')
    return (
        '<!DOCTYPE html><html lang="cs"><head><meta charset="utf-8">'
        '<title>Oznámení o smluvní pokutě</title>'
        + _PDF_CSS + _OZNAMENI_CSS
        + '</head><body><div class="sankce-doc">'
        + oznameni
        + _vystaveni_pdf_html(rows)
        + '</div></body></html>'
    )


async def _render_pdf_bytes_batch(html_list: list, status_cb=None,
                                  on_wait=None, on_start=None) -> list:
    """Vyrenderuje seznam (filename, html) přes Playwright/Chromium do PDF.
    V jeden okamžik renderuje JEN JEDEN tisk (modulový semafor) — jinak by tři
    souběžní uživatelé spustili tři Chromia naráz a sežrali serveru paměť.
    Další volání čekají ve frontě (event loop běží dál, portál nezamrzne).

    Callbacky pro UX:
      * on_wait()              — voláno, když se musí čekat ve frontě
      * on_start()             — voláno hned po získání semaforu (start renderu)
      * status_cb(i,total,fn)  — voláno před každým dokumentem v dávce

    Vrací list (filename, pdf_bytes). Vyhazuje RuntimeError s instrukcí, když
    Playwright/Chromium nejsou nainstalované."""
    try:
        from playwright.async_api import async_playwright  # noqa
    except ImportError as e:
        raise RuntimeError(
            'Knihovna Playwright není nainstalovaná na serveru.\n'
            'Doinstalujte:\n'
            '  pip install playwright\n'
            '  playwright install chromium\n'
            '  (Linux: sudo playwright install-deps chromium)\n'
            f'Detail: {e}'
        )

    sem = _get_render_lock()
    # Pokud je semafor obsazený, dej UI vědět, že čekáme ve frontě
    if sem.locked() and on_wait:
        try:
            on_wait()
        except Exception:
            pass

    async with sem:
        if on_start:
            try:
                on_start()
            except Exception:
                pass
        out = []
        async with async_playwright() as p:
            try:
                browser = await p.chromium.launch()
            except Exception as e:
                raise RuntimeError(
                    f'Nepodařilo se spustit Chromium: {e}\n'
                    'Patrně chybí stažený prohlížeč; spusťte na serveru:\n'
                    '  playwright install chromium\n'
                    '  (Linux: sudo playwright install-deps chromium)'
                )
            try:
                for i, (fname, html) in enumerate(html_list, 1):
                    if status_cb:
                        try:
                            status_cb(i, len(html_list), fname)
                        except Exception:
                            pass
                    page = await browser.new_page()
                    try:
                        await page.set_content(html, wait_until='load')
                        pdf = await page.pdf(
                            format='A4',
                            margin={'top': '10mm', 'right': '10mm',
                                    'bottom': '10mm', 'left': '10mm'},
                            print_background=True,
                        )
                        out.append((fname, pdf))
                    finally:
                        await page.close()
            finally:
                await browser.close()
        return out


# JS formátovače / renderery (drží se zvyklostí modulu Výsledky)
_MONEY_FMT = (
    "function(p){"
    "if(p.value===null||p.value===undefined||p.value==='')return '';"
    "var n=parseFloat(p.value);if(isNaN(n))return '';"
    "return new Intl.NumberFormat('cs-CZ',{minimumFractionDigits:2,maximumFractionDigits:2}).format(n);"
    "}"
)
_NUM_FMT = (
    "function(p){"
    "if(p.value===null||p.value===undefined||p.value==='')return '';"
    "var n=parseFloat(p.value);if(isNaN(n))return p.value;"
    "return new Intl.NumberFormat('cs-CZ',{minimumFractionDigits:Number.isInteger(n)?0:2,maximumFractionDigits:2}).format(n);"
    "}"
)
_EYE_RENDERER = (
    "function(p){"
    "if(p.node&&p.node.rowPinned)return '';"
    "return '<span title=\"Historie změn řádku\" style=\"cursor:pointer;font-size:15px;opacity:.65\">👁️</span>';"
    "}"
)
_STAV_STYLE = (
    "function(p){"
    "var v=p.value;"
    "if(v==='Nová data')return{backgroundColor:'#fce7f3',color:'#9d174d',fontWeight:'600'};"
    "if(v==='Nevyfakturovat')return{backgroundColor:'#fee2e2',color:'#991b1b',fontWeight:'600'};"
    "if(v==='Rozpracováno')return{backgroundColor:'#fef9c3',color:'#854d0e',fontWeight:'600'};"
    "if(v==='Nákup')return{backgroundColor:'#ffedd5',color:'#9a3412',fontWeight:'600'};"
    "if(v==='Provoz')return{backgroundColor:'#ede9fe',color:'#5b21b6',fontWeight:'600'};"
    "if(v==='Fakturovat')return{backgroundColor:'#dbeafe',color:'#1e40af',fontWeight:'600'};"
    "if(v==='Stornovat')return{backgroundColor:'#fecaca',color:'#991b1b',fontWeight:'600'};"
    "if(v==='Odevzdáno účtárně')return{backgroundColor:'#dcfce7',color:'#166534',fontWeight:'600'};"
    "return null;}"
)
_STAV2_STYLE = (
    "function(p){"
    "var v=p.value;"
    "if(v==='V procesu')return{backgroundColor:'#fef9c3',color:'#854d0e',fontWeight:'600'};"
    "if(v==='Uzavřeno')return{backgroundColor:'#dcfce7',color:'#166534',fontWeight:'600'};"
    "return null;}"
)
_GRID_STYLE = 'height: calc(100vh - 360px); min-height: 420px'
_AUTOSIZE_FIT = (
    "function(p){var api=p.api;if(!api||!api.autoSizeAllColumns||!api.sizeColumnsToFit)return;"
    "requestAnimationFrame(function(){api.autoSizeAllColumns();"
    "var cols=api.getColumns?api.getColumns():null;if(!cols)return;"
    "var lim=cols.filter(function(c){return c.isVisible();})"
    ".map(function(c){return{key:c.getColId(),minWidth:Math.ceil(c.getActualWidth())};});"
    "api.sizeColumnsToFit({columnLimits:lim});});}"
)
_PINNED_TOTAL_STYLE = (
    "function(p){"
    "if(p.node&&p.node.rowPinned==='bottom')"
    "return{fontWeight:'700',backgroundColor:'#f1f5f9',borderTop:'2px solid #cbd5e1'};"
    "return null;}"
)

# Indikátor diskuze (chat) k řádku: ikona 💬 + odznáček s počtem zpráv.
# Nepřečtené „svítí" červeně, přečtené jsou šedé, prázdné vlákno je decentní.
_CHAT_RENDERER = (
    "function(p){"
    "if(p.node&&p.node.rowPinned)return '';"
    "var d=p.data||{};var n=d._chat_pocet||0;var u=!!d._chat_unread;"
    "var wrap='position:relative;display:inline-block;cursor:pointer;font-size:15px;'+(u?'':'opacity:.55;');"
    "var s='<span title=\"Diskuze k případu\" style=\"'+wrap+'\">💬';"
    "if(n>0){var bg=u?'#dc2626':'#94a3b8';"
    "s+='<span style=\"position:absolute;top:-7px;right:-10px;background:'+bg+';color:#fff;"
    "font-size:9px;font-weight:700;line-height:14px;min-width:15px;height:15px;padding:0 3px;"
    "border-radius:8px;text-align:center;box-shadow:0 0 0 2px #fff;\">'+n+'</span>';}"
    "s+='</span>';return s;}"
)
# Stabilní id řádku (DB id) — nutné pro cílené applyTransaction při živém pollingu chatu.
_GET_ROW_ID = "function(p){var d=p.data||{};return ''+(d.id!=null?d.id:(d.row_hash||''));}"


def _make_filter_recalc_js(soucet_cols, sleva_col: str = None) -> str:
    """JS callback pro `:onFilterChanged` v AG Gridu — přepočte připnutý
    součtový řádek („CELKEM") dole podle aktuálně viditelných (filtrovaných)
    řádků. Pure-client, žádný round-trip do Pythonu — reaguje i na floating
    filtry v hlavičkách sloupců (lupa), které Python nevidí. Volá se i ručně
    přes `api.onFilterChanged()` po změně dat / stavu řádku.

    Když je zadán `sleva_col`, „Hodn. sankce" v součtu se sníží o slevu daného
    řádku: Σ hodn_sankce*(1 − sleva). Ceny v jednotlivých řádcích zůstávají beze
    změny — korekce je jen v součtovém řádku."""
    cols_json = json.dumps(sorted(soucet_cols))
    sleva_json = json.dumps(sleva_col) if sleva_col else 'null'
    return (
        "function(p){"
        "var api=p.api;if(!api||!api.forEachNodeAfterFilter)return;"
        f"var COLS={cols_json};var SLEVA={sleva_json};"
        "var tot={};COLS.forEach(function(k){tot[k]=0;});"
        "var cnt=0;var hsDisc=0;"
        "api.forEachNodeAfterFilter(function(n){"
        "if(n.rowPinned)return;"
        "cnt++;var d=n.data||{};"
        "COLS.forEach(function(k){"
        "var v=parseFloat(d[k]);"
        "if(!isNaN(v))tot[k]+=v;"
        "});"
        "if(SLEVA){var hs=parseFloat(d['hodn_sankce']);if(isNaN(hs))hs=0;"
        "var sl=parseFloat(d[SLEVA]);if(isNaN(sl))sl=0;"
        "hsDisc+=hs*(1-sl);}"
        "});"
        "if(SLEVA){tot['hodn_sankce']=hsDisc;}"
        "tot['obdobi']='CELKEM';"
        "tot['jmeno_dodavatele']=cnt+' pol.';"
        "api.setGridOption('pinnedBottomRowData',[tot]);"
        "}"
    )


# Formátovač slevy: podíl (0,05) zobraz i jako procenta („0,05 = 5 %"); na
# součtovém řádku prázdné (sleva se promítá do „Hodn. sankce").
_SLEVA_FMT = (
    "function(p){"
    "if(p.node&&p.node.rowPinned)return '';"
    "if(p.value===null||p.value===undefined||p.value==='')return '';"
    "var n=parseFloat(p.value);if(isNaN(n)||n===0)return '';"
    "var pod=new Intl.NumberFormat('cs-CZ',{minimumFractionDigits:0,maximumFractionDigits:4}).format(n);"
    "var pct=new Intl.NumberFormat('cs-CZ',{minimumFractionDigits:0,maximumFractionDigits:2}).format(n*100);"
    "return pod+' ('+pct+' %)';"
    "}"
)


# Předgenerované callbacky per sestava — set sčítaných sloupců se v běhu nemění.
_FILTER_RECALC_ZAMITNUTE = _make_filter_recalc_js(_SOUCET_ZAMITNUTE)
_FILTER_RECALC_VYSTAVENI = _make_filter_recalc_js(_SOUCET_VYSTAVENI, sleva_col='sleva')


# ── Souhrnný pohled po dodavatelích (agregace nad stejnými daty) ──
# Když má dodavatel v podkladových řádcích víc různých stavů, nezobrazí se
# žádný konkrétní — buňka ukáže „různé". Přepsáním se stav nastaví na VŠECHNY
# jeho (aktuálně filtrované) řádky najednou.
_MIX_LABEL = '— různé —'


def _mix_style(base_js: str) -> str:
    """Obalí existující cellStyle callback o vzhled pro smíšenou hodnotu."""
    return ("function(p){if(p.value===%s)return{backgroundColor:'#f1f5f9',"
            "color:'#64748b',fontStyle:'italic'};return (%s)(p);}"
            % (json.dumps(_MIX_LABEL), base_js))


_SOUHRN_STAV_STYLE = _mix_style(_STAV_STYLE)
_SOUHRN_STAV2_STYLE = _mix_style(_STAV2_STYLE)
_FILTER_RECALC_SOUHRN = (
    "function(p){"
    "var api=p.api;if(!api||!api.forEachNodeAfterFilter)return;"
    "var COLS=['pocet','sankce_celkem','sleva_castka','sankce_uznana'];"
    "var tot={};COLS.forEach(function(k){tot[k]=0;});var cnt=0;"
    "api.forEachNodeAfterFilter(function(n){"
    "if(n.rowPinned)return;cnt++;var d=n.data||{};"
    "COLS.forEach(function(k){var v=parseFloat(d[k]);if(!isNaN(v))tot[k]+=v;});"
    "});"
    "tot['ico']='CELKEM';"
    "tot['jmeno_dodavatele']=cnt+' dod.';"
    "api.setGridOption('pinnedBottomRowData',[tot]);"
    "}"
)


# =========================================================
# INICIALIZACE DATABÁZE
# =========================================================
def inicializace_sankce_db():
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS sankce_zamitnute (
                id INT AUTO_INCREMENT PRIMARY KEY,
                obdobi VARCHAR(60),
                obdobi_od DATE, obdobi_do DATE,
                row_hash VARCHAR(40),
                nase_cislo VARCHAR(20),
                ico VARCHAR(30), jmeno_dodavatele VARCHAR(255),
                kod_zbozi VARCHAR(40), nazev_zbozi VARCHAR(255),
                datum_pozadovano VARCHAR(30), typ_adresy VARCHAR(60), k2 VARCHAR(40),
                cislo_objednavky VARCHAR(40), id_pobocky VARCHAR(20),
                dodavatel VARCHAR(60), nakupci VARCHAR(120),
                objednano_mj DOUBLE, dodano_mj DOUBLE, odmitnuto_mj DOUBLE,
                obj_cena DOUBLE, typ_sankce VARCHAR(60),
                hodn_sankce DOUBLE, odmitnuto_kc_celkem DOUBLE,
                poznamka TEXT,
                import_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                imported_by VARCHAR(255),
                INDEX idx_obdobi (obdobi), INDEX idx_hash (row_hash)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS sankce_vystaveni (
                id INT AUTO_INCREMENT PRIMARY KEY,
                obdobi VARCHAR(60),
                obdobi_od DATE, obdobi_do DATE,
                row_hash VARCHAR(40),
                nase_cislo VARCHAR(20),
                ico VARCHAR(30), jmeno_dodavatele VARCHAR(255),
                kod_zbozi VARCHAR(40), nazev_zbozi VARCHAR(255),
                cislo_objednavky VARCHAR(40), id_pobocky VARCHAR(20),
                objednano_mj DOUBLE, dodano_mj DOUBLE, dod_pozde_mj DOUBLE,
                obj_cena DOUBLE, hodn_sankce DOUBLE,
                nakupci_pob VARCHAR(120),
                stav VARCHAR(20) DEFAULT 'nevyfakturovano',
                stav2 VARCHAR(20) DEFAULT 'v_procesu',
                sleva DOUBLE DEFAULT NULL,
                poznamka TEXT,
                import_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                imported_by VARCHAR(255),
                INDEX idx_obdobi (obdobi), INDEX idx_hash (row_hash)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS sankce_audit (
                id INT AUTO_INCREMENT PRIMARY KEY,
                tabulka VARCHAR(20),
                row_hash VARCHAR(40),
                radek_id INT,
                pole VARCHAR(30),
                stara_hodnota TEXT, nova_hodnota TEXT,
                user_id INT, jmeno VARCHAR(255),
                kdy DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_audit (tabulka, row_hash)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        # Diskuze (chat) k jednotlivým případům — klíč (tabulka, row_hash) jako audit,
        # takže vlákno přežije re-import stejného období.
        cur.execute("""
            CREATE TABLE IF NOT EXISTS sankce_chat (
                id INT AUTO_INCREMENT PRIMARY KEY,
                tabulka VARCHAR(20),
                row_hash VARCHAR(40),
                user_id INT, jmeno VARCHAR(255),
                zprava TEXT,
                kdy DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_chat (tabulka, row_hash)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        # Stav přečtení vlákna per uživatel (pro „svítící" upozornění na nové zprávy).
        cur.execute("""
            CREATE TABLE IF NOT EXISTS sankce_chat_precteno (
                tabulka VARCHAR(20),
                row_hash VARCHAR(40),
                user_id INT,
                precteno_id INT DEFAULT 0,
                PRIMARY KEY (tabulka, row_hash, user_id)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        # Zálohy / body obnovení sestavy „Sankce k vystavení" — celá tabulka jako
        # JSON snímek. Druh 'auto' (hodinová) / 'rucni' (ruční bod obnovy).
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS {_ZALOHA_TABLE} (
                id INT AUTO_INCREMENT PRIMARY KEY,
                typ VARCHAR(20) NOT NULL,
                vytvoreno DATETIME DEFAULT CURRENT_TIMESTAMP,
                vytvoril VARCHAR(255),
                pocet_radku INT DEFAULT 0,
                data LONGTEXT,
                INDEX idx_typ (typ), INDEX idx_vytvoreno (vytvoreno)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        # Tikety: předání skupiny řádků „Sankce k vystavení" nákupčímu / provozu.
        # Klíč skupiny = dodavatel (IČO) + kód nákupčího; provozní tiket je per dodavatel.
        cur.execute("""
            CREATE TABLE IF NOT EXISTS sankce_tikety (
                id INT AUTO_INCREMENT PRIMARY KEY,
                typ VARCHAR(10) NOT NULL,
                stav VARCHAR(20) NOT NULL,
                kod_nakupci VARCHAR(20),
                ico VARCHAR(30),
                jmeno_dodavatele VARCHAR(255),
                obdobi VARCHAR(120),
                poznamka TEXT,
                predal VARCHAR(255),
                predano_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                zmeneno DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                zdroj_tiket INT DEFAULT NULL,
                INDEX idx_tiket_stav (stav), INDEX idx_tiket_kod (kod_nakupci)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        # Položky tiketu = odkaz na řádky sankce_vystaveni + rozhodnutí řešitele.
        cur.execute("""
            CREATE TABLE IF NOT EXISTS sankce_tiket_radky (
                id INT AUTO_INCREMENT PRIMARY KEY,
                tiket_id INT NOT NULL,
                radek_id INT,
                row_hash VARCHAR(40),
                rozhodnuti VARCHAR(20) DEFAULT NULL,
                INDEX idx_tr_hash (row_hash),
                UNIQUE KEY uniq_tr (tiket_id, radek_id)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        # Migrace: doplň sloupec „naše pořadové číslo" do starších tabulek + dočísluj.
        for _t in ('sankce_zamitnute', 'sankce_vystaveni'):
            cur.execute("SELECT COUNT(*) FROM information_schema.COLUMNS "
                        "WHERE TABLE_SCHEMA=DATABASE() AND TABLE_NAME=%s AND COLUMN_NAME='nase_cislo'",
                        (_t,))
            if cur.fetchone()[0] == 0:
                cur.execute(f'ALTER TABLE {_t} ADD COLUMN nase_cislo VARCHAR(20) AFTER row_hash')
        # Migrace: druhý stav „Aktivita 2" do sankce_vystaveni (V procesu / Uzavřeno).
        cur.execute("SELECT COUNT(*) FROM information_schema.COLUMNS "
                    "WHERE TABLE_SCHEMA=DATABASE() AND TABLE_NAME='sankce_vystaveni' "
                    "AND COLUMN_NAME='stav2'")
        if cur.fetchone()[0] == 0:
            cur.execute("ALTER TABLE sankce_vystaveni ADD COLUMN stav2 VARCHAR(20) "
                        "DEFAULT 'v_procesu' AFTER stav")
        # Migrace: sleva na sankci (podíl, 0.05 = 5 %) — promítá se jen do součtu.
        cur.execute("SELECT COUNT(*) FROM information_schema.COLUMNS "
                    "WHERE TABLE_SCHEMA=DATABASE() AND TABLE_NAME='sankce_vystaveni' "
                    "AND COLUMN_NAME='sleva'")
        if cur.fetchone()[0] == 0:
            cur.execute("ALTER TABLE sankce_vystaveni ADD COLUMN sleva DOUBLE "
                        "DEFAULT NULL AFTER hodn_sankce")
        # Migrace: nákupčí pobočky (sloupec L exportu) do sankce_vystaveni.
        cur.execute("SELECT COUNT(*) FROM information_schema.COLUMNS "
                    "WHERE TABLE_SCHEMA=DATABASE() AND TABLE_NAME='sankce_vystaveni' "
                    "AND COLUMN_NAME='nakupci_pob'")
        if cur.fetchone()[0] == 0:
            cur.execute("ALTER TABLE sankce_vystaveni ADD COLUMN nakupci_pob VARCHAR(120) "
                        "DEFAULT NULL AFTER hodn_sankce")
        conn.commit()
        cur.close()
        for _t in ('sankce_zamitnute', 'sankce_vystaveni'):
            _backfill_cisla(_t)
    except Exception as e:
        print(f'Chyba při inicializaci DB Sankcí: {e}')
    finally:
        conn.close()


def _backfill_cisla(tabulka: str):
    """Přidělí „naše pořadové číslo" řádkům, které ho ještě nemají (rok z období,
    čítač per rok, chronologicky). Po prvním proběhnutí je to už jen prázdný SELECT."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.execute(f"SELECT id, YEAR(obdobi_od) FROM {tabulka} "
                    f"WHERE nase_cislo IS NULL OR nase_cislo='' ORDER BY obdobi_od, id")
        chybi = cur.fetchall()
        if not chybi:
            cur.close()
            return
        cur.execute(f"SELECT LEFT(nase_cislo,4), MAX(CAST(SUBSTRING(nase_cislo,5) AS UNSIGNED)) "
                    f"FROM {tabulka} WHERE nase_cislo IS NOT NULL AND nase_cislo<>'' "
                    f"GROUP BY LEFT(nase_cislo,4)")
        maxseq = {r[0]: int(r[1] or 0) for r in cur.fetchall()}
        dnes_rok = datetime.date.today().year
        upd = []
        for rid, rok in chybi:
            ys = str(rok or dnes_rok)
            nxt = maxseq.get(ys, 0) + 1
            maxseq[ys] = nxt
            upd.append((f'{ys}{nxt:05d}', rid))
        cur.executemany(f'UPDATE {tabulka} SET nase_cislo=%s WHERE id=%s', upd)
        conn.commit()
        cur.close()
    except Exception as e:
        print(f'[sankce] _backfill_cisla({tabulka}) error: {e}')
    finally:
        conn.close()


# =========================================================
# AUDIT (očičko)
# =========================================================
def zapis_audit(tabulka, row_hash, radek_id, pole, stara, nova, user_id, jmeno):
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.execute(
            'INSERT INTO sankce_audit '
            '(tabulka,row_hash,radek_id,pole,stara_hodnota,nova_hodnota,user_id,jmeno) '
            'VALUES (%s,%s,%s,%s,%s,%s,%s,%s)',
            (tabulka, row_hash, radek_id, pole,
             None if stara is None else str(stara),
             None if nova is None else str(nova),
             user_id, jmeno or ''),
        )
        conn.commit(); cur.close()
    except Exception as e:
        print(f'[sankce] zapis_audit error: {e}')
    finally:
        conn.close()


def nacti_audit(tabulka, row_hash) -> list:
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(
            'SELECT pole,stara_hodnota,nova_hodnota,jmeno,kdy FROM sankce_audit '
            'WHERE tabulka=%s AND row_hash=%s ORDER BY kdy DESC, id DESC',
            (tabulka, row_hash),
        )
        return cur.fetchall()
    finally:
        conn.close()


# =========================================================
# CHAT / DISKUZE K PŘÍPADU (per řádek, klíč tabulka+row_hash)
# =========================================================
def _nacti_chat_zpravy(tabulka, row_hash) -> list:
    """Zprávy vlákna chronologicky (nejstarší nahoře)."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute('SELECT id,user_id,jmeno,zprava,kdy FROM sankce_chat '
                    'WHERE tabulka=%s AND row_hash=%s ORDER BY id ASC',
                    (tabulka, row_hash))
        return cur.fetchall()
    finally:
        conn.close()


def _chat_max_id(tabulka, row_hash) -> int:
    """Id poslední zprávy ve vlákně (0 = žádná)."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return 0
    try:
        cur = conn.cursor()
        cur.execute('SELECT MAX(id) FROM sankce_chat WHERE tabulka=%s AND row_hash=%s',
                    (tabulka, row_hash))
        r = cur.fetchone()
        return int(r[0]) if r and r[0] else 0
    finally:
        conn.close()


def _oznac_precteno(tabulka, row_hash, user_id):
    """Označí vlákno pro uživatele jako přečtené až po poslední zprávu."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.execute('SELECT MAX(id) FROM sankce_chat WHERE tabulka=%s AND row_hash=%s',
                    (tabulka, row_hash))
        r = cur.fetchone()
        max_id = int(r[0]) if r and r[0] else 0
        cur.execute('INSERT INTO sankce_chat_precteno (tabulka,row_hash,user_id,precteno_id) '
                    'VALUES (%s,%s,%s,%s) ON DUPLICATE KEY UPDATE precteno_id=VALUES(precteno_id)',
                    (tabulka, row_hash, user_id, max_id))
        conn.commit(); cur.close()
    except Exception as e:
        print(f'[sankce] _oznac_precteno error: {e}')
    finally:
        conn.close()


def _pridej_chat(tabulka, row_hash, user_id, jmeno, zprava) -> int:
    """Vloží zprávu a rovnou ji autorovi označí jako přečtenou. Vrací id zprávy."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return 0
    try:
        cur = conn.cursor()
        cur.execute('INSERT INTO sankce_chat (tabulka,row_hash,user_id,jmeno,zprava) '
                    'VALUES (%s,%s,%s,%s,%s)', (tabulka, row_hash, user_id, jmeno or '', zprava))
        new_id = cur.lastrowid
        cur.execute('INSERT INTO sankce_chat_precteno (tabulka,row_hash,user_id,precteno_id) '
                    'VALUES (%s,%s,%s,%s) ON DUPLICATE KEY UPDATE precteno_id=VALUES(precteno_id)',
                    (tabulka, row_hash, user_id, new_id))
        conn.commit(); cur.close()
        return new_id
    except Exception as e:
        print(f'[sankce] _pridej_chat error: {e}')
        try:
            conn.rollback()
        except Exception:
            pass
        return 0
    finally:
        conn.close()


def _smaz_chat(tabulka, row_hash, msg_id, user_id, muze_mazat_vse=False) -> bool:
    """Smaže zprávu vlákna. Vlastní zprávu smí smazat autor, cizí jen ten, kdo
    `muze_mazat_vse` (správce / analytik). Vrací True při úspěchu."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return False
    try:
        cur = conn.cursor()
        if muze_mazat_vse:
            cur.execute('DELETE FROM sankce_chat WHERE id=%s AND tabulka=%s AND row_hash=%s',
                        (msg_id, tabulka, row_hash))
        else:
            cur.execute('DELETE FROM sankce_chat WHERE id=%s AND tabulka=%s AND row_hash=%s '
                        'AND user_id=%s', (msg_id, tabulka, row_hash, user_id))
        smazano = cur.rowcount > 0
        conn.commit(); cur.close()
        return smazano
    except Exception as e:
        print(f'[sankce] _smaz_chat error: {e}')
        try:
            conn.rollback()
        except Exception:
            pass
        return False
    finally:
        conn.close()


# Práva, kterým „cinkne" zvoneček při novém komentáři — podle sestavy.
_KOMENTAR_PRAVA = {
    'sankce_vystaveni': ('vse', 'sankce_analytik', 'sankce_ucetni', 'sankce_ctenar'),
    'sankce_zamitnute': ('vse', 'sankce_analytik', 'sankce_nakup', 'sankce_ctenar'),
    'sankce_tikety':    ('vse', 'sankce_analytik', 'sankce_ucetni'),
}
_KOMENTAR_SESTAVA = {
    'sankce_vystaveni': 'Sankce k vystavení',
    'sankce_zamitnute': 'Zamítnuté dodávky',
    'sankce_tikety':    'Tiket',
}


def _notifikuj_novy_komentar(tabulka, row_hash, autor_id, autor_jmeno, popis,
                             prava_navic=()):
    """Po napsání komentáře k případu „cinkne" oznámení do zvonečku všem, kdo na
    danou sestavu mají právo (kromě autora). Volá se mimo event loop (DB dotazy),
    takže se neblokuje odeslání zprávy. `prava_navic` = adresáti navíc (u tiketu
    jeho řešitel — nákupčí / provoz / kontrola)."""
    try:
        prava = tuple(_KOMENTAR_PRAVA.get(tabulka) or ()) + tuple(prava_navic or ())
        if not prava:
            return
        prijemci = intranet_data.ziskej_uzivatele_s_pravem(*prava)  # {id: jmeno}
        if not prijemci:
            return
        sestava = _KOMENTAR_SESTAVA.get(tabulka, 'Sankce')
        popis_txt = (popis or '').strip(' –') or 'případ'
        text = f'💬 {autor_jmeno} přidal(a) komentář ({sestava}): {popis_txt}'
        for uid in prijemci:
            try:
                if uid is None or int(uid) == int(autor_id):
                    continue
            except (TypeError, ValueError):
                continue
            intranet_notifikace.pridej(uid, text, 'info')
    except Exception as e:
        print(f'[sankce] _notifikuj_novy_komentar error: {e}')


def _chat_stav_radku(tabulka, row_hash, user_id) -> dict:
    """Stav jednoho vlákna pro uživatele: {'pocet': n, 'unread': bool}."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return {'pocet': 0, 'unread': False}
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute('SELECT COUNT(*) AS pocet, MAX(id) AS max_id FROM sankce_chat '
                    'WHERE tabulka=%s AND row_hash=%s', (tabulka, row_hash))
        a = cur.fetchone() or {}
        pocet = a.get('pocet') or 0
        max_id = a.get('max_id') or 0
        cur.execute('SELECT precteno_id FROM sankce_chat_precteno '
                    'WHERE tabulka=%s AND row_hash=%s AND user_id=%s', (tabulka, row_hash, user_id))
        p = cur.fetchone()
        read_id = (p.get('precteno_id') if p else 0) or 0
        return {'pocet': int(pocet), 'unread': bool(max_id and max_id > read_id)}
    finally:
        conn.close()


def _nacti_chat_stav(tabulka, user_id) -> dict:
    """Stav VŠECH vláken sestavy pro uživatele: {row_hash: {'pocet','unread'}}.
    Jeden průchod pro celý grid (kvůli indikátorům i pollingu)."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return {}
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute('SELECT row_hash, COUNT(*) AS pocet, MAX(id) AS max_id FROM sankce_chat '
                    'WHERE tabulka=%s GROUP BY row_hash', (tabulka,))
        agg = {r['row_hash']: (r['pocet'] or 0, r['max_id'] or 0) for r in cur.fetchall()}
        cur.execute('SELECT row_hash, precteno_id FROM sankce_chat_precteno '
                    'WHERE tabulka=%s AND user_id=%s', (tabulka, user_id))
        read = {r['row_hash']: (r['precteno_id'] or 0) for r in cur.fetchall()}
        out = {}
        for rh, (pocet, max_id) in agg.items():
            out[rh] = {'pocet': int(pocet), 'unread': bool(max_id > read.get(rh, 0))}
        return out
    finally:
        conn.close()


# =========================================================
# ZÁLOHY / BODY OBNOVENÍ („Sankce k vystavení")
# =========================================================
# Celá tabulka sankce_vystaveni (všechny sloupce) se ukládá jako JSON snímek do
# sankce_vystaveni_zalohy. Druhy: 'auto' (hodinová automatická záloha) a 'rucni'
# (ruční bod obnovy). Zálohy NEJDE stahovat — slouží jen k obnovení (přehrání) dat.
def _json_default(o):
    """Serializace hodnot z DB do JSON (datum/čas → string, Decimal → float)."""
    if isinstance(o, datetime.datetime):
        return o.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(o, datetime.date):
        return o.isoformat()
    try:
        import decimal
        if isinstance(o, decimal.Decimal):
            return float(o)
    except Exception:
        pass
    return str(o)


def _vytvor_zalohu(typ: str, vytvoril: str, skip_if_empty: bool = False):
    """Vytvoří snímek celé tabulky sankce_vystaveni a uloží ho jako zálohu typu
    'auto' / 'rucni'. U automatické zálohy umí přeskočit prázdnou tabulku
    (skip_if_empty) a po uložení prořeže staré AUTO zálohy. Ruční body obnovy se
    nemažou. Vrací (id|None, počet_řádků, chyba|None)."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return None, 0, 'Není spojení s databází.'
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute('SELECT * FROM sankce_vystaveni ORDER BY id ASC')
        rows = cur.fetchall()
        cur.close()
        if skip_if_empty and not rows:
            return None, 0, None
        data_json = json.dumps(rows, ensure_ascii=False, default=_json_default)
        cur = conn.cursor()
        cur.execute(
            f'INSERT INTO {_ZALOHA_TABLE} (typ, vytvoril, pocet_radku, data) '
            'VALUES (%s,%s,%s,%s)',
            (typ, vytvoril or '', len(rows), data_json),
        )
        new_id = cur.lastrowid
        if typ == 'auto':
            # prořež staré automatické zálohy (ruční body obnovy zůstávají)
            cur.execute(
                f'DELETE FROM {_ZALOHA_TABLE} WHERE typ=%s AND id NOT IN '
                f'(SELECT id FROM (SELECT id FROM {_ZALOHA_TABLE} WHERE typ=%s '
                f'ORDER BY id DESC LIMIT %s) keep)',
                ('auto', 'auto', _ZALOHA_AUTO_MAX),
            )
        conn.commit()
        cur.close()
        return new_id, len(rows), None
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return None, 0, str(e)
    finally:
        conn.close()


def _nacti_zalohy() -> list:
    """Seznam záloh (jen metadata — bez datového blobu), nejnovější nahoře."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(
            f'SELECT id, typ, vytvoreno, vytvoril, pocet_radku '
            f'FROM {_ZALOHA_TABLE} ORDER BY vytvoreno DESC, id DESC'
        )
        return cur.fetchall()
    finally:
        conn.close()


def _nacti_zaloha_data(zaloha_id: int):
    """Vrátí JSON snímek (string) zvolené zálohy nebo None, když neexistuje."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return None
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(f'SELECT data FROM {_ZALOHA_TABLE} WHERE id=%s', (zaloha_id,))
        z = cur.fetchone()
        return (z.get('data') if z else None)
    finally:
        conn.close()


def _posledni_auto_zaloha_cas():
    """Čas poslední automatické zálohy (datetime) nebo None."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return None
    try:
        cur = conn.cursor()
        cur.execute(f"SELECT MAX(vytvoreno) FROM {_ZALOHA_TABLE} WHERE typ='auto'")
        r = cur.fetchone()
        return r[0] if r and r[0] else None
    finally:
        conn.close()


def _obnov_zalohu(zaloha_id: int, vytvoril: str):
    """Přehraje data: nejdřív pojistně zazálohuje AKTUÁLNÍ stav (auto), pak nahradí
    celý obsah tabulky sankce_vystaveni daty ze zvolené zálohy. Díky pojistné záloze
    jde obnovu vrátit zpět. Vrací (ok, počet_obnovených_řádků, chyba|None)."""
    raw = _nacti_zaloha_data(zaloha_id)
    if raw is None:
        return False, 0, 'Záloha nebyla nalezena (možná byla mezitím odstraněna).'
    try:
        rows = json.loads(raw or '[]')
    except Exception as e:
        return False, 0, f'Záloha je poškozená: {e}'

    # Pojistka: aktuální stav ulož jako automatickou zálohu, ať jde obnova vrátit zpět.
    try:
        _vytvor_zalohu('auto', f'{vytvoril} (před obnovou)', skip_if_empty=True)
    except Exception:
        pass

    conn = intranet_data.get_db_connection()
    if not conn:
        return False, 0, 'Není spojení s databází.'
    try:
        cur = conn.cursor()
        cur.execute('DELETE FROM sankce_vystaveni')
        if rows:
            cols = list(rows[0].keys())
            col_sql = ','.join('`' + c + '`' for c in cols)
            ph = ','.join(['%s'] * len(cols))
            vals = [[r.get(c) for c in cols] for r in rows]
            cur.executemany(
                f'INSERT INTO sankce_vystaveni ({col_sql}) VALUES ({ph})', vals)
        conn.commit()
        cur.close()
        return True, len(rows), None
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return False, 0, str(e)
    finally:
        conn.close()


def _smaz_zalohu(zaloha_id: int):
    """Smaže zálohu — POUZE ruční bod obnovy (`typ='rucni'`). Automatické zálohy
    chrání podmínka v SQL, takže je nelze smazat ani omylem. Vrací (ok, chyba|None);
    ok=False bez chyby = nešlo o ruční bod (auto / už neexistuje)."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return False, 'Není spojení s databází.'
    try:
        cur = conn.cursor()
        cur.execute(f"DELETE FROM {_ZALOHA_TABLE} WHERE id=%s AND typ='rucni'",
                    (zaloha_id,))
        smazano = cur.rowcount
        conn.commit()
        cur.close()
        return (smazano > 0), None
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return False, str(e)
    finally:
        conn.close()


def _smaz_zalohy_hromadne(ids: list, povolit_auto: bool = False):
    """Hromadně smaže zálohy podle seznamu id. Když povolit_auto=False, smaže jen
    ruční body obnovy (auto zálohy zůstanou chráněné); povolit_auto=True (hlavní
    administrátor) smaže i automatické. Vrací (počet_smazaných, chyba|None)."""
    ids = [int(i) for i in ids if i is not None]
    if not ids:
        return 0, None
    conn = intranet_data.get_db_connection()
    if not conn:
        return 0, 'Není spojení s databází.'
    try:
        cur = conn.cursor()
        ph = ','.join(['%s'] * len(ids))
        if povolit_auto:
            cur.execute(f'DELETE FROM {_ZALOHA_TABLE} WHERE id IN ({ph})', tuple(ids))
        else:
            cur.execute(f"DELETE FROM {_ZALOHA_TABLE} WHERE typ='rucni' AND id IN ({ph})",
                        tuple(ids))
        smazano = cur.rowcount
        conn.commit()
        cur.close()
        return smazano, None
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return 0, str(e)
    finally:
        conn.close()


async def bg_sankce_zaloha():
    """Serverová smyčka: každé 4 hodiny (zarovnáno na 00/04/08/12/16/20 h) zazálohuje
    celou tabulku sankce_vystaveni jako automatickou zálohu. Stav drží DB (čas
    poslední auto zálohy), takže záloha přežije restart, neudělá se dvakrát v jednom
    4hodinovém slotu a běží jen jednou pro celý server. Drží se posledních
    _ZALOHA_AUTO_MAX (10) auto záloh — při překročení padá vždy ta nejstarší.
    Prázdnou tabulku přeskakuje."""
    try:
        await asyncio.to_thread(inicializace_sankce_db)
    except Exception as e:
        print(f'[sankce] bg_sankce_zaloha init error: {e}')
    while True:
        try:
            posledni = await asyncio.to_thread(_posledni_auto_zaloha_cas)
            now = datetime.datetime.now()
            # začátek aktuálního 4hodinového slotu (00:00, 04:00, 08:00, …)
            slot_h = (now.hour // _ZALOHA_INTERVAL_H) * _ZALOHA_INTERVAL_H
            tento_slot = now.replace(hour=slot_h, minute=0, second=0, microsecond=0)
            if posledni is None or posledni < tento_slot:
                _id, cnt, err = await asyncio.to_thread(
                    _vytvor_zalohu, 'auto', 'Systém', True)
                if err:
                    print(f'[sankce] automatická záloha CHYBA: {err}')
                elif _id is not None:
                    print(f'[sankce] automatická záloha OK (#{_id}, {cnt} řádků)')
        except Exception as e:
            print(f'[sankce] bg_sankce_zaloha error: {e}')
        await asyncio.sleep(60)


app.on_startup(lambda: asyncio.create_task(bg_sankce_zaloha()))


# =========================================================
# ZÁMEK OBDOBÍ (globální výchozí filtr pro Zamítnuté)
# =========================================================
def _ziskej_zamcene_obdobi() -> str:
    return intranet_data.nacti_nastaveni_intranetu().get(_NAST_KLIC_ZAMEK, '')


def _uloz_zamcene_obdobi(obdobi: str):
    n = intranet_data.nacti_nastaveni_intranetu()
    n[_NAST_KLIC_ZAMEK] = obdobi or ''
    intranet_data.uloz_nastaveni_intranetu(n)


# =========================================================
# NAČÍTÁNÍ DAT
# =========================================================
def _nacti(tabulka: str) -> list:
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(f'SELECT * FROM {tabulka} ORDER BY obdobi_od DESC, id ASC')
        rows = cur.fetchall()
        for r in rows:
            # Datum/čas sloupce do gridu nepatří (JSON serializace) — odstraníme je.
            for k in ('obdobi_od', 'obdobi_do', 'import_at', 'imported_by'):
                r.pop(k, None)
            if 'stav' in r:
                r['stav_label'] = STAV_LABEL.get(r.get('stav'), 'Nová data')
            if 'stav2' in r:
                r['stav2_label'] = STAV2_LABEL.get(r.get('stav2'), STAV2_LABEL[STAV2_DEFAULT])
        return rows
    finally:
        conn.close()


def _seznam_obdobi(tabulka: str) -> list:
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor()
        cur.execute(f'SELECT obdobi FROM {tabulka} WHERE obdobi IS NOT NULL '
                    f'GROUP BY obdobi, obdobi_od ORDER BY obdobi_od DESC')
        return [r[0] for r in cur.fetchall() if r[0]]
    finally:
        conn.close()


def _seznam_obdobi_detail(tabulka: str) -> list:
    """Existující období i s ISO daty OD/DO (nejnovější nahoře) — pro výběr období
    v import dialogu, ať se nemusí psát ručně. Vrací [{'obdobi','od','do'}]."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(f'SELECT obdobi, MAX(obdobi_od) AS od, MAX(obdobi_do) AS do '
                    f'FROM {tabulka} WHERE obdobi IS NOT NULL '
                    f'GROUP BY obdobi ORDER BY MAX(obdobi_od) DESC')
        out = []
        for r in cur.fetchall():
            if not r.get('obdobi'):
                continue
            od = r.get('od'); do = r.get('do')
            out.append({
                'obdobi': r['obdobi'],
                'od': od.isoformat() if hasattr(od, 'isoformat') else (od or ''),
                'do': do.isoformat() if hasattr(do, 'isoformat') else (do or ''),
            })
        return out
    finally:
        conn.close()


# =========================================================
# IMPORT (list DATA)
# =========================================================
def _importuj_sync(raw: bytes, tabulka: str, mapa: dict, cisla: set,
                   obdobi: str, od_iso: str, do_iso: str, user_name: str,
                   pripoj: bool = False):
    """Naimportuje list DATA. Dva režimy:
      • pripoj=False (výchozí) — NAHRADÍ celou dávku stejného období a zachová ručně
        zadané poznámky (a u 'Sankce k vystavení' i stavy) párováním přes row_hash;
      • pripoj=True — jen DOPLNÍ k existujícím datům NOVÉ řádky (nové row_hash) ve stavu
        „Rozpracováno". Stará dávka se nemaže, existující řádky (vč. stavu/poznámky)
        zůstanou beze změny; řádky, které už v daném období jsou, se přeskočí.
    Vrací (počet_vložených, počet_přeskočených, chyba|None)."""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as e:
        return 0, 0, f'Soubor nelze otevřít: {e}'

    if 'DATA' not in wb.sheetnames:
        return 0, 0, 'Soubor neobsahuje list „DATA".'
    ws = wb['DATA']

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return 0, 0, 'List DATA je prázdný.'

    # index sloupce → DB pole
    idx_na_pole = {}
    for i, h in enumerate(header):
        pole = mapa.get(_norm(h))
        if pole:
            idx_na_pole[i] = pole
    # Nákupčí (pob.) přibyl do exportu dodatečně – starší soubory smí přijít bez něj.
    chybi = set(mapa.values()) - set(idx_na_pole.values()) - _VOLITELNE
    if chybi:
        return 0, 0, 'V listu DATA chybí sloupce: ' + ', '.join(sorted(chybi))

    je_vystaveni = (tabulka == 'sankce_vystaveni')
    zaznamy = []
    for r in rows_iter:
        if r is None or all(c is None or c == '' for c in r):
            continue
        radek = {}
        for i, pole in idx_na_pole.items():
            v = r[i] if i < len(r) else None
            radek[pole] = _f(v) if pole in cisla else _s(v)
        # přeskoč úplně prázdné (bez IČO i názvu)
        if not radek.get('ico') and not radek.get('nazev_zbozi'):
            continue
        # row_hash z rozlišujících polí
        if je_vystaveni:
            casti = [radek.get('ico'), radek.get('kod_zbozi'), radek.get('cislo_objednavky'),
                     radek.get('id_pobocky'), radek.get('hodn_sankce')]
        else:
            casti = [radek.get('ico'), radek.get('kod_zbozi'), radek.get('cislo_objednavky'),
                     radek.get('id_pobocky'), radek.get('datum_pozadovano'),
                     radek.get('odmitnuto_mj'), radek.get('hodn_sankce')]
        radek['row_hash'] = _row_hash(tabulka, obdobi, casti)
        zaznamy.append(radek)

    if not zaznamy:
        return 0, 0, 'V listu DATA nejsou žádné datové řádky.'

    conn = intranet_data.get_db_connection()
    if not conn:
        return 0, 0, 'Chyba připojení k databázi.'
    try:
        cur = conn.cursor(dictionary=True)
        if pripoj:
            # Režim PŘIDÁNÍ: dávku NEMAŽEME. Zjisti, co už v daném období je, ať
            # doplníme jen nové řádky (existující — vč. jejich stavu — neměníme).
            existujici = set()
            cur.execute(f'SELECT row_hash FROM {tabulka} WHERE obdobi=%s', (obdobi,))
            for x in cur.fetchall():
                existujici.add(x['row_hash'])
            rucni = {}
        else:
            # Režim NAHRAZENÍ: zachovej ručně zadané hodnoty (vč. „našeho čísla").
            existujici = set()
            rucni = {}
            if je_vystaveni:
                cur.execute('SELECT row_hash,stav,stav2,sleva,poznamka,nase_cislo FROM sankce_vystaveni WHERE obdobi=%s', (obdobi,))
                for x in cur.fetchall():
                    rucni[x['row_hash']] = {'stav': x.get('stav'), 'stav2': x.get('stav2'),
                                            'sleva': x.get('sleva'),
                                            'poznamka': x.get('poznamka'),
                                            'nase_cislo': x.get('nase_cislo')}
            else:
                cur.execute('SELECT row_hash,poznamka,nase_cislo FROM sankce_zamitnute WHERE obdobi=%s', (obdobi,))
                for x in cur.fetchall():
                    rucni[x['row_hash']] = {'poznamka': x.get('poznamka'),
                                            'nase_cislo': x.get('nase_cislo')}

        # „Naše pořadové číslo": rok z období + další volné pořadí v rámci roku
        # (počítáno PŘED smazáním, takže nová čísla nikdy nekolidují se zachovanými).
        rok = (od_iso or '')[:4] or str(datetime.date.today().year)
        cur.execute(f"SELECT MAX(CAST(SUBSTRING(nase_cislo,5) AS UNSIGNED)) AS mx FROM {tabulka} "
                    f"WHERE nase_cislo LIKE %s", (rok + '%',))
        _mx = cur.fetchone()
        next_seq = int((_mx.get('mx') if _mx else 0) or 0)

        # smaž starou dávku téhož období JEN v režimu nahrazení
        cur2 = conn.cursor()
        if not pripoj:
            cur2.execute(f'DELETE FROM {tabulka} WHERE obdobi=%s', (obdobi,))

        # vlož nové řádky
        pole_data = list(mapa.values())
        if je_vystaveni:
            sloupce = ['obdobi', 'obdobi_od', 'obdobi_do', 'row_hash', 'nase_cislo'] + pole_data + \
                      ['stav', 'stav2', 'sleva', 'poznamka', 'imported_by']
        else:
            sloupce = ['obdobi', 'obdobi_od', 'obdobi_do', 'row_hash', 'nase_cislo'] + pole_data + \
                      ['poznamka', 'imported_by']
        placeholders = ','.join(['%s'] * len(sloupce))
        sql = f"INSERT INTO {tabulka} ({','.join(sloupce)}) VALUES ({placeholders})"

        davka = []
        preskoceno = 0
        videno = set()   # dedup i v rámci jednoho souboru (v režimu přidání)
        for radek in zaznamy:
            rh = radek['row_hash']
            # V režimu přidání přeskoč, co už v období je (nebo se v souboru opakuje).
            if pripoj and (rh in existujici or rh in videno):
                preskoceno += 1
                continue
            videno.add(rh)
            zachov = rucni.get(rh, {})
            cislo = zachov.get('nase_cislo')
            if not cislo:
                next_seq += 1
                cislo = f'{rok}{next_seq:05d}'
            zaklad = [obdobi, od_iso or None, do_iso or None, rh, cislo] + \
                     [radek.get(p) for p in pole_data]
            if je_vystaveni:
                # Re-import (nahrazení) zachová předchozí stav řádku; nový řádek startuje
                # ve „Nová data". Dodatečně PŘIDANÝ dodavatel (pripoj) startuje rovnou ve
                # stavu „Rozpracováno".
                vychozi_stav = 'rozpracovano' if pripoj else 'nova_data'
                zaklad += [zachov.get('stav') or vychozi_stav,
                           zachov.get('stav2') or STAV2_DEFAULT,
                           zachov.get('sleva'),
                           zachov.get('poznamka'), user_name]
            else:
                zaklad += [zachov.get('poznamka'), user_name]
            davka.append(tuple(zaklad))
        if davka:
            cur2.executemany(sql, davka)
        conn.commit()
        cur.close(); cur2.close()
        return len(davka), preskoceno, None
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return 0, 0, f'Chyba zápisu do databáze: {e}'
    finally:
        conn.close()


def _smaz_data(tabulka: str, obdobi=None) -> tuple:
    """Nevratně smaže data sestavy včetně jejich historie změn (audit):
      • obdobi=None → VŠECHNA data sestavy,
      • obdobi='…'  → jen dané období (row_hash nese období, takže audit se
        maže bezpečně jen pro toto období).
    `tabulka` je vnitřní konstanta. Vrací (počet_smazaných_řádků, chyba|None)."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return 0, 'Chyba připojení k databázi.'
    try:
        cur = conn.cursor()
        if obdobi is None:
            cur.execute(f'SELECT COUNT(*) FROM {tabulka}')
            pocet = cur.fetchone()[0]
            cur.execute('DELETE FROM sankce_audit WHERE tabulka=%s', (tabulka,))
            cur.execute(f'DELETE FROM {tabulka}')
        else:
            cur.execute(f'SELECT COUNT(*) FROM {tabulka} WHERE obdobi=%s', (obdobi,))
            pocet = cur.fetchone()[0]
            # audit smaž dřív, dokud řádky ještě existují pro poddotaz (jiná tabulka)
            cur.execute(f'DELETE FROM sankce_audit WHERE tabulka=%s AND row_hash IN '
                        f'(SELECT row_hash FROM {tabulka} WHERE obdobi=%s)', (tabulka, obdobi))
            cur.execute(f'DELETE FROM {tabulka} WHERE obdobi=%s', (obdobi,))
        conn.commit()
        cur.close()
        return pocet, None
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return 0, f'Chyba mazání: {e}'
    finally:
        conn.close()


def _smaz_radky(tabulka: str, radek_ids: list) -> tuple:
    """Smaže JEDEN nebo VÍC řádků (podle id). Historii změn (audit) k jejich
    row_hash uklidí jen u těch otisků, které už v sestavě nezůstaly. Vrací
    (počet_smazaných, chyba|None)."""
    ids = []
    for x in (radek_ids or []):
        try:
            ids.append(int(x))
        except (TypeError, ValueError):
            pass
    if not ids:
        return 0, None
    conn = intranet_data.get_db_connection()
    if not conn:
        return 0, 'Chyba připojení k databázi.'
    try:
        cur = conn.cursor()
        ph = ','.join(['%s'] * len(ids))
        # otisky dotčených řádků (pro pozdější úklid auditu)
        cur.execute(f'SELECT DISTINCT row_hash FROM {tabulka} WHERE id IN ({ph})', tuple(ids))
        hashes = [r[0] for r in cur.fetchall() if r[0]]
        cur.execute(f'DELETE FROM {tabulka} WHERE id IN ({ph})', tuple(ids))
        smazano = cur.rowcount
        # audit smaž jen u otisků, ke kterým už žádný řádek nezůstal
        for rh in hashes:
            cur.execute(f'SELECT COUNT(*) FROM {tabulka} WHERE row_hash=%s', (rh,))
            if cur.fetchone()[0] == 0:
                cur.execute('DELETE FROM sankce_audit WHERE tabulka=%s AND row_hash=%s',
                            (tabulka, rh))
        conn.commit()
        cur.close()
        return smazano, None
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return 0, f'Chyba mazání: {e}'
    finally:
        conn.close()


# =========================================================
# DIALOG HISTORIE ŘÁDKU (očičko)
# =========================================================
_POLE_LABEL = {'stav': 'Stav', 'stav2': 'Aktivita 2', 'sleva': 'Sleva na sankci',
               'poznamka': 'Poznámka', 'tiket': 'Tiket', 'rozhodnuti': 'Rozhodnutí',
               'tiket_stav': 'Stav tiketu'}


def _zobraz_historii(tabulka: str, row_hash: str, popis: str):
    zaznamy = nacti_audit(tabulka, row_hash)
    with ui.dialog() as dlg, ui.card().classes('p-5 gap-3') \
            .style('min-width: 560px; max-width: 820px; max-height: 80vh; overflow-y: auto'):
        with ui.row().classes('items-center gap-2'):
            ui.icon('history', color='teal')
            ui.label('Historie změn řádku').classes('text-lg font-bold text-gray-800')
        if popis:
            ui.label(popis).classes('text-sm text-gray-500 -mt-2')
        if not zaznamy:
            ui.label('U tohoto řádku zatím nebyla zaznamenána žádná změna.') \
                .classes('text-sm text-gray-500 italic py-4')
        else:
            columns = [
                {'name': 'cas', 'label': 'Čas', 'field': 'cas', 'align': 'left'},
                {'name': 'kdo', 'label': 'Kdo', 'field': 'kdo', 'align': 'left'},
                {'name': 'pole', 'label': 'Pole', 'field': 'pole', 'align': 'left'},
                {'name': 'zmena', 'label': 'Změna', 'field': 'zmena', 'align': 'left'},
            ]
            t_rows = []
            for z in zaznamy:
                cas = z.get('kdy')
                cas_txt = cas.strftime('%d.%m.%Y %H:%M') if hasattr(cas, 'strftime') else str(cas or '')
                pole = z.get('pole', '')
                if pole == 'stav':
                    stara = STAV_LABEL.get(z.get('stara_hodnota'), z.get('stara_hodnota'))
                    nova = STAV_LABEL.get(z.get('nova_hodnota'), z.get('nova_hodnota'))
                elif pole == 'stav2':
                    stara = STAV2_LABEL.get(z.get('stara_hodnota'), z.get('stara_hodnota'))
                    nova = STAV2_LABEL.get(z.get('nova_hodnota'), z.get('nova_hodnota'))
                else:
                    stara = z.get('stara_hodnota')
                    nova = z.get('nova_hodnota')
                t_rows.append({
                    'cas': cas_txt,
                    'kdo': z.get('jmeno') or '—',
                    'pole': _POLE_LABEL.get(pole, pole),
                    'zmena': f'{stara or "—"} → {nova or "—"}',
                })
            ui.table(columns=columns, rows=t_rows, row_key='cas') \
                .classes('w-full').props('dense flat bordered')
            ui.label(f'Celkem změn: {len(zaznamy)}').classes('text-xs text-gray-400')
        with ui.row().classes('justify-end w-full'):
            ui.button('Zavřít', on_click=dlg.close).props('flat no-caps')
    dlg.open()


# =========================================================
# DIALOG OBNOVENÍ DAT (zálohy — „Sankce k vystavení")
# =========================================================
def _potvrd_smaz_zaloha_dialog(zaloha: dict, user_name: str, refresh_list):
    """Potvrzení smazání RUČNÍHO bodu obnovy (automatické zálohy smazat nelze).
    Po smazání se jen překreslí seznam záloh v otevřeném dialogu (živá data zůstávají)."""
    cas = zaloha.get('vytvoreno')
    cas_txt = cas.strftime('%d.%m.%Y %H:%M') if hasattr(cas, 'strftime') else str(cas or '')
    with ui.dialog() as dlg, ui.card().classes('p-6 rounded-2xl gap-3') \
            .style('min-width: 420px; max-width: 540px'):
        with ui.row().classes('items-center gap-2'):
            ui.icon('delete_forever', color='red').classes('text-2xl')
            ui.label('Smazat ruční bod obnovy?').classes('text-xl font-bold text-gray-800')
        ui.label(f'Ruční bod obnovy z {cas_txt} '
                 f'({zaloha.get("pocet_radku") or 0} řádků) bude nevratně odstraněn.') \
            .classes('text-sm text-gray-700 font-medium')
        ui.label('Smaže se jen tato záloha — aktuální data v sestavě ani ostatní zálohy '
                 'se nijak nezmění.').classes('text-sm text-gray-500')

        async def _potvrd():
            ok, err = await asyncio.to_thread(_smaz_zalohu, zaloha.get('id'))
            if err:
                ui.notify(f'Mazání se nezdařilo: {err}',
                          type='negative', position='top', timeout=8000)
                return
            if not ok:
                ui.notify('Bod obnovy se nepodařilo najít (možná už byl smazán).',
                          type='warning', position='top', timeout=5000)
            else:
                intranet_logger.log_activity(
                    user_name, 'Sankce',
                    f'Smazání ručního bodu obnovy (k vystavení) #{zaloha.get("id")}')
                ui.notify('Ruční bod obnovy byl smazán.',
                          type='positive', position='top-right', timeout=4000)
            dlg.close()
            refresh_list()

        with ui.row().classes('w-full justify-end mt-2 gap-2'):
            ui.button('Zrušit', on_click=dlg.close).props('flat no-caps')
            ui.button('Smazat', icon='delete_forever', on_click=_potvrd) \
                .props('unelevated no-caps') \
                .classes('bg-red-600 hover:bg-red-700 text-white font-semibold '
                         'rounded-lg shadow-md px-5')
    dlg.open()


def _potvrd_hromadne_smazani_dialog(ids: list, user_name: str, refresh_list, vybrane: set):
    """Potvrzení HROMADNÉHO smazání záloh (jen hlavní administrátor) — smaže i
    automatické zálohy. Po smazání vyčistí výběr a překreslí seznam záloh."""
    n = len(ids)
    with ui.dialog() as dlg, ui.card().classes('p-6 rounded-2xl gap-3') \
            .style('min-width: 420px; max-width: 540px'):
        with ui.row().classes('items-center gap-2'):
            ui.icon('delete_sweep', color='red').classes('text-2xl')
            ui.label('Smazat vybrané zálohy?').classes('text-xl font-bold text-gray-800')
        ui.label(f'Bude nevratně odstraněno označených záloh: {n} (včetně automatických). '
                 'Aktuální data v sestavě se nijak nezmění.') \
            .classes('text-sm text-gray-700 font-medium')

        async def _potvrd():
            pocet, err = await asyncio.to_thread(_smaz_zalohy_hromadne, ids, True)
            if err:
                ui.notify(f'Mazání se nezdařilo: {err}',
                          type='negative', position='top', timeout=8000)
                return
            intranet_logger.log_activity(
                user_name, 'Sankce',
                f'Hromadné smazání záloh (k vystavení) — {pocet} ks')
            ui.notify(f'Smazáno {pocet} záloh.',
                      type='positive', position='top-right', timeout=4000)
            vybrane.clear()
            dlg.close()
            refresh_list()

        with ui.row().classes('w-full justify-end mt-2 gap-2'):
            ui.button('Zrušit', on_click=dlg.close).props('flat no-caps')
            ui.button('Smazat', icon='delete_sweep', on_click=_potvrd) \
                .props('unelevated no-caps') \
                .classes('bg-red-600 hover:bg-red-700 text-white font-semibold '
                         'rounded-lg shadow-md px-5')
    dlg.open()


def _potvrd_obnova_dialog(zaloha: dict, user_name: str, refresh_fn, close_parent=None):
    """Potvrzení „Opravdu přehrát data?" — obnova přepíše celý přehled.
    Po potvrzení („Ano") se data přehrají, zavřou se oba dialogy a sestava se obnoví."""
    cas = zaloha.get('vytvoreno')
    cas_txt = cas.strftime('%d.%m.%Y %H:%M') if hasattr(cas, 'strftime') else str(cas or '')
    typ_lbl = _ZALOHA_TYP_LABEL.get(zaloha.get('typ'), zaloha.get('typ') or '')
    with ui.dialog() as dlg, ui.card().classes('p-6 rounded-2xl gap-3') \
            .style('min-width: 440px; max-width: 580px'):
        with ui.row().classes('items-center gap-2'):
            ui.icon('restore', color='amber-8').classes('text-2xl')
            ui.label('Opravdu přehrát data?').classes('text-xl font-bold text-gray-800')
        ui.label(f'Záloha: {typ_lbl} z {cas_txt} '
                 f'({zaloha.get("pocet_radku") or 0} řádků).') \
            .classes('text-sm text-gray-700 font-medium')
        ui.label('Aktuální obsah sestavy „Sankce k vystavení" bude nahrazen daty z této '
                 'zálohy. Pro jistotu se současný stav před přehráním sám zazálohuje '
                 '(automatická záloha), takže obnovu lze vrátit zpět.') \
            .classes('text-sm text-gray-500')

        async def _potvrd():
            ok, cnt, err = await asyncio.to_thread(_obnov_zalohu, zaloha.get('id'), user_name)
            if err:
                ui.notify(f'Obnova se nezdařila: {err}',
                          type='negative', position='top', timeout=9000)
                return
            intranet_logger.log_activity(
                user_name, 'Sankce',
                f'Obnova dat (k vystavení) ze zálohy #{zaloha.get("id")} — {cnt} řádků')
            ui.notify(f'Data byla obnovena ({cnt} řádků).',
                      type='positive', position='top-right', timeout=5000)
            dlg.close()
            if close_parent:
                close_parent()
            refresh_fn()

        with ui.row().classes('w-full justify-end mt-2 gap-2'):
            ui.button('Zrušit', on_click=dlg.close).props('flat no-caps')
            ui.button('Ano, přehrát', icon='restore', on_click=_potvrd) \
                .props('unelevated no-caps') \
                .classes('bg-amber-600 hover:bg-amber-700 text-white font-semibold '
                         'rounded-lg shadow-md px-5')
    dlg.open()


def _obnova_button(user_name: str, refresh_fn, text: str = None, je_hlavni_admin: bool = False):
    """Tlačítko pro otevření dialogu obnovení dat ze zálohy. Bez textu = kompaktní
    ikonka (do lišty); s textem = popisné tlačítko (do prázdného stavu, kde je obnova
    hlavní akcí). Dostupné účetní i analytikovi; hlavní administrátor (`je_hlavni_admin`)
    navíc může zálohy hromadně mazat."""
    b = ui.button(text or '', icon='restore',
                  on_click=lambda: _otevri_obnova_dialog(user_name, refresh_fn, je_hlavni_admin))
    if text:
        b.props('outline color=teal no-caps').classes('font-semibold')
    else:
        b.props('outline color=teal dense')
    b.tooltip('Obnovit starší data ze zálohy nebo vytvořit ruční bod obnovení. '
              'Přehled se každé 4 hodiny automaticky zálohuje.')
    return b


def _otevri_obnova_dialog(user_name: str, refresh_fn, je_hlavni_admin: bool = False):
    """Dialog „Obnovení dat" pro Sankce k vystavení: seznam záloh (automatické +
    ruční body obnovy) s možností přehrát data a tlačítkem pro ruční bod obnovy.
    Hlavní administrátor navíc může označit více záloh (i automatických) a smazat
    je najednou. Zálohy nejdou stahovat — slouží jen k obnovení."""
    vybrane = set()            # id záloh označených k hromadnému smazání (hlavní admin)
    stav_ui = {'btn_smaz': None}

    def _aktualizuj_smaz_btn():
        b = stav_ui['btn_smaz']
        if not b:
            return
        n = len(vybrane)
        b.set_text(f'Smazat vybrané ({n})')
        b.set_enabled(n > 0)

    with ui.dialog() as dlg, ui.card().classes('p-5 gap-3 rounded-2xl') \
            .style('min-width: 640px; max-width: 880px; max-height: 84vh; overflow-y: auto'):
        with ui.row().classes('items-center gap-2 w-full'):
            ui.icon('restore', color='teal').classes('text-2xl')
            ui.label('Obnovení dat — Sankce k vystavení') \
                .classes('text-xl font-bold text-gray-800')
            ui.space()
            ui.button(icon='close', on_click=dlg.close) \
                .props('flat round dense color=grey-7')
        ui.label('Přehled se každé 4 hodiny automaticky zálohuje (drží se posledních '
                 f'{_ZALOHA_AUTO_MAX} záloh, nejstarší se přepisuje). Kliknutím na '
                 '„Obnovit" přehrajete data z vybrané zálohy (nutné potvrzení). Zálohy '
                 'nelze stahovat, slouží jen k obnovení.') \
            .classes('text-xs text-gray-500 -mt-1')

        @ui.refreshable
        def _seznam():
            zalohy = _nacti_zalohy()
            # z výběru zahoď id, která už neexistují (např. po prořezání auto záloh)
            vybrane.intersection_update({z.get('id') for z in zalohy})
            _aktualizuj_smaz_btn()
            if not zalohy:
                with ui.column().classes('items-center py-10 gap-2 w-full'):
                    ui.icon('history', size='3rem', color='grey-4')
                    ui.label('Zatím není k dispozici žádná záloha.') \
                        .classes('text-sm text-gray-400')
                return
            with ui.column().classes('w-full gap-0'):
                with ui.row().classes('w-full items-center text-xs font-bold '
                                      'text-gray-500 px-2 pb-1'):
                    if je_hlavni_admin:
                        ui.label('').classes('w-8')
                    ui.label('Datum a čas').classes('w-40')
                    ui.label('Typ').classes('w-44')
                    ui.label('Vytvořil').classes('flex-1')
                    ui.label('Řádků').classes('w-16 text-right')
                    ui.label('').classes('w-36')
                for z in zalohy:
                    cas = z.get('vytvoreno')
                    cas_txt = cas.strftime('%d.%m.%Y %H:%M') if hasattr(cas, 'strftime') \
                        else str(cas or '')
                    je_rucni = z.get('typ') == 'rucni'
                    typ_lbl = _ZALOHA_TYP_LABEL.get(z.get('typ'), z.get('typ') or '')
                    with ui.row().classes('w-full items-center px-2 py-1 '
                                          'hover:bg-gray-50 border-b border-gray-100'):
                        if je_hlavni_admin:
                            with ui.element('div').classes('w-8'):
                                cb = ui.checkbox(value=(z.get('id') in vybrane)).props('dense')

                                def _toggle(e, _id=z.get('id')):
                                    if e.value:
                                        vybrane.add(_id)
                                    else:
                                        vybrane.discard(_id)
                                    _aktualizuj_smaz_btn()
                                cb.on_value_change(_toggle)
                        ui.label(cas_txt).classes('w-40 text-sm text-gray-700 font-medium')
                        with ui.element('div').classes('w-44'):
                            ui.badge(typ_lbl) \
                                .props(f'color={"teal" if je_rucni else "blue-grey"}') \
                                .classes('text-xs px-2 py-1')
                        ui.label(z.get('vytvoril') or '—').classes('flex-1 text-sm text-gray-600')
                        ui.label(str(z.get('pocet_radku') or 0)) \
                            .classes('w-16 text-right text-sm text-gray-600')
                        with ui.row().classes('w-36 justify-end items-center gap-1 flex-nowrap'):
                            if je_rucni:
                                ui.button(icon='delete_outline',
                                          on_click=lambda _z=z: _potvrd_smaz_zaloha_dialog(
                                              _z, user_name, _seznam.refresh)) \
                                    .props('flat round dense color=red') \
                                    .tooltip('Smazat tento ruční bod obnovy')
                            ui.button('Obnovit', icon='restore',
                                      on_click=lambda _z=z: _potvrd_obnova_dialog(
                                          _z, user_name, refresh_fn, dlg.close)) \
                                .props('outline color=teal dense no-caps')
        _seznam()

        async def _novy_bod():
            _id, cnt, err = await asyncio.to_thread(_vytvor_zalohu, 'rucni', user_name, False)
            if err:
                ui.notify(f'Bod obnovení se nepodařilo vytvořit: {err}',
                          type='negative', position='top', timeout=8000)
                return
            intranet_logger.log_activity(user_name, 'Sankce',
                                         f'Ruční bod obnovy (k vystavení) — {cnt} řádků')
            ui.notify(f'Ruční bod obnovy vytvořen ({cnt} řádků).',
                      type='positive', position='top-right', timeout=4000)
            _seznam.refresh()

        def _smaz_vybrane():
            ids = sorted(vybrane)
            if not ids:
                return
            _potvrd_hromadne_smazani_dialog(ids, user_name, _seznam.refresh, vybrane)

        with ui.row().classes('w-full justify-between items-center mt-2'):
            with ui.row().classes('items-center gap-2'):
                ui.button('Vytvořit bod obnovení', icon='add_circle', on_click=_novy_bod) \
                    .props('unelevated no-caps') \
                    .classes('bg-teal-600 hover:bg-teal-700 text-white font-semibold '
                             'rounded-lg shadow-md px-5')
                if je_hlavni_admin:
                    stav_ui['btn_smaz'] = ui.button(
                        'Smazat vybrané (0)', icon='delete_sweep', on_click=_smaz_vybrane) \
                        .props('outline color=red no-caps') \
                        .tooltip('Smaže všechny zaškrtnuté zálohy najednou '
                                 '(včetně automatických). Pouze hlavní administrátor.')
                    _aktualizuj_smaz_btn()
            ui.button('Zavřít', on_click=dlg.close).props('flat no-caps')
    dlg.open()


# =========================================================
# IMPORTNÍ PANEL (jen analytik)
# =========================================================
def _otevri_import_dialog(tabulka: str, mapa: dict, cisla: set, user_name: str, refresh_fn,
                          pripoj_mode: bool = False):
    """Dialog pro nahrání listu DATA. Dva režimy:
      • pripoj_mode=False („Nahrát data") — zadá se období (OD/DO) a dávka období se
        NAHRADÍ (ručně zadané stavy/poznámky se zachovají přes row_hash);
      • pripoj_mode=True („Přidat k datům") — vybere se EXISTUJÍCÍ období (nebo nové) a
        jen se DOPLNÍ nové řádky ve stavu „Rozpracováno" (stará data se nepřepíšou).

    Bytes vybraného souboru si držíme my (`auto_upload` proběhne jen jednou), takže když
    chybí období, soubor se neztratí: stačí ho doplnit a kliknout na „Importovat"."""
    drzeny = {'raw': None, 'name': ''}   # poslední načtený soubor čekající na období

    with ui.dialog() as dlg, ui.card().classes('p-6 rounded-2xl gap-3') \
            .style('min-width: 480px; max-width: 560px'):
        with ui.row().classes('items-center gap-2'):
            ui.icon('group_add' if pripoj_mode else 'upload_file', color='primary').classes('text-2xl')
            ui.label('Přidat k existujícím datům' if pripoj_mode else 'Nahrát data (list DATA)') \
                .classes('text-xl font-bold text-gray-800')
        if pripoj_mode:
            ui.label('Vyberte období, ke kterému data patří (nebo vytvořte nové). Doplní se '
                     'jen NOVÉ řádky (noví dodavatelé / nové objednávky) ve stavu „Rozpracováno"; '
                     'stará data se nepřepíšou a co už v období je, se přeskočí.') \
                .classes('text-xs text-gray-500 -mt-1')
        else:
            ui.label('Zadejte období, za které data jsou — podle něj se pozná, co je nové. '
                     'Re-import stejného období dávku nahradí, ručně zadané poznámky' +
                     (' a stavy' if tabulka == 'sankce_vystaveni' else '') +
                     ' zůstanou zachovány.') \
                .classes('text-xs text-gray-500 -mt-1')

        # Období: v režimu PŘIDÁNÍ se vybírá z existujících (+ „Nové období"); v režimu
        # NAHRÁNÍ se zadává klasicky přes OD/DO.
        sel_obdobi = None
        NOVE_OBD = '__nove__'
        obd_mapa = {}
        if pripoj_mode:
            _obd_detail = _seznam_obdobi_detail(tabulka)
            obd_mapa = {o['obdobi']: (o['od'], o['do']) for o in _obd_detail}
            obd_options = {NOVE_OBD: '➕ Nové období (zadat datum)'}
            for _o in _obd_detail:
                obd_options[_o['obdobi']] = _o['obdobi']
            sel_obdobi = ui.select(
                obd_options,
                value=(_obd_detail[0]['obdobi'] if _obd_detail else NOVE_OBD),
                label='Období') \
                .props('outlined dense options-dense').classes('w-full')

        with ui.row().classes('items-end gap-3 w-full') as row_datum:
            inp_od = ui.input('Období OD').props('type=date outlined dense').classes('flex-1')
            inp_do = ui.input('Období DO').props('type=date outlined dense').classes('flex-1')
        hint_datum = ui.label('Necháte-li prázdné, zkusím období načíst z názvu souboru.') \
            .classes('text-xs text-gray-400 -mt-1')

        # Dropdown řídí, jestli jsou datová pole vidět (jen v režimu přidání).
        if sel_obdobi is not None:
            def _on_obdobi_change(e):
                je_nove = (e.value == NOVE_OBD)
                row_datum.set_visibility(je_nove)
                hint_datum.set_visibility(je_nove)
            sel_obdobi.on_value_change(_on_obdobi_change)
            _je_nove0 = (sel_obdobi.value == NOVE_OBD)
            row_datum.set_visibility(_je_nove0)
            hint_datum.set_visibility(_je_nove0)

        async def _zpracuj(raw: bytes, name: str):
            """Zjistí období (z výběru u přidání, jinak z polí/názvu souboru) a naimportuje
            držené bytes. Když chybí období, soubor podrží a vyzve k jeho doplnění."""
            if sel_obdobi is not None and sel_obdobi.value and sel_obdobi.value != NOVE_OBD:
                # přidání k už existujícímu období — datum vezmeme z něj
                obdobi = sel_obdobi.value
                od_iso, do_iso = obd_mapa.get(obdobi, ('', ''))
            else:
                # nové období — datum z polí, jinak zkus z názvu souboru
                od_iso = (inp_od.value or '').strip()
                do_iso = (inp_do.value or '').strip()
                if not (od_iso and do_iso):
                    p_od, p_do = _parse_obdobi_z_nazvu(name)
                    od_iso = od_iso or p_od
                    do_iso = do_iso or p_do
                if not (od_iso and do_iso):
                    drzeny['raw'] = raw
                    drzeny['name'] = name
                    _vyzva = ('Vyberte období, nebo doplňte datum OD i DO'
                              if pripoj_mode else 'Doplňte období OD i DO')
                    stav_lbl.set_text(f'Soubor „{name}" je načten. {_vyzva} '
                                      'a klikněte na „Importovat".')
                    stav_lbl.set_visibility(True)
                    btn_import.set_visibility(True)
                    ui.notify(f'{_vyzva} a klikněte na „Importovat" '
                              '(nepodařilo se ho odvodit z názvu souboru).',
                              type='warning', timeout=7000)
                    return
                obdobi = _obdobi_label(od_iso, do_iso)

            pripoj = pripoj_mode
            count, skipped, err = await asyncio.to_thread(
                _importuj_sync, raw, tabulka, mapa, cisla, obdobi, od_iso, do_iso, user_name, pripoj)
            if err:
                ui.notify(f'Import se nezdařil: {err}', type='negative', timeout=10000)
                return

            if pripoj and count == 0:
                # nic nového — podrž soubor, ať lze upravit období a zkusit znovu
                drzeny['raw'] = raw
                drzeny['name'] = name
                stav_lbl.set_text(f'Soubor „{name}": žádný nový řádek — vše už v období '
                                  f'{obdobi} je (přeskočeno {skipped}). Zkontrolujte období.')
                stav_lbl.set_visibility(True)
                btn_import.set_visibility(True)
                ui.notify('Nepřidán žádný nový řádek — vše z tohoto souboru už v období '
                          f'{obdobi} je (přeskočeno {skipped}).',
                          type='warning', position='top', timeout=8000)
                return

            # Zamítnuté: automaticky nastav nově naimportované období jako zamčené
            if tabulka == 'sankce_zamitnute':
                _uloz_zamcene_obdobi(obdobi)

            if pripoj:
                zprava = (f'Přidáno {count} nových řádků (stav „Rozpracováno") k období {obdobi}.')
                if skipped:
                    zprava += f' Přeskočeno {skipped} již existujících.'
                intranet_logger.log_activity(
                    user_name, 'Sankce',
                    f'Doplnění dodavatelů (k vystavení) – období {obdobi}: '
                    f'+{count} řádků (přeskočeno {skipped})')
            else:
                zprava = f'Import dokončen — období {obdobi}, načteno {count} řádků.'
                intranet_logger.log_activity(
                    user_name, 'Sankce',
                    f'Import {"Sankce k vystavení" if tabulka=="sankce_vystaveni" else "Zamítnuté dodávky"} '
                    f'– období {obdobi}: {count} řádků')
            ui.notify(zprava, type='positive', position='top-right', timeout=6000)
            drzeny['raw'] = None
            dlg.close()
            refresh_fn()

        async def _on_upload(e):
            zdroj = None
            for attr in ('content', 'file', 'stream', 'data', 'file_obj'):
                val = getattr(e, attr, None)
                if val is not None and hasattr(val, 'read'):
                    zdroj = val
                    break
            if zdroj is None:
                ui.notify('Nepodařilo se načíst obsah souboru.', type='negative')
                up.reset()
                return
            try:
                raw = zdroj.read()
                if inspect.isawaitable(raw):
                    raw = await raw
            except Exception as exc:
                ui.notify(f'Chyba čtení souboru: {exc}', type='negative')
                up.reset()
                return
            # Bytes si držíme sami; widget hned uvolníme, ať jde příště nahrát znovu.
            up.reset()
            await _zpracuj(raw, getattr(e, 'name', '') or drzeny['name'])

        async def _on_import_click():
            if not drzeny['raw']:
                ui.notify('Nejdřív vyberte soubor .xlsx.', type='warning')
                return
            await _zpracuj(drzeny['raw'], drzeny['name'])

        up = ui.upload(on_upload=_on_upload, auto_upload=True, max_file_size=50_000_000,
                       label='Vybrat .xlsx soubor').props('accept=.xlsx').classes('w-full')

        stav_lbl = ui.label('').classes('text-xs text-amber-700 font-medium')
        stav_lbl.set_visibility(False)

        with ui.row().classes('w-full justify-end mt-1 gap-2'):
            btn_import = ui.button('Importovat', icon='cloud_upload', on_click=_on_import_click) \
                .props('unelevated no-caps') \
                .classes('bg-blue-600 hover:bg-blue-700 text-white font-semibold rounded-lg shadow-md px-5')
            btn_import.set_visibility(False)
            ui.button('Zavřít', on_click=dlg.close).props('flat no-caps')
    dlg.open()


def _import_button(tabulka: str, mapa: dict, cisla: set, user_name: str, refresh_fn):
    """Elegantní tlačítko „Nahrát data" (vpravo nahoře) — otevře import dialog
    v režimu nahrání (nahradí dávku období)."""
    ui.button('Nahrát data', icon='upload_file',
              on_click=lambda: _otevri_import_dialog(tabulka, mapa, cisla, user_name, refresh_fn)) \
        .props('unelevated no-caps') \
        .classes('bg-blue-600 hover:bg-blue-700 text-white font-semibold rounded-lg shadow-md px-5')


def _pripoj_button(tabulka: str, mapa: dict, cisla: set, user_name: str, refresh_fn):
    """Tlačítko „Přidat k datům" (vedle „Nahrát data") — import dalších dodavatelů
    k existujícímu období; nové řádky spadnou do stavu „Rozpracováno"."""
    ui.button('Přidat k datům', icon='group_add',
              on_click=lambda: _otevri_import_dialog(tabulka, mapa, cisla, user_name,
                                                     refresh_fn, pripoj_mode=True)) \
        .props('unelevated no-caps') \
        .classes('bg-teal-600 hover:bg-teal-700 text-white font-semibold rounded-lg shadow-md px-5') \
        .tooltip('Doplní k vybranému období jen nové dodavatele/řádky ve stavu „Rozpracováno". '
                 'Stávající data se nepřepíšou.')


# =========================================================
# MAZÁNÍ DAT (jen analytik) — výběr období / vše, s potvrzením
# =========================================================
def _otevri_smazat_dialog(tabulka: str, nazev: str, user_name: str, refresh_fn):
    """Dialog před nevratným smazáním dat sestavy — lze zvolit konkrétní období,
    nebo všechna data. Výběr + tlačítko „Smazat" je zároveň potvrzení."""
    obdobi_list = _seznam_obdobi(tabulka)
    ma_stavy = (tabulka == 'sankce_vystaveni')
    VSE = '(všechna období)'

    with ui.dialog() as dlg, ui.card().classes('p-6 rounded-2xl gap-3') \
            .style('min-width: 460px; max-width: 560px'):
        with ui.row().classes('items-center gap-2'):
            ui.icon('warning', color='red').classes('text-2xl')
            ui.label('Smazat data sestavy').classes('text-xl font-bold text-gray-800')
        ui.label(f'Sestava „{nazev}". Vyberte, co se má nevratně smazat — celá sestava, '
                 'nebo jen jedno období. Smaže se i historie změn (👁)' +
                 (' a ručně zadané stavy/poznámky.' if ma_stavy else ' a ručně zadané poznámky.')) \
            .classes('text-sm text-gray-600')

        sel = ui.select([VSE] + obdobi_list, value=VSE, label='Co smazat') \
            .props('outlined dense options-dense').classes('w-full')

        warn = ui.label('').classes('text-sm font-medium text-red-700')

        def _refresh_warn():
            if sel.value == VSE:
                warn.set_text('⚠ Smaže VŠECHNA data sestavy (všechna období).')
            else:
                warn.set_text(f'⚠ Smaže pouze období „{sel.value}".')
        sel.on_value_change(lambda e: _refresh_warn())
        _refresh_warn()

        async def _potvrd():
            obd = None if sel.value == VSE else sel.value
            pocet, err = await asyncio.to_thread(_smaz_data, tabulka, obd)
            if err:
                ui.notify(f'Mazání se nezdařilo: {err}', type='negative', timeout=10000)
                return
            # Zamítnuté: zruš globálně zamčené období, smazali-li jsme vše nebo právě jeho
            if tabulka == 'sankce_zamitnute' and (obd is None or _ziskej_zamcene_obdobi() == obd):
                _uloz_zamcene_obdobi('')
            kde = 'všechna data' if obd is None else f'období {obd}'
            intranet_logger.log_activity(user_name, 'Sankce',
                                         f'Smazání ({nazev}) – {kde}: {pocet} řádků')
            ui.notify(f'Smazáno {pocet} řádků sestavy „{nazev}" ({kde}).',
                      type='positive', position='top-right', timeout=6000)
            dlg.close()
            refresh_fn()

        with ui.row().classes('w-full justify-end mt-2 gap-2'):
            ui.button('Zrušit', on_click=dlg.close).props('flat no-caps')
            ui.button('Smazat', icon='delete_forever', on_click=_potvrd) \
                .props('unelevated no-caps') \
                .classes('bg-red-600 hover:bg-red-700 text-white font-semibold rounded-lg shadow-md px-5')
    dlg.open()


def _smazat_button(tabulka: str, nazev: str, user_name: str, refresh_fn):
    """Tlačítko „Vymazat data" (vpravo nahoře) — otevře dialog s výběrem období/vše."""
    ui.button('Vymazat data', icon='delete_forever',
              on_click=lambda: _otevri_smazat_dialog(tabulka, nazev, user_name, refresh_fn)) \
        .props('outline color=red dense no-caps') \
        .tooltip('Smaže vybrané období nebo všechna naimportovaná data této sestavy')


# =========================================================
# MAZÁNÍ JEDNOTLIVÉHO ŘÁDKU (Delete / kontextové menu) — s potvrzením
# =========================================================
def _otevri_smazat_radky_dialog(tabulka: str, radek_ids: list, popis: str,
                                user_name: str, refresh_fn):
    """Potvrzovací dialog před smazáním jednoho NEBO více řádků (klávesa Delete
    nebo pravým tlačítkem → „Smazat tento řádek" / „Smazat označené")."""
    nazev = 'Sankce k vystavení' if tabulka == 'sankce_vystaveni' else 'Zamítnuté dodávky'
    n = len(radek_ids)
    je_vic = n > 1
    with ui.dialog() as dlg, ui.card().classes('p-6 rounded-2xl gap-3') \
            .style('min-width: 420px; max-width: 520px'):
        with ui.row().classes('items-center gap-2'):
            ui.icon('delete_forever', color='red').classes('text-2xl')
            ui.label('Opravdu smazat označené řádky?' if je_vic else 'Opravdu smazat tento řádek?') \
                .classes('text-xl font-bold text-gray-800')
        if je_vic:
            ui.label(f'Bude nevratně odstraněno {n} řádků ze sestavy „{nazev}" '
                     'včetně jejich historie změn.').classes('text-sm text-gray-700 font-medium')
        else:
            if popis.strip(' –'):
                ui.label(popis).classes('text-sm text-gray-700 font-medium')
            ui.label(f'Řádek bude nevratně odstraněn ze sestavy „{nazev}" včetně své historie změn.') \
                .classes('text-sm text-gray-500')

        async def _potvrd():
            pocet, err = await asyncio.to_thread(_smaz_radky, tabulka, radek_ids)
            if err:
                ui.notify(f'Mazání se nezdařilo: {err}', type='negative', timeout=8000)
                return
            if not pocet:
                ui.notify('Řádky se nepodařilo najít (možná už byly smazány).', type='warning')
                dlg.close()
                refresh_fn()
                return
            intranet_logger.log_activity(
                user_name, 'Sankce',
                (f'Smazání {pocet} řádků ({nazev})' if pocet > 1
                 else f'Smazání řádku ({nazev}) #{radek_ids[0]}'))
            ui.notify(f'Smazáno {pocet} řádků.' if pocet > 1 else 'Řádek byl smazán.',
                      type='positive', position='top-right', timeout=4000)
            dlg.close()
            refresh_fn()

        with ui.row().classes('w-full justify-end mt-2 gap-2'):
            ui.button('Zrušit', on_click=dlg.close).props('flat no-caps')
            ui.button('Smazat označené' if je_vic else 'Smazat', icon='delete_forever',
                      on_click=_potvrd) \
                .props('unelevated no-caps') \
                .classes('bg-red-600 hover:bg-red-700 text-white font-semibold rounded-lg shadow-md px-5')
    dlg.open()


def _grid_mazani_js(tabulka: str) -> dict:
    """Vrací grid-options callbacky (`:onCellKeyDown` + `:onCellContextMenu`),
    které pošlou do Pythonu event 'sankce_radek_del' s polem `ids`. Delete smaže
    OZNAČENÉ řádky (zaškrtnuté vlevo), jinak zaměřený řádek; pravé tlačítko nabídne
    „Smazat označené (N)" (jsou-li nějaké) a „Smazat tento řádek". Připnutý součtový
    řádek „Celkem" (bez id) se ignoruje."""
    t = json.dumps(tabulka)  # bezpečný JS literál názvu tabulky
    # payload pro 1 řádek (s popisem) a pro N označených řádků
    payload1 = ("function(d){return{tabulka:" + t + ",ids:[d.id],"
                "popis:((d.jmeno_dodavatele||'')+(d.nazev_zbozi?(' – '+d.nazev_zbozi):''))};}")
    sel_ids = ("function(api){var s=(api&&api.getSelectedRows)?api.getSelectedRows():[];"
               "return s.filter(function(r){return r&&r.id;}).map(function(r){return r.id;});}")
    on_key = (
        "function(p){"
        "var e=p.event;if(!e)return;"
        "if(e.key!=='Delete'&&e.key!=='Del')return;"
        "if(p.node&&p.node.rowPinned)return;"
        "if(p.api&&p.api.getEditingCells&&p.api.getEditingCells().length)return;"
        "var ids=(" + sel_ids + ")(p.api);"
        "if(ids.length>0){e.preventDefault();"
        "emitEvent('sankce_radek_del',{tabulka:" + t + ",ids:ids,popis:''});return;}"
        "var d=p.data||{};if(!d.id)return;"
        "e.preventDefault();"
        "emitEvent('sankce_radek_del',(" + payload1 + ")(d));"
        "}"
    )
    on_ctx = (
        "function(p){"
        "if(p.node&&p.node.rowPinned)return;"
        "var d=p.data||{};if(!d.id)return;"
        "var ev=p.event;"  # nativní menu potlačí grid-option preventDefaultOnContextMenu
        "var ids=(" + sel_ids + ")(p.api);"
        "var old=document.getElementById('sankce-ctx-menu');if(old)old.remove();"
        "var x=ev?ev.clientX:0,y=ev?ev.clientY:0;"
        "var m=document.createElement('div');m.id='sankce-ctx-menu';"
        "m.style.cssText='position:fixed;z-index:99999;background:#fff;border:1px solid #e5e7eb;"
        "box-shadow:0 10px 28px rgba(0,0,0,.18);border-radius:10px;padding:4px;min-width:200px;';"
        "m.style.left=x+'px';m.style.top=y+'px';"
        "function zavri(){m.remove();document.removeEventListener('mousedown',ven,true);"
        "document.removeEventListener('keydown',esc,true);}"
        "function ven(e2){if(!m.contains(e2.target))zavri();}"
        "function esc(e2){if(e2.key==='Escape')zavri();}"
        "function pridej(text,payload){"
        "var it=document.createElement('div');it.textContent=text;"
        "it.style.cssText='padding:8px 12px;cursor:pointer;border-radius:7px;color:#b91c1c;"
        "font-weight:600;font-size:14px;white-space:nowrap;';"
        "it.onmouseenter=function(){it.style.background='#fee2e2';};"
        "it.onmouseleave=function(){it.style.background='transparent';};"
        "it.onclick=function(){emitEvent('sankce_radek_del',payload);zavri();};"
        "m.appendChild(it);}"
        "if(ids.length>0){pridej('\U0001f5d1  Smazat označené ('+ids.length+')',"
        "{tabulka:" + t + ",ids:ids,popis:''});}"
        "pridej('\U0001f5d1  Smazat tento řádek',(" + payload1 + ")(d));"
        "document.body.appendChild(m);"
        "var r=m.getBoundingClientRect();"
        "if(r.right>window.innerWidth)m.style.left=(window.innerWidth-r.width-6)+'px';"
        "if(r.bottom>window.innerHeight)m.style.top=(window.innerHeight-r.height-6)+'px';"
        "setTimeout(function(){document.addEventListener('mousedown',ven,true);"
        "document.addEventListener('keydown',esc,true);},0);"
        "}"
    )
    # preventDefaultOnContextMenu = AG Grid sám potlačí nativní menu prohlížeče
    # (ruční preventDefault v callbacku nestačí – jiný/pasivní listener).
    return {'preventDefaultOnContextMenu': True,
            ':onCellKeyDown': on_key, ':onCellContextMenu': on_ctx}


def _zaregistruj_mazani_radku(user_name, vsechna_prava):
    """Jednorázově (na klienta) zaregistruje obsluhu mazání jednotlivých řádků
    spouštěnou z gridu (event 'sankce_radek_del' z injectnutého JS). Mazat smí
    jen analytik / 'vse'; sám handler oprávnění ještě jednou ověří."""
    try:
        if app.storage.client.get('_sankce_radek_del_on'):
            return
        app.storage.client['_sankce_radek_del_on'] = True
    except Exception:
        # Mimo klientský kontext (např. detached background task) NEregistrujeme —
        # jinak by ui.on() navěsil duplicitní listener na persistentní layout a
        # klient by hlásil „Event listeners changed after initial definition".
        return

    ma_vse = 'vse' in vsechna_prava
    je_analytik = ma_vse or 'sankce_analytik' in vsechna_prava

    def _on_del(e):
        a = e.args
        if isinstance(a, (list, tuple)):
            a = a[0] if a else None
        if not isinstance(a, dict):
            return
        if not je_analytik:
            ui.notify('Nemáte oprávnění mazat data.', type='warning')
            return
        tabulka = a.get('tabulka')
        ids = a.get('ids')
        if ids is None and a.get('id') is not None:
            ids = [a.get('id')]
        cisla = []
        for x in (ids or []):
            try:
                cisla.append(int(x))
            except (TypeError, ValueError):
                pass
        if tabulka not in ('sankce_zamitnute', 'sankce_vystaveni') or not cisla:
            return
        refresh_fn = (_vykresli_zamitnute.refresh if tabulka == 'sankce_zamitnute'
                      else _vykresli_vystaveni.refresh)
        _otevri_smazat_radky_dialog(tabulka, cisla, a.get('popis') or '', user_name, refresh_fn)

    ui.on('sankce_radek_del', _on_del)


def _col_poradi() -> dict:
    """Levý připnutý sloupec „Poř. č." — TRVALÉ naše pořadové číslo (rok+pořadí,
    např. 202600001), přidělené při importu a uložené v DB (sloupec nase_cislo).
    Slouží i jako variabilní symbol na výzvě k vystavení sankce."""
    return {'headerName': 'Poř. č.', 'field': 'nase_cislo', 'colId': '_poradi',
            'width': 104, 'minWidth': 96, 'maxWidth': 130, 'pinned': 'left',
            'sortable': True, 'editable': False, 'resizable': False, 'filter': False,
            'suppressSizeToFit': True, 'suppressAutoSize': True,
            'cellStyle': {'fontFamily': 'monospace', 'fontSize': '12px',
                          'color': '#334155', 'fontWeight': '600', 'textAlign': 'center'},
            'headerTooltip': 'Naše pořadové číslo (trvalé; variabilní symbol)'}


def _col_chat() -> dict:
    """Levý připnutý sloupec „💬" — indikátor diskuze k případu (klik otevře chat)."""
    return {'headerName': '', 'field': '_chat', 'width': 50, 'minWidth': 50, 'maxWidth': 56,
            'pinned': 'left', 'sortable': False, 'editable': False, 'resizable': False,
            'filter': False, 'suppressSizeToFit': True, 'suppressAutoSize': True,
            ':cellRenderer': _CHAT_RENDERER,
            'cellStyle': {'textAlign': 'center', 'cursor': 'pointer', 'padding': '0'},
            'headerTooltip': 'Diskuze k případu (chat)'}


# =========================================================
# DIALOG CHATU (diskuze k případu)
# =========================================================
def _otevri_chat(tabulka, row_hash, popis, user_id, user_name, on_badge,
                 muze_mazat_vse=False, prava_navic=()):
    """Okno diskuze k jednomu případu. `on_badge()` překreslí indikátor v gridu.
    Vlákno se otevřením označí jako přečtené; dokud je okno otevřené, nové zprávy
    se průběžně dotahují (~5 s). `muze_mazat_vse` = smí mazat i cizí zprávy."""
    _oznac_precteno(tabulka, row_hash, user_id)
    on_badge()
    stav = {'max_id': _chat_max_id(tabulka, row_hash)}

    def _smaz(z):
        async def _potvrd():
            with ui.dialog() as d, ui.card():
                ui.label('Opravdu smazat tuto zprávu?').classes('text-sm')
                with ui.row().classes('w-full justify-end gap-2'):
                    ui.button('Zrušit', on_click=lambda: d.submit(False)).props('flat')
                    ui.button('Smazat', on_click=lambda: d.submit(True)) \
                        .props('unelevated color=negative')
            if await d:
                if _smaz_chat(tabulka, row_hash, z.get('id'), user_id, muze_mazat_vse):
                    intranet_logger.log_activity(user_name, 'Sankce',
                                                 f'Smazán komentář ({tabulka}) #{row_hash[:8]}')
                    _vykresli()
                    on_badge()
                else:
                    ui.notify('Zprávu se nepodařilo smazat.', type='negative')
        return _potvrd

    def _bublina(z):
        moje = (z.get('user_id') == user_id)
        muze_smazat = moje or muze_mazat_vse
        cas = z.get('kdy')
        cas_txt = cas.strftime('%d.%m.%Y %H:%M') if hasattr(cas, 'strftime') else str(cas or '')
        with ui.row().classes('w-full items-center group ' +
                              ('justify-end' if moje else 'justify-start')):
            if moje and muze_smazat:
                ui.button(icon='delete', on_click=_smaz(z)) \
                    .props('flat round dense size=sm color=grey-6') \
                    .classes('opacity-0 group-hover:opacity-100').tooltip('Smazat zprávu')
            with ui.column().classes('gap-0').style('max-width:78%'):
                if not moje:
                    ui.label(z.get('jmeno') or '—').classes('text-xs font-semibold text-gray-600 px-1')
                bg = 'background:#2563eb;color:#fff' if moje else 'background:#f1f5f9;color:#1e293b'
                with ui.element('div').classes('rounded-2xl px-3 py-2').style(bg):
                    ui.label(z.get('zprava') or '').classes('text-sm') \
                        .style('white-space:pre-wrap;word-break:break-word')
                ui.label(cas_txt).classes('text-gray-400 px-1 ' + ('self-end' if moje else '')) \
                    .style('font-size:10px')
            if not moje and muze_smazat:
                ui.button(icon='delete', on_click=_smaz(z)) \
                    .props('flat round dense size=sm color=grey-6') \
                    .classes('opacity-0 group-hover:opacity-100').tooltip('Smazat zprávu')

    with ui.dialog() as dlg, ui.card().classes('p-0 rounded-2xl gap-0') \
            .style('min-width:560px;max-width:680px'):
        with ui.row().classes('items-center gap-2 px-5 pt-4 pb-2 w-full'):
            ui.icon('forum', color='primary').classes('text-2xl')
            with ui.column().classes('gap-0'):
                ui.label('Diskuze k případu').classes('text-lg font-bold text-gray-800')
                if popis.strip(' –'):
                    ui.label(popis).classes('text-xs text-gray-500')
            ui.space()
            ui.button(icon='close', on_click=dlg.close).props('flat round dense color=grey-7')
        ui.separator()
        box = ui.scroll_area().classes('w-full').style('height:46vh')

        def _vykresli(scroll=False):
            zpravy = _nacti_chat_zpravy(tabulka, row_hash)
            stav['max_id'] = zpravy[-1]['id'] if zpravy else 0
            box.clear()
            with box:
                with ui.column().classes('w-full gap-2 px-3 py-2'):
                    if not zpravy:
                        with ui.column().classes('items-center w-full py-10 gap-2'):
                            ui.icon('chat_bubble_outline', size='2.5rem', color='grey-4')
                            ui.label('Zatím tu nikdo nic nenapsal.').classes('text-sm text-gray-400')
                    for z in zpravy:
                        _bublina(z)
            if scroll:
                ui.timer(0.05, lambda: box.scroll_to(percent=1.0), once=True)

        ui.separator()
        with ui.row().classes('w-full items-end gap-2 px-4 py-3'):
            inp = ui.textarea(placeholder='Napiš zprávu… (Ctrl+Enter odešle)') \
                .props('outlined autogrow dense input-style=max-height:120px').classes('flex-1')

            async def _send():
                txt = (inp.value or '').strip()
                if not txt:
                    return
                if _pridej_chat(tabulka, row_hash, user_id, user_name, txt):
                    inp.value = ''
                    intranet_logger.log_activity(user_name, 'Sankce',
                                                 f'Chat ({tabulka}) #{row_hash[:8]}')
                    _vykresli(scroll=True)
                    on_badge()
                    # „cinkni" do zvonečku všem s právy na tuto sestavu (kromě autora)
                    asyncio.create_task(asyncio.to_thread(
                        _notifikuj_novy_komentar, tabulka, row_hash,
                        user_id, user_name, popis, tuple(prava_navic or ())))
                else:
                    ui.notify('Zprávu se nepodařilo uložit.', type='negative')

            inp.on('keydown.ctrl.enter', _send)
            ui.button(icon='send', on_click=_send).props('round unelevated color=primary') \
                .tooltip('Odeslat (Ctrl+Enter)')

        _vykresli(scroll=True)

        async def _tick():
            if not dlg.value:          # dialog zavřen → nic nedělej
                return
            try:
                # dotaz ve vlákně — polling neblokuje event loop
                novy = await asyncio.to_thread(_chat_max_id, tabulka, row_hash)
                if not dlg.value:      # dialog se mezitím zavřel
                    return
                if novy != stav['max_id']:
                    await asyncio.to_thread(_oznac_precteno, tabulka, row_hash, user_id)
                    _vykresli(scroll=True)
                    on_badge()
            except Exception:
                pass
        chat_timer = ui.timer(5.0, _tick)
        # po zavření okna polling zastav (jinak by běžel dál na pozadí)
        dlg.on('hide', lambda: chat_timer.cancel())
    dlg.open()


# =========================================================
# POHLED: ZAMÍTNUTÉ DODÁVKY DODAVATELEM
# =========================================================
def _col_defs_zamitnute(muze_psat_pozn: bool) -> list:
    # Poznámku smí psát Nákup, ale NIKDY na připnutém řádku „Celkem" (nemá id).
    _edit = ("function(p){return %s && !(p.node&&p.node.rowPinned);}"
             % ('true' if muze_psat_pozn else 'false'))
    cols = [
        {'headerName': '', 'field': '_eye', 'width': 46, 'minWidth': 46, 'maxWidth': 46,
         'pinned': 'left', 'sortable': False, 'editable': False, 'resizable': False,
         'suppressSizeToFit': True, 'suppressAutoSize': True,
         ':cellRenderer': _EYE_RENDERER,
         'cellStyle': {'textAlign': 'center', 'cursor': 'pointer', 'padding': '0'},
         'headerTooltip': 'Historie změn řádku'},
        _col_chat(),
        _col_poradi(),
        {'headerName': 'Období', 'field': 'obdobi', 'width': 150, 'pinned': 'left', 'sortable': True,
         'cellStyle': {'fontSize': '12px', 'color': '#475569'}},
        {'headerName': 'IČO', 'field': 'ico', 'width': 95, 'sortable': True,
         'cellStyle': {'fontFamily': 'monospace', 'fontSize': '12px'}},
        {'headerName': 'Dodavatel', 'field': 'jmeno_dodavatele', 'width': 200, 'sortable': True},
        {'headerName': 'Kód zboží', 'field': 'kod_zbozi', 'width': 100,
         'cellStyle': {'fontFamily': 'monospace', 'fontSize': '12px'}},
        {'headerName': 'Název zboží', 'field': 'nazev_zbozi', 'width': 260, 'sortable': True},
        {'headerName': 'Datum pož.', 'field': 'datum_pozadovano', 'width': 110},
        {'headerName': 'Typ adresy', 'field': 'typ_adresy', 'width': 100},
        {'headerName': 'K2', 'field': 'k2', 'width': 70},
        {'headerName': 'Č.obj.', 'field': 'cislo_objednavky', 'width': 110},
        {'headerName': 'Pobočka', 'field': 'id_pobocky', 'width': 90, 'sortable': True},
        {'headerName': 'Dodavatel(k.)', 'field': 'dodavatel', 'width': 100},
        {'headerName': 'Nákupčí', 'field': 'nakupci', 'width': 100},
        {'headerName': 'Objedn. MJ', 'field': 'objednano_mj', 'width': 100, 'type': 'numericColumn', ':valueFormatter': _NUM_FMT},
        {'headerName': 'Dodáno MJ', 'field': 'dodano_mj', 'width': 100, 'type': 'numericColumn', ':valueFormatter': _NUM_FMT},
        {'headerName': 'Odmít. MJ', 'field': 'odmitnuto_mj', 'width': 100, 'type': 'numericColumn', ':valueFormatter': _NUM_FMT},
        {'headerName': 'Obj.-cena', 'field': 'obj_cena', 'width': 110, 'type': 'numericColumn', ':valueFormatter': _MONEY_FMT},
        {'headerName': 'Typ sankce', 'field': 'typ_sankce', 'width': 130, 'sortable': True},
        {'headerName': 'Hodn. sankce', 'field': 'hodn_sankce', 'width': 120, 'type': 'numericColumn',
         ':valueFormatter': _MONEY_FMT, 'cellStyle': {'fontWeight': 'bold'}},
        {'headerName': 'Odmít. Kč celkem', 'field': 'odmitnuto_kc_celkem', 'width': 140, 'type': 'numericColumn', ':valueFormatter': _MONEY_FMT},
        {'headerName': 'Poznámka', 'field': 'poznamka', 'width': 220,
         ':editable': _edit, 'cellEditor': 'agLargeTextCellEditor',
         'cellEditorPopup': True,
         'cellStyle': {'backgroundColor': '#fffbeb'} if muze_psat_pozn else None,
         'headerTooltip': 'Poznámku může psát Nákup'},
    ]
    return cols


@refreshable_na_klienta
async def _vykresli_zamitnute(user_id, user_name, vsechna_prava):
    ma_vse = 'vse' in vsechna_prava
    je_analytik = ma_vse or 'sankce_analytik' in vsechna_prava
    muze_psat = ma_vse or 'sankce_nakup' in vsechna_prava

    vsechny = await asyncio.to_thread(_nacti, 'sankce_zamitnute')
    if not vsechny:
        if je_analytik:
            with ui.row().classes('w-full justify-end mb-2'):
                _import_button('sankce_zamitnute', _MAPA_ZAMITNUTE, _CISLA_ZAMITNUTE,
                               user_name, _vykresli_zamitnute.refresh)
        with ui.column().classes('items-center py-16 gap-3 w-full'):
            ui.icon('block', size='4rem', color='grey-4')
            ui.label('Zatím nejsou naimportována žádná data.').classes('text-xl text-gray-400 font-bold')
            if je_analytik:
                ui.label('Nahrajte list DATA tlačítkem „Nahrát data" vpravo nahoře.').classes('text-sm text-gray-400')
        return

    # diskuze (chat) — indikátory nepřečtených zpráv pro tohoto uživatele
    _chat_stav = await asyncio.to_thread(_nacti_chat_stav, 'sankce_zamitnute', user_id)
    for _r in vsechny:
        _st = _chat_stav.get(_r.get('row_hash'))
        _r['_chat_pocet'] = _st['pocet'] if _st else 0
        _r['_chat_unread'] = _st['unread'] if _st else False

    obdobi_list = _seznam_obdobi('sankce_zamitnute')
    zamcene = _ziskej_zamcene_obdobi()
    if zamcene not in obdobi_list:
        zamcene = obdobi_list[0] if obdobi_list else ''

    stav = {'obdobi': zamcene, 'vse': False}

    def _zobrazene():
        if stav['vse']:
            return vsechny
        return [r for r in vsechny if r.get('obdobi') == stav['obdobi']]

    def _celkem_row(data):
        """Připnutý součtový řádek „Celkem" — sečte množstevní a hodnotové sloupce
        (Objednáno/Dodáno/Odmít. MJ, Hodn. sankce, Odmít. Kč celkem); Obj.-cenu ne."""
        radek = {p: sum(_f(r.get(p)) or 0 for r in data) for p in _SOUCET_ZAMITNUTE}
        radek['obdobi'] = 'CELKEM'
        radek['jmeno_dodavatele'] = f'{len(data)} pol.'
        return radek

    async def _export():
        """Export do .xlsx přesně toho, co je právě vidět (období + filtry
        v hlavičkách + řazení). Čísla jdou do sešitu jako čísla."""
        data = _zobrazene()
        ids = await _viditelne_ids(grid)
        data, dle_gridu = _serad_dle_ids(data, ids)
        if not dle_gridu:
            ui.notify('Filtry v hlavičkách sloupců se nepodařilo přečíst — '
                      'exportuji vše za zvolené období.',
                      type='warning', position='top', timeout=6000, multi_line=True)
        if not data:
            ui.notify('Aktuální filtr nevrací žádné řádky — není co exportovat.',
                      type='warning', position='top', timeout=6000)
            return
        await _export_xlsx(_EXP_ZAMITNUTE, data, _celkem_row(data),
                           'zamitnute_dodavky', 'Zamítnuté dodávky')

    def _aplikuj():
        data = _zobrazene()
        grid.options['rowData'] = data
        grid.options['pinnedBottomRowData'] = [_celkem_row(data)]
        grid.update()
        # Pokud uživatel měl aktivní AG Grid filtr v hlavičce, re-aplikuj ho
        # na nová data a nech `:onFilterChanged` přepočítat součtový řádek.
        grid.run_grid_method('onFilterChanged')
        info.set_text(f'Zobrazeno řádků: {len(data)} '
                      + ('(všechna období)' if stav['vse'] else f'(období {stav["obdobi"]})'))

    # ── Ovládací lišta (filtr období + zámek) ──
    with ui.row().classes('w-full items-center gap-3 mb-2 flex-wrap'):
        ui.icon('lock' if not stav['vse'] else 'lock_open',
                color='amber-8' if not stav['vse'] else 'grey').classes('text-xl')
        sel_obdobi = ui.select(obdobi_list, value=zamcene, label='Období') \
            .props('outlined dense options-dense').classes('w-72')

        def _on_obdobi(e):
            stav['obdobi'] = e.value
            _aplikuj()
        sel_obdobi.on_value_change(_on_obdobi)

        sw_vse = ui.switch('Zobrazit všechna období', value=False)

        def _on_vse(e):
            stav['vse'] = bool(e.value)
            sel_obdobi.set_enabled(not stav['vse'])
            _aplikuj()
        sw_vse.on_value_change(_on_vse)

        if je_analytik:
            def _zamkni():
                if not sel_obdobi.value:
                    return ui.notify('Nejdřív vyberte období.', type='warning')
                _uloz_zamcene_obdobi(sel_obdobi.value)
                intranet_logger.log_activity(user_name, 'Sankce',
                                             f'Nastaven zámek období (Zamítnuté): {sel_obdobi.value}')
                ui.notify(f'Výchozí (zamčené) období pro všechny: {sel_obdobi.value}', type='positive')
            ui.button('Zamknout jako výchozí', icon='push_pin', on_click=_zamkni) \
                .props('outline color=amber-8 dense no-caps') \
                .tooltip('Toto období se všem zobrazí jako výchozí po otevření sestavy')

        ui.space()
        info = ui.label('').classes('text-sm text-gray-500')
        ui.button(icon='download', text='Export', on_click=_export) \
            .props('color=secondary outline dense no-caps') \
            .tooltip('Stáhne .xlsx s tím, co je právě vidět (období, filtry v hlavičkách '
                     'i řazení). Částky a množství jsou skutečná čísla — jdou rovnou '
                     'sčítat; IČO a čísla objednávek zůstanou textem s vedoucími nulami.')
        if je_analytik:
            _import_button('sankce_zamitnute', _MAPA_ZAMITNUTE, _CISLA_ZAMITNUTE,
                           user_name, _vykresli_zamitnute.refresh)
            _smazat_button('sankce_zamitnute', 'Zamítnuté dodávky',
                           user_name, _vykresli_zamitnute.refresh)

    if je_analytik:
        ui.label('🗑 Jednotlivý řádek smažete klávesou Delete (po kliknutí na něj) nebo '
                 'pravým tlačítkem → „Smazat řádek" — vždy se zobrazí potvrzení.') \
            .classes('text-xs text-gray-500 mb-1')
    if not muze_psat and not je_analytik:
        ui.label('👁 Čtenář: sestavu vidíte jen pro čtení — filtrovat, řadit, exportovat '
                 'a psát do diskuze (💬) můžete, měnit hodnoty ne.') \
            .classes('text-xs text-gray-500 mb-1')

    _opts_z = {
        'columnDefs': _col_defs_zamitnute(muze_psat),
        'rowData': _zobrazene(),
        'pinnedBottomRowData': [_celkem_row(_zobrazene())],
        'defaultColDef': {'resizable': True, 'sortable': False, 'filter': True},
        'rowHeight': 32,
        'singleClickEdit': True,
        'stopEditingWhenCellsLoseFocus': True,
        'suppressMovableColumns': True,
        ':getRowStyle': _PINNED_TOTAL_STYLE,
        ':onFirstDataRendered': _AUTOSIZE_FIT,
        ':onGridSizeChanged': _AUTOSIZE_FIT,
        ':onFilterChanged': _FILTER_RECALC_ZAMITNUTE,
        ':getRowId': _GET_ROW_ID,
    }
    if je_analytik:
        _opts_z.update(_grid_mazani_js('sankce_zamitnute'))
    grid = ui.aggrid(_opts_z).classes('w-full').style(_GRID_STYLE)

    info.set_text(f'Zobrazeno řádků: {len(_zobrazene())} (období {stav["obdobi"]})')

    # diskuze (chat): překreslení indikátoru jednoho řádku v gridu
    def _chat_badge(row_hash):
        st = _chat_stav_radku('sankce_zamitnute', row_hash, user_id)
        upd = []
        for r in vsechny:
            if r.get('row_hash') == row_hash:
                r['_chat_pocet'] = st['pocet']; r['_chat_unread'] = st['unread']
                upd.append(r)
        if upd:
            grid.run_grid_method('applyTransaction', {'update': upd})

    # očičko (historie) + chat (diskuze)
    def _on_click(e):
        a = e.args or {}
        col = a.get('colId')
        d = a.get('data') or {}
        rh = d.get('row_hash')
        if not rh:
            return
        popis = f"{d.get('jmeno_dodavatele','')} – {d.get('nazev_zbozi','')}"
        if col == '_eye':
            _zobraz_historii('sankce_zamitnute', rh, popis)
        elif col == '_chat':
            _otevri_chat('sankce_zamitnute', rh, popis, user_id, user_name,
                         lambda: _chat_badge(rh),
                         muze_mazat_vse=je_analytik or 'sankce_ucetni' in vsechna_prava)
    grid.on('cellClicked', _on_click)

    # živé „svítící" upozornění na nové zprávy (~20 s)
    async def _poll_chat():
        try:
            stav_chat = await asyncio.to_thread(_nacti_chat_stav, 'sankce_zamitnute', user_id)
            zmeneno = []
            for r in vsechny:
                st = stav_chat.get(r.get('row_hash')) or {'pocet': 0, 'unread': False}
                if r.get('_chat_pocet') != st['pocet'] or r.get('_chat_unread') != st['unread']:
                    r['_chat_pocet'] = st['pocet']; r['_chat_unread'] = st['unread']
                    zmeneno.append(r)
            if zmeneno:
                grid.run_grid_method('applyTransaction', {'update': zmeneno})
        except Exception:
            pass
    ui.timer(20.0, _poll_chat)

    # editace poznámky
    if muze_psat:
        def _on_change(e):
            a = e.args or {}
            if a.get('colId') != 'poznamka':
                return
            d = a.get('data') or {}
            rid = d.get('id'); rh = d.get('row_hash')
            nova = a.get('newValue'); stara = a.get('oldValue')
            if not rid or str(nova or '') == str(stara or ''):
                return
            conn = intranet_data.get_db_connection()
            if conn:
                cur = conn.cursor()
                cur.execute('UPDATE sankce_zamitnute SET poznamka=%s WHERE id=%s', (nova, rid))
                conn.commit(); cur.close(); conn.close()
            zapis_audit('sankce_zamitnute', rh, rid, 'poznamka', stara, nova, user_id, user_name)
            for r in vsechny:
                if r.get('id') == rid:
                    r['poznamka'] = nova
                    break
            intranet_logger.log_activity(user_name, 'Sankce', f'Poznámka (Zamítnuté) #{rid}')
        grid.on('cellValueChanged', _on_change)


# =========================================================
# POHLED: SANKCE K VYSTAVENÍ
# =========================================================
def _col_defs_vystaveni(je_ucetni: bool) -> list:
    # Editovatelnost účtárny, ale NIKDY na připnutém řádku „Celkem" (nemá id —
    # editace by se neuložila a jen mátla).
    _edit = ("function(p){return %s && !(p.node&&p.node.rowPinned);}"
             % ('true' if je_ucetni else 'false'))
    cols = [
        {'headerName': '', 'field': '_sel', 'width': 44, 'minWidth': 44, 'maxWidth': 44,
         'pinned': 'left', 'checkboxSelection': True, 'headerCheckboxSelection': True,
         'headerCheckboxSelectionFilteredOnly': True,
         'sortable': False, 'editable': False, 'resizable': False, 'filter': False,
         'suppressSizeToFit': True, 'suppressAutoSize': True, 'suppressMovable': True,
         'headerTooltip': 'Označení řádků k tisku (rozsah: Shift+klik)'},
        {'headerName': '', 'field': '_eye', 'width': 46, 'minWidth': 46, 'maxWidth': 46,
         'pinned': 'left', 'sortable': False, 'editable': False, 'resizable': False,
         'suppressSizeToFit': True, 'suppressAutoSize': True,
         ':cellRenderer': _EYE_RENDERER,
         'cellStyle': {'textAlign': 'center', 'cursor': 'pointer', 'padding': '0'},
         'headerTooltip': 'Historie změn řádku'},
        _col_chat(),
        _col_poradi(),
        {'headerName': 'Stav', 'field': 'stav_label', 'width': 160, 'pinned': 'left', 'sortable': True,
         ':editable': _edit, 'cellEditor': 'agSelectCellEditor',
         'cellEditorParams': {'values': list(STAV_LABEL.values())},
         ':cellStyle': _STAV_STYLE,
         'headerTooltip': 'Stav mění účtárna' if je_ucetni else 'Stav nastavuje účtárna'},
        {'headerName': 'Období', 'field': 'obdobi', 'width': 150, 'pinned': 'left', 'sortable': True,
         'cellStyle': {'fontSize': '12px', 'color': '#475569'}},
        {'headerName': 'IČO', 'field': 'ico', 'width': 95, 'sortable': True,
         'cellStyle': {'fontFamily': 'monospace', 'fontSize': '12px'}},
        {'headerName': 'Dodavatel', 'field': 'jmeno_dodavatele', 'width': 220, 'sortable': True},
        {'headerName': 'Kód zboží', 'field': 'kod_zbozi', 'width': 100,
         'cellStyle': {'fontFamily': 'monospace', 'fontSize': '12px'}},
        {'headerName': 'Název zboží', 'field': 'nazev_zbozi', 'width': 280, 'sortable': True},
        {'headerName': 'Č.obj.', 'field': 'cislo_objednavky', 'width': 110},
        {'headerName': 'Pobočka', 'field': 'id_pobocky', 'width': 90, 'sortable': True},
        {'headerName': 'Objedn. MJ', 'field': 'objednano_mj', 'width': 100, 'type': 'numericColumn', ':valueFormatter': _NUM_FMT},
        {'headerName': 'Dodáno MJ', 'field': 'dodano_mj', 'width': 100, 'type': 'numericColumn', ':valueFormatter': _NUM_FMT},
        {'headerName': 'Dod. pozdě MJ', 'field': 'dod_pozde_mj', 'width': 110, 'type': 'numericColumn', ':valueFormatter': _NUM_FMT},
        {'headerName': 'Obj.-cena', 'field': 'obj_cena', 'width': 110, 'type': 'numericColumn', ':valueFormatter': _MONEY_FMT},
        {'headerName': 'Hodn. sankce', 'field': 'hodn_sankce', 'width': 120, 'type': 'numericColumn',
         ':valueFormatter': _MONEY_FMT, 'cellStyle': {'fontWeight': 'bold'}},
        {'headerName': 'Nákupčí (pob.)', 'field': 'nakupci_pob', 'width': 110, 'sortable': True},
        {'headerName': 'Sleva na sankci', 'field': 'sleva', 'width': 130, 'sortable': True,
         ':editable': _edit, ':valueFormatter': _SLEVA_FMT,
         'cellStyle': ({'textAlign': 'right', 'backgroundColor': '#fffbeb'} if je_ucetni
                       else {'textAlign': 'right'}),
         'headerTooltip': 'Sleva na sankci jako podíl: 0,05 = 5 %. Cena v řádku zůstává '
                          'beze změny, sleva se promítne až do součtu „Hodn. sankce" '
                          'v řádku CELKEM. Lze nastavit hromadně na označené řádky.'},
        {'headerName': 'Poznámka', 'field': 'poznamka', 'width': 240,
         ':editable': _edit, 'cellEditor': 'agLargeTextCellEditor', 'cellEditorPopup': True,
         'cellStyle': {'backgroundColor': '#fffbeb'} if je_ucetni else None,
         'headerTooltip': 'Poznámku píše účtárna'},
        {'headerName': 'Aktivita 2', 'field': 'stav2_label', 'width': 130, 'sortable': True,
         ':editable': _edit, 'cellEditor': 'agSelectCellEditor',
         'cellEditorParams': {'values': list(STAV2_LABEL.values())},
         ':cellStyle': _STAV2_STYLE,
         'headerTooltip': 'Druhý stav (V procesu / Uzavřeno) — mění účtárna'
                          if je_ucetni else 'Druhý stav (V procesu / Uzavřeno)'},
    ]
    return cols


def _col_defs_souhrn(je_ucetni: bool) -> list:
    """Sloupce souhrnného pohledu — jeden řádek = jeden dodavatel (IČ + název).
    Žádný rozpad na kód zboží ani čísla objednávek; hodnoty jsou součty přes
    aktuálně zobrazené (vyfiltrované) řádky detailní sestavy."""
    _edit = ("function(p){return %s && !(p.node&&p.node.rowPinned);}"
             % ('true' if je_ucetni else 'false'))
    return [
        {'headerName': '', 'field': '_sel', 'width': 44, 'minWidth': 44, 'maxWidth': 44,
         'pinned': 'left', 'checkboxSelection': True, 'headerCheckboxSelection': True,
         'headerCheckboxSelectionFilteredOnly': True,
         'sortable': False, 'editable': False, 'resizable': False, 'filter': False,
         'suppressSizeToFit': True, 'suppressAutoSize': True, 'suppressMovable': True,
         'headerTooltip': 'Označení dodavatelů k tisku (rozsah: Shift+klik)'},
        {'headerName': 'IČ', 'field': 'ico', 'width': 110, 'pinned': 'left', 'sortable': True,
         'cellStyle': {'fontFamily': 'monospace', 'fontSize': '12px'}},
        {'headerName': 'Dodavatel', 'field': 'jmeno_dodavatele', 'width': 300, 'pinned': 'left',
         'sortable': True, 'cellStyle': {'fontWeight': '600'}},
        {'headerName': 'Období', 'field': 'obdobi', 'width': 150, 'sortable': True,
         'cellStyle': {'fontSize': '12px', 'color': '#475569'}},
        {'headerName': 'Nákupčí (pob.)', 'field': 'nakupci_pob', 'width': 120, 'sortable': True,
         'tooltipField': 'nakupci_pob',
         'headerTooltip': 'Nákupčí pobočky ze zobrazených řádků dodavatele; '
                          'víc hodnot = dodavatel spadá pod víc nákupčích'},
        {'headerName': 'Položek', 'field': 'pocet', 'width': 95, 'sortable': True,
         'type': 'numericColumn', ':valueFormatter': _NUM_FMT,
         'headerTooltip': 'Počet řádků (položek) dodavatele v aktuálním filtru'},
        {'headerName': 'Sankce celkem', 'field': 'sankce_celkem', 'width': 145, 'sortable': True,
         'type': 'numericColumn', ':valueFormatter': _MONEY_FMT,
         'headerTooltip': 'Σ „Hodn. sankce" před slevou'},
        {'headerName': 'Sleva', 'field': 'sleva_castka', 'width': 120, 'sortable': True,
         'type': 'numericColumn', ':valueFormatter': _MONEY_FMT,
         'cellStyle': {'color': '#b45309'},
         'headerTooltip': 'Σ poskytnutých slev v Kč (sankce celkem − sankce uznaná)'},
        {'headerName': 'Sankce uznaná', 'field': 'sankce_uznana', 'width': 155, 'sortable': True,
         'type': 'numericColumn', ':valueFormatter': _MONEY_FMT,
         'cellStyle': {'fontWeight': 'bold'},
         'headerTooltip': 'Σ „Hodn. sankce" po slevě — částka, která jde dodavateli'},
        {'headerName': 'Stav', 'field': 'stav_label', 'width': 175, 'sortable': True,
         ':editable': _edit, 'cellEditor': 'agSelectCellEditor',
         'cellEditorParams': {'values': list(STAV_LABEL.values())},
         ':cellStyle': _SOUHRN_STAV_STYLE,
         'headerTooltip': ('Stav dodavatele. „%s" = řádky mají různé stavy; '
                           'výběrem se stav nastaví na VŠECHNY jeho zobrazené řádky.'
                           % _MIX_LABEL) if je_ucetni else
                          ('Stav dodavatele („%s" = řádky mají různé stavy)' % _MIX_LABEL)},
        {'headerName': 'Aktivita 2', 'field': 'stav2_label', 'width': 150, 'sortable': True,
         ':editable': _edit, 'cellEditor': 'agSelectCellEditor',
         'cellEditorParams': {'values': list(STAV2_LABEL.values())},
         ':cellStyle': _SOUHRN_STAV2_STYLE,
         'headerTooltip': ('Druhý stav dodavatele. „%s" = řádky se liší; výběrem '
                           'se nastaví na všechny jeho zobrazené řádky.' % _MIX_LABEL)
                          if je_ucetni else
                          ('Druhý stav dodavatele („%s" = řádky se liší)' % _MIX_LABEL)},
    ]


@refreshable_na_klienta
async def _vykresli_vystaveni(user_id, user_name, vsechna_prava):
    ma_vse = 'vse' in vsechna_prava
    je_analytik = ma_vse or 'sankce_analytik' in vsechna_prava
    je_ucetni = ma_vse or 'sankce_ucetni' in vsechna_prava

    vsechny = await asyncio.to_thread(_nacti, 'sankce_vystaveni')
    if not vsechny:
        # I v prázdném stavu nech dostupné obnovení dat ze zálohy — po smazání dat
        # je to hlavní cesta, jak je vrátit zpět (pro účetní i analytika, ne pro čtenáře).
        with ui.row().classes('w-full justify-end mb-2 gap-2'):
            if je_ucetni or je_analytik:
                _obnova_button(user_name, _vykresli_vystaveni.refresh, text='Obnovit data',
                               je_hlavni_admin=ma_vse)
            if je_analytik:
                _import_button('sankce_vystaveni', _MAPA_VYSTAVENI, _CISLA_VYSTAVENI,
                               user_name, _vykresli_vystaveni.refresh)
        with ui.column().classes('items-center py-16 gap-3 w-full'):
            ui.icon('receipt_long', size='4rem', color='grey-4')
            ui.label('Zatím nejsou naimportována žádná data.').classes('text-xl text-gray-400 font-bold')
            if je_analytik:
                ui.label('Nahrajte list DATA tlačítkem „Nahrát data" vpravo nahoře, '
                         'nebo obnovte dřívější stav tlačítkem „Obnovit data".') \
                    .classes('text-sm text-gray-400')
            else:
                ui.label('Dřívější stav můžete vrátit tlačítkem „Obnovit data" vpravo nahoře.') \
                    .classes('text-sm text-gray-400')
        return

    # Diskuze (chat) — indikátory nepřečtených zpráv pro tohoto uživatele.
    # Řádek předaný na tiket sdílí vlákno s tiketem (klíč T{id}), aby se diskuze
    # k případu nerozpadla na dvě poloviny.
    def _chat_prehled():
        return (_mapa_tiketu_radku(),
                {'sankce_vystaveni': _nacti_chat_stav('sankce_vystaveni', user_id),
                 'sankce_tikety': _nacti_chat_stav('sankce_tikety', user_id)})

    def _uprav_chat(r, tik, stavy) -> bool:
        """Nastaví řádku klíč vlákna + indikátory; vrací True při změně."""
        tid = tik.get(r.get('row_hash'))
        tab, rh = (('sankce_tikety', _tiket_rh(tid)) if tid
                   else ('sankce_vystaveni', r.get('row_hash')))
        st = stavy.get(tab, {}).get(rh) or {'pocet': 0, 'unread': False}
        zmena = (r.get('_chat_pocet') != st['pocet']
                 or r.get('_chat_unread') != st['unread'])
        r['_chat_tab'], r['_chat_rh'] = tab, rh
        r['_chat_pocet'], r['_chat_unread'] = st['pocet'], st['unread']
        return zmena

    _tik, _chat_stav = await asyncio.to_thread(_chat_prehled)
    for _r in vsechny:
        _uprav_chat(_r, _tik, _chat_stav)

    obdobi_list = _seznam_obdobi('sankce_vystaveni')
    VSE_OBD = '(všechna období)'
    # Filtr období je defaultně na posledním (nejnovějším) nahraném období.
    _vychozi_obd = obdobi_list[0] if obdobi_list else None
    stav = {'stavy': list(STAV_DEFAULT), 'stavy2': list(STAV2_LABEL.keys()),
            'obdobi': _vychozi_obd, 'jen_komentar': False}
    # Který ze dvou pohledů na TATÁŽ data je zobrazený: 'radky' (detail) /
    # 'souhrn' (agregace po dodavatelích). Volba se pamatuje mezi návštěvami.
    pohled = {'v': ('souhrn' if app.storage.user.get('sankce_vystaveni_pohled') == 'souhrn'
                    else 'radky')}

    def _zobrazene():
        data = [r for r in vsechny if r.get('stav') in stav['stavy']]
        data = [r for r in data if (r.get('stav2') or STAV2_DEFAULT) in stav['stavy2']]
        if stav['obdobi']:
            data = [r for r in data if r.get('obdobi') == stav['obdobi']]
        if stav['jen_komentar']:
            data = [r for r in data if (r.get('_chat_pocet') or 0) > 0]
        return data

    def _soucet(data):
        # Σ sankce PO SLEVĚ: hodn_sankce*(1−sleva). Bez slevy (NULL/0) = beze změny.
        return sum((_f(r.get('hodn_sankce')) or 0) * (1 - (_f(r.get('sleva')) or 0)) for r in data)

    def _celkem_row(data):
        """Připnutý součtový řádek „Celkem" — sečte množstevní a hodnotové sloupce
        (Objednáno/Dodáno/Dod. pozdě MJ, Hodn. sankce); Obj.-cenu (cena za 1 MJ) ne.
        „Hodn. sankce" je po slevě: Σ hodn_sankce*(1−sleva) — ceny v řádcích zůstávají."""
        radek = {p: sum(_f(r.get(p)) or 0 for r in data) for p in _SOUCET_VYSTAVENI}
        radek['hodn_sankce'] = sum((_f(r.get('hodn_sankce')) or 0) * (1 - (_f(r.get('sleva')) or 0))
                                   for r in data)
        radek['obdobi'] = 'CELKEM'
        radek['jmeno_dodavatele'] = f'{len(data)} pol.'
        return radek

    # ── Souhrnný pohled: agregace týchž (vyfiltrovaných) řádků po dodavatelích ──
    def _gid(r) -> str:
        """Klíč skupiny = IČ + název dodavatele (bez rozpadu na zboží/objednávky)."""
        return f"{(r.get('ico') or '').strip()}|{(r.get('jmeno_dodavatele') or '').strip()}"

    def _radky_skupiny(gid: str, data=None) -> list:
        """Podkladové řádky jednoho dodavatele — jen ty aktuálně zobrazené."""
        return [r for r in (_zobrazene() if data is None else data) if _gid(r) == gid]

    def _souhrn_data(data=None) -> list:
        data = _zobrazene() if data is None else data
        skup = OrderedDict()
        for r in data:
            k = _gid(r)
            g = skup.get(k)
            if g is None:
                g = skup[k] = {'id': k,
                               'ico': (r.get('ico') or '').strip(),
                               'jmeno_dodavatele': (r.get('jmeno_dodavatele') or '').strip(),
                               'pocet': 0, 'sankce_celkem': 0.0, 'sankce_uznana': 0.0,
                               'sleva_castka': 0.0,
                               '_stavy': set(), '_stavy2': set(), '_obd': set(),
                               '_nak': set()}
            hs = _f(r.get('hodn_sankce')) or 0.0
            sl = _f(r.get('sleva')) or 0.0
            g['pocet'] += 1
            g['sankce_celkem'] += hs
            g['sankce_uznana'] += hs * (1 - sl)
            g['_stavy'].add(r.get('stav'))
            g['_stavy2'].add(r.get('stav2') or STAV2_DEFAULT)
            g['_obd'].add(r.get('obdobi') or '')
            if _s(r.get('nakupci_pob')):
                g['_nak'].add(_s(r.get('nakupci_pob')))
        out = []
        for g in skup.values():
            g['sleva_castka'] = g['sankce_celkem'] - g['sankce_uznana']
            st = g.pop('_stavy'); st2 = g.pop('_stavy2'); obd = sorted(x for x in g.pop('_obd') if x)
            g['stav_label'] = STAV_LABEL.get(next(iter(st)), '') if len(st) == 1 else _MIX_LABEL
            g['stav2_label'] = STAV2_LABEL.get(next(iter(st2)), '') if len(st2) == 1 else _MIX_LABEL
            g['obdobi'] = obd[0] if len(obd) == 1 else (f'{len(obd)} období' if obd else '')
            g['nakupci_pob'] = ', '.join(sorted(g.pop('_nak')))
            out.append(g)
        out.sort(key=lambda x: -x['sankce_uznana'])
        return out

    def _souhrn_celkem(souhrn_rows) -> dict:
        radek = {p: sum(_f(g.get(p)) or 0 for g in souhrn_rows)
                 for p in ('pocet', 'sankce_celkem', 'sleva_castka', 'sankce_uznana')}
        radek['ico'] = 'CELKEM'
        radek['jmeno_dodavatele'] = f'{len(souhrn_rows)} dod.'
        return radek

    async def _export():
        """Export do .xlsx podle právě zvoleného pohledu — detailní řádky, nebo
        souhrn po dodavatelích. Respektuje filtry v hlavičkách i řazení."""
        je_souhrn = pohled['v'] == 'souhrn'
        _g = souhrn if je_souhrn else grid
        data = _souhrn_data() if je_souhrn else _zobrazene()
        ids = await _viditelne_ids(_g)
        data, dle_gridu = _serad_dle_ids(data, ids)
        if not dle_gridu:
            ui.notify('Filtry v hlavičkách sloupců se nepodařilo přečíst — '
                      'exportuji vše podle filtrů nahoře.',
                      type='warning', position='top', timeout=6000, multi_line=True)
        if not data:
            ui.notify('Aktuální filtr nevrací žádné řádky — není co exportovat.',
                      type='warning', position='top', timeout=6000)
            return
        if je_souhrn:
            await _export_xlsx(_EXP_SOUHRN, data, _souhrn_celkem(data),
                               'sankce_dodavatele', 'Dodavatelé (souhrn)')
        else:
            await _export_xlsx(_EXP_VYSTAVENI, data, _celkem_row(data),
                               'sankce_k_vystaveni', 'Sankce k vystavení')

    def _aplikuj_souhrn(data=None):
        """Přepočte souhrnný grid z aktuálních dat. `setGridOption` (ne update())
        — díky `getRowId` jde o immutable update, takže označení dodavatelů
        k tisku, řazení i filtry v hlavičkách zůstanou."""
        s = _souhrn_data(data)
        souhrn.options['rowData'] = s
        souhrn.options['pinnedBottomRowData'] = [_souhrn_celkem(s)]
        souhrn.run_grid_method('setGridOption', 'rowData', s)
        souhrn.run_grid_method('onFilterChanged')
        return s

    def _info_text(data=None):
        data = _zobrazene() if data is None else data
        n_dod = len({_gid(r) for r in data})
        return (f'Zobrazeno řádků: {len(data)} • dodavatelů: {n_dod} • Σ sankce: '
                f'{_soucet(data):,.2f} Kč'.replace(',', ' ').replace('.', ','))

    async def _aplikuj():
        data = _zobrazene()
        # Zachovej aktivní filtry v hlavičkách sloupců (lupa) — grid.update() jinak
        # grid plně překreslí a filtr (např. hledaného dodavatele) by zmizel.
        try:
            fm = await grid.run_grid_method('getFilterModel', timeout=5)
        except Exception:
            fm = None
        grid.options['rowData'] = data
        grid.options['pinnedBottomRowData'] = [_celkem_row(data)]
        grid.update()
        if fm:
            grid.run_grid_method('setFilterModel', fm)
        # Re-aplikuj filtr na nová data a nech `:onFilterChanged` přepočítat součet.
        grid.run_grid_method('onFilterChanged')
        # Souhrn po dodavatelích běží nad týmiž daty — drž ho v synchronu.
        _aplikuj_souhrn(data)
        info.set_text(_info_text(data))

    # ── Ovládací lišta (filtr stavu se zámkem + volitelně období) ──
    with ui.row().classes('w-full items-center gap-3 mb-2 flex-wrap'):
        ui.icon('lock', color='amber-8').classes('text-xl') \
            .tooltip('Výchozí filtr: Nová data + Nevyfakturovat + Rozpracováno + Nákup + Provoz')
        sel_stav = ui.select(
            dict(STAV_LABEL),
            value=list(STAV_DEFAULT), multiple=True, label='Stav') \
            .props('outlined dense options-dense').classes('w-72')

        async def _on_stav(e):
            stav['stavy'] = list(e.value or [])
            await _aplikuj()
        sel_stav.on_value_change(_on_stav)

        def _vse_stavy():
            sel_stav.set_value(list(STAV_LABEL.keys()))
        ui.button('Zobrazit vše', icon='lock_open', on_click=_vse_stavy) \
            .props('outline color=grey dense no-caps') \
            .tooltip('Zruší výchozí filtr a zobrazí i uzavřené řádky (Fakturovat, Odevzdáno účtárně)')

        def _vychozi_stavy():
            sel_stav.set_value(list(STAV_DEFAULT))
        ui.button('Výchozí', icon='lock', on_click=_vychozi_stavy) \
            .props('outline color=amber-8 dense no-caps') \
            .tooltip('Zpět na Nová data + Nevyfakturovat + Rozpracováno + Nákup + Provoz')

        sel_stav2 = ui.select(
            dict(STAV2_LABEL),
            value=list(STAV2_LABEL.keys()), multiple=True, label='Aktivita 2') \
            .props('outlined dense options-dense').classes('w-48') \
            .tooltip('Filtr druhého stavu (V procesu / Uzavřeno)')

        async def _on_stav2(e):
            stav['stavy2'] = list(e.value or [])
            await _aplikuj()
        sel_stav2.on_value_change(_on_stav2)

        sel_obd = ui.select([VSE_OBD] + obdobi_list, value=_vychozi_obd or VSE_OBD, label='Období') \
            .props('outlined dense options-dense').classes('w-60')

        async def _on_obd(e):
            stav['obdobi'] = None if e.value == VSE_OBD else e.value
            await _aplikuj()
        sel_obd.on_value_change(_on_obd)

        sw_koment = ui.switch('Jen s komentářem', value=False) \
            .tooltip('Zobrazí pouze řádky, ke kterým už někdo napsal komentář v diskuzi (💬)')

        async def _on_koment(e):
            stav['jen_komentar'] = bool(e.value)
            await _aplikuj()
        sw_koment.on_value_change(_on_koment)

        ui.space()
        info = ui.label('').classes('text-sm text-gray-500')

        async def _tisk_podkladu(jen_podklad: bool = False):
            """Hromadný tisk: seskupí označené řádky podle dodavatele (IČO + jméno),
            pro každého vyrenderuje samostatné PDF (oznámení + podklad) na serveru
            (Playwright/Chromium). Jeden dodavatel → 1 PDF; více dodavatelů → ZIP.
            `jen_podklad=True` vynechá 1. list (oznámení)."""
            # Z prohlížeče tahám JEN ID označených řádků (ne celá data) — celá data
            # by u stovek řádků přesáhla limit WS payloadu („Payload size exceeds…")
            # i defaultní 1s timeout. Kompletní řádky pak vyberu z `vsechny` v paměti.
            # V souhrnném pohledu je označený „řádek" celý dodavatel → rozbalím ho
            # zpět na jeho podkladové (aktuálně zobrazené) řádky.
            je_souhrn = pohled['v'] == 'souhrn'
            _g = souhrn if je_souhrn else grid
            try:
                sel_ids = await ui.run_javascript(
                    f'const c=getElement({_g.id});'
                    "return (c&&c.run_grid_method)?"
                    "c.run_grid_method('getSelectedRows').map(r=>r.id):[];",
                    timeout=30,
                )
            except TimeoutError:
                ui.notify('Nepodařilo se získat označené řádky z prohlížeče (timeout). '
                          'Zkuste akci zopakovat.',
                          type='negative', position='top', timeout=10000, multi_line=True)
                return
            sel_ids = {str(x) for x in (sel_ids or []) if x is not None}
            if not sel_ids:
                ui.notify(('Označte dodavatele k tisku (zaškrtnutím vlevo).' if je_souhrn else
                           'Označte řádky k tisku (zaškrtnutím vlevo; souvislý rozsah označíte Shift+klikem).'),
                          type='warning', position='top', timeout=8000)
                return
            if je_souhrn:
                sel = [r for r in _zobrazene() if _gid(r) in sel_ids]
            else:
                sel = [r for r in vsechny if str(r.get('id')) in sel_ids]
            if not sel:
                ui.notify('Označené řádky se nepodařilo spárovat s daty (zkuste obnovit stránku).',
                          type='negative', position='top', timeout=9000)
                return

            # Seskup po dodavatelích A PODLE SLEVY (klíč = IČO + jméno + podíl slevy),
            # aby každé oznámení mělo jednotnou slevu: když dodavatel nemá slevu úplně
            # u všech řádků (nebo má různé podíly), rozdělí se do samostatných dokumentů.
            by_dod = OrderedDict()
            for r in sel:
                sl = round(_f(r.get('sleva')) or 0.0, 6)
                klic = ((r.get('ico') or '').strip(),
                        (r.get('jmeno_dodavatele') or '').strip(),
                        sl)
                by_dod.setdefault(klic, []).append(r)
            n_dok = len(by_dod)                                  # počet dokumentů
            n_dod = len({(k[0], k[1]) for k in by_dod})          # počet dodavatelů
            n_pol = len(sel)

            # sestav (filename, html) per dokument — bezpečné názvy souborů
            pouzite = set()
            html_list = []
            for (_ico, dod, sl), rows in by_dod.items():
                nazev = dod or 'dodavatel'
                base_txt = ('Podklad_sankce_' if jen_podklad
                            else 'Oznameni_o_smluvni_pokute_') + nazev
                if sl and sl > 0:
                    pct = _cz_num(sl * 100).replace(',', '_').replace(' ', '')
                    base_txt += f'_sleva_{pct}pct'
                base = _safe_filename(base_txt)
                fname = base + '.pdf'
                # ošetři duplicitní názvy (stejné safe-jméno)
                k = 2
                while fname in pouzite:
                    fname = f'{base}_{k}.pdf'; k += 1
                pouzite.add(fname)
                html_list.append((fname, _slep_dokument(rows, jen_podklad)))

            # průběžná notifikace s progress-barem
            notif = ui.notification(
                f'Připravuji {n_dok} PDF…',
                type='ongoing', position='top-right', spinner=True, timeout=None)

            def _on_wait():
                # render je obsazený jiným uživatelem → portál pojede dál, jen čekáme
                try:
                    notif.message = ('Čekám ve frontě — jiný uživatel právě generuje '
                                     'PDF. Portál i moje stránka pojedou dál.')
                except Exception:
                    pass

            def _on_start():
                try:
                    notif.message = f'Generuji PDF na serveru… (0 / {n_dok})'
                except Exception:
                    pass

            def _cb(i, total, fname):
                try:
                    notif.message = f'Generuji PDF… ({i} / {total})'
                except Exception:
                    pass

            try:
                pdfs = await _render_pdf_bytes_batch(
                    html_list, status_cb=_cb, on_wait=_on_wait, on_start=_on_start)
            except RuntimeError as e:
                notif.dismiss()
                ui.notify(str(e), type='negative', position='top', timeout=15000, multi_line=True)
                return
            except Exception as e:
                notif.dismiss()
                ui.notify(f'Generování PDF selhalo: {e}',
                          type='negative', position='top', timeout=15000, multi_line=True)
                return

            # doručení uživateli — VŽDY přes HTTP (ne WS), aby velký ZIP/PDF prošel
            if len(pdfs) == 1:
                fname, pdf_bytes = pdfs[0]
                _stahni_pres_http(pdf_bytes, fname)
                hl = f'PDF připraveno: {fname}'
            else:
                buf = io.BytesIO()
                with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
                    for fname, pdf_bytes in pdfs:
                        zf.writestr(fname, pdf_bytes)
                zip_name = _safe_filename(
                    ('Podklad_sankce_' if jen_podklad else 'Oznameni_o_smluvni_pokute_')
                    + datetime.date.today().isoformat()) + '.zip'
                _stahni_pres_http(buf.getvalue(), zip_name)
                hl = (f'ZIP připraven: {zip_name} ({len(pdfs)} dokumentů'
                      + (f' / {n_dod} dodavatelů' if n_dod != len(pdfs) else '') + ')')

            notif.dismiss()
            ui.notify(hl, type='positive', position='top-right', timeout=8000)
            intranet_logger.log_activity(user_name, 'Sankce',
                f'Tisk {"podkladu" if jen_podklad else "oznámení"} (k vystavení) '
                f'— {n_dod} dod. / {n_dok} dok., {n_pol} pol.')

        async def _predat_na_tiket():
            """Předání označených řádků k rozhodnutí: založí tikety (dodavatel ×
            kód nákupčího) a přepne řádky na stav Nákup / Provoz. Funguje nad
            označenými řádky i nad označenými dodavateli v souhrnu."""
            je_souhrn = pohled['v'] == 'souhrn'
            _g = souhrn if je_souhrn else grid
            try:
                sel_ids = await ui.run_javascript(
                    f'const c=getElement({_g.id});'
                    "return (c&&c.run_grid_method)?"
                    "c.run_grid_method('getSelectedRows').map(r=>r.id):[];",
                    timeout=30,
                )
            except TimeoutError:
                ui.notify('Nepodařilo se získat označené řádky z prohlížeče (timeout).',
                          type='negative', position='top', timeout=10000)
                return
            sel_ids = {str(x) for x in (sel_ids or []) if x is not None}
            if not sel_ids:
                ui.notify(('Označte dodavatele k předání (zaškrtnutím vlevo).' if je_souhrn else
                           'Označte řádky k předání (zaškrtnutím vlevo).'),
                          type='warning', position='top', timeout=8000)
                return
            if je_souhrn:
                sel = [r for r in _zobrazene() if _gid(r) in sel_ids]
            else:
                sel = [r for r in vsechny if str(r.get('id')) in sel_ids]
            if not sel:
                ui.notify('Označené řádky se nepodařilo spárovat s daty.',
                          type='negative', position='top', timeout=9000)
                return

            volba = {'typ': 'nakup'}
            with ui.dialog() as d, ui.card().classes('p-6 rounded-2xl gap-3') \
                    .style('min-width:520px;max-width:640px'):
                ui.label('Předat na tiket').classes('text-xl font-bold text-gray-800')
                ui.label(f'Označeno {len(sel)} řádků. Vznikne jeden tiket na dodavatele '
                         'a kód nákupčího; provozní tiket je vždy jeden na dodavatele.') \
                    .classes('text-sm text-gray-500')
                tg = ui.toggle({'nakup': '🛒 Nákup', 'provoz': '🏭 Provoz'}, value='nakup') \
                    .props('no-caps unelevated')
                nahled = ui.label('').classes('text-sm text-gray-700 font-medium')
                pozn = ui.textarea(label='Poznámka pro řešitele (nepovinná)') \
                    .props('outlined autogrow').classes('w-full')

                # Řádek už předaný (stav Nákup / Provoz) by dostal druhý tiket —
                # účtárnu na to upozorni, ať nevzniknou dvě fronty nad týmiž daty.
                uz_predane = len([r for r in sel if r.get('stav') in ('nakup', 'provoz')])

                def _nahled():
                    skup = _seskup_do_tiketu(sel, volba['typ'])
                    bez = sum(len(v) for k, v in skup.items()
                              if volba['typ'] == 'nakup' and not k[2])
                    txt = f'Vznikne {len(skup)} tiketů z {len(sel)} řádků.'
                    if bez:
                        txt += f' Pozor: {bez} řádků nemá kód nákupčího — takový tiket '\
                               'uvidí jen účtárna a admin.'
                    if uz_predane:
                        txt += f' Pozor: {uz_predane} řádků už je předaných (stav Nákup / '\
                               'Provoz) — dostanou další tiket.'
                    nahled.set_text(txt)
                tg.on_value_change(lambda e: (volba.__setitem__('typ', e.value or 'nakup'),
                                              _nahled()))
                _nahled()

                async def _potvrd():
                    typ = volba['typ']
                    txt = (pozn.value or '').strip()

                    def _prace():
                        zal = _zaloz_tikety(sel, typ, txt, user_id, user_name)
                        for tid, kod, dod, n in zal:
                            prava = (('sankce_tiket_provoz',) if typ == 'provoz'
                                     else (KOD_PRAVO.get(kod),))
                            _notifikuj_tiket(
                                prava, f'Sankce – nový tiket {_tiket_cislo(tid)} ({dod})',
                                f'{user_name} vám předal(a) {_radku(n)} dodavatele {dod} '
                                f'k rozhodnutí.'
                                + (f'\n\nPoznámka: {txt}' if txt else ''))
                        return zal
                    d.close()
                    zal = await asyncio.to_thread(_prace)
                    if not zal:
                        ui.notify('Tikety se nepodařilo založit.',
                                  type='negative', position='top', timeout=8000)
                        return
                    intranet_logger.log_activity(
                        user_name, 'Sankce',
                        f'Předání na tiket ({TIKET_TYP_LABEL[typ]}): '
                        f'{len(zal)} tiketů / {len(sel)} řádků')
                    ui.notify(f'Založeno {len(zal)} tiketů ({len(sel)} řádků) — '
                              f'{TIKET_TYP_LABEL[typ]}.',
                              type='positive', position='top', timeout=7000)
                    _vykresli_vystaveni.refresh()

                with ui.row().classes('w-full justify-end gap-2 mt-2'):
                    ui.button('Zrušit', on_click=d.close).props('flat no-caps')
                    ui.button('Předat', icon='send', on_click=_potvrd) \
                        .props('unelevated no-caps') \
                        .classes('bg-emerald-600 hover:bg-emerald-700 text-white '
                                 'font-semibold rounded-lg shadow-md px-5')
            d.open()

        # Předání na tiket smí jen účtárna (vlastník fronty podkladů).
        if je_ucetni:
            ui.button('Předat na tiket', icon='confirmation_number',
                      on_click=_predat_na_tiket) \
                .props('unelevated no-caps') \
                .classes('bg-emerald-600 hover:bg-emerald-700 text-white font-semibold '
                         'rounded-lg shadow-md px-5') \
                .tooltip('Označené řádky předá nákupčímu (podle kódu) nebo provozu — '
                         'vznikne tiket a řádky přejdou do stavu Nákup / Provoz.')
        # Tisk oznámení i obnova ze zálohy jsou „ostré" operace → čtenář je nedostane.
        if je_ucetni or je_analytik:
            with ui.button('Tisk', icon='print') \
                    .props('unelevated no-caps') \
                    .classes('bg-rose-600 hover:bg-rose-700 text-white font-semibold rounded-lg shadow-md px-5'):
                with ui.menu():
                    ui.menu_item('Kompletní sestava',
                                 on_click=lambda: _tisk_podkladu(False))
                    ui.menu_item('Pouze podklad',
                                 on_click=lambda: _tisk_podkladu(True))
            _obnova_button(user_name, _vykresli_vystaveni.refresh, je_hlavni_admin=ma_vse)
        ui.button(icon='download', text='Export', on_click=_export) \
            .props('color=secondary outline dense no-caps') \
            .tooltip('Stáhne .xlsx s tím, co je právě vidět — podle zvoleného pohledu '
                     '(Detail / Dodavatelé), včetně filtrů v hlavičkách a řazení. Částky '
                     'a množství jsou skutečná čísla; IČO a čísla objednávek zůstanou '
                     'textem s vedoucími nulami.')
        if je_analytik:
            _import_button('sankce_vystaveni', _MAPA_VYSTAVENI, _CISLA_VYSTAVENI,
                           user_name, _vykresli_vystaveni.refresh)
            _pripoj_button('sankce_vystaveni', _MAPA_VYSTAVENI, _CISLA_VYSTAVENI,
                           user_name, _vykresli_vystaveni.refresh)
            _smazat_button('sankce_vystaveni', 'Sankce k vystavení',
                           user_name, _vykresli_vystaveni.refresh)

    if je_ucetni or je_analytik:
        ui.label('🖨️ Tisk oznámení dodavateli: zaškrtněte řádky vlevo (Shift+klik = souvislý '
                 'rozsah). Pro každého dodavatele vznikne 1. list „Oznámení o sankci" '
                 '(předvyplněné: dodavatel, IČO, var. symbol = rozsah našich čísel, dnešní '
                 'datum, období, celková částka) + podklad. Řádky s RŮZNOU slevou se rozdělí '
                 'do samostatných oznámení (každé má jednotnou slevu). JEDEN dokument → PDF; '
                 'VÍCE → ZIP.').classes('text-xs text-gray-500 mb-1')
    else:
        ui.label('👁 Čtenář: sestavu vidíte jen pro čtení — filtrovat, řadit, exportovat '
                 'a psát do diskuze (💬) můžete, měnit hodnoty ne.') \
            .classes('text-xs text-gray-500 mb-1')
    if je_ucetni:
        ui.label('💡 Účtárna: stav i poznámku změníte přímo v buňce (klikem). '
                 'Každá změna se zapíše do historie (👁).').classes('text-xs text-gray-500 mb-1')
    if je_analytik:
        ui.label('🗑 Jednotlivý řádek smažete klávesou Delete (po kliknutí na něj) nebo '
                 'pravým tlačítkem → „Smazat řádek" — vždy se zobrazí potvrzení.') \
            .classes('text-xs text-gray-500 mb-1')

    # Přepínač dvou pohledů na TATÁŽ data (filtry nahoře platí pro oba):
    #   • Řádky      – původní detail (kód zboží, čísla objednávek…)
    #   • Dodavatelé – souhrn po IČ/názvu, bez rozpadu na položky
    with ui.row().classes('items-center gap-3 mb-2'):
        prep = ui.toggle({'radky': '📋 Řádky (detail)', 'souhrn': '🏢 Dodavatelé (souhrn)'},
                         value=pohled['v']).props('no-caps dense unelevated')
        lbl_pohled = ui.label('').classes('text-xs text-gray-500')

    _opts_v = {
        'columnDefs': _col_defs_vystaveni(je_ucetni),
        'rowData': _zobrazene(),
        'pinnedBottomRowData': [_celkem_row(_zobrazene())],
        'defaultColDef': {'resizable': True, 'sortable': False, 'filter': True},
        'rowHeight': 32,
        'singleClickEdit': True,
        'stopEditingWhenCellsLoseFocus': True,
        'suppressMovableColumns': True,
        'rowSelection': 'multiple',
        'suppressRowClickSelection': True,
        ':getRowStyle': _PINNED_TOTAL_STYLE,
        ':onFirstDataRendered': _AUTOSIZE_FIT,
        ':onGridSizeChanged': _AUTOSIZE_FIT,
        ':onFilterChanged': _FILTER_RECALC_VYSTAVENI,
        ':getRowId': _GET_ROW_ID,
    }
    if je_analytik:
        _opts_v.update(_grid_mazani_js('sankce_vystaveni'))
    grid = ui.aggrid(_opts_v).classes('w-full').style(_GRID_STYLE)

    _d0 = _zobrazene()
    _s0 = _souhrn_data(_d0)
    souhrn = ui.aggrid({
        'columnDefs': _col_defs_souhrn(je_ucetni),
        'rowData': _s0,
        'pinnedBottomRowData': [_souhrn_celkem(_s0)],
        'defaultColDef': {'resizable': True, 'sortable': False, 'filter': True},
        'rowHeight': 32,
        'singleClickEdit': True,
        'stopEditingWhenCellsLoseFocus': True,
        'suppressMovableColumns': True,
        'rowSelection': 'multiple',
        'suppressRowClickSelection': True,
        ':getRowStyle': _PINNED_TOTAL_STYLE,
        ':onFirstDataRendered': _AUTOSIZE_FIT,
        ':onGridSizeChanged': _AUTOSIZE_FIT,
        ':onFilterChanged': _FILTER_RECALC_SOUHRN,
        ':getRowId': "function(p){var d=p.data||{};return ''+(d.id!=null?d.id:'');}",
    }).classes('w-full').style(_GRID_STYLE)

    def _prepni(v: str):
        pohled['v'] = v
        app.storage.user['sankce_vystaveni_pohled'] = v
        je_s = v == 'souhrn'
        grid.set_visibility(not je_s)
        souhrn.set_visibility(je_s)
        lbl_pohled.set_text(
            'Jeden řádek = jeden dodavatel (součty přes zobrazené položky). '
            'Tisk i změna stavu se propíší do detailu.' if je_s else
            'Detailní řádky včetně kódu zboží a čísel objednávek.')
        # Grid schovaný přes display:none má nulovou šířku — po zobrazení
        # ho nech znovu rozvrhnout, jinak zůstanou sloupce slepené vlevo.
        (souhrn if je_s else grid).run_grid_method('sizeColumnsToFit')

    prep.on_value_change(lambda e: _prepni(e.value or 'radky'))
    _prepni(pohled['v'])

    info.set_text(_info_text(_d0))

    if je_ucetni:
        async def _on_change_souhrn(e):
            """Změna stavu v souhrnu = změna na VŠECH zobrazených řádcích dodavatele
            (a na všech označených dodavatelích, je-li jich zaškrtnuto víc)."""
            a = e.args or {}
            field = a.get('colId') or ''
            d = a.get('data') or {}
            gid = d.get('id')
            if not gid or field not in ('stav_label', 'stav2_label'):
                return
            nova = a.get('newValue')
            je_stav = field == 'stav_label'
            kod_novy = (STAV_LABEL_REV if je_stav else STAV2_LABEL_REV).get(nova)
            if not kod_novy:      # včetně volby „— různé —"
                _aplikuj_souhrn()
                return
            sloupec = 'stav' if je_stav else 'stav2'

            try:
                sel_ids = await ui.run_javascript(
                    f'const c=getElement({souhrn.id});'
                    "return (c&&c.run_grid_method)?"
                    "c.run_grid_method('getSelectedRows').map(r=>r.id):[];",
                    timeout=5,
                )
            except Exception:
                sel_ids = []
            sel_ids = {str(x) for x in (sel_ids or []) if x is not None}
            cile = {str(gid)}
            if str(gid) in sel_ids and len(sel_ids) > 1:
                cile = sel_ids

            data = _zobrazene()
            cilove = [r for r in data if _gid(r) in cile
                      and (r.get(sloupec) or (None if je_stav else STAV2_DEFAULT)) != kod_novy]
            if not cilove:
                _aplikuj_souhrn(data)
                return

            conn = intranet_data.get_db_connection()
            if conn:
                cur = conn.cursor()
                cur.executemany(
                    f'UPDATE sankce_vystaveni SET {sloupec}=%s WHERE id=%s',
                    [(kod_novy, r.get('id')) for r in cilove],
                )
                conn.commit(); cur.close(); conn.close()

            for r in cilove:
                zapis_audit('sankce_vystaveni', r.get('row_hash'), r.get('id'),
                            sloupec, r.get(sloupec), kod_novy, user_id, user_name)
                r[sloupec] = kod_novy
                r[field] = nova

            intranet_logger.log_activity(
                user_name, 'Sankce',
                f'Souhrn dodavatelů: {"stav" if je_stav else "aktivita 2"} → {nova} '
                f'({len(cile)} dod. / {len(cilove)} řádků)')
            ui.notify(f'„{nova}" nastaveno u {len(cilove)} řádků '
                      f'({len(cile)} dodavatel(ů)).',
                      type='positive', position='top', timeout=4000)

            # propiš do obou pohledů: detail překresli transakcí (nezmizí označení),
            # souhrn přepočti z čerstvých dat
            grid.run_grid_method('applyTransaction', {'update': cilove})
            grid.run_grid_method('onFilterChanged')
            data = _zobrazene()
            celk = [_celkem_row(data)]
            grid.options['pinnedBottomRowData'] = celk
            grid.run_grid_method('setGridOption', 'pinnedBottomRowData', celk)
            _aplikuj_souhrn(data)
            info.set_text(_info_text(data))

        souhrn.on('cellValueChanged', _on_change_souhrn)

    # diskuze (chat): překreslení indikátoru jednoho řádku v gridu
    def _chat_badge(row_hash):
        upd = []
        for r in vsechny:
            if r.get('row_hash') == row_hash:
                st = _chat_stav_radku(r.get('_chat_tab') or 'sankce_vystaveni',
                                      r.get('_chat_rh') or row_hash, user_id)
                r['_chat_pocet'] = st['pocet']; r['_chat_unread'] = st['unread']
                upd.append(r)
        if upd:
            grid.run_grid_method('applyTransaction', {'update': upd})

    # očičko (historie) + chat (diskuze)
    def _on_click(e):
        a = e.args or {}
        col = a.get('colId')
        d = a.get('data') or {}
        rh = d.get('row_hash')
        if not rh:
            return
        popis = f"{d.get('jmeno_dodavatele','')} – {d.get('nazev_zbozi','')}"
        if col == '_eye':
            _zobraz_historii('sankce_vystaveni', rh, popis)
        elif col == '_chat':
            _otevri_chat(d.get('_chat_tab') or 'sankce_vystaveni',
                         d.get('_chat_rh') or rh, popis, user_id, user_name,
                         lambda: _chat_badge(rh),
                         muze_mazat_vse=je_analytik or je_ucetni)
    grid.on('cellClicked', _on_click)

    # živé „svítící" upozornění na nové zprávy (~20 s)
    async def _poll_chat():
        try:
            tik, stav_chat = await asyncio.to_thread(_chat_prehled)
            zmeneno = [r for r in vsechny if _uprav_chat(r, tik, stav_chat)]
            if zmeneno:
                grid.run_grid_method('applyTransaction', {'update': zmeneno})
                # Při aktivním filtru „jen s komentářem" musí nově okomentovaný
                # řádek do gridu přibýt (applyTransaction sám nové řádky nepřidá).
                if stav.get('jen_komentar'):
                    await _aplikuj()
        except Exception:
            pass
    ui.timer(20.0, _poll_chat)

    if je_ucetni:
        async def _on_change(e):
            a = e.args or {}
            field = a.get('colId') or ''
            d = a.get('data') or {}
            rid = d.get('id'); rh = d.get('row_hash')
            if not rid:
                return
            nova = a.get('newValue'); stara = a.get('oldValue')

            # Zachovej aktivní filtry v hlavičkách sloupců (lupa), ať se po změně
            # stavu/slevy nevypnou a uživatel nemusel filtrovaného dodavatele hledat znovu.
            if field in ('stav_label', 'stav2_label', 'sleva'):
                try:
                    _filtr_model = await grid.run_grid_method('getFilterModel', timeout=5)
                except Exception:
                    _filtr_model = None
            else:
                _filtr_model = None

            def _obnov_filtr():
                if _filtr_model:
                    grid.run_grid_method('setFilterModel', _filtr_model)
                    grid.run_grid_method('onFilterChanged')

            if field == 'stav_label':
                kod_novy = STAV_LABEL_REV.get(nova)
                kod_stary = STAV_LABEL_REV.get(stara)
                if not kod_novy or kod_novy == kod_stary:
                    return

                # Pokud má uživatel označeno více řádků (checkboxy vlevo, příp.
                # Shift+klik na rozsah) a edituje stav na jednom z nich, aplikuj
                # stejný stav na všechny označené.
                try:
                    sel_ids = await ui.run_javascript(
                        f'const c=getElement({grid.id});'
                        "return (c&&c.run_grid_method)?"
                        "c.run_grid_method('getSelectedRows').map(r=>r.id):[];",
                        timeout=5,
                    )
                except Exception:
                    sel_ids = []
                sel_ids = {str(x) for x in (sel_ids or []) if x is not None}

                cilove_ids = {str(rid)}
                if str(rid) in sel_ids and len(sel_ids) > 1:
                    cilove_ids = sel_ids

                cilove = [r for r in vsechny
                          if str(r.get('id')) in cilove_ids and r.get('stav') != kod_novy]
                if not cilove:
                    return

                conn = intranet_data.get_db_connection()
                if conn:
                    cur = conn.cursor()
                    cur.executemany(
                        'UPDATE sankce_vystaveni SET stav=%s WHERE id=%s',
                        [(kod_novy, r.get('id')) for r in cilove],
                    )
                    conn.commit(); cur.close(); conn.close()

                for r in cilove:
                    zapis_audit('sankce_vystaveni', r.get('row_hash'), r.get('id'),
                                'stav', r.get('stav'), kod_novy, user_id, user_name)
                    r['stav'] = kod_novy
                    r['stav_label'] = nova

                if len(cilove) == 1:
                    intranet_logger.log_activity(user_name, 'Sankce',
                                                 f'Změna stavu (k vystavení) #{cilove[0].get("id")} → {nova}')
                else:
                    intranet_logger.log_activity(user_name, 'Sankce',
                                                 f'Hromadná změna stavu (k vystavení) — {len(cilove)} řádků → {nova}')
                    ui.notify(f'Stav „{nova}" nastaven u {len(cilove)} označených řádků.',
                              type='positive', position='top', timeout=4000)

                # překresli VŠECHNY změněné řádky — včetně editovaného. AG Grid sice
                # při commitu editoru zapíše novou hodnotu do data['stav_label'], ale
                # data['stav'] zůstane staré a cellStyle se nepřepočte; explicitní
                # update sjednotí stav buňky se zbytkem.
                grid.run_grid_method('applyTransaction', {'update': cilove})

                # přepočet součtu (řádek může vypadnout z filtru až po překreslení)
                _d = _zobrazene()
                info.set_text(_info_text(_d))
                _aplikuj_souhrn(_d)     # drž souhrn dodavatelů v synchronu
                # přepočítej i připnutý řádek „Celkem" (bez plného překreslení, ať
                # nezmizí označení řádků k tisku)
                celk = [_celkem_row(_d)]
                grid.options['pinnedBottomRowData'] = celk
                grid.run_grid_method('setGridOption', 'pinnedBottomRowData', celk)
                # když má uživatel aktivní AG Grid filtr (např. na Stav), nechej
                # ho znovu projet a `:onFilterChanged` přepočte součet jen přes
                # aktuálně viditelné řádky.
                grid.run_grid_method('onFilterChanged')
                _obnov_filtr()

            elif field == 'stav2_label':
                kod_novy = STAV2_LABEL_REV.get(nova)
                kod_stary = STAV2_LABEL_REV.get(stara)
                if not kod_novy or kod_novy == kod_stary:
                    return

                # Hromadné nastavení na všechny označené řádky (jako u hlavní roletky).
                try:
                    sel_ids = await ui.run_javascript(
                        f'const c=getElement({grid.id});'
                        "return (c&&c.run_grid_method)?"
                        "c.run_grid_method('getSelectedRows').map(r=>r.id):[];",
                        timeout=5,
                    )
                except Exception:
                    sel_ids = []
                sel_ids = {str(x) for x in (sel_ids or []) if x is not None}

                cilove_ids = {str(rid)}
                if str(rid) in sel_ids and len(sel_ids) > 1:
                    cilove_ids = sel_ids

                cilove = [r for r in vsechny
                          if str(r.get('id')) in cilove_ids and r.get('stav2') != kod_novy]
                if not cilove:
                    return

                conn = intranet_data.get_db_connection()
                if conn:
                    cur = conn.cursor()
                    cur.executemany(
                        'UPDATE sankce_vystaveni SET stav2=%s WHERE id=%s',
                        [(kod_novy, r.get('id')) for r in cilove],
                    )
                    conn.commit(); cur.close(); conn.close()

                for r in cilove:
                    zapis_audit('sankce_vystaveni', r.get('row_hash'), r.get('id'),
                                'stav2', r.get('stav2'), kod_novy, user_id, user_name)
                    r['stav2'] = kod_novy
                    r['stav2_label'] = nova

                if len(cilove) == 1:
                    intranet_logger.log_activity(user_name, 'Sankce',
                                                 f'Změna Aktivity 2 (k vystavení) #{cilove[0].get("id")} → {nova}')
                else:
                    intranet_logger.log_activity(user_name, 'Sankce',
                                                 f'Hromadná změna Aktivity 2 (k vystavení) — {len(cilove)} řádků → {nova}')
                    ui.notify(f'„Aktivita 2 = {nova}" nastaveno u {len(cilove)} označených řádků.',
                              type='positive', position='top', timeout=4000)

                grid.run_grid_method('applyTransaction', {'update': cilove})
                _aplikuj_souhrn()       # drž souhrn dodavatelů v synchronu
                _obnov_filtr()

            elif field == 'sleva':
                # Podíl slevy (0,05 = 5 %). Cena v řádku se NEMĚNÍ; promítne se až
                # do součtu „Hodn. sankce" v řádku CELKEM (server i JS přepočet).
                nova_val = _f(nova)
                if nova_val is None:
                    nova_val = 0.0
                if nova_val < 0:
                    nova_val = 0.0
                if nova_val > 1:
                    nova_val = 1.0
                    ui.notify('Sleva se zadává jako podíl: 0,05 = 5 %. Hodnota nad 1 (100 %) '
                              'byla omezena na 1.', type='warning', position='top', timeout=6000)

                # Hromadné nastavení na všechny označené řádky (jako u stavu).
                try:
                    sel_ids = await ui.run_javascript(
                        f'const c=getElement({grid.id});'
                        "return (c&&c.run_grid_method)?"
                        "c.run_grid_method('getSelectedRows').map(r=>r.id):[];",
                        timeout=5,
                    )
                except Exception:
                    sel_ids = []
                sel_ids = {str(x) for x in (sel_ids or []) if x is not None}

                cilove_ids = {str(rid)}
                if str(rid) in sel_ids and len(sel_ids) > 1:
                    cilove_ids = sel_ids

                cilove = [r for r in vsechny
                          if str(r.get('id')) in cilove_ids
                          and (_f(r.get('sleva')) or 0.0) != nova_val]
                if not cilove:
                    # Hodnota se nezměnila — jen srovnej zobrazení editované buňky.
                    grid.run_grid_method('applyTransaction', {'update':
                        [r for r in vsechny if str(r.get('id')) == str(rid)]})
                    _obnov_filtr()
                    return

                # NULL místo 0 do DB, ať „bez slevy" zůstane prázdné.
                db_val = nova_val if nova_val else None
                conn = intranet_data.get_db_connection()
                if conn:
                    cur = conn.cursor()
                    cur.executemany(
                        'UPDATE sankce_vystaveni SET sleva=%s WHERE id=%s',
                        [(db_val, r.get('id')) for r in cilove],
                    )
                    conn.commit(); cur.close(); conn.close()

                for r in cilove:
                    zapis_audit('sankce_vystaveni', r.get('row_hash'), r.get('id'),
                                'sleva', r.get('sleva'), db_val, user_id, user_name)
                    r['sleva'] = db_val

                if len(cilove) == 1:
                    intranet_logger.log_activity(user_name, 'Sankce',
                                                 f'Sleva na sankci (k vystavení) #{cilove[0].get("id")} → {nova_val}')
                else:
                    intranet_logger.log_activity(user_name, 'Sankce',
                                                 f'Hromadná sleva na sankci (k vystavení) — {len(cilove)} řádků → {nova_val}')
                    ui.notify(f'Sleva {nova_val} nastavena u {len(cilove)} označených řádků.',
                              type='positive', position='top', timeout=4000)

                grid.run_grid_method('applyTransaction', {'update': cilove})

                # přepočet součtu (Σ i připnutý CELKEM jsou po slevě)
                _d = _zobrazene()
                info.set_text(_info_text(_d))
                _aplikuj_souhrn(_d)     # drž souhrn dodavatelů v synchronu
                celk = [_celkem_row(_d)]
                grid.options['pinnedBottomRowData'] = celk
                grid.run_grid_method('setGridOption', 'pinnedBottomRowData', celk)
                grid.run_grid_method('onFilterChanged')
                _obnov_filtr()

            elif field == 'poznamka':
                if str(nova or '') == str(stara or ''):
                    return
                conn = intranet_data.get_db_connection()
                if conn:
                    cur = conn.cursor()
                    cur.execute('UPDATE sankce_vystaveni SET poznamka=%s WHERE id=%s', (nova, rid))
                    conn.commit(); cur.close(); conn.close()
                zapis_audit('sankce_vystaveni', rh, rid, 'poznamka', stara, nova, user_id, user_name)
                for r in vsechny:
                    if r.get('id') == rid:
                        r['poznamka'] = nova
                        break
                intranet_logger.log_activity(user_name, 'Sankce', f'Poznámka (k vystavení) #{rid}')
        grid.on('cellValueChanged', _on_change)


# =========================================================
# TIKETY (předání sankcí nákupčímu / provozu)
# =========================================================
# Tiket = balík řádků „Sankce k vystavení", který účtárna předá k rozhodnutí.
# Seskupení: dodavatel + kód nákupčího (provozní tiket kód neřeší → 1 na dodavatele).
# Data řádků mění až POTVRZENÉ rozhodnutí; storno navíc čeká na druhotnou kontrolu.
TIKET_TYP_LABEL = {'nakup': 'Nákup', 'provoz': 'Provoz'}
TIKET_STAV_LABEL = {
    'nakup':         'U nákupčího',
    'provoz':        'U provozu',
    'castecne':      'Částečně rozhodnuto',
    'vyfakturovano': 'K fakturaci',
    'storno_ceka':   'Storno – ke kontrole',
    'storno':        'Stornováno',
    'abnormalita':   'Abnormalita',
    'uzavreno':      'Uzavřeno',
}
# „Otevřené" = někdo na nich ještě má něco udělat (výchozí filtr seznamu tiketů).
TIKET_STAV_OTEVRENE = ('nakup', 'provoz', 'castecne', 'storno_ceka', 'abnormalita')
ROZ_LABEL = {'vyfakturovat': 'Fakturovat', 'storno': 'Stornovat', 'provoz': 'Na provoz'}
ROZ_LABEL_REV = {v: k for k, v in ROZ_LABEL.items()}
# Kódy nákupčích z listu DATA („Nákupčí (pob.)") → individuální právo na tikety.
KODY_NAKUPCI = ['DR', 'SK', 'CK', 'VI', 'LT', 'NP', 'RD', 'HV', 'UP', 'KO', 'ML', 'OZ', 'VN']
KOD_PRAVO = {k: 'sankce_tiket_' + k.lower() for k in KODY_NAKUPCI}

_SANKCE_URL = 'https://analytikasys.jip-napoje.cz/sankce'


def _radku(n: int) -> str:
    """„1 řádek" / „3 řádky" / „5 řádků" — česky do e-mailů."""
    n = int(n or 0)
    return f'{n} ' + ('řádek' if n == 1 else 'řádky' if 2 <= n <= 4 else 'řádků')


_TIKET_STAV_STYLE = (
    "function(p){"
    "var v=p.value;"
    "if(v==='U nákupčího')return{backgroundColor:'#fef9c3',color:'#854d0e',fontWeight:'600'};"
    "if(v==='U provozu')return{backgroundColor:'#ede9fe',color:'#5b21b6',fontWeight:'600'};"
    "if(v==='Částečně rozhodnuto')return{backgroundColor:'#ccfbf1',color:'#115e59',fontWeight:'600'};"
    "if(v==='K fakturaci')return{backgroundColor:'#dbeafe',color:'#1e40af',fontWeight:'600'};"
    "if(v==='Storno – ke kontrole')return{backgroundColor:'#ffedd5',color:'#9a3412',fontWeight:'600'};"
    "if(v==='Stornováno')return{backgroundColor:'#fee2e2',color:'#991b1b',fontWeight:'600'};"
    "if(v==='Abnormalita')return{backgroundColor:'#fae8ff',color:'#86198f',fontWeight:'600'};"
    "if(v==='Uzavřeno')return{backgroundColor:'#dcfce7',color:'#166534',fontWeight:'600'};"
    "return null;}"
)
_ROZ_STYLE = (
    "function(p){"
    "var v=p.value;"
    "if(v==='Fakturovat')return{backgroundColor:'#dbeafe',color:'#1e40af',fontWeight:'600'};"
    "if(v==='Stornovat')return{backgroundColor:'#fee2e2',color:'#991b1b',fontWeight:'600'};"
    "if(v==='Na provoz')return{backgroundColor:'#ede9fe',color:'#5b21b6',fontWeight:'600'};"
    "return null;}"
)


# ── čistá logika (bez DB — testovatelná) ────────────────────────────────────
def _kod_nakupci(r) -> str:
    """Kód nákupčího řádku (sloupec „Nákupčí (pob.)"). Vezme první token, který
    odpovídá číselníku; jinak vrátí hodnotu tak, jak je — ať se řádky různých
    nákupčích neslijí do jednoho tiketu."""
    v = _s(r.get('nakupci_pob')).upper()
    if not v:
        return ''
    for tok in re.split(r'[^A-Z0-9]+', v):
        if tok in KOD_PRAVO:
            return tok
    return v[:20]


def _seskup_do_tiketu(radky: list, typ: str) -> OrderedDict:
    """Rozdělení řádků do tiketů: klíč = (IČO, dodavatel, kód nákupčího).
    Dodavatel s řádky DR i CK dostane 2 nákupní tikety, každý svému nákupčímu.
    Provozní tiket kód neřeší → jeden na dodavatele."""
    skup = OrderedDict()
    for r in radky:
        kod = _kod_nakupci(r) if typ == 'nakup' else ''
        skup.setdefault((_s(r.get('ico')), _s(r.get('jmeno_dodavatele')), kod), []).append(r)
    return skup


def _tiket_stav_z_rozhodnuti(hodnoty) -> str:
    """Stav tiketu po odeslání: jednotné rozhodnutí → odpovídající stav, různá
    rozhodnutí → „částečně". Bez rozhodnutí → None (nic se neposílá)."""
    ruzna = {h for h in hodnoty if h}
    if not ruzna:
        return None
    if len(ruzna) > 1:
        return 'castecne'
    return {'vyfakturovat': 'vyfakturovano', 'storno': 'storno_ceka',
            'provoz': 'provoz'}[ruzna.pop()]


def _tiket_cislo(tid) -> str:
    return f'T{int(tid):05d}'


def _tiket_rh(tid) -> str:
    """Klíč vlákna diskuze i historie tiketu (sdílí ho všechny jeho řádky)."""
    return f'T{int(tid)}'


def _tiket_prava(t: dict) -> tuple:
    """Kdo tiket zrovna řeší (adresáti notifikací)."""
    if _s(t.get('stav')) == 'storno_ceka':
        return ('sankce_tiket_kontrola',)
    if _s(t.get('typ')) == 'provoz':
        return ('sankce_tiket_provoz',)
    p = KOD_PRAVO.get(_s(t.get('kod_nakupci')).upper())
    return (p,) if p else ()


def _viditelne_tikety(tikety: list, vsechna_prava) -> list:
    """Účtárna, analytik, čtenář a admin vidí všechno; ostatní jen tikety, které
    jsou směrované na jejich právo (nákupčí vidí i své storno u kontroly)."""
    if {'vse', 'sankce_ucetni', 'sankce_analytik', 'sankce_ctenar'} & set(vsechna_prava):
        return list(tikety)
    out = []
    for t in tikety:
        prava = set()
        kod_p = KOD_PRAVO.get(_s(t.get('kod_nakupci')).upper())
        if kod_p:
            prava.add(kod_p)
        if _s(t.get('typ')) == 'provoz':
            prava.add('sankce_tiket_provoz')
        if _s(t.get('stav')) == 'storno_ceka':
            prava.add('sankce_tiket_kontrola')
        if prava & set(vsechna_prava):
            out.append(t)
    return out


# ── DB vrstva ───────────────────────────────────────────────────────────────
def _zapis_audit_bulk(zaznamy: list):
    """Hromadný zápis do sankce_audit jedním spojením. Položka =
    (tabulka, row_hash, radek_id, pole, stara, nova, user_id, jmeno)."""
    if not zaznamy:
        return
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.executemany(
            'INSERT INTO sankce_audit '
            '(tabulka,row_hash,radek_id,pole,stara_hodnota,nova_hodnota,user_id,jmeno) '
            'VALUES (%s,%s,%s,%s,%s,%s,%s,%s)',
            [(t, rh, rid, pole,
              None if st is None else str(st), None if no is None else str(no),
              uid, jm or '') for (t, rh, rid, pole, st, no, uid, jm) in zaznamy])
        conn.commit(); cur.close()
    except Exception as e:
        print(f'[sankce] _zapis_audit_bulk error: {e}')
    finally:
        conn.close()


def _nastav_stav_radku(ids: list, stav: str):
    """Hromadná změna stavu řádků sankce_vystaveni (bez auditu — ten píše volající)."""
    ids = [i for i in ids if i]
    if not ids:
        return
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.executemany('UPDATE sankce_vystaveni SET stav=%s WHERE id=%s',
                        [(stav, i) for i in ids])
        conn.commit(); cur.close()
    finally:
        conn.close()


def _zaloz_tikety(radky: list, typ: str, poznamka: str, user_id, user_name,
                  zdroj_tiket=None) -> list:
    """Založí tikety nad označenými řádky a přepne jejich data na Nákup/Provoz.
    Vrací [(tiket_id, kod_nakupci, jmeno_dodavatele, pocet_radku)]."""
    skup = _seskup_do_tiketu(radky, typ)
    if not skup:
        return []
    novy_stav = 'nakup' if typ == 'nakup' else 'provoz'
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    zalozene, audit = [], []
    try:
        cur = conn.cursor()
        for (ico, dod, kod), rows in skup.items():
            obd = sorted({_s(r.get('obdobi')) for r in rows if _s(r.get('obdobi'))})
            obd_txt = obd[0] if len(obd) == 1 else (f'{len(obd)} období' if obd else '')
            cur.execute(
                'INSERT INTO sankce_tikety '
                '(typ,stav,kod_nakupci,ico,jmeno_dodavatele,obdobi,poznamka,predal,zdroj_tiket) '
                'VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)',
                (typ, novy_stav, kod, ico, dod, obd_txt, poznamka or '', user_name, zdroj_tiket))
            tid = cur.lastrowid
            cur.executemany(
                'INSERT IGNORE INTO sankce_tiket_radky (tiket_id,radek_id,row_hash) '
                'VALUES (%s,%s,%s)', [(tid, r.get('id'), r.get('row_hash')) for r in rows])
            cur.executemany('UPDATE sankce_vystaveni SET stav=%s WHERE id=%s',
                            [(novy_stav, r.get('id')) for r in rows])
            conn.commit()
            cislo = _tiket_cislo(tid)
            audit.append(('sankce_tikety', _tiket_rh(tid), tid, 'tiket_stav',
                          None, TIKET_STAV_LABEL[novy_stav], user_id, user_name))
            for r in rows:
                audit.append(('sankce_vystaveni', r.get('row_hash'), r.get('id'), 'stav',
                              STAV_LABEL.get(r.get('stav'), r.get('stav')),
                              STAV_LABEL[novy_stav], user_id, user_name))
                audit.append(('sankce_vystaveni', r.get('row_hash'), r.get('id'), 'tiket',
                              None, f'{cislo} ({TIKET_TYP_LABEL.get(typ, typ)})',
                              user_id, user_name))
            zalozene.append((tid, kod, dod, len(rows)))
        cur.close()
    except Exception as e:
        print(f'[sankce] _zaloz_tikety error: {e}')
    finally:
        conn.close()
    _zapis_audit_bulk(audit)
    return zalozene


def _nacti_tikety() -> list:
    """Seznam tiketů s dopočtem počtu položek a částky (Σ sankce po slevě).
    Částka se počítá živě z dat — po změně slevy sedí i ve starém tiketu."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT t.*,
              (SELECT COUNT(*) FROM sankce_tiket_radky r WHERE r.tiket_id=t.id) AS pocet_radku,
              (SELECT COALESCE(SUM(v.hodn_sankce*(1-COALESCE(v.sleva,0))),0)
                 FROM sankce_tiket_radky r JOIN sankce_vystaveni v ON v.id=r.radek_id
                WHERE r.tiket_id=t.id) AS castka
            FROM sankce_tikety t ORDER BY t.id DESC
        """)
        rows = cur.fetchall()
        for t in rows:
            t['cislo'] = _tiket_cislo(t['id'])
            t['stav_label'] = TIKET_STAV_LABEL.get(t.get('stav'), t.get('stav') or '')
            t['typ_label'] = TIKET_TYP_LABEL.get(t.get('typ'), t.get('typ') or '')
            t['castka'] = float(t.get('castka') or 0)
            for k in ('predano_at', 'zmeneno'):
                d = t.pop(k, None)
                t[k + '_txt'] = d.strftime('%d.%m.%Y %H:%M') if hasattr(d, 'strftime') else ''
        return rows
    finally:
        conn.close()


def _nacti_tiket_radky(tid) -> list:
    """Položky tiketu i s aktuálními daty řádku (řádek smazaný z dat se přeskočí)."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute('SELECT v.*, r.id AS tr_id, r.rozhodnuti FROM sankce_tiket_radky r '
                    'JOIN sankce_vystaveni v ON v.id=r.radek_id '
                    'WHERE r.tiket_id=%s ORDER BY r.id', (tid,))
        rows = cur.fetchall()
        for r in rows:
            for k in ('obdobi_od', 'obdobi_do', 'import_at', 'imported_by'):
                r.pop(k, None)
            r['stav_label'] = STAV_LABEL.get(r.get('stav'), 'Nová data')
            r['rozhodnuti_label'] = ROZ_LABEL.get(r.get('rozhodnuti'), '')
        return rows
    finally:
        conn.close()


def _mapa_tiketu_radku() -> dict:
    """{row_hash: tiket_id} — poslední (nejnovější) tiket řádku vyhrává. Slouží
    ke sdílení vlákna diskuze mezi řádkem sestavy a jeho tiketem."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return {}
    try:
        cur = conn.cursor()
        cur.execute('SELECT row_hash, MAX(tiket_id) FROM sankce_tiket_radky '
                    'WHERE row_hash IS NOT NULL GROUP BY row_hash')
        return {r[0]: r[1] for r in cur.fetchall()}
    finally:
        conn.close()


def _uloz_rozhodnuti(dvojice: list):
    """Uloží rozhodnutí k položkám tiketu. Položka = (rozhodnuti|None, tr_id)."""
    if not dvojice:
        return
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.executemany('UPDATE sankce_tiket_radky SET rozhodnuti=%s WHERE id=%s', dvojice)
        conn.commit(); cur.close()
    finally:
        conn.close()


def _uprav_tiket(tid, stav=None, poznamka=None):
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        if stav is not None:
            cur.execute('UPDATE sankce_tikety SET stav=%s WHERE id=%s', (stav, tid))
        if poznamka is not None:
            cur.execute('UPDATE sankce_tikety SET poznamka=%s WHERE id=%s', (poznamka, tid))
        conn.commit(); cur.close()
    finally:
        conn.close()


def _notifikuj_tiket(prava, predmet: str, text: str):
    """Zvoneček + e-mail všem s daným právem. Běží mimo event loop (DB + SMTP).
    Pozdrav, odkaz do modulu a patičku doplní tahle funkce — volající píše jen věc."""
    prava = tuple(p for p in prava if p)
    if not prava:
        return
    telo = (f'Dobrý den,\n\n{text}\n\n'
            f'Otevřít modul Sankce → Tikety: {_SANKCE_URL}\n\n'
            'Tato zpráva byla odeslána automaticky systémem MojeJIPka.')
    try:
        for uid in (intranet_data.ziskej_uzivatele_s_pravem(*prava) or {}):
            try:
                intranet_notifikace.pridej(uid, predmet, 'info')
            except Exception:
                pass
        for em in (intranet_data.ziskej_emaily_s_pravem(*prava) or []):
            try:
                intranet_emaily.odesli_upozorneni_email(em, predmet, telo)
            except Exception as e:
                print(f'[sankce] mail tiketu ({em}): {e}')
    except Exception as e:
        print(f'[sankce] _notifikuj_tiket error: {e}')


def _odesli_rozhodnuti(tiket: dict, radky: list, user_id, user_name) -> tuple:
    """Odeslání rozhodnutí řešitelem. „Fakturovat" a „Na provoz" se promítnou do
    dat hned (na provoz navíc vznikne navazující provozní tiket), „Stornovat"
    čeká na palec druhotné kontroly. Vrací (novy_stav_tiketu, hlaska)."""
    tid = tiket.get('id')
    novy = _tiket_stav_z_rozhodnuti([r.get('rozhodnuti') for r in radky])
    if not novy:
        return None, 'Není vyplněné žádné rozhodnutí.'
    rh, cislo = _tiket_rh(tid), _tiket_cislo(tid)
    _uloz_rozhodnuti([(r.get('rozhodnuti') or None, r.get('tr_id')) for r in radky])

    fakt = [r for r in radky if r.get('rozhodnuti') == 'vyfakturovat']
    prov = [r for r in radky if r.get('rozhodnuti') == 'provoz']
    stor = [r for r in radky if r.get('rozhodnuti') == 'storno']

    audit = [('sankce_tikety', rh, tid, 'tiket_stav',
              TIKET_STAV_LABEL.get(tiket.get('stav')), TIKET_STAV_LABEL[novy],
              user_id, user_name)]
    for r in radky:
        if r.get('rozhodnuti'):
            audit.append(('sankce_tikety', rh, tid, 'rozhodnuti', None,
                          f"{r.get('nase_cislo') or r.get('id')} → "
                          f"{ROZ_LABEL[r['rozhodnuti']]}", user_id, user_name))
    if fakt:
        _nastav_stav_radku([r.get('id') for r in fakt], 'vyfakturovano')
        for r in fakt:
            audit.append(('sankce_vystaveni', r.get('row_hash'), r.get('id'), 'stav',
                          STAV_LABEL.get(r.get('stav')), STAV_LABEL['vyfakturovano'],
                          user_id, user_name))
    _zapis_audit_bulk(audit)

    if prov:
        _zaloz_tikety(prov, 'provoz', f'Postoupeno z tiketu {cislo}',
                      user_id, user_name, zdroj_tiket=tid)
    _uprav_tiket(tid, stav=novy)

    dod = _s(tiket.get('jmeno_dodavatele'))
    if fakt:
        _notifikuj_tiket(('vse', 'sankce_ucetni'),
                         f'Sankce – {cislo} k fakturaci ({dod})',
                         f'{user_name} rozhodl(a) o fakturaci {_radku(len(fakt))} '
                         f'dodavatele {dod}. Řádky jsou ve stavu „Fakturovat".')
    if prov:
        _notifikuj_tiket(('sankce_tiket_provoz',),
                         f'Sankce – nový tiket pro provoz ({dod})',
                         f'{user_name} postoupil(a) {_radku(len(prov))} dodavatele {dod} '
                         f'na provoz (z tiketu {cislo}).')
    if stor:
        _notifikuj_tiket(('sankce_tiket_kontrola',),
                         f'Sankce – {cislo} ke schválení storna ({dod})',
                         f'{user_name} navrhuje stornovat {_radku(len(stor))} dodavatele '
                         f'{dod}. Bez vašeho schválení se data nemění.')
    return novy, (f'Rozhodnutí odesláno: {TIKET_STAV_LABEL[novy]} '
                  f'(fakturovat {len(fakt)}, storno {len(stor)}, provoz {len(prov)}).')


def _palec_storno(tiket: dict, radky: list, schvaleno: bool, user_id, user_name) -> str:
    """Druhotná kontrola storna. Schváleno → řádky přejdou na „Stornovat" a jde
    mail účtárně. Zamítnuto → data se nemění a tiket se vrací řešiteli."""
    tid = tiket.get('id')
    rh, cislo = _tiket_rh(tid), _tiket_cislo(tid)
    stor = [r for r in radky if r.get('rozhodnuti') == 'storno']
    dod = _s(tiket.get('jmeno_dodavatele'))
    if schvaleno:
        _nastav_stav_radku([r.get('id') for r in stor], 'storno')
        audit = [('sankce_vystaveni', r.get('row_hash'), r.get('id'), 'stav',
                  STAV_LABEL.get(r.get('stav')), STAV_LABEL['storno'], user_id, user_name)
                 for r in stor]
        audit.append(('sankce_tikety', rh, tid, 'tiket_stav',
                      TIKET_STAV_LABEL.get(tiket.get('stav')), TIKET_STAV_LABEL['storno'],
                      user_id, user_name))
        _zapis_audit_bulk(audit)
        _uprav_tiket(tid, stav='storno')
        _notifikuj_tiket(('vse', 'sankce_ucetni'),
                         f'Sankce – {cislo} storno schváleno ({dod})',
                         f'{user_name} schválil(a) storno {_radku(len(stor))} dodavatele '
                         f'{dod}. Řádky jsou ve stavu „Stornovat".')
        return f'Storno schváleno — {_radku(len(stor))} přešlo na „Stornovat".'

    zpet = 'provoz' if _s(tiket.get('typ')) == 'provoz' else 'nakup'
    _uloz_rozhodnuti([(None, r.get('tr_id')) for r in stor])
    _uprav_tiket(tid, stav=zpet)
    _zapis_audit_bulk([('sankce_tikety', rh, tid, 'tiket_stav',
                        TIKET_STAV_LABEL.get(tiket.get('stav')), TIKET_STAV_LABEL[zpet],
                        user_id, user_name)])
    _notifikuj_tiket(_tiket_prava({'typ': tiket.get('typ'), 'stav': zpet,
                                   'kod_nakupci': tiket.get('kod_nakupci')}),
                     f'Sankce – {cislo} storno vráceno k přepracování ({dod})',
                     f'{user_name} storno neschválil(a). Tiket {cislo} je zpět u vás '
                     f'a data se nezměnila. Důvod najdete v diskuzi tiketu.')
    return 'Storno zamítnuto — tiket se vrátil řešiteli, data se nezměnila.'


# ── UI: seznam tiketů + detail ──────────────────────────────────────────────
def _col_defs_tikety() -> list:
    return [
        _col_chat(),
        {'headerName': '', 'field': '_eye', 'width': 46, 'minWidth': 46, 'maxWidth': 46,
         'pinned': 'left', 'sortable': False, 'editable': False, 'resizable': False,
         'filter': False, 'suppressSizeToFit': True, 'suppressAutoSize': True,
         ':cellRenderer': _EYE_RENDERER,
         'cellStyle': {'textAlign': 'center', 'cursor': 'pointer', 'padding': '0'},
         'headerTooltip': 'Historie tiketu'},
        {'headerName': 'Tiket', 'field': 'cislo', 'width': 100, 'pinned': 'left',
         'sortable': True, 'cellStyle': {'fontFamily': 'monospace', 'fontWeight': '600',
                                         'color': '#334155'}},
        {'headerName': 'Stav', 'field': 'stav_label', 'width': 180, 'sortable': True,
         ':cellStyle': _TIKET_STAV_STYLE},
        {'headerName': 'Směr', 'field': 'typ_label', 'width': 95, 'sortable': True},
        {'headerName': 'Nákupčí', 'field': 'kod_nakupci', 'width': 100, 'sortable': True,
         'cellStyle': {'fontFamily': 'monospace'},
         'headerTooltip': 'Kód nákupčího (vlastník tiketu)'},
        {'headerName': 'IČO', 'field': 'ico', 'width': 100, 'sortable': True,
         'cellStyle': {'fontFamily': 'monospace', 'fontSize': '12px', 'color': '#475569'}},
        {'headerName': 'Dodavatel', 'field': 'jmeno_dodavatele', 'width': 280, 'sortable': True,
         'cellStyle': {'fontWeight': '600'}},
        {'headerName': 'Období', 'field': 'obdobi', 'width': 140, 'sortable': True},
        {'headerName': 'Položek', 'field': 'pocet_radku', 'width': 100, 'sortable': True,
         'type': 'numericColumn', ':valueFormatter': _NUM_FMT},
        {'headerName': 'Sankce celkem', 'field': 'castka', 'width': 150, 'sortable': True,
         'type': 'numericColumn', ':valueFormatter': _MONEY_FMT,
         'cellStyle': {'fontWeight': 'bold'},
         'headerTooltip': 'Σ Hodn. sankce po slevě (počítá se živě z dat)'},
        {'headerName': 'Předáno dne', 'field': 'predano_at_txt', 'width': 140, 'sortable': True},
        {'headerName': 'Předal', 'field': 'predal', 'width': 170, 'sortable': True},
        {'headerName': 'Poznámka', 'field': 'poznamka', 'width': 240, 'sortable': True},
    ]


def _col_defs_tiket_radky(volby: list, editovatelne: bool) -> list:
    return [
        {'headerName': 'Poř. č.', 'field': 'nase_cislo', 'width': 104, 'pinned': 'left',
         'cellStyle': {'fontFamily': 'monospace', 'fontSize': '12px'}},
        {'headerName': 'Kód zboží', 'field': 'kod_zbozi', 'width': 110,
         'cellStyle': {'fontFamily': 'monospace', 'fontSize': '12px'}},
        {'headerName': 'Název zboží', 'field': 'nazev_zbozi', 'width': 260},
        {'headerName': 'Č. objednávky', 'field': 'cislo_objednavky', 'width': 120},
        {'headerName': 'Období', 'field': 'obdobi', 'width': 120},
        {'headerName': 'Dod. pozdě MJ', 'field': 'dod_pozde_mj', 'width': 120,
         'type': 'numericColumn', ':valueFormatter': _NUM_FMT},
        {'headerName': 'Hodn. sankce', 'field': 'hodn_sankce', 'width': 130,
         'type': 'numericColumn', ':valueFormatter': _MONEY_FMT,
         'cellStyle': {'fontWeight': 'bold'}},
        {'headerName': 'Stav dat', 'field': 'stav_label', 'width': 150,
         ':cellStyle': _STAV_STYLE, 'editable': False},
        {'headerName': 'Rozhodnutí', 'field': 'rozhodnuti_label', 'width': 150,
         'editable': editovatelne, 'cellEditor': 'agSelectCellEditor',
         'cellEditorParams': {'values': volby}, ':cellStyle': _ROZ_STYLE,
         'headerTooltip': 'Co s řádkem uděláme (klikem změníte)'},
    ]


def _otevri_tiket(t: dict, user_id, user_name, vsechna_prava, refresh):
    """Detail tiketu: položky s rozhodnutím + akce podle role (řešitel / druhotná
    kontrola / účtárna). Data se mění výhradně přes tlačítka tady."""
    tid = t.get('id')
    rh, cislo = _tiket_rh(tid), _tiket_cislo(tid)
    stav = _s(t.get('stav'))
    typ = _s(t.get('typ'))
    dod = _s(t.get('jmeno_dodavatele'))
    ma_vse = 'vse' in vsechna_prava
    je_ucetni = ma_vse or 'sankce_ucetni' in vsechna_prava
    resitel = ('sankce_tiket_provoz',) if typ == 'provoz' else \
              (KOD_PRAVO.get(_s(t.get('kod_nakupci')).upper()),)
    muze_resit = ma_vse or bool({p for p in resitel if p} & set(vsechna_prava))
    muze_palec = ma_vse or 'sankce_tiket_kontrola' in vsechna_prava
    rozhoduje_se = stav in ('nakup', 'provoz', 'castecne')
    editovatelne = muze_resit and rozhoduje_se

    radky = _nacti_tiket_radky(tid)
    volby = ['', ROZ_LABEL['vyfakturovat'], ROZ_LABEL['storno']]
    if typ == 'nakup':
        volby.append(ROZ_LABEL['provoz'])

    with ui.dialog() as dlg, ui.card().classes('p-0 rounded-2xl gap-0') \
            .style('min-width:960px;max-width:96vw'):
        with ui.row().classes('items-center gap-3 px-5 pt-4 pb-2 w-full'):
            ui.icon('confirmation_number', color='primary').classes('text-2xl')
            with ui.column().classes('gap-0'):
                ui.label(f'Tiket {cislo} — {dod or "dodavatel"}') \
                    .classes('text-lg font-bold text-gray-800')
                ui.label(f'{TIKET_TYP_LABEL.get(typ, typ)}'
                         + (f' / {t.get("kod_nakupci")}' if t.get('kod_nakupci') else '')
                         + f' · období {t.get("obdobi") or "—"}'
                         + f' · předal {t.get("predal") or "—"} {t.get("predano_at_txt") or ""}') \
                    .classes('text-xs text-gray-500')
            ui.space()
            ui.label(TIKET_STAV_LABEL.get(stav, stav)).classes(
                'px-3 py-1 rounded-lg text-sm font-bold bg-gray-100 text-gray-700')
            ui.button(icon='close', on_click=dlg.close).props('flat round dense color=grey-7')
        if _s(t.get('poznamka')):
            ui.label(f'📝 {t.get("poznamka")}').classes('text-xs text-gray-500 px-5 pb-1')
        ui.separator()

        g = ui.aggrid({
            'columnDefs': _col_defs_tiket_radky(volby, editovatelne),
            'rowData': radky,
            'defaultColDef': {'resizable': True, 'sortable': True, 'filter': True},
            'rowHeight': 32,
            'singleClickEdit': True,
            'stopEditingWhenCellsLoseFocus': True,
            ':getRowId': "function(p){var d=p.data||{};return ''+(d.tr_id!=null?d.tr_id:'');}",
        }).classes('w-full').style('height:46vh')

        def _on_roz(e):
            a = e.args or {}
            if (a.get('colId') or '') != 'rozhodnuti_label':
                return
            d = a.get('data') or {}
            kod = ROZ_LABEL_REV.get(a.get('newValue') or '')
            for r in radky:
                if r.get('tr_id') == d.get('tr_id'):
                    r['rozhodnuti'] = kod
                    r['rozhodnuti_label'] = ROZ_LABEL.get(kod, '')
        g.on('cellValueChanged', _on_roz)

        def _vse(kod):
            for r in radky:
                r['rozhodnuti'] = kod
                r['rozhodnuti_label'] = ROZ_LABEL.get(kod, '')
            g.run_grid_method('setGridOption', 'rowData', radky)

        async def _odeslat():
            novy, hlaska = await asyncio.to_thread(
                _odesli_rozhodnuti, t, radky, user_id, user_name)
            if not novy:
                ui.notify(hlaska, type='warning', position='top', timeout=6000)
                return
            intranet_logger.log_activity(user_name, 'Sankce',
                                         f'Tiket {cislo}: rozhodnutí → {TIKET_STAV_LABEL[novy]}')
            ui.notify(hlaska, type='positive', position='top', timeout=6000, multi_line=True)
            dlg.close()
            refresh()

        async def _palec(schvaleno: bool):
            hlaska = await asyncio.to_thread(
                _palec_storno, t, radky, schvaleno, user_id, user_name)
            intranet_logger.log_activity(
                user_name, 'Sankce',
                f'Tiket {cislo}: storno {"schváleno" if schvaleno else "zamítnuto"}')
            ui.notify(hlaska, type='positive' if schvaleno else 'warning',
                      position='top', timeout=7000, multi_line=True)
            dlg.close()
            refresh()

        def _abnormalita():
            """Netypický případ — tiket se odloží a data zůstanou beze změny."""
            with ui.dialog() as d2, ui.card().classes('p-5 gap-3').style('min-width:520px'):
                ui.label('Označit případ jako abnormalitu').classes('text-lg font-bold')
                ui.label('Data se nezmění. Popište, co je na případu nestandardní — '
                         'text se uloží do diskuze tiketu.').classes('text-sm text-gray-500')
                duvod = ui.textarea(placeholder='Důvod…').props('outlined autogrow') \
                    .classes('w-full')

                async def _ok():
                    txt = (duvod.value or '').strip()
                    if not txt:
                        ui.notify('Vyplňte důvod.', type='warning', position='top')
                        return

                    def _prace():
                        _uprav_tiket(tid, stav='abnormalita')
                        _pridej_chat('sankce_tikety', rh, user_id, user_name,
                                     f'⚠️ Abnormalita: {txt}')
                        _zapis_audit_bulk([('sankce_tikety', rh, tid, 'tiket_stav',
                                            TIKET_STAV_LABEL.get(stav),
                                            TIKET_STAV_LABEL['abnormalita'],
                                            user_id, user_name)])
                        _notifikuj_tiket(('vse', 'sankce_ucetni', 'sankce_analytik'),
                                         f'Sankce – {cislo} označen jako abnormalita ({dod})',
                                         f'{user_name}: {txt}')
                    await asyncio.to_thread(_prace)
                    intranet_logger.log_activity(user_name, 'Sankce',
                                                 f'Tiket {cislo}: abnormalita')
                    ui.notify('Tiket označen jako abnormalita, data beze změny.',
                              type='warning', position='top', timeout=6000)
                    d2.close(); dlg.close(); refresh()

                with ui.row().classes('w-full justify-end gap-2'):
                    ui.button('Zrušit', on_click=d2.close).props('flat no-caps')
                    ui.button('Označit', on_click=_ok).props('unelevated color=warning no-caps')
            d2.open()

        async def _uzavrit():
            await asyncio.to_thread(_uprav_tiket, tid, 'uzavreno')
            await asyncio.to_thread(_zapis_audit_bulk, [
                ('sankce_tikety', rh, tid, 'tiket_stav', TIKET_STAV_LABEL.get(stav),
                 TIKET_STAV_LABEL['uzavreno'], user_id, user_name)])
            intranet_logger.log_activity(user_name, 'Sankce', f'Tiket {cislo}: uzavřen')
            ui.notify('Tiket uzavřen.', type='positive', position='top-right', timeout=4000)
            dlg.close(); refresh()

        ui.separator()
        with ui.row().classes('w-full items-center gap-2 px-4 py-3 flex-wrap'):
            ui.button('Historie', icon='history',
                      on_click=lambda: _zobraz_historii('sankce_tikety', rh,
                                                        f'Tiket {cislo} — {dod}')) \
                .props('outline color=teal dense no-caps')
            ui.button('Diskuze', icon='forum',
                      on_click=lambda: _otevri_chat(
                          'sankce_tikety', rh, f'Tiket {cislo} — {dod}', user_id, user_name,
                          lambda: None, muze_mazat_vse=je_ucetni,
                          prava_navic=_tiket_prava(t))) \
                .props('outline color=primary dense no-caps')
            ui.space()
            if editovatelne:
                ui.button('Vše fakturovat', icon='done_all',
                          on_click=lambda: _vse('vyfakturovat')) \
                    .props('outline color=blue-8 dense no-caps')
                ui.button('Vše stornovat', icon='block',
                          on_click=lambda: _vse('storno')) \
                    .props('outline color=red-7 dense no-caps')
                if typ == 'nakup':
                    ui.button('Vše na provoz', icon='engineering',
                              on_click=lambda: _vse('provoz')) \
                        .props('outline color=purple-7 dense no-caps')
                ui.button('Abnormalita', icon='report_problem', on_click=_abnormalita) \
                    .props('outline color=orange-8 dense no-caps') \
                    .tooltip('Netypický případ — data se nemění, řeší účtárna')
                ui.button('Odeslat rozhodnutí', icon='send', on_click=_odeslat) \
                    .props('unelevated no-caps') \
                    .classes('bg-emerald-600 hover:bg-emerald-700 text-white font-semibold '
                             'rounded-lg shadow-md px-5')
            if muze_palec and stav == 'storno_ceka':
                ui.button('Vrátit řešiteli', icon='thumb_down',
                          on_click=lambda: _palec(False)) \
                    .props('outline color=red-7 dense no-caps') \
                    .tooltip('Storno neschváleno — data se nemění')
                ui.button('Schválit storno', icon='thumb_up',
                          on_click=lambda: _palec(True)) \
                    .props('unelevated no-caps') \
                    .classes('bg-emerald-600 hover:bg-emerald-700 text-white font-semibold '
                             'rounded-lg shadow-md px-5')
            if je_ucetni and stav in ('vyfakturovano', 'storno', 'abnormalita'):
                ui.button('Uzavřít tiket', icon='task_alt', on_click=_uzavrit) \
                    .props('unelevated color=green-8 no-caps')
            if not (editovatelne or (muze_palec and stav == 'storno_ceka')):
                ui.label('Tiket je jen ke čtení (není ve vaší frontě).') \
                    .classes('text-xs text-gray-400')
    dlg.open()


@refreshable_na_klienta
async def _vykresli_tikety(user_id, user_name, vsechna_prava):
    vsechny = _viditelne_tikety(await asyncio.to_thread(_nacti_tikety), vsechna_prava)
    if not vsechny:
        with ui.column().classes('items-center py-16 gap-3 w-full'):
            ui.icon('confirmation_number', size='4rem', color='grey-4')
            ui.label('Žádné tikety ve vaší frontě.').classes('text-xl text-gray-400 font-bold')
            ui.label('Tiket vzniká v sestavě „Sankce k vystavení" tlačítkem „Předat na tiket".') \
                .classes('text-sm text-gray-400')
        return

    chat = await asyncio.to_thread(_nacti_chat_stav, 'sankce_tikety', user_id)
    for t in vsechny:
        st = chat.get(_tiket_rh(t['id']))
        t['_chat_pocet'] = st['pocet'] if st else 0
        t['_chat_unread'] = st['unread'] if st else False

    filtr = {'jen_otevrene': True}

    def _zobrazene():
        if filtr['jen_otevrene']:
            return [t for t in vsechny if t.get('stav') in TIKET_STAV_OTEVRENE]
        return list(vsechny)

    with ui.row().classes('w-full items-center gap-3 mb-2 flex-wrap'):
        sw = ui.switch('Jen otevřené', value=True) \
            .tooltip('Skryje uzavřené a vyřešené tikety (stornováno / k fakturaci / uzavřeno)')
        ui.space()
        info = ui.label('').classes('text-sm text-gray-500')

    grid = ui.aggrid({
        'columnDefs': _col_defs_tikety(),
        'rowData': _zobrazene(),
        'defaultColDef': {'resizable': True, 'sortable': False, 'filter': True},
        'rowHeight': 32,
        'suppressMovableColumns': True,
        ':onFirstDataRendered': _AUTOSIZE_FIT,
        ':onGridSizeChanged': _AUTOSIZE_FIT,
        ':getRowId': _GET_ROW_ID,
    }).classes('w-full').style(_GRID_STYLE)

    def _info(data):
        castka = sum(_f(t.get('castka')) or 0 for t in data)
        return f'Tiketů: {len(data)} · sankce celkem: {_cz_money(castka)}'

    def _aplikuj():
        data = _zobrazene()
        grid.options['rowData'] = data
        grid.run_grid_method('setGridOption', 'rowData', data)
        info.set_text(_info(data))
    sw.on_value_change(lambda e: (filtr.__setitem__('jen_otevrene', bool(e.value)), _aplikuj()))
    info.set_text(_info(_zobrazene()))

    def _chat_badge(t):
        st = _chat_stav_radku('sankce_tikety', _tiket_rh(t['id']), user_id)
        t['_chat_pocet'] = st['pocet']; t['_chat_unread'] = st['unread']
        grid.run_grid_method('applyTransaction', {'update': [t]})

    def _on_click(e):
        a = e.args or {}
        col = a.get('colId')
        d = a.get('data') or {}
        tid = d.get('id')
        if not tid:
            return
        t = next((x for x in vsechny if x.get('id') == tid), None)
        if not t:
            return
        popis = f'Tiket {_tiket_cislo(tid)} — {_s(t.get("jmeno_dodavatele"))}'
        if col == '_eye':
            _zobraz_historii('sankce_tikety', _tiket_rh(tid), popis)
        elif col == '_chat':
            _otevri_chat('sankce_tikety', _tiket_rh(tid), popis, user_id, user_name,
                         lambda: _chat_badge(t),
                         muze_mazat_vse='vse' in vsechna_prava or 'sankce_ucetni' in vsechna_prava,
                         prava_navic=_tiket_prava(t))
        else:
            _otevri_tiket(t, user_id, user_name, vsechna_prava, _vykresli_tikety.refresh)
    grid.on('cellClicked', _on_click)

    ui.label('💡 Klikem na řádek otevřete detail tiketu (rozhodnutí o jednotlivých '
             'položkách). 👁 = historie, 💬 = diskuze — vlákno je sdílené s řádky '
             'sestavy „Sankce k vystavení".').classes('text-xs text-gray-500 mt-1')


# =========================================================
# VSTUPNÍ OBRAZOVKA — DVĚ DLAŽDICE
# =========================================================
def _dlazdice(emoji, nadpis, barva_border, barva_btn, on_click):
    """Klasická čtvercová dlaždice (stejný styl jako na hlavní nástěnce)."""
    with ui.card().classes(
        'w-80 h-72 items-center justify-center shadow-xl hover:scale-105 '
        'transition-transform duration-300 cursor-pointer bg-white rounded-2xl '
        f'border {barva_border}'
    ).on('click', on_click):
        ui.label(emoji).classes('text-7xl mb-6')
        ui.label(nadpis).classes('text-2xl font-bold text-gray-800 mb-4 text-center')
        ui.button('Otevřít', on_click=on_click).classes(
            f'{barva_btn} text-white font-bold py-3 px-8 rounded-lg shadow-md')


@refreshable_na_klienta
async def vykresli_sankce(user_id, user_name, vsechna_prava):
    # DB dotazy běží ve vlákně — nedrží event loop celého serveru
    await asyncio.to_thread(inicializace_sankce_db)
    # Pozn.: globální handler 'sankce_radek_del' se registruje JIŽ při stavbě
    # stránky (intranet.py, před flushem), NE zde — jinak by ui.on() při každém
    # líném renderu / refresh navěsil nový listener na persistentní layout a
    # klient by hlásil „Event listeners changed after initial definition".

    ma_vse = 'vse' in vsechna_prava
    je_ctenar = 'sankce_ctenar' in vsechna_prava
    vidi_vystaveni = ma_vse or je_ctenar or 'sankce_ucetni' in vsechna_prava or 'sankce_analytik' in vsechna_prava
    vidi_zamitnute = ma_vse or je_ctenar or 'sankce_nakup' in vsechna_prava or 'sankce_analytik' in vsechna_prava
    vidi_tikety = vidi_vystaveni or bool(
        {'sankce_tiket_provoz', 'sankce_tiket_kontrola', *KOD_PRAVO.values()} & set(vsechna_prava))

    pohled = app.storage.user.get('sankce_pohled')
    if pohled == 'vystaveni' and not vidi_vystaveni:
        pohled = None
    if pohled == 'zamitnute' and not vidi_zamitnute:
        pohled = None
    if pohled == 'tikety' and not vidi_tikety:
        pohled = None

    # ── Hlavička ──
    with ui.row().classes('w-full items-center gap-3 mb-6'):
        if pohled:
            def _zpet():
                app.storage.user['sankce_pohled'] = None
                vykresli_sankce.refresh()
            ui.button(icon='arrow_back', on_click=_zpet).props('flat round color=grey-7') \
                .tooltip('Zpět na přehled Sankcí')
        ui.icon('gavel', size='2.2rem').classes('text-rose-600')
        with ui.column().classes('gap-0'):
            ui.label('Sankce').classes('text-3xl font-extrabold text-gray-800')
            nadpis = {'zamitnute': 'Zamítnuté dodávky dodavatelem',
                      'vystaveni': 'Sankce k vystavení',
                      'tikety': 'Tikety — rozhodnutí o sankcích'}.get(pohled,
                      'Přehled sankcí vůči dodavatelům')
            ui.label(nadpis).classes('text-sm text-gray-500')

    # ── Obsah ──
    if pohled == 'zamitnute':
        await _vykresli_zamitnute(user_id, user_name, vsechna_prava)
        return
    if pohled == 'vystaveni':
        await _vykresli_vystaveni(user_id, user_name, vsechna_prava)
        return
    if pohled == 'tikety':
        await _vykresli_tikety(user_id, user_name, vsechna_prava)
        return

    if not (vidi_vystaveni or vidi_zamitnute or vidi_tikety):
        with ui.column().classes('items-center py-20 gap-3 w-full'):
            ui.icon('lock', size='4rem', color='grey-4')
            ui.label('Nemáte přístup k žádné sestavě modulu Sankce.').classes('text-lg text-gray-400')
        return

    with ui.row().classes('w-full gap-8 flex-wrap pt-4'):
        if vidi_zamitnute:
            def _otevri_z():
                app.storage.user['sankce_pohled'] = 'zamitnute'
                vykresli_sankce.refresh()
            _dlazdice('🚫', 'Zamítnuté dodávky dodavatelem',
                      'border-orange-200', 'bg-orange-600 hover:bg-orange-700', _otevri_z)
        if vidi_vystaveni:
            def _otevri_v():
                app.storage.user['sankce_pohled'] = 'vystaveni'
                vykresli_sankce.refresh()
            _dlazdice('🧾', 'Sankce k vystavení',
                      'border-rose-200', 'bg-rose-600 hover:bg-rose-700', _otevri_v)
        if vidi_tikety:
            def _otevri_t():
                app.storage.user['sankce_pohled'] = 'tikety'
                vykresli_sankce.refresh()
            _dlazdice('🎫', 'Tikety (nákup / provoz)',
                      'border-emerald-200', 'bg-emerald-600 hover:bg-emerald-700', _otevri_t)
