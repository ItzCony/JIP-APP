"""Modul „Formuláře ASM" — dlaždice „Změna zákazníků na OZ/ASM" + „Data".

Funkčně klon modulu Cenopřípad (IND ceny), přizpůsobený nové doméně: ASM (a
vedoucí poboček) podávají žádost o přeřazení zákazníků na jiného OZ/ASM nebo o
změnu provize. Žádost tvoří frontu, kterou zpracovává Office (odd. správa
zákaznických dat); sporné případy řeší Správce.

Dlaždice:
  • „Změna zákazníků na OZ/ASM" — formulář + fronta + workflow.
  • „Data" — nahrávání číselníků (Dealer.xlsx + Kontaktní údaje VO.xlsb týdně,
    GIST_obrat zakaznik.xlsx měsíčně), nahrává Vkladatel (AO).

Mapování zdrojových dat (potvrzeno):
  • Dealer.xlsx → číselník OZ: Číslo OZ = sloupec „Klíč", Jméno OZ = sloupec
    „Popis" (očištěný); číselník ASM = unikátní hodnoty sloupce „Hodnota".
  • Kontaktní údaje VO.xlsb → po zadání IČO se dotáhnou kontaktní sloupce
    (párování dle NÁZVU v hlavičce); „Původní provize" = 4. znak sloupce „Dealer".
    Nové sloupce: „Limit salda" → Aktuální kreditní limit, „Splatnost" → Splatnost 1,
    „Splatnost2" → Splatnost 2 (prázdné = převzít Splatnost 1). Řádky s „Typ platiče"
    = „vyřazen" se importují také; při shodě IČ 10 má ale přednost aktivní řádek.
  • GIST_obrat zakaznik.xlsx → obraty zákazníků dle IČO (dva bloky: 3 měsíce / 1 měsíc,
    blok se pozná podle popisku nad hlavičkou) → pole Obrat 3 měsíce / Obrat 1 měsíc.
"""
import asyncio
import datetime
import io
import json

from nicegui import ui, app, context

import intranet_data
import intranet_schuzky
import intranet_emaily
import intranet_logger


# ============================================================================
# Konstanty
# ============================================================================
# Stavy workflow → (popisek, barva quasar badge)
_STAV_BADGE = {
    "odeslano":         ("Odesláno / ve frontě", "blue"),
    "vraceno_oprava":   ("Vráceno k opravě", "orange"),
    "vraceno_zpet":     ("Vráceno zpět na Office", "purple"),
    "u_spravce":        ("U správce", "indigo"),
    "spravce_schvalil": ("Správce schválil", "teal"),
    "zpracovano":       ("Zpracováno", "green"),
    "uzavreno":         ("Uzavřeno", "grey-8"),
    "zamitnuto":        ("Zamítnuto", "red"),
    "stornovano":       ("Stornováno", "grey"),
}

# Stavy, kdy je případ uzavřený/ukončený (nelze měnit ani stornovat).
_STAVY_KONCOVE = {"uzavreno", "zamitnuto", "stornovano"}

# Kontaktní sloupce dotahované z asm_kontakty (klíč v DB/řádku, popisek do UI/exportu).
# Pořadí = pořadí zobrazení i pořadí v exportu.
# PŮVODNÍ (K. dodací) sloupce z Kontaktních údajů VO — dlaždice „Změna zákazníků".
_KONTAKT_SLOUPCE = [
    ("jmeno",         "Jméno"),
    ("k_jmeno",       "K.Jméno"),
    ("k_jmeno2",      "K.jméno 2"),
    ("k_ulice",       "K.Ulice"),
    ("k_mesto",       "K.Město"),
    ("k_psc",         "K.PSČ"),
    ("dealer",        "Dealer"),
    ("id_kontakt",    "ID"),
    ("prodejni_doba", "Prodejní doba 6"),
]

# ----------------------------------------------------------------------------
# Konfigurace „unifikovaných" formulářů (Navýšení limitů / Změna dodacích listů).
# Jeden zákazník na případ (žádná tabulka řádků). Data se ukládají do asm_pripady.
# data_json. Pole se renderují i čtou generickou funkcí dle této konfigurace.
#
# typ pole: "ico" (spustí dotažení z kontaktů), "auto" (read-only, dotažené dle IČO),
#           "text", "number", "textarea", "select" (s options).
# ----------------------------------------------------------------------------
_REZIM_FAKTURACE = ["týdenní", "dekádní", "čtrnáctidenní", "měsíční"]

# Společná sekce 1 (identifikace zákazníka) — pro oba nové formuláře.
_SEKCE_IDENT_POLE = [
    ("ico",            "IČO zákazníka",                      "ico"),
    ("pobocka",        "Pobočka (Město)",                    "auto"),
    ("fakt_nazev",     "Fakturační název zákazníka",         "auto"),
    ("provozovna",     "Název místa dodání (Provozovna)",    "auto"),
    ("adresa",         "Adresa místa dodání",                "auto"),
]
_SEKCE_IDENT = ("1. Obecné informace a identifikace", _SEKCE_IDENT_POLE)

# Varianty přesně dle vzorového formuláře (formuláře navýšení limitů…xlsm).
_PRAVNI_FORMA = ["právnická osoba", "fyzická osoba",
                 "příspěvková organizace/státní instituce"]
# Sekce 1 pro „Navýšení limitů a splatností" — navíc Právní forma a Smlouva se zajištěním.
_SEKCE_IDENT_LIMITY = ("1. Obecné informace a identifikace", _SEKCE_IDENT_POLE + [
    ("pravni_forma",      "Právní forma",                     "select", _PRAVNI_FORMA),
    ("smlouva_zajisteni", "Smlouva se zajištěním / ručením",  "select", ["ANO", "NE"]),
])

_FORMULARE = {
    "limity": {
        "nazev": "Navýšení limitů a splatností", "emoji": "💳", "barva": "border-amber-200",
        "sekce": [
            _SEKCE_IDENT_LIMITY,
            ("3. Současný stav a ekonomické ukazatele", [
                ("akt_limit",      "Aktuální kreditní limit",                       "auto"),
                ("splatnost1_akt", "Splatnost 1 (zboží) - aktuální [ve dnech]",     "auto"),
                ("splatnost2_akt", "Splatnost 2 (cigarety) - aktuální [ve dnech]",  "auto"),
                ("obrat_1m",       "Obrat za poslední měsíc (všechny doklady, celé IČ 8)",  "auto"),
                ("obrat_3m",       "Obrat za poslední 3 měsíce (nehotovostní, celé IČ 8)",  "auto"),
                ("dluh",           "Dluh po splatnosti + 7 dnů (saldo)",            "text"),
            ]),
            ("4. Požadovaná změna limitu / splatnosti", [
                ("navrh_kredit",     "Návrh nového kreditu (limitu)",                "text"),
                ("navrh_splatnost1", "Návrh nové splatnosti 1 (zboží) [ve dnech]",   "text"),
                ("navrh_splatnost2", "Návrh nové splatnosti 2 (cigarety) [ve dnech]", "text"),
                ("zduvodneni",       "Zdůvodnění požadované změny / žádosti",         "textarea"),
            ]),
        ],
    },
    "dodaci": {
        "nazev": "Změna splatnosti dodacích listů", "emoji": "🚚", "barva": "border-cyan-200",
        "sekce": [
            _SEKCE_IDENT,
            ("2. Specifické údaje pro Dodací listy (DL)", [
                ("rezim_fakturace", "Režim fakturace",                 "select", _REZIM_FAKTURACE),
                ("el_fakturace",    "Zřízená el. fakturace (ANO/NE)",  "select", ["ANO", "NE"]),
            ]),
            # Spodní část (dle zadání): jen potřebná pole — limit/obraty/dluh/návrh kreditu
            # se zde nevyplňují. Splatnosti se dotahují podle IČO z Kontaktních údajů VO.
            ("3. Současný stav (splatnosti)", [
                ("splatnost1_akt", "Splatnost 1 (zboží) - aktuální [ve dnech]",     "auto"),
                ("splatnost2_akt", "Splatnost 2 (cigarety) - aktuální [ve dnech]",  "auto"),
            ]),
            ("4. Požadovaná změna splatnosti", [
                ("navrh_splatnost1", "Návrh nové splatnosti 1 (zboží) [ve dnech]",    "text"),
                ("navrh_splatnost2", "Návrh nové splatnosti 2 (cigarety) [ve dnech]", "text"),
                ("zduvodneni",       "Zdůvodnění požadované změny / žádosti",         "textarea"),
            ]),
        ],
    },
}

# Mapování dotažených (auto) polí na sloupce z asm_kontakty.
def _cislo_cz(v):
    """Číslo na český formát (mezera jako oddělovač tisíců, bez zbytečných desetin).
    Nečíselné/prázdné vrátí jako čistý text."""
    if v is None or v == "":
        return ""
    try:
        f = float(v)
    except (TypeError, ValueError):
        return str(v).strip()
    if f == int(f):
        s = f"{int(f):,}".replace(",", " ")
    else:
        s = f"{f:,.2f}".replace(",", " ").replace(".", ",")
    return s


# Obrat nad tuto hranici žadatel NEVIDÍ přesně — zobrazí se jen „> 2 mil. Kč".
# Office obchod (a správce) vidí přesnou hodnotu (viz detail případu).
_OBRAT_MASK_PRAH = 2_000_000
_OBRAT_MASK_TEXT = "> 2 mil. Kč"


def _obrat_zobraz(val, maskovat):
    """Obrat do formuláře. maskovat=True (žadatel) → nad 2 mil. Kč jen „> 2 mil. Kč",
    jinak přesné číslo (CZ formát)."""
    if maskovat:
        try:
            if val is not None and float(val) > _OBRAT_MASK_PRAH:
                return _OBRAT_MASK_TEXT
        except (TypeError, ValueError):
            pass
    return _cislo_cz(val)


def _obrat_vs_limit(obrat, limit):
    """Porovnání obratu s aktuálním kreditním limitem — do detailu případu,
    kde se částky obratu nezobrazují. Chybějící/nečitelná hodnota → „—".
    Parsuje se formát z _cislo_cz (mezera = tisíce, čárka = desetiny)."""
    def _f(v):
        try:
            return float(str(v).replace(" ", "").replace(" ", "").replace(",", "."))
        except (TypeError, ValueError):
            return None
    o, l = _f(obrat), _f(limit)
    if o is None or l is None:
        return "—"
    if o > l:
        return "Obrat je vyšší než aktuální kreditní limit"
    if o < l:
        return "Obrat je menší než aktuální kreditní limit"
    return "Obrat je stejný jako aktuální kreditní limit"


def _ico_auto_hodnoty(k, o=None, maskovat_obrat=False):
    """Z kontaktu (dict z _kontakty_dle_ico) + obratů (dict z _obrat_dle_ico)
    sestaví auto pole formulářů (sekce 1 i ekonomické ukazatele sekce 3).
    maskovat_obrat=True → obrat nad 2 mil. Kč se žadateli zamaskuje („> 2 mil. Kč")."""
    out = {"pobocka": "", "fakt_nazev": "", "provozovna": "", "adresa": "",
           "akt_limit": "", "splatnost1_akt": "", "splatnost2_akt": "",
           "obrat_1m": "", "obrat_3m": ""}
    if k:
        # Provozovna/adresa místa dodání: přednostně K.* (dodací) z Kontaktních údajů VO2,
        # fallback na fakturační Ulice/Město/PSČ + Jméno, když K.* nejsou vyplněné.
        adresa = " ".join(x for x in (k.get("k_ulice"), k.get("k_mesto"), k.get("k_psc"))
                          if x and str(x).strip())
        if not adresa:
            adresa = " ".join(x for x in (k.get("ulice"), k.get("mesto"), k.get("psc"))
                              if x and str(x).strip())
        provozovna = (k.get("k_jmeno") or "").strip() or (k.get("jmeno") or "")
        # Splatnost 2: když prázdná, použij Splatnost 1 (dle zadání).
        spl1 = k.get("splatnost1")
        spl2 = k.get("splatnost2")
        if spl2 is None or str(spl2).strip() == "":
            spl2 = spl1
        out.update({
            "pobocka":        k.get("prodejni_doba") or "",
            "fakt_nazev":     k.get("jmeno") or "",
            "provozovna":     provozovna,
            "adresa":         adresa,
            "akt_limit":      _cislo_cz(k.get("limit_salda")),
            "splatnost1_akt": _cislo_cz(spl1),
            "splatnost2_akt": _cislo_cz(spl2),
        })
    if o:
        out["obrat_1m"] = _obrat_zobraz(o.get("obrat_1m"), maskovat_obrat)
        out["obrat_3m"] = _obrat_zobraz(o.get("obrat_3m"), maskovat_obrat)
    return out


def _formular_pole(formular):
    """Seznam všech (key, label, typ, options?) napříč sekcemi daného formuláře."""
    out = []
    for _, pola in _FORMULARE[formular]["sekce"]:
        out.extend(pola)
    return out


# ============================================================================
# Práva / viditelnost
# ============================================================================
def _je_spravce(p):
    return ("vse" in p or "asm_spravce" in p or "asm_spravce_bez_emailu" in p)


def _je_office(p):
    return _je_spravce(p) or "asm_office" in p


def _je_zadatel(p):
    return _je_spravce(p) or _je_office(p) or "asm_zadatel" in p


def _vidi_import(p):
    """Smí do dlaždice „Data" (nahrávání číselníků). Správce nebo vkladatel."""
    return _je_spravce(p) or "asm_vkladatel" in p


def _vidi_vsechny_pripady(p):
    """Office/správce vidí celou frontu; žadatel jen své případy."""
    return _je_spravce(p) or "asm_office" in p


def _ma_pristup(p):
    """Brána modulu — má uživatel přístup k dlaždici „Formuláře ASM"?"""
    return (_je_zadatel(p) or _vidi_import(p))


# ============================================================================
# Inicializace DB
# ============================================================================
# Schéma + migrace se řeší jen JEDNOU za běh procesu. Bez toho běžely desítky
# dotazů do information_schema (kontrola sloupců/collation/indexů) při KAŽDÉM
# renderu ASM → načítání trvalo vteřiny. Po prvním úspěchu se přeskočí.
_ASM_DB_INIT_HOTOVO = False


def inicializace_asm_db():
    global _ASM_DB_INIT_HOTOVO
    if _ASM_DB_INIT_HOTOVO:
        return
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        # Číselník OZ/ASM (Dealer.xlsx): klic = číslo OZ, jmeno = Popis, asm = Hodnota
        cur.execute("""
            CREATE TABLE IF NOT EXISTS asm_dealer (
                klic VARCHAR(40) COLLATE utf8mb4_bin PRIMARY KEY,
                jmeno VARCHAR(255),
                asm VARCHAR(120),
                email VARCHAR(255),
                INDEX idx_asm (asm)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        # Kontaktní údaje VO (xlsx) — fakturační sada, klíč IČ 10 (plné), lookup dle IČ 8.
        # POZOR: ico = utf8mb4_bin (case/diakritika SENSITIVE). Výchozí
        # utf8mb4_unicode_ci je necitlivá → IČO jako '12345-Á' a '12345-A'
        # (různí zákazníci) by kolidovaly v PRIMARY KEY (chyba 1062).
        # ico = IČ 10 (plné, vč. přípony), ico8 = IČ 8 (zleva 8) pro dotahování ve formulářích.
        cur.execute("""
            CREATE TABLE IF NOT EXISTS asm_kontakty (
                ico VARCHAR(40) COLLATE utf8mb4_bin PRIMARY KEY,
                ico8 VARCHAR(20) COLLATE utf8mb4_bin,
                jmeno VARCHAR(255),
                ulice VARCHAR(255),
                mesto VARCHAR(255),
                psc VARCHAR(40),
                dealer VARCHAR(40),
                id_kontakt VARCHAR(40),
                prodejni_doba VARCHAR(120),
                konst_sym VARCHAR(40),
                limit_salda VARCHAR(40),
                splatnost1 VARCHAR(40),
                splatnost2 VARCHAR(40),
                typ_platice VARCHAR(120),
                INDEX idx_ico8 (ico8)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        # Kontaktní údaje pro dlaždici „Změna zákazníků na OZ/ASM" — PŮVODNÍ formát
        # (K. dodací sloupce, dotahování dle plného IČO). Samostatná tabulka a import,
        # aby koexistovala s novým fakturačním formátem unifikovaných formulářů.
        cur.execute("""
            CREATE TABLE IF NOT EXISTS asm_kontakty_oz (
                ico VARCHAR(40) COLLATE utf8mb4_bin PRIMARY KEY,
                jmeno VARCHAR(255),
                k_jmeno VARCHAR(255),
                k_jmeno2 VARCHAR(255),
                k_ulice VARCHAR(255),
                k_mesto VARCHAR(255),
                k_psc VARCHAR(40),
                dealer VARCHAR(40),
                id_kontakt VARCHAR(40),
                prodejni_doba VARCHAR(120),
                limit_salda VARCHAR(40),
                splatnost1 VARCHAR(40),
                splatnost2 VARCHAR(40),
                typ_platice VARCHAR(120)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        # Obraty zákazníků (GIST_obrat zakaznik.xlsx) — měsíční nahrávání, párování dle IČO.
        cur.execute("""
            CREATE TABLE IF NOT EXISTS asm_obrat (
                ico VARCHAR(40) COLLATE utf8mb4_bin PRIMARY KEY,
                obrat_1m DOUBLE,
                obrat_3m DOUBLE
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        # Obraty po jednotlivých IČ 10 (bloky „IČO" v témže GIST souboru, sloupce G–H / J–K).
        cur.execute("""
            CREATE TABLE IF NOT EXISTS asm_obrat_ic10 (
                ico VARCHAR(40) COLLATE utf8mb4_bin PRIMARY KEY,
                obrat_1m DOUBLE,
                obrat_3m DOUBLE
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        # Log nahrávek číselníků
        cur.execute("""
            CREATE TABLE IF NOT EXISTS asm_import (
                id INT AUTO_INCREMENT PRIMARY KEY,
                kdy DATETIME DEFAULT CURRENT_TIMESTAMP,
                uzivatel VARCHAR(255),
                pocet_dealer INT DEFAULT 0,
                pocet_kontakty INT DEFAULT 0,
                pocet_obrat INT DEFAULT 0,
                soubor_dealer VARCHAR(255),
                soubor_kontakty VARCHAR(255),
                soubor_obrat VARCHAR(255),
                INDEX idx_kdy (kdy)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        # Případy (hlavička formuláře)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS asm_pripady (
                id INT AUTO_INCREMENT PRIMARY KEY,
                cislo VARCHAR(20),
                formular VARCHAR(30) DEFAULT 'oz_zmena',
                data_json TEXT,
                duvod_zmeny VARCHAR(1000),
                asm_jmeno VARCHAR(255),
                zmena_asm_jmeno VARCHAR(255),
                novy_asm_jmeno VARCHAR(255),
                cislo_oz VARCHAR(40), jmeno_oz VARCHAR(255),
                cislo_novy_oz VARCHAR(40), jmeno_novy_oz VARCHAR(255),
                datum_zmeny_od VARCHAR(20),
                zakaznik_v_regionu_oz VARCHAR(5),
                oduvodneni VARCHAR(1000),
                zadavatel_id INT, zadavatel_jmeno VARCHAR(255),
                datum_zadani DATETIME DEFAULT CURRENT_TIMESTAMP,
                stav VARCHAR(30) DEFAULT 'odeslano',
                pocet_radku INT DEFAULT 0,
                poznamka VARCHAR(1000),
                zamitnuti_duvod VARCHAR(1000),
                storno_duvod VARCHAR(1000),
                aktualizovano DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                INDEX idx_zadavatel (zadavatel_id), INDEX idx_stav (stav)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        # Řádky případu (detailní tabulka)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS asm_radky (
                id INT AUTO_INCREMENT PRIMARY KEY,
                pripad_id INT, poradi INT,
                ico VARCHAR(40),
                jmeno VARCHAR(255), k_jmeno VARCHAR(255), k_jmeno2 VARCHAR(255),
                k_ulice VARCHAR(255), k_mesto VARCHAR(255), k_psc VARCHAR(40),
                dealer VARCHAR(40), id_kontakt VARCHAR(40), prodejni_doba VARCHAR(120),
                puvodni_provize VARCHAR(10),
                nova_provize VARCHAR(40),
                oz2 VARCHAR(120),
                datum_od VARCHAR(20), datum_do VARCHAR(20),
                INDEX idx_pripad (pripad_id)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        # Historie/průběh případu („očičko")
        cur.execute("""
            CREATE TABLE IF NOT EXISTS asm_historie (
                id INT AUTO_INCREMENT PRIMARY KEY,
                pripad_id INT,
                akce VARCHAR(80),
                detail VARCHAR(1000),
                kdo VARCHAR(255),
                kdy DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_pripad (pripad_id)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        # Migrace: nové sloupce případů (storno + formulář/data_json pro další typy formulářů)
        for _col, _typ in (("storno_duvod", "VARCHAR(1000)"),
                           ("formular", "VARCHAR(30) DEFAULT 'oz_zmena'"),
                           ("data_json", "TEXT"),
                           # Interní poznámka Office → správce při postoupení ke schválení.
                           # Záměrně NE `poznamka` — ta je sdílená s „Vrátit k opravě“
                           # a zobrazuje se i žadateli (viz _spolecne_info).
                           ("spravce_pozn", "VARCHAR(1000)")):
            cur.execute("SELECT COUNT(*) FROM information_schema.COLUMNS WHERE "
                        "TABLE_SCHEMA=DATABASE() AND TABLE_NAME='asm_pripady' AND "
                        "COLUMN_NAME=%s", (_col,))
            if cur.fetchone()[0] == 0:
                cur.execute(f"ALTER TABLE asm_pripady ADD COLUMN {_col} {_typ}")
        # Migrace: ico na utf8mb4_bin, pokud byla tabulka vytvořena dříve s ci
        # collation (kolize IČO s diakritikou v PRIMARY KEY).
        cur.execute("SELECT COLLATION_NAME FROM information_schema.COLUMNS WHERE "
                    "TABLE_SCHEMA=DATABASE() AND TABLE_NAME='asm_kontakty' AND "
                    "COLUMN_NAME='ico'")
        _r = cur.fetchone()
        if _r and _r[0] and _r[0] != "utf8mb4_bin":
            cur.execute("ALTER TABLE asm_kontakty MODIFY ico VARCHAR(40) "
                        "COLLATE utf8mb4_bin")
        # Migrace: asm_dealer.klic na utf8mb4_bin — stejný důvod (klíče '…-E' vs '…-É'
        # kolidují v necitlivé ci → chyba 1062 Duplicate entry při importu dealeru).
        cur.execute("SELECT COLLATION_NAME FROM information_schema.COLUMNS WHERE "
                    "TABLE_SCHEMA=DATABASE() AND TABLE_NAME='asm_dealer' AND "
                    "COLUMN_NAME='klic'")
        _rd = cur.fetchone()
        if _rd and _rd[0] and _rd[0] != "utf8mb4_bin":
            cur.execute("ALTER TABLE asm_dealer MODIFY klic VARCHAR(40) COLLATE utf8mb4_bin")
        # Migrace: nové sloupce kontaktů (fakturační sada + IČ 8). Nový formát souboru
        # „Kontaktní údaje VO" zrušil „K." (dodací) sloupce → přemapováno na fakturační.
        def _ma_sloupec(tab, col):
            cur.execute("SELECT COUNT(*) FROM information_schema.COLUMNS WHERE "
                        "TABLE_SCHEMA=DATABASE() AND TABLE_NAME=%s AND COLUMN_NAME=%s",
                        (tab, col))
            return cur.fetchone()[0] > 0
        for _col, _typ in (("ico8", "VARCHAR(20) COLLATE utf8mb4_bin"),
                           ("ulice", "VARCHAR(255)"), ("mesto", "VARCHAR(255)"),
                           ("psc", "VARCHAR(40)"), ("konst_sym", "VARCHAR(40)"),
                           ("limit_salda", "VARCHAR(40)"), ("splatnost1", "VARCHAR(40)"),
                           ("splatnost2", "VARCHAR(40)"), ("typ_platice", "VARCHAR(120)")):
            if not _ma_sloupec("asm_kontakty", _col):
                cur.execute(f"ALTER TABLE asm_kontakty ADD COLUMN {_col} {_typ}")
        # Migrace: asm_dealer.email (firemní e-mail OZ, dopočítaný ze jména při importu).
        # Naplní se až při nejbližším nahrání Dealer.xlsx po nasazení.
        if not _ma_sloupec("asm_dealer", "email"):
            cur.execute("ALTER TABLE asm_dealer ADD COLUMN email VARCHAR(255)")
        # „K." (dodací) sloupce — plní se z Kontaktních údajů VO2 (K.Jméno, K.Ulice, …).
        for _col, _typ in (("k_jmeno", "VARCHAR(255)"), ("k_ulice", "VARCHAR(255)"),
                           ("k_mesto", "VARCHAR(255)"), ("k_psc", "VARCHAR(40)")):
            if not _ma_sloupec("asm_kontakty", _col):
                cur.execute(f"ALTER TABLE asm_kontakty ADD COLUMN {_col} {_typ}")
        # Index na IČ 8 (lookup ve formulářích), pokud chybí.
        cur.execute("SELECT COUNT(*) FROM information_schema.STATISTICS WHERE "
                    "TABLE_SCHEMA=DATABASE() AND TABLE_NAME='asm_kontakty' AND "
                    "INDEX_NAME='idx_ico8'")
        if cur.fetchone()[0] == 0 and _ma_sloupec("asm_kontakty", "ico8"):
            cur.execute("ALTER TABLE asm_kontakty ADD INDEX idx_ico8 (ico8)")
        # Migrace: řádky případů (Změna zákazníků) — VRÁCENO na původní K. sloupce.
        # Kdyby DB byla v mezistavu (přejmenováno na ulice/mesto/psc), vrať zpět (zachová data).
        for _novy, _stary, _typ in (("ulice", "k_ulice", "VARCHAR(255)"),
                                    ("mesto", "k_mesto", "VARCHAR(255)"),
                                    ("psc", "k_psc", "VARCHAR(40)")):
            if _ma_sloupec("asm_radky", _novy) and not _ma_sloupec("asm_radky", _stary):
                cur.execute(f"ALTER TABLE asm_radky CHANGE COLUMN {_novy} {_stary} {_typ}")
            elif not _ma_sloupec("asm_radky", _stary):
                cur.execute(f"ALTER TABLE asm_radky ADD COLUMN {_stary} {_typ}")
        for _col, _typ in (("k_jmeno", "VARCHAR(255)"), ("k_jmeno2", "VARCHAR(255)")):
            if not _ma_sloupec("asm_radky", _col):
                cur.execute(f"ALTER TABLE asm_radky ADD COLUMN {_col} {_typ}")
        # Migrace: log importu — sloupce pro obraty (GIST)
        for _col, _typ in (("pocet_obrat", "INT DEFAULT 0"),
                           ("soubor_obrat", "VARCHAR(255)")):
            cur.execute("SELECT COUNT(*) FROM information_schema.COLUMNS WHERE "
                        "TABLE_SCHEMA=DATABASE() AND TABLE_NAME='asm_import' AND "
                        "COLUMN_NAME=%s", (_col,))
            if cur.fetchone()[0] == 0:
                cur.execute(f"ALTER TABLE asm_import ADD COLUMN {_col} {_typ}")
        conn.commit()
        cur.close()
        _ASM_DB_INIT_HOTOVO = True   # migrace proběhly → příště přeskoč (výkon)
    except Exception as e:
        print(f"Chyba při inicializaci DB Formuláře ASM: {e}")
    finally:
        conn.close()


# ============================================================================
# Pomocné funkce
# ============================================================================
def _str(v, n=255):
    return None if v is None else str(v).strip()[:n]


def _ico_text(v):
    """IČO jako text — z floatu udělá celé číslo bez '.0', jinak ořízne mezery."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _dt_cz(dt):
    if not dt:
        return ""
    if hasattr(dt, "strftime"):
        return dt.strftime("%d.%m.%Y %H:%M")
    return str(dt)


# ============================================================================
# Import číselníků (dlaždice „Data") — běží v threadu
# ============================================================================
def _parse_dealer(raw_bytes):
    """Dealer (.xlsx i .xlsb) → [(klic, jmeno, asm, email)]. Hlavička na ř.5 (Klíč/Popis/
    Hodnota), data od ř.6. Jméno OZ = část Popisu před prvním '/' (odstraní koncové kódy).
    E-mail OZ se dopočítá ze jména (jmeno.prijmeni@<firemní doména>) už při importu —
    formulář ho pak jen čte z DB. Bez příjmení / bez domény zůstane prázdný."""
    out, viděno = [], set()
    domena = _smtp_domena()
    for i, row in enumerate(_iter_sheet_rows(raw_bytes)):
        if i < 5:          # hlavička na ř.5 (index 4), data od ř.6 (index 5)
            continue
        if not row:
            continue
        klic = _ico_text(row[0]) if len(row) > 0 else ""
        if not klic:
            continue
        popis = (str(row[1]).strip() if len(row) > 1 and row[1] is not None else "")
        jmeno = popis.split("/", 1)[0].strip() if popis else ""
        asm = (str(row[2]).strip() if len(row) > 2 and row[2] is not None else "")
        if klic in viděno:
            continue
        viděno.add(klic)
        email = _email_z_jmena(jmeno, domena)
        out.append((klic[:40], jmeno[:255], asm[:120], email[:255]))
    return out


def _norm_hlavicka(s):
    """Název sloupce hlavičky → kanonický tvar (lower, bez diakritiky, jednoduché mezery)."""
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = (s.replace("č", "c").replace("ř", "r").replace("š", "s").replace("ž", "z")
          .replace("ý", "y").replace("á", "a").replace("í", "i").replace("é", "e")
          .replace("ě", "e").replace("ú", "u").replace("ů", "u").replace("ó", "o")
          .replace("ď", "d").replace("ť", "t").replace("ň", "n"))
    return " ".join(s.split())


def _je_vyrazen(typ_platice):
    """True, když „Typ platiče" znamená vyřazeného zákazníka."""
    return "vyrazen" in _norm_hlavicka(typ_platice)


# Kanonický (bez diakritiky) název sloupce → klíč v DB. Pořadí v souboru je libovolné.
_VO_SLOUPCE = {
    "ico":            "ico",           # IČ 10 (plné); IČ 8 se dopočítá zleva 8 znaků
    "jmeno":          "jmeno",         # fakturační název
    "ulice":          "ulice",
    "mesto":          "mesto",
    "psc":            "psc",
    "dealer":         "dealer",
    "id":             "id_kontakt",
    "prodejni doba 6": "prodejni_doba",
    "prodejni kanal 6": "prodejni_doba",   # alternativní název téhož sloupce
    # K. dodací (provozovna) — jméno/adresa místa dodání z Kontaktních údajů VO2.
    "k.jmeno":        "k_jmeno",
    "k.ulice":        "k_ulice",
    "k.mesto":        "k_mesto",
    "k.psc":          "k_psc",
    "konst.sym.":     "konst_sym",
    "typ platice":    "typ_platice",
    "splatnost":      "splatnost1",
    "splatnost2":     "splatnost2",
    "limit salda":    "limit_salda",
}


def _iter_sheet_rows(raw_bytes):
    """Řádky prvního listu — zvládne .xlsx (openpyxl) i .xlsb (pyxlsb).
    Yielduje list hodnot buněk. Formát se detekuje pokusem o openpyxl; při selhání
    (např. „File contains no valid workbook part" u .xlsb) se použije pyxlsb."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
    except Exception:
        wb = None
    if wb is not None:
        try:
            for r in wb.active.iter_rows(values_only=True):
                yield list(r)
        finally:
            wb.close()
        return
    # pyxlsb: sh.rows() vrací JEN neprázdné buňky → nesmíme je slévat prostým
    # [c.v for c in row] (posunuly by se sloupce). Poskládáme dle indexu buňky c.c,
    # prázdné pozice = None (drží zarovnání sloupců stejně jako openpyxl).
    from pyxlsb import open_workbook
    with open_workbook(io.BytesIO(raw_bytes)) as wb2:
        with wb2.get_sheet(wb2.sheets[0]) as sh:
            for row in sh.rows():
                bunky = {c.c: c.v for c in row}
                yield [bunky.get(i) for i in range(max(bunky) + 1)] if bunky else []


def _parse_kontakty(raw_bytes):
    """Kontaktní údaje VO (.xlsx i .xlsb) → [dict]. Sloupce se párují podle NÁZVU v hlavičce
    (odolné vůči přeházení/doplnění sloupců). Hlavička = první řádek obsahující „IČO".
    IČO ve souboru je IČ 10 (plné) → ukládá se i IČ 8 (zleva 8) pro dotahování.
    Dedup dle plného IČ 10 (víc variant pod jedním IČ 8 se ZACHOVÁ — pro výběr v bublině).
    Vyřazení platiči („Typ platiče" = „vyřazen") se importují také; při shodě IČ 10 má
    ale přednost aktivní řádek — vyřazený se použije, jen když aktivní neexistuje."""
    out, poradi, vyr = [], {}, {}   # out, {ico: index v out}, {ico: je_vyrazen}
    col_map = None          # {db_klic: index sloupce}
    for vals in _iter_sheet_rows(raw_bytes):
        if col_map is None:
            norm = [_norm_hlavicka(v) for v in vals]
            if "ico" not in norm:
                continue    # ještě nejsme na hlavičce (preambule nad daty)
            col_map = {}
            for j, nm in enumerate(norm):
                klic = _VO_SLOUPCE.get(nm)
                if klic and klic not in col_map:
                    col_map[klic] = j
            continue        # hlavičku samotnou nezpracovávat jako data
        def g(klic):
            j = col_map.get(klic)
            return vals[j] if j is not None and j < len(vals) else None
        ico = _ico_text(g("ico"))
        if not ico:
            continue
        je_vyr = _je_vyrazen(g("typ_platice"))
        idx = poradi.get(ico)
        # Duplicitu nahradíme jen tehdy, když aktivní řádek střídá dříve načtený vyřazený.
        if idx is not None and not (vyr[ico] and not je_vyr):
            continue
        radek = {
            "ico":           ico[:40],
            "ico8":          ico[:8],
            "jmeno":         _str(g("jmeno")),
            "ulice":         _str(g("ulice")),
            "mesto":         _str(g("mesto")),
            "psc":           _str(g("psc"), 40),
            "dealer":        _str(g("dealer"), 40),
            "id_kontakt":    _str(g("id_kontakt"), 40),
            "prodejni_doba": _str(g("prodejni_doba"), 120),
            "k_jmeno":       _str(g("k_jmeno")),
            "k_ulice":       _str(g("k_ulice")),
            "k_mesto":       _str(g("k_mesto")),
            "k_psc":         _str(g("k_psc"), 40),
            "konst_sym":     _str(g("konst_sym"), 40),
            "limit_salda":   _ico_text(g("limit_salda"))[:40],
            "splatnost1":    _ico_text(g("splatnost1"))[:40],
            "splatnost2":    _ico_text(g("splatnost2"))[:40],
            "typ_platice":   _str(g("typ_platice"), 120),
        }
        vyr[ico] = je_vyr
        if idx is None:
            poradi[ico] = len(out)
            out.append(radek)
        else:
            out[idx] = radek
    return out


# Mapování PŮVODNÍHO souboru kontaktů (dlaždice Změna OZ) — K. dodací sloupce.
_VO_SLOUPCE_OZ = {
    "ico":            "ico",
    "limit salda":    "limit_salda",
    "splatnost":      "splatnost1",
    "splatnost2":     "splatnost2",
    "jmeno":          "jmeno",
    "k.jmeno":        "k_jmeno",
    "k.jmeno 2":      "k_jmeno2",
    "typ platice":    "typ_platice",
    "k.ulice":        "k_ulice",
    "k.mesto":        "k_mesto",
    "k.psc":          "k_psc",
    "dealer":         "dealer",
    "id":             "id_kontakt",
    "prodejni doba 6": "prodejni_doba",
}


def _parse_kontakty_oz(raw_bytes):
    """PŮVODNÍ Kontaktní údaje VO (K. dodací sloupce) pro dlaždici „Změna zákazníků" →
    [dict]. Dotahování dle plného IČO. Dedup dle IČO; vyřazení platiči se importují,
    ale při shodě IČO má přednost aktivní řádek."""
    out, poradi, vyr = [], {}, {}   # out, {ico: index v out}, {ico: je_vyrazen}
    col_map = None
    for vals in _iter_sheet_rows(raw_bytes):
        if col_map is None:
            norm = [_norm_hlavicka(v) for v in vals]
            if "ico" not in norm:
                continue
            col_map = {}
            for j, nm in enumerate(norm):
                klic = _VO_SLOUPCE_OZ.get(nm)
                if klic and klic not in col_map:
                    col_map[klic] = j
            continue
        def g(klic):
            j = col_map.get(klic)
            return vals[j] if j is not None and j < len(vals) else None
        ico = _ico_text(g("ico"))
        if not ico:
            continue
        je_vyr = _je_vyrazen(g("typ_platice"))
        idx = poradi.get(ico)
        # Duplicitu nahradíme jen tehdy, když aktivní řádek střídá dříve načtený vyřazený.
        if idx is not None and not (vyr[ico] and not je_vyr):
            continue
        radek = {
            "ico":           ico[:40],
            "jmeno":         _str(g("jmeno")),
            "k_jmeno":       _str(g("k_jmeno")),
            "k_jmeno2":      _str(g("k_jmeno2")),
            "k_ulice":       _str(g("k_ulice")),
            "k_mesto":       _str(g("k_mesto")),
            "k_psc":         _str(g("k_psc"), 40),
            "dealer":        _str(g("dealer"), 40),
            "id_kontakt":    _str(g("id_kontakt"), 40),
            "prodejni_doba": _str(g("prodejni_doba"), 120),
            "limit_salda":   _ico_text(g("limit_salda"))[:40],
            "splatnost1":    _ico_text(g("splatnost1"))[:40],
            "splatnost2":    _ico_text(g("splatnost2"))[:40],
            "typ_platice":   _str(g("typ_platice"), 120),
        }
        vyr[ico] = je_vyr
        if idx is None:
            poradi[ico] = len(out)
            out.append(radek)
        else:
            out[idx] = radek
    return out


def _parse_gist(raw_bytes):
    """GIST_obrat zakaznik.xlsx → (dle_ic8, dle_ic10), obojí
    {ico: {'obrat_1m':float|None, 'obrat_3m':float|None}}.
    Bloky s hlavičkou „IČO 8" jsou za IČ 8, bloky s hlavičkou „IČO" za IČ 10.
    Období bloku se určí podle popisku („3 měsíce" / „1 měsíc") nad hlavičkou
    — je tak jedno, ve kterých sloupcích se zrovna nachází."""
    rows = list(_iter_sheet_rows(raw_bytes))   # .xlsx i .xlsb
    if not rows:
        return {}, {}

    def _per_z_popisku(txt):
        t = _norm_hlavicka(txt)
        if not t or "mesic" not in t:
            return None
        if t.startswith("3") or "3 m" in t:
            return "obrat_3m"
        if t.startswith("1") or "1 m" in t:
            return "obrat_1m"
        return None

    # Nový formát: hlavička „IČO 8" (hodnoty jsou přímo IČ 8 = zleva 8 znaků).
    def _je_ico_hlavicka(c):
        return _norm_hlavicka(c) in ("ico 8", "ico8")

    # Najdi řádek s hlavičkou „IČO 8" (může jich být víc ve více sloupcích).
    hdr_idx = None
    for i, r in enumerate(rows):
        if any(_je_ico_hlavicka(c) for c in r):
            hdr_idx = i
            break
    if hdr_idx is None:
        return {}, {}

    # Bloky: sloupec s hlavičkou „IČO 8" → data za IČ 8, s hlavičkou „IČO" → za IČ 10.
    # Každý blok = (per, cil, ico_col, obrat_col).
    bloky = []
    hdr = rows[hdr_idx]
    popisky = rows[hdr_idx - 1] if hdr_idx > 0 else []
    poradi = {"ic8": 0, "ic10": 0}
    for j, c in enumerate(hdr):
        if _je_ico_hlavicka(c):
            cil = "ic8"
        elif _norm_hlavicka(c) == "ico":
            cil = "ic10"
        else:
            continue
        per = _per_z_popisku(popisky[j]) if j < len(popisky) else None
        if per is None:
            # fallback: pořadí bloků daného typu (1. = 3m, 2. = 1m) dle layoutu souboru
            per = "obrat_3m" if poradi[cil] == 0 else "obrat_1m"
        poradi[cil] += 1
        bloky.append((per, cil, j, j + 1))

    out = {"ic8": {}, "ic10": {}}
    for r in rows[hdr_idx + 1:]:
        for per, cil, ic, oc in bloky:
            ico = _ico_text(r[ic]) if ic < len(r) else ""
            if not ico:
                continue
            val = r[oc] if oc < len(r) else None
            try:
                val = float(val) if val not in (None, "") else None
            except (TypeError, ValueError):
                val = None
            rec = out[cil].setdefault(ico[:40], {"obrat_1m": None, "obrat_3m": None})
            rec[per] = val
    return out["ic8"], out["ic10"]


def _importuj_sync(dealer_raw, dealer_name, kontakty_raw, kontakty_name, user_name,
                   gist_raw=None, gist_name="", kontakty_oz_raw=None, kontakty_oz_name=""):
    """Plná náhrada vybraných číselníků. Lze nahrát jeden i víc najednou.
    Vrací (pocet_dealer, pocet_kontakty, pocet_obrat, chyba|None)."""
    if not dealer_raw and not kontakty_raw and not gist_raw and not kontakty_oz_raw:
        return 0, 0, 0, "Nevybrán žádný soubor k importu."

    dealer_rows = None
    kontakt_rows = None
    kontakt_oz_rows = None
    gist_map = None
    gist_map_ic10 = None
    try:
        if dealer_raw:
            dealer_rows = _parse_dealer(dealer_raw)
        if kontakty_raw:
            kontakt_rows = _parse_kontakty(kontakty_raw)
        if kontakty_oz_raw:
            kontakt_oz_rows = _parse_kontakty_oz(kontakty_oz_raw)
        if gist_raw:
            gist_map, gist_map_ic10 = _parse_gist(gist_raw)
    except Exception as e:
        return 0, 0, 0, f"Chyba čtení souboru: {e}"

    conn = intranet_data.get_db_connection()
    if not conn:
        return 0, 0, 0, "Není připojení k databázi."
    try:
        cur = conn.cursor()
        if dealer_rows is not None:
            cur.execute("DELETE FROM asm_dealer")
            cur.executemany(
                "INSERT INTO asm_dealer (klic, jmeno, asm, email) VALUES (%s,%s,%s,%s)",
                dealer_rows)
        if kontakt_rows is not None:
            cur.execute("DELETE FROM asm_kontakty")
            cur.executemany(
                "INSERT INTO asm_kontakty (ico, ico8, jmeno, ulice, mesto, psc, "
                "k_jmeno, k_ulice, k_mesto, k_psc, "
                "dealer, id_kontakt, prodejni_doba, konst_sym, "
                "limit_salda, splatnost1, splatnost2, typ_platice) "
                "VALUES (%(ico)s,%(ico8)s,%(jmeno)s,%(ulice)s,%(mesto)s,%(psc)s,"
                "%(k_jmeno)s,%(k_ulice)s,%(k_mesto)s,%(k_psc)s,"
                "%(dealer)s,%(id_kontakt)s,%(prodejni_doba)s,%(konst_sym)s,"
                "%(limit_salda)s,%(splatnost1)s,%(splatnost2)s,%(typ_platice)s)",
                kontakt_rows)
        if kontakt_oz_rows is not None:
            cur.execute("DELETE FROM asm_kontakty_oz")
            cur.executemany(
                "INSERT INTO asm_kontakty_oz (ico, jmeno, k_jmeno, k_jmeno2, k_ulice, "
                "k_mesto, k_psc, dealer, id_kontakt, prodejni_doba, "
                "limit_salda, splatnost1, splatnost2, typ_platice) "
                "VALUES (%(ico)s,%(jmeno)s,%(k_jmeno)s,%(k_jmeno2)s,%(k_ulice)s,"
                "%(k_mesto)s,%(k_psc)s,%(dealer)s,%(id_kontakt)s,%(prodejni_doba)s,"
                "%(limit_salda)s,%(splatnost1)s,%(splatnost2)s,%(typ_platice)s)",
                kontakt_oz_rows)
        if gist_map is not None:
            cur.execute("DELETE FROM asm_obrat")
            gist_rows = [(ico, v.get("obrat_1m"), v.get("obrat_3m"))
                         for ico, v in gist_map.items()]
            if gist_rows:
                cur.executemany(
                    "INSERT INTO asm_obrat (ico, obrat_1m, obrat_3m) VALUES (%s,%s,%s)",
                    gist_rows)
            # Obraty po IČ 10 (stejný soubor, bloky „IČO")
            cur.execute("DELETE FROM asm_obrat_ic10")
            gist_rows10 = [(ico, v.get("obrat_1m"), v.get("obrat_3m"))
                           for ico, v in (gist_map_ic10 or {}).items()]
            if gist_rows10:
                cur.executemany(
                    "INSERT INTO asm_obrat_ic10 (ico, obrat_1m, obrat_3m) "
                    "VALUES (%s,%s,%s)",
                    gist_rows10)
        # Počty aktuálně v DB
        cur.execute("SELECT COUNT(*) FROM asm_dealer")
        pd = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM asm_kontakty")
        pk = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM asm_obrat")
        po = cur.fetchone()[0]
        cur.execute(
            "INSERT INTO asm_import (uzivatel, pocet_dealer, pocet_kontakty, pocet_obrat, "
            "soubor_dealer, soubor_kontakty, soubor_obrat) VALUES (%s,%s,%s,%s,%s,%s,%s)",
            (user_name[:255], pd, pk, po,
             (dealer_name or None) if dealer_rows is not None else None,
             (kontakty_name or None) if kontakt_rows is not None else None,
             (gist_name or None) if gist_map is not None else None))
        conn.commit()
        cur.close()
        return pd, pk, po, None
    except Exception as e:
        conn.rollback()
        return 0, 0, 0, f"Chyba zápisu do DB: {e}"
    finally:
        conn.close()


def pocty_v_db():
    """Vrací (pocet_dealer, pocet_kontakty, pocet_obrat)."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return 0, 0, 0
    try:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM asm_dealer")
        pd = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM asm_kontakty")
        pk = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM asm_obrat")
        po = cur.fetchone()[0]
        cur.close()
        return pd, pk, po
    except Exception:
        return 0, 0, 0
    finally:
        conn.close()


def posledni_import():
    conn = intranet_data.get_db_connection()
    if not conn:
        return None
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT * FROM asm_import ORDER BY id DESC LIMIT 1")
        r = cur.fetchone()
        cur.close()
        return r
    except Exception:
        return None
    finally:
        conn.close()


# ============================================================================
# Číselníky (čtení z DB)
# ============================================================================
def _oz_ciselnik():
    """[(klic, jmeno)] pro výběr OZ."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor()
        cur.execute("SELECT klic, jmeno FROM asm_dealer ORDER BY CAST(klic AS UNSIGNED), klic")
        out = cur.fetchall()
        cur.close()
        return [(k, j or "") for k, j in out]
    except Exception as e:
        print(f"[asm] _oz_ciselnik: {e}")
        return []
    finally:
        conn.close()


def _oz_ciselnik_lide():
    """[(klic, jmeno, email)] jen OZ-lidé (jméno = křestní + příjmení).
    E-mail se dopočítá ze jména za běhu (ne z uloženého sloupce — ten bývá prázdný,
    dokud neproběhne import / není nastavená doména). Bez domény zůstane e-mail "".
    """
    domena = _smtp_domena()
    out = []
    for k, j in _oz_ciselnik():
        j = (j or "").strip()
        if len(j.split()) < 2:   # člověk = aspoň dvě slova (křestní + příjmení)
            continue
        out.append((k, j, _email_z_jmena(j, domena)))
    out.sort(key=lambda t: t[1].lower())
    return out


def _asm_ciselnik():
    """Unikátní hodnoty sloupce ASM."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor()
        cur.execute("SELECT DISTINCT asm FROM asm_dealer WHERE asm IS NOT NULL AND asm<>'' ORDER BY asm")
        out = [r[0] for r in cur.fetchall()]
        cur.close()
        return out
    except Exception as e:
        print(f"[asm] _asm_ciselnik: {e}")
        return []
    finally:
        conn.close()


def _kontakty_dle_ico(ico_list):
    """Dotáhne kontaktní údaje (PŮVODNÍ formát pro dlaždici Změna OZ) pro seznam IČO
    z asm_kontakty_oz. Vrací {ico: dict}. 'puvodni_provize' = 4. znak sloupce Dealer."""
    ico_list = [i for i in dict.fromkeys(ico_list) if i]
    if not ico_list:
        return {}
    conn = intranet_data.get_db_connection()
    if not conn:
        return {}
    try:
        cur = conn.cursor(dictionary=True)
        ph = ",".join(["%s"] * len(ico_list))
        cur.execute(f"SELECT * FROM asm_kontakty_oz WHERE ico IN ({ph})", tuple(ico_list))
        out = {}
        for r in cur.fetchall():
            dealer = r.get("dealer") or ""
            r["puvodni_provize"] = dealer[3] if len(dealer) >= 4 else ""
            out[r["ico"]] = r
        cur.close()
        return out
    except Exception as e:
        print(f"[asm] _kontakty_dle_ico: {e}")
        return {}
    finally:
        conn.close()


def _kod_pobocky(prodejni_doba):
    """Kód pobočky z „Prodejní kanál 6" („034Q JIP Plzeň" → „034").
    Bez rozpoznatelného 3místného čísla vrací "" — takový řádek se nefiltruje."""
    kod = str(prodejni_doba or "").strip()[:3]
    return kod if len(kod) == 3 and kod.isdigit() else ""


def _pobocka_uzivatele(user_id):
    """Pobočka přihlášeného uživatele (user.pobocka, VARCHAR(3)) nebo None."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return None
    try:
        cur = conn.cursor()
        cur.execute("SELECT pobocka FROM user WHERE iduser=%s", (user_id,))
        r = cur.fetchone()
        cur.close()
        return (r[0] or None) if r else None
    except Exception as e:
        print(f"[asm] _pobocka_uzivatele: {e}")
        return None
    finally:
        conn.close()


def _kontakty_dle_ico8(ico8):
    """Dotáhne kontakty pro jedno IČ 8 (zleva 8). Vrací (fakt, varianty):
      • fakt = reprezentativní řádek (identita je napříč variantami stejná) nebo None,
      • varianty = VŠECHNY IČ 10 pod tímto IČ 8 (list dict) — pro výběr v bublině.
    'puvodni_provize' = 4. znak sloupce Dealer.
    Když K.* (dodací) údaje v asm_kontakty chybí (soubor VO2 je neobsahoval),
    doplní se z asm_kontakty_oz (původní Kontaktní údaje VO) dle plného IČO —
    k_jmeno, k_ulice, k_mesto, k_psc + prodejni_doba (Prodejní kanál 6)."""
    ico8 = (ico8 or "").strip()
    if not ico8:
        return None, []
    conn = intranet_data.get_db_connection()
    if not conn:
        return None, []
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT * FROM asm_kontakty WHERE ico8=%s ORDER BY ico", (ico8,))
        rows = cur.fetchall()
        # Fallback K.* z asm_kontakty_oz pro řádky, kde v asm_kontakty chybí.
        _FALLBACK_POLE = ("k_jmeno", "k_ulice", "k_mesto", "k_psc", "prodejni_doba")

        def _chybi(v):
            return v is None or not str(v).strip()

        chybejici = [r["ico"] for r in rows
                     if any(_chybi(r.get(p)) for p in _FALLBACK_POLE)]
        if chybejici:
            ph = ",".join(["%s"] * len(chybejici))
            cur.execute(f"SELECT ico, k_jmeno, k_ulice, k_mesto, k_psc, prodejni_doba "
                        f"FROM asm_kontakty_oz WHERE ico IN ({ph})", tuple(chybejici))
            oz = {o["ico"]: o for o in cur.fetchall()}
            for r in rows:
                o = oz.get(r["ico"])
                if not o:
                    continue
                for p in _FALLBACK_POLE:
                    if _chybi(r.get(p)) and not _chybi(o.get(p)):
                        r[p] = o[p]
        cur.close()
        for r in rows:
            dealer = r.get("dealer") or ""
            r["puvodni_provize"] = dealer[3] if len(dealer) >= 4 else ""
        return (rows[0] if rows else None), rows
    except Exception as e:
        print(f"[asm] _kontakty_dle_ico8: {e}")
        return None, []
    finally:
        conn.close()


def _obrat_dle_ico(ico_list):
    """Dotáhne obraty (1m/3m) pro seznam IČ 8 z asm_obrat (GIST je klíčovaný dle IČ 8).
    Vrací {ic: dict}."""
    ico_list = [i for i in dict.fromkeys(ico_list) if i]
    if not ico_list:
        return {}
    conn = intranet_data.get_db_connection()
    if not conn:
        return {}
    try:
        cur = conn.cursor(dictionary=True)
        ph = ",".join(["%s"] * len(ico_list))
        cur.execute(f"SELECT ico, obrat_1m, obrat_3m FROM asm_obrat WHERE ico IN ({ph})",
                    tuple(ico_list))
        out = {r["ico"]: r for r in cur.fetchall()}
        cur.close()
        return out
    except Exception as e:
        print(f"[asm] _obrat_dle_ico: {e}")
        return {}
    finally:
        conn.close()


def _obrat_ic10_dle_ico(ico_list):
    """Dotáhne obraty (1m/3m) pro seznam IČ 10 z asm_obrat_ic10. Vrací {ic: dict}."""
    ico_list = [i for i in dict.fromkeys(ico_list) if i]
    if not ico_list:
        return {}
    conn = intranet_data.get_db_connection()
    if not conn:
        return {}
    try:
        cur = conn.cursor(dictionary=True)
        ph = ",".join(["%s"] * len(ico_list))
        cur.execute(
            f"SELECT ico, obrat_1m, obrat_3m FROM asm_obrat_ic10 WHERE ico IN ({ph})",
            tuple(ico_list))
        out = {r["ico"]: r for r in cur.fetchall()}
        cur.close()
        return out
    except Exception as e:
        print(f"[asm] _obrat_ic10_dle_ico: {e}")
        return {}
    finally:
        conn.close()


# ============================================================================
# E-maily (vzor Cenopřípad)
# ============================================================================
def _emaily_s_pravy(*role_keys):
    klice = [k for k in role_keys if k] + ["vse"]
    return list(dict.fromkeys(intranet_data.ziskej_emaily_s_pravem(*klice)))


def _emaily_office():
    return _emaily_s_pravy("asm_office")


def _emaily_spravce():
    # Varianta `asm_spravce_bez_emailu` se ZÁMĚRNĚ nezahrnuje do příjemců.
    return _emaily_s_pravy("asm_spravce")


def _app_url():
    try:
        u = intranet_data.APP_URL
        return f"{u}/asm" if u else ""
    except Exception:
        return ""


def _pripad_url(pripad_id):
    """Přímý odkaz na konkrétní případ (deep-link), který route `/asm` otevře rovnou
    v detailu (očičko) — příjemce se nemusí proklikávat frontou. Bez app_url vrací ""."""
    if not pripad_id:
        return _app_url()
    try:
        u = intranet_data.APP_URL
        return f"{u}/asm?pripad={int(pripad_id)}" if u else ""
    except Exception:
        return ""


def _posli_emaily_sync(prijemci, predmet, text, pripad_id=None):
    if pripad_id:
        odkaz = _pripad_url(pripad_id)
        popis = "Otevřít případ v portálu"
    else:
        odkaz = _app_url()
        popis = "Otevřít v portálu"
    if odkaz:
        text = f"{text}\n\n{popis}: {odkaz}"
    for p in prijemci:
        try:
            intranet_emaily.odesli_upozorneni_email(p, predmet, text)
        except Exception as e:
            print(f"[asm] e-mail {p}: {e}")


def _odesli_emaily(prijemci, predmet, text, pripad_id=None):
    prijemci = [p for p in dict.fromkeys(prijemci) if p and "@" in p]
    if not prijemci:
        return
    try:
        asyncio.create_task(asyncio.to_thread(_posli_emaily_sync, prijemci, predmet, text, pripad_id))
    except RuntimeError:
        _posli_emaily_sync(prijemci, predmet, text, pripad_id)


def _smtp_domena():
    """Firemní e-mailová doména = část za '@' v odesílací adrese (smtp_user).
    Bez konfigurace vrací "" → e-maily OZ se nedopočítají."""
    try:
        su = (intranet_data.nacti_smtp().get("smtp_user") or "").strip()
    except Exception:
        su = ""
    return su.split("@", 1)[1].lower() if "@" in su else ""


def _email_z_jmena(jmeno, domena=None):
    """„Jan Novák" → jan.novak@<firemní doména>. Diakritika se odstraní.
    Bez příjmení nebo bez domény vrací "" (notifikace se pak zahodí)."""
    if domena is None:
        domena = _smtp_domena()
    if not domena:
        return ""
    prevod = {"č": "c", "ř": "r", "š": "s", "ž": "z", "ý": "y", "á": "a",
              "í": "i", "é": "e", "ě": "e", "ú": "u", "ů": "u", "ó": "o",
              "ď": "d", "ť": "t", "ň": "n"}
    zaklad = "".join(prevod.get(ch, ch) for ch in str(jmeno or "").strip().lower())
    casti = [c for c in "".join(
        ch if ch.isalnum() or ch.isspace() else " " for ch in zaklad).split() if c]
    if len(casti) < 2:
        return ""
    return f"{casti[0]}.{casti[-1]}@{domena}"


def _email_oz(cislo_oz):
    """Firemní e-mail OZ dle čísla OZ (asm_dealer.klic). "" když není/nedopočítán."""
    cislo_oz = (str(cislo_oz or "")).strip()
    if not cislo_oz:
        return ""
    conn = intranet_data.get_db_connection()
    if not conn:
        return ""
    try:
        cur = conn.cursor()
        cur.execute("SELECT email FROM asm_dealer WHERE klic=%s", (cislo_oz,))
        r = cur.fetchone()
        cur.close()
        return (r[0] or "") if r and r[0] and "@" in r[0] else ""
    except Exception as e:
        print(f"[asm] _email_oz: {e}")
        return ""
    finally:
        conn.close()


OZ_MAIL_DOMENA = "jip-potraviny.cz"


def _email_oz_dle_cisla(cislo_oz):
    """Číslo OZ → oz<číslo>@jip-potraviny.cz (přizvaní OZ v „+"; DB sloupec se neřeší).
    "" když je číslo prázdné nebo není alfanumerické."""
    c = str(cislo_oz or "").strip()
    if not c or not c.isalnum():
        return ""
    return f"oz{c.lower()}@{OZ_MAIL_DOMENA}"


def _email_uzivatele(uid):
    if not uid:
        return None
    conn = intranet_data.get_db_connection()
    if not conn:
        return None
    try:
        cur = conn.cursor()
        cur.execute("SELECT email FROM user WHERE iduser=%s", (uid,))
        r = cur.fetchone()
        cur.close()
        return r[0] if r and r[0] and "@" in r[0] else None
    except Exception as e:
        print(f"[asm] _email_uzivatele: {e}")
        return None
    finally:
        conn.close()


def _odesli_emaily_zadateli(pripad, predmet, text):
    em = _email_uzivatele(pripad.get("zadavatel_id"))
    if em:
        _odesli_emaily([em], predmet, text, pripad_id=pripad.get("id"))


def _notifikuj_oz(cisla_oz, cislo, nazev, user_name, detail, pripad_id=None,
                  dle_cisla=False):
    """Upozorní dotčené OZ (původního i nového) na nový/změněný případ.
    `cisla_oz` je seznam čísel OZ; nedopočítané/prázdné e-maily se zahodí.
    `dle_cisla=True` → adresa se skládá jako oz<číslo>@doména (přizvaní OZ v „+"),
    jinak se bere e-mail z asm_dealer."""
    resolver = _email_oz_dle_cisla if dle_cisla else _email_oz
    prijemci = list(dict.fromkeys(em for em in (resolver(c) for c in (cisla_oz or [])) if em))
    if not prijemci:
        return
    _odesli_emaily(
        prijemci,
        f"Formuláře ASM — dotčený případ {cislo} ({nazev})",
        f"Vaše OZ je dotčeno případem, který podal {user_name}.\n\n"
        f"Číslo: {cislo}\nTyp: {nazev}\nDetail: {detail}\n",
        pripad_id=pripad_id,
    )


# ============================================================================
# Historie / průběh případu
# ============================================================================
def zaznam_historie(pripad_id, akce, kdo, detail=None):
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.execute("INSERT INTO asm_historie (pripad_id, akce, detail, kdo) "
                    "VALUES (%s,%s,%s,%s)",
                    (pripad_id, akce[:80], (detail or "")[:1000], (kdo or "")[:255]))
        conn.commit()
        cur.close()
    except Exception as e:
        print(f"[asm] zaznam_historie: {e}")
    finally:
        conn.close()


def nacti_historie(pripad_id):
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT akce, detail, kdo, kdy FROM asm_historie "
                    "WHERE pripad_id=%s ORDER BY id", (pripad_id,))
        out = cur.fetchall()
        cur.close()
        return out
    except Exception:
        return []
    finally:
        conn.close()


# ============================================================================
# Případy — CRUD a workflow
# ============================================================================
def zaloz_pripad(hlavicka, radky, user_id, user_name):
    """Uloží nový případ + řádky. Vrací (pripad_id, cislo) nebo (None, chyba)."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return None, "Není připojení k databázi."
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO asm_pripady
                (duvod_zmeny, asm_jmeno, zmena_asm_jmeno, novy_asm_jmeno,
                 cislo_oz, jmeno_oz, cislo_novy_oz, jmeno_novy_oz,
                 datum_zmeny_od, zakaznik_v_regionu_oz, oduvodneni,
                 zadavatel_id, zadavatel_jmeno, stav, pocet_radku)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'odeslano',%s)
        """, (
            hlavicka.get("duvod_zmeny", "")[:1000],
            hlavicka.get("asm_jmeno", "")[:255],
            hlavicka.get("zmena_asm_jmeno", "")[:255],
            hlavicka.get("novy_asm_jmeno", "")[:255],
            hlavicka.get("cislo_oz", "")[:40], hlavicka.get("jmeno_oz", "")[:255],
            hlavicka.get("cislo_novy_oz", "")[:40], hlavicka.get("jmeno_novy_oz", "")[:255],
            hlavicka.get("datum_zmeny_od", "")[:20],
            hlavicka.get("zakaznik_v_regionu_oz", "")[:5],
            hlavicka.get("oduvodneni", "")[:1000],
            user_id, user_name[:255], len(radky),
        ))
        pid = cur.lastrowid
        cislo = f"ASM{pid:05d}"
        cur.execute("UPDATE asm_pripady SET cislo=%s WHERE id=%s", (cislo, pid))
        if radky:
            data = []
            for i, r in enumerate(radky, 1):
                data.append((
                    pid, i, r.get("ico", "")[:40],
                    _str(r.get("jmeno")), _str(r.get("k_jmeno")), _str(r.get("k_jmeno2")),
                    _str(r.get("k_ulice")), _str(r.get("k_mesto")), _str(r.get("k_psc"), 40),
                    _str(r.get("dealer"), 40), _str(r.get("id_kontakt"), 40),
                    _str(r.get("prodejni_doba"), 120),
                    (r.get("puvodni_provize") or "")[:10],
                    (str(r.get("nova_provize") or ""))[:40],
                    (str(r.get("oz2") or ""))[:120],
                    (str(r.get("datum_od") or ""))[:20],
                    (str(r.get("datum_do") or ""))[:20],
                ))
            cur.executemany("""
                INSERT INTO asm_radky
                    (pripad_id, poradi, ico, jmeno, k_jmeno, k_jmeno2, k_ulice,
                     k_mesto, k_psc, dealer, id_kontakt, prodejni_doba,
                     puvodni_provize, nova_provize, oz2, datum_od, datum_do)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, data)
        conn.commit()
        cur.close()
        return pid, cislo
    except Exception as e:
        conn.rollback()
        print(f"[asm] zaloz_pripad: {e}")
        return None, f"Chyba uložení: {e}"
    finally:
        conn.close()


def nacti_pripady(user_id, prava, formular="oz_zmena"):
    """Seznam případů daného formuláře dle viditelnosti. Office/správce celou frontu,
    žadatel jen své. `formular` filtruje typ formuláře (každá dlaždice má vlastní frontu).
    Starší případy bez vyplněného sloupce formular = 'oz_zmena'."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        kde = "COALESCE(formular,'oz_zmena')=%s"
        if _vidi_vsechny_pripady(prava):
            cur.execute(f"SELECT * FROM asm_pripady WHERE {kde} ORDER BY id DESC", (formular,))
        else:
            cur.execute(f"SELECT * FROM asm_pripady WHERE {kde} AND zadavatel_id=%s "
                        f"ORDER BY id DESC", (formular, user_id))
        out = cur.fetchall()
        cur.close()
        return out
    except Exception as e:
        print(f"[asm] nacti_pripady: {e}")
        return []
    finally:
        conn.close()


def zaloz_pripad_formular(formular, data, user_id, user_name):
    """Uloží nový případ unifikovaného formuláře (limity/dodaci) do data_json.
    Vrací (pripad_id, cislo) nebo (None, chyba)."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return None, "Není připojení k databázi."
    try:
        cur = conn.cursor()
        cur.execute("INSERT INTO asm_pripady (formular, data_json, zadavatel_id, "
                    "zadavatel_jmeno, stav, pocet_radku) VALUES (%s,%s,%s,%s,'odeslano',0)",
                    (formular, json.dumps(data, ensure_ascii=False), user_id, user_name[:255]))
        pid = cur.lastrowid
        cislo = f"ASM{pid:05d}"
        cur.execute("UPDATE asm_pripady SET cislo=%s WHERE id=%s", (cislo, pid))
        conn.commit()
        cur.close()
        return pid, cislo
    except Exception as e:
        conn.rollback()
        print(f"[asm] zaloz_pripad_formular: {e}")
        return None, f"Chyba uložení: {e}"
    finally:
        conn.close()


def uloz_data_formular(pid, data):
    """Aktualizuje data_json existujícího případu (po vrácení k opravě)."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return False
    try:
        cur = conn.cursor()
        cur.execute("UPDATE asm_pripady SET data_json=%s WHERE id=%s",
                    (json.dumps(data, ensure_ascii=False), pid))
        conn.commit()
        cur.close()
        return True
    except Exception as e:
        conn.rollback()
        print(f"[asm] uloz_data_formular: {e}")
        return False
    finally:
        conn.close()


def nacti_pripad(pid):
    conn = intranet_data.get_db_connection()
    if not conn:
        return None
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT * FROM asm_pripady WHERE id=%s", (pid,))
        p = cur.fetchone()
        cur.close()
        return p
    except Exception:
        return None
    finally:
        conn.close()


def nacti_radky(pid):
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT * FROM asm_radky WHERE pripad_id=%s ORDER BY poradi", (pid,))
        out = cur.fetchall()
        cur.close()
        return out
    except Exception:
        return []
    finally:
        conn.close()


def uloz_radky(pid, radky):
    """Uloží upravené řádky existujícího případu (po vrácení k opravě). Aktualizuje
    podle id; aktualizuje i počet řádků případu."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return False
    try:
        cur = conn.cursor()
        for r in radky:
            rid = r.get("id")
            if rid is None:
                continue
            cur.execute("""
                UPDATE asm_radky SET
                    ico=%s, jmeno=%s, k_jmeno=%s, k_jmeno2=%s, k_ulice=%s, k_mesto=%s,
                    k_psc=%s, dealer=%s, id_kontakt=%s, prodejni_doba=%s,
                    puvodni_provize=%s, nova_provize=%s, oz2=%s, datum_od=%s, datum_do=%s
                WHERE id=%s AND pripad_id=%s
            """, (
                (str(r.get("ico") or ""))[:40], _str(r.get("jmeno")), _str(r.get("k_jmeno")),
                _str(r.get("k_jmeno2")), _str(r.get("k_ulice")), _str(r.get("k_mesto")),
                _str(r.get("k_psc"), 40), _str(r.get("dealer"), 40), _str(r.get("id_kontakt"), 40),
                _str(r.get("prodejni_doba"), 120), (str(r.get("puvodni_provize") or ""))[:10],
                (str(r.get("nova_provize") or ""))[:40], (str(r.get("oz2") or ""))[:120],
                (str(r.get("datum_od") or ""))[:20], (str(r.get("datum_do") or ""))[:20],
                rid, pid,
            ))
        cur.execute("UPDATE asm_pripady SET pocet_radku=("
                    "SELECT COUNT(*) FROM asm_radky WHERE pripad_id=%s) WHERE id=%s", (pid, pid))
        conn.commit()
        cur.close()
        return True
    except Exception as e:
        conn.rollback()
        print(f"[asm] uloz_radky: {e}")
        return False
    finally:
        conn.close()


def smaz_pripad(pid):
    """Nevratně smaže případ vč. řádků a historie. Jen pro administrátora modulu."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return False
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM asm_radky WHERE pripad_id=%s", (pid,))
        cur.execute("DELETE FROM asm_historie WHERE pripad_id=%s", (pid,))
        cur.execute("DELETE FROM asm_pripady WHERE id=%s", (pid,))
        conn.commit()
        cur.close()
        return True
    except Exception as e:
        conn.rollback()
        print(f"[asm] smaz_pripad: {e}")
        return False
    finally:
        conn.close()


def zmen_stav(pid, novy_stav, poznamka=None, zamitnuti_duvod=None, storno_duvod=None,
              spravce_pozn=None):
    conn = intranet_data.get_db_connection()
    if not conn:
        return False
    try:
        cur = conn.cursor()
        sety, params = ["stav=%s"], [novy_stav]
        if poznamka is not None:
            sety.append("poznamka=%s")
            params.append(poznamka[:1000])
        if zamitnuti_duvod is not None:
            sety.append("zamitnuti_duvod=%s")
            params.append(zamitnuti_duvod[:1000])
        if storno_duvod is not None:
            sety.append("storno_duvod=%s")
            params.append(storno_duvod[:1000])
        if spravce_pozn is not None:
            sety.append("spravce_pozn=%s")
            params.append(spravce_pozn[:1000])
        params.append(pid)
        cur.execute(f"UPDATE asm_pripady SET {','.join(sety)} WHERE id=%s", params)
        conn.commit()
        cur.close()
        return True
    except Exception as e:
        print(f"[asm] zmen_stav: {e}")
        return False
    finally:
        conn.close()


# ============================================================================
# UI — refresh / navigace
# ============================================================================
def _refresh():
    fn = app.storage.client.get("_asm_refresh")
    if fn:
        fn()


def _nav(kam):
    app.storage.user["asm_pohled"] = kam
    app.storage.user["asm_sekce"] = "seznam"
    _refresh()


def _badge(stav):
    popis, barva = _STAV_BADGE.get(stav, (stav, "grey"))
    ui.badge(popis, color=barva).props("rounded").classes("px-2 py-1 text-xs")


def _date_input(label, value=""):
    """Textové pole DD.MM.RRRR s našeptávacím kalendářem (lze i psát/Ctrl+V)."""
    inp = ui.input(label, value=value).props("outlined dense").classes("w-full")
    with inp.add_slot("append"):
        ui.icon("event").classes("cursor-pointer text-blue-500").on("click", lambda: menu.open())
    with ui.menu().props("no-parent-event") as menu:
        ui.date(mask="DD.MM.YYYY").bind_value(inp).props("today-btn").classes("p-0")
    return inp


# ============================================================================
# Uživatelská příručka (manuál) — 1:1 dle dokumentu
# ============================================================================
ASM_MANUAL_MD = """\
# Formuláře ASM

**Uživatelská příručka**

*Změna zákazníků na OZ/ASM · Navýšení limitů a splatností · Změna dodacích listů*

![Úvodní obrazovka modulu Formuláře ASM s přehledem všech tří formulářů.](/asm_prirucka/image1.png)

*Úvodní obrazovka modulu Formuláře ASM s přehledem všech tří formulářů.*

## Obsah

1. **Úvod** – k čemu modul slouží a kdo s ním pracuje
2. **Vstup do aplikace** – jak se dostat k formulářům
3. **Společné principy** – fronta, stavy, notifikace, export
4. **Změna zákazníků na OZ/ASM** – přeřazení zákazníků a změna provize
5. **Navýšení limitů a splatností** – žádost o limit a splatnost
6. **Změna dodacích listů** – úprava splatnosti dodacích listů
7. **Schvalovací proces (workflow)** – cesta žádosti od ASM po vyřízení
8. **Často kladené otázky** – tipy a řešení běžných situací

## 1. Úvod

Modul „Formuláře ASM“ slouží k elektronickému zadávání a schvalování tří typů žádostí, které ASM (area sales manažeři) potřebují řešit v souvislosti se svými zákazníky:

- **Změna zákazníků na OZ/ASM** – přeřazení zákazníka pod jiného obchodního zástupce (OZ) nebo ASM a úprava provize.
- **Navýšení limitů a splatností** – žádost o navýšení kreditního limitu nebo prodloužení splatnosti faktur.
- **Změna dodacích listů** – úprava splatnosti vázané na dodací listy (DL).

Všechny tři formuláře fungují na stejném principu: ASM zadá žádost, ta se zařadí do fronty, a odpovědné oddělení (Office / správa zákaznických dat) ji zkontroluje, zpracuje, vrátí k opravě nebo postoupí ke schválení nadřízenému správci.

## 2. Vstup do aplikace

Na hlavní obrazovce aplikace klikněte na dlaždici „Formuláře ASM“ a poté na tlačítko „Otevřít přehled“.

![Dlaždice „Formuláře ASM“ na hlavní obrazovce.](/asm_prirucka/image2.png)

*Dlaždice „Formuláře ASM“ na hlavní obrazovce.*

Otevře se rozcestník se třemi dlaždicemi formulářů. Kliknutím na příslušnou dlaždici se otevře seznam (fronta) daného typu žádostí.

![Rozcestník: Změna zákazníků na OZ/ASM, Navýšení limitů a splatností, Změna dodacích listů.](/asm_prirucka/image1.png)

*Rozcestník: Změna zákazníků na OZ/ASM, Navýšení limitů a splatností, Změna dodacích listů.*

## 3. Společné principy

Než přejdeme k jednotlivým formulářům, je dobré znát pravidla, která platí pro všechny tři:

### 3.1 Vytvoření nové žádosti

V seznamu klikněte na tlačítko „+ Nový formulář“ vlevo nahoře. Otevře se prázdný formulář k vyplnění.

### 3.2 Načítání údajů o zákazníkovi podle IČO

Ve formulářích „Navýšení limitů a splatností“ a „Změna dodacích listů“ stačí zadat IČO zákazníka a stisknout Tab/Enter – ostatní identifikační pole (název, adresa, provozovna …) se doplní automaticky z evidence kontaktních údajů.

Ve formuláři „Změna zákazníků na OZ/ASM“ se IČO vkládají rovnou do detailní tabulky (lze vložit více IČO najednou, jedno na řádek, i zkopírovat z Excelu) a tlačítkem „Načíst IČO“ se doplní kontaktní údaje a původní provize ke všem řádkům.

### 3.3 Rozbalovací pole

Pole s šipkou dolů (právní forma, smlouva se zajištěním/ručením, číslo OZ apod.) jsou číselníky – hodnotu vyberte ze seznamu, ručně se nepřepisují.

### 3.4 Odeslání žádosti

Po vyplnění formuláře klikněte na „Odeslat“ vpravo dole. Žádost se zařadí do fronty se stavem „Odesláno / ve frontě“ a automaticky se na ni doplní jméno žadatele a datum zadání.

### 3.5 Stavy žádosti a fronta

Každá žádost prochází několika stavy podle toho, v jaké fázi zpracování se nachází:

| Stav | Co znamená |
| --- | --- |
| **Odesláno / ve frontě** | Formulář byl odeslán a čeká na zpracování na centrále (Office / správce obchodu). |
| **Zpracováno** | Office žádost zkontroloval, schválil a provedl požadovanou změnu. Žadateli přišel e-mail. |
| **Vráceno k opravě** | V žádosti je chyba nebo nejasnost. ASM dostal e-mail s poznámkou, co má opravit, a může formulář upravit a odeslat znovu. |
| **Postoupeno správci** | Office žádost postoupil ke konečnému schválení správci obchodu (např. u sporných/velkých změn). |
| **Zamítnuto** | Správce obchodu žádost zamítl. Žadateli přišel informační e-mail. |
| **Stornováno** | Žádost byla zrušena a dále se nezpracovává. |

### 3.6 Filtrování a vyhledávání

V seznamu žádostí lze filtrovat podle stavu (rozbalovací pole „Stav“) a vyhledávat podle ASM, OZ…..

### 3.7 Detail žádosti

Kliknutím na ikonu oka u řádku v seznamu se otevře detail žádosti se všemi zadanými údaji. V detailu jsou k dispozici akce:

- **Export do Excelu** – stáhne kompletní obsah žádosti do přehledné excelové tabulky.
- **Zpracováno** – označí žádost jako vyřízenou.
- **Vrátit k opravě** – vrátí žádost zpět ASM s poznámkou, co je třeba opravit.
- **Postoupit správci** – posune žádost ke konečnému schválení správci obchodu. Vyžaduje poznámku (co žádost nesplňuje a proč ji přesto schválit); poznámku vidí jen Office a správce, žadateli se nezobrazuje.
- **Stornovat** – žádost se zruší a dál se neřeší.

Dole v detailu je sekce „Průběh případu“, kde je vidět historie všech kroků a poznámek k žádosti.

### 3.8 E-mailové notifikace

Aplikace automaticky posílá e-mail při každé důležité události: odeslání nové žádosti (informace pro Office), vrácení k opravě (informace pro ASM), zpracování (informace pro žadatele), postoupení správci a zamítnutí. Díky tomu nemusíte stav žádosti aktivně hlídat.

## 4. Změna zákazníků na OZ/ASM

Tento formulář použijte, pokud potřebujete přeřadit jednoho nebo více zákazníků pod jiného obchodního zástupce (OZ), případně i pod jiného ASM, a/nebo upravit výši provize.

### 4.1 Otevření formuláře

V dlaždici „Změna zákazníků na OZ/ASM“ klikněte na „+ Nový formulář“.

![Horní část formuláře – hlavička, změna OZ a změna provize.](/asm_prirucka/image3.png)

*Horní část formuláře – hlavička, změna OZ a změna provize.*

### 4.2 Vyplnění hlavičky

- **Důvod změny** – povinné pole, stručně popište důvod přeřazení.
- **Jméno ASM** a **Změna ASM jméno** – doplní se automaticky podle přihlášeného uživatele (z logu).
- **Nový ASM jméno** – vyberte z číselníku, pokud se zákazníci přeřazují i pod jiného ASM.

### 4.3 Změna OZ

- **Číslo OZ** – vyberte původního obchodního zástupce z číselníku, jméno OZ se doplní automaticky.
- **Číslo nový OZ** – vyberte nového obchodního zástupce, jméno se opět doplní automaticky.

### 4.4 Změna provize

- **Datum změny od** – zadejte datem nebo vyberte v kalendáři.
- **Zákazník v regionu OZ** – vyberte ANO/NE (zodpovídá za správnost ASM).
- **Případné odůvodnění** – nepovinné doplňující pole.

### 4.5 Detailní tabulka zákazníků

![Detailní tabulka zákazníků a tlačítka Načíst IČO / Přidat řádek / Smazat vybrané.](/asm_prirucka/image4.png)

*Detailní tabulka zákazníků a tlačítka Načíst IČO / Přidat řádek / Smazat vybrané.*

- Do pole „Seznam IČO“ vložte jedno nebo více IČO (každé na nový řádek). Lze vložit ruční zápis i hromadně zkopírovat sloupec IČO z Excelu pomocí Ctrl+C / Ctrl+V.
- Klikněte na „Načíst IČO“ – do tabulky se pro každé IČO automaticky doplní jméno, město, dealer a původní provize z evidence kontaktních údajů.
- U každého řádku ručně doplňte: Novou provizi, OZ2 (pokud se týká), Datum od a Datum do.
- Tlačítkem „Přidat řádek“ lze přidat další prázdný řádek, tlačítkem „Smazat vybrané“ odstraníte označené řádky.

**Tip:** Tabulka zvládá i velké objemy zákazníků (v praxi i řádově tisíce řádků), všechna pole jdou kopírovat a vkládat klávesovými zkratkami Ctrl+C / Ctrl+V.

### 4.6 Odeslání a sledování žádosti

Po kontrole klikněte na „Odeslat“. Žádost se objeví v seznamu se stavem „Odesláno / ve frontě“, číslem (např. ASM00010) a souhrnem (kdo, odkud kam, počet řádků).

![Seznam žádostí o změnu zákazníků na OZ/ASM.](/asm_prirucka/image5.png)

*Seznam žádostí o změnu zákazníků na OZ/ASM.*

Kliknutím na ikonu oka otevřete detail žádosti se všemi údaji a celou detailní tabulkou zákazníků.

![Detail žádosti – hlavička, detailní tabulka zákazníků a akční tlačítka.](/asm_prirucka/image6.png)

*Detail žádosti – hlavička, detailní tabulka zákazníků a akční tlačítka.*

## 5. Navýšení limitů a splatností

Tento formulář použijte, pokud zákazník potřebuje navýšit kreditní limit (saldo) nebo prodloužit splatnost faktur za zboží či cigarety.

### 5.1 Otevření formuláře

V dlaždici „Navýšení limitů a splatností“ klikněte na „+ Nový formulář“.

![Formulář Navýšení limitů a splatností.](/asm_prirucka/image7.png)

*Formulář Navýšení limitů a splatností.*

### 5.2 Obecné informace a identifikace

Zadejte IČO zákazníka a stiskněte Tab/Enter – automaticky se doplní pobočka, fakturační název, místo dodání, adresa.

### 5.3 Současný stav a ekonomické ukazatele

Tato sekce se doplní automaticky z evidence (aktuální kreditní limit, aktuální splatnost 1 – zboží, aktuální splatnost 2 – cigarety, obrat za poslední měsíc, obrat za poslední 3 měsíce). Slouží jako podklad pro posouzení žádosti.

### 5.4 Požadovaná změna limitu / splatnosti

- **Návrh nového kreditu (limitu)** – zadejte požadovanou výši.
- **Návrh nové splatnosti 1 (zboží)** a **Návrh nové splatnosti 2 (cigarety)** – zadejte počet dnů.
- **Zdůvodnění požadované změny / žádosti** – vysvětlete, proč o navýšení žádáte.

### 5.5 Odeslání a sledování žádosti

Klikněte na „Odeslat“. Žádost se zařadí do fronty pod číslem ve formátu ASM000xx se jménem zákazníka, IČO a aktuálním stavem.

![Seznam žádostí o navýšení limitů a splatností.](/asm_prirucka/image8.png)

*Seznam žádostí o navýšení limitů a splatností.*

V detailu žádosti vidíte zadané i automaticky dotažené hodnoty vedle sebe – usnadňuje to rychlé posouzení Office.

![Detail žádosti o navýšení limitů a splatností.](/asm_prirucka/image9.png)

*Detail žádosti o navýšení limitů a splatností.*

## 6. Změna dodacích listů

Tento formulář použijte, pokud potřebujete upravit splatnost vázanou na dodací listy (DL).

### 6.1 Otevření formuláře

V dlaždici „Změna dodacích listů“ klikněte na „+ Nový formulář“.

![Formulář Změna dodacích listů.](/asm_prirucka/image10.png)

*Formulář Změna dodacích listů.*

### 6.2 Obecné informace a identifikace

Stejně jako u navýšení limitů zadejte IČO zákazníka a stiskněte Tab/Enter – pobočka, fakturační název, místo dodání, adresa se doplní automaticky.

### 6.3 Specifické údaje pro dodací listy (DL)

- **Režim fakturace** – vyberte z číselníku.
- **Zřízená el. fakturace (ANO/NE)** – vyberte z rozbalovacího pole.

### 6.4 Současný stav a požadovaná změna

- **Splatnost 1 (zboží)** a **Splatnost 2 (cigarety)** – aktuální hodnoty (sekce 3).
- **Návrh nové splatnosti 1 a 2** a **Zdůvodnění požadované změny / žádosti** (sekce 4).

### 6.5 Odeslání a sledování žádosti

Klikněte na „Odeslat“. Žádost se zařadí do fronty se stavem „Odesláno / ve frontě“.

![Seznam žádostí o změnu dodacích listů.](/asm_prirucka/image11.png)

*Seznam žádostí o změnu dodacích listů.*

![Detail žádosti o změnu dodacích listů.](/asm_prirucka/image12.png)

*Detail žádosti o změnu dodacích listů.*

## 7. Schvalovací proces (workflow)

Všechny tři formuláře sdílí stejnou logiku schvalování:

- ASM vyplní a odešle formulář → žádost se zařadí do fronty se stavem „Odesláno / ve frontě“ a na Office (oddělení správy zákaznických dat) přijde e-mail o nové žádosti.
- Office žádost zkontroluje. Má tři možnosti:
  - **Zpracovat** – schválí a provede změnu, žadateli odejde e-mail o zpracování; na konci dne se případ uzavírá.
  - **Vrátit k opravě** – pošle žádost zpět ASM s poznámkou, co opravit; ASM e-mail obdrží, opraví formulář a odešle znovu.
  - **Postoupit správci** – pokud je žádost sporná nebo nad rámec běžného schválení, postoupí ji ke konečnému rozhodnutí. Otevře se okno s povinnou poznámkou pro správce – napište, co žádost nesplňuje a proč ji přesto schválit. Poznámka dorazí správci v e-mailu i v detailu případu; žadatel ji nevidí.
- Pokud ASM vrátí opravenou žádost a Office na ní stále vidí nesrovnalost, může ji znovu postoupit správci obchodu.
- Správce obchodu žádost buď zamítne (žadateli odejde e-mail o zamítnutí a žádost končí ve stavu „Zamítnuto“), nebo ji schválí (žadateli odejde e-mail a žádost se vrací na Office jako schválená správcem ke zpracování).

**Pozn.:** Role v procesu: **Žadatel** = ASM / vedoucí poboček. **Office** (správa zákaznických dat) = kontroluje a zpracovává. **Správce obchodu** = konečné schválení sporných případů.

## 8. Často kladené otázky

#### Proč se mi po zadání IČO nedoplnily údaje o zákazníkovi?

Zkontrolujte, že jste po zadání IČO stiskli Tab nebo Enter, překlikli mimo pole (samotné napsání čísla údaje nedotáhne). Pokud se údaje stále nenačtou, zákazník zřejmě není v aktuální evidenci kontaktních údajů aktivní – obraťte se na Office.

#### Mohu do formuláře vložit více zákazníků najednou?

Ano, ale pouze ve formuláři „Změna zákazníků na OZ/ASM“ – do pole „Seznam IČO“ vložte jedno IČO na řádek (lze i hromadně přes Ctrl+V z Excelu) a klikněte na „Načíst IČO“.

#### Jak poznám, že byla moje žádost vyřízena?

Obdržíte e-mail při každé změně stavu žádosti (zpracování, vrácení k opravě, postoupení správci, zamítnutí). Stav lze kdykoliv zkontrolovat i přímo v seznamu žádostí.

#### Co mám dělat, když je žádost vrácena k opravě?

Otevřete detail žádosti, v sekci „Průběh případu“ najdete poznámku, co je třeba opravit, formulář upravte a odešlete znovu.
"""


def _dialog_manual():
    """Zobrazí uživatelskou příručku (markdown) v rolovatelném dialogu."""
    with ui.dialog() as dlg, ui.card().style(
            "max-width: 96vw; width: 900px; max-height: 90vh; overflow-y: auto"):
        with ui.row().classes("w-full items-center gap-2 sticky top-0 bg-white z-10 pb-2 "
                              "border-b border-gray-200"):
            ui.icon("menu_book", size="1.6rem").classes("text-emerald-600")
            ui.label("Uživatelská příručka").classes("text-xl font-bold")
            ui.space()
            ui.button(icon="close", on_click=dlg.close) \
                .props("flat round dense").classes("text-gray-500")
        ui.markdown(ASM_MANUAL_MD).classes("w-full overflow-x-auto text-sm leading-relaxed")
    dlg.open()


# ============================================================================
# UI — vstupní bod a rozcestník
# ============================================================================
async def vykresli_asm(user_id, user_name, vsechna_prava):
    @ui.refreshable
    async def _obsah():
        await _vykresli_asm(user_id, user_name, vsechna_prava)

    app.storage.client["_asm_refresh"] = _obsah.refresh
    await _obsah()


async def _vykresli_asm(user_id, user_name, vsechna_prava):
    # DB dotazy běží ve vlákně — nedrží event loop celého serveru
    await asyncio.to_thread(inicializace_asm_db)
    pohled = app.storage.user.get("asm_pohled")
    if pohled == "data" and not _vidi_import(vsechna_prava):
        pohled = None
    if pohled in ("zmena", "limity", "dodaci") and not _je_zadatel(vsechna_prava):
        pohled = None
    if pohled == "schuzky" and not _vidi_schuzky(vsechna_prava):
        pohled = None

    # Deep-link z e-mailu (/asm?pripad=<id>): přepni na správnou frontu a otevři detail,
    # aby se posuzovatel neproklikával. Zpracuje se jen jednou (pop).
    deep_pid = app.storage.user.pop("asm_deep_pripad", None)
    if deep_pid and _je_zadatel(vsechna_prava):
        dp = await asyncio.to_thread(nacti_pripad, deep_pid)
        if dp:
            _f = dp.get("formular") or "oz_zmena"
            pohled = "zmena" if _f == "oz_zmena" else _f
            app.storage.user["asm_pohled"] = pohled
            ui.timer(0.25, lambda pid=deep_pid: _detail_dialog(pid, user_id, user_name, vsechna_prava),
                     once=True)

    with ui.row().classes("w-full items-center gap-3 mb-6"):
        if pohled:
            ui.button(icon="arrow_back", on_click=lambda: _nav(None)) \
                .props("flat round color=grey-7").tooltip("Zpět na přehled")
        ui.icon("description", size="2.2rem").classes("text-emerald-600")
        with ui.column().classes("gap-0"):
            ui.label("Formuláře ASM").classes("text-3xl font-extrabold text-gray-800")
            podtitul = "Formuláře pro správu zákazníků"
            if pohled == "zmena":
                podtitul = "Změna zákazníků na OZ/ASM"
            elif pohled == "data":
                podtitul = "Data — týdenní import číselníků"
            elif pohled == "schuzky":
                podtitul = "Rezervace ind. schůzky s vedoucími"
            elif pohled in _FORMULARE:
                podtitul = _FORMULARE[pohled]["nazev"]
            ui.label(podtitul).classes("text-sm text-gray-500")
        ui.space()
        # Manuál popisuje formuláře ASM — v sekci schůzek nedává smysl.
        if pohled != "schuzky":
            ui.button("Manuál", icon="menu_book", on_click=_dialog_manual) \
                .props("outline no-caps") \
                .classes("text-emerald-700 font-semibold rounded-lg") \
                .tooltip("Uživatelská příručka modulu Formuláře ASM.")

    if pohled == "schuzky":
        # vlastní modul, jen bydlí v rozcestníku Formulářů ASM (hlavičku dělá ASM)
        intranet_schuzky.vykresli_schuzky(user_id, user_name,
                                          app.storage.user.get("user_email", ""),
                                          vsechna_prava, s_hlavickou=False)
        return
    if pohled == "data":
        _panel_data(user_name)
        return
    if pohled == "zmena":
        await _view_zmena(user_id, user_name, vsechna_prava)
        return
    if pohled in _FORMULARE:
        await _view_formular(pohled, user_id, user_name, vsechna_prava)
        return

    # Rozcestník dlaždic
    # Hláška o datech se týká jen formulářů — kdo sem chodí jen na schůzky, nemá ji vidět.
    if _je_zadatel(vsechna_prava) or _vidi_import(vsechna_prava):
        with ui.row().classes(
                "w-full items-center gap-2 mb-6 px-4 py-3 rounded-lg "
                "bg-emerald-50 border border-emerald-200 text-emerald-800"):
            ui.icon("info", size="1.4rem").classes("text-emerald-600")
            ui.label("Interní data o zákaznících a dealer jsou aktuální vždy "
                     "k poslednímu pátku.").classes("text-sm font-medium")

    with ui.row().classes("w-full gap-6 flex-wrap pt-2"):
        if _je_zadatel(vsechna_prava):
            _tile("👥", "Změna zákazníků na OZ/ASM", "formulář + fronta",
                  "border-emerald-200", lambda: _nav("zmena"))
            _tile("💳", "Navýšení limitů a splatností", "formulář + fronta",
                  "border-amber-200", lambda: _nav("limity"))
            _tile("🚚", "Změna splatnosti dodacích listů", "formulář + fronta",
                  "border-cyan-200", lambda: _nav("dodaci"))
        if _vidi_import(vsechna_prava):
            _tile("⬆️", "Data", "Dealer + Kontaktní údaje VO",
                  "border-blue-200", lambda: _nav("data"))
        if _vidi_schuzky(vsechna_prava):
            _tile("🗓️", "Schůzky s vedoucími", "rezervace individuální schůzky",
                  "border-indigo-200", lambda: _nav("schuzky"))
    if (not _je_zadatel(vsechna_prava) and not _vidi_import(vsechna_prava)
            and not _vidi_schuzky(vsechna_prava)):
        with ui.column().classes("items-center py-20 gap-3 w-full"):
            ui.icon("lock", size="4rem", color="grey-4")
            ui.label("Nemáte přístup k žádné dlaždici modulu Formuláře ASM.") \
                .classes("text-lg text-gray-400")


def _vidi_schuzky(vsechna_prava) -> bool:
    """Dlaždice Schůzky — vlastní modul, jen bydlí v rozcestníku Formulářů ASM."""
    if not ('vse' in vsechna_prava or any(p in vsechna_prava for p in (
            'schuzky_zadatel', 'schuzky_vedouci', 'schuzky_spravce'))):
        return False
    return bool(intranet_data.nacti_nastaveni_intranetu().get('schuzky_zapnuty', True))


def _tile(emoji, nadpis, popis, barva, on_click):
    with ui.card().classes(
            "w-72 h-60 items-center justify-center shadow-xl hover:scale-105 "
            "transition-transform duration-300 cursor-pointer bg-white rounded-2xl border "
            + barva).on("click", on_click):
        ui.label(emoji).classes("text-6xl mb-3")
        ui.label(nadpis).classes("text-xl font-bold text-gray-800 text-center")
        if popis:
            ui.label(popis).classes("text-xs text-gray-500 text-center uppercase tracking-wide")


# ============================================================================
# Dlaždice „Data" — import číselníků
# ============================================================================
async def _precti_upload(e, up, reset=True):
    """Vrátí (raw, jmeno) z upload eventu.

    `reset=False` u multi-uploadů: Quasar `reset()` zahodí frontu a **abortuje
    rozdělané XHR** — při 20 souborech naráz by první dokončený soubor ustřelil
    zbytek (tichá ztráta + ClientDisconnect na serveru).
    """
    import inspect
    zdroj = None
    for attr in ("content", "file", "stream", "data", "file_obj"):
        val = getattr(e, attr, None)
        if val is not None and hasattr(val, "read"):
            zdroj = val
            break
    if zdroj is None:
        ui.notify("Nepodařilo se načíst obsah souboru.", type="negative")
        if reset:
            up.reset()
        return None, None
    try:
        if inspect.iscoroutinefunction(getattr(zdroj, "read", None)):
            raw = await zdroj.read()
        else:
            # Upload je spooled soubor na disku — až 120 MB. Synchronní read()
            # by na tu dobu zastavil event loop všem klientům.
            raw = await asyncio.to_thread(zdroj.read)
            if inspect.isawaitable(raw):
                raw = await raw
    except Exception as exc:
        ui.notify(f"Chyba čtení souboru: {exc}", type="negative")
        if reset:
            up.reset()
        return None, None
    name = (getattr(getattr(e, "file", None), "name", None)
            or getattr(e, "name", None) or "")
    if reset:
        up.reset()
    return raw, name


def _pocet_kontakty_oz() -> int:
    conn = intranet_data.get_db_connection()
    if not conn:
        return 0
    try:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM asm_kontakty_oz")
        n = cur.fetchone()[0]
        cur.close()
        return n
    except Exception:
        return 0
    finally:
        conn.close()


def _panel_data(user_name):
    drzeny = {"dealer_raw": None, "dealer_name": "", "kontakty_raw": None, "kontakty_name": "",
              "gist_raw": None, "gist_name": "",
              "kontakty_oz_raw": None, "kontakty_oz_name": ""}
    pd, pk, po = pocty_v_db()
    pk_oz = _pocet_kontakty_oz()
    posl = posledni_import()
    with ui.card().classes("w-full max-w-3xl p-6 rounded-2xl shadow-lg"):
        ui.label("Import dat").classes("text-2xl font-bold text-gray-800")
        ui.label("Nahrajte „Dealer“ (.xlsx), „Kontaktní údaje VO“ (.xlsx) a/nebo "
                 "„GIST_obrat zakaznik“ (.xlsx). Stačí jen některý. Nahraný číselník se "
                 "kompletně nahradí. Dealer + Kontakty obvykle týdně (pátek), "
                 "GIST obraty na začátku měsíce (za měsíc předchozí).") \
            .classes("text-sm text-gray-500 mb-3")

        with ui.column().classes("w-full bg-slate-50 border border-slate-200 "
                                 "rounded-xl p-3 mb-4 gap-1 text-sm"):
            ui.label("Aktuálně v databázi").classes(
                "text-xs font-semibold text-slate-500 uppercase tracking-wide")
            ui.label(f"📒 Dealer (číselník OZ/ASM): {pd} záznamů")
            ui.label(f"📇 Kontaktní údaje VO (nový – limity/splatnosti): {pk} záznamů")
            ui.label(f"📇 Kontaktní údaje – Změna OZ (původní): {pk_oz} záznamů")
            ui.label(f"📈 Obraty zákazníků (GIST): {po} záznamů")
            if posl:
                ui.label(f"Poslední import: {_dt_cz(posl.get('kdy'))} · {posl.get('uzivatel') or ''}") \
                    .classes("text-gray-400")

        stav = ui.label("").classes("text-sm font-medium")

        def _aktualizuj():
            casti = [s for s, v in (("Dealer", drzeny["dealer_raw"]),
                                    ("Kontakty", drzeny["kontakty_raw"]),
                                    ("Kontakty Změna OZ", drzeny["kontakty_oz_raw"]),
                                    ("GIST obraty", drzeny["gist_raw"])) if v]
            stav.text = ("K importu: " + " + ".join(casti)) if casti \
                else "Vyberte soubor k importu."
            btn.set_enabled(bool(drzeny["dealer_raw"] or drzeny["kontakty_raw"]
                                 or drzeny["kontakty_oz_raw"] or drzeny["gist_raw"]))

        async def _on_dealer(e):
            raw, name = await _precti_upload(e, up_dealer)
            if raw is not None:
                drzeny["dealer_raw"], drzeny["dealer_name"] = raw, name
                _aktualizuj()

        async def _on_kontakty(e):
            raw, name = await _precti_upload(e, up_kontakty)
            if raw is not None:
                drzeny["kontakty_raw"], drzeny["kontakty_name"] = raw, name
                _aktualizuj()

        async def _on_gist(e):
            raw, name = await _precti_upload(e, up_gist)
            if raw is not None:
                drzeny["gist_raw"], drzeny["gist_name"] = raw, name
                _aktualizuj()

        async def _on_kontakty_oz(e):
            raw, name = await _precti_upload(e, up_kontakty_oz)
            if raw is not None:
                drzeny["kontakty_oz_raw"], drzeny["kontakty_oz_name"] = raw, name
                _aktualizuj()

        with ui.row().classes("w-full gap-4 flex-wrap"):
            with ui.column().classes("flex-1 min-w-64"):
                ui.label("Dealer (.xlsx)").classes("text-xs font-semibold text-gray-500")
                up_dealer = ui.upload(on_upload=_on_dealer, auto_upload=True,
                                      max_file_size=20_000_000, label="Vybrat .xlsx") \
                    .props("accept=.xlsx").classes("w-full")
            with ui.column().classes("flex-1 min-w-64"):
                ui.label("Kontaktní údaje VO (.xlsx) — nový (limity/splatnosti)") \
                    .classes("text-xs font-semibold text-gray-500")
                up_kontakty = ui.upload(on_upload=_on_kontakty, auto_upload=True,
                                        max_file_size=80_000_000, label="Vybrat .xlsx") \
                    .props("accept=.xlsx").classes("w-full")
            with ui.column().classes("flex-1 min-w-64"):
                ui.label("Kontaktní údaje – Změna OZ (.xlsx) — původní") \
                    .classes("text-xs font-semibold text-gray-500")
                up_kontakty_oz = ui.upload(on_upload=_on_kontakty_oz, auto_upload=True,
                                           max_file_size=80_000_000, label="Vybrat .xlsx") \
                    .props("accept=.xlsx").classes("w-full")
            with ui.column().classes("flex-1 min-w-64"):
                ui.label("GIST_obrat zakaznik (.xlsx)").classes("text-xs font-semibold text-gray-500")
                up_gist = ui.upload(on_upload=_on_gist, auto_upload=True,
                                    max_file_size=80_000_000, label="Vybrat .xlsx") \
                    .props("accept=.xlsx").classes("w-full")
        spin = ui.spinner(size="lg").classes("text-blue-600")
        spin.set_visibility(False)

        async def _import_click():
            if not (drzeny["dealer_raw"] or drzeny["kontakty_raw"]
                    or drzeny["kontakty_oz_raw"] or drzeny["gist_raw"]):
                ui.notify("Vyberte alespoň jeden soubor.", type="warning")
                return
            btn.set_enabled(False)
            spin.set_visibility(True)
            stav.text = "Importuji… (velký soubor může chvíli trvat)"
            try:
                pd2, pk2, po2, err = await asyncio.to_thread(
                    _importuj_sync, drzeny["dealer_raw"], drzeny["dealer_name"],
                    drzeny["kontakty_raw"], drzeny["kontakty_name"], user_name,
                    drzeny["gist_raw"], drzeny["gist_name"],
                    drzeny["kontakty_oz_raw"], drzeny["kontakty_oz_name"])
            finally:
                spin.set_visibility(False)
            if err:
                ui.notify(f"Import selhal: {err}", type="negative", timeout=9000)
                stav.text = f"❌ {err}"
                btn.set_enabled(True)
            else:
                ui.notify(f"Import OK — {pd2} dealerů, {pk2} kontaktů, {po2} obratů.",
                          type="positive", position="top-right", timeout=6000)
                drzeny["dealer_raw"] = drzeny["kontakty_raw"] = None
                drzeny["kontakty_oz_raw"] = drzeny["gist_raw"] = None
                _refresh()

        with ui.row().classes("w-full justify-end mt-3"):
            btn = ui.button("Importovat data", icon="cloud_upload", on_click=_import_click) \
                .props("unelevated no-caps") \
                .classes("bg-blue-600 hover:bg-blue-700 text-white font-semibold rounded-lg shadow-md px-5")
            btn.set_enabled(False)


# ============================================================================
# Dlaždice „Změna zákazníků na OZ/ASM" — fronta / formulář / detail
# ============================================================================
async def _view_zmena(user_id, user_name, prava):
    sekce = app.storage.user.get("asm_sekce", "seznam")
    if sekce == "novy" and not _je_zadatel(prava):
        sekce = "seznam"

    def _set_sekce(s):
        app.storage.user["asm_sekce"] = s
        _refresh()

    if sekce == "novy":
        _formular_novy(user_id, user_name, prava, lambda: _set_sekce("seznam"))
        return

    # ── Fronta případů ───────────────────────────────────────────────
    with ui.row().classes("w-full items-center gap-3 mb-3"):
        if _je_zadatel(prava):
            ui.button("Nový formulář", icon="add", on_click=lambda: _set_sekce("novy")) \
                .props("unelevated no-caps") \
                .classes("bg-emerald-600 text-white font-semibold rounded-lg")
        ui.space()
        # Filtry
        stav_opts = {"": "Všechny stavy"} | {k: v[0] for k, v in _STAV_BADGE.items()}
        filtr_stav = ui.select(stav_opts, value=app.storage.user.get("asm_f_stav", ""),
                               label="Stav").props("outlined dense").classes("min-w-48")
        filtr_asm = ui.input("Filtr ASM", value=app.storage.user.get("asm_f_asm", "")) \
            .props("outlined dense clearable").classes("min-w-40")
        filtr_oz = ui.input("Filtr OZ", value=app.storage.user.get("asm_f_oz", "")) \
            .props("outlined dense clearable").classes("min-w-40")

    @ui.refreshable
    async def _tabulka():
        # čerstvé stavy při každé obnově; dotaz ve vlákně mimo event loop
        pripady = await asyncio.to_thread(nacti_pripady, user_id, prava)
        fs = (filtr_stav.value or "")
        fa = (filtr_asm.value or "").strip().lower()
        fo = (filtr_oz.value or "").strip().lower()
        app.storage.user["asm_f_stav"] = fs
        app.storage.user["asm_f_asm"] = fa
        app.storage.user["asm_f_oz"] = fo
        rows = []
        for p in pripady:
            if fs and p.get("stav") != fs:
                continue
            asm_txt = " ".join(str(p.get(k) or "") for k in
                               ("asm_jmeno", "zmena_asm_jmeno", "novy_asm_jmeno")).lower()
            if fa and fa not in asm_txt:
                continue
            oz_txt = " ".join(str(p.get(k) or "") for k in
                              ("cislo_oz", "jmeno_oz", "cislo_novy_oz", "jmeno_novy_oz")).lower()
            if fo and fo not in oz_txt:
                continue
            rows.append(p)

        if not rows:
            ui.label("Žádné případy neodpovídají filtru.").classes("text-gray-400 italic p-4")
            return

        with ui.column().classes("w-full gap-2"):
            for p in rows:
                with ui.card().classes("w-full p-3 rounded-xl shadow-sm hover:shadow-md "
                                       "transition-shadow"):
                    with ui.row().classes("w-full items-center gap-3 no-wrap"):
                        with ui.column().classes("gap-0 min-w-28"):
                            ui.label(p.get("cislo") or f"#{p['id']}").classes("font-bold text-gray-800")
                            ui.label(_dt_cz(p.get("datum_zadani"))).classes("text-xs text-gray-400")
                        with ui.column().classes("gap-0 flex-1 min-w-0"):
                            ui.label(p.get("duvod_zmeny") or "(bez důvodu)") \
                                .classes("font-medium text-gray-700 truncate")
                            sub = f"ASM: {p.get('zmena_asm_jmeno') or '—'} → {p.get('novy_asm_jmeno') or '—'}"
                            if p.get("cislo_oz") or p.get("cislo_novy_oz"):
                                sub += f"  •  OZ: {p.get('cislo_oz') or '—'} → {p.get('cislo_novy_oz') or '—'}"
                            sub += f"  •  {p.get('pocet_radku', 0)} řádků  •  {p.get('zadavatel_jmeno') or ''}"
                            ui.label(sub).classes("text-xs text-gray-500 truncate")
                        _badge(p.get("stav"))
                        ui.button(icon="visibility",
                                  on_click=lambda _=None, pid=p["id"]: _detail_dialog(pid, user_id, user_name, prava)) \
                            .props("flat round color=grey-7").tooltip("Detail / průběh")

    for f in (filtr_stav, filtr_asm, filtr_oz):
        f.on("update:model-value", lambda _=None: _tabulka.refresh())
    await _tabulka()

    # Auto-obnova fronty à 10 s: otisk dat se počítá ve vlákně a tabulka se
    # překreslí JEN při skutečné změně dat (dřív plný re-render každých 10 s).
    # Refreshuje JEN tabulku (ne celou stránku) → neruší filtry, otevřený formulář ani dialog.
    _fronta_otisk = {'v': None}

    async def _auto_obnova():
        novy = await asyncio.to_thread(lambda: hash(repr(nacti_pripady(user_id, prava))))
        if _fronta_otisk['v'] is None:
            _fronta_otisk['v'] = novy
            return
        if novy != _fronta_otisk['v']:
            _fronta_otisk['v'] = novy
            _tabulka.refresh()

    ui.timer(10.0, _auto_obnova)


# ============================================================================
# Unifikované formuláře (Navýšení limitů / Změna dodacích listů)
# ============================================================================
def _data_pripadu(p):
    """data_json případu jako dict."""
    try:
        return json.loads(p.get("data_json")) if p.get("data_json") else {}
    except Exception:
        return {}


def _render_form_pole(formular, hodnoty, editable, auto_editovatelne=False,
                      pobocka_filtr=None):
    """Vykreslí pole formuláře po sekcích.
      • editable=True  → vstupy (vč. rolovátek) + auto-dotažení dle IČO; vrací getter.
      • editable=False → jen popisky (read-only); vrací None.
      • auto_editovatelne=True → i „auto" pole (dotažená dle IČO) jsou ručně editovatelná
        (pro režim opravy — když auto-dotažené hodnoty výjimečně nesedí).
      • pobocka_filtr="012" → při dotažení dle IČO se nabídnou jen IČ 10 té pobočky
        (dle „Prod. kanál 6"); None = bez omezení (office/správce, uživatel bez pobočky).
    Sekce 5 (administrativní schválení) se needituje — odvozuje se z fronty (žadatel + datum)."""
    widgets, auto_refs = {}, {}
    # Obraty se nikde nezobrazují jako částka (formulář ani detail) — jen porovnání
    # s aktuálním kreditním limitem. Čísla drží obrat_hodnoty (maskovaná, do data_json)
    # a obrat_real (přesná, skrytá); obrat_widgets jsou readonly pole s porovnáním.
    obrat_widgets = {}
    obrat_hodnoty = {k: hodnoty.get(k, "") for k in ("obrat_1m", "obrat_3m")}
    for nadpis, pola in _FORMULARE[formular]["sekce"]:
        ui.label(nadpis).classes("text-sm font-semibold text-gray-500 uppercase mt-2")
        with ui.row().classes("w-full gap-3 flex-wrap items-start"):
            for spec in pola:
                key, label, typ = spec[0], spec[1], spec[2]
                val = hodnoty.get(key, "")
                if not editable:
                    # Detail případu: částka obratu se nezobrazuje, jen porovnání
                    # s aktuálním kreditním limitem (tabulka IČ 10 níže je beze změny).
                    if key in ("obrat_1m", "obrat_3m"):
                        val = _obrat_vs_limit(hodnoty.get(key + "_real") or val,
                                              hodnoty.get("akt_limit"))
                    with ui.column().classes("gap-0 min-w-56"):
                        ui.label(label).classes("text-xs text-gray-500")
                        ui.label(str(val or "—")).classes("text-gray-800 font-medium break-words")
                    continue
                if key in ("obrat_1m", "obrat_3m"):
                    # I při zadávání žádosti jen porovnání, žádná částka. Hodnota se
                    # dotáhne skrytě (obrat_hodnoty) a uloží do data_json.
                    w = ui.input(label, value="").props("outlined dense readonly") \
                        .classes("min-w-96") \
                        .tooltip("Částka se nezobrazuje — jen porovnání s aktuálním "
                                 "kreditním limitem")
                    obrat_widgets[key] = w
                    continue
                if typ == "select":
                    opts = {o: o for o in spec[3]}
                    w = ui.select(opts, label=label, value=(val or None), with_input=True) \
                        .props("outlined dense").classes("min-w-56")
                elif typ == "textarea":
                    w = ui.textarea(label, value=val).props("outlined autogrow").classes("w-full")
                elif typ == "auto":
                    w = ui.input(label, value=val) \
                        .props("outlined dense" if auto_editovatelne else "outlined dense readonly") \
                        .classes("min-w-56").tooltip(
                            "Dotaženo dle IČO — lze ručně upravit" if auto_editovatelne
                            else "Dotaženo automaticky podle IČO")
                    auto_refs[key] = w
                else:   # ico / text / number
                    w = ui.input(label, value=val).props("outlined dense").classes("min-w-56")
                widgets[key] = w

    # ── Bublina IČ 10 (výběr konkrétních IČ 10 náležejících k zadanému IČ 8) ──
    # Zadává se IČ 8; pod ním může existovat víc IČ 10. Uživatel zaškrtne, kterých
    # se žádost týká (výchozí: všechna). Vybraná se ukládají do data_json.
    ic10 = {"vse": list(hodnoty.get("ico10_varianty") or []),
            "vybrane": set(hodnoty.get("vybrane_ico10") or [])}

    if not editable:
        # Detail (read-only): vypiš vybraná IČ 10.
        if ic10["vybrane"]:
            ui.label("ℹ️ Při výběru konkrétního IČ se řiďte splatností uvedenou u daného IČ níže "
                     "(spl zboží/cigarety). Obraty v polích formuláře jsou za celé IČ 8 "
                     "(všechny provozovny dohromady); sloupce Obrat v tabulce níže "
                     "jsou za jednotlivá IČ 10.") \
                .classes("text-sm text-amber-700 bg-amber-50 border border-amber-200 "
                         "rounded-lg p-3 w-full mt-2")
            ui.label("Vybraná IČ 10").classes("text-sm font-semibold text-gray-500 uppercase mt-2")

            def _bunka_ro(k_val, fallback):
                """K.* hodnota; když chybí, fakturační šedě kurzívou (ať je jasné, co je co)."""
                k_val = (str(k_val or "")).strip()
                if k_val:
                    ui.label(k_val).classes("text-sm")
                elif (str(fallback or "")).strip():
                    ui.label(str(fallback)).classes("text-sm text-gray-400 italic") \
                        .tooltip("K.* údaj chybí — zobrazen fakturační")
                else:
                    ui.label("—").classes("text-sm text-gray-400")

            # Stejné sloupce jako ve formuláři (viz _bublina), bez checkboxu.
            with ui.column().classes("w-full border border-gray-200 rounded-lg p-2 "
                                     "max-h-64 overflow-auto shrink-0"):
                # minmax(max-content,1fr): sloupce se rozloží rovnoměrně od kraje ke kraji,
                # ale žádný nespadne pod šířku obsahu.
                ukaz_obrat = (formular == "limity")
                hlavicky = ["IČ 10", "K.Jméno", "K.Adresa (ulice, město, PSČ)",
                            "Prod. kanál 6", "Akt. kreditní limit", "Typ plátce", "Spl 1/2"]
                if ukaz_obrat:
                    hlavicky += ["Obrat 1 měs.", "Obrat 3 měs."]
                with ui.grid(columns=f"repeat({len(hlavicky)}, "
                                     "minmax(max-content, 1fr))") \
                        .classes("w-full gap-x-6 gap-y-1 items-center"):
                    for h in hlavicky:
                        ui.label(h).classes("text-xs font-semibold text-gray-500 uppercase")
                    zdroj = ic10["vse"] or [{"ico": i} for i in sorted(ic10["vybrane"])]
                    for v in zdroj:
                        if not isinstance(v, dict):
                            v = {"ico": v}
                        ic = v.get("ico")
                        if ic not in ic10["vybrane"]:
                            continue
                        ui.label(str(ic or "")).classes("text-sm font-mono")
                        _bunka_ro(v.get("k_jmeno"), v.get("jmeno"))
                        _bunka_ro(v.get("k_adresa"), v.get("adresa"))
                        ui.label(v.get("prodejni_doba") or "—").classes("text-sm")
                        ui.label(v.get("limit_salda") or "—").classes("text-sm whitespace-nowrap")
                        ui.label(v.get("typ_platice") or "—").classes("text-sm")
                        ui.label(f"{v.get('splatnost1') or '—'}/{v.get('splatnost2') or '—'}") \
                            .classes("text-sm whitespace-nowrap")
                        if ukaz_obrat:
                            ui.label(str(v.get("obrat_1m") or "—")) \
                                .classes("text-sm whitespace-nowrap")
                            ui.label(str(v.get("obrat_3m") or "—")) \
                                .classes("text-sm whitespace-nowrap")
        return None

    ico_w = widgets.get("ico")
    # Skutečný (nemaskovaný) obrat — uloží se skrytě do data_json; nikde se nezobrazuje,
    # slouží jen k porovnání s kreditním limitem (a pro pozdější potřeby office).
    obrat_real = {"obrat_1m": hodnoty.get("obrat_1m_real", ""),
                  "obrat_3m": hodnoty.get("obrat_3m_real", "")}

    def _prepocti_obrat(_=None):
        """Přepíše pole obratů porovnáním s aktuálním kreditním limitem."""
        w_lim = widgets.get("akt_limit")
        lim = w_lim.value if w_lim is not None else hodnoty.get("akt_limit")
        for kk, w in obrat_widgets.items():
            w.value = _obrat_vs_limit(obrat_real.get(kk) or obrat_hodnoty.get(kk), lim)
    _prepocti_obrat()
    if widgets.get("akt_limit") is not None:
        widgets["akt_limit"].on("blur", _prepocti_obrat)   # ruční oprava limitu v režimu opravy

    if formular == "limity":
        ui.label("ℹ️ Částky obratu se nezobrazují — u obou obratů uvidíte jen porovnání "
                 "s aktuálním kreditním limitem. Obraty v řádcích výše jsou za celé IČ 8 "
                 "(všechny provozovny dohromady); v tabulce IČ 10 jsou obraty za "
                 "jednotlivá IČ 10.") \
            .classes("text-xs text-amber-700 bg-amber-50 border border-amber-200 "
                     "rounded-lg p-2 w-full mt-1")
    ui.label("IČ 10 náležející k zadanému IČ 8").classes(
        "text-sm font-semibold text-gray-500 uppercase mt-2")
    ui.label("Zaškrtnutá IČ 10 se uloží k žádosti (výchozí: všechna). "
             "Odškrtněte, pokud se změna netýká všech. Při výběru právě jednoho IČ 10 "
             "se Provozovna a Adresa doplní podle něj.").classes("text-xs text-gray-500")

    # Reprezentativní hodnoty za IČ 8 — návrat, když není vybráno právě jedno IČ 10.
    repr_pole = {"provozovna": hodnoty.get("provozovna", "") or "",
                 "adresa":     hodnoty.get("adresa", "") or ""}

    def _aktualizuj_dle_vyberu():
        """Provozovna + Adresa dle výběru v bublině: právě jedno vybrané IČ 10 →
        údaje té provozovny; jinak reprezentativní řádek IČ 8 (dle zadání).
        Splatnosti se nepropisují — platí infotext u bubliny; obrat je vždy za celé IČ 8."""
        w_prov = auto_refs.get("provozovna")
        w_adr = auto_refs.get("adresa")
        if w_prov is None and w_adr is None:
            return
        v1 = None
        if len(ic10["vybrane"]) == 1:
            ic = next(iter(ic10["vybrane"]))
            v1 = next((v for v in ic10["vse"] if v.get("ico") == ic), None)
        # Přednostně K.* (dodací) údaje provozovny, fallback fakturační (stejně jako bublina).
        if w_prov is not None:
            w_prov.value = ((v1.get("k_jmeno") or "").strip() or (v1.get("jmeno") or "")) \
                if v1 else repr_pole["provozovna"]
        if w_adr is not None:
            w_adr.value = ((v1.get("k_adresa") or "").strip() or (v1.get("adresa") or "")) \
                if v1 else repr_pole["adresa"]

    @ui.refreshable
    def _bublina():
        if not ic10["vse"]:
            ui.label("Zadejte IČO (IČ 8) a opusťte pole (Tab/Enter) — zobrazí se náležející IČ 10.") \
                .classes("text-sm text-gray-400 italic")
            return

        def _bunka(k_val, fallback):
            """K.* hodnota; když chybí, fakturační hodnota šedě kurzívou (ať je jasné, co je co)."""
            k_val = (k_val or "").strip()
            if k_val:
                ui.label(k_val).classes("text-sm")
            elif (fallback or "").strip():
                ui.label(fallback).classes("text-sm text-gray-400 italic") \
                    .tooltip("K.* údaj chybí — zobrazen fakturační")
            else:
                ui.label("—").classes("text-sm text-gray-400")

        with ui.column().classes("w-full border border-gray-200 rounded-lg p-2 "
                                 "max-h-64 overflow-auto shrink-0"):
            # minmax(max-content,1fr): sloupce se rozloží rovnoměrně od kraje ke kraji,
            # ale žádný nespadne pod šířku obsahu.
            ukaz_obrat = (formular == "limity")
            hlavicky = ["", "IČ 10", "K.Jméno", "K.Adresa (ulice, město, PSČ)",
                        "Prod. kanál 6", "Akt. kreditní limit", "Typ plátce", "Spl 1/2"]
            if ukaz_obrat:
                hlavicky += ["Obrat 1 měs.", "Obrat 3 měs."]
            with ui.grid(columns=f"auto repeat({len(hlavicky) - 1}, "
                                 "minmax(max-content, 1fr))") \
                    .classes("w-full gap-x-6 gap-y-1 items-center"):
                for h in hlavicky:
                    ui.label(h).classes("text-xs font-semibold text-gray-500 uppercase")
                for v in ic10["vse"]:
                    ic = v.get("ico")

                    def _tog(e, ic=ic):
                        if e.value:
                            ic10["vybrane"].add(ic)
                        else:
                            ic10["vybrane"].discard(ic)
                        _aktualizuj_dle_vyberu()
                    ui.checkbox(value=(ic in ic10["vybrane"]), on_change=_tog).props("dense")
                    ui.label(str(ic or "")).classes("text-sm font-mono")
                    _bunka(v.get("k_jmeno"), v.get("jmeno"))
                    _bunka(v.get("k_adresa"), v.get("adresa"))
                    ui.label(v.get("prodejni_doba") or "—").classes("text-sm")
                    ui.label(v.get("limit_salda") or "—").classes("text-sm whitespace-nowrap")
                    ui.label(v.get("typ_platice") or "—").classes("text-sm")
                    ui.label(f"{v.get('splatnost1') or '—'}/{v.get('splatnost2') or '—'}") \
                        .classes("text-sm whitespace-nowrap")
                    if ukaz_obrat:
                        ui.label(v.get("obrat_1m") or "—") \
                            .classes("text-sm whitespace-nowrap")
                        ui.label(v.get("obrat_3m") or "—") \
                            .classes("text-sm whitespace-nowrap")
    # Wrapper drží výšku: refreshable container je flex item karty dialogu a scroll
    # box uvnitř nepřispívá do min-content, takže by se jinak dal smrsknout.
    with ui.element("div").classes("w-full shrink-0"):
        _bublina()

    def _dotahni(_=None):
        # IČ 8 = zleva 8 znaků zadaného IČO (=ZLEVA(8)); shodné s ico8 v kontaktech.
        ic8 = (ico_w.value or "").strip()[:8]
        fakt, varianty = _kontakty_dle_ico8(ic8) if ic8 else (None, [])
        # Viditelnost dle pobočky uživatele: „Prod. kanál 6" začíná kódem pobočky
        # („034Q JIP Plzeň" → 034). Řádky bez kódu zůstávají viditelné všem.
        varianty_vse = len(varianty)
        if pobocka_filtr:
            varianty = [v for v in varianty
                        if _kod_pobocky(v.get("prodejni_doba")) in ("", pobocka_filtr)]
            fakt = varianty[0] if varianty else None   # auto-pole z povoleného řádku
        o = _obrat_dle_ico([ic8]).get(ic8) if ic8 else None
        auto = _ico_auto_hodnoty(fakt, o, maskovat_obrat=True)   # žadatel vidí maskovaný obrat
        for kk, w in auto_refs.items():
            w.value = auto.get(kk, "")
        # Obraty jen skrytě do data_json; v poli se ukáže porovnání s limitem.
        obrat_real["obrat_1m"] = _cislo_cz(o.get("obrat_1m")) if o else ""
        obrat_real["obrat_3m"] = _cislo_cz(o.get("obrat_3m")) if o else ""
        obrat_hodnoty["obrat_1m"] = auto.get("obrat_1m", "")
        obrat_hodnoty["obrat_3m"] = auto.get("obrat_3m", "")
        _prepocti_obrat()
        repr_pole["provozovna"] = auto.get("provozovna", "")
        repr_pole["adresa"] = auto.get("adresa", "")
        # Obraty po jednotlivých IČ 10 (GIST bloky „IČO"). Žadateli se maskují
        # stejně jako obrat IČ 8 (> 2 mil. Kč); přesné hodnoty jdou skrytě do
        # data_json (*_real) pro office obchod v detailu.
        o10 = _obrat_ic10_dle_ico([v.get("ico") for v in varianty]) if varianty else {}

        def _o10(ic_, klic, maskovat=True):
            return _obrat_zobraz((o10.get(ic_) or {}).get(klic), maskovat)
        ic10["vse"] = [{"ico": v.get("ico"), "jmeno": v.get("jmeno"),
                        "adresa": " ".join(x for x in (v.get("ulice"), v.get("mesto"),
                                                       v.get("psc"))
                                           if x and str(x).strip()),
                        # Kontaktní údaje VO2 (K. dodací) — jméno/adresa provozovny + prodejní doba.
                        "k_jmeno": v.get("k_jmeno"),
                        "k_adresa": " ".join(x for x in (v.get("k_ulice"), v.get("k_mesto"),
                                                         v.get("k_psc"))
                                             if x and str(x).strip()),
                        "prodejni_doba": v.get("prodejni_doba"),
                        "limit_salda": _cislo_cz(v.get("limit_salda")),
                        "typ_platice": v.get("typ_platice"),
                        "splatnost1": v.get("splatnost1"), "splatnost2": v.get("splatnost2"),
                        "obrat_1m": _o10(v.get("ico"), "obrat_1m"),
                        "obrat_3m": _o10(v.get("ico"), "obrat_3m"),
                        "obrat_1m_real": _o10(v.get("ico"), "obrat_1m", False),
                        "obrat_3m_real": _o10(v.get("ico"), "obrat_3m", False)}
                       for v in varianty]
        ic10["vybrane"] = {v["ico"] for v in ic10["vse"]}   # výchozí: všechna zaškrtnutá
        _bublina.refresh()
        _aktualizuj_dle_vyberu()
        if ic8 and not fakt and varianty_vse:
            ui.notify(f"IČO nepatří vaší pobočce {pobocka_filtr} "
                      f"(0 z {varianty_vse} IČ 10).", type="warning", position="top-right")
        elif ic8 and not fakt:
            ui.notify("IČ 8 nenalezeno v kontaktech.", type="warning", position="top-right")
        elif fakt:
            n = len(ic10["vse"])
            msg = (f"Údaje dotaženy ({n}× IČ 10)." if o
                   else f"Údaje dotaženy ({n}× IČ 10; obraty GIST chybí).")
            if pobocka_filtr and n < varianty_vse:
                msg += f" Zobrazena jen pobočka {pobocka_filtr} ({varianty_vse - n} skryto)."
            ui.notify(msg, type="positive", position="top-right")
    if ico_w:
        ico_w.on("blur", _dotahni)
        ico_w.on("keydown.enter", _dotahni)

    def cti():
        d = {key: (widgets[key].value or "") for key in widgets}
        d.update(obrat_hodnoty)          # obraty jako čísla (v poli je jen porovnání)
        d["vybrane_ico10"] = sorted(ic10["vybrane"])
        d["ico10_varianty"] = ic10["vse"]
        d["obrat_1m_real"] = obrat_real["obrat_1m"]   # skrytý přesný obrat (pro office)
        d["obrat_3m_real"] = obrat_real["obrat_3m"]
        return d
    return cti


def _formular_novy_obecny(formular, user_id, user_name, prava, zpet_fn):
    cfg = _FORMULARE[formular]
    pd, pk, po = pocty_v_db()
    # Číselník OZ pro volitelnou notifikaci dotčených OZ (formulář nemá OZ pole).
    # Jen lidé; label = číslo OZ (sloupec A) + jméno. Číslo v labelu musí být,
    # NiceGUI select filtruje jen podle labelu. Adresa se skládá z čísla (viz náhled).
    _oz_lide = _oz_ciselnik_lide()
    oz_jmena = {k: (j or "") for k, j, _e in _oz_lide}
    oz_opts = {k: " — ".join(x for x in (str(k), j) if x) for k, j, _e in _oz_lide}
    OZ_MIN_ZNAKU = 2   # kratší dotaz = žádné návrhy (celý číselník se nikomu nerozbalí)

    # Notifikace OZ — skrytá v dialogu, otevírá se ikonkou „+“ v hlavičce vedle křížku.
    with ui.dialog() as notif_dialog, ui.card().classes("min-w-[440px] max-w-[560px]"):
        ui.label("Notifikace OZ").classes("text-sm font-semibold text-gray-500 uppercase")
        ui.label("Volitelně vyberte OZ, které se e-mailem upozorní na podání případu. "
                 "Prázdné = nikomu se neposílá.").classes("text-xs text-gray-500")
        # Nabídka startuje prázdná a plní se až podle napsaného textu (min. OZ_MIN_ZNAKU
        # znaků). Vybrané položky musí v options zůstat, jinak by zmizely popisky chipů.
        notif_oz_in = ui.select({}, label="Notifikovat OZ", multiple=True,
                                with_input=True) \
            .props("outlined dense use-chips hide-dropdown-icon "
                   'input-debounce=200 hint="Hledejte podle čísla OZ nebo jména"') \
            .classes("w-full")

        def _oz_naseptavac(e):
            dotaz = str(e.args or "").strip().lower()
            vybrane = {k: oz_opts[k] for k in (notif_oz_in.value or []) if k in oz_opts}
            if len(dotaz) < OZ_MIN_ZNAKU:
                notif_oz_in.set_options(vybrane)
                return
            nalezene = {k: v for k, v in oz_opts.items() if dotaz in v.lower()}
            notif_oz_in.set_options({**vybrane, **nalezene})

        notif_oz_in.on("input-value", _oz_naseptavac)

        # Náhled: kdo je přidaný a na jakou adresu upozornění opravdu odejde.
        nahled = ui.column().classes("w-full gap-1 mt-3")

        def _obnov_nahled(_=None):
            vybrane = list(notif_oz_in.value or [])
            nahled.clear()
            with nahled:
                if not vybrane:
                    ui.label("Nikdo nevybrán — e-mail se neodešle.") \
                        .classes("text-xs text-gray-400")
                    return
                ui.label(f"Upozornění odejde na {len(vybrane)} adres:") \
                    .classes("text-xs font-semibold text-gray-500 uppercase")
                for k in vybrane:
                    adresa = _email_oz_dle_cisla(k)
                    with ui.row().classes("w-full items-center gap-2 bg-gray-50 "
                                          "rounded-lg px-2 py-1"):
                        ui.label(oz_opts.get(k) or str(k)).classes("text-sm text-gray-700")
                        ui.space()
                        ui.label(adresa or "bez adresy — nepošle se") \
                            .classes("text-xs font-mono "
                                     + ("text-blue-700" if adresa else "text-red-500"))

        notif_oz_in.on_value_change(_obnov_nahled)
        _obnov_nahled()
        with ui.row().classes("w-full justify-end mt-2"):
            ui.button("Hotovo", on_click=notif_dialog.close).props("flat color=primary")

    with ui.card().classes("w-full max-w-4xl p-5 rounded-2xl shadow-lg mb-6"):
        with ui.row().classes("w-full items-center gap-3"):
            ui.label(f"Nový formulář — {cfg['nazev']}").classes("text-xl font-bold text-gray-800")
            ui.space()
            with ui.button(icon="add", on_click=notif_dialog.open) \
                    .props("flat round color=grey-7").classes("mr-3") \
                    .tooltip("Notifikovat OZ e-mailem"):
                notif_badge = ui.badge("0").props("floating color=red")
                notif_badge.bind_visibility_from(notif_oz_in, "value", backward=bool)
                notif_badge.bind_text_from(notif_oz_in, "value",
                                           backward=lambda v: str(len(v or [])))
            ui.button(icon="close", on_click=zpet_fn).props("flat round color=grey-7") \
                .tooltip("Zpět na frontu")
        if pk == 0:
            ui.label("⚠️ Kontaktní údaje VO nejsou nahrané — auto-dotažení dle IČO nemusí "
                     "fungovat (kontaktujte Vkladatele).") \
                .classes("text-sm text-amber-700 bg-amber-50 border border-amber-200 rounded-lg p-2 w-full")
        ui.label("Zadejte IČO zákazníka a opusťte pole (Tab/Enter) — údaje se dotáhnou "
                 "z Kontaktních údajů VO. „Dodavatel firma“ se nevyplňuje. Sekce „Administrativní "
                 "schválení“ se doplní automaticky (žadatel + datum) a uvidíte ji ve frontě.") \
            .classes("text-xs text-gray-500")

        # Žadatel vidí jen IČ 10 své pobočky (dle „Prod. kanál 6"). Office/správce
        # a uživatel bez nastavené pobočky vidí všechna.
        _pob = None if _je_office(prava) else _pobocka_uzivatele(user_id)
        cti = _render_form_pole(formular, {}, editable=True, pobocka_filtr=_pob)

        async def _odeslat():
            data = cti()
            if not (data.get("ico") or "").strip():
                ui.notify("Vyplňte IČO zákazníka.", type="warning")
                return
            data["notifikovat_oz"] = list(notif_oz_in.value or [])
            pid, cislo = zaloz_pripad_formular(formular, data, user_id, user_name)
            if not pid:
                ui.notify(f"Uložení selhalo: {cislo}", type="negative")
                return
            zaznam_historie(pid, "Odesláno", user_name, f"Vytvořen formulář „{cfg['nazev']}“")
            p = nacti_pripad(pid)
            _odesli_emaily(_emaily_office(),
                           f"Formuláře ASM — nový případ {cislo} ({cfg['nazev']})",
                           f"Máte nový případ k vyřešení.\n\nČíslo: {cislo}\n"
                           f"Typ: {cfg['nazev']}\nŽadatel: {user_name}\n"
                           f"IČO zákazníka: {data.get('ico')}\n"
                           f"Zákazník: {data.get('fakt_nazev') or ''}", pripad_id=pid)
            # Notifikace vybraných OZ při podání žádosti — adresa oz<číslo>@doména.
            _notifikuj_oz(data.get("notifikovat_oz") or [],
                          cislo, cfg["nazev"], user_name, data.get("ico"), pid,
                          dle_cisla=True)
            intranet_logger.log_activity(user_name, "Podání žádosti",
                                         f"Formuláře ASM: nový případ {cislo} ({cfg['nazev']})")
            ui.notify(f"Případ {cislo} odeslán.", type="positive", position="top")
            app.storage.user["asm_sekce"] = "seznam"
            _refresh()

        with ui.row().classes("w-full justify-end mt-4 gap-2"):
            ui.button("Zrušit", on_click=zpet_fn).props("flat no-caps").classes("text-gray-600")
            ui.button("Odeslat", icon="send", on_click=_odeslat) \
                .props("unelevated no-caps") \
                .classes("bg-emerald-600 hover:bg-emerald-700 text-white font-semibold rounded-lg px-5")


async def _view_formular(formular, user_id, user_name, prava):
    cfg = _FORMULARE[formular]
    sekce = app.storage.user.get("asm_sekce", "seznam")
    if sekce == "novy" and not _je_zadatel(prava):
        sekce = "seznam"

    def _set_sekce(s):
        app.storage.user["asm_sekce"] = s
        _refresh()

    if sekce == "novy":
        _formular_novy_obecny(formular, user_id, user_name, prava, lambda: _set_sekce("seznam"))
        return

    with ui.row().classes("w-full items-center gap-3 mb-3"):
        if _je_zadatel(prava):
            ui.button("Nový formulář", icon="add", on_click=lambda: _set_sekce("novy")) \
                .props("unelevated no-caps") \
                .classes("bg-emerald-600 text-white font-semibold rounded-lg")
        ui.space()
        stav_opts = {"": "Všechny stavy"} | {k: v[0] for k, v in _STAV_BADGE.items()}
        filtr_stav = ui.select(stav_opts, value="", label="Stav") \
            .props("outlined dense").classes("min-w-48")
        filtr_text = ui.input("Hledat (IČO / zákazník)") \
            .props("outlined dense clearable").classes("min-w-56")

    @ui.refreshable
    async def _tabulka():
        # čerstvé stavy při každé obnově; dotaz ve vlákně mimo event loop
        pripady = await asyncio.to_thread(nacti_pripady, user_id, prava, formular)
        fs = (filtr_stav.value or "")
        ft = (filtr_text.value or "").strip().lower()
        rows = []
        for p in pripady:
            if fs and p.get("stav") != fs:
                continue
            d = _data_pripadu(p)
            txt = f"{d.get('ico') or ''} {d.get('fakt_nazev') or ''} {d.get('provozovna') or ''}".lower()
            if ft and ft not in txt:
                continue
            rows.append((p, d))
        if not rows:
            ui.label("Žádné případy neodpovídají filtru.").classes("text-gray-400 italic p-4")
            return
        with ui.column().classes("w-full gap-2"):
            for p, d in rows:
                with ui.card().classes("w-full p-3 rounded-xl shadow-sm hover:shadow-md transition-shadow"):
                    with ui.row().classes("w-full items-center gap-3 no-wrap"):
                        with ui.column().classes("gap-0 min-w-28"):
                            ui.label(p.get("cislo") or f"#{p['id']}").classes("font-bold text-gray-800")
                            ui.label(_dt_cz(p.get("datum_zadani"))).classes("text-xs text-gray-400")
                        with ui.column().classes("gap-0 flex-1 min-w-0"):
                            ui.label(d.get("fakt_nazev") or "(zákazník nevyplněn)") \
                                .classes("font-medium text-gray-700 truncate")
                            sub = f"IČO: {d.get('ico') or '—'}  •  {d.get('pobocka') or ''}  •  {p.get('zadavatel_jmeno') or ''}"
                            ui.label(sub).classes("text-xs text-gray-500 truncate")
                        _badge(p.get("stav"))
                        ui.button(icon="visibility",
                                  on_click=lambda _=None, pid=p["id"]: _detail_dialog(pid, user_id, user_name, prava)) \
                            .props("flat round color=grey-7").tooltip("Detail / průběh")

    for f in (filtr_stav, filtr_text):
        f.on("update:model-value", lambda _=None: _tabulka.refresh())
    await _tabulka()

    # Auto-obnova fronty à 10 s (viz _view_zmena) — otisk ve vlákně, překreslení
    # jen při skutečné změně dat; neruší filtry/formulář/dialog.
    _fronta_otisk = {'v': None}

    async def _auto_obnova():
        novy = await asyncio.to_thread(lambda: hash(repr(nacti_pripady(user_id, prava, formular))))
        if _fronta_otisk['v'] is None:
            _fronta_otisk['v'] = novy
            return
        if novy != _fronta_otisk['v']:
            _fronta_otisk['v'] = novy
            _tabulka.refresh()

    ui.timer(10.0, _auto_obnova)


# ============================================================================
# Formulář nového případu
# ============================================================================
def _grid_col_defs():
    # Všechny sloupce kromě pořadí jsou editovatelné (dvojklik na buňku → úprava),
    # vč. dotažených (Jméno, Město, Dealer, Pův. provize) — lze ručně opravit.
    return [
        # „#" = pořadí dopočítané z indexu řádku (nedrží se v datech → bez přečíslování).
        {"headerName": "#", "colId": "poradi", "width": 60, "editable": False,
         ":valueGetter": "(p) => p.node.rowIndex + 1"},
        {"headerName": "IČO", "field": "ico", "width": 120, "editable": True},
        {"headerName": "Jméno", "field": "jmeno", "width": 200, "editable": True},
        {"headerName": "Město", "field": "k_mesto", "width": 140, "editable": True},
        {"headerName": "Dealer", "field": "dealer", "width": 90, "editable": True},
        {"headerName": "Pův. provize", "field": "puvodni_provize", "width": 110, "editable": True},
        {"headerName": "Nová provize", "field": "nova_provize", "width": 120, "editable": True},
        {"headerName": "OZ2", "field": "oz2", "width": 110, "editable": True},
        {"headerName": "Datum od", "field": "datum_od", "width": 120, "editable": True},
        {"headerName": "Datum do", "field": "datum_do", "width": 120, "editable": True},
    ]


def _formular_novy(user_id, user_name, prava, zpet_fn):
    import time as _t
    _c = _t.perf_counter()
    oz_cis = _oz_ciselnik()
    print(f"[asm PERF] _oz_ciselnik: {_t.perf_counter()-_c:.2f}s ({len(oz_cis)} řádků)")
    _c = _t.perf_counter()
    asm_cis = _asm_ciselnik()
    print(f"[asm PERF] _asm_ciselnik: {_t.perf_counter()-_c:.2f}s ({len(asm_cis)} řádků)")
    _c = _t.perf_counter()
    pd, pk, po = pocty_v_db()
    print(f"[asm PERF] pocty_v_db: {_t.perf_counter()-_c:.2f}s (dealer={pd}, kontakty={pk}, obrat={po})")

    # Mapy pro auto-dotažení jména OZ
    oz_jmena = {k: j for k, j in oz_cis}
    oz_opts = {k: (f"{k} — {j}" if j else k) for k, j in oz_cis}

    # Server-side model řádků (synchronizovaný s gridem přes cellValueChanged).
    # Každý řádek má stabilní `_rid` (getRowId) → AG Grid umí cílené transakce
    # (add/remove/update) bez plného překreslení (žádné probliknutí).
    _rid_seq = {"n": 0}

    def _novy_radek(**kw):
        _rid_seq["n"] += 1
        r = {"_rid": _rid_seq["n"], "ico": "", "jmeno": "", "k_jmeno": "", "k_jmeno2": "",
             "k_ulice": "", "k_mesto": "", "k_psc": "", "dealer": "", "id_kontakt": "",
             "prodejni_doba": "", "puvodni_provize": "", "nova_provize": "",
             "oz2": "", "datum_od": "", "datum_do": ""}
        r.update(kw)
        return r

    rows = [_novy_radek()]

    with ui.card().classes("w-full max-w-5xl p-5 rounded-2xl shadow-lg mb-6"):
        with ui.row().classes("w-full items-center gap-3"):
            ui.label("Nový formulář — Změna zákazníků na OZ/ASM") \
                .classes("text-xl font-bold text-gray-800")
            ui.space()
            ui.button(icon="close", on_click=zpet_fn).props("flat round color=grey-7") \
                .tooltip("Zpět na frontu")

        if pd == 0 or pk == 0:
            ui.label("⚠️ Číselníky nejsou nahrané (Dealer / Kontaktní údaje VO). "
                     "Auto-dotažení a výběry nemusí fungovat — kontaktujte Vkladatele.") \
                .classes("text-sm text-amber-700 bg-amber-50 border border-amber-200 rounded-lg p-2 w-full")

        # ── Hlavička ────────────────────────────────────────────────
        ui.label("Hlavička").classes("text-sm font-semibold text-gray-500 uppercase mt-2")
        duvod_in = ui.textarea("Důvod změny *").props("outlined autogrow") \
            .classes("w-full")
        with ui.row().classes("w-full gap-3 flex-wrap"):
            ui.input("Jméno ASM", value=user_name).props("outlined dense readonly") \
                .classes("flex-1 min-w-56").tooltip("Předvyplněno z přihlášení")
            zmena_asm_in = ui.input("Změna ASM jméno", value=user_name) \
                .props("outlined dense").classes("flex-1 min-w-56")
            novy_asm_in = ui.select(asm_cis, label="Nový ASM jméno", with_input=True) \
                .props("outlined dense").classes("flex-1 min-w-56")

        ui.label("Změna OZ").classes("text-sm font-semibold text-gray-500 uppercase mt-2")
        with ui.row().classes("w-full gap-3 flex-wrap"):
            cislo_oz_in = ui.select(oz_opts, label="Číslo OZ", with_input=True) \
                .props("outlined dense").classes("flex-1 min-w-48")
            jmeno_oz_in = ui.input("Jméno OZ").props("outlined dense").classes("flex-1 min-w-48")
            cislo_novy_oz_in = ui.select(oz_opts, label="Číslo nový OZ", with_input=True) \
                .props("outlined dense").classes("flex-1 min-w-48")
            jmeno_novy_oz_in = ui.input("Jméno nový OZ").props("outlined dense").classes("flex-1 min-w-48")

        def _dotahni_oz(src, dst):
            dst.value = oz_jmena.get(src.value, "")
        cislo_oz_in.on("update:model-value", lambda _=None: _dotahni_oz(cislo_oz_in, jmeno_oz_in))
        cislo_novy_oz_in.on("update:model-value", lambda _=None: _dotahni_oz(cislo_novy_oz_in, jmeno_novy_oz_in))

        ui.label("Změna provize").classes("text-sm font-semibold text-gray-500 uppercase mt-2")
        with ui.row().classes("w-full gap-3 flex-wrap items-start"):
            with ui.column().classes("flex-1 min-w-48"):
                datum_od_in = _date_input("Datum změny od (DD.MM.RRRR)")
            region_in = ui.select({"ANO": "ANO", "NE": "NE"}, label="Zákazník v regionu OZ", value="NE") \
                .props("outlined dense").classes("flex-1 min-w-48")
        oduvodneni_in = ui.textarea("Případné odůvodnění").props("outlined autogrow").classes("w-full")

        # ── Detailní tabulka ────────────────────────────────────────
        ui.separator().classes("my-3")
        ui.label("Detailní tabulka zákazníků").classes("text-sm font-semibold text-gray-500 uppercase")
        ui.label("Vložte IČO do textového pole (jedno na řádek, Ctrl+V z Excelu) a klikněte "
                 "„Načíst IČO“. Kontaktní údaje a původní provize se dotáhnou automaticky; "
                 "ručně doplňte Nová provize, OZ2, Datum od/do přímo v tabulce.") \
            .classes("text-xs text-gray-500")

        ico_paste = ui.textarea("Seznam IČO (jedno na řádek)") \
            .props("outlined autogrow").classes("w-full")

        grid = ui.aggrid({
            "columnDefs": _grid_col_defs(),
            "rowData": rows,
            "rowSelection": "multiple",
            "stopEditingWhenCellsLoseFocus": True,
            "singleClickEdit": False,   # editace buňky dvojklikem
            "defaultColDef": {"resizable": True, "sortable": False,
                              "wrapHeaderText": True, "autoHeaderHeight": True},
            "rowHeight": 32,
            # Bez virtualizace → všechny buňky jsou v DOM, takže Ctrl+C zkopíruje
            # i řádky mimo viditelnou oblast (výběr celého sloupce apod.).
            "suppressRowVirtualisation": True,
            "suppressColumnVirtualisation": True,
            # Stabilní identita řádku → cílené transakce bez probliknutí.
            ":getRowId": "(p) => String(p.data._rid)",
            # Sloupce automaticky na šířku obsahu (vč. záhlaví) — ať se hlavičky neořezávají.
            ":onFirstDataRendered": "(p) => p.api.autoSizeAllColumns()",
            ":onRowDataUpdated": "(p) => p.api.autoSizeAllColumns()",
        }).classes("w-full").style("height: 380px")

        info_lbl = ui.label("").classes("text-xs text-gray-500")

        def _tx(add=None, update=None, remove=None):
            """Cílená AG Grid transakce — překreslí jen dotčené řádky (bez probliknutí)."""
            t = {}
            if add:    t["add"] = add
            if update: t["update"] = update
            if remove: t["remove"] = remove
            if t:
                grid.run_grid_method("applyTransaction", t)

        def _najdi(rid):
            for r in rows:
                if r.get("_rid") == rid:
                    return r
            return None

        def _nacti_ico():
            text = (ico_paste.value or "").strip()
            if not text:
                ui.notify("Vložte alespoň jedno IČO.", type="warning")
                return
            icos = []
            for line in text.replace(",", "\n").replace(";", "\n").splitlines():
                v = line.strip().split("\t")[0].strip()
                if v:
                    icos.append(v)
            if not icos:
                return
            mapa = _kontakty_dle_ico(icos)
            stare = list(rows)        # pro odebrání starých řádků transakcí
            rows.clear()
            nenalezeno = 0
            for ico in icos:
                k = mapa.get(ico)
                r = _novy_radek(ico=ico)
                if k:
                    for key, _ in _KONTAKT_SLOUPCE:
                        r[key] = k.get(key, "") or ""
                    r["puvodni_provize"] = k.get("puvodni_provize", "") or ""
                else:
                    nenalezeno += 1
                rows.append(r)
            # Nahrazení obsahu transakcí (remove staré + add nové) — verze-bezpečné,
            # bez plného přepisu rowData (a tedy bez probliknutí).
            _tx(remove=stare, add=list(rows))
            msg = f"Načteno {len(rows)} řádků."
            if nenalezeno:
                msg += f" {nenalezeno} IČO nenalezeno v číselníku."
            info_lbl.text = msg
            ui.notify(msg, type="positive" if not nenalezeno else "warning", position="top-right")

        def _pridej_radek():
            r = _novy_radek()
            rows.append(r)
            _tx(add=[r])

        async def _smaz_vybrane():
            sel = await grid.get_selected_rows()
            sel_ids = {s.get("_rid") for s in sel}
            if not sel_ids:
                ui.notify("Označte řádky ke smazání.", type="warning")
                return
            smaz = [r for r in rows if r.get("_rid") in sel_ids]
            rows[:] = [r for r in rows if r.get("_rid") not in sel_ids]
            if not rows:
                novy = _novy_radek()
                rows.append(novy)
                _tx(remove=smaz, add=[novy])
            else:
                _tx(remove=smaz)

        def _on_cell_change(e):
            a = e.args or {}
            data = a.get("data") or {}
            rid = data.get("_rid")
            field = a.get("colId") or (a.get("column") or {}).get("colId")
            new_val = data.get(field) if field else None
            r = _najdi(rid)
            if not r:
                return
            if field:
                r[field] = new_val
            # Ruční změna IČO → zkus dotáhnout kontakt (jen update jednoho řádku)
            if field == "ico" and new_val:
                k = _kontakty_dle_ico([str(new_val).strip()]).get(str(new_val).strip())
                if k:
                    for key, _ in _KONTAKT_SLOUPCE:
                        r[key] = k.get(key, "") or ""
                    r["puvodni_provize"] = k.get("puvodni_provize", "") or ""
                    _tx(update=[r])
        grid.on("cellValueChanged", _on_cell_change)

        # ── Excel-like schránka v gridu (Community AG Grid, vlastní JS) ──────
        # Ctrl+C zkopíruje označený obdélník bunk jako TSV, Ctrl+V vloží blok
        # z Excelu od aktivní buňky (vyplní dolů/doprava, řádky se dle potřeby
        # přidají), Ctrl+Shift+šipky rozšíří výběr. Sloupce v pořadí zobrazení:
        _GRID_COLS = ["ico", "jmeno", "k_mesto", "dealer", "puvodni_provize",
                      "nova_provize", "oz2", "datum_od", "datum_do"]

        def _on_grid_paste(e):
            a = e.args or {}
            text = a.get("text") or ""
            try:
                start_r = int(a.get("rowIndex"))
            except (TypeError, ValueError):
                start_r = 0
            col_id = a.get("colId")
            start_c = _GRID_COLS.index(col_id) if col_id in _GRID_COLS else 0
            import csv as _csv
            matrix = [r for r in _csv.reader(io.StringIO(text), delimiter="\t")]
            # VNITŘNÍ prázdné řádky zachovat (respektovat mezery ve sloupci);
            # odříznout jen KONCOVÉ prázdné (z koncového newlinu).
            while matrix and not any((c or "").strip() for c in matrix[-1]):
                matrix.pop()
            if not matrix:
                return

            # ── Fill-down: schránka = jediná buňka (1×1) + označený rozsah více buněk ──
            # Excel chování: rozkopíruj tu jednu hodnotu do CELÉHO označeného výběru.
            sel = a.get("sel") or {}
            def _si(k):
                try:
                    return int(sel.get(k))
                except (TypeError, ValueError):
                    return None
            sr1, sr2, sc1, sc2 = _si("r1"), _si("r2"), _si("c1"), _si("c2")
            je_jedna = len(matrix) == 1 and len(matrix[0]) == 1
            if je_jedna and None not in (sr1, sr2, sc1, sc2) and (sr2 > sr1 or sc2 > sc1):
                val = (matrix[0][0] or "").strip()
                r_od, r_do = max(0, min(sr1, sr2)), min(len(rows) - 1, max(sr1, sr2))
                c_od, c_do = max(0, min(sc1, sc2)), min(len(_GRID_COLS) - 1, max(sc1, sc2))
                touched, zmena_ico = set(), []
                for ri in range(r_od, r_do + 1):
                    for ci in range(c_od, c_do + 1):
                        field = _GRID_COLS[ci]
                        rows[ri][field] = val
                        touched.add(ri)
                        if field == "ico" and val:
                            zmena_ico.append(ri)
                if zmena_ico:
                    mapa = _kontakty_dle_ico([rows[ri]["ico"] for ri in zmena_ico])
                    for ri in zmena_ico:
                        k = mapa.get(rows[ri]["ico"])
                        if k:
                            for key, _ in _KONTAKT_SLOUPCE:
                                rows[ri][key] = k.get(key, "") or ""
                            rows[ri]["puvodni_provize"] = k.get("puvodni_provize", "") or ""
                _tx(update=[rows[ri] for ri in sorted(touched)])
                ui.notify(f"Hodnota rozkopírována do {len(touched)} řádků.",
                          type="positive", position="top-right")
                return

            orig_len = len(rows)
            touched, zmena_ico = set(), []
            for dr, line in enumerate(matrix):
                ri = start_r + dr
                while ri >= len(rows):
                    rows.append(_novy_radek())
                for dc, val in enumerate(line):
                    ci = start_c + dc
                    if ci >= len(_GRID_COLS):
                        break
                    field = _GRID_COLS[ci]
                    rows[ri][field] = (val or "").strip()
                    touched.add(ri)
                    if field == "ico" and rows[ri]["ico"]:
                        zmena_ico.append(ri)
            # Dotáhni kontakty pro vložená IČO (jako při ruční změně IČO).
            if zmena_ico:
                mapa = _kontakty_dle_ico([rows[ri]["ico"] for ri in zmena_ico])
                for ri in zmena_ico:
                    k = mapa.get(rows[ri]["ico"])
                    if k:
                        for key, _ in _KONTAKT_SLOUPCE:
                            rows[ri][key] = k.get(key, "") or ""
                        rows[ri]["puvodni_provize"] = k.get("puvodni_provize", "") or ""
            # Cílená transakce: nové řádky (add) + upravené stávající (update).
            new_rows = rows[orig_len:]
            upd = [rows[ri] for ri in sorted(touched) if ri < orig_len]
            _tx(add=new_rows, update=upd)
            ui.run_javascript(
                f"setTimeout(() => {{ const c = getElement({grid.id});"
                f" const api = c && (c.api || (c.gridOptions && c.gridOptions.api));"
                f" if (api) api.setFocusedCell({start_r}, {json.dumps(col_id or _GRID_COLS[start_c])}); }}, 60);")
            ui.notify(f"Vloženo {len(matrix)} řádků ze schránky.",
                      type="positive", position="top-right")

        def _on_grid_clear(e):
            a = e.args or {}
            try:
                r1, r2 = int(a.get("r1")), int(a.get("r2"))
                c1, c2 = int(a.get("c1")), int(a.get("c2"))
            except (TypeError, ValueError):
                return
            upd = []
            for ri in range(max(0, r1), min(len(rows) - 1, r2) + 1):
                for ci in range(max(0, c1), min(len(_GRID_COLS) - 1, c2) + 1):
                    rows[ri][_GRID_COLS[ci]] = ""
                upd.append(rows[ri])
            _tx(update=upd)
            ui.run_javascript(
                f"setTimeout(() => {{ const c = getElement({grid.id});"
                f" const api = c && (c.api || (c.gridOptions && c.gridOptions.api));"
                f" if (api) api.setFocusedCell({max(0, r1)}, {json.dumps(_GRID_COLS[max(0, c1)])}); }}, 60);")

        def _on_grid_addrow(e):
            _pridej_radek()
            last = len(rows) - 1
            ui.run_javascript(
                f"setTimeout(() => {{ const c = getElement({grid.id});"
                f" const api = c && (c.api || (c.gridOptions && c.gridOptions.api));"
                f" if (api) {{ api.ensureIndexVisible({last}); api.setFocusedCell({last}, 'ico'); }} }}, 60);")

        async def _on_grid_delrow(e):
            a = e.args or {}
            # 1) Přednost mají zaškrtnuté řádky (vícenásobný výběr přes checkboxy).
            sel = await grid.get_selected_rows()
            nf = 0
            if sel:
                sel_ids = {s.get("_rid") for s in sel}
                smaz = [r for r in rows if r.get("_rid") in sel_ids]
                nf = min((i for i, r in enumerate(rows) if r.get("_rid") in sel_ids), default=0)
                rows[:] = [r for r in rows if r.get("_rid") not in sel_ids]
            else:
                # 2) Jinak řádky z výběru buněk (rozsah r1..r2).
                try:
                    r1, r2 = int(a.get("r1")), int(a.get("r2"))
                except (TypeError, ValueError):
                    return
                r1, r2 = max(0, min(r1, r2)), min(len(rows) - 1, max(r1, r2))
                if r1 > r2:
                    return
                smaz = rows[r1:r2 + 1]
                del rows[r1:r2 + 1]
                nf = r1
            add = None
            if not rows:   # nech aspoň jeden prázdný řádek
                novy = _novy_radek()
                rows.append(novy)
                add = [novy]
            _tx(remove=smaz, add=add)
            nf = max(0, min(nf, len(rows) - 1))
            ui.run_javascript(
                f"setTimeout(() => {{ const c = getElement({grid.id});"
                f" const api = c && (c.api || (c.gridOptions && c.gridOptions.api));"
                f" if (api) api.setFocusedCell({nf}, 'ico'); }}, 60);")

        _paste_evt = f"asm_grid_paste_{grid.id}"
        _clear_evt = f"asm_grid_clear_{grid.id}"
        _addrow_evt = f"asm_grid_addrow_{grid.id}"
        _delrow_evt = f"asm_grid_delrow_{grid.id}"
        ui.on(_paste_evt, _on_grid_paste)
        ui.on(_clear_evt, _on_grid_clear)
        ui.on(_addrow_evt, _on_grid_addrow)
        ui.on(_delrow_evt, _on_grid_delrow)
        ui.add_css(".jip-cellsel{background:#bfdbfe !important;}")
        ui.run_javascript(f"""
        setTimeout(() => {{
          const COLS = {json.dumps(_GRID_COLS)};
          const EVT = "{_paste_evt}";
          const EVT_CLEAR = "{_clear_evt}";
          const EVT_ADDROW = "{_addrow_evt}";
          const EVT_DELROW = "{_delrow_evt}";
          const comp = getElement({grid.id});
          const root = comp && comp.$el;
          if (!root || root.__jipClip) return;
          root.__jipClip = true;
          const getCell = (ri, col) =>
            root.querySelector('.ag-row[row-index="' + ri + '"] .ag-cell[col-id="' + col + '"]');
          const curFocus = () => {{
            const fc = root.querySelector('.ag-cell-focus[col-id]');
            if (!fc) return null;
            const rowEl = fc.closest('.ag-row');
            if (!rowEl) return null;
            const ri = parseInt(rowEl.getAttribute('row-index'));
            const ci = COLS.indexOf(fc.getAttribute('col-id'));
            return (isNaN(ri) || ci < 0) ? null : {{ri, ci}};
          }};
          let anchor = null, head = null;
          const clearHi = () => root.querySelectorAll('.jip-cellsel')
                                     .forEach(el => el.classList.remove('jip-cellsel'));
          const applyHi = () => {{
            clearHi();
            if (!anchor || !head) return;
            const r1 = Math.min(anchor.ri, head.ri), r2 = Math.max(anchor.ri, head.ri);
            const c1 = Math.min(anchor.ci, head.ci), c2 = Math.max(anchor.ci, head.ci);
            for (let ri = r1; ri <= r2; ri++)
              for (let ci = c1; ci <= c2; ci++) {{
                const cell = getCell(ri, COLS[ci]);
                if (cell) cell.classList.add('jip-cellsel');
              }}
          }};
          const edituje = () => {{
            const ae = document.activeElement;
            return ae && (ae.tagName === 'INPUT' || ae.tagName === 'TEXTAREA');
          }};
          const inGrid = () => document.body.contains(root) && !edituje() &&
            (root.contains(document.activeElement) || root.querySelector('.ag-cell-focus'));
          // TSV z aktuálního výběru (čte buňky z DOM; virtualizace je vypnutá).
          const buildTSV = () => {{
            if (!anchor || !head) return '';
            const r1 = Math.min(anchor.ri, head.ri), r2 = Math.max(anchor.ri, head.ri);
            const c1 = Math.min(anchor.ci, head.ci), c2 = Math.max(anchor.ci, head.ci);
            const lines = [];
            for (let ri = r1; ri <= r2; ri++) {{
              const cells = [];
              for (let ci = c1; ci <= c2; ci++) {{
                const cell = getCell(ri, COLS[ci]);
                cells.push(cell ? cell.textContent.trim() : '');
              }}
              lines.push(cells.join('\\t'));
            }}
            return lines.join('\\n');
          }};
          // Zápis do schránky — textarea+execCommand (funguje i na http, v rámci gesta).
          // Dočasná textarea ukradne focus → po kopii ho vrátíme na původní buňku,
          // aby uživatel nepřišel o focus/výběr při Ctrl+C.
          const doCopy = (txt) => {{
            const prev = document.activeElement;
            try {{
              const ta = document.createElement('textarea');
              ta.value = txt; ta.style.position = 'fixed'; ta.style.opacity = '0';
              document.body.appendChild(ta); ta.focus(); ta.select();
              document.execCommand('copy'); document.body.removeChild(ta);
            }} catch (_) {{
              try {{ if (navigator.clipboard) navigator.clipboard.writeText(txt); }} catch (e) {{}}
            }}
            try {{ if (prev && prev.focus) prev.focus({{preventScroll: true}}); }} catch (e) {{}}
          }};
          // Kotva výběru se nastaví při kliknutí do buňky (mousedown), NE při každé
          // změně focusu — jinak by ji posun focusu (i z naší navigace) přenastavil.
          root.addEventListener('mousedown', (ev) => {{
            const cell = ev.target.closest && ev.target.closest('.ag-cell[col-id]');
            if (!cell) return;
            const rowEl = cell.closest('.ag-row');
            if (!rowEl) return;
            const ri = parseInt(rowEl.getAttribute('row-index'));
            const ci = COLS.indexOf(cell.getAttribute('col-id'));
            if (!isNaN(ri) && ci >= 0) {{ anchor = {{ri, ci}}; head = {{ri, ci}}; clearHi(); }}
          }});
          // Klik na název sloupce (hlavičku) → označí celý sloupec.
          root.addEventListener('click', (ev) => {{
            const hc = ev.target.closest && ev.target.closest('.ag-header-cell[col-id]');
            if (!hc) return;
            const ci = COLS.indexOf(hc.getAttribute('col-id'));
            if (ci < 0) return;
            const _api = comp && (comp.api || (comp.gridOptions && comp.gridOptions.api));
            const lastRow = Math.max(0, (_api ? _api.getDisplayedRowCount() : 1) - 1);
            anchor = {{ri: 0, ci}};
            head = {{ri: lastRow, ci}};
            applyHi();
          }});
          // Plynulý posun bez Shiftu → posuň kotvu na nově zaměřenou buňku.
          root.addEventListener('keyup', (e) => {{
            if (e.shiftKey || edituje()) return;
            if (!['ArrowUp','ArrowDown','ArrowLeft','ArrowRight'].includes(e.key)) return;
            const f = curFocus(); if (f) {{ anchor = f; head = f; clearHi(); }}
          }});
          // Shift+šipky = rozšíření výběru. CAPTURE fáze + stopPropagation, aby
          // AG Grid sám neposunul focus (jinak by se výběr posouval o buňku).
          // Ctrl + „+"/„−" (přidat/smazat řádek). Na DOCUMENTU v capture fázi a vždy
          // potlačit, dokud je formulář otevřený — jinak při rychlém mačkání (focus
          // krátce mimo grid po překreslení) prohlížeč zoomuje stránku.
          document.addEventListener('keydown', (e) => {{
            if (!document.body.contains(root) || !e.ctrlKey) return;
            const plus  = (e.key === '+' || e.key === '=' || e.code === 'NumpadAdd');
            const minus = (e.key === '-' || e.key === '_' || e.code === 'NumpadSubtract');
            if (!plus && !minus) return;
            e.preventDefault();
            e.stopPropagation();
            if (edituje()) return;   // při editaci buňky/pole jen potlač zoom, neprováděj akci
            if (plus) {{
              emitEvent(EVT_ADDROW, {{}});
            }} else {{
              let r1 = null, r2 = null;
              const a1 = anchor ? anchor.ri : null;
              if (a1 !== null) {{ const h = head ? head.ri : a1; r1 = Math.min(a1, h); r2 = Math.max(a1, h); }}
              emitEvent(EVT_DELROW, {{r1, r2}});
            }}
          }}, true);
          root.addEventListener('keydown', (e) => {{
            if (edituje()) return;
            // Delete/Backspace → smaž všechny označené buňky (ne jen kotvu).
            if ((e.key === 'Delete' || e.key === 'Backspace') && anchor && head) {{
              const r1 = Math.min(anchor.ri, head.ri), r2 = Math.max(anchor.ri, head.ri);
              const c1 = Math.min(anchor.ci, head.ci), c2 = Math.max(anchor.ci, head.ci);
              emitEvent(EVT_CLEAR, {{r1, r2, c1, c2}});
              e.preventDefault();
              e.stopPropagation();
              return;
            }}
            // Ctrl+C → zkopíruj výběr do schránky sami (nativní copy se bez textového
            // výběru nemusí vůbec vyvolat — např. po kliku na hlavičku).
            if (e.ctrlKey && !e.shiftKey && (e.key === 'c' || e.key === 'C') && anchor && head) {{
              doCopy(buildTSV());
              // Vrať AG focus na aktivní roh výběru (doCopy obnoví DOM focus, tímto
              // sedí i vnitřní .ag-cell-focus AG Gridu).
              const _api = comp && (comp.api || (comp.gridOptions && comp.gridOptions.api));
              if (_api && head) {{ try {{ _api.setFocusedCell(head.ri, COLS[head.ci]); }} catch (_e) {{}} }}
              e.preventDefault();
              e.stopPropagation();
              return;
            }}
            if (!e.shiftKey) return;   // Shift + šipky (rozšíření o 1) / Ctrl+Shift+šipky (na kraj)
            if (!['ArrowUp','ArrowDown','ArrowLeft','ArrowRight'].includes(e.key)) return;
            if (!anchor) {{ const f = curFocus(); if (!f) return; anchor = f; head = f; }}
            const _api = comp && (comp.api || (comp.gridOptions && comp.gridOptions.api));
            const lastRow = Math.max(0, (_api ? _api.getDisplayedRowCount() : (head.ri + 1)) - 1);
            const lastCol = COLS.length - 1;
            let ri = head.ri, ci = head.ci;
            if (e.ctrlKey) {{
              // Ctrl+Shift+šipka → výběr od aktivní buňky až na kraj v daném směru.
              if (e.key === 'ArrowUp')    ri = 0;
              if (e.key === 'ArrowDown')  ri = lastRow;
              if (e.key === 'ArrowLeft')  ci = 0;
              if (e.key === 'ArrowRight') ci = lastCol;
            }} else {{
              if (e.key === 'ArrowUp')    ri = Math.max(0, ri - 1);
              if (e.key === 'ArrowDown')  ri = Math.min(lastRow, ri + 1);
              if (e.key === 'ArrowLeft')  ci = Math.max(0, ci - 1);
              if (e.key === 'ArrowRight') ci = Math.min(lastCol, ci + 1);
            }}
            head = {{ri, ci}};
            applyHi();
            e.preventDefault();
            e.stopPropagation();
          }}, true);
          // Fallback: když se nativní copy přece jen vyvolá (textový výběr), použij ho.
          document.addEventListener('copy', (e) => {{
            if (edituje() || !document.body.contains(root) || !anchor || !head) return;
            const maVyber = !!root.querySelector('.jip-cellsel');
            if (!maVyber && !(root.contains(document.activeElement) ||
                              root.querySelector('.ag-cell-focus'))) return;
            e.clipboardData.setData('text/plain', buildTSV());
            e.preventDefault();
          }});
          document.addEventListener('paste', (e) => {{
            if (edituje() || !document.body.contains(root)) return;
            // Cíl ber z kotvy výběru (klik na hlavičku/buňku ji aktualizuje), ne ze
            // staré fokusnuté buňky — jinak by se vkládalo pořád do původního sloupce.
            const maVyber = !!root.querySelector('.jip-cellsel');
            if (!maVyber && !(root.contains(document.activeElement) ||
                              root.querySelector('.ag-cell-focus'))) return;
            const f = anchor || curFocus() || head;
            if (!f) return;
            const txt = (e.clipboardData || window.clipboardData).getData('text');
            if (!txt) return;
            e.preventDefault();
            // Předej i rozsah označeného výběru (pro fill-down jedné hodnoty do více buněk).
            const sel = (anchor && head) ? {{
              r1: Math.min(anchor.ri, head.ri), r2: Math.max(anchor.ri, head.ri),
              c1: Math.min(anchor.ci, head.ci), c2: Math.max(anchor.ci, head.ci)
            }} : null;
            emitEvent(EVT, {{text: txt, rowIndex: f.ri, colId: COLS[f.ci], sel: sel}});
          }});
        }}, 300);
        """)

        with ui.row().classes("w-full gap-2 mt-2 flex-wrap"):
            ui.button("Načíst IČO", icon="download", on_click=_nacti_ico) \
                .props("outline no-caps").classes("text-emerald-700")
            ui.button("Přidat řádek", icon="add", on_click=_pridej_radek) \
                .props("flat no-caps").classes("text-gray-600")
            ui.button("Smazat vybrané", icon="delete", on_click=_smaz_vybrane) \
                .props("flat no-caps").classes("text-red-600")

        # ── Odeslání ────────────────────────────────────────────────
        async def _odeslat():
            if not (duvod_in.value or "").strip():
                ui.notify("Vyplňte „Důvod změny“ (povinné).", type="warning")
                return
            platne = [r for r in rows if (r.get("ico") or "").strip()]
            if not platne:
                ui.notify("Detailní tabulka neobsahuje žádné IČO.", type="warning")
                return
            hlavicka = {
                "duvod_zmeny": (duvod_in.value or "").strip(),
                "asm_jmeno": user_name,
                "zmena_asm_jmeno": (zmena_asm_in.value or "").strip(),
                "novy_asm_jmeno": (novy_asm_in.value or "") or "",
                "cislo_oz": (cislo_oz_in.value or "") or "",
                "jmeno_oz": (jmeno_oz_in.value or "").strip(),
                "cislo_novy_oz": (cislo_novy_oz_in.value or "") or "",
                "jmeno_novy_oz": (jmeno_novy_oz_in.value or "").strip(),
                "datum_zmeny_od": (datum_od_in.value or "").strip(),
                "zakaznik_v_regionu_oz": (region_in.value or "NE"),
                "oduvodneni": (oduvodneni_in.value or "").strip(),
            }
            pid, cislo = zaloz_pripad(hlavicka, platne, user_id, user_name)
            if not pid:
                ui.notify(f"Uložení selhalo: {cislo}", type="negative")
                return
            zaznam_historie(pid, "Odesláno", user_name,
                            f"Vytvořen případ ({len(platne)} řádků)")
            p = nacti_pripad(pid)
            _odesli_emaily(_emaily_office(),
                           f"Formuláře ASM — nový případ {cislo}",
                           f"Máte nový případ k vyřešení.\n\nČíslo: {cislo}\n"
                           f"Žadatel: {user_name}\nDůvod: {hlavicka['duvod_zmeny']}\n"
                           f"Počet řádků: {len(platne)}", pripad_id=pid)
            # Notifikace dotčených OZ (původní + nový) při podání žádosti.
            _notifikuj_oz([hlavicka.get("cislo_oz"), hlavicka.get("cislo_novy_oz")],
                          cislo, "Změna zákazníků na OZ/ASM", user_name,
                          f"{len(platne)} řádků", pid)
            intranet_logger.log_activity(user_name, "Podání žádosti",
                                         f"Formuláře ASM: nový případ {cislo} ({len(platne)} řádků)")
            ui.notify(f"Případ {cislo} odeslán.", type="positive", position="top")
            app.storage.user["asm_sekce"] = "seznam"
            _refresh()

        with ui.row().classes("w-full justify-end mt-4 gap-2"):
            ui.button("Zrušit", on_click=zpet_fn).props("flat no-caps").classes("text-gray-600")
            ui.button("Odeslat", icon="send", on_click=_odeslat) \
                .props("unelevated no-caps") \
                .classes("bg-emerald-600 hover:bg-emerald-700 text-white font-semibold rounded-lg px-5")


# ============================================================================
# Detail případu + workflow
# ============================================================================
def _detail_dialog(pid, user_id, user_name, prava):
    p = nacti_pripad(pid)
    if not p:
        ui.notify("Případ nenalezen.", type="negative")
        return
    radky = nacti_radky(pid)
    historie = nacti_historie(pid)
    je_office = _je_office(prava)
    je_spravce = _je_spravce(prava)
    je_zadatel_vlastnik = (p.get("zadavatel_id") == user_id)
    stav = p.get("stav")

    # Dialog navěsíme na stabilní top-level slot stránky, NE do slotu očičko-tlačítka
    # (to leží uvnitř @ui.refreshable _tabulka; auto-obnova à 10 s by dialog jinak
    #  smazala → samovolné zavírání detailu). Karta a obsah zůstávají dětmi dialogu.
    with context.client.content:
        dlg = ui.dialog()
    with dlg, ui.card().classes("p-5 gap-3").style("min-width: 70vw; max-width: 92vw"):
        with ui.row().classes("w-full items-center gap-3"):
            ui.label(p.get("cislo") or f"#{pid}").classes("text-2xl font-bold text-gray-800")
            _badge(stav)
            ui.space()
            ui.button(icon="close", on_click=dlg.close).props("flat round color=grey-7")

        formular = p.get("formular") or "oz_zmena"
        # Žadatel-vlastník smí po vrácení k opravě editovat (dvojklik / vstupy).
        lze_upravit = je_zadatel_vlastnik and stav == "vraceno_oprava"
        cti_form = None   # getter dat pro unifikované formuláře v režimu úpravy

        def _info(popisek, hodnota):
            with ui.row().classes("gap-2 no-wrap"):
                ui.label(popisek + ":").classes("font-semibold text-gray-500 whitespace-nowrap")
                ui.label(str(hodnota or "—")).classes("text-gray-800 break-words")

        def _spolecne_info():
            if p.get("poznamka"):
                _info("Poznámka", p.get("poznamka"))
            if p.get("zamitnuti_duvod"):
                _info("Důvod zamítnutí", p.get("zamitnuti_duvod"))
            if p.get("storno_duvod"):
                _info("Důvod storna", p.get("storno_duvod"))

        if formular == "oz_zmena":
            # Hlavička
            with ui.element("div").classes("w-full grid gap-x-6 gap-y-1 text-sm") \
                    .style("grid-template-columns: repeat(2, minmax(0,1fr))"):
                _info("Důvod změny", p.get("duvod_zmeny"))
                _info("Žadatel", p.get("zadavatel_jmeno"))
                _info("Jméno ASM", p.get("asm_jmeno"))
                _info("Datum zadání", _dt_cz(p.get("datum_zadani")))
                _info("Změna ASM → Nový ASM", f"{p.get('zmena_asm_jmeno') or '—'} → {p.get('novy_asm_jmeno') or '—'}")
                _info("OZ → Nový OZ", f"{p.get('cislo_oz') or '—'} {('('+p['jmeno_oz']+')') if p.get('jmeno_oz') else ''} → "
                                      f"{p.get('cislo_novy_oz') or '—'} {('('+p['jmeno_novy_oz']+')') if p.get('jmeno_novy_oz') else ''}")
                _info("Datum změny od", p.get("datum_zmeny_od"))
                _info("Zákazník v regionu OZ", p.get("zakaznik_v_regionu_oz"))
                if p.get("oduvodneni"):
                    _info("Případné odůvodnění", p.get("oduvodneni"))
                _spolecne_info()

            ui.label(f"Detailní tabulka ({len(radky)} řádků)"
                     + ("  — dvojklik na buňku = úprava" if lze_upravit else "")) \
                .classes("text-sm font-semibold text-gray-500 uppercase mt-2")
            det_cols = [
                {"headerName": "#", "field": "poradi", "width": 60, "editable": False},
                {"headerName": "IČO", "field": "ico", "width": 120, "editable": lze_upravit},
            ] + [{"headerName": lbl, "field": key, "width": 140, "editable": lze_upravit}
                 for key, lbl in _KONTAKT_SLOUPCE] + [
                {"headerName": "Pův. provize", "field": "puvodni_provize", "width": 110, "editable": lze_upravit},
                {"headerName": "Nová provize", "field": "nova_provize", "width": 120, "editable": lze_upravit},
                {"headerName": "OZ2", "field": "oz2", "width": 110, "editable": lze_upravit},
                {"headerName": "Datum od", "field": "datum_od", "width": 120, "editable": lze_upravit},
                {"headerName": "Datum do", "field": "datum_do", "width": 120, "editable": lze_upravit},
            ]
            det_grid = ui.aggrid({
                "columnDefs": det_cols,
                "rowData": radky,
                "defaultColDef": {"resizable": True, "sortable": True,
                                  "wrapHeaderText": True, "autoHeaderHeight": True},
                # Sloupce roztažené dle nejdelšího textu (spolehlivě už při prvním renderu).
                "autoSizeStrategy": {"type": "fitCellContents"},
                "rowHeight": 30,
                "singleClickEdit": False,   # editace buňky dvojklikem
                "stopEditingWhenCellsLoseFocus": True,
                ":onFirstDataRendered": "(p) => p.api.autoSizeAllColumns()",
                ":onRowDataUpdated": "(p) => p.api.autoSizeAllColumns()",
            }).classes("w-full").style("height: 300px")
            if lze_upravit:
                def _det_cell_change(e):
                    a = e.args or {}
                    data = a.get("data") or {}
                    rid = data.get("id")
                    field = a.get("colId") or (a.get("column") or {}).get("colId")
                    if rid is None or not field:
                        return
                    for rr in radky:
                        if rr.get("id") == rid:
                            rr[field] = data.get(field)
                            break
                det_grid.on("cellValueChanged", _det_cell_change)
        else:
            # Unifikovaný formulář (Navýšení limitů / Změna dodacích listů)
            with ui.element("div").classes("w-full grid gap-x-6 gap-y-1 text-sm mb-1") \
                    .style("grid-template-columns: repeat(2, minmax(0,1fr))"):
                _info("Žadatel (vypracoval)", p.get("zadavatel_jmeno"))
                _info("Datum vypracování", _dt_cz(p.get("datum_zadani")))
                _spolecne_info()
            if lze_upravit:
                ui.label("Úprava formuláře — opravte pole a uložte (rolovátka, IČO dotažení "
                         "i ručně přepsatelná auto-pole fungují).") \
                    .classes("text-sm text-amber-700")
            _fdata = _data_pripadu(p)
            # Office obchod / správce porovnává obrat proti limitu z PŘESNÉ částky
            # (žadatel z maskované). V režimu opravy se nepřepisuje — uložená
            # maskovaná hodnota by se jinak při uložení přepsala přesnou.
            if je_office and formular == "limity" and not lze_upravit:
                _fdata = dict(_fdata)
                for _ok in ("obrat_1m", "obrat_3m"):
                    _real = _fdata.get(_ok + "_real")
                    if _real:
                        _fdata[_ok] = _real
                # Přesné obraty i u jednotlivých IČ 10 v tabulce.
                _var = []
                for _v in (_fdata.get("ico10_varianty") or []):
                    if isinstance(_v, dict):
                        _v = dict(_v)
                        for _ok in ("obrat_1m", "obrat_3m"):
                            if _v.get(_ok + "_real"):
                                _v[_ok] = _v[_ok + "_real"]
                    _var.append(_v)
                _fdata["ico10_varianty"] = _var
            # Read-only detail ukazuje uložený výběr (filtr se neuplatní); při opravě
            # žadatelem se re-dotažení dle IČO omezí na jeho pobočku stejně jako v novém
            # formuláři.
            cti_form = _render_form_pole(
                formular, _fdata, editable=lze_upravit, auto_editovatelne=lze_upravit,
                pobocka_filtr=(None if not lze_upravit or _je_office(prava)
                               else _pobocka_uzivatele(user_id)))

        # Průběh („očičko")
        with ui.expansion("Průběh případu", icon="history").classes("w-full"):
            if not historie:
                ui.label("Zatím bez záznamů.").classes("text-gray-400 italic text-sm")
            for h in historie:
                with ui.row().classes("w-full gap-2 text-sm items-baseline"):
                    ui.label(_dt_cz(h.get("kdy"))).classes("text-gray-400 whitespace-nowrap")
                    ui.label(h.get("akce") or "").classes("font-semibold text-gray-700 whitespace-nowrap")
                    ui.label(f"· {h.get('kdo') or ''}").classes("text-gray-500 whitespace-nowrap")
                    # Poznámka Office → správce je interní: žadateli se v průběhu
                    # ukáže jen holý krok „Postoupeno správci“ bez textu.
                    _interni = (h.get("akce") == "Postoupeno správci"
                                and not (je_office or je_spravce))
                    if h.get("detail") and not _interni:
                        ui.label(f"— {h['detail']}").classes("text-gray-600 break-words")

        # Export
        async def _export():
            if formular == "oz_zmena":
                data, nazev = await asyncio.to_thread(_export_xlsx, p, radky)
            else:
                data, nazev = await asyncio.to_thread(_export_formular_xlsx, p)
            ui.download.content(data, nazev,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # Poznámka od Office pro správce (interní — žadatel ji nevidí)
        if p.get("spravce_pozn") and (je_office or je_spravce):
            with ui.column().classes("w-full gap-1 mt-3 p-3 rounded-lg "
                                     "bg-indigo-50 border border-indigo-200"):
                with ui.row().classes("items-center gap-2"):
                    ui.icon("arrow_upward").classes("text-indigo-700")
                    ui.label("Poznámka správci") \
                        .classes("font-semibold text-indigo-800")
                ui.label(p.get("spravce_pozn")).classes("text-gray-800 break-words whitespace-pre-wrap")

        with ui.row().classes("w-full justify-between items-center mt-3 flex-wrap gap-2"):
            ui.button("Export do Excelu", icon="file_download", on_click=_export) \
                .props("outline no-caps").classes("text-emerald-700")

            # ── Workflow akce ───────────────────────────────────────
            with ui.row().classes("gap-2 flex-wrap"):
                def _akce_email_office(predmet, text):
                    _odesli_emaily(_emaily_office(), predmet, text, pripad_id=pid)

                # Žadatel – případ vrácen k opravě
                if je_zadatel_vlastnik and stav == "vraceno_oprava":
                    def _uloz_opravu():
                        if formular == "oz_zmena":
                            ok = uloz_radky(pid, radky)
                        else:
                            ok = uloz_data_formular(pid, cti_form()) if cti_form else False
                        if not ok:
                            ui.notify("Uložení opravy selhalo.", type="negative"); return
                        zmen_stav(pid, "odeslano")
                        zaznam_historie(pid, "Opraveno žadatelem", user_name,
                                        "Formulář upraven a odeslán zpět ke zpracování")
                        _akce_email_office(f"Formuláře ASM — případ {p.get('cislo')} opraven",
                                           f"Žadatel {user_name} opravil případ {p.get('cislo')} "
                                           f"a odeslal ho zpět ke zpracování.")
                        ui.notify("Oprava uložena a odeslána Office.", type="positive")
                        dlg.close(); _refresh()

                    def _trvam():
                        zmen_stav(pid, "vraceno_zpet")
                        zaznam_historie(pid, "Vráceno zpět na Office", user_name,
                                        "Žadatel trvá na svém")
                        _akce_email_office(f"Formuláře ASM — případ {p.get('cislo')} vrácen zpět",
                                           f"Žadatel {user_name} vrátil případ {p.get('cislo')} "
                                           f"zpět na Office (trvá na svém).")
                        ui.notify("Odesláno zpět na Office.", type="positive")
                        dlg.close(); _refresh()
                    ui.button("Uložit opravu a odeslat", icon="save",
                              on_click=_uloz_opravu).props("unelevated no-caps") \
                        .classes("bg-emerald-600 text-white rounded-lg")
                    ui.button("Trvám na svém – vrátit Office", icon="reply",
                              on_click=_trvam).props("outline no-caps color=purple")

                # Office akce
                if je_office and stav in ("odeslano", "vraceno_zpet", "spravce_schvalil"):
                    def _zpracovano():
                        zmen_stav(pid, "zpracovano")
                        zaznam_historie(pid, "Zpracováno", user_name)
                        _odesli_emaily_zadateli(p, f"Formuláře ASM — případ {p.get('cislo')} zpracován",
                                                f"Váš případ {p.get('cislo')} byl zpracován.")
                        intranet_logger.log_activity(user_name, "Schválení žádosti",
                                                     f"Formuláře ASM: zpracováno {p.get('cislo')}")
                        ui.notify("Označeno jako zpracováno.", type="positive")
                        dlg.close(); _refresh()

                    def _vratit():
                        with ui.dialog() as d2, ui.card().classes("p-4 gap-3").style("min-width:420px"):
                            ui.label("Vrátit k opravě").classes("text-lg font-bold")
                            pozn = ui.textarea("Co má žadatel opravit *").props("outlined autogrow").classes("w-full")
                            def _ok():
                                if not (pozn.value or "").strip():
                                    ui.notify("Vyplňte poznámku.", type="warning"); return
                                zmen_stav(pid, "vraceno_oprava", poznamka=(pozn.value or "").strip())
                                zaznam_historie(pid, "Vráceno k opravě", user_name, (pozn.value or "").strip())
                                _odesli_emaily_zadateli(p, f"Formuláře ASM — případ {p.get('cislo')} vrácen k opravě",
                                                        f"Váš případ {p.get('cislo')} byl vrácen k opravě.\n\n"
                                                        f"Poznámka: {(pozn.value or '').strip()}")
                                ui.notify("Vráceno žadateli.", type="positive")
                                d2.close(); dlg.close(); _refresh()
                            with ui.row().classes("w-full justify-end gap-2"):
                                ui.button("Zrušit", on_click=d2.close).props("flat no-caps")
                                ui.button("Vrátit", on_click=_ok).props("unelevated no-caps color=orange")
                        d2.open()

                    def _postoupit():
                        with ui.dialog() as d3, ui.card().classes("p-4 gap-3").style("min-width:460px"):
                            ui.label("Postoupit správci ke schválení").classes("text-lg font-bold")
                            ui.label("Napište, co případ nesplňuje a proč ho má správce přesto "
                                     "schválit. Poznámku uvidí jen Office a správce, žadateli "
                                     "se nezobrazuje.") \
                                .classes("text-sm text-gray-500")
                            pozn3 = ui.textarea("Poznámka správci *") \
                                .props("outlined autogrow").classes("w-full")

                            def _ok3():
                                txt = (pozn3.value or "").strip()
                                if not txt:
                                    ui.notify("Vyplňte poznámku správci.", type="warning"); return
                                zmen_stav(pid, "u_spravce", spravce_pozn=txt)
                                zaznam_historie(pid, "Postoupeno správci", user_name, txt)
                                _odesli_emaily(_emaily_spravce(),
                                               f"Formuláře ASM — případ {p.get('cislo')} ke schválení",
                                               f"Office ({user_name}) postoupil případ "
                                               f"{p.get('cislo')} k rozhodnutí.\n\n"
                                               f"Poznámka správci:\n{txt}", pripad_id=pid)
                                ui.notify("Postoupeno správci.", type="positive")
                                d3.close(); dlg.close(); _refresh()

                            with ui.row().classes("w-full justify-end gap-2"):
                                ui.button("Zrušit", on_click=d3.close).props("flat no-caps")
                                ui.button("Postoupit", on_click=_ok3) \
                                    .props("unelevated no-caps color=indigo")
                        d3.open()

                    ui.button("Zpracováno", icon="task_alt", on_click=_zpracovano) \
                        .props("unelevated no-caps").classes("bg-green-600 text-white rounded-lg")
                    ui.button("Vrátit k opravě", icon="undo", on_click=_vratit) \
                        .props("outline no-caps color=orange")
                    ui.button("Postoupit správci", icon="arrow_upward", on_click=_postoupit) \
                        .props("outline no-caps color=indigo")

                # Office – uzavření
                if je_office and stav == "zpracovano":
                    def _uzavrit():
                        zmen_stav(pid, "uzavreno")
                        zaznam_historie(pid, "Uzavřeno", user_name)
                        ui.notify("Případ uzavřen.", type="positive")
                        dlg.close(); _refresh()
                    ui.button("Uzavřít", icon="lock", on_click=_uzavrit) \
                        .props("unelevated no-caps").classes("bg-gray-700 text-white rounded-lg")

                # Správce akce
                if je_spravce and stav == "u_spravce":
                    def _schvalit():
                        zmen_stav(pid, "spravce_schvalil")
                        zaznam_historie(pid, "Správce schválil", user_name)
                        _odesli_emaily_zadateli(p, f"Formuláře ASM — případ {p.get('cislo')} schválen",
                                                f"Váš případ {p.get('cislo')} byl správcem schválen (nahráno).")
                        _odesli_emaily(_emaily_office(),
                                       f"Formuláře ASM — případ {p.get('cislo')} správce schválil",
                                       f"Správce schválil případ {p.get('cislo')}. Proveďte realizaci.", pripad_id=pid)
                        intranet_logger.log_activity(user_name, "Schválení žádosti",
                                                     f"Formuláře ASM: správce schválil {p.get('cislo')}")
                        ui.notify("Schváleno, vráceno na Office.", type="positive")
                        dlg.close(); _refresh()

                    def _zamitnout():
                        with ui.dialog() as d3, ui.card().classes("p-4 gap-3").style("min-width:420px"):
                            ui.label("Zamítnout případ").classes("text-lg font-bold")
                            duv = ui.textarea("Důvod zamítnutí *").props("outlined autogrow").classes("w-full")
                            def _ok():
                                if not (duv.value or "").strip():
                                    ui.notify("Vyplňte důvod.", type="warning"); return
                                zmen_stav(pid, "zamitnuto", zamitnuti_duvod=(duv.value or "").strip())
                                zaznam_historie(pid, "Zamítnuto", user_name, (duv.value or "").strip())
                                _odesli_emaily_zadateli(p, f"Formuláře ASM — případ {p.get('cislo')} zamítnut",
                                                        f"Váš případ {p.get('cislo')} byl zamítnut.\n\n"
                                                        f"Důvod: {(duv.value or '').strip()}")
                                intranet_logger.log_activity(user_name, "Zamítnutí žádosti",
                                                             f"Formuláře ASM: zamítnuto {p.get('cislo')}")
                                ui.notify("Případ zamítnut.", type="positive")
                                d3.close(); dlg.close(); _refresh()
                            with ui.row().classes("w-full justify-end gap-2"):
                                ui.button("Zrušit", on_click=d3.close).props("flat no-caps")
                                ui.button("Zamítnout", on_click=_ok).props("unelevated no-caps color=red")
                        d3.open()

                    ui.button("Schválit (palec nahoru)", icon="thumb_up", on_click=_schvalit) \
                        .props("unelevated no-caps").classes("bg-teal-600 text-white rounded-lg")
                    ui.button("Zamítnout", icon="thumb_down", on_click=_zamitnout) \
                        .props("outline no-caps color=red")

                # Storno — vlastník nebo Office/Správce, dokud případ není v koncovém stavu
                if stav not in _STAVY_KONCOVE and (je_zadatel_vlastnik or je_office or je_spravce):
                    def _storno():
                        with ui.dialog() as d4, ui.card().classes("p-4 gap-3").style("min-width:420px"):
                            ui.label("Stornovat formulář").classes("text-lg font-bold")
                            ui.label("Případ se uzavře jako stornovaný a nebude dál zpracován.") \
                                .classes("text-sm text-gray-500")
                            duv = ui.textarea("Důvod storna *").props("outlined autogrow").classes("w-full")
                            def _ok():
                                if not (duv.value or "").strip():
                                    ui.notify("Vyplňte důvod.", type="warning"); return
                                zmen_stav(pid, "stornovano", storno_duvod=(duv.value or "").strip())
                                zaznam_historie(pid, "Stornováno", user_name, (duv.value or "").strip())
                                # Žadateli i Office dej vědět (kdo nestornoval).
                                _odesli_emaily_zadateli(p, f"Formuláře ASM — případ {p.get('cislo')} stornován",
                                                        f"Případ {p.get('cislo')} byl stornován.\n\n"
                                                        f"Důvod: {(duv.value or '').strip()}")
                                if not je_zadatel_vlastnik:
                                    pass
                                else:
                                    _odesli_emaily(_emaily_office(),
                                                   f"Formuláře ASM — případ {p.get('cislo')} stornován žadatelem",
                                                   f"Žadatel {user_name} stornoval případ {p.get('cislo')}.\n\n"
                                                   f"Důvod: {(duv.value or '').strip()}", pripad_id=pid)
                                intranet_logger.log_activity(user_name, "Stornování žádosti",
                                                             f"Formuláře ASM: stornováno {p.get('cislo')}")
                                ui.notify("Případ stornován.", type="positive")
                                d4.close(); dlg.close(); _refresh()
                            with ui.row().classes("w-full justify-end gap-2"):
                                ui.button("Zrušit", on_click=d4.close).props("flat no-caps")
                                ui.button("Stornovat", on_click=_ok).props("unelevated no-caps color=grey-8")
                        d4.open()
                    ui.button("Stornovat", icon="cancel", on_click=_storno) \
                        .props("outline no-caps color=grey-7")

                # Smazání případu — JEN administrátor modulu (správce / vse). Nevratné.
                if je_spravce:
                    def _smaz():
                        with ui.dialog() as d5, ui.card().classes("p-4 gap-3").style("min-width:420px"):
                            ui.label("Smazat případ").classes("text-lg font-bold text-red-700")
                            ui.label(f"Případ {p.get('cislo')} bude NEVRATNĚ smazán včetně "
                                     "všech řádků a historie. Tuto akci nelze vrátit zpět.") \
                                .classes("text-sm text-gray-600")
                            def _ok():
                                if not smaz_pripad(pid):
                                    ui.notify("Smazání selhalo.", type="negative"); return
                                intranet_logger.log_activity(user_name, "Stornování žádosti",
                                                             f"Formuláře ASM: SMAZÁN případ {p.get('cislo')}")
                                ui.notify(f"Případ {p.get('cislo')} smazán.", type="positive")
                                d5.close(); dlg.close(); _refresh()
                            with ui.row().classes("w-full justify-end gap-2"):
                                ui.button("Zrušit", on_click=d5.close).props("flat no-caps")
                                ui.button("Smazat nevratně", on_click=_ok) \
                                    .props("unelevated no-caps color=red")
                        d5.open()
                    ui.button("Smazat", icon="delete_forever", on_click=_smaz) \
                        .props("outline no-caps color=red") \
                        .tooltip("Nevratně smaže celý případ — jen administrátor modulu.")
    dlg.open()


# ============================================================================
# Export do Excelu (IČO jako text)
# ============================================================================
def _export_formular_xlsx(p):
    """Export unifikovaného formuláře (Navýšení limitů / Změna dod. listů) do .xlsx.
    Dvousloupcový přehled Položka / Hodnota po sekcích. Vrací (bytes, nazev)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    formular = p.get("formular") or "oz_zmena"
    cfg = _FORMULARE.get(formular, {"nazev": formular, "sekce": []})
    data = _data_pripadu(p)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formulář ASM"
    bold = Font(bold=True)
    sek_font = Font(bold=True, color="1F4E78")
    sek_fill = PatternFill("solid", fgColor="DDEBF7")
    ico_cell = None

    r = 1
    ws.cell(row=r, column=1, value="Typ formuláře").font = bold
    ws.cell(row=r, column=2, value=cfg["nazev"]); r += 1
    ws.cell(row=r, column=1, value="Číslo případu").font = bold
    ws.cell(row=r, column=2, value=p.get("cislo")); r += 1
    ws.cell(row=r, column=1, value="Stav").font = bold
    ws.cell(row=r, column=2, value=_STAV_BADGE.get(p.get("stav"), (p.get("stav"),))[0]); r += 1
    r += 1
    for nadpis, pola in cfg["sekce"]:
        c1 = ws.cell(row=r, column=1, value=nadpis); c1.font = sek_font; c1.fill = sek_fill
        ws.cell(row=r, column=2, value="").fill = sek_fill
        r += 1
        for spec in pola:
            key, label = spec[0], spec[1]
            ws.cell(row=r, column=1, value=label).font = bold
            val = data.get(key, "")
            cell = ws.cell(row=r, column=2, value=(val if val not in (None, "") else "—"))
            if key == "ico":
                cell.number_format = "@"   # IČO jako text
                cell.value = str(val) if val not in (None, "") else "—"
            r += 1
    # Sekce 5 — administrativní (z fronty)
    c1 = ws.cell(row=r, column=1, value="5. Administrativní schválení")
    c1.font = sek_font; c1.fill = sek_fill
    ws.cell(row=r, column=2, value="").fill = sek_fill
    r += 1
    for lbl, val in (("Vypracoval", p.get("zadavatel_jmeno")),
                     ("Datum vypracování žádosti", _dt_cz(p.get("datum_zadani")))):
        ws.cell(row=r, column=1, value=lbl).font = bold
        ws.cell(row=r, column=2, value=val or "—"); r += 1
    for lbl, key in (("Poznámka", "poznamka"), ("Důvod zamítnutí", "zamitnuti_duvod"),
                     ("Důvod storna", "storno_duvod")):
        if p.get(key):
            ws.cell(row=r, column=1, value=lbl).font = bold
            ws.cell(row=r, column=2, value=p.get(key)); r += 1

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 50
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read(), f"{p.get('cislo') or 'formular_asm'}.xlsx"


def _export_xlsx(p, radky):
    """Sestaví .xlsx (běží v threadu). Vrací (bytes, nazev_souboru)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formulář ASM"

    bold = Font(bold=True)
    hdr_fill = PatternFill("solid", fgColor="DDEBF7")
    hdr_font = Font(bold=True, color="1F4E78")

    # Hlavička formuláře
    info = [
        ("Číslo případu", p.get("cislo")),
        ("Stav", _STAV_BADGE.get(p.get("stav"), (p.get("stav"),))[0]),
        ("Důvod změny", p.get("duvod_zmeny")),
        ("Jméno ASM", p.get("asm_jmeno")),
        ("Změna ASM jméno", p.get("zmena_asm_jmeno")),
        ("Nový ASM jméno", p.get("novy_asm_jmeno")),
        ("Číslo OZ", p.get("cislo_oz")), ("Jméno OZ", p.get("jmeno_oz")),
        ("Číslo nový OZ", p.get("cislo_novy_oz")), ("Jméno nový OZ", p.get("jmeno_novy_oz")),
        ("Datum změny od", p.get("datum_zmeny_od")),
        ("Zákazník v regionu OZ", p.get("zakaznik_v_regionu_oz")),
        ("Případné odůvodnění", p.get("oduvodneni")),
        ("Žadatel", p.get("zadavatel_jmeno")),
        ("Datum zadání", _dt_cz(p.get("datum_zadani"))),
        ("Poznámka", p.get("poznamka")),
        ("Důvod zamítnutí", p.get("zamitnuti_duvod")),
    ]
    r = 1
    for popisek, hodnota in info:
        ws.cell(row=r, column=1, value=popisek).font = bold
        ws.cell(row=r, column=2, value=(hodnota if hodnota not in (None, "") else "—"))
        r += 1

    # Tabulka řádků
    r += 1
    sloupce = (["#", "IČO"]
               + [lbl for _, lbl in _KONTAKT_SLOUPCE]
               + ["Původní provize", "Nová provize", "OZ2", "Datum od", "Datum do"])
    klice = (["poradi", "ico"]
             + [key for key, _ in _KONTAKT_SLOUPCE]
             + ["puvodni_provize", "nova_provize", "oz2", "datum_od", "datum_do"])
    hlav_radek = r
    for c, lbl in enumerate(sloupce, 1):
        cell = ws.cell(row=r, column=c, value=lbl)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
    ico_col = klice.index("ico") + 1
    for rd in radky:
        r += 1
        for c, key in enumerate(klice, 1):
            val = rd.get(key, "")
            cell = ws.cell(row=r, column=c, value=val)
            if c == ico_col:
                cell.number_format = "@"   # IČO jako text — drží počáteční nuly
                cell.value = "" if val is None else str(val)

    # Šířky sloupců
    for c, lbl in enumerate(sloupce, 1):
        ws.column_dimensions[ws.cell(row=hlav_radek, column=c).column_letter].width = max(12, min(40, len(str(lbl)) + 4))
    ws.column_dimensions["A"].width = 24

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nazev = f"{p.get('cislo') or 'formular_asm'}.xlsx"
    return buf.read(), nazev
