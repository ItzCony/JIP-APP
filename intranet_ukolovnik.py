from nicegui import ui, app, background_tasks, events as ui_events
import asyncio
import intranet_data
import intranet_static
import intranet_logger
import intranet_notifikace
import datetime
import os
import uuid
import calendar
import tempfile
import time
from collections import OrderedDict

UKOL_PRILOHY_DIR = 'ukol_prilohy'
PROJEKT_PRILOHY_DIR = 'projekt_prilohy'
os.makedirs(UKOL_PRILOHY_DIR, exist_ok=True)
os.makedirs(PROJEKT_PRILOHY_DIR, exist_ok=True)
intranet_static.chranene_soubory('/ukol_prilohy', UKOL_PRILOHY_DIR)
intranet_static.chranene_soubory('/projekt_prilohy', PROJEKT_PRILOHY_DIR)

# =========================================================
# KONSTANTY
# =========================================================
STAVY_UKOL = {
    'Zadáno':       ('bg-red-100 text-red-800 border-red-300',       '#ef4444'),
    'Rozpracováno': ('bg-orange-100 text-orange-800 border-orange-300', '#f97316'),
    'Pozastaveno':  ('bg-yellow-100 text-yellow-800 border-yellow-300', '#eab308'),
    'Hotovo':       ('bg-green-100 text-green-800 border-green-300',   '#22c55e'),
    'Zrušen':       ('bg-gray-200 text-gray-600 border-gray-400',      '#9ca3af'),
}

PRIORITY_BARVY = {
    'Nízká':    'text-gray-500',
    'Normální': 'text-blue-600',
    'Vysoká':   'text-orange-600',
    'Kritická': 'text-red-700 font-extrabold',
}

BARVY_STAVU_PORADA = {
    'Plánovaná': 'bg-blue-100 text-blue-800 border-blue-300',
    'Probíhá':   'bg-orange-100 text-orange-800 border-orange-300',
    'Dokončená': 'bg-green-100 text-green-800 border-green-300',
    'Zrušená':   'bg-gray-200 text-gray-600 border-gray-400',
}

MESICE_CZ = ['', 'Leden', 'Únor', 'Březen', 'Duben', 'Květen', 'Červen',
             'Červenec', 'Srpen', 'Září', 'Říjen', 'Listopad', 'Prosinec']
DNY_CZ = ['Po', 'Út', 'St', 'Čt', 'Pá', 'So', 'Ne']
DNY_CZ_PLNE = ['Pondělí', 'Úterý', 'Středa', 'Čtvrtek', 'Pátek', 'Sobota', 'Neděle']

# Týdenní osa kalendáře
KAL_HODINA_PX = 52          # výška jedné hodiny v týdenní ose
KAL_HOD_OD, KAL_HOD_DO = 6, 19   # výchozí rozsah osy (rozšíří se podle záznamů)
KAL_BARVY_OSOB = ['#2563eb', '#0ea5e9', '#a855f7', '#059669',
                  '#f59e0b', '#e11d48', '#7c3aed', '#0891b2']

# =========================================================
# INICIALIZACE DB
# =========================================================
def inicializace_ukolovnik_db():
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor(dictionary=True)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS ukolovnik_porady (
                id INT AUTO_INCREMENT PRIMARY KEY,
                nazev VARCHAR(255) NOT NULL,
                popis TEXT,
                datum DATE NOT NULL,
                cas_od TIME, cas_do TIME,
                misto VARCHAR(255),
                moderator_id INT, moderator_jmeno VARCHAR(255),
                zapisovatel_id INT, zapisovatel_jmeno VARCHAR(255),
                stav VARCHAR(50) DEFAULT 'Plánovaná',
                vytvoreno DATETIME DEFAULT CURRENT_TIMESTAMP,
                vytvoril_id INT, vytvoril_jmeno VARCHAR(255)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS ukolovnik_porada_ucastnici (
                id INT AUTO_INCREMENT PRIMARY KEY,
                porada_id INT NOT NULL, user_id INT NOT NULL, jmeno VARCHAR(255),
                UNIQUE KEY uq_porada_user (porada_id, user_id)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS ukolovnik_porada_poznamky (
                id INT AUTO_INCREMENT PRIMARY KEY,
                porada_id INT NOT NULL, user_id INT, jmeno_autora VARCHAR(255),
                text TEXT, vytvoreno DATETIME DEFAULT CURRENT_TIMESTAMP
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS ukolovnik_ukoly (
                id INT AUTO_INCREMENT PRIMARY KEY,
                nazev VARCHAR(255) NOT NULL,
                popis TEXT,
                porada_id INT DEFAULT NULL,
                porada_nazev VARCHAR(255),
                porada_datum DATE,
                prirazen_id INT, prirazen_jmeno VARCHAR(255),
                zadal_id INT, zadal_jmeno VARCHAR(255),
                termin DATE,
                odhad_hodin FLOAT DEFAULT NULL,
                priorita VARCHAR(50) DEFAULT 'Normální',
                stav VARCHAR(50) DEFAULT 'Zadáno',
                vytvoreno DATETIME DEFAULT CURRENT_TIMESTAMP
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)

        # Migrace: stará jména stavů → nová
        cur.execute("UPDATE ukolovnik_ukoly SET stav='Zadáno' WHERE stav='Nový'")
        cur.execute("UPDATE ukolovnik_ukoly SET stav='Rozpracováno' WHERE stav='Probíhá'")

        # Přidej odhad_hodin pokud chybí
        cur.execute("SHOW COLUMNS FROM ukolovnik_ukoly LIKE 'odhad_hodin'")
        if not cur.fetchone():
            cur.execute("ALTER TABLE ukolovnik_ukoly ADD COLUMN odhad_hodin FLOAT DEFAULT NULL")

        cur.execute("""
            CREATE TABLE IF NOT EXISTS ukolovnik_ukol_poznamky (
                id INT AUTO_INCREMENT PRIMARY KEY,
                ukol_id INT NOT NULL, user_id INT, jmeno_autora VARCHAR(255),
                text TEXT, vytvoreno DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_ukol (ukol_id)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS ukolovnik_ukol_prilohy (
                id VARCHAR(64) PRIMARY KEY,
                ukol_id INT NOT NULL, user_id INT, jmeno_autora VARCHAR(255),
                soubor_nazev VARCHAR(255), soubor_cesta VARCHAR(512),
                vytvoreno DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_ukol (ukol_id)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS ukolovnik_ukol_plany (
                id INT AUTO_INCREMENT PRIMARY KEY,
                ukol_id INT NOT NULL, popis VARCHAR(500),
                hotovo TINYINT(1) DEFAULT 0,
                termin DATE DEFAULT NULL,
                poradi INT DEFAULT 0,
                INDEX idx_ukol (ukol_id)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS ukolovnik_cas_zaznamy (
                id INT AUTO_INCREMENT PRIMARY KEY,
                ukol_id INT NOT NULL, user_id INT, jmeno VARCHAR(255),
                cas_start DATETIME NOT NULL,
                cas_konec DATETIME DEFAULT NULL,
                trvani_minut INT DEFAULT NULL,
                INDEX idx_ukol (ukol_id),
                INDEX idx_user (user_id)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)

        # ── PROJEKTY ──────────────────────────────────────────────────────────
        cur.execute("""
            CREATE TABLE IF NOT EXISTS ukolovnik_projekty (
                id INT AUTO_INCREMENT PRIMARY KEY,
                nazev VARCHAR(255) NOT NULL,
                popis TEXT,
                termin DATE,
                stav VARCHAR(50) DEFAULT 'Aktivní',
                vytvoril_id INT, vytvoril_jmeno VARCHAR(255),
                vytvoreno DATETIME DEFAULT CURRENT_TIMESTAMP,
                prevedeno_na_ukol TINYINT(1) DEFAULT 0,
                ukol_id INT DEFAULT NULL
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS ukolovnik_projekt_clenove (
                id INT AUTO_INCREMENT PRIMARY KEY,
                projekt_id INT NOT NULL, user_id INT NOT NULL, jmeno VARCHAR(255),
                role VARCHAR(50) DEFAULT 'spolupracovnik',
                UNIQUE KEY uq_pj_user (projekt_id, user_id)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS ukolovnik_projekt_komentare (
                id INT AUTO_INCREMENT PRIMARY KEY,
                projekt_id INT NOT NULL, user_id INT, jmeno_autora VARCHAR(255),
                text TEXT, vytvoreno DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_pj (projekt_id)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS ukolovnik_projekt_prilohy (
                id VARCHAR(64) PRIMARY KEY,
                projekt_id INT NOT NULL, user_id INT, jmeno_autora VARCHAR(255),
                soubor_nazev VARCHAR(255), soubor_cesta VARCHAR(512),
                vytvoreno DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_pj (projekt_id)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)

        # Migrace: nové sloupce pro úkoly
        for col, defn in [
            ('oddeleni',      "VARCHAR(500) DEFAULT NULL"),
            ('typ_ukolu',     "VARCHAR(20) DEFAULT 'Pravidelný'"),
            ('dokonceno_datum', "DATETIME DEFAULT NULL"),
            ('zobrazit_od',   "DATE DEFAULT NULL"),
            ('opakovani_id',  "VARCHAR(64) DEFAULT NULL"),
        ]:
            cur.execute(f"SHOW COLUMNS FROM ukolovnik_ukoly LIKE '{col}'")
            if not cur.fetchone():
                cur.execute(f"ALTER TABLE ukolovnik_ukoly ADD COLUMN {col} {defn}")

        # Migrace: oddělení pro projekty
        cur.execute("SHOW COLUMNS FROM ukolovnik_projekty LIKE 'oddeleni'")
        if not cur.fetchone():
            cur.execute("ALTER TABLE ukolovnik_projekty ADD COLUMN oddeleni VARCHAR(500) DEFAULT NULL")

        # Přidej projekt_id / projekt_nazev do ukoly pokud chybí
        cur.execute("SHOW COLUMNS FROM ukolovnik_ukoly LIKE 'projekt_id'")
        if not cur.fetchone():
            cur.execute("ALTER TABLE ukolovnik_ukoly ADD COLUMN projekt_id INT DEFAULT NULL")
            cur.execute("ALTER TABLE ukolovnik_ukoly ADD COLUMN projekt_nazev VARCHAR(255) DEFAULT NULL")

        # Individuální ruční pořadí úkolů (drag&drop v rámci dne) — per uživatel
        cur.execute("""
            CREATE TABLE IF NOT EXISTS ukolovnik_poradi (
                user_id INT NOT NULL,
                ukol_id INT NOT NULL,
                poradi  INT NOT NULL DEFAULT 0,
                PRIMARY KEY (user_id, ukol_id)
            )
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS ukolovnik_log (
                id INT AUTO_INCREMENT PRIMARY KEY,
                typ VARCHAR(20) NOT NULL,
                ref_id INT NOT NULL,
                user_id INT,
                user_name VARCHAR(255),
                akce VARCHAR(500) NOT NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_typ_ref (typ, ref_id),
                INDEX idx_created (created_at)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS ukolovnik_sledovaci (
                id INT AUTO_INCREMENT PRIMARY KEY,
                typ VARCHAR(20) NOT NULL,
                ref_id INT NOT NULL,
                user_id INT NOT NULL,
                user_name VARCHAR(255),
                added_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                UNIQUE KEY uq_sled (typ, ref_id, user_id),
                INDEX idx_user (user_id)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)

        conn.commit()
        cur.close()
    except Exception as e:
        print(f"Chyba při inicializaci DB Úkolovníku: {e}")
    finally:
        conn.close()


# =========================================================
# POMOCNÉ FUNKCE
# =========================================================
def _ziskej_oddeleni_options() -> dict:
    """Vrátí {id_name: nazev} pro všechna oddělení, seřazeno."""
    try:
        vsechna = intranet_data.ziskej_vsechna_oddeleni()
        return {k: k for k in sorted(vsechna.keys())}
    except Exception:
        return {}


def _oddeleni_chips(oddeleni_str: str, max_chips: int = 3) -> list:
    """Rozdělí csv string na list oddělení."""
    if not oddeleni_str:
        return []
    return [o.strip() for o in oddeleni_str.split(',') if o.strip()][:max_chips]


def _ziskej_uzivatele_options():
    vsichni = intranet_data.ziskej_vsechny_uzivatele()
    opts = {}
    for _, u in vsichni.items():
        if u.get('aktivni', True):
            jmeno = f"{u.get('jmeno', '')} {u.get('prijmeni', '')}".strip()
            opts[u['id']] = jmeno
    return dict(sorted(opts.items(), key=lambda x: x[1]))


def _ziskej_viditelne_uzivatele(user_id: int, vsechna_prava) -> dict:
    """
    Vrátí {id: jmeno} uživatelů viditelných pro daného uživatele.
    Admin vidí všechny; hlavní vedoucí oddělení vidí jen své oddělení; ostatní vidí všechny.
    """
    is_admin = 'vse' in vsechna_prava or 'ukolovnik_admin' in vsechna_prava
    if not is_admin and _je_hlavni_vedouci(vsechna_prava):
        dept = _ziskej_uzivatele_oddeleni(vsechna_prava)
        if dept:
            return {u['id']: u['jmeno'] for u in dept}
    return _ziskej_uzivatele_options()


def _format_odhad(odhad_hodin):
    """Naformátuje odhad uložený v hodinách na lidsky čitelný text v hodinách/minutách."""
    if not odhad_hodin:
        return '—'
    celkem_min = int(round(float(odhad_hodin) * 60))
    if celkem_min <= 0:
        return '—'
    h, m = divmod(celkem_min, 60)
    if h and m:
        return f'{h} h {m} min'
    if h:
        return f'{h} h'
    return f'{m} min'


def _dny_do_terminu(termin):
    """Vrátí (delta_dní, css_třídy, text). Kladné = zbývá, záporné = po termínu."""
    if not termin:
        return None, 'text-gray-400', '—'
    dnes = datetime.date.today()
    delta = (termin - dnes).days
    if delta > 7:
        cls = 'text-green-600 font-bold'
        txt = f'{delta} dní'
    elif delta > 3:
        cls = 'text-orange-500 font-bold'
        txt = f'{delta} dní'
    elif delta >= 1:
        cls = 'text-red-500 font-bold'
        txt = f'{delta} dní'
    elif delta == 0:
        cls = 'text-red-700 font-extrabold'
        txt = 'Dnes!'
    else:
        cls = 'text-red-800 font-extrabold'
        txt = f'{abs(delta)} dní po termínu'
    return delta, cls, txt


def _notifikuj_ucastniky_ukolu(ukol_id, vyjma_user_id, zprava, typ='info'):
    """Pošle notifikaci všem relevantním osobám úkolu (přiřazenému + zadavateli + sledovačům)."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT prirazen_id, zadal_id FROM ukolovnik_ukoly WHERE id = %s", (ukol_id,))
        row = cur.fetchone()
        cur.close()
    finally:
        conn.close()
    if not row:
        return
    adresati = set()
    if row.get('prirazen_id'):
        adresati.add(row['prirazen_id'])
    if row.get('zadal_id'):
        adresati.add(row['zadal_id'])
    adresati.update(_ziskej_sledovace_ids('ukol', ukol_id))
    adresati.discard(vyjma_user_id)
    for uid in adresati:
        try:
            intranet_notifikace.pridej(uid, zprava, typ)
        except Exception:
            pass


def _notifikuj_ucastniky_porady(porada_id, vyjma_user_id, zprava, typ='info'):
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT user_id FROM ukolovnik_porada_ucastnici WHERE porada_id = %s", (porada_id,))
        radky = cur.fetchall()
        cur.execute("SELECT moderator_id, vytvoril_id FROM ukolovnik_porady WHERE id = %s", (porada_id,))
        meta = cur.fetchone()
        cur.close()
    finally:
        conn.close()
    adresati = set(r['user_id'] for r in radky)
    if meta:
        if meta.get('moderator_id'):
            adresati.add(meta['moderator_id'])
        if meta.get('vytvoril_id'):
            adresati.add(meta['vytvoril_id'])
    adresati.update(_ziskej_sledovace_ids('porada', porada_id))
    adresati.discard(vyjma_user_id)
    for uid in adresati:
        try:
            intranet_notifikace.pridej(uid, zprava, typ)
        except Exception:
            pass


# =========================================================
# AKTIVITNÍ LOG + SLEDOVÁNÍ
# =========================================================
# Živé per-úkol překreslení karet: client.id → async handler (registruje _vykresli_ukoly)
_UK_ZIVE_RADKY: dict = {}


def _bump_data_verze(typ=None, ref_id=None):
    """Zvýší globální revizi dat úkolovníku → ostatní připojení klienti se při
    nejbližším tiknutí timeru živě překreslí (viz timer ve `vykresli_ukolovnik`).

    Je-li známo, čeho se změna týkala (typ + ref_id), poznačí ji do kapované
    fronty `ukolovnik_zmeny` — klienti si podle ní živě obnoví i právě otevřené
    detaily (náhled úkolu se překreslí s čerstvými daty).

    Aktérovi zároveň poznačí jeho vlastní verzi do `app.storage.user`, aby se mu
    view nepřekreslilo podruhé — svůj refresh už dostal přes `on_refresh`.
    """
    try:
        nova = app.storage.general.get('ukolovnik_data_verze', 0) + 1
        app.storage.general['ukolovnik_data_verze'] = nova
        if typ is not None and ref_id is not None:
            try:
                zmeny = list(app.storage.general.get('ukolovnik_zmeny', []))[-49:]
                zmeny.append({'v': nova, 'typ': str(typ), 'ref_id': int(ref_id)})
                app.storage.general['ukolovnik_zmeny'] = zmeny
            except Exception:
                pass
        try:
            app.storage.user['ukolovnik_data_verze_vlastni'] = nova
        except Exception:
            pass
    except Exception:
        pass


def _log(typ: str, ref_id: int, user_id, user_name: str, akce: str):
    conn = intranet_data.get_db_connection()
    if not conn: return
    try:
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO ukolovnik_log (typ, ref_id, user_id, user_name, akce) VALUES (%s,%s,%s,%s,%s)",
            (typ, int(ref_id), user_id, str(user_name)[:255], str(akce)[:500])
        )
        conn.commit()
        # Po každé zalogované mutaci posuň revizi → ostatní klienti se živě překreslí.
        _bump_data_verze(typ, ref_id)
    except Exception as e:
        print(f'[Log] {e}')
    finally:
        cur.close(); conn.close()


def _ziskej_log(typ: str, ref_id: int, limit: int = 150) -> list:
    conn = intranet_data.get_db_connection()
    if not conn: return []
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(
            "SELECT * FROM ukolovnik_log WHERE typ=%s AND ref_id=%s ORDER BY created_at DESC LIMIT %s",
            (typ, int(ref_id), int(limit))
        )
        return cur.fetchall()
    except Exception:
        return []
    finally:
        cur.close(); conn.close()


def _ziskej_sledovace_ids(typ: str, ref_id: int) -> set:
    conn = intranet_data.get_db_connection()
    if not conn: return set()
    try:
        cur = conn.cursor()
        cur.execute("SELECT user_id FROM ukolovnik_sledovaci WHERE typ=%s AND ref_id=%s", (typ, int(ref_id)))
        return {r[0] for r in cur.fetchall()}
    except Exception:
        return set()
    finally:
        cur.close(); conn.close()


def _je_sledovac(typ: str, ref_id: int, user_id) -> bool:
    conn = intranet_data.get_db_connection()
    if not conn: return False
    try:
        cur = conn.cursor()
        cur.execute("SELECT id FROM ukolovnik_sledovaci WHERE typ=%s AND ref_id=%s AND user_id=%s",
                    (typ, int(ref_id), int(user_id)))
        return bool(cur.fetchone())
    except Exception:
        return False
    finally:
        cur.close(); conn.close()


def _toggle_sledovat(typ: str, ref_id: int, user_id, user_name: str) -> bool:
    """Přidá nebo odebere sledovače. Vrátí True pokud nyní sleduje."""
    conn = intranet_data.get_db_connection()
    if not conn: return False
    try:
        cur = conn.cursor()
        cur.execute("SELECT id FROM ukolovnik_sledovaci WHERE typ=%s AND ref_id=%s AND user_id=%s",
                    (typ, int(ref_id), int(user_id)))
        if cur.fetchone():
            cur.execute("DELETE FROM ukolovnik_sledovaci WHERE typ=%s AND ref_id=%s AND user_id=%s",
                        (typ, int(ref_id), int(user_id)))
            conn.commit()
            return False
        else:
            cur.execute("INSERT IGNORE INTO ukolovnik_sledovaci (typ, ref_id, user_id, user_name) VALUES (%s,%s,%s,%s)",
                        (typ, int(ref_id), int(user_id), str(user_name)))
            conn.commit()
            _log(typ, ref_id, user_id, user_name, 'Přidal/a se jako sledovač')
            return True
    except Exception:
        return False
    finally:
        cur.close(); conn.close()


def _dialog_log(typ: str, ref_id: int, nazev: str, current_user_id, current_user_name: str):
    """Zobrazí modal s timeline logem a tlačítkem Sledovat."""
    zaznamy = _ziskej_log(typ, ref_id)

    IKONY_AKCI = [
        ('Vytvořil', '🆕'), ('Zahájil', '▶️'), ('Pozastavil', '⏸️'), ('Dokončil', '✅'),
        ('Splnil', '☑️'), ('Převedl', '🔀'), ('Upravil', '✏️'), ('Nahrál', '📎'),
        ('Odeslal', '💬'), ('Změnil', '🔄'), ('Přidal', '➕'), ('Sledovač', '👁'),
    ]

    def _ikona_akce(akce_text: str) -> str:
        for klic, ikona in IKONY_AKCI:
            if klic in akce_text:
                return ikona
        return '📋'

    typ_lbl = {'ukol': 'Úkol', 'projekt': 'Projekt', 'porada': 'Porada'}.get(typ, typ)

    with _dialog_kotva(), ui.dialog() as dlg, ui.card().classes('w-full max-w-lg p-0 rounded-xl overflow-hidden'):

        with ui.row().classes('w-full items-center justify-between px-5 py-3 bg-gray-800 text-white shrink-0'):
            with ui.column().classes('gap-0'):
                ui.label('Historie aktivit').classes('text-base font-extrabold')
                ui.label(f'{typ_lbl}: {nazev[:55]}').classes('text-xs text-gray-300')
            ui.button(icon='close', on_click=dlg.close).props('flat round dense').classes('text-white')

        sleduje_stav = [_je_sledovac(typ, ref_id, current_user_id)]
        btn_container = ui.row().classes('w-full items-center justify-between px-5 py-2 bg-gray-50 border-b border-gray-200 shrink-0')

        def _render_sled_btn():
            btn_container.clear()
            with btn_container:
                ui.label(f'{len(zaznamy)} záznamů  •  {len(_ziskej_sledovace_ids(typ, ref_id))} sledovatelů'
                         ).classes('text-xs text-gray-500')
                def _toggle():
                    nova = _toggle_sledovat(typ, ref_id, current_user_id, current_user_name)
                    sleduje_stav[0] = nova
                    ui.notify('Sledujete!' if nova else 'Sledování zrušeno.',
                              type='positive' if nova else 'info', position='top-right', timeout=1500)
                    _render_sled_btn()
                lbl = '👁 Sleduji' if sleduje_stav[0] else '👁 Sledovat'
                cls = 'bg-blue-600 text-white font-bold px-4 text-xs h-8' if sleduje_stav[0] else 'bg-gray-200 text-gray-700 font-bold px-4 text-xs h-8'
                ui.button(lbl, on_click=_toggle).classes(cls)

        _render_sled_btn()

        with ui.scroll_area().classes('w-full').style('max-height:520px'):
            with ui.column().classes('w-full px-4 py-3 gap-0'):
                if not zaznamy:
                    ui.label('Zatím žádná aktivita.').classes('text-gray-400 italic text-sm text-center py-10')
                else:
                    for i, z in enumerate(zaznamy):
                        je_posledni = (i == len(zaznamy) - 1)
                        ikona = _ikona_akce(z['akce'])
                        dt = z['created_at'].strftime('%d.%m.%Y %H:%M')
                        with ui.row().classes('w-full items-start gap-3'):
                            with ui.column().classes('items-center shrink-0').style('width:28px'):
                                with ui.element('div').classes(
                                    'w-7 h-7 rounded-full bg-gray-100 border-2 border-gray-300 '
                                    'flex items-center justify-center shrink-0'
                                ):
                                    ui.label(ikona).style('font-size:13px;line-height:1')
                                if not je_posledni:
                                    ui.element('div').classes('w-0.5 bg-gray-200').style('height:24px')
                            with ui.column().classes('flex-1 pb-3 gap-0 min-w-0'):
                                ui.label(z['akce']).classes('text-sm font-bold text-gray-800 leading-tight')
                                with ui.row().classes('gap-2 items-center mt-0.5'):
                                    ui.label(z.get('user_name') or '—').classes('text-xs text-blue-600 font-medium')
                                    ui.label(dt).classes('text-[10px] text-gray-400')

    dlg.open()


def _ikona_souboru(nazev):
    ext = os.path.splitext(nazev)[1].lower()
    if ext == '.pdf':
        return 'picture_as_pdf', 'text-red-500'
    if ext in ('.xlsx', '.xls', '.csv'):
        return 'table_chart', 'text-green-600'
    if ext in ('.jpg', '.jpeg', '.png', '.gif', '.webp'):
        return 'image', 'text-blue-500'
    if ext in ('.doc', '.docx'):
        return 'description', 'text-blue-700'
    return 'attach_file', 'text-gray-500'


def _ziskej_ucastniky_porady(porada_id):
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT * FROM ukolovnik_porada_ucastnici WHERE porada_id = %s", (porada_id,))
        return cur.fetchall()
    finally:
        cur.close(); conn.close()


def _ziskej_ukoly_porady(porada_id):
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT * FROM ukolovnik_ukoly WHERE porada_id = %s ORDER BY vytvoreno ASC", (porada_id,))
        return cur.fetchall()
    finally:
        cur.close(); conn.close()


# =========================================================
# TIMER: ZAHÁJIT / POZASTAVIT / DOKONČIT
# =========================================================
def _ziskej_aktivni_zaznam(ukol_id, user_id):
    """Vrátí otevřený časový záznam nebo None."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return None
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(
            "SELECT * FROM ukolovnik_cas_zaznamy WHERE ukol_id=%s AND user_id=%s AND cas_konec IS NULL ORDER BY cas_start DESC LIMIT 1",
            (ukol_id, user_id)
        )
        return cur.fetchone()
    finally:
        cur.close(); conn.close()


def _ziskej_aktivni_ukol_ids(user_id) -> set:
    """IDs úkolů, u kterých má uživatel právě běžící časový záznam.

    Jeden dotaz pro celý seznam úkolů — náhrada za dřívější dotaz per řádek."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return set()
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT DISTINCT ukol_id FROM ukolovnik_cas_zaznamy WHERE user_id=%s AND cas_konec IS NULL",
            (user_id,)
        )
        return {r[0] for r in cur.fetchall()}
    finally:
        cur.close(); conn.close()


def _zahaj_ukol(ukol_id, user_id, user_name):
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO ukolovnik_cas_zaznamy (ukol_id, user_id, jmeno, cas_start) VALUES (%s,%s,%s,%s)",
            (ukol_id, user_id, user_name, datetime.datetime.now())
        )
        cur.execute("UPDATE ukolovnik_ukoly SET stav='Rozpracováno' WHERE id=%s", (ukol_id,))
        conn.commit()
    finally:
        cur.close(); conn.close()
    intranet_logger.log_activity(user_name, "Úkolovník", f"Zahájen úkol #{ukol_id}")
    _log('ukol', ukol_id, user_id, user_name, 'Zahájil/a práci')
    _notifikuj_ucastniky_ukolu(ukol_id, user_id, f"⏱️ {user_name} zahájil/a zpracování úkolu.", 'info')


def _pozastav_ukol(ukol_id, user_id, user_name):
    now = datetime.datetime.now()
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(
            "SELECT id, cas_start FROM ukolovnik_cas_zaznamy WHERE ukol_id=%s AND user_id=%s AND cas_konec IS NULL ORDER BY cas_start DESC LIMIT 1",
            (ukol_id, user_id)
        )
        zaznam = cur.fetchone()
        if zaznam:
            delta_min = max(1, int((now - zaznam['cas_start']).total_seconds() / 60))
            cur.execute(
                "UPDATE ukolovnik_cas_zaznamy SET cas_konec=%s, trvani_minut=%s WHERE id=%s",
                (now, delta_min, zaznam['id'])
            )
        cur.execute("UPDATE ukolovnik_ukoly SET stav='Pozastaveno' WHERE id=%s", (ukol_id,))
        conn.commit()
    finally:
        cur.close(); conn.close()
    intranet_logger.log_activity(user_name, "Úkolovník", f"Pozastaven úkol #{ukol_id}")
    _log('ukol', ukol_id, user_id, user_name, 'Pozastavil/a práci')
    _notifikuj_ucastniky_ukolu(ukol_id, user_id, f"⏸️ {user_name} pozastavil/a zpracování úkolu.", 'info')


def _dokoncit_ukol(ukol_id, user_id, user_name):
    now = datetime.datetime.now()
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(
            "SELECT id, cas_start FROM ukolovnik_cas_zaznamy WHERE ukol_id=%s AND user_id=%s AND cas_konec IS NULL ORDER BY cas_start DESC LIMIT 1",
            (ukol_id, user_id)
        )
        zaznam = cur.fetchone()
        if zaznam:
            delta_min = max(1, int((now - zaznam['cas_start']).total_seconds() / 60))
            cur.execute(
                "UPDATE ukolovnik_cas_zaznamy SET cas_konec=%s, trvani_minut=%s WHERE id=%s",
                (now, delta_min, zaznam['id'])
            )
        cur.execute("UPDATE ukolovnik_ukoly SET stav='Hotovo', dokonceno_datum=%s WHERE id=%s", (now, ukol_id))
        conn.commit()
    finally:
        cur.close(); conn.close()
    intranet_logger.log_activity(user_name, "Úkolovník", f"Dokončen úkol #{ukol_id}")
    _log('ukol', ukol_id, user_id, user_name, 'Dokončil/a práci a označil/a úkol za hotovo')
    _notifikuj_ucastniky_ukolu(ukol_id, user_id, f"✅ {user_name} dokončil/a úkol!", 'success')


# =========================================================
# SMAZÁNÍ — pomocné funkce
# =========================================================

def _potvrdit_smazat(zprava, on_potvrdit):
    """Otevře malý potvrzovací dialog. on_potvrdit se zavolá při potvrzení."""
    with ui.dialog() as dlg, ui.card().classes('p-6 rounded-xl max-w-sm w-full'):
        ui.label(zprava).classes('text-gray-800 font-bold mb-4 text-center')
        with ui.row().classes('w-full justify-center gap-4'):
            ui.button('Smazat', icon='delete', on_click=lambda: [dlg.close(), on_potvrdit()]).classes('bg-red-600 hover:bg-red-700 text-white font-bold px-6')
            ui.button('Zrušit', on_click=dlg.close).classes('bg-gray-200 hover:bg-gray-300 text-gray-700 font-bold px-6')
    dlg.open()


def _smazat_ukol_db(ukol_id):
    """Cascade delete úkolu ze všech tabulek. Vrátí True při úspěchu."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return False
    try:
        cur = conn.cursor()
        cur.execute("SELECT soubor_cesta FROM ukolovnik_ukol_prilohy WHERE ukol_id=%s", (ukol_id,))
        prilohy_cesty = [r[0] for r in cur.fetchall()]
        cur.execute("DELETE FROM ukolovnik_ukol_poznamky WHERE ukol_id=%s", (ukol_id,))
        cur.execute("DELETE FROM ukolovnik_ukol_prilohy WHERE ukol_id=%s", (ukol_id,))
        cur.execute("DELETE FROM ukolovnik_ukol_plany WHERE ukol_id=%s", (ukol_id,))
        cur.execute("DELETE FROM ukolovnik_cas_zaznamy WHERE ukol_id=%s", (ukol_id,))
        cur.execute("DELETE FROM ukolovnik_sledovaci WHERE typ='ukol' AND ref_id=%s", (ukol_id,))
        cur.execute("DELETE FROM ukolovnik_log WHERE typ='ukol' AND ref_id=%s", (ukol_id,))
        cur.execute("DELETE FROM ukolovnik_poradi WHERE ukol_id=%s", (ukol_id,))
        cur.execute("DELETE FROM ukolovnik_ukoly WHERE id=%s", (ukol_id,))
        conn.commit()
        _bump_data_verze('ukol', ukol_id)  # živě promítnout smazání ostatním klientům
        for cesta in prilohy_cesty:
            try:
                if cesta and os.path.isfile(cesta):
                    os.remove(cesta)
            except Exception:
                pass
        return True
    except Exception as e:
        print(f"[_smazat_ukol_db] {e}")
        return False
    finally:
        cur.close(); conn.close()


def _smazat_poradu_db(porada_id):
    """Cascade delete porady (včetně všech úkolů, zápisu, účastníků, logu)."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return False
    try:
        cur = conn.cursor()
        cur.execute("SELECT id FROM ukolovnik_ukoly WHERE porada_id=%s", (porada_id,))
        ukol_ids = [r[0] for r in cur.fetchall()]
        cur.close(); conn.close()
        for uid in ukol_ids:
            _smazat_ukol_db(uid)
        conn2 = intranet_data.get_db_connection()
        if not conn2:
            return False
        cur2 = conn2.cursor()
        cur2.execute("DELETE FROM ukolovnik_porada_poznamky WHERE porada_id=%s", (porada_id,))
        cur2.execute("DELETE FROM ukolovnik_porada_ucastnici WHERE porada_id=%s", (porada_id,))
        cur2.execute("DELETE FROM ukolovnik_sledovaci WHERE typ='porada' AND ref_id=%s", (porada_id,))
        cur2.execute("DELETE FROM ukolovnik_log WHERE typ='porada' AND ref_id=%s", (porada_id,))
        cur2.execute("DELETE FROM ukolovnik_porady WHERE id=%s", (porada_id,))
        conn2.commit()
        _bump_data_verze('porada', porada_id)  # živě promítnout smazání ostatním klientům
        return True
    except Exception as e:
        print(f"[_smazat_poradu_db] {e}")
        return False
    finally:
        try: cur2.close(); conn2.close()
        except Exception: pass


def _smazat_projekt_db(projekt_id):
    """Cascade delete projektu (členové, komentáře, přílohy, sledovači, log).
    Úkoly vytvořené z projektu zůstávají, jen se odpojí (projekt_id → NULL)."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return False
    try:
        cur = conn.cursor()
        cur.execute("SELECT soubor_cesta FROM ukolovnik_projekt_prilohy WHERE projekt_id=%s", (projekt_id,))
        prilohy_cesty = [r[0] for r in cur.fetchall()]
        cur.execute("DELETE FROM ukolovnik_projekt_komentare WHERE projekt_id=%s", (projekt_id,))
        cur.execute("DELETE FROM ukolovnik_projekt_prilohy WHERE projekt_id=%s", (projekt_id,))
        cur.execute("DELETE FROM ukolovnik_projekt_clenove WHERE projekt_id=%s", (projekt_id,))
        cur.execute("DELETE FROM ukolovnik_sledovaci WHERE typ='projekt' AND ref_id=%s", (projekt_id,))
        cur.execute("DELETE FROM ukolovnik_log WHERE typ='projekt' AND ref_id=%s", (projekt_id,))
        cur.execute("UPDATE ukolovnik_ukoly SET projekt_id=NULL WHERE projekt_id=%s", (projekt_id,))
        cur.execute("DELETE FROM ukolovnik_projekty WHERE id=%s", (projekt_id,))
        conn.commit()
        for cesta in prilohy_cesty:
            try:
                if cesta and os.path.isfile(cesta):
                    os.remove(cesta)
            except Exception:
                pass
        _bump_data_verze('projekt', projekt_id)  # živě promítnout smazání ostatním klientům
        return True
    except Exception as e:
        print(f"[_smazat_projekt_db] {e}")
        return False
    finally:
        cur.close(); conn.close()


# =========================================================
# DIALOG: DETAIL ÚKOLU — plnohodnotný
# =========================================================
def _dialog_detail_ukolu(ukol_id, user_id, user_name, vsechna_prava=None, on_refresh=None, dialog_anchor=None):
    if vsechna_prava is None:
        vsechna_prava = []

    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT * FROM ukolovnik_ukoly WHERE id = %s", (ukol_id,))
        uk = cur.fetchone()
    finally:
        cur.close(); conn.close()

    if not uk:
        ui.notify('Úkol nenalezen.', type='negative')
        return

    is_admin = vsechna_prava and ('vse' in vsechna_prava or 'ukolovnik_admin' in vsechna_prava)
    muze_editovat = is_admin or uk.get('prirazen_id') == user_id or uk.get('zadal_id') == user_id
    muze_editovat_meta = is_admin or uk.get('zadal_id') == user_id

    _navrhy = {}   # rozepsané texty přežijí živé překreslení obsahu
    with _dialog_kotva(dialog_anchor), \
         ui.dialog().props('persistent no-refocus no-shake').on('close', lambda: _dialog_guard_close()) as dlg, \
            ui.card().classes('rounded-xl p-0 bg-gray-100 flex flex-col overflow-hidden').style('width:75vw;max-width:75vw;height:75vh').on('keydown.escape', lambda: dlg.close()):
        # Tělo dialogu jako refreshable — živá synchronizace překresluje jen
        # obsah na místě, okno zůstává otevřené (žádné close+reopen).
        @ui.refreshable
        def _telo():
            conn = intranet_data.get_db_connection()
            if not conn:
                return
            try:
                cur = conn.cursor(dictionary=True)
                cur.execute("SELECT * FROM ukolovnik_ukoly WHERE id = %s", (ukol_id,))
                uk = cur.fetchone()
            finally:
                cur.close(); conn.close()
            if not uk:
                ui.notify('Úkol byl mezitím smazán jiným uživatelem.', type='warning')
                dlg.close()
                return
            is_admin = vsechna_prava and ('vse' in vsechna_prava or 'ukolovnik_admin' in vsechna_prava)
            muze_editovat = is_admin or uk.get('prirazen_id') == user_id or uk.get('zadal_id') == user_id
            muze_editovat_meta = is_admin or uk.get('zadal_id') == user_id

            # ── Hlavička ──────────────────────────────────────────────────────────
            b_stav, _ = STAVY_UKOL.get(uk['stav'], ('bg-gray-100 text-gray-800 border-gray-300', '#ccc'))
            termin_str = uk['termin'].strftime('%d.%m.%Y') if uk.get('termin') else '—'
            delta, dt_cls, dt_txt = _dny_do_terminu(uk.get('termin'))
            p_barva = PRIORITY_BARVY.get(uk.get('priorita', 'Normální'), 'text-blue-600')

            with ui.row().classes('w-full items-center justify-between px-6 py-4 bg-white border-b border-gray-200 shrink-0 shadow-sm flex-wrap gap-3'):
                with ui.column().classes('gap-0.5 flex-1 min-w-0'):
                    ui.label(uk['nazev']).classes('text-2xl font-black text-gray-800 line-clamp-1')
                    if uk.get('porada_id'):
                        pd_str = uk['porada_datum'].strftime('%d.%m.%Y') if uk.get('porada_datum') else ''
                        with ui.row().classes('items-center gap-1'):
                            ui.icon('meeting_room', size='xs', color='blue-400')
                            ui.label(f"Z porady: {uk.get('porada_nazev','')} ({pd_str})").classes('text-xs text-blue-600 font-bold')

                with ui.row().classes('items-center gap-2 shrink-0 flex-wrap'):
                    ui.label(uk['stav']).classes(f'px-3 py-1 rounded-full text-xs font-bold border uppercase tracking-wider {b_stav}')
                    if uk.get('termin'):
                        ui.label(dt_txt).classes(f'text-xs {dt_cls}')
                    if muze_editovat_meta:
                        def _smazat_ukol_click():
                            def _provest():
                                ok = _smazat_ukol_db(ukol_id)
                                if ok:
                                    intranet_logger.log_activity(user_name, "Úkolovník", f"Smazán úkol #{ukol_id}")
                                    ui.notify('Úkol byl smazán.', type='positive', position='top')
                                    dlg.close()
                                    if on_refresh: on_refresh()
                                else:
                                    ui.notify('Chyba při mazání.', type='negative', position='top')
                            _potvrdit_smazat(f'Opravdu smazat úkol „{uk["nazev"]}"?\nTato akce je nevratná.', _provest)
                        ui.button(icon='delete_forever', on_click=_smazat_ukol_click).props('flat round dense').classes('text-red-400 hover:text-red-600').tooltip('Smazat úkol')

                    # ── Duplikovat úkol ──
                    def _duplikovat_ukol_click():
                        with ui.dialog() as dup_dlg, ui.card().classes('p-5 min-w-[320px] rounded-xl'):
                            ui.label('📋 Duplikovat úkol').classes('text-lg font-extrabold text-gray-800 mb-2')
                            ui.label(f'„{uk["nazev"][:60]}"').classes('text-sm text-gray-500 mb-3')
                            dup_datum = ui.input('Nový termín (DD.MM.RRRR)').classes('w-full bg-white mb-1').props('outlined')
                            dup_datum.value = datetime.date.today().strftime('%d.%m.%Y')
                            with dup_datum.add_slot('append'):
                                ui.icon('event').classes('cursor-pointer text-blue-500').on('click', lambda: _dup_menu.open())
                            with ui.menu().props('no-parent-event') as _dup_menu:
                                ui.date(mask='DD.MM.YYYY').bind_value(dup_datum).props('today-btn').classes('p-0')
                            def _provest_duplikaci():
                                try:
                                    novy_termin = datetime.datetime.strptime(dup_datum.value.strip(), '%d.%m.%Y').date()
                                except Exception:
                                    ui.notify('Neplatný formát data.', type='warning'); return
                                c = intranet_data.get_db_connection()
                                cu = c.cursor()
                                cu.execute("""
                                    INSERT INTO ukolovnik_ukoly
                                        (nazev, popis, porada_id, porada_nazev, porada_datum, prirazen_id, prirazen_jmeno,
                                         zadal_id, zadal_jmeno, termin, odhad_hodin, priorita, oddeleni, typ_ukolu)
                                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                                """, (uk['nazev'], uk.get('popis', ''),
                                      uk.get('porada_id'), uk.get('porada_nazev'), uk.get('porada_datum'),
                                      uk.get('prirazen_id'), uk.get('prirazen_jmeno'),
                                      user_id, user_name, novy_termin,
                                      uk.get('odhad_hodin'), uk.get('priorita', 'Normální'),
                                      uk.get('oddeleni'), uk.get('typ_ukolu', 'Pravidelný')))
                                nova_id = cu.lastrowid
                                c.commit(); cu.close(); c.close()
                                intranet_logger.log_activity(user_name, "Úkolovník", f"Duplikace úkolu #{uk['id']} → #{nova_id}")
                                _log('ukol', nova_id, user_id, user_name, f'Duplikováno z úkolu #{uk["id"]}: {uk["nazev"][:60]}')
                                if uk.get('prirazen_id') and uk['prirazen_id'] != user_id:
                                    intranet_notifikace.pridej(uk['prirazen_id'],
                                        f"📋 {user_name} ti duplikoval/a úkol: '{uk['nazev'][:50]}' → {novy_termin.strftime('%d.%m.%Y')}", 'info')
                                ui.notify(f'Úkol duplikován na {novy_termin.strftime("%d.%m.%Y")}.', type='positive', position='top')
                                dup_dlg.close()
                                if on_refresh: on_refresh()
                            with ui.row().classes('w-full justify-end gap-2 mt-3'):
                                ui.button('Zrušit', on_click=dup_dlg.close).classes('bg-gray-400 text-white font-bold px-4')
                                ui.button('Duplikovat', icon='content_copy', on_click=_provest_duplikaci).classes('bg-blue-600 text-white font-bold px-4')
                        dup_dlg.open()
                    ui.button(icon='content_copy', on_click=_duplikovat_ukol_click).props('flat round dense').classes('text-blue-400 hover:text-blue-600').tooltip('Duplikovat úkol')

                    ui.button(icon='close', on_click=dlg.close).props('flat round dense').classes('text-gray-400 hover:text-red-500')

            # ── Tělo: 3 sloupce ──────────────────────────────────────────────────
            with ui.row().classes('w-full flex-1 overflow-hidden'):

                # Levý panel: meta + timer + poznámky + přílohy
                with ui.column().classes('flex-1 overflow-y-auto p-5 gap-4 bg-white border-r border-gray-100'):

                    # Meta informace
                    with ui.card().classes('w-full p-4 bg-gray-50 border border-gray-200 rounded-xl'):
                        with ui.grid(columns=2).classes('w-full gap-x-6 gap-y-2 text-sm'):
                            ui.label('Přiděleno:').classes('text-gray-500 font-bold')
                            ui.label(uk.get('prirazen_jmeno') or '—').classes('text-gray-800 font-bold')
                            ui.label('Zadal:').classes('text-gray-500 font-bold')
                            ui.label(uk.get('zadal_jmeno') or '—').classes('text-gray-800')
                            ui.label('Termín:').classes('text-gray-500 font-bold')
                            ui.label(termin_str).classes(dt_cls)
                            if uk.get('dokonceno_datum'):
                                ui.label('Dokončeno:').classes('text-gray-500 font-bold')
                                dok_str = uk['dokonceno_datum'].strftime('%d.%m.%Y %H:%M')
                                if uk.get('termin'):
                                    _d_diff = (uk['dokonceno_datum'].date() - uk['termin']).days
                                    if _d_diff > 0:
                                        dok_str += f'  (+{_d_diff} dní po termínu)'
                                        ui.label(dok_str).classes('text-red-600 font-bold')
                                    elif _d_diff == 0:
                                        ui.label(dok_str + '  (včas)').classes('text-green-600 font-bold')
                                    else:
                                        ui.label(dok_str + f'  ({abs(_d_diff)} dní před termínem)').classes('text-green-600 font-bold')
                                else:
                                    ui.label(dok_str).classes('text-gray-800')
                            ui.label('Typ:').classes('text-gray-500 font-bold')
                            _typ = uk.get('typ_ukolu', 'Pravidelný')
                            _typ_cls = 'text-orange-600 font-extrabold' if _typ == 'Mimořádný' else 'text-blue-600 font-bold'
                            ui.label(('⚡ ' if _typ == 'Mimořádný' else '🔄 ') + _typ).classes(_typ_cls)
                            ui.label('Odhad práce:').classes('text-gray-500 font-bold')
                            ui.label(_format_odhad(uk.get('odhad_hodin'))).classes('text-gray-800')
                            ui.label('Priorita:').classes('text-gray-500 font-bold')
                            ui.label(uk.get('priorita', 'Normální')).classes(p_barva + ' font-bold')
                            ui.label('Vytvořeno:').classes('text-gray-500 font-bold')
                            ui.label(uk['vytvoreno'].strftime('%d.%m.%Y %H:%M')).classes('text-gray-800')
                            if uk.get('oddeleni'):
                                ui.label('Oddělení:').classes('text-gray-500 font-bold')
                                with ui.row().classes('gap-1 flex-wrap'):
                                    for _odd in _oddeleni_chips(uk['oddeleni'], max_chips=10):
                                        ui.label(_odd).classes('bg-indigo-100 text-indigo-700 text-[10px] px-2 py-0.5 rounded-full font-bold border border-indigo-200')

                            # Další termín opakovaného úkolu
                            if uk.get('opakovani_id'):
                                _conn_dt = intranet_data.get_db_connection()
                                _dalsi_t = None
                                if _conn_dt:
                                    try:
                                        _cur_dt = _conn_dt.cursor(dictionary=True)
                                        _cur_dt.execute("""
                                            SELECT MIN(termin) as dalsi FROM ukolovnik_ukoly
                                            WHERE opakovani_id=%s AND termin > CURDATE()
                                              AND stav NOT IN ('Hotovo','Zrušen')
                                        """, (uk['opakovani_id'],))
                                        _r = _cur_dt.fetchone()
                                        _dalsi_t = _r['dalsi'] if _r else None
                                    finally:
                                        _cur_dt.close(); _conn_dt.close()
                                ui.label('Opakovaný úkol:').classes('text-gray-500 font-bold')
                                if _dalsi_t:
                                    _dn_t = DNY_CZ_PLNE[_dalsi_t.weekday()]
                                    ui.label(f'🔁 Další termín: {_dn_t} {_dalsi_t.strftime("%d.%m.%Y")}').classes('text-purple-600 font-bold')
                                else:
                                    ui.label('🔁 Žádný další výskyt').classes('text-gray-400')

                        if uk.get('popis'):
                            ui.separator().classes('my-2')
                            ui.label('Popis:').classes('text-xs font-bold text-gray-500 uppercase')
                            ui.label(uk['popis']).classes('text-sm text-gray-700 whitespace-pre-wrap mt-1 bg-white p-2 rounded border border-gray-100')

                    # ── EDITACE METADAT (jen zadavatel/admin) ─────────────────────
                    if muze_editovat_meta:
                        _meta_titulek = '✏️ Upravit přiřazení / termín / typ / oddělení' if is_admin else '✏️ Upravit přiřazení / termín / typ'
                        with ui.expansion(_meta_titulek, icon='edit').classes('w-full border border-gray-200 rounded-xl bg-white'):
                            # oddělení smí měnit jen správce modulu
                            oddeleni_opts_d = _ziskej_oddeleni_options() if is_admin else {}
                            curr_odd = [o.strip() for o in (uk.get('oddeleni') or '').split(',') if o.strip()]
                            # Osoby, na které lze úkol přehodit (dle oprávnění)
                            uziv_opts_d = _ziskej_viditelne_uzivatele(user_id, vsechna_prava)
                            if uk.get('prirazen_id') and uk['prirazen_id'] not in uziv_opts_d:
                                uziv_opts_d = {uk['prirazen_id']: uk.get('prirazen_jmeno') or '—', **uziv_opts_d}
                            with ui.column().classes('w-full gap-3 p-3'):
                                edit_prirazen = ui.select(
                                    uziv_opts_d, label='Přiděleno (vlastník úkolu)',
                                    value=uk.get('prirazen_id'), with_input=True
                                ).classes('w-full bg-white').props('outlined dense')
                                edit_termin = ui.input('Termín (DD.MM.RRRR)', value=termin_str if uk.get('termin') else '').classes('w-full bg-white').props('outlined dense')
                                with edit_termin.add_slot('append'):
                                    ui.icon('event').classes('cursor-pointer text-blue-500').on('click', lambda: edit_termin_menu.open())
                                with ui.menu().props('no-parent-event') as edit_termin_menu:
                                    ui.date(mask='DD.MM.YYYY').bind_value(edit_termin).props('today-btn').classes('p-0')
                                edit_typ = ui.select(
                                    {'Pravidelný': '🔄 Pravidelný', 'Mimořádný': '⚡ Mimořádný'},
                                    label='Typ', value=uk.get('typ_ukolu', 'Pravidelný')
                                ).classes('w-full bg-white').props('outlined dense')
                                edit_odhad = ui.number('Odhad (minuty)', value=int(round((uk.get('odhad_hodin') or 0) * 60)), min=0, step=15).classes('w-full bg-white').props('outlined dense')
                                if oddeleni_opts_d:
                                    edit_odd = ui.select(
                                        oddeleni_opts_d, label='Oddělení',
                                        multiple=True, value=curr_odd
                                    ).classes('w-full bg-white').props('outlined dense use-chips')
                                else:
                                    edit_odd = None

                                def _ulozit_meta():
                                    new_termin = None
                                    try:
                                        new_termin = datetime.datetime.strptime(edit_termin.value.strip(), '%d.%m.%Y').date()
                                    except Exception:
                                        pass
                                    if edit_odd is not None:
                                        new_odd = ','.join(edit_odd.value) if edit_odd.value else None
                                    else:
                                        new_odd = uk.get('oddeleni')  # nesprávce pole nevidí → zachovej stávající
                                    new_odhad = (float(edit_odhad.value) / 60.0) if edit_odhad.value else None  # minuty → hodiny
                                    # Nový vlastník úkolu
                                    novy_prirazen_id = edit_prirazen.value
                                    stary_prirazen_id = uk.get('prirazen_id')
                                    zmena_vlastnika = bool(novy_prirazen_id) and novy_prirazen_id != stary_prirazen_id
                                    novy_prirazen_jmeno = uziv_opts_d.get(novy_prirazen_id, uk.get('prirazen_jmeno'))
                                    c = intranet_data.get_db_connection()
                                    cu = c.cursor()
                                    cu.execute(
                                        "UPDATE ukolovnik_ukoly SET termin=%s, typ_ukolu=%s, oddeleni=%s, odhad_hodin=%s, "
                                        "prirazen_id=%s, prirazen_jmeno=%s WHERE id=%s",
                                        (new_termin, edit_typ.value, new_odd, new_odhad,
                                         novy_prirazen_id, novy_prirazen_jmeno, ukol_id)
                                    )
                                    # Předelegování → ukonči případný běžící časovač původního zpracovatele.
                                    # Jinak by mu úkol zůstal viset v reálném kalendáři s neukončeným záznamem
                                    # (cas_konec IS NULL) a časovač by běžel donekonečna.
                                    if zmena_vlastnika and stary_prirazen_id:
                                        _now = datetime.datetime.now()
                                        cu.execute(
                                            "SELECT id, cas_start FROM ukolovnik_cas_zaznamy "
                                            "WHERE ukol_id=%s AND user_id=%s AND cas_konec IS NULL "
                                            "ORDER BY cas_start DESC LIMIT 1",
                                            (ukol_id, stary_prirazen_id))
                                        _bezici = cu.fetchone()
                                        if _bezici:
                                            _zid, _zstart = _bezici[0], _bezici[1]
                                            _delta_min = max(1, int((_now - _zstart).total_seconds() / 60))
                                            cu.execute(
                                                "UPDATE ukolovnik_cas_zaznamy SET cas_konec=%s, trvani_minut=%s WHERE id=%s",
                                                (_now, _delta_min, _zid))
                                    c.commit(); cu.close(); c.close()
                                    _log('ukol', ukol_id, user_id, user_name,
                                         f'Upravil/a metadata: termín={new_termin}, typ={edit_typ.value}')
                                    # Změna termínu → notifikace „z X na Y", jinak obecná úprava
                                    stary_termin = uk.get('termin')
                                    if isinstance(stary_termin, datetime.datetime):
                                        stary_termin = stary_termin.date()
                                    if new_termin != stary_termin:
                                        _z  = stary_termin.strftime('%d.%m.%Y') if stary_termin else '—'
                                        _na = new_termin.strftime('%d.%m.%Y') if new_termin else '—'
                                        _notifikuj_ucastniky_ukolu(ukol_id, user_id,
                                            f"✏️ {user_name} změnil/a termín úkolu '{uk['nazev'][:40]}' z {_z} na {_na}.")
                                    else:
                                        _notifikuj_ucastniky_ukolu(ukol_id, user_id,
                                            f"✏️ {user_name} upravil/a úkol '{uk['nazev'][:40]}'.")
                                    if zmena_vlastnika:
                                        _log('ukol', ukol_id, user_id, user_name,
                                             f'Přehodil/a úkol na: {novy_prirazen_jmeno}')
                                        _term = f" (termín: {new_termin.strftime('%d.%m.%Y')})" if new_termin else ''
                                        intranet_notifikace.pridej(novy_prirazen_id,
                                            f"📌 {user_name} ti přidělil/a úkol: '{uk['nazev'][:50]}'{_term}", 'info')
                                    ui.notify('Uloženo.', type='positive', position='top-right')
                                    dlg.close()
                                    if on_refresh: on_refresh()

                                ui.button('Uložit změny', icon='save', on_click=_ulozit_meta).classes('bg-blue-600 text-white font-bold px-5 mt-1')

                    # ── TIMER ──────────────────────────────────────────────────────
                    if muze_editovat and uk['stav'] != 'Hotovo' and uk['stav'] != 'Zrušen':
                        with ui.card().classes('w-full p-4 bg-blue-50 border border-blue-200 rounded-xl'):
                            ui.label('Sledování času').classes('text-sm font-extrabold text-blue-800 mb-2')

                            aktivni = _ziskej_aktivni_zaznam(ukol_id, user_id)

                            # Celkový odpracovaný čas
                            c2 = intranet_data.get_db_connection()
                            total_min = 0
                            if c2:
                                try:
                                    cu = c2.cursor(dictionary=True)
                                    cu.execute("SELECT SUM(trvani_minut) as total FROM ukolovnik_cas_zaznamy WHERE ukol_id=%s AND cas_konec IS NOT NULL", (ukol_id,))
                                    row2 = cu.fetchone()
                                    total_min = row2['total'] or 0
                                finally:
                                    cu.close(); c2.close()

                            if total_min:
                                h, m = divmod(int(total_min), 60)
                                ui.label(f'Odpracováno celkem: {h}h {m}min').classes('text-xs text-blue-700 mb-2')

                            if aktivni:
                                delta_sec = int((datetime.datetime.now() - aktivni['cas_start']).total_seconds())
                                h2, rem = divmod(delta_sec, 3600)
                                m2, _ = divmod(rem, 60)
                                ui.label(f'⏱ Běží od {aktivni["cas_start"].strftime("%H:%M")} (cca {h2}h {m2}min)').classes('text-xs font-bold text-orange-600 mb-2')

                                with ui.row().classes('gap-2'):
                                    def _pozastav():
                                        _pozastav_ukol(ukol_id, user_id, user_name)
                                        dlg.close()
                                        if on_refresh: on_refresh()

                                    def _dokoncit():
                                        _dokoncit_ukol(ukol_id, user_id, user_name)
                                        dlg.close()
                                        if on_refresh: on_refresh()

                                    ui.button('Pozastavit', icon='pause', on_click=_pozastav).classes('bg-yellow-500 text-white font-bold px-4')
                                    ui.button('Dokončit', icon='check_circle', on_click=_dokoncit).classes('bg-green-600 text-white font-bold px-4')
                            else:
                                def _zahaj():
                                    _zahaj_ukol(ukol_id, user_id, user_name)
                                    dlg.close()
                                    if on_refresh: on_refresh()

                                btn_label = 'Zahájit úkol' if uk['stav'] == 'Zadáno' else 'Pokračovat v úkolu'
                                ui.button(btn_label, icon='play_arrow', on_click=_zahaj).classes('bg-blue-600 text-white font-bold px-6 shadow-md')

                            # Tlačítka změny stavu (manuální) — se správou timeru
                            if muze_editovat:
                                with ui.row().classes('mt-3 gap-2 pt-3 border-t border-blue-100 flex-wrap items-center'):
                                    ui.label('Stav:').classes('text-xs font-bold text-blue-700')
                                    for s in ['Zadáno', 'Rozpracováno', 'Pozastaveno', 'Hotovo', 'Zrušen']:
                                        def _zmen(ns=s):
                                            _zmenit_stav_ukolu(uk, ns, user_id, user_name)
                                            dlg.close()
                                            if on_refresh: on_refresh()
                                        b_cls, _ = STAVY_UKOL.get(s, ('', ''))
                                        ui.button(s, on_click=_zmen).props('size=xs outline').classes('font-bold text-xs')

                    # ── POZNÁMKY ──────────────────────────────────────────────────
                    ui.label('Poznámky').classes('text-base font-extrabold text-gray-700')

                    # Vlastní scrollbar — při větším počtu poznámek se roluje jen tento
                    # blok, ne celý panel (jinak nový komentář „uteče" pod přílohy).
                    pozn_container = ui.column().classes('w-full gap-2 max-h-[340px] overflow-y-auto pr-1')

                    def nacti_poznamky():
                        pozn_container.clear()
                        c = intranet_data.get_db_connection()
                        poznamky = []
                        if c:
                            try:
                                cu = c.cursor(dictionary=True)
                                cu.execute("SELECT * FROM ukolovnik_ukol_poznamky WHERE ukol_id=%s ORDER BY vytvoreno ASC", (ukol_id,))
                                poznamky = cu.fetchall()
                            finally:
                                cu.close(); c.close()
                        with pozn_container:
                            if not poznamky:
                                ui.label('Zatím žádné poznámky.').classes('text-gray-400 italic text-xs')
                            for p in poznamky:
                                je_muj = p['user_id'] == user_id
                                bc = 'bg-blue-50 border-blue-200' if je_muj else 'bg-gray-50 border-gray-200'
                                with ui.card().classes(f'w-full p-3 border rounded-lg {bc}'):
                                    with ui.row().classes('w-full justify-between items-center mb-1'):
                                        ui.label(p['jmeno_autora']).classes('text-xs font-bold text-gray-600')
                                        ui.label(p['vytvoreno'].strftime('%d.%m.%Y %H:%M')).classes('text-xs text-gray-400')
                                    ui.label(p['text']).classes('text-sm text-gray-800 whitespace-pre-wrap')

                    nacti_poznamky()

                    with ui.row().classes('w-full items-end gap-2'):
                        nova_poz = ui.textarea('Nová poznámka...').classes('flex-1 bg-gray-50').props('outlined autogrow').bind_value(_navrhy, 'poznamka')

                        def pridat_poznamku():
                            txt = nova_poz.value.strip() if nova_poz.value else ''
                            if not txt: return
                            c = intranet_data.get_db_connection()
                            cu = c.cursor()
                            cu.execute(
                                "INSERT INTO ukolovnik_ukol_poznamky (ukol_id, user_id, jmeno_autora, text) VALUES (%s,%s,%s,%s)",
                                (ukol_id, user_id, user_name, txt)
                            )
                            c.commit(); cu.close(); c.close()
                            intranet_logger.log_activity(user_name, "Úkolovník", f"Poznámka k úkolu #{ukol_id}")
                            _log('ukol', ukol_id, user_id, user_name, 'Přidal/a poznámku')
                            _notifikuj_ucastniky_ukolu(ukol_id, user_id, f"💬 {user_name} přidal/a poznámku k úkolu '{uk['nazev'][:40]}'.")
                            nova_poz.value = ''
                            nacti_poznamky()

                        ui.button(icon='send', on_click=pridat_poznamku).classes('bg-blue-600 text-white h-12 w-12 rounded-xl')

                    # ── PŘÍLOHY ──────────────────────────────────────────────────
                    ui.label('Přílohy').classes('text-base font-extrabold text-gray-700 mt-2')

                    prilohy_container = ui.column().classes('w-full gap-2')

                    def nacti_prilohy():
                        prilohy_container.clear()
                        c = intranet_data.get_db_connection()
                        prilohy = []
                        if c:
                            try:
                                cu = c.cursor(dictionary=True)
                                cu.execute("SELECT * FROM ukolovnik_ukol_prilohy WHERE ukol_id=%s ORDER BY vytvoreno ASC", (ukol_id,))
                                prilohy = cu.fetchall()
                            finally:
                                cu.close(); c.close()
                        with prilohy_container:
                            if not prilohy:
                                ui.label('Žádné přílohy.').classes('text-gray-400 italic text-xs')
                            for pr in prilohy:
                                ikona, ik_cls = _ikona_souboru(pr['soubor_nazev'])
                                with ui.row().classes('w-full items-center gap-2 p-2 bg-gray-50 border border-gray-200 rounded-lg'):
                                    ui.icon(ikona, size='sm').classes(ik_cls)
                                    ui.link(pr['soubor_nazev'], f"/ukol_prilohy/{os.path.basename(pr['soubor_cesta'])}").classes('text-sm text-blue-600 hover:underline flex-1 truncate').props('target=_blank')
                                    ui.label(pr.get('jmeno_autora', '')).classes('text-xs text-gray-400')
                                    ui.label(pr['vytvoreno'].strftime('%d.%m.%Y')).classes('text-xs text-gray-400')

                    nacti_prilohy()

                    async def zpracuj_upload(e: ui_events.UploadEventArguments):
                        try:
                            soubor_id = str(uuid.uuid4())
                            nazev = e.file.name
                            ext = os.path.splitext(nazev)[1]
                            cesta = os.path.join(UKOL_PRILOHY_DIR, f'{soubor_id}{ext}')
                            os.makedirs(UKOL_PRILOHY_DIR, exist_ok=True)
                            obsah = await e.file.read()
                            with open(cesta, 'wb') as f:
                                f.write(obsah)
                            c = intranet_data.get_db_connection()
                            cu = c.cursor()
                            cu.execute(
                                "INSERT INTO ukolovnik_ukol_prilohy (id, ukol_id, user_id, jmeno_autora, soubor_nazev, soubor_cesta) VALUES (%s,%s,%s,%s,%s,%s)",
                                (soubor_id, ukol_id, user_id, user_name, nazev, cesta)
                            )
                            c.commit(); cu.close(); c.close()
                            intranet_logger.log_activity(user_name, "Úkolovník", f"Příloha k úkolu #{ukol_id}: {nazev}")
                            _log('ukol', ukol_id, user_id, user_name, f'Nahrál/a přílohu: {nazev}')
                            _notifikuj_ucastniky_ukolu(ukol_id, user_id, f"📎 {user_name} přidal/a přílohu '{nazev}' k úkolu '{uk['nazev'][:40]}'.")
                            ui.notify(f'Nahráno: {nazev}', type='positive')
                            nacti_prilohy()
                        except Exception as ex:
                            ui.notify(f'Chyba nahrávání: {ex}', type='negative')

                    if uk['stav'] not in ('Hotovo', 'Zrušen'):
                        ui.upload(
                            label='Nahrát přílohu (PDF, Excel, obrázek…)',
                            on_upload=zpracuj_upload,
                            auto_upload=True,
                            multiple=True
                        ).props('accept=".pdf,.xlsx,.xls,.csv,.jpg,.jpeg,.png,.gif,.doc,.docx" flat color=blue-grey').classes('w-full mt-1')
                    else:
                        ui.label('Nahrávání příloh je uzamčeno — úkol je dokončen.').classes('text-xs text-gray-400 italic mt-1')

                # Pravý panel: plány + časový log
                with ui.column().classes('w-80 shrink-0 overflow-y-auto p-5 gap-4 bg-gray-50 border-l border-gray-200'):

                    # ── PLÁNY / CHECKLISTY ─────────────────────────────────────
                    ui.label('Plán úkolu').classes('text-base font-extrabold text-gray-700')

                    plany_container = ui.column().classes('w-full gap-2')

                    def nacti_plany():
                        plany_container.clear()
                        c = intranet_data.get_db_connection()
                        plany = []
                        if c:
                            try:
                                cu = c.cursor(dictionary=True)
                                cu.execute("SELECT * FROM ukolovnik_ukol_plany WHERE ukol_id=%s ORDER BY poradi, id", (ukol_id,))
                                plany = cu.fetchall()
                            finally:
                                cu.close(); c.close()
                        with plany_container:
                            if not plany:
                                ui.label('Přidejte kroky plánu.').classes('text-gray-400 italic text-xs')
                            for pl in plany:
                                hotovo_val = bool(pl['hotovo'])
                                text_cls = 'line-through text-gray-400' if hotovo_val else 'text-gray-800'
                                termin_pl = pl['termin'].strftime('%d.%m.') if pl.get('termin') else ''

                                def _toggle_plan(pid=pl['id'], curr=hotovo_val, popis=pl['popis']):
                                    c2 = intranet_data.get_db_connection()
                                    cu2 = c2.cursor()
                                    cu2.execute("UPDATE ukolovnik_ukol_plany SET hotovo=%s WHERE id=%s", (0 if curr else 1, pid))
                                    c2.commit(); cu2.close(); c2.close()
                                    akce_pl = f'Splnil/a krok: {popis[:80]}' if not curr else f'Odznačil/a krok: {popis[:80]}'
                                    _log('ukol', ukol_id, user_id, user_name, akce_pl)
                                    nacti_plany()

                                def _smazat_plan(pid=pl['id']):
                                    c2 = intranet_data.get_db_connection()
                                    cu2 = c2.cursor()
                                    cu2.execute("DELETE FROM ukolovnik_ukol_plany WHERE id=%s", (pid,))
                                    c2.commit(); cu2.close(); c2.close()
                                    nacti_plany()

                                with ui.row().classes('w-full items-center gap-2 p-2 bg-white border border-gray-200 rounded-lg'):
                                    ui.button(
                                        icon='check_box' if hotovo_val else 'check_box_outline_blank',
                                        on_click=lambda pll=pl: _toggle_plan(pll['id'], bool(pll['hotovo']))
                                    ).props('flat round dense').classes('text-green-500' if hotovo_val else 'text-gray-300')
                                    with ui.column().classes('flex-1 gap-0'):
                                        ui.label(pl['popis']).classes(f'text-sm {text_cls}')
                                        if termin_pl:
                                            ui.label(f'do {termin_pl}').classes('text-xs text-gray-400')
                                    ui.button(icon='close', on_click=lambda pll=pl: _smazat_plan(pll['id'])).props('flat round dense size=xs').classes('text-gray-300 hover:text-red-400')

                    nacti_plany()

                    # Přidat krok plánu
                    with ui.row().classes('w-full gap-1 items-end'):
                        novy_plan_text = ui.input('Nový krok…').classes('flex-1 text-xs').props('outlined dense').bind_value(_navrhy, 'plan_text')
                        novy_plan_termin = ui.input('Termín').classes('w-20 text-xs').props('outlined dense placeholder="DD.MM."').bind_value(_navrhy, 'plan_termin')

                        def pridat_plan():
                            txt = novy_plan_text.value.strip() if novy_plan_text.value else ''
                            if not txt: return
                            termin_p = None
                            if novy_plan_termin.value and novy_plan_termin.value.strip():
                                try:
                                    rok = datetime.date.today().year
                                    termin_p = datetime.datetime.strptime(f"{novy_plan_termin.value.strip()}.{rok}", '%d.%m.%Y').date()
                                except Exception:
                                    pass
                            c = intranet_data.get_db_connection()
                            cu = c.cursor()
                            cu.execute("INSERT INTO ukolovnik_ukol_plany (ukol_id, popis, termin) VALUES (%s,%s,%s)", (ukol_id, txt, termin_p))
                            c.commit(); cu.close(); c.close()
                            _log('ukol', ukol_id, user_id, user_name, f'Přidal/a krok plánu: {txt[:80]}')
                            novy_plan_text.value = ''
                            novy_plan_termin.value = ''
                            nacti_plany()

                        ui.button(icon='add', on_click=pridat_plan).props('round dense').classes('bg-blue-600 text-white')

                    ui.separator().classes('my-2')

                    # ── ČASOVÝ LOG ─────────────────────────────────────────────
                    ui.label('Záznamy práce').classes('text-sm font-extrabold text-gray-700 mt-1')

                    c3 = intranet_data.get_db_connection()
                    zaznamy = []
                    if c3:
                        try:
                            cu3 = c3.cursor(dictionary=True)
                            cu3.execute("SELECT * FROM ukolovnik_cas_zaznamy WHERE ukol_id=%s ORDER BY cas_start DESC LIMIT 20", (ukol_id,))
                            zaznamy = cu3.fetchall()
                        finally:
                            cu3.close(); c3.close()

                    if not zaznamy:
                        ui.label('Žádné záznamy.').classes('text-gray-400 italic text-xs')
                    else:
                        for z in zaznamy:
                            start_s = z['cas_start'].strftime('%d.%m. %H:%M')
                            if z.get('trvani_minut'):
                                h, m = divmod(z['trvani_minut'], 60)
                                doba = f'{h}h {m}min'
                            elif not z.get('cas_konec'):
                                doba = '⏱ běží'
                            else:
                                doba = '?'
                            with ui.row().classes('w-full justify-between text-xs text-gray-600 py-1 border-b border-gray-100'):
                                ui.label(f'{z["jmeno"]}').classes('font-bold truncate max-w-[100px]')
                                ui.label(start_s)
                                ui.label(doba).classes('text-blue-600 font-bold')

        _telo()

    _dialog_guard_open()
    # Evidence pro živou obnovu: změní-li úkol jiný uživatel, obsah otevřeného
    # náhledu se dotčenému klientovi překreslí na místě (okno se nezavírá).
    _registruj_otevreny_detail('ukol', ukol_id, dlg, _telo.refresh)
    dlg.open()


def _vypocti_vyskyty(frekvence: str, dny_tydne: list, den_mesice: int,
                     od: datetime.date, do: datetime.date) -> list:
    """Vrátí seřazený seznam dat výskytů opakovaného úkolu v intervalu [od, do]."""
    vyskyty = []
    if frekvence == 'tyden':
        wd_set = set(int(d) for d in (dny_tydne or []))
        if not wd_set:
            return []
        den = od
        while den <= do:
            if den.weekday() in wd_set:
                vyskyty.append(den)
            den += datetime.timedelta(days=1)
    elif frekvence == 'mesic':
        if not den_mesice:
            return []
        rok, mes = od.year, od.month
        while True:
            posl = calendar.monthrange(rok, mes)[1]
            d = min(int(den_mesice), posl)  # ošetři měsíce kratší než 31
            datum = datetime.date(rok, mes, d)
            if datum > do:
                break
            if datum >= od:
                vyskyty.append(datum)
            mes += 1
            if mes > 12:
                mes = 1; rok += 1
    return vyskyty


# =========================================================
# DIALOG: NOVÝ ÚKOL
# =========================================================
def _dialog_novy_ukol(user_id, user_name, vsechna_prava=None, porada_id=None, porada_nazev=None, porada_datum=None, on_refresh=None, dialog_anchor=None):
    if vsechna_prava is None:
        vsechna_prava = []

    is_admin   = 'vse' in vsechna_prava or 'ukolovnik_admin' in vsechna_prava
    je_vedouci = _je_hlavni_vedouci(vsechna_prava)

    # ── Výběr uživatelů a viditelnost oddělení ────────────────────────────────
    if is_admin:
        uziv_opts  = _ziskej_uzivatele_options()
        pouze_sobe = False
    elif je_vedouci:
        dept       = _ziskej_uzivatele_oddeleni(vsechna_prava)
        uziv_opts  = {u['id']: u['jmeno'] for u in dept} if dept else _ziskej_uzivatele_options()
        pouze_sobe = False
    elif 'ukolovnik_ukoly_zadat_oddeleni' in vsechna_prava:
        dept       = _ziskej_uziv_sveho_oddeleni(user_id)
        uziv_opts  = {u['id']: u['jmeno'] for u in dept} if dept else {user_id: user_name}
        pouze_sobe = False
    else:
        uziv_opts  = {user_id: user_name}
        pouze_sobe = True

    # Ujisti se, že aktuální uživatel je vždy v nabídce
    if user_id not in uziv_opts:
        uziv_opts[user_id] = user_name

    uziv_select = {v: v for v in uziv_opts.values()}

    with _dialog_kotva(dialog_anchor), \
         ui.dialog().props('persistent no-esc-dismiss no-backdrop-dismiss no-route-dismiss no-refocus no-shake').on('close', lambda: _dialog_guard_close()) as dlg, \
            ui.card().classes('w-full max-w-2xl p-6 rounded-xl').on('keydown.escape.stop.prevent', lambda: None):
        if porada_nazev:
            ui.label(f'Nový úkol z porady: {porada_nazev}').classes('text-xl font-bold text-blue-800 mb-1')
            if porada_datum:
                ui.label(porada_datum.strftime('%d.%m.%Y')).classes('text-xs text-gray-500 mb-3')
        else:
            ui.label('Nový úkol').classes('text-2xl font-bold text-blue-800 mb-4')

        nazev = ui.input('Název úkolu').classes('w-full mb-3 bg-white').props('outlined')
        popis = ui.textarea('Popis (volitelně)').classes('w-full mb-3 bg-white').props('outlined rows=2')

        _default_uziv = user_name if user_name in uziv_select else next(iter(uziv_select), None)

        # Omezení oddělení — zúží nabídku přidělených osob; pole vidí jen správce modulu
        oddeleni_opts = _ziskej_oddeleni_options() if is_admin else {}
        oddeleni_sel = None

        with ui.row().classes('w-full gap-3 mb-3 flex-wrap'):
            if pouze_sobe:
                ui.input('Přidělit komu', value=user_name).classes('flex-1 bg-white').props('outlined readonly')
                priradit = None
            else:
                priradit = ui.select(uziv_select, label='Přidělit komu', value=_default_uziv).classes('flex-1 bg-white').props('outlined')
            priorita = ui.select(
                {'Nízká': 'Nízká', 'Normální': 'Normální', 'Vysoká': 'Vysoká', 'Kritická': 'Kritická'},
                label='Priorita', value='Normální'
            ).classes('w-36 bg-white').props('outlined')

        if oddeleni_opts and priradit is not None:
            def _filtruj_dle_oddeleni():
                vybrana = oddeleni_sel.value or []
                if not vybrana:
                    nove = dict(uziv_select)
                else:
                    povolene_ids = {u['id'] for u in _ziskej_uziv_dle_oddeleni(vybrana)}
                    nove = {jm: jm for uid, jm in uziv_opts.items() if uid in povolene_ids}
                soucasna = priradit.value if priradit.value in nove else next(iter(nove), None)
                priradit.set_options(nove, value=soucasna)

            with ui.row().classes('w-full gap-3 mb-3 flex-wrap'):
                oddeleni_sel = ui.select(
                    oddeleni_opts, label='Omezit na oddělení (volitelné)', multiple=True
                ).classes('flex-1 bg-white').props('outlined use-chips')
                oddeleni_sel.on('update:model-value', lambda: _filtruj_dle_oddeleni())

        with ui.row().classes('w-full gap-3 mb-3 flex-wrap'):
            typ_ukolu_sel = ui.select(
                {'Pravidelný': '🔄 Pravidelný', 'Mimořádný': '⚡ Mimořádný'},
                label='Typ úkolu', value='Pravidelný'
            ).classes('w-44 bg-white').props('outlined')

        with ui.row().classes('w-full gap-3 mb-4 flex-wrap'):
            termin_inp = ui.input('Termín (DD.MM.RRRR)').classes('flex-1 bg-white').props('outlined')
            termin_inp.value = datetime.date.today().strftime('%d.%m.%Y')
            with termin_inp.add_slot('append'):
                ui.icon('event').classes('cursor-pointer text-blue-500').on('click', lambda: termin_menu.open())
            with ui.menu().props('no-parent-event') as termin_menu:
                ui.date(mask='DD.MM.YYYY').bind_value(termin_inp).props('today-btn').classes('p-0')
            odhad_inp = ui.number('Odhad (minuty)', min=0, step=15).classes('w-36 bg-white').props('outlined')

        # ── Opakování úkolu ───────────────────────────────────────────────────
        opakovat_sw = ui.switch('🔁 Opakovat úkol').classes('mb-1')
        opak_box = ui.column().classes('w-full gap-3 mb-3 p-3 bg-blue-50 border border-blue-200 rounded-xl')
        with opak_box:
            ui.label('Úkol se předgeneruje pro zvolené období (datumy včetně) a každý výskyt se zobrazí až ve svůj den.').classes('text-xs text-gray-500')
            with ui.row().classes('w-full gap-3 flex-wrap items-end'):
                opak_frekvence = ui.select(
                    {'tyden': 'Týdně', 'mesic': 'Měsíčně'}, label='Frekvence', value='tyden'
                ).classes('w-40 bg-white').props('outlined dense')
                opak_dny = ui.select(
                    {i: DNY_CZ[i] for i in range(7)}, label='Dny v týdnu', multiple=True, value=[0]
                ).classes('flex-1 bg-white').props('outlined dense use-chips')
                opak_den_mesice = ui.number('Den v měsíci', min=1, max=31, value=1, step=1).classes('w-36 bg-white').props('outlined dense')
            with ui.row().classes('w-full gap-3 flex-wrap items-end'):
                opak_od_inp = ui.input('Začátek opakování (DD.MM.RRRR)').classes('flex-1 bg-white').props('outlined dense')
                opak_od_inp.value = datetime.date.today().strftime('%d.%m.%Y')
                with opak_od_inp.add_slot('append'):
                    ui.icon('event').classes('cursor-pointer text-blue-500').on('click', lambda: opak_od_menu.open())
                with ui.menu().props('no-parent-event') as opak_od_menu:
                    ui.date(mask='DD.MM.YYYY').bind_value(opak_od_inp).props('today-btn').classes('p-0')
                opak_do_inp = ui.input('Konec opakování (DD.MM.RRRR)').classes('flex-1 bg-white').props('outlined dense')
                opak_do_inp.value = (datetime.date.today() + datetime.timedelta(days=92)).strftime('%d.%m.%Y')
                with opak_do_inp.add_slot('append'):
                    ui.icon('event').classes('cursor-pointer text-blue-500').on('click', lambda: opak_do_menu.open())
                with ui.menu().props('no-parent-event') as opak_do_menu:
                    ui.date(mask='DD.MM.YYYY').bind_value(opak_do_inp).props('today-btn').classes('p-0')
        opak_box.bind_visibility_from(opakovat_sw, 'value')

        def _opak_prepni_pole():
            je_tyden = opak_frekvence.value == 'tyden'
            opak_dny.set_visibility(je_tyden)
            opak_den_mesice.set_visibility(not je_tyden)
        opak_frekvence.on('update:model-value', lambda: _opak_prepni_pole())
        _opak_prepni_pole()

        def ulozit():
            if not nazev.value or not nazev.value.strip():
                ui.notify('Vyplňte název!', type='warning'); return
            termin_date = None
            try:
                termin_date = datetime.datetime.strptime(termin_inp.value.strip(), '%d.%m.%Y').date()
            except Exception:
                pass
            if pouze_sobe or priradit is None:
                prirazen_jmeno = user_name
                prirazen_id    = user_id
            else:
                prirazen_jmeno = priradit.value
                prirazen_id    = next((uid for uid, jm in uziv_opts.items() if jm == prirazen_jmeno), None)
            odhad = (float(odhad_inp.value) / 60.0) if odhad_inp.value else None  # minuty → hodiny
            odd_val = ','.join(oddeleni_sel.value) if oddeleni_sel and oddeleni_sel.value else None
            typ_val = typ_ukolu_sel.value
            nazev_val = nazev.value.strip()
            popis_val = popis.value.strip() if popis.value else ''

            # ── Opakovaný úkol → předgeneruj výskyty ─────────────────────────
            if opakovat_sw.value:
                try:
                    opak_od = datetime.datetime.strptime(opak_od_inp.value.strip(), '%d.%m.%Y').date()
                except Exception:
                    ui.notify('Neplatný začátek opakování.', type='warning'); return
                try:
                    opak_do = datetime.datetime.strptime(opak_do_inp.value.strip(), '%d.%m.%Y').date()
                except Exception:
                    ui.notify('Neplatný konec opakování.', type='warning'); return
                if opak_do < opak_od:
                    ui.notify('Konec opakování nesmí být dříve než začátek.', type='warning'); return
                vyskyty = _vypocti_vyskyty(
                    opak_frekvence.value,
                    [int(d) for d in (opak_dny.value or [])],
                    int(opak_den_mesice.value or 0),
                    opak_od, opak_do
                )
                if not vyskyty:
                    ui.notify('Vyberte alespoň jeden den opakování.', type='warning'); return
                if len(vyskyty) > 1000:
                    ui.notify(f'Příliš mnoho výskytů ({len(vyskyty)}). Zkraťte období opakování.', type='warning'); return
                opak_uid = str(uuid.uuid4())
                c = intranet_data.get_db_connection()
                cu = c.cursor()
                for d in vyskyty:
                    cu.execute("""
                        INSERT INTO ukolovnik_ukoly
                            (nazev, popis, porada_id, porada_nazev, porada_datum, prirazen_id, prirazen_jmeno,
                             zadal_id, zadal_jmeno, termin, odhad_hodin, priorita, oddeleni, typ_ukolu,
                             zobrazit_od, opakovani_id)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """, (nazev_val, popis_val, porada_id, porada_nazev, porada_datum,
                          prirazen_id, prirazen_jmeno, user_id, user_name,
                          d, odhad, priorita.value, odd_val, typ_val, d, opak_uid))
                c.commit(); cu.close(); c.close()
                _bump_data_verze()  # živě promítnout nové úkoly ostatním klientům
                intranet_logger.log_activity(user_name, "Úkolovník", f"Vytvořen opakovaný úkol ({len(vyskyty)}×): {nazev_val[:40]}")
                if prirazen_id and prirazen_id != user_id:
                    intranet_notifikace.pridej(prirazen_id, f"🔁 {user_name} ti přidělil/a opakovaný úkol: '{nazev_val[:50]}' ({len(vyskyty)} výskytů, první {vyskyty[0].strftime('%d.%m.%Y')})", 'info')
                ui.notify(f'Opakovaný úkol vytvořen ({len(vyskyty)} výskytů).', type='positive', position='top')
                dlg.close()
                if on_refresh: on_refresh()
                return

            c = intranet_data.get_db_connection()
            cu = c.cursor()
            cu.execute("""
                INSERT INTO ukolovnik_ukoly
                    (nazev, popis, porada_id, porada_nazev, porada_datum, prirazen_id, prirazen_jmeno,
                     zadal_id, zadal_jmeno, termin, odhad_hodin, priorita, oddeleni, typ_ukolu)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (nazev_val, popis_val,
                  porada_id, porada_nazev, porada_datum,
                  prirazen_id, prirazen_jmeno, user_id, user_name,
                  termin_date, odhad, priorita.value, odd_val, typ_val))
            nova_id = cu.lastrowid
            c.commit(); cu.close(); c.close()

            intranet_logger.log_activity(user_name, "Úkolovník", f"Vytvořen úkol: {nazev_val[:40]}")
            _log('ukol', nova_id, user_id, user_name, f'Vytvořil/a úkol: {nazev_val[:80]}')
            if prirazen_id and prirazen_id != user_id:
                _term = f" (termín: {termin_date.strftime('%d.%m.%Y')})" if termin_date else ''
                intranet_notifikace.pridej(prirazen_id, f"📋 {user_name} ti přidělil/a nový úkol: '{nazev_val[:50]}'{_term}", 'info')

            ui.notify('Úkol vytvořen.', type='positive', position='top')
            _dialog_guard_close()
            dlg.close()
            if on_refresh: on_refresh()

        def _zrusit():
            _dialog_guard_close()
            dlg.close()

        with ui.row().classes('w-full justify-between mt-2'):
            ui.button('Zrušit', on_click=_zrusit).classes('bg-gray-400 text-white font-bold px-6')
            ui.button('Vytvořit úkol', icon='add_task', on_click=ulozit).classes('bg-green-600 text-white font-bold px-6 shadow-md')

    _dialog_guard_open()
    dlg.open()


# =========================================================
# GUARD: Ochrana dialogů před refreshem
# =========================================================
# Pokud je otevřen dialog (nový úkol / detail), refresh se odloží
# a provede až po zavření dialogu.  Stav je per-session (app.storage.user).


def _refresh_ukoly_klient():
    """Překreslí seznam úkolů POUZE pro aktuálního klienta.

    `_vykresli_ukoly` je modulová @ui.refreshable, takže `.refresh()` by překreslila
    instance VŠECH připojených klientů — to ostatním uživatelům zavírá právě otevřené
    dialogy (nový úkol / detail). Proto refreshujeme jen target patřící tomuto klientovi.
    """
    try:
        klient = ui.context.client
    except Exception:
        klient = None
    if klient is None:
        _vykresli_ukoly.refresh()
        return
    _vykresli_ukoly.prune()
    for target in list(_vykresli_ukoly.targets):
        if target.container.is_deleted or target.container.client is not klient:
            continue
        # Kontejner tu ZÁMĚRNĚ nečistíme — sekce si ho smaže sama až PO načtení
        # dat (viz clear v tělech sekcí). Jinak by mezi smazáním a dokončením
        # async načítání zela prázdná stránka → bliknutí + skok scrollu nahoru.
        vysledek = target.run(_vykresli_ukoly.func)
        # async refreshable vrací coroutine — bez naplánování by se nic nevykreslilo
        if asyncio.iscoroutine(vysledek):
            background_tasks.create(vysledek)


def _safe_refresh():
    """Překreslí přehled úkolů jen pro aktuálního klienta.

    Refresh je bezpečný i s otevřeným dialogem: dialogy jsou přes `dlg.move()`
    přesunuté do kotvy MIMO refreshable kontejner, takže je překreslení nezničí.
    Žádné odkládání proto není potřeba — stav se promítne okamžitě.
    """
    _refresh_ukoly_klient()


def _refresh_refreshable_klient(rf):
    """Překreslí danou modulovou @ui.refreshable POUZE pro aktuálního klienta.

    Syrový `rf.refresh()` by překreslil instance VŠECH připojených klientů (a tím
    by jim zavíral otevřené dialogy) — používá se proto v živé synchronizaci, kde
    si každý klient ve svém timeru překresluje výhradně své vlastní sekce.
    """
    try:
        klient = ui.context.client
    except Exception:
        klient = None
    if klient is None:
        try:
            rf.refresh()
        except Exception:
            pass
        return
    try:
        rf.prune()
        for target in list(rf.targets):
            if target.container.is_deleted or target.container.client is not klient:
                continue
            # Bez clear() předem — sekce si kontejner smaže sama až s daty v ruce
            # (zamezuje bliknutí prázdné stránky a skoku scrollu během načítání).
            vysledek = target.run(rf.func)
            # async refreshable vrací coroutine — bez naplánování by se nic nevykreslilo
            if asyncio.iscoroutine(vysledek):
                background_tasks.create(vysledek)
    except Exception:
        pass


# Klient-scoped refreshe jednotlivých sekcí. Oproti nativnímu X.refresh():
#   • překreslí JEN aktuálního klienta (nezavírá dialogy ostatním — ti se
#     dozvědí o změně přes živou synchronizaci ukolovnik_data_verze),
#   • nečistí kontejner předem — sekce si ho smaže sama až s daty v ruce,
#     takže přestavba proběhne bez bliknutí a skoku scrollu.
def _porady_refresh_klient():
    _refresh_refreshable_klient(_vykresli_porady)

def _kalendar_refresh_klient():
    _refresh_refreshable_klient(_vykresli_kalendar)

def _kapacita_refresh_klient():
    _refresh_refreshable_klient(_vykresli_kapacita)

def _statistika_refresh_klient():
    _refresh_refreshable_klient(_vykresli_statistika)

def _projekty_refresh_klient():
    _refresh_refreshable_klient(_vykresli_projekty)


def _dialog_kotva(dialog_anchor=None):
    """Kontext pro VYTVOŘENÍ dialogu mimo refreshable kontejnery sekcí.

    NiceGUI si při konstrukci dialogu vloží do právě aktivního slotu
    neviditelného „kanárka" (weakref.finalize → dialog.delete()): smaže-li se
    slot, smaže se i dialog — bez ohledu na pozdější `dlg.move()`. Dialog je
    proto nutné VYTVOŘIT ve slotu, který překreslení sekcí přežije: explicitní
    kotva, jinak kořenový content klienta."""
    cil = dialog_anchor
    if cil is not None and getattr(cil, 'is_deleted', False):
        cil = None
    if cil is None:
        try:
            cil = ui.context.client.content
        except Exception:
            cil = None
    if cil is None:
        from contextlib import nullcontext
        return nullcontext()
    return cil


# Otevřené detaily per klient — po změně dat jiným uživatelem se přes tuto
# evidenci právě zobrazený náhled překreslí na místě (bez zavření okna).
_OTEVRENE_DETAILY: dict = {}   # client.id -> [{'typ', 'ref_id', 'dlg', 'refresh', 'odlozeno'}]


def _registruj_otevreny_detail(typ, ref_id, dlg, refresh):
    """Zaeviduje otevřený detail pro živou obnovu. `refresh` = bezparametrická
    funkce, která překreslí obsah dialogu na místě (okno zůstává otevřené)."""
    try:
        klient = ui.context.client
    except Exception:
        return
    try:   # úklid evidence po odpojených klientech
        from nicegui import Client as _Client
        for cid in list(_OTEVRENE_DETAILY):
            if cid not in _Client.instances:
                _OTEVRENE_DETAILY.pop(cid, None)
    except Exception:
        pass
    _OTEVRENE_DETAILY.setdefault(klient.id, []).append(
        {'typ': str(typ), 'ref_id': int(ref_id), 'dlg': dlg, 'refresh': refresh,
         'odlozeno': False})


def _ma_otevreny_vnoreny_dialog(dlg):
    """True, má-li detail otevřené podřízené okno (editace, potvrzení, …).
    Překreslení obsahu by ho zavřelo — obnova se v tom případě odkládá."""
    try:
        for el in dlg.descendants():
            if isinstance(el, ui.dialog) and el.value:
                return True
    except Exception:
        pass
    return False


def _obnov_otevrene_detaily(klient, zmeny):
    """Překreslí NA MÍSTĚ obsah otevřených detailů daného klienta, kterých se
    týkají změny [{'typ', 'ref_id'}, …]. Okno se nezavírá (zavře ho až smazání
    záznamu uvnitř _telo). Má-li detail otevřené podřízené okno, obnova se
    odloží (vlajka 'odlozeno') a dožene při dalším ticku. Zavřené/smazané
    dialogy z evidence průběžně odstraňuje."""
    zaznamy = _OTEVRENE_DETAILY.get(klient.id)
    if not zaznamy:
        return
    dotcene = {(z.get('typ'), z.get('ref_id')) for z in zmeny}
    zbyle = []
    for z in zaznamy:
        dlg = z['dlg']
        if getattr(dlg, 'is_deleted', False) or not dlg.value:
            continue   # zavřený/smazaný → jen vyčistit z evidence
        if (z['typ'], z['ref_id']) in dotcene or z.get('odlozeno'):
            if _ma_otevreny_vnoreny_dialog(dlg):
                z['odlozeno'] = True   # rozdělaná editace — nechat na později
            else:
                z['odlozeno'] = False
                try:
                    z['refresh']()
                except Exception as e:
                    print(f'[ukolovnik] živá obnova detailu selhala: {e}')
        zbyle.append(z)
    if zbyle:
        _OTEVRENE_DETAILY[klient.id] = zbyle
    else:
        _OTEVRENE_DETAILY.pop(klient.id, None)


def _dialog_guard_open():
    """Ponecháno pro zpětnou kompatibilitu (dialogy přežijí refresh přes dlg.move)."""
    pass


def _dialog_guard_close():
    """Ponecháno pro zpětnou kompatibilitu (dialogy přežijí refresh přes dlg.move)."""
    pass


# =========================================================
# RUČNÍ POŘADÍ ÚKOLŮ (drag & drop v rámci dne) — individuální per uživatel
# =========================================================
def _nacti_poradi_uziv(user_id) -> dict:
    """Vrátí {ukol_id: poradi} pro daného uživatele."""
    res = {}
    conn = intranet_data.get_db_connection()
    if not conn:
        return res
    try:
        cur = conn.cursor()
        cur.execute("SELECT ukol_id, poradi FROM ukolovnik_poradi WHERE user_id=%s", (user_id,))
        for uid, p in cur.fetchall():
            res[uid] = p
    finally:
        cur.close(); conn.close()
    return res


def _uloz_poradi_dne(user_id, ordered_ids: list):
    """Uloží pořadí (0..n) pro seznam úkolů jednoho dne pro daného uživatele."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        for idx, uid in enumerate(ordered_ids):
            cur.execute("""
                INSERT INTO ukolovnik_poradi (user_id, ukol_id, poradi)
                VALUES (%s, %s, %s)
                ON DUPLICATE KEY UPDATE poradi = VALUES(poradi)
            """, (user_id, uid, idx))
        conn.commit()
    finally:
        cur.close(); conn.close()


# =========================================================
# SEKCE: ÚKOLY — rozšířené řádky
# =========================================================
def _nacti_ukoly_seznam(user_id, is_admin, nacti_oddeleni, odd_user_ids) -> list:
    """Hlavní seznam úkolů dle viditelnosti uživatele — čistý DB loader pro vlákno."""
    conn = intranet_data.get_db_connection()
    ukoly = []
    if conn:
        try:
            cur = conn.cursor(dictionary=True)
            if is_admin:
                cur.execute("""
                    SELECT * FROM ukolovnik_ukoly
                    ORDER BY vytvoreno DESC
                """)
            elif nacti_oddeleni:
                # Vidí úkoly celého oddělení + sledované
                placeholders = ','.join(['%s'] * len(odd_user_ids))
                cur.execute(f"""
                    SELECT DISTINCT u.* FROM ukolovnik_ukoly u
                    LEFT JOIN ukolovnik_sledovaci s ON s.typ='ukol' AND s.ref_id=u.id AND s.user_id=%s
                    WHERE (u.prirazen_id IN ({placeholders})
                       OR u.zadal_id IN ({placeholders})
                       OR s.id IS NOT NULL)
                    ORDER BY
                        CASE u.stav WHEN 'Hotovo' THEN 9 WHEN 'Zrušen' THEN 10 ELSE 0 END,
                        CASE u.priorita WHEN 'Kritická' THEN 0 WHEN 'Vysoká' THEN 1 WHEN 'Normální' THEN 2 ELSE 3 END,
                        u.termin ASC, u.vytvoreno DESC
                """, [user_id] + odd_user_ids + odd_user_ids)
            else:
                cur.execute("""
                    SELECT DISTINCT u.* FROM ukolovnik_ukoly u
                    LEFT JOIN ukolovnik_sledovaci s ON s.typ='ukol' AND s.ref_id=u.id AND s.user_id=%s
                    WHERE (u.prirazen_id=%s OR u.zadal_id=%s OR s.id IS NOT NULL)
                    ORDER BY
                        CASE u.stav WHEN 'Hotovo' THEN 9 WHEN 'Zrušen' THEN 10 ELSE 0 END,
                        CASE u.priorita WHEN 'Kritická' THEN 0 WHEN 'Vysoká' THEN 1 WHEN 'Normální' THEN 2 ELSE 3 END,
                        u.termin ASC, u.vytvoreno DESC
                """, (user_id, user_id, user_id))
            ukoly = cur.fetchall()
        finally:
            cur.close(); conn.close()
    return ukoly


def _nacti_dalsi_terminy_opak(opak_ids) -> dict:
    """{opakovani_id: nejbližší budoucí termín} pro badge „Další termín"."""
    vysledek = {}
    if not opak_ids:
        return vysledek
    conn = intranet_data.get_db_connection()
    if not conn:
        return vysledek
    try:
        cur = conn.cursor(dictionary=True)
        ph = ','.join(['%s'] * len(opak_ids))
        cur.execute(f"""
            SELECT opakovani_id, MIN(termin) as dalsi
            FROM ukolovnik_ukoly
            WHERE opakovani_id IN ({ph})
              AND termin > CURDATE()
              AND stav NOT IN ('Hotovo', 'Zrušen')
            GROUP BY opakovani_id
        """, list(opak_ids))
        for r in cur.fetchall():
            vysledek[r['opakovani_id']] = r['dalsi']
    finally:
        cur.close(); conn.close()
    return vysledek


# =========================================================
# SEKCE: EXPORT ÚKOLŮ DO XLSX
# =========================================================
_EXPORT_COLS = [
    ('nazev', 'Název'), ('stav', 'Stav'), ('priorita', 'Priorita'),
    ('typ_ukolu', 'Typ'), ('oddeleni', 'Oddělení'),
    ('prirazen_jmeno', 'Přiděleno'), ('zadal_jmeno', 'Zadal'),
    ('termin', 'Termín'), ('vytvoreno', 'Vytvořeno'),
    ('dokonceno_datum', 'Dokončeno'), ('odhad_hodin', 'Odhad (h)'),
    ('projekt_nazev', 'Projekt'), ('porada_nazev', 'Porada'), ('popis', 'Popis'),
]

# Whitelist — hodnota jde do SQL bez uvozovek, nesmí přijít z klienta jinak než přes klíč.
_EXPORT_DATUM_SLOUPCE = {
    'termin':    'termin',
    'vytvoreno': 'DATE(vytvoreno)',
    'dokonceno': 'DATE(dokonceno_datum)',
}

_EXPORT_TMP_DIR = os.path.join(tempfile.gettempdir(), 'jip_ukolovnik')


def _parse_mesic(hodnota: str) -> datetime.date:
    """'2026-03' → date(2026, 3, 1). Vyhodí ValueError na nesmysl."""
    rok, mesic = (int(x) for x in str(hodnota).split('-')[:2])
    return datetime.date(rok, mesic, 1)


def _konec_mesice(d: datetime.date) -> datetime.date:
    return datetime.date(d.year, d.month, calendar.monthrange(d.year, d.month)[1])


def _export_scope_ids(user_id: int, vsechna_prava):
    """Koho smí uživatel exportovat. None = bez omezení (správce)."""
    if 'vse' in vsechna_prava or 'ukolovnik_admin' in vsechna_prava:
        return None
    lide = _ziskej_uzivatele_oddeleni(vsechna_prava)          # hlavní vedoucí oddělení
    if not lide and 'ukolovnik_ukoly_oddeleni_vse' in vsechna_prava:
        lide = _ziskej_uziv_sveho_oddeleni(user_id)
    if not lide:
        lide = _ziskej_podrizene(user_id)                     # liniový vedoucí
    ids = {u['id'] for u in lide}
    ids.add(user_id)                                          # sebe vždy
    return sorted(ids)


def _export_ukoly_data(od, do, stavy=None, osoba_id=None, oddeleni=None,
                       dle='termin', scope_ids=None) -> list:
    """Úkoly pro export. od/do = datetime.date (včetně). scope_ids=None → bez omezení."""
    kde = [f"{_EXPORT_DATUM_SLOUPCE.get(dle, 'termin')} BETWEEN %s AND %s"]
    par = [od, do]
    if stavy:
        kde.append(f"stav IN ({','.join(['%s'] * len(stavy))})")
        par += list(stavy)
    if osoba_id:
        oid = int(osoba_id)
        if scope_ids is not None and oid not in scope_ids:
            return []                                          # mimo oprávnění → nic
        kde.append("prirazen_id = %s")
        par.append(oid)
    elif scope_ids is not None:
        ph = ','.join(['%s'] * len(scope_ids))
        kde.append(f"(prirazen_id IN ({ph}) OR zadal_id IN ({ph}))")
        par += list(scope_ids) + list(scope_ids)

    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(
            f"SELECT {', '.join(c for c, _ in _EXPORT_COLS)} FROM ukolovnik_ukoly "
            f"WHERE {' AND '.join(kde)} ORDER BY termin, vytvoreno", par)
        radky = cur.fetchall()
    finally:
        cur.close(); conn.close()

    if oddeleni:
        # sloupec je CSV s mezerami („Obchod, IT") → porovnání v Pythonu, ne FIND_IN_SET
        radky = [r for r in radky
                 if oddeleni in _oddeleni_chips(r.get('oddeleni'), max_chips=99)]
    return radky


def _export_popis_parametru(od, do, stavy=None, osoba_jmeno=None, oddeleni=None,
                            dle='termin', pocet=None, kdo=None) -> list:
    """Dvojice (popisek, hodnota) do hlavičky sešitu — co přesně je v exportu."""
    dle_lbl = {'termin': 'termín úkolu', 'vytvoreno': 'datum vytvoření',
               'dokonceno': 'datum dokončení'}.get(dle, dle)
    radky = [
        ('Období', f'{MESICE_CZ[od.month]} {od.year} – {MESICE_CZ[do.month]} {do.year}'
                   f'  ({od:%d.%m.%Y} – {do:%d.%m.%Y}, dle: {dle_lbl})'),
        ('Stav úkolů', ', '.join(stavy) if stavy else 'všechny'),
        ('Oddělení', oddeleni or 'všechna'),
        ('Osoba (přiděleno)', osoba_jmeno or 'všichni'),
        ('Počet úkolů', pocet),
        ('Exportoval', f'{kdo or "?"} — {datetime.datetime.now():%d.%m.%Y %H:%M}'),
    ]
    return radky


def _export_ukoly_xlsx(radky: list, cesta: str, parametry: list = None) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    ws.title = 'Úkoly'

    # Hlavička s parametry exportu; tabulka začíná až pod ní (offset drží auto_filter/freeze).
    tucne = Font(bold=True)
    if parametry:
        nadpis = ws.cell(1, 1, 'Export úkolů z Úkolovníku')
        nadpis.font = Font(bold=True, size=13)
        for i, (lbl, hodnota) in enumerate(parametry, start=2):
            ws.cell(i, 1, f'{lbl}:').font = tucne
            ws.cell(i, 2, hodnota if hodnota is not None else '')
        prvni_radek = len(parametry) + 3          # +1 nadpis, +1 prázdný, +1 hlavička
    else:
        prvni_radek = 1

    for i, (_, lbl) in enumerate(_EXPORT_COLS, 1):
        ws.cell(prvni_radek, i, lbl).font = tucne
    ws.freeze_panes = f'A{prvni_radek + 1}'
    for r in radky:
        ws.append([r.get(k) for k, _ in _EXPORT_COLS])

    posledni = prvni_radek + len(radky)
    for i, (k, lbl) in enumerate(_EXPORT_COLS, 1):
        pismeno = ws.cell(prvni_radek, i).column_letter
        ws.column_dimensions[pismeno].width = 60 if k == 'popis' else max(12, min(32, len(lbl) + 8))
        if k in ('termin', 'vytvoreno', 'dokonceno_datum'):
            for radek in range(prvni_radek + 1, posledni + 1):
                ws.cell(radek, i).number_format = 'DD.MM.YYYY'
    ws.auto_filter.ref = f'A{prvni_radek}:{ws.cell(prvni_radek, len(_EXPORT_COLS)).column_letter}{posledni}'
    wb.save(cesta)
    return cesta


def _export_tmp_cesta(jmeno: str) -> str:
    """Cesta v temp adresáři; při každém exportu smaže staré soubory (obsahují firemní data)."""
    os.makedirs(_EXPORT_TMP_DIR, exist_ok=True)
    ted = time.time()
    for f in os.listdir(_EXPORT_TMP_DIR):
        stara = os.path.join(_EXPORT_TMP_DIR, f)
        try:
            if os.path.isfile(stara) and ted - os.path.getmtime(stara) > 1800:
                os.remove(stara)
        except OSError:
            pass
    return os.path.join(_EXPORT_TMP_DIR, f'{int(ted * 1000)}_{jmeno}')


def _dialog_export_ukoly(user_id, user_name, vsechna_prava, predvolby=None):
    """Dialog: období měsíců + stav + oddělení/osoba → .xlsx ke stažení."""
    predvolby = predvolby or {}
    is_admin = 'vse' in vsechna_prava or 'ukolovnik_admin' in vsechna_prava
    dnes = datetime.date.today()
    ted_mesic = f'{dnes:%Y-%m}'

    osoba_opts = {'': 'Vše'}
    for uid, jm in _ziskej_viditelne_uzivatele(user_id, vsechna_prava).items():
        osoba_opts[str(uid)] = jm
    odd_opts = {'': 'Vše', **_ziskej_oddeleni_options()} if is_admin else {}

    with ui.dialog() as dlg, ui.card().classes('w-[540px] max-w-full p-5 gap-3'):
        ui.label('Export úkolů do Excelu').classes('text-lg font-bold text-gray-800')

        with ui.row().classes('w-full gap-2 no-wrap'):
            od_in = ui.input('Od měsíce', value=ted_mesic).props('type=month dense outlined').classes('flex-1')
            do_in = ui.input('Do měsíce', value=ted_mesic).props('type=month dense outlined').classes('flex-1')
        dle_in = ui.select({'termin': 'Termín úkolu', 'vytvoreno': 'Datum vytvoření',
                            'dokonceno': 'Datum dokončení'},
                           value='termin', label='Období počítat dle').classes('w-full').props('dense outlined')

        _stav_pred = [predvolby['stav']] if predvolby.get('stav') in STAVY_UKOL else []
        stav_in = ui.select(list(STAVY_UKOL.keys()), multiple=True, value=_stav_pred,
                            label='Stav (prázdné = všechny)').classes('w-full').props('dense outlined use-chips')

        odd_in = None
        if odd_opts:
            odd_in = ui.select(odd_opts, value=predvolby.get('odd') or '',
                               label='Oddělení').classes('w-full').props('dense outlined')
        _osoba_pred = predvolby.get('osoba') or ''
        osoba_in = ui.select(osoba_opts, value=_osoba_pred if _osoba_pred in osoba_opts else '',
                             label='Osoba (přiděleno)').classes('w-full').props('dense outlined')

        stav_lbl = ui.label('').classes('text-xs text-gray-500')

        async def _spust():
            try:
                od = _parse_mesic(od_in.value)
                do = _konec_mesice(_parse_mesic(do_in.value))
            except (ValueError, AttributeError, TypeError):
                ui.notify('Vyber platné období (měsíc od / do).', type='warning'); return
            if do < od:
                ui.notify('Konec období je před začátkem.', type='warning'); return

            stav_lbl.text = 'Připravuji export…'
            scope = await asyncio.to_thread(_export_scope_ids, user_id, vsechna_prava)
            radky = await asyncio.to_thread(
                _export_ukoly_data, od, do, list(stav_in.value or []),
                osoba_in.value or None, (odd_in.value if odd_in else None) or None,
                dle_in.value, scope)
            if not radky:
                stav_lbl.text = ''
                ui.notify('Žádné úkoly neodpovídají zvolenému filtru.', type='info'); return

            parametry = _export_popis_parametru(
                od, do, list(stav_in.value or []),
                osoba_opts.get(str(osoba_in.value or '')) if osoba_in.value else None,
                (odd_in.value if odd_in else None) or None,
                dle_in.value, len(radky), user_name)

            jmeno = f'ukoly_{od:%Y-%m}_{do:%Y-%m}.xlsx'
            cesta = await asyncio.to_thread(_export_tmp_cesta, jmeno)
            await asyncio.to_thread(_export_ukoly_xlsx, radky, cesta, parametry)
            # Přes HTTP, ne WebSocket — ui.download.content padá nad ~1 MB na engine.io limitu.
            ui.download.file(cesta, jmeno)
            intranet_logger.log_activity(
                user_name, 'Úkolovník',
                f'Export úkolů {od:%Y-%m}–{do:%Y-%m} ({len(radky)} řádků)')
            stav_lbl.text = ''
            dlg.close()

            async def _uklid():
                await asyncio.sleep(300)
                try:
                    os.remove(cesta)
                except OSError:
                    pass
            background_tasks.create(_uklid())

        with ui.row().classes('w-full justify-end gap-2 mt-1'):
            ui.button('Zrušit', on_click=dlg.close).props('flat color=grey')
            ui.button('Stáhnout XLSX', icon='download', on_click=_spust).props('color=green-7 unelevated')
    dlg.open()


@ui.refreshable
async def _vykresli_ukoly(user_id, user_name, vsechna_prava, dialog_anchor=None, ukoly_stav=None):
    is_admin = 'vse' in vsechna_prava or 'ukolovnik_admin' in vsechna_prava
    # Stav pohledu (stránka, řazení, sbalené dny, pohled) drží rodič → přežije refresh
    if ukoly_stav is None:
        ukoly_stav = {}
    _rezim = ukoly_stav.setdefault('rezim', {'kanban': True})  # výchozí kanban; False = seznam

    # Uživatelé oddělení pro vedoucí
    uzivatele_oddeleni = [] if is_admin else await asyncio.to_thread(_ziskej_uzivatele_oddeleni, vsechna_prava)
    je_vedouci_uk = is_admin or _je_hlavni_vedouci(vsechna_prava)
    # Právo „vidět všechny úkoly v oddělení" (jen čtení, bez manažerských funkcí)
    vidi_cele_oddeleni = (not is_admin) and ('ukolovnik_ukoly_oddeleni_vse' in vsechna_prava)
    # IDs uživatelů v oddělení (pro vedoucí) — zahrnuje samotného uživatele
    if uzivatele_oddeleni:
        odd_user_ids = [u['id'] for u in uzivatele_oddeleni]
    elif vidi_cele_oddeleni:
        odd_user_ids = [u['id'] for u in await asyncio.to_thread(_ziskej_uziv_sveho_oddeleni, user_id)] or [user_id]
    else:
        odd_user_ids = [user_id]
    # Načti úkoly celého oddělení, pokud je uživatel vedoucí NEBO má právo na čtení oddělení
    nacti_oddeleni = (je_vedouci_uk and bool(uzivatele_oddeleni)) or vidi_cele_oddeleni
    # Kdo vidí úkoly i jiných lidí → má smysl detailní sloupec „Přiděleno" (jako správce)
    muze_videt_prideleno = is_admin or nacti_oddeleni

    # DB dotazy běží ve vlákně — nedrží event loop celého serveru
    ukoly = await asyncio.to_thread(_nacti_ukoly_seznam, user_id, is_admin, nacti_oddeleni, odd_user_ids)

    # Nejbližší další termín opakovaného úkolu (pro badge „Další termín" u každého výskytu)
    opak_ids = set(u.get('opakovani_id') for u in ukoly if u.get('opakovani_id'))
    _dalsi_termin_opak = await asyncio.to_thread(_nacti_dalsi_terminy_opak, opak_ids)

    # Úkoly s právě běžícím časovačem uživatele — jedním dotazem pro celý seznam
    # (nahrazuje dřívější dotaz per řádek)
    _aktivni_timer_ids = await asyncio.to_thread(_ziskej_aktivni_ukol_ids, user_id)

    dnes = datetime.date.today()

    # Individuální ruční pořadí úkolů tohoto uživatele (drag & drop v rámci dne)
    _poradi_uziv = await asyncio.to_thread(_nacti_poradi_uziv, user_id)

    # Podřízení pro filtr osob — vedoucí vidí lidi svého oddělení (bez sebe)
    podrizeni_uk = [u for u in uzivatele_oddeleni if u['id'] != user_id] if je_vedouci_uk else []
    if is_admin and not podrizeni_uk:
        uziv_all = await asyncio.to_thread(_ziskej_uzivatele_options)
        podrizeni_uk = [{'id': uid, 'jmeno': jm} for uid, jm in uziv_all.items() if uid != user_id]

    # Oddělení pro filtr správce — načíst PŘED stavbou UI (atomická přestavba)
    odd_opts_admin = (await asyncio.to_thread(_ziskej_oddeleni_options)) if is_admin else {}

    # Konec datové fáze: starý obsah smažeme až teď, s daty v ruce — smazání
    # i postavení odejde klientovi v jednom kroku → žádné bliknutí ani skok
    # scrollu (refresh helpery kontejner předem nečistí).
    ui.context.slot.parent.clear()

    with ui.row().classes('w-full items-center justify-between mb-3'):
        with ui.column().classes('gap-0'):
            nadpis = ui.label('Přehled úkolů').classes('text-3xl font-extrabold text-gray-800')
            podnadpis = ui.label('Všechny přidělené i zadané úkoly, vč. úkolů z porad.').classes('text-gray-500 text-sm')
        with ui.row().classes('items-center gap-2'):
            btn_hotove = ui.button('Hotové / zrušené úkoly', icon='check_circle',
                      on_click=lambda: _prepni_hotove()
                      ).props('flat').classes('text-gray-600 font-bold h-10 px-4 rounded-xl border border-gray-200 hover:bg-gray-50')
            btn_novy = ui.button('Nový úkol', icon='add_task',
                      on_click=lambda: _dialog_novy_ukol(user_id, user_name, vsechna_prava, on_refresh=_safe_refresh, dialog_anchor=dialog_anchor)
                      ).classes('bg-green-600 text-white font-bold h-10 px-5 rounded-xl shadow-md')

    # ── FILTRY ──────────────────────────────────────────────────────────────────
    # Hodnoty filtrů drží rodič (ukoly_stav) — stejně jako stránka, řazení a pohled.
    # Bez toho by je změna stavu úkolu (→ _safe_refresh → přestavba sekce) shodila
    # zpět na výchozí, včetně termínu zpět na „Dnes".
    _filtry = ukoly_stav.setdefault('filtry', {
        'stav': '', 'typ': '', 'termin': 'dnes', 'zdroj': '',
        'moje': '', 'osoba': '', 'odd': ''})
    with ui.row().classes('w-full gap-2 mb-1 flex-wrap items-center'):
        filtr_stav = ui.select(
            {'': 'Stav: Vše', 'Zadáno': 'Zadáno', 'Rozpracováno': 'Rozpracováno',
             'Pozastaveno': 'Pozastaveno'},
            value=_filtry['stav'], label=''
        ).classes('w-44').props('dense outlined')
        filtr_typ = ui.select(
            {'': 'Typ: Vše', 'Pravidelný': '🔄 Pravidelný', 'Mimořádný': '⚡ Mimořádný'},
            value=_filtry['typ'], label=''
        ).classes('w-40').props('dense outlined')
        filtr_termin = ui.select(
            {'': 'Termín: Vše', 'po': '🔴 Po termínu', 'dnes': '🟡 Dnes', 'tyden': '🟠 Tento týden', 'budouci': '🟢 Budoucí'},
            value=_filtry['termin'], label=''
        ).classes('w-44').props('dense outlined')
        filtr_zdroj = ui.select(
            {'': 'Zdroj: Vše', 'porada': 'Z porady', 'samostatne': 'Samostatné'},
            value=_filtry['zdroj'], label=''
        ).classes('w-40').props('dense outlined')

    with ui.row().classes('w-full gap-2 mb-3 flex-wrap items-center'):
        filtr_moje = ui.select(
            {'': 'Přiřazení: Vše', 'prirazen': 'Přiděleno mně', 'zadal': 'Zadal jsem'},
            value=_filtry['moje'], label=''
        ).classes('w-44').props('dense outlined')
        # Filtr osoby — jen pro vedoucí
        if je_vedouci_uk:
            osoba_opts_uk = {'': 'Osoba: Vše'}
            for p in podrizeni_uk:
                osoba_opts_uk[str(p['id'])] = p['jmeno']
            # uložená osoba už nemusí být v podřízených (změna struktury) → fallback Vše
            filtr_osoba = ui.select(osoba_opts_uk,
                                    value=_filtry['osoba'] if _filtry['osoba'] in osoba_opts_uk else '',
                                    label='').classes('w-44').props('dense outlined')
        else:
            filtr_osoba = None
        # Filtr oddělení — pouze pro správce (načteno v datové fázi)
        odd_opts = odd_opts_admin
        if odd_opts:
            filtr_odd = ui.select({'': 'Oddělení: Vše', **odd_opts},
                                  value=_filtry['odd'] if _filtry['odd'] in odd_opts else '',
                                  label='').classes('w-48').props('dense outlined')
        else:
            filtr_odd = None

        ui.button('Export', icon='download', on_click=lambda: _dialog_export_ukoly(
            user_id, user_name, vsechna_prava,
            {'stav': filtr_stav.value,
             'osoba': filtr_osoba.value if filtr_osoba else '',
             'odd': filtr_odd.value if filtr_odd else ''})
        ).props('flat dense color=green-7').classes('text-xs')

        ui.element('div').classes('flex-1')

        def _prepni_rezim():
            _rezim['kanban'] = not _rezim['kanban']
            _aplikuj_rezim_btn()
            prekresli()

        btn_rezim = ui.button(on_click=_prepni_rezim).props('outline size=sm')

        def _aplikuj_rezim_btn():
            btn_rezim.text = 'Seznam' if _rezim['kanban'] else 'Kanban'
            btn_rezim.props('icon=view_list' if _rezim['kanban'] else 'icon=view_kanban')
            btn_rezim.update()

        _aplikuj_rezim_btn()

    # Hlavičková řádka tabulky
    def _ikona_razeni(pole):
        if _razeni['pole'] != pole: return 'unfold_more'
        return 'keyboard_arrow_up' if _razeni['smer'] == 'asc' else 'keyboard_arrow_down'

    def _prepni_razeni(pole):
        # Cyklus: výchozí (ruční pořadí) → vzestupně → sestupně → zpět na výchozí
        if _razeni['pole'] != pole:
            _razeni['pole'] = pole
            _razeni['smer'] = 'asc'
        elif _razeni['smer'] == 'asc':
            _razeni['smer'] = 'desc'
        else:
            _razeni['pole'] = None   # zpět na výchozí zobrazení → znovu povolí drag&drop
            _razeni['smer'] = 'asc'
        _stranka['p'] = 1
        prekresli()

    hlavicka_row = ui.row().classes('w-full items-center border-b border-gray-200 mb-1')

    seznam_box = ui.column().classes('w-full gap-1')

    def _filtruj_termin(u, ft):
        if not ft: return True
        t = u.get('termin')
        if ft == 'po':
            return t and (t - dnes).days < 0 and u['stav'] not in ('Hotovo', 'Zrušen')
        if ft == 'dnes':
            return t and t == dnes
        if ft == 'tyden':
            return t and 0 <= (t - dnes).days <= 7
        if ft == 'budouci':
            return t and (t - dnes).days > 7
        return True

    STRANKA_VEL = 30
    # Perzistentní stav napříč refreshi (uložení úkolu nevrací uživatele na 1. stránku)
    _stranka = ukoly_stav.setdefault('stranka', {'p': 1})
    _razeni = ukoly_stav.setdefault('razeni', {'pole': None, 'smer': 'asc'})  # pole: None|'termin'|'prirazen'
    _sbalene_dny: set = ukoly_stav.setdefault('sbalene_dny', set())
    _inicializovano = ukoly_stav.setdefault('inicializovano', {'hodnota': False})
    _pohled = ukoly_stav.setdefault('pohled', {'hotove': False})  # False = přehled (bez hotových/zrušených), True = archiv

    # Drag & drop přeřazení v rámci dne (jen ve výchozím zobrazení dle termínu)
    _dragged = {'id': None, 'termin': None}
    _zobr = {'list': []}  # poslední vykreslený seznam (pro reorder)

    def _on_drop(cilovy_uk):
        src_id = _dragged.get('id')
        src_termin = _dragged.get('termin')
        _dragged['id'] = None
        if not src_id or src_id == cilovy_uk['id']:
            return
        if src_termin != cilovy_uk.get('termin'):
            ui.notify('Úkoly lze přesouvat jen v rámci stejného dne.', type='warning', position='top')
            return
        termin = cilovy_uk.get('termin')
        den_ids = [u['id'] for u in _zobr['list'] if u.get('termin') == termin]
        if src_id not in den_ids or cilovy_uk['id'] not in den_ids:
            return
        den_ids.remove(src_id)
        cil_idx = den_ids.index(cilovy_uk['id'])
        den_ids.insert(cil_idx, src_id)  # vlož před cílový úkol
        _uloz_poradi_dne(user_id, den_ids)
        for i, uid in enumerate(den_ids):
            _poradi_uziv[uid] = i
        prekresli()

    # ── Nový model: živé per-úkol karty + kanban ──────────────────────
    _karty_reg: dict = {}   # ukol_id → (obal, termin, stav) právě vykreslených karet
    _stav_render = {'drag': True, 'prid': 'detail'}   # stav posledního prekresli() pro rebuild karty

    def _vykresli_kartu_telo(uk, termin_obj):
        b_stav_cls, _ = STAVY_UKOL.get(uk['stav'], ('bg-gray-100 text-gray-700 border-gray-300', '#ccc'))
        je_hotovo  = uk['stav'] == 'Hotovo'
        je_zrusen  = uk['stav'] == 'Zrušen'
        je_muj     = uk.get('prirazen_id') == user_id
        termin_str = termin_obj.strftime('%d.%m.%Y') if termin_obj else '—'
        delta, dt_cls, dt_txt = _dny_do_terminu(termin_obj)
        po_terminu = delta is not None and delta < 0 and not je_hotovo
        aktivni_timer = uk['id'] in _aktivni_timer_ids
        typ_uk = uk.get('typ_ukolu', 'Pravidelný')
        odd_chips = _oddeleni_chips(uk.get('oddeleni') or '', max_chips=2)

        # Levý pruh
        if je_hotovo:        barva_p = 'bg-green-400'
        elif po_terminu:     barva_p = 'bg-red-500'
        elif aktivni_timer:  barva_p = 'bg-orange-400'
        elif je_muj:         barva_p = 'bg-blue-400'
        else:                barva_p = 'bg-gray-300'

        opacity_cls = 'opacity-60' if je_hotovo or je_zrusen else ''

        # Zbývá / Dokončení
        if je_hotovo and uk.get('dokonceno_datum'):
            dok_d = uk['dokonceno_datum'].date() if hasattr(uk['dokonceno_datum'], 'date') else uk['dokonceno_datum']
            zbyvá_txt = f"✅ {uk['dokonceno_datum'].strftime('%d.%m.%Y')}"
            if termin_obj:
                diff = (dok_d - termin_obj).days
                if diff > 0:
                    zbyvá_cls = 'text-red-600 font-bold'
                    zbyvá_txt += f' (+{diff}d)'
                elif diff == 0:
                    zbyvá_cls = 'text-green-600 font-bold'
                else:
                    zbyvá_cls = 'text-green-600 font-bold'
                    zbyvá_txt += f' (-{abs(diff)}d)'
            else:
                zbyvá_cls = 'text-green-600 font-bold'
        elif je_hotovo:
            zbyvá_txt, zbyvá_cls = '✅ Hotovo', 'text-green-600 font-bold'
        else:
            zbyvá_txt = dt_txt if termin_obj else ''
            zbyvá_cls = dt_cls

        _karta = ui.card().classes(
            f'w-full p-0 overflow-hidden rounded-xl border border-gray-100 bg-white '
            f'hover:shadow-md transition-shadow cursor-pointer {opacity_cls}'
        ).on('click', lambda u=uk: _dialog_detail_ukolu(u['id'], user_id, user_name, vsechna_prava, on_refresh=_safe_refresh, dialog_anchor=dialog_anchor))
        if _stav_render['drag']:
            # Karta je cílová zóna pro upuštění (drop)
            _karta.on('dragover.prevent', lambda: None)
            _karta.on('drop', lambda e, u=uk: _on_drop(u))
        with _karta:
            with ui.row().classes('w-full items-stretch'):
                ui.element('div').classes(f'w-1.5 shrink-0 self-stretch {barva_p}')

                with ui.column().classes('flex-1 gap-0'):

                    # ── Řádek 1: hlavní info ──────────────────────────
                    with ui.row().classes('w-full px-4 py-3 items-center gap-3'):

                        # Úchyt pro přetažení (drag) — jen ve výchozím zobrazení
                        if _stav_render['drag']:
                            _grip = ui.element('div').classes(
                                'cursor-move shrink-0 flex items-center text-gray-300 hover:text-blue-500'
                            ).props('draggable=true')
                            _grip.on('dragstart', lambda e, u=uk: _dragged.update(id=u['id'], termin=u.get('termin')))
                            _grip.on('dragend', lambda: _dragged.update(id=None))
                            _grip.on('click.stop', lambda: None)
                            with _grip:
                                ui.icon('drag_indicator', size='sm')
                                ui.tooltip('Přetažením změníte pořadí v rámci dne')

                        # Název + zdroj
                        with ui.column().classes('flex-1 min-w-0 gap-0.5'):
                            nazev_cls = 'text-base font-bold text-gray-800' + (' line-through text-gray-400' if je_hotovo else '')
                            ui.label(uk['nazev']).classes(f'{nazev_cls} line-clamp-1')
                            with ui.row().classes('gap-1 flex-wrap items-center'):
                                if uk.get('porada_id'):
                                    pd = uk['porada_datum'].strftime('%d.%m.%Y') if uk.get('porada_datum') else ''
                                    ui.label(f'📋 Porada · {pd}').classes('text-xs text-blue-500 font-bold')
                                elif uk.get('projekt_nazev'):
                                    ui.label(f'🚀 {uk["projekt_nazev"]}').classes('text-xs text-indigo-500 font-bold')
                                for _odd in odd_chips:
                                    ui.label(_odd).classes('text-[10px] bg-indigo-50 text-indigo-600 border border-indigo-200 px-1.5 py-0 rounded-full font-bold')
                                # Další termín opakovaného úkolu
                                _dt_next = _dalsi_termin_opak.get(uk.get('opakovani_id')) if uk.get('opakovani_id') else None
                                if _dt_next and _dt_next != termin_obj:
                                    ui.label(f'🔁 Další termín: {_dt_next.strftime("%d.%m.%Y")}').classes(
                                        'text-[10px] bg-purple-50 text-purple-600 border border-purple-200 px-1.5 py-0 rounded-full font-bold'
                                    )

                        # Přiděleno
                        if _stav_render['prid'] == 'detail':
                            with ui.column().classes('w-44 shrink-0 gap-0'):
                                with ui.row().classes('items-center gap-1'):
                                    ui.icon('person', size='xs').classes('text-blue-400 shrink-0')
                                    ui.label(uk.get('prirazen_jmeno', '—')).classes('text-sm font-bold text-blue-700 truncate')
                                if uk.get('zadal_jmeno') and uk.get('zadal_jmeno') != uk.get('prirazen_jmeno'):
                                    ui.label(f"Zadal: {uk['zadal_jmeno']}").classes('text-xs text-gray-400 truncate')
                        elif _stav_render['prid'] == 'simple':
                            ui.label(uk.get('prirazen_jmeno', '—')).classes(
                                ('text-sm font-bold text-blue-700' if je_muj else 'text-sm text-gray-500') +
                                ' w-36 shrink-0 truncate hidden md:block'
                            )
                        # _prid_mode == 'none' → sloupec „Přiděleno" se nevykreslí

                        # Termín
                        ui.label(termin_str).classes(
                            ('text-sm font-bold text-red-600' if po_terminu else 'text-sm text-gray-600') +
                            ' w-32 shrink-0'
                        )

                        # Zbývá / Dokončení
                        ui.label(zbyvá_txt).classes(f'text-sm {zbyvá_cls} w-40 shrink-0')

                        # Ikony
                        with ui.row().classes('items-center gap-0.5 shrink-0 w-20 justify-end').on('click.stop', lambda: None):
                            if aktivni_timer:
                                ui.icon('timer', size='sm', color='orange').classes('shrink-0')
                            ui.button(icon='visibility').props('flat round dense size=sm').classes(
                                'text-gray-300 hover:text-blue-500 shrink-0'
                            ).on('click.stop', lambda u=uk: _dialog_log('ukol', u['id'], u['nazev'], user_id, user_name))
                            if is_admin or uk.get('zadal_id') == user_id:
                                def _del_u(u=uk):
                                    def _provest():
                                        ok = _smazat_ukol_db(u['id'])
                                        if ok:
                                            intranet_logger.log_activity(user_name, "Úkolovník", f"Smazán úkol #{u['id']}: {u['nazev']}")
                                            ui.notify('Úkol smazán.', type='positive', position='top')
                                            _safe_refresh()
                                        else:
                                            ui.notify('Chyba při mazání.', type='negative', position='top')
                                    _potvrdit_smazat(f'Opravdu smazat úkol „{u["nazev"]}"?\nTato akce je nevratná.', _provest)
                                ui.button(icon='delete_outline').props('flat round dense size=sm').classes(
                                    'text-gray-300 hover:text-red-500 shrink-0'
                                ).on('click.stop', lambda u=uk: _del_u(u))

                    # ── Řádek 2: stavová tlačítka + typ ───────────────
                    muze_stav = (is_admin or uk.get('prirazen_id') == user_id or uk.get('zadal_id') == user_id) and not je_hotovo and not je_zrusen
                    with ui.row().classes('w-full px-4 pb-2 pt-0 items-center gap-1.5 border-t border-gray-50').on('click.stop', lambda: None):
                        for _sn in STAVY_UKOL:
                            _sc, _ = STAVY_UKOL[_sn]
                            _akt = uk['stav'] == _sn
                            if muze_stav:
                                _btn_cls = (
                                    f'{_sc} border font-extrabold cursor-pointer'
                                    if _akt else
                                    'bg-white text-gray-300 border border-gray-200 cursor-pointer hover:bg-gray-50 hover:text-gray-600'
                                )
                                with ui.element('div').classes(
                                    f'px-2.5 py-1 rounded text-xs {_btn_cls} select-none'
                                ).on('click.stop', lambda sn=_sn, u=uk: _zmenit_stav_ukolu(u, sn, user_id, user_name)):
                                    ui.label(_sn)
                            elif _akt:
                                with ui.element('div').classes(f'px-2.5 py-1 rounded text-xs font-bold border {_sc}'):
                                    ui.label(_sn)
                        ui.element('div').classes('flex-1')
                        if typ_uk == 'Mimořádný':
                            ui.label('⚡ Mimořádný').classes('text-xs font-extrabold text-orange-700 bg-orange-50 border border-orange-300 px-2 py-0.5 rounded')
                        else:
                            ui.label('🔄 Pravidelný').classes('text-xs font-bold text-blue-600 bg-blue-50 border border-blue-200 px-2 py-0.5 rounded')

    def _vykresli_kartu(uk, termin_obj):
        _obal = ui.element('div').classes('w-full')
        _karty_reg[uk['id']] = (_obal, termin_obj, uk['stav'])
        with _obal:
            _vykresli_kartu_telo(uk, termin_obj)

    def _vykresli_kanban(zobrazene):
        """Kanban nástěnka — sloupce dle stavu, přetažením karty se mění stav."""
        _kb_tazeny = {'id': None}

        def _dok_datum(u):
            d = u.get('dokonceno_datum')
            if not d:
                return None
            return d.date() if hasattr(d, 'date') else d

        # Sloupec Hotovo: dokončené za posledních 14 dní
        # (ponytail: bez ostatních filtrů — jednodušší a pro nástěnku stačí)
        _ned = dnes - datetime.timedelta(days=14)
        hotove_ned = [u for u in ukoly if u['stav'] == 'Hotovo' and (_dok_datum(u) or datetime.date.min) >= _ned]
        sloupce = [('Zadáno', zobrazene), ('Rozpracováno', zobrazene), ('Pozastaveno', zobrazene), ('Hotovo', hotove_ned)]

        with seznam_box:
            with ui.row().classes('w-full gap-3 items-stretch flex-nowrap overflow-x-auto pb-2 mt-2'):
                for _st, _zdroj in sloupce:
                    _sc, _hex = STAVY_UKOL.get(_st, ('bg-gray-100 text-gray-700 border-gray-300', '#ccc'))
                    _kus = [u for u in _zdroj if u['stav'] == _st]

                    def _kb_drop(st=_st):
                        uid = _kb_tazeny.get('id')
                        _kb_tazeny['id'] = None
                        if not uid:
                            return
                        u = next((x for x in ukoly if x['id'] == uid), None)
                        if u is None or u['stav'] == st:
                            return
                        if not (is_admin or u.get('prirazen_id') == user_id or u.get('zadal_id') == user_id):
                            ui.notify('Stav může měnit jen řešitel, zadavatel nebo správce.', type='warning', position='top')
                            return
                        _zmenit_stav_ukolu(u, st, user_id, user_name)

                    _sl = ui.column().classes('flex-1 min-w-[260px] gap-2 rounded-xl bg-gray-50 border border-gray-200 p-2 min-h-[280px]')
                    _sl.on('dragover.prevent', lambda: None)
                    _sl.on('drop', lambda e, st=_st: _kb_drop(st))
                    with _sl:
                        with ui.row().classes('w-full items-center gap-2 px-1 pt-1'):
                            ui.element('div').classes('w-2.5 h-2.5 rounded-full').style(f'background:{_hex}')
                            ui.label(_st).classes('text-sm font-extrabold text-gray-700 flex-1')
                            ui.badge(str(len(_kus)), color='grey-4').props('rounded').classes('text-gray-700')
                        for u in _kus:
                            _delta, _dt_cls, _dt_txt = _dny_do_terminu(u.get('termin'))
                            _po_t = _delta is not None and _delta < 0 and u['stav'] != 'Hotovo'
                            _kk = ui.card().classes('w-full p-3 gap-1 rounded-lg border border-gray-200 bg-white hover:shadow-md cursor-pointer')
                            _kk.props('draggable=true')
                            _kk.on('dragstart', lambda e, uu=u: _kb_tazeny.update(id=uu['id']))
                            _kk.on('dragend', lambda: _kb_tazeny.update(id=None))
                            _kk.on('click', lambda uu=u: _dialog_detail_ukolu(uu['id'], user_id, user_name, vsechna_prava, on_refresh=_safe_refresh, dialog_anchor=dialog_anchor))
                            with _kk:
                                ui.label(u['nazev']).classes('text-sm font-bold text-gray-800 line-clamp-2')
                                with ui.row().classes('w-full items-center gap-2'):
                                    ui.icon('person', size='xs').classes('text-blue-400')
                                    ui.label(u.get('prirazen_jmeno') or '—').classes('text-xs text-blue-700 font-bold truncate flex-1')
                                    if u.get('termin'):
                                        ui.label(u['termin'].strftime('%d.%m.')).classes('text-xs font-bold ' + ('text-red-600' if _po_t else 'text-gray-500'))
                                if u.get('porada_id'):
                                    ui.label('📋 Porada').classes('text-[10px] text-blue-500 font-bold')
                                elif u.get('projekt_nazev'):
                                    ui.label(f'🚀 {u["projekt_nazev"]}').classes('text-[10px] text-indigo-500 font-bold')
                        if not _kus:
                            ui.label('Přetáhněte sem úkol').classes('text-xs text-gray-300 italic mx-auto my-4')

    def prekresli():
        hlavicka_row.clear()
        seznam_box.clear()
        _karty_reg.clear()
        fs  = filtr_stav.value
        ft  = filtr_termin.value
        fz  = filtr_zdroj.value
        fm  = filtr_moje.value
        fty = filtr_typ.value
        fo  = filtr_osoba.value if filtr_osoba else ''
        fodd = filtr_odd.value if filtr_odd else ''

        # Režim sloupce „Přiděleno":
        #   'detail' = ikona + jméno + „Zadal:" (správce nebo právo na oddělení)
        #   'simple' = jednoduchý štítek se jménem (běžný uživatel)
        #   'none'   = skryto (má právo na oddělení, ale filtruje „Přiděleno mně")
        if muze_videt_prideleno:
            _prid_mode = 'none' if fm == 'prirazen' else 'detail'
        else:
            _prid_mode = 'simple'

        with hlavicka_row:
            ui.element('div').classes('w-1.5 shrink-0')
            with ui.row().classes('flex-1 px-4 py-1.5 items-center gap-3 text-xs font-bold text-gray-400 uppercase tracking-wider'):
                ui.label('Název úkolu').classes('flex-1')
                if _prid_mode != 'none':
                    _p_cls = ('w-44' if _prid_mode == 'detail' else 'w-36 hidden md:block') + ' shrink-0'
                    _akt_p = _razeni['pole'] == 'prirazen'
                    with ui.row().classes(f'{_p_cls} items-center gap-0.5 cursor-pointer hover:text-gray-600 select-none ' + ('text-blue-600' if _akt_p else '')).on('click', lambda: _prepni_razeni('prirazen')):
                        ui.label('Přiděleno')
                        ui.icon(_ikona_razeni('prirazen'), size='xs')
                _akt_t = _razeni['pole'] == 'termin'
                with ui.row().classes('w-32 shrink-0 items-center gap-0.5 cursor-pointer hover:text-gray-600 select-none ' + ('text-blue-600' if _akt_t else '')).on('click', lambda: _prepni_razeni('termin')):
                    ui.label('Termín')
                    ui.icon(_ikona_razeni('termin'), size='xs')
                ui.label('Zbývá / Dokončení').classes('w-40 shrink-0')
                ui.label('').classes('w-20 shrink-0')

        # Pohled „Hotové / zrušené" → jen dokončené a zrušené; výchozí přehled → bez nich
        zobrazene = [u for u in ukoly if
            ((u['stav'] in ('Hotovo', 'Zrušen')) if _pohled['hotove'] else (u['stav'] not in ('Hotovo', 'Zrušen'))) and
            (not fs  or u['stav'] == fs) and
            (not fty or u.get('typ_ukolu', 'Pravidelný') == fty) and
            (not fz  or (fz == 'porada' and u.get('porada_id')) or (fz == 'samostatne' and not u.get('porada_id'))) and
            (not fm  or (fm == 'prirazen' and u.get('prirazen_id') == user_id) or (fm == 'zadal' and u.get('zadal_id') == user_id)) and
            (not fo  or str(u.get('prirazen_id','')) == fo or str(u.get('zadal_id','')) == fo) and
            (not fodd or fodd in (u.get('oddeleni') or '')) and
            _filtruj_termin(u, ft)
        ]

        # Řazení (výchozí vždy dle termínu, aby oddělovače dnů fungovaly)
        desc = (_razeni['smer'] == 'desc')
        if _razeni['pole'] == 'prirazen':
            zobrazene.sort(key=lambda u: ((u.get('prirazen_jmeno') or '').lower(), u.get('termin') or datetime.date.max), reverse=desc)
        elif _razeni['pole'] is None:
            # Výchozí zobrazení → termín, pak ruční pořadí uživatele (drag & drop v rámci dne).
            # Úkoly bez uloženého pořadí spadnou na konec dne, stabilní sort zachová původní řazení.
            zobrazene.sort(key=lambda u: (u.get('termin') or datetime.date.max, _poradi_uziv.get(u['id'], 10**9)))
        else:
            # Explicitní řazení dle termínu (klik na hlavičku) → ruční pořadí se potlačí
            zobrazene.sort(key=lambda u: u.get('termin') or datetime.date.max, reverse=desc)

        _zobr['list'] = zobrazene

        if _rezim['kanban'] and not _pohled['hotove']:
            hlavicka_row.clear()   # sloupcová hlavička patří jen seznamu
            _vykresli_kanban(zobrazene)
            return

        # Drag & drop přeřazení je aktivní jen ve výchozím zobrazení (řazení dle termínu)
        _drag_enabled = (_razeni['pole'] is None)
        _stav_render.update(drag=_drag_enabled, prid=_prid_mode)

        # Při prvním vykreslení sbal vše kromě dnešního dne (v pohledu Hotové ponech vše rozbalené)
        if not _inicializovano['hodnota']:
            _inicializovano['hodnota'] = True
            if not _pohled['hotove']:
                vsechny_terminy = {u.get('termin') for u in zobrazene}
                _sbalene_dny.update(t for t in vsechny_terminy if t != dnes)

        with seznam_box:
            if not zobrazene:
                with ui.card().classes('w-full p-8 items-center bg-gray-50 border border-dashed border-gray-200 rounded-xl'):
                    ui.icon('task_alt', size='3rem', color='gray-300').classes('mb-2')
                    if _pohled['hotove']:
                        ui.label('Zatím žádné dokončené ani zrušené úkoly.').classes('text-gray-500 font-bold')
                    else:
                        ui.label('Žádné úkoly odpovídají filtru.').classes('text-gray-500 font-bold')
                return

            # Stránkování po CELÝCH dnech — jeden den se nikdy nerozdělí mezi dvě stránky.
            # Cílová velikost je STRANKA_VEL úkolů; den s více úkoly dostane vlastní stránku.
            celkem = len(zobrazene)
            # Seskup po dnech (dict zachytí i nesouvislé výskyty téhož dne, např. při řazení dle osoby)
            _skupiny_dict: OrderedDict = OrderedDict()
            for u in zobrazene:
                _t = u.get('termin')
                if _t not in _skupiny_dict:
                    _skupiny_dict[_t] = []
                _skupiny_dict[_t].append(u)
            _skupiny = list(_skupiny_dict.items())
            # Naskládej dny do stránek bez překročení STRANKA_VEL (den větší než limit = vlastní stránka)
            _stranky = []
            _akt, _akt_pocet = [], 0
            for _t, _items in _skupiny:
                if _akt and _akt_pocet + len(_items) > STRANKA_VEL:
                    _stranky.append(_akt); _akt, _akt_pocet = [], 0
                _akt.append((_t, _items)); _akt_pocet += len(_items)
            if _akt:
                _stranky.append(_akt)
            pocet_stran = max(1, len(_stranky))
            _stranka['p'] = max(1, min(_stranka['p'], pocet_stran))
            # Kumulativně: stránky 1.._stranka['p'] (scroll jen přidává další)
            _akt_skupiny = [sk for _pg in _stranky[:_stranka['p']] for sk in _pg]
            strana_ukoly = [u for _, _items in _akt_skupiny for u in _items]

            _posledni_termin = object()  # unikátní sentinel
            for uk in strana_ukoly:
                termin_obj = uk.get('termin')

                # ── Oddělovač dnů ─────────────────────────────────────────────
                if termin_obj != _posledni_termin:
                    _posledni_termin = termin_obj
                    if termin_obj:
                        _dn = DNY_CZ_PLNE[termin_obj.weekday()]
                        _dt = termin_obj.strftime('%d.%m.%Y')
                        if termin_obj == dnes:
                            _sep_cls = 'bg-blue-50 border-l-4 border-blue-500'
                            _sep_txt_cls = 'text-blue-700'
                            _sep_ico = 'today'
                            _sep_label = f'{_dn} · {_dt}  — Dnes'
                        elif termin_obj < dnes:
                            _sep_cls = 'bg-red-50 border-l-4 border-red-400'
                            _sep_txt_cls = 'text-red-600'
                            _sep_ico = 'warning'
                            _sep_label = f'{_dn} · {_dt}  — Po termínu'
                        else:
                            _sep_cls = 'bg-gray-50 border-l-4 border-gray-300'
                            _sep_txt_cls = 'text-gray-600'
                            _sep_ico = 'event'
                            _sep_label = f'{_dn} · {_dt}'
                    else:
                        _sep_cls = 'bg-gray-50 border-l-4 border-gray-200'
                        _sep_txt_cls = 'text-gray-400'
                        _sep_ico = 'event_busy'
                        _sep_label = 'Bez termínu'

                    _je_sbaleno = termin_obj in _sbalene_dny
                    _pocet_dne  = sum(1 for u in strana_ukoly if u.get('termin') == termin_obj)

                    def _toggle_den(td=termin_obj):
                        if td in _sbalene_dny:
                            _sbalene_dny.discard(td)
                        else:
                            _sbalene_dny.add(td)
                        prekresli()

                    with ui.row().classes(
                        f'w-full px-4 py-1.5 items-center gap-2 rounded-lg mt-3 cursor-pointer '
                        f'select-none hover:brightness-95 {_sep_cls}'
                    ).on('click', _toggle_den):
                        ui.icon('chevron_right' if _je_sbaleno else 'expand_more', size='xs').classes(_sep_txt_cls)
                        ui.icon(_sep_ico, size='xs').classes(_sep_txt_cls)
                        ui.label(_sep_label).classes(f'text-sm font-bold {_sep_txt_cls} flex-1')
                        if _je_sbaleno:
                            ui.badge(str(_pocet_dne), color='grey').classes('text-xs')

                if termin_obj in _sbalene_dny:
                    continue

                _vykresli_kartu(uk, termin_obj)

            # ── Průběžné donačítání (nekonečný scroll) ────────────────────────
            with ui.row().classes('w-full items-center justify-center gap-2 mt-3 pt-2'):
                ui.label(f'{len(strana_ukoly)} z {celkem}').classes('text-xs text-gray-400 mr-2')
                if _stranka['p'] < pocet_stran:
                    def _nacti_dalsi():
                        _stranka['p'] += 1
                        prekresli()
                    _btn_dalsi = ui.button('Načíst další', on_click=_nacti_dalsi).props('outline size=sm icon=expand_more')
                    # Jakmile tlačítko vjede do viewportu, načte se další dávka sama
                    ui.run_javascript(
                        f"new IntersectionObserver((es,o)=>es.forEach(e=>{{if(e.isIntersecting){{o.disconnect();"
                        f"const el=document.getElementById('c{_btn_dalsi.id}');el&&el.click();}}}}),"
                        f"{{rootMargin:'200px'}}).observe(document.getElementById('c{_btn_dalsi.id}'));"
                    )

    async def _obnov_zive_radky(ref_ids) -> bool:
        """Živá per-úkol synchronizace: překreslí jen dotčené karty.

        True  = obslouženo lokálně (sekce nepotřebuje plný refresh; scroll,
                rozbalené dny i otevřené dialogy zůstávají netknuté),
        False = nutný plný refresh sekce (změnilo se členství seznamu)."""
        stare_ids = {u['id'] for u in ukoly}
        try:
            cerstve = await asyncio.to_thread(_nacti_ukoly_seznam, user_id, is_admin, nacti_oddeleni, odd_user_ids)
        except Exception:
            return False
        zmena_clenstvi = {u['id'] for u in cerstve} != stare_ids
        ukoly[:] = cerstve
        if _rezim['kanban'] and not _pohled['hotove']:
            prekresli()   # kanban = levné lokální překreslení z čerstvých dat
            return True
        if zmena_clenstvi:
            return False   # přibyl / zmizel úkol → plný refresh sekce
        mapa = {u['id']: u for u in cerstve}
        for rid in ref_ids:
            u = mapa.get(rid)
            if u is None:
                continue   # změněný úkol není v mém výřezu
            reg = _karty_reg.get(rid)
            if reg is None:
                prekresli()   # kartu nemám (filtr / sbalený den / nenačtená dávka)
                return True
            _obal, _termin, _stav = reg
            if u.get('termin') != _termin or ((u['stav'] in ('Hotovo', 'Zrušen')) != (_stav in ('Hotovo', 'Zrušen'))):
                prekresli()   # přeskupení dnů / (od)archivace → lokální překreslení
                return True
            _obal.clear()
            with _obal:
                _vykresli_kartu_telo(u, u.get('termin'))
            _karty_reg[rid] = (_obal, u.get('termin'), u['stav'])
        return True

    try:
        _kl_zive = ui.context.client
        if _kl_zive is not None:
            _UK_ZIVE_RADKY[_kl_zive.id] = _obnov_zive_radky
            _kl_zive.on_disconnect(lambda: _UK_ZIVE_RADKY.pop(_kl_zive.id, None))
    except Exception:
        pass

    def _aplikuj_pohled():
        """Promítne aktuální stav pohledu (přehled/archiv) do tlačítek a nadpisů."""
        if _pohled['hotove']:
            btn_hotove.text = 'Zpět na přehled'
            btn_hotove.props('icon=arrow_back')
            btn_novy.set_visibility(False)
            nadpis.set_text('Hotové / zrušené úkoly')
            podnadpis.set_text('Archiv dokončených a zrušených úkolů.')
        else:
            btn_hotove.text = 'Hotové / zrušené úkoly'
            btn_hotove.props('icon=check_circle')
            btn_novy.set_visibility(True)
            nadpis.set_text('Přehled úkolů')
            podnadpis.set_text('Všechny přidělené i zadané úkoly, vč. úkolů z porad.')
        btn_hotove.update()

    def _prepni_hotove():
        _pohled['hotove'] = not _pohled['hotove']
        _stranka['p'] = 1
        _sbalene_dny.clear()
        _inicializovano['hodnota'] = False  # přepočítej sbalení dnů pro nový pohled
        _aplikuj_pohled()
        prekresli()

    _aplikuj_pohled()  # promítni perzistentní stav pohledu i po refreshi
    prekresli()

    def _filtr_zmena():
        # zapiš do rodičovského stavu, ať přestavba sekce (změna stavu úkolu) filtry nesmaže
        _filtry.update(stav=filtr_stav.value, typ=filtr_typ.value,
                       termin=filtr_termin.value, zdroj=filtr_zdroj.value,
                       moje=filtr_moje.value,
                       osoba=filtr_osoba.value if filtr_osoba else '',
                       odd=filtr_odd.value if filtr_odd else '')
        _stranka['p'] = 1
        prekresli()
    for _f in [filtr_stav, filtr_typ, filtr_termin, filtr_zdroj, filtr_moje]:
        _f.on('update:model-value', lambda: _filtr_zmena())
    if filtr_osoba:
        filtr_osoba.on('update:model-value', lambda: _filtr_zmena())
    if filtr_odd:
        filtr_odd.on('update:model-value', lambda: _filtr_zmena())


def _rychla_zmena_stavu(uk, user_id, user_name):
    """Cyklická změna stavu kliknutím na badge."""
    poradi = ['Zadáno', 'Rozpracováno', 'Pozastaveno', 'Hotovo', 'Zrušen']
    aktualni = uk['stav']
    idx = poradi.index(aktualni) if aktualni in poradi else 0
    novy = poradi[(idx + 1) % len(poradi)]
    _zmenit_stav_ukolu(uk, novy, user_id, user_name)


def _zmenit_stav_ukolu(uk, novy_stav, user_id, user_name):
    """Přímá změna stavu úkolu — včetně správy timeru (jako v detailu)."""
    if uk['stav'] == novy_stav:
        return
    ukol_id = uk['id']

    if novy_stav == 'Rozpracováno':
        # Zahájí timer + nastaví stav
        _zahaj_ukol(ukol_id, user_id, user_name)
    elif novy_stav == 'Pozastaveno':
        # Zastaví timer + nastaví stav
        _pozastav_ukol(ukol_id, user_id, user_name)
    elif novy_stav == 'Hotovo':
        # Zastaví timer + nastaví stav + dokonceno_datum
        _dokoncit_ukol(ukol_id, user_id, user_name)
    else:
        # Zadáno / Zrušen — zastavit běžící timer, pokud existuje, a změnit stav
        now = datetime.datetime.now()
        c = intranet_data.get_db_connection()
        if c:
            try:
                cu = c.cursor(dictionary=True)
                cu.execute(
                    "SELECT id, cas_start FROM ukolovnik_cas_zaznamy "
                    "WHERE ukol_id=%s AND user_id=%s AND cas_konec IS NULL "
                    "ORDER BY cas_start DESC LIMIT 1", (ukol_id, user_id))
                zaznam = cu.fetchone()
                if zaznam:
                    delta_min = max(1, int((now - zaznam['cas_start']).total_seconds() / 60))
                    cu.execute("UPDATE ukolovnik_cas_zaznamy SET cas_konec=%s, trvani_minut=%s WHERE id=%s",
                               (now, delta_min, zaznam['id']))
                cu.execute("UPDATE ukolovnik_ukoly SET stav=%s, dokonceno_datum=NULL WHERE id=%s",
                           (novy_stav, ukol_id))
                c.commit()
            finally:
                cu.close(); c.close()
        intranet_logger.log_activity(user_name, "Úkolovník", f"Úkol #{ukol_id} stav→{novy_stav}")
        _log('ukol', ukol_id, user_id, user_name, f'Změnil/a stav na: {novy_stav}')
        _notifikuj_ucastniky_ukolu(ukol_id, user_id,
            f"📌 {user_name} změnil/a stav úkolu '{uk['nazev'][:40]}' na: {novy_stav}")

    ui.notify(novy_stav, type='positive', position='top-right', timeout=1500)
    _safe_refresh()


# =========================================================
# SEKCE: KALENDÁŘ
# =========================================================
def _barva_osoby(jmeno) -> str:
    """Stabilní barva osoby (stejné jméno = stejná barva napříč relacemi)."""
    return KAL_BARVY_OSOB[sum(map(ord, jmeno or '')) % len(KAL_BARVY_OSOB)]


def _rozvrstvi_bloky(bloky):
    """Rozvrství překrývající se bloky vedle sebe.

    bloky   = [(zacatek, konec, data)]  (zacatek/konec = porovnatelná čísla, např. minuty)
    vrací   = [(sloupec, pocet_sloupcu, data)]
    """
    vysledek, skupina, sloupce = [], [], []

    def _uzavri():
        n = max(len(sloupce), 1)
        for sl, d in skupina:
            vysledek.append((sl, n, d))
        skupina.clear(); sloupce.clear()

    for zac, kon, d in sorted(bloky, key=lambda x: (x[0], x[1])):
        if sloupce and zac >= max(sloupce):
            _uzavri()                       # žádný překryv → nová skupina, zase plná šířka
        sl = next((i for i, k in enumerate(sloupce) if k <= zac), len(sloupce))
        if sl == len(sloupce):
            sloupce.append(kon)
        else:
            sloupce[sl] = kon
        skupina.append((sl, d))
    _uzavri()
    return vysledek


def _nacti_udalosti_kalendar(user_id, typ, od, do, vsechna_prava=None):
    """Načte události kalendáře pro rozsah dat. Vrací dict[datetime.date, list].

    Rozsah viditelnosti je stejný jako v přehledu úkolů:
      • admin                                  → všechny úkoly,
      • hlavní vedoucí / právo „vidět oddělení" → úkoly celého oddělení + sledované,
      • běžný uživatel                          → vlastní (přidělené/zadané) + sledované.
    """
    vsechna_prava = vsechna_prava or []
    is_admin = 'vse' in vsechna_prava or 'ukolovnik_admin' in vsechna_prava
    uzivatele_oddeleni = [] if is_admin else _ziskej_uzivatele_oddeleni(vsechna_prava)
    je_vedouci_uk = is_admin or _je_hlavni_vedouci(vsechna_prava)
    vidi_cele_oddeleni = (not is_admin) and ('ukolovnik_ukoly_oddeleni_vse' in vsechna_prava)
    if uzivatele_oddeleni:
        odd_user_ids = [u['id'] for u in uzivatele_oddeleni]
    elif vidi_cele_oddeleni:
        odd_user_ids = [u['id'] for u in _ziskej_uziv_sveho_oddeleni(user_id)] or [user_id]
    else:
        odd_user_ids = [user_id]
    nacti_oddeleni = (je_vedouci_uk and bool(uzivatele_oddeleni)) or vidi_cele_oddeleni

    udalosti: dict[datetime.date, list] = {}
    conn = intranet_data.get_db_connection()
    if not conn:
        return udalosti
    try:
        cur = conn.cursor(dictionary=True)
        if typ == 'plan':
            # Plánovací: úkoly dle termínu
            if is_admin:
                cur.execute("""
                    SELECT id, nazev, stav, priorita, termin, prirazen_jmeno, zadal_jmeno, odhad_hodin, porada_nazev
                    FROM ukolovnik_ukoly
                    WHERE termin BETWEEN %s AND %s
                """, (od, do))
            elif nacti_oddeleni:
                ph = ','.join(['%s'] * len(odd_user_ids))
                cur.execute(f"""
                    SELECT DISTINCT u.id, u.nazev, u.stav, u.priorita, u.termin,
                           u.prirazen_jmeno, u.zadal_jmeno, u.odhad_hodin, u.porada_nazev
                    FROM ukolovnik_ukoly u
                    LEFT JOIN ukolovnik_sledovaci s ON s.typ='ukol' AND s.ref_id=u.id AND s.user_id=%s
                    WHERE u.termin BETWEEN %s AND %s
                      AND (u.prirazen_id IN ({ph}) OR u.zadal_id IN ({ph}) OR s.id IS NOT NULL)
                """, [user_id, od, do] + odd_user_ids + odd_user_ids)
            else:
                cur.execute("""
                    SELECT DISTINCT u.id, u.nazev, u.stav, u.priorita, u.termin,
                           u.prirazen_jmeno, u.zadal_jmeno, u.odhad_hodin, u.porada_nazev
                    FROM ukolovnik_ukoly u
                    LEFT JOIN ukolovnik_sledovaci s ON s.typ='ukol' AND s.ref_id=u.id AND s.user_id=%s
                    WHERE u.termin BETWEEN %s AND %s
                      AND (u.prirazen_id=%s OR u.zadal_id=%s OR s.id IS NOT NULL)
                """, (user_id, od, do, user_id, user_id))
            for row in cur.fetchall():
                t = row['termin']
                d = t.date() if isinstance(t, datetime.datetime) else t   # klíč vždy datetime.date
                udalosti.setdefault(d, []).append({'typ': 'ukol', **row})
        else:
            # Reálný: časové záznamy — admin/vedoucí vidí záznamy celého oddělení
            if is_admin:
                cur.execute("""
                    SELECT z.id, z.cas_start, z.cas_konec, z.trvani_minut, z.jmeno,
                           u.id as ukol_id, u.nazev, u.stav, u.prirazen_jmeno, u.zadal_jmeno
                    FROM ukolovnik_cas_zaznamy z
                    JOIN ukolovnik_ukoly u ON z.ukol_id = u.id
                    WHERE DATE(z.cas_start) BETWEEN %s AND %s
                """, (od, do))
            elif nacti_oddeleni:
                ph = ','.join(['%s'] * len(odd_user_ids))
                cur.execute(f"""
                    SELECT z.id, z.cas_start, z.cas_konec, z.trvani_minut, z.jmeno,
                           u.id as ukol_id, u.nazev, u.stav, u.prirazen_jmeno, u.zadal_jmeno
                    FROM ukolovnik_cas_zaznamy z
                    JOIN ukolovnik_ukoly u ON z.ukol_id = u.id
                    WHERE z.user_id IN ({ph})
                      AND DATE(z.cas_start) BETWEEN %s AND %s
                """, odd_user_ids + [od, do])
            else:
                cur.execute("""
                    SELECT z.id, z.cas_start, z.cas_konec, z.trvani_minut, z.jmeno,
                           u.id as ukol_id, u.nazev, u.stav, u.prirazen_jmeno, u.zadal_jmeno
                    FROM ukolovnik_cas_zaznamy z
                    JOIN ukolovnik_ukoly u ON z.ukol_id = u.id
                    WHERE z.user_id=%s
                      AND DATE(z.cas_start) BETWEEN %s AND %s
                """, (user_id, od, do))
            for row in cur.fetchall():
                cs = row['cas_start']
                d = cs.date() if hasattr(cs, 'date') else cs
                udalosti.setdefault(d, []).append({'typ': 'zaznam', **row})
    finally:
        cur.close(); conn.close()
    return udalosti


@ui.refreshable
async def _vykresli_kalendar(user_id, user_name, vsechna_prava, mesic_stav: dict):
    """mesic_stav = {'rok': int, 'mesic': int, 'typ': 'plan'|'real', 'nahled': 'mesic'|'tyden'|'den', 'den': int}"""
    dnes = datetime.date.today()
    rok = mesic_stav.get('rok', dnes.year)
    mesic = mesic_stav.get('mesic', dnes.month)
    typ = mesic_stav.get('typ', 'plan')
    nahled = mesic_stav.get('nahled', 'mesic')
    den = mesic_stav.get('den', dnes.day)
    # Ošetři přetečení dne při změně měsíce
    posledni_den = calendar.monthrange(rok, mesic)[1]
    if den > posledni_den:
        den = posledni_den

    # Datová fáze: události pro zvolený rozsah načíst PŘED stavbou UI
    if nahled == 'den':
        _rozsah_od = _rozsah_do = datetime.date(rok, mesic, den)
    elif nahled == 'tyden':
        _pondeli = datetime.date(rok, mesic, den)
        _pondeli -= datetime.timedelta(days=_pondeli.weekday())
        _rozsah_od, _rozsah_do = _pondeli, _pondeli + datetime.timedelta(days=6)
    else:
        _rozsah_od = datetime.date(rok, mesic, 1)
        _rozsah_do = datetime.date(rok, mesic, posledni_den)

    if nahled == 'tyden':
        # Týdenní osa ukazuje obojí naráz: termíny v pásu nahoře, odpracovaný čas v ose
        udalosti_plan, udalosti_real = await asyncio.gather(
            asyncio.to_thread(_nacti_udalosti_kalendar, user_id, 'plan', _rozsah_od, _rozsah_do, vsechna_prava),
            asyncio.to_thread(_nacti_udalosti_kalendar, user_id, 'real', _rozsah_od, _rozsah_do, vsechna_prava))
        udalosti = udalosti_plan
    else:
        udalosti = await asyncio.to_thread(
            _nacti_udalosti_kalendar, user_id, typ, _rozsah_od, _rozsah_do, vsechna_prava)

    # Starý obsah smazat až s daty v ruce → přestavba bez bliknutí a skoku scrollu
    ui.context.slot.parent.clear()

    def _posun(delta):
        _n = mesic_stav.get('nahled', 'mesic')
        if _n in ('den', 'tyden'):
            try:
                d_akt = datetime.date(mesic_stav['rok'], mesic_stav['mesic'], mesic_stav.get('den', dnes.day))
            except ValueError:
                d_akt = datetime.date(mesic_stav['rok'], mesic_stav['mesic'],
                                      calendar.monthrange(mesic_stav['rok'], mesic_stav['mesic'])[1])
            d_novy = d_akt + datetime.timedelta(days=delta * (7 if _n == 'tyden' else 1))
            mesic_stav.update({'rok': d_novy.year, 'mesic': d_novy.month, 'den': d_novy.day})
        else:
            m = mesic_stav['mesic'] + delta
            r = mesic_stav['rok']
            if m < 1:  m = 12; r -= 1
            if m > 12: m = 1;  r += 1
            mesic_stav.update({'rok': r, 'mesic': m})
        _kalendar_refresh_klient()

    with ui.row().classes('w-full items-center justify-between mb-4'):
        with ui.column().classes('gap-0'):
            ui.label('Kalendář').classes('text-3xl font-extrabold text-gray-800')
            ui.label('Plánovací = termíny úkolů  •  Reálný stav = odpracovaný čas').classes('text-xs text-gray-400')

        with ui.row().classes('items-center gap-3'):
            # Segmentovaný přepínač náhledů
            with ui.row().classes('items-center gap-0 bg-gray-100 rounded-lg p-0.5'):
                for _kod, _popis in (('mesic', 'Měsíc'), ('tyden', 'Týden'), ('den', 'Den')):
                    _akt = (_kod == nahled)
                    _cls = ('px-3 py-1 rounded-md text-xs font-bold cursor-pointer transition-colors '
                            + ('bg-white text-blue-700 shadow-sm' if _akt else 'text-gray-500 hover:text-gray-700'))
                    ui.label(_popis).classes(_cls).on(
                        'click', lambda k=_kod: [mesic_stav.update({'nahled': k}), _kalendar_refresh_klient()])
            if nahled != 'tyden':   # týdenní osa ukazuje termíny i odpracovaný čas naráz
                ui.select({'plan': 'Plánovací', 'real': 'Reálný stav'}, value=typ,
                          on_change=lambda e: [mesic_stav.update({'typ': e.value}), _kalendar_refresh_klient()]
                          ).classes('w-36').props('dense outlined')
            ui.button(icon='chevron_left', on_click=lambda: _posun(-1)).props('flat round dense')
            if nahled == 'den':
                d_lbl = datetime.date(rok, mesic, den)
                ui.label(f'{DNY_CZ_PLNE[d_lbl.weekday()]} {den}. {MESICE_CZ[mesic]} {rok}').classes('text-xl font-black text-gray-700 min-w-[16rem] text-center')
            elif nahled == 'tyden':
                with ui.column().classes('gap-0 min-w-[16rem] items-center'):
                    ui.label(f'{_rozsah_od.day}. {MESICE_CZ[_rozsah_od.month][:3].lower()} – '
                             f'{_rozsah_do.day}. {MESICE_CZ[_rozsah_do.month][:3].lower()} {_rozsah_do.year}'
                             ).classes('text-xl font-black text-gray-700 leading-tight')
                    ui.label(f'{_rozsah_od.isocalendar()[1]}. týden').classes('text-[11px] text-gray-400 font-bold')
            else:
                ui.label(f'{MESICE_CZ[mesic]} {rok}').classes('text-xl font-black text-gray-700 w-36 text-center')
            ui.button(icon='chevron_right', on_click=lambda: _posun(1)).props('flat round dense')
            ui.button('Dnes', on_click=lambda: [mesic_stav.update({'rok': dnes.year, 'mesic': dnes.month, 'den': dnes.day}), _kalendar_refresh_klient()]).props('flat dense').classes('text-xs text-blue-600')

    def _klik_ukol(uid):
        _dialog_detail_ukolu(uid, user_id, user_name, vsechna_prava, on_refresh=_kalendar_refresh_klient)

    # ---------- DENNÍ NÁHLED ----------
    if nahled == 'den':
        datum = datetime.date(rok, mesic, den)
        ev_dne = udalosti.get(datum, [])

        if not ev_dne:
            with ui.card().classes('w-full p-10 items-center justify-center bg-gray-50 border border-dashed border-gray-200 rounded-xl'):
                ui.icon('event_available', size='3rem').classes('text-gray-300')
                ui.label('Žádné události pro tento den.').classes('text-lg text-gray-500 font-bold')
            return

        # Seřaď podle času (reálný náhled) / názvu
        if typ == 'real':
            ev_dne = sorted(ev_dne, key=lambda e: e.get('cas_start') or datetime.datetime.min)

        with ui.column().classes('w-full gap-2'):
            for ev in ev_dne:
                if ev['typ'] == 'ukol':
                    _, color = STAVY_UKOL.get(ev['stav'], ('', '#9ca3af'))
                    with ui.card().classes('w-full p-3 rounded-xl bg-white border border-gray-100 hover:shadow-md transition-shadow cursor-pointer') \
                            .style(f'border-left:5px solid {color}') \
                            .on('click', lambda uid=ev['id']: _klik_ukol(uid)):
                        with ui.row().classes('w-full items-start justify-between gap-3 no-wrap'):
                            with ui.column().classes('gap-0.5 flex-grow min-w-0'):
                                ui.label(ev['nazev']).classes('text-base font-bold text-gray-800 whitespace-normal break-words')
                                meta = f"Přiděleno: {ev.get('prirazen_jmeno','—')}  •  Zadal: {ev.get('zadal_jmeno','—')}"
                                if ev.get('porada_nazev'):
                                    meta += f"  •  Porada: {ev['porada_nazev']}"
                                ui.label(meta).classes('text-xs text-gray-500 whitespace-normal break-words')
                            with ui.column().classes('items-end gap-1 shrink-0'):
                                ui.label(ev['stav']).classes('px-2 py-0.5 rounded-full text-[11px] font-bold').style(f'background:{color}20; color:{color}')
                                if ev.get('odhad_hodin'):
                                    ui.label(f"⏱ {_format_odhad(ev['odhad_hodin'])}").classes('text-xs text-gray-400')
                else:
                    h, m = divmod(ev.get('trvani_minut') or 0, 60)
                    cas_s = ev['cas_start'].strftime('%H:%M') if ev.get('cas_start') else ''
                    cas_k = ev['cas_konec'].strftime('%H:%M') if ev.get('cas_konec') else ''
                    rozsah = f'{cas_s}–{cas_k}' if cas_k else cas_s
                    trvani = f'{h}h {m}m' if ev.get('trvani_minut') else ''
                    with ui.card().classes('w-full p-3 rounded-xl bg-white border border-gray-100 hover:shadow-md transition-shadow cursor-pointer') \
                            .style('border-left:5px solid #22c55e') \
                            .on('click', lambda uid=ev['ukol_id']: _klik_ukol(uid)):
                        with ui.row().classes('w-full items-start justify-between gap-3 no-wrap'):
                            with ui.column().classes('gap-0.5 flex-grow min-w-0'):
                                ui.label(ev.get('nazev', '')).classes('text-base font-bold text-gray-800 whitespace-normal break-words')
                                ui.label(f"{ev.get('jmeno','')}  •  {ev.get('prirazen_jmeno','—')}").classes('text-xs text-gray-500 whitespace-normal break-words')
                            with ui.column().classes('items-end gap-1 shrink-0'):
                                ui.label(rozsah).classes('text-sm font-bold text-green-700')
                                if trvani:
                                    ui.label(trvani).classes('text-xs text-gray-400')
        return

    # ---------- TÝDENNÍ NÁHLED (časová osa) ----------
    if nahled == 'tyden':
        dny_tydne = [_rozsah_od + datetime.timedelta(days=i) for i in range(7)]

        # Bloky odpracovaného času po dnech + rozsah osy podle dat
        bloky_dne, h_min, h_max = {}, KAL_HOD_OD, KAL_HOD_DO
        for d in dny_tydne:
            polozky = []
            for ev in udalosti_real.get(d, []):
                cs = ev.get('cas_start')
                if not cs:
                    continue
                zac = cs.hour * 60 + cs.minute
                ck = ev.get('cas_konec')
                if ck and ck > cs:
                    kon = zac + int((ck - cs).total_seconds() // 60)
                else:
                    kon = zac + int(ev.get('trvani_minut') or 30)
                kon = min(max(kon, zac + 15), 24 * 60)
                polozky.append((zac, kon, ev))
                h_min = min(h_min, zac // 60)
                h_max = max(h_max, -(-kon // 60))          # zaokrouhlení nahoru
            bloky_dne[d] = _rozvrstvi_bloky(polozky)

        h_min, h_max = max(0, h_min), min(24, max(h_max, h_min + 1))
        pocet_hodin = h_max - h_min
        vyska_px = pocet_hodin * KAL_HODINA_PX
        mrizka = 'display:grid; grid-template-columns:60px repeat(7, minmax(0, 1fr));'

        with ui.card().classes('w-full p-0 bg-white border border-gray-100 rounded-xl shadow-sm overflow-hidden'):
            # ── Záhlaví dnů ─────────────────────────────────────────────────
            with ui.element('div').classes('w-full border-b border-gray-100').style(mrizka):
                ui.element('div').classes('border-r border-gray-100')
                for d in dny_tydne:
                    je_dnes = d == dnes
                    je_vikend = d.weekday() >= 5
                    bunka = ui.element('div').classes(
                        'py-2 text-center border-r border-gray-100 cursor-pointer hover:bg-blue-50 transition-colors '
                        + ('bg-blue-50' if je_dnes else ('bg-gray-50' if je_vikend else '')))
                    bunka.on('click', lambda dd=d: [mesic_stav.update(
                        {'nahled': 'den', 'rok': dd.year, 'mesic': dd.month, 'den': dd.day}),
                        _kalendar_refresh_klient()])
                    with bunka:
                        ui.label(DNY_CZ[d.weekday()]).classes(
                            'block text-[11px] font-bold uppercase tracking-wide '
                            + ('text-blue-600' if je_dnes else 'text-gray-400'))
                        ui.label(str(d.day)).classes(
                            'block text-lg font-black leading-tight '
                            + ('text-blue-700' if je_dnes else ('text-gray-400' if je_vikend else 'text-gray-700')))
                        ui.tooltip('Otevřít denní náhled')

            # ── Celodenní pás: termíny úkolů ────────────────────────────────
            with ui.element('div').classes('w-full border-b border-gray-100 bg-gray-50').style(mrizka):
                with ui.element('div').classes('border-r border-gray-100 flex items-center justify-end pr-2 py-1'):
                    ui.label('Termíny').classes('text-[10px] font-bold text-gray-400 uppercase')
                for d in dny_tydne:
                    ukoly_dne = [e for e in udalosti_plan.get(d, []) if e.get('typ') == 'ukol']
                    with ui.element('div').classes('border-r border-gray-100 p-1 min-h-[34px] flex flex-col gap-0.5'):
                        for ev in ukoly_dne[:3]:
                            _, color = STAVY_UKOL.get(ev['stav'], ('', '#9ca3af'))
                            with ui.element('div').classes(
                                    'px-1.5 py-0.5 rounded text-[10px] font-bold leading-tight truncate '
                                    'cursor-pointer hover:opacity-80'
                            ).style(f'background:{color}20; border-left:3px solid {color}; color:{color}') \
                                    .on('click', lambda uid=ev['id']: _klik_ukol(uid)):
                                ui.label(ev['nazev']).classes('block truncate')
                                ui.tooltip(f"{ev['nazev']}\nPřiděleno: {ev.get('prirazen_jmeno','—')}\n"
                                           f"Stav: {ev['stav']}"
                                           + (f"\nOdhad: {_format_odhad(ev['odhad_hodin'])}" if ev.get('odhad_hodin') else ''))
                        if len(ukoly_dne) > 3:
                            ui.label(f'+{len(ukoly_dne)-3} další').classes(
                                'text-[9px] text-gray-400 font-bold cursor-pointer hover:text-blue-600').on(
                                'click', lambda dd=d: [mesic_stav.update(
                                    {'nahled': 'den', 'rok': dd.year, 'mesic': dd.month, 'den': dd.day}),
                                    _kalendar_refresh_klient()])

            # ── Časová osa se záznamy ───────────────────────────────────────
            with ui.element('div').classes('w-full overflow-y-auto').style('max-height:560px'):
                with ui.element('div').classes('w-full').style(mrizka):
                    # Hodinový sloupec
                    with ui.element('div').classes('relative border-r border-gray-100').style(f'height:{vyska_px}px'):
                        for i in range(pocet_hodin):
                            with ui.element('div').classes('absolute right-2 text-[10px] font-bold text-gray-400').style(
                                    f'top:{max(i * KAL_HODINA_PX - 6, 0)}px'):   # první popisek neuříznout
                                ui.label(f'{h_min + i}:00')

                    for d in dny_tydne:
                        je_vikend = d.weekday() >= 5
                        with ui.element('div').classes(
                                'relative border-r border-gray-100 ' + ('bg-gray-50' if je_vikend else '')
                        ).style(f'height:{vyska_px}px'):
                            # Hodinové linky
                            for i in range(1, pocet_hodin):
                                ui.element('div').classes('absolute left-0 right-0 border-t border-gray-100').style(
                                    f'top:{i * KAL_HODINA_PX}px')

                            # Ukazatel „teď"
                            if d == dnes:
                                ted = datetime.datetime.now()
                                pozice = (ted.hour * 60 + ted.minute - h_min * 60) / 60 * KAL_HODINA_PX
                                if 0 <= pozice <= vyska_px:
                                    with ui.element('div').classes('absolute left-0 right-0 z-20').style(
                                            f'top:{pozice:.0f}px; border-top:2px solid #ef4444'):
                                        ui.element('div').classes('absolute rounded-full').style(
                                            'left:-4px; top:-5px; width:8px; height:8px; background:#ef4444')

                            # Bloky záznamů
                            for sl, poc, ev in bloky_dne.get(d, []):
                                cs, ck = ev.get('cas_start'), ev.get('cas_konec')
                                zac = cs.hour * 60 + cs.minute
                                kon = (zac + int((ck - cs).total_seconds() // 60)) if (ck and ck > cs) \
                                    else zac + int(ev.get('trvani_minut') or 30)
                                kon = min(max(kon, zac + 15), 24 * 60)
                                top = (zac - h_min * 60) / 60 * KAL_HODINA_PX
                                vyska = max((kon - zac) / 60 * KAL_HODINA_PX - 3, 20)
                                barva = _barva_osoby(ev.get('jmeno'))
                                cas_s = cs.strftime('%H:%M')
                                cas_k = ck.strftime('%H:%M') if ck else ''
                                hh, mm = divmod(ev.get('trvani_minut') or (kon - zac), 60)
                                trvani = f'{hh}h {mm}m' if hh else f'{mm}m'
                                with ui.element('div').classes(
                                        'absolute rounded-md px-1.5 py-0.5 overflow-hidden cursor-pointer '
                                        'hover:shadow-md hover:z-10 transition-shadow'
                                ).style(
                                    f'top:{top:.0f}px; height:{vyska:.0f}px; '
                                    f'left:calc({sl / poc * 100:.2f}% + 2px); width:calc({100 / poc:.2f}% - 5px); '
                                    f'background:{barva}1a; border-left:3px solid {barva}; color:{barva}'
                                ).on('click', lambda uid=ev['ukol_id']: _klik_ukol(uid)):
                                    ui.label(f'{cas_s} · {trvani}').classes('block text-[10px] font-black leading-tight truncate')
                                    if vyska >= 34:
                                        ui.label(ev.get('nazev', '')).classes(
                                            'block text-[10px] font-bold leading-tight text-gray-700 line-clamp-2')
                                    if vyska >= 60 and ev.get('jmeno'):
                                        ui.label(ev['jmeno']).classes('block text-[9px] text-gray-500 truncate')
                                    ui.tooltip(f"{ev.get('nazev','')}\n{ev.get('jmeno','')}\n"
                                               f"{cas_s}{('–' + cas_k) if cas_k else ''} • {trvani}")

            # ── Součty po dnech ─────────────────────────────────────────────
            with ui.element('div').classes('w-full border-t border-gray-100 bg-gray-50').style(mrizka):
                with ui.element('div').classes('border-r border-gray-100 flex items-center justify-end pr-2 py-1.5'):
                    ui.label('Součet').classes('text-[10px] font-bold text-gray-400 uppercase')
                for d in dny_tydne:
                    minuty = sum((e.get('trvani_minut') or 0) for e in udalosti_real.get(d, []))
                    hod = minuty / 60
                    podil = min(hod / 8 * 100, 100)
                    barva_bar = '#ef4444' if hod > 9 else ('#f59e0b' if hod > 8 else '#22c55e')
                    with ui.element('div').classes('border-r border-gray-100 px-2 py-1.5'):
                        ui.label(f'{hod:.1f} h'.replace('.', ',') if minuty else '—').classes(
                            'block text-center text-xs font-black ' + ('text-gray-700' if minuty else 'text-gray-300'))
                        with ui.element('div').classes('w-full rounded-full mt-1').style('height:3px; background:#e5e7eb'):
                            if minuty:
                                ui.element('div').classes('rounded-full').style(
                                    f'height:3px; width:{podil:.0f}%; background:{barva_bar}')

        # ── Legenda osob ────────────────────────────────────────────────────
        osoby = sorted({e.get('jmeno') for d in dny_tydne for e in udalosti_real.get(d, []) if e.get('jmeno')})
        with ui.row().classes('w-full items-center gap-4 mt-3 px-1 flex-wrap'):
            ui.label('Odpracovaný čas:').classes('text-[11px] font-bold text-gray-400 uppercase')
            for jm in osoby:
                with ui.row().classes('items-center gap-1.5'):
                    ui.element('div').classes('rounded-full').style(
                        f'width:10px; height:10px; background:{_barva_osoby(jm)}')
                    ui.label(jm).classes('text-xs text-gray-600 font-bold')
            if not osoby:
                ui.label('V tomto týdnu nejsou žádné časové záznamy.').classes('text-xs text-gray-400')
        return

    # ---------- MĚSÍČNÍ NÁHLED ----------
    prvni = datetime.date(rok, mesic, 1)
    posledni = datetime.date(rok, mesic, posledni_den)
    # Mřížka vždy celé týdny — dny sousedních měsíců se ukážou ztlumeně
    zacatek_mrizky = prvni - datetime.timedelta(days=prvni.weekday())
    konec_mrizky = posledni + datetime.timedelta(days=6 - posledni.weekday())
    pocet_tydnu = ((konec_mrizky - zacatek_mrizky).days + 1) // 7
    mrizka_m = 'display:grid; grid-template-columns:34px repeat(7, minmax(0, 1fr)); gap:2px;'

    # Hlavička dnů
    with ui.element('div').classes('w-full mb-0.5').style(mrizka_m):
        ui.element('div')
        for dn in DNY_CZ:
            ui.label(dn).classes('text-center text-xs font-bold text-gray-500 py-1')

    # Dny v mřížce (po týdnech, s číslem týdne v levém sloupci)
    with ui.element('div').classes('w-full').style(mrizka_m):
        for t in range(pocet_tydnu):
            pondeli_t = zacatek_mrizky + datetime.timedelta(days=t * 7)
            with ui.element('div').classes('flex items-start justify-center pt-1.5'):
                ui.label(str(pondeli_t.isocalendar()[1])).classes(
                    'text-[10px] font-bold text-gray-300 cursor-pointer hover:text-blue-600').on(
                    'click', lambda dd=pondeli_t: [mesic_stav.update(
                        {'nahled': 'tyden', 'rok': dd.year, 'mesic': dd.month, 'den': dd.day}),
                        _kalendar_refresh_klient()]).tooltip('Zobrazit týden')

            for i in range(7):
                datum = pondeli_t + datetime.timedelta(days=i)
                d = datum.day
                je_dnes = datum == dnes
                je_vikend = datum.weekday() >= 5
                mimo_mesic = datum.month != mesic
                ev_dne = [] if mimo_mesic else udalosti.get(datum, [])

                if mimo_mesic:
                    with ui.card().classes('min-h-[96px] p-1.5 rounded-lg bg-gray-50 border border-gray-100 overflow-hidden'):
                        ui.label(str(d)).classes('text-xs text-gray-300 font-bold')
                    continue

                bg = 'bg-blue-50 border-blue-300' if je_dnes else ('bg-gray-50' if je_vikend else 'bg-white')
                border = 'border-2 border-blue-400' if je_dnes else 'border border-gray-100'

                with ui.card().classes(f'min-h-[96px] p-1.5 rounded-lg {bg} {border} overflow-hidden hover:shadow-md transition-shadow'):
                    cislo_cls = 'text-xs font-extrabold text-blue-700 mb-1' if je_dnes else ('text-xs text-gray-400 mb-1' if je_vikend else 'text-xs font-bold text-gray-600 mb-1')
                    # Číslo dne otevře denní náhled
                    ui.label(str(d)).classes(cislo_cls + ' cursor-pointer hover:text-blue-600') \
                        .on('click', lambda dd=d: [mesic_stav.update({'nahled': 'den', 'den': dd}), _kalendar_refresh_klient()]) \
                        .tooltip('Zobrazit den')

                    for ev in ev_dne[:4]:
                        if ev['typ'] == 'ukol':
                            _, color = STAVY_UKOL.get(ev['stav'], ('bg-gray-200 text-gray-600 border-gray-300', '#9ca3af'))

                            with ui.element('div').classes('w-full px-1 py-0.5 rounded text-[10px] font-bold leading-tight cursor-pointer mb-0.5 hover:opacity-80').style(f'background:{color}20; border-left:3px solid {color}; color:{color}').on('click', lambda uid=ev['id']: _klik_ukol(uid)):
                                ui.label(ev['nazev']).classes('block whitespace-normal break-words line-clamp-2')
                                ui.tooltip(
                                    f"{ev['nazev']}\n"
                                    f"Přiděleno: {ev.get('prirazen_jmeno','—')}\n"
                                    f"Stav: {ev['stav']}"
                                    + (f"\nOdhad: {_format_odhad(ev['odhad_hodin'])}" if ev.get('odhad_hodin') else "")
                                )
                        else:
                            h, m = divmod(ev.get('trvani_minut') or 0, 60)
                            cas_s = ev['cas_start'].strftime('%H:%M') if ev.get('cas_start') else ''
                            trvani = f'{h}h{m}m' if ev.get('trvani_minut') else '⏱'

                            with ui.element('div').classes('w-full px-1 py-0.5 rounded text-[10px] font-bold leading-tight cursor-pointer mb-0.5 hover:opacity-80').style('background:#dcfce7;border-left:3px solid #22c55e;color:#166534').on('click', lambda uid=ev['ukol_id']: _klik_ukol(uid)):
                                ui.label(f'{cas_s} {trvani}').classes('block whitespace-normal break-words')
                                ui.tooltip(f"{ev.get('nazev','')}\n{ev.get('jmeno','')}\n{cas_s} • {trvani}")

                    if len(ev_dne) > 4:
                        ui.label(f'+{len(ev_dne)-4} další').classes('text-[9px] text-gray-400 cursor-pointer hover:text-blue-600') \
                            .on('click', lambda dd=d: [mesic_stav.update({'nahled': 'den', 'den': dd}), _kalendar_refresh_klient()])


# =========================================================
# DIALOG: DETAIL PORADY
# =========================================================
def _dialog_detail_porady(porada_id, user_id, user_name, vsechna_prava, on_refresh):
    conn = intranet_data.get_db_connection()
    if not conn: return
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT * FROM ukolovnik_porady WHERE id=%s", (porada_id,))
        porada = cur.fetchone()
    finally:
        cur.close(); conn.close()
    if not porada:
        ui.notify('Porada nenalezena.', type='negative'); return

    is_admin = vsechna_prava and ('vse' in vsechna_prava or 'ukolovnik_admin' in vsechna_prava)
    muze_editovat = is_admin or porada.get('vytvoril_id') == user_id or porada.get('moderator_id') == user_id

    _navrhy = {}   # rozepsané texty přežijí živé překreslení obsahu
    with _dialog_kotva(), \
         ui.dialog().props('maximized persistent no-refocus no-shake') as dlg, \
            ui.card().classes('w-full h-full rounded-none p-0 bg-gray-100 flex flex-col overflow-hidden').on('keydown.escape', lambda: dlg.close()):
        # Tělo dialogu jako refreshable — živá synchronizace překresluje jen
        # obsah na místě, okno zůstává otevřené (žádné close+reopen).
        @ui.refreshable
        def _telo():
            conn = intranet_data.get_db_connection()
            if not conn:
                return
            try:
                cur = conn.cursor(dictionary=True)
                cur.execute("SELECT * FROM ukolovnik_porady WHERE id=%s", (porada_id,))
                porada = cur.fetchone()
            finally:
                cur.close(); conn.close()
            if not porada:
                ui.notify('Porada byla mezitím smazána jiným uživatelem.', type='warning')
                dlg.close()
                return
            is_admin = vsechna_prava and ('vse' in vsechna_prava or 'ukolovnik_admin' in vsechna_prava)
            muze_editovat = is_admin or porada.get('vytvoril_id') == user_id or porada.get('moderator_id') == user_id

            with ui.row().classes('w-full items-center justify-between px-6 py-4 bg-white border-b border-gray-200 shrink-0 shadow-sm'):
                with ui.column().classes('gap-0'):
                    ui.label(porada['nazev']).classes('text-2xl font-black text-gray-800')
                    datum_str = porada['datum'].strftime('%d.%m.%Y') if porada.get('datum') else ''
                    cas_od = str(porada['cas_od'])[:5] if porada.get('cas_od') else ''
                    cas_do = str(porada['cas_do'])[:5] if porada.get('cas_do') else ''
                    meta = datum_str + (f'  •  {cas_od}–{cas_do}' if cas_od else '') + (f'  •  {porada["misto"]}' if porada.get('misto') else '')
                    ui.label(meta).classes('text-sm text-gray-500')

                with ui.row().classes('items-center gap-3'):
                    b_stav = BARVY_STAVU_PORADA.get(porada['stav'], 'bg-gray-100 text-gray-800 border-gray-300')
                    ui.label(porada['stav']).classes(f'px-3 py-1 rounded-full text-xs font-bold border uppercase tracking-wider {b_stav}')
                    if muze_editovat:
                        def _smazat_poradu_click():
                            def _provest():
                                ok = _smazat_poradu_db(porada_id)
                                if ok:
                                    intranet_logger.log_activity(user_name, "Úkolovník", f"Smazána porada #{porada_id}: {porada['nazev']}")
                                    ui.notify('Porada byla smazána.', type='positive', position='top')
                                    dlg.close()
                                    on_refresh()
                                else:
                                    ui.notify('Chyba při mazání.', type='negative', position='top')
                            _potvrdit_smazat(f'Opravdu smazat poradu „{porada["nazev"]}"?\nBudou smazány i všechny zápisy a úkoly z porady.', _provest)
                        ui.button(icon='delete_forever', on_click=_smazat_poradu_click).props('flat round dense').classes('text-red-400 hover:text-red-600').tooltip('Smazat poradu')
                    ui.button(icon='close', on_click=dlg.close).props('flat round dense').classes('text-gray-400 hover:text-red-500')

            with ui.row().classes('w-full flex-1 overflow-hidden gap-0'):

                with ui.column().classes('flex-1 overflow-y-auto p-6 gap-4 bg-white border-r border-gray-200'):

                    with ui.card().classes('w-full p-4 bg-gray-50 border border-gray-200 rounded-xl'):
                        ucastnici = _ziskej_ucastniky_porady(porada_id)
                        ucastnici_jmena = ', '.join(u['jmeno'] for u in ucastnici) or '—'
                        with ui.grid(columns=2).classes('w-full gap-x-6 gap-y-2 text-sm'):
                            ui.label('Moderátor:').classes('text-gray-500 font-bold')
                            ui.label(porada.get('moderator_jmeno') or '—').classes('text-gray-800')
                            ui.label('Zapisovatel:').classes('text-gray-500 font-bold')
                            ui.label(porada.get('zapisovatel_jmeno') or '—').classes('text-gray-800')
                            ui.label('Účastníci:').classes('text-gray-500 font-bold')
                            ui.label(ucastnici_jmena).classes('text-gray-800')
                            ui.label('Vytvořil:').classes('text-gray-500 font-bold')
                            ui.label(f"{porada.get('vytvoril_jmeno','—')} ({porada['vytvoreno'].strftime('%d.%m.%Y %H:%M')})").classes('text-gray-800')

                        if porada.get('popis'):
                            ui.separator().classes('my-2')
                            ui.label('Program / popis:').classes('text-xs font-bold text-gray-500 uppercase')
                            ui.label(porada['popis']).classes('text-gray-700 whitespace-pre-wrap text-sm mt-1')

                        if muze_editovat:
                            with ui.row().classes('mt-3 gap-2 flex-wrap'):
                                for s in ['Plánovaná', 'Probíhá', 'Dokončená', 'Zrušená']:
                                    def _zmen_stav(st=s):
                                        c = intranet_data.get_db_connection()
                                        cu = c.cursor()
                                        cu.execute("UPDATE ukolovnik_porady SET stav=%s WHERE id=%s", (st, porada_id))
                                        c.commit(); cu.close(); c.close()
                                        intranet_logger.log_activity(user_name, "Úkolovník", f"Stav porady #{porada_id}→{st}")
                                        _log('porada', porada_id, user_id, user_name, f'Změnil/a stav na: {st}')
                                        _notifikuj_ucastniky_porady(porada_id, user_id, f"📅 Stav porady '{porada['nazev'][:40]}' byl změněn na: {st}")
                                        ui.notify(f'Stav: {st}', type='positive')
                                        dlg.close(); on_refresh()
                                    ui.button(s, on_click=_zmen_stav).props('size=sm outline').classes('font-bold')

                    ui.label('Zápis z porady').classes('text-lg font-extrabold text-gray-700 mt-2')
                    poznamky_box = ui.column().classes('w-full gap-3')

                    def nacti_poz():
                        poznamky_box.clear()
                        c = intranet_data.get_db_connection()
                        pz = []
                        if c:
                            try:
                                cu = c.cursor(dictionary=True)
                                cu.execute("SELECT * FROM ukolovnik_porada_poznamky WHERE porada_id=%s ORDER BY vytvoreno ASC", (porada_id,))
                                pz = cu.fetchall()
                            finally:
                                cu.close(); c.close()
                        with poznamky_box:
                            if not pz:
                                ui.label('Zatím žádný zápis.').classes('text-gray-400 italic text-sm')
                            for p in pz:
                                je_muj = p['user_id'] == user_id
                                bc = 'bg-blue-50 border-blue-200' if je_muj else 'bg-gray-50 border-gray-200'
                                with ui.card().classes(f'w-full p-3 border rounded-lg {bc}'):
                                    with ui.row().classes('w-full justify-between items-center mb-1'):
                                        ui.label(p['jmeno_autora']).classes('text-xs font-bold text-gray-600')
                                        ui.label(p['vytvoreno'].strftime('%d.%m.%Y %H:%M')).classes('text-xs text-gray-400')
                                    ui.label(p['text']).classes('text-sm text-gray-800 whitespace-pre-wrap')

                    nacti_poz()

                    with ui.row().classes('w-full items-end gap-2 mt-2'):
                        nova_p = ui.textarea('Přidat poznámku / zápis...').classes('flex-1 bg-gray-50').props('outlined autogrow').bind_value(_navrhy, 'zapis')

                        def pridat_poz():
                            txt = nova_p.value.strip() if nova_p.value else ''
                            if not txt: return
                            c = intranet_data.get_db_connection()
                            cu = c.cursor()
                            cu.execute("INSERT INTO ukolovnik_porada_poznamky (porada_id, user_id, jmeno_autora, text) VALUES (%s,%s,%s,%s)",
                                       (porada_id, user_id, user_name, txt))
                            c.commit(); cu.close(); c.close()
                            intranet_logger.log_activity(user_name, "Úkolovník", f"Poznámka k poradě #{porada_id}")
                            _log('porada', porada_id, user_id, user_name, 'Přidal/a zápis z porady')
                            _notifikuj_ucastniky_porady(porada_id, user_id, f"✏️ {user_name} přidal/a zápis k poradě '{porada['nazev'][:40]}'.")
                            nova_p.value = ''
                            nacti_poz()

                        ui.button(icon='send', on_click=pridat_poz).classes('bg-blue-600 text-white h-14 w-14 rounded-xl shadow-md')

                # Pravý panel: úkoly porady
                with ui.column().classes('w-96 shrink-0 overflow-y-auto p-5 gap-4 bg-gray-50'):
                    with ui.row().classes('w-full justify-between items-center'):
                        ui.label('Úkoly z porady').classes('text-lg font-extrabold text-gray-700')
                        if muze_editovat:
                            ui.button(icon='add', on_click=lambda: _dialog_novy_ukol(
                                user_id, user_name, vsechna_prava,
                                porada_id=porada_id, porada_nazev=porada['nazev'],
                                porada_datum=porada['datum'],
                                on_refresh=nacti_ukoly_p
                            )).props('round dense').classes('bg-green-600 text-white shadow')

                    ukoly_box = ui.column().classes('w-full gap-2')

                    def nacti_ukoly_p():
                        ukoly_box.clear()
                        ukoly = _ziskej_ukoly_porady(porada_id)
                        with ukoly_box:
                            if not ukoly:
                                ui.label('Žádné úkoly.').classes('text-gray-400 italic text-sm')
                                return
                            for uk in ukoly:
                                b_uk, _ = STAVY_UKOL.get(uk['stav'], ('bg-gray-100 text-gray-700 border-gray-300', ''))
                                termin_str = uk['termin'].strftime('%d.%m.%Y') if uk.get('termin') else '—'
                                delta, dt_cls, dt_txt = _dny_do_terminu(uk.get('termin'))

                                def _det(u=uk):
                                    _dialog_detail_ukolu(u['id'], user_id, user_name, vsechna_prava,
                                                         on_refresh=lambda: [nacti_ukoly_p(), on_refresh()])

                                je_hotovo = uk['stav'] == 'Hotovo'
                                t_cls = 'line-through text-gray-400' if je_hotovo else 'text-gray-800 font-bold'

                                with ui.card().classes('w-full p-3 border border-gray-200 rounded-xl bg-white hover:shadow-md cursor-pointer').on('click', lambda u=uk: _det(u)):
                                    with ui.row().classes('w-full justify-between items-start gap-1'):
                                        with ui.column().classes('flex-1 gap-0.5'):
                                            ui.label(uk['nazev']).classes(f'text-sm {t_cls} line-clamp-2')
                                            ui.label(f'Pro: {uk.get("prirazen_jmeno","—")}').classes('text-xs text-gray-500')
                                            with ui.row().classes('gap-2'):
                                                ui.label(f'Termín: {termin_str}').classes('text-xs text-gray-400')
                                                if uk.get('termin') and not je_hotovo:
                                                    ui.label(dt_txt).classes(f'text-xs {dt_cls}')
                                        ui.label(uk['stav']).classes(f'px-2 py-0.5 rounded text-[10px] font-bold border {b_uk} shrink-0')

                    nacti_ukoly_p()

        _telo()

    _registruj_otevreny_detail('porada', porada_id, dlg, _telo.refresh)
    dlg.open()


# =========================================================
# DIALOG: NOVÁ PORADA
# =========================================================
def _dialog_nova_porada(user_id, user_name, on_refresh, vsechna_prava=None):
    if vsechna_prava is None:
        vsechna_prava = []

    is_admin_p = 'vse' in vsechna_prava or 'ukolovnik_admin' in vsechna_prava
    vidi_vsechny = is_admin_p or 'ukolovnik_porady_vsichni' in vsechna_prava

    # Moderátor / Zapisovatel — vedoucí oddělení, nebo všichni s příslušným právem
    if vidi_vsechny:
        mod_zap_opts = _ziskej_uzivatele_options()
    else:
        mod_zap_opts = _ziskej_vedouci_vsech_oddeleni()
        # Přidej sebe, ať se může vybrat jako moderátor/zapisovatel
        if user_id not in mod_zap_opts:
            mod_zap_opts[user_id] = user_name

    # Účastníci — stejné pravidlo
    ucast_opts = _ziskej_uzivatele_options() if vidi_vsechny else _ziskej_viditelne_uzivatele(user_id, vsechna_prava)

    mod_zap_select = {v: v for v in mod_zap_opts.values()}
    # Pro zpětné dohledání id → jmeno při ukládání
    uziv_opts = {**mod_zap_opts, **ucast_opts}
    uziv_select = {v: v for v in uziv_opts.values()}

    with _dialog_kotva(), ui.dialog() as dlg, ui.card().classes('w-full max-w-3xl p-6 rounded-xl'):
        ui.label('Nová porada').classes('text-2xl font-bold text-blue-800 mb-4')

        nazev = ui.input('Název / téma porady').classes('w-full mb-3 bg-white').props('outlined')
        popis = ui.textarea('Program a popis tématu').classes('w-full mb-3 bg-white').props('outlined rows=3')

        with ui.row().classes('w-full gap-3 mb-3 flex-wrap'):
            datum_inp = ui.input('Datum (DD.MM.RRRR)').classes('flex-1 bg-white').props('outlined')
            datum_inp.value = datetime.date.today().strftime('%d.%m.%Y')
            cas_od = ui.input('Čas od (HH:MM)').classes('flex-1 bg-white').props('outlined')
            cas_do = ui.input('Čas do (HH:MM)').classes('flex-1 bg-white').props('outlined')

            def normalizuj_cas(inp):
                """Převede '5' → '05:00', '9:5' → '09:05', '930' → '09:30' apod."""
                v = (inp.value or '').strip()
                if not v:
                    return
                # Oddělovač dvojtečkou
                if ':' in v:
                    casti = v.split(':', 1)
                    h = casti[0].strip().zfill(2)
                    m = casti[1].strip().zfill(2)
                else:
                    # Číslo bez dvojtečky: 1–2 číslice = hodiny, 3–4 = HHMM
                    if len(v) <= 2:
                        h, m = v.zfill(2), '00'
                    else:
                        h, m = v[:-2].zfill(2), v[-2:]
                try:
                    if 0 <= int(h) <= 23 and 0 <= int(m) <= 59:
                        inp.value = f'{h}:{m}'
                except ValueError:
                    pass

            cas_od.on('blur', lambda: normalizuj_cas(cas_od))
            cas_do.on('blur', lambda: normalizuj_cas(cas_do))

        _default_mod_zap = user_name if user_name in mod_zap_select else next(iter(mod_zap_select), None)

        with ui.row().classes('w-full gap-3 mb-3 flex-wrap'):
            misto = ui.input('Místo konání').classes('flex-1 bg-white').props('outlined')
            moderator_sel = ui.select(mod_zap_select, label='Moderátor', value=_default_mod_zap).classes('flex-1 bg-white').props('outlined')
            zapisovatel_sel = ui.select(mod_zap_select, label='Zapisovatel', value=_default_mod_zap).classes('flex-1 bg-white').props('outlined')

        ui.label('Účastníci porady').classes('text-sm font-bold text-gray-600 mb-1')

        ucastnici_vyber: dict = {user_id: user_name}
        jmeno_na_id = {jm: uid for uid, jm in ucast_opts.items()}
        vsechna_jmena = sorted(ucast_opts.values())

        chipy = ui.row().classes('w-full gap-2 flex-wrap min-h-8 p-2 border border-gray-200 rounded-lg bg-gray-50 mb-2')

        def prekresli_ucastniky():
            chipy.clear()
            with chipy:
                for uid, jm in sorted(ucastnici_vyber.items(), key=lambda x: x[1]):
                    with ui.row().classes('items-center gap-0.5 bg-blue-100 text-blue-800 px-2 py-0.5 rounded-full'):
                        ui.label(jm).classes('text-xs font-semibold')
                        ui.button(icon='close',
                                  on_click=lambda u=uid: [ucastnici_vyber.pop(u, None), prekresli_ucastniky()]) \
                            .props('flat round dense size=xs').classes('text-blue-400 ml-0.5')

        prekresli_ucastniky()

        with ui.row().classes('w-full gap-2 items-center mb-3'):
            inp_ucastnik = ui.input(
                placeholder='Hledat a přidat účastníka…',
                autocomplete=vsechna_jmena,
            ).classes('flex-1 bg-white').props('outlined dense clearable')

            def pridat_ucastnika():
                jm = (inp_ucastnik.value or '').strip()
                uid = jmeno_na_id.get(jm)
                if not uid:
                    if jm:
                        ui.notify('Uživatel nenalezen — vyberte ze seznamu.', type='warning')
                    return
                if uid in ucastnici_vyber:
                    ui.notify(f'{jm} je již přidán/a.', type='info')
                    inp_ucastnik.value = ''
                    return
                ucastnici_vyber[uid] = jm
                inp_ucastnik.value = ''
                prekresli_ucastniky()

            inp_ucastnik.on('keydown.enter', pridat_ucastnika)
            ui.button(icon='person_add', on_click=pridat_ucastnika) \
                .props('flat round').classes('text-blue-600').tooltip('Přidat účastníka (Enter)')

        def ulozit():
            if not nazev.value or not nazev.value.strip():
                ui.notify('Vyplňte název!', type='warning'); return
            try:
                datum_date = datetime.datetime.strptime(datum_inp.value.strip(), '%d.%m.%Y').date()
            except Exception:
                ui.notify('Neplatný formát data!', type='warning'); return

            mod_jmeno = moderator_sel.value
            mod_id = next((uid for uid, jm in uziv_opts.items() if jm == mod_jmeno), None)
            zap_jmeno = zapisovatel_sel.value
            zap_id = next((uid for uid, jm in uziv_opts.items() if jm == zap_jmeno), None)

            c = intranet_data.get_db_connection()
            cu = c.cursor()
            cu.execute("""
                INSERT INTO ukolovnik_porady
                    (nazev, popis, datum, cas_od, cas_do, misto, moderator_id, moderator_jmeno,
                     zapisovatel_id, zapisovatel_jmeno, vytvoril_id, vytvoril_jmeno)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (nazev.value.strip(), popis.value.strip() if popis.value else '',
                  datum_date, cas_od.value.strip() or None, cas_do.value.strip() or None,
                  misto.value.strip() if misto.value else '',
                  mod_id, mod_jmeno, zap_id, zap_jmeno, user_id, user_name))
            nova_id = cu.lastrowid

            vybrany_ucastnici = []
            for uid, jmeno in ucastnici_vyber.items():
                try:
                    cu.execute("INSERT IGNORE INTO ukolovnik_porada_ucastnici (porada_id, user_id, jmeno) VALUES (%s,%s,%s)",
                               (nova_id, uid, jmeno))
                    if uid != user_id:
                        vybrany_ucastnici.append(uid)
                except Exception:
                    pass
            c.commit(); cu.close(); c.close()

            intranet_logger.log_activity(user_name, "Úkolovník", f"Vytvořena porada: {nazev.value[:40]}")
            _log('porada', nova_id, user_id, user_name, f'Vytvořil/a poradu: {nazev.value[:80]}')
            for uid in vybrany_ucastnici:
                try:
                    intranet_notifikace.pridej(uid, f"📅 {user_name} tě pozval/a na poradu: '{nazev.value[:50]}' ({datum_date.strftime('%d.%m.%Y')})", 'info')
                except Exception:
                    pass
            ui.notify('Porada vytvořena.', type='positive', position='top')
            dlg.close(); on_refresh()

        with ui.row().classes('w-full justify-between mt-4'):
            ui.button('Zrušit', on_click=dlg.close).classes('bg-gray-400 text-white font-bold px-6')
            ui.button('Vytvořit poradu', icon='event', on_click=ulozit).classes('bg-blue-600 text-white font-bold px-6 shadow-md')

    dlg.open()


# =========================================================
# SEKCE: PORADY
# =========================================================
def _nacti_porady_seznam(user_id, is_admin, je_vedouci, odd_user_ids, ma_oddeleni) -> list:
    """Seznam porad dle viditelnosti uživatele — čistý DB loader pro vlákno."""
    conn = intranet_data.get_db_connection()
    porady = []
    if conn:
        try:
            cur = conn.cursor(dictionary=True)
            if is_admin:
                cur.execute("SELECT * FROM ukolovnik_porady ORDER BY datum DESC, vytvoreno DESC")
            elif je_vedouci and ma_oddeleni:
                placeholders = ','.join(['%s'] * len(odd_user_ids))
                cur.execute(f"""
                    SELECT DISTINCT p.* FROM ukolovnik_porady p
                    LEFT JOIN ukolovnik_porada_ucastnici pu ON p.id=pu.porada_id
                    LEFT JOIN ukolovnik_sledovaci s ON s.typ='porada' AND s.ref_id=p.id AND s.user_id=%s
                    WHERE p.vytvoril_id IN ({placeholders})
                       OR p.moderator_id IN ({placeholders})
                       OR pu.user_id IN ({placeholders})
                       OR s.id IS NOT NULL
                    ORDER BY p.datum DESC, p.vytvoreno DESC
                """, [user_id] + odd_user_ids * 3)
            else:
                cur.execute("""
                    SELECT DISTINCT p.* FROM ukolovnik_porady p
                    LEFT JOIN ukolovnik_porada_ucastnici u ON p.id=u.porada_id
                    LEFT JOIN ukolovnik_sledovaci s ON s.typ='porada' AND s.ref_id=p.id AND s.user_id=%s
                    WHERE p.vytvoril_id=%s OR p.moderator_id=%s OR u.user_id=%s OR s.id IS NOT NULL
                    ORDER BY p.datum DESC, p.vytvoreno DESC
                """, (user_id, user_id, user_id, user_id))
            porady = cur.fetchall()
        finally:
            cur.close(); conn.close()
    return porady


@ui.refreshable
async def _vykresli_porady(user_id, user_name, vsechna_prava):
    is_admin = 'vse' in vsechna_prava or 'ukolovnik_admin' in vsechna_prava

    uzivatele_oddeleni = [] if is_admin else await asyncio.to_thread(_ziskej_uzivatele_oddeleni, vsechna_prava)
    je_vedouci = is_admin or _je_hlavni_vedouci(vsechna_prava)
    odd_user_ids = [u['id'] for u in uzivatele_oddeleni] if uzivatele_oddeleni else [user_id]

    porady = await asyncio.to_thread(
        _nacti_porady_seznam, user_id, is_admin, je_vedouci, odd_user_ids, bool(uzivatele_oddeleni))

    # Starý obsah smazat až s daty v ruce → přestavba bez bliknutí a skoku scrollu
    ui.context.slot.parent.clear()

    je_admin_p = 'vse' in vsechna_prava or 'ukolovnik_admin' in vsechna_prava
    muze_zadat_poradu = je_admin_p or 'ukolovnik_porady_zadat' in vsechna_prava

    with ui.row().classes('w-full items-center justify-between mb-6'):
        with ui.column().classes('gap-1'):
            ui.label('Plánování porad').classes('text-3xl font-extrabold text-gray-800')
            ui.label('Plánujte porady, veďte zápisy a zadávejte úkoly přímo z porady.').classes('text-gray-500 text-sm')
        if muze_zadat_poradu:
            ui.button('Nová porada', icon='add_circle',
                      on_click=lambda: _dialog_nova_porada(user_id, user_name, _porady_refresh_klient, vsechna_prava=vsechna_prava)
                      ).classes('bg-blue-600 text-white font-bold h-10 px-5 rounded-xl shadow-md')

    if not porady:
        with ui.card().classes('w-full p-12 items-center justify-center bg-gray-50 border border-dashed border-gray-200 rounded-xl'):
            ui.icon('meeting_room', size='4rem', color='gray-400').classes('mb-4')
            ui.label('Zatím žádné porady.').classes('text-xl text-gray-500 font-bold')
        return

    with ui.row().classes('w-full gap-3 mb-4 flex-wrap items-center'):
        filtr_stav = ui.select(
            {'': 'Všechny', 'Plánovaná': 'Plánované', 'Probíhá': 'Probíhající',
             'Dokončená': 'Dokončené', 'Zrušená': 'Zrušené'},
            value='', label=''
        ).classes('w-40').props('dense outlined')

    seznam = ui.column().classes('w-full gap-3')

    def prekresli():
        seznam.clear()
        filtr = filtr_stav.value or None
        with seznam:
            for p in porady:
                if filtr and p['stav'] != filtr: continue
                b_stav = BARVY_STAVU_PORADA.get(p['stav'], 'bg-gray-100 text-gray-800 border-gray-300')
                datum_str = p['datum'].strftime('%d.%m.%Y') if p.get('datum') else '—'
                cas_od = str(p['cas_od'])[:5] if p.get('cas_od') else ''
                cas_do = str(p['cas_do'])[:5] if p.get('cas_do') else ''
                cas_str = f'{cas_od}–{cas_do}' if cas_od else ''

                c2 = intranet_data.get_db_connection()
                pocet_uk, pocet_hot = 0, 0
                if c2:
                    try:
                        cu = c2.cursor(dictionary=True)
                        cu.execute("SELECT stav FROM ukolovnik_ukoly WHERE porada_id=%s", (p['id'],))
                        ul = cu.fetchall()
                        pocet_uk = len(ul)
                        pocet_hot = sum(1 for u in ul if u['stav'] == 'Hotovo')
                    finally:
                        cu.close(); c2.close()

                barva_pruhu = {'Plánovaná': 'bg-blue-500', 'Probíhá': 'bg-orange-500',
                               'Dokončená': 'bg-green-500', 'Zrušená': 'bg-gray-400'}.get(p['stav'], 'bg-gray-300')

                with ui.card().classes('w-full p-0 overflow-hidden rounded-xl border border-gray-100 hover:shadow-lg transition-shadow cursor-pointer bg-white').on('click', lambda pid=p['id']: _dialog_detail_porady(pid, user_id, user_name, vsechna_prava, on_refresh=_porady_refresh_klient)):
                    with ui.row().classes('w-full items-stretch'):
                        ui.element('div').classes(f'w-1.5 shrink-0 {barva_pruhu}')
                        with ui.column().classes('flex-1 p-4 gap-1'):
                            with ui.row().classes('w-full justify-between items-start'):
                                ui.label(p['nazev']).classes('font-black text-lg text-gray-800 flex-1 line-clamp-1')
                                ui.label(p['stav']).classes(f'px-2 py-0.5 rounded text-[10px] font-bold border whitespace-nowrap ml-2 {b_stav}')
                                ui.button(icon='visibility').props('flat round dense size=xs').classes(
                                    'text-gray-300 hover:text-blue-500 shrink-0 ml-1'
                                ).on('click.stop', lambda pp=p: _dialog_log('porada', pp['id'], pp['nazev'], user_id, user_name))
                                if is_admin or p.get('vytvoril_id') == user_id:
                                    def _del_p(pp=p):
                                        def _provest():
                                            ok = _smazat_poradu_db(pp['id'])
                                            if ok:
                                                intranet_logger.log_activity(user_name, "Úkolovník", f"Smazána porada #{pp['id']}: {pp['nazev']}")
                                                ui.notify('Porada smazána.', type='positive', position='top')
                                                _porady_refresh_klient()
                                            else:
                                                ui.notify('Chyba při mazání.', type='negative', position='top')
                                        _potvrdit_smazat(f'Opravdu smazat poradu „{pp["nazev"]}"?\nBudou smazány i všechny zápisy a úkoly z porady.', _provest)
                                    ui.button(icon='delete_outline').props('flat round dense size=xs').classes(
                                        'text-gray-300 hover:text-red-500 shrink-0 ml-1'
                                    ).on('click.stop', lambda pp=p: _del_p(pp))
                            with ui.row().classes('w-full gap-4 text-sm text-gray-500 items-center flex-wrap'):
                                with ui.row().classes('items-center gap-1'):
                                    ui.icon('calendar_today', size='xs')
                                    ui.label(datum_str).classes('font-bold text-gray-700')
                                if cas_str:
                                    with ui.row().classes('items-center gap-1'):
                                        ui.icon('schedule', size='xs'); ui.label(cas_str)
                                if p.get('misto'):
                                    with ui.row().classes('items-center gap-1'):
                                        ui.icon('location_on', size='xs'); ui.label(p['misto'])
                                with ui.row().classes('items-center gap-1'):
                                    ui.icon('person', size='xs'); ui.label(f'Moderátor: {p.get("moderator_jmeno","—")}')
                            with ui.row().classes('w-full justify-between items-center mt-1 pt-2 border-t border-gray-50'):
                                ui.label(f'Zapisovatel: {p.get("zapisovatel_jmeno","—")}').classes('text-xs text-gray-400')
                                if pocet_uk > 0:
                                    barva_uk = 'text-green-600' if pocet_hot == pocet_uk else 'text-orange-600'
                                    ui.label(f'Úkoly: {pocet_hot}/{pocet_uk}').classes(f'text-xs font-bold {barva_uk}')
                                else:
                                    ui.label('Bez úkolů').classes('text-xs text-gray-300')

    prekresli()
    filtr_stav.on('update:model-value', lambda: prekresli())


# =========================================================
# KAPACITA ODDĚLENÍ — pomocné funkce
# =========================================================
def _ziskej_podrizene(manager_id: int) -> list:
    """Vrátí seznam {'id': int, 'jmeno': str} podřízených daného vedoucího."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT u.iduser AS id, CONCAT(u.name, ' ', u.surname) AS jmeno
            FROM user_manager um
            JOIN user u ON um.user_id = u.iduser
            WHERE um.manager_id = %s AND u.is_active = 1
            ORDER BY u.surname, u.name
        """, (manager_id,))
        return cur.fetchall()
    except Exception:
        return []
    finally:
        cur.close(); conn.close()


def _je_hlavni_vedouci(vsechna_prava) -> bool:
    """True pokud má uživatel právo 'hlavni_vedouci_*' pro alespoň jedno oddělení."""
    return any(p.startswith('hlavni_vedouci_') for p in (vsechna_prava or []))


def _ziskej_nazvy_oddeleni_vedouciho(vsechna_prava) -> list:
    """Vrátí list lowercase názvů oddělení, kde je uživatel hlavním vedoucím."""
    return [p[len('hlavni_vedouci_'):] for p in (vsechna_prava or []) if p.startswith('hlavni_vedouci_')]


def _ziskej_uzivatele_oddeleni(vsechna_prava) -> list:
    """
    Vrátí seznam {'id': int, 'jmeno': str} všech aktivních uživatelů
    v odděleních, kde je aktuální uživatel nastaven jako Hlavní vedoucí.
    Prázdný seznam = uživatel není hlavním vedoucím žádného oddělení.
    """
    dept_names = _ziskej_nazvy_oddeleni_vedouciho(vsechna_prava)
    if not dept_names:
        return []
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        placeholders = ','.join(['%s'] * len(dept_names))
        cur.execute(f"""
            SELECT DISTINCT u.iduser AS id, CONCAT(u.name, ' ', u.surname) AS jmeno
            FROM department d
            JOIN department_To_user dtu ON d.iddepartment = dtu.department_iddepartment
            JOIN user u ON dtu.user_iduser = u.iduser
            WHERE LOWER(d.name) IN ({placeholders}) AND u.is_active = 1
            ORDER BY jmeno
        """, dept_names)
        return cur.fetchall()
    except Exception as e:
        print(f'[Ukolovnik] ziskej_uzivatele_oddeleni: {e}')
        return []
    finally:
        cur.close(); conn.close()


# Ponecháno pro zpětnou kompatibilitu s _vykresli_ukoly filtr osoby (admini)
def _je_vedouci_oddeleni(vsechna_prava) -> bool:
    return _je_hlavni_vedouci(vsechna_prava)


def _ziskej_uziv_sveho_oddeleni(user_id: int) -> list:
    """
    Vrátí {'id', 'jmeno'} všech aktivních uživatelů ve stejném oddělení jako user_id.
    Používá členství v oddělení (department_To_user), ne roli vedoucího.
    """
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT DISTINCT u.iduser AS id, CONCAT(u.name, ' ', u.surname) AS jmeno
            FROM department_To_user dtu1
            JOIN department_To_user dtu2
                ON dtu1.department_iddepartment = dtu2.department_iddepartment
            JOIN user u ON dtu2.user_iduser = u.iduser
            WHERE dtu1.user_iduser = %s AND u.is_active = 1
            ORDER BY jmeno
        """, (user_id,))
        return cur.fetchall()
    except Exception as e:
        print(f'[Ukolovnik] ziskej_uziv_sveho_oddeleni: {e}')
        return []
    finally:
        cur.close(); conn.close()


def _ziskej_uziv_dle_oddeleni(dept_names: list) -> list:
    """Vrátí {'id', 'jmeno'} aktivních uživatelů patřících do zadaných oddělení (podle názvu)."""
    if not dept_names:
        return []
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        placeholders = ','.join(['%s'] * len(dept_names))
        cur.execute(f"""
            SELECT DISTINCT u.iduser AS id, CONCAT(u.name, ' ', u.surname) AS jmeno
            FROM department d
            JOIN department_To_user dtu ON d.iddepartment = dtu.department_iddepartment
            JOIN user u ON dtu.user_iduser = u.iduser
            WHERE LOWER(d.name) IN ({placeholders}) AND u.is_active = 1
            ORDER BY jmeno
        """, [str(n).lower() for n in dept_names])
        return cur.fetchall()
    except Exception as e:
        print(f'[Ukolovnik] ziskej_uziv_dle_oddeleni: {e}')
        return []
    finally:
        cur.close(); conn.close()


def _ziskej_vedouci_vsech_oddeleni() -> dict:
    """
    Vrátí {id: jmeno} aktivních uživatelů, kteří jsou:
    - Hlavní vedoucí jakéhokoliv oddělení (právo hlavni_vedouci_*), nebo
    - Členové oddělení označeného jako Majitelé (department.is_majitele = 1).
    """
    result: dict = {}
    try:
        # 1) Hlavní vedoucí
        vsechna_odd = intranet_data.ziskej_vsechna_oddeleni()
        prava_vedoucich = [f'hlavni_vedouci_{n.lower()}' for n in vsechna_odd.keys()]
        if prava_vedoucich:
            vedouci = intranet_data.ziskej_uzivatele_s_pravem(*prava_vedoucich)
            if isinstance(vedouci, dict):
                result.update(vedouci)

        # 2) Majitelé — členové oddělení s is_majitele = 1
        conn = intranet_data.get_db_connection()
        if conn:
            try:
                cur = conn.cursor(dictionary=True)
                cur.execute("""
                    SELECT DISTINCT u.iduser AS id, CONCAT(u.name, ' ', u.surname) AS jmeno
                    FROM department d
                    JOIN department_To_user dtu ON d.iddepartment = dtu.department_iddepartment
                    JOIN user u ON dtu.user_iduser = u.iduser
                    WHERE d.is_majitele = 1 AND u.is_active = 1
                    ORDER BY jmeno
                """)
                for row in cur.fetchall():
                    result[row['id']] = row['jmeno']
            finally:
                cur.close(); conn.close()
    except Exception as e:
        print(f'[Ukolovnik] ziskej_vedouci_vsech_oddeleni: {e}')
    return result


def _vypocti_vytizeni(user_ids: list, datum_od: datetime.date, datum_do: datetime.date) -> dict:
    """
    Vrátí {user_id: {datum: {'planovane_h': float, 'realne_h': float, 'ukoly': list}}}.
    Plánované hodiny = celý odhad_hodin úkolu se započítá do dne jeho termínu.
    Reálné hodiny = sum(trvani_minut)/60 z time records.
    """
    if not user_ids:
        return {}

    placeholders = ','.join(['%s'] * len(user_ids))
    conn = intranet_data.get_db_connection()
    if not conn:
        return {}

    result: dict = {uid: {} for uid in user_ids}

    try:
        cur = conn.cursor(dictionary=True)

        # Úkoly aktivní v daném rozsahu
        cur.execute(f"""
            SELECT id, nazev, prirazen_id, termin, odhad_hodin, stav, vytvoreno
            FROM ukolovnik_ukoly
            WHERE prirazen_id IN ({placeholders})
              AND stav NOT IN ('Zrušen')
              AND (termin IS NULL OR termin >= %s)
              AND DATE(vytvoreno) <= %s
        """, user_ids + [datum_od, datum_do])
        ukoly = cur.fetchall()

        # Reálné záznamy
        cur.execute(f"""
            SELECT user_id, DATE(cas_start) AS den, SUM(trvani_minut)/60.0 AS hodiny
            FROM ukolovnik_cas_zaznamy
            WHERE user_id IN ({placeholders})
              AND DATE(cas_start) BETWEEN %s AND %s
              AND cas_konec IS NOT NULL
            GROUP BY user_id, DATE(cas_start)
        """, user_ids + [datum_od, datum_do])
        realne = cur.fetchall()

    finally:
        cur.close(); conn.close()

    # Inicializace dnů
    delta = (datum_do - datum_od).days + 1
    vsechny_dny = [datum_od + datetime.timedelta(days=i) for i in range(delta)]
    for uid in user_ids:
        for den in vsechny_dny:
            result[uid][den] = {'planovane_h': 0.0, 'realne_h': 0.0, 'ukoly': []}

    # Plánované hodiny z úkolů — celý odhad se započítá pouze do dne termínu úkolu
    for uk in ukoly:
        uid = uk['prirazen_id']
        termin = uk.get('termin')
        if uid not in result or not termin:
            continue
        # Úkol bez termínu se do plánovaného vytížení nezapočítává.
        if termin in result[uid]:
            odhad = float(uk.get('odhad_hodin') or 0)
            result[uid][termin]['planovane_h'] += odhad
            result[uid][termin]['ukoly'].append(uk['nazev'])

    # Reálné hodiny
    for r in realne:
        uid = r['user_id']
        den = r['den']
        if uid in result and den in result[uid]:
            result[uid][den]['realne_h'] = float(r['hodiny'] or 0)

    return result


def _barva_vytizeni(procenta: float) -> tuple:
    """Vrátí (bg_css, text_css, label) dle procent vytížení."""
    if procenta <= 0:
        return 'bg-gray-100', 'text-gray-400', '—'
    elif procenta <= 50:
        return 'bg-green-100', 'text-green-700', f'{int(procenta)} %'
    elif procenta <= 80:
        return 'bg-lime-100', 'text-lime-700', f'{int(procenta)} %'
    elif procenta <= 100:
        return 'bg-orange-100', 'text-orange-700', f'{int(procenta)} %'
    else:
        return 'bg-red-100', 'text-red-700', f'{int(procenta)} %'


# =========================================================
# KAPACITA ODDĚLENÍ — hlavní sekce
# =========================================================
def _nacti_kapacita_statistiky(zobrazeni, uid_list, dnes, stat_od, stat_do) -> list:
    """Statistiky kapacit (počty úkolů + odpracovaný čas) pro zadané období.

    Dva agregační dotazy pro všechny osoby najednou — čistý DB loader pro vlákno."""
    conn = intranet_data.get_db_connection()
    statistiky = []
    if conn:
        try:
            cur = conn.cursor(dictionary=True)
            placeholders = ','.join(['%s'] * len(uid_list))

            # Počty úkolů
            cur.execute(f"""
                SELECT prirazen_id,
                    COUNT(*) AS celkem,
                    SUM(CASE WHEN stav='Hotovo' THEN 1 ELSE 0 END) AS hotovo,
                    SUM(CASE WHEN stav NOT IN ('Hotovo','Zrušen') AND termin < %s THEN 1 ELSE 0 END) AS po_terminu,
                    SUM(CASE WHEN stav NOT IN ('Hotovo','Zrušen') THEN 1 ELSE 0 END) AS otevrene,
                    COALESCE(SUM(odhad_hodin), 0) AS sum_odhad
                FROM ukolovnik_ukoly
                WHERE prirazen_id IN ({placeholders})
                  AND DATE(vytvoreno) <= %s
                  AND (termin IS NULL OR termin >= %s)
                GROUP BY prirazen_id
            """, [dnes] + uid_list + [stat_do, stat_od])
            ukol_stats = {r['prirazen_id']: r for r in cur.fetchall()}

            # Reálný odpracovaný čas
            cur.execute(f"""
                SELECT user_id, SUM(trvani_minut)/60.0 AS real_h, COUNT(*) AS zaznamu
                FROM ukolovnik_cas_zaznamy
                WHERE user_id IN ({placeholders})
                  AND DATE(cas_start) BETWEEN %s AND %s
                  AND cas_konec IS NOT NULL
                GROUP BY user_id
            """, uid_list + [stat_od, stat_do])
            real_stats = {r['user_id']: r for r in cur.fetchall()}

            for osoba in zobrazeni:
                oid = osoba['id']
                uk = ukol_stats.get(oid, {})
                re = real_stats.get(oid, {})
                pracovnich_dni = sum(
                    1 for i in range((stat_do - stat_od).days + 1)
                    if (stat_od + datetime.timedelta(days=i)).weekday() < 5
                )
                kapacita_h = pracovnich_dni * 8
                real_h = float(re.get('real_h') or 0)
                odhad_h = float(uk.get('sum_odhad') or 0)
                vytiz_plan = min(int((odhad_h / kapacita_h * 100) if kapacita_h else 0), 999)
                vytiz_real = min(int((real_h / kapacita_h * 100) if kapacita_h else 0), 999)

                statistiky.append({
                    'jmeno': osoba['jmeno'],
                    'celkem': int(uk.get('celkem') or 0),
                    'hotovo': int(uk.get('hotovo') or 0),
                    'otevrene': int(uk.get('otevrene') or 0),
                    'po_terminu': int(uk.get('po_terminu') or 0),
                    'odhad_h': odhad_h,
                    'real_h': real_h,
                    'kapacita_h': kapacita_h,
                    'vytiz_plan': vytiz_plan,
                    'vytiz_real': vytiz_real,
                })
        finally:
            cur.close(); conn.close()
    return statistiky


@ui.refreshable
async def _vykresli_kapacita(user_id, user_name, vsechna_prava, kap_stav: dict):
    """
    kap_stav = {
        'pohled': 'moje'|'osoba'|'oddeleni',
        'vybrany_uid': int|None,
        'tyden_offset': int,        # 0 = aktuální týden
        'stat_od': str (DD.MM.RRRR),
        'stat_do': str,
        'zobraz_stat': bool,
    }
    """
    is_admin = 'vse' in vsechna_prava or 'ukolovnik_admin' in vsechna_prava

    if is_admin:
        uziv_opts = await asyncio.to_thread(_ziskej_uzivatele_options)
        podrizeni = [{'id': uid, 'jmeno': jmeno} for uid, jmeno in uziv_opts.items()]
    elif _je_hlavni_vedouci(vsechna_prava):
        uzivatele_oddeleni = await asyncio.to_thread(_ziskej_uzivatele_oddeleni, vsechna_prava)
        podrizeni = [u for u in uzivatele_oddeleni if u['id'] != user_id]
    else:
        podrizeni = await asyncio.to_thread(_ziskej_podrizene, user_id)

    je_vedouci = bool(podrizeni) or is_admin

    # Starý obsah smazat až s daty v ruce → přestavba bez bliknutí a skoku scrollu
    ui.context.slot.parent.clear()

    dnes = datetime.date.today()

    # ── Hlavička ──────────────────────────────────────────────────────────────
    with ui.row().classes('w-full items-center justify-between mb-4 flex-wrap gap-3'):
        with ui.column().classes('gap-0'):
            ui.label('Kapacita oddělení').classes('text-3xl font-extrabold text-gray-800')
            ui.label('Přehled vytížení pracovníků dle plánovaných úkolů a skutečně odpracovaného času.').classes('text-xs text-gray-400')

        with ui.row().classes('items-center gap-2'):
            # Přecvakávač pohledu (jen pro vedoucí)
            if je_vedouci:
                def _set_pohled(p):
                    kap_stav['pohled'] = p
                    _kapacita_refresh_klient()

                pohled = kap_stav.get('pohled', 'moje')
                with ui.button_group():
                    for key, label in [('moje', 'Moje'), ('osoba', 'Osoba'), ('oddeleni', 'Oddělení')]:
                        if pohled == key:
                            ui.button(label, on_click=lambda k=key: _set_pohled(k)).props('no-caps unelevated color=blue-7')
                        else:
                            ui.button(label, on_click=lambda k=key: _set_pohled(k)).props('no-caps outline color=blue-7')

            # Statistika tlačítko
            def _toggle_stat():
                kap_stav['zobraz_stat'] = not kap_stav.get('zobraz_stat', False)
                _kapacita_refresh_klient()
            stat_active = 'bg-indigo-600 text-white' if kap_stav.get('zobraz_stat') else 'bg-white text-indigo-600 border border-indigo-300'
            ui.button('Statistiky oddělení', icon='bar_chart', on_click=_toggle_stat).classes(f'font-bold px-4 rounded-xl {stat_active}')

    # ── Selector osoby (pohled = 'osoba') ─────────────────────────────────────
    pohled = kap_stav.get('pohled', 'moje')
    if je_vedouci and pohled == 'osoba' and podrizeni:
        osoba_opts = {str(p['id']): p['jmeno'] for p in podrizeni}
        vybrany_str = str(kap_stav.get('vybrany_uid') or next(iter(osoba_opts)))
        def _set_osoba(e):
            kap_stav['vybrany_uid'] = int(e.value)
            _kapacita_refresh_klient()
        ui.select(osoba_opts, value=vybrany_str, label='Vybrat pracovníka', on_change=_set_osoba).classes('w-64 mb-3').props('outlined dense')

    # ── Sestavení seznamu zobrazených uživatelů ────────────────────────────────
    if not je_vedouci or pohled == 'moje':
        zobrazeni = [{'id': user_id, 'jmeno': user_name}]
    elif pohled == 'osoba':
        vid = kap_stav.get('vybrany_uid') or (podrizeni[0]['id'] if podrizeni else user_id)
        zobrazeni = [p for p in podrizeni if p['id'] == vid] or [{'id': user_id, 'jmeno': user_name}]
    else:  # oddeleni
        zobrazeni = podrizeni if podrizeni else [{'id': user_id, 'jmeno': user_name}]

    uid_list = [p['id'] for p in zobrazeni]

    # ── GANTT — týdenní pohled ─────────────────────────────────────────────────
    offset = kap_stav.get('tyden_offset', 0)
    # Pondělí aktuálního / posunutého týdne
    pondeli = dnes - datetime.timedelta(days=dnes.weekday()) + datetime.timedelta(weeks=offset)
    nedele = pondeli + datetime.timedelta(days=6)

    with ui.card().classes('w-full p-4 bg-white border border-gray-100 rounded-xl shadow-sm mb-4'):
        # Navigace týdne
        with ui.row().classes('w-full items-center justify-between mb-4'):
            with ui.row().classes('items-center gap-2'):
                ui.button(icon='chevron_left', on_click=lambda: [kap_stav.update({'tyden_offset': offset - 1}), _kapacita_refresh_klient()]).props('flat round dense')
                tydne_label = f'{pondeli.strftime("%d.%m.")} – {nedele.strftime("%d.%m.%Y")}'
                ui.label(tydne_label).classes('text-base font-black text-gray-700 w-44 text-center')
                ui.button(icon='chevron_right', on_click=lambda: [kap_stav.update({'tyden_offset': offset + 1}), _kapacita_refresh_klient()]).props('flat round dense')
                if offset != 0:
                    ui.button('Dnes', on_click=lambda: [kap_stav.update({'tyden_offset': 0}), _kapacita_refresh_klient()]).props('flat dense').classes('text-xs text-blue-600')

            # Legenda
            with ui.row().classes('items-center gap-3 text-xs text-gray-500'):
                for bg, tc, label in [('bg-green-100', 'text-green-700', '0–50 %'),
                                      ('bg-lime-100', 'text-lime-700', '51–80 %'),
                                      ('bg-orange-100', 'text-orange-700', '81–100 %'),
                                      ('bg-red-100', 'text-red-700', '> 100 %')]:
                    with ui.row().classes('items-center gap-1'):
                        ui.element('div').classes(f'w-3 h-3 rounded {bg}')
                        ui.label(label).classes(tc)

        # Výpočet dat
        vytizeni = _vypocti_vytizeni(uid_list, pondeli, nedele)

        # Záhlaví: dny
        dny_tydne = [pondeli + datetime.timedelta(days=i) for i in range(7)]
        with ui.element('div').classes('w-full overflow-x-auto'):
            with ui.element('table').classes('w-full border-collapse text-sm'):
                # Hlavička
                with ui.element('thead'):
                    with ui.element('tr'):
                        with ui.element('th').classes('text-left p-2 text-gray-500 font-bold w-36 border-b border-gray-100'):
                            ui.label('Pracovník')
                        for den in dny_tydne:
                            je_dnes = den == dnes
                            je_vikend = den.weekday() >= 5
                            den_cls = 'text-blue-700 font-extrabold' if je_dnes else ('text-gray-400' if je_vikend else 'text-gray-600 font-bold')
                            th = ui.element('th').classes(f'text-center p-2 border-b border-gray-100 w-20 {den_cls}')
                            with th:
                                ui.label(DNY_CZ[den.weekday()]).classes('block text-xs')
                                ui.label(den.strftime('%d.%m.')).classes('block text-xs')

                # Řádky pracovníků
                with ui.element('tbody'):
                    for osoba in zobrazeni:
                        oid = osoba['id']
                        with ui.element('tr').classes('hover:bg-gray-50'):
                            td_jmeno = ui.element('td').classes('p-2 font-bold text-gray-700 border-b border-gray-50 w-36')
                            with td_jmeno:
                                ui.label(osoba['jmeno']).classes('text-sm truncate max-w-[140px] block')

                            for den in dny_tydne:
                                den_data = vytizeni.get(oid, {}).get(den, {'planovane_h': 0, 'realne_h': 0, 'ukoly': []})
                                plan_h = den_data['planovane_h']
                                real_h = den_data['realne_h']
                                je_vikend = den.weekday() >= 5

                                if je_vikend:
                                    td = ui.element('td').classes('p-1 border-b border-gray-50 bg-gray-50 w-20')
                                    with td:
                                        ui.label('').classes('block')
                                    continue

                                # Využij reálné hodiny pokud existují, jinak plánované
                                zobraz_h = real_h if real_h > 0 else plan_h
                                procenta = (zobraz_h / 8.0) * 100
                                bg, tc, label = _barva_vytizeni(procenta)

                                td = ui.element('td').classes(f'p-1 border-b border-gray-50 text-center w-20')
                                with td:
                                    tooltip_text = f'Plánováno: {plan_h:.1f}h\nSkutečně: {real_h:.1f}h'
                                    if den_data['ukoly']:
                                        tooltip_text += '\nÚkoly:\n' + '\n'.join(f'• {u[:30]}' for u in den_data['ukoly'][:5])

                                    with ui.element('div').classes(f'rounded-md px-1 py-1.5 {bg} cursor-default'):
                                        ui.label(label if zobraz_h > 0 else '—').classes(f'text-xs font-bold {tc} block')
                                        if real_h > 0:
                                            ui.label(f'{real_h:.1f}h').classes('text-[10px] text-gray-500 block')
                                        elif plan_h > 0:
                                            ui.label(f'{plan_h:.1f}h').classes('text-[10px] text-gray-400 block')
                                        ui.tooltip(tooltip_text)

    # ── STATISTIKY ────────────────────────────────────────────────────────────
    if kap_stav.get('zobraz_stat'):
        with ui.card().classes('w-full p-5 bg-white border border-gray-100 rounded-xl shadow-sm'):
            ui.label('Statistiky vytížení').classes('text-lg font-extrabold text-gray-800 mb-4')

            # Výběr období
            with ui.row().classes('w-full gap-3 mb-5 items-end flex-wrap'):
                stat_od_inp = ui.input('Od (DD.MM.RRRR)', value=kap_stav.get('stat_od', (dnes - datetime.timedelta(days=29)).strftime('%d.%m.%Y'))).classes('w-40').props('outlined dense')
                stat_do_inp = ui.input('Do (DD.MM.RRRR)', value=kap_stav.get('stat_do', dnes.strftime('%d.%m.%Y'))).classes('w-40').props('outlined dense')

                def nacti_statistiky():
                    kap_stav['stat_od'] = stat_od_inp.value
                    kap_stav['stat_do'] = stat_do_inp.value
                    _kapacita_refresh_klient()

                ui.button('Zobrazit', icon='search', on_click=nacti_statistiky).classes('bg-indigo-600 text-white font-bold px-5 rounded-xl')

            # Výpočet statistik pro zadané období
            try:
                stat_od = datetime.datetime.strptime(kap_stav.get('stat_od', (dnes - datetime.timedelta(days=29)).strftime('%d.%m.%Y')), '%d.%m.%Y').date()
                stat_do = datetime.datetime.strptime(kap_stav.get('stat_do', dnes.strftime('%d.%m.%Y')), '%d.%m.%Y').date()
            except Exception:
                stat_od = dnes - datetime.timedelta(days=29)
                stat_do = dnes

            statistiky = await asyncio.to_thread(
                _nacti_kapacita_statistiky, zobrazeni, uid_list, dnes, stat_od, stat_do)

            if not statistiky:
                ui.label('Žádná data pro toto období.').classes('text-gray-400 italic')
            else:
                # Tabulka statistik
                with ui.element('div').classes('w-full overflow-x-auto'):
                    with ui.element('table').classes('w-full text-sm border-collapse'):
                        # Záhlaví
                        with ui.element('thead'):
                            with ui.element('tr').classes('border-b-2 border-gray-200'):
                                for col, w in [
                                    ('Pracovník', 'w-40'), ('Otevřené úkoly', 'w-28'), ('Hotovo', 'w-20'),
                                    ('Po termínu', 'w-24'), ('Odhad (h)', 'w-24'),
                                    ('Odpracováno (h)', 'w-32'), ('Kapacita', 'w-24'),
                                    ('Vytížení plán', 'w-28'), ('Vytížení reál', 'w-28'),
                                ]:
                                    with ui.element('th').classes(f'p-2 text-left text-xs font-bold text-gray-500 uppercase {w}'):
                                        ui.label(col)

                        with ui.element('tbody'):
                            for s in statistiky:
                                _, bg_plan, _ = _barva_vytizeni(s['vytiz_plan'])
                                _, bg_real, _ = _barva_vytizeni(s['vytiz_real'])
                                po_term_cls = 'text-red-600 font-bold' if s['po_terminu'] > 0 else 'text-gray-500'

                                with ui.element('tr').classes('border-b border-gray-100 hover:bg-gray-50'):
                                    with ui.element('td').classes('p-2 font-bold text-gray-800'): ui.label(s['jmeno'])
                                    with ui.element('td').classes('p-2 text-center text-gray-700'): ui.label(str(s['otevrene']))
                                    with ui.element('td').classes('p-2 text-center text-green-600 font-bold'): ui.label(str(s['hotovo']))
                                    with ui.element('td').classes(f'p-2 text-center {po_term_cls}'): ui.label(str(s['po_terminu']))
                                    with ui.element('td').classes('p-2 text-center text-gray-600'): ui.label(f"{s['odhad_h']:.1f}")
                                    with ui.element('td').classes('p-2 text-center text-blue-600 font-bold'): ui.label(f"{s['real_h']:.1f}")
                                    with ui.element('td').classes('p-2 text-center text-gray-500'): ui.label(f"{s['kapacita_h']:.0f} h")

                                    # Progress bar — plánované vytížení
                                    with ui.element('td').classes('p-2'):
                                        bg_p, tc_p, _ = _barva_vytizeni(s['vytiz_plan'])
                                        with ui.element('div').classes('w-full bg-gray-100 rounded-full h-4 overflow-hidden'):
                                            fill = min(s['vytiz_plan'], 100)
                                            ui.element('div').classes(f'h-full rounded-full {bg_p}').style(f'width:{fill}%')
                                        ui.label(f"{s['vytiz_plan']} %").classes(f'text-xs {tc_p} font-bold text-center block mt-0.5')

                                    # Progress bar — reálné vytížení
                                    with ui.element('td').classes('p-2'):
                                        bg_r, tc_r, _ = _barva_vytizeni(s['vytiz_real'])
                                        with ui.element('div').classes('w-full bg-gray-100 rounded-full h-4 overflow-hidden'):
                                            fill = min(s['vytiz_real'], 100)
                                            ui.element('div').classes(f'h-full rounded-full {bg_r}').style(f'width:{fill}%')
                                        ui.label(f"{s['vytiz_real']} %").classes(f'text-xs {tc_r} font-bold text-center block mt-0.5')


# =========================================================
# STATISTIKA — dashboardové funkce
# =========================================================
def _nacti_statistiku_osoby(uid: int, obdobi: str = 'tyden') -> dict:
    """Vrátí kompletní statistiku pro jednu osobu.

    obdobi: 'tyden' (výchozí) | 'mesic' — určuje, za jaké období se počítají
    časově omezené metriky (hotovo, odpracováno, porady, blížící se termíny)."""
    dnes = datetime.date.today()
    tyden_od = dnes - datetime.timedelta(days=dnes.weekday())   # pondělí aktuálního týdne
    tyden_pa = tyden_od + datetime.timedelta(days=4)            # pátek aktuálního týdne
    mesic_od = dnes.replace(day=1)
    # poslední den aktuálního měsíce
    if dnes.month == 12:
        mesic_do = dnes.replace(day=31)
    else:
        mesic_do = dnes.replace(month=dnes.month + 1, day=1) - datetime.timedelta(days=1)
    # Hranice pro období-závislé metriky (od kdy počítáme „za období")
    obdobi_od = tyden_od if obdobi == 'tyden' else mesic_od
    obdobi_do = (tyden_od + datetime.timedelta(days=6)) if obdobi == 'tyden' else mesic_do
    # Předchozí srovnávací období (pro trend „vs. minulý týden/měsíc")
    if obdobi == 'tyden':
        prev_od = tyden_od - datetime.timedelta(days=7)
        prev_do = tyden_od - datetime.timedelta(days=1)
    else:
        prev_do = mesic_od - datetime.timedelta(days=1)
        prev_od = prev_do.replace(day=1)

    conn = intranet_data.get_db_connection()
    if not conn:
        return {}
    try:
        cur = conn.cursor(dictionary=True)

        # Úkoly dle stavu (všechny přidělené)
        cur.execute("""
            SELECT stav, COUNT(*) AS pocet
            FROM ukolovnik_ukoly
            WHERE prirazen_id=%s
            GROUP BY stav
        """, (uid,))
        stavy_raw = {r['stav']: r['pocet'] for r in cur.fetchall()}

        # Termíny
        cur.execute("""
            SELECT
                SUM(CASE WHEN termin < %s AND stav NOT IN ('Hotovo','Zrušen') THEN 1 ELSE 0 END) AS po_terminu,
                SUM(CASE WHEN termin = %s AND stav NOT IN ('Hotovo','Zrušen') THEN 1 ELSE 0 END) AS dnes_termin,
                SUM(CASE WHEN termin BETWEEN %s AND %s AND stav NOT IN ('Hotovo','Zrušen') THEN 1 ELSE 0 END) AS tento_tyden,
                SUM(CASE WHEN termin BETWEEN %s AND %s AND stav NOT IN ('Hotovo','Zrušen') THEN 1 ELSE 0 END) AS tyden_po_pa,
                SUM(CASE WHEN termin BETWEEN %s AND %s AND stav NOT IN ('Hotovo','Zrušen') THEN 1 ELSE 0 END) AS mesic_termin,
                SUM(CASE WHEN stav NOT IN ('Hotovo','Zrušen') THEN 1 ELSE 0 END) AS otevrene
            FROM ukolovnik_ukoly WHERE prirazen_id=%s
        """, (dnes, dnes, dnes + datetime.timedelta(days=1), dnes + datetime.timedelta(days=7),
              tyden_od, tyden_pa, mesic_od, mesic_do, uid))
        terminy = cur.fetchone() or {}

        # Hotovo za zvolené období + kvalita plnění termínů.
        # Datum dokončení: dokonceno_datum (novější záznamy), u starých fallback
        # na vytvoreno, aby historie nevypadla ze statistiky úplně.
        cur.execute("""
            SELECT
                COUNT(*) AS n,
                SUM(CASE WHEN termin IS NOT NULL THEN 1 ELSE 0 END) AS s_terminem,
                SUM(CASE WHEN termin IS NOT NULL
                          AND DATE(COALESCE(dokonceno_datum, vytvoreno)) <= termin
                         THEN 1 ELSE 0 END) AS vcas
            FROM ukolovnik_ukoly
            WHERE prirazen_id=%s AND stav='Hotovo'
              AND DATE(COALESCE(dokonceno_datum, vytvoreno)) BETWEEN %s AND %s
        """, (uid, obdobi_od, obdobi_do))
        _hot = cur.fetchone() or {}
        hotovo_obdobi = int(_hot.get('n') or 0)
        hotovo_s_term = int(_hot.get('s_terminem') or 0)
        hotovo_vcas = int(_hot.get('vcas') or 0)

        # Hotovo v předchozím období — pro trendovou šipku
        cur.execute("""
            SELECT COUNT(*) AS n FROM ukolovnik_ukoly
            WHERE prirazen_id=%s AND stav='Hotovo'
              AND DATE(COALESCE(dokonceno_datum, vytvoreno)) BETWEEN %s AND %s
        """, (uid, prev_od, prev_do))
        hotovo_predchozi = int((cur.fetchone() or {}).get('n') or 0)

        # Otevřené úkoly bez termínu (riziko „zapadnutí")
        cur.execute("""
            SELECT COUNT(*) AS n FROM ukolovnik_ukoly
            WHERE prirazen_id=%s AND stav NOT IN ('Hotovo','Zrušen') AND termin IS NULL
        """, (uid,))
        bez_terminu = int((cur.fetchone() or {}).get('n') or 0)

        # Odpracováno za zvolené období — po dnech (pro sloupcový graf i součet)
        cur.execute("""
            SELECT DATE(cas_start) AS d, COALESCE(SUM(trvani_minut)/60.0, 0) AS hodiny
            FROM ukolovnik_cas_zaznamy
            WHERE user_id=%s AND DATE(cas_start) BETWEEN %s AND %s AND cas_konec IS NOT NULL
            GROUP BY DATE(cas_start)
        """, (uid, obdobi_od, obdobi_do))
        po_dnech = {}
        for r in cur.fetchall():
            _d = r['d']
            if hasattr(_d, 'date'):
                _d = _d.date()
            po_dnech[_d] = float(r['hodiny'] or 0)
        odprac_obdobi = sum(po_dnech.values())

        # Projekty
        cur.execute("""
            SELECT COUNT(DISTINCT p.id) AS n FROM ukolovnik_projekty p
            LEFT JOIN ukolovnik_projekt_clenove c ON p.id=c.projekt_id
            WHERE (p.vytvoril_id=%s OR c.user_id=%s) AND p.stav='Aktivní'
        """, (uid, uid))
        aktivni_projekty = (cur.fetchone() or {}).get('n', 0)

        # Porady za zvolené období
        cur.execute("""
            SELECT COUNT(DISTINCT p.id) AS n FROM ukolovnik_porady p
            LEFT JOIN ukolovnik_porada_ucastnici u ON p.id=u.porada_id
            WHERE (p.vytvoril_id=%s OR p.moderator_id=%s OR u.user_id=%s) AND p.datum >= %s
        """, (uid, uid, uid, obdobi_od))
        porady_obdobi = (cur.fetchone() or {}).get('n', 0)

        # Porady celkem (jakékoli, kde je osoba součástí) — pro rozhodnutí o zobrazení dlaždice
        cur.execute("""
            SELECT COUNT(DISTINCT p.id) AS n FROM ukolovnik_porady p
            LEFT JOIN ukolovnik_porada_ucastnici u ON p.id=u.porada_id
            WHERE (p.vytvoril_id=%s OR p.moderator_id=%s OR u.user_id=%s)
        """, (uid, uid, uid))
        porady_celkem = (cur.fetchone() or {}).get('n', 0)

        # Nedávno dokončené (posledních 5)
        cur.execute("""
            SELECT id, nazev, termin,
                   DATE(COALESCE(dokonceno_datum, vytvoreno)) AS created_date
            FROM ukolovnik_ukoly
            WHERE prirazen_id=%s AND stav='Hotovo'
            ORDER BY COALESCE(dokonceno_datum, vytvoreno) DESC LIMIT 5
        """, (uid,))
        nedavno_hotovo = cur.fetchall()

        # Nadcházející termíny (příštích 14 dní)
        cur.execute("""
            SELECT id, nazev, termin, stav, priorita
            FROM ukolovnik_ukoly
            WHERE prirazen_id=%s AND stav NOT IN ('Hotovo','Zrušen')
              AND termin BETWEEN %s AND %s
            ORDER BY termin ASC LIMIT 7
        """, (uid, dnes, dnes + datetime.timedelta(days=14)))
        nadchazejici = cur.fetchall()

        # Úkoly po termínu (pro panel „Vyžaduje pozornost")
        cur.execute("""
            SELECT id, nazev, termin, priorita
            FROM ukolovnik_ukoly
            WHERE prirazen_id=%s AND stav NOT IN ('Hotovo','Zrušen') AND termin < %s
            ORDER BY termin ASC LIMIT 7
        """, (uid, dnes))
        po_termine_ukoly = cur.fetchall()

    finally:
        cur.close(); conn.close()

    # ── Odvozené metriky (mimo DB) ────────────────────────────────────────────
    # Série odpracovaných hodin: týden = po dnech, měsíc = po kalendářních týdnech
    if obdobi == 'tyden':
        serie_hodin = [
            {'lbl': DNY_CZ[i],
             'h': po_dnech.get(tyden_od + datetime.timedelta(days=i), 0.0),
             'akt': (tyden_od + datetime.timedelta(days=i)) == dnes}
            for i in range(7)
        ]
    else:
        serie_hodin = []
        _b_od, _idx = mesic_od, 1
        while _b_od <= mesic_do:
            _b_do = min(_b_od + datetime.timedelta(days=6 - _b_od.weekday()), mesic_do)
            serie_hodin.append({
                'lbl': f'{_idx}. t',
                'h': sum(v for k, v in po_dnech.items() if _b_od <= k <= _b_do),
                'akt': _b_od <= dnes <= _b_do,
            })
            _b_od, _idx = _b_do + datetime.timedelta(days=1), _idx + 1

    # Kapacita období = pracovní dny × 8 h (víkendy se nepočítají)
    _prac_dni = sum(1 for i in range((obdobi_do - obdobi_od).days + 1)
                    if (obdobi_od + datetime.timedelta(days=i)).weekday() < 5)
    kapacita_h = _prac_dni * 8

    return {
        'zadano':       stavy_raw.get('Zadáno', 0),
        'rozpracovano': stavy_raw.get('Rozpracováno', 0),
        'pozastaveno':  stavy_raw.get('Pozastaveno', 0),
        'hotovo_celk':  stavy_raw.get('Hotovo', 0),
        'zrusen':       stavy_raw.get('Zrušen', 0),
        'otevrene':     int(terminy.get('otevrene') or 0),
        'po_terminu':   int(terminy.get('po_terminu') or 0),
        'dnes_termin':  int(terminy.get('dnes_termin') or 0),
        'tento_tyden':  int(terminy.get('tento_tyden') or 0),
        'tyden_po_pa':  int(terminy.get('tyden_po_pa') or 0),
        'mesic_termin': int(terminy.get('mesic_termin') or 0),
        'obdobi':        obdobi,
        'hotovo_obdobi': int(hotovo_obdobi),
        'odprac_obdobi': odprac_obdobi,
        'aktivni_proj': int(aktivni_projekty),
        'porady_obdobi': int(porady_obdobi),
        'porady_celkem': int(porady_celkem),
        'nedavno_hotovo': nedavno_hotovo,
        'nadchazejici':   nadchazejici,
        # ── metriky pro dashboard vedoucího ──
        'hotovo_predchozi': hotovo_predchozi,
        'hotovo_vcas':      hotovo_vcas,
        'hotovo_s_term':    hotovo_s_term,
        'plneni_pct':       int(round(hotovo_vcas / hotovo_s_term * 100)) if hotovo_s_term else None,
        'bez_terminu':      bez_terminu,
        'kapacita_h':       kapacita_h,
        'vytiz_pct':        int(round(odprac_obdobi / kapacita_h * 100)) if kapacita_h else 0,
        'serie_hodin':      serie_hodin,
        'po_termine_ukoly': po_termine_ukoly,
    }


# ── Stavební prvky statistiky (sdílené napříč pohledy) ───────────────────────
_STAT_PRIO_HEX = {'Nízká': '#94a3b8', 'Normální': '#2563eb',
                  'Vysoká': '#ea580c', 'Kritická': '#b91c1c'}


def _stat_iniciály(jmeno: str) -> str:
    casti = [c for c in (jmeno or '').replace('.', ' ').split() if c[:1].isalpha()]
    if not casti:
        return '?'
    if len(casti) == 1:
        return casti[0][:2].upper()
    return (casti[0][:1] + casti[-1][:1]).upper()


def _stat_avatar(jmeno: str, px: int = 34):
    """Barevný čtvereček s iniciálami — stejná barva jako v kalendáři."""
    with ui.element('div').style(
            f'width:{px}px;height:{px}px;border-radius:{max(7, px // 3)}px;'
            f'background:{_barva_osoby(jmeno)};color:#fff;flex-shrink:0;'
            f'display:flex;align-items:center;justify-content:center;'
            f'font-weight:800;font-size:{max(10, int(px * 0.36))}px;letter-spacing:.4px'):
        ui.label(_stat_iniciály(jmeno))


def _stat_seg(volby, aktivni, on_set):
    """Segmentovaný přepínač ve stejném stylu jako v kalendáři."""
    with ui.row().classes('items-center gap-0 bg-gray-100 rounded-lg p-0.5'):
        for kod, popis in volby:
            akt = (kod == aktivni)
            cls = ('px-3 py-1 rounded-md text-xs font-bold cursor-pointer transition-colors '
                   + ('bg-white text-blue-700 shadow-sm' if akt else 'text-gray-500 hover:text-gray-700'))
            ui.label(popis).classes(cls).on('click', lambda k=kod: on_set(k))


def _stat_pruh(pct: int, barva: str, vyska: int = 6):
    """Vodorovný ukazatel 0–100 % (přetečení se vizuálně ořízne na 100 %)."""
    with ui.element('div').classes('w-full rounded-full overflow-hidden').style(
            f'height:{vyska}px;background:#eef2f7'):
        ui.element('div').classes('h-full rounded-full transition-all').style(
            f'width:{max(0, min(pct, 100))}%;background:{barva}')


def _stat_barva_vytizeni(pct: int) -> str:
    if pct >= 100: return '#e11d48'     # přetížení
    if pct >= 85:  return '#059669'     # ideál
    if pct >= 60:  return '#2563eb'     # v pořádku
    return '#94a3b8'                    # málo vykázáno


def _stat_kpi(nazev: str, hodnota, podtext: str = '', barva: str = '#1f2937',
              trend: int = None, alert: bool = False):
    """Jedna dlaždice KPI pásu: nadpis, velké číslo, popisek, volitelný trend."""
    ram = 'border-red-200 bg-red-50' if alert else 'border-gray-100 bg-white'
    with ui.card().classes(f'flex-1 min-w-[148px] p-4 rounded-xl border {ram} shadow-none gap-0'):
        ui.label(nazev.upper()).classes('text-[10px] font-extrabold text-gray-400 tracking-wider mb-1')
        with ui.row().classes('items-baseline gap-2 no-wrap'):
            ui.label(str(hodnota)).classes('text-[34px] font-black leading-none').style(f'color:{barva}')
            if trend is not None and trend != 0:
                kladny = trend > 0
                chip = 'text-green-700 bg-green-50' if kladny else 'text-red-700 bg-red-50'
                ui.label(f"{'▲' if kladny else '▼'} {abs(trend)}").classes(
                    f'text-[10px] font-extrabold px-1.5 py-0.5 rounded {chip}')
        ui.label(podtext or ' ').classes('text-[11px] text-gray-400 mt-1 truncate')


def _stat_sloupce(serie, nadpis: str, podnadpis: str = ''):
    """Sloupcový graf odpracovaných hodin (aktuální den/týden zvýrazněn)."""
    with ui.card().classes('w-full p-5 rounded-xl border border-gray-100 bg-white shadow-none'):
        with ui.row().classes('w-full justify-between items-center mb-4'):
            ui.label(nadpis).classes('text-base font-extrabold text-gray-700')
            if podnadpis:
                ui.label(podnadpis).classes('text-xs text-gray-400')
        max_h = max([s['h'] for s in serie] or [0]) or 1
        with ui.row().classes('w-full items-end gap-2 no-wrap').style('height:120px'):
            for s in serie:
                with ui.column().classes('flex-1 items-center gap-1 h-full justify-end'):
                    ui.label(f"{s['h']:.1f}".replace('.0', '') if s['h'] else '').classes(
                        'text-[10px] font-bold text-gray-400')
                    ui.element('div').classes('w-full rounded-t-md transition-all').style(
                        f"height:{max(2, int(s['h'] / max_h * 78))}px;"
                        f"background:{'#2563eb' if s['akt'] else '#cbd5e1'}")
                    ui.label(s['lbl']).classes(
                        'text-[11px] font-bold ' + ('text-blue-700' if s['akt'] else 'text-gray-400'))


def _vykresli_stat_karty(st: dict, jmeno: str, zobraz_projekty: bool = True, zobraz_porady: bool = True,
                         zobraz_kpi: bool = True):
    """Vykreslí dashboard pro jednu osobu.

    zobraz_projekty: skryje dlaždici „Aktivní projekty", pokud uživatel nemá
    právo na projekty nebo není členem žádného projektu.
    zobraz_porady: skryje dlaždici „Porady", pokud uživatel nemá právo na
    porady nebo není součástí žádné porady.
    zobraz_kpi: vypne úvodní řádek s velkými čísly (statistika si dnes kreslí
    vlastní KPI pás, aby se čísla neopakovala dvakrát pod sebou)."""
    dnes = datetime.date.today()
    obdobi = st.get('obdobi', 'tyden')
    je_tyden = obdobi == 'tyden'
    obd_lbl = 'týden' if je_tyden else 'měsíc'          # do závorky („(týden)")
    obd_lbl2 = 'tento týden' if je_tyden else 'tento měsíc'

    # ── Řádek 1: Klíčová čísla ───────────────────────────────────────────────
    if zobraz_kpi:
        with ui.row().classes('w-full gap-4 mb-4 flex-wrap'):

            def _karta_cislo(ikona, barva_bg, barva_ico, nazev, hodnota, podtext='', alert=False):
                alert_cls = 'ring-2 ring-red-400' if alert else ''
                with ui.card().classes(f'flex-1 min-w-[150px] p-4 rounded-xl border border-gray-100 {barva_bg} {alert_cls}'):
                    with ui.row().classes('w-full justify-between items-start mb-2'):
                        ui.icon(ikona, size='md').classes(barva_ico)
                        if alert and int(hodnota) > 0:
                            ui.badge(str(hodnota), color='red').classes('text-xs font-bold')
                    ui.label(str(hodnota)).classes('text-4xl font-black text-gray-800 leading-none')
                    ui.label(nazev).classes('text-sm font-bold text-gray-600 mt-1')
                    if podtext:
                        ui.label(podtext).classes('text-xs text-gray-400 mt-0.5')

            _karta_cislo('assignment', 'bg-blue-50', 'text-blue-500',
                         'Otevřené úkoly', st['otevrene'], f"Zadáno: {st['zadano']}  Rozpr.: {st['rozpracovano']}")
            _karta_cislo('warning', 'bg-red-50', 'text-red-500',
                         'Po termínu', st['po_terminu'], 'Nutná okamžitá akce', alert=st['po_terminu'] > 0)
            _karta_cislo('today', 'bg-orange-50', 'text-orange-500',
                         f'Termín {obd_lbl2}',
                         st['tyden_po_pa'] if je_tyden else st['mesic_termin'],
                         'Pondělí–pátek' if je_tyden else 'Do konce měsíce')
            _karta_cislo('check_circle', 'bg-green-50', 'text-green-500',
                         f'Hotovo {obd_lbl2}', st['hotovo_obdobi'], 'Dokončené úkoly')
            _karta_cislo('timer', 'bg-indigo-50', 'text-indigo-500',
                         f'Odpracováno ({obd_lbl})', f"{st['odprac_obdobi']:.1f}h", 'Skutečně zaznamenáno')
            if zobraz_porady:
                _karta_cislo('groups', 'bg-teal-50', 'text-teal-500',
                             f'Porady {obd_lbl2}', st['porady_obdobi'], 'Účast / moderace')
            if zobraz_projekty:
                _karta_cislo('rocket_launch', 'bg-purple-50', 'text-purple-500',
                             'Aktivní projekty', st['aktivni_proj'], 'Projekty, kde jste členem')

    # ── Řádek 2: Grafy ────────────────────────────────────────────────────────
    with ui.row().classes('w-full gap-4 mb-4 flex-wrap items-stretch'):

        # Donut: úkoly dle stavu
        with ui.card().classes('flex-1 min-w-[260px] p-5 rounded-xl border border-gray-100 bg-white'):
            ui.label('Úkoly dle stavu').classes('text-base font-extrabold text-gray-700 mb-4')
            celkem = max(st['zadano'] + st['rozpracovano'] + st['pozastaveno'] + st['hotovo_celk'] + st['zrusen'], 1)
            segmenty = [
                (st['zadano'],       '#ef4444', 'Zadáno'),
                (st['rozpracovano'], '#f97316', 'Rozpracováno'),
                (st['pozastaveno'],  '#eab308', 'Pozastaveno'),
                (st['hotovo_celk'],  '#22c55e', 'Hotovo'),
                (st['zrusen'],       '#9ca3af', 'Zrušen'),
            ]
            # Conic-gradient donut
            grad_parts = []
            uhel = 0
            for pocet, barva, _ in segmenty:
                pct = pocet / celkem * 360
                if pct > 0:
                    grad_parts.append(f'{barva} {uhel:.1f}deg {uhel + pct:.1f}deg')
                uhel += pct
            gradient = ', '.join(grad_parts) if grad_parts else '#e5e7eb 0deg 360deg'

            with ui.row().classes('w-full gap-4 items-center'):
                # Donut SVG
                with ui.element('div').classes('relative shrink-0').style('width:100px;height:100px'):
                    ui.element('div').style(
                        f'width:100px;height:100px;border-radius:50%;'
                        f'background:conic-gradient({gradient});'
                        f'position:relative;'
                    )
                    with ui.element('div').style(
                        'position:absolute;top:50%;left:50%;transform:translate(-50%,-50%);'
                        'width:60px;height:60px;background:white;border-radius:50%;'
                        'display:flex;align-items:center;justify-content:center;'
                        'font-weight:900;font-size:18px;color:#374151;'
                    ):
                        ui.label(str(celkem))

                with ui.column().classes('gap-1.5 flex-1'):
                    for pocet, barva, label in segmenty:
                        if pocet == 0: continue
                        pct_lbl = f'{pocet/celkem*100:.2f}'.replace('.', ',') + ' %'
                        with ui.row().classes('w-full items-center gap-2'):
                            ui.element('div').style(f'width:10px;height:10px;border-radius:50%;background:{barva};flex-shrink:0')
                            ui.label(label).classes('text-xs text-gray-600 flex-1')
                            ui.label(f'{pocet}').classes('text-xs font-bold text-gray-800')
                            ui.label(pct_lbl).classes('text-xs text-gray-400 w-12 text-right')

        # Donut: úkoly dle priority
        with ui.card().classes('flex-1 min-w-[260px] p-5 rounded-xl border border-gray-100 bg-white'):
            ui.label('Úkoly dle termínu').classes('text-base font-extrabold text-gray-700 mb-4')

            term_data = [
                (st['po_terminu'],                        '#dc2626', 'Po termínu'),
                (st['dnes_termin'],                       '#ea580c', 'Dnes'),
                (st['tento_tyden'],                       '#d97706', 'Tento týden'),
                (max(st['otevrene'] - st['po_terminu'] - st['dnes_termin'] - st['tento_tyden'], 0), '#16a34a', 'V pořádku'),
            ]
            celkem_t = max(sum(x[0] for x in term_data), 1)
            grad_t = []
            uhel_t = 0
            for pocet, barva, _ in term_data:
                pct = pocet / celkem_t * 360
                if pct > 0:
                    grad_t.append(f'{barva} {uhel_t:.1f}deg {uhel_t + pct:.1f}deg')
                uhel_t += pct
            gradient_t = ', '.join(grad_t) if grad_t else '#e5e7eb 0deg 360deg'

            with ui.row().classes('w-full gap-4 items-center'):
                with ui.element('div').classes('relative shrink-0').style('width:100px;height:100px'):
                    ui.element('div').style(
                        f'width:100px;height:100px;border-radius:50%;'
                        f'background:conic-gradient({gradient_t});position:relative;'
                    )
                    with ui.element('div').style(
                        'position:absolute;top:50%;left:50%;transform:translate(-50%,-50%);'
                        'width:60px;height:60px;background:white;border-radius:50%;'
                        'display:flex;align-items:center;justify-content:center;'
                        'font-weight:900;font-size:16px;color:#374151;'
                    ):
                        ui.label(str(st['otevrene']))

                with ui.column().classes('gap-1.5 flex-1'):
                    for pocet, barva, label in term_data:
                        with ui.row().classes('w-full items-center gap-2'):
                            ui.element('div').style(f'width:10px;height:10px;border-radius:50%;background:{barva};flex-shrink:0')
                            ui.label(label).classes('text-xs text-gray-600 flex-1')
                            ui.label(str(pocet)).classes('text-xs font-bold text-gray-800')
                            pct_t = f'{pocet/celkem_t*100:.2f}'.replace('.', ',') + ' %'
                            ui.label(pct_t).classes('text-xs text-gray-400 w-12 text-right')

        # Progress bar úkolů dle priority
        with ui.card().classes('flex-1 min-w-[260px] p-5 rounded-xl border border-gray-100 bg-white'):
            ui.label('Nedokončené úkoly').classes('text-base font-extrabold text-gray-700 mb-4')
            # Stav Zadáno + Rozpracováno + Pozastaveno
            for label, pocet, barva, bg in [
                ('Zadáno',       st['zadano'],       'bg-red-500',    'bg-red-50'),
                ('Rozpracováno', st['rozpracovano'], 'bg-orange-500', 'bg-orange-50'),
                ('Pozastaveno',  st['pozastaveno'],  'bg-yellow-400', 'bg-yellow-50'),
            ]:
                celk2 = max(st['otevrene'], 1)
                pct2 = int(pocet / celk2 * 100)
                with ui.column().classes(f'w-full p-3 rounded-lg {bg} mb-2 gap-1'):
                    with ui.row().classes('w-full justify-between items-center'):
                        ui.label(label).classes('text-xs font-bold text-gray-700')
                        ui.label(f'{pocet} úkolů  ({pct2} %)').classes('text-xs font-bold text-gray-500')
                    with ui.element('div').classes('w-full bg-white rounded-full h-2 overflow-hidden'):
                        ui.element('div').classes(f'h-full rounded-full {barva} transition-all').style(f'width:{pct2}%')

    # ── Řádek 3: Seznamy ──────────────────────────────────────────────────────
    with ui.row().classes('w-full gap-4 flex-wrap items-stretch'):

        # Nadcházející termíny
        with ui.card().classes('flex-1 min-w-[260px] p-5 rounded-xl border border-gray-100 bg-white'):
            with ui.row().classes('w-full justify-between items-center mb-3'):
                ui.label('Nadcházející termíny').classes('text-base font-extrabold text-gray-700')
                ui.label('příštích 14 dní').classes('text-xs text-gray-400')
            if not st['nadchazejici']:
                ui.label('Žádné blížící se termíny.').classes('text-xs text-gray-400 italic')
            for uk in st['nadchazejici']:
                termin_d = uk.get('termin')
                delta_d = (termin_d - dnes).days if termin_d else None
                if delta_d is not None and delta_d < 0:
                    tc = 'text-red-600 font-bold'
                    t_txt = f'{abs(delta_d)} dní po termínu'
                elif delta_d == 0:
                    tc = 'text-red-500 font-bold'; t_txt = 'Dnes'
                elif delta_d and delta_d <= 3:
                    tc = 'text-orange-500 font-bold'; t_txt = f'Za {delta_d} dny'
                else:
                    tc = 'text-gray-500'; t_txt = termin_d.strftime('%d.%m.%Y') if termin_d else '—'
                p_barva = PRIORITY_BARVY.get(uk.get('priorita', 'Normální'), 'text-blue-600')
                with ui.row().classes('w-full items-center gap-2 py-1.5 border-b border-gray-50'):
                    ui.label(uk['nazev']).classes('text-sm text-gray-800 flex-1 truncate')
                    ui.label(uk.get('priorita', '')).classes(f'text-[10px] {p_barva} shrink-0')
                    ui.label(t_txt).classes(f'text-xs {tc} shrink-0 w-28 text-right')

        # Nedávno dokončeno
        with ui.card().classes('flex-1 min-w-[260px] p-5 rounded-xl border border-gray-100 bg-white'):
            with ui.row().classes('w-full justify-between items-center mb-3'):
                ui.label('Nedávno dokončeno').classes('text-base font-extrabold text-gray-700')
                ui.label('posledních 5').classes('text-xs text-gray-400')
            if not st['nedavno_hotovo']:
                ui.label('Zatím žádné dokončené úkoly.').classes('text-xs text-gray-400 italic')
            for uk in st['nedavno_hotovo']:
                with ui.row().classes('w-full items-center gap-2 py-1.5 border-b border-gray-50'):
                    ui.icon('check_circle', size='xs', color='green').classes('shrink-0')
                    ui.label(uk['nazev']).classes('text-sm text-gray-700 flex-1 truncate')
                    d_str = uk.get('created_date')
                    if d_str:
                        ui.label(d_str.strftime('%d.%m.') if hasattr(d_str, 'strftime') else str(d_str)[:5]).classes('text-xs text-gray-400 shrink-0')


@ui.refreshable
async def _vykresli_statistika(user_id, user_name, vsechna_prava, stat_stav: dict):
    """
    stat_stav = {'pohled': 'moje'|'osoba'|'oddeleni', 'vybrany_uid': int|None,
                 'obdobi': 'tyden'|'mesic', 'drill_uid': int|None,
                 'sort': str, 'sort_dir': 'asc'|'desc'}
    """
    is_admin = 'vse' in vsechna_prava or 'ukolovnik_admin' in vsechna_prava
    obdobi = stat_stav.get('obdobi', 'tyden')
    je_tyden = obdobi == 'tyden'
    obd_lbl = 'týden' if je_tyden else 'měsíc'
    dnes = datetime.date.today()
    # Dlaždice „Aktivní projekty"/„Porady" mají smysl jen pro toho, kdo má
    # odpovídající právo (a zároveň je součástí aspoň jednoho projektu/porady).
    viditelne_sekce = _uk_viditelne_sekce(vsechna_prava)
    ma_projekty_pravo = 'projekty' in viditelne_sekce
    ma_porady_pravo = 'porady' in viditelne_sekce

    if is_admin:
        uziv_opts2 = await asyncio.to_thread(_ziskej_uzivatele_options)
        podrizeni = [{'id': uid, 'jmeno': jm} for uid, jm in uziv_opts2.items()]
    elif _je_hlavni_vedouci(vsechna_prava):
        uzivatele_oddeleni_s = await asyncio.to_thread(_ziskej_uzivatele_oddeleni, vsechna_prava)
        podrizeni = [u for u in uzivatele_oddeleni_s if u['id'] != user_id]
    else:
        podrizeni = await asyncio.to_thread(_ziskej_podrizene, user_id)

    je_vedouci = bool(podrizeni) or is_admin

    # Starý obsah smazat až s daty v ruce → přestavba bez bliknutí a skoku scrollu
    ui.context.slot.parent.clear()

    def _uprav(**zmeny):
        stat_stav.update(zmeny)
        _statistika_refresh_klient()

    pohled = stat_stav.get('pohled', 'moje')
    if not je_vedouci:
        pohled = 'moje'

    # ── Hlavička ──────────────────────────────────────────────────────────────
    with ui.row().classes('w-full items-center justify-between mb-5 flex-wrap gap-3'):
        with ui.column().classes('gap-0'):
            ui.label('Statistika').classes('text-3xl font-extrabold text-gray-800')
            _obd_popis = ('Aktuální týden' if je_tyden else f'{MESICE_CZ[dnes.month]} {dnes.year}')
            ui.label(f'Výkonnost a vytížení • {_obd_popis}').classes('text-xs text-gray-400')

        with ui.row().classes('items-center gap-2 flex-wrap'):
            _stat_seg([('tyden', 'Týden'), ('mesic', 'Měsíc')], obdobi,
                      lambda k: _uprav(obdobi=k))
            if je_vedouci:
                _stat_seg([('moje', 'Moje'), ('osoba', 'Osoba'), ('oddeleni', 'Oddělení')], pohled,
                          lambda k: _uprav(pohled=k, drill_uid=None))
                ui.button('Export XLSX', icon='download',
                          on_click=lambda: _dialog_export_ukoly(user_id, user_name, vsechna_prava)
                          ).props('outline dense color=green-7').classes('text-xs')

    # Selector osoby
    if je_vedouci and pohled == 'osoba' and podrizeni:
        osoba_opts2 = {str(p['id']): p['jmeno'] for p in podrizeni}
        vyb_str = str(stat_stav.get('vybrany_uid') or next(iter(osoba_opts2)))
        def _set_o(e):
            stat_stav['vybrany_uid'] = int(e.value)
            _statistika_refresh_klient()
        ui.select(osoba_opts2, value=vyb_str, label='Vybrat pracovníka', on_change=_set_o).classes('w-64 mb-4').props('outlined dense')

    def _barva_plneni(pct):
        if pct is None: return '#1f2937'
        return '#059669' if pct >= 90 else '#d97706' if pct >= 75 else '#e11d48'

    # ── Pohled: ODDĚLENÍ — srovnávací tabulka týmu + detail pracovníka ───────
    if je_vedouci and pohled == 'oddeleni' and podrizeni:
        stat_osob = await asyncio.to_thread(
            lambda: {p['id']: _nacti_statistiku_osoby(p['id'], obdobi) for p in podrizeni})

        radky = []
        for p in podrizeni:
            s = stat_osob.get(p['id']) or {}
            radky.append({
                'id': p['id'], 'jmeno': p['jmeno'],
                'otevrene':   s.get('otevrene', 0),
                'po_terminu': s.get('po_terminu', 0),
                'hotovo':     s.get('hotovo_obdobi', 0),
                'hodiny':     float(s.get('odprac_obdobi') or 0),
                'kapacita':   float(s.get('kapacita_h') or 0),
                'vytiz':      s.get('vytiz_pct', 0),
                'plneni':     s.get('plneni_pct'),
                'bez_term':   s.get('bez_terminu', 0),
            })

        c_otevrene = sum(r['otevrene'] for r in radky)
        c_po_term  = sum(r['po_terminu'] for r in radky)
        c_hotovo   = sum(r['hotovo'] for r in radky)
        c_predch   = sum((stat_osob.get(r['id']) or {}).get('hotovo_predchozi', 0) for r in radky)
        c_hodiny   = sum(r['hodiny'] for r in radky)
        c_kapacita = sum(r['kapacita'] for r in radky)
        c_vcas     = sum((stat_osob.get(r['id']) or {}).get('hotovo_vcas', 0) for r in radky)
        c_sterm    = sum((stat_osob.get(r['id']) or {}).get('hotovo_s_term', 0) for r in radky)
        c_bez_term = sum(r['bez_term'] for r in radky)
        c_plneni   = int(round(c_vcas / c_sterm * 100)) if c_sterm else None
        c_vytiz    = int(round(c_hodiny / c_kapacita * 100)) if c_kapacita else 0
        v_riziku   = sum(1 for r in radky if r['po_terminu'] > 0)

        # KPI pás oddělení
        with ui.row().classes('w-full gap-3 mb-5 flex-wrap'):
            _stat_kpi('Otevřené úkoly', c_otevrene,
                      f'{c_bez_term} bez termínu' if c_bez_term else 'všechny mají termín')
            _stat_kpi('Po termínu', c_po_term,
                      f'u {v_riziku} z {len(radky)} pracovníků' if c_po_term else 'vše v termínu',
                      barva='#e11d48' if c_po_term else '#059669', alert=c_po_term > 0)
            _stat_kpi(f'Hotovo / {obd_lbl}', c_hotovo, f'předchozí {obd_lbl}: {c_predch}',
                      trend=c_hotovo - c_predch)
            _stat_kpi('Plnění termínů', '—' if c_plneni is None else f'{c_plneni} %',
                      f'{c_vcas} z {c_sterm} včas' if c_sterm else 'bez úkolů s termínem',
                      barva=_barva_plneni(c_plneni))
            _stat_kpi('Odpracováno', f'{c_hodiny:.0f} h', f'kapacita {c_kapacita:.0f} h')
            _stat_kpi('Vytížení týmu', f'{c_vytiz} %', 'podíl vykázaného času',
                      barva=_stat_barva_vytizeni(c_vytiz))

        # Řazení
        sort_kl  = stat_stav.get('sort') or 'po_terminu'
        sort_dir = stat_stav.get('sort_dir') or 'desc'

        def _prepni_sort(k):
            if stat_stav.get('sort') == k:
                _uprav(sort_dir=('asc' if sort_dir == 'desc' else 'desc'))
            else:
                _uprav(sort=k, sort_dir=('asc' if k == 'jmeno' else 'desc'))

        def _klic(r):
            if sort_kl == 'jmeno':
                return (r['jmeno'] or '').lower()
            v = r.get(sort_kl)
            return -1 if v is None else v

        radky.sort(key=_klic, reverse=(sort_dir == 'desc'))

        drill_uid = stat_stav.get('drill_uid')
        if drill_uid not in stat_osob:
            drill_uid = None

        MRIZKA = ('display:grid;align-items:center;gap:8px;'
                  'grid-template-columns:minmax(160px,1fr) 74px 84px 128px 78px 74px 76px')

        with ui.row().classes('w-full gap-4 items-start flex-wrap'):
            # ── Srovnávací tabulka ───────────────────────────────────────────
            with ui.card().classes('flex-1 min-w-[560px] p-0 rounded-xl border border-gray-100 '
                                   'bg-white shadow-none overflow-hidden'):
                with ui.row().classes('w-full items-center justify-between px-4 py-3 border-b border-gray-100'):
                    ui.label(f'Tým — {len(radky)} pracovníků').classes('text-base font-extrabold text-gray-700')
                    ui.label('Klikněte na řádek pro detail').classes('text-[11px] text-gray-400')

                with ui.element('div').classes('w-full px-4 py-2 bg-gray-50 border-b border-gray-100').style(MRIZKA):
                    for kl, popis, zar in [('jmeno', 'Pracovník', 'flex-start'), ('otevrene', 'Otevř.', 'flex-end'),
                                           ('po_terminu', 'Po term.', 'flex-end'), ('vytiz', 'Vytížení', 'flex-start'),
                                           ('hodiny', 'Hodiny', 'flex-end'), ('hotovo', 'Hotovo', 'flex-end'),
                                           ('plneni', 'Plnění', 'flex-end')]:
                        akt = (kl == sort_kl)
                        with ui.element('div').classes('cursor-pointer select-none').style(
                                f'display:flex;align-items:center;gap:3px;justify-content:{zar}').on(
                                'click', lambda k=kl: _prepni_sort(k)):
                            ui.label(popis).classes('text-[10px] font-extrabold tracking-wider uppercase '
                                                    + ('text-blue-700' if akt else 'text-gray-400'))
                            if akt:
                                ui.label('▼' if sort_dir == 'desc' else '▲').classes('text-[8px] text-blue-700')

                for r in radky:
                    vyb = (r['id'] == drill_uid)
                    zvyr = 'background:#eff6ff;box-shadow:inset 3px 0 0 #2563eb;' if vyb else ''
                    with ui.element('div').classes(
                            'w-full px-4 py-2 border-b border-gray-50 cursor-pointer '
                            'hover:bg-gray-50 transition-colors').style(MRIZKA + ';' + zvyr).on(
                            'click', lambda i=r['id']: _uprav(drill_uid=(None if i == drill_uid else i))):
                        with ui.element('div').style('display:flex;align-items:center;gap:10px;min-width:0'):
                            _stat_avatar(r['jmeno'], 30)
                            ui.label(r['jmeno']).classes('text-sm font-bold text-gray-700 truncate')
                        ui.label(str(r['otevrene'])).classes('text-sm font-bold text-gray-600 text-right')
                        if r['po_terminu']:
                            with ui.element('div').style('display:flex;justify-content:flex-end'):
                                ui.label(str(r['po_terminu'])).classes(
                                    'text-xs font-extrabold text-red-700 bg-red-50 rounded px-2 py-0.5')
                        else:
                            ui.label('—').classes('text-sm text-gray-300 text-right')
                        with ui.element('div').style('display:flex;align-items:center;gap:7px'):
                            _stat_pruh(r['vytiz'], _stat_barva_vytizeni(r['vytiz']))
                            ui.label(f"{r['vytiz']}%").classes('text-[11px] font-bold text-gray-500 shrink-0')
                        ui.label(f"{r['hodiny']:.1f} h").classes('text-sm font-bold text-gray-600 text-right')
                        ui.label(str(r['hotovo'])).classes('text-sm font-bold text-gray-600 text-right')
                        ui.label('—' if r['plneni'] is None else f"{r['plneni']}%").classes(
                            'text-sm font-extrabold text-right').style(
                            'color:#d1d5db' if r['plneni'] is None else f'color:{_barva_plneni(r["plneni"])}')

            # ── Boční panel: detail pracovníka / co vyžaduje pozornost ───────
            with ui.column().classes('w-[330px] min-w-[300px] gap-4'):
                if drill_uid:
                    s = stat_osob[drill_uid]
                    jm = next((p['jmeno'] for p in podrizeni if p['id'] == drill_uid), '')
                    with ui.card().classes('w-full p-4 rounded-xl border border-gray-100 bg-white shadow-none gap-0'):
                        with ui.row().classes('w-full items-center gap-3 mb-3 no-wrap'):
                            _stat_avatar(jm, 40)
                            with ui.column().classes('gap-0 flex-1 min-w-0'):
                                ui.label(jm).classes('text-base font-extrabold text-gray-800 truncate')
                                ui.label(f"{s.get('otevrene', 0)} otevřených • {s.get('hotovo_obdobi', 0)} hotovo "
                                         f"za {obd_lbl}").classes('text-[11px] text-gray-400')
                            ui.icon('close', size='sm').classes(
                                'text-gray-300 cursor-pointer hover:text-gray-600').on(
                                'click', lambda: _uprav(drill_uid=None))

                        with ui.element('div').classes('w-full mb-3').style(
                                'display:grid;grid-template-columns:1fr 1fr;gap:8px'):
                            for lbl_m, hod_m, bar_m in [
                                ('Vytížení', f"{s.get('vytiz_pct', 0)} %", _stat_barva_vytizeni(s.get('vytiz_pct', 0))),
                                ('Odpracováno', f"{float(s.get('odprac_obdobi') or 0):.1f} h", '#1f2937'),
                                ('Plnění termínů', '—' if s.get('plneni_pct') is None else f"{s['plneni_pct']} %",
                                 _barva_plneni(s.get('plneni_pct'))),
                                ('Po termínu', s.get('po_terminu', 0),
                                 '#e11d48' if s.get('po_terminu') else '#059669'),
                            ]:
                                with ui.element('div').classes('rounded-lg bg-gray-50 px-3 py-2'):
                                    ui.label(lbl_m.upper()).classes(
                                        'text-[9px] font-extrabold text-gray-400 tracking-wider')
                                    ui.label(str(hod_m)).classes('text-lg font-black leading-tight').style(
                                        f'color:{bar_m}')

                        serie_d = s.get('serie_hodin') or []
                        if serie_d:
                            ui.label('HODINY PO DNECH' if je_tyden else 'HODINY PO TÝDNECH').classes(
                                'text-[10px] font-extrabold text-gray-400 tracking-wider mb-1')
                            max_d = max([x['h'] for x in serie_d] or [0]) or 1
                            with ui.row().classes('w-full items-end gap-1 no-wrap mb-3').style('height:54px'):
                                for x in serie_d:
                                    with ui.column().classes('flex-1 items-center gap-0.5 h-full justify-end'):
                                        ui.element('div').classes('w-full rounded-t').style(
                                            f"height:{max(2, int(x['h'] / max_d * 36))}px;"
                                            f"background:{'#2563eb' if x['akt'] else '#dbe3ec'}").tooltip(
                                            f"{x['lbl']}: {x['h']:.1f} h")
                                        ui.label(str(x['lbl'])[:2]).classes(
                                            'text-[9px] font-bold ' + ('text-blue-700' if x['akt'] else 'text-gray-400'))

                        po_term = s.get('po_termine_ukoly') or []
                        if po_term:
                            ui.label('PO TERMÍNU').classes(
                                'text-[10px] font-extrabold text-red-500 tracking-wider mb-1')
                            for u in po_term[:4]:
                                dni = (dnes - u['termin']).days if u.get('termin') else 0
                                with ui.element('div').classes(
                                        'w-full flex items-center gap-2 py-1 border-b border-gray-50'):
                                    ui.element('div').style(
                                        'width:6px;height:6px;border-radius:3px;background:#e11d48;flex-shrink:0')
                                    ui.label(u['nazev']).classes('text-xs text-gray-600 truncate flex-1')
                                    ui.label(f'{dni} d').classes('text-[10px] font-bold text-red-600 shrink-0')

                        nadch = s.get('nadchazejici') or []
                        if nadch:
                            ui.label('NEJBLIŽŠÍ TERMÍNY').classes(
                                'text-[10px] font-extrabold text-gray-400 tracking-wider mt-3 mb-1')
                            for u in nadch[:4]:
                                with ui.element('div').classes(
                                        'w-full flex items-center gap-2 py-1 border-b border-gray-50'):
                                    ui.element('div').style(
                                        f"width:6px;height:6px;border-radius:3px;flex-shrink:0;"
                                        f"background:{_STAT_PRIO_HEX.get(u.get('priorita'), '#94a3b8')}")
                                    ui.label(u['nazev']).classes('text-xs text-gray-600 truncate flex-1')
                                    ui.label(u['termin'].strftime('%d.%m.') if u.get('termin') else '').classes(
                                        'text-[10px] font-bold text-gray-400 shrink-0')

                        if not po_term and not nadch:
                            ui.label('Žádné úkoly po termínu ani v nejbližších 14 dnech.').classes(
                                'text-xs text-gray-400 italic py-2')

                        ui.button('Otevřít celý profil', icon='open_in_new',
                                  on_click=lambda i=drill_uid: _uprav(pohled='osoba', vybrany_uid=i)
                                  ).props('flat no-caps dense color=primary').classes('mt-3 self-start')
                else:
                    # Panel „Vyžaduje pozornost"
                    riz = sorted([r for r in radky if r['po_terminu'] > 0],
                                 key=lambda r: -r['po_terminu'])
                    bez_vykazu = [r for r in radky if r['hodiny'] <= 0]
                    with ui.card().classes('w-full p-4 rounded-xl border border-gray-100 bg-white shadow-none gap-0'):
                        ui.label('Vyžaduje pozornost').classes('text-base font-extrabold text-gray-700 mb-3')
                        if riz:
                            ui.label('ÚKOLY PO TERMÍNU').classes(
                                'text-[10px] font-extrabold text-red-500 tracking-wider mb-1')
                            for r in riz[:5]:
                                with ui.element('div').classes(
                                        'w-full flex items-center gap-2 py-1.5 border-b border-gray-50 '
                                        'cursor-pointer hover:bg-gray-50 rounded').on(
                                        'click', lambda i=r['id']: _uprav(drill_uid=i)):
                                    _stat_avatar(r['jmeno'], 24)
                                    ui.label(r['jmeno']).classes('text-xs font-bold text-gray-600 truncate flex-1')
                                    ui.label(str(r['po_terminu'])).classes(
                                        'text-[11px] font-extrabold text-red-700 bg-red-50 rounded px-1.5 shrink-0')
                        if bez_vykazu:
                            ui.label(f'BEZ VYKÁZANÉHO ČASU ({obd_lbl})').classes(
                                'text-[10px] font-extrabold text-amber-600 tracking-wider mt-3 mb-1')
                            for r in bez_vykazu[:5]:
                                with ui.element('div').classes(
                                        'w-full flex items-center gap-2 py-1.5 border-b border-gray-50 '
                                        'cursor-pointer hover:bg-gray-50 rounded').on(
                                        'click', lambda i=r['id']: _uprav(drill_uid=i)):
                                    _stat_avatar(r['jmeno'], 24)
                                    ui.label(r['jmeno']).classes('text-xs font-bold text-gray-600 truncate flex-1')
                                    ui.label('0 h').classes('text-[11px] font-bold text-amber-700 shrink-0')
                        if not riz and not bez_vykazu:
                            with ui.column().classes('w-full items-center py-6 gap-1'):
                                ui.icon('verified', size='lg').classes('text-green-500')
                                ui.label('Tým je zcela v termínu').classes('text-sm font-bold text-gray-600')
                                ui.label('Žádné úkoly po termínu, čas vykázán.').classes('text-xs text-gray-400')

                    with ui.card().classes('w-full p-4 rounded-xl border border-gray-100 bg-white shadow-none gap-0'):
                        ui.label('Rozložení práce').classes('text-base font-extrabold text-gray-700 mb-3')
                        for r in sorted(radky, key=lambda x: -x['hodiny'])[:8]:
                            with ui.element('div').classes('w-full mb-2'):
                                with ui.element('div').classes('w-full flex justify-between items-center mb-0.5'):
                                    ui.label(r['jmeno']).classes('text-[11px] font-bold text-gray-500 truncate')
                                    ui.label(f"{r['hodiny']:.1f} h").classes('text-[11px] font-bold text-gray-400')
                                _stat_pruh(int(r['hodiny'] / max(1.0, max(x['hodiny'] for x in radky)) * 100),
                                           _stat_barva_vytizeni(r['vytiz']), 5)

    else:
        # ── Pohled: MOJE / KONKRÉTNÍ OSOBA ──────────────────────────────────
        if pohled == 'osoba' and je_vedouci:
            vid = stat_stav.get('vybrany_uid') or (podrizeni[0]['id'] if podrizeni else user_id)
            jmeno_v = next((p['jmeno'] for p in podrizeni if p['id'] == vid), user_name)
        else:
            vid = user_id
            jmeno_v = user_name

        st_data = await asyncio.to_thread(_nacti_statistiku_osoby, vid, obdobi)
        if not st_data:
            ui.label('Statistiku se nepodařilo načíst — zkuste obnovit stránku.').classes(
                'text-sm text-gray-500 italic py-6')
            return

        with ui.row().classes('w-full items-center gap-3 mb-4 no-wrap'):
            _stat_avatar(jmeno_v, 42)
            with ui.column().classes('gap-0'):
                ui.label(jmeno_v).classes('text-xl font-extrabold text-gray-800')
                ui.label(f"{st_data.get('otevrene', 0)} otevřených úkolů • "
                         f"{st_data.get('bez_terminu', 0)} bez termínu").classes('text-xs text-gray-400')

        _tr = st_data.get('hotovo_obdobi', 0) - st_data.get('hotovo_predchozi', 0)
        _pl = st_data.get('plneni_pct')
        _vy = st_data.get('vytiz_pct', 0)
        with ui.row().classes('w-full gap-3 mb-5 flex-wrap'):
            _stat_kpi('Otevřené úkoly', st_data.get('otevrene', 0),
                      f"zadáno {st_data.get('zadano', 0)} • rozpracováno {st_data.get('rozpracovano', 0)}")
            _stat_kpi('Po termínu', st_data.get('po_terminu', 0),
                      'vyžaduje okamžitou akci' if st_data.get('po_terminu') else 'vše v termínu',
                      barva='#e11d48' if st_data.get('po_terminu') else '#059669',
                      alert=st_data.get('po_terminu', 0) > 0)
            _stat_kpi(f'Termín / {obd_lbl}',
                      st_data.get('tyden_po_pa', 0) if je_tyden else st_data.get('mesic_termin', 0),
                      'pondělí–pátek' if je_tyden else 'do konce měsíce')
            _stat_kpi(f'Hotovo / {obd_lbl}', st_data.get('hotovo_obdobi', 0),
                      f"předchozí {obd_lbl}: {st_data.get('hotovo_predchozi', 0)}", trend=_tr)
            _stat_kpi('Plnění termínů', '—' if _pl is None else f'{_pl} %',
                      f"{st_data.get('hotovo_vcas', 0)} z {st_data.get('hotovo_s_term', 0)} včas"
                      if st_data.get('hotovo_s_term') else 'bez úkolů s termínem',
                      barva=_barva_plneni(_pl))
            _stat_kpi('Odpracováno', f"{float(st_data.get('odprac_obdobi') or 0):.1f} h",
                      f"kapacita {float(st_data.get('kapacita_h') or 0):.0f} h")
            _stat_kpi('Vytížení', f'{_vy} %', 'podíl vykázaného času',
                      barva=_stat_barva_vytizeni(_vy))
            if ma_porady_pravo and st_data.get('porady_celkem'):
                _stat_kpi(f'Porady / {obd_lbl}', st_data.get('porady_obdobi', 0),
                          f"celkem {st_data.get('porady_celkem', 0)}")
            if ma_projekty_pravo and st_data.get('aktivni_proj'):
                _stat_kpi('Aktivní projekty', st_data.get('aktivni_proj', 0), 'kde jste členem')

        serie = st_data.get('serie_hodin') or []
        if serie:
            with ui.element('div').classes('w-full mb-4'):
                _stat_sloupce(serie, 'Odpracované hodiny',
                              'aktuální týden' if je_tyden else f'{MESICE_CZ[dnes.month]} {dnes.year}')

        _vykresli_stat_karty(st_data, jmeno_v,
                             zobraz_projekty=ma_projekty_pravo and st_data['aktivni_proj'] > 0,
                             zobraz_porady=ma_porady_pravo and st_data['porady_celkem'] > 0,
                             zobraz_kpi=False)


# =========================================================
# PROJEKTY — pomocné funkce
# =========================================================
STAVY_PROJEKT = {
    'Aktivní':    'bg-blue-100 text-blue-800 border-blue-300',
    'Pozastavený':'bg-yellow-100 text-yellow-800 border-yellow-300',
    'Dokončený':  'bg-green-100 text-green-800 border-green-300',
    'Zrušený':    'bg-gray-200 text-gray-600 border-gray-400',
}
ROLE_PROJEKT = {
    'vlastnik':       ('👑 Vlastník',    'text-yellow-700'),
    'spolupracovnik': ('🤝 Spolupracovník', 'text-blue-700'),
    'komentar':       ('💬 Komentátor',  'text-gray-500'),
}


def _notifikuj_projekt(projekt_id, vyjma_uid, zprava, typ='info'):
    conn = intranet_data.get_db_connection()
    if not conn: return
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT user_id FROM ukolovnik_projekt_clenove WHERE projekt_id=%s", (projekt_id,))
        uids = {r['user_id'] for r in cur.fetchall()}
        cur.execute("SELECT vytvoril_id FROM ukolovnik_projekty WHERE id=%s", (projekt_id,))
        row = cur.fetchone()
        if row and row.get('vytvoril_id'):
            uids.add(row['vytvoril_id'])
    finally:
        cur.close(); conn.close()
    uids.update(_ziskej_sledovace_ids('projekt', projekt_id))
    uids.discard(vyjma_uid)
    for uid in uids:
        try: intranet_notifikace.pridej(uid, zprava, typ)
        except Exception: pass


def _je_clen_projektu(projekt_id, user_id):
    conn = intranet_data.get_db_connection()
    if not conn: return False
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT id FROM ukolovnik_projekt_clenove WHERE projekt_id=%s AND user_id=%s", (projekt_id, user_id))
        return bool(cur.fetchone())
    finally:
        cur.close(); conn.close()


# =========================================================
# DIALOG: NOVÝ / UPRAVIT PROJEKT
# =========================================================
def _dialog_novy_projekt(user_id, user_name, on_refresh, projekt=None, vsechna_prava=None):
    """projekt = dict pro editaci, None pro nový."""
    if vsechna_prava is None:
        vsechna_prava = []
    uziv_opts = _ziskej_viditelne_uzivatele(user_id, vsechna_prava)

    with _dialog_kotva(), ui.dialog() as dlg, ui.card().classes('w-full max-w-3xl p-6 rounded-xl'):
        ui.label('Upravit projekt' if projekt else 'Nový projekt').classes('text-2xl font-bold text-blue-800 mb-4')

        nazev = ui.input('Název projektu').classes('w-full mb-3 bg-white').props('outlined')
        if projekt: nazev.value = projekt.get('nazev', '')

        popis = ui.textarea('Popis / cíl projektu').classes('w-full mb-3 bg-white').props('outlined rows=3')
        if projekt: popis.value = projekt.get('popis', '') or ''

        with ui.row().classes('w-full gap-3 mb-3 flex-wrap'):
            termin_inp = ui.input('Termín (DD.MM.RRRR)').classes('flex-1 bg-white').props('outlined')
            if projekt and projekt.get('termin'):
                termin_inp.value = projekt['termin'].strftime('%d.%m.%Y')
            stav_sel = ui.select(
                {k: k for k in STAVY_PROJEKT}, label='Stav', value=(projekt['stav'] if projekt else 'Aktivní')
            ).classes('w-40 bg-white').props('outlined')

        _odd_proj_opts = _ziskej_oddeleni_options()
        _curr_odd_proj = [o.strip() for o in (projekt.get('oddeleni') or '').split(',') if o.strip()] if projekt else []
        if _odd_proj_opts:
            odd_proj_sel = ui.select(
                _odd_proj_opts, label='Zúčastněná oddělení',
                multiple=True, value=_curr_odd_proj
            ).classes('w-full bg-white mb-3').props('outlined use-chips')

        # Členové (jen při novém projektu — při editaci se spravují v detailu)
        if not projekt:
            ui.label('Přidat členy').classes('text-sm font-bold text-gray-600 mb-1')
            cbs: dict = {}
            role_opts = {'spolupracovnik': 'Spolupracovník', 'komentar': 'Komentátor'}
            with ui.scroll_area().classes('w-full border border-gray-200 rounded-lg p-3 bg-gray-50').style('max-height:150px'):
                with ui.column().classes('w-full gap-1'):
                    for uid, jmeno in uziv_opts.items():
                        if uid == user_id: continue
                        with ui.row().classes('w-full items-center gap-2'):
                            cb = ui.checkbox(jmeno).classes('text-sm flex-1')
                            role_sel = ui.select(role_opts, value='spolupracovnik').classes('w-36').props('dense outlined')
                            cbs[uid] = (jmeno, cb, role_sel)

        def ulozit():
            if not nazev.value or not nazev.value.strip():
                ui.notify('Vyplňte název!', type='warning'); return
            termin_date = None
            if termin_inp.value and termin_inp.value.strip():
                try: termin_date = datetime.datetime.strptime(termin_inp.value.strip(), '%d.%m.%Y').date()
                except Exception: pass
            odd_proj_val = ','.join(odd_proj_sel.value) if ('odd_proj_sel' in dir()) and odd_proj_sel and odd_proj_sel.value else None

            c = intranet_data.get_db_connection()
            cu = c.cursor()
            if projekt:
                cu.execute("UPDATE ukolovnik_projekty SET nazev=%s, popis=%s, termin=%s, stav=%s, oddeleni=%s WHERE id=%s",
                           (nazev.value.strip(), popis.value.strip() if popis.value else '',
                            termin_date, stav_sel.value, odd_proj_val, projekt['id']))
                c.commit(); cu.close(); c.close()
                intranet_logger.log_activity(user_name, "Projekty", f"Upraven projekt #{projekt['id']}: {nazev.value[:40]}")
                _log('projekt', projekt['id'], user_id, user_name, f'Upravil/a projekt: {nazev.value[:80]}')
                _notifikuj_projekt(projekt['id'], user_id, f"✏️ {user_name} upravil/a projekt '{nazev.value[:50]}'.")
            else:
                cu.execute("INSERT INTO ukolovnik_projekty (nazev, popis, termin, stav, vytvoril_id, vytvoril_jmeno, oddeleni) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                           (nazev.value.strip(), popis.value.strip() if popis.value else '',
                            termin_date, stav_sel.value, user_id, user_name, odd_proj_val))
                nid = cu.lastrowid
                # Vlastník = tvůrce
                cu.execute("INSERT IGNORE INTO ukolovnik_projekt_clenove (projekt_id, user_id, jmeno, role) VALUES (%s,%s,%s,'vlastnik')",
                           (nid, user_id, user_name))
                # Přidaní členové
                pozvaní = []
                for uid, (jmeno, cb, role_sel) in cbs.items():
                    if cb.value:
                        cu.execute("INSERT IGNORE INTO ukolovnik_projekt_clenove (projekt_id, user_id, jmeno, role) VALUES (%s,%s,%s,%s)",
                                   (nid, uid, jmeno, role_sel.value))
                        pozvaní.append(uid)
                c.commit(); cu.close(); c.close()
                intranet_logger.log_activity(user_name, "Projekty", f"Vytvořen projekt: {nazev.value[:40]}")
                _log('projekt', nid, user_id, user_name, f'Vytvořil/a projekt: {nazev.value[:80]}')
                for uid in pozvaní:
                    try: intranet_notifikace.pridej(uid, f"🚀 {user_name} tě přidal/a do projektu: '{nazev.value[:50]}'", 'info')
                    except Exception: pass

            ui.notify('Uloženo.', type='positive', position='top')
            dlg.close(); on_refresh()

        with ui.row().classes('w-full justify-between mt-4'):
            ui.button('Zrušit', on_click=dlg.close).classes('bg-gray-400 text-white font-bold px-6')
            ui.button('Uložit projekt', icon='save', on_click=ulozit).classes('bg-blue-600 text-white font-bold px-6 shadow-md')

    dlg.open()


# =========================================================
# DIALOG: DETAIL PROJEKTU
# =========================================================
def _dialog_detail_projektu(projekt_id, user_id, user_name, vsechna_prava, on_refresh):
    conn = intranet_data.get_db_connection()
    if not conn: return
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT * FROM ukolovnik_projekty WHERE id=%s", (projekt_id,))
        proj = cur.fetchone()
    finally:
        cur.close(); conn.close()
    if not proj:
        ui.notify('Projekt nenalezen.', type='negative'); return

    is_admin = vsechna_prava and ('vse' in vsechna_prava or 'ukolovnik_admin' in vsechna_prava)
    je_vlastnik = proj.get('vytvoril_id') == user_id or is_admin or _je_clen_projektu(projekt_id, user_id)

    uziv_opts = _ziskej_viditelne_uzivatele(user_id, vsechna_prava or [])

    _navrhy = {}   # rozepsané texty přežijí živé překreslení obsahu
    with _dialog_kotva(), \
         ui.dialog().props('maximized persistent no-refocus no-shake') as dlg, \
            ui.card().classes('w-full h-full rounded-none p-0 bg-gray-100 flex flex-col overflow-hidden').on('keydown.escape', lambda: dlg.close()):
        # Tělo dialogu jako refreshable — živá synchronizace překresluje jen
        # obsah na místě, okno zůstává otevřené (žádné close+reopen).
        @ui.refreshable
        def _telo():
            conn = intranet_data.get_db_connection()
            if not conn:
                return
            try:
                cur = conn.cursor(dictionary=True)
                cur.execute("SELECT * FROM ukolovnik_projekty WHERE id=%s", (projekt_id,))
                proj = cur.fetchone()
            finally:
                cur.close(); conn.close()
            if not proj:
                ui.notify('Projekt byl mezitím smazán jiným uživatelem.', type='warning')
                dlg.close()
                return
            is_admin = vsechna_prava and ('vse' in vsechna_prava or 'ukolovnik_admin' in vsechna_prava)
            je_vlastnik = proj.get('vytvoril_id') == user_id or is_admin or _je_clen_projektu(projekt_id, user_id)
            uziv_opts = _ziskej_viditelne_uzivatele(user_id, vsechna_prava or [])

            # ── Hlavička ──────────────────────────────────────────────────────────
            b_stav = STAVY_PROJEKT.get(proj['stav'], 'bg-gray-100 text-gray-800 border-gray-300')
            termin_str = proj['termin'].strftime('%d.%m.%Y') if proj.get('termin') else '—'
            delta, dt_cls, dt_txt = _dny_do_terminu(proj.get('termin'))

            with ui.row().classes('w-full items-center justify-between px-6 py-4 bg-white border-b border-gray-200 shrink-0 shadow-sm flex-wrap gap-2'):
                with ui.column().classes('gap-0.5 flex-1 min-w-0'):
                    with ui.row().classes('items-center gap-2'):
                        ui.icon('rocket_launch', color='blue-600', size='md')
                        ui.label(proj['nazev']).classes('text-2xl font-black text-gray-800 line-clamp-1')
                    _meta_txt = f"Termín: {termin_str}  •  Vytvořil: {proj.get('vytvoril_jmeno','—')}  •  {proj['vytvoreno'].strftime('%d.%m.%Y')}"
                    ui.label(_meta_txt).classes('text-xs text-gray-400')
                    if proj.get('oddeleni'):
                        with ui.row().classes('gap-1 flex-wrap mt-0.5'):
                            for _od in _oddeleni_chips(proj['oddeleni'], max_chips=5):
                                ui.label(_od).classes('bg-indigo-100 text-indigo-700 text-[10px] px-2 py-0 rounded-full font-bold border border-indigo-200')
                with ui.row().classes('items-center gap-2 shrink-0 flex-wrap'):
                    ui.label(proj['stav']).classes(f'px-3 py-1 rounded-full text-xs font-bold border uppercase tracking-wider {b_stav}')
                    if proj.get('termin') and proj['stav'] not in ('Dokončený', 'Zrušený'):
                        ui.label(dt_txt).classes(f'text-xs {dt_cls}')
                    if proj.get('prevedeno_na_ukol'):
                        ui.label('✅ Zadán v úkolech').classes('text-xs bg-green-50 border border-green-200 text-green-700 px-2 py-1 rounded-full font-bold')
                    ui.button(icon='visibility', on_click=lambda: _dialog_log('projekt', projekt_id, proj['nazev'], user_id, user_name)).props('flat round dense').classes('text-gray-400 hover:text-blue-500')
                    if je_vlastnik:
                        ui.button(icon='edit', on_click=lambda: [dlg.close(), _dialog_novy_projekt(user_id, user_name, on_refresh, projekt=proj, vsechna_prava=vsechna_prava)]).props('flat round dense').classes('text-gray-400 hover:text-blue-500')
                    if is_admin or proj.get('vytvoril_id') == user_id:
                        def _smazat_projekt_click():
                            def _provest():
                                if _smazat_projekt_db(projekt_id):
                                    intranet_logger.log_activity(user_name, "Projekty", f"Smazán projekt #{projekt_id}: {proj['nazev']}")
                                    ui.notify('Projekt byl smazán.', type='positive', position='top')
                                    dlg.close()
                                    if on_refresh: on_refresh()
                                else:
                                    ui.notify('Chyba při mazání projektu.', type='negative', position='top')
                            _potvrdit_smazat(f'Opravdu smazat projekt „{proj["nazev"]}"?\nTato akce je nevratná.', _provest)
                        ui.button(icon='delete_forever', on_click=_smazat_projekt_click).props('flat round dense').classes('text-red-400 hover:text-red-600').tooltip('Smazat projekt')
                    ui.button(icon='close', on_click=dlg.close).props('flat round dense').classes('text-gray-400 hover:text-red-500')

            # ── Tělo ──────────────────────────────────────────────────────────────
            with ui.row().classes('w-full flex-1 overflow-hidden gap-0'):

                # ── LEVÝ PANEL: info + chat ────────────────────────────────────────
                with ui.column().classes('flex-1 overflow-y-auto p-5 gap-4 bg-white border-r border-gray-100'):

                    # Popis
                    if proj.get('popis'):
                        with ui.card().classes('w-full p-4 bg-gray-50 border border-gray-100 rounded-xl'):
                            ui.label('Popis / cíl projektu').classes('text-xs font-bold text-gray-400 uppercase mb-1')
                            ui.label(proj['popis']).classes('text-gray-800 whitespace-pre-wrap text-sm')

                    # Stav — rychlá změna
                    if je_vlastnik:
                        with ui.row().classes('items-center gap-2 flex-wrap'):
                            ui.label('Stav:').classes('text-xs font-bold text-gray-500')
                            for s in STAVY_PROJEKT:
                                def _zmen_stav(st=s):
                                    c = intranet_data.get_db_connection()
                                    cu = c.cursor()
                                    cu.execute("UPDATE ukolovnik_projekty SET stav=%s WHERE id=%s", (st, projekt_id))
                                    c.commit(); cu.close(); c.close()
                                    intranet_logger.log_activity(user_name, "Projekty", f"Stav projektu #{projekt_id}→{st}")
                                    _log('projekt', projekt_id, user_id, user_name, f'Změnil/a stav na: {st}')
                                    _notifikuj_projekt(projekt_id, user_id, f"📊 Stav projektu '{proj['nazev'][:40]}' změněn na: {st}")
                                    ui.notify(f'Stav: {st}', type='positive')
                                    dlg.close(); on_refresh()
                                aktiv = proj['stav'] == s
                                ui.button(s, on_click=_zmen_stav).props(f'size=sm {"" if aktiv else "outline"}').classes('font-bold ' + ('bg-blue-600 text-white' if aktiv else ''))

                    # ── CHAT ──────────────────────────────────────────────────────
                    ui.label('Chat / komentáře').classes('text-base font-extrabold text-gray-700 mt-2')

                    chat_box = ui.column().classes('w-full gap-2')

                    def nacti_chat():
                        chat_box.clear()
                        c = intranet_data.get_db_connection()
                        zpravy = []
                        if c:
                            try:
                                cu = c.cursor(dictionary=True)
                                cu.execute("SELECT * FROM ukolovnik_projekt_komentare WHERE projekt_id=%s ORDER BY vytvoreno ASC", (projekt_id,))
                                zpravy = cu.fetchall()
                            finally:
                                cu.close(); c.close()
                        with chat_box:
                            if not zpravy:
                                ui.label('Zatím žádné zprávy.').classes('text-gray-400 italic text-xs')
                            for z in zpravy:
                                je_muj = z['user_id'] == user_id
                                align = 'items-end' if je_muj else 'items-start'
                                bg = 'bg-blue-600 text-white' if je_muj else 'bg-gray-100 text-gray-800'
                                with ui.column().classes(f'w-full {align} gap-0.5'):
                                    if not je_muj:
                                        ui.label(z['jmeno_autora']).classes('text-xs font-bold text-gray-500 ml-1')
                                    with ui.element('div').classes(f'max-w-[80%] px-4 py-2 rounded-2xl {bg} {"rounded-br-sm" if je_muj else "rounded-bl-sm"}'):
                                        ui.label(z['text']).classes('text-sm whitespace-pre-wrap')
                                    ui.label(z['vytvoreno'].strftime('%d.%m.%Y %H:%M')).classes('text-[10px] text-gray-400 mx-1')

                    nacti_chat()

                    # Vstup do chatu
                    with ui.row().classes('w-full items-end gap-2 mt-2 sticky bottom-0 bg-white pt-2'):
                        nova_zprava = ui.textarea('Napište zprávu...').classes('flex-1 bg-gray-50').props('outlined autogrow').bind_value(_navrhy, 'zprava')

                        def odeslat_chat():
                            txt = nova_zprava.value.strip() if nova_zprava.value else ''
                            if not txt: return
                            c = intranet_data.get_db_connection()
                            cu = c.cursor()
                            cu.execute("INSERT INTO ukolovnik_projekt_komentare (projekt_id, user_id, jmeno_autora, text) VALUES (%s,%s,%s,%s)",
                                       (projekt_id, user_id, user_name, txt))
                            c.commit(); cu.close(); c.close()
                            intranet_logger.log_activity(user_name, "Projekty", f"Zpráva v projektu #{projekt_id}")
                            _log('projekt', projekt_id, user_id, user_name, 'Odeslal/a zprávu v chatu')
                            _notifikuj_projekt(projekt_id, user_id, f"💬 {user_name} napsal/a v projektu '{proj['nazev'][:40]}'.")
                            nova_zprava.value = ''
                            nacti_chat()

                        nova_zprava.on('keydown.ctrl.enter', odeslat_chat)
                        ui.button(icon='send', on_click=odeslat_chat).classes('bg-blue-600 text-white h-12 w-12 rounded-xl')

                # ── PRAVÝ PANEL: členové + přílohy + úkoly + převod ───────────────
                with ui.column().classes('w-96 shrink-0 overflow-y-auto p-5 gap-5 bg-gray-50 border-l border-gray-200'):

                    # ── PŘEVOD NA ÚKOL ─────────────────────────────────────────────
                    if je_vlastnik and not proj.get('prevedeno_na_ukol'):
                        with ui.card().classes('w-full p-4 bg-gradient-to-r from-blue-50 to-indigo-50 border border-blue-200 rounded-xl'):
                            ui.label('Převést projekt na úkol').classes('text-sm font-extrabold text-blue-800 mb-2')
                            ui.label('Projekt se promítne do sekce Úkolů a bude mu přidělen pracovník a termín.').classes('text-xs text-blue-600 mb-3')

                            uziv_sel_opts = {v: v for v in uziv_opts.values()}
                            _def_uziv = user_name if user_name in uziv_sel_opts else next(iter(uziv_sel_opts), None)
                            prevod_priradit = ui.select(uziv_sel_opts, label='Přidělit komu', value=_def_uziv).classes('w-full bg-white mb-2').props('outlined dense')
                            prevod_termin = ui.input('Termín (DD.MM.RRRR)').classes('w-full bg-white mb-3').props('outlined dense')
                            if proj.get('termin'):
                                prevod_termin.value = proj['termin'].strftime('%d.%m.%Y')

                            def prevest_na_ukol():
                                termin_date = None
                                if prevod_termin.value and prevod_termin.value.strip():
                                    try: termin_date = datetime.datetime.strptime(prevod_termin.value.strip(), '%d.%m.%Y').date()
                                    except Exception: pass
                                prirazen_jmeno = prevod_priradit.value
                                prirazen_id = next((uid for uid, jm in uziv_opts.items() if jm == prirazen_jmeno), None)

                                c = intranet_data.get_db_connection()
                                cu = c.cursor()
                                cu.execute("""
                                    INSERT INTO ukolovnik_ukoly
                                        (nazev, popis, projekt_id, projekt_nazev, prirazen_id, prirazen_jmeno,
                                         zadal_id, zadal_jmeno, termin)
                                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                                """, (proj['nazev'], proj.get('popis') or '', projekt_id, proj['nazev'],
                                      prirazen_id, prirazen_jmeno, user_id, user_name, termin_date))
                                ukol_id = cu.lastrowid
                                cu.execute("UPDATE ukolovnik_projekty SET prevedeno_na_ukol=1, ukol_id=%s WHERE id=%s",
                                           (ukol_id, projekt_id))
                                # Systémová zpráva do chatu projektu
                                cu.execute("INSERT INTO ukolovnik_projekt_komentare (projekt_id, user_id, jmeno_autora, text) VALUES (%s,%s,%s,%s)",
                                           (projekt_id, user_id, '🤖 Systém',
                                            f'Projekt byl zadán jako úkol pracovníkovi {prirazen_jmeno}' +
                                            (f' s termínem {termin_date.strftime("%d.%m.%Y")}' if termin_date else '') + '.'))
                                c.commit(); cu.close(); c.close()

                                intranet_logger.log_activity(user_name, "Projekty", f"Projekt #{projekt_id} převeden na úkol #{ukol_id}")
                                _log('projekt', projekt_id, user_id, user_name, f'Převedl/a projekt na úkol (přidělen: {prirazen_jmeno})')
                                _log('ukol', ukol_id, user_id, user_name, f'Vytvořil/a úkol z projektu: {proj["nazev"][:80]}')
                                _notifikuj_projekt(projekt_id, user_id,
                                                   f"📋 Projekt '{proj['nazev'][:40]}' byl zadán jako úkol pracovníkovi {prirazen_jmeno}.", 'success')
                                if prirazen_id and prirazen_id != user_id:
                                    try: intranet_notifikace.pridej(prirazen_id, f"📋 {user_name} ti přidělil/a úkol z projektu: '{proj['nazev'][:50]}'" + (f" (termín: {termin_date.strftime('%d.%m.%Y')})" if termin_date else ''), 'info')
                                    except Exception: pass
                                ui.notify('Projekt zadán v úkolech!', type='positive', position='top')
                                dlg.close(); on_refresh()

                            ui.button('Zadat jako úkol', icon='task_alt', on_click=prevest_na_ukol).classes('w-full bg-blue-600 text-white font-bold rounded-xl shadow-md')

                    elif proj.get('prevedeno_na_ukol') and proj.get('ukol_id'):
                        with ui.card().classes('w-full p-3 bg-green-50 border border-green-200 rounded-xl'):
                            ui.label('✅ Projekt je zadán v úkolech').classes('text-sm font-bold text-green-700')
                            def _otevrit_ukol(uid=proj['ukol_id']):
                                dlg.close()
                                _dialog_detail_ukolu(uid, user_id, user_name, vsechna_prava, on_refresh=on_refresh)
                            ui.button('Otevřít úkol', icon='open_in_new', on_click=_otevrit_ukol).props('flat dense').classes('text-green-600 text-xs')

                    # ── ČLENOVÉ ────────────────────────────────────────────────────
                    with ui.row().classes('w-full justify-between items-center'):
                        ui.label('Členové projektu').classes('text-sm font-extrabold text-gray-700')
                        if je_vlastnik:
                            def _pridat_clena_dialog():
                                with ui.dialog() as dlg2, ui.card().classes('p-5 w-80 rounded-xl'):
                                    ui.label('Přidat člena').classes('font-bold text-lg mb-3')
                                    uziv_sel2 = {v: v for v in uziv_opts.values()}
                                    _def2 = next(iter(uziv_sel2), None)
                                    novy_clen = ui.select(uziv_sel2, label='Pracovník', value=_def2).classes('w-full mb-2').props('outlined')
                                    novy_role = ui.select({'spolupracovnik': 'Spolupracovník', 'komentar': 'Komentátor'}, label='Role', value='spolupracovnik').classes('w-full mb-3').props('outlined')
                                    def _ulozit_clena():
                                        jm = novy_clen.value
                                        uid2 = next((uid for uid, j in uziv_opts.items() if j == jm), None)
                                        if not uid2: return
                                        c = intranet_data.get_db_connection()
                                        cu = c.cursor()
                                        cu.execute("INSERT IGNORE INTO ukolovnik_projekt_clenove (projekt_id, user_id, jmeno, role) VALUES (%s,%s,%s,%s)",
                                                   (projekt_id, uid2, jm, novy_role.value))
                                        c.commit(); cu.close(); c.close()
                                        _log('projekt', projekt_id, user_id, user_name, f'Přidal/a člena: {jm} (role: {novy_role.value})')
                                        try: intranet_notifikace.pridej(uid2, f"🚀 {user_name} tě přidal/a do projektu '{proj['nazev'][:50]}'", 'info')
                                        except Exception: pass
                                        dlg2.close()
                                        nacti_cleny()
                                    with ui.row().classes('w-full justify-between'):
                                        ui.button('Zrušit', on_click=dlg2.close).classes('bg-gray-400 text-white font-bold')
                                        ui.button('Přidat', on_click=_ulozit_clena).classes('bg-blue-600 text-white font-bold')
                                dlg2.open()
                            ui.button(icon='person_add', on_click=_pridat_clena_dialog).props('round dense flat').classes('text-blue-500')

                    clenove_box = ui.column().classes('w-full gap-1')

                    def nacti_cleny():
                        clenove_box.clear()
                        c = intranet_data.get_db_connection()
                        clenove = []
                        if c:
                            try:
                                cu = c.cursor(dictionary=True)
                                cu.execute("SELECT * FROM ukolovnik_projekt_clenove WHERE projekt_id=%s ORDER BY role, jmeno", (projekt_id,))
                                clenove = cu.fetchall()
                            finally:
                                cu.close(); c.close()
                        with clenove_box:
                            if not clenove:
                                ui.label('Žádní členové.').classes('text-xs text-gray-400 italic')
                            for cl in clenove:
                                role_lbl, role_cls = ROLE_PROJEKT.get(cl['role'], ('👤', 'text-gray-600'))
                                with ui.row().classes('w-full items-center gap-2 p-2 bg-white border border-gray-100 rounded-lg'):
                                    ui.label(role_lbl).classes(f'text-xs {role_cls}')
                                    ui.label(cl['jmeno']).classes('text-sm text-gray-800 flex-1 font-bold')
                                    if je_vlastnik and cl['user_id'] != user_id:
                                        def _odebrat(cid=cl['id'], jmeno=cl['jmeno']):
                                            c2 = intranet_data.get_db_connection()
                                            cu2 = c2.cursor()
                                            cu2.execute("DELETE FROM ukolovnik_projekt_clenove WHERE id=%s", (cid,))
                                            c2.commit(); cu2.close(); c2.close()
                                            nacti_cleny()
                                        ui.button(icon='close', on_click=_odebrat).props('flat round dense size=xs').classes('text-gray-300 hover:text-red-400')

                    nacti_cleny()

                    # ── PŘÍLOHY ────────────────────────────────────────────────────
                    ui.separator().classes('my-1')
                    ui.label('Přílohy').classes('text-sm font-extrabold text-gray-700')

                    prilohy_box = ui.column().classes('w-full gap-1')

                    def nacti_prilohy_proj():
                        prilohy_box.clear()
                        c = intranet_data.get_db_connection()
                        prilohy = []
                        if c:
                            try:
                                cu = c.cursor(dictionary=True)
                                cu.execute("SELECT * FROM ukolovnik_projekt_prilohy WHERE projekt_id=%s ORDER BY vytvoreno ASC", (projekt_id,))
                                prilohy = cu.fetchall()
                            finally:
                                cu.close(); c.close()
                        with prilohy_box:
                            if not prilohy:
                                ui.label('Žádné přílohy.').classes('text-xs text-gray-400 italic')
                            for pr in prilohy:
                                ikona, ik_cls = _ikona_souboru(pr['soubor_nazev'])
                                with ui.row().classes('w-full items-center gap-2 p-2 bg-white border border-gray-100 rounded-lg'):
                                    ui.icon(ikona, size='xs').classes(ik_cls)
                                    ui.link(pr['soubor_nazev'],
                                            f"/projekt_prilohy/{os.path.basename(pr['soubor_cesta'])}").classes('text-xs text-blue-600 hover:underline flex-1 truncate').props('target=_blank')
                                    ui.label(pr['vytvoreno'].strftime('%d.%m.')).classes('text-[10px] text-gray-400 shrink-0')

                    nacti_prilohy_proj()

                    async def zpracuj_upload_proj(e: ui_events.UploadEventArguments):
                        try:
                            sid = str(uuid.uuid4())
                            nazev = e.file.name
                            ext = os.path.splitext(nazev)[1]
                            cesta = os.path.join(PROJEKT_PRILOHY_DIR, f'{sid}{ext}')
                            os.makedirs(PROJEKT_PRILOHY_DIR, exist_ok=True)
                            obsah = await e.file.read()
                            with open(cesta, 'wb') as f: f.write(obsah)
                            c = intranet_data.get_db_connection()
                            cu = c.cursor()
                            cu.execute("INSERT INTO ukolovnik_projekt_prilohy (id, projekt_id, user_id, jmeno_autora, soubor_nazev, soubor_cesta) VALUES (%s,%s,%s,%s,%s,%s)",
                                       (sid, projekt_id, user_id, user_name, nazev, cesta))
                            c.commit(); cu.close(); c.close()
                            intranet_logger.log_activity(user_name, "Projekty", f"Příloha k projektu #{projekt_id}: {nazev}")
                            _log('projekt', projekt_id, user_id, user_name, f'Nahrál/a přílohu: {nazev}')
                            _notifikuj_projekt(projekt_id, user_id, f"📎 {user_name} přidal/a přílohu '{nazev}' do projektu '{proj['nazev'][:40]}'.")
                            ui.notify(f'Nahráno: {nazev}', type='positive')
                            nacti_prilohy_proj()
                        except Exception as ex:
                            ui.notify(f'Chyba: {ex}', type='negative')

                    ui.upload(label='Nahrát přílohu', on_upload=zpracuj_upload_proj, auto_upload=True, multiple=True
                              ).props('accept=".pdf,.xlsx,.xls,.csv,.jpg,.jpeg,.png,.doc,.docx" flat color=blue-grey size=sm').classes('w-full mt-1')

                    # ── ÚKOLY PROJEKTU ─────────────────────────────────────────────
                    ui.separator().classes('my-1')
                    ui.label('Úkoly projektu').classes('text-sm font-extrabold text-gray-700')

                    ukoly_proj_box = ui.column().classes('w-full gap-1')

                    def nacti_ukoly_proj():
                        ukoly_proj_box.clear()
                        c = intranet_data.get_db_connection()
                        ukoly = []
                        if c:
                            try:
                                cu = c.cursor(dictionary=True)
                                cu.execute("SELECT * FROM ukolovnik_ukoly WHERE projekt_id=%s ORDER BY vytvoreno DESC", (projekt_id,))
                                ukoly = cu.fetchall()
                            finally:
                                cu.close(); c.close()
                        with ukoly_proj_box:
                            if not ukoly:
                                ui.label('Žádné úkoly.').classes('text-xs text-gray-400 italic')
                            for uk in ukoly:
                                b_uk, _ = STAVY_UKOL.get(uk['stav'], ('bg-gray-100 text-gray-700 border-gray-300', ''))
                                def _det_uk(u=uk):
                                    _dialog_detail_ukolu(u['id'], user_id, user_name, vsechna_prava,
                                                         on_refresh=lambda: [nacti_ukoly_proj(), on_refresh()])
                                je_h = uk['stav'] == 'Hotovo'
                                with ui.row().classes('w-full items-center gap-2 p-2 bg-white border border-gray-100 rounded-lg cursor-pointer hover:shadow-sm').on('click', lambda u=uk: _det_uk(u)):
                                    ui.label(uk['nazev']).classes('text-xs flex-1 truncate ' + ('line-through text-gray-400' if je_h else 'text-gray-800 font-bold'))
                                    ui.label(uk['stav']).classes(f'text-[10px] font-bold border px-1 rounded {b_uk} shrink-0')

                    nacti_ukoly_proj()

        _telo()

    _registruj_otevreny_detail('projekt', projekt_id, dlg, _telo.refresh)
    dlg.open()


# =========================================================
# SEKCE: PROJEKTY — seznam
# =========================================================
def _nacti_projekty_seznam(user_id, is_admin, je_vedouci, odd_user_ids, ma_oddeleni) -> list:
    """Seznam projektů dle viditelnosti uživatele — čistý DB loader pro vlákno."""
    conn = intranet_data.get_db_connection()
    projekty = []
    if conn:
        try:
            cur = conn.cursor(dictionary=True)
            if is_admin:
                cur.execute("SELECT * FROM ukolovnik_projekty ORDER BY vytvoreno DESC")
            elif je_vedouci and ma_oddeleni:
                placeholders = ','.join(['%s'] * len(odd_user_ids))
                cur.execute(f"""
                    SELECT DISTINCT p.* FROM ukolovnik_projekty p
                    LEFT JOIN ukolovnik_projekt_clenove c ON p.id = c.projekt_id
                    LEFT JOIN ukolovnik_sledovaci s ON s.typ='projekt' AND s.ref_id=p.id AND s.user_id=%s
                    WHERE p.vytvoril_id IN ({placeholders})
                       OR c.user_id IN ({placeholders})
                       OR s.id IS NOT NULL
                    ORDER BY p.vytvoreno DESC
                """, [user_id] + odd_user_ids * 2)
            else:
                cur.execute("""
                    SELECT DISTINCT p.* FROM ukolovnik_projekty p
                    LEFT JOIN ukolovnik_projekt_clenove c ON p.id = c.projekt_id
                    LEFT JOIN ukolovnik_sledovaci s ON s.typ='projekt' AND s.ref_id=p.id AND s.user_id=%s
                    WHERE p.vytvoril_id=%s OR c.user_id=%s OR s.id IS NOT NULL
                    ORDER BY p.vytvoreno DESC
                """, (user_id, user_id, user_id))
            projekty = cur.fetchall()
        finally:
            cur.close(); conn.close()
    return projekty


@ui.refreshable
async def _vykresli_projekty(user_id, user_name, vsechna_prava):
    is_admin = 'vse' in vsechna_prava or 'ukolovnik_admin' in vsechna_prava

    uzivatele_oddeleni = [] if is_admin else await asyncio.to_thread(_ziskej_uzivatele_oddeleni, vsechna_prava)
    je_vedouci = is_admin or _je_hlavni_vedouci(vsechna_prava)
    odd_user_ids = [u['id'] for u in uzivatele_oddeleni] if uzivatele_oddeleni else [user_id]

    projekty = await asyncio.to_thread(
        _nacti_projekty_seznam, user_id, is_admin, je_vedouci, odd_user_ids, bool(uzivatele_oddeleni))

    # Starý obsah smazat až s daty v ruce → přestavba bez bliknutí a skoku scrollu
    ui.context.slot.parent.clear()

    with ui.row().classes('w-full items-center justify-between mb-5'):
        with ui.column().classes('gap-0'):
            ui.label('Projekty').classes('text-3xl font-extrabold text-gray-800')
            ui.label('Projektové nápady i realizace — od A do Z, se spolupracovníky, chatem a přílohami.').classes('text-xs text-gray-400')
        ui.button('Nový projekt', icon='rocket_launch',
                  on_click=lambda: _dialog_novy_projekt(user_id, user_name, _projekty_refresh_klient, vsechna_prava=vsechna_prava)
                  ).classes('bg-indigo-600 text-white font-bold h-10 px-5 rounded-xl shadow-md')

    # Filtry
    with ui.row().classes('w-full gap-2 mb-4 flex-wrap items-center'):
        filtr_stav = ui.select(
            {'': 'Všechny stavy', **{k: k for k in STAVY_PROJEKT}},
            value='', label=''
        ).classes('w-44').props('dense outlined')
        filtr_moje = ui.select(
            {'': 'Všechny', 'moje': 'Moje projekty', 'clen': 'Jsem člen'},
            value='', label=''
        ).classes('w-44').props('dense outlined')

    if not projekty:
        with ui.card().classes('w-full p-12 items-center bg-gray-50 border border-dashed border-gray-200 rounded-xl'):
            ui.icon('rocket_launch', size='4rem', color='gray-400').classes('mb-4')
            ui.label('Zatím žádné projekty.').classes('text-xl text-gray-500 font-bold')
        return

    karty_box = ui.element('div').classes('w-full grid gap-4 grid-cols-1 md:grid-cols-2 xl:grid-cols-3')

    def prekresli_projekty():
        karty_box.clear()
        fs = filtr_stav.value
        fm = filtr_moje.value
        zobrazene = [p for p in projekty if
                     (not fs or p['stav'] == fs) and
                     (not fm or (fm == 'moje' and p.get('vytvoril_id') == user_id) or fm == 'clen')]
        with karty_box:
            if not zobrazene:
                ui.label('Žádné projekty odpovídají filtru.').classes('text-gray-400 italic col-span-3')
                return
            for p in zobrazene:
                b_stav = STAVY_PROJEKT.get(p['stav'], 'bg-gray-100 text-gray-800 border-gray-300')
                termin_str = p['termin'].strftime('%d.%m.%Y') if p.get('termin') else '—'
                delta, dt_cls, dt_txt = _dny_do_terminu(p.get('termin'))
                je_prevedeno = bool(p.get('prevedeno_na_ukol'))

                # Počet zpráv + členů
                c2 = intranet_data.get_db_connection()
                pocet_zpr = pocet_cl = pocet_pr = 0
                if c2:
                    try:
                        cu = c2.cursor(dictionary=True)
                        cu.execute("SELECT COUNT(*) AS n FROM ukolovnik_projekt_komentare WHERE projekt_id=%s", (p['id'],))
                        pocet_zpr = cu.fetchone()['n']
                        cu.execute("SELECT COUNT(*) AS n FROM ukolovnik_projekt_clenove WHERE projekt_id=%s", (p['id'],))
                        pocet_cl = cu.fetchone()['n']
                        cu.execute("SELECT COUNT(*) AS n FROM ukolovnik_projekt_prilohy WHERE projekt_id=%s", (p['id'],))
                        pocet_pr = cu.fetchone()['n']
                    finally:
                        cu.close(); c2.close()

                barva_pruhu = {
                    'Aktivní': 'bg-indigo-500', 'Pozastavený': 'bg-yellow-400',
                    'Dokončený': 'bg-green-500', 'Zrušený': 'bg-gray-400'
                }.get(p['stav'], 'bg-gray-300')

                def _otevrit(pid=p['id']):
                    _dialog_detail_projektu(pid, user_id, user_name, vsechna_prava, on_refresh=_projekty_refresh_klient)

                with ui.card().classes('w-full p-0 overflow-hidden rounded-xl border border-gray-100 bg-white hover:shadow-xl transition-shadow cursor-pointer').on('click', lambda pid=p['id']: _otevrit(pid)):
                    # Barevný pruh nahoře
                    ui.element('div').classes(f'w-full h-1.5 {barva_pruhu}')
                    with ui.column().classes('p-4 gap-2'):
                        with ui.row().classes('w-full justify-between items-start'):
                            with ui.column().classes('flex-1 gap-0.5 min-w-0'):
                                with ui.row().classes('items-center gap-1'):
                                    ui.icon('rocket_launch', size='xs', color='indigo-400')
                                    ui.label(p['nazev']).classes('font-black text-base text-gray-800 line-clamp-1')
                                ui.label(f"Vytvořil: {p.get('vytvoril_jmeno','—')}").classes('text-xs text-gray-400')
                            with ui.column().classes('items-end gap-1 shrink-0'):
                                ui.label(p['stav']).classes(f'px-2 py-0.5 rounded text-[10px] font-bold border {b_stav}')
                                if je_prevedeno:
                                    ui.label('✅ V úkolech').classes('text-[10px] text-green-600 font-bold')
                                with ui.row().classes('items-center gap-0'):
                                    ui.button(icon='visibility').props('flat round dense size=xs').classes(
                                        'text-gray-300 hover:text-blue-500'
                                    ).on('click.stop', lambda pp=p: _dialog_log('projekt', pp['id'], pp['nazev'], user_id, user_name))
                                    if is_admin or p.get('vytvoril_id') == user_id:
                                        def _del_p(pp=p):
                                            def _provest():
                                                if _smazat_projekt_db(pp['id']):
                                                    intranet_logger.log_activity(user_name, "Projekty", f"Smazán projekt #{pp['id']}: {pp['nazev']}")
                                                    ui.notify('Projekt smazán.', type='positive', position='top')
                                                    _projekty_refresh_klient()
                                                else:
                                                    ui.notify('Chyba při mazání projektu.', type='negative', position='top')
                                            _potvrdit_smazat(f'Opravdu smazat projekt „{pp["nazev"]}"?\nTato akce je nevratná.', _provest)
                                        ui.button(icon='delete_outline').props('flat round dense size=xs').classes(
                                            'text-gray-300 hover:text-red-500'
                                        ).on('click.stop', lambda pp=p: _del_p(pp))

                        if p.get('popis'):
                            ui.label(p['popis']).classes('text-xs text-gray-500 line-clamp-2')

                        with ui.row().classes('w-full justify-between items-center pt-2 border-t border-gray-50 flex-wrap gap-1'):
                            with ui.row().classes('items-center gap-1 text-xs text-gray-400'):
                                ui.icon('calendar_today', size='xs')
                                ui.label(termin_str).classes(dt_cls if p.get('termin') and p['stav'] == 'Aktivní' else 'text-gray-400')
                                if delta is not None and p['stav'] == 'Aktivní':
                                    ui.label(f'· {dt_txt}').classes(f'text-xs {dt_cls}')
                            with ui.row().classes('items-center gap-3 text-xs text-gray-400'):
                                with ui.row().classes('items-center gap-0.5'):
                                    ui.icon('people', size='xs'); ui.label(str(pocet_cl))
                                with ui.row().classes('items-center gap-0.5'):
                                    ui.icon('chat', size='xs'); ui.label(str(pocet_zpr))
                                with ui.row().classes('items-center gap-0.5'):
                                    ui.icon('attach_file', size='xs'); ui.label(str(pocet_pr))

    prekresli_projekty()
    filtr_stav.on('update:model-value', lambda: prekresli_projekty())
    filtr_moje.on('update:model-value', lambda: prekresli_projekty())


# =========================================================
# HLAVNÍ FUNKCE
# =========================================================
_UKOLOVNIK_SEKCE = ['porady', 'ukoly', 'kalendar', 'kapacita', 'projekty', 'statistika']

def _uk_viditelne_sekce(vsechna_prava) -> set:
    """Vrátí set názvů záložek, které má uživatel právo vidět."""
    if 'vse' in vsechna_prava or 'ukolovnik_admin' in vsechna_prava:
        return set(_UKOLOVNIK_SEKCE)
    return {s for s in _UKOLOVNIK_SEKCE if f'ukolovnik_{s}' in vsechna_prava}


async def vykresli_ukolovnik(user_id, user_name, vsechna_prava):
    await asyncio.to_thread(inicializace_ukolovnik_db)

    # Cross-browser drag&drop: nastaví dataTransfer při startu tažení (nutné pro Firefox).
    # Window-guard zabrání vícenásobné registraci posluchače.
    ui.add_body_html("""
    <script>
    if (!window.__uk_dnd_init) {
        window.__uk_dnd_init = true;
        document.addEventListener('dragstart', function(e){
            var t = e.target;
            if (t && t.getAttribute && t.getAttribute('draggable') === 'true') {
                try { e.dataTransfer.setData('text/plain', ''); e.dataTransfer.effectAllowed = 'move'; } catch (_) {}
            }
        }, true);
    }
    </script>
    """)

    dnes = datetime.date.today()
    mesic_stav = {'rok': dnes.year, 'mesic': dnes.month, 'typ': 'plan', 'nahled': 'tyden', 'den': dnes.day}
    kap_stav = {
        'pohled': 'moje', 'vybrany_uid': None, 'tyden_offset': 0,
        'stat_od': (dnes - datetime.timedelta(days=29)).strftime('%d.%m.%Y'),
        'stat_do': dnes.strftime('%d.%m.%Y'),
        'zobraz_stat': False,
    }
    stat_stav = {'pohled': 'moje', 'vybrany_uid': None, 'obdobi': 'tyden',
                 'drill_uid': None, 'sort': 'po_terminu', 'sort_dir': 'desc'}
    # Stav přehledu úkolů (stránka, řazení, sbalené dny, pohled) — drží se napříč refreshi
    ukoly_stav: dict = {}

    viditelne = _uk_viditelne_sekce(vsechna_prava)
    prvni = next((s for s in _UKOLOVNIK_SEKCE if s in viditelne), None)

    if not prvni:
        with ui.card().classes('w-full p-12 items-center bg-gray-50 border border-dashed border-gray-200 rounded-xl'):
            ui.icon('lock', size='4rem', color='gray-400').classes('mb-4')
            ui.label('Nemáte přístup k žádné sekci modulu Porady a úkoly.').classes('text-xl text-gray-500 font-bold')
        return

    _TAB_DEF = [
        ('porady',     'Porady a zápisy',                'meeting_room'),
        ('ukoly',      'Zadávání úkolů a přehled úkolů', 'task_alt'),
        ('kalendar',   'Kalendář',                       'calendar_month'),
        ('kapacita',   'Kapacita oddělení',              'groups'),
        ('projekty',   'Projekty',                       'rocket_launch'),
        ('statistika', 'Statistika',                     'bar_chart'),
    ]

    # Aktivní sekce se pamatuje napříč reloady stránky (stejně jako hlavní
    # navigace intranetu). Pokud uložená sekce už není viditelná (změna práv),
    # spadne to zpět na první dostupnou.
    ulozena_sekce = app.storage.user.get('ukolovnik_sekce')
    if ulozena_sekce not in viditelne:
        ulozena_sekce = prvni
    app.storage.user['ukolovnik_sekce'] = ulozena_sekce

    with ui.tabs().classes('w-full mb-6').bind_value(app.storage.user, 'ukolovnik_sekce') as tabs_uk:
        for key, label, icon in _TAB_DEF:
            if key in viditelne:
                ui.tab(key, label=label, icon=icon)

    with ui.tab_panels(tabs_uk, value=ulozena_sekce).bind_value(app.storage.user, 'ukolovnik_sekce').classes('w-full bg-transparent p-0'):
        if 'porady' in viditelne:
            with ui.tab_panel('porady'):
                await _vykresli_porady(user_id, user_name, vsechna_prava)
        if 'ukoly' in viditelne:
            with ui.tab_panel('ukoly'):
                # Kontejner pro dialogy MIMO refreshable – přežije _vykresli_ukoly.refresh()
                _dlg_anchor = ui.element('div')
                await _vykresli_ukoly(user_id, user_name, vsechna_prava, _dlg_anchor, ukoly_stav)
        if 'kalendar' in viditelne:
            with ui.tab_panel('kalendar'):
                await _vykresli_kalendar(user_id, user_name, vsechna_prava, mesic_stav)
        if 'kapacita' in viditelne:
            with ui.tab_panel('kapacita'):
                await _vykresli_kapacita(user_id, user_name, vsechna_prava, kap_stav)
        if 'projekty' in viditelne:
            with ui.tab_panel('projekty'):
                await _vykresli_projekty(user_id, user_name, vsechna_prava)
        if 'statistika' in viditelne:
            with ui.tab_panel('statistika'):
                await _vykresli_statistika(user_id, user_name, vsechna_prava, stat_stav)

    # ── Živá synchronizace mezi uživateli ────────────────────────────────────
    # Každá mutace úkolu (přehození, vytvoření, změna stavu…) posune globální
    # revizi (viz _bump_data_verze v _log). Každý klient si ji tady poll-uje a při
    # změně překreslí SVÉ sekce. Tím se např. přehození úkolu projeví novému
    # i původnímu zpracovateli v reálném čase, bez obnovení stránky.
    try:
        _klient_uk = ui.context.client
    except Exception:
        _klient_uk = None
    try:
        _init_data_verze = app.storage.general.get('ukolovnik_data_verze', 0)
    except Exception:
        _init_data_verze = 0
    app.storage.user.setdefault('ukolovnik_data_verze_vlastni', _init_data_verze)

    _ZIVE_SEKCE = {
        'porady': _vykresli_porady, 'ukoly': _vykresli_ukoly, 'kalendar': _vykresli_kalendar,
        'kapacita': _vykresli_kapacita, 'projekty': _vykresli_projekty, 'statistika': _vykresli_statistika,
    }

    async def _kontrola_data_verze():
        if _klient_uk is None or not _klient_uk.has_socket_connection:
            return
        try:
            gv = app.storage.general.get('ukolovnik_data_verze', 0)
        except Exception:
            return
        stara = app.storage.user.get('ukolovnik_data_verze_vlastni', _init_data_verze)
        if gv == stara:
            # žádná změna od posledně — jen případné odložené obnovy detailů
            _obnov_otevrene_detaily(_klient_uk, [])
            return
        app.storage.user['ukolovnik_data_verze_vlastni'] = gv
        try:
            zmeny = [z for z in app.storage.general.get('ukolovnik_zmeny', [])
                     if z.get('v', 0) > stara]
        except Exception:
            zmeny = []
        # Rychlá cesta: samé změny úkolů (a fronta je pokrývá celé) → živě se
        # překreslí jen dotčené karty; seznam nepřijde o scroll ani o dialogy.
        _zvladnuto = False
        if (zmeny and (gv - stara) <= len(zmeny) and 'ukoly' in viditelne
                and all(z.get('typ') == 'ukol' and z.get('ref_id') is not None for z in zmeny)):
            _handler = _UK_ZIVE_RADKY.get(_klient_uk.id)
            if _handler is not None:
                try:
                    _zvladnuto = bool(await _handler({int(z['ref_id']) for z in zmeny}))
                except Exception:
                    _zvladnuto = False
        for _sekce, _rf in _ZIVE_SEKCE.items():
            if _sekce == 'ukoly' and _zvladnuto:
                continue
            if _sekce in viditelne:
                _refresh_refreshable_klient(_rf)
        # Živá obnova otevřených náhledů: obsah detailu, který mezitím někdo
        # změnil, se překreslí na místě (smazaný záznam detail zavře s hláškou).
        _obnov_otevrene_detaily(_klient_uk, zmeny)

    if _klient_uk is not None and _klient_uk.has_socket_connection:
        ui.timer(2.0, _kontrola_data_verze)
