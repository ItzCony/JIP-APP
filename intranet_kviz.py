from nicegui import ui, app, Client
import pandas as pd
import os
import time
import json
import zipfile
import asyncio
from openpyxl.styles import PatternFill

import intranet_data
import intranet_logger
import intranet_session

aktivni_studenti = {}
dokoncene_testy = []
SOUBOR_NASTAVENI = 'nastaveni_kvizu.json'

# ─── výchozí nastavení jedné skupiny (OZ nebo ASM) ───────────────────────────
_SKUPINA_DEFAULT = {
    'pocet_otazek': 20,
    'casovy_limit_minuty': 30,
    'min_procento_uspesnosti': 80,
    'okruh_1_nazev': 'Okruh 1', 'okruh_1_pocet': 20, 'okruh_1_zapnut': True,
    'okruh_2_nazev': 'Okruh 2', 'okruh_2_pocet': 0,  'okruh_2_zapnut': False,
    'okruh_3_nazev': 'Okruh 3', 'okruh_3_pocet': 0,  'okruh_3_zapnut': False,
}

config = {
    'aktivni_skupina': 'oz',   # globální nastavení — který kvíz dostanou studenti
    'oz':  {**_SKUPINA_DEFAULT},
    'asm': {**_SKUPINA_DEFAULT},
}

# ─── načtení uloženého nastavení (s migrací ze starého flat formátu) ──────────
if os.path.exists(SOUBOR_NASTAVENI):
    try:
        with open(SOUBOR_NASTAVENI, 'r', encoding='utf-8') as f:
            nacteno = json.load(f)
        if 'oz' in nacteno and 'asm' in nacteno:
            # nový vnořený formát
            config['aktivni_skupina'] = nacteno.get('aktivni_skupina', 'oz')
            for typ in ('oz', 'asm'):
                for k in _SKUPINA_DEFAULT:
                    config[typ][k] = nacteno[typ].get(k, _SKUPINA_DEFAULT[k])
        elif 'pocet_otazek' in nacteno:
            # starý plochý formát — přenes do oz
            for k in _SKUPINA_DEFAULT:
                if k in nacteno:
                    config['oz'][k] = nacteno[k]
    except Exception:
        pass

# ─── soubory s otázkami pro každou skupinu ───────────────────────────────────
SOUBORY_OKRUHU = {
    'oz':  ['otazky.csv',     'otazky2.csv',     'otazky3.csv'],
    'asm': ['otazky_asm.csv', 'otazky2_asm.csv', 'otazky3_asm.csv'],
}


def get_max_questions(soubor):
    if not os.path.exists(soubor): return 0
    try:
        with open(soubor, 'r', encoding='utf-8', errors='ignore') as f:
            return max(0, sum(1 for line in f if line.strip()) - 1)
    except: return 0


def _nacti_csv(soubor):
    try: df = pd.read_csv(soubor, sep=';', encoding='utf-8')
    except UnicodeDecodeError: df = pd.read_csv(soubor, sep=';', encoding='cp1250')
    if len(df.columns) == 6: df.columns = ['OTÁZKA', 'A', 'B', 'C', 'D', 'správná odpověď']
    return df


def uloz_nastaveni(typ):
    """Uloží nastavení kvízu pro danou skupinu ('oz' nebo 'asm')."""
    c = config[typ]
    try:
        c['pocet_otazek']            = int(c['pocet_otazek'] or 1)
        c['casovy_limit_minuty']     = int(c['casovy_limit_minuty'] or 1)
        c['min_procento_uspesnosti'] = int(c['min_procento_uspesnosti'] or 0)
        c['okruh_1_pocet']           = int(c['okruh_1_pocet'] or 0)
        c['okruh_2_pocet']           = int(c['okruh_2_pocet'] or 0)
        c['okruh_3_pocet']           = int(c['okruh_3_pocet'] or 0)
        soucet = sum(
            c[f'okruh_{i}_pocet']
            for i in (1, 2, 3)
            if c.get(f'okruh_{i}_zapnut', i == 1)
        )
        if soucet != c['pocet_otazek']:
            ui.notify(
                f'Součet otázek z okruhů ({soucet}) se nerovná celkovému počtu '
                f'({c["pocet_otazek"]})! Opravte nastavení.',
                type='negative', position='top', timeout=6000)
            return
        intranet_data.zapis_json_atomicky(SOUBOR_NASTAVENI, config, indent=4)
        label = 'ASM' if typ == 'asm' else 'OZ'
        ui.notify(f'Nastavení {label} bylo úspěšně uloženo!', type='positive', position='top')
    except Exception as e:
        ui.notify(f'Chyba při ukládání: {e}', type='negative', position='top')


def nacti_otazky(typ='oz', pocet_otazek=20):
    """Načte otázky pro danou skupinu ('oz' nebo 'asm')."""
    c       = config[typ]
    soubory = SOUBORY_OKRUHU[typ]
    zapnute = [c.get('okruh_1_zapnut', True), c.get('okruh_2_zapnut', False), c.get('okruh_3_zapnut', False)]
    pocty   = [
        int(c['okruh_1_pocet'] or 0) if zapnute[0] else 0,
        int(c['okruh_2_pocet'] or 0) if zapnute[1] else 0,
        int(c['okruh_3_pocet'] or 0) if zapnute[2] else 0,
    ]
    soucet = sum(pocty)
    pouzit_okruhy = soucet > 0 and soucet == int(pocet_otazek)

    if pouzit_okruhy:
        vsechny = []
        for i, (soubor, pocet) in enumerate(zip(soubory, pocty), 1):
            if pocet <= 0:
                continue
            if not os.path.exists(soubor):
                ui.notify(f'Chybí soubor {soubor} pro Okruh {i}! Nahrajte jej v administraci.', type='negative', position='top')
                return []
            df = _nacti_csv(soubor)
            if len(df) < pocet:
                ui.notify(f'Okruh {i}: soubor má jen {len(df)} otázek, požadováno {pocet}.', type='negative', position='top')
                return []
            vsechny.extend(df.sample(n=pocet).to_dict('records'))
        import random; random.shuffle(vsechny)
        return vsechny
    else:
        soubor = soubory[0]
        if not os.path.exists(soubor):
            ui.notify(f'Chybí soubor {soubor}! Nahrajte jej v administraci.', type='negative', position='top')
            return []
        df = _nacti_csv(soubor)
        pocet_int = int(min(pocet_otazek, len(df)))
        return df.sample(n=pocet_int).to_dict('records')

def _nav_na_prehled():
    with ui.dialog().props('maximized persistent transition-show="fade" transition-hide="fade"') as _dlg_back:
        with ui.column().classes('w-full h-full items-center justify-center bg-gray-50 m-0 p-0'):
            ui.spinner('dots', size='5rem', color='blue').classes('mb-4')
            ui.label('Otevírám modul').classes('text-sm font-bold text-gray-500 uppercase tracking-widest')
            ui.label('Přehled').classes('text-3xl font-black text-blue-900')
    _dlg_back.open()
    ui.timer(0.2, lambda: ui.navigate.to('/'), once=True)

@ui.page('/kviz')
def index_page(client: Client):

    ui.add_head_html('''
        <style>
            @media print {
                .skryt-pri-tisku { display: none !important; }
                body { background-color: white !important; background-image: none !important; }
                .q-dialog__backdrop { display: none !important; }
            }
            body.kviz-pozadi {
                margin: 0 !important;
                padding: 0 !important;
                overflow: hidden !important;
                background-image: url("/static/pozadi.jpg");
                background-size: cover;
                background-position: center;
                background-repeat: no-repeat;
                height: 100vh;
                width: 100vw;
            }
            /* Plynulý přechod při načtení kvízu */
            #q-app { animation: _kviz_fadein 0.3s ease-out; }
            @keyframes _kviz_fadein { from { opacity: 0; } to { opacity: 1; } }
        </style>
    ''')
    ui.query('body').classes('kviz-pozadi')

    user_id = app.storage.user.get('user_id')
    user_name = app.storage.user.get('user_name')

    if not user_id:
        app.storage.user['redirect_to'] = '/kviz'
        with ui.column().classes('w-full max-w-2xl mx-auto items-center mt-32'):
            ui.label('🚫 Nejste přihlášen(a)').classes('text-5xl font-bold text-red-600 mb-6')
            ui.label('Přesměrovávám na bezpečné přihlášení...').classes('text-2xl text-gray-700 mb-6')
            ui.spinner(size='xl')
            ui.timer(1.5, _nav_na_prehled, once=True)
        return

    # Eviduj živé připojení i pro samostatnou stránku /kviz (single-session
    # + odhlašování po zavření prohlížeče), ať návrat sem nezpůsobí výpadek relace.
    _user_email = str(app.storage.user.get('user_email', '')).lower()
    _token = app.storage.user.get('login_token')
    if _token:
        intranet_session.registruj_pripojeni(client, _token, _user_email, user_name)

    vsechna_prava = intranet_data.ziskej_prava_uzivatele(user_id)
    is_admin = 'vse' in vsechna_prava
    ma_pristup_k_vysledkum = "vystup_osobni" in vsechna_prava or "vystup_vse" in vsechna_prava or "vse" in vsechna_prava

    # Skupina kvízu je řízena globálním nastavením administrátora
    kviz_typ = config.get('aktivni_skupina', 'oz')

    nastaveni = intranet_data.nacti_nastaveni_intranetu()

    # Modul kvízu je záměrně VYJMUT z automatického odhlašování po nečinnosti –
    # uživatel nesmí být odhlášen uprostřed vyplňování testu. Žádný idle-timer zde neběží.

    if not nastaveni.get('kviz_zapnuty', True) and not is_admin:
        with ui.column().classes('w-full max-w-2xl mx-auto items-center mt-32'):
            ui.label('🚫 MODUL JE VYPNUTÝ').classes('text-5xl font-bold text-orange-600 mb-6')
            ui.label('Zkouškový kvíz je momentálně administrátorem deaktivován.').classes('text-2xl text-gray-700 text-center')
            ui.button('Zpět na Přehled', on_click=_nav_na_prehled).classes('mt-10 bg-blue-600 text-white font-bold px-8 py-3 rounded-lg shadow-md')
        return

    if 'quiz_initialized' not in app.storage.browser or app.storage.browser.get('jmeno') != user_name:
        app.storage.browser.update({
            'quiz_initialized': True, 'jmeno': user_name, 'zobrazit_admin_panel': False,
            'otazky': [], 'aktualni_index': 0, 'body': 0, 'zbyva_sekund': 0,
            'bezi_test': False, 'vybrana_odpoved': None, 'vysledek': None, 'historie_odpovedi': []
        })

    state = app.storage.browser
    local_ui = {'cas_label': None}

    def prepni_admin_panel(zobrazit):
        state['zobrazit_admin_panel'] = zobrazit
        hlavni_rozhrani.refresh()

    @ui.refreshable
    def hlavni_rozhrani():
        zarovnani = 'justify-start pt-12' if state.get('zobrazit_admin_panel') else 'justify-center'

        with ui.column().classes(f'w-full max-w-6xl mx-auto items-center {zarovnani} h-screen p-4 overflow-y-auto'):

            if state['zobrazit_admin_panel'] and is_admin:
                with ui.row().classes('w-full justify-between items-center mb-6 pb-4 border-b-2 border-gray-200 dark:border-gray-700 skryt-pri-tisku'):
                    with ui.card().classes('p-2 bg-white/90 backdrop-blur-sm rounded-xl'):
                        ui.label('⚙️ Správa Kvízu').classes('text-4xl font-extrabold text-gray-800')
                    ui.button('Zpět na test', on_click=lambda: prepni_admin_panel(False)).classes('bg-blue-500 hover:bg-blue-600 text-lg rounded-2xl shadow-md')

                with ui.row().classes('w-full items-start no-wrap skryt-pri-tisku pb-10'):
                    with ui.card().classes('w-64 shrink-0 shadow-xl rounded-2xl border border-gray-100 bg-white/95 backdrop-blur-sm mr-6 p-2'):
                        with ui.tabs().props('vertical active-color="primary" indicator-color="primary"').classes('w-full text-left') as tabs:
                            tab_monitor = ui.tab('📡 Live Monitor').classes('justify-start')
                            tab_parametry = ui.tab('⚙️ Nastavení').classes('justify-start')
                            tab_otazky = ui.tab('📝 Otázky (CSV)').classes('justify-start')
                            tab_soubory = ui.tab('📁 Excel Výstupy').classes('justify-start')

                    with ui.column().classes('flex-1 w-full'):
                        with ui.tab_panels(
                            tabs,
                            value=state.get('admin_aktivni_tab', tab_monitor)
                        ).classes('w-full bg-transparent p-0').bind_value(state, 'admin_aktivni_tab'):

                            with ui.tab_panel(tab_monitor):
                                with ui.card().classes('w-full p-6 shadow-xl rounded-2xl border border-gray-100 bg-white/95 backdrop-blur-sm mb-6'):
                                    ui.label('▶ Aktuálně píšící studenti').classes('text-2xl font-bold mb-4 text-purple-600')
                                    monitor_kontejner = ui.column().classes('w-full')
                                    def obnov_monitor():
                                        monitor_kontejner.clear()

                                        aktualni_cas = time.time()
                                        mrtve_duse = [k for k, v in aktivni_studenti.items() if aktualni_cas - v.get('last_seen', aktualni_cas) > 5]
                                        for md in mrtve_duse:
                                            del aktivni_studenti[md]

                                        with monitor_kontejner:
                                            if not aktivni_studenti:
                                                ui.label('Zatím nikdo nedělá test.').classes('text-gray-500 italic')
                                            else:
                                                sloupce = [{'name': 'jmeno', 'label': 'Jméno studenta', 'field': 'jmeno', 'align': 'left'}, {'name': 'otazka', 'label': 'Aktuální otázka', 'field': 'otazka', 'align': 'center'}, {'name': 'cas', 'label': 'Zbývající čas', 'field': 'cas', 'align': 'center'}]
                                                data = []
                                                for st in aktivni_studenti.values():
                                                    m, s = divmod(int(st['zbyva_sekund']), 60)
                                                    data.append({'jmeno': st['jmeno'], 'otazka': f"{st['aktualni_index'] + 1} / {st['celkem_otazek']}", 'cas': f"{m:02d}:{s:02d}"})
                                                ui.table(columns=sloupce, rows=data).classes('w-full')
                                    ui.timer(3.0, obnov_monitor)

                                with ui.card().classes('w-full p-6 shadow-xl rounded-2xl border border-gray-100 bg-white/95 backdrop-blur-sm'):
                                    ui.label('✅ Dokončené testy').classes('text-2xl font-bold mb-4 text-green-600')
                                    dokoncene_kontejner = ui.column().classes('w-full')
                                    def obnov_dokoncene():
                                        dokoncene_kontejner.clear()
                                        with dokoncene_kontejner:
                                            if not dokoncene_testy: ui.label('Zatím nikdo nedokončil test.').classes('text-gray-500 italic')
                                            else:
                                                cols_dokonceno = [{'name': 'jmeno', 'label': 'Jméno', 'field': 'jmeno', 'align': 'left'}, {'name': 'cas', 'label': 'Odevzdáno', 'field': 'cas_odevzdani', 'align': 'center'}, {'name': 'uspesnost', 'label': 'Úspěšnost', 'field': 'uspesnost', 'align': 'center'}, {'name': 'skore', 'label': 'Skóre', 'field': 'skore', 'align': 'center'}, {'name': 'stav', 'label': 'Stav testu', 'field': 'vysledek', 'align': 'center'}]
                                                ui.table(columns=cols_dokonceno, rows=dokoncene_testy).classes('w-full')
                                    ui.timer(5.0, obnov_dokoncene)

                            with ui.tab_panel(tab_parametry):
                                with ui.card().classes('w-full p-8 shadow-xl rounded-2xl border border-gray-100 bg-white/95 backdrop-blur-sm'):

                                    # ── AKTUÁLNÍ SKUPINA (globální přepínač) ─
                                    with ui.card().classes('w-full p-5 mb-6 border-2 border-blue-300 rounded-2xl bg-blue-50'):
                                        with ui.row().classes('items-center gap-4 flex-wrap'):
                                            with ui.column().classes('gap-1'):
                                                ui.label('Aktuální skupina').classes('text-lg font-extrabold text-blue-800')

                                            def _zmen_aktivni_skupinu(hodnota):
                                                config['aktivni_skupina'] = hodnota
                                                try:
                                                    intranet_data.zapis_json_atomicky(SOUBOR_NASTAVENI, config, indent=4)
                                                    ui.notify(f'Aktivní skupina změněna na {hodnota.upper()}', type='positive', position='top')
                                                except Exception as _e:
                                                    ui.notify(f'Chyba při ukládání: {_e}', type='negative', position='top')
                                                hlavni_rozhrani.refresh()

                                            ui.toggle(
                                                {'oz': '🧑‍💼 OZ', 'asm': '👔 ASM'},
                                                value=config.get('aktivni_skupina', 'oz'),
                                                on_change=lambda e: _zmen_aktivni_skupinu(e.value)
                                            ).classes('font-extrabold text-lg')

                                    # ── toggle pro editaci nastavení (ASM / OZ)
                                    if 'admin_kviz_typ' not in state:
                                        state['admin_kviz_typ'] = 'oz'

                                    def _prepni_admin_typ(hodnota):
                                        state['admin_kviz_typ'] = hodnota
                                        hlavni_rozhrani.refresh()

                                    with ui.row().classes('items-center gap-4 mb-6'):
                                        ui.label('Skupina:').classes('font-bold text-gray-700 text-lg')
                                        ui.toggle(
                                            {'oz': '🧑‍💼 OZ', 'asm': '👔 ASM'},
                                            value=state.get('admin_kviz_typ', 'oz'),
                                            on_change=lambda e: _prepni_admin_typ(e.value)
                                        ).classes('font-bold')

                                    edit_typ = state.get('admin_kviz_typ', 'oz')
                                    c = config[edit_typ]
                                    soubory_skupiny = SOUBORY_OKRUHU[edit_typ]
                                    barva_skupiny = 'text-blue-700' if edit_typ == 'oz' else 'text-indigo-700'

                                    ui.label(f'Celkový počet otázek v testu ({edit_typ.upper()}):').classes(f'text-lg font-medium {barva_skupiny}')
                                    with ui.row().classes('w-full items-center mb-8'):
                                        ui.slider(min=1, max=500, step=1).classes('w-2/3').bind_value(c, 'pocet_otazek')
                                        ui.number(min=1, max=500, step=1, format='%d').classes('w-24 ml-4 bg-white').bind_value(c, 'pocet_otazek')

                                    ui.separator().classes('mb-6')
                                    ui.label('Rozdělení otázek do okruhů').classes(f'text-xl font-bold {barva_skupiny} mb-1')
                                    ui.label('Součet otázek z okruhů musí odpovídat celkovému počtu výše.').classes('text-sm text-gray-500 mb-4')

                                    soucet_label = ui.label().classes('text-base font-semibold mb-4')

                                    def aktualizuj_soucet_label(et=edit_typ, cc=c, sl=soucet_label):
                                        try:
                                            s = sum([
                                                int(cc['okruh_1_pocet'] or 0) if cc.get('okruh_1_zapnut', True)  else 0,
                                                int(cc['okruh_2_pocet'] or 0) if cc.get('okruh_2_zapnut', False) else 0,
                                                int(cc['okruh_3_pocet'] or 0) if cc.get('okruh_3_zapnut', False) else 0,
                                            ])
                                            celkem = int(cc['pocet_otazek'] or 1)
                                            ok = s == celkem
                                            sl.set_text(f'Součet zapnutých okruhů: {s}  {"✅ Odpovídá" if ok else f"⚠️ Neodpovídá celkovému počtu {celkem}"}')
                                            sl.classes(replace='text-base font-semibold mb-4 ' + ('text-green-600' if ok else 'text-red-600'))
                                        except Exception: pass

                                    for idx, (nazev_key, pocet_key, zapnut_key, soubor) in enumerate([
                                        ('okruh_1_nazev', 'okruh_1_pocet', 'okruh_1_zapnut', soubory_skupiny[0]),
                                        ('okruh_2_nazev', 'okruh_2_pocet', 'okruh_2_zapnut', soubory_skupiny[1]),
                                        ('okruh_3_nazev', 'okruh_3_pocet', 'okruh_3_zapnut', soubory_skupiny[2]),
                                    ], 1):
                                        max_ok = get_max_questions(soubor)
                                        soubor_stav = f'✅ {max_ok} otázek' if max_ok > 0 else '❌ soubor nenáhrán'
                                        zapnut = c.get(zapnut_key, idx == 1)
                                        with ui.card().classes('w-full p-4 mb-3 border border-gray-200 rounded-xl ' + ('bg-white' if zapnut else 'bg-gray-100 opacity-70')):
                                            with ui.row().classes('w-full items-center gap-4 mb-2'):
                                                tog = ui.switch(f'Okruh {idx}', value=zapnut).bind_value(c, zapnut_key).classes('font-bold text-gray-700')
                                                tog.on('update:model-value', lambda _: aktualizuj_soucet_label())
                                                ui.input(placeholder=f'Název okruhu {idx}').classes('flex-1 bg-white').bind_value(c, nazev_key).bind_enabled_from(c, zapnut_key)
                                                ui.badge(soubor_stav, color='green' if max_ok > 0 else 'red').classes('shrink-0')
                                            with ui.row().classes('w-full items-center gap-3'):
                                                ui.label('Počet losovaných otázek:').classes('text-sm text-gray-600 w-48 shrink-0')
                                                ui.number(min=0, max=max(max_ok, 500), step=1, format='%d').classes('w-24 bg-white').bind_value(c, pocet_key).bind_enabled_from(c, zapnut_key).on('update:model-value', lambda _: aktualizuj_soucet_label())

                                    aktualizuj_soucet_label()
                                    ui.separator().classes('my-6')

                                    ui.label(f'Časový limit testu — {edit_typ.upper()} (min):').classes(f'text-lg font-medium {barva_skupiny}')
                                    with ui.row().classes('w-full items-center mb-6'):
                                        ui.slider(min=1, max=120, step=1).classes('w-2/3 text-blue-500').bind_value(c, 'casovy_limit_minuty')
                                        ui.number(min=1, max=120, step=1, format='%d').classes('w-24 ml-4 bg-white').bind_value(c, 'casovy_limit_minuty')
                                    ui.label(f'Minimální úspěšnost — {edit_typ.upper()} (%):').classes(f'text-lg font-medium {barva_skupiny}')
                                    with ui.row().classes('w-full items-center mb-6'):
                                        ui.slider(min=0, max=100, step=1).classes('w-2/3 text-green-500').bind_value(c, 'min_procento_uspesnosti')
                                        ui.number(min=0, max=100, step=1, format='%d').classes('w-24 ml-4 bg-white').bind_value(c, 'min_procento_uspesnosti')
                                    ui.button(
                                        f'💾 Uložit nastavení {edit_typ.upper()}',
                                        on_click=lambda t=edit_typ: uloz_nastaveni(t)
                                    ).classes('w-64 h-12 mt-4 bg-green-500 hover:bg-green-600 text-white font-bold rounded-2xl shadow-md')

                            with ui.tab_panel(tab_otazky):
                                with ui.card().classes('w-full p-8 shadow-xl rounded-2xl border border-gray-100 bg-white/95 backdrop-blur-sm'):
                                    ui.label('📝 Nahrání souborů s otázkami').classes('text-2xl font-bold mb-2 text-blue-600')
                                    ui.label('Každý okruh má vlastní soubor CSV. Starý soubor bude přepsán. Změny se projeví v dalším testu.').classes('mb-6 text-gray-500')

                                    async def _nahrat_do_souboru(e, cilovy_soubor):
                                        nazev = getattr(e, 'name', '')
                                        if nazev and not nazev.lower().endswith('.csv'):
                                            ui.notify('CHYBA: Lze nahrávat POUZE soubory ve formátu .csv!', type='negative', position='top')
                                            return
                                        try:
                                            obsah = None
                                            zdroj = getattr(e, 'content', None) or getattr(e, 'file', None) or getattr(e, 'stream', None)
                                            if not zdroj:
                                                for attr_name in dir(e):
                                                    attr_val = getattr(e, attr_name)
                                                    if hasattr(attr_val, 'read'):
                                                        zdroj = attr_val; break
                                            if zdroj and hasattr(zdroj, 'read'):
                                                cteni = zdroj.read()
                                                obsah = await cteni if asyncio.iscoroutine(cteni) else cteni
                                            if obsah is None:
                                                return ui.notify('CHYBA VERZE: Nelze načíst data souboru.', type='negative', position='top')
                                            with open(cilovy_soubor, 'wb') as f: f.write(obsah)
                                            ui.notify(f'Soubor {cilovy_soubor} byl úspěšně nahrán!', type='positive', position='top')
                                        except Exception as err:
                                            ui.notify(f'Chyba: {err}', type='negative', position='top')

                                    def _vykresli_okruhy_skupiny(typ, barva_nadpisu, nazev_skupiny):
                                        soubory_sk = SOUBORY_OKRUHU[typ]
                                        c_sk = config[typ]
                                        with ui.expansion(
                                            f'{nazev_skupiny} — okruhy otázek',
                                            icon='folder_open'
                                        ).classes(f'w-full border border-gray-200 rounded-xl mb-4 font-bold {barva_nadpisu}').props('default-opened'):
                                            for idx, (soubor, nazev_key) in enumerate([
                                                (soubory_sk[0], 'okruh_1_nazev'),
                                                (soubory_sk[1], 'okruh_2_nazev'),
                                                (soubory_sk[2], 'okruh_3_nazev'),
                                            ], 1):
                                                max_ok = get_max_questions(soubor)
                                                nazev_okruhu = c_sk.get(nazev_key, f'Okruh {idx}')
                                                stav = f'✅ Nahrán — {max_ok} otázek' if max_ok > 0 else '❌ Soubor nenáhrán'
                                                barva_stavu = 'text-green-600' if max_ok > 0 else 'text-red-500'
                                                with ui.card().classes('w-full p-5 mb-3 border border-gray-200 rounded-xl bg-gray-50'):
                                                    with ui.row().classes('w-full items-center justify-between mb-3'):
                                                        ui.label(f'Okruh {idx} — {nazev_okruhu}').classes('text-lg font-bold text-gray-800')
                                                        ui.label(stav).classes(f'text-sm font-semibold {barva_stavu}')
                                                    ui.label(f'Soubor: {soubor}').classes('text-xs text-gray-400 mb-2')
                                                    ui.upload(
                                                        on_upload=lambda e, s=soubor: _nahrat_do_souboru(e, s),
                                                        label=f'Vyberte CSV pro Okruh {idx}',
                                                        auto_upload=True
                                                    ).props('accept=".csv"').classes('max-w-md w-full')

                                    _vykresli_okruhy_skupiny('oz',  'text-blue-700',   '🧑‍💼 OZ')
                                    _vykresli_okruhy_skupiny('asm', 'text-indigo-700', '👔 ASM')

                            with ui.tab_panel(tab_soubory):
                                with ui.card().classes('w-full p-8 shadow-xl rounded-2xl border border-gray-100 bg-white/95 backdrop-blur-sm'):
                                    slozka = "Vysledky_Web"
                                    os.makedirs(slozka, exist_ok=True)
                                    vybrane_soubory = set()
                                    def checkbox_zmena(e, soubor): vybrane_soubory.add(soubor) if e.value else vybrane_soubory.discard(soubor)
                                    def smazat_vybrane():
                                        if not vybrane_soubory: return ui.notify('Nejprve zaškrtněte testy k vymazání.', type='warning', position='top')
                                        for f in vybrane_soubory:
                                            try: os.remove(os.path.join(slozka, f))
                                            except Exception: pass
                                        ui.notify('Vybrané testy byly smazány.', type='positive', position='top'); obnovit_seznam_souboru()
                                    def stahnout_vybrane():
                                        if not vybrane_soubory: return ui.notify('Nejprve zaškrtněte testy ke stažení.', type='warning', position='top')
                                        zip_path = os.path.join(slozka, "Vybrane_testy_export.zip")
                                        try:
                                            with zipfile.ZipFile(zip_path, 'w') as zf:
                                                for f in vybrane_soubory:
                                                    cesta = os.path.join(slozka, f)
                                                    if os.path.exists(cesta): zf.write(cesta, arcname=f)
                                            ui.download(zip_path); ui.notify('Stahování ZIP archivu začalo.', type='positive', position='top')
                                        except Exception as e: ui.notify(f'Chyba ZIP: {e}', type='negative', position='top')
                                    def vytisknout_web(cesta_k_souboru):
                                        try:
                                            df_s = pd.read_excel(cesta_k_souboru, sheet_name='Souhrn')
                                            df_a = pd.read_excel(cesta_k_souboru, sheet_name='Analýza')
                                            with ui.dialog() as print_dlg, ui.card().classes('w-full max-w-4xl bg-white text-black p-8'):
                                                ui.label('Zkušební protokol - Výsledek testu').classes('text-3xl font-bold mb-6 text-center')
                                                cols_s = [{'name': c, 'label': c, 'field': c, 'align': 'left'} for c in df_s.columns]
                                                ui.table(columns=cols_s, rows=df_s.to_dict('records')).classes('w-full mb-8 shadow-sm')
                                                ui.label('Analýza odpovědí').classes('text-2xl font-bold mb-4')
                                                cols_a = [{'name': c, 'label': c, 'field': c, 'align': 'left'} for c in df_a.columns]
                                                ui.table(columns=cols_a, rows=df_a.to_dict('records')).classes('w-full mb-8 shadow-sm')
                                                with ui.row().classes('w-full justify-between mt-4 skryt-pri-tisku'):
                                                    ui.button('Zavřít náhled', on_click=print_dlg.close).classes('bg-gray-500 hover:bg-gray-600 text-white')
                                                    ui.button('🖨️ Fyzicky Vytisknout', on_click=lambda: ui.run_javascript('window.print()')).classes('bg-orange-500 hover:bg-orange-600 text-white font-bold px-8')
                                            print_dlg.open()
                                        except Exception as e: ui.notify(f'Chyba při přípravě tisku: {e}', type='negative')

                                    with ui.row().classes('w-full items-center justify-between mb-6 pb-4 border-b border-gray-200'):
                                        ui.label('📁 Odevzdané testy').classes('text-2xl font-bold text-blue-600')
                                        with ui.row().classes('space-x-4'):
                                            ui.button('🔄 Obnovit', color='blue', on_click=lambda: obnovit_seznam_souboru()).classes('font-bold shadow-md')
                                            ui.button('🗑️ Smazat', color='red', on_click=smazat_vybrane).classes('font-bold shadow-md')
                                            ui.button('📦 Stáhnout (ZIP)', color='green', on_click=stahnout_vybrane).classes('font-bold shadow-md')

                                    kontejner_souboru = ui.column().classes('w-full space-y-2')
                                    def obnovit_seznam_souboru():
                                        vybrane_soubory.clear(); kontejner_souboru.clear()
                                        with kontejner_souboru:
                                            soubory = [f for f in os.listdir(slozka) if f.startswith('Vysledek_') and f.endswith('.xlsx')]
                                            if not soubory: ui.label('Zatím nebyly odevzdány žádné testy.').classes('text-gray-500 italic text-lg')
                                            else:
                                                for f in sorted(soubory, reverse=True)[:50]:
                                                    cesta_k_souboru = os.path.join(slozka, f)
                                                    with ui.row().classes('w-full items-center justify-between border-b border-gray-100 p-2 hover:bg-gray-50'):
                                                        with ui.row().classes('items-center space-x-4'):
                                                            ui.checkbox(on_change=lambda e, f_name=f: checkbox_zmena(e, f_name))
                                                            ui.label(f"📄 {f}").classes('text-lg text-gray-700 font-medium')
                                                        with ui.row().classes('space-x-2'):
                                                            ui.button('🖨️ Tisk', on_click=lambda c=cesta_k_souboru: vytisknout_web(c)).classes('bg-orange-500 text-white px-3 py-1 shadow-sm hover:bg-orange-600')
                                                            ui.button('📥 Excel', on_click=lambda c=cesta_k_souboru: ui.download(c)).classes('bg-blue-500 text-white px-3 py-1 shadow-sm hover:bg-blue-600')
                                    obnovit_seznam_souboru()

            # --- B) BĚŽÍCÍ TEST ---
            elif state['bezi_test']:
                with ui.column().classes('w-full max-w-4xl mt-12 skryt-pri-tisku'):
                    with ui.row().classes('w-full justify-between items-center mb-8 border-b-2 pb-4 border-gray-200'):
                        ui.label(f'Otázka {state["aktualni_index"] + 1} z {len(state["otazky"])}').classes('text-2xl text-gray-600 font-bold')
                        local_ui['cas_label'] = ui.label().classes('text-3xl font-bold tracking-widest')
                        def aktualizuj_zobrazeni_casu():
                            if local_ui['cas_label']:
                                s = int(state['zbyva_sekund'])
                                local_ui['cas_label'].set_text(f'⏱ {s // 60:02d}:{s % 60:02d}')
                                local_ui['cas_label'].classes(replace='text-3xl font-bold tracking-widest text-red-600' if s <= 600 else 'text-3xl font-bold tracking-widest text-blue-600')
                        aktualizuj_zobrazeni_casu()

                    aktualni_q = state['otazky'][state['aktualni_index']]
                    ui.label(aktualni_q['OTÁZKA']).classes('text-4xl font-extrabold mb-10 text-center text-gray-800 w-full skryt-pri-tisku')
                    moznosti = {'A': f"A)  {aktualni_q['A']}", 'B': f"B)  {aktualni_q['B']}", 'C': f"C)  {aktualni_q['C']}", 'D': f"D)  {aktualni_q['D']}"}
                    radio = ui.radio(moznosti).classes('text-2xl mb-12 space-y-6 w-full ml-12 text-gray-800 font-medium skryt-pri-tisku').bind_value(state, 'vybrana_odpoved')

                    def dalsi_otazka():
                        if not state['vybrana_odpoved']: return ui.notify('Vyberte odpověď!', type='warning', position='top')
                        spravna, v = str(aktualni_q['správná odpověď']).strip().upper(), state['vybrana_odpoved']
                        if v == spravna: state['body'] += 1; hodnoceni = "Správně"
                        else: hodnoceni = "Špatně"

                        state['historie_odpovedi'] = state['historie_odpovedi'] + [{"Pořadí v testu": state['aktualni_index'] + 1, "Otázka": aktualni_q['OTÁZKA'], "Tvoje volba": f"{v}: {aktualni_q.get(v, '')}", "Hodnocení": hodnoceni, "Správná odpověď": "-" if v == spravna else f"{spravna}: {aktualni_q.get(spravna, '')}"}]
                        state['vybrana_odpoved'], state['aktualni_index'] = None, state['aktualni_index'] + 1
                        if state['jmeno'] in aktivni_studenti: aktivni_studenti[state['jmeno']]['aktualni_index'] = state['aktualni_index']

                        if state['aktualni_index'] < len(state['otazky']): hlavni_rozhrani.refresh()
                        else: ukoncit_test()

                    text_tlacitka = 'Odevzdat test' if state['aktualni_index'] == len(state['otazky']) - 1 else 'Další otázka'
                    ui.button(text_tlacitka, on_click=dalsi_otazka).classes('w-[25rem] h-20 text-3xl font-bold bg-blue-500 hover:bg-blue-600 text-white rounded-xl shadow-lg mx-auto block skryt-pri-tisku')

            # --- C) VÝSLEDKY PO TESTU ---
            elif state['vysledek'] is not None:
                with ui.column().classes('w-full max-w-4xl items-center mt-12 skryt-pri-tisku'):
                    ui.label('ZPRACOVÁNÍ TESTU DOKONČENO').classes('text-4xl font-extrabold mb-8 text-gray-800 skryt-pri-tisku')
                    celkem = len(state['otazky'])
                    uspesnost = (state['body'] / celkem * 100) if celkem > 0 else 0
                    min_usp = int(config[kviz_typ]['min_procento_uspesnosti'])
                    barva = 'text-green-500' if uspesnost >= min_usp else 'text-red-600'
                    stav_text = "ÚSPĚŠNĚ SPLNĚNO" if uspesnost >= min_usp else "NESPLNĚNO"

                    ui.label(stav_text).classes(f'text-5xl font-black mb-4 {barva} skryt-pri-tisku')
                    ui.label(f'Úspěšnost: {uspesnost:.2f} % (Požadováno {min_usp} %)').classes(f'text-4xl font-bold mb-4 {barva} skryt-pri-tisku')
                    ui.label(f'Získané body: {state["body"]} z {celkem}').classes('text-3xl mb-12 text-gray-600 font-bold skryt-pri-tisku')

                    with ui.dialog().props('maximized transition-show="slide-up" transition-hide="slide-down"') as vysledky_dlg:
                        with ui.card().classes('w-full h-full max-w-full m-0 p-8 bg-white overflow-y-auto'):
                            ui.label('Zkušební protokol - Výsledek testu').classes('text-4xl font-bold mb-6 text-center text-gray-800')

                            df_s = [{"Jméno": state['jmeno'], "Stav": state.get('konec_stav', ''), "Úspěšnost": f"{state.get('konec_proc', 0):.2f}%", "Bodů": f"{state['body']}/{celkem}", "Trvání": f"{state.get('konec_m', 0)}m {state.get('konec_s', 0)}s"}]
                            cols_s = [{'name': c, 'label': c, 'field': c, 'align': 'left'} for c in df_s[0].keys()]
                            ui.table(columns=cols_s, rows=df_s).classes('w-full mb-8 shadow-sm text-lg')

                            ui.label('Analýza odpovědí').classes('text-3xl font-bold mb-4 text-gray-800')
                            if state['historie_odpovedi']:
                                cols_a = [{'name': c, 'label': c, 'field': c, 'align': 'left'} for c in state['historie_odpovedi'][0].keys()]
                                ui.table(columns=cols_a, rows=state['historie_odpovedi']).classes('w-full mb-8 shadow-sm text-base')

                            with ui.row().classes('w-full justify-center mt-8 skryt-pri-tisku gap-4'):
                                ui.button('Zavřít výsledky', on_click=vysledky_dlg.close).classes('bg-gray-500 hover:bg-gray-600 text-white px-8 h-14 text-lg font-bold shadow-md')
                                ui.button('🖨️ Vytisknout', on_click=lambda: ui.run_javascript('window.print()')).classes('bg-orange-500 hover:bg-orange-600 text-white font-bold px-8 h-14 text-lg shadow-md ml-4')

                    with ui.row().classes('gap-4 mt-8'):
                        ui.button('Zpět na Přehled', on_click=_nav_na_prehled).classes('bg-blue-600 hover:bg-blue-700 text-white font-bold h-16 px-8 text-xl shadow-md')

                        if ma_pristup_k_vysledkum:
                            ui.button('📊 Zobrazit mé výsledky', on_click=vysledky_dlg.open).classes('bg-green-600 hover:bg-green-700 text-white font-bold h-16 px-8 text-xl shadow-md')

            else:
                with ui.column().classes('w-full max-w-3xl items-center mt-20 skryt-pri-tisku'):
                    try: ui.image('logo.png').classes('w-64 mb-6 skryt-pri-tisku')
                    except Exception: pass

                    ui.label('Testovací systém JIP').classes('text-3xl font-bold text-blue-600 mb-8 text-center skryt-pri-tisku')

                    ui.label(f'Přihlášen(a): {state["jmeno"]}').classes('text-2xl font-bold mb-4 text-gray-700 skryt-pri-tisku')
                    skupina_label = '👔 ASM' if kviz_typ == 'asm' else '🧑‍💼 OZ'
                    ui.label(f'Skupina: {skupina_label}').classes('text-base font-semibold text-gray-500 mb-2 skryt-pri-tisku')
                    ui.label(f'Čeká vás {config[kviz_typ]["pocet_otazek"]} otázek. Na vypracování máte {config[kviz_typ]["casovy_limit_minuty"]} minut.').classes('text-lg text-gray-500 mb-8 skryt-pri-tisku')

                    with ui.row().classes('gap-4 items-center'):
                        ui.button('Zpět na Přehled', on_click=_nav_na_prehled).classes('h-20 text-xl font-bold bg-gray-500 hover:bg-gray-600 text-white shadow-md rounded-2xl skryt-pri-tisku')
                        ui.button('SPUSTIT TEST', on_click=spustit_test).classes('w-[25rem] h-20 text-3xl font-bold bg-blue-500 hover:bg-blue-600 text-white shadow-xl rounded-2xl skryt-pri-tisku')

                    if is_admin:
                        ui.button('⚙️ Správa Kvízu', on_click=lambda: prepni_admin_panel(True)).classes('mt-12 w-64 h-12 text-lg bg-gray-800 hover:bg-gray-700 text-white shadow-md rounded-2xl skryt-pri-tisku')

    def tik_casovace():
        if state.get('bezi_test') and state.get('zbyva_sekund', 0) > 0:
            state['zbyva_sekund'] -= 1
            if local_ui['cas_label']:
                s = int(state['zbyva_sekund'])
                local_ui['cas_label'].set_text(f'⏱ {s // 60:02d}:{s % 60:02d}')
                local_ui['cas_label'].classes(replace='text-3xl font-bold tracking-widest text-red-600' if s <= 600 else 'text-3xl font-bold tracking-widest text-blue-600')

            # Přidáno zaznamenání aktuálního času (last_seen), kdy se student naposledy ozval
            if state['jmeno'] not in aktivni_studenti:
                aktivni_studenti[state['jmeno']] = {'jmeno': state['jmeno'], 'aktualni_index': state['aktualni_index'], 'celkem_otazek': len(state['otazky']), 'zbyva_sekund': state['zbyva_sekund'], 'last_seen': time.time()}
            else:
                aktivni_studenti[state['jmeno']]['zbyva_sekund'] = state['zbyva_sekund']
                aktivni_studenti[state['jmeno']]['last_seen'] = time.time()

            if state['zbyva_sekund'] <= 0:
                ui.notify('Čas vypršel! Odevzdáno.', type='negative')
                ukoncit_test()

    ui.timer(1.0, tik_casovace)

    def spustit_test():
        otazky = nacti_otazky(typ=kviz_typ, pocet_otazek=int(config[kviz_typ]['pocet_otazek']))
        if not otazky: return
        state['otazky'], state['aktualni_index'], state['body'], state['zbyva_sekund'] = otazky, 0, 0, int(config[kviz_typ]['casovy_limit_minuty']) * 60
        state['vybrana_odpoved'], state['vysledek'], state['historie_odpovedi'], state['bezi_test'] = None, None, [], True

        aktivni_studenti[state['jmeno']] = {'jmeno': state['jmeno'], 'aktualni_index': 0, 'celkem_otazek': len(otazky), 'zbyva_sekund': state['zbyva_sekund'], 'last_seen': time.time()}

        hlavni_rozhrani.refresh()

    def ukoncit_test():
        state['bezi_test'] = False; state['vysledek'] = True
        if state['jmeno'] in aktivni_studenti: del aktivni_studenti[state['jmeno']]
        trvani = (int(config[kviz_typ]['casovy_limit_minuty']) * 60) - int(state['zbyva_sekund'])
        m, s = divmod(trvani, 60)
        celkem = len(state['otazky'])
        proc = (state['body'] / celkem) * 100 if celkem > 0 else 0
        stav_testu = "DOKONČENO" if int(state['zbyva_sekund']) > 0 else "ČAS VYPRŠEL"

        state['konec_m'] = m
        state['konec_s'] = s
        state['konec_proc'] = proc
        state['konec_stav'] = stav_testu

        dokoncene_testy.insert(0, {'jmeno': state['jmeno'], 'cas_odevzdani': time.strftime('%H:%M:%S'), 'vysledek': stav_testu, 'skore': f"{state['body']}/{celkem}", 'uspesnost': f"{proc:.2f}%"})
        slozka = "Vysledky_Web"
        os.makedirs(slozka, exist_ok=True)
        cesta = os.path.join(slozka, f"Vysledek_{state['jmeno']}_{time.strftime('%H-%M-%S')}.xlsx")
        try:
            df_s = pd.DataFrame([{"Jméno": state['jmeno'], "Stav": stav_testu, "Úspěšnost": f"{proc:.2f}%", "Bodů": f"{state['body']}/{celkem}", "Trvání": f"{m}m {s}s"}])
            with pd.ExcelWriter(cesta, engine='openpyxl') as w:
                df_s.to_excel(w, sheet_name='Souhrn', index=False)
                if state['historie_odpovedi']: pd.DataFrame(state['historie_odpovedi']).to_excel(w, sheet_name='Analýza', index=False)
                w.sheets['Souhrn']['B2'].fill = PatternFill(start_color="FF00FF00" if proc >= int(config[kviz_typ]['min_procento_uspesnosti']) else "00E02319", fill_type="solid")
        except Exception: pass

        if intranet_data.nacti_mysql().get("enabled"):
            intranet_data.uloz_vysledek_kvizu(
                app.storage.user.get('user_id'),
                stav_testu,
                f"{proc:.2f}%",
                f"{state['body']}/{celkem}",
                f"{m}m {s}s",
                state['historie_odpovedi']
            )

        hlavni_rozhrani.refresh()

    hlavni_rozhrani()


@ui.refreshable
def vykresli_vystup_kviz(user_name, vsechna_prava):
    if "vse" not in vsechna_prava and "vystup_vse" not in vsechna_prava and "vystup_osobni" not in vsechna_prava:
        ui.label('Přístup odepřen').classes('text-red-500 text-xl')
        return

    muze_mazat = "vse" in vsechna_prava or "vystup_vse" in vsechna_prava

    with ui.row().classes('w-full justify-between items-center mb-8'):
        ui.label('Výsledky znalostního testu').classes('text-4xl font-extrabold text-gray-800')

        async def export_kviz_excel():
            ui.notify('Generuji Excel export, malý moment...', type='info', icon='hourglass_empty')
            def _export():
                conn = intranet_data.get_db_connection()
                if not conn: raise ValueError('Nelze se připojit k databázi.')
                cursor = conn.cursor(dictionary=True)
                if "vse" in vsechna_prava or "vystup_vse" in vsechna_prava:
                    cursor.execute("SELECT u.name AS Jméno, u.surname AS Příjmení, v.stav_testu AS Stav, v.uspesnost AS Úspěšnost, v.body AS Body, v.doba_trvani AS Trvání, v.datum AS Odevzdáno FROM vysledky_kvizu v JOIN user u ON v.user_iduser = u.iduser ORDER BY v.datum DESC")
                else:
                    u_id = next((u['id'] for u in intranet_data.ziskej_vsechny_uzivatele().values() if f"{u['jmeno']} {u['prijmeni']}" == user_name), None)
                    cursor.execute("SELECT u.name AS Jméno, u.surname AS Příjmení, v.stav_testu AS Stav, v.uspesnost AS Úspěšnost, v.body AS Body, v.doba_trvani AS Trvání, v.datum AS Odevzdáno FROM vysledky_kvizu v JOIN user u ON v.user_iduser = u.iduser WHERE v.user_iduser = %s ORDER BY v.datum DESC", (u_id,))
                data = cursor.fetchall()
                cursor.close()
                conn.close()

                if not data: raise ValueError('Žádná data k exportu.')
                df = pd.DataFrame(data)
                if 'Odevzdáno' in df.columns: df['Odevzdáno'] = pd.to_datetime(df['Odevzdáno']).dt.tz_localize(None)
                os.makedirs("Exporty_Kviz", exist_ok=True)
                cesta = os.path.join("Exporty_Kviz", f"Export_Kviz_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")
                with pd.ExcelWriter(cesta, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Výsledky', index=False)
                return cesta

            try:
                cesta = await asyncio.to_thread(_export)
                ui.download(cesta)
                intranet_logger.log_activity(user_name, "Kvíz", "Stažen hromadný export výsledků do Excelu")
                ui.notify('Excel úspěšně stažen.', type='positive', position='top')
            except ValueError as ve: ui.notify(str(ve), type='warning', position='top')
            except Exception as e: ui.notify(f'Chyba při exportu: {e}', type='negative', position='top')

        ui.button('📥 Exportovat vše do Excelu', on_click=export_kviz_excel).classes('bg-green-600 hover:bg-green-700 text-white font-bold h-[3.5rem] px-6 shadow-md')

    conn = intranet_data.get_db_connection()
    if not conn:
        ui.label('Chyba databáze').classes('text-red-500 text-xl')
        return

    cursor = conn.cursor(dictionary=True)
    if "vse" in vsechna_prava or "vystup_vse" in vsechna_prava:
        cursor.execute("SELECT v.id, u.name, u.surname, v.stav_testu, v.uspesnost, v.body, v.doba_trvani, v.datum FROM vysledky_kvizu v JOIN user u ON v.user_iduser = u.iduser ORDER BY v.datum DESC")
    else:
        u_id = next((u['id'] for u in intranet_data.ziskej_vsechny_uzivatele().values() if f"{u['jmeno']} {u['prijmeni']}" == user_name), None)
        cursor.execute("SELECT v.id, u.name, u.surname, v.stav_testu, v.uspesnost, v.body, v.doba_trvani, v.datum FROM vysledky_kvizu v JOIN user u ON v.user_iduser = u.iduser WHERE v.user_iduser = %s ORDER BY v.datum DESC", (u_id,))

    vysledky = cursor.fetchall()
    cursor.close(); conn.close()

    if not vysledky:
        ui.label('Zatím nejsou k dispozici žádné výsledky.').classes('text-gray-500 italic text-xl')
        return

    with ui.card().classes('w-full p-8 shadow-md bg-white rounded-xl'):
        def ukaz_detail(vid):
            conn = intranet_data.get_db_connection()
            if not conn: return
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT * FROM zaznamy_odpovedi WHERE vysledek_id = %s ORDER BY poradi", (vid,))
            odpovedi = cursor.fetchall()
            cursor.close(); conn.close()

            with ui.dialog() as dlg, ui.card().classes('w-full max-w-4xl p-6 rounded-xl max-h-[80vh] overflow-y-auto'):
                ui.label('Detail odpovědí').classes('text-2xl font-bold mb-4')
                if not odpovedi: ui.label('Detail odpovědí není k dispozici.').classes('text-gray-500 italic')
                else:
                    for o in odpovedi:
                        barva = 'bg-green-50 border-green-200' if o['hodnoceni'] == 'Správně' else 'bg-red-50 border-red-200'
                        ikona = '✅' if o['hodnoceni'] == 'Správně' else '❌'
                        with ui.card().classes(f'w-full p-4 mb-2 shadow-sm border {barva}'):
                            ui.label(f"Otázka {o['poradi']}: {o['otazka']}").classes('font-bold text-lg mb-2')
                            ui.label(f"{ikona} Vaše volba: {o['tvoje_volba']}").classes('mb-1')
                            if o['hodnoceni'] != 'Správně': ui.label(f"👉 Správná odpověď: {o['spravna_odpoved']}").classes('font-bold text-green-700')
                ui.button('Zavřít', on_click=dlg.close).classes('mt-4 bg-gray-500 text-white font-bold w-full')
            dlg.open()

        def smazat_test(vid):
            with ui.dialog() as dlg, ui.card().classes('p-6 rounded-xl w-full max-w-sm'):
                ui.label('Smazání výsledku').classes('text-xl font-bold mb-4 text-red-600')
                ui.label('Opravdu chcete tento výsledek nenávratně smazat?').classes('mb-6 text-gray-700')
                def potvrdit():
                    conn = intranet_data.get_db_connection()
                    if conn:
                        try:
                            cursor = conn.cursor()
                            cursor.execute("DELETE FROM zaznamy_odpovedi WHERE vysledek_id = %s", (vid,))
                            cursor.execute("DELETE FROM vysledky_kvizu WHERE id = %s", (vid,))
                            conn.commit(); cursor.close(); conn.close()
                            intranet_logger.log_activity(user_name, "Kvíz", f"Smazán výsledek testu (ID: {vid})")
                            ui.notify('Výsledek smazán.', type='positive', position='top')
                            dlg.close(); vykresli_vystup_kviz.refresh()
                        except Exception as e: ui.notify(f'Chyba při mazání: {e}', type='negative', position='top')
                with ui.row().classes('w-full justify-between'):
                    ui.button('Zrušit', on_click=dlg.close).classes('bg-gray-400 text-white font-bold')
                    ui.button('Smazat', on_click=potvrdit).classes('bg-red-600 text-white font-bold shadow-md')
            dlg.open()

        for v in vysledky:
            barva = 'border-green-500 bg-green-50' if v['stav_testu'] == 'Úspěšný' else 'border-red-500 bg-red-50'
            with ui.row().classes(f'w-full justify-between items-center p-4 mb-4 border-l-8 shadow-sm rounded {barva}'):
                with ui.column().classes('gap-1'):
                    ui.label(f"{v['name']} {v['surname']}").classes('font-bold text-xl text-gray-800')
                    ui.label(f"Odevzdáno: {v['datum'].strftime('%d.%m.%Y %H:%M:%S')}").classes('text-sm text-gray-600')
                with ui.column().classes('items-end gap-1'):
                    ui.label(f"{v['uspesnost']} ({v['body']})").classes('font-black text-2xl')
                    ui.label(f"Trvání: {v['doba_trvani']}").classes('text-sm text-gray-600')

                with ui.row().classes('gap-2'):
                    ui.button('Detail chyb', icon='search', on_click=lambda vid=v['id']: ukaz_detail(vid)).classes('bg-blue-500 hover:bg-blue-600 text-white shadow-md')
                    async def stahni_excel(vid=v['id'], v_data=v):
                        ui.notify('Generuji Excel export...', type='info')
                        def _zpracuj_excel():
                            conn = intranet_data.get_db_connection()
                            cursor = conn.cursor(dictionary=True)
                            cursor.execute("SELECT poradi AS `Pořadí v testu`, otazka AS `Otázka`, tvoje_volba AS `Tvoje volba`, spravna_odpoved AS `Správná odpověď`, hodnoceni AS `Hodnocení` FROM zaznamy_odpovedi WHERE vysledek_id = %s ORDER BY poradi", (vid,))
                            odpovedi = cursor.fetchall()
                            cursor.close(); conn.close()
                            os.makedirs("Exporty_Kviz", exist_ok=True)
                            bezpecne_jmeno = f"{v_data['name']}_{v_data['surname']}".replace(" ", "_")
                            format_data = v_data['datum'].strftime('%Y%m%d_%H%M%S')
                            cesta = os.path.join("Exporty_Kviz", f"Vysledek_{bezpecne_jmeno}_{format_data}.xlsx")
                            df_s = pd.DataFrame([{"Jméno": f"{v_data['name']} {v_data['surname']}", "Stav": v_data['stav_testu'], "Úspěšnost": v_data['uspesnost'], "Bodů": v_data['body'], "Trvání": v_data['doba_trvani']}])
                            with pd.ExcelWriter(cesta, engine='openpyxl') as w:
                                df_s.to_excel(w, sheet_name='Souhrn', index=False)
                                if odpovedi: pd.DataFrame(odpovedi).to_excel(w, sheet_name='Analýza', index=False)
                                barva_fill = "FF00FF00" if v_data['stav_testu'] == 'Úspěšný' else "00E02319"
                                w.sheets['Souhrn']['B2'].fill = PatternFill(start_color=barva_fill, fill_type="solid")
                            return cesta
                        try:
                            cesta_k_souboru = await asyncio.to_thread(_zpracuj_excel)
                            ui.download(cesta_k_souboru)
                            intranet_logger.log_activity(user_name, "Kvíz", f"Stažen detailní export výsledku pro: {v_data['name']} {v_data['surname']}")
                        except Exception as e: ui.notify(f'Chyba při tvorbě Excelu: {e}', type='negative')
                    ui.button('Excel', icon='download', on_click=stahni_excel).classes('bg-green-600 hover:bg-green-700 text-white shadow-md')
                    if muze_mazat: ui.button('Smazat', icon='delete', color='red', on_click=lambda vid=v['id']: smazat_test(vid)).classes('shadow-md')