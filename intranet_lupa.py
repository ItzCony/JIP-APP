"""Lupou na obchod — analytika prodejů ASM.

Etapa 1: schéma, dávkový import GIST prodejů, párování ASM ↔ uživatel, dlaždice.

Přístup se NEŘÍDÍ per-ASM právy. Odvozuje se:
  ASM     → příjmení v datech se páruje s user.surname (bez diakritiky)
  Vedoucí → má právo hlavni_vedouci_{oddělení} → vidí všechna ASM v tom oddělení
  Extra   → lupa_extra    (vše + komentáře, záložka 3)
  Vkladatel → lupa_vkladatel (jen import)
  Admin   → lupa_admin / vse
"""

import asyncio
import datetime
import html
import io
import os
import re
import tempfile
import threading
import time
import unicodedata

from nicegui import ui, app

import intranet_data
import intranet_logger
import intranet_notifikace

# ============================================================================
# Konstanty
# ============================================================================

# Import běží paralelně. Kolidují jen dvě místa a ta se hlídají zvlášť:
#   1) partice (asm, rok, mesic) — dva importy téhož ASM by si rozbily DELETE+INSERT
#      → _zamek_asm(), jedno ASM naráz. Různá ASM sahají na disjunktní řádky.
#   2) dimenze (produkt, zákazník, dealer) — sdílené klíče napříč ASM, souběžné
#      ON DUPLICATE KEY UPDATE = deadlock → _DIMENZE_ZAMEK, zápis i commit uvnitř.
_IMPORT_SOUBEZNE = 3   # každý běh drží XLSX v paměti a jedno DB spojení
_IMPORT_SEM = asyncio.Semaphore(_IMPORT_SOUBEZNE)
_ASM_ZAMKY = {}
# Běžící importy napříč celým procesem: id(stav) -> stav. Drží je task na event
# loopu, ne klient — odchod ze stránky import nezastaví, jen přijde o UI.
_BEZICI_IMPORTY = {}
_DIMENZE_ZAMEK = threading.Lock()   # vlákna, ne event loop — drží ho importuj_soubor
_DAVKA = 2000        # řádků na jeden executemany
_COMMIT_PO = 50_000  # commit po N řádcích (aby neroste undo log)
_CACHE_TTL = 600.0   # s — mapy ASM↔uživatel↔oddělení

_DB_INIT_HOTOVO = False

# Sloupce zdrojového xlsx. Párujeme podle NÁZVU v hlavičce, ne podle pozice —
# export mezi verzemi mění pořadí i psaní ("K. jméno" → "K jméno",
# "Produkt - kód" → "Produkt - Kód") a nový sloupec "K ulice" přibyl doprostřed.
_SLOUPCE = {
    'asm':                'asm',
    'dealer':             'dealer',
    'dealer jmeno':       'dealer_jmeno',
    'mesic':              'mesic',
    'ico':                'ico',
    'jmeno':              'jmeno',
    'k jmeno':            'k_jmeno',
    'k mesto':            'k_mesto',
    'k ulice':            'k_ulice',
    'produkt kod':        'kod',
    'produkt nazev':      'nazev',
    'dodavatel popis':    'dodavatel',
    'obrat v mj':         'obrat_mj',
    'obrat v kc bez dph': 'obrat_kc',
}

_POVINNE = ('asm', 'mesic', 'ico', 'kod', 'obrat_mj', 'obrat_kc')

_MESIC_RE = re.compile(r'^\s*(\d{4})\s*M\s*(\d{1,2})\s*$', re.I)


# ============================================================================
# Pomocné
# ============================================================================

def _bez_diakritiky(text):
    t = unicodedata.normalize('NFD', str(text or ''))
    return ''.join(ch for ch in t if unicodedata.category(ch) != 'Mn')


def slug(text):
    """'Korábečný' → 'KORABECNY'. Klíč pro párování ASM ↔ příjmení."""
    return re.sub(r'[^A-Z0-9]+', '', _bez_diakritiky(text).upper())


def zaklad_asm(hodnota):
    """'BRAUN-VO' → 'BRAUN'. Sufixy -K / -VO / -ALG patří pod stejné ASM."""
    return slug(str(hodnota or '').split('-')[0])


def klic_ze_souboru(nazev_souboru):
    """'data prodeje zak3 – Weinberger.xlsx' → 'WEINBERGER'.

    Klíč bereme z názvu souboru, ne z dat: kód v datech je useknutý na 8 znaků,
    takže Weinberger má v jednom souboru 'WEINBERGER' i 'WEINBERG-K'. Název
    souboru je zároveň příjmení, na které se ASM páruje s uživatelem.
    """
    zaklad = re.split(r'[–—-]', str(nazev_souboru or '').rsplit('.', 1)[0])[-1]
    return slug(zaklad)


def _sedi_asm(zaklad, klic):
    """'WEINBERG' vs 'WEINBERGER' = totéž ASM (useknutý kód)."""
    return bool(zaklad) and (zaklad.startswith(klic) or klic.startswith(zaklad))


def _norm_hlavicka(text):
    """'Obrat v Kč bez DPH' → 'obrat v kc bez dph'."""
    t = _bez_diakritiky(text).lower()
    return re.sub(r'\s+', ' ', re.sub(r'[^a-z0-9]+', ' ', t)).strip()


def _cislo(hodnota):
    """Číslo z buňky. Excel dává float; textový export dává '1 234,50' / nbsp."""
    if hodnota is None or hodnota == '':
        return 0.0
    if isinstance(hodnota, (int, float)):
        return float(hodnota)
    t = str(hodnota).replace('\xa0', '').replace(' ', '').replace(',', '.')
    try:
        return float(t)
    except ValueError:
        return 0.0


def _obdobi(hodnota):
    """'2026M01' → (2026, 1). Nerozpoznané → (None, None)."""
    m = _MESIC_RE.match(str(hodnota or ''))
    if not m:
        return None, None
    mes = int(m.group(2))
    return (int(m.group(1)), mes) if 1 <= mes <= 12 else (None, None)


# ============================================================================
# Schéma
# ============================================================================

def inicializace_db():
    global _DB_INIT_HOTOVO
    if _DB_INIT_HOTOVO:
        return
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        # Fakta. ~7,1 M řádků za 20 souborů → texty jdou do dimenzí, tady jen klíče.
        cur.execute("""
            CREATE TABLE IF NOT EXISTS lupa_obrat (
                id       BIGINT AUTO_INCREMENT PRIMARY KEY,
                asm      VARCHAR(40) NOT NULL,
                dealer   VARCHAR(40) NOT NULL DEFAULT '',
                ico      VARCHAR(40) COLLATE utf8mb4_bin NOT NULL,
                kod      VARCHAR(40) NOT NULL DEFAULT '',
                rok      SMALLINT NOT NULL,
                mesic    TINYINT NOT NULL,
                obrat_mj DOUBLE DEFAULT 0,
                obrat_kc DOUBLE DEFAULT 0,
                INDEX idx_asm_obd (asm, rok, mesic),
                INDEX idx_ico_obd (ico, rok, mesic),
                INDEX idx_dealer (dealer, rok, mesic),
                INDEX idx_kod (kod)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci ENGINE=InnoDB
        """)
        # Dimenze. ico má utf8mb4_bin shodně s asm_kontakty_oz — jinak JOIN nesedne.
        cur.execute("""
            CREATE TABLE IF NOT EXISTS lupa_zakaznik (
                ico     VARCHAR(40) COLLATE utf8mb4_bin PRIMARY KEY,
                jmeno   VARCHAR(255),
                k_jmeno VARCHAR(255),
                k_ulice VARCHAR(255),
                k_mesto VARCHAR(255),
                asm     VARCHAR(40),
                INDEX idx_asm (asm)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci ENGINE=InnoDB
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS lupa_produkt (
                kod       VARCHAR(40) PRIMARY KEY,
                nazev     VARCHAR(255),
                dodavatel VARCHAR(255),
                INDEX idx_dod (dodavatel)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci ENGINE=InnoDB
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS lupa_dealer (
                dealer VARCHAR(40) PRIMARY KEY,
                jmeno  VARCHAR(255),
                asm    VARCHAR(40),
                INDEX idx_asm (asm)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci ENGINE=InnoDB
        """)
        # Souhrn per ASM — dlaždice se z něj kreslí bez dotazu do 7M tabulky.
        # user_id = spárovaný obchoďák (auto dle příjmení, admin může přepsat).
        cur.execute("""
            CREATE TABLE IF NOT EXISTS lupa_asm (
                asm           VARCHAR(40) PRIMARY KEY,
                user_id       INT DEFAULT NULL,
                radku         BIGINT DEFAULT 0,
                zakazniku     INT DEFAULT 0,
                obdobi_od     VARCHAR(7) DEFAULT '',
                obdobi_do     VARCHAR(7) DEFAULT '',
                obrat_kc      DOUBLE DEFAULT 0,
                aktualizovano DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_user (user_id)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci ENGINE=InnoDB
        """)
        # Poznámky ASM. Vážou se na ASM, ne na uživatele — když se obchoďák
        # vymění, poznámky k území zůstanou nástupci.
        cur.execute("""
            CREATE TABLE IF NOT EXISTS lupa_poznamka (
                id        INT AUTO_INCREMENT PRIMARY KEY,
                asm       VARCHAR(40) NOT NULL,
                user_id   INT,
                uzivatel  VARCHAR(200),
                text      TEXT,
                vytvoreno DATETIME DEFAULT CURRENT_TIMESTAMP,
                upraveno  DATETIME DEFAULT NULL,
                INDEX idx_asm (asm, id)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci ENGINE=InnoDB
        """)
        # Komentáře extra čtenářů k poznámkám. Mazání poznámky je uklidí ručně —
        # cizí klíč nedáváme, zbytek modulu ho taky nemá.
        cur.execute("""
            CREATE TABLE IF NOT EXISTS lupa_komentar (
                id          INT AUTO_INCREMENT PRIMARY KEY,
                poznamka_id INT NOT NULL,
                user_id     INT,
                uzivatel    VARCHAR(200),
                text        TEXT,
                vytvoreno   DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_poznamka (poznamka_id, id)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci ENGINE=InnoDB
        """)
        conn.commit()
        cur.close()
        _DB_INIT_HOTOVO = True
    except Exception as e:
        print(f'[lupa] inicializace_db: {e}')
    finally:
        if conn:
            conn.close()


_AKCE_NAZVY = {
    'nahled': 'Náhled odběratelů',
    'export_pdf': 'Export PDF — odběratelé',
    'export_xlsx': 'Export XLSX — odběratelé',
    'export_xlsx_detail': 'Export XLSX — detailní pohled',
    'obraty_nahled': 'Náhled obratů',
    'obraty_export_pdf': 'Export PDF — obraty',
    'obraty_export_xlsx': 'Export XLSX — obraty',
    'import': 'Import dat',
    'poznamka': 'Poznámka',
    'poznamka_edit': 'Úprava poznámky',
    'poznamka_smaz': 'Smazání poznámky',
    'komentar': 'Komentář',
    'komentar_smaz': 'Smazání komentáře',
}


def zapis_log(user_id, uzivatel, akce, asm='', parametry='', pocet_radku=0):
    """Audit výjezdů do centrální audit konzole (úroveň „Lupa").

    Soubory se neukládají, jen nastavení, se kterým se export vyjížděl.
    `user_id` konzole nezná — identifikuje jménem, drží se kvůli volajícím."""
    casti = [_AKCE_NAZVY.get(akce, akce)]
    if asm:
        casti.append(f'ASM {asm.capitalize()}')
    if pocet_radku:
        casti.append(f'{pocet_radku} řádků')
    if parametry:
        casti.append(str(parametry)[:2000])
    intranet_logger.log_activity(uzivatel, 'Lupa', ' · '.join(casti))


# ============================================================================
# Poznámky a komentáře
# ============================================================================

def _zapis(sql, par):
    """Jednorázový zápis. Vrací lastrowid, nebo None když DB neběží."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return None
    try:
        cur = conn.cursor()
        cur.execute(sql, par)
        conn.commit()
        nove = cur.lastrowid
        cur.close()
        return nove
    except Exception as e:
        print(f'[lupa] zapis: {e}')
        return None
    finally:
        conn.close()


def poznamky(asm):
    """Poznámky ASM od nejnovější, každá s komentáři od nejstaršího."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute('SELECT * FROM lupa_poznamka WHERE asm=%s ORDER BY id DESC', (asm,))
        radky = cur.fetchall()
        for p in radky:
            p['komentare'] = []
        if radky:
            dle_id = {p['id']: p for p in radky}
            cur.execute(
                'SELECT * FROM lupa_komentar WHERE poznamka_id IN ({}) ORDER BY id'
                .format(','.join(['%s'] * len(dle_id))), tuple(dle_id))
            for k in cur.fetchall():
                dle_id[k['poznamka_id']]['komentare'].append(k)
        cur.close()
        return radky
    except Exception as e:
        print(f'[lupa] poznamky: {e}')
        return []
    finally:
        conn.close()


def pridej_poznamku(asm, user_id, uzivatel, text):
    return _zapis('INSERT INTO lupa_poznamka (asm, user_id, uzivatel, text) '
                  'VALUES (%s,%s,%s,%s)', (asm, user_id, uzivatel, text[:60000]))


def uprav_poznamku(pid, text):
    _zapis('UPDATE lupa_poznamka SET text=%s, upraveno=NOW() WHERE id=%s',
           (text[:60000], pid))


def smaz_poznamku(pid):
    _zapis('DELETE FROM lupa_komentar WHERE poznamka_id=%s', (pid,))
    _zapis('DELETE FROM lupa_poznamka WHERE id=%s', (pid,))


def pridej_komentar(pid, user_id, uzivatel, text):
    return _zapis('INSERT INTO lupa_komentar (poznamka_id, user_id, uzivatel, text) '
                  'VALUES (%s,%s,%s,%s)', (pid, user_id, uzivatel, text[:60000]))


def smaz_komentar(kid):
    _zapis('DELETE FROM lupa_komentar WHERE id=%s', (kid,))


# ============================================================================
# Import
# ============================================================================

def _radky_xlsx(raw_bytes, stav=None):
    """Řádky prvního listu. read_only → drží konstantní paměť i u 900k řádků."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        if stav is not None:
            stav['celkem'] = ws.max_row   # z dimenze listu, None když ji soubor nemá
        for r in ws.iter_rows(values_only=True):
            yield r
    finally:
        wb.close()


def _najdi_hlavicku(radek):
    """Mapa {interní_název: index} pokud je řádek hlavička, jinak None."""
    if not radek:
        return None
    mapa = {}
    for i, bunka in enumerate(radek):
        klic = _SLOUPCE.get(_norm_hlavicka(bunka))
        if klic and klic not in mapa:
            mapa[klic] = i
    return mapa if all(k in mapa for k in _POVINNE) else None


def importuj_soubor(raw_bytes, nazev_souboru, stav=None):
    """Jeden soubor = jedno ASM. Vrací dict s výsledkem.

    Replace je částečný: maže se jen (asm, rok, měsíc), které soubor obsahuje.
    Dohrání nového měsíce tedy nesmaže historii, reimport téhož měsíce přepíše.

    `stav` je volitelný dict pro průběh — běžíme ve vlákně, UI ho čte timerem.
    Zápis do dictu je pod GIL atomický, jiná synchronizace netřeba.
    """
    stav = stav if stav is not None else {}
    stav.update(faze='otevírám soubor', hotovo=0, celkem=None)
    klic = klic_ze_souboru(nazev_souboru)
    vysledek = {'soubor': nazev_souboru, 'asm': klic, 'radku': 0, 'preskoceno': 0,
                'cizi': 0, 'obdobi': [], 'chyba': None}
    if not klic:
        vysledek['chyba'] = ('Z názvu souboru nejde určit ASM. Očekávám tvar '
                             '„data prodeje zak3 – Příjmení.xlsx".')
        return vysledek
    inicializace_db()
    conn = intranet_data.get_db_connection()
    if not conn:
        vysledek['chyba'] = 'Není spojení s databází.'
        return vysledek

    zakaznici, produkty, dealeri = {}, {}, {}
    smazane_partice = set()
    davka = []
    od_commitu = 0
    cur = None
    try:
        cur = conn.cursor()
        mapa = None
        stav['faze'] = 'načítám řádky'
        for i, radek in enumerate(_radky_xlsx(raw_bytes, stav), 1):
            stav['hotovo'] = i
            if mapa is None:
                mapa = _najdi_hlavicku(radek)
                continue

            def bunka(klic):
                i = mapa.get(klic)
                return radek[i] if i is not None and i < len(radek) else None

            asm_raw = bunka('asm')
            if not asm_raw:
                continue
            rok, mesic = _obdobi(bunka('mesic'))
            ico = str(bunka('ico') or '').strip()
            if rok is None or not ico:
                vysledek['preskoceno'] += 1   # součtové řádky, prázdné patičky
                continue

            if not _sedi_asm(zaklad_asm(asm_raw), klic):
                vysledek['cizi'] += 1   # cizí ASM v souboru — nemícháme
                continue

            asm = klic
            kod = str(bunka('kod') or '').strip()
            dealer = str(bunka('dealer') or '').strip()

            partice = (asm, rok, mesic)
            if partice not in smazane_partice:
                cur.execute('DELETE FROM lupa_obrat WHERE asm=%s AND rok=%s AND mesic=%s',
                            partice)
                conn.commit()
                smazane_partice.add(partice)

            zakaznici.setdefault(ico, (
                str(bunka('jmeno') or '')[:255], str(bunka('k_jmeno') or '')[:255],
                str(bunka('k_ulice') or '')[:255], str(bunka('k_mesto') or '')[:255], asm))
            if kod:
                produkty.setdefault(kod, (str(bunka('nazev') or '')[:255],
                                          str(bunka('dodavatel') or '')[:255]))
            if dealer:
                dealeri.setdefault(dealer, (str(bunka('dealer_jmeno') or '')[:255], asm))

            davka.append((asm, dealer[:40], ico[:40], kod[:40], rok, mesic,
                          _cislo(bunka('obrat_mj')), _cislo(bunka('obrat_kc'))))
            if len(davka) >= _DAVKA:
                _uloz_davku(cur, davka)
                vysledek['radku'] += len(davka)
                od_commitu += len(davka)
                davka = []
                if od_commitu >= _COMMIT_PO:
                    conn.commit()
                    od_commitu = 0

        if mapa is None:
            vysledek['chyba'] = ('Nenalezena hlavička. Očekávám sloupce ASM, Měsíc, '
                                 'IČO, Produkt - Kód, Obrat v MJ, Obrat v Kč bez DPH.')
            return vysledek
        if davka:
            _uloz_davku(cur, davka)
            vysledek['radku'] += len(davka)
        conn.commit()

        if not vysledek['radku']:
            vysledek['chyba'] = ('Soubor neobsahuje žádná data pro ASM '
                                 f'{klic} (přeskočeno {vysledek["cizi"]} cizích řádků).')
            return vysledek

        stav['faze'] = 'ukládám číselníky'
        _uloz_dimenze(conn, cur, zakaznici, produkty, dealeri)

        vysledek['obdobi'] = sorted({f'{r}M{m:02d}' for _, r, m in smazane_partice})
        stav['faze'] = 'přepočítávám souhrn'
        _prepocti_souhrn(cur, conn, klic)
        conn.commit()
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        vysledek['chyba'] = str(e)
        print(f'[lupa] import {nazev_souboru}: {e}')
    finally:
        if cur:
            try:
                cur.close()
            except Exception:
                pass
        conn.close()
    vymaz_cache()
    return vysledek


def _uloz_davku(cur, davka):
    cur.executemany(
        'INSERT INTO lupa_obrat (asm, dealer, ico, kod, rok, mesic, obrat_mj, obrat_kc) '
        'VALUES (%s,%s,%s,%s,%s,%s,%s,%s)', davka)


def _uloz_dimenze(conn, cur, zakaznici, produkty, dealeri):
    """Sdílené číselníky — katalog produktů má každé ASM stejný.

    Souběžné ON DUPLICATE KEY UPDATE nad týmiž klíči = deadlock, proto jeden
    zapisovatel naráz. Commit patří dovnitř, jinak držíme zámky řádků i po
    uvolnění mutexu. Sekce je krátká — desítky tisíc klíčů, ne miliony faktů.
    """
    with _DIMENZE_ZAMEK:
        if zakaznici:
            cur.executemany(
                'INSERT INTO lupa_zakaznik (ico, jmeno, k_jmeno, k_ulice, k_mesto, asm) '
                'VALUES (%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE '
                'jmeno=VALUES(jmeno), k_jmeno=VALUES(k_jmeno), k_ulice=VALUES(k_ulice), '
                'k_mesto=VALUES(k_mesto), asm=VALUES(asm)',
                [(k[:40],) + v for k, v in sorted(zakaznici.items())])
        if produkty:
            cur.executemany(
                'INSERT INTO lupa_produkt (kod, nazev, dodavatel) VALUES (%s,%s,%s) '
                'ON DUPLICATE KEY UPDATE nazev=VALUES(nazev), dodavatel=VALUES(dodavatel)',
                [(k[:40],) + v for k, v in sorted(produkty.items())])
        if dealeri:
            cur.executemany(
                'INSERT INTO lupa_dealer (dealer, jmeno, asm) VALUES (%s,%s,%s) '
                'ON DUPLICATE KEY UPDATE jmeno=VALUES(jmeno), asm=VALUES(asm)',
                [(k[:40],) + v for k, v in sorted(dealeri.items())])
        conn.commit()


def _prepocti_souhrn(cur, conn, asm):
    cur.execute(
        'SELECT COUNT(*), COUNT(DISTINCT ico), MIN(rok*100+mesic), MAX(rok*100+mesic), '
        'COALESCE(SUM(obrat_kc),0) FROM lupa_obrat WHERE asm=%s', (asm,))
    radku, zak, obd_od, obd_do, obrat = cur.fetchone()

    def _fmt(v):
        return f'{v // 100}M{v % 100:02d}' if v else ''

    cur.execute(
        'INSERT INTO lupa_asm (asm, user_id, radku, zakazniku, obdobi_od, obdobi_do, '
        'obrat_kc, aktualizovano) VALUES (%s,%s,%s,%s,%s,%s,%s,NOW()) '
        'ON DUPLICATE KEY UPDATE radku=VALUES(radku), zakazniku=VALUES(zakazniku), '
        'obdobi_od=VALUES(obdobi_od), obdobi_do=VALUES(obdobi_do), '
        'obrat_kc=VALUES(obrat_kc), aktualizovano=NOW(), '
        # user_id přepíšeme jen když ještě není — ruční volbu admina nepřebíjíme.
        'user_id=COALESCE(user_id, VALUES(user_id))',
        (asm, _uzivatel_dle_prijmeni(cur, asm), radku, zak,
         _fmt(obd_od), _fmt(obd_do), float(obrat or 0)))


def _uzivatel_dle_prijmeni(cur, asm):
    """ASM 'KORABECNY' → iduser Miroslava Korábečného. Nejednoznačné → None."""
    cur.execute('SELECT iduser, surname FROM user')
    nalezeni = [r[0] for r in cur.fetchall() if slug(r[1]) == asm]
    return nalezeni[0] if len(nalezeni) == 1 else None


# ============================================================================
# Práva a viditelnost
# ============================================================================

_CACHE = {'cas': 0.0, 'asm_user': {}, 'user_odd': {}, 'odd_users': {}}


def vymaz_cache():
    _CACHE['cas'] = 0.0


def _mapy():
    """{asm: user_id}, {user_id: {oddělení}}, {oddělení: {user_id}} — 10min cache."""
    if time.time() - _CACHE['cas'] < _CACHE_TTL and _CACHE['asm_user']:
        return _CACHE['asm_user'], _CACHE['user_odd'], _CACHE['odd_users']
    inicializace_db()
    asm_user, user_odd, odd_users = {}, {}, {}
    conn = intranet_data.get_db_connection()
    if conn:
        try:
            cur = conn.cursor()
            cur.execute('SELECT asm, user_id FROM lupa_asm')
            asm_user = {r[0]: r[1] for r in cur.fetchall()}
            cur.execute('SELECT dtu.user_iduser, LOWER(d.name) FROM department_To_user dtu '
                        'JOIN department d ON d.iddepartment = dtu.department_iddepartment')
            for uid, odd in cur.fetchall():
                user_odd.setdefault(uid, set()).add(odd)
                odd_users.setdefault(odd, set()).add(uid)
            cur.close()
        except Exception as e:
            print(f'[lupa] _mapy: {e}')
        finally:
            conn.close()
    _CACHE.update({'cas': time.time(), 'asm_user': asm_user,
                   'user_odd': user_odd, 'odd_users': odd_users})
    return asm_user, user_odd, odd_users


def je_admin(prava):
    return 'vse' in prava or 'lupa_admin' in prava


def je_extra(prava):
    return je_admin(prava) or 'lupa_extra' in prava


def je_vkladatel(prava):
    return je_admin(prava) or 'lupa_vkladatel' in prava


def pristupna_asm(user_id, prava):
    """Seznam ASM, které uživatel vidí. Admin/extra = všechna."""
    asm_user, _, odd_users = _mapy()
    if je_extra(prava):
        return sorted(asm_user)

    videna = {a for a, uid in asm_user.items() if uid and uid == user_id}

    # Vedoucí oddělení vidí všechna ASM svých lidí.
    vedeni = {p[len('hlavni_vedouci_'):] for p in prava if p.startswith('hlavni_vedouci_')}
    if vedeni:
        clenove = set()
        for odd in vedeni:
            clenove |= odd_users.get(odd, set())
        videna |= {a for a, uid in asm_user.items() if uid in clenove}
    return sorted(videna)


def ma_pristup(user_id, prava):
    """Brána modulu. Vkladatel se dovnitř dostane i bez jediného vlastního ASM."""
    if je_extra(prava) or je_vkladatel(prava):
        return True
    return bool(pristupna_asm(user_id, prava))


def je_moje_asm(user_id, asm, prava):
    """Kdo píše poznámky: spárovaný obchoďák daného ASM (a admin).

    Extra čtenář ani vedoucí sem nepíšou — ti komentují. Poznámky jsou hlas
    toho, kdo území drží."""
    if je_admin(prava):
        return True
    asm_user, _, _ = _mapy()
    return bool(user_id) and asm_user.get(asm) == user_id


def nesparovana_asm():
    asm_user, _, _ = _mapy()
    return sorted(a for a, uid in asm_user.items() if not uid)


def _adminska_id():
    """Uživatelé s právem lupa_admin / vse — cíl notifikací o nespárovaných ASM."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor()
        # Právo může viset na uživateli, pozici i oddělení — stejné tři kanály
        # jako ziskej_prava_uzivatele(). „vse" je v ostrých datech jen na oddělení.
        cur.execute("""
            SELECT utp.user_iduser FROM user_To_privileges utp
                JOIN privileges p ON p.idprivileges = utp.privileges_idprivileges
                WHERE p.name IN ('vse', 'lupa_admin')
            UNION
            SELECT utj.user_iduser FROM user_To_jobPosition utj
                JOIN jobPosition_To_privileges jtp
                    ON utj.jobPosition_idjobPosition = jtp.jobPosition_idjobPosition
                JOIN privileges p ON p.idprivileges = jtp.privileges_idprivileges
                WHERE p.name IN ('vse', 'lupa_admin')
            UNION
            SELECT dtu.user_iduser FROM department_To_user dtu
                JOIN department_To_privileges dtp
                    ON dtu.department_iddepartment = dtp.department_iddepartment
                JOIN privileges p ON p.idprivileges = dtp.privileges_idprivileges
                WHERE p.name IN ('vse', 'lupa_admin')
        """)
        return [r[0] for r in cur.fetchall()]
    finally:
        conn.close()


# ============================================================================
# UI
# ============================================================================

_KARTA = ('w-72 h-56 items-center justify-center shadow-xl hover:scale-105 '
          'transition-transform duration-300 cursor-pointer bg-white rounded-2xl '
          'border border-indigo-200')


def _kc(hodnota):
    return f'{hodnota:,.0f}'.replace(',', ' ') + ' Kč'


def vykresli_lupa(user_id, user_name, vsechna_prava):
    inicializace_db()
    stav_klic = f'lupa_pohled_{user_id}'

    @ui.refreshable
    def obsah():
        pohled = app.storage.user.get(stav_klic)
        moje = pristupna_asm(user_id, vsechna_prava)

        if pohled == 'import' and not je_vkladatel(vsechna_prava):
            pohled = None
        if pohled and pohled not in moje and pohled != 'import':
            pohled = None

        if pohled == 'import':
            _zpet_hlavicka('Import dat', obsah, stav_klic)
            _vykresli_import(user_id, user_name, vsechna_prava)
            return
        if pohled:
            _zpet_hlavicka(f'ASM {pohled.capitalize()}', obsah, stav_klic)
            _vykresli_detail(pohled, user_id, user_name, vsechna_prava)
            return

        ui.label('🔍 Lupou na obchod').classes('text-3xl font-bold text-gray-800 mb-1')
        ui.label('Obratová analytika ASM, obchodních zástupců a zákazníků.') \
            .classes('text-gray-500 mb-6')

        souhrny = _nacti_souhrny(moje)
        if not souhrny and not je_vkladatel(vsechna_prava):
            with ui.column().classes('items-center py-20 gap-3 w-full'):
                ui.icon('insights', size='5rem', color='grey-3')
                ui.label('Žádná data').classes('text-xl font-semibold text-gray-400')
                ui.label('Zatím vám není přiřazeno žádné ASM.').classes('text-gray-400')
            return

        with ui.row().classes('gap-6 flex-wrap items-stretch'):
            for s in souhrny:
                _dlazdice_asm(s, stav_klic, obsah)
            if je_vkladatel(vsechna_prava):
                _dlazdice_import(stav_klic, obsah)

    obsah()


def _zpet_hlavicka(titulek, refreshable, stav_klic):
    def zpet():
        app.storage.user[stav_klic] = None
        refreshable.refresh()

    with ui.row().classes('items-center gap-3 mb-4'):
        ui.button(icon='arrow_back', on_click=zpet).props('flat round') \
            .tooltip('Zpět na výběr ASM')
        ui.label(titulek).classes('text-2xl font-bold text-gray-800')


def _nacti_souhrny(asmy):
    if not asmy:
        return []
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        ph = ','.join(['%s'] * len(asmy))
        cur.execute(f'SELECT * FROM lupa_asm WHERE asm IN ({ph}) ORDER BY asm', tuple(asmy))
        return cur.fetchall()
    except Exception as e:
        print(f'[lupa] _nacti_souhrny: {e}')
        return []
    finally:
        conn.close()


def _dlazdice_asm(s, stav_klic, refreshable):
    def otevri():
        app.storage.user[stav_klic] = s['asm']
        refreshable.refresh()

    with ui.card().classes(_KARTA).on('click', otevri):
        ui.label('🔍').classes('text-5xl mb-2')
        ui.label(s['asm'].capitalize()) \
            .classes('text-xl font-bold text-gray-800 text-center w-full')


def _dlazdice_import(stav_klic, refreshable):
    def otevri():
        app.storage.user[stav_klic] = 'import'
        refreshable.refresh()

    with ui.card().classes(_KARTA.replace('border-indigo-200', 'border-amber-200')) \
            .on('click', otevri):
        ui.label('⬆️').classes('text-5xl mb-2')
        ui.label('Import dat').classes('text-xl font-bold text-gray-800')
        ui.label('GIST prodeje — dávka .xlsx').classes('text-sm text-gray-500 mt-1')


# ============================================================================
# Záložka „Lupni odběratele" — dotazy
# ============================================================================

_STROP_NAHLED = 5000     # řádků do aggridu (souhrn po zákazníkovi bývá do 2 tis.)
_STROP_PDF = 3000        # řádků do PDF — Chromium nad to renderuje minuty


def _dotaz(sql, params=(), slovnik=False):
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    cur = None
    try:
        cur = conn.cursor(dictionary=True) if slovnik else conn.cursor()
        cur.execute(sql, params)
        return cur.fetchall()
    except Exception as e:
        print(f'[lupa] dotaz: {e}')
        return []
    finally:
        if cur:
            cur.close()
        conn.close()


def _obd_int(txt):
    """'2026M08' → 202608. Neplatný tvar → None (= filtr se neuplatní)."""
    m = _MESIC_RE.match(str(txt or ''))
    return int(m.group(1)) * 100 + int(m.group(2)) if m else None


def _volby_obdobi(asm):
    """Měsíce, které ASM v datech opravdu má."""
    return [f'{r}M{m:02d}' for r, m in _dotaz(
        'SELECT DISTINCT rok, mesic FROM lupa_obrat WHERE asm=%s ORDER BY rok, mesic',
        (asm,))]


def _volby_zakazniku(asm):
    """{ičo: 'Jméno · Město · IČO'} pro ui.select. Bereme z dimenze, ne z faktů.
    IČO je v labelu schválně — with_input hledá v labelu, takže jde vyhledat i psaním IČ."""
    return {
        r[0]: ' · '.join([x for x in (r[1], r[2]) if x] + [r[0]])
        for r in _dotaz('SELECT ico, jmeno, k_mesto FROM lupa_zakaznik '
                        'WHERE asm=%s ORDER BY jmeno', (asm,))
    }


def _najdi_ica(text, zak_volby, vybrane=()):
    """Nalepený seznam IČ → (klíče do filtru, nespárované tokeny).
    IČ z GISTu mívají příponu (`05783704J`), proto bereme i shodu na začátku —
    ale až od 4 znaků, aby „12“ nevybralo půl ASM."""
    vybrane = list(vybrane)
    neznama = []
    for t in [t.upper() for t in re.split(r'[^0-9A-Za-z]+', text or '') if t]:
        shody = [k for k in zak_volby if k.upper() == t]
        if not shody and len(t) >= 4:
            shody = [k for k in zak_volby if k.upper().startswith(t)]
        if not shody:
            neznama.append(t)
        vybrane += [k for k in shody if k not in vybrane]
    return vybrane, neznama


def _pole_ica(sel_zak, zak_volby):
    """Hromadné vložení IČ do filtru zákazníků — nalepení seznamu z Excelu."""
    pole = ui.input(label='Vložit IČ') \
        .props('dense outlined clearable').style('min-width: 200px') \
        .tooltip('Nalep IČ oddělená čárkou, mezerou nebo novým řádkem a stiskni Enter.')

    def _pridej():
        if not (pole.value or '').strip():
            return
        vybrane, neznama = _najdi_ica(pole.value, zak_volby, sel_zak.value or [])
        sel_zak.value = vybrane
        pole.value = ''
        if neznama:
            ui.notify('V tomto ASM nenalezeno: ' + ', '.join(neznama[:10]),
                      type='warning', multi_line=True)
        else:
            ui.notify(f'Ve filtru je {len(vybrane)} zákazníků.', type='positive')

    pole.on('keydown.enter', _pridej)
    return pole


def _volby_dodavatelu():
    """Dodavatelé napříč všemi ASM. Per-ASM seznam by znamenal JOIN přes stovky
    tisíc řádků faktů při každém otevření záložky — 3 s za trochu kratší nabídku."""
    return [r[0] for r in _dotaz(
        'SELECT DISTINCT dodavatel FROM lupa_produkt '
        "WHERE dodavatel <> '' ORDER BY dodavatel") if r[0]]


def _filtr_sql(asm, filtr, produkt_uz_pripojen=False):
    """WHERE pro `lupa_obrat o`. Vrací (sql, params, je_potreba_join_produktu)."""
    kde, par = ['o.asm = %s'], [asm]
    od, do = _obd_int(filtr.get('od')), _obd_int(filtr.get('do'))
    if od:
        kde.append('o.rok*100+o.mesic >= %s')
        par.append(od)
    if do:
        kde.append('o.rok*100+o.mesic <= %s')
        par.append(do)
    ica = [i for i in (filtr.get('ico') or []) if i]
    if ica:
        kde.append('o.ico IN (%s)' % ','.join(['%s'] * len(ica)))
        par += ica
    dod = [d for d in (filtr.get('dodavatel') or []) if d]
    if dod:
        kde.append('p.dodavatel IN (%s)' % ','.join(['%s'] * len(dod)))
        par += dod
    return ' AND '.join(kde), par, bool(dod) and not produkt_uz_pripojen


def _souhrn_odberatelu(asm, filtr):
    """Řádek = zákazník za zvolené období. Dimenze (jméno, adresa) joinujeme až
    NA VÝSLEDEK agregace — GROUP BY přes VARCHAR(255) jména a adresy dělal
    filesort nad 600 tis. řádků a trval 15 s místo 2,4 s.
    S filtrem 'rozpad' je řádek zákazník × produkt (řádků násobně víc)."""
    kde, par, join_p = _filtr_sql(asm, filtr)
    jp = 'JOIN lupa_produkt p ON p.kod = o.kod' if join_p else ''
    if filtr.get('rozpad'):
        return _dotaz(f"""
            SELECT a.ico, z.jmeno, z.k_jmeno, z.k_ulice, z.k_mesto,
                   a.kod, p.nazev, p.dodavatel, a.mj, a.kc
            FROM (SELECT o.ico, o.kod,
                         SUM(o.obrat_mj) AS mj, SUM(o.obrat_kc) AS kc
                  FROM lupa_obrat o {jp}
                  WHERE {kde}
                  GROUP BY o.ico, o.kod) a
            LEFT JOIN lupa_zakaznik z ON z.ico = a.ico
            LEFT JOIN lupa_produkt p ON p.kod = a.kod
            ORDER BY a.kc DESC
        """, par, slovnik=True)
    return _dotaz(f"""
        SELECT a.ico, z.jmeno, z.k_jmeno, z.k_ulice, z.k_mesto,
               a.polozek, a.mj, a.kc
        FROM (SELECT o.ico, COUNT(DISTINCT o.kod) AS polozek,
                     SUM(o.obrat_mj) AS mj, SUM(o.obrat_kc) AS kc
              FROM lupa_obrat o {jp}
              WHERE {kde}
              GROUP BY o.ico) a
        LEFT JOIN lupa_zakaznik z ON z.ico = a.ico
        ORDER BY a.kc DESC
    """, par, slovnik=True)


def _produkty_zakaznika(asm, ico, filtr):
    """Rozklik řádku: co ten zákazník ve zvoleném období bral, po produktech."""
    f = dict(filtr, ico=[ico])
    kde, par, _ = _filtr_sql(asm, f, produkt_uz_pripojen=True)
    return _dotaz(f"""
        SELECT o.kod, p.nazev, p.dodavatel,
               SUM(o.obrat_mj) AS mj, SUM(o.obrat_kc) AS kc
        FROM lupa_obrat o
        LEFT JOIN lupa_produkt p ON p.kod = o.kod
        WHERE {kde}
        GROUP BY o.kod, p.nazev, p.dodavatel
        ORDER BY kc DESC
        LIMIT {_STROP_NAHLED}
    """, par, slovnik=True)


# ============================================================================
# Záložka "Lupni obraty" — časové řady, OZ, propady, TOP položky
# ============================================================================

_TOP_ZAK_GRAF = 10      # kolik zákazníků kreslit do grafu vývoje
_TOP_POLOZEK = 300      # TOP N produktů do tabulky/exportu
_PROPAD_OKNO = 3        # kolik předchozích měsíců tvoří srovnávací základ

_MESICE_KR = ['Led', 'Úno', 'Bře', 'Dub', 'Kvě', 'Čvn',
              'Čvc', 'Srp', 'Zář', 'Říj', 'Lis', 'Pro']

# Odstíny pro série grafů (roky, OZ, zákazníci) — dost odlišné i po tisku.
_GRAF_PALETA = [
    '#4338CA', '#EA580C', '#059669', '#7C3AED', '#0891B2', '#DC2626',
    '#65A30D', '#DB2777', '#0284C7', '#B45309', '#4F46E5', '#15803D',
]


def _mesic_txt(klic):
    """202408 → '08/2024'."""
    return f'{klic % 100:02d}/{klic // 100}'


def _rada_dealeru(asm, filtr):
    """Obrat po měsících × OZ. Jeden dotaz pokrývá i celkovou řadu ASM —
    součet přes OZ je totéž a šetří druhý průchod tabulkou."""
    kde, par, join_p = _filtr_sql(asm, filtr)
    jp = 'JOIN lupa_produkt p ON p.kod = o.kod' if join_p else ''
    return _dotaz(f"""
        SELECT o.rok, o.mesic, o.dealer,
               COALESCE(d.jmeno, o.dealer) AS dealer_jmeno,
               SUM(o.obrat_mj) AS mj, SUM(o.obrat_kc) AS kc
        FROM lupa_obrat o {jp}
        LEFT JOIN lupa_dealer d ON d.dealer = o.dealer
        WHERE {kde}
        GROUP BY o.rok, o.mesic, o.dealer, COALESCE(d.jmeno, o.dealer)
        ORDER BY o.rok, o.mesic
    """, par, slovnik=True)


def _rada_zakazniku(asm, filtr):
    """Obrat po měsících × zákazník. Řádů je ~počet zákazníků × počet měsíců
    (jednotky desítek tisíc) — vývoj v čase i propady dopočítáme v Pythonu."""
    kde, par, join_p = _filtr_sql(asm, filtr)
    jp = 'JOIN lupa_produkt p ON p.kod = o.kod' if join_p else ''
    return _dotaz(f"""
        SELECT o.rok, o.mesic, o.ico, z.jmeno, SUM(o.obrat_kc) AS kc
        FROM lupa_obrat o {jp}
        LEFT JOIN lupa_zakaznik z ON z.ico = o.ico
        WHERE {kde}
        GROUP BY o.rok, o.mesic, o.ico, z.jmeno
        ORDER BY o.rok, o.mesic
    """, par, slovnik=True)


def _top_produkty(asm, filtr, limit=_TOP_POLOZEK):
    """TOP odebrané položky za celé zvolené období."""
    kde, par, _ = _filtr_sql(asm, filtr, produkt_uz_pripojen=True)
    return _dotaz(f"""
        SELECT o.kod, p.nazev, p.dodavatel,
               COUNT(DISTINCT o.ico) AS zakazniku,
               SUM(o.obrat_mj) AS mj, SUM(o.obrat_kc) AS kc
        FROM lupa_obrat o
        LEFT JOIN lupa_produkt p ON p.kod = o.kod
        WHERE {kde}
        GROUP BY o.kod, p.nazev, p.dodavatel
        ORDER BY kc DESC
        LIMIT {int(limit)}
    """, par, slovnik=True)


def _propady(zak, mesice, prah_pct, prah_kc, okno=_PROPAD_OKNO):
    """Poslední měsíc v datech proti průměru předchozích `okno` měsíců.

    Zákazník bez obratu v posledním měsíci = pokles 100 % ("ztracený") — právě
    ti se v běžném součtovém pohledu ztratí, proto je hlídáme zvlášť."""
    if len(mesice) < 2:
        return []
    posl = mesice[-1]
    predch = mesice[-1 - okno:-1]
    ven = []
    for ico, z in zak.items():
        zaklad = sum(z['kc'].get(m, 0.0) for m in predch) / len(predch)
        if zaklad <= 0:
            continue
        akt = z['kc'].get(posl, 0.0)
        rozdil = akt - zaklad
        pct = rozdil / zaklad * 100.0
        if -pct < prah_pct or -rozdil < prah_kc:
            continue
        ven.append({'ico': ico, 'jmeno': z['jmeno'] or ico, 'zaklad': zaklad,
                    'akt': akt, 'rozdil': rozdil, 'pct': pct,
                    'stav': 'ztracený' if akt <= 0 else 'propad'})
    # Při stejné korunové ztrátě je horší ten, kdo propadl relativně víc —
    # zákazník na nule je jiný problém než zákazník, co jen ubral.
    ven.sort(key=lambda r: (r['rozdil'], r['pct']))
    return ven


def _obraty_prehled(asm, filtr, prah_pct, prah_kc):
    """Vše pro záložku obratů: 3 dotazy, zbytek dopočet nad načtenými řádky."""
    dealeri = _rada_dealeru(asm, filtr)
    zakaznici = _rada_zakazniku(asm, filtr)
    produkty = _top_produkty(asm, filtr)

    mes_kc, mes_mj = {}, {}
    oz = {}
    for r in dealeri:
        k = r['rok'] * 100 + r['mesic']
        kc, mj = float(r['kc'] or 0), float(r['mj'] or 0)
        mes_kc[k] = mes_kc.get(k, 0.0) + kc
        mes_mj[k] = mes_mj.get(k, 0.0) + mj
        d = oz.setdefault(r['dealer'], {'dealer': r['dealer'],
                                        'dealer_jmeno': r['dealer_jmeno'],
                                        'kc': 0.0, 'mj': 0.0, 'rada': {}})
        d['kc'] += kc
        d['mj'] += mj
        d['rada'][k] = d['rada'].get(k, 0.0) + kc

    zak, mes_zak = {}, {}
    for r in zakaznici:
        k = r['rok'] * 100 + r['mesic']
        kc = float(r['kc'] or 0)
        z = zak.setdefault(r['ico'], {'jmeno': r['jmeno'], 'celkem': 0.0, 'kc': {}})
        z['celkem'] += kc
        z['kc'][k] = z['kc'].get(k, 0.0) + kc
        mes_zak.setdefault(k, set()).add(r['ico'])

    mesice = sorted(mes_kc)
    celkem_kc = sum(mes_kc.values())
    radky_mesice = [{'klic': k, 'mesic': _mesic_txt(k), 'kc': mes_kc[k],
                     'mj': mes_mj.get(k, 0.0),
                     'zakazniku': len(mes_zak.get(k, ()))} for k in mesice]
    oz_radky = sorted(oz.values(), key=lambda d: d['kc'], reverse=True)
    for d in oz_radky:
        d['podil'] = d['kc'] / celkem_kc * 100.0 if celkem_kc else 0.0

    top_zak = sorted(zak.items(), key=lambda p: p[1]['celkem'], reverse=True)
    top_zak = top_zak[:_TOP_ZAK_GRAF]

    posl = mesice[-1] if mesice else None
    predch = mesice[-2] if len(mesice) > 1 else None
    lonske = (posl - 100) if posl else None
    return {
        'mesice': mesice,
        'radky_mesice': radky_mesice,
        'oz': oz_radky,
        'zak_serie': [{'name': (z['jmeno'] or ico), 'ico': ico,
                       'data': [round(z['kc'].get(m, 0.0), 2) for m in mesice]}
                      for ico, z in top_zak],
        'propady': _propady(zak, mesice, prah_pct, prah_kc),
        'produkty': produkty,
        'kpi': {
            'celkem_kc': celkem_kc,
            'celkem_mj': sum(mes_mj.values()),
            'zakazniku': len(zak),
            'oz': len(oz),
            'posl_mesic': _mesic_txt(posl) if posl else '—',
            'posl_kc': mes_kc.get(posl, 0.0) if posl else 0.0,
            'mom': _zmena_pct(mes_kc.get(posl), mes_kc.get(predch)),
            'yoy': _zmena_pct(mes_kc.get(posl), mes_kc.get(lonske)),
        },
    }


def _zmena_pct(nove, stare):
    if not stare or nove is None:
        return None
    return (nove - stare) / stare * 100.0


_DETAIL_SQL = """
    SELECT o.asm, o.dealer, d.jmeno AS dealer_jmeno,
           CONCAT(o.rok, 'M', LPAD(o.mesic, 2, '0')) AS mesic,
           o.ico, z.jmeno, z.k_jmeno, z.k_ulice, z.k_mesto,
           o.kod, p.nazev, p.dodavatel, o.obrat_mj, o.obrat_kc
    FROM lupa_obrat o
    LEFT JOIN lupa_dealer d ON d.dealer = o.dealer
    LEFT JOIN lupa_zakaznik z ON z.ico = o.ico
    LEFT JOIN lupa_produkt p ON p.kod = o.kod
    WHERE {kde}
    ORDER BY o.ico, o.rok, o.mesic, o.kod
"""


def _souhrn_total(rows, rozpad=False):
    return {
        'jmeno': (f'CELKEM {len(rows)} řádků' if rozpad
                  else f'CELKEM {len(rows)} zákazníků'),
        'polozek': sum(int(r.get('polozek') or 0) for r in rows),
        'mj': sum(float(r['mj'] or 0) for r in rows),
        'kc': sum(float(r['kc'] or 0) for r in rows),
    }


# ============================================================================
# Exporty (XLSX streamem, PDF přes Chromium)
# ============================================================================

_COLS_SOUHRN = [
    ('IČO', 'ico', 'text', 16),
    ('Zákazník', 'jmeno', 'text', 38),
    ('K. jméno', 'k_jmeno', 'text', 30),
    ('K. ulice', 'k_ulice', 'text', 28),
    ('K. město', 'k_mesto', 'text', 20),
    ('Položek', 'polozek', 'int', 10),
    ('Obrat v MJ', 'mj', 'num', 14),
    ('Obrat v Kč bez DPH', 'kc', 'money', 20),
]

_COLS_SOUHRN_ROZPAD = [
    ('IČO', 'ico', 'text', 16),
    ('Zákazník', 'jmeno', 'text', 38),
    ('K. jméno', 'k_jmeno', 'text', 30),
    ('K. ulice', 'k_ulice', 'text', 28),
    ('K. město', 'k_mesto', 'text', 20),
    ('Produkt - Kód', 'kod', 'text', 16),
    ('Produkt - název', 'nazev', 'text', 40),
    ('Dodavatel - popis', 'dodavatel', 'text', 30),
    ('Obrat v MJ', 'mj', 'num', 14),
    ('Obrat v Kč bez DPH', 'kc', 'money', 20),
]

_COLS_DETAIL = [
    ('ASM', 'asm', 'text', 14),
    ('Dealer', 'dealer', 'text', 12),
    ('Dealer jméno', 'dealer_jmeno', 'text', 30),
    ('Měsíc', 'mesic', 'text', 11),
    ('IČO', 'ico', 'text', 16),
    ('Jméno', 'jmeno', 'text', 38),
    ('K. jméno', 'k_jmeno', 'text', 30),
    ('K. ulice', 'k_ulice', 'text', 28),
    ('K. město', 'k_mesto', 'text', 20),
    ('Produkt - Kód', 'kod', 'text', 16),
    ('Produkt - název', 'nazev', 'text', 40),
    ('Dodavatel - popis', 'dodavatel', 'text', 30),
    ('Obrat v MJ', 'obrat_mj', 'num', 14),
    ('Obrat v Kč bez DPH', 'obrat_kc', 'money', 20),
]

_COLS_MESICE = [
    ('Měsíc', 'mesic', 'text', 12),
    ('Obrat v MJ', 'mj', 'num', 14),
    ('Obrat v Kč bez DPH', 'kc', 'money', 20),
    ('Zákazníků', 'zakazniku', 'int', 12),
]

_COLS_OZ = [
    ('OZ', 'dealer', 'text', 12),
    ('Jméno', 'dealer_jmeno', 'text', 32),
    ('Obrat v MJ', 'mj', 'num', 14),
    ('Obrat v Kč bez DPH', 'kc', 'money', 20),
    ('Podíl %', 'podil', 'num', 10),
]

_COLS_PROPADY = [
    ('IČO', 'ico', 'text', 16),
    ('Zákazník', 'jmeno', 'text', 38),
    ('Průměr předchozích Kč', 'zaklad', 'money', 22),
    ('Poslední měsíc Kč', 'akt', 'money', 20),
    ('Rozdíl Kč', 'rozdil', 'money', 18),
    ('Změna %', 'pct', 'num', 12),
    ('Stav', 'stav', 'text', 12),
]

_COLS_POLOZKY = [
    ('Kód', 'kod', 'text', 16),
    ('Produkt', 'nazev', 'text', 40),
    ('Dodavatel', 'dodavatel', 'text', 30),
    ('Zákazníků', 'zakazniku', 'int', 12),
    ('Obrat v MJ', 'mj', 'num', 14),
    ('Obrat v Kč bez DPH', 'kc', 'money', 20),
]

_XLSX_FMT = {'text': '@', 'int': '#,##0', 'num': '#,##0.###', 'money': '#,##0.00'}

_TMP_DIR = os.path.join(tempfile.gettempdir(), 'jip_lupa')


def _xlsx_hodnota(v, typ):
    if v is None or v == '':
        return None
    if typ in ('int', 'num', 'money'):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None
    return str(v)


def _xlsx_list(wb, sheet, cols, radky, total=None):
    """Zapíše jeden list write_only sešitu; `radky` smí být iterátor."""
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet((sheet or 'Data')[:31])
    for i, (_n, _f, _t, sirka) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = sirka

    hlavicka = []
    for nadpis, _f, _t, _w in cols:
        c = WriteOnlyCell(ws, value=nadpis)
        c.font = Font(bold=True, color='FFFFFF', size=11)
        c.fill = PatternFill('solid', fgColor='3730A3')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        hlavicka.append(c)
    ws.append(hlavicka)
    ws.freeze_panes = 'A2'

    pocet = 0
    for r in radky:
        bunky = []
        for _n, field, typ, _w in cols:
            c = WriteOnlyCell(ws, value=_xlsx_hodnota(r.get(field), typ))
            c.number_format = _XLSX_FMT.get(typ, 'General')
            bunky.append(c)
        ws.append(bunky)
        pocet += 1

    ws.auto_filter.ref = f'A1:{get_column_letter(len(cols))}{pocet + 1}'

    if total:
        # Prázdný řádek a až pod ním součet — mimo rozsah autofiltru, aby ho
        # filtrování v Excelu neschovalo a nepočítalo mezi data.
        ws.append([])
        radek = []
        for _n, field, typ, _w in cols:
            c = WriteOnlyCell(ws, value=_xlsx_hodnota(total.get(field), typ))
            c.number_format = _XLSX_FMT.get(typ, 'General')
            c.font = Font(bold=True)
            radek.append(c)
        ws.append(radek)

    return pocet


def _xlsx_na_disk(cesta, cols, radky, total=None, sheet='Data'):
    """Streamovaný zápis .xlsx (openpyxl write_only) rovnou na disk.

    Proč ne běžný Workbook jako v modulu Sankce: detailní výjezd jednoho ASM je
    až 600 tis. řádků × 14 sloupců. Klasický sešit drží každou buňku jako objekt
    v paměti (~3 GB) — write_only zapisuje průběžně a paměť zůstává konstantní.
    `radky` je proto iterátor, ne list."""
    from openpyxl import Workbook
    wb = Workbook(write_only=True)
    pocet = _xlsx_list(wb, sheet, cols, radky, total)
    wb.save(cesta)
    return pocet


def _xlsx_listy_na_disk(cesta, listy):
    """Jeden sešit, víc listů: listy = [(nazev, cols, radky, total)]."""
    from openpyxl import Workbook
    wb = Workbook(write_only=True)
    for nazev, cols, radky, total in listy:
        _xlsx_list(wb, nazev, cols, radky, total)
    wb.save(cesta)


def _detail_iter(conn, asm, filtr, davka=2000):
    """Nebufferovaný kurzor → generátor dictů. Řádky tečou z DB rovnou do sešitu."""
    kde, par, _ = _filtr_sql(asm, filtr, produkt_uz_pripojen=True)
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(_DETAIL_SQL.format(kde=kde), par)
        while True:
            radky = cur.fetchmany(davka)
            if not radky:
                return
            for r in radky:
                yield r
    finally:
        cur.close()


def _export_detail_na_disk(asm, filtr, cesta):
    conn = intranet_data.get_db_connection()
    if not conn:
        raise RuntimeError('Není spojení s databází.')
    try:
        return _xlsx_na_disk(cesta, _COLS_DETAIL, _detail_iter(conn, asm, filtr),
                             sheet='Detail')
    finally:
        conn.close()


def _uklid_tmp(max_age_s=1800):
    """Smaže dočasné exporty starší než ~30 min (obsahují obchodní data)."""
    try:
        ted = time.time()
        for f in os.listdir(_TMP_DIR):
            cesta = os.path.join(_TMP_DIR, f)
            try:
                if os.path.isfile(cesta) and ted - os.path.getmtime(cesta) > max_age_s:
                    os.remove(cesta)
            except OSError:
                pass
    except FileNotFoundError:
        pass


def _tmp_cesta(jmeno):
    os.makedirs(_TMP_DIR, exist_ok=True)
    _uklid_tmp()
    return os.path.join(_TMP_DIR, f'{int(time.time() * 1000)}_{jmeno}')


def _stahni_soubor(cesta, jmeno):
    """Přes HTTP, ne přes WebSocket — ui.download.content posílá bajty socketem
    a nad ~1 MB padá na engine.io limitu. Takhle jde klientovi jen URL."""
    ui.download.file(cesta, jmeno)

    async def _pozdeji():
        await asyncio.sleep(300)
        try:
            os.remove(cesta)
        except OSError:
            pass

    try:
        asyncio.create_task(_pozdeji())
    except RuntimeError:
        pass   # mimo event loop — uklidí to _uklid_tmp při příštím exportu


def _stahni_bajty(data, jmeno):
    cesta = _tmp_cesta(jmeno)
    with open(cesta, 'wb') as fh:
        fh.write(data)
    _stahni_soubor(cesta, jmeno)


def _bezpecne_jmeno(s):
    s = unicodedata.normalize('NFKD', str(s or ''))
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return re.sub(r'[^A-Za-z0-9_\- ]', '', s).strip().replace(' ', '_') or 'lupa'


def _cz(hodnota, desetin=2):
    """Číslo v českém tvaru: 1 234 567,89 (mezera = NBSP, aby se nezlomilo)."""
    try:
        s = f'{float(hodnota or 0):,.{desetin}f}'
    except (TypeError, ValueError):
        return ''
    return s.replace(',', ' ').replace('.', ',')


_PDF_CSS = """
<style>
  .lupa-doc { font-family: Arial, sans-serif; color:#111; width:190mm; }
  .lupa-doc .hdr { display:flex; justify-content:space-between; align-items:flex-start;
        border-bottom:2px solid #3730A3; padding-bottom:8px; margin-bottom:10px; }
  .lupa-doc h1 { font-size:18px; margin:0 0 2px; color:#312e81; }
  .lupa-doc h2 { font-size:13px; margin:14px 0 5px; color:#312e81;
        page-break-after:avoid; }
  .lupa-doc .firma { font-size:11px; color:#374151; }
  .lupa-doc .meta { font-size:11px; color:#374151; line-height:1.55; text-align:right; }
  .lupa-doc .meta b { color:#111; }
  .lupa-doc .filtr { font-size:11px; color:#374151; margin-bottom:8px; }
  .lupa-doc table.t { width:100%; border-collapse:collapse; font-size:10px;
        page-break-inside:auto; }
  .lupa-doc table.t thead { display:table-header-group; }
  .lupa-doc table.t tr { page-break-inside:avoid; break-inside:avoid; }
  .lupa-doc table.t th { background:#e0e7ff; color:#312e81; border:1px solid #c7d2fe;
        padding:5px 6px; text-align:left; }
  .lupa-doc table.t td { border:1px solid #e5e7eb; padding:4px 6px; vertical-align:top; }
  .lupa-doc table.t td.r { text-align:right; white-space:nowrap; }
  .lupa-doc table.t tr:nth-child(even) td { background:#fafafa; }
  .lupa-doc table.t tr.celkem td { background:#eef2ff; font-weight:bold;
        border-top:2px solid #3730A3; }
  .lupa-doc .pozn { margin-top:14px; font-size:9px; color:#6b7280; }
</style>
"""


def _popis_filtru(filtr, zakaznici_volby):
    casti = []
    od, do = filtr.get('od'), filtr.get('do')
    casti.append(f'Období: {od or "od začátku"} – {do or "do konce"}')
    ica = filtr.get('ico') or []
    if ica:
        jmena = [zakaznici_volby.get(i, i) for i in ica[:5]]
        casti.append('Zákazníci: ' + ', '.join(jmena)
                     + (f' (+{len(ica) - 5} dalších)' if len(ica) > 5 else ''))
    dod = filtr.get('dodavatel') or []
    if dod:
        casti.append('Dodavatelé: ' + ', '.join(dod[:5])
                     + (f' (+{len(dod) - 5} dalších)' if len(dod) > 5 else ''))
    if filtr.get('rozpad'):
        casti.append('Rozpad po produktech')
    return ' · '.join(casti)


def _pdf_html(asm, rows, total, popis_filtru, uzivatel, rozpad=False):
    e = html.escape
    cols = _COLS_SOUHRN_ROZPAD if rozpad else _COLS_SOUHRN
    des = {'int': 0, 'num': 1, 'money': 2}
    hlavicky = ''.join(f'<th>{e(n)}</th>' for n, _f, _t, _w in cols)
    textovych = sum(1 for _n, _f, t, _w in cols if t == 'text')

    def _bunka(r, field, typ):
        if typ in des:
            return f'<td class="r">{_cz(r.get(field), des[typ])}</td>'
        return f'<td>{e(str(r.get(field) or ""))}</td>'

    telo = ['<tr>' + ''.join(_bunka(r, f, t) for _n, f, t, _w in cols) + '</tr>'
            for r in rows]
    celkem = (f'<td colspan="{textovych}">'
              f'{e(str(total.get("jmeno") or "CELKEM"))}</td>'
              + ''.join(_bunka(total, f, t) for _n, f, t, _w in cols
                        if t != 'text'))

    return (
        '<!DOCTYPE html><html lang="cs"><head><meta charset="utf-8">'
        f'<title>Lupou na obchod — {e(asm)}</title>{_PDF_CSS}</head><body>'
        '<div class="lupa-doc">'
        '<div class="hdr"><div>'
        f'<h1>Lupou na obchod — odběratelé</h1>'
        f'<div class="firma">JIP východočeská, a.s. · ASM {e(asm.capitalize())}</div>'
        '</div><div class="meta">'
        f'<b>Vytiskl:</b> {e(uzivatel)}<br>'
        f'<b>Datum:</b> {datetime.datetime.now():%d.%m.%Y %H:%M}<br>'
        f'<b>{"Řádků" if rozpad else "Zákazníků"}:</b> {len(rows)}'
        '</div></div>'
        f'<div class="filtr">{e(popis_filtru)}</div>'
        '<table class="t"><thead><tr>' + hlavicky + '</tr></thead>'
        '<tbody>' + ''.join(telo)
        + '<tr class="celkem">' + celkem
        + '</tr></tbody></table>'
        '<div class="pozn">Obraty bez DPH. Sestava vychází z dat GIST nahraných '
        'do modulu Lupou na obchod; zdrojem pravdy zůstává GIST.</div>'
        '</div></body></html>'
    )


def _pdf_tabulka(nadpis, cols, rows, limit=60):
    """Obecná tabulka do PDF podle `cols` (nadpis, field, typ, šířka)."""
    e = html.escape
    if not rows:
        return f'<h2>{e(nadpis)}</h2><div class="filtr">Žádná data.</div>'
    des = {'int': 0, 'num': 1, 'money': 2}
    hlav = ''.join(f'<th>{e(n)}</th>' for n, _f, _t, _w in cols)
    telo = []
    for r in rows[:limit]:
        bunky = []
        for _n, field, typ, _w in cols:
            v = r.get(field)
            if typ in des:
                bunky.append(f'<td class="r">{_cz(v, des[typ])}</td>')
            else:
                bunky.append(f'<td>{e(str(v or ""))}</td>')
        telo.append('<tr>' + ''.join(bunky) + '</tr>')
    vic = (f'<div class="pozn">Zobrazeno prvních {limit} z {len(rows)} řádků — '
           'úplný seznam je v XLSX.</div>') if len(rows) > limit else ''
    return (f'<h2>{e(nadpis)}</h2>'
            '<table class="t"><thead><tr>' + hlav + '</tr></thead><tbody>'
            + ''.join(telo) + '</tbody></table>' + vic)


def _pdf_obraty_html(asm, data, popis_filtru, uzivatel):
    e = html.escape
    k = data['kpi']
    zmena = lambda v: '—' if v is None else f'{v:+.1f} %'.replace('.', ',')
    return (
        '<!DOCTYPE html><html lang="cs"><head><meta charset="utf-8">'
        f'<title>Lupou na obchod — obraty {e(asm)}</title>{_PDF_CSS}</head><body>'
        '<div class="lupa-doc">'
        '<div class="hdr"><div>'
        '<h1>Lupou na obchod — obraty</h1>'
        f'<div class="firma">JIP východočeská, a.s. · ASM {e(asm.capitalize())}</div>'
        '</div><div class="meta">'
        f'<b>Vytiskl:</b> {e(uzivatel)}<br>'
        f'<b>Datum:</b> {datetime.datetime.now():%d.%m.%Y %H:%M}<br>'
        f'<b>Obrat celkem:</b> {_cz(k["celkem_kc"])} Kč'
        '</div></div>'
        f'<div class="filtr">{e(popis_filtru)}</div>'
        f'<div class="filtr">Zákazníků: {k["zakazniku"]} · OZ: {k["oz"]} · '
        f'Poslední měsíc {e(k["posl_mesic"])}: {_cz(k["posl_kc"])} Kč '
        f'(meziměsíčně {zmena(k["mom"])}, meziročně {zmena(k["yoy"])})</div>'
        + _pdf_tabulka('Obrat po měsících', _COLS_MESICE, data['radky_mesice'], 40)
        + _pdf_tabulka('Výsledky OZ', _COLS_OZ, data['oz'], 40)
        + _pdf_tabulka('Propady zákazníků', _COLS_PROPADY, data['propady'])
        + _pdf_tabulka('TOP odebrané položky', _COLS_POLOZKY, data['produkty'])
        + '<div class="pozn">Obraty bez DPH. Poslední měsíc může být neúplný, '
        'pokud import proběhl v jeho průběhu. Zdrojem pravdy zůstává GIST.</div>'
        '</div></body></html>'
    )


async def _render_pdf(html_text, jmeno, notifikace=None):
    """Render deleguje na modul Sankce — na serveru je jedno sdílené Chromium
    chráněné semaforem. Vlastní launch by znamenal dvě Chromia naráz a dvojnásobek
    paměti při souběžném tisku."""
    import intranet_sankce
    pdfs = await intranet_sankce._render_pdf_bytes_batch(
        [(jmeno, html_text)],
        on_wait=lambda: notifikace and notifikace.__setattr__(
            'message', 'Čekám ve frontě na tiskárnu…'),
        on_start=lambda: notifikace and notifikace.__setattr__(
            'message', 'Renderuji PDF…'))
    return pdfs[0][1]


def _cas_txt(hodnota):
    return hodnota.strftime('%d.%m.%Y %H:%M') if hodnota else ''


def _vykresli_poznamky(asm, user_id, user_name, vsechna_prava):
    admin = je_admin(vsechna_prava)
    smi_psat = je_moje_asm(user_id, asm, vsechna_prava)

    def smi_smazat(radek):
        return admin or (bool(user_id) and radek.get('user_id') == user_id)

    @ui.refreshable
    def seznam():
        radky = poznamky(asm)
        if not radky:
            with ui.column().classes('items-center py-16 gap-3 w-full'):
                ui.icon('sticky_note_2', size='4rem', color='grey-3')
                ui.label('Zatím žádná poznámka.' if smi_psat else
                         'ASM zatím nic nezapsal.') \
                    .classes('text-lg font-semibold text-gray-400')
            return

        for p in radky:
            with ui.card().classes('w-full shadow-sm rounded-xl mb-3'):
                with ui.row().classes('w-full items-center gap-2'):
                    ui.icon('person', color='indigo-4')
                    ui.label(p['uzivatel'] or '—').classes('font-semibold text-gray-800')
                    ui.label('· ' + _cas_txt(p['vytvoreno'])).classes('text-xs text-gray-500')
                    if p['upraveno']:
                        ui.label(f"(upraveno {_cas_txt(p['upraveno'])})") \
                            .classes('text-xs text-gray-400 italic')
                    ui.space()
                    if smi_smazat(p):
                        ui.button(icon='edit', on_click=lambda _, p=p: _dialog_edit(p)) \
                            .props('flat round dense color=grey-7').tooltip('Upravit')
                        ui.button(icon='delete',
                                  on_click=lambda _, p=p: _smaz_poznamku(p)) \
                            .props('flat round dense color=red-5').tooltip('Smazat')
                ui.label(p['text'] or '').classes('whitespace-pre-wrap text-gray-800 mt-1')

                with ui.column().classes('w-full gap-2 mt-2 pl-4 border-l-4 border-gray-200'):
                    for k in p['komentare']:
                        with ui.row().classes('w-full items-start gap-2'):
                            ui.icon('reply', size='1rem', color='grey-5').classes('mt-1')
                            with ui.column().classes('gap-0 grow'):
                                with ui.row().classes('items-center gap-2'):
                                    ui.label(k['uzivatel'] or '—') \
                                        .classes('text-sm font-semibold text-gray-700')
                                    ui.label(_cas_txt(k['vytvoreno'])) \
                                        .classes('text-xs text-gray-500')
                                ui.label(k['text'] or '') \
                                    .classes('whitespace-pre-wrap text-sm text-gray-700')
                            if smi_smazat(k):
                                ui.button(icon='close',
                                          on_click=lambda _, k=k: _smaz_komentar(k)) \
                                    .props('flat round dense size=sm color=grey-5')

                    with ui.row().classes('w-full items-center gap-2'):
                        pole = ui.input(placeholder='Napsat komentář…') \
                            .props('dense outlined').classes('grow')
                        pole.on('keydown.enter', lambda _, p=p, pole=pole: _pridej_komentar(p, pole))
                        ui.button(icon='send', on_click=lambda _, p=p, pole=pole:
                                  _pridej_komentar(p, pole)) \
                            .props('flat round dense color=indigo-6')

    def _pridej_komentar(p, pole):
        text = (pole.value or '').strip()
        if not text:
            return
        pridej_komentar(p['id'], user_id, user_name, text)
        zapis_log(user_id, user_name, 'komentar', asm, f"poznamka {p['id']}")
        # Autor poznámky se o zpětné vazbě dozví, i když zrovna není v modulu.
        if p.get('user_id') and p['user_id'] != user_id:
            intranet_notifikace.pridej(
                p['user_id'], f'{user_name} okomentoval vaši poznámku v Lupou na obchod ({asm}).')
        pole.value = ''
        seznam.refresh()

    def _smaz_komentar(k):
        smaz_komentar(k['id'])
        zapis_log(user_id, user_name, 'komentar_smaz', asm, str(k['id']))
        seznam.refresh()

    def _smaz_poznamku(p):
        smaz_poznamku(p['id'])
        zapis_log(user_id, user_name, 'poznamka_smaz', asm, str(p['id']))
        ui.notify('Poznámka smazána.', type='positive')
        seznam.refresh()

    def _dialog_edit(p):
        with ui.dialog() as dlg, ui.card().classes('w-full max-w-2xl'):
            ui.label('Úprava poznámky').classes('text-lg font-bold text-gray-800')
            pole = ui.textarea(value=p['text'] or '').props('outlined autogrow').classes('w-full')
            with ui.row().classes('w-full justify-end gap-2'):
                ui.button('Zavřít', on_click=dlg.close).props('flat')

                def uloz():
                    text = (pole.value or '').strip()
                    if not text:
                        ui.notify('Prázdnou poznámku neuložím.', type='warning')
                        return
                    uprav_poznamku(p['id'], text)
                    zapis_log(user_id, user_name, 'poznamka_edit', asm, str(p['id']))
                    dlg.close()
                    seznam.refresh()

                ui.button('Uložit', on_click=uloz).props('color=indigo-6')
        dlg.open()

    ui.label('Moje poznámky').classes('text-3xl font-bold text-gray-800 mb-1')
    ui.label('Postřehy k území. Extra čtenáři je vidí a mohou přidat komentář.') \
        .classes('text-gray-500 mb-4')

    if smi_psat:
        with ui.card().classes('w-full shadow-sm rounded-xl mb-4'):
            nova = ui.textarea(placeholder='Co je nového u zákazníků, kde je příležitost…') \
                .props('outlined autogrow').classes('w-full')

            def pridej():
                text = (nova.value or '').strip()
                if not text:
                    ui.notify('Napište nejdřív text poznámky.', type='warning')
                    return
                pridej_poznamku(asm, user_id, user_name, text)
                zapis_log(user_id, user_name, 'poznamka', asm, pocet_radku=len(text))
                nova.value = ''
                ui.notify('Poznámka uložena.', type='positive')
                seznam.refresh()

            with ui.row().classes('w-full justify-end'):
                ui.button('Přidat poznámku', icon='add', on_click=pridej) \
                    .props('unelevated color=indigo-7')
    else:
        ui.label('Psát poznámky může obchodní zástupce daného ASM — vy můžete komentovat.') \
            .classes('text-xs text-gray-500 mb-3')

    seznam()


def _vykresli_detail(asm, user_id, user_name, vsechna_prava):
    with ui.tabs().props('align=left active-color=primary indicator-color=primary') \
            .classes('w-full border-b border-gray-200') as taby:
        ui.tab('odberatele', label='📄 Lupni odběratele')
        ui.tab('obraty', label='📊 Lupni obraty')
        ui.tab('poznamky', label='📝 Moje poznámky')
    with ui.tab_panels(taby, value='odberatele').classes('w-full pt-4'):
        with ui.tab_panel('odberatele'):
            _vykresli_odberatele(asm, user_id, user_name, vsechna_prava)
        with ui.tab_panel('obraty'):
            _vykresli_obraty(asm, user_id, user_name, vsechna_prava)
        with ui.tab_panel('poznamky'):
            _vykresli_poznamky(asm, user_id, user_name, vsechna_prava)


# Formátování čísel v gridu necháváme na prohlížeči (Intl) — stejný vzor jako
# ve Výsledcích poboček; hodnoty tak zůstávají čísla a dají se řadit.
def _fmt_js(desetin):
    return (
        "function(p){"
        "if(p.value===null||p.value===undefined||p.value==='')return '—';"
        "var n=parseFloat(p.value);if(isNaN(n))return '—';"
        "return new Intl.NumberFormat('cs-CZ',{minimumFractionDigits:%d,"
        "maximumFractionDigits:%d}).format(n);}" % (desetin, desetin)
    )


_GRID_ZAKLAD = {
    'defaultColDef': {'resizable': True, 'sortable': True, 'wrapHeaderText': True,
                      'autoHeaderHeight': True},
    'rowHeight': 32,
    'suppressMovableColumns': True,
    ':getRowStyle': ("function(p){if(p.node&&p.node.rowPinned)"
                     "return{fontWeight:'bold',backgroundColor:'#e0e7ff'};return null;}"),
}


def _grid(opts):
    """AG Grid 34 odmítá colDef.flex spolu s gridOptions.autoSizeStrategy, kterou
    NiceGUI přidává při auto_size_columns=True (default) — grid se pak nevykreslí.
    Šířky řešíme přes flex/minWidth, takže strategii vypínáme."""
    return ui.aggrid(opts, auto_size_columns=False)


def _bez_klienta(fn):
    """Načtení a exporty běží desítky sekund až minuty. Klient mezitím může zmizet
    (reload, odchod ze stránky, refresh dlaždic) a každé další UI volání —
    ui.download, ui.notify, notification.dismiss — pak hodí RuntimeError o smazaném
    elementu. Doručovat není komu, takže jen tichý log místo tří tracebacků."""
    async def obal(*args):
        try:
            await fn()
        except RuntimeError as e:
            if 'has been deleted' not in str(e):
                raise
            print(f'[lupa] {fn.__name__}: klient odešel před dokončením')
    return obal


def _vykresli_odberatele(asm, user_id, user_name, vsechna_prava):
    filtr_klic = f'lupa_filtr_{user_id}_{asm}'
    ulozeny = app.storage.user.get(filtr_klic) or {}
    obdobi = _volby_obdobi(asm)
    zak_volby = _volby_zakazniku(asm)
    dod_volby = _volby_dodavatelu()

    stav = {'rows': [], 'total': {}, 'filtr': {}, 'popis': ''}

    def _zaklad_jmena(co):
        return (f'lupa_{_bezpecne_jmeno(asm)}_{co}_'
                f'{datetime.datetime.now():%Y-%m-%d_%H%M}')

    _pruh_importu(asm)

    with ui.row().classes('w-full items-end gap-3 mb-2 flex-wrap'):
        sel_od = ui.select(obdobi, label='Od měsíce', clearable=True,
                           value=ulozeny.get('od') if ulozeny.get('od') in obdobi else None) \
            .props('dense outlined options-dense').style('min-width: 140px')
        sel_do = ui.select(obdobi, label='Do měsíce', clearable=True,
                           value=ulozeny.get('do') if ulozeny.get('do') in obdobi else None) \
            .props('dense outlined options-dense').style('min-width: 140px')
        sel_zak = ui.select(
            zak_volby, label='Zákazníci', multiple=True, with_input=True, clearable=True,
            value=[i for i in (ulozeny.get('ico') or []) if i in zak_volby]) \
            .props('dense outlined options-dense use-chips').style('min-width: 340px') \
            .tooltip('Prázdné = všichni zákazníci ASM. Pište jméno nebo IČO.')
        _pole_ica(sel_zak, zak_volby)
        sel_dod = ui.select(
            dod_volby, label='Dodavatelé', multiple=True, with_input=True, clearable=True,
            value=[d for d in (ulozeny.get('dodavatel') or []) if d in dod_volby]) \
            .props('dense outlined options-dense use-chips').style('min-width: 280px') \
            .tooltip('Prázdné = všichni dodavatelé.')
        chk_rozpad = ui.checkbox('Rozpad po produktech',
                                 value=bool(ulozeny.get('rozpad'))) \
            .props('dense color=indigo-7') \
            .tooltip('Řádek = zákazník × produkt, s kódem, názvem a dodavatelem. '
                     'Řádků je násobně víc — na PDF platí strop '
                     f'{_STROP_PDF} řádků.')

    def _soucasny_filtr():
        return {'od': sel_od.value, 'do': sel_do.value,
                'ico': list(sel_zak.value or []), 'dodavatel': list(sel_dod.value or []),
                'rozpad': bool(chk_rozpad.value)}

    with ui.row().classes('w-full items-center gap-3 mb-3 flex-wrap'):
        btn_nacti = ui.button('Načíst', icon='search') \
            .props('unelevated color=indigo-7')
        btn_pdf = ui.button('PDF souhrn', icon='picture_as_pdf') \
            .props('outline color=indigo-7')
        btn_xlsx = ui.button('XLSX souhrn', icon='table_view') \
            .props('outline color=green-7')
        btn_detail = ui.button('Detailní pohled (Excel)', icon='download') \
            .props('outline color=green-9') \
            .tooltip('Surové řádky dle filtru — zákazník × produkt × měsíc. '
                     'U celého ASM jde o stovky tisíc řádků a export trvá minuty.')

    for b in (btn_pdf, btn_xlsx, btn_detail):
        b.disable()

    vysledek = ui.column().classes('w-full')

    with vysledek:
        with ui.column().classes('items-center py-16 gap-3 w-full'):
            ui.icon('filter_alt', size='4rem', color='grey-3')
            ui.label('Zvolte filtr a klikněte na Načíst') \
                .classes('text-lg font-semibold text-gray-400')

    async def otevri_zakaznika(e):
        radek = (e.args or {}).get('data') or {}
        ico = radek.get('ico')
        if not ico or radek.get('_celkem') or _blokuje_import(asm):
            return
        produkty = await asyncio.to_thread(_produkty_zakaznika, asm, ico, stav['filtr'])
        with ui.dialog() as dlg, ui.card().classes('w-full max-w-6xl'):
            ui.label(f"{radek.get('jmeno') or ico} · {ico}") \
                .classes('text-xl font-bold text-gray-800')
            ui.label(f"{radek.get('k_ulice') or ''} {radek.get('k_mesto') or ''} · "
                     f"{stav['popis']}").classes('text-xs text-gray-500 mb-2')
            _grid({
                'columnDefs': [
                    {'headerName': 'Kód', 'field': 'kod', 'width': 130},
                    {'headerName': 'Produkt', 'field': 'nazev', 'flex': 2, 'minWidth': 240},
                    {'headerName': 'Dodavatel', 'field': 'dodavatel', 'flex': 1,
                     'minWidth': 180},
                    {'headerName': 'Obrat MJ', 'field': 'mj', 'width': 130,
                     'type': 'numericColumn', ':valueFormatter': _fmt_js(1)},
                    {'headerName': 'Obrat Kč', 'field': 'kc', 'width': 150,
                     'type': 'numericColumn', ':valueFormatter': _fmt_js(2)},
                ],
                'rowData': produkty,
                'pinnedBottomRowData': [{
                    'nazev': f'CELKEM {len(produkty)} položek',
                    'mj': sum(float(p['mj'] or 0) for p in produkty),
                    'kc': sum(float(p['kc'] or 0) for p in produkty)}],
                **_GRID_ZAKLAD,
            }).classes('w-full').style('height: 60vh')
            with ui.row().classes('w-full justify-end mt-2'):
                ui.button('Zavřít', on_click=dlg.close).props('flat')
        dlg.open()

    def vykresli_vysledek():
        rows, total = stav['rows'], stav['total']
        vysledek.clear()
        with vysledek:
            if not rows:
                with ui.column().classes('items-center py-16 gap-3 w-full'):
                    ui.icon('search_off', size='4rem', color='grey-3')
                    ui.label('Filtru nic neodpovídá') \
                        .classes('text-lg font-semibold text-gray-400')
                return
            zobrazene = rows[:_STROP_NAHLED]
            if len(rows) > _STROP_NAHLED:
                ui.label(f'Náhled ukazuje prvních {_STROP_NAHLED} řádků z '
                         f'{len(rows)} — exporty obsahují vše.') \
                    .classes('text-xs text-amber-700 mb-1')
            rozpad = bool(stav['filtr'].get('rozpad'))
            grid = _grid({
                'columnDefs': [
                    {'headerName': 'IČO', 'field': 'ico', 'width': 130},
                    {'headerName': 'Zákazník', 'field': 'jmeno', 'flex': 2,
                     'minWidth': 220},
                    {'headerName': 'K. jméno', 'field': 'k_jmeno', 'flex': 1,
                     'minWidth': 150},
                    {'headerName': 'K. ulice', 'field': 'k_ulice', 'flex': 1,
                     'minWidth': 150},
                    {'headerName': 'K. město', 'field': 'k_mesto', 'width': 150},
                ] + ([
                    {'headerName': 'Produkt - Kód', 'field': 'kod', 'width': 130},
                    {'headerName': 'Produkt - název', 'field': 'nazev', 'flex': 2,
                     'minWidth': 220},
                    {'headerName': 'Dodavatel - popis', 'field': 'dodavatel',
                     'flex': 1, 'minWidth': 170},
                ] if rozpad else [
                    {'headerName': 'Položek', 'field': 'polozek', 'width': 110,
                     'type': 'numericColumn', ':valueFormatter': _fmt_js(0)},
                ]) + [
                    {'headerName': 'Obrat MJ', 'field': 'mj', 'width': 130,
                     'type': 'numericColumn', ':valueFormatter': _fmt_js(1)},
                    {'headerName': 'Obrat Kč bez DPH', 'field': 'kc', 'width': 170,
                     'type': 'numericColumn', ':valueFormatter': _fmt_js(2)},
                ],
                'rowData': zobrazene,
                'pinnedBottomRowData': [dict(total, _celkem=True)],
                **_GRID_ZAKLAD,
            }).classes('w-full').style('height: calc(100vh - 430px); min-height: 380px')
            grid.on('rowClicked', otevri_zakaznika)
            ui.label('Klikem na řádek zobrazíte, co zákazník ve zvoleném období bral.') \
                .classes('text-xs text-gray-400 mt-1')

    async def nacti():
        if _blokuje_import(asm):
            return
        filtr = _soucasny_filtr()
        app.storage.user[filtr_klic] = filtr
        btn_nacti.disable()
        chk_rozpad.disable()
        pozn = ui.notification('Načítám data…', spinner=True, timeout=None)
        try:
            rows = await asyncio.to_thread(_souhrn_odberatelu, asm, filtr)
        finally:
            pozn.dismiss()
            btn_nacti.enable()
            chk_rozpad.enable()
        rozpad = bool(filtr.get('rozpad'))
        stav.update({'rows': rows, 'total': _souhrn_total(rows, rozpad),
                     'filtr': filtr, 'popis': _popis_filtru(filtr, zak_volby)})
        for b in (btn_pdf, btn_xlsx, btn_detail):
            b.set_enabled(bool(rows))
        vykresli_vysledek()
        await asyncio.to_thread(zapis_log, user_id, user_name, 'nahled', asm,
                                stav['popis'], len(rows))

    async def export_pdf():
        rows = stav['rows']
        if len(rows) > _STROP_PDF:
            ui.notify(f'PDF zvládne {_STROP_PDF} řádků, filtr jich vrací {len(rows)}. '
                      'Zpřísněte filtr, nebo použijte XLSX.',
                      type='warning', multi_line=True)
            return
        jmeno = _zaklad_jmena('odberatele') + '.pdf'
        pozn = ui.notification('Připravuji PDF…', spinner=True, timeout=None)
        try:
            html_text = _pdf_html(asm, rows, stav['total'], stav['popis'], user_name,
                                  bool(stav['filtr'].get('rozpad')))
            data = await _render_pdf(html_text, jmeno, pozn)
            _stahni_bajty(data, jmeno)
            await asyncio.to_thread(zapis_log, user_id, user_name, 'export_pdf', asm,
                                    stav['popis'], len(rows))
        except Exception as e:
            ui.notify(f'PDF se nepodařilo vytvořit: {e}', type='negative',
                      multi_line=True, timeout=15000)
        finally:
            pozn.dismiss()

    async def export_xlsx():
        jmeno = _zaklad_jmena('odberatele') + '.xlsx'
        cesta = _tmp_cesta(jmeno)
        pozn = ui.notification('Sestavuji XLSX…', spinner=True, timeout=None)
        try:
            cols = (_COLS_SOUHRN_ROZPAD if stav['filtr'].get('rozpad')
                    else _COLS_SOUHRN)
            await asyncio.to_thread(_xlsx_na_disk, cesta, cols, stav['rows'],
                                    stav['total'], 'Odběratelé')
            _stahni_soubor(cesta, jmeno)
            await asyncio.to_thread(zapis_log, user_id, user_name, 'export_xlsx', asm,
                                    stav['popis'], len(stav['rows']))
        except Exception as e:
            ui.notify(f'Export selhal: {e}', type='negative', multi_line=True)
        finally:
            pozn.dismiss()

    async def export_detail():
        if _blokuje_import(asm):
            return
        jmeno = _zaklad_jmena('detail') + '.xlsx'
        cesta = _tmp_cesta(jmeno)
        pozn = ui.notification('Sestavuji detailní výjezd — u celého ASM to trvá '
                               'i pár minut…', spinner=True, timeout=None)
        btn_detail.disable()
        try:
            pocet = await asyncio.to_thread(_export_detail_na_disk, asm,
                                            stav['filtr'], cesta)
            _stahni_soubor(cesta, jmeno)
            ui.notify(f'Hotovo — {pocet} řádků.', type='positive')
            await asyncio.to_thread(zapis_log, user_id, user_name, 'export_xlsx_detail',
                                    asm, stav['popis'], pocet)
        except Exception as e:
            ui.notify(f'Export selhal: {e}', type='negative', multi_line=True)
        finally:
            pozn.dismiss()
            btn_detail.enable()

    async def prepni_rozpad():
        # data na obrazovce = jiná granularita než nový stav přepínače → přenačíst
        if stav['rows']:
            await nacti()

    btn_nacti.on_click(_bez_klienta(nacti))
    btn_pdf.on_click(_bez_klienta(export_pdf))
    btn_xlsx.on_click(_bez_klienta(export_xlsx))
    btn_detail.on_click(_bez_klienta(export_detail))
    chk_rozpad.on_value_change(_bez_klienta(prepni_rozpad))


def _graf_zaklad(nadpis, legenda_vpravo=False):
    return {
        'title': {'text': nadpis, 'left': 'center', 'top': 6,
                  'textStyle': {'fontSize': 14}},
        'tooltip': {'trigger': 'axis'},
        'legend': ({'type': 'scroll', 'orient': 'vertical', 'right': 0, 'top': 40,
                    'bottom': 20} if legenda_vpravo else
                   {'type': 'scroll', 'top': 30}),
        'grid': {'top': 70 if not legenda_vpravo else 40, 'left': 80,
                 'right': 220 if legenda_vpravo else 24, 'bottom': 40},
        'yAxis': {'type': 'value'},
    }


def _graf_yoy_option(radky_mesice):
    """Osa X = měsíc 1–12, série = rok. Meziroční srovnání na první pohled."""
    kc = {r['klic']: r['kc'] for r in radky_mesice}
    roky = sorted({k // 100 for k in kc})
    opt = _graf_zaklad('Obrat ASM po měsících (Kč bez DPH, meziroční srovnání)')
    opt['xAxis'] = {'type': 'category', 'data': _MESICE_KR}
    opt['series'] = [{
        'name': str(rok), 'type': 'bar', 'emphasis': {'focus': 'series'},
        'itemStyle': {'color': _GRAF_PALETA[i % len(_GRAF_PALETA)]},
        'data': [round(kc[rok * 100 + m]) if (rok * 100 + m) in kc else None
                 for m in range(1, 13)],
    } for i, rok in enumerate(roky)]
    return opt


def _graf_rada_option(nadpis, mesice, serie):
    """Chronologická spojnice přes zvolené období; série = OZ nebo zákazník."""
    opt = _graf_zaklad(nadpis, legenda_vpravo=True)
    opt['xAxis'] = {'type': 'category', 'data': [_mesic_txt(m) for m in mesice]}
    opt['series'] = [{
        'name': s['name'], 'type': 'line', 'smooth': True, 'connectNulls': True,
        'emphasis': {'focus': 'series'},
        'itemStyle': {'color': _GRAF_PALETA[i % len(_GRAF_PALETA)]},
        'lineStyle': {'color': _GRAF_PALETA[i % len(_GRAF_PALETA)]},
        'data': s['data'],
    } for i, s in enumerate(serie)]
    return opt


def _graf_polozky_option(produkty, n=15):
    """Vodorovné sloupce — názvy produktů se na osu X nevejdou."""
    top = list(reversed(produkty[:n]))
    opt = _graf_zaklad(f'TOP {len(top)} položek podle obratu (Kč bez DPH)')
    opt['legend'] = {'show': False}
    opt['tooltip'] = {'trigger': 'axis', 'axis': 'shadow'}
    opt['grid'] = {'top': 40, 'left': 240, 'right': 40, 'bottom': 30}
    opt['xAxis'] = {'type': 'value'}
    opt['yAxis'] = {'type': 'category',
                    'data': [(p['nazev'] or p['kod'])[:38] for p in top],
                    'axisLabel': {'fontSize': 10}}
    opt['series'] = [{'type': 'bar', 'itemStyle': {'color': '#4338CA'},
                      'data': [round(float(p['kc'] or 0)) for p in top]}]
    return opt


def _vykresli_obraty(asm, user_id, user_name, vsechna_prava):
    # Filtr sdílíme s odběrateli — je to týž pohled na táž data, jen jinak
    # vykreslený; dvě sady filtrů by uživatel jen držel v synchronizaci ručně.
    filtr_klic = f'lupa_filtr_{user_id}_{asm}'
    prah_klic = f'lupa_propad_prahy_{user_id}'
    ulozeny = app.storage.user.get(filtr_klic) or {}
    prahy = app.storage.user.get(prah_klic) or {}
    obdobi = _volby_obdobi(asm)
    zak_volby = _volby_zakazniku(asm)
    dod_volby = _volby_dodavatelu()

    stav = {'data': None, 'filtr': {}, 'popis': ''}

    def _zaklad_jmena(co):
        return (f'lupa_{_bezpecne_jmeno(asm)}_{co}_'
                f'{datetime.datetime.now():%Y-%m-%d_%H%M}')

    _pruh_importu(asm)

    with ui.row().classes('w-full items-end gap-3 mb-2 flex-wrap'):
        sel_od = ui.select(obdobi, label='Od měsíce', clearable=True,
                           value=ulozeny.get('od') if ulozeny.get('od') in obdobi else None) \
            .props('dense outlined options-dense').style('min-width: 140px')
        sel_do = ui.select(obdobi, label='Do měsíce', clearable=True,
                           value=ulozeny.get('do') if ulozeny.get('do') in obdobi else None) \
            .props('dense outlined options-dense').style('min-width: 140px')
        sel_zak = ui.select(
            zak_volby, label='Zákazníci', multiple=True, with_input=True, clearable=True,
            value=[i for i in (ulozeny.get('ico') or []) if i in zak_volby]) \
            .props('dense outlined options-dense use-chips').style('min-width: 340px') \
            .tooltip('Prázdné = všichni zákazníci ASM. Pište jméno nebo IČO.')
        _pole_ica(sel_zak, zak_volby)
        sel_dod = ui.select(
            dod_volby, label='Dodavatelé', multiple=True, with_input=True, clearable=True,
            value=[d for d in (ulozeny.get('dodavatel') or []) if d in dod_volby]) \
            .props('dense outlined options-dense use-chips').style('min-width: 280px') \
            .tooltip('Prázdné = všichni dodavatelé.')
        num_pct = ui.number(label='Propad od %', value=prahy.get('pct', 25),
                            min=0, max=100, step=5, format='%.0f') \
            .props('dense outlined').style('min-width: 120px') \
            .tooltip('Pokles posledního měsíce proti průměru předchozích tří.')
        num_kc = ui.number(label='… a od Kč', value=prahy.get('kc', 10000),
                           min=0, step=1000, format='%.0f') \
            .props('dense outlined').style('min-width: 130px') \
            .tooltip('Filtruje drobné poklesy, které nestojí za telefonát.')

    def _soucasny_filtr():
        return {'od': sel_od.value, 'do': sel_do.value,
                'ico': list(sel_zak.value or []), 'dodavatel': list(sel_dod.value or [])}

    with ui.row().classes('w-full items-center gap-3 mb-3 flex-wrap'):
        btn_nacti = ui.button('Načíst', icon='insights').props('unelevated color=indigo-7')
        btn_pdf = ui.button('PDF přehled', icon='picture_as_pdf') \
            .props('outline color=indigo-7')
        btn_xlsx = ui.button('XLSX (4 listy)', icon='table_view') \
            .props('outline color=green-7') \
            .tooltip('Měsíce, OZ, propady, TOP položky — každý na vlastním listu.')
        info = ui.label('').classes('text-sm text-gray-500')

    for b in (btn_pdf, btn_xlsx):
        b.disable()

    vysledek = ui.column().classes('w-full')

    with vysledek:
        with ui.column().classes('items-center py-16 gap-3 w-full'):
            ui.icon('query_stats', size='4rem', color='grey-3')
            ui.label('Zvolte filtr a klikněte na Načíst') \
                .classes('text-lg font-semibold text-gray-400')

    def _dlazdice(nadpis, hodnota, podtitul='', barva='indigo'):
        with ui.card().classes(f'flex-1 min-w-[200px] border-l-4 border-{barva}-500 '
                               'shadow-sm rounded-xl'):
            ui.label(nadpis).classes('text-xs uppercase text-gray-500')
            ui.label(hodnota).classes(f'text-2xl font-bold text-{barva}-700')
            if podtitul:
                ui.label(podtitul).classes('text-xs text-gray-500')

    def _zmena_txt(v):
        return '—' if v is None else f'{v:+.1f} %'.replace('.', ',')

    def _zmena_barva(v):
        if v is None:
            return 'gray'
        return 'green' if v >= 0 else 'red'

    def vykresli_vysledek():
        data = stav['data']
        vysledek.clear()
        with vysledek:
            if not data or not data['mesice']:
                with ui.column().classes('items-center py-16 gap-3 w-full'):
                    ui.icon('search_off', size='4rem', color='grey-3')
                    ui.label('Filtru nic neodpovídá') \
                        .classes('text-lg font-semibold text-gray-400')
                return
            k = data['kpi']
            with ui.row().classes('w-full gap-3 flex-wrap mb-2'):
                _dlazdice('Obrat celkem', _kc(k['celkem_kc']),
                          f"{len(data['mesice'])} měsíců · {k['oz']} OZ")
                _dlazdice('Zákazníků', f"{k['zakazniku']}",
                          's obratem ve zvoleném období', 'cyan')
                _dlazdice(f"Poslední měsíc ({k['posl_mesic']})", _kc(k['posl_kc']),
                          f"meziměsíčně {_zmena_txt(k['mom'])}",
                          _zmena_barva(k['mom']))
                _dlazdice('Meziročně', _zmena_txt(k['yoy']),
                          f"{k['posl_mesic']} proti stejnému měsíci loni",
                          _zmena_barva(k['yoy']))
            ui.label('Poslední měsíc v datech může být neúplný, pokud import '
                     'proběhl v jeho průběhu — meziměsíční srovnání pak podhodnocuje.') \
                .classes('text-xs text-gray-400 italic mb-2')

            with ui.card().classes('w-full shadow-sm rounded-xl mb-3'):
                ui.echart(_graf_yoy_option(data['radky_mesice'])) \
                    .classes('w-full').style('height: 340px')

            propady = data['propady']
            with ui.card().classes('w-full shadow-sm rounded-xl mb-3 '
                                   + ('border-l-4 border-red-500' if propady else '')):
                with ui.row().classes('items-center gap-2 mb-1'):
                    ui.icon('warning' if propady else 'check_circle',
                            color='red-6' if propady else 'green-6')
                    ui.label(f'Propady zákazníků — {len(propady)} nad prahem'
                             if propady else 'Propady zákazníků — nic nad prahem') \
                        .classes('text-lg font-semibold text-gray-800')
                ui.label(f'Poslední měsíc {k["posl_mesic"]} proti průměru '
                         f'předchozích {_PROPAD_OKNO} měsíců.') \
                    .classes('text-xs text-gray-500 mb-2')
                if propady:
                    _grid({
                        'columnDefs': [
                            {'headerName': 'IČO', 'field': 'ico', 'width': 120},
                            {'headerName': 'Zákazník', 'field': 'jmeno', 'flex': 2,
                             'minWidth': 220},
                            {'headerName': 'Průměr předchozích Kč', 'field': 'zaklad',
                             'width': 180, 'type': 'numericColumn',
                             ':valueFormatter': _fmt_js(0)},
                            {'headerName': 'Poslední měsíc Kč', 'field': 'akt',
                             'width': 160, 'type': 'numericColumn',
                             ':valueFormatter': _fmt_js(0)},
                            {'headerName': 'Rozdíl Kč', 'field': 'rozdil', 'width': 150,
                             'type': 'numericColumn', ':valueFormatter': _fmt_js(0)},
                            {'headerName': 'Změna %', 'field': 'pct', 'width': 120,
                             'type': 'numericColumn', ':valueFormatter': _fmt_js(1)},
                            {'headerName': 'Stav', 'field': 'stav', 'width': 120},
                        ],
                        'rowData': propady[:_STROP_NAHLED],
                        **_GRID_ZAKLAD,
                    }).classes('w-full').style('height: 320px')

            with ui.card().classes('w-full shadow-sm rounded-xl mb-3'):
                ui.echart(_graf_rada_option(
                    'Obrat OZ v čase (Kč bez DPH)', data['mesice'],
                    [{'name': d['dealer_jmeno'],
                      'data': [round(d['rada'].get(m, 0.0), 2) for m in data['mesice']]}
                     for d in data['oz']])).classes('w-full').style('height: 360px')
                _grid({
                    'columnDefs': [
                        {'headerName': 'OZ', 'field': 'dealer', 'width': 110},
                        {'headerName': 'Jméno', 'field': 'dealer_jmeno', 'flex': 1,
                         'minWidth': 200},
                        {'headerName': 'Obrat MJ', 'field': 'mj', 'width': 140,
                         'type': 'numericColumn', ':valueFormatter': _fmt_js(1)},
                        {'headerName': 'Obrat Kč', 'field': 'kc', 'width': 160,
                         'type': 'numericColumn', ':valueFormatter': _fmt_js(2)},
                        {'headerName': 'Podíl %', 'field': 'podil', 'width': 110,
                         'type': 'numericColumn', ':valueFormatter': _fmt_js(1)},
                    ],
                    'rowData': data['oz'],
                    'pinnedBottomRowData': [{
                        'dealer_jmeno': f"CELKEM {len(data['oz'])} OZ",
                        'mj': k['celkem_mj'], 'kc': k['celkem_kc'], 'podil': 100.0}],
                    **_GRID_ZAKLAD,
                }).classes('w-full mt-2').style('height: 260px')

            with ui.card().classes('w-full shadow-sm rounded-xl mb-3'):
                ui.echart(_graf_rada_option(
                    f'Vývoj obratů — TOP {len(data["zak_serie"])} zákazníků',
                    data['mesice'], data['zak_serie'])) \
                    .classes('w-full').style('height: 380px')

            with ui.card().classes('w-full shadow-sm rounded-xl'):
                ui.echart(_graf_polozky_option(data['produkty'])) \
                    .classes('w-full').style('height: 420px')
                _grid({
                    'columnDefs': [
                        {'headerName': 'Kód', 'field': 'kod', 'width': 130},
                        {'headerName': 'Produkt', 'field': 'nazev', 'flex': 2,
                         'minWidth': 240},
                        {'headerName': 'Dodavatel', 'field': 'dodavatel', 'flex': 1,
                         'minWidth': 180},
                        {'headerName': 'Zákazníků', 'field': 'zakazniku', 'width': 120,
                         'type': 'numericColumn', ':valueFormatter': _fmt_js(0)},
                        {'headerName': 'Obrat MJ', 'field': 'mj', 'width': 130,
                         'type': 'numericColumn', ':valueFormatter': _fmt_js(1)},
                        {'headerName': 'Obrat Kč', 'field': 'kc', 'width': 150,
                         'type': 'numericColumn', ':valueFormatter': _fmt_js(2)},
                    ],
                    'rowData': data['produkty'],
                    **_GRID_ZAKLAD,
                }).classes('w-full mt-2').style('height: 320px')

    async def nacti():
        if _blokuje_import(asm):
            return
        filtr = _soucasny_filtr()
        prah_pct = float(num_pct.value or 0)
        prah_kc = float(num_kc.value or 0)
        app.storage.user[filtr_klic] = filtr
        app.storage.user[prah_klic] = {'pct': prah_pct, 'kc': prah_kc}
        btn_nacti.disable()
        pozn = ui.notification('Počítám obraty…', spinner=True, timeout=None)
        try:
            data = await asyncio.to_thread(_obraty_prehled, asm, filtr,
                                           prah_pct, prah_kc)
        finally:
            pozn.dismiss()
            btn_nacti.enable()
        stav.update({'data': data, 'filtr': filtr,
                     'popis': _popis_filtru(filtr, zak_volby)})
        k = data['kpi']
        info.set_text(f"{_kc(k['celkem_kc'])} · {k['zakazniku']} zákazníků · "
                      f"{len(data['propady'])} propadů")
        for b in (btn_pdf, btn_xlsx):
            b.set_enabled(bool(data['mesice']))
        vykresli_vysledek()
        if data['propady']:
            ui.notify(f"{len(data['propady'])} zákazníků propadlo v měsíci "
                      f"{k['posl_mesic']} — viz červená karta.",
                      type='warning', multi_line=True, timeout=8000)
        await asyncio.to_thread(zapis_log, user_id, user_name, 'obraty_nahled', asm,
                                stav['popis'], len(data['mesice']))

    async def export_pdf():
        data = stav['data']
        jmeno = _zaklad_jmena('obraty') + '.pdf'
        pozn = ui.notification('Připravuji PDF…', spinner=True, timeout=None)
        try:
            html_text = _pdf_obraty_html(asm, data, stav['popis'], user_name)
            bajty = await _render_pdf(html_text, jmeno, pozn)
            _stahni_bajty(bajty, jmeno)
            await asyncio.to_thread(zapis_log, user_id, user_name, 'obraty_export_pdf',
                                    asm, stav['popis'], len(data['propady']))
        except Exception as e:
            ui.notify(f'PDF se nepodařilo vytvořit: {e}', type='negative',
                      multi_line=True, timeout=15000)
        finally:
            pozn.dismiss()

    async def export_xlsx():
        data = stav['data']
        jmeno = _zaklad_jmena('obraty') + '.xlsx'
        cesta = _tmp_cesta(jmeno)
        pozn = ui.notification('Sestavuji XLSX…', spinner=True, timeout=None)
        try:
            k = data['kpi']
            listy = [
                ('Měsíce', _COLS_MESICE, data['radky_mesice'],
                 {'mesic': 'CELKEM', 'mj': k['celkem_mj'], 'kc': k['celkem_kc']}),
                ('OZ', _COLS_OZ, data['oz'],
                 {'dealer_jmeno': f"CELKEM {len(data['oz'])} OZ",
                  'mj': k['celkem_mj'], 'kc': k['celkem_kc'], 'podil': 100.0}),
                ('Propady', _COLS_PROPADY, data['propady'],
                 {'jmeno': f"CELKEM {len(data['propady'])} zákazníků",
                  'rozdil': sum(p['rozdil'] for p in data['propady'])}),
                ('TOP položky', _COLS_POLOZKY, data['produkty'], None),
            ]
            await asyncio.to_thread(_xlsx_listy_na_disk, cesta, listy)
            _stahni_soubor(cesta, jmeno)
            await asyncio.to_thread(zapis_log, user_id, user_name, 'obraty_export_xlsx',
                                    asm, stav['popis'], len(data['radky_mesice']))
        except Exception as e:
            ui.notify(f'Export selhal: {e}', type='negative', multi_line=True)
        finally:
            pozn.dismiss()

    btn_nacti.on_click(_bez_klienta(nacti))
    btn_pdf.on_click(_bez_klienta(export_pdf))
    btn_xlsx.on_click(_bez_klienta(export_xlsx))


def _zamek_asm(klic):
    """Jedno ASM = jeden import naráz. Zámků je tolik co ASM, tj. desítky."""
    zamek = _ASM_ZAMKY.get(klic)
    if zamek is None:
        zamek = _ASM_ZAMKY[klic] = asyncio.Lock()   # event loop → bez race
    return zamek


def bezi_import(asm=None):
    """Běží import daného ASM (nebo jakýkoli)? Čte se z každé relace."""
    if not _BEZICI_IMPORTY:
        return False
    if asm is None:
        return True
    hledany = (asm or '').upper()
    return any((s.get('asm') or '').upper() == hledany
               for s in list(_BEZICI_IMPORTY.values()))


def _blokuje_import(asm):
    """DELETE+INSERT má data rozpůlená — dotazy během něj nepouštíme."""
    if not bezi_import(asm):
        return False
    ui.notify('Probíhá import dat, zkuste to později.', type='warning')
    return True


def _pruh_importu(asm):
    """Hláška nad filtrem, dokud běží import daného ASM."""
    pruh = ui.row().classes('w-full items-center gap-2 mb-3 px-3 py-2 rounded-lg '
                            'bg-amber-50 border border-amber-200')
    with pruh:
        ui.spinner(size='sm', color='amber-8')
        ui.label('Probíhá import dat, zkuste to později.') \
            .classes('text-sm font-semibold text-amber-800')
    pruh.set_visibility(bezi_import(asm))
    ui.timer(1.5, lambda: pruh.set_visibility(bezi_import(asm)))


def _tise(fn):
    """UI update po odchodu ze stránky nesmí shodit import běžící na pozadí."""
    try:
        fn()
    except Exception:
        pass


def _vykresli_import(user_id, user_name, vsechna_prava):
    import intranet_asm

    ui.label('Vyberte jeden nebo více souborů „data prodeje zak3 – *.xlsx". '
             f'Zpracují se paralelně (až {_IMPORT_SOUBEZNE} naráz), '
             'jeden soubor = jedno ASM.') \
        .classes('text-gray-600 mb-2')
    ui.label('Reimport stejného měsíce data přepíše, nový měsíc se přidá.') \
        .classes('text-xs text-gray-400 mb-4')

    protokol = ui.column().classes('w-full max-w-2xl gap-2 mt-4')

    def slot_prubehu(stav, cizi=False):
        """Řádek s progress barem. Timer jen čte stav, který plní import thread."""
        with protokol:
            radek_ui = ui.column().classes('w-full gap-1')
        with radek_ui:
            with ui.row().classes('w-full items-center justify-between gap-3'):
                ui.label(stav['nazev']).classes('text-sm text-gray-700 truncate')
                popis = ui.label('ve frontě…').classes('text-xs text-gray-500 '
                                                       'whitespace-nowrap')
            bar = ui.linear_progress(value=0, show_value=False, size='8px') \
                .props('rounded instant-feedback color=indigo-7')

        def tik():
            if cizi and id(stav) not in _BEZICI_IMPORTY:
                # Import z jiné relace doběhl — výsledek zná jeho vlastní záložka.
                radek_ui.clear()
                with radek_ui:
                    ui.label(f"✅ {stav['nazev']} — import na pozadí dokončen") \
                        .classes('text-sm text-green-700')
                timer.cancel()
                return
            celkem, hotovo = stav['celkem'], stav['hotovo']
            if celkem:
                # Poslední fáze (číselníky, souhrn) běží po dočtení řádků — bar
                # by stál na 100 % a tvářil se zaseknutě. Držíme ho pod stropem.
                bar.value = min(hotovo / celkem, 1.0) * 0.95
                popis.text = f"{stav['faze']} · {bar.value:.0%}"
            else:
                bar.props('indeterminate')
                popis.text = f"{stav['faze']}… {hotovo:,} řádků".replace(',', ' ')

        timer = ui.timer(0.3, tik)
        return radek_ui, timer

    # Import doběhne i po odchodu ze stránky — po návratu ukážeme, co jede dál.
    for bezici in list(_BEZICI_IMPORTY.values()):
        slot_prubehu(bezici, cizi=True)

    async def zpracuj(e):
        # reset=False: reset() abortuje ostatní běžící uploady z téže dávky
        raw, nazev = await intranet_asm._precti_upload(e, up, reset=False)
        if raw is None:
            return
        # Paralelně, ale nejvýš _IMPORT_SOUBEZNE naráz a jedno ASM jen jednou —
        # platí napříč všemi klienty, lokální příznak by hlídal jen jednu záložku.
        klic = klic_ze_souboru(nazev) or nazev
        zamek = _zamek_asm(klic)
        stav = {'faze': 've frontě', 'hotovo': 0, 'celkem': None,
                'nazev': nazev, 'asm': klic}
        _BEZICI_IMPORTY[id(stav)] = stav   # zámek filtrů + stopa pro ostatní relace
        radek_ui, timer = slot_prubehu(stav)
        try:
            async with _IMPORT_SEM, zamek:
                vysledek = await asyncio.to_thread(importuj_soubor, raw, nazev, stav)
        finally:
            _BEZICI_IMPORTY.pop(id(stav), None)
            _tise(timer.cancel)
            _tise(radek_ui.clear)   # bar dosloužil, slot souboru zůstává na místě

        # Od téhle chvíle jen UI — klient mezitím mohl odejít, chyby polykáme.
        if vysledek['chyba']:
            def selhalo():
                with radek_ui:
                    ui.label(f"❌ {nazev}: {vysledek['chyba']}") \
                        .classes('text-sm text-red-600')
                ui.notify(f"{nazev}: {vysledek['chyba']}", type='negative')
            _tise(selhalo)
            return
        zapis_log(user_id, user_name, 'import', vysledek['asm'],
                  f"{nazev} · období {', '.join(vysledek['obdobi'])}",
                  vysledek['radku'])
        await asyncio.to_thread(_ohlas_nesparovane)

        def hotovo():
            with radek_ui:
                ui.label(f"✅ {nazev} → ASM {vysledek['asm']}, "
                         f"{vysledek['radku']:,} řádků, období "
                         f"{', '.join(vysledek['obdobi'])}"
                         .replace(',', ' ')).classes('text-sm text-green-700')
            # Překreslit smí jen karta níž. Refresh celé záložky by zrušil
            # ui.upload a s ním i soubory, které se ještě nahrávají
            # (ClientDisconnect), plus by smazal protokol dokončených importů.
            karta_nesparovane.refresh()
        _tise(hotovo)

    def odmitnuto(e):
        # Bez toho zmizí soubor odmítnutý klientem (velikost, přípona) beze slova.
        ui.notify(f'Soubor odmítnut (velikost / typ): {len(e.args or [])} ks — nahrajte znovu.',
                  type='negative', multi_line=True)

    up = ui.upload(on_upload=zpracuj, on_rejected=odmitnuto, auto_upload=True, multiple=True,
                   max_file_size=120_000_000, label='Vybrat .xlsx (lze více najednou)') \
        .props('accept=.xlsx multiple').classes('w-full max-w-2xl')

    @ui.refreshable
    def karta_nesparovane():
        nesparovane = nesparovana_asm()
        if not nesparovane or not je_admin(vsechna_prava):
            return
        with ui.card().classes('w-full max-w-2xl mt-6 p-4 bg-amber-50 border '
                               'border-amber-200 rounded-xl'):
            ui.label('⚠️ ASM bez spárovaného uživatele').classes('font-bold text-amber-800')
            ui.label(', '.join(nesparovane)).classes('text-sm text-amber-700')
            ui.label('Data se nahrála, ale dlaždici zatím vidí jen admin a extra čtenář. '
                     'Zkontrolujte příjmení uživatele ve Správě uživatelů.') \
                .classes('text-xs text-amber-600 mt-1')

    karta_nesparovane()


def _ohlas_nesparovane():
    nesparovane = nesparovana_asm()
    if not nesparovane:
        return
    zprava = ('⚠️ Lupou na obchod: bez spárovaného uživatele zůstávají ASM '
              + ', '.join(nesparovane))
    for uid in _adminska_id():
        try:
            intranet_notifikace.pridej(uid, zprava, 'warning')
        except Exception:
            pass
