# -*- coding: utf-8 -*-
"""Modul Vizitky a e-mailové podpisy — žádanky o výrobu / změnu.

Workflow:
  • Žadatel (automaticky každý přihlášený uživatel) vyplní formulář (Nová vizitka /
    Nový e-mail podpis) a odešle ho k realizaci. Vidí jen své žádosti.
  • Realizátor / grafik (právo `vizitky_realizator`) vidí všechny žádosti, realizuje je
    (vloží podklady ke stažení) a spravuje číselníky (oddělení / pobočky / adresy).
  • `vse` = admin (dědí obojí).

Stavy žádosti:
  nova → rozpracovano → vyrizeno      a u vyřízené:  vyrizeno → zmena → vyrizeno

Každá žádost má historii (👁 očičko) a e-mailové notifikace na všechny zúčastněné.
Vzory: intranet_helpdesk.py (workflow karet) + intranet_sankce.py (historie/očičko).
"""

from nicegui import ui, app
import intranet_data
import intranet_logger
import intranet_emaily
import datetime
import json
import os
import re
import shutil
import asyncio
import base64
import unicodedata

try:
    import vizitky_seed
except Exception:  # pragma: no cover — bez seedu jen prázdné číselníky
    vizitky_seed = None


# =========================================================
# KONSTANTY
# =========================================================
PODKLADY_DIR = 'Vizitky_Podklady'
PODKLADY_URL = '/vizitky_podklady'
# Podklady zpřístupníme i přes URL (media route) kvůli náhledu v prohlížeči.
try:
    os.makedirs(PODKLADY_DIR, exist_ok=True)
    app.add_media_files(PODKLADY_URL, PODKLADY_DIR)
except Exception as _e:
    print(f'[vizitky] add_media_files: {_e}')

# Statické obrázky pro návod na nastavení podpisu (Outlook + Thunderbird).
NAVOD_DIR = 'Vizitky_Navod'
NAVOD_URL = '/vizitky_navod'
try:
    os.makedirs(NAVOD_DIR, exist_ok=True)
    app.add_media_files(NAVOD_URL, NAVOD_DIR)
except Exception as _e:
    print(f'[vizitky] add_media_files (navod): {_e}')

# Export do Excelu (formulář pro skript grafika)
EXPORT_DIR = 'Vizitky_Export'
SABLONA_NAZEV = 'Formulář pro výrobu vizitek a podpisu.xlsx'   # stejný název jako originál (čte ho skript grafika)
SHEET_NAZEV = 'Formulář pro výrobu vizitek a p'                # název listu jako v originále (max 31 znaků)
SABLONA_SOUBOR = 'vizitky_sablona.xlsx'                        # kopie zaslané šablony — plní se 1:1
QR_CESTA = 'U:\\JIP\\vizitky\\NOVÉ_2024\\QR_code\\'            # pevná cesta v poli QR kód (needitovatelné)

TYP_LABEL = {'email': 'E-mailový podpis', 'vizitka': 'Vizitka'}
TYP_IKONA = {'email': 'alternate_email', 'vizitka': 'badge'}

STAV_LABEL = {
    'nova':         'Nová',
    'rozpracovano': 'Rozpracováno',
    'vyrizeno':     'Vyřízeno',
    'zmena':        'Žádost o změnu',
}
# Badge (štítek stavu)
STAV_BADGE = {
    'nova':         'bg-red-100 text-red-800 border-red-300',
    'rozpracovano': 'bg-amber-100 text-amber-800 border-amber-300',
    'vyrizeno':     'bg-green-100 text-green-800 border-green-300',
    'zmena':        'bg-purple-100 text-purple-800 border-purple-300',
}
# Barevné odlišení karty (levý okraj + pozadí) — vyřízené ztlumené
STAV_KARTA = {
    'nova':         'border-l-4 border-red-400 bg-white',
    'rozpracovano': 'border-l-4 border-amber-400 bg-white',
    'vyrizeno':     'border-l-4 border-green-400 bg-green-50/40 opacity-90',
    'zmena':        'border-l-4 border-purple-500 bg-purple-50/40',
}

# Pole formuláře pro zobrazení v detailu (klíč, popisek, je_bool)
FIELD_DISPLAY = [
    ('organizace',     'Organizace',          False),
    ('jmeno_prijmeni', 'Jméno a příjmení',    False),
    ('oddeleni',       'Oddělení',            False),
    ('telefon',        'Telefonní číslo',     False),
    ('email',          'E-mail',              False),
    ('adresa_ulice',   'Adresa – ulice',      False),
    ('adresa_psc',     'Adresa – PSČ',        False),
    ('adresa_mesto',   'Adresa – město',      False),
    ('web',            'Internetová adresa',  False),
    ('kod_pobocky',    'Kód pobočky',         False),
    ('qr_code',        'QR kód',              False),
    ('email_podpis',   'Příznak: e-mail podpis', True),
    ('vizitka_tisk',   'Příznak: vizitka tisk',  True),
]
# „Zóny" pro žádost o změnu = textová pole (bez booleanů)
ZONY = [(k, lbl) for (k, lbl, je_bool) in FIELD_DISPLAY if not je_bool]
ZONY_LABEL = dict(ZONY)
ZADOST_SLOUPCE = {k for (k, _l, _b) in FIELD_DISPLAY}  # whitelist názvů sloupců

INFO_ZMENA_EMAIL = (
    'Dohledejte prosím svou žádost o e-mailový podpis, ve které chcete udělat změnu, '
    'a na ní klikněte na „Zažádat o změnu". Vyberte, co chcete změnit a za co, a odešlete '
    'žádost ke změně.\n\nPokud jste podpis nezadávali ještě přes tento portál, musíte žádost '
    'vytvořit kompletně znovu a do poznámky uvést „změna".'
)
INFO_ZMENA_VIZITKA = (
    'Dohledejte prosím svou žádost o vizitku, ve které chcete udělat změnu, a na žádosti '
    'klikněte na „Zažádat o změnu". Vyberte, co chcete změnit a za co, a odešlete žádost ke '
    'změně.\n\nPokud jste vizitku nezadávali ještě přes tento portál, musíte žádost vytvořit '
    'kompletně znovu a do poznámky uvést „změna".'
)


# =========================================================
# POMOCNÉ FUNKCE
# =========================================================
def _s(v) -> str:
    return '' if v is None else str(v).strip()


def _safe_filename(s: str) -> str:
    s = unicodedata.normalize('NFKD', str(s or ''))
    s = ''.join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r'[^A-Za-z0-9_.\- ]', '', s).strip().replace(' ', '_')
    return s or 'soubor'


def _cas(dt) -> str:
    return dt.strftime('%d.%m.%Y %H:%M') if hasattr(dt, 'strftime') else _s(dt)


# =========================================================
# INICIALIZACE DATABÁZE + SEED ČÍSELNÍKŮ
# =========================================================
def inicializace_vizitky_db():
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS vizitky_zadosti (
                id INT AUTO_INCREMENT PRIMARY KEY,
                user_id INT,
                jmeno_zadavatele VARCHAR(255),
                email_zadavatele VARCHAR(255),
                typ VARCHAR(20),
                stav VARCHAR(20) DEFAULT 'nova',
                organizace VARCHAR(255),
                jmeno_prijmeni VARCHAR(255),
                oddeleni VARCHAR(255),
                telefon VARCHAR(60),
                email VARCHAR(255),
                adresa_ulice VARCHAR(255),
                adresa_psc VARCHAR(30),
                adresa_mesto VARCHAR(120),
                web VARCHAR(255),
                kod_pobocky VARCHAR(60),
                qr_code VARCHAR(500),
                email_podpis TINYINT(1) DEFAULT 0,
                vizitka_tisk TINYINT(1) DEFAULT 0,
                poznamka TEXT,
                zmena_pozadavky TEXT,
                zmena_poznamka TEXT,
                realizator_poznamka TEXT,
                realizoval_jmeno VARCHAR(255),
                vytvoreno DATETIME DEFAULT CURRENT_TIMESTAMP,
                aktualizovano DATETIME DEFAULT CURRENT_TIMESTAMP,
                vyrizeno_at DATETIME,
                INDEX idx_user (user_id), INDEX idx_stav (stav)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS vizitky_podklady (
                id INT AUTO_INCREMENT PRIMARY KEY,
                zadost_id INT,
                nazev_souboru VARCHAR(255),
                cesta VARCHAR(500),
                nahrano_kym VARCHAR(255),
                nahrano_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_zadost (zadost_id)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS vizitky_log (
                id INT AUTO_INCREMENT PRIMARY KEY,
                zadost_id INT,
                akce VARCHAR(60),
                detail TEXT,
                user_id INT,
                jmeno VARCHAR(255),
                kdy DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_zadost (zadost_id)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        # Číselníky (editovatelné v aplikaci)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS vizitky_oddeleni (
                id INT AUTO_INCREMENT PRIMARY KEY,
                nazev VARCHAR(191) UNIQUE
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS vizitky_pobocky (
                id INT AUTO_INCREMENT PRIMARY KEY,
                kod VARCHAR(60) UNIQUE
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS vizitky_adresy (
                id INT AUTO_INCREMENT PRIMARY KEY,
                ulice VARCHAR(255),
                psc VARCHAR(30),
                mesto VARCHAR(120)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        conn.commit()
        cur.close()
    except Exception as e:
        print(f'Chyba při inicializaci DB Vizitky: {e}')
    finally:
        conn.close()

    _seed_ciselniky()


def _seed_ciselniky():
    """Naplní číselníky ze seed dat POUZE pokud jsou prázdné (první start)."""
    if vizitky_seed is None:
        return
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM vizitky_oddeleni")
        if cur.fetchone()[0] == 0:
            data = [(x,) for x in getattr(vizitky_seed, 'SEED_ODDELENI', [])]
            if data:
                cur.executemany("INSERT IGNORE INTO vizitky_oddeleni (nazev) VALUES (%s)", data)
        cur.execute("SELECT COUNT(*) FROM vizitky_pobocky")
        if cur.fetchone()[0] == 0:
            data = [(x,) for x in getattr(vizitky_seed, 'SEED_POBOCKY', [])]
            if data:
                cur.executemany("INSERT IGNORE INTO vizitky_pobocky (kod) VALUES (%s)", data)
        cur.execute("SELECT COUNT(*) FROM vizitky_adresy")
        if cur.fetchone()[0] == 0:
            data = [(u, p, m) for (u, p, m) in getattr(vizitky_seed, 'SEED_ADRESY', [])]
            if data:
                cur.executemany("INSERT INTO vizitky_adresy (ulice, psc, mesto) VALUES (%s,%s,%s)", data)
        conn.commit()
        cur.close()
    except Exception as e:
        print(f'Chyba při seedu číselníků Vizitky: {e}')
    finally:
        conn.close()


# =========================================================
# ČÍSELNÍKY
# =========================================================
def nacti_oddeleni() -> list:
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor()
        cur.execute("SELECT nazev FROM vizitky_oddeleni ORDER BY nazev ASC")
        return [r[0] for r in cur.fetchall() if r[0]]
    except Exception as e:
        print(f'[vizitky] nacti_oddeleni: {e}')
        return []
    finally:
        conn.close()


def nacti_pobocky() -> list:
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor()
        cur.execute("SELECT kod FROM vizitky_pobocky ORDER BY kod ASC")
        return [r[0] for r in cur.fetchall() if r[0]]
    except Exception as e:
        print(f'[vizitky] nacti_pobocky: {e}')
        return []
    finally:
        conn.close()


def nacti_adresy() -> list:
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT id, ulice, psc, mesto FROM vizitky_adresy ORDER BY mesto, ulice")
        return cur.fetchall()
    except Exception as e:
        print(f'[vizitky] nacti_adresy: {e}')
        return []
    finally:
        conn.close()


def _ciselnik_pridej(tabulka: str, sloupce: tuple, hodnoty: tuple) -> tuple:
    conn = intranet_data.get_db_connection()
    if not conn:
        return False, 'Chyba připojení k databázi.'
    try:
        cur = conn.cursor()
        col = ','.join(sloupce)
        ph = ','.join(['%s'] * len(hodnoty))
        cur.execute(f"INSERT IGNORE INTO {tabulka} ({col}) VALUES ({ph})", hodnoty)
        conn.commit()
        cur.close()
        return True, None
    except Exception as e:
        return False, f'Chyba zápisu: {e}'
    finally:
        conn.close()


def _ciselnik_smaz(tabulka: str, radek_id: int) -> bool:
    conn = intranet_data.get_db_connection()
    if not conn:
        return False
    try:
        cur = conn.cursor()
        cur.execute(f"DELETE FROM {tabulka} WHERE id=%s", (radek_id,))
        conn.commit()
        cur.close()
        return True
    except Exception as e:
        print(f'[vizitky] smaz ciselnik: {e}')
        return False
    finally:
        conn.close()


def _smaz_oddeleni_dle_nazvu(nazev: str):
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM vizitky_oddeleni WHERE nazev=%s", (nazev,))
        conn.commit()
        cur.close()
    finally:
        conn.close()


def _smaz_pobocku_dle_kodu(kod: str):
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM vizitky_pobocky WHERE kod=%s", (kod,))
        conn.commit()
        cur.close()
    finally:
        conn.close()


# =========================================================
# ŽÁDOSTI
# =========================================================
def nacti_zadosti(user_id=None) -> list:
    """user_id=None → všechny (realizátor); jinak jen žádosti daného žadatele."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        if user_id is None:
            cur.execute(
                "SELECT * FROM vizitky_zadosti ORDER BY "
                "CASE stav WHEN 'nova' THEN 1 WHEN 'zmena' THEN 2 "
                "WHEN 'rozpracovano' THEN 3 ELSE 4 END, vytvoreno DESC")
        else:
            cur.execute("SELECT * FROM vizitky_zadosti WHERE user_id=%s ORDER BY vytvoreno DESC",
                        (user_id,))
        return cur.fetchall()
    finally:
        conn.close()


def nacti_zadost(zadost_id: int):
    conn = intranet_data.get_db_connection()
    if not conn:
        return None
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT * FROM vizitky_zadosti WHERE id=%s", (zadost_id,))
        return cur.fetchone()
    finally:
        conn.close()


def zaloz_zadost(data: dict) -> int:
    conn = intranet_data.get_db_connection()
    if not conn:
        return 0
    try:
        cur = conn.cursor()
        sloupce = ['user_id', 'jmeno_zadavatele', 'email_zadavatele', 'typ', 'stav',
                   'organizace', 'jmeno_prijmeni', 'oddeleni', 'telefon', 'email',
                   'adresa_ulice', 'adresa_psc', 'adresa_mesto', 'web', 'kod_pobocky',
                   'qr_code', 'email_podpis', 'vizitka_tisk', 'poznamka']
        ph = ','.join(['%s'] * len(sloupce))
        hodnoty = tuple(data.get(s) for s in sloupce)
        cur.execute(f"INSERT INTO vizitky_zadosti ({','.join(sloupce)}) VALUES ({ph})", hodnoty)
        conn.commit()
        nove_id = cur.lastrowid
        cur.close()
        return nove_id
    except Exception as e:
        print(f'[vizitky] zaloz_zadost: {e}')
        return 0
    finally:
        conn.close()


def oznac_rozpracovano(zadost_id: int):
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.execute("UPDATE vizitky_zadosti SET stav='rozpracovano', aktualizovano=NOW() "
                    "WHERE id=%s AND stav='nova'", (zadost_id,))
        conn.commit()
        cur.close()
    finally:
        conn.close()


def uloz_zadost_o_zmenu(zadost_id: int, zmeny: dict, poznamka: str):
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.execute(
            "UPDATE vizitky_zadosti SET stav='zmena', zmena_pozadavky=%s, zmena_poznamka=%s, "
            "aktualizovano=NOW() WHERE id=%s",
            (json.dumps(zmeny, ensure_ascii=False), poznamka, zadost_id))
        conn.commit()
        cur.close()
    finally:
        conn.close()


def dokonci_realizaci(zadost_id: int, realizator_poznamka: str, realizoval_jmeno: str,
                      aplikuj_zmeny=None):
    """Nastaví stav 'vyrizeno'. Pokud aplikuj_zmeny, zapíše nové hodnoty do polí
    žádosti a vyčistí požadavek na změnu (whitelist sloupců)."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        set_casti = ["stav='vyrizeno'", "realizator_poznamka=%s", "realizoval_jmeno=%s",
                     "vyrizeno_at=NOW()", "aktualizovano=NOW()"]
        hodnoty = [realizator_poznamka, realizoval_jmeno]
        if aplikuj_zmeny:
            for k, v in aplikuj_zmeny.items():
                if k in ZADOST_SLOUPCE:
                    set_casti.append(f"{k}=%s")
                    hodnoty.append(v)
            set_casti.append("zmena_pozadavky=NULL")
        hodnoty.append(zadost_id)
        cur.execute(f"UPDATE vizitky_zadosti SET {', '.join(set_casti)} WHERE id=%s", tuple(hodnoty))
        conn.commit()
        cur.close()
    finally:
        conn.close()


def uloz_upravu_zadosti(zadost_id: int, data: dict) -> bool:
    """Realizátor opraví obsah žádosti (whitelist polí). Stav žádosti nemění."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return False
    try:
        cur = conn.cursor()
        set_casti, hodnoty = [], []
        for k, v in data.items():
            if k in ZADOST_SLOUPCE:
                set_casti.append(f"{k}=%s")
                hodnoty.append(v)
        if not set_casti:
            return False
        set_casti.append("aktualizovano=NOW()")
        hodnoty.append(zadost_id)
        cur.execute(f"UPDATE vizitky_zadosti SET {', '.join(set_casti)} WHERE id=%s", tuple(hodnoty))
        conn.commit()
        cur.close()
        return True
    except Exception as e:
        print(f'[vizitky] uloz_upravu_zadosti: {e}')
        return False
    finally:
        conn.close()


def smaz_zadost(zadost_id: int) -> bool:
    """Nevratně smaže žádost: soubory podkladů z disku + řádky v podkladech, logu a žádosti."""
    try:
        slozka = os.path.join(PODKLADY_DIR, str(zadost_id))
        if os.path.isdir(slozka):
            shutil.rmtree(slozka, ignore_errors=True)
    except Exception as e:
        print(f'[vizitky] smaz_zadost soubory: {e}')

    conn = intranet_data.get_db_connection()
    if not conn:
        return False
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM vizitky_podklady WHERE zadost_id=%s", (zadost_id,))
        cur.execute("DELETE FROM vizitky_log WHERE zadost_id=%s", (zadost_id,))
        cur.execute("DELETE FROM vizitky_zadosti WHERE id=%s", (zadost_id,))
        conn.commit()
        cur.close()
        return True
    except Exception as e:
        print(f'[vizitky] smaz_zadost: {e}')
        return False
    finally:
        conn.close()


# =========================================================
# LOG (očičko)
# =========================================================
def zapis_log(zadost_id: int, akce: str, detail: str, user_id, jmeno: str):
    conn = intranet_data.get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO vizitky_log (zadost_id, akce, detail, user_id, jmeno) "
            "VALUES (%s,%s,%s,%s,%s)",
            (zadost_id, akce, detail, user_id, jmeno or ''))
        conn.commit()
        cur.close()
    except Exception as e:
        print(f'[vizitky] zapis_log: {e}')
    finally:
        conn.close()


def nacti_log(zadost_id: int) -> list:
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT akce, detail, jmeno, kdy FROM vizitky_log "
                    "WHERE zadost_id=%s ORDER BY kdy DESC, id DESC", (zadost_id,))
        return cur.fetchall()
    finally:
        conn.close()


# =========================================================
# PODKLADY (soubory od grafika)
# =========================================================
def uloz_podklad(zadost_id: int, nazev: str, raw: bytes, kym: str) -> bool:
    try:
        slozka = os.path.join(PODKLADY_DIR, str(zadost_id))
        os.makedirs(slozka, exist_ok=True)
        stamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        cesta = os.path.join(slozka, f"{stamp}_{_safe_filename(nazev)}")
        with open(cesta, 'wb') as f:
            f.write(raw)
    except Exception as e:
        print(f'[vizitky] uloz_podklad zápis: {e}')
        return False

    conn = intranet_data.get_db_connection()
    if not conn:
        return False
    try:
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO vizitky_podklady (zadost_id, nazev_souboru, cesta, nahrano_kym) "
            "VALUES (%s,%s,%s,%s)", (zadost_id, nazev, cesta, kym))
        conn.commit()
        cur.close()
        return True
    except Exception as e:
        print(f'[vizitky] uloz_podklad DB: {e}')
        return False
    finally:
        conn.close()


def nacti_podklady(zadost_id: int) -> list:
    conn = intranet_data.get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT id, nazev_souboru, cesta, nahrano_kym, nahrano_at "
                    "FROM vizitky_podklady WHERE zadost_id=%s ORDER BY nahrano_at ASC", (zadost_id,))
        return cur.fetchall()
    finally:
        conn.close()


def _podklad_url(cesta: str) -> str:
    """URL pro náhled podkladu v prohlížeči (přes media route)."""
    try:
        rel = os.path.relpath(cesta, PODKLADY_DIR).replace('\\', '/')
        return f'{PODKLADY_URL}/{rel}'
    except Exception:
        return ''


def smaz_podklad(podklad_id: int) -> bool:
    """Smaže jeden podklad (soubor z disku + řádek v DB)."""
    conn = intranet_data.get_db_connection()
    if not conn:
        return False
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT cesta FROM vizitky_podklady WHERE id=%s", (podklad_id,))
        r = cur.fetchone()
        if r and r.get('cesta'):
            try:
                if os.path.isfile(r['cesta']):
                    os.remove(r['cesta'])
            except Exception as e:
                print(f'[vizitky] smaz_podklad soubor: {e}')
        cur.execute("DELETE FROM vizitky_podklady WHERE id=%s", (podklad_id,))
        conn.commit()
        cur.close()
        return True
    except Exception as e:
        print(f'[vizitky] smaz_podklad: {e}')
        return False
    finally:
        conn.close()


# =========================================================
# E-MAILY (best-effort, na pozadí)
# =========================================================
def _emaily_realizatoru() -> list:
    # EFEKTIVNÍ práva (vč. zděděných z oddělení/pozice) — viz
    # intranet_data.ziskej_emaily_s_pravem. Dřív se bralo ziskej_vsechny_uzivatele()
    # ['prava'] = jen OSOBNÍ práva → komu byl 'vizitky_realizator' zděděný z oddělení,
    # tomu notifikace NECHODILY.
    try:
        return list(dict.fromkeys(
            intranet_data.ziskej_emaily_s_pravem('vizitky_realizator', 'vse')))
    except Exception as e:
        print(f'[vizitky] _emaily_realizatoru: {e}')
        return []


def _app_url() -> str:
    """Plný odkaz na modul Vizitky (z env JIPKA_APP_URL); '' pokud není nastaven."""
    try:
        u = intranet_data.APP_URL
        return f'{u}/vizitky' if u else ''
    except Exception:
        return ''


def _email_html(text: str, odkaz: str) -> str:
    """Obalí prostý text do HTML s klikacím tlačítkem na modul (je-li odkaz znám)."""
    import html as _h
    telo = _h.escape(text).replace('\n', '<br>')
    tlacitko = ''
    if odkaz:
        tlacitko = (
            f'<div style="margin:26px 0"><a href="{odkaz}" '
            'style="background:#0284c7;color:#ffffff;text-decoration:none;font-weight:bold;'
            'padding:12px 24px;border-radius:8px;display:inline-block">'
            '🪪 Otevřít v portálu MOJEJIPka</a></div>'
            f'<div style="font-size:12px;color:#888">Nefunguje tlačítko? Otevřete: '
            f'<a href="{odkaz}" style="color:#0284c7">{odkaz}</a></div>'
        )
    return (
        '<!DOCTYPE html><html><body style="font-family:Arial,sans-serif;font-size:14px;'
        'color:#222;line-height:1.6;max-width:620px;margin:0 auto;padding:24px">'
        f'{telo}{tlacitko}'
        '<hr style="border:none;border-top:1px solid #eee;margin:24px 0">'
        '<div style="font-size:12px;color:#999">Automatická zpráva z portálu MOJEJIPka — '
        'modul „Vizitky a podpisy".</div>'
        '</body></html>'
    )


def _posli_email_sync(prijemci, predmet, html_obsah):
    for p in prijemci:
        try:
            intranet_emaily.odesli_html_email(p, predmet, html_obsah)
        except Exception as e:
            print(f'[vizitky] e-mail {p}: {e}')


def _odesli_emaily(prijemci, predmet, text):
    prijemci = [p for p in dict.fromkeys(prijemci) if p and '@' in p]
    if not prijemci:
        return
    html_obsah = _email_html(text, _app_url())
    try:
        asyncio.create_task(asyncio.to_thread(_posli_email_sync, prijemci, predmet, html_obsah))
    except RuntimeError:
        _posli_email_sync(prijemci, predmet, html_obsah)


# =========================================================
# EXPORT DO EXCELU (formulář pro skript grafika)
# =========================================================
def vygeneruj_excel(z: dict) -> str:
    """Naplní KOPII zaslané šablony (1:1 – stejný list, styl i sloupce) daty žádosti
    do 1. datového řádku a zafajfkuje „email podpis" / „vizitka tisk".
    Vrací cestu k souboru na serveru (ke stažení se použije pevný název `SABLONA_NAZEV`).

    Mapování sloupců (dle šablony): A organizace · B Jméno a Příjmení (jedna buňka) ·
    C oddělení · D telefon · E e-mail · F ulice · G PSČ · H město · I web ·
    J kód pobočky · K @QR_Code · L email podpis · M vizitka tisk."""
    import openpyxl, warnings
    with warnings.catch_warnings():
        # Šablona má pokročilé „data validation" (rozbalovací seznamy), které openpyxl neumí
        # přenést. Je to neškodné — jen potlačíme varování, ať nezahlcuje log serveru.
        warnings.filterwarnings('ignore', message='Data Validation extension is not supported')
        wb = openpyxl.load_workbook(SABLONA_SOUBOR)
    ws = wb[SHEET_NAZEV] if SHEET_NAZEV in wb.sheetnames else wb.worksheets[0]
    R = 2  # první datový řádek pod hlavičkou
    ws.cell(row=R, column=1,  value=_s(z.get('organizace')))
    ws.cell(row=R, column=2,  value=_s(z.get('jmeno_prijmeni')))   # 1:1 – jedna buňka
    ws.cell(row=R, column=3,  value=_s(z.get('oddeleni')))
    ws.cell(row=R, column=4,  value=_s(z.get('telefon')))
    ws.cell(row=R, column=5,  value=_s(z.get('email')))
    ws.cell(row=R, column=6,  value=_s(z.get('adresa_ulice')))
    ws.cell(row=R, column=7,  value=_s(z.get('adresa_psc')))
    ws.cell(row=R, column=8,  value=_s(z.get('adresa_mesto')))
    ws.cell(row=R, column=9,  value=_s(z.get('web')))
    ws.cell(row=R, column=10, value=_s(z.get('kod_pobocky')))
    if _s(z.get('qr_code')):                                       # jinak ponecháme cestu ze šablony
        ws.cell(row=R, column=11, value=_s(z.get('qr_code')))
    ws.cell(row=R, column=12, value=bool(z.get('email_podpis')))   # fajfka „email podpis"
    ws.cell(row=R, column=13, value=bool(z.get('vizitka_tisk')))   # fajfka „vizitka tisk"

    os.makedirs(EXPORT_DIR, exist_ok=True)
    cesta = os.path.join(EXPORT_DIR, f'vizitky_zadost_{z.get("id")}.xlsx')
    wb.save(cesta)
    return cesta


async def _stahni_excel(z: dict):
    try:
        cesta = await asyncio.to_thread(vygeneruj_excel, z)
        ui.download.file(cesta, SABLONA_NAZEV)
    except Exception as e:
        ui.notify(f'Nepodařilo se vytvořit Excel: {e}', type='negative')


# =========================================================
# GENEROVÁNÍ E-MAILOVÉHO PODPISU (.htm + .txt)
# Logika 1:1 přejata ze skriptu grafika (původně běžel ručně v U:\JIP\vizitky\…).
# Modul ji nyní vytvoří přímo z dat žádosti — bez Excelu a síťového disku.
# =========================================================
_TITULY = {
    'Bc.', 'BcA.', 'Ing.', 'Mgr.', 'MgA.', 'PhDr.', 'JUDr.', 'RNDr.',
    'MUDr.', 'MDDr.', 'MVDr.', 'PharmDr.', 'ThDr.',
    'prof.', 'doc.', 'Dr.', 'PaedDr.', 'Ph.D.', 'Th.D.',
}


def _format_phone(phone: str) -> str:
    """'722661659' / '+420 722 661 659' / '00420 722-661-659' → '+420 722 661 659'."""
    if not phone:
        return ''
    digits = re.sub(r'\D', '', phone)
    if digits.startswith('00420'):
        digits = digits[5:]
    elif digits.startswith('420'):
        digits = digits[3:]
    if not digits:
        return ''
    groups = [digits[i:i + 3] for i in range(0, len(digits), 3)]
    return '+420 ' + ' '.join(groups)


def _split_name(full_name: str):
    """Rozdělí jméno s tituly: 'Bc. Dávid Deák' → ('Bc. Dávid', 'DEÁK');
    'Jana Nováková, Ph.D.' → ('Jana', 'NOVÁKOVÁ, Ph.D.'). Příjmení velkými písmeny."""
    if not full_name:
        return '', ''
    parts = full_name.strip().replace(',', '').split()
    titul_pred = titul_za = ''
    if parts and parts[0] in _TITULY:
        titul_pred = parts[0]
        parts = parts[1:]
    if parts and parts[-1].rstrip(',') in _TITULY:
        titul_za = parts[-1].rstrip(',')
        parts = parts[:-1]
    if not parts:
        return full_name, ''
    if len(parts) == 1:
        first, last = parts[0].capitalize(), ''
    else:
        first = parts[0].capitalize()
        last = ' '.join(parts[1:]).upper()
    if titul_pred:
        first = f'{titul_pred} {first}'
    if titul_za:
        last = f'{last}, {titul_za}'
    return first, last


def _escape_entities(text: str) -> str:
    """Všechny znaky na HTML entity (kompatibilita se starými Outlook klienty)."""
    return ''.join(f'&#{ord(c)};' for c in text) if text else ''


_TXT_JIP = (
    "{jmeno}\n{funkce}\n{telefon}\n{email}\n\n"
    "JIP východočeská, a.s.\n{ulice}\n{psc} {mesto}\nČeská republika\nIČ: 27464822\n\n"
    "www.jip-potraviny.cz\n"
)
_TXT_PLUS = (
    "{jmeno}\n{funkce}\n\n{telefon}\n{email}\n\n"
    "PLUS JIP s.r.o.\nHrádecká 1260\n342 01 Sušice, Česká republika\nIČ: 45353450\n\n"
    "Kancelář: Hradišťská 407\n533 52 Staré Hradiště, Pardubice\n\n"
    "www.plus-potraviny.cz\nČlen skupiny JIP.\n"
)
_HTML_JIP = """<meta charset="UTF-8">
<table style="width:620px; font-family:Calibri,sans-serif; font-size:11pt; color:#000;">
  <tr>
    <td style="width:170px; border-right:1px solid #000; text-align:center;">
      <a href="http://www.jip-potraviny.cz/"><img border="0" width="120" height="60" src="http://www2.jip-potraviny.cz/podpis/logo2.png" alt="JIP logo"></a><br>
      <div style="font-size:12pt; text-align:center; margin-top:4px;"><b>JIP. J&#237;dlo i Pit&#237;</b><br><b>na doma i pro prof&#237;ky.</b></div><br>
      <a href="http://www.jip-potraviny.cz/"><img width="23" src="http://www2.jip-potraviny.cz/podpis/ww.png"></a>
      <a href="http://www.facebook.com/jip.potraviny"><img width="23" src="http://www2.jip-potraviny.cz/podpis/fb.png"></a>
      <a href="http://www.instagram.com/jip_potraviny"><img width="23" src="http://www2.jip-potraviny.cz/podpis/ig.png"></a>
      <a href="https://www.linkedin.com/company/jip-skupina/"><img width="23" src="http://www2.jip-potraviny.cz/podpis/in.png"></a>
      <a href="https://www.youtube.com/channel/UCOzCMaLtLMxlIw2iRDoiTnA/"><img width="23" src="http://www2.jip-potraviny.cz/podpis/yt.png"></a><br>
      <div style="font-size:11pt; font-weight:bold; text-align:center; margin-top:4px;">
        <a href="http://www.jip-potraviny.cz" style="color:#000; text-decoration:underline;">www.jip-potraviny.cz</a>
      </div>
    </td>
    <td style="width:420px; padding-left:20px;">
      <b style="font-size:13pt;">{jmeno} {prijmeni}</b><br>
      <span style="font-size:10pt;">{funkce}</span><br><br>
      <a href="tel:{telefon}" style="color:#0563C1; text-decoration:underline;">{telefon}</a><br>
      <a href="mailto:{email}" style="color:#0563C1; text-decoration:underline;">{email}</a><br><br>
      JIP v&#253;chodo&#269;esk&#225;, a.s.<br>
      {ulice}<br>
      {psc} {mesto}<br>
      &#268;esk&#225; republika<br>
      I&#268;: 27464822
    </td>
  </tr>
  <tr><td colspan="2" style="text-align:center; padding-top:10px;">
    <a href="http://www2.jip-potraviny.cz/podpis/odkaz-PHA.php"><img border="0" src="http://www2.jip-potraviny.cz/podpis/JIP-PHA.png"></a>
  </td></tr>
</table>
"""
_HTML_PLUS = """<meta charset="UTF-8">
<table style="width:620px; font-family:Calibri,sans-serif; font-size:11pt; color:#000;">
  <tr>
    <td style="width:170px; border-right:1px solid #000; text-align:center;">
      <a href="https://www.plus-potraviny.cz/"><img border="0" width="120" height="22" src="http://www2.jip-potraviny.cz/podpis/PLUS-logo.png" alt="PLUS logo"></a><br>
      <div style="font-size:11pt; font-weight:bold; text-align:center; margin-top:4px;">
        <a href="https://www.plus-potraviny.cz" style="color:#000; text-decoration:underline;">www.plus-potraviny.cz</a>
      </div>
    </td>
    <td style="width:420px; padding-left:20px;">
      <b style="font-size:13pt;">{jmeno} {prijmeni}</b><br>
      <span style="font-size:10pt;">{funkce}</span><br><br>
      <a href="tel:{telefon}" style="color:#0563C1; text-decoration:underline;">{telefon}</a><br>
      <a href="mailto:{email}" style="color:#0563C1; text-decoration:underline;">{email}</a><br><br>
      PLUS JIP s.r.o.<br>
      Hr&#225;deck&#225; 1260<br>
      342 01 Su&#353;ice, &#268;esk&#225; republika<br>
      I&#268;: 45353450<br><br>
      Kancel&#225;&#345;: Hradi&#353;t&#353;k&#225; 407<br>
      533 52 Star&#233; Hradi&#353;t&#283;, Pardubice
    </td>
  </tr>
  <tr><td colspan="2" style="text-align:center; padding-top:10px;">
    <a href="https://www.plus-potraviny.cz/"><img border="0" src="http://www2.jip-potraviny.cz/podpis/PLUS.png" alt="PLUS banner"></a>
  </td></tr>
</table>
"""


def vygeneruj_podpis(z: dict) -> dict:
    """Sestaví e-mailový podpis z dat žádosti. Vrací {'jmeno','prijmeni','txt','html','fname_base','je_plus'}.
    Rozhodnutí JIP × PLUS je dle `kod_pobocky` (PLUS = PLUS šablona, jinak JIP)."""
    jmeno, prijmeni = _split_name(_s(z.get('jmeno_prijmeni')))
    funkce = _s(z.get('oddeleni'))
    telefon = _format_phone(_s(z.get('telefon')))
    email = _s(z.get('email'))
    ulice = _s(z.get('adresa_ulice'))
    psc = _s(z.get('adresa_psc'))
    mesto = _s(z.get('adresa_mesto'))
    je_plus = _s(z.get('kod_pobocky')).upper() == 'PLUS'

    if je_plus:
        txt = _TXT_PLUS.format(jmeno=f'{jmeno} {prijmeni}'.strip(), funkce=funkce,
                               telefon=telefon, email=email)
        html_obsah = _HTML_PLUS.format(
            jmeno=_escape_entities(jmeno), prijmeni=_escape_entities(prijmeni),
            funkce=_escape_entities(funkce), telefon=_escape_entities(telefon),
            email=_escape_entities(email))
    else:
        txt = _TXT_JIP.format(jmeno=f'{jmeno} {prijmeni}'.strip(), funkce=funkce,
                              telefon=telefon, email=email,
                              ulice=ulice, psc=psc, mesto=mesto)
        html_obsah = _HTML_JIP.format(
            jmeno=_escape_entities(jmeno), prijmeni=_escape_entities(prijmeni),
            funkce=_escape_entities(funkce), telefon=_escape_entities(telefon),
            email=_escape_entities(email), ulice=_escape_entities(ulice),
            psc=_escape_entities(psc), mesto=_escape_entities(mesto))

    fname_base = _safe_filename(f'{jmeno}_{prijmeni}'.strip('_'))
    return {'jmeno': jmeno, 'prijmeni': prijmeni, 'txt': txt, 'html': html_obsah,
            'fname_base': fname_base, 'je_plus': je_plus}


def uloz_podpis_do_podkladu(z: dict, kym: str) -> tuple:
    """Vygeneruje podpis a uloží `.htm` + `.txt` přímo do podkladů žádosti.
    Vrací (htm_nazev, txt_nazev) nebo (None, None)."""
    p = vygeneruj_podpis(z)
    htm_name = f'{p["fname_base"]}.htm'
    txt_name = f'{p["fname_base"]}.txt'
    ok1 = uloz_podklad(z['id'], htm_name, p['html'].encode('utf-8'), kym)
    ok2 = uloz_podklad(z['id'], txt_name, p['txt'].encode('utf-8'), kym)
    if ok1 and ok2:
        return htm_name, txt_name
    return None, None


# =========================================================
# DIALOG: NÁHLED + ULOŽENÍ PODPISU (realizátor)
# =========================================================
def _dialog_podpis(z, user_id, user_name, prekresli):
    p = vygeneruj_podpis(z)

    with ui.dialog() as dlg, ui.card().classes('w-full max-w-4xl p-0 rounded-xl overflow-hidden bg-gray-50'):
        with ui.row().classes('w-full items-center justify-between px-6 py-4 bg-blue-600'):
            with ui.row().classes('items-center gap-2'):
                ui.icon('alternate_email', color='white')
                ui.label(f'E-mailový podpis #{z["id"]} — {"PLUS" if p["je_plus"] else "JIP"} šablona') \
                    .classes('text-xl font-bold text-white')
            ui.button(icon='close', on_click=dlg.close).props('flat round dense').classes('text-white')

        with ui.column().classes('w-full p-6 gap-4').style('max-height: 65vh; overflow-y: auto'):
            ui.label('Náhled HTML podpisu').classes('text-xs font-bold text-gray-500 uppercase')
            with ui.card().classes('w-full p-4 bg-white rounded-lg border border-gray-200'):
                ui.html(p['html'])

            ui.label('Textová verze (.txt)').classes('text-xs font-bold text-gray-500 uppercase mt-2')
            with ui.card().classes('w-full p-4 bg-gray-100 rounded-lg border border-gray-200'):
                ui.label(p['txt']).classes('text-sm text-gray-800 whitespace-pre-wrap font-mono')

            with ui.row().classes('w-full text-xs text-gray-500 italic mt-1'):
                ui.label(f'Soubory budou pojmenovány: {p["fname_base"]}.htm a {p["fname_base"]}.txt')

        with ui.row().classes('w-full justify-end gap-2 px-6 py-3 bg-white border-t'):
            def stahni_htm():
                tmp = os.path.join(EXPORT_DIR, f'{p["fname_base"]}.htm')
                os.makedirs(EXPORT_DIR, exist_ok=True)
                with open(tmp, 'w', encoding='utf-8') as f:
                    f.write(p['html'])
                ui.download.file(tmp, f'{p["fname_base"]}.htm')

            def stahni_txt():
                tmp = os.path.join(EXPORT_DIR, f'{p["fname_base"]}.txt')
                os.makedirs(EXPORT_DIR, exist_ok=True)
                with open(tmp, 'w', encoding='utf-8') as f:
                    f.write(p['txt'])
                ui.download.file(tmp, f'{p["fname_base"]}.txt')

            async def uloz_do_zadosti():
                htm, txt = await asyncio.to_thread(uloz_podpis_do_podkladu, z, user_name)
                if htm and txt:
                    zapis_log(z['id'], 'Vygenerován podpis', f'{htm}, {txt}', user_id, user_name)
                    intranet_logger.log_activity(user_name, 'Vizitky',
                                                 f'Vygenerován e-mailový podpis #{z["id"]}')
                    ui.notify('Podpis vložen do podkladů žádosti.', type='positive', position='top')
                    dlg.close()
                    prekresli()
                else:
                    ui.notify('Uložení podpisu se nezdařilo.', type='negative')

            ui.button('Stáhnout .htm', icon='download', on_click=stahni_htm) \
                .props('outline no-caps color=primary')
            ui.button('Stáhnout .txt', icon='download', on_click=stahni_txt) \
                .props('outline no-caps color=primary')
            ui.button('Přidat do podkladů žádosti', icon='save', on_click=uloz_do_zadosti) \
                .props('unelevated no-caps').classes('bg-green-600 hover:bg-green-700 text-white font-bold')
            ui.button('Zavřít', on_click=dlg.close).props('flat no-caps').classes('text-gray-600')
    dlg.open()


# =========================================================
# DIALOG: NÁVOD NA NASTAVENÍ PODPISU (Outlook + Thunderbird)
# =========================================================
# =========================================================
# NÁVOD → PDF (server-side render přes Playwright/Chromium)
# Stejný vzor jako tisk v intranet_sankce.py: max 1 Chromium naráz (semafor),
# obrázky vložené jako data: URI (set_content nemá base URL, jinak by se nenačetly).
# =========================================================
_navod_render_lock = None


def _get_navod_render_lock() -> asyncio.Semaphore:
    """Globální semafor — serializuje render (lazy init kvůli importu bez event loopu)."""
    global _navod_render_lock
    if _navod_render_lock is None:
        _navod_render_lock = asyncio.Semaphore(1)
    return _navod_render_lock


def _navod_img_data_uri(filename: str) -> str:
    """Načte obrázek návodu z disku a vrátí ho jako data: URI pro vložení do PDF."""
    try:
        with open(os.path.join(NAVOD_DIR, filename), 'rb') as f:
            raw = f.read()
        ext = os.path.splitext(filename)[1].lower().lstrip('.')
        mime = 'image/jpeg' if ext in ('jpg', 'jpeg') else f'image/{ext}'
        return f'data:{mime};base64,' + base64.b64encode(raw).decode('ascii')
    except Exception as e:
        print(f'[vizitky] _navod_img_data_uri {filename}: {e}')
        return ''


def _navod_pdf_html() -> str:
    """Samostatné (offline) HTML návodu k renderu do PDF — stejný obsah jako dialog,
    obrázky vložené napevno, aby je Chromium zobrazil i bez serveru."""
    img1 = _navod_img_data_uri('image1.png')
    img2 = _navod_img_data_uri('image2.png')
    img3 = _navod_img_data_uri('image3.png')

    def _img(uri, maxw):
        return (f'<img src="{uri}" style="display:block;width:100%;max-width:{maxw};'
                'height:auto;margin:10px auto;border:1px solid #d1d5db;border-radius:6px">'
                ) if uri else ''

    return (
        '<!DOCTYPE html><html lang="cs"><head><meta charset="utf-8">'
        '<title>Návod na nastavení podpisu</title><style>'
        'body{font-family:Arial,Helvetica,sans-serif;color:#1f2937;font-size:12pt;'
        'line-height:1.5;margin:0}'
        'h1{font-size:21pt;color:#1e3a8a;margin:0 0 6px}'
        'h2{font-size:16pt;color:#1e3a8a;border-bottom:2px solid #bfdbfe;'
        'padding-bottom:5px;margin:24px 0 10px}'
        '.lead{color:#374151;margin:0 0 6px}'
        '.klient{page-break-inside:avoid}'
        '.krok{display:flex;gap:10px;margin:9px 0;page-break-inside:avoid}'
        '.cislo{font-weight:bold;color:#1d4ed8;flex:0 0 22px}'
        '.kod{font-family:Consolas,monospace;font-size:10.5pt;background:#f3f4f6;'
        'border:1px solid #d1d5db;border-radius:4px;padding:2px 6px;display:inline-block;margin-top:3px}'
        '.key{font-family:Consolas,monospace;font-size:10.5pt;border:1px solid #9ca3af;'
        'border-radius:4px;background:#f3f4f6;padding:1px 6px;margin:0 2px}'
        '</style></head><body>'
        '<h1>Návod na nastavení podpisu</h1>'
        '<p class="lead">Tento návod obsahuje postupy pro dva e-mailové klienty — '
        'Microsoft Outlook a Mozilla Thunderbird. Postupujte podle návodu pro Váš '
        'e-mailový klient.</p>'
        # ── Microsoft Outlook ──
        '<div class="klient"><h2>Microsoft Outlook</h2>'
        '<div class="krok"><div class="cislo">1.</div><div>Stiskněte klávesovou zkratku '
        '<span class="key">⊞ Win</span> + <span class="key">R</span> na klávesnici.</div></div>'
        '<div class="krok"><div class="cislo">2.</div><div>Do okna Spustit vložte tuto '
        'cestu a klikněte na OK (otevře se Vám příslušná složka):<br>'
        '<span class="kod">%userprofile%\\appdata\\Roaming\\Microsoft\\Signatures</span></div></div>'
        '<div class="krok"><div class="cislo">3.</div><div>Vložte obě přílohy do této '
        'složky.</div></div>'
        '<div class="krok"><div class="cislo">4.</div><div>Restartujte Outlook. Podpis '
        'bude nyní dostupný v nabídce „Podpis" v horní liště nového mailu, kde ho můžete '
        'použít na všechny nové zprávy, případně i odpovědi.</div></div>'
        + _img(img1, '520px') + '</div>'
        # ── Mozilla Thunderbird ──
        '<div class="klient"><h2>Mozilla Thunderbird</h2>'
        '<div class="krok"><div class="cislo">1.</div><div>Uložte pouze soubor z přílohy '
        'tohoto mailu s koncovkou .htm do libovolného umístění v počítači.</div></div>'
        '<div class="krok"><div class="cislo">2.</div><div>V Thunderbirdu klikněte pravým '
        'tlačítkem na Váš účet a vyberte možnost „Nastavení".' + _img(img2, '300px') + '</div></div>'
        '<div class="krok"><div class="cislo">3.</div><div>V novém okně zaškrtněte '
        '„Připojit podpis ze souboru (text, HTML nebo obrázek)".</div></div>'
        '<div class="krok"><div class="cislo">4.</div><div>Tlačítkem „Vybrat…" vyberte '
        'soubor .htm z umístění dle bodu č. 1.' + _img(img3, '460px') + '</div></div>'
        '<div class="krok"><div class="cislo">5.</div><div><b>Hotovo, podpis se bude '
        'používat na všechny nové e-maily.</b></div></div>'
        '</div></body></html>'
    )


async def _vyrenderuj_navod_pdf() -> str:
    """Vyrenderuje návod do PDF (Playwright/Chromium) a vrátí cestu k souboru na serveru.
    Render serializuje semafor (max 1 Chromium naráz). RuntimeError, když chybí Playwright."""
    try:
        from playwright.async_api import async_playwright  # noqa
    except ImportError as e:
        raise RuntimeError(
            'Knihovna Playwright není nainstalovaná na serveru.\n'
            'Doinstalujte:\n  pip install playwright\n  playwright install chromium\n'
            f'Detail: {e}')

    html = _navod_pdf_html()
    async with _get_navod_render_lock():
        async with async_playwright() as p:
            try:
                browser = await p.chromium.launch()
            except Exception as e:
                raise RuntimeError(
                    f'Nepodařilo se spustit Chromium: {e}\n'
                    'Spusťte na serveru: playwright install chromium')
            try:
                page = await browser.new_page()
                try:
                    await page.set_content(html, wait_until='load')
                    pdf = await page.pdf(
                        format='A4',
                        margin={'top': '12mm', 'right': '12mm',
                                'bottom': '12mm', 'left': '12mm'},
                        print_background=True)
                finally:
                    await page.close()
            finally:
                await browser.close()

    os.makedirs(EXPORT_DIR, exist_ok=True)
    cesta = os.path.join(EXPORT_DIR, 'Navod_na_nastaveni_podpisu.pdf')
    with open(cesta, 'wb') as f:
        f.write(pdf)
    return cesta


async def _stahni_navod_pdf():
    """Tlačítko „Stáhnout návod" — vyrenderuje PDF na serveru a pošle ho ke stažení."""
    notif = ui.notification('Připravuji PDF návodu…', type='ongoing',
                            position='top-right', spinner=True, timeout=None)
    try:
        cesta = await _vyrenderuj_navod_pdf()
        ui.download.file(cesta, 'Návod na nastavení podpisu.pdf')
        notif.dismiss()
    except Exception as e:
        notif.dismiss()
        ui.notify(f'Nepodařilo se vytvořit PDF návodu: {e}',
                  type='negative', position='top', timeout=15000, multi_line=True)


def _dialog_navod_podpis():
    with ui.dialog() as dlg, ui.card().classes(
            'w-full max-w-4xl p-0 rounded-xl overflow-hidden bg-gray-50'):
        # Hlavička dialogu
        with ui.row().classes('w-full items-center justify-between px-6 py-4 bg-blue-700'):
            with ui.row().classes('items-center gap-2'):
                ui.icon('menu_book', color='white')
                ui.label('Návod na nastavení podpisu').classes('text-xl font-bold text-white')
            ui.button(icon='close', on_click=dlg.close).props('flat round dense').classes('text-white')

        with ui.column().classes('w-full p-6 gap-4').style('max-height: 75vh; overflow-y: auto'):
            ui.label(
                'Tento návod obsahuje postupy pro dva e-mailové klienty — Microsoft Outlook '
                'a Mozilla Thunderbird. Postupujte podle návodu pro Váš e-mailový klient.'
            ).classes('text-gray-700')

            # ── Microsoft Outlook ────────────────────────────────
            with ui.card().classes('w-full p-5 bg-white rounded-lg border border-gray-200 gap-3'):
                with ui.row().classes('items-center gap-2'):
                    ui.icon('mail_outline', color='blue-700')
                    ui.label('Microsoft Outlook').classes('text-2xl font-bold text-blue-900')

                with ui.column().classes('gap-2 pl-1'):
                    with ui.row().classes('items-start gap-3'):
                        ui.label('1.').classes('font-bold text-blue-700 w-6 shrink-0')
                        ui.html(
                            'Stiskněte klávesovou zkratku '
                            '<span class="inline-block px-2 py-0.5 mx-1 border border-gray-400 '
                            'rounded bg-gray-100 font-mono text-sm">⊞ Win</span> + '
                            '<span class="inline-block px-2 py-0.5 mx-1 border border-gray-400 '
                            'rounded bg-gray-100 font-mono text-sm">R</span> na klávesnici.'
                        ).classes('text-gray-800')
                    with ui.row().classes('items-start gap-3'):
                        ui.label('2.').classes('font-bold text-blue-700 w-6 shrink-0')
                        with ui.column().classes('gap-1'):
                            ui.label('Do okna Spustit vložte tuto cestu a klikněte na OK '
                                     '(otevře se Vám příslušná složka):').classes('text-gray-800')
                            ui.label('%userprofile%\\appdata\\Roaming\\Microsoft\\Signatures') \
                                .classes('font-mono text-sm bg-gray-100 border border-gray-300 '
                                         'rounded px-2 py-1 select-all')
                    with ui.row().classes('items-start gap-3'):
                        ui.label('3.').classes('font-bold text-blue-700 w-6 shrink-0')
                        ui.label('Vložte obě přílohy do této složky.').classes('text-gray-800')
                    with ui.row().classes('items-start gap-3'):
                        ui.label('4.').classes('font-bold text-blue-700 w-6 shrink-0')
                        ui.label('Restartujte Outlook. Podpis bude nyní dostupný v nabídce '
                                 '„Podpis" v horní liště nového mailu, kde ho můžete použít '
                                 'na všechny nové zprávy, případně i odpovědi.') \
                            .classes('text-gray-800')

                ui.image(f'{NAVOD_URL}/image1.png') \
                    .classes('w-full max-w-2xl mx-auto rounded border border-gray-200 mt-2')

            # ── Mozilla Thunderbird ──────────────────────────────
            with ui.card().classes('w-full p-5 bg-white rounded-lg border border-gray-200 gap-3'):
                with ui.row().classes('items-center gap-2'):
                    ui.icon('mail', color='blue-700')
                    ui.label('Mozilla Thunderbird').classes('text-2xl font-bold text-blue-900')

                with ui.column().classes('gap-3 pl-1'):
                    with ui.row().classes('items-start gap-3'):
                        ui.label('1.').classes('font-bold text-blue-700 w-6 shrink-0')
                        ui.label('Uložte pouze soubor z přílohy tohoto mailu s koncovkou '
                                 '.htm do libovolného umístění v počítači.').classes('text-gray-800')
                    with ui.row().classes('items-start gap-3'):
                        ui.label('2.').classes('font-bold text-blue-700 w-6 shrink-0')
                        with ui.column().classes('gap-2'):
                            ui.label('V Thunderbirdu klikněte pravým tlačítkem na Váš účet '
                                     'a vyberte možnost „Nastavení".').classes('text-gray-800')
                            ui.image(f'{NAVOD_URL}/image2.png') \
                                .classes('max-w-sm rounded border border-gray-200')
                    with ui.row().classes('items-start gap-3'):
                        ui.label('3.').classes('font-bold text-blue-700 w-6 shrink-0')
                        ui.label('V novém okně zaškrtněte „Připojit podpis ze souboru '
                                 '(text, HTML nebo obrázek)".').classes('text-gray-800')
                    with ui.row().classes('items-start gap-3'):
                        ui.label('4.').classes('font-bold text-blue-700 w-6 shrink-0')
                        with ui.column().classes('gap-2'):
                            ui.label('Tlačítkem „Vybrat…" vyberte soubor .htm z umístění '
                                     'dle bodu č. 1.').classes('text-gray-800')
                            ui.image(f'{NAVOD_URL}/image3.png') \
                                .classes('w-full max-w-2xl rounded border border-gray-200')
                    with ui.row().classes('items-start gap-3'):
                        ui.label('5.').classes('font-bold text-blue-700 w-6 shrink-0')
                        ui.label('Hotovo, podpis se bude používat na všechny nové e-maily.') \
                            .classes('text-gray-800 font-semibold')

        with ui.row().classes('w-full justify-between items-center gap-2 px-6 py-3 bg-white border-t'):
            ui.button('Stáhnout návod (PDF)', icon='download',
                      on_click=_stahni_navod_pdf).props('unelevated no-caps') \
                .classes('bg-blue-700 hover:bg-blue-800 text-white font-bold') \
                .tooltip('Vygeneruje a stáhne tento návod jako PDF')
            ui.button('Zavřít', on_click=dlg.close).props('unelevated no-caps') \
                .classes('bg-gray-600 hover:bg-gray-700 text-white')
    dlg.open()


# =========================================================
# DIALOG: HISTORIE (👁 očičko)
# =========================================================
def _dialog_historie(z: dict):
    zaznamy = nacti_log(z['id'])
    with ui.dialog() as dlg, ui.card().classes('p-5 gap-3') \
            .style('min-width: 540px; max-width: 760px; max-height: 80vh; overflow-y: auto'):
        with ui.row().classes('items-center gap-2'):
            ui.icon('history', color='teal')
            ui.label(f'Historie žádosti #{z["id"]}').classes('text-lg font-bold text-gray-800')
        if not zaznamy:
            ui.label('U této žádosti zatím není žádný záznam.').classes('text-sm text-gray-500 italic py-4')
        else:
            for r in zaznamy:
                with ui.row().classes('w-full items-start gap-3 border-b border-gray-100 py-2'):
                    ui.label(_cas(r.get('kdy'))).classes('text-xs text-gray-400 font-mono whitespace-nowrap w-32 shrink-0')
                    with ui.column().classes('gap-0 flex-1 min-w-0'):
                        with ui.row().classes('items-center gap-2'):
                            ui.label(r.get('akce') or '').classes(
                                'text-xs font-bold px-2 py-0.5 rounded bg-teal-100 text-teal-800')
                            ui.label(r.get('jmeno') or '—').classes('text-xs text-gray-500')
                        if r.get('detail'):
                            ui.label(r['detail']).classes('text-sm text-gray-700 whitespace-pre-wrap mt-0.5')
        with ui.row().classes('justify-end w-full'):
            ui.button('Zavřít', on_click=dlg.close).props('flat no-caps')
    dlg.open()


# =========================================================
# DIALOG: SPRÁVA ČÍSELNÍKŮ (jen realizátor)
# =========================================================
def _dialog_ciselniky(user_name, prekresli):
    with ui.dialog() as dlg, ui.card().classes('p-0 rounded-xl w-full max-w-3xl overflow-hidden'):
        with ui.row().classes('w-full items-center justify-between px-6 py-4 bg-gray-100 border-b'):
            with ui.row().classes('items-center gap-2'):
                ui.icon('tune', color='primary')
                ui.label('Správa číselníků').classes('text-xl font-bold text-gray-800')
            ui.button(icon='close', on_click=dlg.close).props('flat round dense').classes('text-gray-500')

        with ui.tabs().classes('w-full') as taby:
            t_odd = ui.tab('Oddělení')
            t_pob = ui.tab('Pobočky')
            t_adr = ui.tab('Adresy')
        with ui.tab_panels(taby, value=t_odd).classes('w-full bg-white').style('max-height: 62vh; overflow-y: auto'):
            # --- ODDĚLENÍ ---
            with ui.tab_panel(t_odd):
                box = ui.column().classes('w-full gap-1')

                def prekresli_odd():
                    box.clear()
                    with box:
                        with ui.row().classes('w-full items-end gap-2 mb-2'):
                            inp = ui.input('Nové oddělení').props('outlined dense').classes('flex-1')

                            def pridej():
                                if not _s(inp.value):
                                    return ui.notify('Zadejte název oddělení.', type='warning')
                                ok, err = _ciselnik_pridej('vizitky_oddeleni', ('nazev',), (_s(inp.value),))
                                if err:
                                    return ui.notify(err, type='negative')
                                ui.notify('Přidáno.', type='positive')
                                prekresli_odd()
                            ui.button('Přidat', icon='add', on_click=pridej).props('unelevated no-caps').classes('bg-blue-600 text-white')
                        for nazev in nacti_oddeleni():
                            with ui.row().classes('w-full items-center justify-between px-3 py-1.5 rounded hover:bg-gray-50 border-b border-gray-50'):
                                ui.label(nazev).classes('text-sm text-gray-700')
                                ui.button(icon='delete', on_click=lambda n=nazev: (_smaz_oddeleni_dle_nazvu(n), prekresli_odd())) \
                                    .props('flat round dense color=red size=sm')
                prekresli_odd()
            # --- POBOČKY ---
            with ui.tab_panel(t_pob):
                box2 = ui.column().classes('w-full gap-1')

                def prekresli_pob():
                    box2.clear()
                    with box2:
                        with ui.row().classes('w-full items-end gap-2 mb-2'):
                            inp = ui.input('Nový kód pobočky').props('outlined dense').classes('flex-1')

                            def pridej():
                                if not _s(inp.value):
                                    return ui.notify('Zadejte kód pobočky.', type='warning')
                                ok, err = _ciselnik_pridej('vizitky_pobocky', ('kod',), (_s(inp.value),))
                                if err:
                                    return ui.notify(err, type='negative')
                                ui.notify('Přidáno.', type='positive')
                                prekresli_pob()
                            ui.button('Přidat', icon='add', on_click=pridej).props('unelevated no-caps').classes('bg-blue-600 text-white')
                        for kod in nacti_pobocky():
                            with ui.row().classes('w-full items-center justify-between px-3 py-1.5 rounded hover:bg-gray-50 border-b border-gray-50'):
                                ui.label(kod).classes('text-sm text-gray-700 font-mono')
                                ui.button(icon='delete', on_click=lambda k=kod: (_smaz_pobocku_dle_kodu(k), prekresli_pob())) \
                                    .props('flat round dense color=red size=sm')
                prekresli_pob()
            # --- ADRESY ---
            with ui.tab_panel(t_adr):
                box3 = ui.column().classes('w-full gap-1')

                def prekresli_adr():
                    box3.clear()
                    with box3:
                        with ui.row().classes('w-full items-end gap-2 mb-2'):
                            i_ul = ui.input('Ulice').props('outlined dense').classes('flex-1')
                            i_ps = ui.input('PSČ').props('outlined dense').classes('w-28')
                            i_me = ui.input('Město').props('outlined dense').classes('flex-1')

                            def pridej():
                                if not _s(i_ul.value):
                                    return ui.notify('Zadejte alespoň ulici.', type='warning')
                                ok, err = _ciselnik_pridej('vizitky_adresy', ('ulice', 'psc', 'mesto'),
                                                           (_s(i_ul.value), _s(i_ps.value), _s(i_me.value)))
                                if err:
                                    return ui.notify(err, type='negative')
                                ui.notify('Přidáno.', type='positive')
                                prekresli_adr()
                            ui.button('Přidat', icon='add', on_click=pridej).props('unelevated no-caps').classes('bg-blue-600 text-white')
                        for a in nacti_adresy():
                            with ui.row().classes('w-full items-center justify-between px-3 py-1.5 rounded hover:bg-gray-50 border-b border-gray-50'):
                                ui.label(f"{a['ulice']}, {a['psc']} {a['mesto']}").classes('text-sm text-gray-700')
                                ui.button(icon='delete', on_click=lambda i=a['id']: (_ciselnik_smaz('vizitky_adresy', i), prekresli_adr())) \
                                    .props('flat round dense color=red size=sm')
                prekresli_adr()

        with ui.row().classes('w-full justify-end px-6 py-3 bg-gray-50 border-t'):
            ui.button('Hotovo', on_click=lambda: (dlg.close(), prekresli())).props('unelevated no-caps').classes('bg-gray-700 text-white')
    dlg.open()


# =========================================================
# DIALOG: NOVÁ ŽÁDOST (formulář)
# =========================================================
def _dialog_formular(typ, user_id, user_name, user_email, prekresli):
    oddeleni_list = nacti_oddeleni()
    pobocky_list = nacti_pobocky()
    adresy_list = nacti_adresy()

    with ui.dialog() as dlg, ui.card().classes('w-full max-w-2xl p-0 rounded-xl overflow-hidden'):
        with ui.row().classes('w-full items-center gap-3 px-6 py-4 bg-blue-600'):
            ui.icon(TYP_IKONA.get(typ, 'badge'), color='white')
            ui.label(f'Nová žádost — {TYP_LABEL.get(typ, typ)}').classes('text-xl font-bold text-white')

        with ui.column().classes('w-full p-6 gap-3 bg-white').style('max-height: 68vh; overflow-y: auto'):
            i_org = ui.input('Organizace').props('outlined dense disable').classes('w-full')
            i_jmeno = ui.input('Jméno a příjmení', value=user_name or '').props('outlined dense').classes('w-full')
            i_odd = ui.select(oddeleni_list, label='Oddělení') \
                .props('outlined dense options-dense behavior=menu') \
                .classes('w-full').tooltip('Vyberte oddělení z rozbalovacího seznamu')
            with ui.row().classes('w-full gap-3'):
                i_tel = ui.input('Telefonní číslo').props('outlined dense').classes('flex-1')
                i_email = ui.input('E-mail', value=user_email or '').props('outlined dense').classes('flex-1')

            # Výběr adresy z číselníku → předvyplní pole
            adr_options = {a['id']: f"{a['ulice']}, {a['psc']} {a['mesto']}" for a in adresy_list}
            adr_map = {a['id']: a for a in adresy_list}
            sel_adr = ui.select(adr_options, label='Vybrat adresu z číselníku (předvyplní)', with_input=True) \
                .props('outlined dense clearable').classes('w-full')
            with ui.row().classes('w-full gap-3'):
                i_ulice = ui.input('Adresa – ulice').props('outlined dense').classes('flex-1')
                i_psc = ui.input('PSČ').props('outlined dense').classes('w-28')
                i_mesto = ui.input('Město').props('outlined dense').classes('flex-1')

            def _autofill_adr(e):
                a = adr_map.get(e.value)
                if a:
                    i_ulice.value = a['ulice']
                    i_psc.value = a['psc']
                    i_mesto.value = a['mesto']
            sel_adr.on_value_change(_autofill_adr)

            with ui.row().classes('w-full gap-3'):
                i_web = ui.input('Internetová adresa').props('outlined dense disable').classes('flex-1')
                i_pob = ui.select(pobocky_list, label='Kód pobočky', with_input=True, new_value_mode='add-unique') \
                    .props('outlined dense clearable').classes('flex-1')
            i_qr = ui.input('QR kód (pevná cesta — needitovatelné)', value=QR_CESTA) \
                .props('outlined dense disable').classes('w-full') \
                .tooltip('Pevná cesta, kterou grafik použije; needitovatelné.')

            with ui.row().classes('w-full gap-6 mt-1'):
                ch_podpis = ui.checkbox('E-mailový podpis', value=(typ == 'email'))
                ch_vizitka = ui.checkbox('Vizitka tisk', value=(typ == 'vizitka'))

            i_pozn = ui.textarea('Poznámka').props('outlined autogrow').classes('w-full')

        with ui.row().classes('w-full justify-between px-6 py-3 bg-gray-50 border-t'):
            ui.button('Zrušit', on_click=dlg.close).props('flat no-caps').classes('text-gray-600')

            async def odeslat():
                if not _s(i_jmeno.value) or not _s(i_odd.value):
                    return ui.notify('Vyplňte alespoň jméno a oddělení.', type='warning')
                data = {
                    'user_id': user_id,
                    'jmeno_zadavatele': user_name,
                    'email_zadavatele': user_email,
                    'typ': typ,
                    'stav': 'nova',
                    'organizace': _s(i_org.value),
                    'jmeno_prijmeni': _s(i_jmeno.value),
                    'oddeleni': _s(i_odd.value),
                    'telefon': _s(i_tel.value),
                    'email': _s(i_email.value),
                    'adresa_ulice': _s(i_ulice.value),
                    'adresa_psc': _s(i_psc.value),
                    'adresa_mesto': _s(i_mesto.value),
                    'web': _s(i_web.value),
                    'kod_pobocky': _s(i_pob.value),
                    'qr_code': QR_CESTA,
                    'email_podpis': 1 if ch_podpis.value else 0,
                    'vizitka_tisk': 1 if ch_vizitka.value else 0,
                    'poznamka': _s(i_pozn.value),
                }
                nove_id = await asyncio.to_thread(zaloz_zadost, data)
                if not nove_id:
                    return ui.notify('Žádost se nepodařilo uložit (chyba DB).', type='negative')
                zapis_log(nove_id, 'Vytvořeno', f'{TYP_LABEL.get(typ, typ)} — {data["jmeno_prijmeni"]}',
                          user_id, user_name)
                intranet_logger.log_activity(user_name, 'Vizitky',
                                             f'Nová žádost #{nove_id} ({TYP_LABEL.get(typ, typ)})')
                _odesli_emaily(
                    _emaily_realizatoru(),
                    f'Nová žádost o realizaci — {TYP_LABEL.get(typ, typ)} (#{nove_id})',
                    f'Uživatel {user_name} odeslal novou žádost o realizaci: '
                    f'{TYP_LABEL.get(typ, typ)} pro „{data["jmeno_prijmeni"]}".\n\n'
                    f'Přihlaste se do firemního portálu → modul „Vizitky a podpisy" a žádost zrealizujte.')
                ui.notify('Žádost byla odeslána k realizaci.', type='positive', position='top')
                dlg.close()
                prekresli()

            ui.button('Odeslat k realizaci', icon='send', on_click=odeslat) \
                .props('unelevated no-caps').classes('bg-green-600 hover:bg-green-700 text-white font-bold px-6 shadow-md')
    dlg.open()


# =========================================================
# DIALOG: VÝBĚR TYPU NOVÉ ŽÁDOSTI (4 volby)
# =========================================================
def _dialog_nova(user_id, user_name, user_email, prekresli):
    with ui.dialog() as dlg, ui.card().classes('w-full max-w-2xl p-6 rounded-xl gap-4'):
        ui.label('Co si přejete vytvořit?').classes('text-2xl font-bold text-gray-800')

        def info(text):
            with ui.dialog() as d2, ui.card().classes('w-full max-w-lg p-6 rounded-xl gap-3'):
                with ui.row().classes('items-center gap-2'):
                    ui.icon('info', color='amber-600')
                    ui.label('Jak postupovat při změně').classes('text-lg font-bold text-gray-800')
                ui.label(text).classes('text-sm text-gray-700 whitespace-pre-wrap')
                with ui.row().classes('w-full justify-end'):
                    ui.button('Rozumím', on_click=d2.close).props('unelevated no-caps').classes('bg-blue-600 text-white')
            d2.open()

        def volba_novy(typ):
            dlg.close()
            _dialog_formular(typ, user_id, user_name, user_email, prekresli)

        with ui.grid(columns=2).classes('w-full gap-3'):
            with ui.card().classes('p-4 cursor-pointer hover:shadow-lg border border-gray-200 rounded-xl items-center text-center gap-2') \
                    .on('click', lambda: volba_novy('email')):
                ui.icon('alternate_email', size='2.5rem', color='blue-600')
                ui.label('Nový e-mail podpis').classes('font-bold text-gray-800')
                ui.label('Vyplnit formulář').classes('text-xs text-gray-500')
            with ui.card().classes('p-4 cursor-pointer hover:shadow-lg border border-gray-200 rounded-xl items-center text-center gap-2') \
                    .on('click', lambda: volba_novy('vizitka')):
                ui.icon('badge', size='2.5rem', color='blue-600')
                ui.label('Nová vizitka').classes('font-bold text-gray-800')
                ui.label('Vyplnit formulář').classes('text-xs text-gray-500')
            with ui.card().classes('p-4 cursor-pointer hover:shadow-lg border border-amber-200 bg-amber-50 rounded-xl items-center text-center gap-2') \
                    .on('click', lambda: info(INFO_ZMENA_EMAIL)):
                ui.icon('edit', size='2.5rem', color='amber-600')
                ui.label('Změna v e-mail podpisu').classes('font-bold text-gray-800')
                ui.label('Jak na to').classes('text-xs text-gray-500')
            with ui.card().classes('p-4 cursor-pointer hover:shadow-lg border border-amber-200 bg-amber-50 rounded-xl items-center text-center gap-2') \
                    .on('click', lambda: info(INFO_ZMENA_VIZITKA)):
                ui.icon('edit', size='2.5rem', color='amber-600')
                ui.label('Změna ve vizitce').classes('font-bold text-gray-800')
                ui.label('Jak na to').classes('text-xs text-gray-500')

        with ui.row().classes('w-full justify-end'):
            ui.button('Zavřít', on_click=dlg.close).props('flat no-caps').classes('text-gray-600')
    dlg.open()


# =========================================================
# DIALOG: ŽÁDOST O ZMĚNU (žadatel u vyřízené žádosti)
# =========================================================
def _dialog_zmena(z, user_id, user_name, prekresli, zavri_detail):
    with ui.dialog() as dlg, ui.card().classes('w-full max-w-2xl p-6 rounded-xl gap-3'):
        with ui.row().classes('items-center gap-2'):
            ui.icon('published_with_changes', color='purple')
            ui.label(f'Žádost o změnu #{z["id"]}').classes('text-xl font-bold text-gray-800')
        ui.label('Vyberte zónu (pole), kterou chcete změnit, a napište novou hodnotu. '
                 'Vybraná pole se realizátorovi zvýrazní červeně.').classes('text-sm text-gray-500')

        sel = ui.select(ZONY_LABEL, multiple=True, label='Které zóny chcete změnit?') \
            .props('outlined dense use-chips').classes('w-full')
        pole_box = ui.column().classes('w-full gap-2')
        nove = {}

        def prekresli_pole():
            pole_box.clear()
            nove.clear()
            with pole_box:
                for k in (sel.value or []):
                    nove[k] = ui.input(f'Nová hodnota — {ZONY_LABEL.get(k, k)}',
                                       value=_s(z.get(k))).props('outlined dense').classes('w-full')
        sel.on_value_change(lambda e: prekresli_pole())

        i_pozn = ui.textarea('Poznámka ke změně').props('outlined autogrow').classes('w-full')

        with ui.row().classes('w-full justify-between mt-2'):
            ui.button('Zrušit', on_click=dlg.close).props('flat no-caps').classes('text-gray-600')

            async def odeslat():
                zmeny = {k: _s(inp.value) for k, inp in nove.items()}
                if not zmeny:
                    return ui.notify('Vyberte alespoň jednu zónu ke změně.', type='warning')
                await asyncio.to_thread(uloz_zadost_o_zmenu, z['id'], zmeny, _s(i_pozn.value))
                detail_txt = '; '.join(f'{ZONY_LABEL.get(k, k)} → {v}' for k, v in zmeny.items())
                zapis_log(z['id'], 'Žádost o změnu', detail_txt, user_id, user_name)
                intranet_logger.log_activity(user_name, 'Vizitky', f'Žádost o změnu #{z["id"]}')
                _odesli_emaily(
                    _emaily_realizatoru() + [z.get('email_zadavatele')],
                    f'Žádost o změnu — {TYP_LABEL.get(z.get("typ"), "")} (#{z["id"]})',
                    f'Uživatel {user_name} vytvořil žádost o změnu u žádosti #{z["id"]}.\n\n'
                    f'Požadované změny:\n{detail_txt}\n\nPoznámka: {_s(i_pozn.value) or "—"}')
                ui.notify('Žádost o změnu byla odeslána.', type='positive', position='top')
                dlg.close()
                zavri_detail()
                prekresli()

            ui.button('Odeslat žádost o změnu', icon='send', on_click=odeslat) \
                .props('unelevated no-caps').classes('bg-purple-600 hover:bg-purple-700 text-white font-bold px-6')
    dlg.open()


# =========================================================
# DIALOG: DOKONČENÍ REALIZACE (realizátor)
# =========================================================
def _dialog_dokonceni(z, user_id, user_name, prekresli, zavri_detail):
    je_zmena = (z.get('stav') == 'zmena')
    zmeny = {}
    if je_zmena:
        try:
            zmeny = json.loads(z.get('zmena_pozadavky') or '{}')
        except Exception:
            zmeny = {}
    with ui.dialog() as dlg, ui.card().classes('w-full max-w-xl p-6 rounded-xl gap-3'):
        with ui.row().classes('items-center gap-2'):
            ui.icon('task_alt', color='green')
            ui.label(('Provést změnu a potvrdit realizaci' if je_zmena else 'Dokončit realizaci')
                     + f' #{z["id"]}').classes('text-xl font-bold text-gray-800')

        if je_zmena and zmeny:
            with ui.column().classes('w-full gap-1 p-3 rounded-lg bg-purple-50 border border-purple-200'):
                ui.label('Požadované změny (po potvrzení se zapíšou do žádosti):').classes('text-xs font-bold text-purple-700 uppercase')
                for k, v in zmeny.items():
                    ui.label(f'• {ZONY_LABEL.get(k, k)}: {_s(z.get(k)) or "—"} → {v}').classes('text-sm text-purple-900')

        ui.label('Vložte podklady (hotové soubory ke stažení žadatelem):').classes('text-sm font-bold text-gray-700 mt-1')

        podklady_box = ui.column().classes('w-full gap-1')

        def _prekresli_podklady():
            podklady_box.clear()
            seznam = nacti_podklady(z['id'])
            with podklady_box:
                if not seznam:
                    ui.label('Zatím nevloženy žádné podklady.').classes('text-xs text-gray-400 italic')
                for p in seznam:
                    with ui.row().classes('w-full items-center justify-between px-3 py-1.5 bg-green-50 rounded-lg border border-green-100'):
                        with ui.row().classes('items-center gap-2 min-w-0'):
                            ui.icon('check_circle', color='green-600', size='sm')
                            ui.label(p['nazev_souboru']).classes('text-sm text-gray-800 truncate')
                        with ui.row().classes('items-center gap-0 shrink-0'):
                            ui.button(icon='visibility', on_click=lambda pp=p: ui.navigate.to(_podklad_url(pp['cesta']), new_tab=True)) \
                                .props('flat round dense size=sm').classes('text-blue-600').tooltip('Zobrazit (náhled)')
                            ui.button(icon='delete', on_click=lambda pp=p: (smaz_podklad(pp['id']), _prekresli_podklady())) \
                                .props('flat round dense size=sm color=red').tooltip('Odebrat')

        async def _on_upload(e):
            try:
                f = e.file  # NiceGUI 3.x: FileUpload (.name, .content_type, async read())
                raw = await f.read()
                ok = await asyncio.to_thread(uloz_podklad, z['id'], f.name, raw, user_name)
                if ok:
                    ui.notify(f'Soubor „{f.name}" nahrán.', type='positive')
                    _prekresli_podklady()
                else:
                    ui.notify('Soubor se nepodařilo uložit.', type='negative')
            except Exception as exc:
                ui.notify(f'Chyba nahrávání: {exc}', type='negative')

        ui.upload(on_upload=_on_upload, auto_upload=True, multiple=True, max_file_size=50_000_000,
                  label='Vybrat podklady').classes('w-full')
        _prekresli_podklady()

        i_pozn = ui.textarea('Poznámka pro žadatele (nepovinné)').props('outlined autogrow').classes('w-full')

        with ui.row().classes('w-full justify-between mt-2'):
            ui.button('Zrušit', on_click=dlg.close).props('flat no-caps').classes('text-gray-600')

            async def odeslat():
                await asyncio.to_thread(dokonci_realizaci, z['id'], _s(i_pozn.value), user_name,
                                        zmeny if je_zmena else None)
                if je_zmena:
                    detail_txt = '; '.join(f'{ZONY_LABEL.get(k, k)} → {v}' for k, v in zmeny.items())
                    zapis_log(z['id'], 'Změna realizována', detail_txt, user_id, user_name)
                    predmet = f'Změna byla realizována — {TYP_LABEL.get(z.get("typ"), "")} (#{z["id"]})'
                else:
                    zapis_log(z['id'], 'Realizováno',
                              f'Nahráno podkladů: {len(nacti_podklady(z["id"]))}', user_id, user_name)
                    predmet = f'Vaše žádost byla vyřízena — {TYP_LABEL.get(z.get("typ"), "")} (#{z["id"]})'
                intranet_logger.log_activity(user_name, 'Vizitky', f'Realizováno #{z["id"]}')
                _odesli_emaily(
                    [z.get('email_zadavatele')], predmet,
                    f'Vaše žádost #{z["id"]} ({TYP_LABEL.get(z.get("typ"), "")}) byla vyřízena.\n\n'
                    f'Přihlaste se do firemního portálu → modul „Vizitky a podpisy", '
                    f'otevřete žádost a stáhněte si podklady.'
                    + (f'\n\nPoznámka realizátora: {_s(i_pozn.value)}' if _s(i_pozn.value) else ''))
                ui.notify('Realizace dokončena, žadatel byl informován e-mailem.', type='positive', position='top')
                dlg.close()
                zavri_detail()
                prekresli()

            ui.button('Odeslat (dokončit)', icon='send', on_click=odeslat) \
                .props('unelevated no-caps').classes('bg-green-600 hover:bg-green-700 text-white font-bold px-6 shadow-md')
    dlg.open()


# =========================================================
# DIALOG: ÚPRAVA ŽÁDOSTI (jen realizátor) — oprava špatně vyplněných údajů
# =========================================================
def _dialog_uprava(z, user_id, user_name, prekresli, zavri_detail):
    """Realizátor opraví obsah žádosti (např. když ji žadatel vyplnil chybně).
    Mění jen údaje žádosti — stav ani workflow se nedotýká. Změny jdou do historie."""
    oddeleni_list = nacti_oddeleni()
    pobocky_list = nacti_pobocky()
    adresy_list = nacti_adresy()
    # Aktuální hodnoty musí být ve výběrech, i kdyby v číselníku nebyly (jinak by se nezobrazily).
    if _s(z.get('oddeleni')) and _s(z.get('oddeleni')) not in oddeleni_list:
        oddeleni_list = oddeleni_list + [_s(z.get('oddeleni'))]
    if _s(z.get('kod_pobocky')) and _s(z.get('kod_pobocky')) not in pobocky_list:
        pobocky_list = pobocky_list + [_s(z.get('kod_pobocky'))]

    with ui.dialog() as dlg, ui.card().classes('w-full max-w-2xl p-0 rounded-xl overflow-hidden'):
        with ui.row().classes('w-full items-center gap-3 px-6 py-4 bg-amber-600'):
            ui.icon('edit', color='white')
            ui.label(f'Upravit žádost #{z["id"]}').classes('text-xl font-bold text-white')

        with ui.column().classes('w-full p-6 gap-3 bg-white').style('max-height: 68vh; overflow-y: auto'):
            ui.label('Opravte nesprávně vyplněné údaje. Změny se uloží do žádosti a zapíší do historie (👁).') \
                .classes('text-sm text-gray-500')

            i_jmeno = ui.input('Jméno a příjmení', value=_s(z.get('jmeno_prijmeni'))) \
                .props('outlined dense').classes('w-full')
            i_odd = ui.select(oddeleni_list, label='Oddělení (pozice)',
                              value=_s(z.get('oddeleni')) or None,
                              with_input=True, new_value_mode='add-unique') \
                .props('outlined dense options-dense behavior=menu clearable').classes('w-full') \
                .tooltip('Pozice / funkce zobrazená v podpisu')
            with ui.row().classes('w-full gap-3'):
                i_tel = ui.input('Telefonní číslo', value=_s(z.get('telefon'))).props('outlined dense').classes('flex-1')
                i_email = ui.input('E-mail', value=_s(z.get('email'))).props('outlined dense').classes('flex-1')

            adr_options = {a['id']: f"{a['ulice']}, {a['psc']} {a['mesto']}" for a in adresy_list}
            adr_map = {a['id']: a for a in adresy_list}
            sel_adr = ui.select(adr_options, label='Vybrat adresu z číselníku (předvyplní)', with_input=True) \
                .props('outlined dense clearable').classes('w-full')
            with ui.row().classes('w-full gap-3'):
                i_ulice = ui.input('Adresa – ulice', value=_s(z.get('adresa_ulice'))).props('outlined dense').classes('flex-1')
                i_psc = ui.input('PSČ', value=_s(z.get('adresa_psc'))).props('outlined dense').classes('w-28')
                i_mesto = ui.input('Město', value=_s(z.get('adresa_mesto'))).props('outlined dense').classes('flex-1')

            def _autofill_adr(e):
                a = adr_map.get(e.value)
                if a:
                    i_ulice.value = a['ulice']
                    i_psc.value = a['psc']
                    i_mesto.value = a['mesto']
            sel_adr.on_value_change(_autofill_adr)

            i_pob = ui.select(pobocky_list, label='Kód pobočky',
                              value=_s(z.get('kod_pobocky')) or None,
                              with_input=True, new_value_mode='add-unique') \
                .props('outlined dense clearable').classes('w-full')

            with ui.row().classes('w-full gap-6 mt-1'):
                ch_podpis = ui.checkbox('E-mailový podpis', value=bool(z.get('email_podpis')))
                ch_vizitka = ui.checkbox('Vizitka tisk', value=bool(z.get('vizitka_tisk')))

            i_pozn = ui.textarea('Poznámka', value=_s(z.get('poznamka'))).props('outlined autogrow').classes('w-full')

        with ui.row().classes('w-full justify-between px-6 py-3 bg-gray-50 border-t'):
            ui.button('Zrušit', on_click=dlg.close).props('flat no-caps').classes('text-gray-600')

            async def uloz():
                if not _s(i_jmeno.value):
                    return ui.notify('Jméno a příjmení nesmí být prázdné.', type='warning')
                nova_data = {
                    'jmeno_prijmeni': _s(i_jmeno.value),
                    'oddeleni': _s(i_odd.value),
                    'telefon': _s(i_tel.value),
                    'email': _s(i_email.value),
                    'adresa_ulice': _s(i_ulice.value),
                    'adresa_psc': _s(i_psc.value),
                    'adresa_mesto': _s(i_mesto.value),
                    'kod_pobocky': _s(i_pob.value),
                    'email_podpis': 1 if ch_podpis.value else 0,
                    'vizitka_tisk': 1 if ch_vizitka.value else 0,
                    'poznamka': _s(i_pozn.value),
                }
                # Diff pro historii (čitelný popis, co se změnilo)
                popisky = {k: lbl for (k, lbl, _b) in FIELD_DISPLAY}
                popisky['poznamka'] = 'Poznámka'
                zmeny_log = []
                for k, nv in nova_data.items():
                    if k in ('email_podpis', 'vizitka_tisk'):
                        stara = 1 if z.get(k) else 0
                        if stara != nv:
                            zmeny_log.append(f'{popisky.get(k, k)}: '
                                             f'{"Ano" if stara else "Ne"} → {"Ano" if nv else "Ne"}')
                    else:
                        stara = _s(z.get(k))
                        if stara != nv:
                            zmeny_log.append(f'{popisky.get(k, k)}: {stara or "—"} → {nv or "—"}')
                if not zmeny_log:
                    ui.notify('Nebyla provedena žádná změna.', type='info')
                    return dlg.close()
                ok = await asyncio.to_thread(uloz_upravu_zadosti, z['id'], nova_data)
                if not ok:
                    return ui.notify('Úpravu se nepodařilo uložit (chyba DB).', type='negative')
                zapis_log(z['id'], 'Úprava realizátorem', '; '.join(zmeny_log), user_id, user_name)
                intranet_logger.log_activity(user_name, 'Vizitky', f'Úprava žádosti #{z["id"]}')
                ui.notify('Žádost byla upravena.', type='positive', position='top')
                dlg.close()
                zavri_detail()
                prekresli()

            ui.button('Uložit úpravy', icon='save', on_click=uloz) \
                .props('unelevated no-caps').classes('bg-amber-600 hover:bg-amber-700 text-white font-bold px-6 shadow-md')
    dlg.open()


# =========================================================
# DIALOG: SMAZÁNÍ ŽÁDOSTI (jen realizátor) — s potvrzením
# =========================================================
def _potvrd_smazani_zadosti(z, user_id, user_name, prekresli, zavri_detail):
    with ui.dialog() as dlg, ui.card().classes('p-6 rounded-xl gap-3 w-full max-w-md'):
        with ui.row().classes('items-center gap-2'):
            ui.icon('delete_forever', color='red').classes('text-2xl')
            ui.label(f'Smazat žádost #{z["id"]}?').classes('text-xl font-bold text-gray-800')
        ui.label('Žádost, vložené podklady i historie (👁) budou nevratně odstraněny.') \
            .classes('text-sm text-gray-600')

        async def proved():
            ok = await asyncio.to_thread(smaz_zadost, z['id'])
            if not ok:
                return ui.notify('Smazání se nezdařilo (chyba DB).', type='negative')
            intranet_logger.log_activity(user_name, 'Vizitky',
                                         f'Smazána žádost #{z["id"]} ({TYP_LABEL.get(z.get("typ"), "")})')
            ui.notify('Žádost byla smazána.', type='positive', position='top')
            dlg.close()
            zavri_detail()
            prekresli()

        with ui.row().classes('w-full justify-end gap-2 mt-2'):
            ui.button('Zrušit', on_click=dlg.close).props('flat no-caps').classes('text-gray-600')
            ui.button('Smazat', icon='delete_forever', on_click=proved) \
                .props('unelevated no-caps').classes('bg-red-600 hover:bg-red-700 text-white font-bold px-6')
    dlg.open()


# =========================================================
# DIALOG: DETAIL ŽÁDOSTI
# =========================================================
def _dialog_detail(z, user_id, user_name, je_realizator, prekresli):
    je_vlastnik = (z.get('user_id') == user_id)
    try:
        zmeny = json.loads(z.get('zmena_pozadavky') or '{}')
    except Exception:
        zmeny = {}

    # Scrolluje celá karta (pevný strop 90vh → přizpůsobí se monitoru, nikdy nepřeteče
    # mimo obrazovku). Hlavička a patička jsou „sticky", takže zůstanou vždy viditelné.
    # POZOR: třída `overflow-hidden` se NESMÍ použít — NiceGUI ji renderuje jako
    # `!important`, čímž přebije inline `overflow-y:auto` a obsah by se ořezával bez
    # scrollbaru. Overflow proto řešíme jen přes inline .style() (viz [[mojejipka_nicegui_dialog_scroll]]).
    with ui.dialog() as dlg, ui.card().classes(
            'w-full max-w-3xl p-0 rounded-xl bg-gray-50') \
            .style('max-height: 90vh; overflow-y: auto; overflow-x: hidden'):
        # Hlavička — přilepená nahoře
        with ui.column().classes('w-full px-6 py-4 bg-white border-b gap-1 sticky top-0 z-10'):
            with ui.row().classes('w-full items-center justify-between'):
                with ui.row().classes('items-center gap-2'):
                    ui.icon(TYP_IKONA.get(z.get('typ'), 'badge'), color='blue-600')
                    ui.label(f'{TYP_LABEL.get(z.get("typ"), z.get("typ"))} #{z["id"]}').classes('text-2xl font-black text-gray-800')
                with ui.row().classes('items-center gap-2'):
                    ui.label(STAV_LABEL.get(z.get('stav'), z.get('stav'))).classes(
                        f'px-3 py-1 rounded-full text-xs font-bold border uppercase {STAV_BADGE.get(z.get("stav"), "")}')
                    ui.button(icon='visibility', on_click=lambda: _dialog_historie(z)) \
                        .props('flat round dense').classes('text-teal-600').tooltip('Historie (log)')
                    ui.button(icon='close', on_click=dlg.close).props('flat round dense').classes('text-gray-500')
            with ui.row().classes('w-full gap-4 text-sm text-gray-500'):
                ui.label(f'Žadatel: {z.get("jmeno_zadavatele") or "—"}').classes('font-bold')
                ui.label(f'Vytvořeno: {_cas(z.get("vytvoreno"))}')
                if z.get('vyrizeno_at'):
                    ui.label(f'Vyřízeno: {_cas(z.get("vyrizeno_at"))}')

        # Tělo — bez vlastního scrollu, roluje celá karta (viz max-height na kartě).
        with ui.column().classes('w-full p-6 gap-4'):
            if z.get('stav') == 'zmena' and zmeny:
                with ui.row().classes('w-full items-center gap-2 p-3 rounded-lg bg-purple-100 border border-purple-300'):
                    ui.icon('priority_high', color='purple')
                    ui.label('Žadatel požaduje změnu zvýrazněných polí (červeně).').classes('text-sm font-bold text-purple-800')

            # Pole formuláře
            with ui.card().classes('w-full p-0 rounded-xl overflow-hidden border border-gray-200'):
                for i, (k, lbl, je_bool) in enumerate(FIELD_DISPLAY):
                    hodnota = z.get(k)
                    if je_bool:
                        zobraz = 'Ano' if hodnota else 'Ne'
                    else:
                        zobraz = _s(hodnota) or '—'
                    je_zmenene = k in zmeny
                    radek_bg = 'bg-red-50' if je_zmenene else ('bg-white' if i % 2 == 0 else 'bg-gray-50/60')
                    with ui.row().classes(f'w-full items-start gap-4 px-4 py-2 border-b border-gray-100 {radek_bg}'):
                        ui.label(lbl).classes('text-xs font-bold text-gray-500 uppercase w-44 shrink-0 pt-0.5')
                        with ui.column().classes('gap-0 flex-1 min-w-0'):
                            ui.label(zobraz).classes(
                                'text-sm ' + ('text-red-700 font-bold line-through' if je_zmenene else 'text-gray-800'))
                            if je_zmenene:
                                ui.label(f'➜ požadováno: {zmeny.get(k)}').classes('text-sm text-red-700 font-bold')

            if _s(z.get('poznamka')):
                with ui.column().classes('w-full gap-1'):
                    ui.label('Poznámka žadatele').classes('text-xs font-bold text-gray-500 uppercase')
                    ui.label(z['poznamka']).classes('text-sm text-gray-800 bg-white p-3 rounded-lg border border-gray-100 whitespace-pre-wrap')
            if z.get('stav') == 'zmena' and _s(z.get('zmena_poznamka')):
                with ui.column().classes('w-full gap-1'):
                    ui.label('Poznámka ke změně').classes('text-xs font-bold text-purple-600 uppercase')
                    ui.label(z['zmena_poznamka']).classes('text-sm text-purple-900 bg-purple-50 p-3 rounded-lg border border-purple-100 whitespace-pre-wrap')
            if _s(z.get('realizator_poznamka')):
                with ui.column().classes('w-full gap-1'):
                    ui.label('Poznámka realizátora').classes('text-xs font-bold text-green-600 uppercase')
                    ui.label(z['realizator_poznamka']).classes('text-sm text-green-900 bg-green-50 p-3 rounded-lg border border-green-100 whitespace-pre-wrap')

            # Podklady — náhled a stažení
            podklady = nacti_podklady(z['id'])
            if podklady:
                with ui.column().classes('w-full gap-1'):
                    ui.label('Podklady').classes('text-xs font-bold text-gray-500 uppercase')
                    for p in podklady:
                        with ui.row().classes('w-full items-center justify-between px-3 py-2 bg-white rounded-lg border border-gray-100'):
                            with ui.row().classes('items-center gap-2 min-w-0'):
                                ui.icon('attach_file', color='gray-500')
                                ui.label(p['nazev_souboru']).classes('text-sm text-gray-800 truncate')
                            with ui.row().classes('items-center gap-1 shrink-0'):
                                ui.button('Zobrazit', icon='visibility',
                                          on_click=lambda pp=p: ui.navigate.to(_podklad_url(pp['cesta']), new_tab=True)) \
                                    .props('outline no-caps size=sm color=primary')
                                ui.button('Stáhnout', icon='download',
                                          on_click=lambda pp=p: ui.download.file(pp['cesta'], pp['nazev_souboru'])) \
                                    .props('outline no-caps size=sm color=primary')

        # Patička — akce; přilepená dolů, vždy viditelná (tlačítka se případně zalomí)
        with ui.row().classes('w-full justify-end gap-2 px-6 py-3 bg-white border-t sticky bottom-0 z-10 flex-wrap'):
            if je_realizator:
                ui.button('Smazat žádost', icon='delete',
                          on_click=lambda: _potvrd_smazani_zadosti(z, user_id, user_name, prekresli, dlg.close)) \
                    .props('outline no-caps color=red').classes('mr-auto').tooltip('Nevratně smazat tuto žádost')
            if je_realizator and z.get('stav') != 'vyrizeno':
                ui.button('Upravit', icon='edit',
                          on_click=lambda: _dialog_uprava(z, user_id, user_name, prekresli, dlg.close)) \
                    .props('outline no-caps color=amber-8').tooltip('Opravit nesprávně vyplněné údaje žádosti')
            if je_realizator and z.get('stav') == 'nova':
                ui.button('Převzít (rozpracováno)', icon='play_arrow',
                          on_click=lambda: (oznac_rozpracovano(z['id']),
                                            zapis_log(z['id'], 'Rozpracováno', '', user_id, user_name),
                                            dlg.close(), prekresli())) \
                    .props('outline no-caps color=amber-8')
            if je_realizator and z.get('stav') in ('nova', 'rozpracovano'):
                ui.button('Dokončit realizaci', icon='task_alt',
                          on_click=lambda: _dialog_dokonceni(z, user_id, user_name, prekresli, dlg.close)) \
                    .props('unelevated no-caps').classes('bg-green-600 hover:bg-green-700 text-white font-bold')
            if je_realizator and z.get('stav') == 'zmena':
                ui.button('Provést změnu a potvrdit', icon='published_with_changes',
                          on_click=lambda: _dialog_dokonceni(z, user_id, user_name, prekresli, dlg.close)) \
                    .props('unelevated no-caps').classes('bg-green-600 hover:bg-green-700 text-white font-bold')
            if je_vlastnik and z.get('stav') == 'vyrizeno':
                ui.button('Zažádat o změnu', icon='published_with_changes',
                          on_click=lambda: _dialog_zmena(z, user_id, user_name, prekresli, dlg.close)) \
                    .props('unelevated no-caps').classes('bg-purple-600 hover:bg-purple-700 text-white font-bold')
            if je_realizator and z.get('email_podpis'):
                ui.button('Vygenerovat e-mailový podpis', icon='alternate_email',
                          on_click=lambda: _dialog_podpis(z, user_id, user_name, prekresli)) \
                    .props('unelevated no-caps').classes('bg-blue-600 hover:bg-blue-700 text-white font-bold') \
                    .tooltip('Vytvoří .htm a .txt podpis přímo z dat žádosti a vloží do podkladů')
            if je_realizator:
                ui.button('Stáhnout Excel', icon='grid_on', on_click=lambda: _stahni_excel(z)) \
                    .props('outline no-caps color=green').tooltip('Vygeneruje .xlsx pro skript (název = šablona, jméno/příjmení v jedné buňce)')
            ui.button('Zavřít', on_click=dlg.close).props('flat no-caps').classes('text-gray-600')
    dlg.open()


# =========================================================
# HLAVNÍ VYKRESLOVACÍ FUNKCE
# =========================================================
@ui.refreshable
def vykresli_vizitky(user_id, user_name, user_email, vsechna_prava):
    inicializace_vizitky_db()

    je_realizator = 'vse' in vsechna_prava or 'vizitky_realizator' in vsechna_prava
    je_zadatel = True  # žadatelem je automaticky každý přihlášený uživatel

    # Sdílené reference na UI prvky (naplní se až níže při stavbě UI). Pomocné funkce
    # na ně sahají přes tento slovník, takže jsou definované PŘED hlavičkou a případná
    # chyba při stavbě lišty nezpůsobí pád tlačítek v hlavičce (NameError na closure).
    refs = {}

    def _projde_filtrem(z) -> bool:
        f_jmeno = refs.get('f_jmeno'); f_stav = refs.get('f_stav'); f_odd = refs.get('f_odd')
        f_od = refs.get('f_od'); f_do = refs.get('f_do')
        if f_jmeno is not None and _s(f_jmeno.value):
            q = f_jmeno.value.lower()
            if q not in (_s(z.get('jmeno_zadavatele')).lower() + ' ' + _s(z.get('jmeno_prijmeni')).lower()):
                return False
        if f_stav is not None and f_stav.value and z.get('stav') != f_stav.value:
            return False
        if f_odd is not None and f_odd.value and _s(z.get('oddeleni')) != f_odd.value:
            return False
        vyt = z.get('vytvoreno')
        den = vyt.date() if hasattr(vyt, 'date') else None
        if den and f_od is not None and _s(f_od.value):
            try:
                if den < datetime.date.fromisoformat(f_od.value):
                    return False
            except ValueError:
                pass
        if den and f_do is not None and _s(f_do.value):
            try:
                if den > datetime.date.fromisoformat(f_do.value):
                    return False
            except ValueError:
                pass
        return True

    def _prekresli_seznam():
        seznam_box = refs.get('seznam_box')
        if seznam_box is None:
            return
        seznam_box.clear()
        zadosti = nacti_zadosti(None if je_realizator else user_id)
        zadosti = [z for z in zadosti if _projde_filtrem(z)]
        with seznam_box:
            if not zadosti:
                with ui.card().classes('w-full p-12 items-center justify-center bg-gray-50 border border-dashed border-gray-200 rounded-xl'):
                    ui.icon('badge', size='4rem', color='gray-400').classes('mb-2')
                    ui.label('Žádné žádosti k zobrazení.').classes('text-xl text-gray-500 font-bold')
                return
            with ui.grid(columns=1).classes('w-full gap-4 lg:grid-cols-2'):
                for z in zadosti:
                    karta_cls = STAV_KARTA.get(z.get('stav'), 'border-l-4 border-gray-300 bg-white')
                    with ui.card().classes(f'w-full p-4 rounded-xl shadow-sm hover:shadow-lg transition-shadow {karta_cls}'):
                        with ui.row().classes('w-full items-start justify-between'):
                            with ui.row().classes('items-center gap-2 min-w-0 cursor-pointer flex-1') \
                                    .on('click', lambda zz=z: _dialog_detail(zz, user_id, user_name, je_realizator, _prekresli_seznam)):
                                ui.icon(TYP_IKONA.get(z.get('typ'), 'badge'), color='blue-600')
                                with ui.column().classes('gap-0 min-w-0'):
                                    ui.label(f'{TYP_LABEL.get(z.get("typ"), z.get("typ"))} #{z["id"]}').classes('font-bold text-gray-800 truncate')
                                    ui.label(_s(z.get('jmeno_prijmeni')) or _s(z.get('jmeno_zadavatele'))).classes('text-sm text-gray-500 truncate')
                            with ui.column().classes('items-end gap-1 shrink-0'):
                                ui.label(STAV_LABEL.get(z.get('stav'), z.get('stav'))).classes(
                                    f'px-2 py-0.5 rounded text-[10px] font-bold uppercase border whitespace-nowrap {STAV_BADGE.get(z.get("stav"), "")}')
                                with ui.row().classes('items-center gap-0'):
                                    ui.button(icon='visibility', on_click=lambda zz=z: _dialog_historie(zz)) \
                                        .props('flat round dense size=sm').classes('text-teal-600').tooltip('Historie (log)')
                                    if je_realizator:
                                        ui.button(icon='grid_on', on_click=lambda zz=z: _stahni_excel(zz)) \
                                            .props('flat round dense size=sm color=green').tooltip('Stáhnout Excel (pro skript)')
                                        ui.button(icon='delete',
                                                  on_click=lambda zz=z: _potvrd_smazani_zadosti(zz, user_id, user_name, _prekresli_seznam, lambda: None)) \
                                            .props('flat round dense size=sm color=red').tooltip('Smazat žádost')
                        with ui.row().classes('w-full items-center justify-between mt-2 pt-2 border-t border-gray-100 text-xs text-gray-400 cursor-pointer') \
                                .on('click', lambda zz=z: _dialog_detail(zz, user_id, user_name, je_realizator, _prekresli_seznam)):
                            ui.label(f'🙋 {_s(z.get("jmeno_zadavatele")) or "—"}')
                            if _s(z.get('oddeleni')):
                                ui.label(f'🏢 {z.get("oddeleni")}').classes('truncate')
                            ui.label(f'🕒 {_cas(z.get("vytvoreno"))}')

    def _reset_filtr():
        for k in ('f_jmeno', 'f_stav', 'f_odd'):
            el = refs.get(k)
            if el is not None:
                el.value = ''
        for k in ('f_od', 'f_do'):
            el = refs.get(k)
            if el is not None:
                el.value = None
        _prekresli_seznam()

    # ── Hlavička ──────────────────────────────────────────────
    with ui.row().classes('w-full items-center justify-between mb-4'):
        with ui.column().classes('gap-1'):
            ui.label('Vizitky a e-mailové podpisy').classes('text-4xl font-extrabold text-gray-800')
            ui.label('Žádanky o výrobu či změnu vizitky nebo e-mailového podpisu.').classes('text-gray-500 text-sm')
        with ui.row().classes('items-center gap-2'):
            if je_realizator:
                ui.button('Číselníky', icon='tune', on_click=lambda: _dialog_ciselniky(user_name, vykresli_vizitky.refresh)) \
                    .props('outline no-caps').classes('text-gray-600').tooltip('Správa oddělení, poboček a adres')
            if je_zadatel:
                ui.button('Nová žádost', icon='add_circle',
                          on_click=lambda: _dialog_nova(user_id, user_name, user_email, _prekresli_seznam)) \
                    .classes('bg-green-600 hover:bg-green-700 text-white font-bold h-12 px-6 shadow-md rounded-xl')

    # ── Velké tlačítko – NÁVOD NA NASTAVENÍ PODPISU ──────────
    with ui.row().classes('w-full mb-6'):
        ui.button('NÁVOD NA NASTAVENÍ PODPISU', icon='menu_book',
                  on_click=lambda: _dialog_navod_podpis()) \
            .props('unelevated no-caps') \
            .classes('w-full bg-gradient-to-r from-blue-700 to-blue-500 hover:from-blue-800 '
                     'hover:to-blue-600 text-white font-extrabold text-lg h-14 shadow-md rounded-xl') \
            .tooltip('Otevře návod pro Microsoft Outlook a Mozilla Thunderbird')

    # ── Filtrační lišta ───────────────────────────────────────
    oddeleni_filtr = [''] + nacti_oddeleni()
    with ui.row().classes('w-full items-end gap-3 p-4 mb-4 bg-gray-50 rounded-xl border border-gray-200 flex-wrap'):
        f_jmeno = ui.input('Jméno žadatele / na vizitce').props('outlined dense clearable').classes('flex-1 min-w-[180px]')
        f_od = ui.input('Datum od').props('type=date outlined dense clearable').classes('w-40')
        f_do = ui.input('Datum do').props('type=date outlined dense clearable').classes('w-40')
        f_stav = ui.select({'': 'Vše', 'nova': 'Nová', 'rozpracovano': 'Rozpracováno',
                            'vyrizeno': 'Vyřízeno', 'zmena': 'Žádost o změnu'},
                           value='', label='Stav').props('outlined dense').classes('w-44')
        f_odd = ui.select(oddeleni_filtr, value='', label='Oddělení', with_input=True) \
            .props('outlined dense clearable').classes('w-56')
        ui.button('Filtrovat', icon='filter_alt', on_click=lambda: _prekresli_seznam()) \
            .props('unelevated no-caps').classes('bg-blue-600 text-white')
        ui.button(icon='restart_alt', on_click=lambda: _reset_filtr()) \
            .props('flat round dense').classes('text-gray-500').tooltip('Zrušit filtry')

    refs.update({
        'f_jmeno': f_jmeno, 'f_od': f_od, 'f_do': f_do,
        'f_stav': f_stav, 'f_odd': f_odd,
        'seznam_box': ui.column().classes('w-full gap-4'),
    })
    _prekresli_seznam()
